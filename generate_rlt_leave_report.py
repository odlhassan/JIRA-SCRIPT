"""
Generate a dedicated leave intelligence report for Jira project RLT.
"""
from __future__ import annotations

import argparse
import html
import json
import os
import re
import sqlite3
import uuid
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook

from jira_incremental_cache import (
    apply_overlap,
    bootstrap_default_checkpoint,
    get_cached_issue_payloads,
    get_cached_worklogs_for_subtasks,
    get_db_path,
    get_or_init_checkpoint,
    init_db,
    mark_full_sync,
    mark_missing_issues_deleted,
    needs_full_sync,
    parse_iso_utc,
    record_pipeline_artifact,
    set_checkpoint,
    upsert_issue_index,
    upsert_issue_payloads,
    upsert_worklog_payload,
    utc_now_iso,
)
from jira_client import BASE_URL, get_session

DEFAULT_PROJECT_KEY = "RLT"
DEFAULT_PROJECT_NAME = "RnD Leave Tracker"
DEFAULT_XLSX_OUT = "rlt_leave_report.xlsx"
DEFAULT_HTML_OUT = "rlt_leave_report.html"
DEFAULT_MD_OUT = "RLT_LEAVE_REPORT.md"
DEFAULT_WINDOW = "prev-current-next"
DEFAULT_START_DATE_FIELD = "customfield_10133"
LEAVE_TYPE_FIELD = "customfield_10584"
DEFAULT_STANDARD_HOURS_PER_DAY = 8.0
DEFAULT_RAMADAN_HOURS_PER_DAY = 6.5

PLANNED_KEYWORDS = ("planned", "annual", "casual", "considered in roadmap & queued")
UNPLANNED_KEYWORDS = ("sick", "emergency", "unplanned")
RLT_PIPELINE_NAME = "rlt_leave_report"
DEFAULT_OVERLAP_MINUTES = 5
DEFAULT_FORCE_FULL_SYNC_DAYS = 7
DEFAULT_BOOTSTRAP_DAYS = 365


@dataclass
class WorklogRow:
    issue_key: str
    started_raw: str
    started_date: str
    author: str
    hours_logged: float


@dataclass
class SubtaskRow:
    issue_key: str
    issue_id: str
    summary: str
    status: str
    assignee: str
    parent_task_key: str
    parent_task_assignee: str
    created: str
    updated: str
    start_date: str
    due_date: str
    original_estimate_hours: float
    timespent_hours: float
    leave_type_raw: str
    leave_classification: str
    total_worklog_hours: float
    planned_date_for_bucket: str
    clubbed_leave: str
    no_entry_flag: str


def _is_incremental_disabled(enable_incremental: bool) -> bool:
    if enable_incremental:
        return False
    return (_to_text(os.getenv("JIRA_INCREMENTAL_DISABLE", "1")) or "1") == "1"


def _get_overlap_minutes() -> int:
    raw = _to_text(os.getenv("JIRA_INCREMENTAL_OVERLAP_MINUTES", str(DEFAULT_OVERLAP_MINUTES)))
    try:
        return max(int(raw), 0)
    except ValueError:
        return DEFAULT_OVERLAP_MINUTES


def _get_force_full_days() -> int:
    raw = _to_text(os.getenv("JIRA_FORCE_FULL_SYNC_DAYS", str(DEFAULT_FORCE_FULL_SYNC_DAYS)))
    try:
        return max(int(raw), 1)
    except ValueError:
        return DEFAULT_FORCE_FULL_SYNC_DAYS


def _get_bootstrap_days() -> int:
    raw = _to_text(os.getenv("JIRA_INCREMENTAL_BOOTSTRAP_DAYS", str(DEFAULT_BOOTSTRAP_DAYS)))
    try:
        return max(int(raw), 1)
    except ValueError:
        return DEFAULT_BOOTSTRAP_DAYS


def _to_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _to_float_text(value: Any, default: float) -> float:
    text = _to_text(value)
    if not text:
        return default
    try:
        return float(text)
    except ValueError:
        return default


def _day_hours_profile_from_env() -> dict[str, Any]:
    standard_hours = _to_float_text(os.getenv("RLT_STANDARD_HOURS_PER_DAY"), DEFAULT_STANDARD_HOURS_PER_DAY)
    ramadan_hours = _to_float_text(os.getenv("RLT_RAMADAN_HOURS_PER_DAY"), DEFAULT_RAMADAN_HOURS_PER_DAY)
    standard_hours = standard_hours if standard_hours > 0 else DEFAULT_STANDARD_HOURS_PER_DAY
    ramadan_hours = ramadan_hours if ramadan_hours > 0 else DEFAULT_RAMADAN_HOURS_PER_DAY

    ramadan_start = _to_text(os.getenv("RLT_RAMADAN_START_DATE"))
    ramadan_end = _to_text(os.getenv("RLT_RAMADAN_END_DATE"))
    if bool(ramadan_start) != bool(ramadan_end):
        ramadan_start = ""
        ramadan_end = ""
    if ramadan_start and ramadan_end:
        start_value = parse_iso_date(ramadan_start)
        end_value = parse_iso_date(ramadan_end)
        if not start_value or not end_value or end_value < start_value:
            ramadan_start = ""
            ramadan_end = ""

    return {
        "standard_hours_per_day": standard_hours,
        "ramadan_hours_per_day": ramadan_hours,
        "ramadan_start_date": ramadan_start,
        "ramadan_end_date": ramadan_end,
    }


def _hours_per_day_for_iso_day(iso_day: str, profile: dict[str, Any]) -> float:
    day_value = parse_iso_date(iso_day)
    if not day_value:
        return float(profile.get("standard_hours_per_day", DEFAULT_STANDARD_HOURS_PER_DAY))
    start_text = _to_text(profile.get("ramadan_start_date"))
    end_text = _to_text(profile.get("ramadan_end_date"))
    if start_text and end_text:
        start_day = parse_iso_date(start_text)
        end_day = parse_iso_date(end_text)
        if start_day and end_day and start_day <= day_value <= end_day:
            return float(profile.get("ramadan_hours_per_day", DEFAULT_RAMADAN_HOURS_PER_DAY))
    return float(profile.get("standard_hours_per_day", DEFAULT_STANDARD_HOURS_PER_DAY))


def _daily_hours_to_days(daily_hours_by_day: dict[str, float], profile: dict[str, Any]) -> float:
    days = 0.0
    for iso_day, hours in daily_hours_by_day.items():
        day_hours = _hours_per_day_for_iso_day(iso_day, profile)
        if day_hours <= 0:
            continue
        days += float(hours or 0.0) / day_hours
    return round(days, 2)


def _seconds_to_hours(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    try:
        return round(float(value) / 3600.0, 2)
    except (TypeError, ValueError):
        return 0.0


def _month_bounds(year: int, month: int) -> tuple[date, date]:
    start = date(year, month, 1)
    if month == 12:
        next_start = date(year + 1, 1, 1)
    else:
        next_start = date(year, month + 1, 1)
    return start, next_start - timedelta(days=1)


def _shift_month(year: int, month: int, delta: int) -> tuple[int, int]:
    idx = (year * 12 + (month - 1)) + delta
    new_year = idx // 12
    new_month = (idx % 12) + 1
    return new_year, new_month


def parse_iso_date(value: Any) -> date | None:
    text = _to_text(value)
    if not text:
        return None
    if len(text) >= 10:
        candidate = text[:10]
        try:
            return date.fromisoformat(candidate)
        except ValueError:
            pass
    for fmt in (
        "%Y-%m-%dT%H:%M:%S.%f%z",
        "%Y-%m-%dT%H:%M:%S%z",
        "%Y-%m-%dT%H:%M:%S.%f",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%d",
    ):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00")).date()
    except ValueError:
        return None


MONTH_NAME_TO_NUMBER = {
    "jan": 1,
    "january": 1,
    "feb": 2,
    "february": 2,
    "mar": 3,
    "march": 3,
    "apr": 4,
    "april": 4,
    "may": 5,
    "jun": 6,
    "june": 6,
    "jul": 7,
    "july": 7,
    "aug": 8,
    "august": 8,
    "sep": 9,
    "sept": 9,
    "september": 9,
    "oct": 10,
    "october": 10,
    "nov": 11,
    "november": 11,
    "dec": 12,
    "december": 12,
}


def _parse_day_month_year_token(token: str, default_year: int | None = None) -> date | None:
    match = re.search(r"(?i)\b(\d{1,2})\s+([a-z]{3,9})(?:\s+(\d{4}))?\b", _to_text(token))
    if not match:
        return None
    day = int(match.group(1))
    month_text = _to_text(match.group(2)).lower()
    month = MONTH_NAME_TO_NUMBER.get(month_text)
    year = int(match.group(3)) if match.group(3) else default_year
    if not month or not year:
        return None
    try:
        return date(year, month, day)
    except ValueError:
        return None


def infer_date_range_from_summary(summary: str) -> tuple[str, str]:
    text = _to_text(summary)
    if not text:
        return "", ""
    match = re.search(
        r"(?i)\b(\d{1,2}\s+[a-z]{3,9}(?:\s+\d{4})?)\s*(?:-|to)\s*(\d{1,2}\s+[a-z]{3,9}(?:\s+\d{4})?)\b",
        text,
    )
    if not match:
        return "", ""
    left = _to_text(match.group(1))
    right = _to_text(match.group(2))
    right_year_match = re.search(r"\b(\d{4})\b", right)
    left_year_match = re.search(r"\b(\d{4})\b", left)
    right_year = int(right_year_match.group(1)) if right_year_match else None
    left_year = int(left_year_match.group(1)) if left_year_match else None
    if left_year is None:
        left_year = right_year
    if right_year is None:
        right_year = left_year
    left_date = _parse_day_month_year_token(left, default_year=left_year)
    right_date = _parse_day_month_year_token(right, default_year=right_year)
    if not left_date or not right_date:
        return "", ""
    if right_date < left_date:
        return "", ""
    return left_date.isoformat(), right_date.isoformat()


def normalize_subtask_dates(start_date: str, due_date: str, summary: str) -> tuple[str, str]:
    normalized_start = _to_text(start_date)
    normalized_due = _to_text(due_date)
    if normalized_start and normalized_due:
        return normalized_start, normalized_due
    inferred_start, inferred_due = infer_date_range_from_summary(summary)
    if not inferred_start or not inferred_due:
        return normalized_start, normalized_due
    if not normalized_start:
        normalized_start = inferred_start
    if not normalized_due:
        normalized_due = inferred_due
    return normalized_start, normalized_due


def iso_week_code(iso_date: str) -> str:
    dt = date.fromisoformat(iso_date)
    iso = dt.isocalendar()
    return f"{iso.year}-W{iso.week:02d}"


def month_code(iso_date: str) -> str:
    dt = date.fromisoformat(iso_date)
    return f"{dt.year:04d}-{dt.month:02d}"


def resolve_window_range(
    window: str = DEFAULT_WINDOW,
    from_date: str = "",
    to_date: str = "",
    today: date | None = None,
) -> tuple[str, str]:
    if from_date and to_date:
        return from_date, to_date

    now = today or datetime.now(timezone.utc).date()
    prev_y, prev_m = _shift_month(now.year, now.month, -1)
    next_y, next_m = _shift_month(now.year, now.month, 1)
    from_d, _ = _month_bounds(prev_y, prev_m)
    _, to_d = _month_bounds(next_y, next_m)
    return from_d.isoformat(), to_d.isoformat()


def leave_type_text(value: Any) -> str:
    if not value:
        return ""
    if isinstance(value, dict):
        parent = _to_text(value.get("value") or value.get("name"))
        child = value.get("child")
        if isinstance(child, dict):
            child_text = _to_text(child.get("value") or child.get("name"))
            if child_text:
                return f"{parent} > {child_text}" if parent else child_text
        return parent
    return _to_text(value)


def classify_leave(leave_type_raw: str, status: str, summary: str) -> str:
    leave_type_low = _to_text(leave_type_raw).lower()
    if "unplanned" in leave_type_low:
        return "Unplanned"
    if "planned" in leave_type_low:
        return "Planned"

    text = f"{_to_text(status)} {_to_text(summary)}".lower()
    if any(token in text for token in UNPLANNED_KEYWORDS):
        return "Unplanned"
    if any(token in text for token in PLANNED_KEYWORDS):
        return "Planned"
    return "Unknown"


def choose_planned_date(start_date: str, due_date: str) -> str:
    start_d = parse_iso_date(start_date)
    due_d = parse_iso_date(due_date)
    if start_d:
        return start_d.isoformat()
    if due_d:
        return due_d.isoformat()
    return ""


def is_clubbed_leave(logged_hours: float, estimate_hours: float, start_date: str, due_date: str) -> bool:
    if logged_hours > DEFAULT_STANDARD_HOURS_PER_DAY:
        return True
    if estimate_hours > DEFAULT_STANDARD_HOURS_PER_DAY:
        return True
    start_d = parse_iso_date(start_date)
    due_d = parse_iso_date(due_date)
    return bool(start_d and due_d and due_d > start_d)


def is_defective_no_entry(
    classification: str,
    logged_hours: float,
    estimate_hours: float,
    start_date: str,
    due_date: str,
) -> bool:
    if classification != "Planned" or logged_hours > 0:
        return False
    has_date = bool(parse_iso_date(start_date) or parse_iso_date(due_date))
    has_estimate = estimate_hours > 0
    return not (has_date and has_estimate)


def _fetch_issues(session, jql: str, fields: list[str]) -> list[dict]:
    url = f"{BASE_URL}/rest/api/3/search/jql"
    issues: list[dict] = []
    token = None
    while True:
        payload: dict[str, Any] = {"jql": jql, "maxResults": 100, "fields": fields}
        if token:
            payload["nextPageToken"] = token
        response = session.post(url, json=payload)
        response.raise_for_status()
        data = response.json()
        issues.extend(data.get("issues", []) or [])
        token = data.get("nextPageToken")
        if not token:
            return issues


def _fetch_issues_by_keys(session, issue_keys: list[str], fields: list[str]) -> list[dict]:
    if not issue_keys:
        return []
    rows: list[dict] = []
    for i in range(0, len(issue_keys), 500):
        chunk = issue_keys[i : i + 500]
        keys_clause = ", ".join(f'"{k}"' for k in chunk)
        jql = f"key in ({keys_clause})"
        rows.extend(_fetch_issues(session, jql, fields))
    return rows


def _candidate_rows_from_issues(issues: list[dict]) -> list[dict]:
    now_utc = utc_now_iso()
    rows: list[dict] = []
    for issue in issues:
        fields = issue.get("fields", {}) or {}
        rows.append(
            {
                "issue_id": _to_text(issue.get("id")),
                "issue_key": _to_text(issue.get("key")),
                "updated_utc": _to_text(fields.get("updated")),
                "issue_type": _to_text((fields.get("issuetype") or {}).get("name")),
                "project_key": _to_text((fields.get("project") or {}).get("key")),
                "last_seen_utc": now_utc,
                "is_deleted": 0,
            }
        )
    return [row for row in rows if row["issue_id"] and row["issue_key"]]


def _classify_candidates(
    conn: sqlite3.Connection,
    candidates: list[dict],
) -> tuple[list[str], int, int]:
    issue_ids = [_to_text(item.get("issue_id")) for item in candidates if _to_text(item.get("issue_id"))]
    existing_updated: dict[str, str] = {}
    for offset in range(0, len(issue_ids), 900):
        chunk = issue_ids[offset : offset + 900]
        if not chunk:
            continue
        placeholders = ",".join("?" for _ in chunk)
        rows = conn.execute(
            f"SELECT issue_id, updated_utc FROM issue_index WHERE issue_id IN ({placeholders})",
            tuple(chunk),
        ).fetchall()
        for row in rows:
            existing_updated[_to_text(row[0])] = _to_text(row[1])

    changed: list[str] = []
    seen: set[str] = set()
    new_count = 0
    changed_existing_count = 0
    for item in candidates:
        issue_id = _to_text(item.get("issue_id"))
        issue_key = _to_text(item.get("issue_key"))
        updated_utc = _to_text(item.get("updated_utc"))
        if not issue_id or not issue_key:
            continue
        if issue_id not in existing_updated:
            new_count += 1
            if issue_key not in seen:
                changed.append(issue_key)
                seen.add(issue_key)
            continue
        if existing_updated.get(issue_id, "") != updated_utc:
            changed_existing_count += 1
            if issue_key not in seen:
                changed.append(issue_key)
                seen.add(issue_key)
    return changed, new_count, changed_existing_count


def _fetch_worklogs_for_issue(session, issue_key: str) -> list[dict]:
    url = f"{BASE_URL}/rest/api/3/issue/{issue_key}/worklog"
    start_at = 0
    out: list[dict] = []
    while True:
        response = session.get(url, params={"startAt": start_at, "maxResults": 100})
        response.raise_for_status()
        data = response.json()
        rows = data.get("worklogs", []) or []
        out.extend(rows)
        start_at += len(rows)
        total = int(data.get("total", len(out)) or 0)
        if not rows or start_at >= total:
            return out


def _normalize_worklogs(issue_key: str, logs: list[dict]) -> list[WorklogRow]:
    out: list[WorklogRow] = []
    for log in logs:
        started_raw = _to_text(log.get("started"))
        started = parse_iso_date(started_raw)
        hours = _seconds_to_hours(log.get("timeSpentSeconds"))
        if not started or hours <= 0:
            continue
        out.append(
            WorklogRow(
                issue_key=issue_key,
                started_raw=started_raw,
                started_date=started.isoformat(),
                author=_to_text((log.get("author") or {}).get("displayName")) or "Unknown",
                hours_logged=round(hours, 2),
            )
        )
    return out


def _in_date_window(iso_date: str, from_date: str, to_date: str) -> bool:
    if not iso_date:
        return False
    d = date.fromisoformat(iso_date)
    return date.fromisoformat(from_date) <= d <= date.fromisoformat(to_date)


def _weekday_dates_inclusive(start_iso: str, due_iso: str) -> list[str]:
    start = parse_iso_date(start_iso)
    due = parse_iso_date(due_iso)
    if not start or not due or due < start:
        return []
    out: list[str] = []
    current = start
    while current <= due:
        if current.weekday() < 5:  # Mon-Fri
            out.append(current.isoformat())
        current += timedelta(days=1)
    return out


def _calendar_dates_inclusive(start_iso: str, due_iso: str) -> list[str]:
    start = parse_iso_date(start_iso)
    due = parse_iso_date(due_iso)
    if not start or not due or due < start:
        return []
    out: list[str] = []
    current = start
    while current <= due:
        out.append(current.isoformat())
        current += timedelta(days=1)
    return out


def _split_hours_evenly(total_hours: float, slots: int) -> list[float]:
    if total_hours <= 0 or slots <= 0:
        return []
    base = round(total_hours / slots, 2)
    values = [base for _ in range(slots)]
    values[-1] = round(total_hours - sum(values[:-1]), 2)
    return values


def _distribute_hours_with_daily_cap(total_hours: float, slots: int, per_day_cap: float) -> list[float]:
    if total_hours <= 0 or slots <= 0 or per_day_cap <= 0:
        return []
    remaining = round(total_hours, 2)
    out: list[float] = []
    for _ in range(slots):
        if remaining <= 0:
            out.append(0.0)
            continue
        allocation = min(per_day_cap, remaining)
        allocation = round(allocation, 2)
        out.append(allocation)
        remaining = round(remaining - allocation, 2)
    return out


def _should_redistribute_subtask(
    subtask: SubtaskRow,
    total_logged_hours: float,
    weekday_dates: list[str],
    day_hours_profile: dict[str, Any],
) -> bool:
    if total_logged_hours <= DEFAULT_STANDARD_HOURS_PER_DAY:
        return False
    if len(weekday_dates) < 1:
        return False
    start = parse_iso_date(subtask.start_date)
    due = parse_iso_date(subtask.due_date)
    if not start or not due or due < start:
        return False
    return True


def _should_redistribute_estimate(subtask: SubtaskRow) -> bool:
    start = parse_iso_date(subtask.start_date)
    due = parse_iso_date(subtask.due_date)
    if not start or not due or due < start:
        return False
    return subtask.original_estimate_hours > DEFAULT_STANDARD_HOURS_PER_DAY


def _redistribute_continuous_leave_subtasks(
    subtasks: list[SubtaskRow],
    day_hours_profile: dict[str, Any],
) -> list[SubtaskRow]:
    out: list[SubtaskRow] = []
    for subtask in subtasks:
        calendar_dates = _calendar_dates_inclusive(subtask.start_date, subtask.due_date)
        if not _should_redistribute_estimate(subtask) or not calendar_dates:
            out.append(subtask)
            continue

        estimate_parts = _distribute_hours_with_daily_cap(
            round(subtask.original_estimate_hours, 2),
            len(calendar_dates),
            DEFAULT_STANDARD_HOURS_PER_DAY,
        )
        worklog_parts = _distribute_hours_with_daily_cap(
            round(subtask.total_worklog_hours, 2),
            len(calendar_dates),
            DEFAULT_STANDARD_HOURS_PER_DAY,
        )
        timespent_parts = _distribute_hours_with_daily_cap(
            round(subtask.timespent_hours, 2),
            len(calendar_dates),
            DEFAULT_STANDARD_HOURS_PER_DAY,
        )
        for i, day in enumerate(calendar_dates):
            out.append(
                SubtaskRow(
                    issue_key=subtask.issue_key,
                    issue_id=subtask.issue_id,
                    summary=subtask.summary,
                    status=subtask.status,
                    assignee=subtask.assignee,
                    parent_task_key=subtask.parent_task_key,
                    parent_task_assignee=subtask.parent_task_assignee,
                    created=subtask.created,
                    updated=subtask.updated,
                    start_date=day,
                    due_date=day,
                    original_estimate_hours=estimate_parts[i] if i < len(estimate_parts) else 0.0,
                    timespent_hours=timespent_parts[i] if i < len(timespent_parts) else 0.0,
                    leave_type_raw=subtask.leave_type_raw,
                    leave_classification=subtask.leave_classification,
                    total_worklog_hours=worklog_parts[i] if i < len(worklog_parts) else 0.0,
                    planned_date_for_bucket=day,
                    clubbed_leave=subtask.clubbed_leave,
                    no_entry_flag=subtask.no_entry_flag,
                )
            )
    return out


def _redistribute_continuous_leave_worklogs(
    subtasks: list[SubtaskRow],
    worklogs: list[WorklogRow],
    day_hours_profile: dict[str, Any],
) -> list[WorklogRow]:
    subtask_by_key = {s.issue_key: s for s in subtasks}
    logs_by_issue: dict[str, list[WorklogRow]] = defaultdict(list)
    for log in worklogs:
        logs_by_issue[log.issue_key].append(log)

    redistributed: list[WorklogRow] = []
    for issue_key, issue_logs in logs_by_issue.items():
        subtask = subtask_by_key.get(issue_key)
        if not subtask:
            redistributed.extend(issue_logs)
            continue

        weekday_dates = _weekday_dates_inclusive(subtask.start_date, subtask.due_date)
        total_logged_hours = round(sum(w.hours_logged for w in issue_logs), 2)
        if not _should_redistribute_subtask(subtask, total_logged_hours, weekday_dates, day_hours_profile):
            redistributed.extend(issue_logs)
            continue

        distributed_hours = _distribute_hours_with_daily_cap(
            total_logged_hours,
            len(weekday_dates),
            DEFAULT_STANDARD_HOURS_PER_DAY,
        )
        author = issue_logs[0].author if issue_logs else "Unknown"
        for day, hours in zip(weekday_dates, distributed_hours):
            if hours <= 0:
                continue
            redistributed.append(
                WorklogRow(
                    issue_key=issue_key,
                    started_raw=f"redistributed:{subtask.start_date}->{subtask.due_date}",
                    started_date=day,
                    author=author,
                    hours_logged=round(hours, 2),
                )
            )

    return sorted(redistributed, key=lambda w: (w.started_date, w.issue_key, w.author))


def _compute_aggregates(
    subtasks: list[SubtaskRow],
    worklogs: list[WorklogRow],
    from_date: str,
    to_date: str,
    day_hours_profile: dict[str, Any],
) -> dict[str, Any]:
    by_assignee: dict[str, dict[str, float]] = defaultdict(
        lambda: {
            "planned_taken_hours": 0.0,
            "unplanned_taken_hours": 0.0,
            "planned_not_taken_hours": 0.0,
            "planned_not_taken_no_entry_count": 0.0,
            "unknown_subtasks_count": 0.0,
        }
    )
    weekly: dict[tuple[str, str], dict[str, float]] = defaultdict(
        lambda: {
            "planned_taken_hours": 0.0,
            "unplanned_taken_hours": 0.0,
            "planned_not_taken_hours": 0.0,
        }
    )
    daily: dict[tuple[str, str], dict[str, float]] = defaultdict(
        lambda: {
            "planned_taken_hours": 0.0,
            "unplanned_taken_hours": 0.0,
            "planned_not_taken_hours": 0.0,
        }
    )
    monthly: dict[tuple[str, str], dict[str, float]] = defaultdict(
        lambda: {
            "planned_taken_hours": 0.0,
            "unplanned_taken_hours": 0.0,
            "planned_not_taken_hours": 0.0,
        }
    )

    defective: list[dict[str, Any]] = []
    clubbed: list[dict[str, Any]] = []
    subtask_by_key = {s.issue_key: s for s in subtasks}
    assignee_planned_taken_daily: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    assignee_unplanned_taken_daily: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    assignee_planned_not_taken_daily: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))

    for subtask in subtasks:
        assignee = subtask.assignee
        if subtask.leave_classification == "Unknown":
            by_assignee[assignee]["unknown_subtasks_count"] += 1
            defective.append(
                {
                    "issue_key": subtask.issue_key,
                    "assignee": assignee,
                    "summary": subtask.summary,
                    "status": subtask.status,
                    "leave_classification": "Unknown",
                    "reason": "Unknown leave classification",
                    "planned_dates": subtask.planned_date_for_bucket or "No Entry",
                    "original_estimate_hours": subtask.original_estimate_hours or "No Entry",
                }
            )

        if subtask.clubbed_leave == "Yes":
            clubbed.append(
                {
                    "issue_key": subtask.issue_key,
                    "assignee": assignee,
                    "summary": subtask.summary,
                    "leave_classification": subtask.leave_classification,
                    "status": subtask.status,
                    "logged_hours": subtask.total_worklog_hours,
                    "estimate_hours": subtask.original_estimate_hours,
                    "start_date": subtask.start_date or "No Entry",
                    "due_date": subtask.due_date or "No Entry",
                }
            )

        if subtask.no_entry_flag == "Yes":
            by_assignee[assignee]["planned_not_taken_no_entry_count"] += 1
            defective.append(
                {
                    "issue_key": subtask.issue_key,
                    "assignee": assignee,
                    "summary": subtask.summary,
                    "status": subtask.status,
                    "leave_classification": subtask.leave_classification,
                    "reason": "No Entry for planned date and/or original estimate",
                    "planned_dates": subtask.planned_date_for_bucket or "No Entry",
                    "original_estimate_hours": subtask.original_estimate_hours if subtask.original_estimate_hours > 0 else "No Entry",
                }
            )

        if (
            subtask.leave_classification == "Planned"
            and subtask.total_worklog_hours <= 0
            and subtask.no_entry_flag != "Yes"
        ):
            planned_hours = subtask.original_estimate_hours if subtask.original_estimate_hours > 0 else 0.0
            calendar_dates = _calendar_dates_inclusive(subtask.start_date, subtask.due_date)
            distributed = _should_redistribute_estimate(subtask) and bool(calendar_dates)
            targets = (
                list(
                    zip(
                        calendar_dates,
                        _distribute_hours_with_daily_cap(
                            planned_hours,
                            len(calendar_dates),
                            DEFAULT_STANDARD_HOURS_PER_DAY,
                        ),
                    )
                )
                if distributed
                else [(subtask.planned_date_for_bucket, planned_hours)]
            )
            for planned_date, planned_part in targets:
                if not planned_date or planned_part <= 0:
                    continue
                if not _in_date_window(planned_date, from_date, to_date):
                    continue
                by_assignee[assignee]["planned_not_taken_hours"] += planned_part
                daily[(assignee, planned_date)]["planned_not_taken_hours"] += planned_part
                assignee_planned_not_taken_daily[assignee][planned_date] += planned_part
                weekly[(assignee, iso_week_code(planned_date))]["planned_not_taken_hours"] += planned_part
                monthly[(assignee, month_code(planned_date))]["planned_not_taken_hours"] += planned_part

    for log in worklogs:
        subtask = subtask_by_key.get(log.issue_key)
        if not subtask:
            continue
        if not _in_date_window(log.started_date, from_date, to_date):
            continue
        if subtask.leave_classification not in ("Planned", "Unplanned"):
            continue

        assignee = subtask.assignee
        week_key = iso_week_code(log.started_date)
        month_key = month_code(log.started_date)
        if subtask.leave_classification == "Planned":
            by_assignee[assignee]["planned_taken_hours"] += log.hours_logged
            daily[(assignee, log.started_date)]["planned_taken_hours"] += log.hours_logged
            assignee_planned_taken_daily[assignee][log.started_date] += log.hours_logged
            weekly[(assignee, week_key)]["planned_taken_hours"] += log.hours_logged
            monthly[(assignee, month_key)]["planned_taken_hours"] += log.hours_logged
        else:
            by_assignee[assignee]["unplanned_taken_hours"] += log.hours_logged
            daily[(assignee, log.started_date)]["unplanned_taken_hours"] += log.hours_logged
            assignee_unplanned_taken_daily[assignee][log.started_date] += log.hours_logged
            weekly[(assignee, week_key)]["unplanned_taken_hours"] += log.hours_logged
            monthly[(assignee, month_key)]["unplanned_taken_hours"] += log.hours_logged

    assignee_summary = []
    for assignee, m in sorted(by_assignee.items(), key=lambda x: x[0].lower()):
        p = round(m["planned_taken_hours"], 2)
        u = round(m["unplanned_taken_hours"], 2)
        f = round(m["planned_not_taken_hours"], 2)
        assignee_summary.append(
            {
                "assignee": assignee,
                "planned_taken_hours": p,
                "unplanned_taken_hours": u,
                "planned_not_taken_hours": f,
                "planned_not_taken_no_entry_count": int(m["planned_not_taken_no_entry_count"]),
                "unknown_subtasks_count": int(m["unknown_subtasks_count"]),
                "planned_taken_days": _daily_hours_to_days(assignee_planned_taken_daily[assignee], day_hours_profile),
                "unplanned_taken_days": _daily_hours_to_days(assignee_unplanned_taken_daily[assignee], day_hours_profile),
                "planned_not_taken_days": _daily_hours_to_days(assignee_planned_not_taken_daily[assignee], day_hours_profile),
            }
        )

    weekly_rows = []
    for (assignee, period), m in sorted(weekly.items(), key=lambda x: (x[0][1], x[0][0].lower())):
        total = m["planned_taken_hours"] + m["unplanned_taken_hours"] + m["planned_not_taken_hours"]
        weekly_rows.append(
            {
                "assignee": assignee,
                "period_week": period,
                "planned_taken_hours": round(m["planned_taken_hours"], 2),
                "unplanned_taken_hours": round(m["unplanned_taken_hours"], 2),
                "planned_not_taken_hours": round(m["planned_not_taken_hours"], 2),
                "total_hours": round(total, 2),
            }
        )

    daily_rows = []
    for (assignee, period), m in sorted(daily.items(), key=lambda x: (x[0][1], x[0][0].lower())):
        total = m["planned_taken_hours"] + m["unplanned_taken_hours"] + m["planned_not_taken_hours"]
        daily_rows.append(
            {
                "assignee": assignee,
                "period_day": period,
                "planned_taken_hours": round(m["planned_taken_hours"], 2),
                "unplanned_taken_hours": round(m["unplanned_taken_hours"], 2),
                "planned_not_taken_hours": round(m["planned_not_taken_hours"], 2),
                "total_hours": round(total, 2),
            }
        )

    monthly_rows = []
    for (assignee, period), m in sorted(monthly.items(), key=lambda x: (x[0][1], x[0][0].lower())):
        total = m["planned_taken_hours"] + m["unplanned_taken_hours"] + m["planned_not_taken_hours"]
        monthly_rows.append(
            {
                "assignee": assignee,
                "period_month": period,
                "planned_taken_hours": round(m["planned_taken_hours"], 2),
                "unplanned_taken_hours": round(m["unplanned_taken_hours"], 2),
                "planned_not_taken_hours": round(m["planned_not_taken_hours"], 2),
                "total_hours": round(total, 2),
            }
        )

    return {
        "assignee_summary": assignee_summary,
        "daily": daily_rows,
        "weekly": weekly_rows,
        "monthly": monthly_rows,
        "defective": sorted(defective, key=lambda r: (r["assignee"].lower(), r["issue_key"])),
        "clubbed": sorted(clubbed, key=lambda r: (r["assignee"].lower(), r["issue_key"])),
    }


def _sheet_append_rows(ws, headers: list[str], rows: list[dict[str, Any]]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append([row.get(col, "") for col in headers])


def _write_xlsx(
    output_path: Path,
    subtasks: list[SubtaskRow],
    distributed_subtasks: list[SubtaskRow],
    worklogs: list[WorklogRow],
    aggregates: dict[str, Any],
) -> None:
    wb = Workbook()
    ws_raw = wb.active
    ws_raw.title = "Raw_Subtasks"
    raw_headers = [
        "issue_key", "issue_id", "summary", "status", "assignee", "parent_task_key", "parent_task_assignee",
        "created", "updated", "start_date", "due_date", "original_estimate_hours", "timespent_hours",
        "leave_type_raw", "leave_classification", "total_worklog_hours", "planned_date_for_bucket", "clubbed_leave", "no_entry_flag",
    ]
    ws_raw.append(raw_headers)
    for s in subtasks:
        ws_raw.append([
            s.issue_key, s.issue_id, s.summary, s.status, s.assignee, s.parent_task_key, s.parent_task_assignee,
            s.created, s.updated, s.start_date, s.due_date, s.original_estimate_hours, s.timespent_hours,
            s.leave_type_raw, s.leave_classification, s.total_worklog_hours, s.planned_date_for_bucket, s.clubbed_leave, s.no_entry_flag,
        ])

    ws_distributed = wb.create_sheet("Subtasks_Distributed")
    ws_distributed.append(raw_headers)
    for s in distributed_subtasks:
        ws_distributed.append([
            s.issue_key, s.issue_id, s.summary, s.status, s.assignee, s.parent_task_key, s.parent_task_assignee,
            s.created, s.updated, s.start_date, s.due_date, s.original_estimate_hours, s.timespent_hours,
            s.leave_type_raw, s.leave_classification, s.total_worklog_hours, s.planned_date_for_bucket, s.clubbed_leave, s.no_entry_flag,
        ])

    ws_worklogs = wb.create_sheet("Worklogs_Normalized")
    ws_worklogs.append(["issue_key", "started_raw", "started_date", "author", "hours_logged"])
    for w in worklogs:
        ws_worklogs.append([w.issue_key, w.started_raw, w.started_date, w.author, w.hours_logged])

    ws_assignee = wb.create_sheet("Assignee_Summary")
    _sheet_append_rows(ws_assignee, [
        "assignee", "planned_taken_hours", "unplanned_taken_hours", "planned_not_taken_hours",
        "planned_not_taken_no_entry_count", "unknown_subtasks_count", "planned_taken_days",
        "unplanned_taken_days", "planned_not_taken_days",
    ], aggregates["assignee_summary"])

    ws_daily = wb.create_sheet("Daily_Assignee")
    _sheet_append_rows(ws_daily, [
        "assignee", "period_day", "planned_taken_hours", "unplanned_taken_hours", "planned_not_taken_hours", "total_hours",
    ], aggregates["daily"])

    ws_weekly = wb.create_sheet("Weekly_Assignee")
    _sheet_append_rows(ws_weekly, [
        "assignee", "period_week", "planned_taken_hours", "unplanned_taken_hours", "planned_not_taken_hours", "total_hours",
    ], aggregates["weekly"])

    ws_monthly = wb.create_sheet("Monthly_Assignee")
    _sheet_append_rows(ws_monthly, [
        "assignee", "period_month", "planned_taken_hours", "unplanned_taken_hours", "planned_not_taken_hours", "total_hours",
    ], aggregates["monthly"])

    ws_defective = wb.create_sheet("Defective_NoEntry")
    _sheet_append_rows(ws_defective, [
        "issue_key", "assignee", "summary", "status", "leave_classification", "reason", "planned_dates", "original_estimate_hours",
    ], aggregates["defective"])

    ws_clubbed = wb.create_sheet("Clubbed_Leaves")
    _sheet_append_rows(ws_clubbed, [
        "issue_key", "assignee", "summary", "leave_classification", "status", "logged_hours", "estimate_hours", "start_date", "due_date",
    ], aggregates["clubbed"])

    wb.save(output_path)

def _markdown_table(headers: list[str], rows: list[list[str]]) -> str:
    head = "| " + " | ".join(headers) + " |"
    sep = "| " + " | ".join("---" for _ in headers) + " |"
    body = ["| " + " | ".join(row) + " |" for row in rows]
    return "\n".join([head, sep, *body]) if body else "\n".join([head, sep])


def _write_md(output_path: Path, project_key: str, project_name: str, from_date: str, to_date: str, aggregates: dict[str, Any]) -> None:
    summary = aggregates["assignee_summary"]
    total_planned_taken = round(sum(r["planned_taken_hours"] for r in summary), 2)
    total_unplanned_taken = round(sum(r["unplanned_taken_hours"] for r in summary), 2)
    total_planned_not_taken = round(sum(r["planned_not_taken_hours"] for r in summary), 2)
    total_no_entry = int(sum(r["planned_not_taken_no_entry_count"] for r in summary))

    rows = [
        [
            _to_text(r["assignee"]),
            f'{r["planned_taken_hours"]:.2f}',
            f'{r["unplanned_taken_hours"]:.2f}',
            f'{r["planned_not_taken_hours"]:.2f}',
            _to_text(r["planned_not_taken_no_entry_count"]),
            _to_text(r["unknown_subtasks_count"]),
        ]
        for r in summary
    ]

    text = "\n".join([
        f"# {project_key} Leave Intelligence Report",
        "",
        "## Project and Window",
        f"- Project Key: `{project_key}`",
        f"- Project Name: `{project_name}`",
        f"- Reporting Window: `{from_date}` to `{to_date}`",
        "",
        "## Executive Summary",
        f"- Planned Taken (hours): `{total_planned_taken:.2f}`",
        f"- Unplanned Taken (hours): `{total_unplanned_taken:.2f}`",
        f"- Planned Not Yet Taken (hours): `{total_planned_not_taken:.2f}`",
        f"- Planned Not Yet Taken (No Entry count): `{total_no_entry}`",
        f"- Defective subtasks listed: `{len(aggregates['defective'])}`",
        f"- Clubbed leave subtasks: `{len(aggregates['clubbed'])}`",
        "",
        "## Assignee-wise Summary",
        _markdown_table([
            "Assignee", "Planned Taken (h)", "Unplanned Taken (h)", "Planned Not Yet Taken (h)", "No Entry Count", "Unknown Count",
        ], rows),
        "",
        "## Defective and No Entry",
        "- `No Entry` means planned leave subtask is missing planned date and/or original estimate while no hours are logged.",
        "- Unknown classification subtasks are excluded from planned/unplanned totals.",
        "",
        "## Clubbed Leave",
        "- Clubbed leave means one subtask represents more than one day (for example logged/estimated hours > 8 or multi-day date span).",
        "",
        "## Data-Quality Notes",
        "- Month/week forecasts use Jira date fields only.",
        "- Subtasks without Jira dates are not bucketed into week/month and are reported as data-quality issues.",
        "- Hours are primary; days are derived by date-aware hours/day (Ramadan dates use Ramadan hours/day; other dates use standard hours/day).",
    ])
    output_path.write_text(text, encoding="utf-8")


def _json_default(value: Any):
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return str(value)


def _html_table(headers: list[str], rows: list[dict[str, Any]], table_id: str) -> str:
    head = "".join(f"<th>{html.escape(col)}</th>" for col in headers)
    body_rows = []
    for row in rows:
        cells = []
        for col in headers:
            text = _to_text(row.get(col, ""))
            css = "warn-badge" if text == "No Entry" else ""
            cells.append(f'<td class="{css}">{html.escape(text)}</td>')
        body_rows.append("<tr>" + "".join(cells) + "</tr>")
    body = "".join(body_rows) if body_rows else f'<tr><td colspan="{len(headers)}">No rows</td></tr>'
    return f'<table id="{table_id}"><thead><tr>{head}</tr></thead><tbody>{body}</tbody></table>'


def _write_html(output_path: Path, project_key: str, project_name: str, from_date: str, to_date: str, subtasks: list[SubtaskRow], aggregates: dict[str, Any]) -> None:
    day_profile = _day_hours_profile_from_env()
    payload = {
        "subtasks": [s.__dict__ for s in subtasks],
        "assignee_summary": aggregates["assignee_summary"],
        "daily": aggregates["daily"],
        "weekly": aggregates["weekly"],
        "monthly": aggregates["monthly"],
        "defective": aggregates["defective"],
        "clubbed": aggregates["clubbed"],
        "day_profile": day_profile,
    }
    payload_json = json.dumps(payload, default=_json_default)

    total_planned_taken = round(sum(r["planned_taken_hours"] for r in aggregates["assignee_summary"]), 2)
    total_unplanned_taken = round(sum(r["unplanned_taken_hours"] for r in aggregates["assignee_summary"]), 2)
    total_taken = round(total_planned_taken + total_unplanned_taken, 2)
    total_taken_days = round(
        sum(r.get("planned_taken_days", 0.0) + r.get("unplanned_taken_days", 0.0) for r in aggregates["assignee_summary"]),
        2,
    )
    total_future = round(sum(r["planned_not_taken_hours"] for r in aggregates["assignee_summary"]), 2)
    total_no_entry = int(sum(r["planned_not_taken_no_entry_count"] for r in aggregates["assignee_summary"]))

    summary_headers = [
        "Assignee",
        "Planned Taken (h)",
        "Unplanned Taken (h)",
        "Planned Not Yet Taken (h)",
        "No Entry Count",
        "Unknown Count",
    ]
    summary_rows = [
        {
            "Assignee": r["assignee"],
            "Planned Taken (h)": f'{r["planned_taken_hours"]:.2f}',
            "Unplanned Taken (h)": f'{r["unplanned_taken_hours"]:.2f}',
            "Planned Not Yet Taken (h)": f'{r["planned_not_taken_hours"]:.2f}',
            "No Entry Count": r["planned_not_taken_no_entry_count"],
            "Unknown Count": r["unknown_subtasks_count"],
        }
        for r in aggregates["assignee_summary"]
    ]
    weekly_rows = [
        {
            "Assignee": r["assignee"],
            "Week": r["period_week"],
            "Planned Taken (h)": f'{r["planned_taken_hours"]:.2f}',
            "Unplanned Taken (h)": f'{r["unplanned_taken_hours"]:.2f}',
            "Future Planned (h)": f'{r["planned_not_taken_hours"]:.2f}',
            "Total (h)": f'{r["total_hours"]:.2f}',
        }
        for r in aggregates["weekly"]
    ]
    daily_rows = [
        {
            "Assignee": r["assignee"],
            "Day": r["period_day"],
            "Planned Taken (h)": f'{r["planned_taken_hours"]:.2f}',
            "Unplanned Taken (h)": f'{r["unplanned_taken_hours"]:.2f}',
            "Future Planned (h)": f'{r["planned_not_taken_hours"]:.2f}',
            "Total (h)": f'{r["total_hours"]:.2f}',
        }
        for r in aggregates["daily"]
    ]
    monthly_rows = [
        {
            "Assignee": r["assignee"],
            "Month": r["period_month"],
            "Planned Taken (h)": f'{r["planned_taken_hours"]:.2f}',
            "Unplanned Taken (h)": f'{r["unplanned_taken_hours"]:.2f}',
            "Future Planned (h)": f'{r["planned_not_taken_hours"]:.2f}',
            "Total (h)": f'{r["total_hours"]:.2f}',
        }
        for r in aggregates["monthly"]
    ]

    doc = f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{html.escape(project_key)} Leave Intelligence Report</title>
  <link rel="stylesheet" href="https://fonts.googleapis.com/css2?family=Material+Symbols+Outlined:opsz,wght,FILL,GRAD@20..48,500,0,0">
  <style>
    :root {{
      --bg: #eef3ef;
      --panel: #ffffff;
      --ink: #1a2a1f;
      --muted: #5b6f62;
      --line: #d7e4da;
      --accent: #0f766e;
      --planned: #14532d;
      --unplanned: #9a3412;
      --future: #1d4ed8;
      --warn-bg: #fff1e8;
      --warn-ink: #9a3412;
      --kpi-capacity-bg: #fee2e2;
      --kpi-capacity-line: #fca5a5;
      --kpi-capacity-ink: #7f1d1d;
      --kpi-planned-bg: #dbeafe;
      --kpi-planned-line: #93c5fd;
      --kpi-planned-ink: #1e3a8a;
      --kpi-actual-bg: #f3e8ff;
      --kpi-actual-line: #d8b4fe;
      --kpi-actual-ink: #a855f7;
      --kpi-gap-bg: #e0f2fe;
      --kpi-gap-line: #7dd3fc;
      --kpi-gap-ink: #2563eb;
      --kpi-leaves-bg: #fecaca;
      --kpi-leaves-line: #f87171;
      --kpi-leaves-ink: #b91c1c;
      --kpi-capacity-gap-bg: #fef2f2;
      --kpi-capacity-gap-line: #fecaca;
      --kpi-capacity-gap-ink: #ef4444;
    }}
    * {{ box-sizing: border-box; }}
    body {{ margin: 0; color: var(--ink); background: linear-gradient(140deg, #f6f9f7, var(--bg)); font: 14px/1.45 "Segoe UI", Tahoma, sans-serif; }}
    .wrap {{ max-width: 1380px; margin: 0 auto; padding: 18px; }}
    .hero {{ border: 1px solid var(--line); border-radius: 16px; background: radial-gradient(circle at top right, #e9f5ef 0, #fff 45%); padding: 18px; }}
    .hero h1 {{ margin: 0; font-size: 32px; letter-spacing: .2px; }}
    .hero-meta {{ margin-top: 6px; color: var(--muted); font-weight: 600; }}
    .stats {{ margin-top: 14px; display: grid; grid-template-columns: repeat(5, minmax(140px, 1fr)); gap: 10px; }}
    .stat {{ border: 1px solid var(--line); border-radius: 12px; background: #fff; padding: 10px 12px; }}
    .stat .k {{ color: var(--muted); font-size: 12px; }}
    .stat .k {{
      display: inline-flex;
      align-items: center;
      gap: 6px;
    }}
    .stat-info {{
      position: relative;
      display: inline-flex;
      align-items: center;
      justify-content: center;
      width: 15px;
      height: 15px;
      border-radius: 999px;
      border: 1px solid #94a3b8;
      color: #334155;
      background: #f8fafc;
      font-size: 10px;
      line-height: 1;
      font-weight: 700;
      cursor: default;
      user-select: none;
    }}
    .stat-info-tip {{
      position: absolute;
      left: 50%;
      top: calc(100% + 8px);
      transform: translateX(-50%);
      min-width: 230px;
      max-width: 320px;
      padding: 8px 10px;
      border-radius: 8px;
      border: 1px solid #cbd5e1;
      background: #fff;
      color: #1f2937;
      font-size: 12px;
      line-height: 1.35;
      box-shadow: 0 10px 24px rgba(15, 23, 42, 0.18);
      opacity: 0;
      visibility: hidden;
      pointer-events: none;
      z-index: 30;
      transition: opacity 0.14s ease;
    }}
    .stat-info:hover .stat-info-tip,
    .stat-info:focus .stat-info-tip,
    .stat-info:focus-visible .stat-info-tip {{
      opacity: 1;
      visibility: visible;
    }}
    .stat .v {{ font-size: 24px; font-weight: 800; line-height: 1.1; }}
    .stat .s {{ margin-top: 4px; font-size: 12px; color: var(--muted); font-weight: 700; }}
    .stat.total {{ background: var(--kpi-capacity-bg); border-color: var(--kpi-capacity-line); }}
    .stat.planned {{ background: var(--kpi-planned-bg); border-color: var(--kpi-planned-line); }}
    .stat.unplanned {{ background: var(--kpi-actual-bg); border-color: var(--kpi-actual-line); }}
    .stat.future {{ background: var(--kpi-gap-bg); border-color: var(--kpi-gap-line); }}
    .stat.leaves {{ background: var(--kpi-leaves-bg); border-color: var(--kpi-leaves-line); }}
    .stat.warn {{ background: var(--kpi-capacity-gap-bg); border-color: var(--kpi-capacity-gap-line); }}
    .stat.total .v {{ color: var(--kpi-capacity-ink); }}
    .stat.planned .v {{ color: var(--kpi-planned-ink); }}
    .stat.unplanned .v {{ color: var(--kpi-actual-ink); }}
    .stat.future .v {{ color: var(--kpi-gap-ink); }}
    .stat.leaves .v {{ color: var(--kpi-leaves-ink); }}
    .stat.warn .v {{ color: var(--kpi-capacity-gap-ink); }}
    .controls {{ margin-top: 14px; border: 1px solid var(--line); border-radius: 12px; background: var(--panel); padding: 12px; display: flex; flex-wrap: wrap; gap: 10px; align-items: end; }}
    label {{ font-size: 12px; color: var(--muted); font-weight: 700; }}
    select, input {{ margin-top: 4px; display: block; min-width: 220px; border: 1px solid #c8d8cd; border-radius: 9px; background: #fff; padding: 7px 9px; color: var(--ink); }}
    .grid {{ display: grid; gap: 12px; margin-top: 12px; }}
    .panel {{ border: 1px solid var(--line); border-radius: 12px; background: var(--panel); padding: 10px; overflow: auto; }}
    .panel h2 {{ margin: 2px 0 10px 0; font-size: 18px; }}
    .tabs {{ display: flex; gap: 8px; margin-bottom: 10px; }}
    .tab-btn {{ border: 1px solid #c8d8cd; background: #f8fcfa; color: #234033; border-radius: 9px; padding: 6px 12px; cursor: pointer; font-weight: 700; }}
    .tab-btn.active {{ background: #e2f2e9; border-color: #8fc2a7; }}
    .tab-pane {{ display: none; }}
    .tab-pane.active {{ display: block; }}
    table {{ width: 100%; border-collapse: collapse; font-size: 13px; }}
    thead th {{ position: sticky; top: 0; background: #edf6f0; z-index: 1; }}
    th, td {{ border-bottom: 1px solid #e6eee8; padding: 7px 8px; text-align: left; white-space: nowrap; }}
    .warn-badge {{ background: var(--warn-bg); color: var(--warn-ink); font-weight: 700; border-radius: 6px; padding: 2px 6px; }}
    .hidden {{ display: none; }}
    @media (max-width: 980px) {{
      .stats {{ grid-template-columns: 1fr 1fr; }}
      select, input {{ min-width: 150px; width: 100%; }}
    }}
  </style>
  <link rel="stylesheet" href="shared-nav.css">
</head>
<body>
  <div class="wrap">
    <section class="hero">
      <h1>{html.escape(project_key)} · {html.escape(project_name)}</h1>
      <div class="hero-meta">Window: <span id="window-label">{html.escape(from_date)} to {html.escape(to_date)}</span></div>
      <div class="stats">
        <div class="stat total"><div class="k">Total Taken <span class="stat-info" tabindex="0">i<span class="stat-info-tip">Planned Taken + Unplanned Taken within selected date range. Days are derived from the configured day-hours profile.</span></span></div><div class="v" id="stat-total-taken-hours">{total_taken:.2f}h</div><div class="s" id="stat-total-taken-days">{total_taken_days:.2f}d</div></div>
        <div class="stat planned"><div class="k">Planned Taken <span class="stat-info" tabindex="0">i<span class="stat-info-tip">Logged leave hours classified as Planned within selected date range.</span></span></div><div class="v" id="stat-planned-taken-hours">{total_planned_taken:.2f}h</div></div>
        <div class="stat unplanned"><div class="k">Unplanned Taken <span class="stat-info" tabindex="0">i<span class="stat-info-tip">Logged leave hours classified as Unplanned within selected date range.</span></span></div><div class="v" id="stat-unplanned-taken-hours">{total_unplanned_taken:.2f}h</div></div>
        <div class="stat leaves"><div class="k">Total Leaves Planned <span class="stat-info" tabindex="0">i<span class="stat-info-tip">Planned Taken + Future Planned within selected date range.</span></span></div><div class="v" id="stat-total-planned-leaves-hours">0.00h</div></div>
        <div class="stat future"><div class="k">Future Planned <span class="stat-info" tabindex="0">i<span class="stat-info-tip">Planned leave hours not yet logged in the selected range.</span></span></div><div class="v" id="stat-future-hours">{total_future:.2f}h</div></div>
        <div class="stat warn"><div class="k">No Entry <span class="stat-info" tabindex="0">i<span class="stat-info-tip">Count of planned leave items missing required planned date and/or estimate metadata.</span></span></div><div class="v" id="stat-no-entry">{total_no_entry}</div></div>
      </div>
    </section>

    <section class="controls">
      <label>From<input id="from-date" type="date"></label>
      <label>To<input id="to-date" type="date"></label>
      <label>Assignee<select id="assignee"><option value="">All</option></select></label>
      <label>Leave Type<select id="leave"><option value="">All</option><option>Planned</option><option>Unplanned</option><option>Unknown</option></select></label>
      <label>Status<input id="status" placeholder="contains..."></label>
    </section>

    <div class="grid">
      <section class="panel"><h2>Assignee KPI</h2>{_html_table(summary_headers, summary_rows, "assignee-table")}</section>
      <section class="panel">
        <h2>Time Perspective</h2>
        <div class="tabs">
          <button class="tab-btn active" data-tab="daily">Daily</button>
          <button class="tab-btn" data-tab="weekly">Weekly</button>
          <button class="tab-btn" data-tab="monthly">Monthly</button>
        </div>
        <div id="tab-daily" class="tab-pane active">{_html_table(list(daily_rows[0].keys()) if daily_rows else ["Assignee","Day","Planned Taken (h)","Unplanned Taken (h)","Future Planned (h)","Total (h)"], daily_rows, "daily-table")}</div>
        <div id="tab-weekly" class="tab-pane">{_html_table(list(weekly_rows[0].keys()) if weekly_rows else ["Assignee","Week","Planned Taken (h)","Unplanned Taken (h)","Future Planned (h)","Total (h)"], weekly_rows, "weekly-table")}</div>
        <div id="tab-monthly" class="tab-pane">{_html_table(list(monthly_rows[0].keys()) if monthly_rows else ["Assignee","Month","Planned Taken (h)","Unplanned Taken (h)","Future Planned (h)","Total (h)"], monthly_rows, "monthly-table")}</div>
      </section>
      <section class="panel"><h2>Defective / No Entry</h2>{_html_table(["issue_key","assignee","summary","status","leave_classification","reason","planned_dates","original_estimate_hours"], aggregates['defective'], "defective-table")}</section>
      <section class="panel"><h2>Clubbed Leave Subtasks</h2>{_html_table(["issue_key","assignee","summary","leave_classification","status","logged_hours","estimate_hours","start_date","due_date"], aggregates['clubbed'], "clubbed-table")}</section>
    </div>
  </div>
<script>
const payload={payload_json};
const assigneeEl=document.getElementById('assignee');
const leaveEl=document.getElementById('leave');
const statusEl=document.getElementById('status');
const fromDateEl=document.getElementById('from-date');
const toDateEl=document.getElementById('to-date');
const windowLabelEl=document.getElementById('window-label');
const statTotalTakenHoursEl=document.getElementById('stat-total-taken-hours');
const statTotalTakenDaysEl=document.getElementById('stat-total-taken-days');
const statPlannedTakenHoursEl=document.getElementById('stat-planned-taken-hours');
const statUnplannedTakenHoursEl=document.getElementById('stat-unplanned-taken-hours');
const statTotalPlannedLeavesHoursEl=document.getElementById('stat-total-planned-leaves-hours');
const statFutureHoursEl=document.getElementById('stat-future-hours');
const statNoEntryEl=document.getElementById('stat-no-entry');
const tabButtons = Array.from(document.querySelectorAll('.tab-btn'));
const subtasks = Array.isArray(payload.subtasks) ? payload.subtasks : [];
const dailyData = Array.isArray(payload.daily) ? payload.daily : [];
const weeklyData = Array.isArray(payload.weekly) ? payload.weekly : [];
const monthlyData = Array.isArray(payload.monthly) ? payload.monthly : [];
const defectiveData = Array.isArray(payload.defective) ? payload.defective : [];
const clubbedData = Array.isArray(payload.clubbed) ? payload.clubbed : [];
const dayProfile = payload.day_profile || {{}};
const subtaskByKey = new Map(subtasks.map((x) => [String(x.issue_key || '').toUpperCase(), x]));

function toDateObj(iso) {{
  if(!iso) return null;
  const d = new Date(`${{iso}}T00:00:00`);
  return Number.isNaN(d.getTime()) ? null : d;
}}

function hoursText(value) {{
  return `${{Number(value || 0).toFixed(2)}}h`;
}}

function daysText(value) {{
  return `${{Number(value || 0).toFixed(2)}}d`;
}}

function isoWeekToRange(code) {{
  const m = /^([0-9]{{4}})-W([0-9]{{2}})$/.exec(String(code || ''));
  if(!m) return {{ start: null, end: null }};
  const year = Number(m[1]);
  const week = Number(m[2]);
  const jan4 = new Date(year, 0, 4);
  const jan4Dow = (jan4.getDay() + 6) % 7;
  const mondayWeek1 = new Date(year, 0, 4 - jan4Dow);
  const monday = new Date(mondayWeek1);
  monday.setDate(mondayWeek1.getDate() + ((week - 1) * 7));
  const sunday = new Date(monday);
  sunday.setDate(monday.getDate() + 6);
  return {{ start: monday, end: sunday }};
}}

function monthToRange(code) {{
  const m = /^([0-9]{{4}})-([0-9]{{2}})$/.exec(String(code || ''));
  if(!m) return {{ start: null, end: null }};
  const year = Number(m[1]);
  const month = Number(m[2]) - 1;
  const start = new Date(year, month, 1);
  const end = new Date(year, month + 1, 0);
  return {{ start, end }};
}}

function inRangeDay(iso, fromDate, toDate) {{
  const d = toDateObj(iso);
  if(!d) return false;
  return d >= fromDate && d <= toDate;
}}

function dayHoursForDate(isoDay) {{
  const d = toDateObj(isoDay);
  const standard = Number(dayProfile.standard_hours_per_day || 8);
  const ramadan = Number(dayProfile.ramadan_hours_per_day || 6.5);
  if(!d) return standard;
  const rStart = toDateObj(dayProfile.ramadan_start_date || '');
  const rEnd = toDateObj(dayProfile.ramadan_end_date || '');
  if(rStart && rEnd && d >= rStart && d <= rEnd) return ramadan;
  return standard;
}}

function overlapsRange(start, end, fromDate, toDate) {{
  if(!start || !end) return false;
  return start <= toDate && end >= fromDate;
}}

function inRangeWeek(code, fromDate, toDate) {{
  const r = isoWeekToRange(code);
  return overlapsRange(r.start, r.end, fromDate, toDate);
}}

function inRangeMonth(code, fromDate, toDate) {{
  const r = monthToRange(code);
  return overlapsRange(r.start, r.end, fromDate, toDate);
}}

function renderTable(tableId, headers, rows) {{
  const table = document.getElementById(tableId);
  if(!table) return;
  const thead = table.querySelector('thead');
  const tbody = table.querySelector('tbody');
  if(thead) {{
    thead.innerHTML = `<tr>${{headers.map((h) => `<th>${{h}}</th>`).join('')}}</tr>`;
  }}
  if(!tbody) return;
  if(!rows.length) {{
    tbody.innerHTML = `<tr><td colspan="${{headers.length}}">No rows</td></tr>`;
    return;
  }}
  tbody.innerHTML = rows.map((row) => {{
    const cells = headers.map((h) => {{
      const value = row[h] ?? '';
      const cls = String(value) === 'No Entry' ? ' class="warn-badge"' : '';
      return `<td${{cls}}>${{String(value)}}</td>`;
    }}).join('');
    return `<tr>${{cells}}</tr>`;
  }}).join('');
}}

function init(){{
 const names=[...new Set((payload.subtasks||[]).map(x=>x.assignee||'Unassigned'))].sort((a,b)=>a.localeCompare(b));
 for(const n of names){{const o=document.createElement('option');o.value=n;o.textContent=n;assigneeEl.appendChild(o);}}
 const todayIso = new Date().toISOString().slice(0, 10);
 fromDateEl.value = '2026-01-01';
 toDateEl.value = todayIso;
}}
function switchTab(tabName){{
  for(const b of tabButtons){{
    b.classList.toggle('active', b.dataset.tab === tabName);
  }}
  for(const pane of document.querySelectorAll('.tab-pane')){{
    pane.classList.toggle('active', pane.id === `tab-${{tabName}}`);
  }}
}}
function apply(){{
 const fromDate = toDateObj(fromDateEl.value || '2026-01-01');
 const toDate = toDateObj(toDateEl.value || new Date().toISOString().slice(0,10));
 if(!fromDate || !toDate) return;
 const fromUsed = fromDate <= toDate ? fromDate : toDate;
 const toUsed = fromDate <= toDate ? toDate : fromDate;
 const a=(assigneeEl.value||'').toLowerCase();
 const l=(leaveEl.value||'').toLowerCase();
 const s=(statusEl.value||'').toLowerCase();

 const dailyFilteredRaw = dailyData.filter((row) => {{
   const day = String(row.period_day || '');
   if(!inRangeDay(day, fromUsed, toUsed)) return false;
   if(a && String(row.assignee || '').toLowerCase() !== a) return false;
   return true;
 }});

 const assigneeMap = new Map();
 for(const row of dailyFilteredRaw) {{
   const name = String(row.assignee || 'Unassigned');
   const current = assigneeMap.get(name) || {{ planned: 0, unplanned: 0, future: 0 }};
   current.planned += Number(row.planned_taken_hours || 0);
   current.unplanned += Number(row.unplanned_taken_hours || 0);
   current.future += Number(row.planned_not_taken_hours || 0);
   assigneeMap.set(name, current);
 }}
 const assigneeRows = Array.from(assigneeMap.entries()).sort((x,y)=>x[0].localeCompare(y[0])).map(([name, v]) => {{
   const planned = Number(v.planned.toFixed(2));
   const unplanned = Number(v.unplanned.toFixed(2));
   const future = Number(v.future.toFixed(2));
   return {{
     "Assignee": name,
     "Planned Taken (h)": planned.toFixed(2),
     "Unplanned Taken (h)": unplanned.toFixed(2),
     "Planned Not Yet Taken (h)": future.toFixed(2),
     "No Entry Count": "0",
     "Unknown Count": "0",
   }};
 }});

 const dailyRows = dailyFilteredRaw
   .sort((x,y)=>String(x.period_day||'').localeCompare(String(y.period_day||'')) || String(x.assignee||'').localeCompare(String(y.assignee||'')))
   .map((r)=>({{
     "Assignee": String(r.assignee || ''),
     "Day": String(r.period_day || ''),
     "Planned Taken (h)": Number(r.planned_taken_hours || 0).toFixed(2),
     "Unplanned Taken (h)": Number(r.unplanned_taken_hours || 0).toFixed(2),
     "Future Planned (h)": Number(r.planned_not_taken_hours || 0).toFixed(2),
     "Total (h)": Number(r.total_hours || 0).toFixed(2),
   }}));

 const weeklyRows = weeklyData
   .filter((r)=>inRangeWeek(String(r.period_week || ''), fromUsed, toUsed) && (!a || String(r.assignee || '').toLowerCase() === a))
   .sort((x,y)=>String(x.period_week||'').localeCompare(String(y.period_week||'')) || String(x.assignee||'').localeCompare(String(y.assignee||'')))
   .map((r)=>({{
     "Assignee": String(r.assignee || ''),
     "Week": String(r.period_week || ''),
     "Planned Taken (h)": Number(r.planned_taken_hours || 0).toFixed(2),
     "Unplanned Taken (h)": Number(r.unplanned_taken_hours || 0).toFixed(2),
     "Future Planned (h)": Number(r.planned_not_taken_hours || 0).toFixed(2),
     "Total (h)": Number(r.total_hours || 0).toFixed(2),
   }}));

 const monthlyRows = monthlyData
   .filter((r)=>inRangeMonth(String(r.period_month || ''), fromUsed, toUsed) && (!a || String(r.assignee || '').toLowerCase() === a))
   .sort((x,y)=>String(x.period_month||'').localeCompare(String(y.period_month||'')) || String(x.assignee||'').localeCompare(String(y.assignee||'')))
   .map((r)=>({{
     "Assignee": String(r.assignee || ''),
     "Month": String(r.period_month || ''),
     "Planned Taken (h)": Number(r.planned_taken_hours || 0).toFixed(2),
     "Unplanned Taken (h)": Number(r.unplanned_taken_hours || 0).toFixed(2),
     "Future Planned (h)": Number(r.planned_not_taken_hours || 0).toFixed(2),
     "Total (h)": Number(r.total_hours || 0).toFixed(2),
   }}));

 const inRangeSubtask = (subtask) => {{
   const plannedDate = String(subtask.planned_date_for_bucket || '').slice(0,10);
   const startDate = String(subtask.start_date || '').slice(0,10);
   const dueDate = String(subtask.due_date || '').slice(0,10);
   const plannedObj = toDateObj(plannedDate);
   const startObj = toDateObj(startDate);
   const dueObj = toDateObj(dueDate);
   if(plannedObj) return plannedObj >= fromUsed && plannedObj <= toUsed;
   if(startObj || dueObj) {{
     const rowStart = startObj || dueObj;
     const rowEnd = dueObj || startObj;
     return Boolean(rowStart && rowEnd && rowStart <= toUsed && rowEnd >= fromUsed);
   }}
   return true;
 }};

 const defectiveRows = defectiveData.filter((row) => {{
   const subtask = subtaskByKey.get(String(row.issue_key || '').toUpperCase()) || {{}};
   if(a && String(subtask.assignee || '').toLowerCase() !== a) return false;
   if(l && String(subtask.leave_classification || '').toLowerCase() !== l) return false;
   if(s && !String(subtask.status || '').toLowerCase().includes(s)) return false;
   if(!inRangeSubtask(subtask)) return false;
   return true;
 }});

 const clubbedRows = clubbedData.filter((row) => {{
   const subtask = subtaskByKey.get(String(row.issue_key || '').toUpperCase()) || {{}};
   if(a && String(subtask.assignee || '').toLowerCase() !== a) return false;
   if(l && String(subtask.leave_classification || '').toLowerCase() !== l) return false;
   if(s && !String(subtask.status || '').toLowerCase().includes(s)) return false;
   if(!inRangeSubtask(subtask)) return false;
   return true;
 }});

 const plannedTaken = dailyFilteredRaw.reduce((acc, row) => acc + Number(row.planned_taken_hours || 0), 0);
 const unplannedTaken = dailyFilteredRaw.reduce((acc, row) => acc + Number(row.unplanned_taken_hours || 0), 0);
 const futurePlanned = dailyFilteredRaw.reduce((acc, row) => acc + Number(row.planned_not_taken_hours || 0), 0);
 const totalPlannedLeaves = plannedTaken + futurePlanned;
 const totalTaken = plannedTaken + unplannedTaken;
 const takenByDay = new Map();
 for(const row of dailyFilteredRaw) {{
   const iso = String(row.period_day || '');
   const hours = Number(row.planned_taken_hours || 0) + Number(row.unplanned_taken_hours || 0);
   takenByDay.set(iso, Number(takenByDay.get(iso) || 0) + hours);
 }}
 let totalTakenDays = 0;
 for(const [iso, hours] of takenByDay.entries()) {{
   const hpd = Number(dayHoursForDate(iso) || 0);
   if(hpd > 0) totalTakenDays += hours / hpd;
 }}

 windowLabelEl.textContent = `${{fromDateEl.value || '2026-01-01'}} to ${{toDateEl.value || new Date().toISOString().slice(0,10)}}`;
 statTotalTakenHoursEl.textContent = hoursText(totalTaken);
 statTotalTakenDaysEl.textContent = daysText(totalTakenDays);
 statPlannedTakenHoursEl.textContent = hoursText(plannedTaken);
 statUnplannedTakenHoursEl.textContent = hoursText(unplannedTaken);
 statTotalPlannedLeavesHoursEl.textContent = hoursText(totalPlannedLeaves);
 statFutureHoursEl.textContent = hoursText(futurePlanned);
 statNoEntryEl.textContent = String(defectiveRows.filter((r)=>String(r.reason || '').includes('No Entry')).length);

 renderTable('assignee-table', ["Assignee","Planned Taken (h)","Unplanned Taken (h)","Planned Not Yet Taken (h)","No Entry Count","Unknown Count"], assigneeRows);
 renderTable('daily-table', ["Assignee","Day","Planned Taken (h)","Unplanned Taken (h)","Future Planned (h)","Total (h)"], dailyRows);
 renderTable('weekly-table', ["Assignee","Week","Planned Taken (h)","Unplanned Taken (h)","Future Planned (h)","Total (h)"], weeklyRows);
 renderTable('monthly-table', ["Assignee","Month","Planned Taken (h)","Unplanned Taken (h)","Future Planned (h)","Total (h)"], monthlyRows);
 renderTable('defective-table', ["issue_key","assignee","summary","status","leave_classification","reason","planned_dates","original_estimate_hours"], defectiveRows);
 renderTable('clubbed-table', ["issue_key","assignee","summary","leave_classification","status","logged_hours","estimate_hours","start_date","due_date"], clubbedRows);
}}
init();
for(const btn of tabButtons){{
  btn.addEventListener('click', () => switchTab(btn.dataset.tab));
}}
assigneeEl.addEventListener('change',apply);
leaveEl.addEventListener('change',apply);
statusEl.addEventListener('input',apply);
fromDateEl.addEventListener('change',apply);
toDateEl.addEventListener('change',apply);
apply();
</script>
<script src="shared-nav.js"></script>
</body>
</html>"""
    output_path.write_text(doc, encoding="utf-8")


def _normalize_subtasks(issues: list[dict], task_assignee_by_key: dict[str, str], worklogs_by_issue: dict[str, list[WorklogRow]], start_date_field_id: str) -> list[SubtaskRow]:
    out: list[SubtaskRow] = []
    for issue in issues:
        key = _to_text(issue.get("key"))
        fields = issue.get("fields", {}) or {}
        summary = _to_text(fields.get("summary"))
        status = _to_text((fields.get("status") or {}).get("name"))
        assignee = _to_text((fields.get("assignee") or {}).get("displayName")) or "Unassigned"
        parent_key = _to_text((fields.get("parent") or {}).get("key"))
        parent_assignee = task_assignee_by_key.get(parent_key, "Unassigned")
        start_date = _to_text(fields.get(start_date_field_id))
        due_date = _to_text(fields.get("duedate"))
        start_date, due_date = normalize_subtask_dates(start_date, due_date, summary)
        estimate_h = _seconds_to_hours(fields.get("timeoriginalestimate"))
        leave_raw = leave_type_text(fields.get(LEAVE_TYPE_FIELD))
        classification = classify_leave(leave_raw, status, summary)
        issue_worklogs = worklogs_by_issue.get(key, [])
        logged_h = round(sum(w.hours_logged for w in issue_worklogs), 2)
        out.append(
            SubtaskRow(
                issue_key=key,
                issue_id=_to_text(issue.get("id")),
                summary=summary,
                status=status,
                assignee=assignee,
                parent_task_key=parent_key,
                parent_task_assignee=parent_assignee,
                created=_to_text(fields.get("created")),
                updated=_to_text(fields.get("updated")),
                start_date=start_date,
                due_date=due_date,
                original_estimate_hours=estimate_h,
                timespent_hours=_seconds_to_hours(fields.get("timespent")),
                leave_type_raw=leave_raw,
                leave_classification=classification,
                total_worklog_hours=logged_h,
                planned_date_for_bucket=choose_planned_date(start_date, due_date),
                clubbed_leave="Yes" if is_clubbed_leave(logged_h, estimate_h, start_date, due_date) else "No",
                no_entry_flag="Yes" if is_defective_no_entry(classification, logged_h, estimate_h, start_date, due_date) else "No",
            )
        )
    return sorted(out, key=lambda s: s.issue_key)


def run_report(
    project_key: str,
    project_name: str,
    from_date: str,
    to_date: str,
    xlsx_out: Path,
    html_out: Path,
    md_out: Path,
    start_date_field_id: str = DEFAULT_START_DATE_FIELD,
    enable_incremental: bool = False,
) -> dict[str, int]:
    run_started = datetime.now(timezone.utc)
    run_id = uuid.uuid4().hex

    session = get_session()
    incremental_disabled = _is_incremental_disabled(enable_incremental)
    overlap_minutes = _get_overlap_minutes()
    force_full_days = _get_force_full_days()
    bootstrap_days = _get_bootstrap_days()
    db_path = get_db_path()
    db_path.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(db_path)
    init_db(conn)

    default_checkpoint = bootstrap_default_checkpoint(bootstrap_days)
    checkpoint = get_or_init_checkpoint(conn, RLT_PIPELINE_NAME, default_checkpoint)
    now_utc = utc_now_iso()
    force_full_sync = incremental_disabled or needs_full_sync(conn, RLT_PIPELINE_NAME, now_utc, force_full_days)
    from_updated = None if force_full_sync else apply_overlap(checkpoint, overlap_minutes)

    discovery_fields = ["project", "issuetype", "updated"]
    discovery_jql = f'project = {project_key} AND issuetype in ("Task", "Sub-task", "Subtask")'
    if from_updated:
        from_updated_jql = parse_iso_utc(from_updated).strftime("%Y-%m-%d %H:%M")
        discovery_jql += f' AND updated >= "{from_updated_jql}"'
    discovery_jql += " ORDER BY updated ASC"

    discovered = _fetch_issues(session, discovery_jql, discovery_fields)
    candidates = _candidate_rows_from_issues(discovered)
    changed_issue_keys, new_issue_count, changed_existing_count = _classify_candidates(conn, candidates)
    upsert_issue_index(conn, candidates)

    active_ids = {row["issue_id"] for row in candidates}
    if force_full_sync:
        mark_missing_issues_deleted(conn, [project_key], ["Task", "Sub-task", "Subtask"], active_ids)

    detail_fetch_keys = sorted({row["issue_key"] for row in candidates} if force_full_sync else set(changed_issue_keys))
    detail_fields = [
        "project",
        "issuetype",
        "summary",
        "status",
        "assignee",
        "parent",
        "created",
        "updated",
        "duedate",
        "timeoriginalestimate",
        "timespent",
        LEAVE_TYPE_FIELD,
        start_date_field_id,
    ]
    if detail_fetch_keys:
        detailed = _fetch_issues_by_keys(session, detail_fetch_keys, detail_fields)
        payload_rows = []
        index_rows = []
        now_seen = utc_now_iso()
        for issue in detailed:
            issue_id = _to_text(issue.get("id"))
            issue_key = _to_text(issue.get("key"))
            fields = issue.get("fields", {}) or {}
            if not issue_id or not issue_key:
                continue
            payload_rows.append(
                {
                    "issue_id": issue_id,
                    "issue_key": issue_key,
                    "updated_utc": _to_text(fields.get("updated")),
                    "payload": issue,
                }
            )
            index_rows.append(
                {
                    "issue_id": issue_id,
                    "issue_key": issue_key,
                    "updated_utc": _to_text(fields.get("updated")),
                    "issue_type": _to_text((fields.get("issuetype") or {}).get("name")),
                    "project_key": _to_text((fields.get("project") or {}).get("key")),
                    "last_seen_utc": now_seen,
                    "is_deleted": 0,
                }
            )
        upsert_issue_payloads(conn, payload_rows)
        upsert_issue_index(conn, index_rows)

    cached = get_cached_issue_payloads(conn, project_keys=[project_key], issue_types=["Task", "Sub-task", "Subtask"])
    tasks = [i for i in cached if _to_text((i.get("fields", {}).get("issuetype") or {}).get("name")).lower() == "task"]
    subtasks_raw = [i for i in cached if "sub" in _to_text((i.get("fields", {}).get("issuetype") or {}).get("name")).lower()]

    task_assignees = {
        _to_text(t.get("key")): (_to_text(((t.get("fields") or {}).get("assignee") or {}).get("displayName")) or "Unassigned")
        for t in tasks
    }

    subtask_keys = sorted({_to_text(i.get("key")) for i in subtasks_raw if _to_text(i.get("key"))})
    cached_worklogs = get_cached_worklogs_for_subtasks(conn, subtask_keys)
    changed_set = set(detail_fetch_keys)
    worklog_fetch_keys = set(subtask_keys if force_full_sync else [k for k in subtask_keys if k in changed_set])
    worklog_fetch_keys.update(k for k in subtask_keys if k not in cached_worklogs)

    worklogs_by_issue: dict[str, list[WorklogRow]] = {}
    all_worklogs: list[WorklogRow] = []
    worklog_fetch_count = 0
    for item in subtasks_raw:
        key = _to_text(item.get("key"))
        if not key:
            continue
        issue_id = _to_text(item.get("id"))
        if key in worklog_fetch_keys:
            raw_logs = _fetch_worklogs_for_issue(session, key)
            max_updated = ""
            for wl in raw_logs:
                updated = _to_text(wl.get("updated"))
                if updated and (not max_updated or parse_iso_utc(updated) > parse_iso_utc(max_updated)):
                    max_updated = updated
            upsert_worklog_payload(conn, issue_key=key, issue_id=issue_id, worklogs=raw_logs, worklog_updated_utc=max_updated or None)
            cached_worklogs[key] = raw_logs
            worklog_fetch_count += 1
        rows = _normalize_worklogs(key, cached_worklogs.get(key, []))
        worklogs_by_issue[key] = rows
        all_worklogs.extend(rows)

    max_updated_seen = ""
    for row in candidates:
        updated_utc = _to_text(row.get("updated_utc"))
        if not updated_utc:
            continue
        if not max_updated_seen or parse_iso_utc(updated_utc) > parse_iso_utc(max_updated_seen):
            max_updated_seen = updated_utc
    if max_updated_seen:
        set_checkpoint(conn, RLT_PIPELINE_NAME, max_updated_seen)
    if force_full_sync:
        mark_full_sync(conn, RLT_PIPELINE_NAME, utc_now_iso())

    subtasks = _normalize_subtasks(subtasks_raw, task_assignees, worklogs_by_issue, start_date_field_id)
    day_hours_profile = _day_hours_profile_from_env()
    distributed_subtasks = _redistribute_continuous_leave_subtasks(subtasks, day_hours_profile)
    redistributed_worklogs = _redistribute_continuous_leave_worklogs(subtasks, all_worklogs, day_hours_profile)
    aggr = _compute_aggregates(subtasks, redistributed_worklogs, from_date, to_date, day_hours_profile)

    _write_xlsx(xlsx_out, subtasks, distributed_subtasks, redistributed_worklogs, aggr)
    _write_html(html_out, project_key, project_name, from_date, to_date, subtasks, aggr)
    _write_md(md_out, project_key, project_name, from_date, to_date, aggr)

    run_ended = datetime.now(timezone.utc)
    record_pipeline_artifact(
        conn,
        {
            "run_id": run_id,
            "pipeline": RLT_PIPELINE_NAME,
            "started_at_utc": run_started.replace(microsecond=0).isoformat().replace("+00:00", "Z"),
            "ended_at_utc": run_ended.replace(microsecond=0).isoformat().replace("+00:00", "Z"),
            "issues_scanned": len(candidates),
            "issues_changed": changed_existing_count,
            "new_issues": new_issue_count,
            "detail_fetches": len(detail_fetch_keys),
            "worklog_fetches": worklog_fetch_count,
            "duration_ms": int((run_ended - run_started).total_seconds() * 1000),
        },
    )
    conn.close()

    return {
        "tasks": len(tasks),
        "subtasks": len(subtasks),
        "worklogs": len(redistributed_worklogs),
        "assignees": len(aggr["assignee_summary"]),
        "defective": len(aggr["defective"]),
        "clubbed": len(aggr["clubbed"]),
        "issues_scanned": len(candidates),
        "issues_changed": changed_existing_count,
        "new_issues": new_issue_count,
        "detail_fetches": len(detail_fetch_keys),
        "worklog_fetches": worklog_fetch_count,
    }


def _parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Generate RLT leave intelligence report artifacts.")
    p.add_argument("--project-key", default=DEFAULT_PROJECT_KEY)
    p.add_argument("--project-name", default=DEFAULT_PROJECT_NAME)
    p.add_argument("--from-date", default="")
    p.add_argument("--to-date", default="")
    p.add_argument("--window", default=DEFAULT_WINDOW)
    p.add_argument("--xlsx-out", default=DEFAULT_XLSX_OUT)
    p.add_argument("--html-out", default=DEFAULT_HTML_OUT)
    p.add_argument("--md-out", default=DEFAULT_MD_OUT)
    p.add_argument("--start-date-field-id", default=DEFAULT_START_DATE_FIELD)
    p.add_argument(
        "--incremental",
        action="store_true",
        help="Enable smart incremental fetch (default: full fetch).",
    )
    return p.parse_args()


def main() -> None:
    args = _parse_args()
    from_date, to_date = resolve_window_range(args.window, _to_text(args.from_date), _to_text(args.to_date))

    base = Path(__file__).resolve().parent
    xlsx_out = Path(args.xlsx_out)
    html_out = Path(args.html_out)
    md_out = Path(args.md_out)
    if not xlsx_out.is_absolute():
        xlsx_out = base / xlsx_out
    if not html_out.is_absolute():
        html_out = base / html_out
    if not md_out.is_absolute():
        md_out = base / md_out

    print(f"Generating leave report for {args.project_key} ({args.project_name})")
    print(f"Window: {from_date} -> {to_date}")
    stats = run_report(
        project_key=args.project_key,
        project_name=args.project_name,
        from_date=from_date,
        to_date=to_date,
        xlsx_out=xlsx_out,
        html_out=html_out,
        md_out=md_out,
        start_date_field_id=args.start_date_field_id,
        enable_incremental=args.incremental,
    )
    print(
        "Report generated: "
        f"tasks={stats['tasks']}, subtasks={stats['subtasks']}, worklogs={stats['worklogs']}, "
        f"assignees={stats['assignees']}, defective={stats['defective']}, clubbed={stats['clubbed']}"
    )
    print(f"Excel: {xlsx_out}")
    print(f"HTML:  {html_out}")
    print(f"MD:    {md_out}")


if __name__ == "__main__":
    main()

