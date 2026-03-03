from __future__ import annotations

from collections import defaultdict
from datetime import date, datetime, timezone
import io
import json
import os
import random
import re
import shutil
import sqlite3
import subprocess
import sys
import threading
import time
import uuid
from pathlib import Path

from flask import Flask, jsonify, redirect, request, send_file
from openpyxl import load_workbook
from generate_assignee_hours_report import (
    DEFAULT_CAPACITY_DB,
    DEFAULT_LEAVE_REPORT_INPUT_XLSX,
    DEFAULT_SUMMARY_OUTPUT_XLSX,
    _delete_capacity_settings,
    _hours_to_days_over_range,
    _init_capacity_db,
    _list_capacity_profiles,
    _load_capacity_settings,
    _load_leave_metrics,
    _read_summary_xlsx,
    _save_capacity_settings,
    _to_text,
    calculate_capacity_metrics,
)
from generate_employee_performance_report import (
    DEFAULT_PERFORMANCE_SETTINGS,
    _delete_performance_team,
    _init_performance_settings_db,
    _list_performance_teams,
    _load_performance_settings,
    _normalize_performance_settings,
    _save_performance_settings,
    _save_performance_team,
)
from manage_fields_registry import (
    create_manage_field,
    init_manage_fields_db,
    load_manage_fields,
    restore_manage_field,
    soft_delete_manage_field,
    update_manage_field,
)
from managed_projects_registry import (
    create_managed_project,
    deterministic_color_for_project_key,
    init_managed_projects_db,
    list_managed_projects,
    normalize_project_key,
    parse_project_keys_from_env,
    restore_managed_project,
    seed_managed_projects,
    soft_delete_managed_project,
    update_managed_project,
)
from report_entity_registry import (
    REPORT_ENTITY_GLOBAL_SETTING_KEYS,
    init_report_entities_db,
    load_report_entities,
    load_report_entity_global_settings,
    reset_report_entities_to_defaults,
    save_report_entities,
    save_report_entity_global_settings,
)
from ipp_meeting_utils import resolve_jira_end_date_field_ids, resolve_jira_start_date_field_id
from jira_client import BASE_URL, extract_jira_key_from_url, get_session
from planned_actual_table_view_service import (
    DEFAULT_RETENTION_DAYS as PACTV_DEFAULT_RETENTION_DAYS,
    VALID_MODES as PACTV_VALID_MODES,
    build_snapshot_payload as pactv_build_snapshot_payload,
    begin_queued_run as pactv_begin_queued_run,
    create_run as pactv_create_run,
    diff_snapshots as pactv_diff_snapshots,
    export_csv_bytes as pactv_export_csv_bytes,
    export_xlsx_bytes as pactv_export_xlsx_bytes,
    finish_run_failed as pactv_finish_run_failed,
    finish_run_success as pactv_finish_run_success,
    get_run as pactv_get_run,
    has_running_run_for_scope as pactv_has_running_run_for_scope,
    init_db as pactv_init_db,
    is_cancel_requested as pactv_is_cancel_requested,
    list_history as pactv_list_history,
    list_queue as pactv_list_queue,
    mark_run_canceled as pactv_mark_run_canceled,
    load_snapshot_by_filter as pactv_load_snapshot_by_filter,
    load_snapshot_by_id as pactv_load_snapshot_by_id,
    load_sync_state as pactv_load_sync_state,
    load_ui_settings as pactv_load_ui_settings,
    make_filter as pactv_make_filter,
    prune_old_data as pactv_prune_old_data,
    pin_official_snapshot as pactv_pin_official_snapshot,
    request_cancel as pactv_request_cancel,
    save_snapshot as pactv_save_snapshot,
    save_snapshot_event as pactv_save_snapshot_event,
    save_source_audit as pactv_save_source_audit,
    save_sync_state as pactv_save_sync_state,
    save_ui_settings as pactv_save_ui_settings,
    unpin_official_snapshot as pactv_unpin_official_snapshot,
    update_run_progress as pactv_update_run_progress,
)


REPORT_FILENAME_TO_ID: dict[str, str] = {
    "dashboard.html": "dashboard",
    "nested_view_report.html": "nested_view",
    "missed_entries.html": "missed_entries",
    "assignee_hours_report.html": "assignee_hours",
    "rnd_data_story.html": "rnd_data_story",
    "planned_rmis_report.html": "planned_rmis",
    "gantt_chart_report.html": "gantt_chart",
    "phase_rmi_gantt_report.html": "phase_rmi_gantt",
    "ipp_meeting_dashboard.html": "ipp_meeting_dashboard",
    "rlt_leave_report.html": "rlt_leave_report",
    "leaves_planned_calendar.html": "leaves_planned_calendar",
    "employee_performance_report.html": "employee_performance",
    "planned_vs_dispensed_report.html": "planned_vs_dispensed",
    "planned_actual_table_view.html": "planned_actual_table_view",
}

REPORT_REFRESH_CHAINS: dict[str, list[str]] = {
    "dashboard": [
        "run_all_exports.py",
        "fetch_jira_dashboard.py",
    ],
    "nested_view": [
        "run_all_exports.py",
        "generate_nested_view_html.py",
    ],
    "missed_entries": [
        "run_all_exports.py",
        "generate_missed_entries_html.py",
    ],
    "assignee_hours": [
        "run_all_exports.py",
        "generate_rlt_leave_report.py",
        "generate_assignee_hours_report.py",
    ],
    "rnd_data_story": [
        "run_all_exports.py",
        "generate_rlt_leave_report.py",
        "generate_assignee_hours_report.py",
        "generate_rnd_data_story.py",
    ],
    "planned_rmis": [
        "run_all_exports.py",
        "generate_planned_rmis_html.py",
    ],
    "gantt_chart": [
        "run_all_exports.py",
        "generate_gantt_chart_html.py",
    ],
    "phase_rmi_gantt": [
        "run_all_exports.py",
        "generate_phase_rmi_gantt_html.py",
    ],
    "ipp_meeting_dashboard": [
        "export_ipp_phase_breakdown.py",
        "generate_ipp_meeting_dashboard.py",
    ],
    "rlt_leave_report": [
        "generate_rlt_leave_report.py",
    ],
    "leaves_planned_calendar": [
        "generate_rlt_leave_report.py",
        "generate_leaves_planned_calendar_html.py",
    ],
    "employee_performance": [
        "run_all_exports.py",
        "generate_rlt_leave_report.py",
        "generate_employee_performance_report.py",
    ],
    # This report is API-driven (atomic Jira fetches), so no batch refresh scripts.
    "planned_vs_dispensed": [],
    # This report is API-driven and DB-backed.
    "planned_actual_table_view": [],
}

REFRESH_WIDGET_MARKER = "codex-refresh-widget-v2"
REFRESH_WIDGET_START = "<!-- codex-refresh-widget-start -->"
REFRESH_WIDGET_END = "<!-- codex-refresh-widget-end -->"
INFO_DRAWER_MARKER = "codex-info-drawer-v1"
INFO_DRAWER_START = "<!-- codex-info-drawer-start -->"
INFO_DRAWER_END = "<!-- codex-info-drawer-end -->"
CAPACITY_SETTINGS_ROUTE = "/settings/capacity"
PERFORMANCE_SETTINGS_ROUTE = "/settings/performance"
REPORT_ENTITIES_SETTINGS_ROUTE = "/settings/report-entities"
MANAGE_FIELDS_SETTINGS_ROUTE = "/settings/manage-fields"
PROJECTS_SETTINGS_ROUTE = "/settings/projects"
EPICS_MANAGEMENT_SETTINGS_ROUTE = "/settings/epics-management"
EPICS_DROPDOWN_OPTIONS_SETTINGS_ROUTE = "/settings/epics-dropdown-options"
EPIC_PHASES_SETTINGS_ROUTE = "/settings/epic-phases"
DASHBOARD_RISK_SETTINGS_ROUTE = "/settings/dashboard-risk"
PAGE_CATEGORIES_SETTINGS_ROUTE = "/settings/page-categories"

STATIC_REPORT_NAV_ITEMS: list[dict[str, object]] = [
    {"page_key": "dashboard", "title": "Dashboard", "href": "/dashboard.html", "icon": "space_dashboard", "file": "dashboard.html", "default_nav_order": 10, "page_type": "report"},
    {"page_key": "nested_view_report", "title": "Nested View Report", "href": "/nested_view_report.html", "icon": "account_tree", "file": "nested_view_report.html", "default_nav_order": 20, "page_type": "report"},
    {"page_key": "missed_entries_report", "title": "Missed Entries Report", "href": "/missed_entries.html", "icon": "event_busy", "file": "missed_entries.html", "default_nav_order": 30, "page_type": "report"},
    {"page_key": "assignee_hours_report", "title": "Assignee Hours Report", "href": "/assignee_hours_report.html", "icon": "schedule", "file": "assignee_hours_report.html", "default_nav_order": 40, "page_type": "report"},
    {"page_key": "employee_performance_report", "title": "Employee Performance", "href": "/employee_performance_report.html", "icon": "monitoring", "file": "employee_performance_report.html", "default_nav_order": 50, "page_type": "report"},
    {"page_key": "rlt_leave_report", "title": "RLT Leave Report", "href": "/rlt_leave_report.html", "icon": "beach_access", "file": "rlt_leave_report.html", "default_nav_order": 60, "page_type": "report"},
    {"page_key": "leaves_planned_calendar", "title": "Leaves Planned Calendar", "href": "/leaves_planned_calendar.html", "icon": "calendar_month", "file": "leaves_planned_calendar.html", "default_nav_order": 70, "page_type": "report"},
    {"page_key": "rnd_data_story", "title": "RnD Data Story", "href": "/rnd_data_story.html", "icon": "auto_stories", "file": "rnd_data_story.html", "default_nav_order": 80, "page_type": "report"},
    {"page_key": "planned_rmis_report", "title": "Planned RMIs", "href": "/planned_rmis_report.html", "icon": "assignment_turned_in", "file": "planned_rmis_report.html", "default_nav_order": 90, "page_type": "report"},
    {"page_key": "phase_rmi_gantt_report", "title": "Phase RMI Gantt", "href": "/phase_rmi_gantt_report.html", "icon": "view_timeline", "file": "phase_rmi_gantt_report.html", "default_nav_order": 100, "page_type": "report"},
    {"page_key": "planned_vs_dispensed_report", "title": "Planned vs Dispensed", "href": "/planned_vs_dispensed_report.html", "icon": "analytics", "file": "planned_vs_dispensed_report.html", "default_nav_order": 110, "page_type": "report"},
    {"page_key": "planned_actual_table_view", "title": "Planned vs Actual Table View", "href": "/planned_actual_table_view.html", "icon": "table_view", "file": "planned_actual_table_view.html", "default_nav_order": 120, "page_type": "report"},
    {"page_key": "ipp_meeting_dashboard", "title": "IPP Meeting Dashboard", "href": "/ipp_meeting_dashboard.html", "icon": "groups", "file": "ipp_meeting_dashboard.html", "default_nav_order": 130, "page_type": "report"},
]

STATIC_ADMIN_NAV_ITEMS: list[dict[str, object]] = [
    {"page_key": "capacity_settings", "title": "Capacity Settings", "href": CAPACITY_SETTINGS_ROUTE, "icon": "tune", "path": CAPACITY_SETTINGS_ROUTE, "default_nav_order": 10, "page_type": "configuration"},
    {"page_key": "performance_settings", "title": "Performance Settings", "href": PERFORMANCE_SETTINGS_ROUTE, "icon": "speed", "path": PERFORMANCE_SETTINGS_ROUTE, "default_nav_order": 20, "page_type": "configuration"},
    {"page_key": "report_entities", "title": "Report Entities", "href": REPORT_ENTITIES_SETTINGS_ROUTE, "icon": "dataset", "path": REPORT_ENTITIES_SETTINGS_ROUTE, "default_nav_order": 30, "page_type": "configuration"},
    {"page_key": "manage_fields", "title": "Manage Fields", "href": MANAGE_FIELDS_SETTINGS_ROUTE, "icon": "list_alt", "path": MANAGE_FIELDS_SETTINGS_ROUTE, "default_nav_order": 40, "page_type": "configuration"},
    {"page_key": "projects", "title": "Projects", "href": PROJECTS_SETTINGS_ROUTE, "icon": "work", "path": PROJECTS_SETTINGS_ROUTE, "default_nav_order": 50, "page_type": "configuration"},
    {"page_key": "epic_dropdowns", "title": "Epic Dropdowns", "href": EPICS_DROPDOWN_OPTIONS_SETTINGS_ROUTE, "icon": "arrow_drop_down_circle", "path": EPICS_DROPDOWN_OPTIONS_SETTINGS_ROUTE, "default_nav_order": 60, "page_type": "configuration"},
    {"page_key": "epic_phases", "title": "Epic Phases", "href": EPIC_PHASES_SETTINGS_ROUTE, "icon": "alt_route", "path": EPIC_PHASES_SETTINGS_ROUTE, "default_nav_order": 70, "page_type": "configuration"},
    {"page_key": "epics_planner", "title": "Epics Planner", "href": EPICS_MANAGEMENT_SETTINGS_ROUTE, "icon": "event_note", "path": EPICS_MANAGEMENT_SETTINGS_ROUTE, "default_nav_order": 80, "page_type": "configuration"},
    {"page_key": "page_categories", "title": "Page Categories", "href": PAGE_CATEGORIES_SETTINGS_ROUTE, "icon": "category", "path": PAGE_CATEGORIES_SETTINGS_ROUTE, "default_nav_order": 90, "page_type": "configuration"},
]


def _settings_nav_items() -> list[tuple[str, str]]:
    return [
        ("Capacity Settings", CAPACITY_SETTINGS_ROUTE),
        ("Performance Settings", PERFORMANCE_SETTINGS_ROUTE),
        ("Dashboard Risk", DASHBOARD_RISK_SETTINGS_ROUTE),
        ("Report Entities", REPORT_ENTITIES_SETTINGS_ROUTE),
        ("Manage Fields", MANAGE_FIELDS_SETTINGS_ROUTE),
        ("Projects", PROJECTS_SETTINGS_ROUTE),
        ("Page Categories", PAGE_CATEGORIES_SETTINGS_ROUTE),
        ("Epic Dropdowns", EPICS_DROPDOWN_OPTIONS_SETTINGS_ROUTE),
        ("Epic Phases", EPIC_PHASES_SETTINGS_ROUTE),
        ("Epics Planner", EPICS_MANAGEMENT_SETTINGS_ROUTE),
    ]


DEFAULT_DASHBOARD_RISK_SETTINGS: dict[str, object] = {
    "indicator_points": {
        "subtask_linear_lag": 3,
        "due_crossed_unresolved": 3,
        "subtask_late_actual_start": 1,
        "start_passed_not_in_progress": 1,
        "inherited_child_risk": 3,
    },
    "thresholds": {
        "can_be_min": 1,
        "medium_min": 2,
        "high_min": 4,
        "at_risk_min": 2,
    },
    "labels": {
        "low": "Low",
        "can_be": "Can Be",
        "medium": "Medium Risk",
        "high": "Highly At Risk",
    },
}


def _default_dashboard_risk_settings() -> dict[str, object]:
    return json.loads(json.dumps(DEFAULT_DASHBOARD_RISK_SETTINGS))


def _coerce_int(value: object, field_name: str, minimum: int = 0, maximum: int = 100) -> int:
    try:
        numeric = int(value)
    except (TypeError, ValueError):
        raise ValueError(f"{field_name} must be an integer.")
    if numeric < minimum or numeric > maximum:
        raise ValueError(f"{field_name} must be between {minimum} and {maximum}.")
    return numeric


def _normalize_dashboard_risk_settings(payload: object) -> dict[str, object]:
    defaults = _default_dashboard_risk_settings()
    raw = payload if isinstance(payload, dict) else {}

    raw_indicator_points = raw.get("indicator_points")
    if not isinstance(raw_indicator_points, dict):
        raw_indicator_points = {}
    indicator_points: dict[str, int] = {}
    for key, default_value in (defaults["indicator_points"] or {}).items():
        indicator_points[key] = _coerce_int(
            raw_indicator_points.get(key, default_value),
            f"indicator_points.{key}",
            minimum=0,
            maximum=30,
        )

    raw_thresholds = raw.get("thresholds")
    if not isinstance(raw_thresholds, dict):
        raw_thresholds = {}
    thresholds = {
        "can_be_min": _coerce_int(raw_thresholds.get("can_be_min", (defaults["thresholds"] or {}).get("can_be_min", 1)), "thresholds.can_be_min", 0, 100),
        "medium_min": _coerce_int(raw_thresholds.get("medium_min", (defaults["thresholds"] or {}).get("medium_min", 2)), "thresholds.medium_min", 0, 100),
        "high_min": _coerce_int(raw_thresholds.get("high_min", (defaults["thresholds"] or {}).get("high_min", 4)), "thresholds.high_min", 0, 100),
        "at_risk_min": _coerce_int(raw_thresholds.get("at_risk_min", (defaults["thresholds"] or {}).get("at_risk_min", 2)), "thresholds.at_risk_min", 0, 100),
    }
    if thresholds["can_be_min"] > thresholds["medium_min"]:
        raise ValueError("thresholds.can_be_min cannot be greater than thresholds.medium_min.")
    if thresholds["medium_min"] > thresholds["high_min"]:
        raise ValueError("thresholds.medium_min cannot be greater than thresholds.high_min.")
    if thresholds["at_risk_min"] < thresholds["medium_min"] or thresholds["at_risk_min"] > thresholds["high_min"]:
        raise ValueError("thresholds.at_risk_min must be between medium_min and high_min.")

    raw_labels = raw.get("labels")
    if not isinstance(raw_labels, dict):
        raw_labels = {}
    labels: dict[str, str] = {}
    for key, default_value in (defaults["labels"] or {}).items():
        label = _to_text(raw_labels.get(key, default_value))
        if not label:
            raise ValueError(f"labels.{key} cannot be empty.")
        if len(label) > 32:
            raise ValueError(f"labels.{key} must be 32 characters or fewer.")
        labels[key] = label

    return {
        "indicator_points": indicator_points,
        "thresholds": thresholds,
        "labels": labels,
    }


def _init_dashboard_risk_settings_db(settings_db_path: Path) -> None:
    settings_db_path.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(settings_db_path)
    conn.row_factory = sqlite3.Row
    try:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS dashboard_risk_settings (
                id INTEGER PRIMARY KEY CHECK(id = 1),
                settings_json TEXT NOT NULL,
                updated_at_utc TEXT NOT NULL DEFAULT ''
            )
            """
        )
        existing = conn.execute("SELECT settings_json FROM dashboard_risk_settings WHERE id = 1").fetchone()
        if not existing:
            default_json = json.dumps(_default_dashboard_risk_settings(), ensure_ascii=True)
            now_utc = datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")
            conn.execute(
                "INSERT INTO dashboard_risk_settings(id, settings_json, updated_at_utc) VALUES(1, ?, ?)",
                (default_json, now_utc),
            )
        conn.commit()
    finally:
        conn.close()


def _load_dashboard_risk_settings(settings_db_path: Path) -> dict[str, object]:
    _init_dashboard_risk_settings_db(settings_db_path)
    conn = sqlite3.connect(settings_db_path)
    try:
        row = conn.execute("SELECT settings_json FROM dashboard_risk_settings WHERE id = 1").fetchone()
        raw_json = _to_text(row[0]) if row else ""
        parsed: object = {}
        if raw_json:
            try:
                parsed = json.loads(raw_json)
            except Exception:
                parsed = {}
        return _normalize_dashboard_risk_settings(parsed)
    finally:
        conn.close()


def _save_dashboard_risk_settings(settings_db_path: Path, payload: object) -> dict[str, object]:
    existing = _load_dashboard_risk_settings(settings_db_path)
    raw = payload if isinstance(payload, dict) else {}
    merged = {
        "indicator_points": {
            **(existing.get("indicator_points") if isinstance(existing.get("indicator_points"), dict) else {}),
            **(raw.get("indicator_points") if isinstance(raw.get("indicator_points"), dict) else {}),
        },
        "thresholds": {
            **(existing.get("thresholds") if isinstance(existing.get("thresholds"), dict) else {}),
            **(raw.get("thresholds") if isinstance(raw.get("thresholds"), dict) else {}),
        },
        "labels": {
            **(existing.get("labels") if isinstance(existing.get("labels"), dict) else {}),
            **(raw.get("labels") if isinstance(raw.get("labels"), dict) else {}),
        },
    }
    normalized = _normalize_dashboard_risk_settings(merged)

    conn = sqlite3.connect(settings_db_path)
    try:
        now_utc = datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")
        conn.execute(
            """
            INSERT INTO dashboard_risk_settings(id, settings_json, updated_at_utc)
            VALUES(1, ?, ?)
            ON CONFLICT(id) DO UPDATE SET settings_json=excluded.settings_json, updated_at_utc=excluded.updated_at_utc
            """,
            (json.dumps(normalized, ensure_ascii=True), now_utc),
        )
        conn.commit()
    finally:
        conn.close()
    return normalized


def _settings_top_nav_html(active_route: str) -> str:
    _ = active_route
    # Admin page-to-page links are intentionally removed from headers.
    # Navigation is handled by the shared left navigation.
    return ""


def _page_catalog() -> list[dict[str, object]]:
    items = STATIC_REPORT_NAV_ITEMS + STATIC_ADMIN_NAV_ITEMS
    out: list[dict[str, object]] = []
    for item in items:
        out.append(
            {
                "page_key": _to_text(item.get("page_key")),
                "title": _to_text(item.get("title")),
                "route_or_file": _to_text(item.get("path") or item.get("file") or item.get("href")),
                "href": _to_text(item.get("href")),
                "icon": _to_text(item.get("icon")),
                "page_type": _to_text(item.get("page_type")),
                "default_nav_order": int(item.get("default_nav_order") or 0),
            }
        )
    return out


def _normalize_page_category_name(value: object) -> str:
    name = _to_text(value)
    if not name:
        raise ValueError("Category name is required.")
    return name


def _normalize_google_icon_name(value: object) -> str:
    icon_name = _to_text(value).lower()
    if not icon_name:
        return "folder"
    if not re.match(r"^[a-z0-9_]+$", icon_name):
        raise ValueError("icon_name must contain only lowercase letters, numbers, and underscores.")
    return icon_name


def _normalize_page_category_payload(payload: object, require_name: bool = True) -> dict[str, object]:
    raw = payload if isinstance(payload, dict) else {}
    name = _to_text(raw.get("name"))
    if require_name:
        name = _normalize_page_category_name(name)
    display_order_raw = raw.get("display_order", 100)
    try:
        display_order = int(display_order_raw)
    except (TypeError, ValueError):
        raise ValueError("display_order must be an integer.")
    return {
        "id": int(raw.get("id") or 0),
        "name": name,
        "icon_name": _normalize_google_icon_name(raw.get("icon_name")),
        "display_in_navigation": 1 if bool(raw.get("display_in_navigation", True)) else 0,
        "display_order": display_order,
        "is_active": 1 if bool(raw.get("is_active", True)) else 0,
    }


def _normalize_page_assignment_payload(payload: object, valid_page_keys: set[str], valid_category_ids: set[int]) -> dict[str, object]:
    raw = payload if isinstance(payload, dict) else {}
    page_key = _to_text(raw.get("page_key"))
    if not page_key:
        raise ValueError("page_key is required for assignments.")
    if page_key not in valid_page_keys:
        raise ValueError(f"Unknown page_key: {page_key}")
    page_type = _to_text(raw.get("page_type")).lower()
    if page_type not in {"report", "configuration"}:
        raise ValueError("page_type must be report or configuration.")
    try:
        category_id = int(raw.get("category_id"))
    except (TypeError, ValueError):
        raise ValueError("category_id must be an integer.")
    if category_id not in valid_category_ids:
        raise ValueError(f"Unknown category_id: {category_id}")
    return {
        "page_key": page_key,
        "page_type": page_type,
        "category_id": category_id,
    }


def _init_page_categories_db(db_path: Path) -> None:
    db_path.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(db_path)
    try:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS page_categories (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL COLLATE NOCASE UNIQUE,
                icon_name TEXT NOT NULL DEFAULT 'folder',
                display_in_navigation INTEGER NOT NULL DEFAULT 1,
                display_order INTEGER NOT NULL DEFAULT 100,
                is_active INTEGER NOT NULL DEFAULT 1,
                created_at_utc TEXT NOT NULL,
                updated_at_utc TEXT NOT NULL
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS page_category_assignments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                page_key TEXT NOT NULL,
                page_type TEXT NOT NULL,
                category_id INTEGER NOT NULL,
                created_at_utc TEXT NOT NULL,
                updated_at_utc TEXT NOT NULL,
                UNIQUE(page_key, category_id),
                FOREIGN KEY(category_id) REFERENCES page_categories(id) ON DELETE CASCADE
            )
            """
        )
        columns = conn.execute("PRAGMA table_info(page_categories)").fetchall()
        column_names = {str(row[1]) for row in columns}
        if "icon_name" not in column_names:
            conn.execute("ALTER TABLE page_categories ADD COLUMN icon_name TEXT NOT NULL DEFAULT 'folder'")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_page_category_assignments_page_key ON page_category_assignments(page_key)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_page_category_assignments_category ON page_category_assignments(category_id)")
        conn.commit()
    finally:
        conn.close()


def _row_to_page_category(row: sqlite3.Row) -> dict[str, object]:
    return {
        "id": int(row["id"]),
        "name": _to_text(row["name"]),
        "icon_name": _normalize_google_icon_name(row["icon_name"]),
        "display_in_navigation": bool(int(row["display_in_navigation"] or 0)),
        "display_order": int(row["display_order"] or 0),
        "is_active": bool(int(row["is_active"] or 0)),
        "created_at_utc": _to_text(row["created_at_utc"]),
        "updated_at_utc": _to_text(row["updated_at_utc"]),
    }


def _load_page_categories(db_path: Path) -> dict[str, object]:
    _init_page_categories_db(db_path)
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    try:
        category_rows = conn.execute(
            """
            SELECT id, name, icon_name, display_in_navigation, display_order, is_active, created_at_utc, updated_at_utc
            FROM page_categories
            WHERE is_active = 1
            ORDER BY display_order ASC, lower(name) ASC, id ASC
            """
        ).fetchall()
        assignment_rows = conn.execute(
            """
            SELECT id, page_key, page_type, category_id, created_at_utc, updated_at_utc
            FROM page_category_assignments
            ORDER BY page_type ASC, page_key ASC, category_id ASC
            """
        ).fetchall()
    finally:
        conn.close()
    categories = [_row_to_page_category(row) for row in category_rows]
    category_ids = {int(item["id"]) for item in categories}
    assignments: list[dict[str, object]] = []
    for row in assignment_rows:
        category_id = int(row["category_id"] or 0)
        if category_id not in category_ids:
            continue
        assignments.append(
            {
                "id": int(row["id"]),
                "page_key": _to_text(row["page_key"]),
                "page_type": _to_text(row["page_type"]),
                "category_id": category_id,
                "created_at_utc": _to_text(row["created_at_utc"]),
                "updated_at_utc": _to_text(row["updated_at_utc"]),
            }
        )
    return {
        "categories": categories,
        "assignments": assignments,
        "page_catalog": _page_catalog(),
        "has_categories": len(categories) > 0,
    }


def _create_page_category(db_path: Path, payload: object) -> dict[str, object]:
    _init_page_categories_db(db_path)
    normalized = _normalize_page_category_payload(payload, require_name=True)
    now = _utc_now_iso()
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    try:
        cur = conn.execute(
            """
            INSERT INTO page_categories(name, icon_name, display_in_navigation, display_order, is_active, created_at_utc, updated_at_utc)
            VALUES(?, ?, ?, ?, ?, ?, ?)
            """,
            (
                normalized["name"],
                normalized["icon_name"],
                normalized["display_in_navigation"],
                normalized["display_order"],
                normalized["is_active"],
                now,
                now,
            ),
        )
        conn.commit()
        category_id = int(cur.lastrowid or 0)
        row = conn.execute(
            """
            SELECT id, name, icon_name, display_in_navigation, display_order, is_active, created_at_utc, updated_at_utc
            FROM page_categories
            WHERE id = ?
            """,
            (category_id,),
        ).fetchone()
    except sqlite3.IntegrityError:
        raise ValueError(f"Category '{normalized['name']}' already exists.")
    finally:
        conn.close()
    if not row:
        raise ValueError("Failed to load created category.")
    return _row_to_page_category(row)


def _update_page_category(db_path: Path, category_id: int, payload: object) -> dict[str, object]:
    _init_page_categories_db(db_path)
    normalized = _normalize_page_category_payload(payload, require_name=True)
    cid = int(category_id)
    now = _utc_now_iso()
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    try:
        existing = conn.execute("SELECT id FROM page_categories WHERE id = ?", (cid,)).fetchone()
        if not existing:
            raise LookupError(f"Category '{cid}' not found.")
        conn.execute(
            """
            UPDATE page_categories
            SET name = ?, icon_name = ?, display_in_navigation = ?, display_order = ?, is_active = ?, updated_at_utc = ?
            WHERE id = ?
            """,
            (
                normalized["name"],
                normalized["icon_name"],
                normalized["display_in_navigation"],
                normalized["display_order"],
                normalized["is_active"],
                now,
                cid,
            ),
        )
        conn.commit()
        row = conn.execute(
            """
            SELECT id, name, icon_name, display_in_navigation, display_order, is_active, created_at_utc, updated_at_utc
            FROM page_categories
            WHERE id = ?
            """,
            (cid,),
        ).fetchone()
    except sqlite3.IntegrityError:
        raise ValueError(f"Category '{normalized['name']}' already exists.")
    finally:
        conn.close()
    if not row:
        raise LookupError(f"Category '{cid}' not found.")
    return _row_to_page_category(row)


def _delete_page_category(db_path: Path, category_id: int) -> int:
    _init_page_categories_db(db_path)
    cid = int(category_id)
    conn = sqlite3.connect(db_path)
    try:
        conn.execute("DELETE FROM page_category_assignments WHERE category_id = ?", (cid,))
        cur = conn.execute("DELETE FROM page_categories WHERE id = ?", (cid,))
        conn.commit()
    finally:
        conn.close()
    if int(cur.rowcount or 0) <= 0:
        raise LookupError(f"Category '{cid}' not found.")
    return cid


def _save_page_categories_payload(db_path: Path, payload: object) -> dict[str, object]:
    _init_page_categories_db(db_path)
    raw = payload if isinstance(payload, dict) else {}
    incoming_categories_raw = raw.get("categories")
    incoming_assignments_raw = raw.get("assignments")
    incoming_categories = incoming_categories_raw if isinstance(incoming_categories_raw, list) else []
    incoming_assignments = incoming_assignments_raw if isinstance(incoming_assignments_raw, list) else []

    now = _utc_now_iso()
    valid_page_keys = {item["page_key"] for item in _page_catalog()}
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    try:
        existing_rows = conn.execute(
            "SELECT id, name, icon_name, display_in_navigation, display_order, is_active FROM page_categories ORDER BY id ASC"
        ).fetchall()
        existing_by_id = {int(row["id"]): row for row in existing_rows}

        keep_ids: set[int] = set()
        for item in incoming_categories:
            normalized = _normalize_page_category_payload(item, require_name=True)
            requested_id = int(normalized.get("id") or 0)
            if requested_id > 0 and requested_id in existing_by_id:
                conn.execute(
                    """
                    UPDATE page_categories
                    SET name = ?, icon_name = ?, display_in_navigation = ?, display_order = ?, is_active = ?, updated_at_utc = ?
                    WHERE id = ?
                    """,
                    (
                        normalized["name"],
                        normalized["icon_name"],
                        normalized["display_in_navigation"],
                        normalized["display_order"],
                        normalized["is_active"],
                        now,
                        requested_id,
                    ),
                )
                keep_ids.add(requested_id)
            else:
                cur = conn.execute(
                    """
                    INSERT INTO page_categories(name, icon_name, display_in_navigation, display_order, is_active, created_at_utc, updated_at_utc)
                    VALUES(?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        normalized["name"],
                        normalized["icon_name"],
                        normalized["display_in_navigation"],
                        normalized["display_order"],
                        normalized["is_active"],
                        now,
                        now,
                    ),
                )
                keep_ids.add(int(cur.lastrowid or 0))

        if incoming_categories:
            existing_ids = set(existing_by_id.keys())
            delete_ids = existing_ids - keep_ids
            for cid in delete_ids:
                conn.execute("DELETE FROM page_category_assignments WHERE category_id = ?", (cid,))
                conn.execute("DELETE FROM page_categories WHERE id = ?", (cid,))

        active_category_rows = conn.execute(
            "SELECT id FROM page_categories WHERE is_active = 1"
        ).fetchall()
        valid_category_ids = {int(row["id"]) for row in active_category_rows}

        conn.execute("DELETE FROM page_category_assignments")
        seen_pairs: set[tuple[str, int]] = set()
        for item in incoming_assignments:
            normalized_assignment = _normalize_page_assignment_payload(item, valid_page_keys, valid_category_ids)
            pair = (str(normalized_assignment["page_key"]), int(normalized_assignment["category_id"]))
            if pair in seen_pairs:
                continue
            seen_pairs.add(pair)
            conn.execute(
                """
                INSERT INTO page_category_assignments(page_key, page_type, category_id, created_at_utc, updated_at_utc)
                VALUES(?, ?, ?, ?, ?)
                """,
                (
                    normalized_assignment["page_key"],
                    normalized_assignment["page_type"],
                    normalized_assignment["category_id"],
                    now,
                    now,
                ),
            )
        conn.commit()
    except sqlite3.IntegrityError as exc:
        raise ValueError(f"Failed to save page categories payload: {exc}")
    finally:
        conn.close()
    return _load_page_categories(db_path)


def _sort_nav_items(items: list[dict[str, object]]) -> list[dict[str, object]]:
    return sorted(items, key=lambda item: (int(item.get("default_nav_order") or 0), _to_text(item.get("title")).casefold()))


def _build_navigation_from_page_categories(db_path: Path) -> dict[str, object]:
    state = _load_page_categories(db_path)
    categories = [item for item in state["categories"] if bool(item.get("is_active"))]
    if not categories:
        return {"enabled": False}

    visible_categories = [
        item for item in categories
        if bool(item.get("display_in_navigation"))
    ]
    visible_categories = sorted(
        visible_categories,
        key=lambda item: (int(item.get("display_order") or 0), _to_text(item.get("name")).casefold(), int(item.get("id") or 0)),
    )
    visible_ids = {int(item.get("id") or 0) for item in visible_categories}

    report_by_key = {str(item["page_key"]): dict(item) for item in STATIC_REPORT_NAV_ITEMS}
    admin_by_key = {str(item["page_key"]): dict(item) for item in STATIC_ADMIN_NAV_ITEMS}
    assignment_map: dict[str, set[int]] = defaultdict(set)
    for row in state["assignments"]:
        category_id = int(row.get("category_id") or 0)
        if category_id not in visible_ids:
            continue
        page_key = _to_text(row.get("page_key"))
        if page_key:
            assignment_map[page_key].add(category_id)

    report_categories: list[dict[str, object]] = []
    report_assigned_keys: set[str] = set()
    for category in visible_categories:
        cid = int(category.get("id") or 0)
        items = [item for key, item in report_by_key.items() if cid in assignment_map.get(key, set())]
        if not items:
            continue
        report_assigned_keys.update({str(item["page_key"]) for item in items})
        report_categories.append(
            {
                "id": cid,
                "name": _to_text(category.get("name")),
                "icon_name": _normalize_google_icon_name(category.get("icon_name")),
                "items": _sort_nav_items(items),
            }
        )

    admin_categories: list[dict[str, object]] = []
    admin_assigned_keys: set[str] = set()
    for category in visible_categories:
        cid = int(category.get("id") or 0)
        items = [item for key, item in admin_by_key.items() if cid in assignment_map.get(key, set())]
        if not items:
            continue
        admin_assigned_keys.update({str(item["page_key"]) for item in items})
        admin_categories.append(
            {
                "id": cid,
                "name": _to_text(category.get("name")),
                "icon_name": _normalize_google_icon_name(category.get("icon_name")),
                "items": _sort_nav_items(items),
            }
        )

    uncategorized_reports = _sort_nav_items([item for key, item in report_by_key.items() if key not in report_assigned_keys])
    uncategorized_admin = _sort_nav_items([item for key, item in admin_by_key.items() if key not in admin_assigned_keys])
    return {
        "enabled": True,
        "reports": {
            "items": uncategorized_reports,
            "categories": report_categories,
        },
        "admin_settings": {
            "items": uncategorized_admin,
            "categories": admin_categories,
        },
    }


def _resolve_capacity_runtime_paths(base_dir: Path) -> dict[str, Path]:
    db_name = os.getenv("JIRA_ASSIGNEE_HOURS_CAPACITY_DB_PATH", DEFAULT_CAPACITY_DB).strip() or DEFAULT_CAPACITY_DB
    leave_name = os.getenv("JIRA_LEAVE_REPORT_XLSX_PATH", DEFAULT_LEAVE_REPORT_INPUT_XLSX).strip() or DEFAULT_LEAVE_REPORT_INPUT_XLSX
    summary_name = (
        os.getenv("JIRA_ASSIGNEE_HOURS_SUMMARY_XLSX_PATH", DEFAULT_SUMMARY_OUTPUT_XLSX).strip()
        or DEFAULT_SUMMARY_OUTPUT_XLSX
    )

    db_path = Path(db_name)
    if not db_path.is_absolute():
        db_path = base_dir / db_path

    leave_report_path = Path(leave_name)
    if not leave_report_path.is_absolute():
        leave_report_path = base_dir / leave_report_path

    summary_path = Path(summary_name)
    if not summary_path.is_absolute():
        summary_path = base_dir / summary_path

    return {
        "db_path": db_path,
        "leave_report_path": leave_report_path,
        "summary_path": summary_path,
    }


def _list_assignees_from_summary(summary_path: Path) -> list[str]:
    try:
        rows = _read_summary_xlsx(summary_path)
    except Exception:
        return []
    names = {
        _to_text(item.get("issue_assignee")) or "Unassigned"
        for item in rows
    }
    return sorted(names, key=lambda s: s.casefold())


def _jira_search_projects(query: str, limit: int = 25) -> list[dict[str, str]]:
    text = _to_text(query)
    max_results = max(1, min(int(limit or 25), 100))
    session = get_session()
    params = {
        "query": text,
        "maxResults": max_results,
        "orderBy": "key",
    }
    response = session.get(f"{BASE_URL}/rest/api/3/project/search", params=params, timeout=(10, 30))
    response.raise_for_status()
    payload = response.json()
    values = payload.get("values", []) if isinstance(payload, dict) else []
    out: list[dict[str, str]] = []
    for item in values:
        key = _to_text((item or {}).get("key")).upper()
        name = _to_text((item or {}).get("name"))
        if not key or not name:
            continue
        out.append({"project_key": key, "project_name": name})
    return out


def _jira_lookup_project_name(project_key: str) -> str:
    key = normalize_project_key(project_key)
    session = get_session()
    response = session.get(f"{BASE_URL}/rest/api/3/project/{key}", timeout=(10, 30))
    response.raise_for_status()
    payload = response.json()
    return _to_text(payload.get("name")) or key


def _capacity_settings_html() -> str:
    return """<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Capacity Profile Settings</title>
  <link rel="stylesheet" href="/shared-nav.css">
  <link rel="stylesheet" href="/material-symbols.css">
  <style>
    :root {
      color-scheme: light;
      --bg: #f3f6fb;
      --card: #ffffff;
      --text: #0f172a;
      --muted: #475569;
      --line: #cbd5e1;
      --brand: #1d4ed8;
      --ok: #065f46;
      --err: #991b1b;
    }
    body {
      margin: 0;
      padding: 20px;
      font-family: "Segoe UI", Tahoma, sans-serif;
      background: var(--bg);
      color: var(--text);
    }
    .card {
      max-width: 1000px;
      margin: 0 auto;
      background: var(--card);
      border: 1px solid var(--line);
      border-radius: 12px;
      padding: 18px;
    }
    .topbar {
      display: flex;
      gap: 8px;
      flex-wrap: wrap;
      align-items: center;
      justify-content: space-between;
      margin-bottom: 10px;
    }
    .title {
      margin: 0;
      font-size: 1.25rem;
    }
    .muted {
      color: var(--muted);
      font-size: 0.92rem;
      margin: 6px 0 0;
    }
    .links {
      display: flex;
      gap: 8px;
      flex-wrap: wrap;
    }
    .btn {
      border: 1px solid #1e40af;
      background: var(--brand);
      color: #fff;
      border-radius: 8px;
      padding: 8px 12px;
      cursor: pointer;
      text-decoration: none;
      font-size: 0.9rem;
      display: inline-block;
    }
    .btn.alt {
      border-color: var(--line);
      background: #fff;
      color: #0f172a;
    }
    .btn.warn {
      border-color: #7f1d1d;
      background: #b91c1c;
      color: #fff;
    }
    .grid {
      display: grid;
      gap: 10px;
      grid-template-columns: repeat(auto-fit, minmax(220px, 1fr));
      margin-top: 14px;
    }
    label {
      display: block;
      font-size: 0.82rem;
      font-weight: 700;
      margin-bottom: 4px;
    }
    input, select {
      width: 100%;
      box-sizing: border-box;
      border: 1px solid var(--line);
      border-radius: 8px;
      padding: 8px;
      font-size: 0.92rem;
    }
    .row {
      display: flex;
      gap: 8px;
      flex-wrap: wrap;
      margin-top: 14px;
    }
    .muted-note {
      color: var(--muted);
      font-size: 0.8rem;
      margin-top: 5px;
    }
    .holiday-tools {
      display: flex;
      gap: 8px;
      margin-bottom: 8px;
      flex-wrap: wrap;
    }
    .holiday-list {
      display: flex;
      gap: 6px;
      flex-wrap: wrap;
      min-height: 34px;
      border: 1px solid var(--line);
      border-radius: 8px;
      padding: 6px 8px;
      align-items: center;
      background: #fff;
    }
    .holiday-list .empty {
      color: #64748b;
      font-size: 0.85rem;
    }
    .holiday-chip {
      display: inline-flex;
      align-items: center;
      gap: 6px;
      background: #e0f2fe;
      border: 1px solid #93c5fd;
      color: #1e3a8a;
      border-radius: 999px;
      padding: 3px 8px;
      font-size: 0.8rem;
    }
    .holiday-chip button {
      border: none;
      background: transparent;
      color: #1e3a8a;
      cursor: pointer;
      padding: 0;
      font-size: 0.78rem;
      line-height: 1;
      font-weight: 700;
    }
    #status {
      margin-top: 10px;
      font-size: 0.9rem;
      min-height: 1.2em;
    }
    #status.ok { color: var(--ok); }
    #status.err { color: var(--err); }
    @media (max-width: 760px) {
      body { padding: 12px; }
      .card { padding: 14px; }
    }
  </style>
</head>
<body>
  <main class="card">
    <!-- template-version: 2026-02-21-holiday-picker-v1 -->
    <div class="topbar">
      <div>
        <h1 class="title">Capacity Profile Settings</h1>
        <p class="muted">Manage capacity profiles here. These profiles can be applied in Assignee Hours and Nested View reports.</p>
        <p class="muted" style="font-size:0.78rem;margin-top:4px;">Template: 2026-02-21-holiday-picker-v1</p>
      </div>
      <div class="links">__SETTINGS_TOP_NAV__</div>
    </div>

    <div class="grid">
      <div>
        <label for="profile-select">Saved Profiles</label>
        <select id="profile-select"></select>
      </div>
      <div>
        <label for="from-date">From Date</label>
        <input id="from-date" type="date">
      </div>
      <div>
        <label for="to-date">To Date</label>
        <input id="to-date" type="date">
      </div>
      <div>
        <label for="employees">Employees</label>
        <input id="employees" type="number" min="0" step="1" value="0">
        <div class="muted-note">Assignees found in data: <strong id="assignee-count-help">0</strong> <button class="btn alt" type="button" id="use-assignee-count" style="margin-left:6px;padding:3px 8px;font-size:.75rem;">Use</button></div>
      </div>
      <div>
        <label for="std-hours">Standard Hours/Day</label>
        <input id="std-hours" type="number" min="0.1" step="0.1" value="8">
      </div>
      <div>
        <label for="ramadan-hours">Ramadan Hours/Day</label>
        <input id="ramadan-hours" type="number" min="0.1" step="0.1" value="6.5">
      </div>
      <div>
        <label for="ramadan-start">Ramadan Start</label>
        <input id="ramadan-start" type="date">
      </div>
      <div>
        <label for="ramadan-end">Ramadan End</label>
        <input id="ramadan-end" type="date">
      </div>
      <div style="grid-column: 1 / -1;">
        <label for="holiday-date-picker">Holiday Dates</label>
        <div class="holiday-tools">
          <input id="holiday-date-picker" type="date">
          <button class="btn alt" type="button" id="holiday-add">Add</button>
          <button class="btn alt" type="button" id="holiday-clear">Clear</button>
        </div>
        <div id="holiday-list" class="holiday-list"></div>
      </div>
    </div>

    <div class="row">
      <button class="btn alt" type="button" id="refresh-btn">Refresh</button>
      <button class="btn alt" type="button" id="new-btn">New</button>
      <button class="btn" type="button" id="save-btn">Save</button>
      <button class="btn warn" type="button" id="delete-btn">Delete</button>
    </div>
    <div id="status"></div>
  </main>
  <script>
    const API_CAPACITY = "/api/capacity";
    const API_PROFILES = "/api/capacity/profiles";
    const API_ASSIGNEE_COUNT = "/api/capacity/assignee-count";

    const profileSelect = document.getElementById("profile-select");
    const fromDate = document.getElementById("from-date");
    const toDate = document.getElementById("to-date");
    const employees = document.getElementById("employees");
    const assigneeCountHelp = document.getElementById("assignee-count-help");
    const useAssigneeCountBtn = document.getElementById("use-assignee-count");
    const stdHours = document.getElementById("std-hours");
    const ramadanHours = document.getElementById("ramadan-hours");
    const ramadanStart = document.getElementById("ramadan-start");
    const ramadanEnd = document.getElementById("ramadan-end");
    const holidayDatePicker = document.getElementById("holiday-date-picker");
    const holidayAddBtn = document.getElementById("holiday-add");
    const holidayClearBtn = document.getElementById("holiday-clear");
    const holidayList = document.getElementById("holiday-list");
    const refreshBtn = document.getElementById("refresh-btn");
    const newBtn = document.getElementById("new-btn");
    const saveBtn = document.getElementById("save-btn");
    const deleteBtn = document.getElementById("delete-btn");
    const statusEl = document.getElementById("status");

    let profiles = [];
    let selectedHolidayDates = [];
    let dataAssigneeCount = 0;

    function setStatus(message, variant) {
      statusEl.textContent = String(message || "");
      statusEl.className = variant === "err" ? "err" : (variant === "ok" ? "ok" : "");
    }

    function profileKey(profile) {
      return String(profile && profile.from_date || "") + "|" + String(profile && profile.to_date || "");
    }

    function normalizeHolidayDates(values) {
      const seen = new Set();
      const out = [];
      for (const raw of Array.isArray(values) ? values : []) {
        const v = String(raw || "").trim();
        if (!v) continue;
        if (!/^\\d{4}-\\d{2}-\\d{2}$/.test(v)) {
          throw new Error("Invalid holiday date: " + v);
        }
        if (!seen.has(v)) {
          seen.add(v);
          out.push(v);
        }
      }
      out.sort();
      return out;
    }

    function renderHolidayList() {
      if (!holidayList) return;
      if (!selectedHolidayDates.length) {
        holidayList.innerHTML = '<span class="empty">No holiday dates selected.</span>';
        return;
      }
      holidayList.innerHTML = selectedHolidayDates
        .map((iso) => '<span class="holiday-chip">' + iso + ' <button type="button" data-remove-holiday="' + iso + '">x</button></span>')
        .join("");
      Array.from(holidayList.querySelectorAll("button[data-remove-holiday]")).forEach((btn) => {
        btn.addEventListener("click", () => {
          const value = String(btn.getAttribute("data-remove-holiday") || "").trim();
          if (!value) return;
          selectedHolidayDates = selectedHolidayDates.filter((item) => item !== value);
          renderHolidayList();
        });
      });
    }

    function addHolidayFromPicker() {
      const value = String((holidayDatePicker && holidayDatePicker.value) || "").trim();
      if (!value) return;
      selectedHolidayDates = normalizeHolidayDates([...selectedHolidayDates, value]);
      if (holidayDatePicker) holidayDatePicker.value = "";
      renderHolidayList();
    }

    function setForm(profile) {
      const p = profile || {};
      fromDate.value = String(p.from_date || "");
      toDate.value = String(p.to_date || "");
      employees.value = String(Number(p.employee_count || 0));
      stdHours.value = String(Number(p.standard_hours_per_day || 8));
      ramadanHours.value = String(Number(p.ramadan_hours_per_day || 6.5));
      ramadanStart.value = String(p.ramadan_start_date || "");
      ramadanEnd.value = String(p.ramadan_end_date || "");
      selectedHolidayDates = normalizeHolidayDates(Array.isArray(p.holiday_dates) ? p.holiday_dates : []);
      renderHolidayList();
    }

    function findProfileByKey(key) {
      const wanted = String(key || "");
      if (!wanted) return null;
      for (const p of profiles) {
        if (profileKey(p) === wanted) return p;
      }
      return null;
    }

    function renderProfiles() {
      const current = String(profileSelect.value || "");
      if (!profiles.length) {
        profileSelect.innerHTML = '<option value="">No saved profiles found</option>';
        profileSelect.disabled = true;
        deleteBtn.disabled = true;
        return;
      }
      const options = ['<option value="">Select a saved profile</option>'];
      for (const p of profiles) {
        const key = profileKey(p);
        const label = String(p.from_date || "-") + " to " + String(p.to_date || "-");
        options.push('<option value="' + key + '">' + label + '</option>');
      }
      profileSelect.innerHTML = options.join("");
      profileSelect.disabled = false;
      deleteBtn.disabled = false;
      if (current && findProfileByKey(current)) {
        profileSelect.value = current;
      }
    }

    function buildPayload() {
      const payload = {
        from_date: String(fromDate.value || "").trim(),
        to_date: String(toDate.value || "").trim(),
        employee_count: Math.round(Number(employees.value || 0)),
        standard_hours_per_day: Number(stdHours.value || 0),
        ramadan_start_date: String(ramadanStart.value || "").trim(),
        ramadan_end_date: String(ramadanEnd.value || "").trim(),
        ramadan_hours_per_day: Number(ramadanHours.value || 0),
        holiday_dates: normalizeHolidayDates(selectedHolidayDates),
      };
      if (!payload.from_date || !payload.to_date) throw new Error("From/To date are required.");
      if (payload.to_date < payload.from_date) throw new Error("To date must be on or after From date.");
      if (payload.employee_count < 0) throw new Error("Employees must be >= 0.");
      if (payload.standard_hours_per_day <= 0 || payload.ramadan_hours_per_day <= 0) {
        throw new Error("Hours/day must be > 0.");
      }
      if (!!payload.ramadan_start_date !== !!payload.ramadan_end_date) {
        throw new Error("Set both Ramadan start and end, or leave both empty.");
      }
      if (payload.ramadan_start_date && payload.ramadan_end_date && payload.ramadan_end_date < payload.ramadan_start_date) {
        throw new Error("Ramadan end must be on or after Ramadan start.");
      }
      return payload;
    }

    async function refreshProfiles() {
      const response = await fetch(API_PROFILES);
      if (!response.ok) throw new Error("Failed to load profiles.");
      const payload = await response.json();
      profiles = Array.isArray(payload && payload.profiles) ? payload.profiles : [];
      renderProfiles();
      setStatus("Profiles refreshed.", "ok");
    }

    async function refreshAssigneeCount() {
      const response = await fetch(API_ASSIGNEE_COUNT);
      if (!response.ok) throw new Error("Failed to load assignee count.");
      const payload = await response.json().catch(() => ({}));
      dataAssigneeCount = Math.max(0, Math.round(Number(payload.assignee_count || 0)));
      assigneeCountHelp.textContent = String(dataAssigneeCount);
    }

    async function saveProfile() {
      const payload = buildPayload();
      const response = await fetch(API_CAPACITY, {
        method: "POST",
        headers: { "Content-Type": "application/json" },
        body: JSON.stringify(payload),
      });
      const body = await response.json().catch(() => ({}));
      if (!response.ok) throw new Error(String(body.error || "Failed to save profile."));
      await refreshProfiles();
      const key = payload.from_date + "|" + payload.to_date;
      profileSelect.value = key;
      setStatus("Profile saved.", "ok");
    }

    async function deleteProfile() {
      const key = String(profileSelect.value || "");
      const profile = findProfileByKey(key);
      if (!profile) throw new Error("Select a saved profile to delete.");
      const ok = window.confirm("Delete profile " + profile.from_date + " to " + profile.to_date + "?");
      if (!ok) return;
      const from = encodeURIComponent(String(profile.from_date || ""));
      const to = encodeURIComponent(String(profile.to_date || ""));
      const response = await fetch(API_CAPACITY + "?from=" + from + "&to=" + to, { method: "DELETE" });
      const body = await response.json().catch(() => ({}));
      if (!response.ok) throw new Error(String(body.error || "Failed to delete profile."));
      await refreshProfiles();
      setForm(null);
      setStatus("Profile deleted.", "ok");
    }

    profileSelect.addEventListener("change", () => {
      const selected = findProfileByKey(profileSelect.value);
      if (selected) {
        setForm(selected);
        setStatus("Loaded profile into editor.", "ok");
      }
    });
    refreshBtn.addEventListener("click", async () => {
      try { await refreshProfiles(); } catch (error) { setStatus(error.message || String(error), "err"); }
    });
    newBtn.addEventListener("click", () => {
      profileSelect.value = "";
      setForm(null);
      setStatus("Editing new profile.", "");
    });
    saveBtn.addEventListener("click", async () => {
      try { await saveProfile(); } catch (error) { setStatus(error.message || String(error), "err"); }
    });
    deleteBtn.addEventListener("click", async () => {
      try { await deleteProfile(); } catch (error) { setStatus(error.message || String(error), "err"); }
    });
    useAssigneeCountBtn.addEventListener("click", () => {
      employees.value = String(dataAssigneeCount);
      setStatus("Employees set from assignee count.", "ok");
    });
    holidayAddBtn.addEventListener("click", addHolidayFromPicker);
    holidayDatePicker.addEventListener("change", addHolidayFromPicker);
    holidayClearBtn.addEventListener("click", () => {
      selectedHolidayDates = [];
      renderHolidayList();
    });

    (async function init() {
      setStatus("Loading profiles...", "");
      renderHolidayList();
      try {
        await Promise.all([refreshProfiles(), refreshAssigneeCount()]);
      } catch (error) {
        setStatus(error.message || String(error), "err");
      }
    })();
  </script>
  <script src="/shared-nav.js"></script>
</body>
</html>
""".replace("__SETTINGS_TOP_NAV__", _settings_top_nav_html(CAPACITY_SETTINGS_ROUTE))


def _performance_settings_html() -> str:
    return """<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Performance Point Settings</title>
  <link rel="stylesheet" href="/shared-nav.css">
  <link rel="stylesheet" href="/material-symbols.css">
  <style>
    :root { --bg:#f3f6fb; --card:#fff; --text:#0f172a; --muted:#475569; --line:#cbd5e1; --brand:#1d4ed8; --ok:#065f46; --err:#991b1b; }
    body { margin:0; padding:20px; font-family:"Segoe UI",Tahoma,sans-serif; background:var(--bg); color:var(--text); }
    .card { width:100%; max-width:none; margin:0; background:var(--card); border:1px solid var(--line); border-radius:12px; padding:18px; }
    .top { display:flex; justify-content:space-between; gap:8px; flex-wrap:wrap; align-items:center; }
    .btn { border:1px solid #1e40af; background:var(--brand); color:#fff; border-radius:8px; padding:8px 12px; cursor:pointer; text-decoration:none; display:inline-block; }
    .btn.alt { background:#fff; color:#0f172a; border-color:var(--line); }
    .grid { display:grid; gap:10px; grid-template-columns:repeat(auto-fit,minmax(220px,1fr)); margin-top:14px; }
    label { display:block; font-size:.82rem; font-weight:700; margin-bottom:4px; }
    .label-row { display:flex; align-items:center; gap:6px; }
    .field-info { font-size:1rem; line-height:1; color:#1d4ed8; cursor:pointer; border-radius:999px; padding:2px; }
    .field-info:hover { background:#dbeafe; }
    .field-group { margin-top:14px; border:1px solid #dbeafe; border-radius:10px; background:#f8fbff; padding:10px; }
    .group-title { font-size:.8rem; font-weight:800; letter-spacing:.02em; text-transform:uppercase; color:#334155; }
    .related-box { border:1px solid #bfdbfe; border-radius:10px; background:#eff6ff; padding:8px; }
    .related-title { margin:0 0 6px; font-size:.78rem; font-weight:800; color:#1e3a8a; text-transform:uppercase; letter-spacing:.02em; }
    .related-grid { display:grid; gap:8px; grid-template-columns:repeat(auto-fit,minmax(200px,1fr)); }
    .perf-field { border:1px solid transparent; border-radius:9px; padding:6px; transition:border-color .16s ease, background-color .16s ease, box-shadow .16s ease; }
    .perf-field.is-highlight { border-color:#f59e0b; background:#fffbeb; box-shadow:0 0 0 2px rgba(245,158,11,.2); }
    input { width:100%; box-sizing:border-box; border:1px solid var(--line); border-radius:8px; padding:8px; font-size:.92rem; }
    .row { display:flex; gap:8px; margin-top:14px; flex-wrap:wrap; }
    .team-wrap { margin-top:16px; padding-top:14px; border-top:1px dashed var(--line); }
    .team-grid { display:grid; gap:10px; grid-template-columns: 1fr 1fr; }
    .team-list { margin-top:10px; border:1px solid var(--line); border-radius:8px; background:#f8fbff; max-height:260px; overflow:auto; padding:6px; }
    .team-item { display:flex; justify-content:space-between; align-items:flex-start; gap:8px; border:1px solid #d9e4f0; background:#fff; border-radius:8px; padding:8px; margin-bottom:6px; }
    .team-item:last-child { margin-bottom:0; }
    .team-name { font-weight:700; font-size:.88rem; }
    .team-members { margin-top:4px; color:var(--muted); font-size:.8rem; }
    .guide { margin-top:16px; border:1px solid #c7d2fe; border-radius:12px; background:linear-gradient(135deg,#eef2ff,#f0fdf4); padding:14px; }
    .guide-title { display:flex; align-items:center; gap:8px; margin:0 0 8px; font-size:1.02rem; }
    .guide-title .material-symbols-outlined { color:#1d4ed8; font-size:1.2rem; }
    .formula-box { border:1px solid #bfdbfe; background:#eff6ff; border-radius:10px; padding:10px; }
    .formula-essence { margin:0 0 8px; font-size:.84rem; color:#1e3a8a; line-height:1.45; }
    .formula-code { margin:0; color:#1e3a8a; font-family:Consolas,Monaco,monospace; font-size:.88rem; line-height:1.55; }
    .formula-code .line { display:block; }
    .formula-code [data-target-fields] { cursor:pointer; }
    .formula-code [data-target-fields].is-active { outline:2px solid rgba(37,99,235,.3); outline-offset:1px; }
    .token-reduce { display:inline-block; padding:1px 5px; border-radius:5px; background:#fee2e2; color:#7f1d1d; font-weight:700; }
    .token-add { display:inline-block; padding:1px 5px; border-radius:5px; background:#dcfce7; color:#166534; font-weight:700; }
    .token-neutral { display:inline-block; padding:1px 5px; border-radius:5px; background:#e2e8f0; color:#1e293b; font-weight:700; }
    .ingredients { display:grid; gap:8px; grid-template-columns:repeat(auto-fit,minmax(220px,1fr)); margin-top:10px; }
    .ingredient { border:1px solid #dbeafe; border-radius:10px; background:#fff; padding:9px; transition:border-color .16s ease, box-shadow .16s ease, transform .16s ease; }
    .ingredient[data-target-fields] { cursor:pointer; }
    .ingredient.is-active { border-color:#2563eb; box-shadow:0 8px 18px rgba(37,99,235,.16); transform:translateY(-1px); }
    .ingredient-top { display:flex; align-items:center; gap:6px; font-size:.83rem; font-weight:700; color:#1e3a8a; }
    .ingredient .material-symbols-outlined { font-size:1rem; color:#2563eb; }
    .ingredient p { margin:4px 0 0; font-size:.82rem; color:var(--muted); }
    .explain-grid { display:grid; gap:8px; grid-template-columns:repeat(auto-fit,minmax(260px,1fr)); margin-top:10px; }
    .case { border:1px solid #d1fae5; border-radius:10px; background:#f8fffb; padding:10px; }
    .case h3 { display:flex; align-items:center; gap:6px; margin:0; font-size:.9rem; color:#065f46; }
    .case .material-symbols-outlined { font-size:1rem; color:#0f766e; }
    .case p { margin:6px 0 0; font-size:.82rem; color:#334155; line-height:1.45; }
    .drawer-backdrop { position:fixed; inset:0; background:rgba(15,23,42,.35); opacity:0; pointer-events:none; transition:opacity .18s ease; z-index:80; }
    .drawer { position:fixed; top:0; right:0; width:min(460px,92vw); height:100vh; background:#ffffff; border-left:1px solid #cbd5e1; box-shadow:-14px 0 30px rgba(15,23,42,.18); transform:translateX(100%); transition:transform .2s ease; z-index:81; display:flex; flex-direction:column; }
    .drawer-head { display:flex; align-items:center; justify-content:space-between; gap:8px; padding:14px 14px 10px; border-bottom:1px solid #e2e8f0; background:#f8fbff; }
    .drawer-title { margin:0; font-size:1rem; color:#0f172a; }
    .drawer-close { border:1px solid #cbd5e1; background:#fff; color:#334155; border-radius:8px; padding:6px 10px; cursor:pointer; }
    .drawer-body { padding:14px; overflow:auto; }
    .drawer-body p { margin:0; color:#334155; line-height:1.55; font-size:.9rem; white-space:pre-wrap; }
    body.drawer-open .drawer-backdrop { opacity:1; pointer-events:auto; }
    body.drawer-open .drawer { transform:translateX(0); }
    #team-assignees { min-height:160px; }
    #status { margin-top:10px; min-height:1.2em; }
    #status.ok { color:var(--ok); } #status.err { color:var(--err); }
  </style>
</head>
<body>
  <main class="card">
    <div class="top">
      <div>
        <h1 style="margin:0;font-size:1.25rem;">Performance Point Settings</h1>
        <p style="margin:6px 0 0;color:var(--muted);font-size:.92rem;">Configure penalty points for Employee Performance report.</p>
      </div>
      <div style="display:flex;gap:8px;flex-wrap:wrap;">""" + _settings_top_nav_html(PERFORMANCE_SETTINGS_ROUTE) + """</div>
    </div>
    <section class="field-group">
      <div class="group-title">Score Foundation</div>
      <div class="grid" style="margin-top:8px;">
        <div class="perf-field" data-field-key="base_score">
          <label for="base-score" class="label-row">Base Score
            <span class="material-symbols-outlined field-info" role="button" tabindex="0" data-info-title="Base Score" data-info-body="Starting performance score before penalties are applied.\n\nMeaning for performance calculation:\n- It is the additive anchor in `Raw Score = base_score - Total Penalty`.\n- Increasing Base Score raises everyone’s starting point.\n- Use carefully so penalties still create meaningful separation." aria-label="Base Score info">info</span>
          </label>
          <input id="base-score" type="number" step="0.1">
        </div>
        <div class="related-box">
          <p class="related-title">Related Bounds</p>
          <div class="related-grid">
            <div class="perf-field" data-field-key="min_score">
              <label for="min-score" class="label-row">Min Score
                <span class="material-symbols-outlined field-info" role="button" tabindex="0" data-info-title="Min Score" data-info-body="Lower clamp boundary.\n\nMeaning for performance calculation:\n- Final score cannot go below this value.\n- Protects against extreme penalties causing unstable negative/very-low outputs.\n- Higher Min Score compresses low-end differentiation." aria-label="Min Score info">info</span>
              </label>
              <input id="min-score" type="number" step="0.1">
            </div>
            <div class="perf-field" data-field-key="max_score">
              <label for="max-score" class="label-row">Max Score
                <span class="material-symbols-outlined field-info" role="button" tabindex="0" data-info-title="Max Score" data-info-body="Upper clamp boundary.\n\nMeaning for performance calculation:\n- Final score cannot exceed this value.\n- Prevents inflated scores when Base Score is high.\n- Lower Max Score compresses top-end differentiation." aria-label="Max Score info">info</span>
              </label>
              <input id="max-score" type="number" step="0.1">
            </div>
          </div>
        </div>
      </div>
    </section>
    <section class="field-group">
      <div class="group-title">Penalty Multipliers</div>
      <div class="grid" style="margin-top:8px;">
        <div class="perf-field" data-field-key="points_per_bug_hour">
          <label for="bug-hour" class="label-row">Points per Bug Hour
            <span class="material-symbols-outlined field-info" role="button" tabindex="0" data-info-title="Points per Bug Hour" data-info-body="Penalty rate applied to bug hours.\n\nMeaning for performance calculation:\n- Contributes to Total Penalty as `bug_hours * points_per_bug_hour`.\n- Higher value makes quality-related rework deduct score faster.\n- Use to reflect seriousness of bug workload in scoring." aria-label="Points per Bug Hour info">info</span>
          </label>
          <input id="bug-hour" type="number" min="0" step="0.01">
        </div>
        <div class="perf-field" data-field-key="points_per_bug_late_hour">
          <label for="bug-late-hour" class="label-row">Points per Bug Late Hour
            <span class="material-symbols-outlined field-info" role="button" tabindex="0" data-info-title="Points per Bug Late Hour" data-info-body="Penalty rate for late bug resolution hours.\n\nMeaning for performance calculation:\n- Contributes as `bug_late_hours * points_per_bug_late_hour`.\n- Emphasizes delay in quality closure and SLA risk.\n- Higher value prioritizes timeliness of bug fixing." aria-label="Points per Bug Late Hour info">info</span>
          </label>
          <input id="bug-late-hour" type="number" min="0" step="0.01">
        </div>
        <div class="perf-field" data-field-key="points_per_unplanned_leave_hour">
          <label for="leave-hour" class="label-row">Points per Unplanned Leave Hour
            <span class="material-symbols-outlined field-info" role="button" tabindex="0" data-info-title="Points per Unplanned Leave Hour" data-info-body="Penalty rate for unplanned leave hours.\n\nMeaning for performance calculation:\n- Contributes as `unplanned_leave_hours * points_per_unplanned_leave_hour`.\n- Captures disruption risk from unexpected availability loss.\n- Higher value increases penalty sensitivity to sudden leave." aria-label="Points per Unplanned Leave Hour info">info</span>
          </label>
          <input id="leave-hour" type="number" min="0" step="0.01">
        </div>
        <div class="perf-field" data-field-key="points_per_subtask_late_hour">
          <label for="subtask-late-hour" class="label-row">Points per Subtask Late Hour
            <span class="material-symbols-outlined field-info" role="button" tabindex="0" data-info-title="Points per Subtask Late Hour" data-info-body="Penalty rate for subtask late hours.\n\nMeaning for performance calculation:\n- Contributes as `subtask_late_hours * points_per_subtask_late_hour`.\n- Converts schedule slippage into measurable score deduction.\n- Higher value penalizes execution delays more strongly." aria-label="Points per Subtask Late Hour info">info</span>
          </label>
          <input id="subtask-late-hour" type="number" min="0" step="0.01">
        </div>
        <div class="perf-field" data-field-key="points_per_estimate_overrun_hour">
          <label for="estimate-hour" class="label-row">Points per Estimate Overrun Hour
            <span class="material-symbols-outlined field-info" role="button" tabindex="0" data-info-title="Points per Estimate Overrun Hour" data-info-body="Penalty rate for estimate overrun hours.\n\nMeaning for performance calculation:\n- Contributes as `estimate_overrun_hours * points_per_estimate_overrun_hour`.\n- Measures planning and estimation discipline.\n- Higher value increases penalty for underestimation/effort drift." aria-label="Points per Estimate Overrun Hour info">info</span>
          </label>
          <input id="estimate-hour" type="number" min="0" step="0.01">
        </div>
        <div class="perf-field" data-field-key="points_per_missed_due_date">
          <label for="missed-due-date" class="label-row">Points per Missed Due Date
            <span class="material-symbols-outlined field-info" role="button" tabindex="0" data-info-title="Points per Missed Due Date" data-info-body="Fixed penalty per missed due date event.\n\nMeaning for performance calculation:\n- Contributes as `missed_due_dates * points_per_missed_due_date`.\n- Encodes delivery commitment reliability in final score.\n- Higher value makes date misses more costly than hour-only signals." aria-label="Points per Missed Due Date info">info</span>
          </label>
          <input id="missed-due-date" type="number" min="0" step="0.01">
        </div>
      </div>
    </section>
    <div class="row">
      <button class="btn alt" type="button" id="reload-btn">Reload</button>
      <button class="btn alt" type="button" id="reset-btn">Reset Defaults</button>
      <button class="btn" type="button" id="save-btn">Save</button>
    </div>
    <section class="guide">
      <h2 class="guide-title"><span class="material-symbols-outlined" aria-hidden="true">calculate</span>Performance Formula Guide</h2>
      <div class="formula-box">
        <p class="formula-essence">Essence: this is a penalty-first model. We start from a base score and deduct penalty points for delivery quality, delay, overrun, and unplanned leave factors. In this formula, only `base_score` adds points; all other scoring ingredients reduce points.</p>
        <p class="formula-code">
          <span class="line">1) Total Penalty =</span>
          <span class="line">   (<span class="token-reduce" data-target-fields="points_per_bug_hour" tabindex="0">bug_hours * points_per_bug_hour</span>) +</span>
          <span class="line">   (<span class="token-reduce" data-target-fields="points_per_bug_late_hour" tabindex="0">bug_late_hours * points_per_bug_late_hour</span>) +</span>
          <span class="line">   (<span class="token-reduce" data-target-fields="points_per_unplanned_leave_hour" tabindex="0">unplanned_leave_hours * points_per_unplanned_leave_hour</span>) +</span>
          <span class="line">   (<span class="token-reduce" data-target-fields="points_per_subtask_late_hour" tabindex="0">subtask_late_hours * points_per_subtask_late_hour</span>) +</span>
          <span class="line">   (<span class="token-reduce" data-target-fields="points_per_estimate_overrun_hour" tabindex="0">estimate_overrun_hours * points_per_estimate_overrun_hour</span>) +</span>
          <span class="line">   (<span class="token-reduce" data-target-fields="points_per_missed_due_date" tabindex="0">missed_due_dates * points_per_missed_due_date</span>)</span>
          <span class="line" style="margin-top:6px;">2) Raw Score = <span class="token-add" data-target-fields="base_score" tabindex="0">base_score</span> - <span class="token-reduce">Total Penalty</span></span>
          <span class="line" style="margin-top:6px;">3) Final Score = <span class="token-neutral" data-target-fields="min_score,max_score" tabindex="0">clamp(Raw Score, min_score, max_score)</span></span>
        </p>
      </div>
      <div class="ingredients">
        <article class="ingredient" data-target-fields="points_per_bug_hour,points_per_bug_late_hour" tabindex="0">
          <div class="ingredient-top"><span class="material-symbols-outlined" aria-hidden="true">bug_report</span>Bug Quality Impact</div>
          <p>Higher bug hours and late bug closure reduce score faster when severity or rework is high.</p>
        </article>
        <article class="ingredient" data-target-fields="points_per_unplanned_leave_hour" tabindex="0">
          <div class="ingredient-top"><span class="material-symbols-outlined" aria-hidden="true">event_busy</span>Unplanned Leave Impact</div>
          <p>Unexpected leave hours apply a direct penalty and reflect sudden execution gaps.</p>
        </article>
        <article class="ingredient" data-target-fields="points_per_subtask_late_hour" tabindex="0">
          <div class="ingredient-top"><span class="material-symbols-outlined" aria-hidden="true">schedule</span>Subtask Delay Impact</div>
          <p>Subtask late hours convert schedule drift into score deductions.</p>
        </article>
        <article class="ingredient" data-target-fields="points_per_estimate_overrun_hour" tabindex="0">
          <div class="ingredient-top"><span class="material-symbols-outlined" aria-hidden="true">timer</span>Estimation Accuracy</div>
          <p>Estimate overrun hours track planning accuracy and penalize larger overruns.</p>
        </article>
        <article class="ingredient" data-target-fields="points_per_missed_due_date" tabindex="0">
          <div class="ingredient-top"><span class="material-symbols-outlined" aria-hidden="true">assignment_late</span>Due Date Breaches</div>
          <p>Each missed due date adds a fixed point deduction for delivery discipline.</p>
        </article>
        <article class="ingredient" data-target-fields="min_score,max_score" tabindex="0">
          <div class="ingredient-top"><span class="material-symbols-outlined" aria-hidden="true">shield</span>Score Bounds</div>
          <p>`min_score` and `max_score` cap extreme outcomes to keep scoring stable.</p>
        </article>
        <article class="ingredient" data-target-fields="base_score" tabindex="0">
          <div class="ingredient-top"><span class="material-symbols-outlined" aria-hidden="true">add_circle</span>Positive Contributor (Applicable)</div>
          <p>`base_score` is the additive anchor. All configurable penalty multipliers on this page are deduction terms.</p>
        </article>
      </div>
      <div class="explain-grid">
        <article class="case">
          <h3><span class="material-symbols-outlined" aria-hidden="true">trending_up</span>Case A: Balanced Delivery</h3>
          <p>Given `base_score=100`, penalty total of `14.5` gives `Raw Score=85.5`. If bounds are `40..100`, final remains `85.5` because it is inside range.</p>
        </article>
        <article class="case">
          <h3><span class="material-symbols-outlined" aria-hidden="true">warning</span>Case B: Heavy Delay and Overrun</h3>
          <p>If penalties rise to `73` with `base_score=100`, `Raw Score=27`. With `min_score=40`, final score is clamped up to `40` to represent floor risk case.</p>
        </article>
        <article class="case">
          <h3><span class="material-symbols-outlined" aria-hidden="true">rule</span>Case C: No Penalty Period</h3>
          <p>When all ingredients are zero, `Total Penalty=0`, so `Raw Score=base_score`. Final equals base score unless base exceeds `max_score`, where it is capped.</p>
        </article>
      </div>
    </section>
    <section class="team-wrap">
      <h2 style="margin:0 0 8px;font-size:1rem;">Team Management</h2>
      <p style="margin:0;color:var(--muted);font-size:.88rem;">Create quick teams by selecting assignees.</p>
      <div class="team-grid">
        <div>
          <label for="team-name">Team Name</label>
          <input id="team-name" type="text" maxlength="80" placeholder="e.g. Alpha Squad">
          <label for="team-leader" style="margin-top:8px;">Team Leader</label>
          <select id="team-leader"></select>
          <div class="row" style="margin-top:8px;">
            <button class="btn" type="button" id="create-team-btn">Create / Update Team</button>
            <button class="btn alt" type="button" id="reload-teams-btn">Reload Teams</button>
          </div>
        </div>
        <div>
          <label for="team-assignees">Assignees (multi-select)</label>
          <select id="team-assignees" multiple></select>
        </div>
      </div>
      <div class="team-list" id="team-list"></div>
    </section>
    <div id="status"></div>
  </main>
  <div id="field-drawer-backdrop" class="drawer-backdrop" aria-hidden="true"></div>
  <aside id="field-drawer" class="drawer" aria-hidden="true" aria-labelledby="field-drawer-title" aria-modal="true" role="dialog">
    <div class="drawer-head">
      <h3 id="field-drawer-title" class="drawer-title">Field Details</h3>
      <button id="field-drawer-close" type="button" class="drawer-close">Close</button>
    </div>
    <div class="drawer-body">
      <p id="field-drawer-body"></p>
    </div>
  </aside>
  <script>
    const DEFAULTS = """ + json.dumps(DEFAULT_PERFORMANCE_SETTINGS) + """;
    const API = "/api/performance/settings";
    const ASSIGNEES_API = "/api/performance/assignees";
    const TEAMS_API = "/api/performance/teams";
    const statusEl = document.getElementById("status");
    const drawerEl = document.getElementById("field-drawer");
    const drawerBackdropEl = document.getElementById("field-drawer-backdrop");
    const drawerTitleEl = document.getElementById("field-drawer-title");
    const drawerBodyEl = document.getElementById("field-drawer-body");
    const drawerCloseEl = document.getElementById("field-drawer-close");
    const fieldInfoButtons = Array.from(document.querySelectorAll(".field-info[data-info-title][data-info-body]"));
    const fields = {
      base_score: document.getElementById("base-score"),
      min_score: document.getElementById("min-score"),
      max_score: document.getElementById("max-score"),
      points_per_bug_hour: document.getElementById("bug-hour"),
      points_per_bug_late_hour: document.getElementById("bug-late-hour"),
      points_per_unplanned_leave_hour: document.getElementById("leave-hour"),
      points_per_subtask_late_hour: document.getElementById("subtask-late-hour"),
      points_per_estimate_overrun_hour: document.getElementById("estimate-hour"),
      points_per_missed_due_date: document.getElementById("missed-due-date")
    };
    const fieldContainers = Array.from(document.querySelectorAll(".perf-field[data-field-key]"));
    const fieldContainerByKey = {};
    fieldContainers.forEach((el) => { fieldContainerByKey[String(el.getAttribute("data-field-key") || "")] = el; });
    const ingredientCards = Array.from(document.querySelectorAll(".ingredient[data-target-fields]"));
    const formulaLinks = Array.from(document.querySelectorAll(".formula-code [data-target-fields]"));
    const interactiveLinks = ingredientCards.concat(formulaLinks);
    const teamNameEl = document.getElementById("team-name");
    const teamLeaderEl = document.getElementById("team-leader");
    const teamAssigneesEl = document.getElementById("team-assignees");
    const teamListEl = document.getElementById("team-list");
    function setStatus(msg, kind) { statusEl.textContent = msg || ""; statusEl.className = kind || ""; }
    function setForm(settings) { for (const k of Object.keys(fields)) fields[k].value = String(Number(settings[k] ?? 0)); }
    function readForm() { const out = {}; for (const k of Object.keys(fields)) out[k] = Number(fields[k].value || 0); return out; }
    function esc(text) { return String(text || "").replace(/&/g, "&amp;").replace(/</g, "&lt;").replace(/>/g, "&gt;"); }
    function openFieldDrawer(title, body) {
      drawerTitleEl.textContent = String(title || "Field Details");
      drawerBodyEl.textContent = String(body || "");
      document.body.classList.add("drawer-open");
      drawerEl.setAttribute("aria-hidden", "false");
      drawerBackdropEl.setAttribute("aria-hidden", "false");
      drawerCloseEl.focus();
    }
    function closeFieldDrawer() {
      document.body.classList.remove("drawer-open");
      drawerEl.setAttribute("aria-hidden", "true");
      drawerBackdropEl.setAttribute("aria-hidden", "true");
    }
    fieldInfoButtons.forEach((btn) => {
      const open = () => openFieldDrawer(btn.getAttribute("data-info-title"), btn.getAttribute("data-info-body"));
      btn.addEventListener("click", open);
      btn.addEventListener("keydown", (event) => {
        if (event.key === "Enter" || event.key === " ") {
          event.preventDefault();
          open();
        }
      });
    });
    drawerCloseEl.addEventListener("click", closeFieldDrawer);
    drawerBackdropEl.addEventListener("click", closeFieldDrawer);
    document.addEventListener("keydown", (event) => {
      if (event.key === "Escape" && document.body.classList.contains("drawer-open")) closeFieldDrawer();
    });
    function clearLinkedHighlights() {
      interactiveLinks.forEach((el) => el.classList.remove("is-active"));
      fieldContainers.forEach((el) => el.classList.remove("is-highlight"));
    }
    function highlightTargetsFromElement(element) {
      clearLinkedHighlights();
      if (!element) return;
      element.classList.add("is-active");
      const targets = String(element.getAttribute("data-target-fields") || "")
        .split(",")
        .map((x) => x.trim())
        .filter(Boolean);
      targets.forEach((key) => {
        const targetEl = fieldContainerByKey[key];
        if (targetEl) targetEl.classList.add("is-highlight");
      });
    }
    interactiveLinks.forEach((el) => {
      el.addEventListener("mouseenter", () => highlightTargetsFromElement(el));
      el.addEventListener("mouseleave", () => clearLinkedHighlights());
      el.addEventListener("focus", () => highlightTargetsFromElement(el));
      el.addEventListener("blur", () => clearLinkedHighlights());
    });
    async function loadSettings() {
      setStatus("Loading...", "");
      const response = await fetch(API);
      const data = await response.json();
      if (!response.ok) throw new Error(data.error || "Failed to load settings.");
      setForm(data.settings || DEFAULTS);
      setStatus("Loaded.", "ok");
    }
    document.getElementById("reload-btn").addEventListener("click", () => { loadSettings().catch((e) => setStatus(e.message || String(e), "err")); });
    document.getElementById("reset-btn").addEventListener("click", () => { setForm(DEFAULTS); setStatus("Defaults loaded in form. Save to persist.", ""); });
    document.getElementById("save-btn").addEventListener("click", async () => {
      try {
        const response = await fetch(API, { method: "POST", headers: { "Content-Type": "application/json" }, body: JSON.stringify(readForm()) });
        const data = await response.json();
        if (!response.ok) throw new Error(data.error || "Failed to save settings.");
        setForm(data.settings || DEFAULTS);
        setStatus("Saved.", "ok");
      } catch (err) {
        setStatus(err.message || String(err), "err");
      }
    });
    async function loadAssignees() {
      const response = await fetch(ASSIGNEES_API);
      const data = await response.json();
      if (!response.ok) throw new Error(data.error || "Failed to load assignees.");
      const assignees = Array.isArray(data.assignees) ? data.assignees : [];
      teamAssigneesEl.innerHTML = assignees.map((name) => '<option value="' + esc(name) + '">' + esc(name) + '</option>').join("");
      teamLeaderEl.innerHTML = '<option value="">Select team leader</option>' + assignees.map((name) => '<option value="' + esc(name) + '">' + esc(name) + '</option>').join("");
    }
    function renderTeams(teams) {
      if (!Array.isArray(teams) || !teams.length) {
        teamListEl.innerHTML = '<div style="color:#64748b;font-size:.85rem;">No teams created yet.</div>';
        return;
      }
      teamListEl.innerHTML = teams.map((team) => {
        const name = String(team.team_name || "");
        const members = Array.isArray(team.assignees) ? team.assignees : [];
        const membersText = members.length ? members.join(", ") : "-";
        const leader = String(team.team_leader || "-");
        return '<div class="team-item"><div><div class="team-name">' + esc(name) + ' <span style="font-weight:600;color:#475569;">(Lead: ' + esc(leader) + ')</span></div><div class="team-members">' + esc(membersText) + '</div></div><button class="btn alt" type="button" data-del-team="' + esc(name) + '">Delete</button></div>';
      }).join("");
      Array.from(teamListEl.querySelectorAll("button[data-del-team]")).forEach((btn) => {
        btn.addEventListener("click", async () => {
          const teamName = String(btn.getAttribute("data-del-team") || "");
          if (!teamName) return;
          try {
            const response = await fetch(TEAMS_API + '/' + encodeURIComponent(teamName), { method: "DELETE" });
            const data = await response.json();
            if (!response.ok) throw new Error(data.error || "Failed to delete team.");
            setStatus("Team deleted.", "ok");
            await loadTeams();
          } catch (err) {
            setStatus(err.message || String(err), "err");
          }
        });
      });
    }
    async function loadTeams() {
      const response = await fetch(TEAMS_API);
      const data = await response.json();
      if (!response.ok) throw new Error(data.error || "Failed to load teams.");
      renderTeams(Array.isArray(data.teams) ? data.teams : []);
    }
    document.getElementById("reload-teams-btn").addEventListener("click", () => {
      Promise.all([loadAssignees(), loadTeams()]).then(() => setStatus("Teams refreshed.", "ok")).catch((e) => setStatus(e.message || String(e), "err"));
    });
    document.getElementById("create-team-btn").addEventListener("click", async () => {
      try {
        const teamName = String(teamNameEl.value || "").trim();
        const teamLeader = String(teamLeaderEl.value || "").trim();
        const assignees = Array.from(teamAssigneesEl.selectedOptions).map((o) => String(o.value || "").trim()).filter(Boolean);
        const response = await fetch(TEAMS_API, {
          method: "POST",
          headers: { "Content-Type": "application/json" },
          body: JSON.stringify({ team_name: teamName, team_leader: teamLeader, assignees }),
        });
        const data = await response.json();
        if (!response.ok) throw new Error(data.error || "Failed to save team.");
        teamNameEl.value = "";
        teamLeaderEl.value = "";
        Array.from(teamAssigneesEl.options).forEach((o) => { o.selected = false; });
        setStatus("Team saved.", "ok");
        await loadTeams();
      } catch (err) {
        setStatus(err.message || String(err), "err");
      }
    });
    Promise.all([loadSettings(), loadAssignees(), loadTeams()]).catch((e) => setStatus(e.message || String(e), "err"));
  </script>
  <script src="/shared-nav.js"></script>
</body>
</html>"""


def _dashboard_risk_settings_html() -> str:
    return """<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Dashboard Risk Settings</title>
  <link rel="stylesheet" href="/shared-nav.css">
  <link rel="stylesheet" href="/material-symbols.css">
  <style>
    :root { --bg:#f8fafc; --card:#ffffff; --ink:#0f172a; --muted:#475569; --line:#dbe6f5; --brand:#0f766e; --ok:#166534; --err:#b91c1c; --chip-low:#166534; --chip-can:#a16207; --chip-med:#c2410c; --chip-high:#b91c1c; }
    * { box-sizing:border-box; }
    body { margin:0; padding:20px; font-family:"Segoe UI",Tahoma,sans-serif; color:var(--ink); background:linear-gradient(180deg,#f8fafc,#eef2ff); }
    .wrap { max-width:1240px; margin:0 auto; display:grid; gap:14px; }
    .card { background:var(--card); border:1px solid var(--line); border-radius:12px; padding:14px; box-shadow:0 8px 20px rgba(15,23,42,.05); }
    .top { display:flex; justify-content:space-between; align-items:flex-start; gap:12px; flex-wrap:wrap; }
    .row { display:flex; gap:8px; flex-wrap:wrap; margin-top:10px; }
    .btn { border:1px solid #0f766e; background:#0f766e; color:#fff; border-radius:8px; padding:8px 12px; cursor:pointer; text-decoration:none; }
    .btn.alt { background:#fff; color:#0f172a; border-color:var(--line); }
    .grid { display:grid; gap:10px; grid-template-columns:repeat(auto-fit,minmax(220px,1fr)); margin-top:10px; }
    .grid.five { grid-template-columns:repeat(auto-fit,minmax(180px,1fr)); }
    label { display:block; font-size:.8rem; font-weight:700; margin-bottom:4px; }
    input { width:100%; border:1px solid var(--line); border-radius:8px; padding:8px; font-size:.92rem; }
    .hint { margin:4px 0 0; font-size:.78rem; color:var(--muted); }
    .explain { margin:0; color:var(--muted); font-size:.9rem; }
    .list { margin:10px 0 0; padding-left:18px; color:#334155; font-size:.88rem; }
    .list li { margin:4px 0; }
    .table-wrap { margin-top:10px; border:1px solid var(--line); border-radius:10px; overflow:auto; }
    table { width:100%; min-width:760px; border-collapse:collapse; }
    th, td { border-bottom:1px solid #e2e8f0; text-align:left; padding:8px; font-size:.85rem; vertical-align:top; }
    th { background:#f1f5f9; text-transform:uppercase; letter-spacing:.03em; font-size:.75rem; color:#334155; }
    .preview-grid { display:grid; gap:10px; grid-template-columns:repeat(auto-fit,minmax(270px,1fr)); margin-top:10px; }
    .sample-card { border:1px solid #d1d5db; border-radius:10px; padding:10px; background:#fff; }
    .sample-top { display:flex; justify-content:space-between; gap:8px; align-items:flex-start; }
    .sample-type { font-size:.72rem; text-transform:uppercase; color:#64748b; font-weight:700; }
    .sample-title { margin:2px 0 0; font-weight:700; font-size:.92rem; }
    .sample-status { margin-top:6px; font-size:.82rem; color:#334155; }
    .chip { display:inline-flex; align-items:center; border-radius:999px; padding:3px 10px; font-size:.78rem; font-weight:700; border:1px solid transparent; white-space:nowrap; }
    .risk-low { color:var(--chip-low); border-color:#86efac; background:#dcfce7; }
    .risk-can-be { color:var(--chip-can); border-color:#fde68a; background:#fef9c3; }
    .risk-medium { color:var(--chip-med); border-color:#fdba74; background:#ffedd5; }
    .risk-high { color:var(--chip-high); border-color:#fca5a5; background:#fee2e2; }
    .reasons { margin-top:8px; font-size:.82rem; color:#334155; }
    .reasons ul { margin:6px 0 0; padding-left:18px; }
    #status { min-height:1.2em; margin-top:8px; font-size:.9rem; }
    #status.ok { color:var(--ok); }
    #status.err { color:var(--err); }
    @media (max-width:760px) { body { padding:10px; } .card { padding:10px; } }
  </style>
</head>
<body>
  <main class="wrap">
    <section class="card">
      <div class="top">
        <div>
          <h1 style="margin:0;font-size:1.24rem;">Dashboard Risk Settings</h1>
          <p class="explain">Configure point weights and level labels used for `AT RISK` chips on dashboard cards. Jira status remains unchanged.</p>
        </div>
        <div class="row">__SETTINGS_TOP_NAV__</div>
      </div>
      <div class="grid">
        <div>
          <label for="p-subtask-linear-lag">Subtask: Linear effort lag points</label>
          <input id="p-subtask-linear-lag" type="number" min="0" step="1">
          <p class="hint">Expected hours to date are higher than logged hours while unresolved.</p>
        </div>
        <div>
          <label for="p-due-crossed">Any item: Due crossed unresolved points</label>
          <input id="p-due-crossed" type="number" min="0" step="1">
          <p class="hint">Current date is past planned end date and item is unresolved.</p>
        </div>
        <div>
          <label for="p-late-start">Subtask: Late actual start points</label>
          <input id="p-late-start" type="number" min="0" step="1">
          <p class="hint">Actual start date is later than planned start date.</p>
        </div>
        <div>
          <label for="p-start-passed-not-progress">Any item: Start passed, not in progress points</label>
          <input id="p-start-passed-not-progress" type="number" min="0" step="1">
          <p class="hint">Current date is past planned start and status is still pre-execution.</p>
        </div>
        <div>
          <label for="p-inherited">Story/Epic: Inherited risky child points</label>
          <input id="p-inherited" type="number" min="0" step="1">
          <p class="hint">Applied when at least one descendant is already marked at-risk.</p>
        </div>
      </div>
      <div class="grid four" style="margin-top:12px;">
        <div>
          <label for="t-can-be">Threshold: Can Be minimum score</label>
          <input id="t-can-be" type="number" min="0" step="1">
        </div>
        <div>
          <label for="t-medium">Threshold: Medium minimum score</label>
          <input id="t-medium" type="number" min="0" step="1">
        </div>
        <div>
          <label for="t-high">Threshold: Highly At Risk minimum score</label>
          <input id="t-high" type="number" min="0" step="1">
        </div>
        <div>
          <label for="t-at-risk">Threshold: `is_at_risk` minimum score</label>
          <input id="t-at-risk" type="number" min="0" step="1">
        </div>
      </div>
      <div class="grid four" style="margin-top:12px;">
        <div><label for="l-low">Label for score below Can Be</label><input id="l-low" type="text" maxlength="32"></div>
        <div><label for="l-can-be">Label for Can Be level</label><input id="l-can-be" type="text" maxlength="32"></div>
        <div><label for="l-medium">Label for Medium level</label><input id="l-medium" type="text" maxlength="32"></div>
        <div><label for="l-high">Label for Highly At Risk level</label><input id="l-high" type="text" maxlength="32"></div>
      </div>
      <div class="row">
        <button class="btn alt" type="button" id="reload-btn">Reload</button>
        <button class="btn alt" type="button" id="reset-btn">Reset Defaults</button>
        <button class="btn" type="button" id="save-btn">Save Settings</button>
      </div>
      <div id="status"></div>
    </section>

    <section class="card">
      <h2 style="margin:0;font-size:1rem;">How these settings impact scoring</h2>
      <div class="table-wrap">
        <table>
          <thead><tr><th>Indicator</th><th>Applies To</th><th>When Triggered</th><th>Impact</th></tr></thead>
          <tbody>
            <tr><td>Linear effort lag</td><td>Subtask / Bug Subtask</td><td>Expected-to-date hours are greater than logged hours and unresolved.</td><td>Pushes active execution items toward Medium quickly.</td></tr>
            <tr><td>Due crossed unresolved</td><td>Subtask / Story / Epic</td><td>Today is greater than planned end while unresolved.</td><td>Strong schedule breach indicator.</td></tr>
            <tr><td>Late actual start</td><td>Subtask / Bug Subtask</td><td>Actual start date is after planned start date.</td><td>Early warning, usually lower severity.</td></tr>
            <tr><td>Start passed not in progress</td><td>Subtask / Story / Epic</td><td>Today is greater than planned start but status is still not in progress.</td><td>Readiness delay signal.</td></tr>
            <tr><td>Inherited child risk</td><td>Story / Epic</td><td>At least one descendant has final score at or above at-risk threshold.</td><td>Propagates risk up hierarchy without summing all children.</td></tr>
          </tbody>
        </table>
      </div>
      <ul class="list">
        <li>Resolved/Done/Closed items are forced to score `0` and level `Low` regardless of dates or hours.</li>
        <li>Story/Epic final score uses `max(self score, strongest child score)` to avoid score inflation.</li>
        <li>`is_at_risk` is true when score is greater than or equal to `at_risk_min`.</li>
      </ul>
    </section>

    <section class="card">
      <h2 style="margin:0;font-size:1rem;">Live Examples</h2>
      <p class="explain">These previews update instantly as you edit points, thresholds, and labels.</p>
      <div id="preview-grid" class="preview-grid"></div>
    </section>
  </main>
  <script>
    const DEFAULTS = """ + json.dumps(_default_dashboard_risk_settings()) + """;
    const API = "/api/dashboard-risk/settings";
    const fields = {
      indicator_points: {
        subtask_linear_lag: document.getElementById("p-subtask-linear-lag"),
        due_crossed_unresolved: document.getElementById("p-due-crossed"),
        subtask_late_actual_start: document.getElementById("p-late-start"),
        start_passed_not_in_progress: document.getElementById("p-start-passed-not-progress"),
        inherited_child_risk: document.getElementById("p-inherited"),
      },
      thresholds: {
        can_be_min: document.getElementById("t-can-be"),
        medium_min: document.getElementById("t-medium"),
        high_min: document.getElementById("t-high"),
        at_risk_min: document.getElementById("t-at-risk"),
      },
      labels: {
        low: document.getElementById("l-low"),
        can_be: document.getElementById("l-can-be"),
        medium: document.getElementById("l-medium"),
        high: document.getElementById("l-high"),
      },
    };
    const statusEl = document.getElementById("status");
    const previewGrid = document.getElementById("preview-grid");

    function esc(text) {
      return String(text || "").replace(/&/g, "&amp;").replace(/</g, "&lt;").replace(/>/g, "&gt;");
    }
    function setStatus(msg, kind) {
      statusEl.textContent = msg || "";
      statusEl.className = kind || "";
    }
    function asInt(v, fallback) {
      const n = Number(v);
      return Number.isFinite(n) ? Math.max(0, Math.round(n)) : fallback;
    }
    function setForm(settings) {
      const use = settings || DEFAULTS;
      for (const [k, el] of Object.entries(fields.indicator_points)) el.value = String(asInt(use.indicator_points?.[k], DEFAULTS.indicator_points[k]));
      for (const [k, el] of Object.entries(fields.thresholds)) el.value = String(asInt(use.thresholds?.[k], DEFAULTS.thresholds[k]));
      for (const [k, el] of Object.entries(fields.labels)) el.value = String(use.labels?.[k] || DEFAULTS.labels[k] || "");
      renderPreviews(readForm());
    }
    function readForm() {
      const out = { indicator_points: {}, thresholds: {}, labels: {} };
      for (const [k, el] of Object.entries(fields.indicator_points)) out.indicator_points[k] = asInt(el.value, DEFAULTS.indicator_points[k]);
      for (const [k, el] of Object.entries(fields.thresholds)) out.thresholds[k] = asInt(el.value, DEFAULTS.thresholds[k]);
      for (const [k, el] of Object.entries(fields.labels)) out.labels[k] = String(el.value || DEFAULTS.labels[k] || "").trim();
      return out;
    }
    function levelFromScore(score, settings) {
      const s = asInt(score, 0);
      const t = settings.thresholds || DEFAULTS.thresholds;
      if (s >= asInt(t.high_min, 4)) return "high";
      if (s >= asInt(t.medium_min, 2)) return "medium";
      if (s >= asInt(t.can_be_min, 1)) return "can_be";
      return "low";
    }
    function classFromLevel(level) {
      if (level === "high") return "risk-high";
      if (level === "medium") return "risk-medium";
      if (level === "can_be") return "risk-can-be";
      return "risk-low";
    }
    function labelFromLevel(level, settings) {
      const labels = settings.labels || DEFAULTS.labels;
      return String(labels[level] || DEFAULTS.labels[level] || level);
    }
    function renderPreviews(settings) {
      const p = settings.indicator_points || DEFAULTS.indicator_points;
      const scenarios = [
        {
          type: "Subtask",
          title: "Under-logged in active window",
          jiraStatus: "In Progress",
          indicators: [{ key: "subtask_linear_lag", label: "Linear effort lag" }],
        },
        {
          type: "Subtask",
          title: "Late actual start only",
          jiraStatus: "In Progress",
          indicators: [{ key: "subtask_late_actual_start", label: "Late actual start" }],
        },
        {
          type: "Story",
          title: "Risk inherited from child subtask",
          jiraStatus: "In Progress",
          indicators: [{ key: "inherited_child_risk", label: "Inherited risky child" }],
        },
        {
          type: "Epic",
          title: "Overdue and not started",
          jiraStatus: "To Do",
          indicators: [
            { key: "due_crossed_unresolved", label: "Due crossed unresolved" },
            { key: "start_passed_not_in_progress", label: "Start passed not in progress" },
          ],
        },
      ];
      previewGrid.innerHTML = scenarios.map((scenario) => {
        const reasons = [];
        let score = 0;
        scenario.indicators.forEach((entry) => {
          const pts = asInt(p[entry.key], 0);
          score += pts;
          reasons.push("+" + pts + " " + entry.label);
        });
        const level = levelFromScore(score, settings);
        const label = labelFromLevel(level, settings);
        return [
          '<article class="sample-card">',
          '<div class="sample-top">',
          '<div><div class="sample-type">' + esc(scenario.type) + '</div><div class="sample-title">' + esc(scenario.title) + '</div></div>',
          '<span class="chip ' + classFromLevel(level) + '">RISK: ' + esc(label) + ' (' + score + ')</span>',
          '</div>',
          '<div class="sample-status"><strong>Jira Status:</strong> ' + esc(scenario.jiraStatus) + '</div>',
          '<div class="reasons"><strong>Indicators:</strong><ul>' + reasons.map((line) => '<li>' + esc(line) + '</li>').join("") + '</ul></div>',
          '</article>',
        ].join("");
      }).join("");
    }
    async function loadSettings() {
      setStatus("Loading...", "");
      const response = await fetch(API);
      const data = await response.json();
      if (!response.ok) throw new Error(data.error || "Failed to load dashboard risk settings.");
      setForm(data.settings || DEFAULTS);
      setStatus("Loaded.", "ok");
    }
    async function saveSettings() {
      setStatus("Saving...", "");
      const response = await fetch(API, {
        method: "POST",
        headers: { "Content-Type": "application/json" },
        body: JSON.stringify(readForm()),
      });
      const data = await response.json();
      if (!response.ok) throw new Error(data.error || "Failed to save dashboard risk settings.");
      setForm(data.settings || DEFAULTS);
      setStatus("Saved.", "ok");
    }
    document.getElementById("reload-btn").addEventListener("click", () => loadSettings().catch((e) => setStatus(e.message || String(e), "err")));
    document.getElementById("reset-btn").addEventListener("click", () => { setForm(DEFAULTS); setStatus("Defaults loaded in form. Save to persist.", ""); });
    document.getElementById("save-btn").addEventListener("click", () => saveSettings().catch((e) => setStatus(e.message || String(e), "err")));
    Object.values(fields.indicator_points).forEach((el) => el.addEventListener("input", () => renderPreviews(readForm())));
    Object.values(fields.thresholds).forEach((el) => el.addEventListener("input", () => renderPreviews(readForm())));
    Object.values(fields.labels).forEach((el) => el.addEventListener("input", () => renderPreviews(readForm())));
    loadSettings().catch((e) => setStatus(e.message || String(e), "err"));
  </script>
  <script src="/shared-nav.js"></script>
</body>
</html>""".replace("__SETTINGS_TOP_NAV__", _settings_top_nav_html(DASHBOARD_RISK_SETTINGS_ROUTE))


def _report_entities_settings_html() -> str:
    return """<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Report Entity Registry</title>
  <link rel="stylesheet" href="/shared-nav.css">
  <link rel="stylesheet" href="/material-symbols.css">
  <style>
    :root { --bg:#eff6ff; --card:#ffffff; --ink:#0f172a; --muted:#475569; --line:#cbd5e1; --brand:#0f766e; --ok:#166534; --err:#b91c1c; --sug:#ccfbf1; }
    * { box-sizing:border-box; }
    body { margin:0; padding:20px; font-family:"Segoe UI",Tahoma,sans-serif; color:var(--ink); background:linear-gradient(180deg,#eff6ff,#f8fafc); }
    .wrap { max-width:1240px; margin:0 auto; }
    .card { background:var(--card); border:1px solid var(--line); border-radius:12px; padding:14px; box-shadow:0 8px 22px rgba(15,23,42,.06); }
    .top { display:flex; justify-content:space-between; align-items:flex-start; gap:10px; flex-wrap:wrap; }
    .btn { border:1px solid #0f766e; background:#0f766e; color:#fff; border-radius:8px; padding:8px 12px; cursor:pointer; text-decoration:none; }
    .btn.alt { background:#fff; color:#0f172a; border-color:var(--line); }
    .row { display:flex; gap:8px; flex-wrap:wrap; margin-top:10px; }
    .grid { display:grid; gap:10px; grid-template-columns:repeat(auto-fit,minmax(220px,1fr)); margin-top:12px; }
    label { display:block; font-size:.8rem; font-weight:700; margin-bottom:4px; }
    input, select, textarea { width:100%; border:1px solid var(--line); border-radius:8px; padding:8px; font-size:.9rem; }
    textarea { min-height:100px; font-family:Consolas,monospace; }
    .table-wrap { margin-top:12px; border:1px solid var(--line); border-radius:10px; overflow:auto; background:#fff; max-height:65vh; }
    table { border-collapse:collapse; min-width:1460px; width:100%; }
    thead th { position:sticky; top:0; background:#ecfeff; border-bottom:1px solid #99f6e4; font-size:.75rem; text-transform:uppercase; letter-spacing:.04em; color:#0f172a; }
    th, td { padding:9px 10px; border-bottom:1px solid #e2e8f0; text-align:left; vertical-align:top; font-size:.85rem; }
    tbody tr:nth-child(even) { background:#f8fafc; }
    .chip { display:inline-block; padding:2px 8px; border-radius:999px; background:#e0f2fe; border:1px solid #bae6fd; margin:1px; font-size:.72rem; }
    .mono { font-family:Consolas,monospace; font-size:.78rem; }
    .formula-preview { max-width:260px; white-space:nowrap; overflow:hidden; text-overflow:ellipsis; }
    #status { margin-top:8px; min-height:1.2em; font-size:.9rem; }
    #status.ok { color:var(--ok); } #status.err { color:var(--err); }
    dialog { width:min(920px,95vw); border:1px solid var(--line); border-radius:12px; padding:0; }
    .modal-head { padding:12px 14px; border-bottom:1px solid var(--line); background:#f0fdfa; font-weight:700; }
    .modal-body { padding:12px 14px; display:grid; gap:10px; }
    .modal-actions { padding:10px 14px; border-top:1px solid var(--line); display:flex; gap:8px; justify-content:flex-end; }
    .formula-toolbar { display:flex; gap:6px; flex-wrap:wrap; }
    .formula-chip { border:1px solid var(--line); background:#fff; color:#0f172a; border-radius:999px; padding:2px 9px; cursor:pointer; font-family:Consolas,monospace; font-size:.8rem; }
    .formula-editor-wrap { position:relative; }
    .formula-suggestions { position:absolute; left:0; right:0; top:100%; z-index:20; border:1px solid var(--line); background:#fff; border-radius:8px; box-shadow:0 8px 18px rgba(15,23,42,.12); max-height:220px; overflow:auto; margin-top:4px; }
    .formula-suggestions.hidden { display:none; }
    .formula-suggestion { display:flex; justify-content:space-between; gap:10px; padding:7px 10px; cursor:pointer; border-bottom:1px solid #f1f5f9; }
    .formula-suggestion:last-child { border-bottom:none; }
    .formula-suggestion.active { background:var(--sug); }
    .formula-kind { color:#0f766e; font-size:.75rem; }
    #formula-validation { min-height:1.2em; font-size:.85rem; }
    #formula-validation.ok { color:var(--ok); }
    #formula-validation.err { color:var(--err); }
    @media (max-width:760px){ body{padding:10px;} .card{padding:10px;} }
  </style>
</head>
<body>
  <main class="wrap">
    <section class="card">
      <div class="top">
        <div>
          <h1 style="margin:0;font-size:1.25rem;">Report Entity Registry</h1>
          <p style="margin:.4rem 0 0;color:var(--muted);font-size:.92rem;">Store and manage report entity definitions in a structured registry. Formulas are configuration-only in this phase.</p>
        </div>
        <div class="row">__SETTINGS_TOP_NAV__</div>
      </div>
      <div class="grid">
        <div><label for="planned-leave-n">Planned Leave Min Notice (days)</label><input id="planned-leave-n" type="number" min="0" step="1"></div>
        <div><label for="planned-leave-apply">Planned Leave Rule Apply From</label><input id="planned-leave-apply" type="date"></div>
        <div><label for="taken-mode">Leave Taken Identification Mode</label><select id="taken-mode"><option value="hours">hours</option><option value="status">status</option></select></div>
        <div><label for="taken-apply">Leave Taken Rule Apply From</label><input id="taken-apply" type="date"></div>
        <div><label for="rmi-resolution">RMI Planned Field Resolution</label><select id="rmi-resolution"><option value="name_lookup">name_lookup</option><option value="field_id">field_id</option><option value="hybrid">hybrid</option></select></div>
        <div><label for="planned-actual-tolerance">Planned vs Actual Equality Tolerance (hours)</label><input id="planned-actual-tolerance" type="number" min="0" step="0.1"></div>
      </div>
      <div class="row">
        <button class="btn alt" type="button" id="reload-btn">Reload</button>
        <button class="btn" type="button" id="save-btn">Save All</button>
        <button class="btn alt" type="button" id="reset-btn">Reset Defaults</button>
      </div>
      <div id="status"></div>
      <div class="table-wrap">
        <table>
          <thead><tr><th>Key</th><th>Label</th><th>Category</th><th>Identity Level</th><th>Output Type</th><th>Source Project</th><th>Formula</th><th>Jira Fields</th><th>Selection Rule</th><th>Completeness Rule</th><th>Updated At</th><th>Edit</th></tr></thead>
          <tbody id="entity-tbody"></tbody>
        </table>
      </div>
    </section>
  </main>
  <dialog id="edit-dialog">
    <div class="modal-head" id="edit-title">Edit Entity</div>
    <div class="modal-body">
      <div class="grid">
        <div><label for="e-label">Label</label><input id="e-label"></div>
        <div><label for="e-category">Category</label><input id="e-category"></div>
        <div><label for="e-identity">Identity Level</label><input id="e-identity"></div>
        <div><label for="e-output">Output Type</label><input id="e-output"></div>
        <div><label for="e-source-project">Source Project</label><input id="e-source-project"></div>
        <div><label for="e-formula-version">Formula Version</label><input id="e-formula-version" type="number" min="1" step="1"></div>
      </div>
      <div>
        <label for="e-formula-expression">Formula Expression</label>
        <div class="formula-toolbar" id="formula-quick-insert">
          <button class="formula-chip" type="button" data-token="+">+</button>
          <button class="formula-chip" type="button" data-token="-">-</button>
          <button class="formula-chip" type="button" data-token="*">*</button>
          <button class="formula-chip" type="button" data-token="/">/</button>
          <button class="formula-chip" type="button" data-token="(">(</button>
          <button class="formula-chip" type="button" data-token=")">)</button>
          <button class="formula-chip" type="button" data-token="sum()">sum()</button>
          <button class="formula-chip" type="button" data-token="count()">count()</button>
          <button class="formula-chip" type="button" data-token="min()">min()</button>
          <button class="formula-chip" type="button" data-token="max()">max()</button>
          <button class="formula-chip" type="button" data-token="average()">average()</button>
        </div>
        <div class="formula-editor-wrap">
          <textarea id="e-formula-expression" spellcheck="false"></textarea>
          <div id="formula-suggestions" class="formula-suggestions hidden"></div>
        </div>
        <div id="formula-validation"></div>
      </div>
      <div><label for="e-formula-meta">Formula Meta JSON</label><textarea id="e-formula-meta"></textarea></div>
      <div><label for="e-definition">Definition</label><textarea id="e-definition"></textarea></div>
      <div><label for="e-issue-types">Source Issue Types JSON</label><textarea id="e-issue-types"></textarea></div>
      <div><label for="e-jira-fields">Jira Fields JSON</label><textarea id="e-jira-fields"></textarea></div>
      <div><label for="e-selection-rule">Selection Rule JSON</label><textarea id="e-selection-rule"></textarea></div>
      <div><label for="e-completeness-rule">Completeness Rule JSON</label><textarea id="e-completeness-rule"></textarea></div>
      <div><label for="e-notes">Admin Notes</label><textarea id="e-notes"></textarea></div>
      <div><label><input type="checkbox" id="e-active" checked> Active</label></div>
    </div>
    <div class="modal-actions">
      <button class="btn alt" type="button" id="cancel-edit">Cancel</button>
      <button class="btn" type="button" id="save-edit">Apply Row Changes</button>
    </div>
  </dialog>
  <script>
    const API = "/api/report-entities";
    const RESET_API = "/api/report-entities/reset";
    const FORMULA_FUNCTIONS = ["sum", "count", "min", "max", "average"];
    const statusEl = document.getElementById("status");
    const tbodyEl = document.getElementById("entity-tbody");
    const dialogEl = document.getElementById("edit-dialog");
    const titleEl = document.getElementById("edit-title");
    const formulaSuggestionsEl = document.getElementById("formula-suggestions");
    const formulaValidationEl = document.getElementById("formula-validation");
    const formEl = {
      plannedLeaveN: document.getElementById("planned-leave-n"), plannedLeaveApply: document.getElementById("planned-leave-apply"),
      takenMode: document.getElementById("taken-mode"), takenApply: document.getElementById("taken-apply"), rmiResolution: document.getElementById("rmi-resolution"),
      plannedActualTolerance: document.getElementById("planned-actual-tolerance"),
      label: document.getElementById("e-label"), category: document.getElementById("e-category"), identity: document.getElementById("e-identity"),
      output: document.getElementById("e-output"), sourceProject: document.getElementById("e-source-project"), definition: document.getElementById("e-definition"),
      issueTypes: document.getElementById("e-issue-types"), jiraFields: document.getElementById("e-jira-fields"), selectionRule: document.getElementById("e-selection-rule"),
      completenessRule: document.getElementById("e-completeness-rule"), notes: document.getElementById("e-notes"), active: document.getElementById("e-active"),
      formulaExpression: document.getElementById("e-formula-expression"), formulaVersion: document.getElementById("e-formula-version"), formulaMeta: document.getElementById("e-formula-meta")
    };
    let entities = [];
    let selectedKey = "";
    let suggestionItems = [];
    let suggestionIndex = -1;
    function setStatus(msg, kind) { statusEl.textContent = String(msg || ""); statusEl.className = kind || ""; }
    function esc(v) { return String(v == null ? "" : v).replace(/&/g,"&amp;").replace(/</g,"&lt;").replace(/>/g,"&gt;"); }
    function compact(v) { return esc(JSON.stringify(v || {}).slice(0, 140)); }
    function chips(v) { const arr = Array.isArray(v) ? v : []; return arr.map((x) => '<span class="chip mono">' + esc(x) + '</span>').join("") || '<span class="mono">[]</span>'; }
    function asFormulaPreview(v) { const t = String(v || "").trim(); return t ? esc(t) : '<span class="mono">-</span>'; }
    function renderRows() {
      tbodyEl.innerHTML = entities.map((e) => '<tr>' +
        '<td class="mono">' + esc(e.entity_key) + '</td><td>' + esc(e.label) + '</td><td>' + esc(e.category) + '</td>' +
        '<td>' + esc(e.identity_level) + '</td><td class="mono">' + esc(e.output_type) + '</td><td>' + esc(e.source_project_key || "-") + '</td>' +
        '<td class="mono formula-preview">' + asFormulaPreview(e.formula_expression) + '</td><td>' + chips(e.jira_fields_json) + '</td><td class="mono">' + compact(e.selection_rule_json) + '</td>' +
        '<td class="mono">' + compact(e.completeness_rule_json) + '</td><td class="mono">' + esc(e.updated_at_utc || "-") + '</td>' +
        '<td><button class="btn alt" type="button" data-edit="' + esc(e.entity_key) + '">Edit</button></td></tr>').join("");
      Array.from(tbodyEl.querySelectorAll("button[data-edit]")).forEach((btn) => btn.addEventListener("click", () => openEdit(String(btn.getAttribute("data-edit") || ""))));
    }
    function toPretty(v) { return JSON.stringify(v == null ? {} : v, null, 2); }
    function parseJson(name, text) { try { return JSON.parse(String(text || "").trim() || "{}"); } catch (_err) { throw new Error("Invalid JSON in " + name); } }
    function parseJsonObject(name, text) {
      const out = parseJson(name, text);
      if (out == null || Array.isArray(out) || typeof out !== "object") throw new Error(name + " must be a JSON object.");
      return out;
    }
    function entityKeysSet() {
      return new Set(entities.map((e) => String(e.entity_key || "").toLowerCase()).filter(Boolean));
    }
    function tokenizeFormula(text) {
      const source = String(text || "");
      const tokens = [];
      let i = 0;
      while (i < source.length) {
        const ch = source[i];
        if (/\\s/.test(ch)) { i += 1; continue; }
        if ("+-*/".includes(ch)) { tokens.push({ t: "op", v: ch, p: i }); i += 1; continue; }
        if (ch === "(") { tokens.push({ t: "lparen", v: ch, p: i }); i += 1; continue; }
        if (ch === ")") { tokens.push({ t: "rparen", v: ch, p: i }); i += 1; continue; }
        if (ch === ",") { tokens.push({ t: "comma", v: ch, p: i }); i += 1; continue; }
        if (/[A-Za-z_]/.test(ch)) {
          const start = i; i += 1;
          while (i < source.length && /[A-Za-z0-9_]/.test(source[i])) i += 1;
          tokens.push({ t: "ident", v: source.slice(start, i), p: start });
          continue;
        }
        throw new Error("Invalid character '" + ch + "' at position " + String(i + 1) + ".");
      }
      tokens.push({ t: "eof", v: "", p: source.length });
      return tokens;
    }
    function validateFormulaClient(formula, currentKey) {
      const text = String(formula || "").trim();
      if (!text) return { ok: true, message: "Formula is empty (allowed)." };
      const known = entityKeysSet();
      const tokens = tokenizeFormula(text);
      let idx = 0;
      function peek() { return tokens[idx]; }
      function consume(expected) {
        const tk = tokens[idx];
        if (expected && tk.t !== expected) throw new Error("Expected " + expected + " at position " + String(tk.p + 1) + ".");
        idx += 1;
        return tk;
      }
      function parseExpr() { parseTerm(); while (peek().t === "op" && (peek().v === "+" || peek().v === "-")) { consume("op"); parseTerm(); } }
      function parseTerm() { parseFactor(); while (peek().t === "op" && (peek().v === "*" || peek().v === "/")) { consume("op"); parseFactor(); } }
      function parseFactor() {
        const tk = peek();
        if (tk.t === "ident") {
          const ident = consume("ident");
          const identKey = String(ident.v || "").toLowerCase();
          if (peek().t === "lparen") {
            if (!FORMULA_FUNCTIONS.includes(identKey)) throw new Error("Unknown function '" + ident.v + "' at position " + String(ident.p + 1) + ".");
            consume("lparen"); parseExpr();
            if (peek().t === "comma") throw new Error("Function '" + ident.v + "' accepts one argument at position " + String(peek().p + 1) + ".");
            consume("rparen");
            return;
          }
          if (!known.has(identKey)) throw new Error("Unknown entity '" + ident.v + "' at position " + String(ident.p + 1) + ".");
          if (currentKey && identKey === String(currentKey).toLowerCase()) throw new Error("Self reference is not allowed.");
          return;
        }
        if (tk.t === "lparen") { consume("lparen"); parseExpr(); consume("rparen"); return; }
        throw new Error("Unexpected token at position " + String(tk.p + 1) + ".");
      }
      parseExpr();
      if (peek().t !== "eof") throw new Error("Unexpected token at position " + String(peek().p + 1) + ".");
      return { ok: true, message: "Formula syntax looks valid." };
    }
    function setFormulaValidation(ok, msg) {
      formulaValidationEl.textContent = String(msg || "");
      formulaValidationEl.className = ok ? "ok" : "err";
    }
    function currentFormulaPrefix() {
      const input = formEl.formulaExpression;
      const cursor = Number(input.selectionStart || 0);
      const left = String(input.value || "").slice(0, cursor);
      const m = left.match(/[A-Za-z_][A-Za-z0-9_]*$/);
      return { cursor, prefix: m ? m[0] : "", prefixStart: m ? cursor - m[0].length : cursor };
    }
    function formulaContext() {
      const input = formEl.formulaExpression;
      const cursor = Number(input.selectionStart || 0);
      const left = String(input.value || "").slice(0, cursor);
      let i = left.length - 1;
      while (i >= 0 && /\\s/.test(left[i])) i -= 1;
      const prev = i >= 0 ? left[i] : "";
      return { expectsOperand: (!prev || "+-*/(,".includes(prev)), prev };
    }
    function scoreMatch(text, query) {
      const t = String(text || "").toLowerCase();
      const q = String(query || "").toLowerCase();
      if (!q) return 1;
      if (t.startsWith(q)) return 4;
      if (t.includes(q)) return 2;
      return 0;
    }
    function buildSuggestions() {
      const ctx = formulaContext();
      const info = currentFormulaPrefix();
      const q = info.prefix.toLowerCase();
      const items = [];
      if (ctx.expectsOperand) {
        FORMULA_FUNCTIONS.forEach((name) => {
          const score = scoreMatch(name, q); if (!score) return;
          items.push({ label: name + "()", kind: "function", insert: name + "()", score: score + 2 });
        });
        entities.forEach((e) => {
          const key = String(e.entity_key || "");
          const label = String(e.label || "");
          const score = Math.max(scoreMatch(key, q), scoreMatch(label, q));
          if (!score) return;
          items.push({ label: key + " - " + label, kind: "entity", insert: key, score: score + 1 });
        });
        if ("(".startsWith(q) || !q) items.push({ label: "(", kind: "group", insert: "(", score: 1 });
      } else {
        ["+", "-", "*", "/", ")"].forEach((op) => {
          if (op.startsWith(q) || !q) items.push({ label: op, kind: "operator", insert: op, score: 1 });
        });
      }
      items.sort((a, b) => b.score - a.score || a.label.localeCompare(b.label));
      return items.slice(0, 20);
    }
    function renderSuggestions(items) {
      suggestionItems = items || [];
      suggestionIndex = suggestionItems.length ? 0 : -1;
      if (!suggestionItems.length) {
        formulaSuggestionsEl.classList.add("hidden");
        formulaSuggestionsEl.innerHTML = "";
        return;
      }
      formulaSuggestionsEl.innerHTML = suggestionItems.map((s, idx) =>
        '<div class="formula-suggestion ' + (idx === suggestionIndex ? "active" : "") + '" data-idx="' + String(idx) + '">' +
        '<span>' + esc(s.label) + '</span><span class="formula-kind">' + esc(s.kind) + '</span></div>').join("");
      formulaSuggestionsEl.classList.remove("hidden");
      Array.from(formulaSuggestionsEl.querySelectorAll(".formula-suggestion")).forEach((node) => {
        node.addEventListener("mousedown", (e) => {
          e.preventDefault();
          const idx = Number(node.getAttribute("data-idx") || -1);
          if (idx >= 0) applySuggestion(suggestionItems[idx]);
        });
      });
    }
    function refreshSuggestionHighlight() {
      Array.from(formulaSuggestionsEl.querySelectorAll(".formula-suggestion")).forEach((node) => {
        const idx = Number(node.getAttribute("data-idx") || -1);
        node.classList.toggle("active", idx === suggestionIndex);
      });
    }
    function applySuggestion(item) {
      if (!item) return;
      const input = formEl.formulaExpression;
      const info = currentFormulaPrefix();
      const before = String(input.value || "").slice(0, info.prefixStart);
      const after = String(input.value || "").slice(Number(input.selectionStart || 0));
      const insert = String(item.insert || "");
      input.value = before + insert + after;
      const caretBase = before.length + insert.length;
      const finalCaret = insert.endsWith("()") ? caretBase - 1 : caretBase;
      input.focus();
      input.setSelectionRange(finalCaret, finalCaret);
      renderSuggestions([]);
      onFormulaInput();
    }
    function onFormulaInput() {
      const raw = String(formEl.formulaExpression.value || "");
      try {
        const out = validateFormulaClient(raw, selectedKey);
        setFormulaValidation(true, out.message);
      } catch (err) {
        setFormulaValidation(false, err.message || String(err));
      }
      renderSuggestions(buildSuggestions());
    }
    function insertToken(token) {
      const input = formEl.formulaExpression;
      const start = Number(input.selectionStart || 0);
      const end = Number(input.selectionEnd || start);
      const value = String(input.value || "");
      const before = value.slice(0, start);
      const after = value.slice(end);
      const t = String(token || "");
      input.value = before + t + after;
      const caret = before.length + t.length - (t.endsWith("()") ? 1 : 0);
      input.focus();
      input.setSelectionRange(caret, caret);
      onFormulaInput();
    }
    function openEdit(entityKey) {
      const row = entities.find((x) => String(x.entity_key) === String(entityKey)); if (!row) return;
      selectedKey = row.entity_key; titleEl.textContent = "Edit: " + row.label + " (" + row.entity_key + ")";
      formEl.label.value = row.label || ""; formEl.category.value = row.category || ""; formEl.identity.value = row.identity_level || "";
      formEl.output.value = row.output_type || ""; formEl.sourceProject.value = row.source_project_key || ""; formEl.definition.value = row.definition_text || "";
      formEl.issueTypes.value = toPretty(row.source_issue_types_json || []); formEl.jiraFields.value = toPretty(row.jira_fields_json || []);
      formEl.selectionRule.value = toPretty(row.selection_rule_json || {}); formEl.completenessRule.value = toPretty(row.completeness_rule_json || {});
      formEl.formulaExpression.value = String(row.formula_expression || "");
      formEl.formulaVersion.value = String(Number(row.formula_version || 1));
      formEl.formulaMeta.value = toPretty(row.formula_meta_json || {});
      formEl.notes.value = row.admin_notes || ""; formEl.active.checked = !!row.is_active; dialogEl.showModal();
      onFormulaInput();
    }
    function applyEdit() {
      const idx = entities.findIndex((x) => String(x.entity_key) === String(selectedKey)); if (idx < 0) return;
      const row = entities[idx];
      const formulaExpression = String(formEl.formulaExpression.value || "").trim();
      const formulaVersion = Number(formEl.formulaVersion.value || 1);
      const formulaMeta = parseJsonObject("Formula Meta JSON", formEl.formulaMeta.value);
      validateFormulaClient(formulaExpression, selectedKey);
      if (!Number.isFinite(formulaVersion) || formulaVersion < 1) throw new Error("Formula Version must be an integer >= 1.");
      row.label = String(formEl.label.value || "").trim(); row.category = String(formEl.category.value || "").trim(); row.identity_level = String(formEl.identity.value || "").trim();
      row.output_type = String(formEl.output.value || "").trim(); row.source_project_key = String(formEl.sourceProject.value || "").trim(); row.definition_text = String(formEl.definition.value || "").trim();
      row.source_issue_types_json = parseJson("Source Issue Types JSON", formEl.issueTypes.value); row.jira_fields_json = parseJson("Jira Fields JSON", formEl.jiraFields.value);
      row.selection_rule_json = parseJson("Selection Rule JSON", formEl.selectionRule.value); row.completeness_rule_json = parseJson("Completeness Rule JSON", formEl.completenessRule.value);
      row.formula_expression = formulaExpression; row.formula_version = Math.floor(formulaVersion); row.formula_meta_json = formulaMeta;
      row.admin_notes = String(formEl.notes.value || "").trim(); row.is_active = !!formEl.active.checked; entities[idx] = row;
      renderRows(); dialogEl.close(); setStatus("Row updated in editor. Click Save All to persist.", "");
    }
    function readGlobalSettings() {
      return {
        planned_leave_min_notice_days: Number(formEl.plannedLeaveN.value || 0),
        planned_leave_rule_apply_from_date: String(formEl.plannedLeaveApply.value || "").trim(),
        leave_taken_identification_mode: String(formEl.takenMode.value || "").trim(),
        leave_taken_rule_apply_from_date: String(formEl.takenApply.value || "").trim(),
        rmi_planned_field_resolution: String(formEl.rmiResolution.value || "").trim(),
        planned_actual_equality_tolerance_hours: Number(formEl.plannedActualTolerance.value || 0)
      };
    }
    function setGlobalSettings(s) {
      const gs = s || {};
      formEl.plannedLeaveN.value = String(Number(gs.planned_leave_min_notice_days || 0));
      formEl.plannedLeaveApply.value = String(gs.planned_leave_rule_apply_from_date || "");
      formEl.takenMode.value = String(gs.leave_taken_identification_mode || "hours");
      formEl.takenApply.value = String(gs.leave_taken_rule_apply_from_date || "");
      formEl.rmiResolution.value = String(gs.rmi_planned_field_resolution || "name_lookup");
      formEl.plannedActualTolerance.value = String(Number(gs.planned_actual_equality_tolerance_hours || 0));
    }
    async function reloadAll() {
      const resp = await fetch(API); const data = await resp.json().catch(() => ({})); if (!resp.ok) throw new Error(data.error || "Failed to load entities.");
      entities = Array.isArray(data.entities) ? data.entities : []; setGlobalSettings(data.global_settings || {}); renderRows(); setStatus("Loaded " + entities.length + " entities.", "ok");
    }
    async function saveAll() {
      const payload = { entities, global_settings: readGlobalSettings() };
      const resp = await fetch(API, { method:"PUT", headers:{ "Content-Type":"application/json" }, body:JSON.stringify(payload) });
      const data = await resp.json().catch(() => ({})); if (!resp.ok) throw new Error(data.error || "Failed to save entities.");
      entities = Array.isArray(data.entities) ? data.entities : []; setGlobalSettings(data.global_settings || {}); renderRows(); setStatus("Saved.", "ok");
    }
    async function resetDefaults() {
      if (!window.confirm("Reset all registry definitions to defaults?")) return;
      const resp = await fetch(RESET_API, { method: "POST" }); const data = await resp.json().catch(() => ({})); if (!resp.ok) throw new Error(data.error || "Reset failed.");
      entities = Array.isArray(data.entities) ? data.entities : []; setGlobalSettings(data.global_settings || {}); renderRows(); setStatus("Reset to defaults.", "ok");
    }
    formEl.formulaExpression.addEventListener("input", onFormulaInput);
    formEl.formulaExpression.addEventListener("keydown", (e) => {
      if (formulaSuggestionsEl.classList.contains("hidden")) return;
      if (!suggestionItems.length) return;
      if (e.key === "ArrowDown") { e.preventDefault(); suggestionIndex = (suggestionIndex + 1) % suggestionItems.length; refreshSuggestionHighlight(); }
      else if (e.key === "ArrowUp") { e.preventDefault(); suggestionIndex = (suggestionIndex - 1 + suggestionItems.length) % suggestionItems.length; refreshSuggestionHighlight(); }
      else if (e.key === "Enter") { e.preventDefault(); if (suggestionIndex >= 0) applySuggestion(suggestionItems[suggestionIndex]); }
      else if (e.key === "Escape") { e.preventDefault(); renderSuggestions([]); }
    });
    formEl.formulaExpression.addEventListener("blur", () => setTimeout(() => renderSuggestions([]), 120));
    document.getElementById("formula-quick-insert").addEventListener("click", (e) => {
      const token = e.target && e.target.getAttribute ? e.target.getAttribute("data-token") : "";
      if (!token) return;
      insertToken(token);
    });
    document.getElementById("reload-btn").addEventListener("click", () => reloadAll().catch((e) => setStatus(e.message || String(e), "err")));
    document.getElementById("save-btn").addEventListener("click", () => saveAll().catch((e) => setStatus(e.message || String(e), "err")));
    document.getElementById("reset-btn").addEventListener("click", () => resetDefaults().catch((e) => setStatus(e.message || String(e), "err")));
    document.getElementById("cancel-edit").addEventListener("click", () => dialogEl.close());
    document.getElementById("save-edit").addEventListener("click", () => { try { applyEdit(); } catch (e) { setStatus(e.message || String(e), "err"); }});
    reloadAll().catch((e) => setStatus(e.message || String(e), "err"));
  </script>
  <script src="/shared-nav.js"></script>
</body>
</html>""".replace("__SETTINGS_TOP_NAV__", _settings_top_nav_html(REPORT_ENTITIES_SETTINGS_ROUTE))


def _manage_fields_settings_html() -> str:
    return """<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Manage Fields</title>
  <link rel="stylesheet" href="/shared-nav.css">
  <link rel="stylesheet" href="/material-symbols.css">
  <style>
    :root { --bg:#f4f8ff; --card:#fff; --ink:#0f172a; --line:#cbd5e1; --muted:#475569; --brand:#1d4ed8; --ok:#166534; --err:#b91c1c; --sug:#dbeafe; }
    * { box-sizing:border-box; }
    body { margin:0; padding:20px; background:linear-gradient(180deg,#eff6ff,#f8fafc); color:var(--ink); font-family:"Segoe UI",Tahoma,sans-serif; }
    .wrap { max-width:1240px; margin:0 auto; }
    .card { background:var(--card); border:1px solid var(--line); border-radius:12px; padding:14px; box-shadow:0 8px 22px rgba(15,23,42,.06); }
    .top { display:flex; justify-content:space-between; align-items:flex-start; gap:10px; flex-wrap:wrap; }
    .btn { border:1px solid #1e40af; background:#1d4ed8; color:#fff; border-radius:8px; padding:8px 12px; cursor:pointer; text-decoration:none; }
    .btn.alt { border-color:var(--line); background:#fff; color:#0f172a; }
    .btn.warn { border-color:#7f1d1d; background:#b91c1c; color:#fff; }
    .row { display:flex; gap:8px; flex-wrap:wrap; margin-top:10px; align-items:center; }
    .table-wrap { margin-top:12px; border:1px solid var(--line); border-radius:10px; overflow:auto; background:#fff; max-height:65vh; }
    table { border-collapse:collapse; min-width:1200px; width:100%; }
    th, td { padding:9px 10px; border-bottom:1px solid #e2e8f0; text-align:left; vertical-align:top; font-size:.85rem; }
    thead th { position:sticky; top:0; background:#eff6ff; text-transform:uppercase; letter-spacing:.04em; font-size:.75rem; }
    tbody tr:nth-child(even){ background:#f8fafc; }
    .mono { font-family:Consolas,monospace; font-size:.78rem; }
    .formula-preview { max-width:320px; white-space:nowrap; overflow:hidden; text-overflow:ellipsis; }
    #mf-status { margin-top:8px; min-height:1.2em; font-size:.9rem; }
    #mf-status.ok { color:var(--ok); } #mf-status.err { color:var(--err); }
    dialog { width:min(940px,96vw); border:1px solid var(--line); border-radius:12px; padding:0; }
    .modal-head { padding:12px 14px; border-bottom:1px solid var(--line); background:#eff6ff; font-weight:700; }
    .modal-body { padding:12px 14px; display:grid; gap:10px; }
    .modal-actions { padding:10px 14px; border-top:1px solid var(--line); display:flex; gap:8px; justify-content:flex-end; }
    .grid { display:grid; gap:10px; grid-template-columns:repeat(auto-fit,minmax(220px,1fr)); }
    label { display:block; font-size:.8rem; font-weight:700; margin-bottom:4px; }
    input, select, textarea { width:100%; border:1px solid var(--line); border-radius:8px; padding:8px; font-size:.9rem; }
    textarea { min-height:96px; font-family:Consolas,monospace; }
    .formula-toolbar { display:flex; gap:6px; flex-wrap:wrap; }
    .formula-chip { border:1px solid var(--line); background:#fff; border-radius:999px; padding:2px 9px; cursor:pointer; font-family:Consolas,monospace; font-size:.8rem; }
    .formula-editor-wrap { position:relative; }
    .formula-suggestions { position:absolute; left:0; right:0; top:100%; z-index:20; border:1px solid var(--line); background:#fff; border-radius:8px; box-shadow:0 8px 18px rgba(15,23,42,.12); max-height:220px; overflow:auto; margin-top:4px; }
    .formula-suggestions.hidden { display:none; }
    .formula-suggestion { display:flex; justify-content:space-between; gap:10px; padding:7px 10px; cursor:pointer; border-bottom:1px solid #f1f5f9; }
    .formula-suggestion.active { background:var(--sug); }
    .formula-kind { color:#1d4ed8; font-size:.75rem; }
    #mf-formula-validation { min-height:1.2em; font-size:.85rem; }
    #mf-formula-validation.ok { color:var(--ok); }
    #mf-formula-validation.err { color:var(--err); }
  </style>
</head>
<body>
  <main class="wrap">
    <section class="card">
      <div class="top">
        <div>
          <h1 style="margin:0;font-size:1.25rem;">Manage Fields</h1>
          <p style="margin:.4rem 0 0;color:var(--muted);font-size:.92rem;">Create and maintain managed fields with formulas referencing report entities.</p>
        </div>
        <div class="row">__SETTINGS_TOP_NAV__</div>
      </div>
      <div class="row">
        <button class="btn alt" type="button" id="mf-reload-btn">Reload</button>
        <button class="btn" type="button" id="mf-create-btn">Create Field</button>
        <label style="display:inline-flex;gap:6px;align-items:center;font-weight:600;font-size:.86rem;">
          <input id="mf-include-inactive" type="checkbox"> Show Inactive
        </label>
      </div>
      <div id="mf-status"></div>
      <div class="table-wrap">
        <table>
          <thead><tr><th>Key</th><th>Label</th><th>Type</th><th>Formula</th><th>Active</th><th>Updated At</th><th>Actions</th></tr></thead>
          <tbody id="mf-tbody"></tbody>
        </table>
      </div>
    </section>
  </main>

  <dialog id="mf-dialog">
    <div class="modal-head" id="mf-dialog-title">Create Field</div>
    <div class="modal-body">
      <div class="grid">
        <div>
          <label for="mf-field-key">Field Key (System Managed)</label>
          <input id="mf-field-key" class="mono" readonly>
          <div style="color:#64748b;font-size:.78rem;margin-top:4px;">Auto-generated from Label. Manual edits are disabled.</div>
        </div>
        <div><label for="mf-label">Label</label><input id="mf-label"></div>
        <div><label for="mf-data-type">Data Type</label><select id="mf-data-type"><option value="number">number</option><option value="text">text</option><option value="date">date</option><option value="boolean">boolean</option></select></div>
        <div><label for="mf-formula-version">Formula Version</label><input id="mf-formula-version" type="number" min="1" step="1"></div>
      </div>
      <div><label for="mf-description">Description</label><textarea id="mf-description"></textarea></div>
      <div>
        <label for="mf-formula-expression">Formula Expression</label>
        <div class="formula-toolbar" id="mf-formula-quick-insert">
          <button class="formula-chip" type="button" data-token="+">+</button>
          <button class="formula-chip" type="button" data-token="-">-</button>
          <button class="formula-chip" type="button" data-token="*">*</button>
          <button class="formula-chip" type="button" data-token="/">/</button>
          <button class="formula-chip" type="button" data-token="(">(</button>
          <button class="formula-chip" type="button" data-token=")">)</button>
          <button class="formula-chip" type="button" data-token="sum()">sum()</button>
          <button class="formula-chip" type="button" data-token="count()">count()</button>
          <button class="formula-chip" type="button" data-token="min()">min()</button>
          <button class="formula-chip" type="button" data-token="max()">max()</button>
          <button class="formula-chip" type="button" data-token="average()">average()</button>
        </div>
        <div class="formula-editor-wrap">
          <textarea id="mf-formula-expression" spellcheck="false"></textarea>
          <div id="mf-formula-suggestions" class="formula-suggestions hidden"></div>
        </div>
        <div id="mf-formula-validation"></div>
      </div>
      <div><label for="mf-formula-meta">Formula Meta JSON</label><textarea id="mf-formula-meta"></textarea></div>
      <div><label><input type="checkbox" id="mf-is-active" checked> Active</label></div>
    </div>
    <div class="modal-actions">
      <button class="btn alt" type="button" id="mf-cancel">Cancel</button>
      <button class="btn" type="button" id="mf-save">Save</button>
    </div>
  </dialog>

  <script>
    const API = "/api/manage-fields";
    const FORMULA_FUNCTIONS = ["sum", "count", "min", "max", "average"];
    const statusEl = document.getElementById("mf-status");
    const tbodyEl = document.getElementById("mf-tbody");
    const includeInactiveEl = document.getElementById("mf-include-inactive");
    const dialogEl = document.getElementById("mf-dialog");
    const dialogTitleEl = document.getElementById("mf-dialog-title");
    const suggestionsEl = document.getElementById("mf-formula-suggestions");
    const validationEl = document.getElementById("mf-formula-validation");
    const formEl = {
      fieldKey: document.getElementById("mf-field-key"),
      label: document.getElementById("mf-label"),
      description: document.getElementById("mf-description"),
      dataType: document.getElementById("mf-data-type"),
      formulaExpression: document.getElementById("mf-formula-expression"),
      formulaVersion: document.getElementById("mf-formula-version"),
      formulaMeta: document.getElementById("mf-formula-meta"),
      isActive: document.getElementById("mf-is-active")
    };
    let rows = [];
    let entityCatalog = [];
    let mode = "create";
    let selectedKey = "";
    let suggestionItems = [];
    let suggestionIndex = -1;

    function setStatus(msg, kind) { statusEl.textContent = String(msg || ""); statusEl.className = kind || ""; }
    function esc(v) { return String(v == null ? "" : v).replace(/&/g,"&amp;").replace(/</g,"&lt;").replace(/>/g,"&gt;"); }
    function toPretty(v) { return JSON.stringify(v == null ? {} : v, null, 2); }
    function parseJson(name, text) { try { return JSON.parse(String(text || "").trim() || "{}"); } catch (_err) { throw new Error("Invalid JSON in " + name); } }
    function parseJsonObject(name, text) { const out = parseJson(name, text); if (!out || Array.isArray(out) || typeof out !== "object") throw new Error(name + " must be a JSON object."); return out; }
    function slugifyFieldKeyFromLabel(label) {
      const base = String(label || "").toLowerCase()
        .replace(/[^a-z0-9]+/g, "_")
        .replace(/_+/g, "_")
        .replace(/^_+|_+$/g, "");
      return base || "field";
    }
    function existingFieldKeysExceptSelected() {
      return new Set(
        rows
          .filter((x) => String(x.field_key || "") !== String(selectedKey || ""))
          .map((x) => String(x.field_key || "").toLowerCase())
          .filter(Boolean)
      );
    }
    function buildUniqueFieldKey(baseKey) {
      const cleanBase = slugifyFieldKeyFromLabel(baseKey);
      const used = existingFieldKeysExceptSelected();
      let candidate = cleanBase;
      let n = 2;
      while (used.has(candidate.toLowerCase())) {
        candidate = cleanBase + "_" + String(n);
        n += 1;
      }
      return candidate;
    }
    function refreshSystemFieldKey() {
      if (mode !== "create") return;
      formEl.fieldKey.value = buildUniqueFieldKey(formEl.label.value);
    }

    function renderRows() {
      tbodyEl.innerHTML = rows.map((r) => '<tr>'
        + '<td class="mono">' + esc(r.field_key) + '</td>'
        + '<td>' + esc(r.label) + '</td>'
        + '<td class="mono">' + esc(r.data_type) + '</td>'
        + '<td class="mono formula-preview">' + (String(r.formula_expression || "").trim() ? esc(r.formula_expression) : '<span class="mono">-</span>') + '</td>'
        + '<td>' + (r.is_active ? "Yes" : "No") + '</td>'
        + '<td class="mono">' + esc(r.updated_at_utc || "-") + '</td>'
        + '<td>'
        + '<button class="btn alt" type="button" data-edit="' + esc(r.field_key) + '">Edit</button> '
        + (r.is_active
          ? '<button class="btn warn" type="button" data-delete="' + esc(r.field_key) + '">Delete</button>'
          : '<button class="btn alt" type="button" data-restore="' + esc(r.field_key) + '">Restore</button>')
        + '</td>'
        + '</tr>').join("");

      Array.from(tbodyEl.querySelectorAll("button[data-edit]")).forEach((btn) => btn.addEventListener("click", () => openEdit(String(btn.getAttribute("data-edit") || ""))));
      Array.from(tbodyEl.querySelectorAll("button[data-delete]")).forEach((btn) => btn.addEventListener("click", () => softDelete(String(btn.getAttribute("data-delete") || ""))));
      Array.from(tbodyEl.querySelectorAll("button[data-restore]")).forEach((btn) => btn.addEventListener("click", () => restore(String(btn.getAttribute("data-restore") || ""))));
    }

    function tokenizeFormula(text) {
      const source = String(text || "");
      const out = []; let i = 0;
      while (i < source.length) {
        const ch = source[i];
        if (/\\s/.test(ch)) { i += 1; continue; }
        if ("+-*/".includes(ch)) { out.push({ t: "op", v: ch, p: i }); i += 1; continue; }
        if (ch === "(") { out.push({ t: "lparen", v: ch, p: i }); i += 1; continue; }
        if (ch === ")") { out.push({ t: "rparen", v: ch, p: i }); i += 1; continue; }
        if (ch === ",") { out.push({ t: "comma", v: ch, p: i }); i += 1; continue; }
        if (/[A-Za-z_]/.test(ch)) {
          const start = i; i += 1;
          while (i < source.length && /[A-Za-z0-9_]/.test(source[i])) i += 1;
          out.push({ t: "ident", v: source.slice(start, i), p: start });
          continue;
        }
        throw new Error("Invalid character '" + ch + "' at position " + String(i + 1) + ".");
      }
      out.push({ t: "eof", v: "", p: source.length });
      return out;
    }

    function validateFormulaClient(formula) {
      const text = String(formula || "").trim();
      if (!text) return { ok: true, message: "Formula is empty (allowed).", references: [] };
      const known = new Set(entityCatalog.map((x) => String(x.entity_key || "").toLowerCase()).filter(Boolean));
      const tokens = tokenizeFormula(text); let idx = 0;
      const refs = new Set();
      function peek() { return tokens[idx]; }
      function consume(expected) { const tk = tokens[idx]; if (expected && tk.t !== expected) throw new Error("Expected " + expected + " at position " + String(tk.p + 1) + "."); idx += 1; return tk; }
      function expr() { term(); while (peek().t === "op" && (peek().v === "+" || peek().v === "-")) { consume("op"); term(); } }
      function term() { factor(); while (peek().t === "op" && (peek().v === "*" || peek().v === "/")) { consume("op"); factor(); } }
      function factor() {
        const tk = peek();
        if (tk.t === "ident") {
          const ident = consume("ident"); const key = String(ident.v || "").toLowerCase();
          if (peek().t === "lparen") {
            if (!FORMULA_FUNCTIONS.includes(key)) throw new Error("Unknown function '" + ident.v + "' at position " + String(ident.p + 1) + ".");
            consume("lparen"); expr(); if (peek().t === "comma") throw new Error("Function '" + ident.v + "' accepts one argument at position " + String(peek().p + 1) + "."); consume("rparen");
            return;
          }
          if (!known.has(key)) throw new Error("Unknown entity '" + ident.v + "' at position " + String(ident.p + 1) + ".");
          refs.add(key);
          return;
        }
        if (tk.t === "lparen") { consume("lparen"); expr(); consume("rparen"); return; }
        throw new Error("Unexpected token at position " + String(tk.p + 1) + ".");
      }
      expr();
      if (peek().t !== "eof") throw new Error("Unexpected token at position " + String(peek().p + 1) + ".");
      return { ok: true, message: "Formula syntax looks valid.", references: Array.from(refs).sort() };
    }
    function setValidation(ok, msg) { validationEl.textContent = String(msg || ""); validationEl.className = ok ? "ok" : "err"; }
    function currentPrefix() { const input = formEl.formulaExpression; const c = Number(input.selectionStart || 0); const left = String(input.value || "").slice(0, c); const m = left.match(/[A-Za-z_][A-Za-z0-9_]*$/); return { cursor: c, prefix: m ? m[0] : "", prefixStart: m ? c - m[0].length : c }; }
    function formulaContext() {
      const input = formEl.formulaExpression;
      const c = Number(input.selectionStart || 0);
      const left = String(input.value || "").slice(0, c);
      const p = currentPrefix();
      let i = p.prefixStart - 1;
      while (i >= 0 && /\\s/.test(left[i])) i -= 1;
      const prev = i >= 0 ? left[i] : "";
      return { expectsOperand: (!prev || "+-*/(,".includes(prev)) };
    }
    function updateFormulaMetaFromReferences(references) {
      let meta = {};
      try {
        const parsed = JSON.parse(String(formEl.formulaMeta.value || "").trim() || "{}");
        if (parsed && typeof parsed === "object" && !Array.isArray(parsed)) meta = parsed;
      } catch (_err) {
        meta = {};
      }
      meta.references = Array.isArray(references) ? references : [];
      formEl.formulaMeta.value = JSON.stringify(meta, null, 2);
    }
    function scoreMatch(text, query) { const t = String(text || "").toLowerCase(); const q = String(query || "").toLowerCase(); if (!q) return 1; if (t.startsWith(q)) return 4; if (t.includes(q)) return 2; return 0; }
    function buildSuggestions() {
      const ctx = formulaContext(); const info = currentPrefix(); const q = info.prefix.toLowerCase(); const items = [];
      if (ctx.expectsOperand) {
        FORMULA_FUNCTIONS.forEach((name) => { const score = scoreMatch(name, q); if (!score) return; items.push({ label: name + "()", kind: "function", insert: name + "()", score: score + 2 }); });
        entityCatalog.forEach((e) => { const key = String(e.entity_key || ""); const label = String(e.label || ""); const score = Math.max(scoreMatch(key, q), scoreMatch(label, q)); if (!score) return; items.push({ label: key + " - " + label, kind: "entity", insert: key, score: score + 1 }); });
      } else {
        ["+", "-", "*", "/", ")"].forEach((op) => { if (op.startsWith(q) || !q) items.push({ label: op, kind: "operator", insert: op, score: 1 }); });
      }
      items.sort((a, b) => b.score - a.score || a.label.localeCompare(b.label));
      return items.slice(0, 20);
    }
    function renderSuggestions(items) {
      suggestionItems = items || []; suggestionIndex = suggestionItems.length ? 0 : -1;
      if (!suggestionItems.length) { suggestionsEl.classList.add("hidden"); suggestionsEl.innerHTML = ""; return; }
      suggestionsEl.innerHTML = suggestionItems.map((s, i) => '<div class="formula-suggestion ' + (i === suggestionIndex ? "active" : "") + '" data-idx="' + String(i) + '"><span>' + esc(s.label) + '</span><span class="formula-kind">' + esc(s.kind) + '</span></div>').join("");
      suggestionsEl.classList.remove("hidden");
      Array.from(suggestionsEl.querySelectorAll(".formula-suggestion")).forEach((node) => node.addEventListener("mousedown", (e) => { e.preventDefault(); const idx = Number(node.getAttribute("data-idx") || -1); if (idx >= 0) applySuggestion(suggestionItems[idx]); }));
    }
    function refreshSuggestionHighlight() { Array.from(suggestionsEl.querySelectorAll(".formula-suggestion")).forEach((node) => { const idx = Number(node.getAttribute("data-idx") || -1); node.classList.toggle("active", idx === suggestionIndex); }); }
    function applySuggestion(item) {
      if (!item) return;
      const input = formEl.formulaExpression; const p = currentPrefix(); const before = String(input.value || "").slice(0, p.prefixStart); const after = String(input.value || "").slice(Number(input.selectionStart || 0)); const ins = String(item.insert || "");
      input.value = before + ins + after; const caret = before.length + ins.length - (ins.endsWith("()") ? 1 : 0);
      input.focus(); input.setSelectionRange(caret, caret); renderSuggestions([]); onFormulaInput();
    }
    function onFormulaInput() {
      try {
        const out = validateFormulaClient(formEl.formulaExpression.value);
        setValidation(true, out.message);
        updateFormulaMetaFromReferences(out.references || []);
      } catch (err) {
        setValidation(false, err.message || String(err));
      }
      renderSuggestions(buildSuggestions());
    }
    function insertToken(token) { const input = formEl.formulaExpression; const s = Number(input.selectionStart || 0); const e = Number(input.selectionEnd || s); const val = String(input.value || ""); const before = val.slice(0, s); const after = val.slice(e); const t = String(token || ""); input.value = before + t + after; const caret = before.length + t.length - (t.endsWith("()") ? 1 : 0); input.focus(); input.setSelectionRange(caret, caret); onFormulaInput(); }

    function resetForm() {
      formEl.fieldKey.value = ""; formEl.fieldKey.readOnly = true; formEl.label.value = ""; formEl.description.value = "";
      formEl.dataType.value = "number"; formEl.formulaExpression.value = ""; formEl.formulaVersion.value = "1"; formEl.formulaMeta.value = "{}"; formEl.isActive.checked = true;
      setValidation(true, "");
      renderSuggestions([]);
      refreshSystemFieldKey();
    }
    function openCreate() { mode = "create"; selectedKey = ""; dialogTitleEl.textContent = "Create Field"; resetForm(); dialogEl.showModal(); }
    function openEdit(fieldKey) {
      const row = rows.find((x) => String(x.field_key) === String(fieldKey)); if (!row) return;
      mode = "edit"; selectedKey = row.field_key; dialogTitleEl.textContent = "Edit: " + row.label + " (" + row.field_key + ")";
      formEl.fieldKey.value = row.field_key || ""; formEl.fieldKey.readOnly = true; formEl.label.value = row.label || ""; formEl.description.value = row.description || "";
      formEl.dataType.value = row.data_type || "number"; formEl.formulaExpression.value = row.formula_expression || ""; formEl.formulaVersion.value = String(Number(row.formula_version || 1));
      formEl.formulaMeta.value = toPretty(row.formula_meta_json || {}); formEl.isActive.checked = !!row.is_active; dialogEl.showModal(); onFormulaInput();
    }
    function readFormPayload() {
      const payload = {
        field_key: String(formEl.fieldKey.value || "").trim(),
        label: String(formEl.label.value || "").trim(),
        description: String(formEl.description.value || "").trim(),
        data_type: String(formEl.dataType.value || "").trim(),
        formula_expression: String(formEl.formulaExpression.value || "").trim(),
        formula_version: Number(formEl.formulaVersion.value || 1),
        formula_meta_json: parseJsonObject("Formula Meta JSON", formEl.formulaMeta.value),
        is_active: !!formEl.isActive.checked
      };
      validateFormulaClient(payload.formula_expression);
      return payload;
    }
    async function loadData() {
      const includeInactive = !!includeInactiveEl.checked;
      const resp = await fetch(API + "?include_inactive=" + (includeInactive ? "1" : "0"));
      const data = await resp.json().catch(() => ({}));
      if (!resp.ok) throw new Error(data.error || "Failed to load managed fields.");
      rows = Array.isArray(data.fields) ? data.fields : [];
      entityCatalog = Array.isArray(data.entity_catalog) ? data.entity_catalog : [];
      renderRows();
      setStatus("Loaded " + rows.length + " fields.", "ok");
    }
    async function saveForm() {
      const payload = readFormPayload();
      const endpoint = mode === "create" ? API : (API + "/" + encodeURIComponent(selectedKey));
      const method = mode === "create" ? "POST" : "PUT";
      const resp = await fetch(endpoint, { method, headers: { "Content-Type": "application/json" }, body: JSON.stringify(payload) });
      const data = await resp.json().catch(() => ({}));
      if (!resp.ok) throw new Error(data.error || "Failed to save field.");
      dialogEl.close();
      await loadData();
      setStatus(mode === "create" ? "Field created." : "Field updated.", "ok");
    }
    async function softDelete(fieldKey) {
      if (!window.confirm("Soft delete this field?")) return;
      const resp = await fetch(API + "/" + encodeURIComponent(fieldKey), { method: "DELETE" });
      const data = await resp.json().catch(() => ({}));
      if (!resp.ok) throw new Error(data.error || "Failed to delete field.");
      await loadData();
      setStatus("Field soft-deleted.", "ok");
    }
    async function restore(fieldKey) {
      const resp = await fetch(API + "/" + encodeURIComponent(fieldKey) + "/restore", { method: "POST" });
      const data = await resp.json().catch(() => ({}));
      if (!resp.ok) throw new Error(data.error || "Failed to restore field.");
      await loadData();
      setStatus("Field restored.", "ok");
    }

    formEl.formulaExpression.addEventListener("input", onFormulaInput);
    formEl.formulaExpression.addEventListener("keydown", (e) => {
      if (suggestionsEl.classList.contains("hidden") || !suggestionItems.length) return;
      if (e.key === "ArrowDown") { e.preventDefault(); suggestionIndex = (suggestionIndex + 1) % suggestionItems.length; refreshSuggestionHighlight(); }
      else if (e.key === "ArrowUp") { e.preventDefault(); suggestionIndex = (suggestionIndex - 1 + suggestionItems.length) % suggestionItems.length; refreshSuggestionHighlight(); }
      else if (e.key === "Enter") { e.preventDefault(); if (suggestionIndex >= 0) applySuggestion(suggestionItems[suggestionIndex]); }
      else if (e.key === "Escape") { e.preventDefault(); renderSuggestions([]); }
    });
    formEl.formulaExpression.addEventListener("blur", () => setTimeout(() => renderSuggestions([]), 120));
    document.getElementById("mf-formula-quick-insert").addEventListener("click", (e) => { const token = e.target && e.target.getAttribute ? e.target.getAttribute("data-token") : ""; if (token) insertToken(token); });
    formEl.label.addEventListener("input", () => refreshSystemFieldKey());
    document.getElementById("mf-reload-btn").addEventListener("click", () => loadData().catch((err) => setStatus(err.message || String(err), "err")));
    document.getElementById("mf-create-btn").addEventListener("click", openCreate);
    includeInactiveEl.addEventListener("change", () => loadData().catch((err) => setStatus(err.message || String(err), "err")));
    document.getElementById("mf-cancel").addEventListener("click", () => dialogEl.close());
    document.getElementById("mf-save").addEventListener("click", () => saveForm().catch((err) => setStatus(err.message || String(err), "err")));
    loadData().catch((err) => setStatus(err.message || String(err), "err"));
  </script>
  <script src="/shared-nav.js"></script>
</body>
</html>""".replace("__SETTINGS_TOP_NAV__", _settings_top_nav_html(MANAGE_FIELDS_SETTINGS_ROUTE))


def _projects_settings_html() -> str:
    return """<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Managed Projects</title>
  <link rel="stylesheet" href="/shared-nav.css">
  <link rel="stylesheet" href="/material-symbols.css">
  <style>
    :root { --bg:#f4f7fc; --card:#fff; --line:#cbd5e1; --text:#0f172a; --muted:#475569; --brand:#1d4ed8; --ok:#166534; --err:#b91c1c; --soft:#eef4ff; }
    body { margin:0; padding:20px; background:linear-gradient(180deg,#f2f6ff,#f8fbff); color:var(--text); font-family:"Segoe UI",Tahoma,sans-serif; }
    .card { max-width:1180px; margin:0 auto; background:var(--card); border:1px solid var(--line); border-radius:12px; padding:16px; }
    .top { display:flex; justify-content:space-between; align-items:flex-start; gap:10px; flex-wrap:wrap; }
    .row { display:flex; gap:8px; flex-wrap:wrap; align-items:center; margin-top:10px; }
    .btn { border:1px solid #1e40af; background:var(--brand); color:#fff; border-radius:8px; padding:8px 12px; cursor:pointer; text-decoration:none; font-size:.86rem; }
    .btn.alt { border-color:var(--line); background:#fff; color:var(--text); }
    .btn.warn { border-color:#7f1d1d; background:#b91c1c; color:#fff; }
    .btn:disabled { opacity:.55; cursor:not-allowed; }
    .grid { display:grid; gap:10px; grid-template-columns:repeat(auto-fit,minmax(220px,1fr)); margin-top:12px; }
    label { display:block; font-size:.82rem; font-weight:700; margin-bottom:4px; }
    input, select { width:100%; border:1px solid var(--line); border-radius:8px; padding:8px; font-size:.92rem; box-sizing:border-box; }
    input[readonly] { background:#f8fafc; color:#334155; }
    .table-wrap { margin-top:12px; border:1px solid var(--line); border-radius:10px; overflow:auto; background:#fff; max-height:58vh; }
    table { border-collapse:collapse; min-width:980px; width:100%; }
    th, td { border-bottom:1px solid #e2e8f0; padding:8px 10px; font-size:.86rem; text-align:left; vertical-align:top; }
    thead th { position:sticky; top:0; background:#eff6ff; font-size:.76rem; text-transform:uppercase; letter-spacing:.04em; }
    .mono { font-family:Consolas,monospace; font-size:.8rem; }
    .swatch { display:inline-block; width:22px; height:14px; border:1px solid #94a3b8; border-radius:4px; vertical-align:middle; margin-right:6px; }
    .hint { color:var(--muted); font-size:.8rem; margin-top:5px; }
    .pill { background:var(--soft); border:1px solid #c7d7ff; color:#1e3a8a; border-radius:999px; padding:3px 8px; font-size:.76rem; font-weight:700; }
    #status { margin-top:10px; min-height:1.2em; font-size:.9rem; }
    #status.ok { color:var(--ok); }
    #status.err { color:var(--err); }
    tbody tr.is-selected { background:#eaf2ff; }
    .active-state { color:#0f766e; font-weight:700; }
    .inactive-state { color:#b45309; font-weight:700; }
    .drawer-overlay { position:fixed; inset:0; background:rgba(15,23,42,.42); opacity:0; pointer-events:none; transition:opacity .18s ease; z-index:70; }
    .drawer-overlay.open { opacity:1; pointer-events:auto; }
    .drawer { position:fixed; top:0; right:0; height:100vh; width:min(460px,94vw); background:#fff; border-left:1px solid var(--line); box-shadow:-18px 0 34px rgba(15,23,42,.18); transform:translateX(100%); transition:transform .2s ease; z-index:80; display:flex; flex-direction:column; }
    .drawer.open { transform:translateX(0); }
    .drawer-head { padding:14px 14px 10px; border-bottom:1px solid var(--line); display:flex; justify-content:space-between; align-items:flex-start; gap:10px; }
    .drawer-title { margin:0; font-size:1rem; }
    .drawer-body { padding:12px 14px; overflow:auto; display:grid; gap:10px; }
    .drawer-foot { padding:12px 14px; border-top:1px solid var(--line); display:flex; gap:8px; flex-wrap:wrap; }
    .drawer-subtitle { margin:.35rem 0 0; color:var(--muted); font-size:.82rem; }
    .hidden-fields { display:none; }
  </style>
</head>
<body>
  <main class="card">
    <div class="top">
      <div>
        <h1 style="margin:0;font-size:1.25rem;">Managed Projects</h1>
        <p style="margin:.45rem 0 0;color:var(--muted);font-size:.92rem;">Manage Jira projects used by export scripts.</p>
        <p class="hint" style="margin-top:4px;">Step 1: search/select a Jira project. Step 2: set display name + color, then save.</p>
      </div>
      <div class="row">__SETTINGS_TOP_NAV__</div>
    </div>

    <div class="grid">
      <div>
        <label for="jira-search">Search Jira Projects</label>
        <input id="jira-search" type="text" placeholder="Type project key or name">
      </div>
      <div>
        <label for="jira-search-results">Search Results (select one)</label>
        <select id="jira-search-results"></select>
        <div class="hint">Pick one result to auto-fill Project Key and Project Name.</div>
      </div>
      <div style="grid-column:1 / -1;">
        <div class="pill">Add New Project opens a right-side drawer modal.</div>
      </div>
    </div>
    <div class="hidden-fields" aria-hidden="true">
      <input id="project-key" readonly class="mono">
      <input id="project-name" readonly>
      <input id="display-name">
      <input id="color-pick" type="color" value="#1D4ED8">
      <input id="color-hex" value="#1D4ED8" class="mono">
    </div>

    <div class="row">
      <button id="search-btn" class="btn alt" type="button">Search Jira</button>
      <span id="search-count" class="pill">0 results</span>
      <button id="new-btn" class="btn alt" type="button">Clear Form</button>
      <button id="add-project-btn" class="btn" type="button">Add New Project</button>
      <span class="pill">Modify existing projects from the table</span>
      <label style="display:inline-flex;align-items:center;gap:6px;font-size:.86rem;font-weight:600;">
        <input id="include-inactive" type="checkbox"> Show Inactive
      </label>
      <button id="reload-btn" class="btn alt" type="button">Reload</button>
      <span id="list-count" class="pill">0 managed projects</span>
    </div>

    <div id="status"></div>

    <div class="table-wrap">
      <table>
        <thead>
          <tr><th>Key</th><th>Project Name</th><th>Display Name</th><th>Color</th><th>Active</th><th>Updated</th><th>Action</th></tr>
        </thead>
        <tbody id="projects-tbody"></tbody>
      </table>
    </div>
  </main>
  <div id="add-drawer-overlay" class="drawer-overlay"></div>
  <aside id="add-drawer" class="drawer" aria-hidden="true">
    <div class="drawer-head">
      <div>
        <h2 class="drawer-title">Add New Project</h2>
        <p class="drawer-subtitle">Save the selected Jira project to managed projects.</p>
      </div>
      <button id="add-drawer-close" class="btn alt" type="button">Close</button>
    </div>
    <div class="drawer-body">
      <div>
        <label for="add-project-key">Project Key</label>
        <input id="add-project-key" readonly class="mono">
      </div>
      <div>
        <label for="add-project-name">Project Name</label>
        <input id="add-project-name" readonly>
      </div>
      <div>
        <label for="add-display-name">Display Name</label>
        <input id="add-display-name">
      </div>
      <div>
        <label for="add-color-hex">Color</label>
        <div class="row" style="margin-top:0;">
          <input id="add-color-pick" type="color" value="#1D4ED8" style="width:64px;padding:2px;height:38px;">
          <input id="add-color-hex" class="mono" value="#1D4ED8">
        </div>
      </div>
      <div id="add-drawer-status" class="hint"></div>
    </div>
    <div class="drawer-foot">
      <button id="add-drawer-save" class="btn" type="button">Add Project</button>
    </div>
  </aside>
  <div id="edit-drawer-overlay" class="drawer-overlay"></div>
  <aside id="edit-drawer" class="drawer" aria-hidden="true">
    <div class="drawer-head">
      <div>
        <h2 class="drawer-title">Modify Project</h2>
        <div id="drawer-key" class="mono" style="margin-top:4px;"></div>
      </div>
      <button id="drawer-close" class="btn alt" type="button">Close</button>
    </div>
    <div class="drawer-body">
      <div>
        <label for="drawer-project-name">Project Name</label>
        <input id="drawer-project-name" readonly>
      </div>
      <div>
        <label for="drawer-display-name">Display Name</label>
        <input id="drawer-display-name">
      </div>
      <div>
        <label for="drawer-color-hex">Color</label>
        <div class="row" style="margin-top:0;">
          <input id="drawer-color-pick" type="color" value="#1D4ED8" style="width:64px;padding:2px;height:38px;">
          <input id="drawer-color-hex" class="mono" value="#1D4ED8">
        </div>
      </div>
      <div id="drawer-status" class="hint"></div>
    </div>
    <div class="drawer-foot">
      <button id="drawer-update" class="btn" type="button">Update</button>
      <button id="drawer-delete" class="btn warn" type="button">Soft Delete</button>
      <button id="drawer-restore" class="btn alt" type="button">Restore</button>
    </div>
  </aside>

  <script>
    const API = "/api/projects";
    const SEARCH_API = "/api/jira/projects/search";
    const statusEl = document.getElementById("status");
    const tbodyEl = document.getElementById("projects-tbody");
    const searchEl = document.getElementById("jira-search");
    const searchResultsEl = document.getElementById("jira-search-results");
    const projectKeyEl = document.getElementById("project-key");
    const projectNameEl = document.getElementById("project-name");
    const displayNameEl = document.getElementById("display-name");
    const colorPickEl = document.getElementById("color-pick");
    const colorHexEl = document.getElementById("color-hex");
    const includeInactiveEl = document.getElementById("include-inactive");
    const searchCountEl = document.getElementById("search-count");
    const listCountEl = document.getElementById("list-count");
    const addProjectBtn = document.getElementById("add-project-btn");
    const addDrawerEl = document.getElementById("add-drawer");
    const addDrawerOverlayEl = document.getElementById("add-drawer-overlay");
    const addProjectKeyEl = document.getElementById("add-project-key");
    const addProjectNameEl = document.getElementById("add-project-name");
    const addDisplayNameEl = document.getElementById("add-display-name");
    const addColorPickEl = document.getElementById("add-color-pick");
    const addColorHexEl = document.getElementById("add-color-hex");
    const addDrawerStatusEl = document.getElementById("add-drawer-status");
    const editDrawerEl = document.getElementById("edit-drawer");
    const editDrawerOverlayEl = document.getElementById("edit-drawer-overlay");
    const drawerKeyEl = document.getElementById("drawer-key");
    const drawerProjectNameEl = document.getElementById("drawer-project-name");
    const drawerDisplayNameEl = document.getElementById("drawer-display-name");
    const drawerColorPickEl = document.getElementById("drawer-color-pick");
    const drawerColorHexEl = document.getElementById("drawer-color-hex");
    const drawerStatusEl = document.getElementById("drawer-status");
    const drawerUpdateBtn = document.getElementById("drawer-update");
    const drawerDeleteBtn = document.getElementById("drawer-delete");
    const drawerRestoreBtn = document.getElementById("drawer-restore");

    let rows = [];
    let searchRows = [];
    let selectedKey = "";

    function setStatus(msg, kind) {
      statusEl.textContent = String(msg || "");
      statusEl.className = kind || "";
    }
    function esc(v) {
      return String(v == null ? "" : v).replace(/&/g, "&amp;").replace(/</g, "&lt;").replace(/>/g, "&gt;");
    }
    function setColor(hex) {
      const value = String(hex || "").toUpperCase();
      colorHexEl.value = value || "#1D4ED8";
      colorPickEl.value = /^#[0-9A-F]{6}$/.test(value) ? value : "#1D4ED8";
    }
    function normalizeColor(hex) {
      const value = String(hex || "").trim().toUpperCase();
      if (!/^#[0-9A-F]{6}$/.test(value)) throw new Error("Color must be #RRGGBB.");
      return value;
    }
    function selectedRow() {
      return rows.find((x) => String(x.project_key || "") === String(selectedKey || "")) || null;
    }
    function updateActionState() {
      const hasSelection = !!String(projectKeyEl.value || "").trim();
      const row = selectedRow();
      addProjectBtn.disabled = !hasSelection || !!row;
    }
    function resetForm() {
      selectedKey = "";
      projectKeyEl.value = "";
      projectNameEl.value = "";
      displayNameEl.value = "";
      setColor("#1D4ED8");
      searchResultsEl.value = "";
      updateActionState();
    }
    function setDrawerColor(hex) {
      const value = String(hex || "").trim().toUpperCase();
      drawerColorHexEl.value = value || "#1D4ED8";
      drawerColorPickEl.value = /^#[0-9A-F]{6}$/.test(value) ? value : "#1D4ED8";
    }
    function setDrawerStatus(message, isError) {
      drawerStatusEl.textContent = String(message || "");
      drawerStatusEl.style.color = isError ? "#b91c1c" : "#475569";
    }
    function setAddDrawerStatus(message, isError) {
      addDrawerStatusEl.textContent = String(message || "");
      addDrawerStatusEl.style.color = isError ? "#b91c1c" : "#475569";
    }
    function setAddDrawerColor(hex) {
      const value = String(hex || "").trim().toUpperCase();
      addColorHexEl.value = value || "#1D4ED8";
      addColorPickEl.value = /^#[0-9A-F]{6}$/.test(value) ? value : "#1D4ED8";
    }
    function openAddDrawer() {
      const projectKey = String(projectKeyEl.value || "").trim().toUpperCase();
      const projectName = String(projectNameEl.value || "").trim();
      if (!projectKey || !projectName) {
        throw new Error("Select a Jira project first.");
      }
      addProjectKeyEl.value = projectKey;
      addProjectNameEl.value = projectName;
      addDisplayNameEl.value = String(displayNameEl.value || "").trim() || projectName;
      setAddDrawerColor(String(colorHexEl.value || "#1D4ED8"));
      setAddDrawerStatus("", false);
      addDrawerEl.classList.add("open");
      addDrawerOverlayEl.classList.add("open");
      addDrawerEl.setAttribute("aria-hidden", "false");
      addDisplayNameEl.focus();
    }
    function closeAddDrawer() {
      addDrawerEl.classList.remove("open");
      addDrawerOverlayEl.classList.remove("open");
      addDrawerEl.setAttribute("aria-hidden", "true");
      setAddDrawerStatus("", false);
    }
    function openEditDrawer(row) {
      if (!row) return;
      closeAddDrawer();
      selectedKey = String(row.project_key || "");
      drawerKeyEl.textContent = selectedKey;
      drawerProjectNameEl.value = String(row.project_name || "");
      drawerDisplayNameEl.value = String(row.display_name || "");
      setDrawerColor(String(row.color_hex || ""));
      drawerUpdateBtn.disabled = false;
      drawerDeleteBtn.disabled = !row.is_active;
      drawerRestoreBtn.disabled = !!row.is_active;
      setDrawerStatus(row.is_active ? "Project is active." : "Project is inactive.", false);
      editDrawerEl.classList.add("open");
      editDrawerOverlayEl.classList.add("open");
      editDrawerEl.setAttribute("aria-hidden", "false");
      renderRows();
      updateActionState();
    }
    function closeEditDrawer() {
      editDrawerEl.classList.remove("open");
      editDrawerOverlayEl.classList.remove("open");
      editDrawerEl.setAttribute("aria-hidden", "true");
      setDrawerStatus("", false);
    }
    function renderSearchResults() {
      const options = ['<option value="">Select Jira project</option>'];
      for (const row of searchRows) {
        const key = String(row.project_key || "");
        const name = String(row.project_name || "");
        options.push('<option value="' + esc(key) + '">' + esc(key + " - " + name) + '</option>');
      }
      searchResultsEl.innerHTML = options.join("");
      searchCountEl.textContent = String(searchRows.length) + " results";
    }
    function selectedSearchProject() {
      const key = String(searchResultsEl.value || "");
      return searchRows.find((x) => String(x.project_key || "") === key) || null;
    }
    function setFormFromSearch() {
      const row = selectedSearchProject();
      if (!row) return;
      selectedKey = "";
      projectKeyEl.value = String(row.project_key || "");
      projectNameEl.value = String(row.project_name || "");
      if (!displayNameEl.value.trim()) displayNameEl.value = String(row.project_name || "");
      updateActionState();
      setStatus("Project selected from Jira. Click Add New Project to continue.", "ok");
    }
    function renderRows() {
      listCountEl.textContent = String(rows.length) + " managed projects";
      tbodyEl.innerHTML = rows.map((r) =>
        '<tr class="' + (String(r.project_key || "") === selectedKey ? "is-selected" : "") + '">'
        + '<td class="mono">' + esc(r.project_key) + '</td>'
        + '<td>' + esc(r.project_name) + '</td>'
        + '<td>' + esc(r.display_name) + '</td>'
        + '<td><span class="swatch" style="background:' + esc(r.color_hex || "#000000") + ';"></span><span class="mono">' + esc(r.color_hex) + '</span></td>'
        + '<td class="' + (r.is_active ? "active-state" : "inactive-state") + '">' + (r.is_active ? "Active" : "Inactive") + '</td>'
        + '<td class="mono">' + esc(r.updated_at_utc || "-") + '</td>'
        + '<td><button class="btn alt" type="button" data-edit="' + esc(r.project_key) + '">Modify</button></td>'
        + '</tr>'
      ).join("");
      Array.from(tbodyEl.querySelectorAll("button[data-edit]")).forEach((btn) => {
        btn.addEventListener("click", () => {
          const key = String(btn.getAttribute("data-edit") || "");
          const row = rows.find((x) => String(x.project_key || "") === key);
          if (!row) return;
          openEditDrawer(row);
          setStatus("Modify drawer opened for " + key + ".", "ok");
        });
      });
    }
    async function searchProjects() {
      const query = encodeURIComponent(String(searchEl.value || "").trim());
      const resp = await fetch(SEARCH_API + "?q=" + query + "&limit=25");
      const body = await resp.json().catch(() => ({}));
      if (!resp.ok) throw new Error(String(body.error || "Failed to search Jira projects."));
      searchRows = Array.isArray(body.projects) ? body.projects : [];
      renderSearchResults();
      setStatus("Loaded " + searchRows.length + " Jira projects.", "ok");
    }
    async function loadProjects() {
      const includeInactive = includeInactiveEl.checked ? "1" : "0";
      const resp = await fetch(API + "?include_inactive=" + includeInactive);
      const body = await resp.json().catch(() => ({}));
      if (!resp.ok) throw new Error(String(body.error || "Failed to load managed projects."));
      rows = Array.isArray(body.projects) ? body.projects : [];
      renderRows();
      updateActionState();
      setStatus("Loaded " + rows.length + " managed projects.", "ok");
    }
    function buildAddDrawerPayload() {
      const projectKey = String(addProjectKeyEl.value || "").trim().toUpperCase();
      const projectName = String(addProjectNameEl.value || "").trim();
      const displayName = String(addDisplayNameEl.value || "").trim();
      if (!projectKey || !projectName) throw new Error("No Jira project selected.");
      if (!displayName) throw new Error("Display Name is required.");
      return {
        project_key: projectKey,
        project_name: projectName,
        display_name: displayName,
        color_hex: normalizeColor(addColorHexEl.value),
      };
    }
    async function addProjectFromDrawer() {
      const payload = buildAddDrawerPayload();
      const resp = await fetch(API, { method: "POST", headers: { "Content-Type": "application/json" }, body: JSON.stringify(payload) });
      const body = await resp.json().catch(() => ({}));
      if (!resp.ok) throw new Error(String(body.error || "Failed to add project."));
      selectedKey = payload.project_key;
      await loadProjects();
      closeAddDrawer();
      setStatus("Project added.", "ok");
    }
    function buildDrawerPayload() {
      const projectKey = String(selectedKey || "").trim().toUpperCase();
      const projectName = String(drawerProjectNameEl.value || "").trim();
      const displayName = String(drawerDisplayNameEl.value || "").trim();
      if (!projectKey || !projectName) throw new Error("No project selected.");
      if (!displayName) throw new Error("Display Name is required.");
      return {
        project_key: projectKey,
        project_name: projectName,
        display_name: displayName,
        color_hex: normalizeColor(drawerColorHexEl.value),
      };
    }
    async function updateProjectFromDrawer() {
      const payload = buildDrawerPayload();
      const resp = await fetch(API + "/" + encodeURIComponent(payload.project_key), { method: "PUT", headers: { "Content-Type": "application/json" }, body: JSON.stringify(payload) });
      const body = await resp.json().catch(() => ({}));
      if (!resp.ok) throw new Error(String(body.error || "Failed to update project."));
      await loadProjects();
      const row = rows.find((x) => String(x.project_key || "") === payload.project_key);
      if (row) openEditDrawer(row);
      setStatus("Project updated.", "ok");
    }
    async function softDeleteProjectFromDrawer() {
      const projectKey = String(selectedKey || "").trim().toUpperCase();
      if (!projectKey) throw new Error("No project selected.");
      if (!window.confirm("Soft delete " + projectKey + "?")) return;
      const resp = await fetch(API + "/" + encodeURIComponent(projectKey), { method: "DELETE" });
      const body = await resp.json().catch(() => ({}));
      if (!resp.ok) throw new Error(String(body.error || "Failed to delete project."));
      await loadProjects();
      const row = rows.find((x) => String(x.project_key || "") === projectKey);
      if (row) {
        openEditDrawer(row);
      } else {
        closeEditDrawer();
      }
      setStatus("Project soft-deleted.", "ok");
    }
    async function restoreProjectFromDrawer() {
      const projectKey = String(selectedKey || "").trim().toUpperCase();
      if (!projectKey) throw new Error("No project selected.");
      const resp = await fetch(API + "/" + encodeURIComponent(projectKey) + "/restore", { method: "POST" });
      const body = await resp.json().catch(() => ({}));
      if (!resp.ok) throw new Error(String(body.error || "Failed to restore project."));
      await loadProjects();
      const row = rows.find((x) => String(x.project_key || "") === projectKey);
      if (row) openEditDrawer(row);
      setStatus("Project restored.", "ok");
    }

    colorPickEl.addEventListener("change", () => setColor(colorPickEl.value));
    colorHexEl.addEventListener("change", () => setColor(colorHexEl.value));
    searchResultsEl.addEventListener("change", setFormFromSearch);
    searchEl.addEventListener("keydown", (event) => {
      if (event.key === "Enter") {
        event.preventDefault();
        searchProjects().catch((e) => setStatus(e.message || String(e), "err"));
      }
    });
    displayNameEl.addEventListener("input", updateActionState);
    document.getElementById("search-btn").addEventListener("click", () => searchProjects().catch((e) => setStatus(e.message || String(e), "err")));
    document.getElementById("reload-btn").addEventListener("click", () => loadProjects().catch((e) => setStatus(e.message || String(e), "err")));
    document.getElementById("new-btn").addEventListener("click", () => { resetForm(); closeEditDrawer(); closeAddDrawer(); setStatus("Ready to add a project.", ""); });
    document.getElementById("add-project-btn").addEventListener("click", () => {
      try {
        openAddDrawer();
      } catch (e) {
        setStatus(e.message || String(e), "err");
      }
    });
    includeInactiveEl.addEventListener("change", () => loadProjects().catch((e) => setStatus(e.message || String(e), "err")));
    addColorPickEl.addEventListener("change", () => setAddDrawerColor(addColorPickEl.value));
    addColorHexEl.addEventListener("change", () => setAddDrawerColor(addColorHexEl.value));
    addDisplayNameEl.addEventListener("input", () => setAddDrawerStatus("", false));
    document.getElementById("add-drawer-save").addEventListener("click", () => addProjectFromDrawer().catch((e) => setAddDrawerStatus(e.message || String(e), true)));
    document.getElementById("add-drawer-close").addEventListener("click", closeAddDrawer);
    addDrawerOverlayEl.addEventListener("click", closeAddDrawer);
    drawerColorPickEl.addEventListener("change", () => setDrawerColor(drawerColorPickEl.value));
    drawerColorHexEl.addEventListener("change", () => setDrawerColor(drawerColorHexEl.value));
    drawerDisplayNameEl.addEventListener("input", () => setDrawerStatus("", false));
    drawerUpdateBtn.addEventListener("click", () => updateProjectFromDrawer().catch((e) => setDrawerStatus(e.message || String(e), true)));
    drawerDeleteBtn.addEventListener("click", () => softDeleteProjectFromDrawer().catch((e) => setDrawerStatus(e.message || String(e), true)));
    drawerRestoreBtn.addEventListener("click", () => restoreProjectFromDrawer().catch((e) => setDrawerStatus(e.message || String(e), true)));
    document.getElementById("drawer-close").addEventListener("click", closeEditDrawer);
    editDrawerOverlayEl.addEventListener("click", closeEditDrawer);

    (async function init() {
      setStatus("Loading...", "");
      resetForm();
      try {
        await Promise.all([searchProjects(), loadProjects()]);
      } catch (e) {
        setStatus(e.message || String(e), "err");
      }
    })();
  </script>
  <script src="/shared-nav.js"></script>
</body>
</html>""".replace("__SETTINGS_TOP_NAV__", _settings_top_nav_html(PROJECTS_SETTINGS_ROUTE))


def _page_categories_settings_html() -> str:
    return """<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Page Categories</title>
  <link rel="stylesheet" href="/shared-nav.css">
  <link rel="stylesheet" href="/material-symbols.css">
  <style>
    :root { --bg:#f5f8ff; --card:#fff; --line:#d6dfef; --ink:#0f172a; --muted:#475569; --brand:#1d4ed8; --ok:#166534; --err:#b91c1c; }
    * { box-sizing:border-box; }
    body { margin:0; padding:20px; background:linear-gradient(180deg,#eef4ff,#f8fbff); color:var(--ink); font-family:"Segoe UI",Tahoma,sans-serif; }
    .wrap { max-width:1240px; margin:0 auto; display:grid; gap:14px; }
    .card { border:1px solid var(--line); border-radius:12px; background:var(--card); padding:14px; box-shadow:0 8px 22px rgba(15,23,42,.05); }
    .top { display:flex; justify-content:space-between; align-items:flex-start; gap:8px; flex-wrap:wrap; }
    .row { display:flex; gap:8px; flex-wrap:wrap; align-items:center; margin-top:10px; }
    .btn { border:1px solid #1e40af; background:var(--brand); color:#fff; border-radius:8px; padding:8px 12px; cursor:pointer; text-decoration:none; font-size:.85rem; }
    .btn.alt { border-color:var(--line); background:#fff; color:var(--ink); }
    .btn.warn { border-color:#7f1d1d; background:#b91c1c; color:#fff; }
    .muted { color:var(--muted); font-size:.88rem; }
    .table-wrap { margin-top:10px; border:1px solid var(--line); border-radius:10px; overflow:auto; background:#fff; }
    table { border-collapse:collapse; width:100%; min-width:760px; }
    th, td { border-bottom:1px solid #e8edf7; padding:8px 10px; text-align:left; font-size:.84rem; vertical-align:top; }
    thead th { position:sticky; top:0; background:#eff6ff; text-transform:uppercase; letter-spacing:.03em; font-size:.73rem; }
    .split { display:grid; gap:14px; grid-template-columns:1fr 1fr; }
    @media (max-width:980px) { .split { grid-template-columns:1fr; } }
    .page-block { border:1px solid var(--line); border-radius:10px; padding:10px; background:#fbfdff; }
    .page-grid { display:grid; gap:8px; }
    .page-row { display:grid; grid-template-columns:240px 1fr; gap:8px; align-items:center; }
    .page-row label { font-weight:600; font-size:.84rem; }
    input[type="text"], input[type="number"], select { width:100%; border:1px solid var(--line); border-radius:8px; padding:7px; font-size:.85rem; }
    select[multiple] { min-height:78px; }
    .icon-btn { display:inline-flex; align-items:center; justify-content:center; width:40px; height:34px; border:1px solid var(--line); background:#fff; border-radius:8px; cursor:pointer; }
    .icon-btn .material-symbols-outlined { font-size:1.15rem; color:#1e40af; }
    .icon-picker { position:fixed; inset:0; background:rgba(15,23,42,.35); display:none; align-items:center; justify-content:center; z-index:1000; }
    .icon-picker.open { display:flex; }
    .icon-picker-card { width:min(820px,94vw); max-height:82vh; overflow:auto; background:#fff; border:1px solid var(--line); border-radius:12px; padding:12px; box-shadow:0 18px 40px rgba(15,23,42,.2); }
    .icon-grid { display:grid; grid-template-columns:repeat(auto-fill,minmax(58px,1fr)); gap:8px; margin-top:10px; }
    .icon-tile { border:1px solid var(--line); background:#fff; border-radius:8px; min-height:50px; display:flex; align-items:center; justify-content:center; cursor:pointer; }
    .icon-tile:hover { border-color:#93c5fd; background:#eff6ff; }
    .icon-tile .material-symbols-outlined { font-size:1.35rem; color:#1e40af; }
    #status { min-height:1.2em; margin-top:8px; font-size:.9rem; }
    #status.ok { color:var(--ok); } #status.err { color:var(--err); }
  </style>
</head>
<body>
  <main class="wrap">
    <section class="card">
      <div class="top">
        <div>
          <h1 style="margin:0;font-size:1.25rem;">Page Categories</h1>
          <p class="muted" style="margin:.45rem 0 0;">Create navigation categories and assign report/configuration pages. Hidden categories are removed from navigation.</p>
        </div>
        <div class="row">__SETTINGS_TOP_NAV__</div>
      </div>
      <div class="row">
        <button id="reload-btn" class="btn alt" type="button">Reload</button>
        <button id="add-category-btn" class="btn alt" type="button">Add Category</button>
        <button id="save-btn" class="btn" type="button">Save All</button>
      </div>
      <div id="status"></div>
    </section>

    <section class="card">
      <h2 style="margin:0;font-size:1.02rem;">Categories</h2>
      <p class="muted" style="margin:.4rem 0 0;">Use Display Order for navigation ordering (lower first).</p>
      <div class="table-wrap">
        <table>
          <thead><tr><th>Name</th><th>Icon</th><th>Display in Navigation</th><th>Display Order</th><th>Active</th><th>Actions</th></tr></thead>
          <tbody id="categories-tbody"></tbody>
        </table>
      </div>
    </section>

    <section class="card">
      <h2 style="margin:0;font-size:1.02rem;">Page Assignments</h2>
      <p class="muted" style="margin:.4rem 0 0;">Each page can belong to multiple categories.</p>
      <div class="split">
        <section class="page-block">
          <h3 style="margin:0 0 8px;font-size:.95rem;">Reports</h3>
          <div id="report-pages" class="page-grid"></div>
        </section>
        <section class="page-block">
          <h3 style="margin:0 0 8px;font-size:.95rem;">Configuration</h3>
          <div id="configuration-pages" class="page-grid"></div>
        </section>
      </div>
    </section>
  </main>

  <div id="icon-picker" class="icon-picker" aria-hidden="true">
    <div class="icon-picker-card">
      <div class="row" style="margin-top:0;justify-content:space-between;">
        <strong>Select Category Icon</strong>
        <button id="icon-picker-close" class="btn alt" type="button">Close</button>
      </div>
      <div class="icon-grid" id="icon-picker-grid"></div>
    </div>
  </div>
  <script>
    const API = "/api/page-categories";
    const CATEGORY_ICON_CHOICES = [
      "folder","category","dashboard","settings","tune","dataset","description","analytics","query_stats","monitoring",
      "account_tree","event_note","calendar_month","inventory_2","workspace_premium","groups","build","construction","engineering","manage_accounts",
      "admin_panel_settings","security","verified","insights","history","list_alt","segment","widgets","view_list","splitscreen",
      "star","flag","bookmark","rocket_launch","lightbulb","fact_check","task","check_circle","label","extension"
    ];
    const statusEl = document.getElementById("status");
    const categoriesTbodyEl = document.getElementById("categories-tbody");
    const reportPagesEl = document.getElementById("report-pages");
    const configurationPagesEl = document.getElementById("configuration-pages");
    const iconPickerEl = document.getElementById("icon-picker");
    const iconPickerGridEl = document.getElementById("icon-picker-grid");

    let categories = [];
    let assignments = [];
    let pageCatalog = [];
    let nextTempId = -1;
    let iconPickerCategoryIndex = -1;

    function esc(v) {
      return String(v == null ? "" : v)
        .replace(/&/g, "&amp;").replace(/</g, "&lt;").replace(/>/g, "&gt;");
    }
    function setStatus(message, kind) {
      statusEl.textContent = String(message || "");
      statusEl.className = kind || "";
    }
    function categoryOptions() {
      return categories
        .filter((c) => c && c.is_active)
        .sort((a, b) => {
          const ao = Number(a.display_order || 0);
          const bo = Number(b.display_order || 0);
          if (ao !== bo) return ao - bo;
          return String(a.name || "").localeCompare(String(b.name || ""), undefined, { sensitivity: "base" });
        });
    }
    function assignmentIdsByPageKey(pageKey) {
      return new Set(
        assignments
          .filter((item) => String(item.page_key || "") === pageKey)
          .map((item) => Number(item.category_id || 0))
          .filter((x) => Number.isInteger(x) && x > 0)
      );
    }
    function normalizeIconName(value) {
      const text = String(value || "").trim().toLowerCase();
      if (!text) return "folder";
      return /^[a-z0-9_]+$/.test(text) ? text : "folder";
    }
    function renderIconPicker() {
      iconPickerGridEl.innerHTML = CATEGORY_ICON_CHOICES.map((iconName) =>
        "<button class='icon-tile' type='button' data-icon-choice='" + esc(iconName) + "' title='" + esc(iconName) + "'>"
          + "<span class='material-symbols-outlined' aria-hidden='true'>" + esc(iconName) + "</span>"
        + "</button>"
      ).join("");
      Array.from(iconPickerGridEl.querySelectorAll("button[data-icon-choice]")).forEach((btn) => {
        btn.addEventListener("click", () => {
          if (iconPickerCategoryIndex < 0 || iconPickerCategoryIndex >= categories.length) return;
          categories[iconPickerCategoryIndex].icon_name = normalizeIconName(btn.getAttribute("data-icon-choice"));
          closeIconPicker();
          renderCategories();
        });
      });
    }
    function openIconPicker(categoryIndex) {
      iconPickerCategoryIndex = categoryIndex;
      iconPickerEl.classList.add("open");
      iconPickerEl.setAttribute("aria-hidden", "false");
      renderIconPicker();
    }
    function closeIconPicker() {
      iconPickerCategoryIndex = -1;
      iconPickerEl.classList.remove("open");
      iconPickerEl.setAttribute("aria-hidden", "true");
    }
    function renderCategories() {
      categoriesTbodyEl.innerHTML = categories.map((row, idx) => {
        const id = Number(row.id || 0);
        const iconName = normalizeIconName(row.icon_name);
        return "<tr>"
          + "<td><input type='text' data-cat-name='" + idx + "' value='" + esc(row.name || "") + "' placeholder='Category name'></td>"
          + "<td><button class='icon-btn' type='button' data-cat-icon='" + idx + "' title='Select icon'><span class='material-symbols-outlined'>" + esc(iconName) + "</span></button></td>"
          + "<td><label><input type='checkbox' data-cat-display='" + idx + "' " + (row.display_in_navigation ? "checked" : "") + "> Show</label></td>"
          + "<td><input type='number' data-cat-order='" + idx + "' value='" + esc(row.display_order == null ? 100 : row.display_order) + "'></td>"
          + "<td><label><input type='checkbox' data-cat-active='" + idx + "' " + (row.is_active ? "checked" : "") + "> Active</label></td>"
          + "<td><button class='btn warn' type='button' data-cat-delete='" + id + "'>Delete</button></td>"
          + "</tr>";
      }).join("");

      Array.from(categoriesTbodyEl.querySelectorAll("input[data-cat-name]")).forEach((el) => {
        el.addEventListener("input", () => {
          const idx = Number(el.getAttribute("data-cat-name") || -1);
          if (idx < 0 || idx >= categories.length) return;
          categories[idx].name = String(el.value || "").trim();
          renderAssignments();
        });
      });
      Array.from(categoriesTbodyEl.querySelectorAll("button[data-cat-icon]")).forEach((btn) => {
        btn.addEventListener("click", () => {
          const idx = Number(btn.getAttribute("data-cat-icon") || -1);
          if (idx < 0 || idx >= categories.length) return;
          openIconPicker(idx);
        });
      });
      Array.from(categoriesTbodyEl.querySelectorAll("input[data-cat-display]")).forEach((el) => {
        el.addEventListener("change", () => {
          const idx = Number(el.getAttribute("data-cat-display") || -1);
          if (idx < 0 || idx >= categories.length) return;
          categories[idx].display_in_navigation = !!el.checked;
        });
      });
      Array.from(categoriesTbodyEl.querySelectorAll("input[data-cat-order]")).forEach((el) => {
        el.addEventListener("input", () => {
          const idx = Number(el.getAttribute("data-cat-order") || -1);
          if (idx < 0 || idx >= categories.length) return;
          categories[idx].display_order = Number(el.value || 100);
          renderAssignments();
        });
      });
      Array.from(categoriesTbodyEl.querySelectorAll("input[data-cat-active]")).forEach((el) => {
        el.addEventListener("change", () => {
          const idx = Number(el.getAttribute("data-cat-active") || -1);
          if (idx < 0 || idx >= categories.length) return;
          categories[idx].is_active = !!el.checked;
          renderAssignments();
        });
      });
      Array.from(categoriesTbodyEl.querySelectorAll("button[data-cat-delete]")).forEach((btn) => {
        btn.addEventListener("click", () => {
          const id = Number(btn.getAttribute("data-cat-delete") || 0);
          categories = categories.filter((x) => Number(x.id || 0) !== id);
          assignments = assignments.filter((x) => Number(x.category_id || 0) !== id);
          renderCategories();
          renderAssignments();
        });
      });
    }
    function renderAssignmentSection(containerEl, type) {
      const pages = pageCatalog
        .filter((x) => String(x.page_type || "") === type)
        .sort((a, b) => Number(a.default_nav_order || 0) - Number(b.default_nav_order || 0));
      const opts = categoryOptions();
      containerEl.innerHTML = pages.map((page) => {
        const key = String(page.page_key || "");
        const selectedIds = assignmentIdsByPageKey(key);
        const optionsHtml = opts.map((cat) => {
          const cid = Number(cat.id || 0);
          const selected = selectedIds.has(cid) ? " selected" : "";
          return "<option value='" + cid + "'" + selected + ">" + esc(cat.name || "") + "</option>";
        }).join("");
        return "<div class='page-row'>"
          + "<label>" + esc(page.title || key) + "</label>"
          + "<select multiple data-page-key='" + esc(key) + "' data-page-type='" + esc(type) + "'>" + optionsHtml + "</select>"
          + "</div>";
      }).join("");
      Array.from(containerEl.querySelectorAll("select[data-page-key]")).forEach((selectEl) => {
        selectEl.addEventListener("change", () => {
          const pageKey = String(selectEl.getAttribute("data-page-key") || "");
          const pageType = String(selectEl.getAttribute("data-page-type") || "");
          const selected = new Set(Array.from(selectEl.selectedOptions).map((op) => Number(op.value || 0)).filter((x) => x > 0));
          assignments = assignments.filter((x) => String(x.page_key || "") !== pageKey);
          for (const cid of selected.values()) {
            assignments.push({ page_key: pageKey, page_type: pageType, category_id: cid });
          }
        });
      });
    }
    function renderAssignments() {
      renderAssignmentSection(reportPagesEl, "report");
      renderAssignmentSection(configurationPagesEl, "configuration");
    }
    function addCategory() {
      categories.push({
        id: nextTempId,
        name: "",
        icon_name: "folder",
        display_in_navigation: true,
        display_order: 100,
        is_active: true
      });
      nextTempId -= 1;
      renderCategories();
      renderAssignments();
    }
    function payloadFromState() {
      const normalizedCategories = categories.map((item) => ({
        id: Number(item.id || 0) > 0 ? Number(item.id || 0) : 0,
        name: String(item.name || "").trim(),
        icon_name: normalizeIconName(item.icon_name),
        display_in_navigation: !!item.display_in_navigation,
        display_order: Number(item.display_order || 100),
        is_active: !!item.is_active,
      }));
      const normalizedAssignments = assignments.map((item) => ({
        page_key: String(item.page_key || ""),
        page_type: String(item.page_type || ""),
        category_id: Number(item.category_id || 0),
      }));
      return { categories: normalizedCategories, assignments: normalizedAssignments };
    }
    async function loadAll() {
      const resp = await fetch(API, { cache: "no-store" });
      const body = await resp.json().catch(() => ({}));
      if (!resp.ok) throw new Error(String(body.error || "Failed to load page categories."));
      categories = (Array.isArray(body.categories) ? body.categories : []).map((item) => ({ ...item, icon_name: normalizeIconName(item.icon_name) }));
      assignments = Array.isArray(body.assignments) ? body.assignments : [];
      pageCatalog = Array.isArray(body.page_catalog) ? body.page_catalog : [];
      renderCategories();
      renderAssignments();
      setStatus("Loaded " + categories.length + " categories.", "ok");
    }
    async function saveAll() {
      const payload = payloadFromState();
      const resp = await fetch(API, {
        method: "PUT",
        headers: { "Content-Type": "application/json" },
        body: JSON.stringify(payload),
      });
      const body = await resp.json().catch(() => ({}));
      if (!resp.ok) throw new Error(String(body.error || "Failed to save page categories."));
      categories = (Array.isArray(body.categories) ? body.categories : []).map((item) => ({ ...item, icon_name: normalizeIconName(item.icon_name) }));
      assignments = Array.isArray(body.assignments) ? body.assignments : [];
      pageCatalog = Array.isArray(body.page_catalog) ? body.page_catalog : pageCatalog;
      renderCategories();
      renderAssignments();
      setStatus("Saved page categories.", "ok");
    }
    document.getElementById("reload-btn").addEventListener("click", () => loadAll().catch((err) => setStatus(err.message || String(err), "err")));
    document.getElementById("add-category-btn").addEventListener("click", addCategory);
    document.getElementById("save-btn").addEventListener("click", () => saveAll().catch((err) => setStatus(err.message || String(err), "err")));
    document.getElementById("icon-picker-close").addEventListener("click", closeIconPicker);
    iconPickerEl.addEventListener("click", (event) => {
      if (event.target === iconPickerEl) closeIconPicker();
    });
    loadAll().catch((err) => setStatus(err.message || String(err), "err"));
  </script>
  <script src="/shared-nav.js"></script>
</body>
</html>""".replace("__SETTINGS_TOP_NAV__", _settings_top_nav_html(PAGE_CATEGORIES_SETTINGS_ROUTE))


def _epics_dropdown_options_settings_html() -> str:
    return """<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Epic Dropdown Options</title>
  <link rel="stylesheet" href="/shared-nav.css">
  <link rel="stylesheet" href="/material-symbols.css">
  <style>
    :root { --bg:#f4f7fc; --card:#fff; --line:#cbd5e1; --text:#0f172a; --muted:#475569; --brand:#1d4ed8; --ok:#166534; --err:#b91c1c; }
    * { box-sizing:border-box; }
    body { margin:0; padding:20px; background:linear-gradient(180deg,#f2f6ff,#f8fbff); color:var(--text); font-family:"Segoe UI",Tahoma,sans-serif; }
    .card { max-width:1180px; margin:0 auto; background:var(--card); border:1px solid var(--line); border-radius:12px; padding:16px; }
    .top { display:flex; justify-content:space-between; align-items:flex-start; gap:10px; flex-wrap:wrap; }
    .row { display:flex; gap:8px; flex-wrap:wrap; align-items:center; margin-top:10px; }
    .grid { display:grid; gap:12px; grid-template-columns:repeat(auto-fit,minmax(360px,1fr)); margin-top:12px; }
    .block { border:1px solid var(--line); border-radius:10px; padding:12px; background:#fff; }
    label { display:block; font-size:.84rem; font-weight:700; margin-bottom:6px; }
    textarea { width:100%; min-height:260px; border:1px solid var(--line); border-radius:8px; padding:8px; font-size:.9rem; line-height:1.35; resize:vertical; }
    .btn { border:1px solid #1e40af; background:var(--brand); color:#fff; border-radius:8px; padding:8px 12px; cursor:pointer; text-decoration:none; font-size:.86rem; }
    .btn.alt { border-color:var(--line); background:#fff; color:var(--text); }
    .hint { color:var(--muted); font-size:.82rem; margin-top:6px; }
    #status { margin-top:10px; min-height:1.2em; font-size:.9rem; }
    #status.ok { color:var(--ok); }
    #status.err { color:var(--err); }
  </style>
</head>
<body>
  <main class="card">
    <div class="top">
      <div>
        <h1 style="margin:0;font-size:1.25rem;">Epic Dropdown Options</h1>
        <p style="margin:.45rem 0 0;color:var(--muted);font-size:.92rem;">Configure shared dropdown values used by Product Categorization, Component, and Plan Status fields.</p>
      </div>
      <div class="row">__SETTINGS_TOP_NAV__</div>
    </div>

    <div class="row">
      <button id="reload-btn" class="btn alt" type="button">Reload</button>
      <button id="save-btn" class="btn" type="button">Save Options</button>
      <span class="hint">One option per line. Blank lines are ignored. Duplicates are removed.</span>
    </div>
    <div id="status"></div>

    <div class="grid">
      <section class="block">
        <label for="product-category-options">Product Categorization Options</label>
        <textarea id="product-category-options" placeholder="Core&#10;Payments&#10;Reporting"></textarea>
      </section>
      <section class="block">
        <label for="component-options">Component Options</label>
        <textarea id="component-options" placeholder="Checkout API&#10;Web Portal&#10;Mobile App"></textarea>
      </section>
      <section class="block">
        <label for="plan-status-options">Plan Status Options</label>
        <textarea id="plan-status-options" placeholder="Planned&#10;Not Planned Yet"></textarea>
      </section>
    </div>
  </main>

  <script>
    const API = "/api/epics-management/dropdown-options";
    const statusEl = document.getElementById("status");
    const productCategoryEl = document.getElementById("product-category-options");
    const componentEl = document.getElementById("component-options");
    const planStatusEl = document.getElementById("plan-status-options");

    function setStatus(msg, kind) {
      statusEl.textContent = String(msg || "");
      statusEl.className = kind || "";
    }
    function parseOptions(text) {
      const seen = new Set();
      const out = [];
      String(text || "").split(/\\r?\\n/).forEach((line) => {
        const value = String(line || "").trim();
        if (!value) return;
        const lower = value.toLowerCase();
        if (seen.has(lower)) return;
        seen.add(lower);
        out.push(value);
      });
      return out;
    }
    function writeOptions(values) {
      return (Array.isArray(values) ? values : []).map((item) => String(item || "").trim()).filter(Boolean).join("\\n");
    }
    function payloadFromForm() {
      return {
        product_category: parseOptions(productCategoryEl.value),
        component: parseOptions(componentEl.value),
        plan_status: parseOptions(planStatusEl.value),
      };
    }
    function setForm(body) {
      productCategoryEl.value = writeOptions(body && body.product_category_options);
      componentEl.value = writeOptions(body && body.component_options);
      planStatusEl.value = writeOptions(body && body.plan_status_options);
    }
    async function loadOptions() {
      const resp = await fetch(API, { cache: "no-store" });
      const body = await resp.json().catch(() => ({}));
      if (!resp.ok) throw new Error(String(body.error || "Failed to load dropdown options."));
      setForm(body);
      setStatus("Loaded dropdown options.", "ok");
    }
    async function saveOptions() {
      const payload = payloadFromForm();
      const resp = await fetch(API, {
        method: "PUT",
        headers: { "Content-Type": "application/json" },
        body: JSON.stringify(payload),
      });
      const body = await resp.json().catch(() => ({}));
      if (!resp.ok) throw new Error(String(body.error || "Failed to save dropdown options."));
      setForm(body);
      setStatus("Saved dropdown options.", "ok");
    }

    document.getElementById("reload-btn").addEventListener("click", () => {
      loadOptions().catch((err) => setStatus(err.message || String(err), "err"));
    });
    document.getElementById("save-btn").addEventListener("click", () => {
      saveOptions().catch((err) => setStatus(err.message || String(err), "err"));
    });

    (async function init() {
      setStatus("Loading...", "");
      try {
        await loadOptions();
      } catch (err) {
        setStatus(err.message || String(err), "err");
      }
    })();
  </script>
  <script src="/shared-nav.js"></script>
</body>
</html>""".replace("__SETTINGS_TOP_NAV__", _settings_top_nav_html(EPICS_DROPDOWN_OPTIONS_SETTINGS_ROUTE))


def _epic_phases_settings_html() -> str:
    return """<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Epic Phases</title>
  <link rel="stylesheet" href="/shared-nav.css">
  <link rel="stylesheet" href="/material-symbols.css">
  <style>
    :root { --bg:#f4f7fc; --card:#fff; --line:#cbd5e1; --text:#0f172a; --muted:#475569; --brand:#1d4ed8; --ok:#166534; --warn:#92400e; --err:#b91c1c; }
    * { box-sizing:border-box; }
    body { margin:0; padding:20px; background:linear-gradient(180deg,#f2f6ff,#f8fbff); color:var(--text); font-family:"Segoe UI",Tahoma,sans-serif; }
    .card { max-width:1240px; margin:0 auto; background:var(--card); border:1px solid var(--line); border-radius:12px; padding:16px; }
    .top { display:flex; justify-content:space-between; align-items:flex-start; gap:10px; flex-wrap:wrap; }
    .row { display:flex; gap:8px; flex-wrap:wrap; align-items:center; margin-top:10px; }
    .grid { display:grid; gap:12px; grid-template-columns:repeat(auto-fit,minmax(320px,1fr)); margin-top:12px; }
    .panel { border:1px solid var(--line); border-radius:10px; padding:12px; background:#fff; }
    label { display:block; font-size:.84rem; font-weight:700; margin-bottom:6px; }
    input, select { width:100%; border:1px solid var(--line); border-radius:8px; padding:8px; font-size:.9rem; }
    .btn { border:1px solid #1e40af; background:var(--brand); color:#fff; border-radius:8px; padding:8px 12px; cursor:pointer; text-decoration:none; font-size:.86rem; }
    .btn.alt { border-color:var(--line); background:#fff; color:var(--text); }
    .btn.small { padding:5px 9px; font-size:.8rem; }
    .icon-btn { width:28px; height:28px; border-radius:8px; border:1px solid var(--line); background:#fff; color:#334155; display:inline-flex; align-items:center; justify-content:center; cursor:pointer; }
    .icon-btn[disabled] { opacity:.45; cursor:not-allowed; }
    .icon-btn.danger { color:#991b1b; border-color:#fecaca; background:#fff7f7; }
    .hint { color:var(--muted); font-size:.82rem; margin-top:6px; }
    #restore-hint { display:none; margin-top:8px; padding:8px; border:1px solid #e2e8f0; border-radius:8px; background:#f8fafc; font-size:.84rem; color:#334155; }
    #status { margin-top:10px; min-height:1.2em; font-size:.9rem; color:var(--muted); }
    #status.ok { color:var(--ok); }
    #status.warn { color:var(--warn); }
    #status.err { color:var(--err); }
    .tabs { display:flex; gap:8px; flex-wrap:wrap; margin-top:12px; }
    .tab-btn { border:1px solid var(--line); background:#fff; color:var(--text); border-radius:999px; padding:6px 12px; cursor:pointer; font-size:.84rem; }
    .tab-btn.active { border-color:#1e40af; background:#eff6ff; color:#1e3a8a; font-weight:700; }
    .table-wrap { margin-top:8px; border:1px solid var(--line); border-radius:10px; overflow:auto; background:#fff; }
    table { width:100%; min-width:860px; border-collapse:separate; border-spacing:0; }
    th, td { border-bottom:1px solid #e2e8f0; border-right:1px solid #e2e8f0; padding:8px 10px; font-size:.83rem; text-align:left; vertical-align:middle; }
    th:last-child, td:last-child { border-right:none; }
    th { background:#f8fbff; text-transform:uppercase; letter-spacing:.03em; font-size:.7rem; }
    code { font-family:ui-monospace,SFMono-Regular,Menlo,Consolas,monospace; font-size:.8rem; }
    .pill { display:inline-flex; align-items:center; padding:2px 8px; border-radius:999px; font-size:.75rem; border:1px solid #cbd5e1; background:#fff; color:#334155; }
    .mono { font-family:ui-monospace,SFMono-Regular,Menlo,Consolas,monospace; }
  </style>
</head>
<body>
  <main class="card">
    <div class="top">
      <div>
        <h1 style="margin:0;font-size:1.25rem;">Manage Epic Phases</h1>
        <p style="margin:.45rem 0 0;color:var(--muted);font-size:.92rem;">Epic Plan Columns are managed here as Epic Phases. Configure order, Jira URL support, and restore deleted phases.</p>
      </div>
      <div class="row">__SETTINGS_TOP_NAV__</div>
    </div>

    <div class="grid">
      <section class="panel">
        <label for="phase-name">Epic Phase Name</label>
        <input id="phase-name" type="text" placeholder="e.g. Security Plan">
        <div id="restore-hint"></div>
        <div class="row">
          <div style="min-width:210px;flex:1;">
            <label for="phase-position" style="margin-bottom:4px;">Insert Position</label>
            <select id="phase-position"></select>
          </div>
          <div style="min-width:220px;display:flex;align-items:flex-end;">
            <label for="phase-jira-enabled" style="display:flex;align-items:center;gap:8px;margin:0;">
              <input id="phase-jira-enabled" type="checkbox" style="width:auto;">
              Enable Jira URL support
            </label>
          </div>
        </div>
        <div class="row">
          <button id="add-phase-btn" class="btn" type="button">Add Epic Phase</button>
          <button id="reload-btn" class="btn alt" type="button">Reload</button>
        </div>
        <p class="hint">Default phases are locked for delete. Use up/down to reorder active phases.</p>
      </section>
    </div>

    <div id="status"></div>
    <div class="tabs" role="tablist" aria-label="Epic phase status tabs">
      <button id="tab-active" class="tab-btn active" type="button" role="tab" aria-selected="true">Active</button>
      <button id="tab-deleted" class="tab-btn" type="button" role="tab" aria-selected="false">Deleted</button>
    </div>
    <div class="table-wrap">
      <table>
        <thead>
          <tr>
            <th>#</th>
            <th>Epic Phase</th>
            <th>Key</th>
            <th>Jira URL</th>
            <th>Type</th>
            <th>Status</th>
            <th>Order</th>
            <th>Action</th>
          </tr>
        </thead>
        <tbody id="phases-tbody"></tbody>
      </table>
    </div>
  </main>

  <script>
    const PLAN_COLUMNS_API = "/api/epics-management/plan-columns";
    const PLAN_COLUMNS_UPDATE_API_BASE = "/api/epics-management/plan-columns";
    const PLAN_COLUMNS_DELETE_API_BASE = "/api/epics-management/plan-columns";
    const PLAN_COLUMNS_RESTORE_API_BASE = "/api/epics-management/plan-columns";
    const PLAN_COLUMNS_ORDER_API = "/api/epics-management/plan-columns/order";

    const statusEl = document.getElementById("status");
    const phaseNameEl = document.getElementById("phase-name");
    const phasePositionEl = document.getElementById("phase-position");
    const phaseJiraEnabledEl = document.getElementById("phase-jira-enabled");
    const restoreHintEl = document.getElementById("restore-hint");
    const phasesTbodyEl = document.getElementById("phases-tbody");
    const tabActiveEl = document.getElementById("tab-active");
    const tabDeletedEl = document.getElementById("tab-deleted");

    let phaseCatalog = [];
    let activeTab = "active";

    function esc(value) {
      return String(value == null ? "" : value).replace(/&/g, "&amp;").replace(/</g, "&lt;").replace(/>/g, "&gt;");
    }
    function setStatus(msg, kind) {
      statusEl.textContent = String(msg || "");
      statusEl.className = kind || "";
    }
    function iconUpSvg() {
      return '<svg viewBox="0 0 24 24" aria-hidden="true"><path fill="currentColor" d="M12 6l6 8h-4v4h-4v-4H6z"/></svg>';
    }
    function iconDownSvg() {
      return '<svg viewBox="0 0 24 24" aria-hidden="true"><path fill="currentColor" d="M12 18l-6-8h4V6h4v4h4z"/></svg>';
    }
    function iconTrashSvg() {
      return '<svg viewBox="0 0 24 24" aria-hidden="true"><path fill="currentColor" d="M9 3h6l1 2h4v2H4V5h4l1-2zm1 6h2v8h-2V9zm4 0h2v8h-2V9zM7 9h2v8H7V9z"/></svg>';
    }
    function getActivePhases() {
      return phaseCatalog.filter((item) => item.is_active);
    }
    function getDeletedPhases() {
      return phaseCatalog.filter((item) => !item.is_active);
    }
    function setActiveTab(tabName) {
      activeTab = tabName === "deleted" ? "deleted" : "active";
      const isActiveTab = activeTab === "active";
      tabActiveEl.classList.toggle("active", isActiveTab);
      tabDeletedEl.classList.toggle("active", !isActiveTab);
      tabActiveEl.setAttribute("aria-selected", isActiveTab ? "true" : "false");
      tabDeletedEl.setAttribute("aria-selected", isActiveTab ? "false" : "true");
      renderPhasesTable();
    }
    function formatPhasePositionLabel(index, phases) {
      if (!phases.length) return "1 (first)";
      if (index <= 0) return "1 (before " + String(phases[0].label || phases[0].key) + ")";
      if (index >= phases.length) return String(phases.length + 1) + " (after " + String(phases[phases.length - 1].label || phases[phases.length - 1].key) + ")";
      return String(index + 1) + " (between " + String(phases[index - 1].label || phases[index - 1].key) + " and " + String(phases[index].label || phases[index].key) + ")";
    }
    function renderPhasePositionOptions() {
      const active = getActivePhases();
      const totalSlots = active.length + 1;
      const current = Math.max(1, Math.min(totalSlots, Number(phasePositionEl.value) || totalSlots));
      const options = [];
      for (let slot = 1; slot <= totalSlots; slot += 1) {
        options.push('<option value="' + slot + '"' + (slot === current ? " selected" : "") + ">" + esc(formatPhasePositionLabel(slot - 1, active)) + "</option>");
      }
      phasePositionEl.innerHTML = options.join("");
      phasePositionEl.value = String(current);
    }
    function refreshRestoreHint() {
      const phaseName = String(phaseNameEl.value || "").trim();
      if (!phaseName) {
        restoreHintEl.style.display = "none";
        restoreHintEl.innerHTML = "";
        return;
      }
      const deleted = phaseCatalog.find((item) =>
        !item.is_active && String(item.label || "").toLowerCase() === phaseName.toLowerCase()
      );
      if (!deleted) {
        restoreHintEl.style.display = "none";
        restoreHintEl.innerHTML = "";
        return;
      }
      restoreHintEl.style.display = "block";
      restoreHintEl.innerHTML =
        'A deleted Epic Phase with this name exists (<code>' + esc(deleted.key) + '</code>). '
        + '<button id="restore-phase-btn" class="btn alt small" type="button" style="margin-left:8px;">Use Existing Deleted Phase</button>';
      const restoreBtn = document.getElementById("restore-phase-btn");
      if (restoreBtn) {
        restoreBtn.addEventListener("click", () => {
          restorePhase(deleted.key).catch((err) => setStatus(err.message || String(err), "err"));
        });
      }
    }
    function renderPhasesTable() {
      const activeOrder = getActivePhases().map((item) => item.key);
      const rows = activeTab === "deleted" ? getDeletedPhases() : getActivePhases();
      tabActiveEl.textContent = "Active (" + String(getActivePhases().length) + ")";
      tabDeletedEl.textContent = "Deleted (" + String(getDeletedPhases().length) + ")";
      const html = rows.map((phase) => {
        const key = String(phase && phase.key || "").trim();
        const label = String(phase && (phase.label || phase.key) || "").trim();
        const isDefault = !!(phase && phase.is_default);
        const isActive = !!(phase && phase.is_active);
        const jiraEnabled = !!(phase && phase.jira_link_enabled);
        const orderIdx = activeOrder.indexOf(key);
        const canMoveUp = isActive && orderIdx > 0;
        const canMoveDown = isActive && orderIdx >= 0 && orderIdx < (activeOrder.length - 1);
        const moveButtons = isActive
          ? '<div style="display:flex;gap:6px;">'
            + '<button class="icon-btn" type="button" data-move-phase="' + esc(key) + '" data-move-dir="up" title="Move up"' + (canMoveUp ? "" : " disabled") + ">" + iconUpSvg() + "</button>"
            + '<button class="icon-btn" type="button" data-move-phase="' + esc(key) + '" data-move-dir="down" title="Move down"' + (canMoveDown ? "" : " disabled") + ">" + iconDownSvg() + "</button>"
            + "</div>"
          : '<span class="mono" style="color:#94a3b8;">-</span>';
        const actionHtml = isActive
          ? (isDefault
              ? '<div style="display:flex;gap:6px;align-items:center;flex-wrap:wrap;"><button class="btn alt small" type="button" data-rename-phase="' + esc(key) + '" data-rename-phase-label="' + esc(label) + '">Rename</button><span class="pill">Locked</span></div>'
              : '<div style="display:flex;gap:6px;align-items:center;flex-wrap:wrap;"><button class="btn alt small" type="button" data-rename-phase="' + esc(key) + '" data-rename-phase-label="' + esc(label) + '">Rename</button><button class="icon-btn danger" type="button" data-delete-phase="' + esc(key) + '" data-delete-phase-label="' + esc(label) + '" title="Delete Epic Phase">' + iconTrashSvg() + "</button></div>")
          : '<button class="btn alt small" type="button" data-restore-phase="' + esc(key) + '">Restore</button>';
        return ""
          + "<tr>"
          + "<td>" + (isActive ? String(orderIdx + 1) : "-") + "</td>"
          + "<td>" + esc(label) + "</td>"
          + "<td><code>" + esc(key) + "</code></td>"
          + "<td>" + (jiraEnabled ? "Enabled" : "Disabled") + "</td>"
          + "<td>" + (isDefault ? "Default" : "Dynamic") + "</td>"
          + "<td>" + (isActive ? '<span class="pill">Active</span>' : '<span class="pill">Deleted</span>') + "</td>"
          + "<td>" + moveButtons + "</td>"
          + "<td>" + actionHtml + "</td>"
          + "</tr>";
      }).join("");
      const emptyLabel = activeTab === "deleted" ? "No deleted Epic Phases found." : "No active Epic Phases found.";
      phasesTbodyEl.innerHTML = html || '<tr><td colspan="8" class="mono" style="color:#94a3b8;">' + emptyLabel + "</td></tr>";
    }
    async function loadCatalog() {
      const resp = await fetch(PLAN_COLUMNS_API + "?include_inactive=1", { cache: "no-store" });
      const body = await resp.json().catch(() => ({}));
      if (!resp.ok) throw new Error(String(body.error || "Failed to load Epic Phases."));
      phaseCatalog = Array.isArray(body.columns) ? body.columns.map((item) => ({
        key: String(item && item.key || "").trim(),
        label: String(item && (item.label || item.key) || "").trim(),
        jira_link_enabled: !!(item && item.jira_link_enabled),
        is_default: !!(item && item.is_default),
        is_active: !!(item && item.is_active),
        sort_order: Number(item && item.sort_order || 0),
      })) : [];
      phaseCatalog.sort((a, b) => (a.sort_order - b.sort_order) || a.label.localeCompare(b.label) || a.key.localeCompare(b.key));
      renderPhasePositionOptions();
      renderPhasesTable();
      refreshRestoreHint();
    }
    async function addPhase() {
      const label = String(phaseNameEl.value || "").trim();
      if (!label) throw new Error("Epic Phase name is required.");
      const insertPosition = Number(phasePositionEl.value || "0");
      const jiraEnabled = !!phaseJiraEnabledEl.checked;
      const resp = await fetch(PLAN_COLUMNS_API, {
        method: "POST",
        headers: { "Content-Type": "application/json" },
        body: JSON.stringify({ label, jira_link_enabled: jiraEnabled, insert_position: insertPosition }),
      });
      const body = await resp.json().catch(() => ({}));
      if (!resp.ok) throw new Error(String(body.error || "Failed to add Epic Phase."));
      phaseNameEl.value = "";
      phaseJiraEnabledEl.checked = false;
      await loadCatalog();
      setStatus("Epic Phase added: " + label, "ok");
    }
    async function deletePhase(columnKey, columnLabel) {
      const key = String(columnKey || "").trim();
      if (!key) throw new Error("Epic Phase key is required.");
      const phase = phaseCatalog.find((item) => item.key === key);
      if (phase && phase.is_default) throw new Error("Default Epic Phases cannot be deleted.");
      const label = String(columnLabel || (phase && (phase.label || phase.key)) || key);
      if (!window.confirm('Delete Epic Phase "' + label + '"?')) return;
      const resp = await fetch(PLAN_COLUMNS_DELETE_API_BASE + "/" + encodeURIComponent(key), { method: "DELETE" });
      const body = await resp.json().catch(() => ({}));
      if (!resp.ok) throw new Error(String(body.error || "Failed to delete Epic Phase."));
      await loadCatalog();
      setStatus("Epic Phase deleted: " + label, "ok");
    }
    async function renamePhase(columnKey, currentLabel) {
      const key = String(columnKey || "").trim();
      if (!key) throw new Error("Epic Phase key is required.");
      const current = String(currentLabel || "").trim();
      const next = String(window.prompt("Rename Epic Phase:", current) || "").trim();
      if (!next || next === current) return;
      const resp = await fetch(PLAN_COLUMNS_UPDATE_API_BASE + "/" + encodeURIComponent(key), {
        method: "PUT",
        headers: { "Content-Type": "application/json" },
        body: JSON.stringify({ label: next }),
      });
      const body = await resp.json().catch(() => ({}));
      if (!resp.ok) throw new Error(String(body.error || "Failed to rename Epic Phase."));
      await loadCatalog();
      setStatus("Epic Phase renamed to: " + next + ". Changes are reflected in Epics Planner on reload.", "ok");
    }
    async function restorePhase(columnKey) {
      const key = String(columnKey || "").trim();
      if (!key) throw new Error("Epic Phase key is required for restore.");
      const resp = await fetch(PLAN_COLUMNS_RESTORE_API_BASE + "/" + encodeURIComponent(key) + "/restore", { method: "POST" });
      const body = await resp.json().catch(() => ({}));
      if (!resp.ok) throw new Error(String(body.error || "Failed to restore Epic Phase."));
      await loadCatalog();
      const label = String(body && body.column && body.column.label || key);
      setStatus("Epic Phase restored: " + label, "ok");
    }
    async function movePhase(columnKey, direction) {
      const key = String(columnKey || "").trim();
      const dir = String(direction || "").toLowerCase();
      if (!key || !dir) return;
      const active = getActivePhases();
      const index = active.findIndex((item) => item.key === key);
      if (index < 0) return;
      const targetIndex = dir === "up" ? index - 1 : index + 1;
      if (targetIndex < 0 || targetIndex >= active.length) return;
      const reordered = active.slice();
      const temp = reordered[index];
      reordered[index] = reordered[targetIndex];
      reordered[targetIndex] = temp;
      const resp = await fetch(PLAN_COLUMNS_ORDER_API, {
        method: "PUT",
        headers: { "Content-Type": "application/json" },
        body: JSON.stringify({ ordered_keys: reordered.map((item) => item.key) }),
      });
      const body = await resp.json().catch(() => ({}));
      if (!resp.ok) throw new Error(String(body.error || "Failed to reorder Epic Phases."));
      await loadCatalog();
      setStatus("Epic Phase order updated.", "ok");
    }

    document.getElementById("reload-btn").addEventListener("click", () => {
      loadCatalog().then(() => setStatus("Epic Phases loaded.", "ok")).catch((err) => setStatus(err.message || String(err), "err"));
    });
    document.getElementById("add-phase-btn").addEventListener("click", () => {
      addPhase().catch((err) => setStatus(err.message || String(err), "err"));
    });
    phaseNameEl.addEventListener("input", refreshRestoreHint);
    tabActiveEl.addEventListener("click", () => setActiveTab("active"));
    tabDeletedEl.addEventListener("click", () => setActiveTab("deleted"));
    phasesTbodyEl.addEventListener("click", (event) => {
      const target = event.target instanceof Element ? event.target : null;
      const renameBtn = target ? target.closest("button[data-rename-phase]") : null;
      if (renameBtn) {
        const key = String(renameBtn.getAttribute("data-rename-phase") || "");
        const label = String(renameBtn.getAttribute("data-rename-phase-label") || "");
        renamePhase(key, label).catch((err) => setStatus(err.message || String(err), "err"));
        return;
      }
      const deleteBtn = target ? target.closest("button[data-delete-phase]") : null;
      if (deleteBtn) {
        const key = String(deleteBtn.getAttribute("data-delete-phase") || "");
        const label = String(deleteBtn.getAttribute("data-delete-phase-label") || key);
        deletePhase(key, label).catch((err) => setStatus(err.message || String(err), "err"));
        return;
      }
      const restoreBtn = target ? target.closest("button[data-restore-phase]") : null;
      if (restoreBtn) {
        const key = String(restoreBtn.getAttribute("data-restore-phase") || "");
        restorePhase(key).catch((err) => setStatus(err.message || String(err), "err"));
        return;
      }
      const moveBtn = target ? target.closest("button[data-move-phase]") : null;
      if (moveBtn) {
        const key = String(moveBtn.getAttribute("data-move-phase") || "");
        const dir = String(moveBtn.getAttribute("data-move-dir") || "");
        movePhase(key, dir).catch((err) => setStatus(err.message || String(err), "err"));
      }
    });

    (async function init() {
      setStatus("Loading Epic Phases...", "");
      try {
        await loadCatalog();
        setStatus("Epic Phases loaded.", "ok");
      } catch (err) {
        setStatus(err.message || String(err), "err");
      }
    })();
  </script>
  <script src="/shared-nav.js"></script>
</body>
</html>""".replace("__SETTINGS_TOP_NAV__", _settings_top_nav_html(EPIC_PHASES_SETTINGS_ROUTE))


def _epics_management_settings_html() -> str:
    return """<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Epics Planner</title>
  <link rel="stylesheet" href="/shared-nav.css">
  <link rel="stylesheet" href="/material-symbols.css">
  <style>
    :root { --bg:#f5f7fb; --card:#fff; --line:#d1d9e8; --text:#0f172a; --muted:#475569; --brand:#1d4ed8; --ok:#166534; --warn:#92400e; --head:#eff6ff; --sticky:#f8fbff; --plan:#eef4ff; --epics-static-min-width:1420px; --epics-plan-col-min-width:170px; --epics-table-min-width:2610px; }
    * { box-sizing:border-box; }
    body { margin:0; padding:0; background:linear-gradient(180deg,#f3f7ff,#f8fbff); color:var(--text); font-family:"Segoe UI",Tahoma,sans-serif; }
    .card { width:100%; max-width:none; margin:0; border:1px solid var(--line); border-left:none; border-right:none; border-radius:0; background:var(--card); padding:16px; }
    .top { display:flex; justify-content:space-between; gap:12px; flex-wrap:wrap; align-items:center; padding:6px 0 18px; border-bottom:1px solid #e2e8f0; margin-bottom:6px; }
    .planner-page-title { display:flex; align-items:center; gap:10px; }
    .planner-page-title .planner-title-icon { font-size:1.6rem; color:var(--brand); font-variation-settings:"FILL" 0,"wght" 500,"GRAD" 0,"opsz" 24; line-height:1; }
    .planner-page-title h1 { margin:0; font-size:1.5rem; font-weight:600; letter-spacing:-0.02em; color:var(--text); }
    .planner-header { margin-top:4px; padding:10px 14px; background:linear-gradient(180deg,#fafbff 0%,#f6f9ff 100%); border:1px solid #e2e8f0; border-radius:8px; box-shadow:0 1px 2px rgba(15,23,42,.04); }
    .planner-header.collapsed { padding:10px 0; background:transparent; border:none; box-shadow:none; }
    .planner-header.collapsed .planner-header-content { display:none; }
    .planner-header-content { padding-top:0; }
    .planner-header-bar { display:flex; flex-wrap:wrap; align-items:center; gap:10px 16px; margin-top:8px; }
    .planner-header-bar .btn, .planner-header-bar .btn.alt { padding:6px 10px; font-size:.8rem; }
    .planner-header-bar .shortcut-chip { padding:3px 8px; font-size:.72rem; }
    .planner-header-bar .shortcut-chip kbd { padding:1px 4px; font-size:.7rem; }
    .planner-header-ipp { margin-left:auto; font-size:.7rem; color:var(--muted); opacity:.85; }
    #header-toggle-btn { font-size:.82rem; padding:6px 10px; }
    .row { display:flex; gap:8px; flex-wrap:wrap; align-items:center; margin-top:10px; }
    .btn { border:1px solid #1e40af; background:var(--brand); color:#fff; border-radius:8px; padding:8px 12px; cursor:pointer; text-decoration:none; font-size:.86rem; }
    .btn.alt { border-color:var(--line); background:#fff; color:var(--text); }
    .btn.small { padding:5px 9px; font-size:.8rem; }
    .muted { color:var(--muted); }
    #status { margin-top:10px; min-height:1.2em; font-size:.9rem; color:var(--muted); }
    #status.ok { color:var(--ok); }
    #status.warn { color:var(--warn); }
    .table-wrap { margin-top:12px; border:1px solid var(--line); border-radius:10px; overflow:auto; background:#fff; max-height:70vh; }
    table { border-collapse:separate; border-spacing:0; min-width:var(--epics-table-min-width); width:max(100%, var(--epics-table-min-width)); }
    th, td { border-bottom:1px solid #e2e8f0; border-right:1px solid #e2e8f0; padding:5px 8px; font-size:.8rem; line-height:1.2; text-align:left; vertical-align:top; background:#fff; }
    th:last-child, td:last-child { border-right:none; }
    thead th { position:sticky; top:0; background:var(--head); font-size:.7rem; text-transform:uppercase; letter-spacing:.04em; z-index:3; }
    thead th:nth-child(4), tbody td:nth-child(4) { position:sticky; left:0; z-index:4; background:var(--sticky); min-width:280px; }
    thead th:nth-child(4) { z-index:5; }
    th:nth-child(1), td:nth-child(1) { min-width:220px; }
    th:nth-child(2), td:nth-child(2) { min-width:220px; }
    th:nth-child(3), td:nth-child(3) { min-width:220px; }
    th:nth-child(4), td:nth-child(4) { min-width:280px; }
    th:nth-child(5), td:nth-child(5) { min-width:220px; }
    th:nth-child(6), td:nth-child(6) { min-width:180px; }
    th:nth-child(7), td:nth-child(7) { min-width:120px; }
    th:nth-child(8), td:nth-child(8) { min-width:150px; white-space:nowrap; }
    th:nth-child(9), td:nth-child(9) { min-width:170px; white-space:nowrap; }
    th:nth-child(10), td:nth-child(10) { min-width:170px; }
    th:nth-child(11), td:nth-child(11) { min-width:260px; }
    th:nth-child(12), td:nth-child(12),
    th:nth-child(13), td:nth-child(13),
    th:nth-child(14), td:nth-child(14),
    th:nth-child(15), td:nth-child(15),
    th:nth-child(16), td:nth-child(16),
    th:nth-child(17), td:nth-child(17),
    th:nth-child(18), td:nth-child(18) { min-width:150px; white-space:nowrap; }
    th:nth-child(19), td:nth-child(19) { min-width:260px; }
    td[contenteditable="true"] { min-width:180px; cursor:text; }
    td[contenteditable="true"]:focus { outline:2px solid #bfdbfe; outline-offset:-2px; background:#f8fbff; }
    td.description-cell { padding:4px 8px; }
    .description-editor { min-width:180px; max-height:60px; overflow-y:auto; white-space:pre-wrap; word-break:break-word; cursor:text; }
    .description-editor:focus { outline:2px solid #bfdbfe; outline-offset:-2px; background:#f8fbff; }
    select { width:100%; border:1px solid #cbd5e1; border-radius:6px; padding:4px 6px; font-size:.78rem; background:#fff; }
    .tree { display:grid; gap:2px; }
    .tree.tree-epic-cell { position:relative; min-height:22px; padding-right:50px; }
    .tree-line { display:flex; align-items:center; gap:6px; white-space:nowrap; }
    .tree-epic { margin-left:6px; color:#1e3a8a; font-weight:600; }
    .tree-title { display:block; max-width:260px; white-space:normal; line-height:1.15; }
    .tree-actions { display:flex; gap:4px; flex-wrap:wrap; margin-top:2px; }
    .tree.tree-epic-cell .tree-actions { position:absolute; top:0; right:0; margin-top:0; flex-wrap:nowrap; }
    .tree-toggle { border:1px solid #cbd5e1; background:#fff; color:#334155; border-radius:6px; width:20px; height:20px; line-height:1; display:inline-flex; align-items:center; justify-content:center; cursor:pointer; }
    .tree-toggle:hover { background:#f1f5f9; }
    .tree-label-project { font-weight:700; color:#0f172a; }
    .tree-label-category { font-weight:600; color:#1e293b; margin-left:14px; }
    .tree-group-total { margin-left:8px; font-size:.72rem; color:#1d4ed8; font-weight:700; background:#e8f1ff; border:1px solid #bfdbfe; border-radius:999px; padding:1px 7px; }
    .group-row td { background:#f8fbff; }
    .group-row.category td { background:#fbfdff; }
    .group-row.component td { background:#fefefe; }
    .missing-categorization { background:#fff7cc !important; }
    .group-plan-total { display:inline-flex; align-items:baseline; gap:3px; color:#1e3a8a; font-size:.74rem; }
    .group-plan-total b { font-size:.8rem; }
    .jira-open, .jira-edit { display:inline-flex; width:20px; height:20px; align-items:center; justify-content:center; border-radius:6px; border:1px solid #bfdbfe; color:#1d4ed8; text-decoration:none; font-weight:700; background:#eff6ff; cursor:pointer; }
    .jira-open:hover, .jira-edit:hover { background:#dbeafe; }
    .jira-open.disabled { border-color:#cbd5e1; color:#94a3b8; background:#f1f5f9; pointer-events:none; }
    .jira-edit { border-color:#cbd5e1; color:#334155; background:#fff; }
    .plan-cell { display:grid; gap:4px; }
    .plan-cell-actions { display:flex; justify-content:flex-end; gap:4px; }
    .plan-btn { width:100%; border:1px solid #bfdbfe; background:var(--plan); color:#1e3a8a; border-radius:8px; text-align:left; padding:6px; cursor:pointer; min-height:42px; max-height:60px; overflow-y:auto; }
    .plan-empty { color:#475569; font-size:.76rem; }
    .plan-summary { font-size:.74rem; line-height:1.2; }
    .plan-summary b { color:#1e3a8a; }
    .plan-main { display:flex; align-items:center; gap:6px; justify-content:space-between; }
    .plan-toggle { width:18px; height:18px; border:1px solid #bfdbfe; border-radius:6px; background:#fff; color:#1e3a8a; font-size:12px; line-height:1; cursor:pointer; display:inline-flex; align-items:center; justify-content:center; padding:0; }
    .plan-dates { margin-top:4px; }
    .plan-col-head { position:sticky; top:0; z-index:3; }
    .plan-col-head, .plan-col-cell { min-width:var(--epics-plan-col-min-width); width:var(--epics-plan-col-min-width); }
    .plan-col-head .plan-head-wrap { position:relative; min-height:20px; display:flex; align-items:center; justify-content:center; }
    .plan-col-head .plan-label { pointer-events:none; }
    .plan-col-head .plan-head-actions { position:absolute; right:2px; top:50%; transform:translateY(-50%); display:flex; gap:3px; }
    .plan-col-head.dragging { opacity:.55; background:#dbeafe; }
    .plan-col-head.drop-target { outline:2px solid #2563eb; outline-offset:-2px; }
    .plan-insert-handle { position:absolute; top:50%; transform:translate(-50%,-50%); left:0; width:18px; height:18px; border-radius:999px; border:1px solid #93c5fd; background:#eff6ff; color:#1d4ed8; font-size:12px; line-height:1; display:none; cursor:pointer; }
    .plan-insert-handle.after { left:100%; }
    .plan-col-head:hover .plan-insert-handle { display:inline-flex; align-items:center; justify-content:center; }
    .icon-btn { display:inline-flex; align-items:center; justify-content:center; width:20px; height:20px; border:1px solid #cbd5e1; border-radius:6px; background:#fff; color:#334155; cursor:pointer; padding:0; }
    .icon-btn:hover { background:#f1f5f9; }
    .icon-btn[disabled] { opacity:.45; cursor:not-allowed; background:#f8fafc; color:#64748b; }
    .icon-btn.danger { border-color:#fecaca; color:#b91c1c; background:#fff5f5; }
    .icon-btn.danger:hover { background:#ffe4e6; }
    .btn.danger { border-color:#fecaca; color:#b91c1c; background:#fff5f5; }
    .btn.danger:hover { background:#ffe4e6; }
    .icon-btn svg { width:12px; height:12px; display:block; }
    .draft-row td { background:#fff7ed; border-top:2px solid #fdba74; }
    .draft-input { width:100%; border:1px dashed #fb923c; border-radius:6px; padding:4px 6px; font-size:.78rem; background:#fff; }
    dialog { width:min(440px,94vw); border:none; border-radius:12px; padding:0; box-shadow:0 16px 40px rgba(15,23,42,.25); }
    dialog::backdrop { background:rgba(15,23,42,.45); }
    .modal-head { padding:14px 14px 10px; border-bottom:1px solid var(--line); }
    .modal-body { padding:12px 14px; display:grid; gap:10px; }
    .modal-foot { padding:12px 14px; border-top:1px solid var(--line); display:flex; gap:8px; justify-content:flex-end; }
    label { display:block; font-size:.8rem; font-weight:700; margin-bottom:4px; }
    input { width:100%; border:1px solid var(--line); border-radius:8px; padding:8px; font-size:.88rem; }
    textarea { width:100%; border:1px solid var(--line); border-radius:8px; padding:8px; font-size:.88rem; font-family:inherit; resize:vertical; min-height:74px; max-height:200px; overflow-y:auto; }
    #manage-columns-dialog { width:min(760px,96vw); border:none; border-radius:12px; padding:0; box-shadow:0 16px 40px rgba(15,23,42,.25); }
    .manage-columns-wrap { max-height:52vh; overflow:auto; border-top:1px solid var(--line); border-bottom:1px solid var(--line); }
    .manage-columns-table { width:100%; min-width:680px; border-collapse:separate; border-spacing:0; }
    .manage-columns-table th, .manage-columns-table td { border-bottom:1px solid #e2e8f0; border-right:1px solid #e2e8f0; padding:8px 10px; font-size:.82rem; vertical-align:middle; }
    .manage-columns-table th:last-child, .manage-columns-table td:last-child { border-right:none; }
    .manage-columns-table th { position:sticky; top:0; z-index:2; background:#f8fbff; text-transform:uppercase; letter-spacing:.03em; font-size:.7rem; }
    .shortcut-chip { display:inline-flex; align-items:center; gap:6px; border:1px solid #cbd5e1; border-radius:999px; background:#ffffff; padding:4px 10px; color:#334155; font-size:.8rem; }
    .shortcut-chip kbd { font-family:ui-monospace,SFMono-Regular,Menlo,Consolas,monospace; font-size:.74rem; border:1px solid #94a3b8; border-bottom-width:2px; border-radius:6px; padding:1px 6px; background:#f8fafc; color:#0f172a; }
    .epic-jump-highlight td { animation:epicJumpPulse 2s ease-out 1; }
    .sync-fetched-highlight td,
    .sync-fetched-highlight td * {
      background-color:#dbeafe !important;
      transition:background-color .18s ease;
    }
    .sync-saved-highlight td,
    .sync-saved-highlight td * {
      background-color:#dcfce7 !important;
      transition:background-color .18s ease;
    }
    @keyframes epicJumpPulse {
      0% { background:#fef3c7; }
      100% { background:#fff; }
    }
  </style>
</head>
<body>
  <main class="card">
    <div class="top">
      <div class="planner-page-title">
        <span class="material-symbols-outlined planner-title-icon" aria-hidden="true">event_note</span>
        <h1>Epics Planner</h1>
      </div>
      <div class="row" style="margin-top:0;">
        <button id="header-toggle-btn" class="btn alt" type="button" aria-expanded="true" aria-controls="planner-header-content">Collapse Header</button>
      </div>
    </div>
    <section id="planner-header" class="planner-header">
      <div id="planner-header-content" class="planner-header-content">
        <div class="row" style="margin-top:0;">__SETTINGS_TOP_NAV__</div>
        <div class="planner-header-bar">
          <button id="reload-btn" class="btn alt" type="button">Reload</button>
          <button id="add-epic-btn" class="btn" type="button">Add Epic</button>
          <button id="add-plan-column-btn" class="btn alt" type="button">Add Phase</button>
          <button id="manage-plan-columns-btn" class="btn alt" type="button">Manage Phases</button>
          <button id="expand-all-btn" class="btn alt" type="button">Expand</button>
          <button id="collapse-all-btn" class="btn alt" type="button">Collapse</button>
          <span class="shortcut-chip" title="Quick add epic"><kbd>Shift</kbd>+<kbd>Tab</kbd></span>
          <span class="planner-header-ipp" title="IPP Meeting Planner column marks epics for the IPP dashboard">IPP</span>
        </div>
      </div>
      <div id="status"></div>
    </section>
    <div class="table-wrap">
      <table>
        <thead>
          <tr id="epics-header-row"></tr>
        </thead>
        <tbody id="epics-tbody"></tbody>
      </table>
    </div>
  </main>

  <dialog id="plan-dialog">
    <div class="modal-head">
      <h2 id="plan-title" style="margin:0;font-size:1rem;">Edit Plan</h2>
      <div id="plan-context" class="muted" style="margin-top:4px;font-size:.8rem;"></div>
    </div>
    <div class="modal-body">
      <div>
        <label for="plan-mandays">Man-days</label>
        <input id="plan-mandays" type="number" min="0" step="0.5" placeholder="e.g. 12">
      </div>
      <div>
        <label for="plan-start">Start Date</label>
        <input id="plan-start" type="date">
      </div>
      <div>
        <label for="plan-due">Due Date</label>
        <input id="plan-due" type="date">
      </div>
    </div>
    <div class="modal-foot">
      <button id="plan-clear" class="btn alt small" type="button">Clear</button>
      <button id="plan-cancel" class="btn alt small" type="button">Cancel</button>
      <button id="plan-save" class="btn small" type="button">Save Plan</button>
    </div>
  </dialog>

  <dialog id="epic-dialog">
    <div class="modal-head">
      <h2 id="epic-dialog-title" style="margin:0;font-size:1rem;">Add New Epic</h2>
      <div id="epic-dialog-subtitle" class="muted" style="margin-top:4px;font-size:.8rem;">Provide epic metadata; you can edit detailed plans and IPP meeting selection after creation.</div>
    </div>
    <div class="modal-body">
      <div><label for="epic-project-select">Project</label><select id="epic-project-select"></select></div>
      <div><label for="epic-product-category">Product Categorization</label><select id="epic-product-category"></select></div>
      <div><label for="epic-component">Component</label><select id="epic-component"></select></div>
      <div><label for="epic-name">Epic Name</label><input id="epic-name" type="text"></div>
      <div><label for="epic-originator">Originator</label><input id="epic-originator" type="text"></div>
      <div><label for="epic-priority">Priority</label><select id="epic-priority"><option>Low</option><option>Medium</option><option>High</option><option>Highest</option></select></div>
      <div><label for="epic-plan-status">Plan Status</label><select id="epic-plan-status"><option>Planned</option><option>Not Planned Yet</option></select></div>
      <div><label for="epic-ipp-meeting-planned">IPP Meeting Planner</label><select id="epic-ipp-meeting-planned"><option>No</option><option>Yes</option></select></div>
      <div><label for="epic-actual-production-date">Actual Production Date</label><input id="epic-actual-production-date" type="date"></div>
      <div><label for="epic-remarks">Remarks</label><input id="epic-remarks" type="text" placeholder="Reason if production exceeded planned date"></div>
      <div><label for="epic-jira-url">Jira URL</label><input id="epic-jira-url" type="url" placeholder="https://..."></div>
      <div><label for="epic-description">Description</label><textarea id="epic-description"></textarea></div>
      <div><label for="epic-research-urs-plan-jira-url">Research/URS Plan Jira URL</label><input id="epic-research-urs-plan-jira-url" type="url" placeholder="https://..."></div>
      <div><label for="epic-dds-plan-jira-url">DDS Plan Jira URL</label><input id="epic-dds-plan-jira-url" type="url" placeholder="https://..."></div>
      <div><label for="epic-development-plan-jira-url">Development Plan Jira URL</label><input id="epic-development-plan-jira-url" type="url" placeholder="https://..."></div>
      <div><label for="epic-sqa-plan-jira-url">SQA Plan Jira URL</label><input id="epic-sqa-plan-jira-url" type="url" placeholder="https://..."></div>
      <div><label for="epic-user-manual-plan-jira-url">User Manual Plan Jira URL</label><input id="epic-user-manual-plan-jira-url" type="url" placeholder="https://..."></div>
      <div><label for="epic-production-plan-jira-url">Production Plan Jira URL</label><input id="epic-production-plan-jira-url" type="url" placeholder="https://..."></div>
      <div id="dynamic-plan-fields"></div>
    </div>
    <div class="modal-foot">
      <button id="epic-cancel" class="btn alt small" type="button">Cancel</button>
      <button id="epic-save" class="btn small" type="button">Create Epic</button>
    </div>
  </dialog>

  <dialog id="plan-column-dialog">
    <div class="modal-head">
      <h2 id="plan-column-title" style="margin:0;font-size:1rem;">Add Plan Column</h2>
      <div id="plan-column-context" class="muted" style="margin-top:4px;font-size:.8rem;">Create a dynamic plan column and choose where it should appear.</div>
    </div>
    <div class="modal-body">
      <div><label for="plan-column-name">Column Name</label><input id="plan-column-name" type="text" placeholder="e.g. Security Plan"></div>
      <div id="plan-column-restore-hint" class="muted" style="display:none;padding:8px;border:1px solid #e2e8f0;border-radius:8px;background:#f8fafc;"></div>
      <div><label for="plan-column-position">Insert Position</label><select id="plan-column-position"></select></div>
      <div><label for="plan-column-jira-enabled" style="display:flex;align-items:center;gap:8px;margin:0;"><input id="plan-column-jira-enabled" type="checkbox" style="width:auto;">Enable Jira URL support</label></div>
    </div>
    <div class="modal-foot">
      <button id="plan-column-cancel" class="btn alt small" type="button">Cancel</button>
      <button id="plan-column-save" class="btn small" type="button">Add Column</button>
    </div>
  </dialog>

  <dialog id="manage-columns-dialog">
    <div class="modal-head">
      <h2 style="margin:0;font-size:1rem;">Manage Columns</h2>
      <div class="muted" style="margin-top:4px;font-size:.8rem;">Dynamic columns can be deleted here. Default columns are locked.</div>
    </div>
    <div class="manage-columns-wrap">
      <table class="manage-columns-table">
        <thead>
          <tr>
            <th>#</th>
            <th>Column</th>
            <th>Key</th>
            <th>Jira URL</th>
            <th>Type</th>
            <th>Delete</th>
          </tr>
        </thead>
        <tbody id="manage-columns-tbody"></tbody>
      </table>
    </div>
    <div class="modal-foot">
      <button id="manage-columns-close" class="btn alt small" type="button">Close</button>
    </div>
  </dialog>

  <script>
    const API = "/api/epics-management/rows";
    const HEADER_COLLAPSE_STORAGE_KEY = "epics-management-header-collapsed";
    const PLAN_COLUMNS_API = "/api/epics-management/plan-columns";
    const PLAN_COLUMNS_DELETE_API_BASE = "/api/epics-management/plan-columns";
    const PLAN_COLUMNS_RESTORE_API_BASE = "/api/epics-management/plan-columns";
    const PLAN_COLUMNS_ORDER_API = "/api/epics-management/plan-columns/order";
    const PROJECTS_API = "/api/projects?include_inactive=0";
    const OPTIONS_API = "/api/epics-management/dropdown-options";
    const STORAGE_KEY = "epics-management-overrides-v1";
    const PRIORITY_OPTIONS = ["Low", "Medium", "High", "Highest"];
    let PLAN_STATUS_OPTIONS = ["Planned", "Not Planned Yet"];
    const IPP_MEETING_PLANNED_OPTIONS = ["No", "Yes"];
    const DEFAULT_PLAN_COLUMNS = [
      { key: "epic_plan", label: "Epic Plan", jira_link_enabled: false, is_default: true },
      { key: "research_urs_plan", label: "Research/URS Plan", jira_link_enabled: true, is_default: true },
      { key: "dds_plan", label: "DDS Plan", jira_link_enabled: true, is_default: true },
      { key: "development_plan", label: "Development Plan", jira_link_enabled: true, is_default: true },
      { key: "sqa_plan", label: "SQA Plan", jira_link_enabled: true, is_default: true },
      { key: "user_manual_plan", label: "User Manual Plan", jira_link_enabled: true, is_default: true },
      { key: "production_plan", label: "Production Plan", jira_link_enabled: true, is_default: true },
    ];
    const EPICS_STATIC_COL_MIN_WIDTH = 1420;
    const EPICS_PLAN_COL_MIN_WIDTH = 170;
    const EPICS_PLAN_COL_MAX_WIDTH = 260;
    const STATIC_PLAN_JIRA_INPUT_KEYS = new Set([
      "research_urs_plan",
      "dds_plan",
      "development_plan",
      "sqa_plan",
      "user_manual_plan",
      "production_plan",
    ]);
    let PLAN_COLUMNS = DEFAULT_PLAN_COLUMNS.slice();
    let PLAN_JIRA_COLUMN_KEYS = new Set(
      PLAN_COLUMNS.filter((item) => STATIC_PLAN_JIRA_INPUT_KEYS.has(item.key)).map((item) => item.key)
    );

    const headerRowEl = document.getElementById("epics-header-row");
    const tbodyEl = document.getElementById("epics-tbody");
    const statusEl = document.getElementById("status");
    const planDialogEl = document.getElementById("plan-dialog");
    const planTitleEl = document.getElementById("plan-title");
    const planContextEl = document.getElementById("plan-context");
    const planMandaysEl = document.getElementById("plan-mandays");
    const planStartEl = document.getElementById("plan-start");
    const planDueEl = document.getElementById("plan-due");
    const epicDialogEl = document.getElementById("epic-dialog");
    const epicDialogTitleEl = document.getElementById("epic-dialog-title");
    const epicDialogSubtitleEl = document.getElementById("epic-dialog-subtitle");
    const epicProjectSelectEl = document.getElementById("epic-project-select");
    const epicProductCategoryEl = document.getElementById("epic-product-category");
    const epicComponentEl = document.getElementById("epic-component");
    const epicNameEl = document.getElementById("epic-name");
    const epicOriginatorEl = document.getElementById("epic-originator");
    const epicPriorityEl = document.getElementById("epic-priority");
    const epicPlanStatusEl = document.getElementById("epic-plan-status");
    const epicIppMeetingPlannedEl = document.getElementById("epic-ipp-meeting-planned");
    const epicActualProductionDateEl = document.getElementById("epic-actual-production-date");
    const epicRemarksEl = document.getElementById("epic-remarks");
    const epicJiraUrlEl = document.getElementById("epic-jira-url");
    const epicDescriptionEl = document.getElementById("epic-description");
    const epicResearchUrsPlanJiraUrlEl = document.getElementById("epic-research-urs-plan-jira-url");
    const epicDdsPlanJiraUrlEl = document.getElementById("epic-dds-plan-jira-url");
    const epicDevelopmentPlanJiraUrlEl = document.getElementById("epic-development-plan-jira-url");
    const epicSqaPlanJiraUrlEl = document.getElementById("epic-sqa-plan-jira-url");
    const epicUserManualPlanJiraUrlEl = document.getElementById("epic-user-manual-plan-jira-url");
    const epicProductionPlanJiraUrlEl = document.getElementById("epic-production-plan-jira-url");
    const dynamicPlanFieldsEl = document.getElementById("dynamic-plan-fields");
    const planColumnDialogEl = document.getElementById("plan-column-dialog");
    const planColumnNameEl = document.getElementById("plan-column-name");
    const planColumnPositionEl = document.getElementById("plan-column-position");
    const planColumnJiraEnabledEl = document.getElementById("plan-column-jira-enabled");
    const planColumnRestoreHintEl = document.getElementById("plan-column-restore-hint");
    const plannerHeaderEl = document.getElementById("planner-header");
    const plannerHeaderToggleBtn = document.getElementById("header-toggle-btn");
    const manageColumnsDialogEl = document.getElementById("manage-columns-dialog");
    const manageColumnsTbodyEl = document.getElementById("manage-columns-tbody");

    let rows = [];
    let managedProjects = [];
    let dropdownOptions = { product_category_options: [], component_options: [], plan_status_options: PLAN_STATUS_OPTIONS.slice() };
    let overrides = {};
    let activePlan = { rowIndex: -1, planKey: "" };
    let activeEpicEditKey = "";
    let activePlanInsertPosition = 0;
    let planDragKey = "";
    let draftEpicRow = null;
    let draftEpicCreateInFlight = false;
    let dynamicPlanInputEls = {};
    let planColumnsCatalog = [];
    const expandedProjects = new Set();
    const expandedCategories = new Set();
    const expandedComponents = new Set();
    const expandedPlanDetails = new Set();
    const autoSaveInFlight = new Map();
    const autoSaveQueued = new Set();
    const deepLinkEpicKey = (() => {
      try {
        const params = new URLSearchParams(window.location.search || "");
        return String(params.get("epic_key") || "").trim().toUpperCase();
      } catch (_) {
        return "";
      }
    })();
    let deepLinkHandled = false;
    let deepLinkMissingWarningShown = false;

    function esc(value) {
      return String(value == null ? "" : value).replace(/&/g, "&amp;").replace(/</g, "&lt;").replace(/>/g, "&gt;");
    }
    function setStatus(message, kind) {
      statusEl.textContent = String(message || "");
      statusEl.className = kind || "";
    }
    function setHeaderCollapsed(collapsed) {
      if (!plannerHeaderEl || !plannerHeaderToggleBtn) return;
      plannerHeaderEl.classList.toggle("collapsed", !!collapsed);
      plannerHeaderToggleBtn.textContent = collapsed ? "Expand Header" : "Collapse Header";
      plannerHeaderToggleBtn.setAttribute("aria-expanded", String(!collapsed));
    }
    function initHeaderToggle() {
      if (!plannerHeaderEl || !plannerHeaderToggleBtn) return;
      const storedCollapsed = localStorage.getItem(HEADER_COLLAPSE_STORAGE_KEY) === "1";
      setHeaderCollapsed(storedCollapsed);
      plannerHeaderToggleBtn.addEventListener("click", () => {
        const nextCollapsed = !plannerHeaderEl.classList.contains("collapsed");
        setHeaderCollapsed(nextCollapsed);
        localStorage.setItem(HEADER_COLLAPSE_STORAGE_KEY, nextCollapsed ? "1" : "0");
      });
    }
    function planColumnTrashIconSvg() {
      return '<svg viewBox="0 0 24 24" aria-hidden="true"><path fill="currentColor" d="M9 3h6l1 2h4v2H4V5h4l1-2zm1 6h2v8h-2V9zm4 0h2v8h-2V9zM7 9h2v8H7V9z"/></svg>';
    }
    function applyPlanColumnLayout() {
      const planCount = Math.max(PLAN_COLUMNS.length, 1);
      const viewportWidth = Math.max(window.innerWidth || 0, 1024);
      const availableForPlans = Math.max(viewportWidth - 240 - EPICS_STATIC_COL_MIN_WIDTH, EPICS_PLAN_COL_MIN_WIDTH * planCount);
      const computedPlanWidth = Math.max(
        EPICS_PLAN_COL_MIN_WIDTH,
        Math.min(EPICS_PLAN_COL_MAX_WIDTH, Math.floor(availableForPlans / planCount))
      );
      const tableMinWidth = EPICS_STATIC_COL_MIN_WIDTH + (computedPlanWidth * planCount);
      document.documentElement.style.setProperty("--epics-plan-col-min-width", String(computedPlanWidth) + "px");
      document.documentElement.style.setProperty("--epics-table-min-width", String(tableMinWidth) + "px");
    }
    function renderManageColumnsTable() {
      if (!manageColumnsTbodyEl) return;
      const html = PLAN_COLUMNS.map((col, index) => {
        const key = String(col && col.key || "").trim();
        const label = String(col && col.label || key).trim();
        const isDefault = !!(col && col.is_default);
        const jiraEnabled = !!(col && col.jira_link_enabled);
        const actionHtml = isDefault
          ? '<span style="display:inline-flex;align-items:center;gap:6px;"><button class="icon-btn danger" type="button" disabled title="Default column is locked">' + planColumnTrashIconSvg() + '</button><span class="muted">Locked</span></span>'
          : '<button class="icon-btn danger" type="button" data-delete-plan-key="' + esc(key) + '" data-delete-plan-label="' + esc(label) + '" title="Delete column">' + planColumnTrashIconSvg() + "</button>";
        return ""
          + "<tr>"
          + "<td>" + String(index + 1) + "</td>"
          + "<td>" + esc(label) + "</td>"
          + "<td><code>" + esc(key) + "</code></td>"
          + "<td>" + (jiraEnabled ? "Enabled" : "Disabled") + "</td>"
          + "<td>" + (isDefault ? "Default" : "Dynamic") + "</td>"
          + "<td>" + actionHtml + "</td>"
          + "</tr>";
      }).join("");
      manageColumnsTbodyEl.innerHTML = html || '<tr><td colspan="6" class="muted">No plan columns.</td></tr>';
    }
    async function deletePlanColumn(columnKey, columnLabel) {
      const key = String(columnKey || "").trim();
      if (!key) throw new Error("Plan column key is required.");
      const column = PLAN_COLUMNS.find((item) => String(item && item.key || "") === key);
      if (column && column.is_default) {
        throw new Error("Default plan columns cannot be deleted.");
      }
      const label = String(columnLabel || (column && (column.label || column.key)) || key);
      if (!window.confirm('Delete plan column "' + label + '"? This removes it from Epics Planner.')) return;
      const resp = await fetch(PLAN_COLUMNS_DELETE_API_BASE + "/" + encodeURIComponent(key), { method: "DELETE" });
      const body = await resp.json().catch(() => ({}));
      if (!resp.ok) throw new Error(String(body.error || "Failed to delete plan column."));
      await Promise.all([loadPlanColumns(), loadPlanColumnsCatalog(), loadRowsFromApi()]);
      renderManageColumnsTable();
      setStatus('Plan column deleted: ' + label, "ok");
    }
    async function loadPlanColumnsCatalog() {
      const resp = await fetch(PLAN_COLUMNS_API + "?include_inactive=1", { cache: "no-store" });
      const body = await resp.json().catch(() => ({}));
      if (!resp.ok) throw new Error(String(body.error || "Failed to load plan columns catalog."));
      planColumnsCatalog = Array.isArray(body.columns) ? body.columns.map((item) => ({
        key: String(item && item.key || "").trim(),
        label: String(item && item.label || item && item.key || "").trim(),
        is_active: !!(item && item.is_active),
      })) : [];
    }
    function refreshPlanColumnRestoreHint() {
      const label = String(planColumnNameEl.value || "").trim();
      if (!label) {
        planColumnRestoreHintEl.style.display = "none";
        planColumnRestoreHintEl.innerHTML = "";
        return;
      }
      const deleted = planColumnsCatalog.find((item) =>
        !item.is_active && String(item.label || "").toLowerCase() === label.toLowerCase()
      );
      if (!deleted) {
        planColumnRestoreHintEl.style.display = "none";
        planColumnRestoreHintEl.innerHTML = "";
        return;
      }
      planColumnRestoreHintEl.style.display = "block";
      planColumnRestoreHintEl.innerHTML =
        'A deleted column with this name exists (<code>' + esc(deleted.key) + '</code>). '
        + '<button id="plan-column-restore-btn" class="btn alt small" type="button" style="margin-left:8px;">Use Existing Deleted Column</button>';
      const restoreBtn = document.getElementById("plan-column-restore-btn");
      if (restoreBtn) {
        restoreBtn.addEventListener("click", () => {
          restoreDeletedPlanColumn(deleted.key).catch((err) => setStatus(err.message || String(err), "warn"));
        });
      }
    }
    async function restoreDeletedPlanColumn(columnKey) {
      const key = String(columnKey || "").trim();
      if (!key) throw new Error("Column key is required for restore.");
      const resp = await fetch(PLAN_COLUMNS_RESTORE_API_BASE + "/" + encodeURIComponent(key) + "/restore", {
        method: "POST",
      });
      const body = await resp.json().catch(() => ({}));
      if (!resp.ok) throw new Error(String(body.error || "Failed to restore column."));
      planColumnDialogEl.close();
      await Promise.all([loadPlanColumns(), loadPlanColumnsCatalog(), loadRowsFromApi()]);
      setStatus("Plan column restored: " + (body.column && body.column.label ? body.column.label : key), "ok");
    }
    function formatPlanPositionLabel(index) {
      if (!PLAN_COLUMNS.length) return "1 (first)";
      if (index <= 0) return "1 (before " + String(PLAN_COLUMNS[0].label || PLAN_COLUMNS[0].key) + ")";
      if (index >= PLAN_COLUMNS.length) return String(PLAN_COLUMNS.length + 1) + " (after " + String(PLAN_COLUMNS[PLAN_COLUMNS.length - 1].label || PLAN_COLUMNS[PLAN_COLUMNS.length - 1].key) + ")";
      return String(index + 1) + " (between " + String(PLAN_COLUMNS[index - 1].label || PLAN_COLUMNS[index - 1].key) + " and " + String(PLAN_COLUMNS[index].label || PLAN_COLUMNS[index].key) + ")";
    }
    function renderPlanColumnPositionOptions(selectedPosition) {
      const options = [];
      const totalSlots = PLAN_COLUMNS.length + 1;
      const selected = Math.max(1, Math.min(totalSlots, Number(selectedPosition) || totalSlots));
      for (let slot = 1; slot <= totalSlots; slot += 1) {
        options.push('<option value="' + slot + '"' + (slot === selected ? " selected" : "") + ">" + esc(formatPlanPositionLabel(slot - 1)) + "</option>");
      }
      planColumnPositionEl.innerHTML = options.join("");
      planColumnPositionEl.value = String(selected);
    }
    function openPlanColumnDialog(insertPosition) {
      activePlanInsertPosition = Number(insertPosition) || (PLAN_COLUMNS.length + 1);
      planColumnNameEl.value = "";
      planColumnJiraEnabledEl.checked = false;
      renderPlanColumnPositionOptions(activePlanInsertPosition);
      refreshPlanColumnRestoreHint();
      planColumnDialogEl.showModal();
      window.setTimeout(() => planColumnNameEl.focus(), 0);
    }
    async function savePlanColumnFromDialog() {
      const label = String(planColumnNameEl.value || "").trim();
      if (!label) throw new Error("Column name is required.");
      const insertPosition = Number(planColumnPositionEl.value || "0");
      const jiraEnabled = !!planColumnJiraEnabledEl.checked;
      const resp = await fetch(PLAN_COLUMNS_API, {
        method: "POST",
        headers: { "Content-Type": "application/json" },
        body: JSON.stringify({ label, jira_link_enabled: jiraEnabled, insert_position: insertPosition }),
      });
      const body = await resp.json().catch(() => ({}));
      if (!resp.ok) throw new Error(String(body.error || "Failed to add plan column."));
      planColumnDialogEl.close();
      await Promise.all([loadPlanColumns(), loadPlanColumnsCatalog(), loadRowsFromApi()]);
      setStatus("Plan column added: " + label, "ok");
    }
    async function persistPlanColumnOrder(orderedColumns) {
      const orderedKeys = orderedColumns.map((item) => String(item.key || "")).filter(Boolean);
      const resp = await fetch(PLAN_COLUMNS_ORDER_API, {
        method: "PUT",
        headers: { "Content-Type": "application/json" },
        body: JSON.stringify({ ordered_keys: orderedKeys }),
      });
      const body = await resp.json().catch(() => ({}));
      if (!resp.ok) throw new Error(String(body.error || "Failed to reorder plan columns."));
      const cols = Array.isArray(body.columns) ? body.columns : [];
      PLAN_COLUMNS = cols.map((item) => ({
        key: String(item && item.key || "").trim(),
        label: String(item && item.label || item && item.key || "").trim(),
        jira_link_enabled: !!(item && item.jira_link_enabled),
        is_default: !!(item && item.is_default),
      })).filter((item) => item.key);
      PLAN_JIRA_COLUMN_KEYS = new Set(
        PLAN_COLUMNS.filter((item) => item && item.jira_link_enabled).map((item) => item.key)
      );
      renderPlanHeaders();
      renderDynamicPlanJiraFields();
      renderManageColumnsTable();
      renderTable();
    }
    function renderPlanHeaders() {
      const planHeaders = PLAN_COLUMNS.map((col, idx) => {
        const deleteBtn = !col.is_default
          ? '<button class="icon-btn danger" type="button" data-delete-plan-key="' + esc(col.key) + '" data-delete-plan-label="' + esc(col.label || col.key || "Plan") + '" title="Delete column">' + planColumnTrashIconSvg() + "</button>"
          : '<button class="icon-btn danger" type="button" disabled title="Default column is locked">' + planColumnTrashIconSvg() + "</button>";
        const insertBeforeBtn = '<button class="plan-insert-handle" type="button" data-insert-position="' + (idx + 1) + '" title="Add column here">+</button>';
        const insertAfterBtn = idx === PLAN_COLUMNS.length - 1
          ? '<button class="plan-insert-handle after" type="button" data-insert-position="' + (PLAN_COLUMNS.length + 1) + '" title="Add column here">+</button>'
          : "";
        return '<th class="plan-col-head" draggable="true" data-plan-key="' + esc(col.key) + '" data-plan-index="' + idx + '"><div class="plan-head-wrap">' + insertBeforeBtn + '<span class="plan-label">' + esc(col.label || col.key || "Plan") + '</span><span class="plan-head-actions">' + deleteBtn + '</span>' + insertAfterBtn + "</div></th>";
      }).join("");
      headerRowEl.innerHTML = ""
        + "<th>Project</th>"
        + "<th>Product Categorization</th>"
        + "<th>Component</th>"
        + "<th>Epic</th>"
        + "<th>Description</th>"
        + "<th>Originator</th>"
        + "<th>Priority</th>"
        + "<th>Plan Status</th>"
        + "<th>IPP Meeting Planner</th>"
        + "<th>Actual Production Date</th>"
        + "<th>Remarks</th>"
        + planHeaders
        + "<th>Actions</th>";
      applyPlanColumnLayout();
      Array.from(headerRowEl.querySelectorAll("button.plan-insert-handle[data-insert-position]")).forEach((btn) => {
        btn.addEventListener("click", (evt) => {
          evt.preventDefault();
          evt.stopPropagation();
          const insertPosition = Number(btn.getAttribute("data-insert-position") || "0");
          openPlanColumnDialog(insertPosition);
        });
      });
      Array.from(headerRowEl.querySelectorAll("button[data-delete-plan-key]")).forEach((btn) => {
        btn.addEventListener("click", (evt) => {
          evt.preventDefault();
          evt.stopPropagation();
          const key = String(btn.getAttribute("data-delete-plan-key") || "");
          const label = String(btn.getAttribute("data-delete-plan-label") || key);
          deletePlanColumn(key, label).catch((err) => setStatus(err.message || String(err), "warn"));
        });
      });
      const planHeadersEls = Array.from(headerRowEl.querySelectorAll("th.plan-col-head[data-plan-key]"));
      planHeadersEls.forEach((thEl) => {
        thEl.addEventListener("dragstart", (evt) => {
          planDragKey = String(thEl.getAttribute("data-plan-key") || "");
          thEl.classList.add("dragging");
          if (evt.dataTransfer) {
            evt.dataTransfer.effectAllowed = "move";
            evt.dataTransfer.setData("text/plain", planDragKey);
          }
        });
        thEl.addEventListener("dragover", (evt) => {
          if (!planDragKey) return;
          evt.preventDefault();
          thEl.classList.add("drop-target");
        });
        thEl.addEventListener("dragleave", () => thEl.classList.remove("drop-target"));
        thEl.addEventListener("drop", (evt) => {
          evt.preventDefault();
          const targetKey = String(thEl.getAttribute("data-plan-key") || "");
          const sourceKey = planDragKey || (evt.dataTransfer ? evt.dataTransfer.getData("text/plain") : "");
          planHeadersEls.forEach((item) => item.classList.remove("drop-target", "dragging"));
          planDragKey = "";
          if (!sourceKey || !targetKey || sourceKey === targetKey) return;
          const current = PLAN_COLUMNS.slice();
          const fromIndex = current.findIndex((item) => item.key === sourceKey);
          const toIndex = current.findIndex((item) => item.key === targetKey);
          if (fromIndex < 0 || toIndex < 0) return;
          const next = current.slice();
          const moved = next.splice(fromIndex, 1)[0];
          next.splice(toIndex, 0, moved);
          PLAN_COLUMNS = next;
          renderPlanHeaders();
          renderDynamicPlanJiraFields();
          renderTable();
          persistPlanColumnOrder(next).catch(async (err) => {
            setStatus(err.message || String(err), "warn");
            await loadPlanColumns();
            renderTable();
          });
        });
        thEl.addEventListener("dragend", () => {
          planDragKey = "";
          planHeadersEls.forEach((item) => item.classList.remove("drop-target", "dragging"));
        });
      });
    }
    function renderDynamicPlanJiraFields() {
      const html = [];
      const nextMap = {};
      for (const col of PLAN_COLUMNS) {
        const key = String(col.key || "").trim();
        if (!key || STATIC_PLAN_JIRA_INPUT_KEYS.has(key) || key === "epic_plan") continue;
        const jiraEnabled = PLAN_JIRA_COLUMN_KEYS.has(key);
        if (!jiraEnabled) continue;
        const inputBase = "epic-dynamic-plan-" + key;
        html.push(
          ''
          + '<div style="border:1px solid #e2e8f0;border-radius:10px;padding:10px;background:#f8fbff;">'
          + '  <div style="font-weight:700;margin-bottom:8px;color:#1e3a8a;">' + esc(String(col.label || key)) + '</div>'
          + '  <div style="display:grid;gap:8px;grid-template-columns:1fr;">'
          + '    <div><label for="' + esc(inputBase + '-jira') + '">Jira URL</label><input id="' + esc(inputBase + '-jira') + '" type="url" placeholder="https://..."></div>'
          + "  </div>"
          + "</div>"
        );
      }
      dynamicPlanFieldsEl.innerHTML = html.join("");
      for (const col of PLAN_COLUMNS) {
        const key = String(col.key || "").trim();
        if (!key || STATIC_PLAN_JIRA_INPUT_KEYS.has(key) || key === "epic_plan") continue;
        if (!PLAN_JIRA_COLUMN_KEYS.has(key)) continue;
        const inputBase = "epic-dynamic-plan-" + key;
        const jiraEl = document.getElementById(inputBase + "-jira");
        if (jiraEl) {
          nextMap[key] = jiraEl;
        }
      }
      dynamicPlanInputEls = nextMap;
    }
    async function loadPlanColumns() {
      const resp = await fetch(PLAN_COLUMNS_API, { cache: "no-store" });
      const body = await resp.json().catch(() => ({}));
      if (!resp.ok) throw new Error(String(body.error || "Failed to load plan columns."));
      const cols = Array.isArray(body.columns) ? body.columns : [];
      PLAN_COLUMNS = cols.length
        ? cols.map((item) => ({
            key: String(item && item.key || "").trim(),
            label: String(item && item.label || item && item.key || "").trim(),
            jira_link_enabled: !!(item && item.jira_link_enabled),
            is_default: !!(item && item.is_default),
          })).filter((item) => item.key)
        : DEFAULT_PLAN_COLUMNS.slice();
      PLAN_JIRA_COLUMN_KEYS = new Set(
        PLAN_COLUMNS.filter((item) => item && item.jira_link_enabled).map((item) => item.key)
      );
      renderPlanHeaders();
      renderDynamicPlanJiraFields();
      renderManageColumnsTable();
      applyPlanColumnLayout();
    }
    function deepClone(input) {
      return JSON.parse(JSON.stringify(input));
    }
    function loadOverrides() {
      try {
        const parsed = JSON.parse(localStorage.getItem(STORAGE_KEY) || "{}");
        return parsed && typeof parsed === "object" ? parsed : {};
      } catch (_) {
        return {};
      }
    }
    function saveOverrides() {
      localStorage.setItem(STORAGE_KEY, JSON.stringify(overrides || {}));
    }
    function mergeRow(base) {
      const id = String(base && base.id || "");
      const ov = id ? (overrides[id] || {}) : {};
      const plans = Object.assign({}, base && base.plans || {}, ov && ov.plans || {});
      const merged = Object.assign({}, base || {}, ov || {});
      merged.plans = plans;
      return merged;
    }
    function ensureRowOverride(row) {
      const id = String(row && row.id || "");
      if (!id) return {};
      if (!overrides[id] || typeof overrides[id] !== "object") overrides[id] = {};
      return overrides[id];
    }
    function normalizePriority(value) {
      const text = String(value || "").trim();
      return PRIORITY_OPTIONS.includes(text) ? text : "Low";
    }
    function defaultPlanStatus() {
      if (PLAN_STATUS_OPTIONS.includes("Not Planned Yet")) return "Not Planned Yet";
      return PLAN_STATUS_OPTIONS[0] || "Not Planned Yet";
    }
    function normalizePlanStatus(value) {
      const text = String(value || "").trim();
      if (text.toLowerCase() === "plan") return "Not Planned Yet";
      return PLAN_STATUS_OPTIONS.includes(text) ? text : defaultPlanStatus();
    }
    function normalizeIppMeetingPlanned(value) {
      const text = String(value || "").trim();
      return IPP_MEETING_PLANNED_OPTIONS.includes(text) ? text : "No";
    }
    function validateJiraUrl(value) {
      const text = String(value || "").trim();
      if (!text) return "";
      if (!/^https?:\\/\\//i.test(text)) throw new Error("Jira URL must start with http:// or https://");
      return text;
    }
    function normalizeEpicKey(value) {
      const text = String(value || "").trim().toUpperCase();
      if (!text) throw new Error("Epic key is required.");
      if (!/^[A-Z0-9]+-\\d+$/.test(text)) throw new Error("Epic key must look like ABC-123.");
      return text;
    }
    function normalizeEpicKeyAllowTmp(value) {
      const text = String(value || "").trim().toUpperCase();
      if (!text) throw new Error("Epic key is required.");
      const jiraStyle = /^[A-Z0-9]+-\\d+$/;
      const tmpStyle = /^TMP-\\d{8}T\\d{6}Z-[A-Z0-9]{6}$/;
      if (jiraStyle.test(text) || tmpStyle.test(text)) return text;
      throw new Error("Epic key must look like ABC-123 or TMP-YYYYMMDDTHHMMSSZ-XXXXXX.");
    }
    function epicKeyFromJiraUrl(url) {
      const text = String(url || "").trim();
      const match = /\\/browse\\/([A-Za-z0-9_-]+-\\d+)/.exec(text);
      return match ? normalizeEpicKey(match[1]) : "";
    }
    function normalizeProjectKey(value) {
      const text = String(value || "").trim().toUpperCase();
      if (!text) return "";
      if (!/^[A-Z0-9_-]+$/.test(text)) return "";
      return text;
    }
    function projectKeyFromEpicKey(epicKey) {
      const text = String(epicKey || "").trim().toUpperCase();
      if (!text) return "";
      const dash = text.indexOf("-");
      const candidate = dash > 0 ? text.slice(0, dash) : text;
      return normalizeProjectKey(candidate);
    }
    function uniqueNonEmptyOptions(values) {
      const out = [];
      const seen = new Set();
      for (const item of Array.isArray(values) ? values : []) {
        const value = String(item == null ? "" : item).trim();
        if (!value) continue;
        const lower = value.toLowerCase();
        if (seen.has(lower)) continue;
        seen.add(lower);
        out.push(value);
      }
      return out;
    }
    function setDropdownSelectOptions(selectEl, configuredValues, selectedValue) {
      if (!selectEl) return;
      const selected = String(selectedValue || "").trim();
      const values = uniqueNonEmptyOptions(configuredValues);
      if (selected && !values.some((item) => item.toLowerCase() === selected.toLowerCase())) {
        values.push(selected);
      }
      const options = ['<option value="">Select</option>'];
      for (const item of values) {
        const isSelected = selected && item.toLowerCase() === selected.toLowerCase();
        options.push('<option value="' + esc(item) + '"' + (isSelected ? " selected" : "") + ">" + esc(item) + "</option>");
      }
      selectEl.innerHTML = options.join("");
      if (selected) selectEl.value = selected;
    }
    function renderCategorizationSelect(field, configuredValues, selectedValue, rowIndex) {
      const selected = String(selectedValue || "").trim();
      const values = uniqueNonEmptyOptions(configuredValues);
      if (selected && !values.some((item) => item.toLowerCase() === selected.toLowerCase())) {
        values.push(selected);
      }
      const options = ['<option value="">Select</option>'];
      for (const item of values) {
        const isSelected = selected && item.toLowerCase() === selected.toLowerCase();
        options.push('<option value="' + esc(item) + '"' + (isSelected ? " selected" : "") + ">" + esc(item) + "</option>");
      }
      return '<select data-row-index="' + rowIndex + '" data-field="' + esc(field) + '">' + options.join("") + "</select>";
    }
    function renderDraftCategorizationSelect(field, configuredValues, selectedValue) {
      const selected = String(selectedValue || "").trim();
      const values = uniqueNonEmptyOptions(configuredValues);
      if (selected && !values.some((item) => item.toLowerCase() === selected.toLowerCase())) {
        values.push(selected);
      }
      const options = ['<option value="">Select</option>'];
      for (const item of values) {
        const isSelected = selected && item.toLowerCase() === selected.toLowerCase();
        options.push('<option value="' + esc(item) + '"' + (isSelected ? " selected" : "") + ">" + esc(item) + "</option>");
      }
      return '<select class="draft-input" data-draft-field="' + esc(field) + '">' + options.join("") + "</select>";
    }
    function renderDraftProjectSelect(selectedKey) {
      const selected = normalizeProjectKey(selectedKey);
      const options = ['<option value="">Select project</option>'];
      for (const item of managedProjects) {
        const key = normalizeProjectKey(item.project_key);
        const name = String(item.display_name || item.project_name || key || "").trim();
        const label = name + (key ? " (" + key + ")" : "");
        options.push('<option value="' + esc(key) + '" data-project-name="' + esc(name) + '"' + (key === selected ? " selected" : "") + ">" + esc(label) + "</option>");
      }
      if (selected && !managedProjects.some((p) => normalizeProjectKey(p.project_key) === selected)) {
        options.push('<option value="' + esc(selected) + '" data-project-name="' + esc(selected) + '" selected>' + esc(selected + " (Unavailable)") + "</option>");
      }
      return '<select class="draft-input draft-project-select" data-draft-field="project_key">' + options.join("") + "</select>";
    }
    async function loadDropdownOptions() {
      const resp = await fetch(OPTIONS_API, { cache: "no-store" });
      const body = await resp.json().catch(() => ({}));
      if (!resp.ok) throw new Error(String(body.error || "Failed to load dropdown options."));
      dropdownOptions = {
        product_category_options: uniqueNonEmptyOptions(body.product_category_options || []),
        component_options: uniqueNonEmptyOptions(body.component_options || []),
        plan_status_options: uniqueNonEmptyOptions(body.plan_status_options || []),
      };
      PLAN_STATUS_OPTIONS = dropdownOptions.plan_status_options.length
        ? dropdownOptions.plan_status_options.slice()
        : ["Planned", "Not Planned Yet"];
      setDropdownSelectOptions(epicProductCategoryEl, dropdownOptions.product_category_options, epicProductCategoryEl.value);
      setDropdownSelectOptions(epicComponentEl, dropdownOptions.component_options, epicComponentEl.value);
      setDropdownSelectOptions(epicPlanStatusEl, PLAN_STATUS_OPTIONS, normalizePlanStatus(epicPlanStatusEl.value));
      if (rows.length) renderTable();
    }
    async function addPlanColumn() {
      openPlanColumnDialog(PLAN_COLUMNS.length + 1);
    }
    function selectedManagedProject() {
      const selectedKey = normalizeProjectKey(epicProjectSelectEl.value);
      return managedProjects.find((p) => normalizeProjectKey(p.project_key) === selectedKey) || null;
    }
    function renderProjectOptions(selectedKey) {
      const selected = normalizeProjectKey(selectedKey);
      const options = ['<option value="">Select managed project</option>'];
      for (const item of managedProjects) {
        const key = normalizeProjectKey(item.project_key);
        const name = String(item.display_name || item.project_name || key || "").trim();
        const label = name + (key ? " (" + key + ")" : "");
        options.push('<option value="' + esc(key) + '"' + (key === selected ? " selected" : "") + ">" + esc(label) + "</option>");
      }
      if (selected && !managedProjects.some((p) => normalizeProjectKey(p.project_key) === selected)) {
        options.push('<option value="' + esc(selected) + '" selected>' + esc(selected + " (Unavailable)") + "</option>");
      }
      epicProjectSelectEl.innerHTML = options.join("");
    }
    async function loadManagedProjects() {
      const resp = await fetch(PROJECTS_API, { cache: "no-store" });
      const body = await resp.json().catch(() => ({}));
      if (!resp.ok) throw new Error(String(body.error || "Failed to load managed projects."));
      managedProjects = Array.isArray(body.projects) ? body.projects : [];
      renderProjectOptions(epicProjectSelectEl.value);
    }
    function toDateValue(value) {
      const text = String(value || "").trim();
      if (!text) return "";
      return /^\\d{4}-\\d{2}-\\d{2}$/.test(text) ? text : "";
    }
    function formatDateDisplay(value) {
      const dateValue = toDateValue(value);
      if (!dateValue) return "-";
      const date = new Date(dateValue + "T00:00:00");
      if (Number.isNaN(date.getTime())) return "-";
      const day = String(date.getDate()).padStart(2, "0");
      const month = date.toLocaleString("en-GB", { month: "short" });
      const year = String(date.getFullYear()).slice(-2);
      return day + " " + month + " " + year;
    }
    function parseManDaysValue(value) {
      if (value == null || value === "") return null;
      const num = Number(value);
      return Number.isFinite(num) ? num : null;
    }
    function formatManDaysValue(value) {
      const rounded = Math.round((Number(value) || 0) * 100) / 100;
      if (Number.isInteger(rounded)) return String(rounded);
      return String(rounded);
    }
    const GROUP_TOTAL_PLAN_KEY = "epic_plan";
    function computeGroupManDaysTotals(rowIndexes) {
      const totals = {};
      PLAN_COLUMNS.forEach((col) => { totals[col.key] = 0; });
      for (const rowIndex of rowIndexes) {
        const row = rows[rowIndex];
        if (!row) continue;
        for (const col of PLAN_COLUMNS) {
          const plan = ((row.plans || {})[col.key]) || {};
          const manDays = parseManDaysValue(plan.man_days);
          if (manDays == null) continue;
          totals[col.key] += manDays;
        }
      }
      const overall = totals[GROUP_TOTAL_PLAN_KEY] || 0;
      return { totals, overall };
    }
    function renderGroupPlanTotalCells(groupTotals) {
      return PLAN_COLUMNS.map((col) => {
        const total = groupTotals && groupTotals.totals ? groupTotals.totals[col.key] : 0;
        return '<td class="plan-col-cell"><div class="group-plan-total"><b>' + esc(formatManDaysValue(total)) + '</b><span>md</span></div></td>';
      }).join("");
    }
    function planCellStateKey(rowIndex, planKey) {
      return String(rowIndex) + "::" + String(planKey || "");
    }
    function planSummary(plan, rowIndex, planKey) {
      const hasManDays = !(plan == null || plan.man_days == null || String(plan.man_days) === "");
      if (!hasManDays) {
        return '<span class="plan-empty">Set plan details</span>';
      }
      const manDays = String(plan.man_days);
      const startIso = toDateValue(plan && plan.start_date);
      const dueIso = toDateValue(plan && plan.due_date);
      const hasAnyDate = !!(startIso || dueIso);
      const stateKey = planCellStateKey(rowIndex, planKey);
      const isExpanded = expandedPlanDetails.has(stateKey);
      const toggle = hasAnyDate
        ? '<button class="plan-toggle" type="button" data-plan-toggle="' + esc(stateKey) + '" aria-label="Toggle dates">' + (isExpanded ? "▾" : "▸") + "</button>"
        : "";
      const dates = hasAnyDate && isExpanded
        ? '<div class="plan-dates">'
            + (startIso ? '<div><b>Start:</b> ' + esc(formatDateDisplay(startIso)) + "</div>" : "")
            + (dueIso ? '<div><b>Due:</b> ' + esc(formatDateDisplay(dueIso)) + "</div>" : "")
          + "</div>"
        : "";
      return '<div class="plan-summary"><div class="plan-main"><div><b>Man-days:</b> ' + esc(manDays) + "</div>" + toggle + "</div>" + dates + "</div>";
    }
    function isPlanJiraEnabled(planKey) {
      return PLAN_JIRA_COLUMN_KEYS.has(String(planKey || ""));
    }
    function planJiraUrl(plan) {
      const candidate = String(plan && plan.jira_url || "").trim();
      if (!candidate) return "";
      try {
        return validateJiraUrl(candidate);
      } catch (_) {
        return "";
      }
    }
    function renderPlanCell(rowIndex, planCol, row) {
      const plan = (row.plans || {})[planCol.key] || {};
      const summary = planSummary(plan, rowIndex, planCol.key);
      if (!isPlanJiraEnabled(planCol.key)) {
        return '<td class="plan-col-cell"><button class="plan-btn" type="button" data-row-index="' + rowIndex + '" data-plan-key="' + esc(planCol.key) + '">' + summary + "</button></td>";
      }
      const jiraUrl = planJiraUrl(plan);
      const hasJira = !!jiraUrl;
      return ''
        + '<td class="plan-col-cell">'
        + '  <div class="plan-cell">'
        + '    <button class="plan-btn" type="button" data-row-index="' + rowIndex + '" data-plan-key="' + esc(planCol.key) + '">' + summary + "</button>"
        + '    <div class="plan-cell-actions">'
        + '      <a class="jira-open ' + (hasJira ? "" : "disabled") + '" href="' + esc(hasJira ? jiraUrl : "#") + '" target="_blank" rel="noopener noreferrer" title="' + (hasJira ? "Open Jira link" : "No Jira link set") + '">J</a>'
        + '      <button class="jira-edit plan-jira-edit" type="button" data-row-index="' + rowIndex + '" data-plan-key="' + esc(planCol.key) + '" title="Set Jira link">E</button>'
        + "    </div>"
        + "  </div>"
        + "</td>";
    }
    function setPlanJiraUrl(rowIndex, planKey) {
      const row = rows[rowIndex];
      const planMeta = PLAN_COLUMNS.find((x) => x.key === planKey);
      if (!row || !planMeta || !isPlanJiraEnabled(planKey)) return;
      row.plans = row.plans || {};
      const currentPlan = row.plans[planKey] || {};
      const current = String(currentPlan.jira_url || "").trim();
      const next = window.prompt("Set Jira link URL for " + planMeta.label + " (leave blank to clear):", current);
      if (next === null) return;
      try {
        const valid = validateJiraUrl(next);
        row.plans[planKey] = {
          man_days: currentPlan.man_days == null ? "" : currentPlan.man_days,
          start_date: String(currentPlan.start_date || ""),
          due_date: String(currentPlan.due_date || ""),
          jira_url: valid,
        };
        const rowOverride = ensureRowOverride(row);
        rowOverride.plans = Object.assign({}, rowOverride.plans || {}, { [planKey]: row.plans[planKey] });
        saveOverrides();
        renderTable();
        setStatus("Saving " + (row.epic_key || row.id || "epic") + "...", "");
        queueAutoPersist(rowIndex, valid ? (planMeta.label + " Jira link saved.") : (planMeta.label + " Jira link cleared."));
      } catch (err) {
        setStatus(err.message || String(err), "warn");
      }
    }
    function renderPrioritySelect(priority, rowIndex) {
      const selected = PRIORITY_OPTIONS.includes(priority) ? priority : "Low";
      return '<select data-row-index="' + rowIndex + '" data-field="priority">' +
        PRIORITY_OPTIONS.map((item) => '<option value="' + esc(item) + '"' + (item === selected ? " selected" : "") + ">" + esc(item) + "</option>").join("") +
        "</select>";
    }
    function renderPlanStatusSelect(planStatus, rowIndex) {
      const selected = normalizePlanStatus(planStatus);
      return '<select data-row-index="' + rowIndex + '" data-field="plan_status">' +
        PLAN_STATUS_OPTIONS.map((item) => '<option value="' + esc(item) + '"' + (item === selected ? " selected" : "") + ">" + esc(item) + "</option>").join("") +
        "</select>";
    }
    function renderIppMeetingPlannedSelect(value, rowIndex) {
      const selected = normalizeIppMeetingPlanned(value);
      return '<select data-row-index="' + rowIndex + '" data-field="ipp_meeting_planned">' +
        IPP_MEETING_PLANNED_OPTIONS.map((item) => '<option value="' + esc(item) + '"' + (item === selected ? " selected" : "") + ">" + esc(item) + "</option>").join("") +
        "</select>";
    }
    function normalizeIsoDateOrBlank(value) {
      const text = String(value || "").trim();
      if (!text) return "";
      return /^\\d{4}-\\d{2}-\\d{2}$/.test(text) ? text : "";
    }
    function renderActualProductionDateInput(value, rowIndex) {
      return '<input type="date" data-row-index="' + rowIndex + '" data-field="actual_production_date" value="' + esc(normalizeIsoDateOrBlank(value)) + '">';
    }
    const UNCATEGORIZED_LABEL = "Uncategorized";
    function projectNodeKey(projectName) {
      return String(projectName || "").trim().toLowerCase();
    }
    function categoryNodeKey(projectName, categoryName) {
      return projectNodeKey(projectName) + "||" + String(categoryName || "").trim().toLowerCase();
    }
    function componentNodeKey(projectName, categoryName, componentName) {
      return categoryNodeKey(projectName, categoryName) + "||" + String(componentName || "").trim().toLowerCase();
    }
    function displayBucketValue(rawValue) {
      const text = String(rawValue || "").trim();
      return text || UNCATEGORIZED_LABEL;
    }
    function findRowIndexByEpicKey(epicKey) {
      const key = String(epicKey || "").trim().toUpperCase();
      if (!key) return -1;
      const byEpicKey = rows.findIndex((item) => String(item.epic_key || item.id || "").trim().toUpperCase() === key);
      if (byEpicKey >= 0) return byEpicKey;
      return rows.findIndex((item) => {
        const jiraUrl = String(item.jira_url || "").trim();
        if (!jiraUrl) return false;
        const parts = jiraUrl.split("/");
        const last = parts[parts.length - 1] || "";
        const bare = String(last.split("?")[0] || "").trim().toUpperCase();
        return bare === key;
      });
    }
    function expandPathForEpicKey(epicKey) {
      const rowIndex = findRowIndexByEpicKey(epicKey);
      if (rowIndex < 0) return false;
      const row = rows[rowIndex];
      const project = String(row.project_name || row.project_key || "-").trim() || "-";
      const category = displayBucketValue(row.product_category);
      const component = displayBucketValue(row.component);
      expandedProjects.add(projectNodeKey(project));
      expandedCategories.add(categoryNodeKey(project, category));
      expandedComponents.add(componentNodeKey(project, category, component));
      return true;
    }
    function jumpToDeepLinkedEpicIfNeeded() {
      if (deepLinkHandled || !deepLinkEpicKey) return;
      let rowEl = Array.from(tbodyEl.querySelectorAll("tr[data-epic-key]")).find((tr) => {
        const rowKey = String(tr.getAttribute("data-epic-key") || "").trim().toUpperCase();
        return rowKey === deepLinkEpicKey;
      });
      if (!rowEl) {
        const rowIndex = findRowIndexByEpicKey(deepLinkEpicKey);
        if (rowIndex >= 0) {
          rowEl = tbodyEl.querySelector('tr[data-row-index="' + String(rowIndex) + '"]');
        }
      }
      if (!rowEl) return;
      deepLinkHandled = true;
      rowEl.classList.remove("epic-jump-highlight");
      void rowEl.offsetWidth;
      rowEl.classList.add("epic-jump-highlight");
      rowEl.scrollIntoView({ behavior: "smooth", block: "center" });
      window.setTimeout(() => rowEl.classList.remove("epic-jump-highlight"), 2000);
    }
    function renderEpicCell(row) {
      const hasJira = !!String(row.jira_url || "").trim();
      return ''
        + '<div class="tree tree-epic-cell">'
        + '  <div class="tree-line tree-epic"><span class="tree-title">' + esc(row.epic_name || row.epic_key || "-") + '</span></div>'
        + '  <div class="tree-actions">'
        + '    <a class="jira-open ' + (hasJira ? "" : "disabled") + '" href="' + esc(hasJira ? row.jira_url : "#") + '" target="_blank" rel="noopener noreferrer" title="' + (hasJira ? "Open Jira link" : "No Jira link set") + '">J</a>'
        + '    <button class="jira-edit" type="button" data-row-index="' + esc(row._row_index) + '" title="Set Jira link">E</button>'
        + '  </div>'
        + '</div>';
    }
    function renderEpicRow(rowIndex) {
      const row = rows[rowIndex];
      const planTds = PLAN_COLUMNS.map((planCol) => renderPlanCell(rowIndex, planCol, row)).join("");
      const categoryRaw = String(row.product_category || "").trim();
      const componentRaw = String(row.component || "").trim();
      const categoryClass = categoryRaw ? "" : " missing-categorization";
      const componentClass = componentRaw ? "" : " missing-categorization";
      const epicKey = String(row.epic_key || row.id || "").trim().toUpperCase();
      const project = String(row.project_name || row.project_key || "-").trim() || "-";
      const category = displayBucketValue(row.product_category);
      const component = displayBucketValue(row.component);
      return ""
        + '<tr data-row-index="' + rowIndex + '" data-epic-key="' + esc(epicKey) + '" data-project-node-key="' + esc(projectNodeKey(project)) + '" data-category-node-key="' + esc(categoryNodeKey(project, category)) + '" data-component-node-key="' + esc(componentNodeKey(project, category, component)) + '">'
        + "<td>" + esc(String(row.project_name || row.project_key || "-").trim() || "-") + "</td>"
        + '<td class="' + categoryClass.trim() + '">' + renderCategorizationSelect("product_category", dropdownOptions.product_category_options, categoryRaw, rowIndex) + "</td>"
        + '<td class="' + componentClass.trim() + '">' + renderCategorizationSelect("component", dropdownOptions.component_options, componentRaw, rowIndex) + "</td>"
        + "<td>" + renderEpicCell(row) + "</td>"
        + '<td class="description-cell"><div class="description-editor" contenteditable="true" data-row-index="' + rowIndex + '" data-field="description">' + esc(row.description || "") + "</div></td>"
        + '<td contenteditable="true" data-row-index="' + rowIndex + '" data-field="originator">' + esc(row.originator || "") + "</td>"
        + "<td>" + renderPrioritySelect(normalizePriority(row.priority), rowIndex) + "</td>"
        + "<td>" + renderPlanStatusSelect(normalizePlanStatus(row.plan_status), rowIndex) + "</td>"
        + "<td>" + renderIppMeetingPlannedSelect(normalizeIppMeetingPlanned(row.ipp_meeting_planned), rowIndex) + "</td>"
        + "<td>" + renderActualProductionDateInput(row.actual_production_date, rowIndex) + "</td>"
        + '<td contenteditable="true" data-row-index="' + rowIndex + '" data-field="remarks">' + esc(row.remarks || "") + "</td>"
        + planTds
        + '<td><div style="display:flex;gap:6px;flex-wrap:wrap;"><button class="btn alt small" type="button" data-edit-row="' + rowIndex + '">Edit</button><button class="btn alt small" type="button" data-save-row="' + rowIndex + '">Save</button><button class="btn alt small" type="button" data-sync-epic-row="' + rowIndex + '">Sync Jira Epic</button><button class="btn alt small danger" type="button" data-delete-epic-row="' + rowIndex + '" title="Delete epic from Epics Planner">Delete</button></div></td>'
        + "</tr>";
    }
    function ensureDraftEpicRow() {
      if (draftEpicRow && typeof draftEpicRow === "object") return draftEpicRow;
      draftEpicRow = {
        id: "draft-epic",
        project_key: "",
        project_name: "",
        product_category: "",
        component: "",
        epic_name: "",
        description: "",
        originator: "",
        priority: "Low",
        plan_status: defaultPlanStatus(),
        ipp_meeting_planned: "No",
        actual_production_date: "",
        remarks: "",
        jira_url: "",
        plans: {},
      };
      return draftEpicRow;
    }
    function renderDraftEpicRow() {
      const draft = ensureDraftEpicRow();
      const planTds = PLAN_COLUMNS.map(() => '<td class="plan-col-cell"><span class="plan-empty">Draft</span></td>').join("");
      return ""
        + '<tr class="draft-row">'
        + "<td>" + renderDraftProjectSelect(draft.project_key || "") + "</td>"
        + "<td>" + renderDraftCategorizationSelect("product_category", dropdownOptions.product_category_options, draft.product_category || "") + "</td>"
        + "<td>" + renderDraftCategorizationSelect("component", dropdownOptions.component_options, draft.component || "") + "</td>"
        + '<td><input id="draft-epic-name" class="draft-input" data-draft-field="epic_name" placeholder="Epic name (required)" value="' + esc(draft.epic_name || "") + '"></td>'
        + '<td><input class="draft-input" data-draft-field="description" placeholder="Description (optional)" value="' + esc(draft.description || "") + '"></td>'
        + '<td><input class="draft-input" data-draft-field="originator" placeholder="Originator (optional)" value="' + esc(draft.originator || "") + '"></td>'
        + '<td><span class="muted">Low</span></td>'
        + '<td><span class="muted">' + esc(defaultPlanStatus()) + "</span></td>"
        + '<td><span class="muted">No</span></td>'
        + '<td><span class="muted">-</span></td>'
        + '<td><span class="muted">Draft row</span></td>'
        + planTds
        + '<td><button class="btn small" type="button" id="save-draft-epic-btn">Save Draft Epic</button></td>'
        + "</tr>";
    }
    async function trySaveUsingVacantTmpKey(payload, conflictBody, options) {
      const opts = options && typeof options === "object" ? options : {};
      const vacantTmpKey = String((conflictBody && conflictBody.vacant_tmp_key) || "").trim().toUpperCase();
      if (!vacantTmpKey) return { reused: false };
      const reusePayload = Object.assign({}, payload, { epic_key: vacantTmpKey });
      const resp = await fetch(API + "/" + encodeURIComponent(vacantTmpKey), {
        method: "PUT",
        headers: { "Content-Type": "application/json" },
        body: JSON.stringify(reusePayload),
      });
      const body = await resp.json().catch(() => ({}));
      if (!resp.ok) {
        throw new Error(String(body.error || "Failed to save using vacant TMP key."));
      }
      if (typeof opts.onSuccess === "function") {
        opts.onSuccess(body, vacantTmpKey);
      }
      return { reused: true, body, vacantTmpKey };
    }
    async function saveDraftEpic() {
      if (!draftEpicRow || draftEpicCreateInFlight) return;
      const epicName = String(draftEpicRow.epic_name || "").trim();
      if (!epicName) {
        setStatus("Epic name is required to create a new epic row.", "warn");
        return;
      }
      draftEpicCreateInFlight = true;
      try {
        const payload = {
          epic_name: epicName,
          project_key: String(draftEpicRow.project_key || "").trim(),
          project_name: String(draftEpicRow.project_name || draftEpicRow.project_key || "").trim(),
          product_category: String(draftEpicRow.product_category || "").trim(),
          component: String(draftEpicRow.component || "").trim(),
          description: String(draftEpicRow.description || "").trim(),
          originator: String(draftEpicRow.originator || "").trim(),
          priority: "Low",
          plan_status: defaultPlanStatus(),
          ipp_meeting_planned: "No",
          jira_url: "",
          plans: {},
        };
        let resp = await fetch(API, {
          method: "POST",
          headers: { "Content-Type": "application/json" },
          body: JSON.stringify(payload),
        });
        let body = await resp.json().catch(() => ({}));
        if (!resp.ok) {
          if (resp.status === 409 && String(body.code || "") === "epic_key_exists" && String(body.vacant_tmp_key || "").trim()) {
            const reused = await trySaveUsingVacantTmpKey(payload, body);
            if (reused.reused) {
              draftEpicRow = null;
              await loadRowsFromApi();
              setStatus("Draft epic saved: " + reused.vacantTmpKey, "ok");
              return;
            }
          }
          resp = await fetch(API, {
            method: "POST",
            headers: { "Content-Type": "application/json" },
            body: JSON.stringify(payload),
          });
          body = await resp.json().catch(() => ({}));
          if (!resp.ok) {
            throw new Error(String(body.error || "Failed to create draft epic."));
          }
        }
        draftEpicRow = null;
        await loadRowsFromApi();
        const createdKey = String((body.row || {}).epic_key || "").trim();
        setStatus("Draft epic created: " + (createdKey || epicName), "ok");
      } finally {
        draftEpicCreateInFlight = false;
      }
    }
    function openDraftEpicRowAndFocus() {
      ensureDraftEpicRow();
      renderTable();
      const input = document.getElementById("draft-epic-name");
      if (input) input.focus();
    }
    function renderTable() {
      rows.forEach((row, index) => { row._row_index = index; });
      const grouped = new Map();
      rows.forEach((row, rowIndex) => {
        const project = String(row.project_name || row.project_key || "-").trim() || "-";
        const category = displayBucketValue(row.product_category);
        const component = displayBucketValue(row.component);
        if (!grouped.has(project)) grouped.set(project, new Map());
        const categoryMap = grouped.get(project);
        if (!categoryMap.has(category)) categoryMap.set(category, new Map());
        const componentMap = categoryMap.get(category);
        if (!componentMap.has(component)) componentMap.set(component, []);
        componentMap.get(component).push(rowIndex);
      });

      const html = [];
      if (!grouped.size) {
        const totalCols = 12 + PLAN_COLUMNS.length;
        tbodyEl.innerHTML = '<tr><td colspan="' + totalCols + '" style="text-align:center;color:#64748b;padding:16px;">No epics found in database.</td></tr>';
        return;
      }
      for (const [project, categoryMap] of grouped.entries()) {
        const pKey = projectNodeKey(project);
        const pExpanded = expandedProjects.has(pKey);
        const projectIndexes = [];
        for (const componentMap of categoryMap.values()) {
          for (const epicIndexes of componentMap.values()) projectIndexes.push(...epicIndexes);
        }
        const projectTotals = computeGroupManDaysTotals(projectIndexes);
        const projectPlanTotalTds = renderGroupPlanTotalCells(projectTotals);
        html.push(
          '<tr class="group-row project">'
          + '<td><div class="tree-line"><button class="tree-toggle" type="button" data-toggle-project="' + esc(pKey) + '">' + (pExpanded ? "-" : "+") + '</button><span class="tree-label-project">' + esc(project) + '</span><span class="tree-group-total">Total: ' + esc(formatManDaysValue(projectTotals.overall)) + ' md</span></div></td>'
          + '<td></td><td></td><td></td><td></td><td></td><td></td><td></td><td></td><td></td><td></td>'
          + projectPlanTotalTds
          + '<td></td>'
          + '</tr>'
        );
        if (!pExpanded) continue;

        for (const [category, componentMap] of categoryMap.entries()) {
          const cKey = categoryNodeKey(project, category);
          const cExpanded = expandedCategories.has(cKey);
          const categoryIndexes = [];
          for (const epicIndexes of componentMap.values()) categoryIndexes.push(...epicIndexes);
          const categoryTotals = computeGroupManDaysTotals(categoryIndexes);
          const categoryPlanTotalTds = renderGroupPlanTotalCells(categoryTotals);
          html.push(
            '<tr class="group-row category">'
            + '<td></td>'
            + '<td><div class="tree-line"><button class="tree-toggle" type="button" data-toggle-category="' + esc(cKey) + '">' + (cExpanded ? "-" : "+") + '</button><span class="tree-label-category">' + esc(category) + '</span><span class="tree-group-total">Total: ' + esc(formatManDaysValue(categoryTotals.overall)) + ' md</span></div></td>'
            + '<td></td><td></td><td></td><td></td><td></td><td></td><td></td><td></td><td></td>'
            + categoryPlanTotalTds
            + '<td></td>'
            + '</tr>'
          );
          if (!cExpanded) continue;

          for (const [component, epicIndexes] of componentMap.entries()) {
            const compKey = componentNodeKey(project, category, component);
            const compExpanded = expandedComponents.has(compKey);
            const componentTotals = computeGroupManDaysTotals(epicIndexes);
            const componentPlanTotalTds = renderGroupPlanTotalCells(componentTotals);
            html.push(
              '<tr class="group-row component">'
              + '<td></td><td></td>'
              + '<td><div class="tree-line"><button class="tree-toggle" type="button" data-toggle-component="' + esc(compKey) + '">' + (compExpanded ? "-" : "+") + '</button><span class="tree-label-category">' + esc(component) + '</span><span class="tree-group-total">Total: ' + esc(formatManDaysValue(componentTotals.overall)) + ' md</span></div></td>'
              + '<td></td><td></td><td></td><td></td><td></td><td></td><td></td><td></td>'
              + componentPlanTotalTds
              + '<td></td>'
              + '</tr>'
            );
            if (!compExpanded) continue;

            for (const rowIndex of epicIndexes) {
              html.push(renderEpicRow(rowIndex));
            }
          }
        }
      }
      if (draftEpicRow) {
        html.push(renderDraftEpicRow());
      }
      tbodyEl.innerHTML = html.join("");
      jumpToDeepLinkedEpicIfNeeded();

      Array.from(tbodyEl.querySelectorAll("button[data-toggle-project]")).forEach((btn) => {
        btn.addEventListener("click", () => {
          const key = String(btn.getAttribute("data-toggle-project") || "");
          if (!key) return;
          if (expandedProjects.has(key)) expandedProjects.delete(key);
          else expandedProjects.add(key);
          renderTable();
        });
      });

      Array.from(tbodyEl.querySelectorAll("button[data-toggle-category]")).forEach((btn) => {
        btn.addEventListener("click", () => {
          const key = String(btn.getAttribute("data-toggle-category") || "");
          if (!key) return;
          if (expandedCategories.has(key)) expandedCategories.delete(key);
          else expandedCategories.add(key);
          renderTable();
        });
      });
      Array.from(tbodyEl.querySelectorAll("button[data-toggle-component]")).forEach((btn) => {
        btn.addEventListener("click", () => {
          const key = String(btn.getAttribute("data-toggle-component") || "");
          if (!key) return;
          if (expandedComponents.has(key)) expandedComponents.delete(key);
          else expandedComponents.add(key);
          renderTable();
        });
      });

      Array.from(tbodyEl.querySelectorAll('[contenteditable="true"][data-row-index][data-field]')).forEach((cell) => {
        cell.addEventListener("blur", () => {
          const rowIndex = Number(cell.getAttribute("data-row-index"));
          const field = String(cell.getAttribute("data-field") || "");
          if (!rows[rowIndex] || !field) return;
          rows[rowIndex][field] = String(cell.textContent || "").trim();
          ensureRowOverride(rows[rowIndex])[field] = rows[rowIndex][field];
          saveOverrides();
          if (field === "product_category" || field === "component") {
            renderTable();
          }
          setStatus("Saving " + (rows[rowIndex].epic_key || rows[rowIndex].id || "epic") + "...", "");
          queueAutoPersist(rowIndex, "Saved " + (rows[rowIndex].epic_key || rows[rowIndex].id || "epic") + ".");
        });
      });

      Array.from(tbodyEl.querySelectorAll("select[data-field='priority']")).forEach((selectEl) => {
        selectEl.addEventListener("change", () => {
          const rowIndex = Number(selectEl.getAttribute("data-row-index"));
          if (!rows[rowIndex]) return;
          const nextValue = normalizePriority(selectEl.value);
          rows[rowIndex].priority = nextValue;
          ensureRowOverride(rows[rowIndex]).priority = nextValue;
          saveOverrides();
          setStatus("Saving " + (rows[rowIndex].epic_key || rows[rowIndex].id || "epic") + "...", "");
          queueAutoPersist(rowIndex, "Priority saved for " + (rows[rowIndex].epic_key || rows[rowIndex].id || "epic") + ".");
        });
      });
      Array.from(tbodyEl.querySelectorAll("select[data-field='product_category']")).forEach((selectEl) => {
        selectEl.addEventListener("change", () => {
          const rowIndex = Number(selectEl.getAttribute("data-row-index"));
          if (!rows[rowIndex]) return;
          const nextValue = String(selectEl.value || "").trim();
          rows[rowIndex].product_category = nextValue;
          ensureRowOverride(rows[rowIndex]).product_category = nextValue;
          saveOverrides();
          renderTable();
          setStatus("Saving " + (rows[rowIndex].epic_key || rows[rowIndex].id || "epic") + "...", "");
          queueAutoPersist(rowIndex, "Product Categorization saved for " + (rows[rowIndex].epic_key || rows[rowIndex].id || "epic") + ".");
        });
      });
      Array.from(tbodyEl.querySelectorAll("select[data-field='component']")).forEach((selectEl) => {
        selectEl.addEventListener("change", () => {
          const rowIndex = Number(selectEl.getAttribute("data-row-index"));
          if (!rows[rowIndex]) return;
          const nextValue = String(selectEl.value || "").trim();
          rows[rowIndex].component = nextValue;
          ensureRowOverride(rows[rowIndex]).component = nextValue;
          saveOverrides();
          renderTable();
          setStatus("Saving " + (rows[rowIndex].epic_key || rows[rowIndex].id || "epic") + "...", "");
          queueAutoPersist(rowIndex, "Component saved for " + (rows[rowIndex].epic_key || rows[rowIndex].id || "epic") + ".");
        });
      });
      Array.from(tbodyEl.querySelectorAll("select[data-field='plan_status']")).forEach((selectEl) => {
        selectEl.addEventListener("change", () => {
          const rowIndex = Number(selectEl.getAttribute("data-row-index"));
          if (!rows[rowIndex]) return;
          const nextValue = normalizePlanStatus(selectEl.value);
          rows[rowIndex].plan_status = nextValue;
          ensureRowOverride(rows[rowIndex]).plan_status = nextValue;
          saveOverrides();
          setStatus("Saving " + (rows[rowIndex].epic_key || rows[rowIndex].id || "epic") + "...", "");
          queueAutoPersist(rowIndex, "Plan status saved for " + (rows[rowIndex].epic_key || rows[rowIndex].id || "epic") + ".");
        });
      });
      Array.from(tbodyEl.querySelectorAll("select[data-field='ipp_meeting_planned']")).forEach((selectEl) => {
        selectEl.addEventListener("change", () => {
          const rowIndex = Number(selectEl.getAttribute("data-row-index"));
          if (!rows[rowIndex]) return;
          const nextValue = normalizeIppMeetingPlanned(selectEl.value);
          rows[rowIndex].ipp_meeting_planned = nextValue;
          ensureRowOverride(rows[rowIndex]).ipp_meeting_planned = nextValue;
          saveOverrides();
          setStatus("Saving " + (rows[rowIndex].epic_key || rows[rowIndex].id || "epic") + "...", "");
          queueAutoPersist(rowIndex, "IPP Meeting Planner saved for " + (rows[rowIndex].epic_key || rows[rowIndex].id || "epic") + ".");
        });
      });
      Array.from(tbodyEl.querySelectorAll("input[data-field='actual_production_date']")).forEach((inputEl) => {
        inputEl.addEventListener("change", () => {
          const rowIndex = Number(inputEl.getAttribute("data-row-index"));
          if (!rows[rowIndex]) return;
          const nextValue = normalizeIsoDateOrBlank(inputEl.value);
          rows[rowIndex].actual_production_date = nextValue;
          ensureRowOverride(rows[rowIndex]).actual_production_date = nextValue;
          saveOverrides();
          setStatus("Saving " + (rows[rowIndex].epic_key || rows[rowIndex].id || "epic") + "...", "");
          queueAutoPersist(rowIndex, "Actual Production Date saved for " + (rows[rowIndex].epic_key || rows[rowIndex].id || "epic") + ".");
        });
      });

      Array.from(tbodyEl.querySelectorAll("button.plan-btn[data-plan-key]")).forEach((btn) => {
        btn.addEventListener("click", () => {
          const rowIndex = Number(btn.getAttribute("data-row-index"));
          const planKey = String(btn.getAttribute("data-plan-key") || "");
          openPlanDialog(rowIndex, planKey);
        });
      });
      Array.from(tbodyEl.querySelectorAll("button[data-plan-toggle]")).forEach((btn) => {
        btn.addEventListener("click", (event) => {
          event.preventDefault();
          event.stopPropagation();
          const key = String(btn.getAttribute("data-plan-toggle") || "");
          if (!key) return;
          if (expandedPlanDetails.has(key)) expandedPlanDetails.delete(key);
          else expandedPlanDetails.add(key);
          renderTable();
        });
      });

      Array.from(tbodyEl.querySelectorAll("button.jira-edit:not(.plan-jira-edit)")).forEach((btn) => {
        btn.addEventListener("click", () => {
          const rowIndex = Number(btn.getAttribute("data-row-index"));
          const row = rows[rowIndex];
          if (!row) return;
          const current = String(row.jira_url || "").trim();
          const next = window.prompt("Set Jira link URL (leave blank to clear):", current);
          if (next === null) return;
          try {
            const valid = validateJiraUrl(next);
            row.jira_url = valid;
            ensureRowOverride(row).jira_url = valid;
            saveOverrides();
            renderTable();
            setStatus("Saving " + (row.epic_key || row.id || "epic") + "...", "");
            queueAutoPersist(rowIndex, valid ? "Jira link saved." : "Jira link cleared and saved.");
          } catch (err) {
            setStatus(err.message || String(err), "warn");
          }
        });
      });
      Array.from(tbodyEl.querySelectorAll("button.plan-jira-edit")).forEach((btn) => {
        btn.addEventListener("click", () => {
          const rowIndex = Number(btn.getAttribute("data-row-index"));
          const planKey = String(btn.getAttribute("data-plan-key") || "");
          setPlanJiraUrl(rowIndex, planKey);
        });
      });

      Array.from(tbodyEl.querySelectorAll("button[data-save-row]")).forEach((btn) => {
        btn.addEventListener("click", () => {
          const rowIndex = Number(btn.getAttribute("data-save-row"));
          persistRow(rowIndex).catch((err) => setStatus(err.message || String(err), "warn"));
        });
      });
      Array.from(tbodyEl.querySelectorAll("button[data-edit-row]")).forEach((btn) => {
        btn.addEventListener("click", () => {
          const rowIndex = Number(btn.getAttribute("data-edit-row"));
          const row = rows[rowIndex];
          if (!row) return;
          openEpicDialogForEdit(row);
        });
      });
      Array.from(tbodyEl.querySelectorAll("button[data-sync-epic-row]")).forEach((btn) => {
        btn.addEventListener("click", () => {
          const rowIndex = Number(btn.getAttribute("data-sync-epic-row"));
          syncRowPlanFromJira(rowIndex).catch((err) => setStatus(err.message || String(err), "warn"));
        });
      });
      Array.from(tbodyEl.querySelectorAll("button[data-delete-epic-row]")).forEach((btn) => {
        btn.addEventListener("click", () => {
          const rowIndex = Number(btn.getAttribute("data-delete-epic-row"));
          deleteEpicRow(rowIndex).catch((err) => setStatus(err.message || String(err), "warn"));
        });
      });
      Array.from(tbodyEl.querySelectorAll("input.draft-input[data-draft-field]")).forEach((inputEl) => {
        inputEl.addEventListener("input", () => {
          if (!draftEpicRow) return;
          const field = String(inputEl.getAttribute("data-draft-field") || "");
          if (!field) return;
          draftEpicRow[field] = String(inputEl.value || "");
        });
      });
      Array.from(tbodyEl.querySelectorAll("select.draft-input[data-draft-field]")).forEach((selectEl) => {
        selectEl.addEventListener("change", () => {
          if (!draftEpicRow) return;
          const field = String(selectEl.getAttribute("data-draft-field") || "");
          if (!field) return;
          draftEpicRow[field] = String(selectEl.value || "");
          if (field === "project_key") {
            const opt = selectEl.options[selectEl.selectedIndex];
            draftEpicRow.project_name = (opt && opt.getAttribute("data-project-name")) || draftEpicRow.project_key || "";
          }
        });
      });
      const draftSaveBtn = document.getElementById("save-draft-epic-btn");
      if (draftSaveBtn) {
        draftSaveBtn.addEventListener("click", () => {
          saveDraftEpic().catch((err) => setStatus(err.message || String(err), "warn"));
        });
      }
    }
    function payloadFromRow(row) {
      const plans = {};
      PLAN_COLUMNS.forEach((col) => {
        plans[col.key] = Object.assign({ man_days: "", start_date: "", due_date: "", jira_url: "" }, (row.plans || {})[col.key] || {});
      });
      return {
        epic_key: String(row.epic_key || row.id || "").toUpperCase(),
        project_key: String(row.project_key || ""),
        project_name: String(row.project_name || ""),
        product_category: String(row.product_category || ""),
        component: String(row.component || ""),
        epic_name: String(row.epic_name || row.epic_key || ""),
        description: String(row.description || ""),
        originator: String(row.originator || ""),
        priority: normalizePriority(row.priority || "Low"),
        plan_status: normalizePlanStatus(row.plan_status || defaultPlanStatus()),
        ipp_meeting_planned: normalizeIppMeetingPlanned(row.ipp_meeting_planned || "No"),
        actual_production_date: normalizeIsoDateOrBlank(row.actual_production_date || ""),
        remarks: String(row.remarks || ""),
        jira_url: validateJiraUrl(row.jira_url || ""),
        plans: plans,
      };
    }
    async function persistRow(rowIndex, options) {
      const opts = options && typeof options === "object" ? options : {};
      const row = rows[rowIndex];
      if (!row) throw new Error("Row not found.");
      const oldId = String(row.id || row.epic_key || "");
      const payload = payloadFromRow(row);
      const resp = await fetch(API + "/" + encodeURIComponent(String(payload.epic_key || "")), {
        method: "PUT",
        headers: { "Content-Type": "application/json" },
        body: JSON.stringify(payload),
      });
      const body = await resp.json().catch(() => ({}));
      if (!resp.ok) throw new Error(String(body.error || "Failed to save epic."));
      const savedRow = mergeRow(body.row || row);
      rows[rowIndex] = savedRow;
      if (oldId) delete overrides[oldId];
      overrides[String(savedRow.id || savedRow.epic_key || "")] = {};
      saveOverrides();
      if (opts.render !== false) renderTable();
      const message = String(opts.successMessage || "").trim() || ("Saved " + (savedRow.epic_key || savedRow.id || "epic") + " to database.");
      setStatus(message, "ok");
    }
    function queueAutoPersist(rowIndex, successMessage) {
      const index = Number(rowIndex);
      if (!Number.isInteger(index) || !rows[index]) return Promise.resolve();
      if (autoSaveInFlight.has(index)) {
        autoSaveQueued.add(index);
        return autoSaveInFlight.get(index);
      }
      const task = (async () => {
        try {
          await persistRow(index, { successMessage });
        } catch (err) {
          setStatus("Auto-save failed: " + (err.message || String(err)), "warn");
        } finally {
          autoSaveInFlight.delete(index);
          if (autoSaveQueued.has(index)) {
            autoSaveQueued.delete(index);
            queueAutoPersist(index, successMessage);
          }
        }
      })();
      autoSaveInFlight.set(index, task);
      return task;
    }
    function flashSyncRowHighlight(rowIndex, mode, durationMs) {
      const rowEl = tbodyEl.querySelector('tr[data-row-index="' + String(rowIndex) + '"]');
      if (!rowEl) return;
      const fetchedClass = "sync-fetched-highlight";
      const savedClass = "sync-saved-highlight";
      rowEl.classList.remove(fetchedClass, savedClass);
      const nextClass = mode === "saved" ? savedClass : fetchedClass;
      rowEl.classList.add(nextClass);
      window.setTimeout(() => {
        rowEl.classList.remove(nextClass);
      }, Math.max(0, Number(durationMs) || 0));
    }
    async function syncRowPlanFromJira(rowIndex) {
      const row = rows[rowIndex];
      if (!row) throw new Error("Row not found.");
      const key = String(row.epic_key || row.id || "").toUpperCase();
      if (!key) throw new Error("Epic key is required to sync.");
      const planJiraLinks = {};
      PLAN_COLUMNS.forEach((col) => {
        const plan = (row.plans || {})[col.key] || {};
        const jiraUrl = String(plan.jira_url || "").trim();
        if (jiraUrl) planJiraLinks[col.key] = jiraUrl;
      });
      const resp = await fetch(API + "/" + encodeURIComponent(key) + "/sync-jira-plan", {
        method: "POST",
        headers: { "Content-Type": "application/json" },
        body: JSON.stringify({
          jira_url: String(row.jira_url || ""),
          plan_jira_links: planJiraLinks,
        }),
      });
      const body = await resp.json().catch(() => ({}));
      if (!resp.ok) throw new Error(String(body.error || "Failed to sync Jira plan."));
      const savedRow = mergeRow(body.row || row);
      const syncedStoryCount = Number(body.synced_story_count || savedRow.synced_story_count || 0);
      rows[rowIndex] = savedRow;
      overrides[String(savedRow.id || savedRow.epic_key || "")] = {};
      saveOverrides();
      renderTable();
      flashSyncRowHighlight(rowIndex, "fetched", 1500);
      await new Promise((resolve) => window.setTimeout(resolve, 1500));
      await persistRow(rowIndex, {
        render: true,
        successMessage: "Synced Jira plan and saved " + (savedRow.epic_key || key) + " plus " + String(syncedStoryCount) + " stories to database.",
      });
      flashSyncRowHighlight(rowIndex, "saved", 2000);
    }
    async function deleteEpicRow(rowIndex) {
      const row = rows[rowIndex];
      if (!row) throw new Error("Row not found.");
      const key = String(row.epic_key || row.id || "").trim().toUpperCase();
      if (!key) throw new Error("Epic key is required to delete.");
      const label = String(row.epic_name || row.epic_key || key).trim() || key;
      if (!window.confirm('Delete epic "' + label + '" (' + key + ') from Epics Planner? This cannot be undone.')) return;
      const resp = await fetch(API + "/" + encodeURIComponent(key), { method: "DELETE" });
      const body = await resp.json().catch(() => ({}));
      if (!resp.ok) throw new Error(String(body.error || "Failed to delete epic."));
      await loadRowsFromApi();
      setStatus("Epic deleted: " + key, "ok");
    }
    function resetEpicCreateForm() {
      renderProjectOptions("");
      epicProjectSelectEl.value = "";
      setDropdownSelectOptions(epicProductCategoryEl, dropdownOptions.product_category_options, "");
      setDropdownSelectOptions(epicComponentEl, dropdownOptions.component_options, "");
      epicNameEl.value = "";
      epicOriginatorEl.value = "";
      epicPriorityEl.value = "Low";
      epicPlanStatusEl.value = defaultPlanStatus();
      epicIppMeetingPlannedEl.value = "No";
      epicActualProductionDateEl.value = "";
      epicRemarksEl.value = "";
      epicJiraUrlEl.value = "";
      epicDescriptionEl.value = "";
      epicResearchUrsPlanJiraUrlEl.value = "";
      epicDdsPlanJiraUrlEl.value = "";
      epicDevelopmentPlanJiraUrlEl.value = "";
      epicSqaPlanJiraUrlEl.value = "";
      epicUserManualPlanJiraUrlEl.value = "";
      epicProductionPlanJiraUrlEl.value = "";
      Object.values(dynamicPlanInputEls).forEach((jiraEl) => {
        jiraEl.value = "";
      });
    }
    function openEpicDialogForCreate() {
      activeEpicEditKey = "";
      epicDialogTitleEl.textContent = "Add New Epic";
      epicDialogSubtitleEl.textContent = "Provide epic metadata; you can edit detailed plans and IPP meeting selection after creation.";
      document.getElementById("epic-save").textContent = "Create Epic";
      resetEpicCreateForm();
      epicDialogEl.showModal();
    }
    function openEpicDialogForEdit(row) {
      activeEpicEditKey = String(row.epic_key || row.id || "").toUpperCase();
      epicDialogTitleEl.textContent = "Edit Epic";
      epicDialogSubtitleEl.textContent = "Update epic metadata and save to database.";
      document.getElementById("epic-save").textContent = "Save Epic";
      renderProjectOptions(String(row.project_key || ""));
      epicProjectSelectEl.value = normalizeProjectKey(row.project_key);
      setDropdownSelectOptions(epicProductCategoryEl, dropdownOptions.product_category_options, String(row.product_category || ""));
      setDropdownSelectOptions(epicComponentEl, dropdownOptions.component_options, String(row.component || ""));
      epicNameEl.value = String(row.epic_name || "");
      epicOriginatorEl.value = String(row.originator || "");
      epicPriorityEl.value = normalizePriority(row.priority || "Low");
      epicPlanStatusEl.value = normalizePlanStatus(row.plan_status || defaultPlanStatus());
      epicIppMeetingPlannedEl.value = normalizeIppMeetingPlanned(row.ipp_meeting_planned || "No");
      epicActualProductionDateEl.value = normalizeIsoDateOrBlank(row.actual_production_date || "");
      epicRemarksEl.value = String(row.remarks || "");
      epicJiraUrlEl.value = String(row.jira_url || "");
      epicDescriptionEl.value = String(row.description || "");
      const plans = row.plans || {};
      epicResearchUrsPlanJiraUrlEl.value = String((plans.research_urs_plan || {}).jira_url || "");
      epicDdsPlanJiraUrlEl.value = String((plans.dds_plan || {}).jira_url || "");
      epicDevelopmentPlanJiraUrlEl.value = String((plans.development_plan || {}).jira_url || "");
      epicSqaPlanJiraUrlEl.value = String((plans.sqa_plan || {}).jira_url || "");
      epicUserManualPlanJiraUrlEl.value = String((plans.user_manual_plan || {}).jira_url || "");
      epicProductionPlanJiraUrlEl.value = String((plans.production_plan || {}).jira_url || "");
      Object.keys(dynamicPlanInputEls).forEach((planKey) => {
        const plan = plans[planKey] || {};
        dynamicPlanInputEls[planKey].value = String(plan.jira_url || "");
      });
      epicDialogEl.showModal();
    }
    function buildEpicCreatePayload() {
      const jiraUrl = validateJiraUrl(epicJiraUrlEl.value);
      const resolvedEpicKey = activeEpicEditKey
        ? normalizeEpicKeyAllowTmp(activeEpicEditKey)
        : epicKeyFromJiraUrl(jiraUrl);
      if (!resolvedEpicKey) {
        throw new Error("Jira URL must include an epic key like /browse/O2-1234.");
      }
      const project = selectedManagedProject();
      let projectKey = project ? normalizeProjectKey(project.project_key) : "";
      if (!projectKey) projectKey = projectKeyFromEpicKey(resolvedEpicKey);
      if (!projectKey) projectKey = "ORPHAN";
      const projectName = project
        ? String(project.display_name || project.project_name || projectKey || "").trim()
        : (projectKey === "ORPHAN" ? "Orphan" : projectKey);
      const editingRow = activeEpicEditKey
        ? rows.find((item) => String(item.epic_key || item.id || "").toUpperCase() === activeEpicEditKey) || null
        : null;
      const basePlans = editingRow && editingRow.plans && typeof editingRow.plans === "object"
        ? deepClone(editingRow.plans)
        : {};
      const planJiraUrls = {
        research_urs_plan: validateJiraUrl(epicResearchUrsPlanJiraUrlEl.value),
        dds_plan: validateJiraUrl(epicDdsPlanJiraUrlEl.value),
        development_plan: validateJiraUrl(epicDevelopmentPlanJiraUrlEl.value),
        sqa_plan: validateJiraUrl(epicSqaPlanJiraUrlEl.value),
        user_manual_plan: validateJiraUrl(epicUserManualPlanJiraUrlEl.value),
        production_plan: validateJiraUrl(epicProductionPlanJiraUrlEl.value),
      };
      const plans = {};
      Object.keys(planJiraUrls).forEach((planKey) => {
        const existing = (basePlans[planKey] && typeof basePlans[planKey] === "object") ? basePlans[planKey] : {};
        plans[planKey] = Object.assign(
          { man_days: "", start_date: "", due_date: "", jira_url: "" },
          existing,
          { jira_url: planJiraUrls[planKey] },
        );
      });
      Object.keys(dynamicPlanInputEls).forEach((planKey) => {
        const jiraInput = dynamicPlanInputEls[planKey];
        const existing = (basePlans[planKey] && typeof basePlans[planKey] === "object") ? basePlans[planKey] : {};
        plans[planKey] = Object.assign(
          { man_days: "", start_date: "", due_date: "", jira_url: "" },
          existing,
          {
            jira_url: validateJiraUrl(jiraInput.value),
          },
        );
      });
      return {
        project_key: projectKey,
        project_name: projectName || projectKey,
        product_category: String(epicProductCategoryEl.value || "").trim(),
        component: String(epicComponentEl.value || "").trim(),
        epic_key: resolvedEpicKey,
        epic_name: String(epicNameEl.value || "").trim(),
        originator: String(epicOriginatorEl.value || "").trim(),
        priority: normalizePriority(epicPriorityEl.value),
        plan_status: normalizePlanStatus(epicPlanStatusEl.value),
        ipp_meeting_planned: normalizeIppMeetingPlanned(epicIppMeetingPlannedEl.value),
        actual_production_date: normalizeIsoDateOrBlank(epicActualProductionDateEl.value),
        remarks: String(epicRemarksEl.value || "").trim(),
        jira_url: jiraUrl,
        description: String(epicDescriptionEl.value || "").trim(),
        plans: plans,
      };
    }
    async function createEpic() {
      const payload = buildEpicCreatePayload();
      const isEdit = !!activeEpicEditKey;
      const endpoint = isEdit ? (API + "/" + encodeURIComponent(activeEpicEditKey)) : API;
      const resp = await fetch(endpoint, {
        method: isEdit ? "PUT" : "POST",
        headers: { "Content-Type": "application/json" },
        body: JSON.stringify(payload),
      });
      const body = await resp.json().catch(() => ({}));
      if (!resp.ok) {
        if (!isEdit && resp.status === 409 && String(body.code || "") === "epic_key_exists" && String(body.vacant_tmp_key || "").trim()) {
          const reused = await trySaveUsingVacantTmpKey(payload, body, {
            onSuccess: () => {
              epicDialogEl.close();
            },
          });
          if (reused.reused) {
            await loadRowsFromApi();
            setStatus("Epic created using vacant TMP key: " + reused.vacantTmpKey, "ok");
            activeEpicEditKey = "";
            return;
          }
        }
        throw new Error(String(body.error || (isEdit ? "Failed to update epic." : "Failed to create epic.")));
      }
      epicDialogEl.close();
      await loadRowsFromApi();
      setStatus((isEdit ? "Epic updated: " : "Epic created: ") + (body.row?.epic_key || payload.epic_key), "ok");
      activeEpicEditKey = "";
    }
    function openPlanDialog(rowIndex, planKey) {
      const row = rows[rowIndex];
      const planMeta = PLAN_COLUMNS.find((x) => x.key === planKey);
      if (!row || !planMeta) return;
      activePlan = { rowIndex, planKey };
      const plan = ((row.plans || {})[planKey]) || {};
      planTitleEl.textContent = "Edit " + planMeta.label;
      planContextEl.textContent = (row.project_name || row.project_key || "-") + " / " + (row.product_category || "-") + " / " + (row.component || "-") + " / " + (row.epic_name || row.epic_key || "-");
      planMandaysEl.value = plan.man_days == null ? "" : String(plan.man_days);
      planStartEl.value = toDateValue(plan.start_date);
      planDueEl.value = toDateValue(plan.due_date);
      planDialogEl.showModal();
    }
    function clearPlanInputs() {
      planMandaysEl.value = "";
      planStartEl.value = "";
      planDueEl.value = "";
    }
    function savePlan() {
      const row = rows[activePlan.rowIndex];
      if (!row || !activePlan.planKey) return;
      const rowIndex = activePlan.rowIndex;
      const manDaysRaw = String(planMandaysEl.value || "").trim();
      const startDate = String(planStartEl.value || "").trim();
      const dueDate = String(planDueEl.value || "").trim();
      if (startDate && dueDate && startDate > dueDate) {
        setStatus("Start date cannot be after due date.", "warn");
        return;
      }
      const manDays = manDaysRaw === "" ? "" : Number(manDaysRaw);
      if (manDaysRaw !== "" && (!Number.isFinite(manDays) || manDays < 0)) {
        setStatus("Man-days must be zero or a positive number.", "warn");
        return;
      }
      row.plans = row.plans || {};
      row.plans[activePlan.planKey] = {
        man_days: manDaysRaw === "" ? "" : manDays,
        start_date: startDate,
        due_date: dueDate,
        jira_url: String(((row.plans || {})[activePlan.planKey] || {}).jira_url || ""),
      };
      const rowOverride = ensureRowOverride(row);
      rowOverride.plans = Object.assign({}, rowOverride.plans || {}, { [activePlan.planKey]: row.plans[activePlan.planKey] });
      saveOverrides();
      planDialogEl.close();
      renderTable();
      setStatus("Saving " + (row.epic_key || row.id || "epic") + "...", "");
      queueAutoPersist(rowIndex, "Plan saved for " + (row.epic_key || row.id || "epic") + ".");
    }
    function clearPlanValue() {
      const row = rows[activePlan.rowIndex];
      if (!row || !activePlan.planKey) return;
      const rowIndex = activePlan.rowIndex;
      row.plans = row.plans || {};
      row.plans[activePlan.planKey] = { man_days: "", start_date: "", due_date: "", jira_url: "" };
      const rowOverride = ensureRowOverride(row);
      rowOverride.plans = Object.assign({}, rowOverride.plans || {}, { [activePlan.planKey]: row.plans[activePlan.planKey] });
      saveOverrides();
      planDialogEl.close();
      renderTable();
      setStatus("Saving " + (row.epic_key || row.id || "epic") + "...", "");
      queueAutoPersist(rowIndex, "Plan cleared for " + (row.epic_key || row.id || "epic") + ".");
    }
    async function loadRowsFromApi() {
      const resp = await fetch(API, { cache: "no-store" });
      const body = await resp.json().catch(() => ({}));
      if (!resp.ok) throw new Error(String(body.error || "Failed to load epics data."));
      const baseRows = Array.isArray(body.rows) ? body.rows : [];
      rows = baseRows.map((row) => mergeRow({
        id: String(row.id || row.epic_key || ""),
        project_key: String(row.project_key || ""),
        project_name: String(row.project_name || row.project_key || ""),
        product_category: String(row.product_category || ""),
        component: String(row.component || ""),
        epic_key: String(row.epic_key || row.id || ""),
        epic_name: String(row.epic_name || row.epic_key || row.id || ""),
        description: String(row.description || ""),
        originator: String(row.originator || ""),
        priority: normalizePriority(row.priority || "Low"),
        plan_status: normalizePlanStatus(row.plan_status || defaultPlanStatus()),
        ipp_meeting_planned: normalizeIppMeetingPlanned(row.ipp_meeting_planned || "No"),
        actual_production_date: normalizeIsoDateOrBlank(row.actual_production_date || ""),
        remarks: String(row.remarks || ""),
        jira_url: String(row.jira_url || ""),
        plans: (row.plans && typeof row.plans === "object") ? row.plans : {},
      }));
      expandedProjects.clear();
      expandedCategories.clear();
      expandedComponents.clear();
      for (const row of rows) {
        const project = String(row.project_name || row.project_key || "-").trim() || "-";
        const category = displayBucketValue(row.product_category);
        const component = displayBucketValue(row.component);
        expandedProjects.add(projectNodeKey(project));
        expandedCategories.add(categoryNodeKey(project, category));
        expandedComponents.add(componentNodeKey(project, category, component));
      }
      if (deepLinkEpicKey) {
        const found = expandPathForEpicKey(deepLinkEpicKey);
        if (!found && !deepLinkHandled) {
          deepLinkHandled = true;
          deepLinkMissingWarningShown = true;
          setStatus("Loaded " + rows.length + " epics. Deep-link epic not found: " + deepLinkEpicKey + ".", "warn");
        }
      }
      renderTable();
      jumpToDeepLinkedEpicIfNeeded();
      const selectedCount = rows.filter((item) => normalizeIppMeetingPlanned(item.ipp_meeting_planned) === "Yes").length;
      if (!deepLinkMissingWarningShown) {
        setStatus("Loaded " + rows.length + " epics from database (" + selectedCount + " selected for IPP Meeting Planner). Use + / - to collapse or expand Project/Product Categorization/Component groups. Yellow cells need categorization data.", "ok");
      }
    }

    document.getElementById("reload-btn").addEventListener("click", () => {
      Promise.all([loadPlanColumns(), loadDropdownOptions(), loadRowsFromApi()]).catch((err) => setStatus(err.message || String(err), "warn"));
    });
    window.addEventListener("resize", () => {
      applyPlanColumnLayout();
    });
    document.getElementById("add-epic-btn").addEventListener("click", () => openEpicDialogForCreate());
    document.getElementById("add-plan-column-btn").addEventListener("click", () => {
      addPlanColumn().catch((err) => setStatus(err.message || String(err), "warn"));
    });
    document.getElementById("manage-plan-columns-btn").addEventListener("click", () => {
      window.location.href = "/settings/epic-phases";
    });
    document.getElementById("plan-column-save").addEventListener("click", () => {
      savePlanColumnFromDialog().catch((err) => setStatus(err.message || String(err), "warn"));
    });
    planColumnNameEl.addEventListener("input", refreshPlanColumnRestoreHint);
    document.getElementById("plan-column-cancel").addEventListener("click", () => planColumnDialogEl.close());
    document.getElementById("manage-columns-close").addEventListener("click", () => manageColumnsDialogEl.close());
    manageColumnsTbodyEl.addEventListener("click", (event) => {
      const target = event.target instanceof Element ? event.target : null;
      const btn = target ? target.closest("button[data-delete-plan-key]") : null;
      if (!btn) return;
      const key = String(btn.getAttribute("data-delete-plan-key") || "");
      const label = String(btn.getAttribute("data-delete-plan-label") || key);
      deletePlanColumn(key, label).catch((err) => setStatus(err.message || String(err), "warn"));
    });
    planColumnDialogEl.addEventListener("close", () => {
      activePlanInsertPosition = 0;
    });
    document.getElementById("epic-save").addEventListener("click", () => {
      createEpic().catch((err) => setStatus(err.message || String(err), "warn"));
    });
    document.getElementById("epic-cancel").addEventListener("click", () => epicDialogEl.close());
    document.getElementById("expand-all-btn").addEventListener("click", () => {
      expandedProjects.clear();
      expandedCategories.clear();
      expandedComponents.clear();
      for (const row of rows) {
        const project = String(row.project_name || row.project_key || "-").trim() || "-";
        const category = displayBucketValue(row.product_category);
        const component = displayBucketValue(row.component);
        expandedProjects.add(projectNodeKey(project));
        expandedCategories.add(categoryNodeKey(project, category));
        expandedComponents.add(componentNodeKey(project, category, component));
      }
      renderTable();
    });
    document.getElementById("collapse-all-btn").addEventListener("click", () => {
      expandedProjects.clear();
      expandedCategories.clear();
      expandedComponents.clear();
      renderTable();
    });
    document.getElementById("plan-save").addEventListener("click", savePlan);
    document.getElementById("plan-clear").addEventListener("click", clearPlanValue);
    document.getElementById("plan-cancel").addEventListener("click", () => planDialogEl.close());
    planDialogEl.addEventListener("close", () => {
      activePlan = { rowIndex: -1, planKey: "" };
      clearPlanInputs();
    });
    document.addEventListener("keydown", (event) => {
      if (event.key !== "Tab" || !event.shiftKey) return;
      if (planDialogEl.open || epicDialogEl.open || planColumnDialogEl.open || manageColumnsDialogEl.open) return;
      const active = document.activeElement;
      if (active && active instanceof Element) {
        const tag = String(active.tagName || "").toUpperCase();
        if (tag === "INPUT" || tag === "SELECT" || tag === "TEXTAREA" || active.isContentEditable || active.closest("#epics-table")) {
          return;
        }
      }
      event.preventDefault();
      openDraftEpicRowAndFocus();
    });
    document.addEventListener("keydown", (event) => {
      if (event.key !== "Escape") return;
      if (!draftEpicRow || draftEpicCreateInFlight) return;
      if (planDialogEl.open || epicDialogEl.open || planColumnDialogEl.open || manageColumnsDialogEl.open) return;
      draftEpicRow = null;
      renderTable();
      setStatus("Draft epic row discarded.", "ok");
    });
    window.addEventListener("resize", applyPlanColumnLayout);

    (async function init() {
      initHeaderToggle();
      overrides = loadOverrides();
      renderPlanHeaders();
      setStatus("Loading epics data...", "");
      try {
        await Promise.all([loadPlanColumns(), loadPlanColumnsCatalog(), loadManagedProjects(), loadDropdownOptions(), loadRowsFromApi()]);
      } catch (err) {
        setStatus(err.message || String(err), "warn");
      }
    })();
  </script>
  <script src="/shared-nav.js"></script>
</body>
</html>""".replace("__SETTINGS_TOP_NAV__", _settings_top_nav_html(EPICS_MANAGEMENT_SETTINGS_ROUTE))


def _resolve_output_html_path(env_var: str, default_name: str, base_dir: Path) -> Path:
    raw_value = (os.getenv(env_var, default_name) or "").strip() or default_name
    path = Path(raw_value)
    if path.is_absolute():
        return path
    return base_dir / path


def _resolve_report_html_sources(base_dir: Path) -> dict[str, Path]:
    return {
        "dashboard.html": base_dir / "dashboard.html",
        "nested_view_report.html": _resolve_output_html_path(
            "JIRA_NESTED_VIEW_HTML_PATH", "nested_view_report.html", base_dir
        ),
        "missed_entries.html": _resolve_output_html_path(
            "JIRA_MISSED_ENTRIES_HTML_PATH", "missed_entries.html", base_dir
        ),
        "assignee_hours_report.html": _resolve_output_html_path(
            "JIRA_ASSIGNEE_HOURS_HTML_PATH", "assignee_hours_report.html", base_dir
        ),
        "rnd_data_story.html": _resolve_output_html_path(
            "JIRA_RND_STORY_HTML_PATH", "rnd_data_story.html", base_dir
        ),
        "planned_rmis_report.html": _resolve_output_html_path(
            "JIRA_PLANNED_RMIS_HTML_PATH", "planned_rmis_report.html", base_dir
        ),
        "gantt_chart_report.html": _resolve_output_html_path(
            "JIRA_GANTT_HTML_PATH", "gantt_chart_report.html", base_dir
        ),
        "phase_rmi_gantt_report.html": _resolve_output_html_path(
            "JIRA_PHASE_GANTT_HTML_PATH", "phase_rmi_gantt_report.html", base_dir
        ),
        "ipp_meeting_dashboard.html": _resolve_output_html_path(
            "IPP_PHASE_DASHBOARD_HTML_PATH", "ipp_meeting_dashboard.html", base_dir
        ),
        "rlt_leave_report.html": _resolve_output_html_path(
            "RLT_LEAVE_REPORT_HTML_PATH", "rlt_leave_report.html", base_dir
        ),
        "leaves_planned_calendar.html": _resolve_output_html_path(
            "JIRA_LEAVES_CALENDAR_HTML_PATH", "leaves_planned_calendar.html", base_dir
        ),
        "employee_performance_report.html": _resolve_output_html_path(
            "JIRA_EMPLOYEE_PERFORMANCE_HTML_PATH", "employee_performance_report.html", base_dir
        ),
        "planned_vs_dispensed_report.html": _resolve_output_html_path(
            "JIRA_PLANNED_VS_DISPENSED_HTML_PATH", "planned_vs_dispensed_report.html", base_dir
        ),
        "planned_actual_table_view.html": _resolve_output_html_path(
            "JIRA_PLANNED_ACTUAL_TABLE_VIEW_HTML_PATH", "planned_actual_table_view.html", base_dir
        ),
    }


def resolve_report_html_dir(base_dir: Path, folder_raw: str) -> Path:
    folder = (folder_raw or "").strip() or "report_html"
    path = Path(folder)
    if not path.is_absolute():
        path = base_dir / path
    return path


def sync_report_html(base_dir: Path, folder_raw: str) -> int:
    target_dir = resolve_report_html_dir(base_dir, folder_raw)
    target_dir.mkdir(parents=True, exist_ok=True)

    moved = 0
    for report_name, source_path in _resolve_report_html_sources(base_dir).items():
        if not source_path.exists() or not source_path.is_file():
            continue
        destination_path = target_dir / report_name
        if source_path.resolve() == destination_path.resolve():
            continue
        if destination_path.exists():
            destination_path.unlink()
        shutil.move(str(source_path), str(destination_path))
        moved += 1
        print(f"[report-html-sync] Moved: {source_path.name} -> {destination_path}")

    # Keep shared nav assets alongside reports so generated pages can always load them.
    for asset_name in ("shared-nav.css", "shared-nav.js", "shared-date-filter.js", "material-symbols.css"):
        source_candidates = [
            base_dir / asset_name,
            base_dir / "report_html" / asset_name,
        ]
        source_asset = next((p for p in source_candidates if p.exists() and p.is_file()), None)
        if not source_asset:
            continue
        destination_asset = target_dir / asset_name
        if source_asset.resolve() == destination_asset.resolve():
            continue
        shutil.copy2(str(source_asset), str(destination_asset))
        print(f"[report-html-sync] Synced asset: {destination_asset.name}")

    # Copy self-hosted font for fast icon loading (no Google CDN dependency).
    fonts_src = base_dir / "report_html" / "fonts"
    fonts_dst = target_dir / "fonts"
    if fonts_src.exists() and fonts_src.is_dir():
        fonts_dst.mkdir(parents=True, exist_ok=True)
        for f in fonts_src.glob("*.woff2"):
            dst = fonts_dst / f.name
            try:
                shutil.copy2(str(f), str(dst))
                print(f"[report-html-sync] Synced font: {f.name}")
            except PermissionError as exc:
                # On Windows, the font file can be momentarily locked by another process
                # (e.g. a running dev server or browser). Skip rather than crash.
                print(
                    f"[report-html-sync] Warning: could not sync font {f.name}: {exc}"
                )

    _materialize_refresh_widgets(target_dir)
    return moved


_LOCAL_MATERIAL_SYMBOLS_LINK = '<link rel="stylesheet" href="/material-symbols.css">'

# Match Google Material Symbols/Symbols Outlined CSS links for replacement with local.
_GOOGLE_MATERIAL_SYMBOLS_PATTERN = re.compile(
    r'<link[^>]+href\s*=\s*["\']https://fonts\.googleapis\.com/(?:css2\?family=Material\+Symbols\+Outlined[^"\']*|icon\?family=Material\+Icons(?:\+Outlined)?)["\'][^>]*>',
    re.I,
)


def _use_local_icons(html: str) -> str:
    """Replace Google font links with self-hosted Material Symbols for fast loading."""
    if "/material-symbols.css" in html:
        # Already using local; just remove any leftover Google links.
        html = _GOOGLE_MATERIAL_SYMBOLS_PATTERN.sub("", html)
        return html
    # Replace first Google Material Symbols/Icons link with local CSS.
    def _repl(m: re.Match[str]) -> str:
        return _LOCAL_MATERIAL_SYMBOLS_LINK

    html = _GOOGLE_MATERIAL_SYMBOLS_PATTERN.sub(_repl, html, count=1)
    # Remove any remaining Google font links (e.g. duplicate or preload).
    html = _GOOGLE_MATERIAL_SYMBOLS_PATTERN.sub("", html)
    # Remove preconnect/preload for Google fonts (no longer needed).
    html = re.sub(
        r'\s*<link\s+rel="(?:preconnect|preload)"[^>]+(?:fonts\.googleapis\.com|fonts\.gstatic\.com)[^>]*>',
        "",
        html,
        flags=re.I,
    )
    # Ensure we have the local stylesheet if we removed Google and none was added.
    needs_icons = (
        "material-symbols-outlined" in html
        or "material-icons-outlined" in html
        or "material-icons" in html
    )
    if "/material-symbols.css" not in html and needs_icons:
        html = re.sub(r"(\s*<head[^>]*>)", r"\1\n  " + _LOCAL_MATERIAL_SYMBOLS_LINK, html, count=1)
    return html


def _inject_refresh_ui(html: str, report_id: str) -> str:
    # Replace any previously injected widget (older/newer versions) to keep one instance.
    html = re.sub(
        r"<!-- codex-refresh-widget-start -->.*?<!-- codex-refresh-widget-end -->",
        "",
        html,
        flags=re.S,
    )
    html = re.sub(
        r"<!-- codex-refresh-widget-v1 -->.*?</script>",
        "",
        html,
        flags=re.S,
    )
    html = _inject_info_drawer_ui(html, report_id)
    if REFRESH_WIDGET_MARKER in html:
        return html
    snippet = f"""
{REFRESH_WIDGET_START}
<!-- {REFRESH_WIDGET_MARKER} -->
<style>
#codex-refresh-wrap {{
  display: flex;
  flex-direction: column;
  gap: 6px;
  margin-top: 8px;
  font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
}}
#codex-refresh-btn {{
  width: 100%;
  display: inline-flex;
  align-items: center;
  justify-content: center;
  gap: 6px;
  box-sizing: border-box;
  border: 1px solid rgba(126, 214, 242, 0.65);
  background: rgba(14, 116, 144, 0.58);
  color: #fff;
  border-radius: 9px;
  padding: 8px 12px;
  cursor: pointer;
  font-size: 13px;
  font-weight: 700;
}}
#codex-refresh-btn[disabled] {{
  cursor: not-allowed;
  opacity: 0.8;
}}
#codex-refresh-btn .material-symbols-outlined {{
  font-size: 18px;
  line-height: 1;
  display: inline-flex;
  align-items: center;
  justify-content: center;
}}
.app-shell.sidebar-collapsed #codex-refresh-btn,
.unified-nav.is-collapsed #codex-refresh-btn {{
  padding-left: 8px;
  padding-right: 8px;
  gap: 0;
  justify-content: center;
}}
.app-shell.sidebar-collapsed #codex-refresh-btn-label,
.unified-nav.is-collapsed #codex-refresh-btn-label {{
  display: none;
}}
.app-shell.sidebar-collapsed #codex-refresh-status,
.unified-nav.is-collapsed #codex-refresh-status {{
  display: none;
}}
#codex-refresh-status {{
  min-height: 14px;
  font-size: 11px;
  color: rgba(219, 246, 255, 0.9);
  text-align: left;
}}
</style>
<div id="codex-refresh-wrap" aria-live="polite">
  <button id="codex-refresh-btn" type="button">
    <span class="material-symbols-outlined" aria-hidden="true">refresh</span>
    <span id="codex-refresh-btn-label">Refresh</span>
  </button>
  <div id="codex-refresh-status"></div>
</div>
<script>
(function () {{
  const reportId = {report_id!r};
  const endpointPath = "/api/report/refresh";
  const wrap = document.getElementById("codex-refresh-wrap");
  const btn = document.getElementById("codex-refresh-btn");
  const btnLabel = document.getElementById("codex-refresh-btn-label");
  const status = document.getElementById("codex-refresh-status");
  if (!wrap || !btn || !btnLabel || !status || !reportId) return;

  function mountRefreshInNav() {{
    const navTargets = [
      ".app-sidebar .sidebar-nav",
      ".unified-nav .unified-nav-list",
      ".sidebar-nav",
      ".unified-nav-list",
    ];
    for (const selector of navTargets) {{
      const target = document.querySelector(selector);
      if (target) {{
        target.appendChild(wrap);
        return true;
      }}
    }}
    return false;
  }}

  if (!mountRefreshInNav()) {{
    const observer = new MutationObserver(function () {{
      if (mountRefreshInNav()) {{
        observer.disconnect();
      }}
    }});
    observer.observe(document.documentElement, {{ childList: true, subtree: true }});
    window.setTimeout(function () {{ observer.disconnect(); }}, 5000);
  }}

  function setBusy(isBusy) {{
    btn.disabled = isBusy;
    btnLabel.textContent = isBusy ? "Refreshing..." : "Refresh";
  }}

  function setStatus(msg) {{
    status.textContent = msg || "";
  }}

  function endpointCandidates() {{
    if (window.location.protocol === "http:" || window.location.protocol === "https:") {{
      return [endpointPath];
    }}
    const savedOrigin = String(localStorage.getItem("codex-report-server-origin") || "").trim();
    const defaults = [
      "http://127.0.0.1:8000",
      "http://localhost:8000",
      "http://127.0.0.1:5000",
      "http://localhost:5000",
    ];
    const origins = [savedOrigin, ...defaults]
      .map((item) => String(item || "").trim().replace(/\\/$/, ""))
      .filter(Boolean);
    return Array.from(new Set(origins)).map((origin) => origin + endpointPath);
  }}

  async function postRefresh(body) {{
    const endpoints = endpointCandidates();
    let lastError = "Failed to fetch";
    for (const endpoint of endpoints) {{
      try {{
        const response = await fetch(endpoint, {{
          method: "POST",
          headers: {{ "Content-Type": "application/json" }},
          body: JSON.stringify(body),
        }});
        if (endpoint.startsWith("http")) {{
          try {{
            localStorage.setItem("codex-report-server-origin", new URL(endpoint).origin);
          }} catch (_err) {{}}
        }}
        return response;
      }} catch (err) {{
        lastError = (err && err.message) ? err.message : String(err);
      }}
    }}
    throw new Error(lastError);
  }}

  btn.addEventListener("click", async function () {{
    setBusy(true);
    setStatus("Running scripts...");
    try {{
      const response = await postRefresh({{ report: reportId }});
      const payload = await response.json().catch(() => ({{}}));
      if (!response.ok || !payload.ok) {{
        const msg = payload.error || "Refresh failed.";
        setStatus(msg);
        alert(msg);
        return;
      }}
      const ts = payload.completed_at || "";
      setStatus(ts ? ("Updated: " + ts) : "Updated.");
      window.location.reload();
    }} catch (err) {{
      const base = (err && err.message) ? err.message : String(err);
      const hint = "If report is opened from file path, use server URL like http://127.0.0.1:8000/dashboard.html";
      const msg = base ? (base + ". " + hint) : hint;
      setStatus(msg);
      alert(msg);
    }} finally {{
      setBusy(false);
    }}
  }});
}})();
</script>
{REFRESH_WIDGET_END}
"""
    if "</body>" in html:
        return html.replace("</body>", snippet + "\n</body>", 1)
    return html + snippet


def _build_report_info_catalog(report_id: str) -> list[dict]:
    shared_capacity_source = [
        "assignee_hours_report.xlsx",
        "nested view.xlsx",
        "rlt_leave_report.xlsx",
        "/api/capacity",
    ]
    by_report: dict[str, list[dict]] = {
        "assignee_hours": [
            {
                "id": "assignee.capacity_subtraction",
                "label": "Capacity Subtraction (Hours)",
                "report": "assignee_hours_report",
                "ui_targets": ["#kpi-capacity-subtraction", "#kpi-capacity-subtraction-days"],
                "definition": "Available capacity remaining after project actual work and leave impact are subtracted.",
                "formula": "Available Capacity - Project Actual Hours - Leave Hours",
                "ingredients": ["available_capacity_hours", "project_actual_hours_non_rlt", "leave_hours_total"],
                "business_validations": ["Date range must be valid.", "Project actual excludes RLT leave project."],
                "field_linkages": ["Total Capacity", "Leave Hours", "Project Actual Hours"],
                "cross_report_linkages": ["nested.capacity_gap", "rnd.investable_more_hours"],
                "data_sources": shared_capacity_source,
                "leadership_interpretation": "Positive value indicates deployable hours; negative value indicates over-commitment.",
            },
            {
                "id": "assignee.project_plan_actual_gap",
                "label": "Project Plan - Actual Hours",
                "report": "assignee_hours_report",
                "ui_targets": ["#kpi-project-gap"],
                "definition": "Difference between planned project effort and actual logged project effort.",
                "formula": "Project Planned Hours - Project Actual Hours",
                "ingredients": ["project_planned_hours_non_rlt", "project_actual_hours_non_rlt"],
                "business_validations": ["Uses selected filter range.", "RLT project is excluded."],
                "field_linkages": ["Project Planned Hours", "Project Actual Hours"],
                "cross_report_linkages": ["nested.hours_required_to_complete", "rnd.hours_required_to_complete_projects"],
                "data_sources": ["1_jira_work_items_export.xlsx", "2_jira_subtask_worklogs.xlsx", "assignee_hours_report.xlsx"],
                "leadership_interpretation": "Higher positive gap means more pending delivery workload.",
            },
        ],
        "nested_view": [
            {
                "id": "nested.capacity_gap",
                "label": "Availability for more work",
                "report": "nested_view_report",
                "ui_targets": ["#score-capacity-gap-card", "#score-capacity-gap"],
                "definition": "Capacity remaining after planned project hours and planned RLT leave estimates.",
                "formula": "Total Capacity (Hours) - Total Planned Projects (Hours) - Total Leaves Planned",
                "ingredients": ["total_capacity_hours", "total_planned_projects_hours_non_rlt", "total_leaves_planned_rlt"],
                "business_validations": ["Date range filters applied.", "RLT exclusion/inclusion rules per card definition."],
                "field_linkages": ["Total Capacity", "Total Planned Projects", "Total Leaves Planned"],
                "cross_report_linkages": ["assignee.capacity_subtraction", "rnd.investable_more_hours"],
                "data_sources": shared_capacity_source,
                "leadership_interpretation": "Primary indicator of additional capacity or deficit.",
            },
            {
                "id": "nested.total_capacity_adjusted",
                "label": "Availability",
                "report": "nested_view_report",
                "ui_targets": ["#score-total-capacity-planned-leaves-adjusted-card"],
                "definition": "Capacity after subtracting planned leave load from total capacity.",
                "formula": "Total Capacity (Hours) - Total Leaves Planned",
                "ingredients": ["total_capacity_hours", "total_leaves_planned_rlt"],
                "business_validations": ["Date range filters applied.", "RLT day-bucketed planned leaves are deducted once."],
                "field_linkages": ["Total Capacity", "Total Leaves Planned"],
                "cross_report_linkages": ["rnd.leave_adjusted_capacity", "rlt.total_taken"],
                "data_sources": ["rlt_leave_report.xlsx", "/api/capacity", "nested view.xlsx"],
                "leadership_interpretation": "Represents practical delivery capacity after planned leave impact.",
            },
        ],
        "rnd_data_story": [
            {
                "id": "rnd.leave_adjusted_capacity",
                "label": "Leave-Adjusted Capacity",
                "report": "rnd_data_story",
                "ui_targets": ["#kpi-capacity-after-leaves"],
                "definition": "Capacity available after all leave categories are deducted.",
                "formula": "Available Capacity - (Planned Taken + Planned Not Taken Yet + Unplanned Taken)",
                "ingredients": ["available_capacity_hours", "planned_taken_hours", "planned_not_taken_hours", "unplanned_taken_hours"],
                "business_validations": ["Department scope fixed to RnD.", "Date range must be valid."],
                "field_linkages": ["Work On Plate", "Investable More Hours"],
                "cross_report_linkages": ["nested.total_capacity_adjusted", "rlt.total_taken"],
                "data_sources": shared_capacity_source,
                "leadership_interpretation": "Leadership baseline for commitment decisions.",
            },
            {
                "id": "rnd.total_leaves_planned",
                "label": "Total Leaves Planned",
                "report": "rnd_data_story",
                "ui_targets": ["#kpi-total-leaves-planned"],
                "definition": "Planned leave load based on day-bucketed leave data within the selected date range.",
                "formula": "Planned Taken + Planned Not Yet Taken",
                "ingredients": ["planned_taken_hours", "planned_not_taken_hours"],
                "business_validations": ["Date range filters applied to day rows.", "Same leave basis is used by capacity scorecards."],
                "field_linkages": ["Capacity", "Availability"],
                "cross_report_linkages": ["rlt.total_planned_leaves", "nested.total_capacity_adjusted"],
                "data_sources": ["rlt_leave_report.xlsx"],
                "leadership_interpretation": "Shows future planned leave effort to account for in planning decisions.",
            },
            {
                "id": "rnd.pending_hours_required",
                "label": "Pending Hours Required",
                "report": "rnd_data_story",
                "ui_targets": ["#kpi-hours-required-projects", "#funnel-hours-required"],
                "definition": "Remaining effort needed to complete scoped epics.",
                "formula": "Sum(max(Epic Original Estimate Hours - Epic Logged Hours, 0))",
                "ingredients": ["epic_original_estimate_hours", "epic_logged_hours"],
                "business_validations": ["Epic date inclusion rule: start OR end in selected range.", "Negative epic residuals clamp to zero."],
                "field_linkages": ["Investable More Hours", "Coverage Gap"],
                "cross_report_linkages": ["assignee.project_plan_actual_gap", "nested.hours_required_to_complete"],
                "data_sources": ["1_jira_work_items_export.xlsx", "2_jira_subtask_worklogs.xlsx"],
                "leadership_interpretation": "Represents demand backlog against available investable hours.",
            },
        ],
        "rlt_leave_report": [
            {
                "id": "rlt.total_taken",
                "label": "Total Taken",
                "report": "rlt_leave_report",
                "ui_targets": ["#stat-total-taken-hours", "#stat-total-taken-days"],
                "definition": "Total leave hours already consumed in selected date range.",
                "formula": "Planned Taken + Unplanned Taken",
                "ingredients": ["planned_taken_hours", "unplanned_taken_hours"],
                "business_validations": ["Hours are primary; days are derived by date-aware hours-per-day logic."],
                "field_linkages": ["Planned Taken", "Unplanned Taken"],
                "cross_report_linkages": ["nested.total_capacity_adjusted", "rnd.leave_adjusted_capacity"],
                "data_sources": ["rlt_leave_report.xlsx"],
                "leadership_interpretation": "Indicates realized leave impact already affecting delivery capacity.",
            },
            {
                "id": "rlt.total_planned_leaves",
                "label": "Total Leaves Planned",
                "report": "rlt_leave_report",
                "ui_targets": ["#stat-total-planned-leaves-hours"],
                "definition": "Total planned leave load in the selected range, combining already consumed planned leave and remaining planned leave.",
                "formula": "Planned Taken + Future Planned",
                "ingredients": ["planned_taken_hours", "planned_not_yet_taken_hours"],
                "business_validations": ["Computed from the same filtered window used by the RLT scorecards."],
                "field_linkages": ["Planned Taken", "Future Planned", "No Entry"],
                "cross_report_linkages": ["nested.total_capacity_adjusted", "nested.capacity_gap"],
                "data_sources": ["rlt_leave_report.xlsx"],
                "leadership_interpretation": "Forward-looking leave demand to reserve against delivery capacity.",
            },
            {
                "id": "rlt.future_planned",
                "label": "Future Planned",
                "report": "rlt_leave_report",
                "ui_targets": ["#stat-future-hours"],
                "definition": "Planned leave hours scheduled but not yet logged in the selected window.",
                "formula": "Sum(planned leave estimates with no matching consumed logs yet)",
                "ingredients": ["planned_not_yet_taken_hours"],
                "business_validations": ["Missing metadata rows are counted under No Entry."],
                "field_linkages": ["No Entry"],
                "cross_report_linkages": ["nested.total_capacity_adjusted", "assignee.leave_hours"],
                "data_sources": ["rlt_leave_report.xlsx"],
                "leadership_interpretation": "Forward-looking leave load that can constrain upcoming delivery capacity.",
            },
        ],
        "employee_performance": [
            {
                "id": "employee.team_avg_score",
                "label": "Team Avg Score",
                "report": "employee_performance_report",
                "ui_targets": ["#kpi-avg"],
                "definition": "Average score across assignees after configured penalty deductions.",
                "formula": "Average(clamp(base_score - total_penalty, min_score, max_score))",
                "ingredients": ["base_score", "penalty_settings", "assignee_penalty_totals"],
                "business_validations": ["Score clamp applies between min and max settings."],
                "field_linkages": ["Total Penalty", "At Risk (<60)"],
                "cross_report_linkages": ["assignee.project_plan_actual_gap", "rlt.total_taken"],
                "data_sources": ["assignee_hours_report.xlsx", "assignee_hours_capacity.db", "/api/performance/settings"],
                "leadership_interpretation": "Summarizes overall execution quality and delivery health.",
            },
            {
                "id": "employee.total_penalty",
                "label": "Total Penalty",
                "report": "employee_performance_report",
                "ui_targets": ["#kpi-pen"],
                "definition": "Aggregate penalty points from bugs, lateness, missed due dates, leave, and overruns.",
                "formula": "Sum(weighted penalty components by configured multipliers)",
                "ingredients": ["bug_hours", "bug_late_hours", "subtask_late_hours", "missed_due_date_count", "estimate_overrun_hours", "unplanned_leave_hours"],
                "business_validations": ["Penalty multipliers loaded from performance settings API."],
                "field_linkages": ["Team Avg Score", "At Risk (<60)"],
                "cross_report_linkages": ["dashboard.delivery_health", "missed_entries.missing_fields"],
                "data_sources": ["assignee_hours_report.xlsx", "rlt_leave_report.xlsx", "/api/performance/settings"],
                "leadership_interpretation": "Higher penalty signals growing delivery risk and coaching needs.",
            },
            {
                "id": "employee.capacity_per_employee",
                "label": "Employee Capacity",
                "report": "employee_performance_report",
                "ui_targets": ["#kpi-capacity"],
                "definition": "Effective per-assignee capacity for selected range after leave adjustment.",
                "formula": "Per-assignee baseline capacity - planned_leave_hours - unplanned_leave_hours",
                "ingredients": ["capacity_profiles", "planned_leave_hours", "unplanned_leave_hours"],
                "business_validations": ["Falls back to weekday*8h model when profile missing."],
                "field_linkages": ["Planned Hours Assigned", "Missed Start Ratio"],
                "cross_report_linkages": ["assignee.capacity_subtraction", "nested.total_capacity_adjusted"],
                "data_sources": ["assignee_hours_capacity.db", "rlt_leave_report.xlsx"],
                "leadership_interpretation": "Shows practical delivery bandwidth per employee.",
            },
            {
                "id": "employee.planned_hours_assigned",
                "label": "Planned Hours Assigned",
                "report": "employee_performance_report",
                "ui_targets": ["#kpi-planned-assigned"],
                "definition": "Total estimate hours assigned to the employee in selected range.",
                "formula": "Sum(original_estimate_hours over assigned in-scope items)",
                "ingredients": ["original_estimate_hours", "assignee", "date_range"],
                "business_validations": ["Rows without estimate contribute 0."],
                "field_linkages": ["Employee Capacity", "Assigned Items (E/S/ST)"],
                "cross_report_linkages": ["nested.total_planned_projects", "assignee.project_plan_actual_gap"],
                "data_sources": ["1_jira_work_items_export.xlsx"],
                "leadership_interpretation": "Indicates planned load committed to each assignee.",
            },
            {
                "id": "employee.assigned_counts",
                "label": "Assigned Item Counts",
                "report": "employee_performance_report",
                "ui_targets": ["#kpi-assigned-counts"],
                "definition": "Count split of assigned Epics, Stories, and Subtasks.",
                "formula": "Count(assigned items by type)",
                "ingredients": ["issue_type", "assignee", "parent_issue_key"],
                "business_validations": ["Unknown types are mapped to subtask bucket."],
                "field_linkages": ["Planned Hours Assigned", "Missed Start Ratio"],
                "cross_report_linkages": ["dashboard.delivery_health", "nested.tree_density"],
                "data_sources": ["1_jira_work_items_export.xlsx"],
                "leadership_interpretation": "Reveals hierarchy mix in assignee workload.",
            },
            {
                "id": "employee.missed_start_ratio",
                "label": "Missed Start Ratio",
                "report": "employee_performance_report",
                "ui_targets": ["#kpi-missed-start-ratio"],
                "definition": "Share of assigned items not started on planned start date.",
                "formula": "missed_start_count / total_assigned_count",
                "ingredients": ["start_date", "worklog_date", "assignee", "assigned_items"],
                "business_validations": ["Late-start context uses planned-start date only."],
                "field_linkages": ["Assigned Items (E/S/ST)", "Employee Capacity"],
                "cross_report_linkages": ["missed_entries.total_missed", "employee.total_penalty"],
                "data_sources": ["1_jira_work_items_export.xlsx", "2_jira_subtask_worklogs.xlsx", "rlt_leave_report.xlsx"],
                "leadership_interpretation": "Highlights schedule-adherence risk at assignee level.",
            },
        ],
        "gantt_chart": [
            {
                "id": "gantt.timeline_window",
                "label": "Timeline Window",
                "report": "gantt_chart_report",
                "ui_targets": ["#gantt-wrap", "#fit-range"],
                "definition": "Visible gantt range derived from planned dates and interactive zoom controls.",
                "formula": "Min/Max scoped planned dates with fit/reset controls",
                "ingredients": ["planned_start_date", "planned_end_date", "zoom_state"],
                "business_validations": ["Rows without valid ranges show missing-date state."],
                "field_linkages": ["Visible Rows"],
                "cross_report_linkages": ["phase_rmi.weekly_load", "nested.total_planned_projects"],
                "data_sources": ["nested view.xlsx", "1_jira_work_items_export.xlsx"],
                "leadership_interpretation": "Highlights schedule spread and overlap across execution layers.",
            }
        ],
        "leaves_planned_calendar": [
            {
                "id": "leave_calendar.intensity_mode",
                "label": "Color Intensity Mode",
                "report": "leaves_planned_calendar",
                "ui_targets": ["#mode-count", "#mode-hours"],
                "definition": "Calendar heatmap coloring can represent leave count or leave hours.",
                "formula": "Cell intensity based on selected mode aggregate per day",
                "ingredients": ["planned_leave_count_per_day", "planned_leave_hours_per_day"],
                "business_validations": ["Month filters must be valid.", "Apply updates rendered grid."],
                "field_linkages": ["Details Panel"],
                "cross_report_linkages": ["rlt.future_planned", "nested.total_capacity_adjusted"],
                "data_sources": ["rlt_leave_report.xlsx"],
                "leadership_interpretation": "Surfaces leave concentration hotspots by date.",
            }
        ],
        "missed_entries": [
            {
                "id": "missed_entries.total_missed",
                "label": "Total Missed Entries",
                "report": "missed_entries",
                "ui_targets": ["#total-missed-count"],
                "definition": "Count of rows missing selected mandatory planning fields.",
                "formula": "Count(rows where any selected missing field condition is true)",
                "ingredients": ["start_date", "end_date", "original_estimate", "missing_filter_selection"],
                "business_validations": ["Date range and field filters both affect totals."],
                "field_linkages": ["Summary table", "Assignee detail accordion"],
                "cross_report_linkages": ["employee.total_penalty", "dashboard.delivery_health"],
                "data_sources": ["1_jira_work_items_export.xlsx"],
                "leadership_interpretation": "Data quality indicator directly tied to planning confidence.",
            }
        ],
        "phase_rmi_gantt": [
            {
                "id": "phase_rmi.weekly_load",
                "label": "Weekly Load Chips",
                "report": "phase_rmi_gantt_report",
                "ui_targets": ["#gantt-root"],
                "definition": "Team lane chips summarize estimated epic workload intensity by week.",
                "formula": "Aggregate team epic man-days (story estimate hours / 8) overlapping each week bucket",
                "ingredients": ["story_assignee_team", "story_parent_epic", "story_estimate_hours", "epic_planned_start", "epic_planned_end"],
                "business_validations": ["Only stories with epic, valid date range, and positive estimate participate in lane load chips."],
                "field_linkages": ["Visible RMIs", "Date range controls"],
                "cross_report_linkages": ["gantt.timeline_window", "ipp.roadmap_geometry"],
                "data_sources": ["1_jira_work_items_export.xlsx", "assignee_hours_capacity.db:performance_teams", "assignee_hours_capacity.db:team_rmi_gantt_items"],
                "leadership_interpretation": "Shows team-level delivery pressure and weekly workload concentration by epic.",
            }
        ],
        "ipp_meeting": [
            {
                "id": "ipp.roadmap_geometry",
                "label": "Roadmap Geometry",
                "report": "ipp_meeting_dashboard",
                "ui_targets": ["#report-product-roadmap", "#rmi-roadmap-grid"],
                "definition": "Roadmap bar and tick positioning comes from computed workbook geometry fields.",
                "formula": "Use precomputed axis span, bar offsets, and week ticks from transformed sheet",
                "ingredients": ["Computed Roadmap Axis*", "Computed Roadmap Bar*", "Computed Phase Geometry JSON"],
                "business_validations": ["Rows flagged invalid are shown with warnings and fallback visuals."],
                "field_linkages": ["Mini gantt", "Phase breakdown cards"],
                "cross_report_linkages": ["phase_rmi.weekly_load", "dashboard.delivery_health"],
                "data_sources": ["ipp_phase_breakdown_*.xlsx"],
                "leadership_interpretation": "Provides meeting-ready schedule shape, slippage visibility, and phase confidence.",
            }
        ],
        "dashboard": [
            {
                "id": "dashboard.delivery_health",
                "label": "Delivery Health Cards",
                "report": "dashboard",
                "ui_targets": [".lane-epics", ".lane-stories", ".lane-subtasks"],
                "definition": "Cards combine schedule dates, status, IPP mismatch flags, and logged effort.",
                "formula": "Issue-level fields + hierarchy rollups + IPP sync flags",
                "ingredients": ["issue status", "planned dates", "actual dates", "total_hours_logged", "IPP sync indicators"],
                "business_validations": ["Hierarchy linkage: epic->story->subtask, including orphan handling."],
                "field_linkages": ["Date filter", "Project filter", "Status filter", "Assignee filters"],
                "cross_report_linkages": ["missed_entries.total_missed", "employee.total_penalty", "ipp.roadmap_geometry"],
                "data_sources": ["1_jira_work_items_export.xlsx", "2_jira_subtask_worklogs.xlsx", "3_jira_subtask_worklog_rollup.xlsx"],
                "leadership_interpretation": "Operational pulse view for execution, data consistency, and risk flags.",
            }
        ],
        "planned_vs_dispensed": [
            {
                "id": "planned_vs_dispensed.project_gap",
                "label": "Planned vs Dispensed Gap",
                "report": "planned_vs_dispensed_report",
                "ui_targets": ["#pvd-comparison-chart", "#pvd-detail-root"],
                "definition": "Compares epic-level planned estimates against subtotal of descendant subtask estimates per project.",
                "formula": "Project Planned (Epic Estimates) - Project Dispensed (Subtask Estimates)",
                "ingredients": ["epic.timeoriginalestimate", "subtask.timeoriginalestimate", "perspective_mode", "date_range"],
                "business_validations": ["Perspective mode must be valid.", "Date range must be valid.", "Hierarchy linkage epic->story->subtask is required."],
                "field_linkages": ["Date Range", "Advanced Filters", "Perspective"],
                "cross_report_linkages": ["assignee.project_plan_actual_gap", "rnd.pending_hours_required"],
                "data_sources": ["/api/planned-vs-dispensed/summary", "/api/planned-vs-dispensed/details"],
                "leadership_interpretation": "Positive gap indicates under-dispensing from planning to actionable subtasks; near-zero gap indicates healthy decomposition.",
            }
        ],
    }
    return by_report.get(report_id, [])


def _inject_info_drawer_ui(html: str, report_id: str) -> str:
    html = re.sub(
        r"<!-- codex-info-drawer-start -->.*?<!-- codex-info-drawer-end -->",
        "",
        html,
        flags=re.S,
    )
    if INFO_DRAWER_MARKER in html:
        return html

    catalog = _build_report_info_catalog(report_id)
    catalog_json = json.dumps(catalog, ensure_ascii=True)
    snippet = f"""
{INFO_DRAWER_START}
<!-- {INFO_DRAWER_MARKER} -->
<style>
.report-info-inline-btn {{
  margin-left: 6px;
  width: 18px;
  height: 18px;
  border-radius: 999px;
  border: 1px solid #94a3b8;
  background: #f8fafc;
  color: #334155;
  font-size: 11px;
  font-weight: 700;
  line-height: 1;
  cursor: pointer;
  display: inline-flex;
  align-items: center;
  justify-content: center;
  vertical-align: middle;
}}
.report-info-inline-btn:hover {{ background: #eef2ff; border-color: #64748b; }}
#report-info-drawer-backdrop {{
  position: fixed; inset: 0; background: rgba(15,23,42,.45); opacity: 0; visibility: hidden;
  transition: opacity .16s ease; z-index: 2500;
}}
#report-info-drawer-backdrop.is-open {{ opacity: 1; visibility: visible; }}
#report-info-drawer {{
  position: fixed; right: 0; top: 0; height: 100vh; width: min(520px, 94vw);
  background: #ffffff; border-left: 1px solid #cbd5e1; box-shadow: -16px 0 30px rgba(15,23,42,.18);
  transform: translateX(105%); transition: transform .18s ease; z-index: 2501; display: flex; flex-direction: column;
}}
#report-info-drawer.is-open {{ transform: translateX(0); }}
.report-info-drawer-head {{
  display: flex; align-items: center; justify-content: space-between; gap: 8px;
  padding: 12px 14px; border-bottom: 1px solid #e2e8f0; background: linear-gradient(180deg,#f8fafc,#f1f5f9);
}}
.report-info-drawer-title {{ margin: 0; font-size: 1rem; color: #0f172a; }}
.report-info-drawer-close {{
  border: 1px solid #94a3b8; background: #fff; color: #334155; border-radius: 8px; padding: 6px 8px; cursor: pointer; font-weight: 700;
}}
#report-info-drawer-body {{ padding: 12px 14px; overflow: auto; color: #1f2937; font-size: .86rem; line-height: 1.45; }}
.report-info-section {{ margin-bottom: 10px; }}
.report-info-section h4 {{ margin: 0 0 4px; font-size: .78rem; text-transform: uppercase; letter-spacing: .04em; color: #475569; }}
.report-info-section p {{ margin: 0; }}
.report-info-list {{ margin: 0; padding-left: 18px; }}
.report-info-list li {{ margin: 0 0 3px; }}
body.report-info-drawer-open {{ overflow: hidden; }}
</style>
<div id="report-info-drawer-backdrop" aria-hidden="true"></div>
<aside id="report-info-drawer" role="dialog" aria-modal="true" aria-hidden="true" aria-labelledby="report-info-drawer-title">
  <div class="report-info-drawer-head">
    <h3 class="report-info-drawer-title" id="report-info-drawer-title">Field Information</h3>
    <button type="button" class="report-info-drawer-close" id="report-info-drawer-close">Close</button>
  </div>
  <div id="report-info-drawer-body"></div>
</aside>
<script>
(function () {{
  const seeded = {catalog_json};
  const byId = Object.create(null);
  function txt(v) {{ return String(v == null ? "" : v).trim(); }}
  function asArr(v) {{ return Array.isArray(v) ? v : []; }}
  function slug(v) {{ return txt(v).toLowerCase().replace(/[^a-z0-9]+/g, "-").replace(/(^-|-$)/g, "") || "info"; }}
  function listHtml(values) {{
    const vals = asArr(values).map((x) => txt(x)).filter(Boolean);
    if (!vals.length) return "<p>-</p>";
    return "<ul class=\\"report-info-list\\">" + vals.map((x) => "<li>" + esc(x) + "</li>").join("") + "</ul>";
  }}
  function esc(v) {{
    return txt(v).replace(/&/g, "&amp;").replace(/</g, "&lt;").replace(/>/g, "&gt;");
  }}
  function addCatalog(item) {{
    if (!item || !txt(item.id)) return;
    byId[item.id] = Object.assign({{
      id: txt(item.id), label: "Field", report: txt(window.__activeReportId || ""),
      ui_targets: [], definition: "", formula: "", ingredients: [],
      business_validations: [], field_linkages: [], cross_report_linkages: [],
      data_sources: [], leadership_interpretation: ""
    }}, item);
  }}
  for (const item of asArr(seeded)) addCatalog(item);
  function ensureInfoButton(host, infoId, label) {{
    if (!host || !txt(infoId)) return;
    if (host.querySelector && host.querySelector('[data-info-id="' + infoId + '"]')) return;
    const btn = document.createElement("button");
    btn.type = "button";
    btn.className = "report-info-inline-btn";
    btn.textContent = "i";
    btn.setAttribute("data-info-id", infoId);
    btn.setAttribute("aria-label", (txt(label) || "Field") + " information");
    host.appendChild(btn);
  }}
  function extractFromTooltips() {{
    const defs = [
      [".score-info", ".score-info-tip"],
      [".kpi-info", ".kpi-info-tip"],
      [".stat-info", ".stat-info-tip"],
      [".card-i-wrap", ".card-i-tip"]
    ];
    for (const pair of defs) {{
      const wrapSel = pair[0];
      const tipSel = pair[1];
      const wraps = Array.from(document.querySelectorAll(wrapSel));
      for (const wrap of wraps) {{
        const tip = wrap.querySelector(tipSel);
        const raw = txt(tip ? tip.textContent : "");
        const parentLabelNode = wrap.closest(".score-label,.label,.k,.story-title,h3,h4,.card,.kpi,.stat,.summary-card");
        const label = txt(parentLabelNode ? parentLabelNode.textContent : "");
        const id = wrap.getAttribute("data-info-id") || slug((window.__activeReportId || "report") + "-" + (label || raw.slice(0, 36)));
        wrap.setAttribute("data-info-id", id);
        wrap.setAttribute("role", "button");
        wrap.setAttribute("tabindex", "0");
        if (tip) tip.style.display = "none";
        if (!byId[id]) {{
          addCatalog({{
            id: id,
            label: label || "Field",
            report: txt(window.__activeReportId || ""),
            ui_targets: [],
            definition: "Auto-captured from report tooltip content.",
            formula: raw.split("\\n")[0] || raw,
            ingredients: [],
            business_validations: [],
            field_linkages: [],
            cross_report_linkages: [],
            data_sources: [],
            leadership_interpretation: raw
          }});
        }}
      }}
    }}
  }}
  extractFromTooltips();
  for (const item of Object.values(byId)) {{
    for (const selector of asArr(item.ui_targets)) {{
      try {{
        const nodes = Array.from(document.querySelectorAll(selector));
        for (const node of nodes) {{
          const target = node.querySelector ? (node.querySelector(".label,.k,.score-label,h2,h3,h4,th") || node) : node;
          ensureInfoButton(target, item.id, item.label);
        }}
      }} catch (_err) {{}}
    }}
  }}
  const backdrop = document.getElementById("report-info-drawer-backdrop");
  const drawer = document.getElementById("report-info-drawer");
  const drawerBody = document.getElementById("report-info-drawer-body");
  const closeBtn = document.getElementById("report-info-drawer-close");
  let activeId = "";
  let lastFocus = null;
  function sectionHtml(title, contentHtml) {{
    return '<section class="report-info-section"><h4>' + esc(title) + '</h4>' + contentHtml + '</section>';
  }}
  function render(item) {{
    if (!item) {{
      drawerBody.innerHTML = "<p>No information is configured for this field yet.</p>";
      return;
    }}
    drawerBody.innerHTML =
      sectionHtml("Definition", "<p>" + esc(item.definition || "-") + "</p>") +
      sectionHtml("Formula", "<p>" + esc(item.formula || "-") + "</p>") +
      sectionHtml("Ingredients", listHtml(item.ingredients)) +
      sectionHtml("Business Validations", listHtml(item.business_validations)) +
      sectionHtml("Linked Fields (This Report)", listHtml(item.field_linkages)) +
      sectionHtml("Cross-Report Linkages", listHtml(item.cross_report_linkages)) +
      sectionHtml("Data Sources", listHtml(item.data_sources)) +
      sectionHtml("Leadership Interpretation", "<p>" + esc(item.leadership_interpretation || "-") + "</p>");
    const title = document.getElementById("report-info-drawer-title");
    if (title) title.textContent = txt(item.label) || "Field Information";
  }}
  function openDrawer(id) {{
    const item = byId[id];
    activeId = id;
    lastFocus = document.activeElement;
    render(item);
    backdrop.classList.add("is-open");
    drawer.classList.add("is-open");
    drawer.setAttribute("aria-hidden", "false");
    document.body.classList.add("report-info-drawer-open");
    if (closeBtn) closeBtn.focus();
  }}
  function closeDrawer() {{
    activeId = "";
    backdrop.classList.remove("is-open");
    drawer.classList.remove("is-open");
    drawer.setAttribute("aria-hidden", "true");
    document.body.classList.remove("report-info-drawer-open");
    if (lastFocus && lastFocus.focus) lastFocus.focus();
  }}
  function triggerFromEvent(e) {{
    const el = e.target.closest("[data-info-id]");
    if (!el) return;
    const id = txt(el.getAttribute("data-info-id"));
    if (!id) return;
    e.preventDefault();
    openDrawer(id);
  }}
  document.addEventListener("click", triggerFromEvent);
  document.addEventListener("keydown", function (e) {{
    if (e.key === "Escape" && activeId) {{
      e.preventDefault();
      closeDrawer();
      return;
    }}
    const inInfo = e.target && e.target.closest && e.target.closest("[data-info-id]");
    if (inInfo && (e.key === "Enter" || e.key === " ")) {{
      e.preventDefault();
      openDrawer(txt(inInfo.getAttribute("data-info-id")));
    }}
    if (!activeId || e.key !== "Tab") return;
    const f = drawer.querySelectorAll("button,[href],[tabindex]:not([tabindex='-1'])");
    const nodes = Array.from(f).filter((x) => !x.disabled);
    if (!nodes.length) return;
    const first = nodes[0];
    const last = nodes[nodes.length - 1];
    if (e.shiftKey && document.activeElement === first) {{ e.preventDefault(); last.focus(); }}
    else if (!e.shiftKey && document.activeElement === last) {{ e.preventDefault(); first.focus(); }}
  }});
  closeBtn.addEventListener("click", closeDrawer);
  backdrop.addEventListener("click", closeDrawer);
  window.reportInfoCatalog = Object.values(byId);
}})();
</script>
{INFO_DRAWER_END}
"""
    if "</body>" in html:
        return html.replace("</body>", snippet + "\n</body>", 1)
    return html + snippet


def _inject_shared_date_filter_script(html: str) -> str:
    if "shared-date-filter.js" in html:
        return html
    if '<script src="/shared-nav.js"></script>' in html:
        return html.replace(
            '<script src="/shared-nav.js"></script>',
            '<script src="/shared-date-filter.js"></script>\n  <script src="/shared-nav.js"></script>',
            1,
        )
    if '<script src="shared-nav.js"></script>' in html:
        return html.replace(
            '<script src="shared-nav.js"></script>',
            '<script src="shared-date-filter.js"></script>\n<script src="shared-nav.js"></script>',
            1,
        )
    if "<script src='shared-nav.js'></script>" in html:
        return html.replace(
            "<script src='shared-nav.js'></script>",
            "<script src='shared-date-filter.js'></script>\n<script src='shared-nav.js'></script>",
            1,
        )
    if "</body>" in html:
        return html.replace("</body>", '<script src="/shared-date-filter.js"></script>\n</body>', 1)
    return html + '\n<script src="/shared-date-filter.js"></script>\n'


def _materialize_refresh_widgets(report_dir: Path) -> None:
    for file_name, report_id in REPORT_FILENAME_TO_ID.items():
        html_path = report_dir / file_name
        if not html_path.exists() or not html_path.is_file():
            continue
        try:
            html = html_path.read_text(encoding="utf-8")
        except OSError as exc:
            print(
                f"[report-html-sync] Warning: could not read {html_path.name}: {exc}"
            )
            continue

        updated = _inject_refresh_ui(html, report_id)
        updated = _inject_shared_date_filter_script(updated)
        if updated == html:
            continue

        try:
            html_path.write_text(updated, encoding="utf-8")
            print(f"[report-html-sync] Added refresh widget: {html_path.name}")
        except OSError as exc:
            # On Windows, a browser/preview process may briefly lock report files.
            # Skip the single file and keep the server startup running.
            print(
                f"[report-html-sync] Warning: could not update {html_path.name}: {exc}"
            )


def _run_script(script_name: str, base_dir: Path) -> tuple[int, str, str]:
    script_path = base_dir / script_name
    if not script_path.exists():
        raise FileNotFoundError(f"Missing script: {script_path}")
    result = subprocess.run(
        [sys.executable, str(script_path)],
        cwd=str(base_dir),
        capture_output=True,
        text=True,
    )
    return result.returncode, result.stdout, result.stderr


def _tail(text: str, lines: int = 30) -> str:
    entries = (text or "").splitlines()
    if len(entries) <= lines:
        return text or ""
    return "\n".join(entries[-lines:])


def _parse_iso_date(value: str) -> date | None:
    text = (value or "").strip()
    if not text:
        return None
    if len(text) >= 10:
        text = text[:10]
    try:
        return date.fromisoformat(text)
    except ValueError:
        return None


def _utc_now_iso_z() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def _round_hours(value: float) -> float:
    return round(float(value), 2)


def _resolve_worklog_xlsx_path(base_dir: Path) -> Path:
    worklog_name = (os.getenv("JIRA_WORKLOG_XLSX_PATH", "2_jira_subtask_worklogs.xlsx") or "").strip() or "2_jira_subtask_worklogs.xlsx"
    path = Path(worklog_name)
    if not path.is_absolute():
        path = base_dir / path
    return path


def _resolve_work_items_xlsx_path(base_dir: Path) -> Path:
    work_items_name = (os.getenv("JIRA_EXPORT_XLSX_PATH", "1_jira_work_items_export.xlsx") or "").strip() or "1_jira_work_items_export.xlsx"
    path = Path(work_items_name)
    if not path.is_absolute():
        path = base_dir / path
    return path


def _priority_for_epics_management(value: object) -> str:
    text = _to_text(value).casefold()
    if text == "highest":
        return "Highest"
    if text == "high":
        return "High"
    if text in {"medium", "meidum"}:
        return "Medium"
    return "Low"


def _plan_status_for_epics_management(value: object) -> str:
    text = _to_text(value).casefold()
    if text == "planned":
        return "Planned"
    if text in {"not planned yet", "not planned", "not_planned_yet", "plan"}:
        return "Not Planned Yet"
    return "Not Planned Yet"


def _ipp_meeting_planned_for_epics_management(value: object) -> str:
    text = _to_text(value).casefold()
    if text in {"yes", "y", "true", "1"}:
        return "Yes"
    return "No"


_EPICS_MANAGEMENT_DEFAULT_PLAN_COLUMNS: tuple[dict[str, object], ...] = (
    {"key": "epic_plan", "label": "Epic Plan", "jira_link_enabled": False, "sort_order": 0},
    {"key": "research_urs_plan", "label": "Research/URS Plan", "jira_link_enabled": True, "sort_order": 1},
    {"key": "dds_plan", "label": "DDS Plan", "jira_link_enabled": True, "sort_order": 2},
    {"key": "development_plan", "label": "Development Plan", "jira_link_enabled": True, "sort_order": 3},
    {"key": "sqa_plan", "label": "SQA Plan", "jira_link_enabled": True, "sort_order": 4},
    {"key": "user_manual_plan", "label": "User Manual Plan", "jira_link_enabled": True, "sort_order": 5},
    {"key": "production_plan", "label": "Production Plan", "jira_link_enabled": True, "sort_order": 6},
)
_EPICS_MANAGEMENT_LEGACY_PLAN_JSON_COLUMN_BY_KEY: dict[str, str] = {
    "epic_plan": "epic_plan_json",
    "research_urs_plan": "research_urs_plan_json",
    "dds_plan": "dds_plan_json",
    "development_plan": "development_plan_json",
    "sqa_plan": "sqa_plan_json",
    "user_manual_plan": "user_manual_plan_json",
    "production_plan": "production_plan_json",
}
_EPICS_MANAGEMENT_DEFAULT_PLAN_KEYS = tuple(
    item["key"] for item in _EPICS_MANAGEMENT_DEFAULT_PLAN_COLUMNS if _to_text(item.get("key"))
)
_EPIC_KEY_PATTERN = re.compile(r"^[A-Z0-9]+-\d+$")
_TMP_EPIC_KEY_PATTERN = re.compile(r"^TMP-\d{8}T\d{6}Z-[A-Z0-9]{6}$")
_EPICS_DROPDOWN_FIELD_KEYS = ("product_category", "component", "plan_status")
_EPICS_DROPDOWN_FIELD_KEY_ALIASES = {
    "product_category": "product_category",
    "product_categorization": "product_category",
    "product categorization": "product_category",
    "component": "component",
    "components": "component",
    "plan_status": "plan_status",
    "plan status": "plan_status",
    "plan_statuses": "plan_status",
    "plan statuses": "plan_status",
}


class _EpicCreateConflictError(ValueError):
    def __init__(self, message: str, *, conflict_epic_key: str = "", vacant_tmp_key: str = "") -> None:
        super().__init__(message)
        self.conflict_epic_key = _to_text(conflict_epic_key).upper()
        self.vacant_tmp_key = _to_text(vacant_tmp_key).upper()


def _normalize_epics_dropdown_field_key(value: object) -> str:
    key = _to_text(value).lower()
    normalized = _EPICS_DROPDOWN_FIELD_KEY_ALIASES.get(key, "")
    if not normalized:
        raise ValueError("field_key must be one of: product_category, component, plan_status.")
    return normalized


def _normalize_epics_dropdown_option_values(values: object) -> list[str]:
    if values is None:
        return []
    if not isinstance(values, list):
        raise ValueError("Dropdown options must be a JSON array of strings.")
    out: list[str] = []
    seen: set[str] = set()
    for raw in values:
        text = _to_text(raw)
        if not text:
            continue
        lower = text.casefold()
        if lower in seen:
            continue
        seen.add(lower)
        out.append(text)
    return out


def _utc_now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def _normalize_plan_column_key(value: object) -> str:
    key = re.sub(r"[^a-z0-9]+", "_", _to_text(value).lower()).strip("_")
    if not key:
        raise ValueError("Plan column key is required.")
    if len(key) > 64:
        raise ValueError("Plan column key must be 64 characters or fewer.")
    return key


def _normalize_plan_column_label(value: object) -> str:
    label = _to_text(value)
    if not label:
        raise ValueError("Plan column label is required.")
    if len(label) > 80:
        raise ValueError("Plan column label must be 80 characters or fewer.")
    return label


def _normalize_plan_column_jira_enabled(value: object) -> int:
    if isinstance(value, bool):
        return 1 if value else 0
    text = _to_text(value).casefold()
    if text in {"1", "true", "yes", "y"}:
        return 1
    return 0


def _normalize_plan_column_insert_position(value: object, total_count: int) -> int:
    if value in (None, ""):
        return max(total_count + 1, 1)
    try:
        pos = int(value)
    except Exception:
        raise ValueError("insert_position must be an integer.")
    if pos < 1 or pos > (total_count + 1):
        raise ValueError(f"insert_position must be between 1 and {total_count + 1}.")
    return pos


def _is_tmp_epic_key(value: object) -> bool:
    return bool(_TMP_EPIC_KEY_PATTERN.match(_to_text(value).upper()))


def _load_epics_plan_columns_from_conn(conn: sqlite3.Connection, include_inactive: bool = False) -> list[dict[str, object]]:
    where_sql = "" if include_inactive else "WHERE is_active = 1"
    rows = conn.execute(
        f"""
        SELECT column_key, label, jira_link_enabled, is_default, is_active, sort_order
        FROM epics_management_plan_columns
        {where_sql}
        ORDER BY sort_order ASC, lower(label) ASC, column_key ASC
        """
    ).fetchall()
    return [
        {
            "key": _to_text(row[0]),
            "label": _to_text(row[1]),
            "jira_link_enabled": bool(int(row[2] or 0)),
            "is_default": bool(int(row[3] or 0)),
            "is_active": bool(int(row[4] or 0)),
            "sort_order": int(row[5] or 0),
        }
        for row in rows
    ]


def _load_epics_plan_columns(settings_db_path: Path, include_inactive: bool = False) -> list[dict[str, object]]:
    _init_epics_management_db(settings_db_path)
    conn = sqlite3.connect(settings_db_path)
    try:
        return _load_epics_plan_columns_from_conn(conn, include_inactive=include_inactive)
    finally:
        conn.close()


def _seed_default_epics_plan_columns(conn: sqlite3.Connection) -> None:
    now_utc = _utc_now_iso()
    for col in _EPICS_MANAGEMENT_DEFAULT_PLAN_COLUMNS:
        key = _to_text(col.get("key"))
        if not key:
            continue
        conn.execute(
            """
            INSERT OR IGNORE INTO epics_management_plan_columns (
                column_key, label, jira_link_enabled, is_default, is_active, sort_order, created_at_utc, updated_at_utc
            ) VALUES (?, ?, ?, 1, 1, ?, ?, ?)
            """,
            (
                key,
                _to_text(col.get("label")),
                1 if bool(col.get("jira_link_enabled")) else 0,
                int(col.get("sort_order") or 0),
                now_utc,
                now_utc,
            ),
        )


def _backfill_legacy_epics_plan_values(conn: sqlite3.Connection, epics_columns: set[str]) -> None:
    now_utc = _utc_now_iso()
    for plan_key, legacy_col in _EPICS_MANAGEMENT_LEGACY_PLAN_JSON_COLUMN_BY_KEY.items():
        if legacy_col not in epics_columns:
            continue
        conn.execute(
            f"""
            INSERT OR IGNORE INTO epics_management_plan_values (epic_key, column_key, plan_json, created_at_utc, updated_at_utc)
            SELECT
                epic_key,
                ?,
                COALESCE(NULLIF(TRIM({legacy_col}), ''), '{{}}'),
                ?,
                ?
            FROM epics_management
            """,
            (plan_key, now_utc, now_utc),
        )


def _create_epics_plan_column(settings_db_path: Path, payload: dict) -> dict[str, object]:
    _init_epics_management_db(settings_db_path)
    raw = payload if isinstance(payload, dict) else {}
    label = _normalize_plan_column_label(raw.get("label"))
    requested_key = _to_text(raw.get("key"))
    base_key = _normalize_plan_column_key(requested_key or label)
    jira_link_enabled = _normalize_plan_column_jira_enabled(raw.get("jira_link_enabled"))

    conn = sqlite3.connect(settings_db_path)
    conn.row_factory = sqlite3.Row
    try:
        existing = conn.execute(
            "SELECT 1 FROM epics_management_plan_columns WHERE is_active=1 AND lower(label)=lower(?)",
            (label,),
        ).fetchone()
        if existing:
            raise ValueError(f"Plan column '{label}' already exists.")

        taken = {
            _to_text(row["column_key"])
            for row in conn.execute("SELECT column_key FROM epics_management_plan_columns").fetchall()
        }
        key = base_key
        suffix = 2
        while key in taken:
            key = f"{base_key}_{suffix}"
            suffix += 1

        active_count = conn.execute(
            "SELECT COUNT(*) FROM epics_management_plan_columns WHERE is_active=1"
        ).fetchone()[0]
        insert_position = _normalize_plan_column_insert_position(raw.get("insert_position"), int(active_count))
        insert_sort_order = max(insert_position - 1, 0)
        conn.execute(
            """
            UPDATE epics_management_plan_columns
            SET sort_order = sort_order + 1, updated_at_utc = ?
            WHERE is_active = 1 AND sort_order >= ?
            """,
            (_utc_now_iso(), insert_sort_order),
        )
        now_utc = _utc_now_iso()
        conn.execute(
            """
            INSERT INTO epics_management_plan_columns (
                column_key, label, jira_link_enabled, is_default, is_active, sort_order, created_at_utc, updated_at_utc
            ) VALUES (?, ?, ?, 0, 1, ?, ?, ?)
            """,
            (key, label, jira_link_enabled, insert_sort_order, now_utc, now_utc),
        )
        conn.commit()
        rows = _load_epics_plan_columns_from_conn(conn, include_inactive=False)
        created = next((item for item in rows if _to_text(item.get("key")) == key), None)
        if not created:
            raise RuntimeError("Failed to load created plan column.")
        return created
    finally:
        conn.close()


def _restore_epics_plan_column(settings_db_path: Path, column_key: str) -> dict[str, object]:
    _init_epics_management_db(settings_db_path)
    key = _to_text(column_key).strip().lower()
    if not key:
        raise ValueError("Plan column key is required.")
    if not re.fullmatch(r"[a-z0-9_]{1,64}", key):
        raise ValueError("Plan column key is invalid.")

    conn = sqlite3.connect(settings_db_path)
    conn.row_factory = sqlite3.Row
    try:
        row = conn.execute(
            """
            SELECT column_key, label, jira_link_enabled, is_default, is_active, sort_order
            FROM epics_management_plan_columns
            WHERE column_key=?
            """,
            (key,),
        ).fetchone()
        if not row:
            raise LookupError(f"Plan column '{key}' not found.")
        if bool(int(row["is_active"] or 0)):
            out = _load_epics_plan_columns_from_conn(conn, include_inactive=False)
            match = next((item for item in out if _to_text(item.get("key")) == key), None)
            if not match:
                raise RuntimeError("Failed to load active plan column.")
            return match

        max_sort = conn.execute(
            "SELECT COALESCE(MAX(sort_order), -1) FROM epics_management_plan_columns WHERE is_active=1"
        ).fetchone()[0]
        now_utc = _utc_now_iso()
        conn.execute(
            """
            UPDATE epics_management_plan_columns
            SET is_active=1, sort_order=?, updated_at_utc=?
            WHERE column_key=?
            """,
            (int(max_sort) + 1, now_utc, key),
        )
        conn.commit()
        out = _load_epics_plan_columns_from_conn(conn, include_inactive=False)
        match = next((item for item in out if _to_text(item.get("key")) == key), None)
        if not match:
            raise RuntimeError("Failed to load restored plan column.")
        return match
    finally:
        conn.close()


def _reorder_epics_plan_columns(settings_db_path: Path, payload: dict) -> list[dict[str, object]]:
    _init_epics_management_db(settings_db_path)
    raw = payload if isinstance(payload, dict) else {}
    ordered_keys_raw = raw.get("ordered_keys")
    if not isinstance(ordered_keys_raw, list) or not ordered_keys_raw:
        raise ValueError("ordered_keys must be a non-empty array of active column keys.")
    ordered_keys = [_to_text(item) for item in ordered_keys_raw if _to_text(item)]
    if len(ordered_keys) != len(ordered_keys_raw):
        raise ValueError("ordered_keys cannot contain blank values.")
    if len(set(ordered_keys)) != len(ordered_keys):
        raise ValueError("ordered_keys cannot contain duplicates.")

    conn = sqlite3.connect(settings_db_path)
    conn.row_factory = sqlite3.Row
    try:
        active_rows = conn.execute(
            "SELECT column_key FROM epics_management_plan_columns WHERE is_active=1 ORDER BY sort_order ASC, column_key ASC"
        ).fetchall()
        active_keys = [_to_text(row["column_key"]) for row in active_rows]
        if set(ordered_keys) != set(active_keys):
            raise ValueError("ordered_keys must contain exactly all active plan column keys.")
        now_utc = _utc_now_iso()
        for idx, key in enumerate(ordered_keys):
            conn.execute(
                """
                UPDATE epics_management_plan_columns
                SET sort_order=?, updated_at_utc=?
                WHERE column_key=?
                """,
                (idx, now_utc, key),
            )
        conn.commit()
        return _load_epics_plan_columns_from_conn(conn, include_inactive=False)
    finally:
        conn.close()


def _update_epics_plan_column(settings_db_path: Path, column_key: str, payload: dict) -> dict[str, object]:
    _init_epics_management_db(settings_db_path)
    key = _to_text(column_key).strip().lower()
    if not key:
        raise ValueError("Plan column key is required.")
    if not re.fullmatch(r"[a-z0-9_]{1,64}", key):
        raise ValueError("Plan column key is invalid.")

    raw = payload if isinstance(payload, dict) else {}
    has_label = "label" in raw
    has_jira_enabled = "jira_link_enabled" in raw
    if not has_label and not has_jira_enabled:
        raise ValueError("At least one field is required: label, jira_link_enabled.")

    conn = sqlite3.connect(settings_db_path)
    conn.row_factory = sqlite3.Row
    try:
        row = conn.execute(
            """
            SELECT column_key, label, jira_link_enabled, is_default, is_active, sort_order
            FROM epics_management_plan_columns
            WHERE column_key=?
            """,
            (key,),
        ).fetchone()
        if not row:
            raise LookupError(f"Plan column '{key}' not found.")

        next_label = _to_text(row["label"])
        if has_label:
            next_label = _normalize_plan_column_label(raw.get("label"))
            duplicate = conn.execute(
                """
                SELECT 1
                FROM epics_management_plan_columns
                WHERE is_active=1 AND column_key<>? AND lower(label)=lower(?)
                """,
                (key, next_label),
            ).fetchone()
            if duplicate:
                raise ValueError(f"Plan column '{next_label}' already exists.")

        next_jira_enabled = int(row["jira_link_enabled"] or 0)
        if has_jira_enabled:
            next_jira_enabled = _normalize_plan_column_jira_enabled(raw.get("jira_link_enabled"))

        now_utc = _utc_now_iso()
        conn.execute(
            """
            UPDATE epics_management_plan_columns
            SET label=?, jira_link_enabled=?, updated_at_utc=?
            WHERE column_key=?
            """,
            (next_label, next_jira_enabled, now_utc, key),
        )
        conn.commit()
        rows = _load_epics_plan_columns_from_conn(conn, include_inactive=True)
        updated = next((item for item in rows if _to_text(item.get("key")) == key), None)
        if not updated:
            raise RuntimeError("Failed to load updated plan column.")
        return updated
    finally:
        conn.close()


def _delete_epics_plan_column(settings_db_path: Path, column_key: str) -> list[dict[str, object]]:
    _init_epics_management_db(settings_db_path)
    key = _to_text(column_key).strip().lower()
    if not key:
        raise ValueError("Plan column key is required.")
    if not re.fullmatch(r"[a-z0-9_]{1,64}", key):
        raise ValueError("Plan column key is invalid.")

    conn = sqlite3.connect(settings_db_path)
    conn.row_factory = sqlite3.Row
    try:
        row = conn.execute(
            """
            SELECT column_key, label, is_default, is_active
            FROM epics_management_plan_columns
            WHERE column_key=?
            """,
            (key,),
        ).fetchone()
        if not row:
            raise LookupError(f"Plan column '{key}' not found.")
        if bool(int(row["is_default"] or 0)):
            raise ValueError("Default plan columns cannot be deleted.")
        if not bool(int(row["is_active"] or 0)):
            return _load_epics_plan_columns_from_conn(conn, include_inactive=False)

        now_utc = _utc_now_iso()
        conn.execute(
            """
            UPDATE epics_management_plan_columns
            SET is_active=0, updated_at_utc=?
            WHERE column_key=?
            """,
            (now_utc, key),
        )

        active_rows = conn.execute(
            """
            SELECT column_key
            FROM epics_management_plan_columns
            WHERE is_active=1
            ORDER BY sort_order ASC, lower(label) ASC, column_key ASC
            """
        ).fetchall()
        for idx, active_row in enumerate(active_rows):
            conn.execute(
                """
                UPDATE epics_management_plan_columns
                SET sort_order=?, updated_at_utc=?
                WHERE column_key=?
                """,
                (idx, now_utc, _to_text(active_row["column_key"])),
            )
        conn.commit()
        return _load_epics_plan_columns_from_conn(conn, include_inactive=False)
    finally:
        conn.close()


def _load_epics_dropdown_options(settings_db_path: Path) -> dict[str, list[str]]:
    _init_epics_management_db(settings_db_path)
    conn = sqlite3.connect(settings_db_path)
    conn.row_factory = sqlite3.Row
    try:
        rows = conn.execute(
            """
            SELECT field_key, option_value
            FROM epics_management_dropdown_options
            WHERE is_active = 1
            ORDER BY field_key ASC, sort_order ASC, lower(option_value) ASC
            """
        ).fetchall()
    finally:
        conn.close()
    grouped: dict[str, list[str]] = {key: [] for key in _EPICS_DROPDOWN_FIELD_KEYS}
    for row in rows:
        field_key = _to_text(row["field_key"])
        option_value = _to_text(row["option_value"])
        if field_key in grouped and option_value:
            grouped[field_key].append(option_value)
    return grouped


def _save_epics_dropdown_options(settings_db_path: Path, payload: dict) -> dict[str, list[str]]:
    _init_epics_management_db(settings_db_path)
    raw = payload if isinstance(payload, dict) else {}
    existing = _load_epics_dropdown_options(settings_db_path)
    updates: dict[str, list[str]] = {}
    for input_key, value in raw.items():
        normalized_key = _normalize_epics_dropdown_field_key(input_key)
        updates[normalized_key] = _normalize_epics_dropdown_option_values(value)
    if not updates:
        raise ValueError("At least one dropdown options list is required.")
    merged = {key: list(existing.get(key, [])) for key in _EPICS_DROPDOWN_FIELD_KEYS}
    for field_key, values in updates.items():
        merged[field_key] = values

    now_utc = datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")
    conn = sqlite3.connect(settings_db_path)
    try:
        for field_key in _EPICS_DROPDOWN_FIELD_KEYS:
            conn.execute(
                "DELETE FROM epics_management_dropdown_options WHERE field_key = ?",
                (field_key,),
            )
            for idx, option in enumerate(merged[field_key]):
                conn.execute(
                    """
                    INSERT INTO epics_management_dropdown_options (
                        field_key, option_value, sort_order, is_active, created_at_utc, updated_at_utc
                    ) VALUES (?, ?, ?, 1, ?, ?)
                    """,
                    (field_key, option, idx, now_utc, now_utc),
                )
        conn.commit()
    finally:
        conn.close()
    return _load_epics_dropdown_options(settings_db_path)


def _init_epics_management_db(settings_db_path: Path) -> None:
    settings_db_path.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(settings_db_path)
    try:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS epics_management (
                epic_key TEXT PRIMARY KEY,
                project_key TEXT NOT NULL,
                project_name TEXT NOT NULL,
                product_category TEXT NOT NULL,
                component TEXT NOT NULL DEFAULT '',
                epic_name TEXT NOT NULL,
                description TEXT NOT NULL DEFAULT '',
                originator TEXT NOT NULL DEFAULT '',
                priority TEXT NOT NULL DEFAULT 'Low',
                plan_status TEXT NOT NULL DEFAULT 'Not Planned Yet',
                ipp_meeting_planned TEXT NOT NULL DEFAULT 'No',
                actual_production_date TEXT NOT NULL DEFAULT '',
                remarks TEXT NOT NULL DEFAULT '',
                jira_url TEXT NOT NULL DEFAULT '',
                epic_plan_json TEXT NOT NULL DEFAULT '{}',
                research_urs_plan_json TEXT NOT NULL DEFAULT '{}',
                dds_plan_json TEXT NOT NULL DEFAULT '{}',
                development_plan_json TEXT NOT NULL DEFAULT '{}',
                sqa_plan_json TEXT NOT NULL DEFAULT '{}',
                user_manual_plan_json TEXT NOT NULL DEFAULT '{}',
                production_plan_json TEXT NOT NULL DEFAULT '{}'
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS epics_management_story_sync (
                story_key TEXT PRIMARY KEY,
                epic_key TEXT NOT NULL,
                project_key TEXT NOT NULL DEFAULT '',
                story_name TEXT NOT NULL DEFAULT '',
                story_status TEXT NOT NULL DEFAULT '',
                jira_url TEXT NOT NULL DEFAULT '',
                start_date TEXT NOT NULL DEFAULT '',
                due_date TEXT NOT NULL DEFAULT '',
                estimate_hours REAL NOT NULL DEFAULT 0,
                payload_json TEXT NOT NULL DEFAULT '{}',
                synced_at_utc TEXT NOT NULL DEFAULT ''
            )
            """
        )
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_epics_management_story_sync_epic_key ON epics_management_story_sync(epic_key)"
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS epics_management_dropdown_options (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                field_key TEXT NOT NULL,
                option_value TEXT NOT NULL,
                sort_order INTEGER NOT NULL DEFAULT 0,
                is_active INTEGER NOT NULL DEFAULT 1,
                created_at_utc TEXT NOT NULL,
                updated_at_utc TEXT NOT NULL
            )
            """
        )
        conn.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_epics_management_dropdown_field
            ON epics_management_dropdown_options(field_key, is_active, sort_order, option_value)
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS epics_management_plan_columns (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                column_key TEXT NOT NULL UNIQUE,
                label TEXT NOT NULL,
                jira_link_enabled INTEGER NOT NULL DEFAULT 0,
                is_default INTEGER NOT NULL DEFAULT 0,
                is_active INTEGER NOT NULL DEFAULT 1,
                sort_order INTEGER NOT NULL DEFAULT 0,
                created_at_utc TEXT NOT NULL,
                updated_at_utc TEXT NOT NULL
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS epics_management_plan_values (
                epic_key TEXT NOT NULL,
                column_key TEXT NOT NULL,
                plan_json TEXT NOT NULL DEFAULT '{}',
                created_at_utc TEXT NOT NULL,
                updated_at_utc TEXT NOT NULL,
                PRIMARY KEY(epic_key, column_key)
            )
            """
        )
        conn.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_epics_management_plan_columns_active
            ON epics_management_plan_columns(is_active, sort_order, label)
            """
        )
        conn.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_epics_management_plan_values_column
            ON epics_management_plan_values(column_key)
            """
        )
        columns = conn.execute("PRAGMA table_info(epics_management)").fetchall()
        names = {str(col[1]) for col in columns}
        col_defaults = {str(col[1]): col[4] for col in columns}
        needs_rebuild = (
            "source_workbook" in names
            or "source_sheet" in names
            or ("created_at_utc" in names and col_defaults.get("created_at_utc") is None)
        )
        if needs_rebuild:
            conn.execute(
                """
                CREATE TABLE epics_management_v2 (
                    epic_key TEXT PRIMARY KEY,
                    project_key TEXT NOT NULL,
                    project_name TEXT NOT NULL,
                    product_category TEXT NOT NULL,
                    component TEXT NOT NULL DEFAULT '',
                    epic_name TEXT NOT NULL,
                    description TEXT NOT NULL DEFAULT '',
                    originator TEXT NOT NULL DEFAULT '',
                    priority TEXT NOT NULL DEFAULT 'Low',
                    plan_status TEXT NOT NULL DEFAULT 'Not Planned Yet',
                    ipp_meeting_planned TEXT NOT NULL DEFAULT 'No',
                    actual_production_date TEXT NOT NULL DEFAULT '',
                    remarks TEXT NOT NULL DEFAULT '',
                    jira_url TEXT NOT NULL DEFAULT '',
                    epic_plan_json TEXT NOT NULL DEFAULT '{}',
                    research_urs_plan_json TEXT NOT NULL DEFAULT '{}',
                    dds_plan_json TEXT NOT NULL DEFAULT '{}',
                    development_plan_json TEXT NOT NULL DEFAULT '{}',
                    sqa_plan_json TEXT NOT NULL DEFAULT '{}',
                    user_manual_plan_json TEXT NOT NULL DEFAULT '{}',
                    production_plan_json TEXT NOT NULL DEFAULT '{}'
                )
                """
            )
            _v2_cols = [
                "epic_key", "project_key", "project_name", "product_category",
                "epic_name", "description", "originator", "priority", "jira_url",
                "epic_plan_json", "research_urs_plan_json", "dds_plan_json",
                "development_plan_json", "sqa_plan_json", "production_plan_json",
            ]
            _v2_opt = {
                "component": "''",
                "plan_status": "'Not Planned Yet'",
                "ipp_meeting_planned": "'No'",
                "actual_production_date": "''",
                "remarks": "''",
                "user_manual_plan_json": "'{}'",
            }
            select_exprs = list(_v2_cols)
            for col_name, fallback in _v2_opt.items():
                if col_name in names:
                    select_exprs.append(col_name)
                else:
                    select_exprs.append(f"{fallback} AS {col_name}")
            conn.execute(
                "INSERT INTO epics_management_v2 ("
                + ", ".join(c if " AS " not in c else c.split(" AS ")[1] for c in select_exprs)
                + ") SELECT "
                + ", ".join(select_exprs)
                + " FROM epics_management"
            )
            conn.execute("DROP TABLE epics_management")
            conn.execute("ALTER TABLE epics_management_v2 RENAME TO epics_management")
            columns = conn.execute("PRAGMA table_info(epics_management)").fetchall()
            names = {str(col[1]) for col in columns}
        if "plan_status" not in names:
            conn.execute("ALTER TABLE epics_management ADD COLUMN plan_status TEXT NOT NULL DEFAULT 'Not Planned Yet'")
        if "ipp_meeting_planned" not in names:
            conn.execute("ALTER TABLE epics_management ADD COLUMN ipp_meeting_planned TEXT NOT NULL DEFAULT 'No'")
        if "actual_production_date" not in names:
            conn.execute("ALTER TABLE epics_management ADD COLUMN actual_production_date TEXT NOT NULL DEFAULT ''")
        if "remarks" not in names:
            conn.execute("ALTER TABLE epics_management ADD COLUMN remarks TEXT NOT NULL DEFAULT ''")
        if "user_manual_plan_json" not in names:
            conn.execute("ALTER TABLE epics_management ADD COLUMN user_manual_plan_json TEXT NOT NULL DEFAULT '{}'")
        if "component" not in names:
            conn.execute("ALTER TABLE epics_management ADD COLUMN component TEXT NOT NULL DEFAULT ''")
        _seed_default_epics_plan_columns(conn)
        _backfill_legacy_epics_plan_values(conn, names)
        conn.commit()
    finally:
        conn.close()


def _jira_adf_to_text(value: object) -> str:
    if isinstance(value, str):
        return value
    if isinstance(value, list):
        parts = [_jira_adf_to_text(item) for item in value]
        return " ".join(part for part in parts if part)
    if isinstance(value, dict):
        parts: list[str] = []
        text = value.get("text")
        if isinstance(text, str) and text.strip():
            parts.append(text.strip())
        content = value.get("content")
        if isinstance(content, list):
            for node in content:
                nested = _jira_adf_to_text(node)
                if nested:
                    parts.append(nested)
        return " ".join(parts)
    return ""


def _upsert_epics_management_story_sync_rows(settings_db_path: Path, epic_key: str, rows: list[dict]) -> int:
    _init_epics_management_db(settings_db_path)
    normalized_epic_key = _normalize_epic_key(epic_key)
    story_rows = [row for row in (rows or []) if _to_text(row.get("story_key"))]
    conn = sqlite3.connect(settings_db_path)
    try:
        now_utc = datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")
        story_keys = sorted({_to_text(row.get("story_key")).upper() for row in story_rows if _to_text(row.get("story_key"))})
        if story_keys:
            placeholders = ",".join("?" for _ in story_keys)
            conn.execute(
                f"DELETE FROM epics_management_story_sync WHERE epic_key=? AND story_key NOT IN ({placeholders})",
                [normalized_epic_key, *story_keys],
            )
        else:
            conn.execute("DELETE FROM epics_management_story_sync WHERE epic_key=?", (normalized_epic_key,))
        for row in story_rows:
            story_key = _to_text(row.get("story_key")).upper()
            if not story_key:
                continue
            conn.execute(
                """
                INSERT INTO epics_management_story_sync (
                    story_key, epic_key, project_key, story_name, story_status, jira_url,
                    start_date, due_date, estimate_hours, payload_json, synced_at_utc
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(story_key) DO UPDATE SET
                    epic_key=excluded.epic_key,
                    project_key=excluded.project_key,
                    story_name=excluded.story_name,
                    story_status=excluded.story_status,
                    jira_url=excluded.jira_url,
                    start_date=excluded.start_date,
                    due_date=excluded.due_date,
                    estimate_hours=excluded.estimate_hours,
                    payload_json=excluded.payload_json,
                    synced_at_utc=excluded.synced_at_utc
                """,
                (
                    story_key,
                    normalized_epic_key,
                    _to_text(row.get("project_key")).upper(),
                    _to_text(row.get("story_name")),
                    _to_text(row.get("story_status")),
                    _to_text(row.get("jira_url")),
                    _to_text(row.get("start_date")),
                    _to_text(row.get("due_date")),
                    float(row.get("estimate_hours") or 0.0),
                    _to_text(row.get("payload_json")) or "{}",
                    now_utc,
                ),
            )
        conn.commit()
    finally:
        conn.close()
    return len(story_rows)


def _normalize_epic_key(value: object) -> str:
    text = _to_text(value).upper()
    if not text:
        raise ValueError("epic_key is required.")
    if not _EPIC_KEY_PATTERN.match(text) and not _TMP_EPIC_KEY_PATTERN.match(text):
        raise ValueError("epic_key must look like ABC-123 or TMP-YYYYMMDDTHHMMSSZ-XXXXXX.")
    return text


def _generate_tmp_epic_key(conn: sqlite3.Connection) -> str:
    for _ in range(50):
        stamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
        suffix = "".join(random.choices("ABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789", k=6))
        candidate = f"TMP-{stamp}-{suffix}"
        exists = conn.execute(
            "SELECT 1 FROM epics_management WHERE epic_key=?",
            (candidate,),
        ).fetchone()
        if not exists:
            return candidate
    raise RuntimeError("Failed to generate a unique temporary epic key.")


def _plan_has_content_for_tmp_reuse(plan_value: object) -> bool:
    try:
        plan = _normalize_epics_management_plan(plan_value)
    except Exception:
        return True
    if plan.get("man_days") not in ("", None):
        return True
    if _to_text(plan.get("start_date")):
        return True
    if _to_text(plan.get("due_date")):
        return True
    if _to_text(plan.get("jira_url")):
        return True
    return False


def _is_vacant_tmp_epic_row_for_reuse(conn: sqlite3.Connection, row: sqlite3.Row) -> bool:
    epic_key = _to_text(row["epic_key"]).upper()
    if not _is_tmp_epic_key(epic_key):
        return False
    project_key = _to_text(row["project_key"]).upper()
    if project_key not in {"", "ORPHAN"}:
        return False
    project_name = _to_text(row["project_name"]).upper()
    if project_name not in {"", "ORPHAN"}:
        return False
    product_category = _to_text(row["product_category"]).upper()
    if product_category not in {"", "ORPHAN"}:
        return False
    if _to_text(row["component"]):
        return False
    epic_name = _to_text(row["epic_name"]).upper()
    if epic_name and epic_name != epic_key:
        return False
    if _to_text(row["description"]) or _to_text(row["originator"]) or _to_text(row["remarks"]) or _to_text(row["jira_url"]):
        return False
    if _to_text(row["actual_production_date"]):
        return False
    if _to_text(row["priority"]).casefold() not in {"", "low"}:
        return False
    if _to_text(row["plan_status"]).casefold() not in {"", "not planned yet", "not planned", "not_planned_yet", "plan"}:
        return False
    if _to_text(row["ipp_meeting_planned"]).casefold() in {"yes", "y", "true", "1"}:
        return False

    legacy_cols = (
        "epic_plan_json",
        "research_urs_plan_json",
        "dds_plan_json",
        "development_plan_json",
        "sqa_plan_json",
        "user_manual_plan_json",
        "production_plan_json",
    )
    for col in legacy_cols:
        if _plan_has_content_for_tmp_reuse(row[col]):
            return False

    plan_value_rows = conn.execute(
        "SELECT plan_json FROM epics_management_plan_values WHERE epic_key=?",
        (epic_key,),
    ).fetchall()
    for plan_value_row in plan_value_rows:
        if _plan_has_content_for_tmp_reuse(plan_value_row[0]):
            return False

    synced_story_row = conn.execute(
        "SELECT 1 FROM epics_management_story_sync WHERE epic_key=? LIMIT 1",
        (epic_key,),
    ).fetchone()
    if synced_story_row:
        return False
    return True


def _find_vacant_tmp_epic_key_for_reuse(conn: sqlite3.Connection, preferred_key: str = "") -> str:
    preferred = _to_text(preferred_key).upper()
    rows = conn.execute(
        """
        SELECT
            epic_key, project_key, project_name, product_category, component, epic_name,
            description, originator, priority, plan_status, ipp_meeting_planned,
            actual_production_date, remarks, jira_url, epic_plan_json, research_urs_plan_json,
            dds_plan_json, development_plan_json, sqa_plan_json, user_manual_plan_json, production_plan_json
        FROM epics_management
        WHERE epic_key LIKE 'TMP-%'
        ORDER BY CASE WHEN epic_key = ? THEN 0 ELSE 1 END, epic_key
        """,
        (preferred,),
    ).fetchall()
    for row in rows:
        if _is_vacant_tmp_epic_row_for_reuse(conn, row):
            return _to_text(row["epic_key"]).upper()
    return ""


def _normalize_epics_management_plan(value: object) -> dict:
    if isinstance(value, str):
        raw = _to_text(value)
        if not raw:
            return {}
        try:
            value = json.loads(raw)
        except Exception:
            return {}
    if not isinstance(value, dict):
        return {}
    man_days_raw = value.get("man_days")
    man_days: object = ""
    if man_days_raw not in (None, ""):
        try:
            parsed = float(man_days_raw)
            if parsed < 0:
                raise ValueError
            man_days = round(parsed, 2)
        except Exception:
            raise ValueError("plan.man_days must be blank or a number >= 0.")
    start_date = _to_text(value.get("start_date"))
    due_date = _to_text(value.get("due_date"))
    if start_date and not _parse_iso_date(start_date):
        raise ValueError("plan.start_date must be ISO date YYYY-MM-DD.")
    if due_date and not _parse_iso_date(due_date):
        raise ValueError("plan.due_date must be ISO date YYYY-MM-DD.")
    if start_date and due_date and start_date > due_date:
        raise ValueError("plan.start_date cannot be after plan.due_date.")
    jira_url = _to_text(value.get("jira_url"))
    if jira_url and not re.match(r"^https?://", jira_url, re.IGNORECASE):
        raise ValueError("plan.jira_url must start with http:// or https://")
    return {
        "man_days": man_days,
        "start_date": start_date,
        "due_date": due_date,
        "jira_url": jira_url,
    }


def _normalize_epics_management_payload(
    payload: dict,
    plan_columns: list[dict[str, object]],
    require_all_fields: bool = True,
) -> dict:
    raw = payload or {}
    epic_key = _normalize_epic_key(raw.get("epic_key"))
    project_key = _to_text(raw.get("project_key")).upper() or _extract_project_key(epic_key)
    if not project_key:
        raise ValueError("project_key is required.")
    project_name = _to_text(raw.get("project_name")) or project_key
    product_category = _to_text(raw.get("product_category"))
    component = _to_text(raw.get("component"))
    epic_name = _to_text(raw.get("epic_name")) or epic_key
    if require_all_fields and not epic_name:
        raise ValueError("epic_name is required.")
    description = _to_text(raw.get("description"))
    originator = _to_text(raw.get("originator"))
    priority = _priority_for_epics_management(raw.get("priority"))
    plan_status = _plan_status_for_epics_management(raw.get("plan_status"))
    ipp_meeting_planned = _ipp_meeting_planned_for_epics_management(raw.get("ipp_meeting_planned"))
    actual_production_date = _to_text(raw.get("actual_production_date"))
    if actual_production_date and not _parse_iso_date(actual_production_date):
        raise ValueError("actual_production_date must be ISO date YYYY-MM-DD.")
    remarks = _to_text(raw.get("remarks"))
    jira_url = _to_text(raw.get("jira_url"))
    if jira_url and not re.match(r"^https?://", jira_url, re.IGNORECASE):
        raise ValueError("jira_url must start with http:// or https://")

    plans_in = raw.get("plans")
    if not isinstance(plans_in, dict):
        plans_in = {}
    plan_columns_by_key = {
        _to_text(col.get("key")): col
        for col in (plan_columns or [])
        if _to_text(col.get("key"))
    }
    unknown_keys = sorted(_to_text(key) for key in plans_in.keys() if _to_text(key) and _to_text(key) not in plan_columns_by_key)
    if unknown_keys:
        raise ValueError("Unknown plan column key(s): " + ", ".join(unknown_keys))

    plans: dict[str, dict] = {}
    for key, column_meta in plan_columns_by_key.items():
        source_value = plans_in.get(key)
        if source_value is None and key in _EPICS_MANAGEMENT_LEGACY_PLAN_JSON_COLUMN_BY_KEY:
            source_value = raw.get(_EPICS_MANAGEMENT_LEGACY_PLAN_JSON_COLUMN_BY_KEY[key], {})
        normalized_plan = _normalize_epics_management_plan(source_value)
        if not bool(column_meta.get("jira_link_enabled")):
            normalized_plan["jira_url"] = ""
        plans[key] = normalized_plan

    return {
        "epic_key": epic_key,
        "project_key": project_key,
        "project_name": project_name,
        "product_category": product_category,
        "component": component,
        "epic_name": epic_name,
        "description": description,
        "originator": originator,
        "priority": priority,
        "plan_status": plan_status,
        "ipp_meeting_planned": ipp_meeting_planned,
        "actual_production_date": actual_production_date,
        "remarks": remarks,
        "jira_url": jira_url,
        "plans": plans,
    }


def _upsert_epics_plan_values_for_row(
    conn: sqlite3.Connection,
    epic_key: str,
    plans: dict[str, dict],
) -> None:
    normalized_epic_key = _normalize_epic_key(epic_key)
    now_utc = _utc_now_iso()
    valid_keys = {
        _to_text(row[0])
        for row in conn.execute(
            "SELECT column_key FROM epics_management_plan_columns WHERE is_active = 1 OR is_default = 1"
        ).fetchall()
    }
    for plan_key, plan_value in (plans or {}).items():
        key = _to_text(plan_key)
        if not key or key not in valid_keys:
            continue
        serialized = json.dumps(_normalize_epics_management_plan(plan_value), ensure_ascii=True)
        conn.execute(
            """
            INSERT INTO epics_management_plan_values (epic_key, column_key, plan_json, created_at_utc, updated_at_utc)
            VALUES (?, ?, ?, ?, ?)
            ON CONFLICT(epic_key, column_key) DO UPDATE SET
                plan_json=excluded.plan_json,
                updated_at_utc=excluded.updated_at_utc
            """,
            (normalized_epic_key, key, serialized, now_utc, now_utc),
        )


def _save_epics_management_row(settings_db_path: Path, payload: dict) -> dict[str, str]:
    _init_epics_management_db(settings_db_path)
    raw_payload = payload if isinstance(payload, dict) else {}
    plan_columns = _load_epics_plan_columns(settings_db_path, include_inactive=False)
    conn = sqlite3.connect(settings_db_path)
    conn.row_factory = sqlite3.Row
    try:
        prepared_payload = dict(raw_payload)
        epic_name_in = _to_text(prepared_payload.get("epic_name"))
        if not epic_name_in:
            raise ValueError("epic_name is required.")
        user_supplied_epic_key = bool(_to_text(prepared_payload.get("epic_key")))
        reuse_vacant_tmp_key = ""
        if not user_supplied_epic_key:
            reuse_vacant_tmp_key = _find_vacant_tmp_epic_key_for_reuse(conn)
            if reuse_vacant_tmp_key:
                prepared_payload["epic_key"] = reuse_vacant_tmp_key
            else:
                prepared_payload["epic_key"] = _generate_tmp_epic_key(conn)
        project_key_in = _to_text(prepared_payload.get("project_key")).upper()
        project_name_in = _to_text(prepared_payload.get("project_name"))
        if not project_key_in:
            prepared_payload["project_key"] = "ORPHAN"
            prepared_payload["project_name"] = "Orphan"
            if not _to_text(prepared_payload.get("product_category")):
                prepared_payload["product_category"] = "Orphan"
        elif not project_name_in:
            prepared_payload["project_name"] = project_key_in

        row = _normalize_epics_management_payload(prepared_payload, plan_columns=plan_columns, require_all_fields=True)

        if reuse_vacant_tmp_key:
            legacy_plans = {
                key: row["plans"].get(key, {})
                for key in _EPICS_MANAGEMENT_DEFAULT_PLAN_KEYS
            }
            conn.execute(
                """
                UPDATE epics_management
                SET project_key=?, project_name=?, product_category=?, component=?, epic_name=?,
                    description=?, originator=?, priority=?, plan_status=?, ipp_meeting_planned=?,
                    actual_production_date=?, remarks=?, jira_url=?,
                    epic_plan_json=?, research_urs_plan_json=?, dds_plan_json=?,
                    development_plan_json=?, sqa_plan_json=?, user_manual_plan_json=?, production_plan_json=?
                WHERE epic_key=?
                """,
                (
                    row["project_key"],
                    row["project_name"],
                    row["product_category"],
                    row["component"],
                    row["epic_name"],
                    row["description"],
                    row["originator"],
                    row["priority"],
                    row["plan_status"],
                    row["ipp_meeting_planned"],
                    row["actual_production_date"],
                    row["remarks"],
                    row["jira_url"],
                    json.dumps(legacy_plans["epic_plan"], ensure_ascii=True),
                    json.dumps(legacy_plans["research_urs_plan"], ensure_ascii=True),
                    json.dumps(legacy_plans["dds_plan"], ensure_ascii=True),
                    json.dumps(legacy_plans["development_plan"], ensure_ascii=True),
                    json.dumps(legacy_plans["sqa_plan"], ensure_ascii=True),
                    json.dumps(legacy_plans["user_manual_plan"], ensure_ascii=True),
                    json.dumps(legacy_plans["production_plan"], ensure_ascii=True),
                    reuse_vacant_tmp_key,
                ),
            )
            _upsert_epics_plan_values_for_row(conn, reuse_vacant_tmp_key, row["plans"])
            conn.commit()
        else:
            for _attempt in range(5):
                legacy_plans = {
                    key: row["plans"].get(key, {})
                    for key in _EPICS_MANAGEMENT_DEFAULT_PLAN_KEYS
                }
                try:
                    conn.execute(
                        """
                        INSERT INTO epics_management (
                            epic_key, project_key, project_name, product_category, component, epic_name,
                            description, originator, priority, plan_status, ipp_meeting_planned, actual_production_date, remarks, jira_url,
                            epic_plan_json, research_urs_plan_json, dds_plan_json,
                            development_plan_json, sqa_plan_json, user_manual_plan_json, production_plan_json
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            row["epic_key"],
                            row["project_key"],
                            row["project_name"],
                            row["product_category"],
                            row["component"],
                            row["epic_name"],
                            row["description"],
                            row["originator"],
                            row["priority"],
                            row["plan_status"],
                            row["ipp_meeting_planned"],
                            row["actual_production_date"],
                            row["remarks"],
                            row["jira_url"],
                            json.dumps(legacy_plans["epic_plan"], ensure_ascii=True),
                            json.dumps(legacy_plans["research_urs_plan"], ensure_ascii=True),
                            json.dumps(legacy_plans["dds_plan"], ensure_ascii=True),
                            json.dumps(legacy_plans["development_plan"], ensure_ascii=True),
                            json.dumps(legacy_plans["sqa_plan"], ensure_ascii=True),
                            json.dumps(legacy_plans["user_manual_plan"], ensure_ascii=True),
                            json.dumps(legacy_plans["production_plan"], ensure_ascii=True),
                        ),
                    )
                    _upsert_epics_plan_values_for_row(conn, row["epic_key"], row["plans"])
                    conn.commit()
                    break
                except sqlite3.IntegrityError:
                    if not _is_tmp_epic_key(row.get("epic_key")):
                        raise
                    if user_supplied_epic_key:
                        conflict_key = _to_text(row.get("epic_key")).upper()
                        vacant_tmp_key = _find_vacant_tmp_epic_key_for_reuse(conn, preferred_key=conflict_key)
                        raise _EpicCreateConflictError(
                            f"Epic '{conflict_key}' already exists.",
                            conflict_epic_key=conflict_key,
                            vacant_tmp_key=vacant_tmp_key,
                        )
                    fallback_vacant = _find_vacant_tmp_epic_key_for_reuse(conn)
                    if fallback_vacant:
                        row["epic_key"] = fallback_vacant
                    else:
                        row["epic_key"] = _generate_tmp_epic_key(conn)
            else:
                raise ValueError("Failed to generate a unique temporary epic key.")
    except _EpicCreateConflictError:
        raise
    except sqlite3.IntegrityError:
        raise ValueError(f"Epic '{row['epic_key']}' already exists.")
    finally:
        conn.close()
    matches = [r for r in _load_epics_management_rows(settings_db_path) if _to_text(r.get("epic_key")).upper() == row["epic_key"]]
    if not matches:
        raise RuntimeError("Failed to load saved epic.")
    return matches[0]


def _update_epics_management_row(settings_db_path: Path, epic_key: str, payload: dict) -> dict[str, str]:
    _init_epics_management_db(settings_db_path)
    key = _normalize_epic_key(epic_key)
    existing = [r for r in _load_epics_management_rows(settings_db_path) if _to_text(r.get("epic_key")).upper() == key]
    if not existing:
        raise LookupError(f"Epic '{key}' not found.")
    base = existing[0]
    plan_columns = _load_epics_plan_columns(settings_db_path, include_inactive=False)
    normalized = _normalize_epics_management_payload(
        {
            **base,
            **(payload or {}),
            "epic_key": key,
            "plans": {
                **(base.get("plans") or {}),
                **((payload or {}).get("plans") or {}),
            },
        },
        plan_columns=plan_columns,
        require_all_fields=True,
    )
    updated_epic_key = key
    jira_key_candidate = _to_text(extract_jira_key_from_url(normalized.get("jira_url"))).upper() if _to_text(normalized.get("jira_url")) else ""
    if _is_tmp_epic_key(key) and jira_key_candidate and jira_key_candidate != key:
        if not _EPIC_KEY_PATTERN.match(jira_key_candidate):
            raise ValueError("Derived Jira key must look like ABC-123.")
        updated_epic_key = jira_key_candidate
    normalized["epic_key"] = updated_epic_key
    conn = sqlite3.connect(settings_db_path)
    try:
        if updated_epic_key != key:
            exists_target = conn.execute("SELECT 1 FROM epics_management WHERE epic_key=?", (updated_epic_key,)).fetchone()
            if exists_target:
                raise ValueError(f"Epic '{updated_epic_key}' already exists.")
        legacy_plans = {
            key_item: normalized["plans"].get(key_item, {})
            for key_item in _EPICS_MANAGEMENT_DEFAULT_PLAN_KEYS
        }
        cur = conn.execute(
            """
            UPDATE epics_management
            SET epic_key=?, project_key=?, project_name=?, product_category=?, component=?, epic_name=?,
                description=?, originator=?, priority=?, plan_status=?, ipp_meeting_planned=?, actual_production_date=?, remarks=?, jira_url=?,
                epic_plan_json=?, research_urs_plan_json=?, dds_plan_json=?,
                development_plan_json=?, sqa_plan_json=?, user_manual_plan_json=?, production_plan_json=?
            WHERE epic_key=?
            """,
            (
                updated_epic_key,
                normalized["project_key"],
                normalized["project_name"],
                normalized["product_category"],
                normalized["component"],
                normalized["epic_name"],
                normalized["description"],
                normalized["originator"],
                normalized["priority"],
                normalized["plan_status"],
                normalized["ipp_meeting_planned"],
                normalized["actual_production_date"],
                normalized["remarks"],
                normalized["jira_url"],
                json.dumps(legacy_plans["epic_plan"], ensure_ascii=True),
                json.dumps(legacy_plans["research_urs_plan"], ensure_ascii=True),
                json.dumps(legacy_plans["dds_plan"], ensure_ascii=True),
                json.dumps(legacy_plans["development_plan"], ensure_ascii=True),
                json.dumps(legacy_plans["sqa_plan"], ensure_ascii=True),
                json.dumps(legacy_plans["user_manual_plan"], ensure_ascii=True),
                json.dumps(legacy_plans["production_plan"], ensure_ascii=True),
                key,
            ),
        )
        if updated_epic_key != key:
            conn.execute(
                "UPDATE epics_management_plan_values SET epic_key=?, updated_at_utc=? WHERE epic_key=?",
                (updated_epic_key, _utc_now_iso(), key),
            )
            conn.execute(
                "UPDATE epics_management_story_sync SET epic_key=?, synced_at_utc=? WHERE epic_key=?",
                (updated_epic_key, _utc_now_iso(), key),
            )
        _upsert_epics_plan_values_for_row(conn, updated_epic_key, normalized["plans"])
        conn.commit()
        if cur.rowcount <= 0:
            raise LookupError(f"Epic '{key}' not found.")
    finally:
        conn.close()
    matches = [
        r
        for r in _load_epics_management_rows(settings_db_path)
        if _to_text(r.get("epic_key")).upper() == updated_epic_key
    ]
    if not matches:
        raise LookupError(f"Epic '{updated_epic_key}' not found.")
    return matches[0]


def _delete_epics_management_row(settings_db_path: Path, epic_key: str) -> str:
    _init_epics_management_db(settings_db_path)
    key = _normalize_epic_key(epic_key)
    existing = [
        r
        for r in _load_epics_management_rows(settings_db_path)
        if _to_text(r.get("epic_key")).upper() == key
    ]
    if not existing:
        raise LookupError(f"Epic '{key}' not found.")
    conn = sqlite3.connect(settings_db_path)
    try:
        conn.execute("DELETE FROM epics_management_plan_values WHERE epic_key=?", (key,))
        conn.execute("DELETE FROM epics_management_story_sync WHERE epic_key=?", (key,))
        conn.execute("DELETE FROM epics_management WHERE epic_key=?", (key,))
        conn.commit()
    finally:
        conn.close()
    return key


def _load_epics_management_rows(settings_db_path: Path) -> list[dict[str, str]]:
    _init_epics_management_db(settings_db_path)

    conn = None
    try:
        conn = sqlite3.connect(settings_db_path)
        conn.row_factory = sqlite3.Row
        table_exists = conn.execute(
            "SELECT 1 FROM sqlite_master WHERE type='table' AND name='epics_management'"
        ).fetchone()
        if not table_exists:
            return []
        rows = conn.execute(
            """
            SELECT
                epic_key, project_key, project_name, product_category, component, epic_name,
                description, originator, priority, plan_status, ipp_meeting_planned, actual_production_date, remarks, jira_url,
                epic_plan_json, research_urs_plan_json, dds_plan_json,
                development_plan_json, sqa_plan_json, user_manual_plan_json, production_plan_json
            FROM epics_management
            ORDER BY lower(project_name) ASC, lower(product_category) ASC, lower(component) ASC, lower(epic_name) ASC, epic_key ASC
            """
        ).fetchall()
        plan_columns = _load_epics_plan_columns_from_conn(conn, include_inactive=False)
        plan_keys = [_to_text(col.get("key")) for col in plan_columns if _to_text(col.get("key"))]
        epic_keys = [_to_text(row["epic_key"]).upper() for row in rows if _to_text(row["epic_key"])]
        plan_values_by_epic_key: dict[str, dict[str, dict]] = {}
        if epic_keys and plan_keys:
            epic_placeholders = ",".join("?" for _ in epic_keys)
            key_placeholders = ",".join("?" for _ in plan_keys)
            value_rows = conn.execute(
                f"""
                SELECT epic_key, column_key, plan_json
                FROM epics_management_plan_values
                WHERE epic_key IN ({epic_placeholders}) AND column_key IN ({key_placeholders})
                """,
                [*epic_keys, *plan_keys],
            ).fetchall()
            for value_row in value_rows:
                key = _to_text(value_row["epic_key"]).upper()
                column_key = _to_text(value_row["column_key"])
                if not key or not column_key:
                    continue
                plan_values_by_epic_key.setdefault(key, {})[column_key] = value_row["plan_json"]
    except Exception:
        return []
    finally:
        if conn is not None:
            try:
                conn.close()
            except Exception:
                pass

    def _safe_json_dict(text: object) -> dict:
        raw = _to_text(text)
        if not raw:
            return {}
        try:
            parsed = json.loads(raw)
            return parsed if isinstance(parsed, dict) else {}
        except Exception:
            return {}

    out: list[dict[str, str]] = []
    for row in rows:
        epic_key = _to_text(row["epic_key"]).upper()
        per_epic_values = plan_values_by_epic_key.get(epic_key, {})
        plans: dict[str, dict] = {}
        for col in plan_columns:
            plan_key = _to_text(col.get("key"))
            if not plan_key:
                continue
            raw_plan = per_epic_values.get(plan_key)
            if raw_plan is None:
                legacy_col = _EPICS_MANAGEMENT_LEGACY_PLAN_JSON_COLUMN_BY_KEY.get(plan_key)
                if legacy_col:
                    raw_plan = row[legacy_col]
            parsed = _safe_json_dict(raw_plan)
            if not bool(col.get("jira_link_enabled")):
                parsed["jira_url"] = ""
            plans[plan_key] = parsed
        out.append(
            {
                "id": epic_key,
                "project_key": _to_text(row["project_key"]).upper(),
                "project_name": _to_text(row["project_name"]) or _to_text(row["project_key"]).upper(),
                "product_category": _to_text(row["product_category"]),
                "component": _to_text(row["component"]),
                "epic_key": epic_key,
                "epic_name": _to_text(row["epic_name"]) or epic_key,
                "description": _to_text(row["description"]),
                "originator": _to_text(row["originator"]),
                "priority": _priority_for_epics_management(row["priority"]),
                "plan_status": _plan_status_for_epics_management(row["plan_status"]),
                "ipp_meeting_planned": _ipp_meeting_planned_for_epics_management(row["ipp_meeting_planned"]),
                "actual_production_date": _to_text(row["actual_production_date"]),
                "remarks": _to_text(row["remarks"]),
                "jira_url": _to_text(row["jira_url"]),
                "plans": plans,
            }
        )
    return out


def _fetch_jira_issues_for_jql(session, jql: str, fields: list[str]) -> list[dict]:
    url = f"{BASE_URL}/rest/api/3/search/jql"
    out: list[dict] = []
    next_page_token = None
    while True:
        payload = {"jql": jql, "maxResults": 100, "fields": fields}
        if next_page_token:
            payload["nextPageToken"] = next_page_token
        response = session.post(url, json=payload, timeout=(10, 90))
        response.raise_for_status()
        data = response.json()
        out.extend(data.get("issues", []))
        next_page_token = data.get("nextPageToken")
        if not next_page_token:
            break
    return out


def _to_seconds(value) -> float:
    try:
        return float(value or 0.0)
    except (TypeError, ValueError):
        return 0.0


def _first_non_empty_issue_field(issue_fields: dict, field_ids: list[str]) -> str:
    for field_id in field_ids:
        value = issue_fields.get(field_id)
        if value is None:
            continue
        text = _to_text(value)
        if text:
            return text
    return ""


def _resolve_epic_key_for_story(issue_fields: dict, valid_epic_keys: set[str]) -> str:
    parent = issue_fields.get("parent") or {}
    parent_key = _to_text(parent.get("key")).upper()
    if parent_key in valid_epic_keys:
        return parent_key
    epic_link = issue_fields.get("customfield_10014")
    if isinstance(epic_link, str):
        link_key = _to_text(epic_link).upper()
        if link_key in valid_epic_keys:
            return link_key
    if isinstance(epic_link, dict):
        link_key = _to_text(epic_link.get("key")).upper()
        if link_key in valid_epic_keys:
            return link_key
    return ""


def _extract_issue_plan_metrics(issue_fields: dict, start_field_id: str, end_field_ids: list[str]) -> tuple[str, str, float]:
    start_iso = ""
    due_iso = ""
    if start_field_id:
        start_date = _parse_iso_date(_to_text(issue_fields.get(start_field_id)))
        if start_date:
            start_iso = start_date.isoformat()
    end_date = _parse_iso_date(_first_non_empty_issue_field(issue_fields, end_field_ids))
    if end_date:
        due_iso = end_date.isoformat()
    estimate_hours = round(max(_to_seconds(issue_fields.get("timeoriginalestimate")) / 3600.0, 0.0), 2)
    return start_iso, due_iso, estimate_hours


def _chunked(items: list[str], size: int = 50) -> list[list[str]]:
    out: list[list[str]] = []
    if size <= 0:
        size = 50
    for i in range(0, len(items), size):
        out.append(items[i:i + size])
    return out


def _jql_quote(value: str) -> str:
    return str(value or "").replace("\\", "\\\\").replace('"', '\\"')


def _project_jql_clause(project_keys: set[str]) -> str:
    normalized = sorted({_to_text(item).upper() for item in (project_keys or set()) if _to_text(item)})
    if not normalized:
        return ""
    quoted = ", ".join(f'"{_jql_quote(key)}"' for key in normalized)
    return f"project in ({quoted})"


def _extract_epic_key_candidate(issue_fields: dict) -> str:
    parent = issue_fields.get("parent") or {}
    parent_key = _to_text(parent.get("key")).upper()
    epic_link = issue_fields.get("customfield_10014")
    if isinstance(epic_link, str):
        link_key = _to_text(epic_link).upper()
        if link_key:
            return link_key
    if isinstance(epic_link, dict):
        link_key = _to_text(epic_link.get("key")).upper()
        if link_key:
            return link_key
    return parent_key


def _issue_overlaps_range(start_iso: str, due_iso: str, from_date: date, to_date: date) -> bool:
    start_parsed = _parse_iso_date(start_iso)
    due_parsed = _parse_iso_date(due_iso)
    return bool(
        (start_parsed and from_date <= start_parsed <= to_date)
        or (due_parsed and from_date <= due_parsed <= to_date)
    )


def _fetch_jira_issues_by_keys(session, issue_keys: list[str], fields: list[str]) -> list[dict]:
    normalized = sorted({_to_text(item).upper() for item in issue_keys if _to_text(item)})
    if not normalized:
        return []
    out: list[dict] = []
    for chunk in _chunked(normalized, 50):
        keys_csv = '", "'.join(_jql_quote(item) for item in chunk)
        out.extend(_fetch_jira_issues_for_jql(session, f'key in ("{keys_csv}")', fields))
    return out


def _fetch_story_issues_for_epics(session, epic_keys: list[str], fields: list[str], project_keys: set[str] | None = None) -> list[dict]:
    normalized = sorted({_to_text(item).upper() for item in epic_keys if _to_text(item)})
    if not normalized:
        return []
    out: list[dict] = []
    project_clause = _project_jql_clause(project_keys or set())
    prefix = f"{project_clause} AND " if project_clause else ""
    for chunk in _chunked(normalized, 30):
        keys_csv = '", "'.join(_jql_quote(item) for item in chunk)
        jql = (
            f'{prefix}issuetype not in subTaskIssueTypes() '
            f'AND (parent in ("{keys_csv}") OR customfield_10014 in ("{keys_csv}"))'
        )
        out.extend(_fetch_jira_issues_for_jql(session, jql, fields))
    return out


def _fetch_subtask_issues_for_stories(session, story_keys: list[str], fields: list[str], project_keys: set[str] | None = None) -> list[dict]:
    normalized = sorted({_to_text(item).upper() for item in story_keys if _to_text(item)})
    if not normalized:
        return []
    out: list[dict] = []
    project_clause = _project_jql_clause(project_keys or set())
    prefix = f"{project_clause} AND " if project_clause else ""
    for chunk in _chunked(normalized, 50):
        keys_csv = '", "'.join(_jql_quote(item) for item in chunk)
        jql = f'{prefix}issueType in subTaskIssueTypes() AND parent in ("{keys_csv}")'
        out.extend(_fetch_jira_issues_for_jql(session, jql, fields))
    return out


def _planned_vs_dispensed_issue_fields(start_field_id: str, end_field_ids: list[str]) -> list[str]:
    fields = [
        "summary",
        "status",
        "assignee",
        "issuetype",
        "project",
        "parent",
        "customfield_10014",
        "timeoriginalestimate",
    ]
    if start_field_id:
        fields.append(start_field_id)
    for field_id in end_field_ids:
        if field_id not in fields:
            fields.append(field_id)
    return fields


def _jira_browse_url(issue_key: str) -> str:
    key = _to_text(issue_key).upper()
    if not key:
        return ""
    base = _to_text(BASE_URL).rstrip("/")
    if not base:
        return ""
    return f"{base}/browse/{key}"


def _normalize_pvd_issue(
    issue: dict,
    start_field_id: str,
    end_field_ids: list[str],
    issue_kind_override: str = "",
) -> dict:
    issue_fields = issue.get("fields", {}) or {}
    issue_key = _to_text(issue.get("key")).upper()
    issue_type_name = _to_text((issue_fields.get("issuetype") or {}).get("name"))
    issue_kind = issue_kind_override or _issue_kind(issue_type_name)
    planned_start, planned_due, estimate_hours = _extract_issue_plan_metrics(issue_fields, start_field_id, end_field_ids)
    assignee_raw = issue_fields.get("assignee") or {}
    if isinstance(assignee_raw, dict):
        assignee = _to_text(assignee_raw.get("displayName")) or _to_text(assignee_raw.get("emailAddress")) or "Unassigned"
    else:
        assignee = "Unassigned"
    parent = issue_fields.get("parent") or {}
    return {
        "issue_key": issue_key,
        "issue_kind": issue_kind,
        "issue_type_name": issue_type_name,
        "project_key": _to_text((issue_fields.get("project") or {}).get("key")).upper() or _extract_project_key(issue_key),
        "project_name": _to_text((issue_fields.get("project") or {}).get("name")),
        "jira_url": _jira_browse_url(issue_key),
        "summary": _to_text(issue_fields.get("summary")) or issue_key,
        "status": _to_text((issue_fields.get("status") or {}).get("name")),
        "assignee": assignee,
        "planned_start": planned_start,
        "planned_due": planned_due,
        "estimate_hours": estimate_hours,
        "parent_key": _to_text(parent.get("key")).upper(),
        "epic_key_candidate": _extract_epic_key_candidate(issue_fields),
    }


def _load_planned_vs_dispensed_hierarchy(
    session,
    from_date: date,
    to_date: date,
    mode: str,
    selected_projects: set[str],
) -> dict[str, object]:
    start_field_id = resolve_jira_start_date_field_id(session, BASE_URL, project_keys=sorted(selected_projects) if selected_projects else None)
    end_field_ids = resolve_jira_end_date_field_ids(session, BASE_URL, project_keys=sorted(selected_projects) if selected_projects else None)
    if "duedate" not in end_field_ids:
        end_field_ids.append("duedate")
    fields = _planned_vs_dispensed_issue_fields(start_field_id, end_field_ids)

    project_clause = _project_jql_clause(selected_projects)
    prefix = f"{project_clause} AND " if project_clause else ""

    scoped_epic_keys: set[str] = set()
    preloaded_epics: dict[str, dict] = {}

    if mode == "planned_dates":
        epic_issues = _fetch_jira_issues_for_jql(session, f"{prefix}issuetype = Epic", fields)
        for issue in epic_issues:
            normalized = _normalize_pvd_issue(issue, start_field_id, end_field_ids, issue_kind_override="epic")
            if _issue_overlaps_range(normalized["planned_start"], normalized["planned_due"], from_date, to_date):
                scoped_epic_keys.add(normalized["issue_key"])
                preloaded_epics[normalized["issue_key"]] = normalized
    else:
        log_jql = (
            f'{prefix}issueType in subTaskIssueTypes() '
            f'AND worklogDate >= "{from_date.isoformat()}" AND worklogDate <= "{to_date.isoformat()}"'
        )
        logged_subtasks = _fetch_jira_issues_for_jql(session, log_jql, fields)
        story_keys = sorted({
            _to_text((issue.get("fields", {}) or {}).get("parent", {}).get("key")).upper()
            for issue in logged_subtasks
            if _to_text((issue.get("fields", {}) or {}).get("parent", {}).get("key"))
        })
        if story_keys:
            story_issues = _fetch_jira_issues_by_keys(session, story_keys, fields)
            for story_issue in story_issues:
                story_fields = story_issue.get("fields", {}) or {}
                epic_key = _extract_epic_key_candidate(story_fields)
                if epic_key:
                    scoped_epic_keys.add(epic_key)

    if not scoped_epic_keys:
        return {
            "start_field_id": start_field_id,
            "end_field_ids": end_field_ids,
            "epics": [],
            "stories": [],
            "subtasks": [],
        }

    scoped_epics = dict(preloaded_epics)
    missing_epics = sorted(scoped_epic_keys - set(scoped_epics.keys()))
    if missing_epics:
        for issue in _fetch_jira_issues_by_keys(session, missing_epics, fields):
            normalized = _normalize_pvd_issue(issue, start_field_id, end_field_ids, issue_kind_override="epic")
            scoped_epics[normalized["issue_key"]] = normalized

    story_issues = _fetch_story_issues_for_epics(session, list(scoped_epics.keys()), fields, project_keys=selected_projects)
    stories: list[dict] = []
    story_to_epic: dict[str, str] = {}
    valid_epic_keys = set(scoped_epics.keys())
    for issue in story_issues:
        normalized = _normalize_pvd_issue(issue, start_field_id, end_field_ids, issue_kind_override="story")
        story_fields = issue.get("fields", {}) or {}
        epic_key = _resolve_epic_key_for_story(story_fields, valid_epic_keys) or normalized["epic_key_candidate"]
        epic_key = _to_text(epic_key).upper()
        if not epic_key or epic_key not in valid_epic_keys:
            continue
        normalized["epic_key"] = epic_key
        story_to_epic[normalized["issue_key"]] = epic_key
        stories.append(normalized)

    subtasks: list[dict] = []
    story_keys = sorted(story_to_epic.keys())
    if story_keys:
        subtask_issues = _fetch_subtask_issues_for_stories(session, story_keys, fields, project_keys=selected_projects)
        for issue in subtask_issues:
            normalized = _normalize_pvd_issue(issue, start_field_id, end_field_ids, issue_kind_override="subtask")
            story_key = _to_text(normalized.get("parent_key")).upper()
            epic_key = story_to_epic.get(story_key, "")
            if not epic_key:
                continue
            normalized["story_key"] = story_key
            normalized["epic_key"] = epic_key
            subtasks.append(normalized)

    return {
        "start_field_id": start_field_id,
        "end_field_ids": end_field_ids,
        "epics": list(scoped_epics.values()),
        "stories": stories,
        "subtasks": subtasks,
    }


def _fetch_subtask_actual_hours_by_keys(session, subtask_keys: list[str]) -> dict[str, float]:
    keys = sorted({_to_text(item).upper() for item in (subtask_keys or []) if _to_text(item)})
    if not keys:
        return {}
    issues = _fetch_jira_issues_by_keys(
        session,
        keys,
        ["timespent", "aggregatetimespent", "issuetype", "summary"],
    )
    out: dict[str, float] = {}
    for issue in issues:
        key = _to_text(issue.get("key")).upper()
        fields = issue.get("fields", {}) or {}
        raw_seconds = fields.get("aggregatetimespent")
        if raw_seconds in (None, ""):
            raw_seconds = fields.get("timespent")
        try:
            seconds = float(raw_seconds or 0.0)
        except Exception:
            seconds = 0.0
        out[key] = _round_hours(seconds / 3600.0)
    for key in keys:
        out.setdefault(key, 0.0)
    return out


def _load_planned_actual_hierarchy_incremental(
    session,
    from_date: date,
    to_date: date,
    mode: str,
    selected_projects: set[str],
    last_successful_fetch_utc: str,
) -> dict[str, object]:
    watermark_date = _to_text(last_successful_fetch_utc)[:10]
    if _parse_iso_date(watermark_date) is None:
        raise ValueError("Incremental watermark is unavailable.")

    start_field_id = resolve_jira_start_date_field_id(
        session, BASE_URL, project_keys=sorted(selected_projects) if selected_projects else None
    )
    end_field_ids = resolve_jira_end_date_field_ids(
        session, BASE_URL, project_keys=sorted(selected_projects) if selected_projects else None
    )
    if "duedate" not in end_field_ids:
        end_field_ids.append("duedate")
    fields = _planned_vs_dispensed_issue_fields(start_field_id, end_field_ids)

    project_clause = _project_jql_clause(selected_projects)
    prefix = f"{project_clause} AND " if project_clause else ""
    changed_jql = (
        f'{prefix}updated >= "{watermark_date}" '
        f'AND (issuetype = Epic OR issuetype = Story OR issueType in subTaskIssueTypes())'
    )
    changed_issues = _fetch_jira_issues_for_jql(session, changed_jql, fields)

    impacted_epics: set[str] = set()
    parent_story_keys: set[str] = set()
    for issue in changed_issues:
        normalized = _normalize_pvd_issue(issue, start_field_id, end_field_ids)
        kind = _to_text(normalized.get("issue_kind")).lower()
        if kind == "epic":
            impacted_epics.add(_to_text(normalized.get("issue_key")).upper())
        elif kind == "story":
            epic_key = _to_text(normalized.get("epic_key_candidate")).upper()
            if epic_key:
                impacted_epics.add(epic_key)
        elif kind == "subtask":
            parent_story = _to_text(normalized.get("parent_key")).upper()
            if parent_story:
                parent_story_keys.add(parent_story)

    if parent_story_keys:
        for story in _fetch_jira_issues_by_keys(session, sorted(parent_story_keys), fields):
            story_fields = story.get("fields", {}) or {}
            epic_key = _extract_epic_key_candidate(story_fields)
            if epic_key:
                impacted_epics.add(epic_key)

    if not impacted_epics:
        return {
            "start_field_id": start_field_id,
            "end_field_ids": end_field_ids,
            "epics": [],
            "stories": [],
            "subtasks": [],
        }

    epic_issues = _fetch_jira_issues_by_keys(session, sorted(impacted_epics), fields)
    scoped_epics: dict[str, dict] = {}
    for issue in epic_issues:
        normalized = _normalize_pvd_issue(issue, start_field_id, end_field_ids, issue_kind_override="epic")
        if mode == "planned_dates":
            if not _issue_overlaps_range(normalized["planned_start"], normalized["planned_due"], from_date, to_date):
                continue
        scoped_epics[normalized["issue_key"]] = normalized

    if not scoped_epics:
        return {
            "start_field_id": start_field_id,
            "end_field_ids": end_field_ids,
            "epics": [],
            "stories": [],
            "subtasks": [],
        }

    story_issues = _fetch_story_issues_for_epics(session, sorted(scoped_epics.keys()), fields, project_keys=selected_projects)
    stories: list[dict] = []
    story_to_epic: dict[str, str] = {}
    valid_epic_keys = set(scoped_epics.keys())
    for issue in story_issues:
        normalized = _normalize_pvd_issue(issue, start_field_id, end_field_ids, issue_kind_override="story")
        story_fields = issue.get("fields", {}) or {}
        epic_key = _resolve_epic_key_for_story(story_fields, valid_epic_keys) or normalized["epic_key_candidate"]
        epic_key = _to_text(epic_key).upper()
        if not epic_key or epic_key not in valid_epic_keys:
            continue
        normalized["epic_key"] = epic_key
        story_to_epic[normalized["issue_key"]] = epic_key
        stories.append(normalized)

    subtasks: list[dict] = []
    story_keys = sorted(story_to_epic.keys())
    if story_keys:
        subtask_issues = _fetch_subtask_issues_for_stories(session, story_keys, fields, project_keys=selected_projects)
        for issue in subtask_issues:
            normalized = _normalize_pvd_issue(issue, start_field_id, end_field_ids, issue_kind_override="subtask")
            story_key = _to_text(normalized.get("parent_key")).upper()
            epic_key = story_to_epic.get(story_key, "")
            if not epic_key:
                continue
            normalized["story_key"] = story_key
            normalized["epic_key"] = epic_key
            subtasks.append(normalized)

    return {
        "start_field_id": start_field_id,
        "end_field_ids": end_field_ids,
        "epics": list(scoped_epics.values()),
        "stories": stories,
        "subtasks": subtasks,
    }


def _story_sync_row_from_issue(
    issue: dict,
    epic_key: str,
    project_key: str,
    start_field_id: str,
    end_field_ids: list[str],
    browse_base: str,
) -> dict | None:
    issue_fields = issue.get("fields", {}) or {}
    issue_type_name = _to_text((issue_fields.get("issuetype") or {}).get("name"))
    if _issue_kind(issue_type_name) != "story":
        return None
    linked_epic = _resolve_epic_key_for_story(issue_fields, {epic_key})
    if linked_epic != epic_key:
        return None
    story_key = _to_text(issue.get("key")).upper()
    if not story_key:
        return None
    story_start, story_due, estimate_hours = _extract_issue_plan_metrics(issue_fields, start_field_id, end_field_ids)
    return {
        "story_key": story_key,
        "project_key": project_key,
        "story_name": _to_text(issue_fields.get("summary")) or story_key,
        "story_status": _to_text((issue_fields.get("status") or {}).get("name")),
        "jira_url": _to_text(f"{browse_base}/{story_key}") if browse_base else _jira_browse_url(story_key),
        "start_date": story_start,
        "due_date": story_due,
        "estimate_hours": estimate_hours,
        "payload_json": json.dumps(issue, ensure_ascii=True),
    }


def _sync_epic_plan_from_jira(
    settings_db_path: Path,
    epic_key: str,
    jira_url_override: str = "",
    plan_jira_overrides: dict[str, str] | None = None,
) -> dict[str, str]:
    key = _normalize_epic_key(epic_key)
    existing_rows = [r for r in _load_epics_management_rows(settings_db_path) if _to_text(r.get("epic_key")).upper() == key]
    if not existing_rows:
        raise LookupError(f"Epic '{key}' not found.")
    existing = existing_rows[0]

    jira_url = _to_text(jira_url_override) or _to_text(existing.get("jira_url"))
    if not jira_url:
        raise ValueError("Jira URL is required for sync.")

    jira_key_from_url = _to_text(extract_jira_key_from_url(jira_url)).upper()
    effective_epic_key = jira_key_from_url or key
    if effective_epic_key != key:
        raise ValueError(f"Jira URL key '{effective_epic_key}' does not match epic '{key}'.")

    project_key = _to_text(existing.get("project_key")).upper() or _extract_project_key(key)
    session = get_session()
    start_field_id = resolve_jira_start_date_field_id(session, BASE_URL, project_keys=[project_key] if project_key else None)
    end_field_ids = resolve_jira_end_date_field_ids(session, BASE_URL, project_keys=[project_key] if project_key else None)
    if "duedate" not in end_field_ids:
        end_field_ids.append("duedate")

    fields = ["issuetype", "parent", "customfield_10014", "timeoriginalestimate", "summary", "description", "status"]
    if start_field_id:
        fields.append(start_field_id)
    for field_id in end_field_ids:
        if field_id not in fields:
            fields.append(field_id)

    epic_issues = _fetch_jira_issues_for_jql(session, f'key in ("{key}")', fields)
    if not epic_issues:
        raise ValueError(f"Epic '{key}' was not found in Jira.")

    children: list[dict] = []
    try:
        children = _fetch_jira_issues_for_jql(
            session,
            f'(parent in ("{key}") OR customfield_10014 in ("{key}"))',
            fields,
        )
    except Exception:
        children = _fetch_jira_issues_for_jql(session, f'parent in ("{key}")', fields)

    epic_name_from_jira = ""
    epic_description_from_jira = ""
    story_rows_by_key: dict[str, dict] = {}
    browse_base = BASE_URL.rstrip("/") + "/browse"
    issue_by_key: dict[str, dict] = {}
    for issue in epic_issues + children:
        issue_key = _to_text(issue.get("key")).upper()
        if issue_key:
            issue_by_key[issue_key] = issue

    for issue in epic_issues:
        issue_fields = issue.get("fields", {}) or {}
        if _to_text(issue.get("key")).upper() == key:
            epic_name_from_jira = _to_text(issue_fields.get("summary")) or epic_name_from_jira
            epic_description_from_jira = _jira_adf_to_text(issue_fields.get("description")) or epic_description_from_jira

    valid_epic_keys = {key}
    for issue in children:
        issue_fields = issue.get("fields", {}) or {}
        linked_epic = _resolve_epic_key_for_story(issue_fields, valid_epic_keys)
        if linked_epic != key:
            continue
        story_row = _story_sync_row_from_issue(
            issue=issue,
            epic_key=key,
            project_key=project_key,
            start_field_id=start_field_id,
            end_field_ids=end_field_ids,
            browse_base=browse_base,
        )
        if story_row:
            story_rows_by_key[_to_text(story_row.get("story_key")).upper()] = story_row

    plan_issue_keys: set[str] = set()
    raw_plan_overrides = plan_jira_overrides if isinstance(plan_jira_overrides, dict) else {}
    normalized_plan_overrides: dict[str, str] = {}
    for raw_key, raw_value in raw_plan_overrides.items():
        plan_key = _to_text(raw_key)
        if not plan_key:
            continue
        jira_value = _to_text(raw_value)
        if jira_value:
            normalized_plan_overrides[plan_key] = jira_value
    plans_in = existing.get("plans") or {}
    plan_issue_refs: list[tuple[str, str, str]] = []
    epic_plan_current = plans_in.get("epic_plan") or {}
    epic_plan_jira_url = (
        _to_text(normalized_plan_overrides.get("epic_plan"))
        or _to_text(epic_plan_current.get("jira_url"))
        or jira_url
    )
    epic_plan_linked_key = _to_text(extract_jira_key_from_url(epic_plan_jira_url)).upper()
    if not epic_plan_linked_key:
        raise ValueError("Invalid Jira URL configured for 'epic_plan'.")
    plan_issue_refs.append(("epic_plan", epic_plan_jira_url, epic_plan_linked_key))
    plan_issue_keys.add(epic_plan_linked_key)

    plan_columns = _load_epics_plan_columns(settings_db_path, include_inactive=False)
    jira_link_plan_keys = [
        _to_text(item.get("key"))
        for item in plan_columns
        if bool(item.get("jira_link_enabled")) and _to_text(item.get("key"))
    ]
    for plan_key in jira_link_plan_keys:
        if plan_key == "epic_plan":
            continue
        current = plans_in.get(plan_key) or {}
        plan_jira_url = _to_text(normalized_plan_overrides.get(plan_key)) or _to_text(current.get("jira_url"))
        if not plan_jira_url:
            continue
        linked_key = _to_text(extract_jira_key_from_url(plan_jira_url)).upper()
        if not linked_key:
            raise ValueError(f"Invalid Jira URL configured for '{plan_key}'.")
        plan_issue_refs.append((plan_key, plan_jira_url, linked_key))
        plan_issue_keys.add(linked_key)

    missing_plan_issue_keys = sorted(key_item for key_item in plan_issue_keys if key_item not in issue_by_key)
    if missing_plan_issue_keys:
        keys_csv = '","'.join(missing_plan_issue_keys)
        fetched_plan_issues = _fetch_jira_issues_for_jql(session, f'key in ("{keys_csv}")', fields)
        for issue in fetched_plan_issues:
            issue_key = _to_text(issue.get("key")).upper()
            if issue_key:
                issue_by_key[issue_key] = issue
        still_missing = sorted(key_item for key_item in missing_plan_issue_keys if key_item not in issue_by_key)
        if still_missing:
            raise ValueError("Failed to fetch Jira issue(s) for plan links: " + ", ".join(still_missing))

    plan_updates: dict[str, dict] = {}
    for plan_key, plan_jira_url, linked_key in plan_issue_refs:
        issue = issue_by_key.get(linked_key) or {}
        issue_fields = issue.get("fields", {}) or {}
        next_item = {
            "man_days": "",
            "start_date": "",
            "due_date": "",
            "jira_url": plan_jira_url,
        }
        plan_start_iso, plan_due_iso, plan_estimate_hours = _extract_issue_plan_metrics(issue_fields, start_field_id, end_field_ids)
        next_item["man_days"] = round(plan_estimate_hours / 8.0, 2)
        next_item["start_date"] = plan_start_iso
        next_item["due_date"] = plan_due_iso
        plan_updates[plan_key] = next_item

        story_row = _story_sync_row_from_issue(
            issue=issue,
            epic_key=key,
            project_key=project_key,
            start_field_id=start_field_id,
            end_field_ids=end_field_ids,
            browse_base=browse_base,
        )
        if story_row:
            story_rows_by_key[_to_text(story_row.get("story_key")).upper()] = story_row

    update_payload: dict[str, object] = {
        "jira_url": jira_url,
        "plans": plan_updates,
    }
    if epic_name_from_jira:
        update_payload["epic_name"] = epic_name_from_jira
    if epic_description_from_jira:
        update_payload["description"] = epic_description_from_jira
    updated_row = _update_epics_management_row(
        settings_db_path,
        key,
        update_payload,
    )
    synced_story_count = _upsert_epics_management_story_sync_rows(
        settings_db_path,
        key,
        list(story_rows_by_key.values()),
    )
    updated_row["synced_story_count"] = synced_story_count
    return updated_row


def _extract_project_key(issue_key: str) -> str:
    text = _to_text(issue_key).upper()
    if "-" not in text:
        return "UNKNOWN"
    prefix = text.split("-", 1)[0].strip()
    return prefix or "UNKNOWN"


def _issue_kind(issue_type: str) -> str:
    value = _to_text(issue_type).lower()
    if "epic" in value:
        return "epic"
    if "sub-task" in value or "subtask" in value:
        return "subtask"
    if "story" in value:
        return "story"
    return "other"


def _iso_week_code(day: date) -> str:
    iso = day.isocalendar()
    return f"{iso.year}-W{iso.week:02d}"


def _to_iso_date_or_blank(value) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    parsed = _parse_iso_date(str(value))
    return parsed.isoformat() if parsed else ""


def _date_in_range(value: str, from_date: date, to_date: date) -> bool:
    parsed = _parse_iso_date(value)
    if not parsed:
        return False
    return from_date <= parsed <= to_date


def _load_work_item_index(work_items_path: Path) -> dict[str, dict[str, str]]:
    out: dict[str, dict[str, str]] = {}
    if not work_items_path.exists() or not work_items_path.is_file():
        return out
    wb = load_workbook(work_items_path, read_only=True, data_only=True)
    try:
        ws = wb.active
        header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header:
            return out
        headers = [str(item or "").strip() for item in header]
        idx = {name: pos for pos, name in enumerate(headers)}
        required = ["issue_key", "jira_issue_type", "parent_issue_key", "project_key"]
        if any(name not in idx for name in required):
            return out
        start_key = "start_date" if "start_date" in idx else ""
        end_key = "end_date" if "end_date" in idx else ""
        for row in ws.iter_rows(min_row=2, values_only=True):
            issue_key = _to_text(row[idx["issue_key"]]).upper()
            if not issue_key:
                continue
            project_key = _to_text(row[idx["project_key"]]).upper() or _extract_project_key(issue_key)
            out[issue_key] = {
                "issue_key": issue_key,
                "project_key": project_key,
                "parent_issue_key": _to_text(row[idx["parent_issue_key"]]).upper(),
                "issue_kind": _issue_kind(_to_text(row[idx["jira_issue_type"]])),
                "planned_start": _to_iso_date_or_blank(row[idx[start_key]]) if start_key else "",
                "planned_end": _to_iso_date_or_blank(row[idx[end_key]]) if end_key else "",
            }
        return out
    finally:
        try:
            wb.close()
        except Exception:
            pass


def _resolve_root_epic(issue_key: str, work_items: dict[str, dict[str, str]]) -> str:
    current = _to_text(issue_key).upper()
    depth = 0
    while current and depth < 12:
        item = work_items.get(current)
        if not item:
            return ""
        kind = _to_text(item.get("issue_kind")).lower()
        if kind == "epic":
            return current
        parent = _to_text(item.get("parent_issue_key")).upper()
        if not parent:
            return ""
        current = parent
        depth += 1
    return ""


def _qualifying_epics_by_planned_range(
    work_items: dict[str, dict[str, str]],
    from_date: date,
    to_date: date,
) -> set[str]:
    qualifying: set[str] = set()
    for issue_key, item in work_items.items():
        if _to_text(item.get("issue_kind")).lower() != "epic":
            continue
        start_in = _date_in_range(_to_text(item.get("planned_start")), from_date, to_date)
        end_in = _date_in_range(_to_text(item.get("planned_end")), from_date, to_date)
        if start_in or end_in:
            qualifying.add(issue_key)
    return qualifying


def _round_dict(values: dict[str, float]) -> dict[str, float]:
    return {key: _round_hours(amount) for key, amount in values.items()}


def _round_nested_dict(values: dict[str, dict[str, float]]) -> dict[str, dict[str, float]]:
    out: dict[str, dict[str, float]] = {}
    for key, mapping in values.items():
        out[key] = _round_dict(mapping)
    return out


def _compute_actual_hours_aggregate(
    worklog_path: Path,
    work_items_path: Path,
    from_date: date,
    to_date: date,
    mode: str,
    selected_projects: set[str] | None = None,
) -> dict[str, object]:
    result: dict[str, object] = {
        "subtask_hours_by_issue": {},
        "epic_hours_by_issue": {},
        "project_hours_by_key": {},
        "assignee_hours_by_period": {"day": {}, "week": {}, "month": {}},
    }
    if not worklog_path.exists() or not worklog_path.is_file():
        return result

    work_items = _load_work_item_index(work_items_path)
    qualifying_epics: set[str] = set()
    if mode == "planned_dates":
        qualifying_epics = _qualifying_epics_by_planned_range(work_items, from_date, to_date)

    subtask_hours = defaultdict(float)
    epic_hours = defaultdict(float)
    project_hours = defaultdict(float)
    assignee_day: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    assignee_week: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    assignee_month: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))

    wb = load_workbook(worklog_path, read_only=True, data_only=True)
    try:
        ws = wb.active
        header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header:
            return result
        headers = [str(item or "").strip() for item in header]
        idx = {name: pos for pos, name in enumerate(headers)}
        required = ["issue_id", "worklog_started", "hours_logged"]
        if any(name not in idx for name in required):
            return result
        for row in ws.iter_rows(min_row=2, values_only=True):
            issue_key = str(row[idx["issue_id"]] or "").strip().upper()
            worklog_started_raw = str(row[idx["worklog_started"]] or "").strip()
            if not issue_key or not worklog_started_raw:
                continue
            worklog_day = _parse_iso_date(worklog_started_raw)
            if worklog_day is None:
                continue
            try:
                hours = float(row[idx["hours_logged"]] or 0.0)
            except (TypeError, ValueError):
                hours = 0.0
            if hours <= 0:
                continue

            epic_key = _to_text(row[idx["parent_epic_id"]]).upper() if "parent_epic_id" in idx else ""
            if not epic_key:
                epic_key = _resolve_root_epic(issue_key, work_items)

            work_item = work_items.get(issue_key, {})
            project_key = _to_text(work_item.get("project_key")).upper() or _extract_project_key(issue_key)
            if epic_key and epic_key in work_items:
                project_key = _to_text(work_items[epic_key].get("project_key")).upper() or project_key
            if not project_key:
                project_key = "UNKNOWN"
            if selected_projects and project_key not in selected_projects:
                continue

            in_selected_range = from_date <= worklog_day <= to_date
            if mode == "log_date":
                include_hours = in_selected_range
            else:
                include_hours = bool(epic_key and epic_key in qualifying_epics)
            if not include_hours:
                continue
            assignee = "Unassigned"
            if "worklog_author" in idx:
                assignee = _to_text(row[idx["worklog_author"]]) or assignee
            if assignee == "Unassigned" and "issue_assignee" in idx:
                assignee = _to_text(row[idx["issue_assignee"]]) or assignee

            subtask_hours[issue_key] += hours
            if epic_key:
                epic_hours[epic_key] += hours
            project_hours[project_key] += hours

            if in_selected_range:
                day_key = worklog_day.isoformat()
                week_key = _iso_week_code(worklog_day)
                month_key = f"{worklog_day.year:04d}-{worklog_day.month:02d}"
                assignee_day[day_key][assignee] += hours
                assignee_week[week_key][assignee] += hours
                assignee_month[month_key][assignee] += hours

        result["subtask_hours_by_issue"] = _round_dict(dict(subtask_hours))
        result["epic_hours_by_issue"] = _round_dict(dict(epic_hours))
        result["project_hours_by_key"] = _round_dict(dict(project_hours))
        result["assignee_hours_by_period"] = {
            "day": _round_nested_dict(dict(assignee_day)),
            "week": _round_nested_dict(dict(assignee_week)),
            "month": _round_nested_dict(dict(assignee_month)),
        }
        return result
    finally:
        try:
            wb.close()
        except Exception:
            pass


def _compute_nested_actual_hours(
    worklog_path: Path,
    work_items_path: Path,
    from_date: date,
    to_date: date,
    mode: str = "log_date",
) -> dict[str, dict[str, float]]:
    aggregate = _compute_actual_hours_aggregate(
        worklog_path=worklog_path,
        work_items_path=work_items_path,
        from_date=from_date,
        to_date=to_date,
        mode=mode,
    )
    return {
        "subtask_hours_by_issue": aggregate.get("subtask_hours_by_issue", {}),
    }


def create_report_server_app(base_dir: Path, folder_raw: str) -> Flask:
    app = Flask(__name__)
    report_dir = resolve_report_html_dir(base_dir, folder_raw)
    report_dir.mkdir(parents=True, exist_ok=True)
    capacity_paths = _resolve_capacity_runtime_paths(base_dir)
    _init_capacity_db(capacity_paths["db_path"])
    _init_performance_settings_db(capacity_paths["db_path"])
    _init_dashboard_risk_settings_db(capacity_paths["db_path"])
    init_report_entities_db(capacity_paths["db_path"])
    init_manage_fields_db(capacity_paths["db_path"])
    init_managed_projects_db(capacity_paths["db_path"])
    _init_page_categories_db(capacity_paths["db_path"])
    pactv_init_db(capacity_paths["db_path"])
    default_project_keys = parse_project_keys_from_env()

    def _resolve_seed_project_name(project_key: str) -> str:
        try:
            return _jira_lookup_project_name(project_key)
        except Exception:
            return project_key

    if default_project_keys:
        try:
            seed_managed_projects(
                capacity_paths["db_path"],
                default_project_keys,
                project_name_resolver=_resolve_seed_project_name,
            )
        except Exception:
            pass
    refresh_lock = threading.Lock()
    actual_hours_cache: dict[tuple[str, str, str, str, str, str, float, float], dict[str, object]] = {}
    pvd_response_cache_version = "v6"
    pactv_jobs_lock = threading.Lock()
    pactv_active_scopes: set[str] = set()
    pactv_pending_by_scope: dict[str, list[dict[str, object]]] = {}

    def _request_roles() -> set[str]:
        direct = _to_text(request.headers.get("X-Role"))
        listed = _to_text(request.headers.get("X-Roles")) or _to_text(request.headers.get("X-User-Role"))
        source = listed or direct
        roles = {
            _to_text(part).lower()
            for part in source.split(",")
            if _to_text(part)
        } if source else set()
        return roles

    def _pactv_require_roles(allowed: set[str]):
        roles = _request_roles()
        if roles.intersection(allowed):
            return None
        return jsonify(
            {
                "ok": False,
                "error": "Forbidden. Missing required role.",
                "required_any_of": sorted(allowed),
            }
        ), 403

    def _is_rlt_managed_project(item: dict) -> bool:
        key = _to_text((item or {}).get("project_key")).upper()
        name = _to_text((item or {}).get("project_name")).lower()
        display = _to_text((item or {}).get("display_name")).lower()
        if key == "RLT":
            return True
        text = f"{name} {display}"
        return "leave tracker" in text and "rnd" in text

    def _managed_project_scope_defaults() -> tuple[set[str], list[str], list[str]]:
        try:
            managed = list_managed_projects(capacity_paths["db_path"], include_inactive=False)
        except Exception:
            managed = []
        managed_keys = sorted({
            _to_text(item.get("project_key")).upper()
            for item in managed
            if _to_text(item.get("project_key"))
        })
        default_selected = sorted({
            _to_text(item.get("project_key")).upper()
            for item in managed
            if _to_text(item.get("project_key")) and not _is_rlt_managed_project(item)
        })
        selected_scope = set(default_selected or managed_keys)
        return selected_scope, managed_keys, default_selected

    def _managed_project_name_map() -> dict[str, str]:
        try:
            managed = list_managed_projects(capacity_paths["db_path"], include_inactive=False)
        except Exception:
            managed = []
        out: dict[str, str] = {}
        for item in managed:
            key = _to_text((item or {}).get("project_key")).upper()
            if not key:
                continue
            name = (
                _to_text((item or {}).get("display_name"))
                or _to_text((item or {}).get("project_name"))
                or key
            )
            out[key] = name
        return out

    def _pactv_rows_with_managed_display_names(rows: list[dict]) -> list[dict]:
        source = rows if isinstance(rows, list) else []
        managed_names = _managed_project_name_map()
        out: list[dict] = []
        for row in source:
            if not isinstance(row, dict):
                out.append(row)
                continue
            item = dict(row)
            project_key = _to_text(item.get("project_key")).upper()
            display_name = managed_names.get(project_key, "")
            if display_name and _to_text(item.get("row_type")).lower() == "project":
                item["project_name"] = display_name
                item["summary"] = display_name
                item["aspect"] = display_name
            out.append(item)
        return out

    def _parse_pactv_request_filters(args_or_payload) -> tuple[str, str, str, set[str], set[str], set[str], object]:
        mode = _to_text((args_or_payload or {}).get("mode")).lower() or "log_date"
        if mode not in PACTV_VALID_MODES:
            raise ValueError("Invalid mode. Expected 'log_date' or 'planned_dates'.")
        if hasattr(args_or_payload, "get"):
            from_raw, to_raw = _resolve_effective_range_from_request(args_or_payload)
        else:
            from_raw = _to_text((args_or_payload or {}).get("from"))
            to_raw = _to_text((args_or_payload or {}).get("to"))
        from_date = _parse_iso_date(from_raw)
        to_date = _parse_iso_date(to_raw)
        if from_date is None or to_date is None:
            raise ValueError("Invalid date format. Expected YYYY-MM-DD.")

        projects_raw = _to_text((args_or_payload or {}).get("projects"))
        statuses_raw = _to_text((args_or_payload or {}).get("statuses"))
        assignees_raw = _to_text((args_or_payload or {}).get("assignees"))
        selected_projects = {
            _to_text(item).upper()
            for item in projects_raw.split(",")
            if _to_text(item)
        } if projects_raw else set()
        selected_statuses = {
            _to_text(item).lower()
            for item in statuses_raw.split(",")
            if _to_text(item)
        } if statuses_raw else set()
        selected_assignees = {
            _to_text(item).lower()
            for item in assignees_raw.split(",")
            if _to_text(item)
        } if assignees_raw else set()
        default_scope, _managed_keys, _default_selected = _managed_project_scope_defaults()
        if not selected_projects:
            selected_projects = set(default_scope)
        return from_raw, to_raw, mode, selected_projects, selected_statuses, selected_assignees, (from_date, to_date)

    def _pactv_scope_key(flt) -> str:
        return "|".join(
            [
                flt.from_date,
                flt.to_date,
                flt.mode,
                flt.projects_scope,
                flt.statuses_scope,
                flt.assignees_scope,
            ]
        )

    def _pactv_managed_options_fallback() -> dict[str, list[str]]:
        try:
            managed = list_managed_projects(capacity_paths["db_path"], include_inactive=False)
        except Exception:
            managed = []
        project_options = []
        seen: set[str] = set()
        for item in managed:
            key = _to_text(item.get("project_key")).upper()
            if not key or key in seen:
                continue
            seen.add(key)
            name = _to_text(item.get("display_name")) or _to_text(item.get("project_name")) or key
            project_options.append({"project_key": key, "project_name": name})
        project_options.sort(key=lambda x: _to_text(x.get("project_key")))
        projects = [_to_text(item.get("project_key")) for item in project_options]
        return {"projects": projects, "project_options": project_options, "statuses": [], "assignees": []}

    def _normalize_pvd_plan_source(raw_value: str) -> str:
        value = _to_text(raw_value).strip().lower()
        if not value:
            return "jira_estimates"
        if value in {"jira_estimates", "epic_planner"}:
            return value
        return ""

    def _pvd_hours_from_man_days(raw_value: object) -> float | None:
        text = _to_text(raw_value)
        if not text:
            return None
        try:
            parsed = float(text)
        except Exception:
            return None
        if parsed <= 0:
            return None
        return _round_hours(parsed * 8.0)

    def _load_pvd_planner_maps() -> tuple[dict[str, dict[str, object]], dict[str, dict[str, object]]]:
        epic_map: dict[str, dict[str, object]] = {}
        story_map: dict[str, dict[str, object]] = {}
        try:
            rows = _load_epics_management_rows(capacity_paths["db_path"])
        except Exception:
            rows = []
        for row in rows:
            epic_key = _to_text((row or {}).get("epic_key")).upper()
            if not epic_key:
                continue
            epic_plan = ((row or {}).get("plans") or {}).get("epic_plan") or {}
            planned_hours = _pvd_hours_from_man_days(epic_plan.get("man_days"))
            planned_start = _to_text(epic_plan.get("start_date"))
            planned_due = _to_text(epic_plan.get("due_date"))
            epic_map[epic_key] = {
                "planned_hours": planned_hours,
                "planned_start": planned_start,
                "planned_due": planned_due,
            }
        conn = sqlite3.connect(capacity_paths["db_path"])
        conn.row_factory = sqlite3.Row
        try:
            story_rows = conn.execute(
                """
                SELECT story_key, start_date, due_date, estimate_hours
                FROM epics_management_story_sync
                """
            ).fetchall()
        except Exception:
            story_rows = []
        finally:
            conn.close()
        for row in story_rows:
            story_key = _to_text(row["story_key"]).upper()
            if not story_key:
                continue
            story_map[story_key] = {
                "planned_hours": _round_hours(float(row["estimate_hours"] or 0.0)),
                "planned_start": _to_text(row["start_date"]),
                "planned_due": _to_text(row["due_date"]),
            }
        return epic_map, story_map

    def _rollup_dispensed_dates(items: list[dict]) -> tuple[str, str]:
        starts: list[date] = []
        dues: list[date] = []
        for item in (items or []):
            start_raw = _to_text((item or {}).get("planned_start"))
            due_raw = _to_text((item or {}).get("planned_due"))
            start_parsed = _parse_iso_date(start_raw)
            due_parsed = _parse_iso_date(due_raw)
            if start_parsed:
                starts.append(start_parsed)
            if due_parsed:
                dues.append(due_parsed)
        start_iso = min(starts).isoformat() if starts else ""
        due_iso = max(dues).isoformat() if dues else ""
        return start_iso, due_iso

    def _init_planned_vs_dispensed_cache_db() -> None:
        conn = sqlite3.connect(capacity_paths["db_path"])
        try:
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS planned_vs_dispensed_cache (
                    cache_key TEXT PRIMARY KEY,
                    from_date TEXT NOT NULL,
                    to_date TEXT NOT NULL,
                    mode TEXT NOT NULL,
                    project_scope TEXT NOT NULL DEFAULT '',
                    payload_json TEXT NOT NULL,
                    fetched_at TEXT NOT NULL DEFAULT (datetime('now'))
                )
                """
            )
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS planned_vs_dispensed_response_cache (
                    cache_key TEXT PRIMARY KEY,
                    endpoint TEXT NOT NULL,
                    from_date TEXT NOT NULL,
                    to_date TEXT NOT NULL,
                    mode TEXT NOT NULL,
                    project_scope TEXT NOT NULL DEFAULT '',
                    statuses_scope TEXT NOT NULL DEFAULT '',
                    assignees_scope TEXT NOT NULL DEFAULT '',
                    project_key TEXT NOT NULL DEFAULT '',
                    payload_json TEXT NOT NULL,
                    fetched_at TEXT NOT NULL DEFAULT (datetime('now'))
                )
                """
            )
            conn.execute(
                """
                CREATE INDEX IF NOT EXISTS idx_pvd_cache_fetched_at
                ON planned_vs_dispensed_cache(fetched_at DESC)
                """
            )
            conn.execute(
                """
                CREATE INDEX IF NOT EXISTS idx_pvd_response_cache_fetched_at
                ON planned_vs_dispensed_response_cache(fetched_at DESC)
                """
            )
            conn.commit()
        finally:
            conn.close()

    def _planned_vs_dispensed_cache_scope(selected_projects: set[str]) -> str:
        return ",".join(sorted({_to_text(item).upper() for item in (selected_projects or set()) if _to_text(item)}))

    def _planned_vs_dispensed_text_scope(values: set[str]) -> str:
        return ",".join(sorted({_to_text(item).lower() for item in (values or set()) if _to_text(item)}))

    def _planned_vs_dispensed_cache_key(
        from_date: date,
        to_date: date,
        mode: str,
        selected_projects: set[str],
    ) -> tuple[str, str]:
        scope = _planned_vs_dispensed_cache_scope(selected_projects)
        return f"{from_date.isoformat()}|{to_date.isoformat()}|{mode}|{scope}", scope

    def _load_planned_vs_dispensed_cache(
        from_date: date,
        to_date: date,
        mode: str,
        selected_projects: set[str],
    ) -> dict[str, object] | None:
        _init_planned_vs_dispensed_cache_db()
        cache_key, _scope = _planned_vs_dispensed_cache_key(from_date, to_date, mode, selected_projects)
        conn = sqlite3.connect(capacity_paths["db_path"])
        conn.row_factory = sqlite3.Row
        try:
            row = conn.execute(
                "SELECT payload_json FROM planned_vs_dispensed_cache WHERE cache_key=?",
                (cache_key,),
            ).fetchone()
            if not row:
                return None
            payload = json.loads(_to_text(row["payload_json"]) or "{}")
            if not isinstance(payload, dict):
                return None
            if not isinstance(payload.get("epics"), list):
                return None
            if not isinstance(payload.get("stories"), list):
                return None
            if not isinstance(payload.get("subtasks"), list):
                return None
            return payload
        except Exception:
            return None
        finally:
            conn.close()

    def _save_planned_vs_dispensed_cache(
        from_date: date,
        to_date: date,
        mode: str,
        selected_projects: set[str],
        hierarchy: dict[str, object],
    ) -> None:
        _init_planned_vs_dispensed_cache_db()
        cache_key, scope = _planned_vs_dispensed_cache_key(from_date, to_date, mode, selected_projects)
        payload_text = json.dumps(hierarchy or {}, ensure_ascii=True, separators=(",", ":"))
        conn = sqlite3.connect(capacity_paths["db_path"])
        try:
            conn.execute(
                """
                INSERT INTO planned_vs_dispensed_cache (
                    cache_key, from_date, to_date, mode, project_scope, payload_json, fetched_at
                ) VALUES (?, ?, ?, ?, ?, ?, datetime('now'))
                ON CONFLICT(cache_key) DO UPDATE SET
                    payload_json=excluded.payload_json,
                    fetched_at=datetime('now')
                """,
                (
                    cache_key,
                    from_date.isoformat(),
                    to_date.isoformat(),
                    mode,
                    scope,
                    payload_text,
                ),
            )
            conn.commit()
        finally:
            conn.close()

    def _get_planned_vs_dispensed_hierarchy_cached(
        from_date: date,
        to_date: date,
        mode: str,
        selected_projects: set[str],
        force_refresh: bool = False,
    ) -> tuple[dict[str, object], str]:
        if not force_refresh:
            cached = _load_planned_vs_dispensed_cache(from_date, to_date, mode, selected_projects)
            if cached is not None:
                return cached, "cache"
        session = get_session()
        hierarchy = _load_planned_vs_dispensed_hierarchy(
            session=session,
            from_date=from_date,
            to_date=to_date,
            mode=mode,
            selected_projects=selected_projects,
        )
        _save_planned_vs_dispensed_cache(from_date, to_date, mode, selected_projects, hierarchy)
        return hierarchy, "jira"

    def _planned_vs_dispensed_response_cache_key(
        endpoint: str,
        from_date: date,
        to_date: date,
        mode: str,
        plan_source: str,
        selected_projects: set[str],
        selected_statuses: set[str],
        selected_assignees: set[str],
        project_key: str = "",
    ) -> tuple[str, str, str, str, str]:
        project_scope = _planned_vs_dispensed_cache_scope(selected_projects)
        statuses_scope = _planned_vs_dispensed_text_scope(selected_statuses)
        assignees_scope = _planned_vs_dispensed_text_scope(selected_assignees)
        pk = _to_text(project_key).upper()
        key = "|".join(
            [
                pvd_response_cache_version,
                endpoint,
                from_date.isoformat(),
                to_date.isoformat(),
                mode,
                plan_source,
                project_scope,
                statuses_scope,
                assignees_scope,
                pk,
            ]
        )
        return key, project_scope, statuses_scope, assignees_scope, pk

    def _load_planned_vs_dispensed_response_cache(
        endpoint: str,
        from_date: date,
        to_date: date,
        mode: str,
        plan_source: str,
        selected_projects: set[str],
        selected_statuses: set[str],
        selected_assignees: set[str],
        project_key: str = "",
    ) -> dict[str, object] | None:
        _init_planned_vs_dispensed_cache_db()
        cache_key, _project_scope, _statuses_scope, _assignees_scope, _pk = _planned_vs_dispensed_response_cache_key(
            endpoint=endpoint,
            from_date=from_date,
            to_date=to_date,
            mode=mode,
            plan_source=plan_source,
            selected_projects=selected_projects,
            selected_statuses=selected_statuses,
            selected_assignees=selected_assignees,
            project_key=project_key,
        )
        conn = sqlite3.connect(capacity_paths["db_path"])
        conn.row_factory = sqlite3.Row
        try:
            row = conn.execute(
                "SELECT payload_json FROM planned_vs_dispensed_response_cache WHERE cache_key=?",
                (cache_key,),
            ).fetchone()
            if not row:
                return None
            payload = json.loads(_to_text(row["payload_json"]) or "{}")
            return payload if isinstance(payload, dict) else None
        except Exception:
            return None
        finally:
            conn.close()

    def _save_planned_vs_dispensed_response_cache(
        endpoint: str,
        from_date: date,
        to_date: date,
        mode: str,
        plan_source: str,
        selected_projects: set[str],
        selected_statuses: set[str],
        selected_assignees: set[str],
        payload: dict[str, object],
        project_key: str = "",
    ) -> None:
        _init_planned_vs_dispensed_cache_db()
        cache_key, project_scope, statuses_scope, assignees_scope, pk = _planned_vs_dispensed_response_cache_key(
            endpoint=endpoint,
            from_date=from_date,
            to_date=to_date,
            mode=mode,
            plan_source=plan_source,
            selected_projects=selected_projects,
            selected_statuses=selected_statuses,
            selected_assignees=selected_assignees,
            project_key=project_key,
        )
        payload_text = json.dumps(payload or {}, ensure_ascii=True, separators=(",", ":"))
        conn = sqlite3.connect(capacity_paths["db_path"])
        try:
            conn.execute(
                """
                INSERT INTO planned_vs_dispensed_response_cache (
                    cache_key, endpoint, from_date, to_date, mode,
                    project_scope, statuses_scope, assignees_scope, project_key,
                    payload_json, fetched_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, datetime('now'))
                ON CONFLICT(cache_key) DO UPDATE SET
                    payload_json=excluded.payload_json,
                    fetched_at=datetime('now')
                """,
                (
                    cache_key,
                    endpoint,
                    from_date.isoformat(),
                    to_date.isoformat(),
                    mode,
                    project_scope,
                    statuses_scope,
                    assignees_scope,
                    pk,
                    payload_text,
                ),
            )
            conn.commit()
        finally:
            conn.close()

    def _init_global_report_date_filter_db(db_path: Path) -> None:
        conn = sqlite3.connect(db_path)
        try:
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS global_report_date_filters (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    from_date TEXT NOT NULL,
                    to_date TEXT NOT NULL,
                    source_page TEXT NOT NULL DEFAULT '',
                    created_at_utc TEXT NOT NULL
                )
                """
            )
            conn.execute(
                """
                CREATE INDEX IF NOT EXISTS idx_global_report_date_filters_created
                ON global_report_date_filters(created_at_utc DESC, id DESC)
                """
            )
            conn.commit()
        finally:
            conn.close()

    def _init_pvd_ui_settings_db(db_path: Path) -> None:
        conn = sqlite3.connect(db_path)
        try:
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS planned_vs_dispensed_ui_settings (
                    id INTEGER PRIMARY KEY CHECK (id = 1),
                    first_column_width_px INTEGER NOT NULL,
                    updated_at_utc TEXT NOT NULL
                )
                """
            )
            conn.commit()
        finally:
            conn.close()

    def _normalize_pvd_first_column_width(raw_value: object) -> int:
        try:
            parsed = int(float(_to_text(raw_value)))
        except Exception:
            raise ValueError("first_column_width_px must be a number.")
        if parsed < 290 or parsed > 900:
            raise ValueError("first_column_width_px must be between 290 and 900.")
        return parsed

    def _load_pvd_ui_settings(db_path: Path) -> dict[str, object]:
        _init_pvd_ui_settings_db(db_path)
        conn = sqlite3.connect(db_path)
        conn.row_factory = sqlite3.Row
        try:
            row = conn.execute(
                """
                SELECT first_column_width_px, updated_at_utc
                FROM planned_vs_dispensed_ui_settings
                WHERE id = 1
                """
            ).fetchone()
        finally:
            conn.close()
        if not row:
            return {"first_column_width_px": 420, "updated_at_utc": ""}
        width_px = int(row["first_column_width_px"] or 420)
        if width_px < 290 or width_px > 900:
            width_px = 420
        return {
            "first_column_width_px": width_px,
            "updated_at_utc": _to_text(row["updated_at_utc"]),
        }

    def _save_pvd_ui_settings(db_path: Path, payload: object) -> dict[str, object]:
        raw = payload if isinstance(payload, dict) else {}
        width_px = _normalize_pvd_first_column_width(raw.get("first_column_width_px"))
        updated_at = _utc_now_iso_z()
        _init_pvd_ui_settings_db(db_path)
        conn = sqlite3.connect(db_path)
        try:
            conn.execute(
                """
                INSERT INTO planned_vs_dispensed_ui_settings(id, first_column_width_px, updated_at_utc)
                VALUES (1, ?, ?)
                ON CONFLICT(id) DO UPDATE SET
                    first_column_width_px = excluded.first_column_width_px,
                    updated_at_utc = excluded.updated_at_utc
                """,
                (width_px, updated_at),
            )
            conn.commit()
        finally:
            conn.close()
        return {"first_column_width_px": width_px, "updated_at_utc": updated_at}

    def _load_latest_global_report_date_filter(db_path: Path) -> dict[str, str] | None:
        _init_global_report_date_filter_db(db_path)
        conn = sqlite3.connect(db_path)
        conn.row_factory = sqlite3.Row
        try:
            row = conn.execute(
                """
                SELECT id, from_date, to_date, source_page, created_at_utc
                FROM global_report_date_filters
                ORDER BY created_at_utc DESC, id DESC
                LIMIT 1
                """
            ).fetchone()
        finally:
            conn.close()
        if not row:
            return None
        from_date = _to_text(row["from_date"])
        to_date = _to_text(row["to_date"])
        from_parsed = _parse_iso_date(from_date)
        to_parsed = _parse_iso_date(to_date)
        if from_parsed is None or to_parsed is None or to_parsed < from_parsed:
            return None
        return {
            "from_date": from_parsed.isoformat(),
            "to_date": to_parsed.isoformat(),
            "source_page": _to_text(row["source_page"]),
            "updated_at_utc": _to_text(row["created_at_utc"]),
        }

    def _save_global_report_date_filter(db_path: Path, from_date: str, to_date: str, source_page: str = "") -> dict[str, str]:
        _init_global_report_date_filter_db(db_path)
        from_parsed = _parse_iso_date(from_date)
        to_parsed = _parse_iso_date(to_date)
        if from_parsed is None or to_parsed is None:
            raise ValueError("Invalid date format. Expected YYYY-MM-DD.")
        if to_parsed < from_parsed:
            raise ValueError("'to_date' must be on or after 'from_date'.")
        created_at = _utc_now_iso_z()
        conn = sqlite3.connect(db_path)
        try:
            conn.execute(
                """
                INSERT INTO global_report_date_filters(from_date, to_date, source_page, created_at_utc)
                VALUES (?, ?, ?, ?)
                """,
                (
                    from_parsed.isoformat(),
                    to_parsed.isoformat(),
                    _to_text(source_page),
                    created_at,
                ),
            )
            conn.commit()
        finally:
            conn.close()
        return {
            "from_date": from_parsed.isoformat(),
            "to_date": to_parsed.isoformat(),
            "source_page": _to_text(source_page),
            "updated_at_utc": created_at,
        }

    def _resolve_effective_range_from_request(
        args,
        *,
        allow_global_fallback: bool = True,
    ) -> tuple[str, str]:
        from_raw = _to_text(args.get("from"))
        to_raw = _to_text(args.get("to"))
        if bool(from_raw) != bool(to_raw):
            raise ValueError("Query params 'from' and 'to' must be provided together.")
        if not from_raw and not to_raw and allow_global_fallback:
            latest = _load_latest_global_report_date_filter(capacity_paths["db_path"])
            if latest:
                return _to_text(latest["from_date"]), _to_text(latest["to_date"])
        if not from_raw or not to_raw:
            raise ValueError("Query params 'from' and 'to' are required.")
        from_parsed = _parse_iso_date(from_raw)
        to_parsed = _parse_iso_date(to_raw)
        if from_parsed is None or to_parsed is None:
            raise ValueError("Invalid date format. Expected YYYY-MM-DD.")
        if to_parsed < from_parsed:
            raise ValueError("'to' must be on or after 'from'.")
        return from_parsed.isoformat(), to_parsed.isoformat()

    _init_global_report_date_filter_db(capacity_paths["db_path"])
    _init_pvd_ui_settings_db(capacity_paths["db_path"])

    @app.after_request
    def add_cors_headers(response):
        response.headers["Access-Control-Allow-Origin"] = "*"
        response.headers["Access-Control-Allow-Headers"] = "Content-Type"
        response.headers["Access-Control-Allow-Methods"] = "GET, POST, PUT, DELETE, OPTIONS"
        req_path = _to_text(getattr(request, "path", ""))
        if req_path.startswith("/api/planned-vs-dispensed/") or req_path.startswith("/api/planned-actual-table-view/"):
            response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
            response.headers["Pragma"] = "no-cache"
            response.headers["Expires"] = "0"
        return response

    @app.route("/")
    def index():
        dashboard_path = report_dir / "dashboard.html"
        if dashboard_path.exists():
            return redirect("/dashboard.html", code=302)
        return redirect("/report_html/", code=302)

    @app.route("/report_html/")
    def list_reports():
        files = sorted([p.name for p in report_dir.glob("*.html")])
        settings_links = "\n".join(
            f'<li><a href="{route}">{label}</a></li>'
            for label, route in _settings_nav_items()
        )
        report_links = "\n".join(
            f'<li><a href="/{name}">{name}</a></li>' for name in files
        )
        if not report_links:
            report_links = "<li>No HTML reports available yet.</li>"
        links = (
            "<h2>Admin Settings</h2><ul>"
            + settings_links
            + "</ul><h2>Reports</h2><ul>"
            + report_links
            + "</ul>"
        )
        return (
            "<!doctype html><html><body>"
            + links
            + "</body></html>"
        )

    @app.route("/api/report/refresh", methods=["POST", "OPTIONS"])
    def refresh_report():
        if request.method == "OPTIONS":
            return ("", 204)
        payload = request.get_json(silent=True) or {}
        report_id = str(payload.get("report", "")).strip()
        if report_id not in REPORT_REFRESH_CHAINS:
            return jsonify(
                {
                    "ok": False,
                    "error": f"Unsupported report id: {report_id}",
                }
            ), 400

        if not refresh_lock.acquire(blocking=False):
            return jsonify(
                {
                    "ok": False,
                    "error": "Another refresh is already running. Try again shortly.",
                }
            ), 409

        started = time.perf_counter()
        steps: list[dict[str, object]] = []
        try:
            for script_name in REPORT_REFRESH_CHAINS[report_id]:
                code, stdout, stderr = _run_script(script_name, base_dir)
                step_data = {
                    "script": script_name,
                    "exit_code": code,
                    "stdout_tail": _tail(stdout),
                    "stderr_tail": _tail(stderr),
                }
                steps.append(step_data)
                if code != 0:
                    duration_sec = round(time.perf_counter() - started, 2)
                    return jsonify(
                        {
                            "ok": False,
                            "report": report_id,
                            "error": f"Step failed: {script_name}",
                            "steps": steps,
                            "duration_sec": duration_sec,
                        }
                    ), 500

            sync_report_html(base_dir, folder_raw)
            duration_sec = round(time.perf_counter() - started, 2)
            return jsonify(
                {
                    "ok": True,
                    "report": report_id,
                    "steps": [step["script"] for step in steps],
                    "duration_sec": duration_sec,
                    "completed_at": time.strftime("%Y-%m-%d %H:%M:%S"),
                }
            )
        finally:
            refresh_lock.release()

    @app.route("/api/report-date-filter", methods=["GET"])
    def get_report_date_filter():
        latest = _load_latest_global_report_date_filter(capacity_paths["db_path"])
        return jsonify({"ok": True, "filter": latest})

    @app.route("/api/report-date-filter", methods=["POST"])
    def save_report_date_filter():
        payload = request.get_json(silent=True) or {}
        try:
            saved = _save_global_report_date_filter(
                capacity_paths["db_path"],
                from_date=_to_text(payload.get("from_date")),
                to_date=_to_text(payload.get("to_date")),
                source_page=_to_text(payload.get("source_page")),
            )
            return jsonify({"ok": True, "filter": saved})
        except ValueError as exc:
            return jsonify({"ok": False, "error": str(exc)}), 400

    @app.route("/api/planned-vs-dispensed/ui-settings", methods=["GET"])
    def get_pvd_ui_settings():
        settings = _load_pvd_ui_settings(capacity_paths["db_path"])
        return jsonify({"ok": True, "settings": settings})

    @app.route("/api/planned-vs-dispensed/ui-settings", methods=["POST"])
    def save_pvd_ui_settings():
        payload = request.get_json(silent=True) or {}
        try:
            settings = _save_pvd_ui_settings(capacity_paths["db_path"], payload)
            return jsonify({"ok": True, "settings": settings})
        except ValueError as exc:
            return jsonify({"ok": False, "error": str(exc)}), 400

    @app.route("/api/capacity", methods=["GET"])
    def get_capacity():
        try:
            from_date, to_date = _resolve_effective_range_from_request(request.args)
            settings = _load_capacity_settings(capacity_paths["db_path"], from_date, to_date)
            result = calculate_capacity_metrics(settings)
            leave = _load_leave_metrics(
                capacity_paths["leave_report_path"],
                result["settings"]["from_date"],
                result["settings"]["to_date"],
                result["settings"],
            )
            remaining = round(
                result["metrics"]["available_capacity_hours"] - leave["taken_hours"] - leave["not_yet_taken_hours"],
                2,
            )
            result["leave_metrics"] = {
                **leave,
                "remaining_balance_hours": remaining,
                "remaining_balance_days": _hours_to_days_over_range(remaining, result["settings"]),
            }
            return jsonify(result)
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400

    @app.route("/api/actual-hours/aggregate", methods=["GET"])
    def actual_hours_aggregate():
        mode = _to_text(request.args.get("mode")).lower() or "log_date"
        report_id = _to_text(request.args.get("report"))
        projects_raw = _to_text(request.args.get("projects"))
        selected_projects = {
            item.strip().upper()
            for item in projects_raw.split(",")
            if item.strip()
        } if projects_raw else None
        try:
            from_raw, to_raw = _resolve_effective_range_from_request(request.args)
        except ValueError as exc:
            return jsonify({"ok": False, "error": str(exc)}), 400
        if mode not in {"log_date", "planned_dates"}:
            return jsonify({"ok": False, "error": "Invalid mode. Expected 'log_date' or 'planned_dates'."}), 400
        from_date = _parse_iso_date(from_raw)
        to_date = _parse_iso_date(to_raw)
        if from_date is None or to_date is None:
            return jsonify({"ok": False, "error": "Invalid date format. Expected YYYY-MM-DD."}), 400

        worklog_path = _resolve_worklog_xlsx_path(base_dir)
        work_items_path = _resolve_work_items_xlsx_path(base_dir)
        worklog_mtime = worklog_path.stat().st_mtime if worklog_path.exists() else 0.0
        work_items_mtime = work_items_path.stat().st_mtime if work_items_path.exists() else 0.0
        cache_key = (
            from_date.isoformat(),
            to_date.isoformat(),
            mode,
            ",".join(sorted(selected_projects or set())),
            str(worklog_path.resolve()),
            str(work_items_path.resolve()),
            worklog_mtime,
            work_items_mtime,
        )
        cached = actual_hours_cache.get(cache_key)
        if cached is None:
            try:
                cached = _compute_actual_hours_aggregate(
                    worklog_path=worklog_path,
                    work_items_path=work_items_path,
                    from_date=from_date,
                    to_date=to_date,
                    mode=mode,
                    selected_projects=selected_projects,
                )
            except Exception as exc:
                return jsonify({"ok": False, "error": f"Failed to aggregate worklogs: {exc}"}), 500
            actual_hours_cache.clear()
            actual_hours_cache[cache_key] = cached
        return jsonify(
            {
                "ok": True,
                "from_date": from_date.isoformat(),
                "to_date": to_date.isoformat(),
                "mode": mode,
                "report": report_id,
                "source_file": str(worklog_path),
                "work_items_file": str(work_items_path),
                **cached,
            }
        )

    @app.route("/api/nested-view/actual-hours", methods=["GET"])
    def nested_view_actual_hours():
        mode = _to_text(request.args.get("mode")).lower() or "log_date"
        try:
            from_raw, to_raw = _resolve_effective_range_from_request(request.args)
        except ValueError as exc:
            return jsonify({"ok": False, "error": str(exc)}), 400
        if mode not in {"log_date", "planned_dates"}:
            return jsonify({"ok": False, "error": "Invalid mode. Expected 'log_date' or 'planned_dates'."}), 400
        from_date = _parse_iso_date(from_raw)
        to_date = _parse_iso_date(to_raw)
        if from_date is None or to_date is None:
            return jsonify({"ok": False, "error": "Invalid date format. Expected YYYY-MM-DD."}), 400

        worklog_path = _resolve_worklog_xlsx_path(base_dir)
        work_items_path = _resolve_work_items_xlsx_path(base_dir)
        worklog_mtime = worklog_path.stat().st_mtime if worklog_path.exists() else 0.0
        work_items_mtime = work_items_path.stat().st_mtime if work_items_path.exists() else 0.0
        cache_key = (
            from_date.isoformat(),
            to_date.isoformat(),
            mode,
            "",
            str(worklog_path.resolve()),
            str(work_items_path.resolve()),
            worklog_mtime,
            work_items_mtime,
        )
        cached = actual_hours_cache.get(cache_key)
        if cached is None:
            try:
                cached = _compute_actual_hours_aggregate(
                    worklog_path=worklog_path,
                    work_items_path=work_items_path,
                    from_date=from_date,
                    to_date=to_date,
                    mode=mode,
                )
            except Exception as exc:
                return jsonify({"ok": False, "error": f"Failed to aggregate worklogs: {exc}"}), 500
            actual_hours_cache.clear()
            actual_hours_cache[cache_key] = cached
        return jsonify(
            {
                "ok": True,
                "from_date": from_date.isoformat(),
                "to_date": to_date.isoformat(),
                "mode": mode,
                "source_file": str(worklog_path),
                "subtask_hours_by_issue": cached.get("subtask_hours_by_issue", {}),
            }
        )

    @app.route("/api/planned-vs-dispensed/summary", methods=["GET"])
    def planned_vs_dispensed_summary():
        mode = _to_text(request.args.get("mode")).lower() or "log_date"
        plan_source = _normalize_pvd_plan_source(_to_text(request.args.get("plan_source")))
        refresh_raw = _to_text(request.args.get("refresh")).lower()
        force_refresh = refresh_raw in {"1", "true", "yes", "y", "on"}
        projects_raw = _to_text(request.args.get("projects"))
        statuses_raw = _to_text(request.args.get("statuses"))
        assignees_raw = _to_text(request.args.get("assignees"))
        try:
            from_raw, to_raw = _resolve_effective_range_from_request(request.args)
        except ValueError as exc:
            return jsonify({"ok": False, "error": str(exc)}), 400
        if mode not in {"log_date", "planned_dates"}:
            return jsonify({"ok": False, "error": "Invalid mode. Expected 'log_date' or 'planned_dates'."}), 400
        if not plan_source:
            return jsonify({"ok": False, "error": "Invalid plan_source. Expected 'jira_estimates' or 'epic_planner'."}), 400
        from_date = _parse_iso_date(from_raw)
        to_date = _parse_iso_date(to_raw)
        if from_date is None or to_date is None:
            return jsonify({"ok": False, "error": "Invalid date format. Expected YYYY-MM-DD."}), 400

        selected_projects_raw = {
            _to_text(item).upper()
            for item in projects_raw.split(",")
            if _to_text(item)
        } if projects_raw else set()
        default_scope, managed_project_keys, default_selected_projects = _managed_project_scope_defaults()
        selected_projects = set(selected_projects_raw or default_scope)
        selected_statuses = {
            _to_text(item).lower()
            for item in statuses_raw.split(",")
            if _to_text(item)
        } if statuses_raw else set()
        selected_assignees = {
            _to_text(item).lower()
            for item in assignees_raw.split(",")
            if _to_text(item)
        } if assignees_raw else set()
        if not force_refresh:
            cached_summary = _load_planned_vs_dispensed_response_cache(
                endpoint="summary",
                from_date=from_date,
                to_date=to_date,
                mode=mode,
                plan_source=plan_source,
                selected_projects=selected_projects,
                selected_statuses=selected_statuses,
                selected_assignees=selected_assignees,
            )
            if cached_summary is not None:
                return jsonify(cached_summary)

        def _match_filter(value: str, selected: set[str]) -> bool:
            if not selected:
                return True
            return _to_text(value).lower() in selected

        try:
            hierarchy, source = _get_planned_vs_dispensed_hierarchy_cached(
                from_date=from_date,
                to_date=to_date,
                mode=mode,
                selected_projects=selected_projects,
                force_refresh=force_refresh,
            )
        except Exception as exc:
            return jsonify({"ok": False, "error": f"Failed to load Jira hierarchy: {exc}"}), 500

        epics = hierarchy.get("epics", []) or []
        stories = hierarchy.get("stories", []) or []
        subtasks = hierarchy.get("subtasks", []) or []
        planner_epic_map: dict[str, dict[str, object]] = {}
        if plan_source == "epic_planner":
            planner_epic_map, _planner_story_map = _load_pvd_planner_maps()
        managed_name_by_key = _managed_project_name_map()
        project_name_by_key: dict[str, str] = {}
        for item in epics + stories + subtasks:
            key = _to_text(item.get("project_key")).upper()
            if not key:
                continue
            name = _to_text(item.get("project_name")) or managed_name_by_key.get(key, key)
            project_name_by_key[key] = name
        story_by_epic: dict[str, list[dict]] = defaultdict(list)
        subtask_by_epic: dict[str, list[dict]] = defaultdict(list)
        for story in stories:
            story_by_epic[_to_text(story.get("epic_key")).upper()].append(story)
        for subtask in subtasks:
            subtask_by_epic[_to_text(subtask.get("epic_key")).upper()].append(subtask)

        project_rows: dict[str, dict[str, object]] = {}
        missing_epics_by_project: dict[str, list[str]] = defaultdict(list)
        for epic in epics:
            if not _match_filter(_to_text(epic.get("status")), selected_statuses):
                continue
            if not _match_filter(_to_text(epic.get("assignee")), selected_assignees):
                continue
            project_key = _to_text(epic.get("project_key")).upper() or "UNKNOWN"
            row = project_rows.setdefault(
                project_key,
                {
                    "project_key": project_key,
                    "project_name": project_name_by_key.get(project_key, managed_name_by_key.get(project_key, project_key)),
                    "planned_epic_hours": 0.0,
                    "dispensed_subtask_hours": 0.0,
                    "remaining_hours": 0.0,
                    "epic_count": 0,
                    "story_count": 0,
                    "subtask_count": 0,
                },
            )
            epic_key = _to_text(epic.get("issue_key")).upper()
            planned_hours = _round_hours(float(epic.get("estimate_hours") or 0.0))
            if plan_source == "epic_planner":
                planner_item = planner_epic_map.get(epic_key) or {}
                planner_hours = planner_item.get("planned_hours")
                if planner_hours is None:
                    missing_epics_by_project[project_key].append(epic_key)
                    planned_hours = 0.0
                else:
                    planned_hours = _round_hours(float(planner_hours or 0.0))
            row["planned_epic_hours"] = float(row["planned_epic_hours"]) + planned_hours
            row["epic_count"] = int(row["epic_count"]) + 1

            filtered_stories = [
                item for item in story_by_epic.get(epic_key, [])
                if _match_filter(_to_text(item.get("status")), selected_statuses)
                and _match_filter(_to_text(item.get("assignee")), selected_assignees)
            ]
            filtered_subtasks = [
                item for item in subtask_by_epic.get(epic_key, [])
                if _match_filter(_to_text(item.get("status")), selected_statuses)
                and _match_filter(_to_text(item.get("assignee")), selected_assignees)
            ]
            row["story_count"] = int(row["story_count"]) + len(filtered_stories)
            row["subtask_count"] = int(row["subtask_count"]) + len(filtered_subtasks)
            row["dispensed_subtask_hours"] = float(row["dispensed_subtask_hours"]) + sum(
                float(item.get("estimate_hours") or 0.0) for item in filtered_subtasks
            )

        rows_out = []
        for row in project_rows.values():
            planned_hours = _round_hours(float(row["planned_epic_hours"]))
            dispensed_hours = _round_hours(float(row["dispensed_subtask_hours"]))
            rows_out.append(
                {
                    **row,
                    "planned_epic_hours": planned_hours,
                    "dispensed_subtask_hours": dispensed_hours,
                    "remaining_hours": _round_hours(planned_hours - dispensed_hours),
                }
            )
        for project_key in sorted(selected_projects):
            if any(_to_text(item.get("project_key")).upper() == project_key for item in rows_out):
                continue
            rows_out.append(
                {
                    "project_key": project_key,
                    "project_name": project_name_by_key.get(project_key, managed_name_by_key.get(project_key, project_key)),
                    "planned_epic_hours": 0.0,
                    "dispensed_subtask_hours": 0.0,
                    "remaining_hours": 0.0,
                    "epic_count": 0,
                    "story_count": 0,
                    "subtask_count": 0,
                }
            )
        rows_out.sort(key=lambda item: (_to_text(item.get("project_key")),))

        project_key_order = managed_project_keys or sorted({
            _to_text(item.get("project_key")).upper()
            for item in epics
            if _to_text(item.get("project_key"))
        })
        project_options = [
            {
                "project_key": key,
                "project_name": project_name_by_key.get(key, managed_name_by_key.get(key, key)),
            }
            for key in project_key_order
        ]
        all_statuses = sorted({
            _to_text(item.get("status"))
            for item in epics + stories + subtasks
            if _to_text(item.get("status"))
        })
        all_assignees = sorted({
            _to_text(item.get("assignee"))
            for item in epics + stories + subtasks
            if _to_text(item.get("assignee"))
        })
        missing_plan_projects = [
            {
                "project_key": key,
                "project_name": project_name_by_key.get(key, managed_name_by_key.get(key, key)),
                "epic_keys": sorted(set(values)),
                "count": len(sorted(set(values))),
            }
            for key, values in sorted(missing_epics_by_project.items())
        ]
        missing_plan = {
            "enabled": plan_source == "epic_planner",
            "message": (
                "The following epics in selected date range do not have planned hours in Epics Planner. "
                "Either switch source or specify plan in Epics Planner."
            ),
            "projects": missing_plan_projects,
            "count": sum(int(item["count"]) for item in missing_plan_projects),
        }

        payload = {
            "ok": True,
            "from_date": from_date.isoformat(),
            "to_date": to_date.isoformat(),
            "mode": mode,
            "plan_source": plan_source,
            "rows": rows_out,
            "selected_projects": sorted(selected_projects),
            "filter_options": {
                "projects": [item.get("project_key") for item in project_options],
                "project_options": project_options,
                "statuses": all_statuses,
                "assignees": all_assignees,
            },
            "default_selected_projects": default_selected_projects,
            "missing_plan": missing_plan,
            "source": source,
        }
        _save_planned_vs_dispensed_response_cache(
            endpoint="summary",
            from_date=from_date,
            to_date=to_date,
            mode=mode,
            plan_source=plan_source,
            selected_projects=selected_projects,
            selected_statuses=selected_statuses,
            selected_assignees=selected_assignees,
            payload=payload,
        )
        return jsonify(payload)

    @app.route("/api/planned-vs-dispensed/details", methods=["GET"])
    def planned_vs_dispensed_details():
        mode = _to_text(request.args.get("mode")).lower() or "log_date"
        plan_source = _normalize_pvd_plan_source(_to_text(request.args.get("plan_source")))
        refresh_raw = _to_text(request.args.get("refresh")).lower()
        force_refresh = refresh_raw in {"1", "true", "yes", "y", "on"}
        project_key = _to_text(request.args.get("project_key")).upper()
        projects_raw = _to_text(request.args.get("projects"))
        statuses_raw = _to_text(request.args.get("statuses"))
        assignees_raw = _to_text(request.args.get("assignees"))
        if not project_key:
            return jsonify({"ok": False, "error": "Query param 'project_key' is required."}), 400
        try:
            from_raw, to_raw = _resolve_effective_range_from_request(request.args)
        except ValueError as exc:
            return jsonify({"ok": False, "error": str(exc)}), 400
        if mode not in {"log_date", "planned_dates"}:
            return jsonify({"ok": False, "error": "Invalid mode. Expected 'log_date' or 'planned_dates'."}), 400
        if not plan_source:
            return jsonify({"ok": False, "error": "Invalid plan_source. Expected 'jira_estimates' or 'epic_planner'."}), 400
        from_date = _parse_iso_date(from_raw)
        to_date = _parse_iso_date(to_raw)
        if from_date is None or to_date is None:
            return jsonify({"ok": False, "error": "Invalid date format. Expected YYYY-MM-DD."}), 400

        selected_projects_raw = {
            _to_text(item).upper()
            for item in projects_raw.split(",")
            if _to_text(item)
        } if projects_raw else set()
        default_scope, _managed_project_keys, _default_selected_projects = _managed_project_scope_defaults()
        selected_projects = set(selected_projects_raw or default_scope)
        selected_projects.add(project_key)
        selected_statuses = {
            _to_text(item).lower()
            for item in statuses_raw.split(",")
            if _to_text(item)
        } if statuses_raw else set()
        selected_assignees = {
            _to_text(item).lower()
            for item in assignees_raw.split(",")
            if _to_text(item)
        } if assignees_raw else set()
        if not force_refresh:
            cached_details = _load_planned_vs_dispensed_response_cache(
                endpoint="details",
                from_date=from_date,
                to_date=to_date,
                mode=mode,
                plan_source=plan_source,
                selected_projects=selected_projects,
                selected_statuses=selected_statuses,
                selected_assignees=selected_assignees,
                project_key=project_key,
            )
            if cached_details is not None:
                return jsonify(cached_details)

        def _match_filter(value: str, selected: set[str]) -> bool:
            if not selected:
                return True
            return _to_text(value).lower() in selected

        try:
            hierarchy, source = _get_planned_vs_dispensed_hierarchy_cached(
                from_date=from_date,
                to_date=to_date,
                mode=mode,
                selected_projects=selected_projects,
                force_refresh=force_refresh,
            )
        except Exception as exc:
            return jsonify({"ok": False, "error": f"Failed to load Jira hierarchy: {exc}"}), 500

        planner_epic_map: dict[str, dict[str, object]] = {}
        planner_story_map: dict[str, dict[str, object]] = {}
        if plan_source == "epic_planner":
            planner_epic_map, planner_story_map = _load_pvd_planner_maps()

        epics = [
            item for item in (hierarchy.get("epics", []) or [])
            if _to_text(item.get("project_key")).upper() == project_key
        ]
        managed_name_by_key = _managed_project_name_map()
        project_name = managed_name_by_key.get(project_key, project_key)
        for item in epics:
            raw_name = _to_text(item.get("project_name"))
            if raw_name:
                project_name = raw_name
                break
        stories = hierarchy.get("stories", []) or []
        subtasks = hierarchy.get("subtasks", []) or []
        stories_by_epic: dict[str, list[dict]] = defaultdict(list)
        subtasks_by_story: dict[str, list[dict]] = defaultdict(list)
        for story in stories:
            stories_by_epic[_to_text(story.get("epic_key")).upper()].append(story)
        for subtask in subtasks:
            subtasks_by_story[_to_text(subtask.get("story_key")).upper()].append(subtask)

        epic_rows: list[dict[str, object]] = []
        total_planned = 0.0
        total_dispensed = 0.0
        missing_epic_keys: list[str] = []

        for epic in epics:
            if not _match_filter(_to_text(epic.get("status")), selected_statuses):
                continue
            if not _match_filter(_to_text(epic.get("assignee")), selected_assignees):
                continue
            epic_key = _to_text(epic.get("issue_key")).upper()
            story_rows: list[dict[str, object]] = []
            epic_dispensed = 0.0
            epic_subtask_count = 0
            source_stories = stories_by_epic.get(epic_key, [])
            source_stories.sort(key=lambda item: (_to_text(item.get("summary")).lower(), _to_text(item.get("issue_key"))))

            for story in source_stories:
                story_key = _to_text(story.get("issue_key")).upper()
                story_subtasks = [
                    sub for sub in subtasks_by_story.get(story_key, [])
                    if _match_filter(_to_text(sub.get("status")), selected_statuses)
                    and _match_filter(_to_text(sub.get("assignee")), selected_assignees)
                ]
                story_subtasks.sort(key=lambda item: (_to_text(item.get("summary")).lower(), _to_text(item.get("issue_key"))))
                story_dispensed_from_subtasks = _round_hours(sum(float(item.get("estimate_hours") or 0.0) for item in story_subtasks))
                story_estimate_hours = _round_hours(float(story.get("estimate_hours") or 0.0))
                story_planned_hours = story_estimate_hours
                story_planned_start = _to_text(story.get("planned_start"))
                story_planned_due = _to_text(story.get("planned_due"))
                if plan_source == "epic_planner":
                    planner_story = planner_story_map.get(story_key) or {}
                    planner_story_hours = planner_story.get("planned_hours")
                    if planner_story_hours is not None:
                        story_planned_hours = _round_hours(float(planner_story_hours or 0.0))
                    else:
                        story_planned_hours = 0.0
                    story_planned_start = _to_text(planner_story.get("planned_start"))
                    story_planned_due = _to_text(planner_story.get("planned_due"))
                story_dispensed = story_dispensed_from_subtasks
                if not story_subtasks and not _match_filter(_to_text(story.get("status")), selected_statuses):
                    continue
                if not story_subtasks and not _match_filter(_to_text(story.get("assignee")), selected_assignees):
                    continue
                story_dispensed_start, story_dispensed_due = _rollup_dispensed_dates(story_subtasks)
                epic_dispensed += story_dispensed
                epic_subtask_count += len(story_subtasks)
                story_rows.append(
                    {
                        "issue_key": _to_text(story.get("issue_key")).upper(),
                        "issue_kind": _to_text(story.get("issue_kind")) or "story",
                        "issue_type_name": _to_text(story.get("issue_type_name")) or "Story",
                        "jira_url": _to_text(story.get("jira_url")) or _jira_browse_url(_to_text(story.get("issue_key")).upper()),
                        "summary": _to_text(story.get("summary")),
                        "assignee": _to_text(story.get("assignee")) or "Unassigned",
                        "status": _to_text(story.get("status")),
                        "planned_start": story_planned_start,
                        "planned_due": story_planned_due,
                        "estimate_hours": story_planned_hours,
                        "planned_hours": story_planned_hours,
                        "dispensed_estimates": story_dispensed,
                        "dispensed_start": story_dispensed_start,
                        "dispensed_due": story_dispensed_due,
                        "remaining": _round_hours(story_planned_hours - story_dispensed),
                        "subtasks": [
                            {
                                "issue_key": _to_text(subtask.get("issue_key")).upper(),
                                "issue_kind": _to_text(subtask.get("issue_kind")) or "subtask",
                                "issue_type_name": _to_text(subtask.get("issue_type_name")) or "Subtask",
                                "jira_url": _to_text(subtask.get("jira_url")) or _jira_browse_url(_to_text(subtask.get("issue_key")).upper()),
                                "summary": _to_text(subtask.get("summary")),
                                "assignee": _to_text(subtask.get("assignee")) or "Unassigned",
                                "status": _to_text(subtask.get("status")),
                                "planned_start": _to_text(subtask.get("planned_start")),
                                "planned_due": _to_text(subtask.get("planned_due")),
                                "estimate_hours": _round_hours(float(subtask.get("estimate_hours") or 0.0)),
                                "planned_hours": None,
                                "dispensed_estimates": _round_hours(float(subtask.get("estimate_hours") or 0.0)),
                                "dispensed_start": _to_text(subtask.get("planned_start")),
                                "dispensed_due": _to_text(subtask.get("planned_due")),
                                "remaining": 0.0,
                            }
                            for subtask in story_subtasks
                        ],
                    }
                )

            epic_planned = _round_hours(float(epic.get("estimate_hours") or 0.0))
            epic_planned_start = _to_text(epic.get("planned_start"))
            epic_planned_due = _to_text(epic.get("planned_due"))
            if plan_source == "epic_planner":
                planner_epic = planner_epic_map.get(epic_key) or {}
                planner_epic_hours = planner_epic.get("planned_hours")
                if planner_epic_hours is None:
                    missing_epic_keys.append(epic_key)
                    epic_planned = 0.0
                else:
                    epic_planned = _round_hours(float(planner_epic_hours or 0.0))
                epic_planned_start = _to_text(planner_epic.get("planned_start"))
                epic_planned_due = _to_text(planner_epic.get("planned_due"))
            epic_dispensed = _round_hours(epic_dispensed)
            epic_dispensed_start, epic_dispensed_due = _rollup_dispensed_dates([
                subtask
                for story_row in story_rows
                for subtask in (story_row.get("subtasks") or [])
            ])
            total_planned += epic_planned
            total_dispensed += epic_dispensed
            epic_rows.append(
                {
                    "issue_key": _to_text(epic.get("issue_key")).upper(),
                    "issue_kind": _to_text(epic.get("issue_kind")) or "epic",
                    "issue_type_name": _to_text(epic.get("issue_type_name")) or "Epic",
                    "jira_url": _to_text(epic.get("jira_url")) or _jira_browse_url(_to_text(epic.get("issue_key")).upper()),
                    "summary": _to_text(epic.get("summary")),
                    "assignee": _to_text(epic.get("assignee")) or "Unassigned",
                    "status": _to_text(epic.get("status")),
                    "planned_start": epic_planned_start,
                    "planned_due": epic_planned_due,
                    "estimate_hours": epic_planned,
                    "planned_hours": epic_planned,
                    "dispensed_estimates": epic_dispensed,
                    "dispensed_start": epic_dispensed_start,
                    "dispensed_due": epic_dispensed_due,
                    "remaining": _round_hours(epic_planned - epic_dispensed),
                    "story_count": len(story_rows),
                    "subtask_count": epic_subtask_count,
                    "stories": story_rows,
                }
            )

        epic_rows.sort(key=lambda item: (_to_text(item.get("summary")).lower(), _to_text(item.get("issue_key"))))
        missing_epic_keys_sorted = sorted(set(missing_epic_keys))
        missing_plan = {
            "enabled": plan_source == "epic_planner",
            "project_key": project_key,
            "project_name": project_name,
            "message": (
                "The following epics in selected date range do not have planned hours in Epics Planner. "
                "Either switch source or specify plan in Epics Planner."
            ),
            "epic_keys": missing_epic_keys_sorted,
            "count": len(missing_epic_keys_sorted),
        }
        payload = {
            "ok": True,
            "project_key": project_key,
            "project_name": project_name,
            "jira_base_url": _to_text(BASE_URL).rstrip("/"),
            "from_date": from_date.isoformat(),
            "to_date": to_date.isoformat(),
            "mode": mode,
            "plan_source": plan_source,
            "totals": {
                "planned_epic_hours": _round_hours(total_planned),
                "dispensed_subtask_hours": _round_hours(total_dispensed),
                "remaining_hours": _round_hours(total_planned - total_dispensed),
                "epic_count": len(epic_rows),
            },
            "epics": epic_rows,
            "missing_plan": missing_plan,
            "source": source,
        }
        _save_planned_vs_dispensed_response_cache(
            endpoint="details",
            from_date=from_date,
            to_date=to_date,
            mode=mode,
            plan_source=plan_source,
            selected_projects=selected_projects,
            selected_statuses=selected_statuses,
            selected_assignees=selected_assignees,
            project_key=project_key,
            payload=payload,
        )
        return jsonify(payload)

    def _run_pactv_refresh(
        run_id: str,
        flt,
        from_date: date,
        to_date: date,
        selected_projects: set[str],
        selected_statuses: set[str],
        selected_assignees: set[str],
        force_full: bool,
        scope_key: str,
    ) -> None:
        source_used = "jira_full"
        stats: dict[str, object] = {}

        def _ensure_not_canceled(step: str) -> None:
            if pactv_is_cancel_requested(capacity_paths["db_path"], run_id):
                pactv_mark_run_canceled(
                    capacity_paths["db_path"],
                    run_id,
                    reason=f"Canceled by user during '{step}'. Rollback applied; previous snapshot retained.",
                )
                raise RuntimeError("run_canceled")

        try:
            pactv_update_run_progress(capacity_paths["db_path"], run_id, "initializing", 5)
            _ensure_not_canceled("initializing")
            session = get_session()
            hierarchy: dict[str, object] = {}
            sync_state = pactv_load_sync_state(capacity_paths["db_path"], scope_key)

            if not force_full and sync_state and _to_text(sync_state.get("last_successful_fetch_utc")):
                pactv_update_run_progress(capacity_paths["db_path"], run_id, "fetching_incremental", 25)
                _ensure_not_canceled("fetching_incremental")
                try:
                    hierarchy = _load_planned_actual_hierarchy_incremental(
                        session=session,
                        from_date=from_date,
                        to_date=to_date,
                        mode=flt.mode,
                        selected_projects=selected_projects,
                        last_successful_fetch_utc=_to_text(sync_state.get("last_successful_fetch_utc")),
                    )
                    source_used = "jira_incremental"
                    if not hierarchy.get("epics") and not hierarchy.get("stories") and not hierarchy.get("subtasks"):
                        _ensure_not_canceled("fallback_fetching_full")
                        hierarchy = _load_planned_vs_dispensed_hierarchy(
                            session=session,
                            from_date=from_date,
                            to_date=to_date,
                            mode=flt.mode,
                            selected_projects=selected_projects,
                        )
                        source_used = "jira_full_fallback"
                except Exception:
                    _ensure_not_canceled("fallback_fetching_full")
                    hierarchy = _load_planned_vs_dispensed_hierarchy(
                        session=session,
                        from_date=from_date,
                        to_date=to_date,
                        mode=flt.mode,
                        selected_projects=selected_projects,
                    )
                    source_used = "jira_full_fallback"
            else:
                pactv_update_run_progress(capacity_paths["db_path"], run_id, "fetching_full", 25)
                _ensure_not_canceled("fetching_full")
                hierarchy = _load_planned_vs_dispensed_hierarchy(
                    session=session,
                    from_date=from_date,
                    to_date=to_date,
                    mode=flt.mode,
                    selected_projects=selected_projects,
                )
                source_used = "jira_full"

            pactv_update_run_progress(capacity_paths["db_path"], run_id, "loading_actuals", 55)
            _ensure_not_canceled("loading_actuals")
            subtask_keys = [
                _to_text(item.get("issue_key")).upper()
                for item in (hierarchy.get("subtasks", []) or [])
                if _to_text(item.get("issue_key"))
            ]
            actual_by_subtask = _fetch_subtask_actual_hours_by_keys(session, subtask_keys)
            pactv_update_run_progress(capacity_paths["db_path"], run_id, "computing", 75)
            _ensure_not_canceled("computing")
            rows, totals, options = pactv_build_snapshot_payload(
                hierarchy=hierarchy,
                actual_hours_by_subtask=actual_by_subtask,
                selected_projects=selected_projects,
                selected_statuses=selected_statuses,
                selected_assignees=selected_assignees,
            )

            snapshot_id = f"pactv-{uuid.uuid4().hex[:12]}"
            watermark_utc = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
            source_payload = {
                "source": source_used,
                "computed_at_utc": watermark_utc,
            }
            pactv_update_run_progress(capacity_paths["db_path"], run_id, "saving", 90)
            _ensure_not_canceled("saving")
            pactv_save_snapshot(
                db_path=capacity_paths["db_path"],
                snapshot_id=snapshot_id,
                flt=flt,
                rows=rows,
                totals=totals,
                options=options,
                source=source_payload,
                watermark_utc=watermark_utc,
            )
            counts = {
                "epics": len(hierarchy.get("epics", []) or []),
                "stories": len(hierarchy.get("stories", []) or []),
                "subtasks": len(hierarchy.get("subtasks", []) or []),
                "rows": len(rows),
            }
            pactv_save_source_audit(
                db_path=capacity_paths["db_path"],
                run_id=run_id,
                snapshot_id=snapshot_id,
                source_type=source_used,
                counts=counts,
                payload={"totals": totals, "counts": counts},
            )
            pactv_save_snapshot_event(
                db_path=capacity_paths["db_path"],
                snapshot_id=snapshot_id,
                run_id=run_id,
                event_type="refresh_success",
                actor="system",
                details={"source": source_used, "counts": counts},
            )
            pactv_save_sync_state(capacity_paths["db_path"], scope_key, watermark_utc, run_id)
            pactv_prune_old_data(capacity_paths["db_path"], retention_days=PACTV_DEFAULT_RETENTION_DAYS)
            stats = {
                "snapshot_id": snapshot_id,
                "counts": counts,
            }
            pactv_finish_run_success(capacity_paths["db_path"], run_id, source_used, stats)
        except Exception as exc:
            if "run_canceled" in str(exc):
                return
            pactv_save_snapshot_event(
                db_path=capacity_paths["db_path"],
                snapshot_id="",
                run_id=run_id,
                event_type="refresh_failed",
                actor="system",
                details={"error": _to_text(exc)},
            )
            pactv_finish_run_failed(
                capacity_paths["db_path"],
                run_id,
                f"Refresh failed: {exc}",
                stats=stats,
            )

    def _start_pactv_job(job: dict[str, object]) -> None:
        def _runner_for_scope(first_job: dict[str, object]) -> None:
            current_job = dict(first_job)
            while True:
                run_id_local = _to_text(current_job.get("run_id"))
                scope_key_local = _to_text(current_job.get("scope_key"))
                if run_id_local:
                    _run_pactv_refresh(
                        run_id=run_id_local,
                        flt=current_job.get("flt"),
                        from_date=current_job.get("from_date"),
                        to_date=current_job.get("to_date"),
                        selected_projects=current_job.get("selected_projects") or set(),
                        selected_statuses=current_job.get("selected_statuses") or set(),
                        selected_assignees=current_job.get("selected_assignees") or set(),
                        force_full=bool(current_job.get("force_full")),
                        scope_key=scope_key_local,
                    )
                with pactv_jobs_lock:
                    queue = pactv_pending_by_scope.get(scope_key_local, [])
                    next_job = None
                    while queue and next_job is None:
                        candidate = queue.pop(0)
                        candidate_run_id = _to_text(candidate.get("run_id"))
                        if pactv_begin_queued_run(capacity_paths["db_path"], candidate_run_id):
                            next_job = dict(candidate)
                        else:
                            pactv_mark_run_canceled(
                                capacity_paths["db_path"],
                                candidate_run_id,
                                reason="Canceled while queued. Rollback applied; previous snapshot retained.",
                            )
                    if queue:
                        pactv_pending_by_scope[scope_key_local] = queue
                    else:
                        pactv_pending_by_scope.pop(scope_key_local, None)
                    if next_job is None:
                        pactv_active_scopes.discard(scope_key_local)
                        break
                current_job = next_job

        threading.Thread(target=_runner_for_scope, args=(job,), daemon=True).start()

    @app.route("/api/planned-actual-table-view/summary", methods=["GET"])
    def planned_actual_table_view_summary():
        try:
            from_raw, to_raw, mode, selected_projects, selected_statuses, selected_assignees, _parsed = _parse_pactv_request_filters(request.args)
        except ValueError as exc:
            return jsonify({"ok": False, "error": str(exc)}), 400

        flt = pactv_make_filter(
            from_raw,
            to_raw,
            mode,
            selected_projects,
            selected_statuses,
            selected_assignees,
        )
        snapshot = pactv_load_snapshot_by_filter(capacity_paths["db_path"], flt)
        if not snapshot:
            return jsonify(
                {
                    "ok": True,
                    "from_date": from_raw,
                    "to_date": to_raw,
                    "mode": mode,
                    "rows": [],
                    "totals": {
                        "planned_hours": 0.0,
                        "actual_hours": 0.0,
                        "variance_hours": 0.0,
                        "project_count": 0,
                        "epic_count": 0,
                        "story_count": 0,
                        "subtask_count": 0,
                    },
                    "filter_options": _pactv_managed_options_fallback(),
                    "selected_projects": sorted(selected_projects),
                    "selected_statuses": sorted(selected_statuses),
                    "selected_assignees": sorted(selected_assignees),
                    "source": "snapshot_db",
                    "needs_refresh": True,
                }
            )
        managed_names = _managed_project_name_map()
        snapshot_options = snapshot.get("options", {}) if isinstance(snapshot.get("options"), dict) else {}
        merged_project_keys = sorted(
            {
                *_managed_project_scope_defaults()[1],
                *[
                    _to_text(item).upper()
                    for item in (snapshot_options.get("projects", []) or [])
                    if _to_text(item)
                ],
                *sorted(selected_projects),
            }
        )
        return jsonify(
            {
                "ok": True,
                "from_date": snapshot.get("from_date"),
                "to_date": snapshot.get("to_date"),
                "mode": snapshot.get("mode"),
                "rows": _pactv_rows_with_managed_display_names(snapshot.get("rows", [])),
                "totals": snapshot.get("totals", {}),
                "filter_options": {
                    **snapshot_options,
                    "projects": merged_project_keys,
                    "project_options": [
                        {
                            "project_key": key,
                            "project_name": (managed_names.get(key) or key),
                        }
                        for key in merged_project_keys
                    ],
                },
                "selected_projects": sorted(selected_projects),
                "selected_statuses": sorted(selected_statuses),
                "selected_assignees": sorted(selected_assignees),
                "snapshot_id": snapshot.get("snapshot_id"),
                "computed_at_utc": snapshot.get("computed_at_utc"),
                "is_official": bool(snapshot.get("is_official")),
                "official_pinned_by": _to_text(snapshot.get("official_pinned_by")),
                "official_pinned_at_utc": _to_text(snapshot.get("official_pinned_at_utc")),
                "lifecycle_state": _to_text(snapshot.get("lifecycle_state")) or "active",
                "source": "snapshot_db",
                "needs_refresh": False,
            }
        )

    @app.route("/api/planned-actual-table-view/refresh", methods=["POST"])
    def planned_actual_table_view_refresh():
        role_error = _pactv_require_roles({"operator", "admin"})
        if role_error is not None:
            return role_error
        payload = request.get_json(silent=True) or {}
        try:
            from_raw, to_raw, mode, selected_projects, selected_statuses, selected_assignees, parsed = _parse_pactv_request_filters(payload)
        except ValueError as exc:
            return jsonify({"ok": False, "error": str(exc)}), 400
        from_date, to_date = parsed
        force_full_raw = _to_text(payload.get("force_full")).lower()
        force_full = force_full_raw in {"1", "true", "yes", "y", "on"}
        run_sync = _to_text(payload.get("run_sync")).lower() in {"1", "true", "yes", "y", "on"}
        max_attempts_raw = _to_text(payload.get("max_attempts"))
        try:
            max_attempts = max(1, min(5, int(max_attempts_raw or "1")))
        except Exception:
            max_attempts = 1

        flt = pactv_make_filter(
            from_raw,
            to_raw,
            mode,
            selected_projects,
            selected_statuses,
            selected_assignees,
        )
        scope_key = _pactv_scope_key(flt)
        run_id = f"pactv-{int(time.time())}-{uuid.uuid4().hex[:8]}"
        job = {
            "run_id": run_id,
            "flt": flt,
            "from_date": from_date,
            "to_date": to_date,
            "selected_projects": selected_projects,
            "selected_statuses": selected_statuses,
            "selected_assignees": selected_assignees,
            "force_full": force_full,
            "scope_key": scope_key,
        }

        should_queue = False
        queue_position = 0
        with pactv_jobs_lock:
            db_running = pactv_has_running_run_for_scope(capacity_paths["db_path"], flt)
            active = scope_key in pactv_active_scopes
            should_queue = db_running or active
            if should_queue:
                pactv_create_run(
                    capacity_paths["db_path"],
                    run_id,
                    flt,
                    force_full,
                    status="queued",
                    progress_step="queued",
                    progress_pct=0,
                    attempt=1,
                    max_attempts=max_attempts,
                )
                queue = pactv_pending_by_scope.setdefault(scope_key, [])
                queue.append(job)
                queue_position = len(queue)
            else:
                pactv_active_scopes.add(scope_key)
                pactv_create_run(
                    capacity_paths["db_path"],
                    run_id,
                    flt,
                    force_full,
                    status="running",
                    progress_step="initializing",
                    progress_pct=1,
                    attempt=1,
                    max_attempts=max_attempts,
                )

        if should_queue:
            return jsonify(
                {
                    "ok": True,
                    "run_id": run_id,
                    "status": "queued",
                    "queue_position": queue_position,
                    "attempt": 1,
                    "max_attempts": max_attempts,
                    "queued_at_utc": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
                }
            ), 202

        if run_sync:
            _run_pactv_refresh(
                run_id=run_id,
                flt=flt,
                from_date=from_date,
                to_date=to_date,
                selected_projects=selected_projects,
                selected_statuses=selected_statuses,
                selected_assignees=selected_assignees,
                force_full=force_full,
                scope_key=scope_key,
            )
            with pactv_jobs_lock:
                pactv_active_scopes.discard(scope_key)
            run_row = pactv_get_run(capacity_paths["db_path"], run_id) or {}
            return jsonify(
                {
                    "ok": True,
                    "run_id": run_id,
                    "status": _to_text(run_row.get("status")) or "unknown",
                    "attempt": int(run_row.get("attempt") or 1),
                    "max_attempts": int(run_row.get("max_attempts") or 1),
                    "queued_at_utc": _to_text(run_row.get("queued_at_utc")),
                }
            )

        _start_pactv_job(job)
        return jsonify({"ok": True, "run_id": run_id, "status": "running", "attempt": 1, "max_attempts": max_attempts, "queued_at_utc": ""}), 202

    @app.route("/api/planned-actual-table-view/refresh/<path:run_id>", methods=["GET"])
    def planned_actual_table_view_refresh_status(run_id: str):
        role_error = _pactv_require_roles({"viewer", "operator", "admin"})
        if role_error is not None:
            return role_error
        row = pactv_get_run(capacity_paths["db_path"], run_id)
        if not row:
            return jsonify({"ok": False, "error": "Run not found."}), 404
        return jsonify(
            {
                "ok": True,
                "run_id": _to_text(row.get("run_id")),
                "status": _to_text(row.get("status")),
                "step": _to_text(row.get("progress_step")),
                "progress": int(row.get("progress_pct") or 0),
                "source": _to_text(row.get("source")),
                "started_at_utc": _to_text(row.get("started_at_utc")),
                "completed_at_utc": _to_text(row.get("completed_at_utc")),
                "queued_at_utc": _to_text(row.get("queued_at_utc")),
                "error": _to_text(row.get("error_text")),
                "attempt": int(row.get("attempt") or 1),
                "max_attempts": int(row.get("max_attempts") or 1),
                "next_retry_at_utc": _to_text(row.get("next_retry_at_utc")),
                "stats": json.loads(_to_text(row.get("stats_json")) or "{}"),
            }
        )

    @app.route("/api/planned-actual-table-view/filter-options", methods=["GET"])
    def planned_actual_table_view_filter_options():
        options = _pactv_managed_options_fallback()
        latest = pactv_list_history(capacity_paths["db_path"], limit=1)
        if latest:
            snapshot_id = _to_text(latest[0].get("snapshot_id"))
            snapshot = pactv_load_snapshot_by_id(capacity_paths["db_path"], snapshot_id)
            if snapshot and isinstance(snapshot.get("options"), dict):
                snap_options = snapshot.get("options") or {}
                options = {
                    **options,
                    **snap_options,
                }
        managed_names = _managed_project_name_map()
        managed_keys = _managed_project_scope_defaults()[1]
        project_keys = options.get("projects", []) if isinstance(options, dict) else []
        merged_keys = sorted(
            {
                *[_to_text(key).upper() for key in (project_keys or []) if _to_text(key)],
                *[_to_text(key).upper() for key in (managed_keys or []) if _to_text(key)],
            }
        )
        options["projects"] = merged_keys
        options["project_options"] = [
            {
                "project_key": _to_text(key).upper(),
                "project_name": managed_names.get(_to_text(key).upper(), _to_text(key).upper()),
            }
            for key in merged_keys
            if _to_text(key)
        ]
        return jsonify({"ok": True, "options": options})

    @app.route("/api/planned-actual-table-view/queue", methods=["GET"])
    def planned_actual_table_view_queue():
        role_error = _pactv_require_roles({"viewer", "operator", "admin"})
        if role_error is not None:
            return role_error
        limit_raw = _to_text(request.args.get("limit"))
        try:
            limit = int(limit_raw) if limit_raw else 20
        except Exception:
            limit = 20
        rows = pactv_list_queue(capacity_paths["db_path"], limit=limit)
        return jsonify({"ok": True, "rows": rows})

    @app.route("/api/planned-actual-table-view/cancel", methods=["POST"])
    def planned_actual_table_view_cancel():
        role_error = _pactv_require_roles({"operator", "admin"})
        if role_error is not None:
            return role_error
        payload = request.get_json(silent=True) or {}
        run_id = _to_text(payload.get("run_id"))
        if not run_id:
            running = [row for row in pactv_list_queue(capacity_paths["db_path"], limit=50) if _to_text(row.get("status")) == "running"]
            if running:
                run_id = _to_text(running[0].get("run_id"))
        if not run_id:
            return jsonify({"ok": False, "error": "run_id is required."}), 400

        row_before = pactv_get_run(capacity_paths["db_path"], run_id)
        if not row_before:
            return jsonify({"ok": False, "error": "Run not found."}), 404
        status_before = _to_text(row_before.get("status")).lower()
        if status_before not in {"queued", "running"}:
            return jsonify({"ok": False, "error": f"Run is not cancelable (status={status_before})."}), 409

        accepted = pactv_request_cancel(capacity_paths["db_path"], run_id)
        if not accepted:
            return jsonify({"ok": False, "error": "Cancel request was not accepted."}), 409

        if status_before == "queued":
            with pactv_jobs_lock:
                for scope_key, jobs in list(pactv_pending_by_scope.items()):
                    filtered = [item for item in jobs if _to_text(item.get("run_id")) != run_id]
                    if filtered:
                        pactv_pending_by_scope[scope_key] = filtered
                    else:
                        pactv_pending_by_scope.pop(scope_key, None)
            pactv_mark_run_canceled(
                capacity_paths["db_path"],
                run_id,
                reason="Canceled while queued. Rollback applied; previous snapshot retained.",
            )
            return jsonify({"ok": True, "run_id": run_id, "status": "canceled", "message": "Queued fetch canceled and rolled back."})

        return jsonify(
            {
                "ok": True,
                "run_id": run_id,
                "status": "cancel_requested",
                "message": "Cancel requested. Running fetch will stop shortly and rollback.",
            }
        )

    @app.route("/api/planned-actual-table-view/history", methods=["GET"])
    def planned_actual_table_view_history():
        role_error = _pactv_require_roles({"viewer", "operator", "admin"})
        if role_error is not None:
            return role_error
        limit_raw = _to_text(request.args.get("limit"))
        try:
            limit = int(limit_raw) if limit_raw else 30
        except Exception:
            limit = 30
        rows = pactv_list_history(capacity_paths["db_path"], limit=limit)
        return jsonify({"ok": True, "rows": rows})

    @app.route("/api/planned-actual-table-view/diff", methods=["GET"])
    def planned_actual_table_view_diff():
        role_error = _pactv_require_roles({"viewer", "operator", "admin"})
        if role_error is not None:
            return role_error
        left_id = _to_text(request.args.get("left_snapshot_id"))
        right_id = _to_text(request.args.get("right_snapshot_id"))
        if not left_id or not right_id:
            return jsonify({"ok": False, "error": "Both left_snapshot_id and right_snapshot_id are required."}), 400
        left = pactv_load_snapshot_by_id(capacity_paths["db_path"], left_id)
        right = pactv_load_snapshot_by_id(capacity_paths["db_path"], right_id)
        if not left or not right:
            return jsonify({"ok": False, "error": "One or both snapshots not found."}), 404
        return jsonify(
            {
                "ok": True,
                "left_snapshot_id": left_id,
                "right_snapshot_id": right_id,
                "delta": pactv_diff_snapshots(left, right),
            }
        )

    @app.route("/api/planned-actual-table-view/export", methods=["POST"])
    def planned_actual_table_view_export():
        role_error = _pactv_require_roles({"viewer", "operator", "admin"})
        if role_error is not None:
            return role_error
        payload = request.get_json(silent=True) or {}
        fmt = _to_text(payload.get("format")).lower() or "csv"
        snapshot_id = _to_text(payload.get("snapshot_id"))
        snapshot = pactv_load_snapshot_by_id(capacity_paths["db_path"], snapshot_id) if snapshot_id else None
        if snapshot is None:
            try:
                from_raw, to_raw, mode, selected_projects, selected_statuses, selected_assignees, _parsed = _parse_pactv_request_filters(payload)
            except ValueError as exc:
                return jsonify({"ok": False, "error": str(exc)}), 400
            flt = pactv_make_filter(
                from_raw,
                to_raw,
                mode,
                selected_projects,
                selected_statuses,
                selected_assignees,
            )
            snapshot = pactv_load_snapshot_by_filter(capacity_paths["db_path"], flt)
        if snapshot is None:
            return jsonify({"ok": False, "error": "Snapshot not found for export."}), 404

        if fmt == "xlsx":
            payload_bytes = pactv_export_xlsx_bytes(snapshot)
            filename = f"planned_actual_table_view_{_to_text(snapshot.get('snapshot_id'))}.xlsx"
            return send_file(
                io.BytesIO(payload_bytes),
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                as_attachment=True,
                download_name=filename,
            )
        if fmt != "csv":
            return jsonify({"ok": False, "error": "Unsupported format. Expected 'csv' or 'xlsx'."}), 400
        payload_bytes = pactv_export_csv_bytes(snapshot)
        filename = f"planned_actual_table_view_{_to_text(snapshot.get('snapshot_id'))}.csv"
        return send_file(
            io.BytesIO(payload_bytes),
            mimetype="text/csv; charset=utf-8",
            as_attachment=True,
            download_name=filename,
        )

    @app.route("/api/planned-actual-table-view/ui-settings", methods=["GET"])
    def planned_actual_table_view_get_ui_settings():
        role_error = _pactv_require_roles({"viewer", "operator", "admin"})
        if role_error is not None:
            return role_error
        return jsonify({"ok": True, "settings": pactv_load_ui_settings(capacity_paths["db_path"])})

    @app.route("/api/planned-actual-table-view/ui-settings", methods=["POST"])
    def planned_actual_table_view_save_ui_settings():
        role_error = _pactv_require_roles({"admin"})
        if role_error is not None:
            return role_error
        payload = request.get_json(silent=True) or {}
        settings = pactv_save_ui_settings(capacity_paths["db_path"], payload if isinstance(payload, dict) else {})
        return jsonify({"ok": True, "settings": settings})

    @app.route("/api/planned-actual-table-view/snapshots/<path:snapshot_id>/pin-official", methods=["POST"])
    def planned_actual_table_view_pin_official(snapshot_id: str):
        role_error = _pactv_require_roles({"admin"})
        if role_error is not None:
            return role_error
        actor = _to_text(request.headers.get("X-Actor")) or _to_text(request.headers.get("X-User")) or "unknown"
        pinned = pactv_pin_official_snapshot(capacity_paths["db_path"], snapshot_id, actor)
        if not pinned:
            return jsonify({"ok": False, "error": "Snapshot not found."}), 404
        return jsonify({"ok": True, "snapshot_id": snapshot_id, "snapshot": pinned})

    @app.route("/api/planned-actual-table-view/snapshots/<path:snapshot_id>/unpin-official", methods=["POST"])
    def planned_actual_table_view_unpin_official(snapshot_id: str):
        role_error = _pactv_require_roles({"admin"})
        if role_error is not None:
            return role_error
        actor = _to_text(request.headers.get("X-Actor")) or _to_text(request.headers.get("X-User")) or "unknown"
        unpinned = pactv_unpin_official_snapshot(capacity_paths["db_path"], snapshot_id, actor)
        if not unpinned:
            return jsonify({"ok": False, "error": "Snapshot not found."}), 404
        return jsonify({"ok": True, "snapshot_id": snapshot_id, "snapshot": unpinned})

    @app.route("/api/capacity", methods=["POST"])
    def save_capacity():
        try:
            payload = request.get_json(silent=True) or {}
            saved = _save_capacity_settings(capacity_paths["db_path"], payload)
            result = calculate_capacity_metrics(saved)
            leave = _load_leave_metrics(
                capacity_paths["leave_report_path"],
                result["settings"]["from_date"],
                result["settings"]["to_date"],
                result["settings"],
            )
            remaining = round(
                result["metrics"]["available_capacity_hours"] - leave["taken_hours"] - leave["not_yet_taken_hours"],
                2,
            )
            result["leave_metrics"] = {
                **leave,
                "remaining_balance_hours": remaining,
                "remaining_balance_days": _hours_to_days_over_range(remaining, result["settings"]),
            }
            return jsonify(result)
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400

    @app.route("/api/capacity", methods=["DELETE"])
    def delete_capacity():
        try:
            from_date = _to_text(request.args.get("from"))
            to_date = _to_text(request.args.get("to"))
            if not from_date or not to_date:
                payload = request.get_json(silent=True) or {}
                from_date = _to_text(payload.get("from_date"))
                to_date = _to_text(payload.get("to_date"))
            if not from_date or not to_date:
                return jsonify({"error": "Range 'from/to' or 'from_date/to_date' is required."}), 400
            deleted = _delete_capacity_settings(capacity_paths["db_path"], from_date, to_date)
            return jsonify(
                {
                    "deleted": bool(deleted),
                    "from_date": from_date,
                    "to_date": to_date,
                }
            )
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400

    @app.route("/api/capacity/calculate", methods=["POST"])
    def calculate_capacity():
        try:
            payload = request.get_json(silent=True) or {}
            result = calculate_capacity_metrics(payload)
            leave = _load_leave_metrics(
                capacity_paths["leave_report_path"],
                result["settings"]["from_date"],
                result["settings"]["to_date"],
                result["settings"],
            )
            remaining = round(
                result["metrics"]["available_capacity_hours"] - leave["taken_hours"] - leave["not_yet_taken_hours"],
                2,
            )
            result["leave_metrics"] = {
                **leave,
                "remaining_balance_hours": remaining,
                "remaining_balance_days": _hours_to_days_over_range(remaining, result["settings"]),
            }
            return jsonify(result)
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400

    @app.route("/api/capacity/profiles", methods=["GET"])
    def list_capacity_profiles():
        return jsonify({"profiles": _list_capacity_profiles(capacity_paths["db_path"])})

    @app.route("/api/capacity/assignee-count", methods=["GET"])
    def capacity_assignee_count():
        try:
            rows = _read_summary_xlsx(capacity_paths["summary_path"])
            assignees = {
                _to_text(item.get("issue_assignee")) or "Unassigned"
                for item in rows
            }
            return jsonify({"assignee_count": len(assignees)})
        except Exception as exc:
            return jsonify({"assignee_count": 0, "error": str(exc)})

    @app.route("/api/performance/settings", methods=["GET"])
    def get_performance_settings():
        try:
            settings = _load_performance_settings(capacity_paths["db_path"])
            return jsonify({"settings": settings, "source": "db"})
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400

    @app.route("/api/performance/settings", methods=["POST"])
    def save_performance_settings():
        try:
            payload = request.get_json(silent=True) or {}
            _normalize_performance_settings(payload, require_all_fields=True)
            saved = _save_performance_settings(capacity_paths["db_path"], payload)
            return jsonify({"settings": saved, "source": "db"})
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400

    @app.route("/api/dashboard-risk/settings", methods=["GET"])
    def get_dashboard_risk_settings():
        try:
            settings = _load_dashboard_risk_settings(capacity_paths["db_path"])
            return jsonify({"settings": settings, "source": "db"})
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400

    @app.route("/api/dashboard-risk/settings", methods=["POST"])
    def save_dashboard_risk_settings():
        try:
            payload = request.get_json(silent=True) or {}
            saved = _save_dashboard_risk_settings(capacity_paths["db_path"], payload)
            return jsonify({"settings": saved, "source": "db"})
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400

    @app.route("/api/performance/assignees", methods=["GET"])
    def list_performance_assignees():
        assignees = _list_assignees_from_summary(capacity_paths["summary_path"])
        return jsonify({"assignees": assignees})

    @app.route("/api/performance/teams", methods=["GET"])
    def list_performance_teams():
        teams = _list_performance_teams(capacity_paths["db_path"])
        return jsonify({"teams": teams})

    @app.route("/api/performance/teams", methods=["POST"])
    def save_performance_team():
        try:
            payload = request.get_json(silent=True) or {}
            team_name = payload.get("team_name")
            team_leader = payload.get("team_leader")
            assignees = payload.get("assignees") or []
            saved = _save_performance_team(capacity_paths["db_path"], team_name, assignees, team_leader)
            return jsonify({"team": saved})
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400

    @app.route("/api/performance/teams/<path:team_name>", methods=["DELETE"])
    def delete_performance_team(team_name: str):
        try:
            deleted = _delete_performance_team(capacity_paths["db_path"], team_name)
            return jsonify({"deleted": bool(deleted), "team_name": team_name})
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400

    @app.route("/api/report-entities", methods=["GET"])
    def get_report_entities():
        try:
            return jsonify(
                {
                    "entities": load_report_entities(capacity_paths["db_path"]),
                    "global_settings": load_report_entity_global_settings(capacity_paths["db_path"]),
                    "source": "db",
                }
            )
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400

    @app.route("/api/report-entities", methods=["PUT"])
    def save_report_entities_api():
        try:
            payload = request.get_json(silent=True) or {}
            entities = payload.get("entities")
            global_settings = payload.get("global_settings")
            if entities is not None:
                save_report_entities(capacity_paths["db_path"], entities)
            if global_settings is not None:
                save_report_entity_global_settings(capacity_paths["db_path"], global_settings)
            return jsonify(
                {
                    "entities": load_report_entities(capacity_paths["db_path"]),
                    "global_settings": load_report_entity_global_settings(capacity_paths["db_path"]),
                    "source": "db",
                }
            )
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400

    @app.route("/api/report-entities/reset", methods=["POST"])
    def reset_report_entities_api():
        try:
            return jsonify(reset_report_entities_to_defaults(capacity_paths["db_path"]))
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400

    @app.route("/api/manage-fields", methods=["GET"])
    def get_manage_fields():
        try:
            include_inactive = _to_text(request.args.get("include_inactive")).lower() in {"1", "true", "yes", "y"}
            fields = load_manage_fields(capacity_paths["db_path"], include_inactive=include_inactive)
            entity_catalog = [
                {"entity_key": item["entity_key"], "label": item["label"]}
                for item in load_report_entities(capacity_paths["db_path"])
            ]
            return jsonify(
                {
                    "fields": fields,
                    "entity_catalog": entity_catalog,
                    "include_inactive": include_inactive,
                    "source": "db",
                }
            )
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400

    @app.route("/api/manage-fields", methods=["POST"])
    def create_manage_field_api():
        try:
            payload = request.get_json(silent=True) or {}
            created = create_manage_field(capacity_paths["db_path"], payload)
            return jsonify({"field": created, "source": "db"})
        except ValueError as exc:
            message = str(exc)
            status = 409 if "already exists" in message else 400
            return jsonify({"error": message}), status

    @app.route("/api/manage-fields/<path:field_key>", methods=["PUT"])
    def update_manage_field_api(field_key: str):
        try:
            payload = request.get_json(silent=True) or {}
            updated = update_manage_field(capacity_paths["db_path"], field_key, payload)
            return jsonify({"field": updated, "source": "db"})
        except LookupError as exc:
            return jsonify({"error": str(exc)}), 404
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400

    @app.route("/api/manage-fields/<path:field_key>", methods=["DELETE"])
    def delete_manage_field_api(field_key: str):
        try:
            deleted = soft_delete_manage_field(capacity_paths["db_path"], field_key)
            return jsonify({"field": deleted, "deleted": True, "source": "db"})
        except LookupError as exc:
            return jsonify({"error": str(exc)}), 404
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400

    @app.route("/api/manage-fields/<path:field_key>/restore", methods=["POST"])
    def restore_manage_field_api(field_key: str):
        try:
            restored = restore_manage_field(capacity_paths["db_path"], field_key)
            return jsonify({"field": restored, "source": "db"})
        except LookupError as exc:
            return jsonify({"error": str(exc)}), 404
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400

    @app.route("/api/projects", methods=["GET"])
    def get_projects():
        try:
            include_inactive = _to_text(request.args.get("include_inactive")).lower() in {"1", "true", "yes", "y"}
            projects = list_managed_projects(capacity_paths["db_path"], include_inactive=include_inactive)
            return jsonify({"projects": projects, "include_inactive": include_inactive, "source": "db"})
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400

    @app.route("/api/projects", methods=["POST"])
    def create_project_api():
        try:
            payload = request.get_json(silent=True) or {}
            created = create_managed_project(capacity_paths["db_path"], payload)
            return jsonify({"project": created, "source": "db"})
        except ValueError as exc:
            message = str(exc)
            status = 409 if "already exists" in message else 400
            return jsonify({"error": message}), status

    @app.route("/api/projects/<path:project_key>", methods=["PUT"])
    def update_project_api(project_key: str):
        try:
            payload = request.get_json(silent=True) or {}
            updated = update_managed_project(capacity_paths["db_path"], project_key, payload)
            return jsonify({"project": updated, "source": "db"})
        except LookupError as exc:
            return jsonify({"error": str(exc)}), 404
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400

    @app.route("/api/projects/<path:project_key>", methods=["DELETE"])
    def delete_project_api(project_key: str):
        try:
            deleted = soft_delete_managed_project(capacity_paths["db_path"], project_key)
            return jsonify({"project": deleted, "deleted": True, "source": "db"})
        except LookupError as exc:
            return jsonify({"error": str(exc)}), 404
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400

    @app.route("/api/projects/<path:project_key>/restore", methods=["POST"])
    def restore_project_api(project_key: str):
        try:
            restored = restore_managed_project(capacity_paths["db_path"], project_key)
            return jsonify({"project": restored, "source": "db"})
        except LookupError as exc:
            return jsonify({"error": str(exc)}), 404
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400

    @app.route("/api/jira/projects/search", methods=["GET"])
    def jira_projects_search_api():
        query = _to_text(request.args.get("q"))
        limit_raw = _to_text(request.args.get("limit")) or "25"
        try:
            limit = int(limit_raw)
        except ValueError:
            limit = 25
        try:
            projects = _jira_search_projects(query, limit=limit)
            return jsonify({"projects": projects, "query": query, "source": "jira"})
        except Exception as exc:
            return jsonify({"error": f"Failed to fetch Jira projects: {exc}"}), 502

    @app.route("/api/page-categories", methods=["GET"])
    def get_page_categories_api():
        try:
            state = _load_page_categories(capacity_paths["db_path"])
            nav_config = _build_navigation_from_page_categories(capacity_paths["db_path"])
            return jsonify(
                {
                    **state,
                    "navigation": nav_config,
                    "source": "db",
                }
            )
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        except Exception as exc:
            return jsonify({"error": f"Failed to load page categories: {exc}"}), 500

    @app.route("/api/page-categories", methods=["PUT"])
    def save_page_categories_api():
        try:
            payload = request.get_json(silent=True) or {}
            state = _save_page_categories_payload(capacity_paths["db_path"], payload)
            nav_config = _build_navigation_from_page_categories(capacity_paths["db_path"])
            return jsonify(
                {
                    **state,
                    "navigation": nav_config,
                    "saved": True,
                    "source": "db",
                }
            )
        except LookupError as exc:
            return jsonify({"error": str(exc)}), 404
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        except Exception as exc:
            return jsonify({"error": f"Failed to save page categories: {exc}"}), 500

    @app.route("/api/page-categories/categories", methods=["POST"])
    def create_page_category_api():
        try:
            payload = request.get_json(silent=True) or {}
            category = _create_page_category(capacity_paths["db_path"], payload)
            state = _load_page_categories(capacity_paths["db_path"])
            return jsonify({"category": category, **state, "created": True, "source": "db"})
        except ValueError as exc:
            message = str(exc)
            status = 409 if "already exists" in message else 400
            return jsonify({"error": message}), status
        except Exception as exc:
            return jsonify({"error": f"Failed to create category: {exc}"}), 500

    @app.route("/api/page-categories/categories/<int:category_id>", methods=["PUT"])
    def update_page_category_api(category_id: int):
        try:
            payload = request.get_json(silent=True) or {}
            category = _update_page_category(capacity_paths["db_path"], category_id, payload)
            state = _load_page_categories(capacity_paths["db_path"])
            return jsonify({"category": category, **state, "updated": True, "source": "db"})
        except LookupError as exc:
            return jsonify({"error": str(exc)}), 404
        except ValueError as exc:
            message = str(exc)
            status = 409 if "already exists" in message else 400
            return jsonify({"error": message}), status
        except Exception as exc:
            return jsonify({"error": f"Failed to update category: {exc}"}), 500

    @app.route("/api/page-categories/categories/<int:category_id>", methods=["DELETE"])
    def delete_page_category_api(category_id: int):
        try:
            deleted_id = _delete_page_category(capacity_paths["db_path"], category_id)
            state = _load_page_categories(capacity_paths["db_path"])
            return jsonify({"deleted": True, "category_id": deleted_id, **state, "source": "db"})
        except LookupError as exc:
            return jsonify({"error": str(exc)}), 404
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        except Exception as exc:
            return jsonify({"error": f"Failed to delete category: {exc}"}), 500

    @app.route("/api/epics-management/dropdown-options", methods=["GET"])
    def get_epics_management_dropdown_options_api():
        try:
            options = _load_epics_dropdown_options(capacity_paths["db_path"])
            return jsonify(
                {
                    "product_category_options": options.get("product_category", []),
                    "component_options": options.get("component", []),
                    "plan_status_options": options.get("plan_status", []),
                    "source": "epics_management_db",
                }
            )
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        except Exception as exc:
            return jsonify({"error": f"Failed to load dropdown options: {exc}"}), 500

    @app.route("/api/epics-management/dropdown-options", methods=["PUT"])
    def save_epics_management_dropdown_options_api():
        try:
            payload = request.get_json(silent=True) or {}
            options = _save_epics_dropdown_options(capacity_paths["db_path"], payload)
            return jsonify(
                {
                    "product_category_options": options.get("product_category", []),
                    "component_options": options.get("component", []),
                    "plan_status_options": options.get("plan_status", []),
                    "source": "epics_management_db",
                }
            )
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        except Exception as exc:
            return jsonify({"error": f"Failed to save dropdown options: {exc}"}), 500

    @app.route("/api/epics-management/plan-columns", methods=["GET"])
    def list_epics_management_plan_columns_api():
        try:
            include_inactive = _to_text(request.args.get("include_inactive")).lower() in {"1", "true", "yes", "y"}
            columns = _load_epics_plan_columns(capacity_paths["db_path"], include_inactive=include_inactive)
            return jsonify({"columns": columns, "include_inactive": include_inactive, "source": "epics_management_db"})
        except Exception as exc:
            return jsonify({"error": f"Failed to load plan columns: {exc}"}), 500

    @app.route("/api/epics-management/plan-columns", methods=["POST"])
    def create_epics_management_plan_column_api():
        try:
            payload = request.get_json(silent=True) or {}
            created = _create_epics_plan_column(capacity_paths["db_path"], payload)
            return jsonify({"column": created, "source": "epics_management_db"}), 201
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        except Exception as exc:
            return jsonify({"error": f"Failed to create plan column: {exc}"}), 500

    @app.route("/api/epics-management/plan-columns/<path:column_key>", methods=["PUT"])
    def update_epics_management_plan_column_api(column_key: str):
        try:
            payload = request.get_json(silent=True) or {}
            updated = _update_epics_plan_column(capacity_paths["db_path"], column_key, payload)
            return jsonify({"column": updated, "source": "epics_management_db"})
        except LookupError as exc:
            return jsonify({"error": str(exc)}), 404
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        except Exception as exc:
            return jsonify({"error": f"Failed to update plan column: {exc}"}), 500

    @app.route("/api/epics-management/plan-columns/order", methods=["PUT"])
    def reorder_epics_management_plan_columns_api():
        try:
            payload = request.get_json(silent=True) or {}
            columns = _reorder_epics_plan_columns(capacity_paths["db_path"], payload)
            return jsonify({"columns": columns, "source": "epics_management_db"})
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        except Exception as exc:
            return jsonify({"error": f"Failed to reorder plan columns: {exc}"}), 500

    @app.route("/api/epics-management/plan-columns/<path:column_key>", methods=["DELETE"])
    def delete_epics_management_plan_column_api(column_key: str):
        try:
            columns = _delete_epics_plan_column(capacity_paths["db_path"], column_key)
            return jsonify({"columns": columns, "deleted": True, "column_key": _to_text(column_key), "source": "epics_management_db"})
        except LookupError as exc:
            return jsonify({"error": str(exc)}), 404
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        except Exception as exc:
            return jsonify({"error": f"Failed to delete plan column: {exc}"}), 500

    @app.route("/api/epics-management/plan-columns/<path:column_key>/restore", methods=["POST"])
    def restore_epics_management_plan_column_api(column_key: str):
        try:
            restored = _restore_epics_plan_column(capacity_paths["db_path"], column_key)
            return jsonify({"column": restored, "restored": True, "source": "epics_management_db"})
        except LookupError as exc:
            return jsonify({"error": str(exc)}), 404
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        except Exception as exc:
            return jsonify({"error": f"Failed to restore plan column: {exc}"}), 500

    @app.route("/api/epics-management/rows", methods=["GET"])
    def epics_management_rows_api():
        try:
            rows = _load_epics_management_rows(settings_db_path=capacity_paths["db_path"])
            return jsonify({"rows": rows, "source": "epics_management_db"})
        except Exception as exc:
            return jsonify({"error": f"Failed to load epics-management rows: {exc}"}), 500

    @app.route("/api/epics-management/rows", methods=["POST"])
    def create_epics_management_row_api():
        try:
            payload = request.get_json(silent=True) or {}
            row = _save_epics_management_row(capacity_paths["db_path"], payload)
            return jsonify({"row": row, "source": "epics_management_db"}), 201
        except _EpicCreateConflictError as exc:
            return jsonify(
                {
                    "error": str(exc),
                    "code": "epic_key_exists",
                    "conflict_epic_key": exc.conflict_epic_key,
                    "vacant_tmp_key": exc.vacant_tmp_key,
                    "can_reuse_vacant_tmp_key": bool(exc.vacant_tmp_key),
                }
            ), 409
        except ValueError as exc:
            message = str(exc)
            status = 409 if "already exists" in message else 400
            return jsonify({"error": message}), status
        except Exception as exc:
            return jsonify({"error": f"Failed to create epic row: {exc}"}), 500

    @app.route("/api/epics-management/rows/<path:epic_key>", methods=["PUT"])
    def update_epics_management_row_api(epic_key: str):
        try:
            payload = request.get_json(silent=True) or {}
            row = _update_epics_management_row(capacity_paths["db_path"], epic_key, payload)
            return jsonify({"row": row, "source": "epics_management_db"})
        except LookupError as exc:
            return jsonify({"error": str(exc)}), 404
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        except Exception as exc:
            return jsonify({"error": f"Failed to update epic row: {exc}"}), 500

    @app.route("/api/epics-management/rows/<path:epic_key>", methods=["DELETE"])
    def delete_epics_management_row_api(epic_key: str):
        try:
            deleted_key = _delete_epics_management_row(capacity_paths["db_path"], epic_key)
            return jsonify({"deleted": True, "epic_key": deleted_key, "source": "epics_management_db"})
        except LookupError as exc:
            return jsonify({"error": str(exc)}), 404
        except Exception as exc:
            return jsonify({"error": f"Failed to delete epic row: {exc}"}), 500

    @app.route("/api/ipp-meeting-dashboard/data", methods=["GET"])
    def ipp_meeting_dashboard_data_api():
        try:
            from generate_ipp_meeting_dashboard import build_payload_from_sources

            payload = build_payload_from_sources(base_dir=base_dir)
            return jsonify(payload)
        except Exception as exc:
            return jsonify({"error": f"Failed to build IPP meeting dashboard data: {exc}"}), 500

    @app.route("/api/epics-management/rows/<path:epic_key>/sync-jira-plan", methods=["POST"])
    def sync_epics_management_row_from_jira_api(epic_key: str):
        try:
            payload = request.get_json(silent=True) or {}
            jira_url = _to_text(payload.get("jira_url"))
            plan_jira_links_raw = payload.get("plan_jira_links")
            plan_jira_links = plan_jira_links_raw if isinstance(plan_jira_links_raw, dict) else {}
            row = _sync_epic_plan_from_jira(
                settings_db_path=capacity_paths["db_path"],
                epic_key=epic_key,
                jira_url_override=jira_url,
                plan_jira_overrides=plan_jira_links,
            )
            return jsonify(
                {
                    "row": row,
                    "source": "jira_sync",
                    "synced_story_count": int(row.get("synced_story_count") or 0),
                }
            )
        except LookupError as exc:
            return jsonify({"error": str(exc)}), 404
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        except Exception as exc:
            return jsonify({"error": f"Failed to sync epic from Jira: {exc}"}), 500

    @app.route(CAPACITY_SETTINGS_ROUTE, methods=["GET"])
    def capacity_settings():
        return _capacity_settings_html()

    @app.route(PERFORMANCE_SETTINGS_ROUTE, methods=["GET"])
    def performance_settings():
        return _performance_settings_html()

    @app.route(DASHBOARD_RISK_SETTINGS_ROUTE, methods=["GET"])
    def dashboard_risk_settings():
        return _dashboard_risk_settings_html()

    @app.route(REPORT_ENTITIES_SETTINGS_ROUTE, methods=["GET"])
    def report_entities_settings():
        return _report_entities_settings_html()

    @app.route(MANAGE_FIELDS_SETTINGS_ROUTE, methods=["GET"])
    def manage_fields_settings():
        return _manage_fields_settings_html()

    @app.route(PROJECTS_SETTINGS_ROUTE, methods=["GET"])
    def projects_settings():
        return _projects_settings_html()

    @app.route(PAGE_CATEGORIES_SETTINGS_ROUTE, methods=["GET"])
    def page_categories_settings():
        return _page_categories_settings_html()

    @app.route(EPICS_DROPDOWN_OPTIONS_SETTINGS_ROUTE, methods=["GET"])
    def epics_dropdown_options_settings():
        return _epics_dropdown_options_settings_html()

    @app.route(EPIC_PHASES_SETTINGS_ROUTE, methods=["GET"])
    def epic_phases_settings():
        return _epic_phases_settings_html()

    @app.route(EPICS_MANAGEMENT_SETTINGS_ROUTE, methods=["GET"])
    def epics_management_settings():
        return _epics_management_settings_html()

    @app.route("/settings/capactiy", methods=["GET"])
    def capacity_settings_typo_redirect():
        return redirect(CAPACITY_SETTINGS_ROUTE, code=302)

    @app.route("/<path:requested_path>")
    def serve_report_asset(requested_path: str):
        target = (report_dir / requested_path).resolve()
        if not target.exists() or not target.is_file():
            return jsonify({"error": "Not found"}), 404
        if report_dir.resolve() not in target.parents and target != report_dir.resolve():
            return jsonify({"error": "Invalid path"}), 400

        if target.suffix.lower() == ".html":
            report_id = REPORT_FILENAME_TO_ID.get(target.name, "")
            html = target.read_text(encoding="utf-8")
            html = _use_local_icons(html)
            html = _inject_shared_date_filter_script(html)
            if report_id:
                html = _inject_refresh_ui(html, report_id)
            return html

        return send_file(target)

    return app


def run_report_server(base_dir: Path, folder_raw: str, host: str, port: int) -> None:
    report_dir = resolve_report_html_dir(base_dir, folder_raw)
    report_dir.mkdir(parents=True, exist_ok=True)
    app = create_report_server_app(base_dir=base_dir, folder_raw=folder_raw)

    dashboard_path = report_dir / "dashboard.html"
    if dashboard_path.exists():
        url = f"http://{host}:{port}/dashboard.html"
    else:
        url = f"http://{host}:{port}/"

    print(f"\n[server] Serving reports from: {report_dir}")
    print(f"[server] Open: {url}")
    print("[server] Press Ctrl+C to stop.")
    app.run(host=host, port=port)
