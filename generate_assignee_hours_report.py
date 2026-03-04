"""
Generate assignee-hours summary artifacts from subtask worklogs.

Inputs:
- 2_jira_subtask_worklogs.xlsx

Outputs:
- assignee_hours_report.xlsx
- assignee_hours_report.html
"""
from __future__ import annotations

import argparse
import json
import os
import re
import sqlite3
from datetime import date, datetime, timedelta, timezone
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

DEFAULT_WORKLOG_INPUT_XLSX = "2_jira_subtask_worklogs.xlsx"
DEFAULT_WORK_ITEMS_INPUT_XLSX = "1_jira_work_items_export.xlsx"
DEFAULT_SUMMARY_OUTPUT_XLSX = "assignee_hours_report.xlsx"
DEFAULT_HTML_OUTPUT = "assignee_hours_report.html"
DEFAULT_CAPACITY_DB = "assignee_hours_capacity.db"
DEFAULT_LEAVE_REPORT_INPUT_XLSX = "rlt_leave_report.xlsx"
SUMMARY_SHEET = "AssigneeHours"
DAY_CROSSTAB_SHEET = "Day Crosstab"
WEEK_CROSSTAB_SHEET = "Week Crosstab"
MONTH_CROSSTAB_SHEET = "Month Crosstab"

SUMMARY_HEADERS = [
    "project_key",
    "worklog_date",
    "period_day",
    "period_week",
    "period_month",
    "issue_assignee",
    "worklog_author",
    "hours_logged",
]


def _resolve_path(value: str, base_dir: Path) -> Path:
    path = Path(value)
    if path.is_absolute():
        return path
    return base_dir / path


def _to_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _to_float(value) -> float:
    if value in (None, ""):
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def extract_project_key(issue_id: str) -> str:
    text = _to_text(issue_id).upper()
    if not text:
        return "UNKNOWN"
    if "-" not in text:
        return "UNKNOWN"
    project = text.split("-", 1)[0].strip()
    if not project:
        return "UNKNOWN"
    if not re.match(r"^[A-Z0-9]+$", project):
        return "UNKNOWN"
    return project


def parse_worklog_date(value) -> str:
    if value in (None, ""):
        return ""

    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()

    text = _to_text(value)
    if not text:
        return ""

    formats = (
        "%Y-%m-%dT%H:%M:%S.%f%z",
        "%Y-%m-%dT%H:%M:%S%z",
        "%Y-%m-%dT%H:%M:%S.%f",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%d",
    )
    for fmt in formats:
        try:
            dt = datetime.strptime(text, fmt)
            return dt.date().isoformat()
        except ValueError:
            continue

    try:
        dt = datetime.fromisoformat(text.replace("Z", "+00:00"))
        return dt.date().isoformat()
    except ValueError:
        return ""


def iso_week_code(iso_date: str) -> str:
    dt = date.fromisoformat(iso_date)
    iso = dt.isocalendar()
    return f"{iso.year}-W{iso.week:02d}"


def iso_week_label(iso_date: str) -> str:
    dt = date.fromisoformat(iso_date)
    iso = dt.isocalendar()
    monday = dt - timedelta(days=iso.weekday - 1)
    sunday = monday + timedelta(days=6)
    return f"{iso.year}-W{iso.week:02d} ({monday:%b %d} - {sunday:%b %d})"


def month_code(iso_date: str) -> str:
    dt = date.fromisoformat(iso_date)
    return f"{dt.year:04d}-{dt.month:02d}"


def _worklog_user(row: dict) -> str:
    return _to_text(row.get("worklog_author")) or _to_text(row.get("issue_assignee")) or "Unassigned"


def _load_worklog_rows(input_path: Path) -> list[dict]:
    if not input_path.exists():
        raise FileNotFoundError(f"Worklog workbook not found: {input_path}")

    wb = load_workbook(input_path, read_only=True, data_only=True)
    ws = wb.active

    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header:
        wb.close()
        raise ValueError("Worklog workbook has no header row.")

    headers = [_to_text(h) for h in header]
    index = {name: idx for idx, name in enumerate(headers)}
    required = ["issue_id", "issue_assignee", "worklog_started", "hours_logged"]
    missing = [name for name in required if name not in index]
    if missing:
        wb.close()
        raise ValueError(f"Worklog workbook missing required columns: {missing}")

    rows: list[dict] = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        issue_id = _to_text(raw[index["issue_id"]])
        worklog_date = parse_worklog_date(raw[index["worklog_started"]])
        hours_logged = _to_float(raw[index["hours_logged"]])
        issue_assignee = _to_text(raw[index["issue_assignee"]]) or "Unassigned"
        worklog_author = _to_text(raw[index["worklog_author"]]) if "worklog_author" in index else ""
        if not worklog_author:
            worklog_author = issue_assignee
        if not issue_id or not worklog_date or hours_logged <= 0:
            continue

        rows.append(
            {
                "project_key": extract_project_key(issue_id),
                "worklog_date": worklog_date,
                "period_day": worklog_date,
                "period_week": iso_week_code(worklog_date),
                "period_month": month_code(worklog_date),
                "issue_assignee": issue_assignee,
                "worklog_author": worklog_author,
                "hours_logged": round(hours_logged, 2),
            }
        )

    wb.close()
    return rows


def period_sort_key(period: str, granularity: str):
    text = _to_text(period)
    if granularity == "day":
        try:
            return (date.fromisoformat(text),)
        except ValueError:
            return (date.max, text)
    if granularity == "week":
        match = re.match(r"^(\d{4})-W(\d{2})$", text)
        if match:
            return (int(match.group(1)), int(match.group(2)))
        return (9999, 99, text)
    if granularity == "month":
        match = re.match(r"^(\d{4})-(\d{2})$", text)
        if match:
            return (int(match.group(1)), int(match.group(2)))
        return (9999, 12, text)
    return (text,)


def day_boundary_flags(iso_day: str) -> dict:
    day = date.fromisoformat(iso_day)
    is_week_end = day.isoweekday() == 7
    is_month_end = (day + timedelta(days=1)).month != day.month
    return {
        "is_week_end": is_week_end,
        "is_month_end": is_month_end,
        "is_both": is_week_end and is_month_end,
    }


def _week_label_from_code(code: str) -> str:
    match = re.match(r"^(\d{4})-W(\d{2})$", _to_text(code))
    if not match:
        return code
    year = int(match.group(1))
    week = int(match.group(2))
    monday = date.fromisocalendar(year, week, 1)
    sunday = monday + timedelta(days=6)
    return f"{code} ({monday:%b %d} - {sunday:%b %d})"


def _period_label(period: str, granularity: str) -> str:
    if granularity == "week":
        return _week_label_from_code(period)
    return period


def _filtered_rows(
    rows: list[dict],
    from_date: str,
    to_date: str,
    selected_projects: set[str] | None = None,
) -> list[dict]:
    from_value = date.fromisoformat(from_date) if from_date else None
    to_value = date.fromisoformat(to_date) if to_date else None
    project_filter = selected_projects or set()
    use_project_filter = bool(project_filter)

    out: list[dict] = []
    for row in rows:
        project_key = _to_text(row.get("project_key")) or "UNKNOWN"
        if use_project_filter and project_key not in project_filter:
            continue

        worklog_date = _to_text(row.get("worklog_date"))
        if not worklog_date:
            continue
        row_date = date.fromisoformat(worklog_date)
        if from_value and row_date < from_value:
            continue
        if to_value and row_date > to_value:
            continue
        out.append(row)
    return out


def _period_value(row: dict, granularity: str) -> str:
    worklog_date = _to_text(row.get("worklog_date"))
    if granularity == "day":
        return _to_text(row.get("period_day")) or worklog_date
    if granularity == "month":
        return _to_text(row.get("period_month")) or month_code(worklog_date)
    return _to_text(row.get("period_week")) or iso_week_code(worklog_date)


def build_crosstab(
    rows: list[dict],
    granularity: str,
    from_date: str,
    to_date: str,
    selected_projects: set[str] | None = None,
) -> dict:
    granularity_value = (granularity or "week").strip().lower()
    if granularity_value not in {"day", "week", "month"}:
        granularity_value = "week"

    filtered = _filtered_rows(rows, from_date, to_date, selected_projects)
    grid: dict[str, dict[str, float]] = {}
    period_set: set[str] = set()
    for row in filtered:
        period = _period_value(row, granularity_value)
        assignee = _worklog_user(row)
        hours = _to_float(row.get("hours_logged"))
        if not period or hours <= 0:
            continue
        period_set.add(period)
        assignee_row = grid.setdefault(assignee, {})
        assignee_row[period] = assignee_row.get(period, 0.0) + hours

    columns = sorted(period_set, key=lambda value: period_sort_key(value, granularity_value))
    column_totals = {column: 0.0 for column in columns}
    row_items: list[dict] = []
    for assignee, values in grid.items():
        mapped: dict[str, float] = {}
        total_hours = 0.0
        for column in columns:
            value = round(values.get(column, 0.0), 2)
            mapped[column] = value
            total_hours += value
            column_totals[column] = round(column_totals[column] + value, 2)
        row_items.append(
            {
                "assignee": assignee,
                "values": mapped,
                "total_hours": round(total_hours, 2),
            }
        )
    row_items.sort(key=lambda item: item["assignee"].lower())
    row_items.sort(key=lambda item: float(item["total_hours"]), reverse=True)

    overall_total = round(sum(item["total_hours"] for item in row_items), 2)
    meta: list[dict] = []
    for column in columns:
        flags = day_boundary_flags(column) if granularity_value == "day" else {}
        meta.append(
            {
                "period": column,
                "label": _period_label(column, granularity_value),
                "is_week_end": bool(flags.get("is_week_end")),
                "is_month_end": bool(flags.get("is_month_end")),
                "is_both": bool(flags.get("is_both")),
            }
        )

    return {
        "granularity": granularity_value,
        "columns": columns,
        "column_meta": meta,
        "row_items": row_items,
        "grand_totals": {
            "columns": {column: round(value, 2) for column, value in column_totals.items()},
            "overall_total": overall_total,
        },
    }


def _autosize_worksheet(ws) -> None:
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        letter = get_column_letter(col_idx)
        for row_idx in range(1, ws.max_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            length = len(str(value)) if value is not None else 0
            if length > max_len:
                max_len = length
        ws.column_dimensions[letter].width = min(max(10, max_len + 2), 48)


def _write_crosstab_sheet(wb: Workbook, title: str, crosstab: dict, highlight_day_boundaries: bool) -> None:
    ws = wb.create_sheet(title=title)

    header_fill = PatternFill(fill_type="solid", fgColor="0F4C5C")
    header_font = Font(color="FFFFFF", bold=True)
    total_fill = PatternFill(fill_type="solid", fgColor="E6EEF5")
    total_font = Font(bold=True)
    week_end_fill = PatternFill(fill_type="solid", fgColor="DBEAFE")
    month_end_fill = PatternFill(fill_type="solid", fgColor="FDE68A")
    both_fill = PatternFill(fill_type="solid", fgColor="FBCFE8")
    boundary_font = Font(color="1F2937", bold=True)
    center = Alignment(horizontal="center", vertical="center")

    meta = crosstab["column_meta"]
    row_items = crosstab["row_items"]
    totals = crosstab["grand_totals"]

    ws.cell(row=1, column=1, value="User")
    for idx, info in enumerate(meta, start=2):
        ws.cell(row=1, column=idx, value=info["label"])
    total_col = len(meta) + 2
    ws.cell(row=1, column=total_col, value="Total")

    for col in range(1, total_col + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    if highlight_day_boundaries:
        for idx, info in enumerate(meta, start=2):
            cell = ws.cell(row=1, column=idx)
            if info["is_both"]:
                cell.fill = both_fill
                cell.font = boundary_font
            elif info["is_month_end"]:
                cell.fill = month_end_fill
                cell.font = boundary_font
            elif info["is_week_end"]:
                cell.fill = week_end_fill
                cell.font = boundary_font

    for row_idx, item in enumerate(row_items, start=2):
        ws.cell(row=row_idx, column=1, value=item["assignee"])
        for col_idx, info in enumerate(meta, start=2):
            value = _to_float(item["values"].get(info["period"], 0))
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.number_format = "0.00"
        total_cell = ws.cell(row=row_idx, column=total_col, value=_to_float(item["total_hours"]))
        total_cell.number_format = "0.00"
        total_cell.fill = total_fill
        total_cell.font = total_font

    total_row = len(row_items) + 2
    ws.cell(row=total_row, column=1, value="Grand Total")
    for col_idx, info in enumerate(meta, start=2):
        total_value = _to_float(totals["columns"].get(info["period"], 0))
        cell = ws.cell(row=total_row, column=col_idx, value=total_value)
        cell.number_format = "0.00"
    overall_cell = ws.cell(row=total_row, column=total_col, value=_to_float(totals["overall_total"]))
    overall_cell.number_format = "0.00"

    for col in range(1, total_col + 1):
        cell = ws.cell(row=total_row, column=col)
        cell.fill = total_fill
        cell.font = total_font

    ws.freeze_panes = "B2"
    _autosize_worksheet(ws)


def _write_summary_xlsx(rows: list[dict], output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = SUMMARY_SHEET
    ws.append(SUMMARY_HEADERS)
    for row in rows:
        ws.append([row.get(header, "") for header in SUMMARY_HEADERS])

    default_from, default_to = _default_range(rows)
    day_crosstab = build_crosstab(rows, "day", default_from, default_to)
    week_crosstab = build_crosstab(rows, "week", default_from, default_to)
    month_crosstab = build_crosstab(rows, "month", default_from, default_to)
    _write_crosstab_sheet(wb, DAY_CROSSTAB_SHEET, day_crosstab, highlight_day_boundaries=True)
    _write_crosstab_sheet(wb, WEEK_CROSSTAB_SHEET, week_crosstab, highlight_day_boundaries=False)
    _write_crosstab_sheet(wb, MONTH_CROSSTAB_SHEET, month_crosstab, highlight_day_boundaries=False)

    wb.save(output_path)


def _read_summary_xlsx(input_path: Path) -> list[dict]:
    if not input_path.exists():
        raise FileNotFoundError(f"Summary workbook not found: {input_path}")

    wb = load_workbook(input_path, read_only=True, data_only=True)
    ws = wb[SUMMARY_SHEET] if SUMMARY_SHEET in wb.sheetnames else wb.active
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header:
        wb.close()
        return []
    headers = [_to_text(h) for h in header]
    required_headers = {"project_key", "worklog_date", "period_day", "period_week", "period_month", "hours_logged"}
    expected_missing = [h for h in required_headers if h not in headers]
    if expected_missing:
        wb.close()
        raise ValueError(f"Summary workbook missing expected columns: {expected_missing}")

    index = {name: headers.index(name) for name in headers}
    rows: list[dict] = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        item = {name: raw[idx] for name, idx in index.items()}
        worklog_date = _to_text(item.get("worklog_date"))
        assignee = _worklog_user(item)
        hours = _to_float(item.get("hours_logged"))
        if not worklog_date or hours <= 0:
            continue
        rows.append(
            {
                "project_key": _to_text(item.get("project_key")) or "UNKNOWN",
                "worklog_date": worklog_date,
                "period_day": _to_text(item.get("period_day")) or worklog_date,
                "period_week": _to_text(item.get("period_week")) or iso_week_code(worklog_date),
                "period_month": _to_text(item.get("period_month")) or month_code(worklog_date),
                "issue_assignee": _to_text(item.get("issue_assignee")) or assignee,
                "worklog_author": _to_text(item.get("worklog_author")) or assignee,
                "hours_logged": round(hours, 2),
            }
        )
    wb.close()
    return rows


def _default_range(rows: list[dict]) -> tuple[str, str]:
    now = datetime.now(timezone.utc).date()
    current_start = date(now.year, now.month, 1)
    if now.month == 12:
        next_month_start = date(now.year + 1, 1, 1)
    else:
        next_month_start = date(now.year, now.month + 1, 1)
    current_end = next_month_start - timedelta(days=1)

    if now.month == 1:
        prev_start = date(now.year - 1, 12, 1)
    else:
        prev_start = date(now.year, now.month - 1, 1)

    if rows:
        min_row_date = min(date.fromisoformat(r["worklog_date"]) for r in rows)
        max_row_date = max(date.fromisoformat(r["worklog_date"]) for r in rows)
        start = max(prev_start, min_row_date)
        end = min(current_end, max_row_date)
        if end < start:
            start, end = min_row_date, max_row_date
        return start.isoformat(), end.isoformat()

    return prev_start.isoformat(), current_end.isoformat()


def _parse_iso_date(value: str, field_name: str) -> date:
    text = _to_text(value)
    if not text:
        raise ValueError(f"Missing required field: {field_name}")
    try:
        return date.fromisoformat(text)
    except ValueError as exc:
        raise ValueError(f"Invalid ISO date for {field_name}: {value}") from exc


def _default_capacity_settings(from_date: str, to_date: str) -> dict:
    return {
        "from_date": from_date,
        "to_date": to_date,
        "employee_count": 0,
        "standard_hours_per_day": 8.0,
        "ramadan_start_date": "",
        "ramadan_end_date": "",
        "ramadan_hours_per_day": 6.5,
        "holiday_dates": [],
    }


def _normalize_capacity_payload(payload: dict, require_all_fields: bool = True) -> dict:
    from_date = _to_text(payload.get("from_date"))
    to_date = _to_text(payload.get("to_date"))
    if require_all_fields and (not from_date or not to_date):
        raise ValueError("Both from_date and to_date are required.")

    from_value = _parse_iso_date(from_date, "from_date")
    to_value = _parse_iso_date(to_date, "to_date")
    if to_value < from_value:
        raise ValueError("to_date must be on or after from_date.")

    employee_count = int(payload.get("employee_count", 0))
    if employee_count < 0:
        raise ValueError("employee_count must be >= 0.")

    standard_hours_per_day = float(payload.get("standard_hours_per_day", 8.0))
    ramadan_hours_per_day = float(payload.get("ramadan_hours_per_day", 6.5))
    if standard_hours_per_day <= 0:
        raise ValueError("standard_hours_per_day must be > 0.")
    if ramadan_hours_per_day <= 0:
        raise ValueError("ramadan_hours_per_day must be > 0.")

    ramadan_start_date = _to_text(payload.get("ramadan_start_date"))
    ramadan_end_date = _to_text(payload.get("ramadan_end_date"))
    if bool(ramadan_start_date) != bool(ramadan_end_date):
        raise ValueError("ramadan_start_date and ramadan_end_date must both be set or both be empty.")

    ramadan_start = None
    ramadan_end = None
    if ramadan_start_date and ramadan_end_date:
        ramadan_start = _parse_iso_date(ramadan_start_date, "ramadan_start_date")
        ramadan_end = _parse_iso_date(ramadan_end_date, "ramadan_end_date")
        if ramadan_end < ramadan_start:
            raise ValueError("ramadan_end_date must be on or after ramadan_start_date.")
        overlap_start = max(from_value, ramadan_start)
        overlap_end = min(to_value, ramadan_end)
        if overlap_end < overlap_start:
            raise ValueError("Ramadan date range must overlap the selected report range.")

    holiday_values = payload.get("holiday_dates", [])
    if isinstance(holiday_values, str):
        holiday_values = [h.strip() for h in holiday_values.split(",") if h.strip()]
    if not isinstance(holiday_values, list):
        raise ValueError("holiday_dates must be a list of ISO date strings.")
    holiday_dates: list[str] = []
    for raw in holiday_values:
        parsed = _parse_iso_date(_to_text(raw), "holiday_dates[]")
        if from_value <= parsed <= to_value:
            holiday_dates.append(parsed.isoformat())
    holiday_dates = sorted(set(holiday_dates))

    return {
        "from_date": from_value.isoformat(),
        "to_date": to_value.isoformat(),
        "employee_count": employee_count,
        "standard_hours_per_day": round(standard_hours_per_day, 2),
        "ramadan_start_date": ramadan_start.isoformat() if ramadan_start else "",
        "ramadan_end_date": ramadan_end.isoformat() if ramadan_end else "",
        "ramadan_hours_per_day": round(ramadan_hours_per_day, 2),
        "holiday_dates": holiday_dates,
    }


def calculate_capacity_metrics(settings: dict) -> dict:
    normalized = _normalize_capacity_payload(settings, require_all_fields=True)
    from_value = date.fromisoformat(normalized["from_date"])
    to_value = date.fromisoformat(normalized["to_date"])
    holiday_set = {date.fromisoformat(item) for item in normalized["holiday_dates"]}

    ramadan_start = date.fromisoformat(normalized["ramadan_start_date"]) if normalized["ramadan_start_date"] else None
    ramadan_end = date.fromisoformat(normalized["ramadan_end_date"]) if normalized["ramadan_end_date"] else None

    total_weekdays = 0
    holiday_weekdays = 0
    ramadan_weekdays = 0
    non_ramadan_weekdays = 0

    cursor = from_value
    while cursor <= to_value:
        if cursor.weekday() < 5:
            total_weekdays += 1
            if cursor in holiday_set:
                holiday_weekdays += 1
            else:
                in_ramadan = bool(ramadan_start and ramadan_end and ramadan_start <= cursor <= ramadan_end)
                if in_ramadan:
                    ramadan_weekdays += 1
                else:
                    non_ramadan_weekdays += 1
        cursor += timedelta(days=1)

    available_capacity_hours = normalized["employee_count"] * (
        non_ramadan_weekdays * normalized["standard_hours_per_day"]
        + ramadan_weekdays * normalized["ramadan_hours_per_day"]
    )

    return {
        "settings": normalized,
        "metrics": {
            "total_weekdays": total_weekdays,
            "holiday_weekdays": holiday_weekdays,
            "ramadan_weekdays": ramadan_weekdays,
            "non_ramadan_weekdays": non_ramadan_weekdays,
            "available_capacity_hours": round(available_capacity_hours, 2),
        },
    }


def _is_ramadan_day(day_value: date, settings: dict) -> bool:
    start_text = _to_text(settings.get("ramadan_start_date"))
    end_text = _to_text(settings.get("ramadan_end_date"))
    if not start_text or not end_text:
        return False
    try:
        start = date.fromisoformat(start_text)
        end = date.fromisoformat(end_text)
    except ValueError:
        return False
    return start <= day_value <= end


def _hours_per_day_for_date(day_value: date, settings: dict) -> float:
    if _is_ramadan_day(day_value, settings):
        return float(settings.get("ramadan_hours_per_day", 6.5) or 6.5)
    return float(settings.get("standard_hours_per_day", 8.0) or 8.0)


def _daily_hours_to_days(daily_hours_by_day: dict[str, float], settings: dict) -> float:
    total_days = 0.0
    for iso_day, hours in daily_hours_by_day.items():
        try:
            day_value = date.fromisoformat(_to_text(iso_day))
        except ValueError:
            continue
        day_hours = _hours_per_day_for_date(day_value, settings)
        if day_hours <= 0:
            continue
        total_days += float(hours or 0.0) / day_hours
    return round(total_days, 2)


def _hours_to_days_over_range(total_hours: float, settings: dict) -> float:
    try:
        from_day = date.fromisoformat(_to_text(settings.get("from_date")))
        to_day = date.fromisoformat(_to_text(settings.get("to_date")))
    except ValueError:
        divisor = float(settings.get("standard_hours_per_day", 8.0) or 8.0)
        return round((float(total_hours or 0.0) / divisor) if divisor > 0 else 0.0, 2)

    holiday_set = set()
    for value in settings.get("holiday_dates", []):
        try:
            holiday_set.add(date.fromisoformat(_to_text(value)))
        except ValueError:
            continue

    workday_hours: list[float] = []
    cursor = from_day
    while cursor <= to_day:
        if cursor.weekday() < 5 and cursor not in holiday_set:
            day_hours = _hours_per_day_for_date(cursor, settings)
            if day_hours > 0:
                workday_hours.append(day_hours)
        cursor += timedelta(days=1)

    if not workday_hours:
        divisor = float(settings.get("standard_hours_per_day", 8.0) or 8.0)
        return round((float(total_hours or 0.0) / divisor) if divisor > 0 else 0.0, 2)

    effective_hours_per_day = sum(workday_hours) / len(workday_hours)
    return round(float(total_hours or 0.0) / effective_hours_per_day, 2) if effective_hours_per_day > 0 else 0.0


def _load_leave_metrics(leave_report_path: Path, from_date: str, to_date: str, settings: dict) -> dict:
    from_value = _parse_iso_date(from_date, "from_date")
    to_value = _parse_iso_date(to_date, "to_date")
    out = {
        "planned_taken_hours": 0.0,
        "unplanned_taken_hours": 0.0,
        "planned_not_taken_hours": 0.0,
        "taken_hours": 0.0,
        "not_yet_taken_hours": 0.0,
        "taken_days": 0.0,
        "not_yet_taken_days": 0.0,
        "source": "unavailable",
    }
    if not leave_report_path.exists():
        return out

    wb = load_workbook(leave_report_path, read_only=True, data_only=True)
    try:
        if "Daily_Assignee" not in wb.sheetnames:
            out["source"] = "missing_daily_sheet"
            return out
        ws = wb["Daily_Assignee"]
        header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header:
            out["source"] = "empty_daily_sheet"
            return out
        headers = [_to_text(h) for h in header]
        required = {"period_day", "planned_taken_hours", "unplanned_taken_hours", "planned_not_taken_hours"}
        if not required.issubset(set(headers)):
            out["source"] = "missing_required_columns"
            return out
        idx = {name: headers.index(name) for name in required}
        planned_taken_daily: dict[str, float] = {}
        unplanned_taken_daily: dict[str, float] = {}
        planned_not_taken_daily: dict[str, float] = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            iso_day = _to_text(row[idx["period_day"]])
            if not iso_day:
                continue
            try:
                day = date.fromisoformat(iso_day)
            except ValueError:
                continue
            if day < from_value or day > to_value:
                continue
            planned_taken_daily[iso_day] = planned_taken_daily.get(iso_day, 0.0) + _to_float(row[idx["planned_taken_hours"]])
            unplanned_taken_daily[iso_day] = unplanned_taken_daily.get(iso_day, 0.0) + _to_float(row[idx["unplanned_taken_hours"]])
            planned_not_taken_daily[iso_day] = planned_not_taken_daily.get(iso_day, 0.0) + _to_float(
                row[idx["planned_not_taken_hours"]]
            )

        planned_taken = round(sum(planned_taken_daily.values()), 2)
        unplanned_taken = round(sum(unplanned_taken_daily.values()), 2)
        planned_not_taken = round(sum(planned_not_taken_daily.values()), 2)
        taken_daily = dict(planned_taken_daily)
        for iso_day, hours in unplanned_taken_daily.items():
            taken_daily[iso_day] = taken_daily.get(iso_day, 0.0) + hours

        taken_hours = round(planned_taken + unplanned_taken, 2)
        not_yet_hours = round(planned_not_taken, 2)
        out.update(
            {
                "planned_taken_hours": planned_taken,
                "unplanned_taken_hours": unplanned_taken,
                "planned_not_taken_hours": planned_not_taken,
                "taken_hours": taken_hours,
                "not_yet_taken_hours": not_yet_hours,
                "taken_days": _daily_hours_to_days(taken_daily, settings),
                "not_yet_taken_days": _daily_hours_to_days(planned_not_taken_daily, settings),
                "source": "rlt_leave_report.xlsx",
            }
        )
        return out
    finally:
        wb.close()


def _init_capacity_db(db_path: Path) -> None:
    conn = sqlite3.connect(db_path)
    try:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS assignee_capacity_settings (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                from_date TEXT NOT NULL,
                to_date TEXT NOT NULL,
                employee_count INTEGER NOT NULL,
                standard_hours_per_day REAL NOT NULL,
                ramadan_start_date TEXT,
                ramadan_end_date TEXT,
                ramadan_hours_per_day REAL NOT NULL,
                holiday_dates_json TEXT NOT NULL,
                created_at_utc TEXT NOT NULL,
                updated_at_utc TEXT NOT NULL,
                UNIQUE(from_date, to_date)
            )
            """
        )
        conn.commit()
    finally:
        conn.close()


def _load_capacity_settings(db_path: Path, from_date: str, to_date: str) -> dict:
    normalized_range = _normalize_capacity_payload(
        {"from_date": from_date, "to_date": to_date, "employee_count": 0, "holiday_dates": []}
    )
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    try:
        row = conn.execute(
            """
            SELECT from_date, to_date, employee_count, standard_hours_per_day,
                   ramadan_start_date, ramadan_end_date, ramadan_hours_per_day, holiday_dates_json
            FROM assignee_capacity_settings
            WHERE from_date = ? AND to_date = ?
            """,
            (normalized_range["from_date"], normalized_range["to_date"]),
        ).fetchone()
    finally:
        conn.close()

    if not row:
        return _default_capacity_settings(normalized_range["from_date"], normalized_range["to_date"])

    holiday_dates = []
    try:
        decoded = json.loads(row["holiday_dates_json"] or "[]")
        if isinstance(decoded, list):
            holiday_dates = [_to_text(item) for item in decoded if _to_text(item)]
    except json.JSONDecodeError:
        holiday_dates = []

    return _normalize_capacity_payload(
        {
            "from_date": row["from_date"],
            "to_date": row["to_date"],
            "employee_count": row["employee_count"],
            "standard_hours_per_day": row["standard_hours_per_day"],
            "ramadan_start_date": row["ramadan_start_date"] or "",
            "ramadan_end_date": row["ramadan_end_date"] or "",
            "ramadan_hours_per_day": row["ramadan_hours_per_day"],
            "holiday_dates": holiday_dates,
        }
    )


def _save_capacity_settings(db_path: Path, payload: dict) -> dict:
    normalized = _normalize_capacity_payload(payload, require_all_fields=True)
    now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
    conn = sqlite3.connect(db_path)
    try:
        conn.execute(
            """
            INSERT INTO assignee_capacity_settings (
                from_date, to_date, employee_count, standard_hours_per_day,
                ramadan_start_date, ramadan_end_date, ramadan_hours_per_day, holiday_dates_json,
                created_at_utc, updated_at_utc
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(from_date, to_date) DO UPDATE SET
                employee_count=excluded.employee_count,
                standard_hours_per_day=excluded.standard_hours_per_day,
                ramadan_start_date=excluded.ramadan_start_date,
                ramadan_end_date=excluded.ramadan_end_date,
                ramadan_hours_per_day=excluded.ramadan_hours_per_day,
                holiday_dates_json=excluded.holiday_dates_json,
                updated_at_utc=excluded.updated_at_utc
            """,
            (
                normalized["from_date"],
                normalized["to_date"],
                normalized["employee_count"],
                normalized["standard_hours_per_day"],
                normalized["ramadan_start_date"] or None,
                normalized["ramadan_end_date"] or None,
                normalized["ramadan_hours_per_day"],
                json.dumps(normalized["holiday_dates"]),
                now,
                now,
            ),
        )
        conn.commit()
    finally:
        conn.close()
    return normalized


def _delete_capacity_settings(db_path: Path, from_date: str, to_date: str) -> bool:
    normalized_range = _normalize_capacity_payload(
        {"from_date": from_date, "to_date": to_date, "employee_count": 0, "holiday_dates": []}
    )
    conn = sqlite3.connect(db_path)
    try:
        cursor = conn.execute(
            "DELETE FROM assignee_capacity_settings WHERE from_date = ? AND to_date = ?",
            (normalized_range["from_date"], normalized_range["to_date"]),
        )
        conn.commit()
        return bool(cursor.rowcount and cursor.rowcount > 0)
    finally:
        conn.close()


def _list_capacity_profiles(db_path: Path) -> list[dict]:
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    try:
        rows = conn.execute(
            """
            SELECT from_date, to_date, employee_count, standard_hours_per_day,
                   ramadan_start_date, ramadan_end_date, ramadan_hours_per_day,
                   holiday_dates_json, updated_at_utc
            FROM assignee_capacity_settings
            ORDER BY updated_at_utc DESC, from_date DESC, to_date DESC
            """
        ).fetchall()
    finally:
        conn.close()
    profiles: list[dict] = []
    for r in rows:
        holidays: list[str] = []
        raw_holidays = _to_text(r["holiday_dates_json"])
        if raw_holidays:
            try:
                decoded = json.loads(raw_holidays)
                if isinstance(decoded, list):
                    holidays = [_to_text(item) for item in decoded if _to_text(item)]
            except json.JSONDecodeError:
                holidays = []
        profiles.append(
            {
                "from_date": _to_text(r["from_date"]),
                "to_date": _to_text(r["to_date"]),
                "employee_count": int(r["employee_count"] or 0),
                "standard_hours_per_day": float(r["standard_hours_per_day"] or 0),
                "ramadan_start_date": _to_text(r["ramadan_start_date"]),
                "ramadan_end_date": _to_text(r["ramadan_end_date"]),
                "ramadan_hours_per_day": float(r["ramadan_hours_per_day"] or 0),
                "holiday_dates": holidays,
                "updated_at_utc": _to_text(r["updated_at_utc"]),
            }
        )
    return profiles


def aggregate_rows(
    rows: list[dict],
    granularity: str,
    from_date: str,
    to_date: str,
    selected_projects: set[str] | None = None,
) -> list[dict]:
    granularity_value = (granularity or "week").strip().lower()
    if granularity_value not in {"day", "week", "month"}:
        granularity_value = "week"

    from_value = date.fromisoformat(from_date) if from_date else None
    to_value = date.fromisoformat(to_date) if to_date else None
    project_filter = selected_projects or set()
    use_project_filter = bool(project_filter)

    grouped: dict[tuple[str, str], float] = {}
    for row in rows:
        project_key = _to_text(row.get("project_key")) or "UNKNOWN"
        if use_project_filter and project_key not in project_filter:
            continue

        worklog_date = _to_text(row.get("worklog_date"))
        if not worklog_date:
            continue
        row_date = date.fromisoformat(worklog_date)
        if from_value and row_date < from_value:
            continue
        if to_value and row_date > to_value:
            continue

        if granularity_value == "day":
            period = _to_text(row.get("period_day")) or worklog_date
        elif granularity_value == "month":
            period = _to_text(row.get("period_month")) or month_code(worklog_date)
        else:
            period = _to_text(row.get("period_week")) or iso_week_code(worklog_date)

        assignee = _worklog_user(row)
        hours = _to_float(row.get("hours_logged"))
        if hours <= 0:
            continue
        key = (period, assignee)
        grouped[key] = grouped.get(key, 0.0) + hours

    out = [
        {"period": period, "assignee": assignee, "total_hours": round(total, 2)}
        for (period, assignee), total in grouped.items()
    ]
    out.sort(key=lambda item: item["assignee"].lower())
    out.sort(key=lambda item: float(item["total_hours"]), reverse=True)
    out.sort(key=lambda item: item["period"], reverse=True)
    return out


def _load_leave_daily_rows(leave_report_path: Path) -> list[dict]:
    if not leave_report_path.exists():
        return []
    wb = load_workbook(leave_report_path, read_only=True, data_only=True)
    try:
        if "Daily_Assignee" not in wb.sheetnames:
            return []
        ws = wb["Daily_Assignee"]
        header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header:
            return []
        headers = [_to_text(h) for h in header]
        required = {"period_day", "planned_taken_hours", "unplanned_taken_hours", "planned_not_taken_hours"}
        if not required.issubset(set(headers)):
            return []
        idx = {name: headers.index(name) for name in required}
        assignee_idx = headers.index("assignee") if "assignee" in headers else -1
        jira_ids_idx = headers.index("jira_task_ids") if "jira_task_ids" in headers else -1
        jira_links_idx = headers.index("jira_task_links") if "jira_task_links" in headers else -1
        out: list[dict] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            period_day = _to_text(row[idx["period_day"]])
            if not period_day:
                continue
            try:
                date.fromisoformat(period_day)
            except ValueError:
                continue
            out.append(
                {
                    "assignee": _to_text(row[assignee_idx]) if assignee_idx >= 0 else "",
                    "period_day": period_day,
                    "planned_taken_hours": round(_to_float(row[idx["planned_taken_hours"]]), 2),
                    "unplanned_taken_hours": round(_to_float(row[idx["unplanned_taken_hours"]]), 2),
                    "planned_not_taken_hours": round(_to_float(row[idx["planned_not_taken_hours"]]), 2),
                    "jira_task_ids": _to_text(row[jira_ids_idx]) if jira_ids_idx >= 0 else "",
                    "jira_task_links": _to_text(row[jira_links_idx]) if jira_links_idx >= 0 else "",
                }
            )
        return out
    finally:
        wb.close()


def _load_leave_subtask_rows(leave_report_path: Path) -> list[dict]:
    if not leave_report_path.exists():
        return []
    wb = load_workbook(leave_report_path, read_only=True, data_only=True)
    try:
        if "Raw_Subtasks" not in wb.sheetnames:
            return []
        ws = wb["Raw_Subtasks"]
        header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header:
            return []
        headers = [_to_text(h) for h in header]
        required = {
            "issue_key",
            "assignee",
            "start_date",
            "due_date",
            "original_estimate_hours",
            "total_worklog_hours",
            "leave_classification",
        }
        if not required.issubset(set(headers)):
            return []
        idx = {name: headers.index(name) for name in required}
        out: list[dict] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            issue_key = _to_text(row[idx["issue_key"]])
            assignee = _to_text(row[idx["assignee"]])
            start_date = _to_text(row[idx["start_date"]])
            due_date = _to_text(row[idx["due_date"]])
            if start_date:
                try:
                    start_date = date.fromisoformat(start_date[:10]).isoformat()
                except ValueError:
                    start_date = ""
            if due_date:
                try:
                    due_date = date.fromisoformat(due_date[:10]).isoformat()
                except ValueError:
                    due_date = ""
            out.append(
                {
                    "issue_key": issue_key,
                    "assignee": assignee,
                    "start_date": start_date,
                    "due_date": due_date,
                    "original_estimate_hours": round(_to_float(row[idx["original_estimate_hours"]]), 2),
                    "total_worklog_hours": round(_to_float(row[idx["total_worklog_hours"]]), 2),
                    "leave_classification": _to_text(row[idx["leave_classification"]]),
                }
            )
        return out
    finally:
        wb.close()


def _load_project_planned_hours_from_work_items(work_items_path: Path) -> float:
    if not work_items_path.exists():
        return 0.0

    wb = load_workbook(work_items_path, read_only=True, data_only=True)
    try:
        ws = wb.active
        header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header:
            return 0.0
        headers = [_to_text(h) for h in header]
        index = {name: idx for idx, name in enumerate(headers)}
        if "project_key" not in index or "original_estimate_hours" not in index:
            return 0.0

        issue_type_col = None
        if "jira_issue_type" in index:
            issue_type_col = "jira_issue_type"
        elif "work_item_type" in index:
            issue_type_col = "work_item_type"
        if not issue_type_col:
            return 0.0

        issue_key_idx = index.get("issue_key")
        seen_issue_keys: set[str] = set()
        total = 0.0

        for row in ws.iter_rows(min_row=2, values_only=True):
            project_key = _to_text(row[index["project_key"]]).upper()
            if not project_key or project_key == "RLT":
                continue

            issue_type = _to_text(row[index[issue_type_col]]).lower()
            if "epic" not in issue_type:
                continue

            if issue_key_idx is not None:
                issue_key = _to_text(row[issue_key_idx]).upper()
                if issue_key and issue_key in seen_issue_keys:
                    continue
                if issue_key:
                    seen_issue_keys.add(issue_key)

            total += _to_float(row[index["original_estimate_hours"]])

        return round(total, 2)
    finally:
        wb.close()


def _load_rlt_leaves_planned_rows_from_work_items(work_items_path: Path) -> list[dict]:
    if not work_items_path.exists():
        return []

    wb = load_workbook(work_items_path, read_only=True, data_only=True)
    try:
        ws = wb.active
        header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header:
            return []
        headers = [_to_text(h) for h in header]
        index = {name: idx for idx, name in enumerate(headers)}
        if "project_key" not in index or "original_estimate_hours" not in index:
            return []

        issue_type_col = None
        if "jira_issue_type" in index:
            issue_type_col = "jira_issue_type"
        elif "work_item_type" in index:
            issue_type_col = "work_item_type"
        if not issue_type_col:
            return []

        issue_key_idx = index.get("issue_key")
        start_idx = index.get("jira_start_date")
        end_idx = index.get("jira_end_date")
        seen_issue_keys: set[str] = set()
        rows: list[dict] = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            project_key = _to_text(row[index["project_key"]]).upper()
            if project_key != "RLT":
                continue

            issue_type = _to_text(row[index[issue_type_col]]).lower()
            if "epic" not in issue_type:
                continue

            issue_key = _to_text(row[issue_key_idx]).upper() if issue_key_idx is not None else ""
            if issue_key:
                if issue_key in seen_issue_keys:
                    continue
                seen_issue_keys.add(issue_key)

            estimate_hours = round(_to_float(row[index["original_estimate_hours"]]), 2)
            if estimate_hours <= 0:
                continue

            rows.append(
                {
                    "issue_key": issue_key,
                    "jira_start_date": _to_text(row[start_idx]) if start_idx is not None else "",
                    "jira_end_date": _to_text(row[end_idx]) if end_idx is not None else "",
                    "original_estimate_hours": estimate_hours,
                }
            )
        return rows
    finally:
        wb.close()


def _normalize_issue_type(value: str) -> str:
    text = _to_text(value).strip().lower()
    if text in {"sub-task", "subtask"}:
        return "subtask"
    if text in {"bug subtask", "bug-subtask", "bugsubtask"}:
        return "bug subtask"
    if text in {"epic", "story"}:
        return text
    return text


def _load_planned_work_items_from_work_items(work_items_path: Path) -> list[dict]:
    if not work_items_path.exists():
        return []
    wb = load_workbook(work_items_path, read_only=True, data_only=True)
    try:
        ws = wb.active
        header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header:
            return []
        headers = [_to_text(h) for h in header]
        index = {name: i for i, name in enumerate(headers)}
        if "issue_key" not in index:
            return []
        issue_type_col = "jira_issue_type" if "jira_issue_type" in index else ("work_item_type" if "work_item_type" in index else "")
        out: list[dict] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            issue_key = _to_text(row[index["issue_key"]]).upper()
            if not issue_key:
                continue
            issue_type = _normalize_issue_type(row[index[issue_type_col]]) if issue_type_col else ""
            if issue_type not in {"epic", "story", "subtask", "bug subtask"}:
                continue
            estimate_hours = round(_to_float(row[index.get("original_estimate_hours", -1)]), 2) if "original_estimate_hours" in index else 0.0
            if estimate_hours <= 0:
                continue
            out.append(
                {
                    "issue_key": issue_key,
                    "project_key": _to_text(row[index.get("project_key", -1)]).upper() if "project_key" in index else _extract_project_key(issue_key),
                    "issue_type": issue_type,
                    "planned_start": _to_text(row[index.get("start_date", -1)]) if "start_date" in index else "",
                    "planned_end": _to_text(row[index.get("end_date", -1)]) if "end_date" in index else "",
                    "original_estimate_hours": estimate_hours,
                }
            )
        return out
    finally:
        wb.close()


def _build_payload(
    rows: list[dict],
    capacity_profiles: list[dict] | None = None,
    leave_daily_rows: list[dict] | None = None,
    leave_subtask_rows: list[dict] | None = None,
    project_planned_hours: float = 0.0,
    rlt_leaves_planned_rows: list[dict] | None = None,
    planned_work_items: list[dict] | None = None,
) -> dict:
    projects = sorted({r.get("project_key", "UNKNOWN") for r in rows})
    default_from, default_to = _default_range(rows)
    return {
        "rows": rows,
        "projects": projects,
        "default_from": default_from,
        "default_to": default_to,
        "capacity_profiles": capacity_profiles or [],
        "leave_daily_rows": leave_daily_rows or [],
        "leave_subtask_rows": leave_subtask_rows or [],
        "project_planned_hours": round(_to_float(project_planned_hours), 2),
        "rlt_leaves_planned_rows": rlt_leaves_planned_rows or [],
        "planned_work_items": planned_work_items or [],
        "generated_at": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC"),
    }


def _build_html(payload: dict) -> str:
    data = json.dumps(payload, ensure_ascii=True)
    return f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Assignee Hours Report</title>
  <link rel="stylesheet" href="https://fonts.googleapis.com/css2?family=Material+Symbols+Outlined:opsz,wght,FILL,GRAD@20..48,500,0,0">
  <style>
    :root {{
      --bg: #eef3f7;
      --panel: #ffffff;
      --line: #d7e0e8;
      --text: #1f2937;
      --muted: #64748b;
      --accent: #0f4c5c;
      --accent2: #1d4ed8;
      --kpi-capacity-bg: #fee2e2;
      --kpi-capacity-line: #fca5a5;
      --kpi-capacity-ink: #7f1d1d;
      --kpi-planned-bg: #dbeafe;
      --kpi-planned-line: #93c5fd;
      --kpi-planned-ink: #1e3a8a;
      --kpi-actual-bg: #f3e8ff;
      --kpi-actual-line: #d8b4fe;
      --kpi-actual-ink: #a855f7;
      --kpi-gap-bg: #e0f2fe;
      --kpi-gap-line: #7dd3fc;
      --kpi-gap-ink: #2563eb;
      --kpi-leaves-bg: #fecaca;
      --kpi-leaves-line: #f87171;
      --kpi-leaves-ink: #b91c1c;
      --kpi-capacity-gap-bg: #fef2f2;
      --kpi-capacity-gap-line: #fecaca;
      --kpi-capacity-gap-ink: #ef4444;
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      font-family: "Segoe UI", Tahoma, Verdana, sans-serif;
      color: var(--text);
      background:
        radial-gradient(900px 250px at 5% -5%, #d8eefa 0%, transparent 60%),
        linear-gradient(180deg, #eff4f8, var(--bg));
    }}
    .page {{ max-width: 1500px; margin: 0 auto; padding: 16px; }}
    .enterprise-header {{
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 12px;
      padding: 14px;
      margin-bottom: 12px;
    }}
    .header-top {{
      display: flex;
      justify-content: space-between;
      gap: 12px;
      flex-wrap: wrap;
      align-items: flex-start;
    }}
    .header-title-wrap {{ min-width: 260px; }}
    .header-actions {{ display: flex; gap: 8px; flex-wrap: wrap; align-items: center; }}
    .status-badge {{
      display: inline-flex;
      align-items: center;
      border-radius: 999px;
      font-size: .78rem;
      font-weight: 700;
      padding: 4px 10px;
      border: 1px solid #bbf7d0;
      background: #f0fdf4;
      color: #166534;
    }}
    .status-badge.warn {{
      border-color: #fde68a;
      background: #fffbeb;
      color: #92400e;
    }}
    .status-badge.risk {{
      border-color: #fecaca;
      background: #fef2f2;
      color: #991b1b;
    }}
    .header-filter-grid {{
      display: grid;
      gap: 10px;
      grid-template-columns: repeat(auto-fit, minmax(240px, 1fr));
      margin-top: 12px;
    }}
    .header-kpi-strip {{
      margin-top: 10px;
      display: grid;
      gap: 10px;
      grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
    }}
    .header-kpi {{
      border: 1px solid #d6e2ee;
      border-radius: 10px;
      padding: 10px;
      background: linear-gradient(180deg, #ffffff, #f8fbff);
    }}
    .header-kpi:nth-child(1) {{
      background: var(--kpi-capacity-bg);
      border-color: var(--kpi-capacity-line);
    }}
    .header-kpi:nth-child(2) {{
      background: var(--kpi-gap-bg);
      border-color: var(--kpi-gap-line);
    }}
    .header-kpi:nth-child(3) {{
      background: var(--kpi-planned-bg);
      border-color: var(--kpi-planned-line);
    }}
    .header-kpi:nth-child(4) {{
      background: var(--kpi-actual-bg);
      border-color: var(--kpi-actual-line);
    }}
    .header-kpi-label {{
      margin: 0 0 4px;
      font-size: .76rem;
      color: #475569;
      text-transform: uppercase;
      font-weight: 700;
      letter-spacing: .03em;
    }}
    .header-kpi-value {{
      margin: 0;
      font-size: 1.2rem;
      font-weight: 800;
      color: #0f172a;
    }}
    .header-kpi:nth-child(1) .header-kpi-value {{ color: var(--kpi-capacity-ink); }}
    .header-kpi:nth-child(2) .header-kpi-value {{ color: var(--kpi-gap-ink); }}
    .header-kpi:nth-child(3) .header-kpi-value {{ color: var(--kpi-planned-ink); }}
    .header-kpi:nth-child(4) .header-kpi-value {{ color: var(--kpi-actual-ink); }}
    .header-context {{
      margin-top: 8px;
      display: flex;
      gap: 8px;
      flex-wrap: wrap;
      align-items: center;
    }}
    .context-note {{
      font-size: .8rem;
      color: #334155;
      border: 1px solid #d6e2ee;
      background: #f8fbfd;
      border-radius: 999px;
      padding: 4px 10px;
      font-weight: 600;
    }}
    .card {{
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 12px;
      padding: 14px;
      margin-bottom: 12px;
    }}
    .title {{ margin: 0 0 6px; font-size: 1.2rem; color: #0b3142; }}
    .meta {{ margin: 0; color: var(--muted); font-size: .9rem; }}
    .controls {{
      display: grid;
      gap: 10px;
      grid-template-columns: repeat(auto-fit, minmax(240px, 1fr));
      margin-top: 12px;
    }}
    .control-block {{
      border: 1px solid var(--line);
      border-radius: 10px;
      padding: 10px;
      background: #f8fbfd;
    }}
    .control-label {{
      display: block;
      font-size: .8rem;
      font-weight: 700;
      color: #334155;
      text-transform: uppercase;
      margin-bottom: 8px;
    }}
    select, input[type="date"] {{
      width: 100%;
      border: 1px solid #b9c7d3;
      border-radius: 8px;
      padding: 7px 9px;
      font-size: .86rem;
      color: #12313f;
      background: #fff;
    }}
    select[multiple] {{ min-height: 120px; }}
    .btn-row {{ display: flex; gap: 8px; flex-wrap: wrap; margin-top: 8px; }}
    .btn {{
      border: 1px solid #255f73;
      background: #0f4c5c;
      color: #fff;
      border-radius: 8px;
      padding: 6px 10px;
      cursor: pointer;
      font-size: .82rem;
      font-weight: 700;
    }}
    .btn.alt {{ background: #eff6ff; border-color: #bfdbfe; color: #1d4ed8; }}
    .summary {{ display: flex; gap: 10px; flex-wrap: wrap; align-items: center; margin: 6px 0 10px; }}
    .chip {{
      display: inline-flex;
      align-items: center;
      border: 1px solid #bfdbfe;
      background: #eff6ff;
      color: #1d4ed8;
      border-radius: 999px;
      padding: 3px 10px;
      font-size: .8rem;
      font-weight: 700;
    }}
    .section-title {{ margin: 0 0 8px; color: #0f4c5c; font-size: 1rem; }}
    .tab-row {{ display: flex; gap: 8px; flex-wrap: wrap; margin-bottom: 10px; }}
    .tab-btn {{
      border: 1px solid #bfdbfe;
      background: #eff6ff;
      color: #1d4ed8;
      border-radius: 8px;
      padding: 6px 10px;
      cursor: pointer;
      font-size: .82rem;
      font-weight: 700;
    }}
    .tab-btn.active {{
      background: #0f4c5c;
      border-color: #0f4c5c;
      color: #fff;
    }}
    .tab-panel {{ display: none; }}
    .tab-panel.active {{ display: block; }}
    table {{ width: max-content; min-width: 100%; border-collapse: collapse; table-layout: auto; }}
    .table-wrap {{ overflow-x: auto; }}
    th {{
      background: var(--accent);
      color: #fff;
      text-align: left;
      padding: 8px 10px;
      font-size: .84rem;
      white-space: nowrap;
    }}
    td {{
      border-top: 1px solid var(--line);
      padding: 8px 10px;
      font-size: .86rem;
      vertical-align: top;
      white-space: nowrap;
      background: #fff;
    }}
    tr:nth-child(even) td {{ background: #fbfdff; }}
    .num {{ text-align: right; white-space: nowrap; }}
    .total-col {{ background: #eef5fc !important; font-weight: 700; }}
    .grand-total td {{ background: #e6eef5 !important; font-weight: 700; }}
    .day-week-end {{ background: #1d4ed8; }}
    .day-month-end {{ background: #b45309; }}
    .day-week-month-end {{ background: #7e22ce; }}
    .empty {{ color: var(--muted); font-style: italic; }}
    .kpi-grid {{
      display: grid;
      grid-template-columns: repeat(auto-fit, minmax(260px, 1fr));
      gap: 10px;
      margin: 10px 0;
    }}
    .kpi-trio {{
      grid-column: 1 / -1;
      display: grid;
      grid-template-columns: repeat(3, minmax(220px, 1fr));
      gap: 10px;
    }}
    @media (max-width: 980px) {{
      .kpi-trio {{
        grid-template-columns: 1fr;
      }}
    }}
    .kpi {{
      border: 1px solid #d6e2ee;
      border-radius: 10px;
      padding: 10px;
      background: linear-gradient(180deg, #ffffff, #f8fbff);
    }}
    .kpi .label {{
      font-size: .78rem;
      text-transform: uppercase;
      color: #4b5563;
      letter-spacing: .03em;
      font-weight: 700;
      display: inline-flex;
      align-items: center;
      gap: 6px;
    }}
    .kpi-info {{
      position: relative;
      display: inline-flex;
      align-items: center;
      justify-content: center;
      width: 15px;
      height: 15px;
      border-radius: 999px;
      border: 1px solid #94a3b8;
      color: #334155;
      background: #f8fafc;
      font-size: 10px;
      line-height: 1;
      font-weight: 700;
      text-transform: none;
      cursor: default;
      user-select: none;
    }}
    .kpi-info-tip {{
      position: absolute;
      left: 50%;
      top: calc(100% + 8px);
      transform: translateX(-50%);
      min-width: 230px;
      max-width: 320px;
      padding: 8px 10px;
      border-radius: 8px;
      border: 1px solid #cbd5e1;
      background: #fff;
      color: #1f2937;
      font-size: 12px;
      line-height: 1.35;
      letter-spacing: normal;
      text-transform: none;
      box-shadow: 0 10px 24px rgba(15, 23, 42, 0.18);
      opacity: 0;
      visibility: hidden;
      pointer-events: none;
      z-index: 25;
      transition: opacity .14s ease;
    }}
    .kpi-info:hover .kpi-info-tip,
    .kpi-info:focus .kpi-info-tip,
    .kpi-info:focus-visible .kpi-info-tip {{
      opacity: 1;
      visibility: visible;
    }}
    .kpi .value {{
      font-size: 1.35rem;
      font-weight: 800;
      margin-top: 4px;
      color: #0f172a;
    }}
    .kpi.blue {{
      background: var(--kpi-capacity-bg);
      border-color: var(--kpi-capacity-line);
    }}
    .kpi.blue .value {{ color: var(--kpi-capacity-ink); }}
    .kpi.warn {{
      background: var(--kpi-leaves-bg);
      border-color: var(--kpi-leaves-line);
    }}
    .kpi.warn .value {{ color: var(--kpi-leaves-ink); }}
    .kpi.good {{
      background: var(--kpi-gap-bg);
      border-color: var(--kpi-gap-line);
    }}
    .kpi.good .value {{ color: var(--kpi-gap-ink); }}
    .kpi.bad {{
      background: var(--kpi-capacity-gap-bg);
      border-color: var(--kpi-capacity-gap-line);
    }}
    .kpi.bad .value {{ color: var(--kpi-capacity-gap-ink); }}
    .kpi-secondary {{ margin-top: 8px; display: none !important; }}
    .kpi-secondary.visible {{ display: none !important; }}
    .math-card .label {{ margin-bottom: 8px; }}
    .math-line {{
      display: grid;
      grid-template-columns: 22px 1fr auto;
      align-items: center;
      gap: 8px;
      margin: 4px 0;
      font-size: .9rem;
    }}
    .math-line .op {{
      color: #334155;
      font-weight: 800;
      text-align: center;
      font-family: "Consolas", "Courier New", monospace;
    }}
    .math-line .name {{ color: #334155; font-weight: 700; }}
    .math-line .num {{
      color: #0f172a;
      font-weight: 800;
      font-family: "Consolas", "Courier New", monospace;
    }}
    .math-line.result {{
      border-top: 2px solid #d6e2ee;
      padding-top: 6px;
      margin-top: 6px;
    }}
    .capacity-row {{
      display: grid;
      gap: 10px;
      grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
      margin-top: 10px;
    }}
    .muted-note {{ color: #64748b; font-size: .82rem; margin-top: 6px; }}
    #capacity-status {{ min-height: 18px; font-size: .82rem; margin-top: 6px; color: #334155; }}
    .assumption {{
      margin-top: 8px;
      font-size: .86rem;
      color: #334155;
      font-weight: 700;
    }}
    .holiday-tools {{ display: flex; gap: 8px; margin-bottom: 8px; }}
    .holiday-list {{
      display: flex;
      gap: 6px;
      flex-wrap: wrap;
      min-height: 30px;
      padding: 6px;
      border: 1px solid #d7e0e8;
      border-radius: 8px;
      background: #f8fbfd;
    }}
    .holiday-chip {{
      display: inline-flex;
      align-items: center;
      gap: 6px;
      border: 1px solid #bfdbfe;
      background: #eff6ff;
      color: #1d4ed8;
      border-radius: 999px;
      padding: 2px 8px;
      font-size: .78rem;
      font-weight: 700;
    }}
    .holiday-chip button {{
      border: none;
      background: transparent;
      color: #1e40af;
      cursor: pointer;
      font-size: .78rem;
      line-height: 1;
      padding: 0;
    }}
    .drawer-overlay {{
      position: fixed;
      inset: 0;
      background: rgba(15, 23, 42, .42);
      z-index: 80;
      opacity: 0;
      visibility: hidden;
      transition: opacity .18s ease;
    }}
    .drawer-overlay.open {{
      opacity: 1;
      visibility: visible;
    }}
    .settings-drawer {{
      position: fixed;
      right: 0;
      top: 0;
      height: 100vh;
      width: min(460px, 94vw);
      background: #fff;
      border-left: 1px solid #d6e2ee;
      box-shadow: -10px 0 30px rgba(15, 23, 42, .2);
      z-index: 90;
      transform: translateX(100%);
      transition: transform .22s ease;
      display: flex;
      flex-direction: column;
    }}
    .settings-drawer.open {{
      transform: translateX(0);
    }}
    .drawer-head {{
      padding: 12px 14px;
      border-bottom: 1px solid #d6e2ee;
      display: flex;
      align-items: center;
      justify-content: space-between;
      gap: 8px;
    }}
    .drawer-title {{
      margin: 0;
      font-size: 1rem;
      color: #0b3142;
      font-weight: 800;
    }}
    .drawer-body {{
      padding: 12px 14px 20px;
      overflow: auto;
      display: grid;
      gap: 12px;
      grid-template-columns: 1fr;
    }}
    .drawer-section {{
      border: 1px solid #d7e0e8;
      border-radius: 10px;
      background: #f8fbfd;
      padding: 10px;
    }}
    .drawer-section-title {{
      margin: 0 0 8px;
      font-size: .78rem;
      text-transform: uppercase;
      color: #355564;
      letter-spacing: .03em;
      font-weight: 800;
    }}
  </style>
  <link rel="stylesheet" href="shared-nav.css">
</head>
<body>
  <div class="page">
    <section class="enterprise-header">
      <div class="header-top">
        <div class="header-title-wrap">
          <h1 class="title">Assignee Hours Report</h1>
          <p class="meta">Generated: <span id="generated-at"></span></p>
        </div>
        <div class="header-actions">
          <button class="btn alt" type="button" id="open-capacity-settings">Open Capacity Settings</button>
          <span class="status-badge" id="capacity-risk-badge">Capacity Healthy</span>
        </div>
      </div>
      <div class="header-filter-grid">
        <div class="control-block">
          <label class="control-label" for="from-date">From</label>
          <input id="from-date" type="date">
          <label class="control-label" for="to-date" style="margin-top:8px;">To</label>
          <input id="to-date" type="date">
          <label class="control-label" for="actual-hours-mode" style="margin-top:8px;">Actual Hours Mode</label>
          <select id="actual-hours-mode">
            <option value="log_date">By Log Date</option>
            <option value="planned_dates">By Planned Dates</option>
          </select>
        </div>
        <div class="control-block">
          <label class="control-label" for="project-select">Projects (Multi-select)</label>
          <select id="project-select" multiple></select>
          <div class="btn-row">
            <button class="btn alt" type="button" id="select-all-projects">Select all</button>
            <button class="btn alt" type="button" id="clear-projects">Clear</button>
          </div>
        </div>
      </div>
      <div class="btn-row">
        <button class="btn" type="button" id="apply">Apply</button>
        <button class="btn alt" type="button" id="reset">Reset</button>
      </div>
      <div class="header-kpi-strip">
        <article class="header-kpi">
          <p class="header-kpi-label">Utilization</p>
          <p class="header-kpi-value" id="header-kpi-utilization">0%</p>
        </article>
        <article class="header-kpi">
          <p class="header-kpi-label">Capacity After Leaves</p>
          <p class="header-kpi-value" id="header-kpi-gap">0h</p>
        </article>
        <article class="header-kpi">
          <p class="header-kpi-label">Plan vs Actual Delta</p>
          <p class="header-kpi-value" id="header-kpi-delta">0h</p>
        </article>
        <article class="header-kpi">
          <p class="header-kpi-label">Project Actual Hours</p>
          <p class="header-kpi-value" id="header-kpi-actual">0h</p>
        </article>
      </div>
      <div class="header-context">
        <span class="context-note" id="capacity-assumption">Employees 0 | Std 8h | Ramadan 6.5h | Range -</span>
        <span class="context-note" id="active-profile-indicator">Profile: Project Defaults</span>
      </div>
    </section>

    <section class="card">
      <h2 class="section-title">Capacity Planning</h2>
      <div class="kpi-grid">
        <div class="kpi" id="kpi-util-card">
          <div class="label">Utilization <span class="kpi-info" tabindex="0">i<span class="kpi-info-tip">Project Actual Hours divided by Total Capacity. Project Actual Hours excludes RLT project leave logs.</span></span></div>
          <div class="value" id="kpi-utilization">0%</div>
        </div>
        <div class="kpi math-card" id="kpi-gap-card">
          <div class="label">Capacity Subtraction (Hours) <span class="kpi-info" tabindex="0">i<span class="kpi-info-tip">Formula: Available Capacity - Project Actual Hours - Leave Hours.</span></span></div>
          <div class="math-line"><span class="op"> </span><span class="name">Available Capacity</span><span class="num" id="kpi-capacity">0h</span></div>
          <div class="math-line"><span class="op">-</span><span class="name">Project Actual Hours</span><span class="num" id="kpi-logged">0h</span></div>
          <div class="math-line"><span class="op">-</span><span class="name">Leave Hours</span><span class="num" id="kpi-leave-total-ref">0h</span></div>
          <div class="math-line result"><span class="op">=</span><span class="name">Capacity After Leaves</span><span class="num" id="kpi-gap">0h</span></div>
        </div>
        <div class="kpi math-card">
          <div class="label">Leave Hours <span class="kpi-info" tabindex="0">i<span class="kpi-info-tip">Total Leaves = Planned Leaves Taken + Planned Leaves Not Taken Yet + Unplanned Leaves Taken.</span></span></div>
          <div class="math-line"><span class="op"> </span><span class="name">Planned Leaves Taken</span><span class="num" id="kpi-planned-leaves-taken">0h</span></div>
          <div class="math-line"><span class="op">+</span><span class="name">Planned Leaves Not Taken Yet</span><span class="num" id="kpi-planned-leaves-not-yet">0h</span></div>
          <div class="math-line"><span class="op">+</span><span class="name">Unplanned Leaves Taken</span><span class="num" id="kpi-unplanned-leaves-taken">0h</span></div>
          <div class="math-line result"><span class="op">=</span><span class="name">Total Leaves</span><span class="num" id="kpi-leave-total">0h</span></div>
        </div>
      </div>
      <div class="kpi-grid kpi-secondary" id="capacity-summary-cards">
        <div class="kpi blue">
          <div class="label">Total Capacity <span class="kpi-info" tabindex="0">i<span class="kpi-info-tip">Raw capacity from selected capacity profile/settings. This card does not subtract leave hours.</span></span></div>
          <div class="value" id="summary-total-capacity">0h</div>
        </div>
        <div class="kpi">
          <div class="label">Leave Hours <span class="kpi-info" tabindex="0">i<span class="kpi-info-tip">All leave impact hours within selected date range.</span></span></div>
          <div class="value" id="summary-leave-hours">0h</div>
        </div>
        <div class="kpi">
          <div class="label">Remaining Capacity <span class="kpi-info" tabindex="0">i<span class="kpi-info-tip">Formula: Total Capacity - Leave Hours.</span></span></div>
          <div class="value" id="summary-remaining-capacity">0h</div>
        </div>
        <div class="kpi">
          <div class="label">Total Employees <span class="kpi-info" tabindex="0">i<span class="kpi-info-tip">Employee count from capacity form/profile used for capacity computation.</span></span></div>
          <div class="value" id="summary-total-employees">0</div>
        </div>
        <div class="kpi-trio">
          <div class="kpi good">
            <div class="label">Project Planned Hours <span class="kpi-info" tabindex="0">i<span class="kpi-info-tip">Sum of Original Estimates for Epic/Story/Subtask items where planned start OR end is in the selected range, excluding RLT (RnD Leave Tracker).</span></span></div>
            <div class="value" id="summary-planned-hours">0h</div>
          </div>
          <div class="kpi">
            <div class="label">Project Actual Hours <span class="kpi-info" tabindex="0">i<span class="kpi-info-tip">Total logged project hours for selected filters excluding project key RLT.</span></span></div>
            <div class="value" id="summary-actual-hours">0h</div>
          </div>
          <div class="kpi" id="summary-delta-card">
            <div class="label">Project Plan - Actual Hours <span class="kpi-info" tabindex="0">i<span class="kpi-info-tip">Formula: Project Planned Hours - Project Actual Hours.</span></span></div>
            <div class="value" id="summary-delta-hours">0h</div>
          </div>
        </div>
      </div>
      <div id="capacity-status"></div>
    </section>

    <section class="card">
      <div class="summary">
        <span class="chip" id="summary-range"></span>
        <span class="chip" id="summary-projects"></span>
        <span class="chip" id="summary-total"></span>
      </div>
      <div class="tab-row">
        <button type="button" class="tab-btn" data-tab="day">Day</button>
        <button type="button" class="tab-btn" data-tab="week">Week</button>
        <button type="button" class="tab-btn active" data-tab="month">Month</button>
      </div>
      <div id="tab-day" class="tab-panel">
        <h2 class="section-title">Day Crosstab</h2>
        <div class="table-wrap" id="day-table-wrap"></div>
      </div>
      <div id="tab-week" class="tab-panel">
        <h2 class="section-title">Week Crosstab</h2>
        <div class="table-wrap" id="week-table-wrap"></div>
      </div>
      <div id="tab-month" class="tab-panel active">
        <h2 class="section-title">Month Crosstab</h2>
        <div class="table-wrap" id="month-table-wrap"></div>
      </div>
    </section>
  </div>
  <div class="drawer-overlay" id="settings-drawer-overlay"></div>
  <aside
    class="settings-drawer"
    id="settings-drawer"
    role="dialog"
    aria-modal="true"
    aria-labelledby="settings-drawer-title"
    aria-hidden="true"
  >
    <div class="drawer-head">
      <h2 class="drawer-title" id="settings-drawer-title">Capacity Settings</h2>
      <button class="btn alt" type="button" id="close-capacity-settings" aria-label="Close capacity settings">Close</button>
    </div>
    <div class="drawer-body">
      <section class="drawer-section">
        <h3 class="drawer-section-title">Workforce</h3>
        <label class="control-label" for="capacity-employees">Employees</label>
        <input id="capacity-employees" type="number" min="0" step="1" value="0">
        <div class="muted-note">Assignees found in data: <strong id="assignee-count-help">0</strong> <button class="btn alt" type="button" id="use-assignee-count" style="margin-left:6px;padding:3px 8px;font-size:.75rem;">Use</button></div>
        <label class="control-label" for="capacity-standard-hours" style="margin-top:8px;">Standard Hours/Day</label>
        <input id="capacity-standard-hours" type="number" min="0.5" step="0.5" value="8">
      </section>
      <section class="drawer-section">
        <h3 class="drawer-section-title">Ramadan Schedule</h3>
        <label class="control-label" for="ramadan-start">Ramadan Start</label>
        <input id="ramadan-start" type="date">
        <label class="control-label" for="ramadan-end" style="margin-top:8px;">Ramadan End</label>
        <input id="ramadan-end" type="date">
        <label class="control-label" for="ramadan-hours" style="margin-top:8px;">Ramadan Hours/Day</label>
        <input id="ramadan-hours" type="number" min="0.5" step="0.5" value="6.5">
      </section>
      <section class="drawer-section">
        <h3 class="drawer-section-title">Holidays</h3>
        <label class="control-label" for="holiday-date-picker">Holiday Dates</label>
        <div class="holiday-tools">
          <input id="holiday-date-picker" type="date">
          <button class="btn alt" type="button" id="holiday-add">Add</button>
          <button class="btn alt" type="button" id="holiday-clear">Clear</button>
        </div>
        <div id="holiday-list" class="holiday-list"></div>
        <div class="muted-note">Optional non-weekend dates inside selected range.</div>
      </section>
      <section class="drawer-section">
        <h3 class="drawer-section-title">Profiles</h3>
        <label class="control-label" for="capacity-profile-select">Reuse Saved Capacity</label>
        <select id="capacity-profile-select"></select>
        <div class="btn-row">
          <button class="btn alt" type="button" id="capacity-profile-refresh">Refresh Profiles</button>
          <button class="btn alt" type="button" id="capacity-profile-apply">Apply Profile To Current Range</button>
        </div>
      </section>
      <section class="drawer-section">
        <h3 class="drawer-section-title">Actions</h3>
        <div class="btn-row">
          <button class="btn" type="button" id="capacity-save">Save Capacity</button>
          <button class="btn alt" type="button" id="capacity-recalc">Recalculate</button>
        </div>
      </section>
    </div>
  </aside>

  <script>
    const payload = {data};
    const rows = Array.isArray(payload.rows) ? payload.rows : [];
    const projectPlannedHoursAllEpics = Number(payload.project_planned_hours || 0);
    const plannedWorkItems = Array.isArray(payload.planned_work_items) ? payload.planned_work_items : [];
    const projects = Array.isArray(payload.projects) ? payload.projects : [];
    const defaults = {{
      from: payload.default_from || "",
      to: payload.default_to || "",
    }};

    const generatedAtEl = document.getElementById("generated-at");
    const openCapacitySettingsEl = document.getElementById("open-capacity-settings");
    const closeCapacitySettingsEl = document.getElementById("close-capacity-settings");
    const settingsDrawerEl = document.getElementById("settings-drawer");
    const settingsDrawerOverlayEl = document.getElementById("settings-drawer-overlay");
    const capacityRiskBadgeEl = document.getElementById("capacity-risk-badge");
    const activeProfileIndicatorEl = document.getElementById("active-profile-indicator");
    const headerKpiUtilizationEl = document.getElementById("header-kpi-utilization");
    const headerKpiGapEl = document.getElementById("header-kpi-gap");
    const headerKpiDeltaEl = document.getElementById("header-kpi-delta");
    const headerKpiActualEl = document.getElementById("header-kpi-actual");
    const fromDateEl = document.getElementById("from-date");
    const toDateEl = document.getElementById("to-date");
    const actualHoursModeEl = document.getElementById("actual-hours-mode");
    const projectSelectEl = document.getElementById("project-select");
    const applyEl = document.getElementById("apply");
    const resetEl = document.getElementById("reset");
    const selectAllProjectsEl = document.getElementById("select-all-projects");
    const clearProjectsEl = document.getElementById("clear-projects");
    const summaryRangeEl = document.getElementById("summary-range");
    const summaryProjectsEl = document.getElementById("summary-projects");
    const summaryTotalEl = document.getElementById("summary-total");
    const dayTableWrapEl = document.getElementById("day-table-wrap");
    const weekTableWrapEl = document.getElementById("week-table-wrap");
    const monthTableWrapEl = document.getElementById("month-table-wrap");
    const tabButtons = Array.from(document.querySelectorAll(".tab-btn"));
    const tabPanels = {{
      day: document.getElementById("tab-day"),
      week: document.getElementById("tab-week"),
      month: document.getElementById("tab-month"),
    }};
    const capacityEmployeesEl = document.getElementById("capacity-employees");
    const assigneeCountHelpEl = document.getElementById("assignee-count-help");
    const useAssigneeCountEl = document.getElementById("use-assignee-count");
    const capacityStandardHoursEl = document.getElementById("capacity-standard-hours");
    const ramadanStartEl = document.getElementById("ramadan-start");
    const ramadanEndEl = document.getElementById("ramadan-end");
    const ramadanHoursEl = document.getElementById("ramadan-hours");
    const holidayDatePickerEl = document.getElementById("holiday-date-picker");
    const holidayAddEl = document.getElementById("holiday-add");
    const holidayClearEl = document.getElementById("holiday-clear");
    const holidayListEl = document.getElementById("holiday-list");
    const capacityProfileSelectEl = document.getElementById("capacity-profile-select");
    const capacityProfileRefreshEl = document.getElementById("capacity-profile-refresh");
    const capacityProfileApplyEl = document.getElementById("capacity-profile-apply");
    const capacitySaveEl = document.getElementById("capacity-save");
    const capacityRecalcEl = document.getElementById("capacity-recalc");
    const capacityStatusEl = document.getElementById("capacity-status");
    const capacityAssumptionEl = document.getElementById("capacity-assumption");
    const kpiLoggedEl = document.getElementById("kpi-logged");
    const kpiCapacityEl = document.getElementById("kpi-capacity");
    const kpiUtilizationEl = document.getElementById("kpi-utilization");
    const kpiGapEl = document.getElementById("kpi-gap");
    const kpiUtilCardEl = document.getElementById("kpi-util-card");
    const kpiGapCardEl = document.getElementById("kpi-gap-card");
    const kpiPlannedLeavesTakenEl = document.getElementById("kpi-planned-leaves-taken");
    const kpiPlannedLeavesNotYetEl = document.getElementById("kpi-planned-leaves-not-yet");
    const kpiUnplannedLeavesTakenEl = document.getElementById("kpi-unplanned-leaves-taken");
    const kpiLeaveTotalEl = document.getElementById("kpi-leave-total");
    const kpiLeaveTotalRefEl = document.getElementById("kpi-leave-total-ref");
    const capacitySummaryCardsEl = document.getElementById("capacity-summary-cards");
    const summaryTotalCapacityEl = document.getElementById("summary-total-capacity");
    const summaryLeaveHoursEl = document.getElementById("summary-leave-hours");
    const summaryRemainingCapacityEl = document.getElementById("summary-remaining-capacity");
    const summaryTotalEmployeesEl = document.getElementById("summary-total-employees");
    const summaryPlannedHoursEl = document.getElementById("summary-planned-hours");
    const summaryActualHoursEl = document.getElementById("summary-actual-hours");
    const summaryDeltaHoursEl = document.getElementById("summary-delta-hours");
    const summaryDeltaCardEl = document.getElementById("summary-delta-card");
    const embeddedCapacityProfiles = Array.isArray(payload.capacity_profiles) ? payload.capacity_profiles : [];
    const leaveDailyRows = Array.isArray(payload.leave_daily_rows) ? payload.leave_daily_rows : [];
    const leaveSubtaskRows = Array.isArray(payload.leave_subtask_rows) ? payload.leave_subtask_rows : [];
    const rltLeavesPlannedRows = Array.isArray(payload.rlt_leaves_planned_rows) ? payload.rlt_leaves_planned_rows : [];
    const nestedScoreTotalCapacityEl = document.getElementById("nested-score-total-capacity");
    const nestedScoreTotalPlannedEl = document.getElementById("nested-score-total-planned");
    const nestedScoreTotalLoggedEl = document.getElementById("nested-score-total-logged");
    const nestedScoreDeltaEl = document.getElementById("nested-score-delta");
    const nestedScoreTotalLeavesTakenEl = document.getElementById("nested-score-total-leaves-taken");
    const nestedScoreTotalLeavesPlannedEl = document.getElementById("nested-score-total-leaves-planned");
    const nestedScoreTotalCapacityPlannedLeavesAdjustedEl = document.getElementById("nested-score-total-capacity-planned-leaves-adjusted");
    const nestedScoreCapacityGapEl = document.getElementById("nested-score-capacity-gap");
    const nestedScoreTotalCapacityTipEl = document.getElementById("nested-score-total-capacity-tip");
    const nestedScoreTotalPlannedTipEl = document.getElementById("nested-score-total-planned-tip");
    const nestedScoreTotalLoggedTipEl = document.getElementById("nested-score-total-logged-tip");
    const nestedScoreDeltaTipEl = document.getElementById("nested-score-delta-tip");
    const nestedScoreTotalLeavesTakenTipEl = document.getElementById("nested-score-total-leaves-taken-tip");
    const nestedScoreTotalLeavesPlannedTipEl = document.getElementById("nested-score-total-leaves-planned-tip");
    const nestedScoreTotalCapacityPlannedLeavesAdjustedTipEl = document.getElementById("nested-score-total-capacity-planned-leaves-adjusted-tip");
    const nestedScoreCapacityGapTipEl = document.getElementById("nested-score-capacity-gap-tip");
    let backendAvailable = false;
    const ACTUAL_MODE_STORAGE_KEY = "actual-hours-mode:assignee-hours";
    const ACTUAL_MODE_DEFAULT = "log_date";
    const ACTUAL_AGG_ENDPOINT = "/api/actual-hours/aggregate";
    const assigneeCount = new Set(rows.map((r) => r.worklog_author || r.issue_assignee || "Unassigned")).size;
    let capacityProfiles = [];
    let selectedHolidayDates = [];
    let currentProjectActualHours = 0;
    let currentProjectPlannedHours = projectPlannedHoursAllEpics;
    let activeProfileLabel = "Project Defaults";
    let lastFocusedBeforeDrawer = null;

    function fillProjects() {{
      projectSelectEl.innerHTML = projects
        .map((p) => `<option value="${{escapeHtml(p)}}">${{escapeHtml(p)}}</option>`)
        .join("");
    }}

    function selectAllProjects() {{
      Array.from(projectSelectEl.options).forEach((opt) => {{ opt.selected = true; }});
    }}

    function selectedProjects() {{
      const selected = Array.from(projectSelectEl.selectedOptions).map((opt) => opt.value);
      if (!selected.length) {{
        return new Set(projects);
      }}
      return new Set(selected);
    }}

    function updateActiveProfileIndicator() {{
      activeProfileIndicatorEl.textContent = `Profile: ${{activeProfileLabel || "Project Defaults"}}`;
    }}

    function openSettingsDrawer() {{
      lastFocusedBeforeDrawer = document.activeElement;
      settingsDrawerEl.classList.add("open");
      settingsDrawerOverlayEl.classList.add("open");
      settingsDrawerEl.setAttribute("aria-hidden", "false");
      document.body.style.overflow = "hidden";
      closeCapacitySettingsEl.focus();
    }}

    function closeSettingsDrawer() {{
      settingsDrawerEl.classList.remove("open");
      settingsDrawerOverlayEl.classList.remove("open");
      settingsDrawerEl.setAttribute("aria-hidden", "true");
      document.body.style.overflow = "";
      if (lastFocusedBeforeDrawer && typeof lastFocusedBeforeDrawer.focus === "function") {{
        lastFocusedBeforeDrawer.focus();
      }}
    }}

    function trapDrawerFocus(event) {{
      if (event.key !== "Tab" || !settingsDrawerEl.classList.contains("open")) return;
      const focusable = Array.from(settingsDrawerEl.querySelectorAll('button, [href], input, select, textarea, [tabindex]:not([tabindex="-1"])'))
        .filter((el) => !el.disabled && el.offsetParent !== null);
      if (!focusable.length) return;
      const first = focusable[0];
      const last = focusable[focusable.length - 1];
      if (event.shiftKey && document.activeElement === first) {{
        event.preventDefault();
        last.focus();
      }} else if (!event.shiftKey && document.activeElement === last) {{
        event.preventDefault();
        first.focus();
      }}
    }}

    function parseIsoDate(value) {{
      if (!value) return null;
      const d = new Date(`${{value}}T00:00:00`);
      if (Number.isNaN(d.getTime())) return null;
      return d;
    }}

    function isDateWithinBounds(dateValue, from, to) {{
      if (!dateValue || !from || !to) {{
        return false;
      }}
      const value = dateValue.getTime();
      return value >= from.getTime() && value <= to.getTime();
    }}

    function plannedWorkHoursInRange(fromIso, toIso, selected) {{
      const from = parseIsoDate(fromIso);
      const to = parseIsoDate(toIso);
      if (!from || !to || to < from) {{
        return 0;
      }}
      let total = 0;
      for (const row of plannedWorkItems) {{
        const projectKey = String(row && row.project_key || "").trim();
        if (!selected.has(projectKey)) {{
          continue;
        }}
        const plannedStart = parseIsoDate(String(row && row.planned_start || ""));
        const plannedEnd = parseIsoDate(String(row && row.planned_end || ""));
        if (!isDateWithinBounds(plannedStart, from, to) && !isDateWithinBounds(plannedEnd, from, to)) {{
          continue;
        }}
        total += Number(row && row.original_estimate_hours || 0);
      }}
      return Number(total.toFixed(2));
    }}

    function getRltLeavesPlannedInRange(fromIso, toIso) {{
      const from = parseIsoDate(fromIso);
      const to = parseIsoDate(toIso);
      if (!from || !to || to < from) {{
        return {{ hours: 0, issue_count: 0 }};
      }}
      let total = 0;
      let issueCount = 0;
      for (const row of rltLeavesPlannedRows) {{
        const hours = Number(row && row.original_estimate_hours || 0);
        if (!Number.isFinite(hours) || hours <= 0) {{
          continue;
        }}
        const start = parseIsoDate(row && row.jira_start_date);
        const end = parseIsoDate(row && row.jira_end_date);
        let include = true;
        if (start || end) {{
          const rowStart = start || end;
          const rowEnd = end || start;
          include = !!rowStart && !!rowEnd && !(rowEnd < from || rowStart > to);
        }}
        if (!include) {{
          continue;
        }}
        total += hours;
        issueCount += 1;
      }}
      return {{
        hours: Number(total.toFixed(2)),
        issue_count: issueCount,
      }};
    }}

    function getLeaveSubtaskPlannedInRange(fromIso, toIso) {{
      const from = parseIsoDate(fromIso);
      const to = parseIsoDate(toIso);
      if (!from || !to || to < from) {{
        return {{ hours: 0, issue_count: 0, uses_leave_subtasks: false }};
      }}
      if (!leaveSubtaskRows.length) {{
        return {{ hours: 0, issue_count: 0, uses_leave_subtasks: false }};
      }}
      let total = 0;
      let issueCount = 0;
      const seenIssueKeys = new Set();
      for (const row of leaveSubtaskRows) {{
        const classification = String(row && row.leave_classification || "").trim().toLowerCase();
        if (classification !== "planned") {{
          continue;
        }}
        const hours = Number(row && row.original_estimate_hours || 0);
        if (!Number.isFinite(hours) || hours <= 0) {{
          continue;
        }}
        const start = parseIsoDate(row && row.start_date);
        const end = parseIsoDate(row && row.due_date);
        let include = true;
        if (start || end) {{
          const rowStart = start || end;
          const rowEnd = end || start;
          include = !!rowStart && !!rowEnd && !(rowEnd < from || rowStart > to);
        }}
        if (!include) {{
          continue;
        }}
        const issueKey = String(row && row.issue_key || "").trim().toUpperCase();
        if (issueKey) {{
          if (seenIssueKeys.has(issueKey)) {{
            continue;
          }}
          seenIssueKeys.add(issueKey);
        }}
        total += hours;
        issueCount += 1;
      }}
      return {{
        hours: Number(total.toFixed(2)),
        issue_count: issueCount,
        uses_leave_subtasks: true,
      }};
    }}

    function escapeHtml(value) {{
      return String(value ?? "")
        .replaceAll("&", "&amp;")
        .replaceAll("<", "&lt;")
        .replaceAll(">", "&gt;")
        .replaceAll('"', "&quot;")
        .replaceAll("'", "&#39;");
    }}

    function hoursText(value) {{
      const n = Number(value || 0);
      return n.toFixed(2).replace(/\\.00$/, "");
    }}

    function weekLabelFromCode(code) {{
      const m = String(code || "").match(/^(\\d{{4}})-W(\\d{{2}})$/);
      if (!m) return code;
      const year = Number(m[1]);
      const week = Number(m[2]);
      const jan4 = new Date(Date.UTC(year, 0, 4));
      const jan4Day = jan4.getUTCDay() === 0 ? 7 : jan4.getUTCDay();
      const monday = new Date(jan4);
      monday.setUTCDate(jan4.getUTCDate() - jan4Day + 1 + (week - 1) * 7);
      const sunday = new Date(monday);
      sunday.setUTCDate(monday.getUTCDate() + 6);
      const fmt = (d) => d.toLocaleDateString(undefined, {{ month: "short", day: "2-digit", timeZone: "UTC" }});
      return `${{code}} (${{fmt(monday)}} - ${{fmt(sunday)}})`;
    }}

    function periodValue(row, granularity) {{
      if (granularity === "day") return row.period_day;
      if (granularity === "month") return row.period_month;
      return row.period_week;
    }}

    function periodLabel(period, granularity) {{
      if (granularity === "week") return weekLabelFromCode(period);
      return period;
    }}

    function comparePeriods(a, b, granularity) {{
      if (granularity === "day" || granularity === "month") {{
        return String(a).localeCompare(String(b));
      }}
      const am = String(a).match(/^(\\d{{4}})-W(\\d{{2}})$/);
      const bm = String(b).match(/^(\\d{{4}})-W(\\d{{2}})$/);
      if (am && bm) {{
        const ay = Number(am[1]);
        const by = Number(bm[1]);
        if (ay !== by) return ay - by;
        return Number(am[2]) - Number(bm[2]);
      }}
      return String(a).localeCompare(String(b));
    }}

    function dayBoundaryFlags(period) {{
      const d = parseIsoDate(period);
      if (!d) {{
        return {{ is_week_end: false, is_month_end: false, is_both: false }};
      }}
      const isWeekEnd = d.getDay() === 0;
      const next = new Date(d);
      next.setDate(d.getDate() + 1);
      const isMonthEnd = next.getMonth() !== d.getMonth();
      return {{
        is_week_end: isWeekEnd,
        is_month_end: isMonthEnd,
        is_both: isWeekEnd && isMonthEnd,
      }};
    }}

    function buildCrosstab(filteredRows, granularity) {{
      const grid = new Map();
      const periodSet = new Set();

      for (const row of filteredRows) {{
        const period = periodValue(row, granularity);
        const assignee = row.worklog_author || row.issue_assignee || "Unassigned";
        const hours = Number(row.hours_logged || 0);
        if (!period || hours <= 0) continue;
        periodSet.add(period);
        if (!grid.has(assignee)) grid.set(assignee, new Map());
        const byPeriod = grid.get(assignee);
        byPeriod.set(period, Number((byPeriod.get(period) || 0) + hours));
      }}

      const columns = Array.from(periodSet).sort((a, b) => comparePeriods(a, b, granularity));
      const rowItems = [];
      const totalsByColumn = new Map(columns.map((c) => [c, 0]));

      for (const [assignee, byPeriod] of grid.entries()) {{
        let rowTotal = 0;
        const values = {{}};
        for (const column of columns) {{
          const value = Number(byPeriod.get(column) || 0);
          values[column] = Number(value.toFixed(2));
          rowTotal += value;
          totalsByColumn.set(column, Number(((totalsByColumn.get(column) || 0) + value).toFixed(2)));
        }}
        rowItems.push({{
          assignee,
          values,
          total_hours: Number(rowTotal.toFixed(2)),
        }});
      }}

      rowItems.sort((a, b) => a.assignee.localeCompare(b.assignee));
      rowItems.sort((a, b) => b.total_hours - a.total_hours);

      const overallTotal = Number(rowItems.reduce((acc, item) => acc + Number(item.total_hours || 0), 0).toFixed(2));
      const columnMeta = columns.map((period) => {{
        const flags = granularity === "day" ? dayBoundaryFlags(period) : {{ is_week_end: false, is_month_end: false, is_both: false }};
        return {{
          period,
          label: periodLabel(period, granularity),
          ...flags,
        }};
      }});

      return {{
        granularity,
        columns,
        column_meta: columnMeta,
        row_items: rowItems,
        grand_totals: {{
          columns: Object.fromEntries(columns.map((column) => [column, Number((totalsByColumn.get(column) || 0).toFixed(2))])),
          overall_total: overallTotal,
        }},
      }};
    }}

    function buildCrosstabFromPeriodMap(periodMap, granularity) {{
      const mapObj = periodMap && typeof periodMap === "object" ? periodMap : {{}};
      const columns = Object.keys(mapObj).sort((a, b) => comparePeriods(a, b, granularity));
      const byAssignee = new Map();
      for (const period of columns) {{
        const assignees = mapObj[period] && typeof mapObj[period] === "object" ? mapObj[period] : {{}};
        for (const [assigneeRaw, hoursRaw] of Object.entries(assignees)) {{
          const assignee = String(assigneeRaw || "").trim() || "Unassigned";
          const hours = Number(hoursRaw || 0);
          if (!Number.isFinite(hours) || hours <= 0) {{
            continue;
          }}
          if (!byAssignee.has(assignee)) {{
            byAssignee.set(assignee, new Map());
          }}
          const periodHours = byAssignee.get(assignee);
          periodHours.set(period, Number(((periodHours.get(period) || 0) + hours).toFixed(2)));
        }}
      }}

      const rowItems = [];
      const totalsByColumn = new Map(columns.map((c) => [c, 0]));
      for (const [assignee, periodHours] of byAssignee.entries()) {{
        let rowTotal = 0;
        const values = {{}};
        for (const column of columns) {{
          const value = Number(periodHours.get(column) || 0);
          values[column] = Number(value.toFixed(2));
          rowTotal += value;
          totalsByColumn.set(column, Number(((totalsByColumn.get(column) || 0) + value).toFixed(2)));
        }}
        rowItems.push({{
          assignee,
          values,
          total_hours: Number(rowTotal.toFixed(2)),
        }});
      }}

      rowItems.sort((a, b) => a.assignee.localeCompare(b.assignee));
      rowItems.sort((a, b) => b.total_hours - a.total_hours);
      const overallTotal = Number(rowItems.reduce((acc, item) => acc + Number(item.total_hours || 0), 0).toFixed(2));
      const columnMeta = columns.map((period) => {{
        const flags = granularity === "day" ? dayBoundaryFlags(period) : {{ is_week_end: false, is_month_end: false, is_both: false }};
        return {{
          period,
          label: periodLabel(period, granularity),
          ...flags,
        }};
      }});

      return {{
        granularity,
        columns,
        column_meta: columnMeta,
        row_items: rowItems,
        grand_totals: {{
          columns: Object.fromEntries(columns.map((column) => [column, Number((totalsByColumn.get(column) || 0).toFixed(2))])),
          overall_total: overallTotal,
        }},
      }};
    }}

    async function fetchActualAggregate(fromIso, toIso, mode, selectedProjectsSet) {{
      const projectList = Array.from(selectedProjectsSet || []).map((x) => String(x || "").trim()).filter(Boolean).join(",");
      const query = `from=${{encodeURIComponent(fromIso)}}&to=${{encodeURIComponent(toIso)}}&mode=${{encodeURIComponent(mode)}}&report=assignee_hours&projects=${{encodeURIComponent(projectList)}}`;
      const response = await fetch(`${{ACTUAL_AGG_ENDPOINT}}?${{query}}`, {{ method: "GET" }});
      const payloadData = await response.json().catch(() => ({{}}));
      if (!response.ok || !payloadData || payloadData.ok === false) {{
        throw new Error(String(payloadData && payloadData.error || "Failed to fetch actual-hour aggregate."));
      }}
      return payloadData;
    }}

    function renderCrosstab(targetEl, crosstab) {{
      const columns = crosstab.columns || [];
      const rowItems = crosstab.row_items || [];
      const columnMeta = crosstab.column_meta || [];
      const grand = crosstab.grand_totals || {{ columns: {{}}, overall_total: 0 }};

      if (!columns.length && !rowItems.length) {{
        targetEl.innerHTML = `<table><tbody><tr><td class="empty">No data for current filters.</td></tr></tbody></table>`;
        return;
      }}

      const headerCells = columnMeta.map((meta) => {{
        let classes = "";
        if (crosstab.granularity === "day") {{
          if (meta.is_both) classes = "day-week-month-end";
          else if (meta.is_month_end) classes = "day-month-end";
          else if (meta.is_week_end) classes = "day-week-end";
        }}
        return `<th class="${{classes}}">${{escapeHtml(meta.label)}}</th>`;
      }}).join("");

      const rowsHtml = rowItems.map((item) => {{
        const valueCells = columns
          .map((column) => `<td class="num">${{escapeHtml(hoursText(item.values[column] || 0))}}</td>`)
          .join("");
        return `<tr><td>${{escapeHtml(item.assignee)}}</td>${{valueCells}}<td class="num total-col">${{escapeHtml(hoursText(item.total_hours))}}</td></tr>`;
      }}).join("");

      const grandCells = columns
        .map((column) => `<td class="num">${{escapeHtml(hoursText((grand.columns || {{}})[column] || 0))}}</td>`)
        .join("");

      targetEl.innerHTML = `
        <table>
          <thead>
            <tr>
              <th>User</th>
              ${{headerCells}}
              <th class="total-col">Total</th>
            </tr>
          </thead>
          <tbody>
            ${{rowsHtml}}
            <tr class="grand-total">
              <td>Grand Total</td>
              ${{grandCells}}
              <td class="num total-col">${{escapeHtml(hoursText(grand.overall_total || 0))}}</td>
            </tr>
          </tbody>
        </table>
      `;
    }}

    function setActiveTab(tabKey) {{
      const key = ["day", "week", "month"].includes(tabKey) ? tabKey : "month";
      for (const btn of tabButtons) {{
        btn.classList.toggle("active", btn.dataset.tab === key);
      }}
      for (const [panelKey, panelEl] of Object.entries(tabPanels)) {{
        panelEl.classList.toggle("active", panelKey === key);
      }}
    }}

    function renderHolidayList() {{
      if (!selectedHolidayDates.length) {{
        holidayListEl.innerHTML = `<span class="empty">No holiday dates selected.</span>`;
        return;
      }}
      holidayListEl.innerHTML = selectedHolidayDates
        .map((iso) => `<span class="holiday-chip">${{escapeHtml(iso)}} <button type="button" data-remove-holiday="${{escapeHtml(iso)}}">x</button></span>`)
        .join("");
      Array.from(holidayListEl.querySelectorAll("button[data-remove-holiday]")).forEach((btn) => {{
        btn.addEventListener("click", () => {{
          const value = btn.getAttribute("data-remove-holiday") || "";
          selectedHolidayDates = selectedHolidayDates.filter((item) => item !== value);
          renderHolidayList();
        }});
      }});
    }}

    function addHolidayFromPicker() {{
      const value = holidayDatePickerEl.value || "";
      if (!/^\\d{{4}}-\\d{{2}}-\\d{{2}}$/.test(value)) return;
      if (!selectedHolidayDates.includes(value)) {{
        selectedHolidayDates.push(value);
        selectedHolidayDates.sort();
      }}
      renderHolidayList();
    }}

    function parseHolidayDates() {{
      return [...selectedHolidayDates];
    }}

    function setCapacityForm(settings) {{
      capacityEmployeesEl.value = settings.employee_count ?? 0;
      capacityStandardHoursEl.value = settings.standard_hours_per_day ?? 8;
      ramadanStartEl.value = settings.ramadan_start_date || "";
      ramadanEndEl.value = settings.ramadan_end_date || "";
      ramadanHoursEl.value = settings.ramadan_hours_per_day ?? 6.5;
      selectedHolidayDates = Array.isArray(settings.holiday_dates) ? [...settings.holiday_dates].sort() : [];
      renderHolidayList();
    }}

    function capacityPayloadFromForm() {{
      return {{
        from_date: fromDateEl.value || defaults.from || "",
        to_date: toDateEl.value || defaults.to || "",
        employee_count: Number(capacityEmployeesEl.value || 0),
        standard_hours_per_day: Number(capacityStandardHoursEl.value || 8),
        ramadan_start_date: ramadanStartEl.value || "",
        ramadan_end_date: ramadanEndEl.value || "",
        ramadan_hours_per_day: Number(ramadanHoursEl.value || 6.5),
        holiday_dates: parseHolidayDates(),
      }};
    }}

    function setCapacityFormFromProfile(profileSettings) {{
      const currentFrom = fromDateEl.value || defaults.from || "";
      const currentTo = toDateEl.value || defaults.to || "";
      setCapacityForm({{
        ...profileSettings,
        from_date: currentFrom,
        to_date: currentTo,
      }});
    }}

    async function recalcCapacityFromCurrentForm() {{
      const payloadNow = capacityPayloadFromForm();
      if (!backendAvailable) {{
        return calculateCapacityClient(payloadNow);
      }}
      const response = await fetch("/api/capacity/calculate", {{
        method: "POST",
        headers: {{ "Content-Type": "application/json" }},
        body: JSON.stringify(payloadNow),
      }});
      const data = await response.json();
      if (!response.ok) {{
        throw new Error(data.error || "Capacity calculation failed.");
      }}
      return data;
    }}

    function renderCapacityProfiles() {{
      if (!capacityProfiles.length) {{
        capacityProfileSelectEl.innerHTML = `<option value="">No saved profiles</option>`;
        return;
      }}
      capacityProfileSelectEl.innerHTML = capacityProfiles
        .map((p) => {{
          const value = `${{p.from_date}}|${{p.to_date}}`;
          const label = `${{p.from_date}} to ${{p.to_date}}  (Emp:${{p.employee_count}}, Std:${{hoursText(p.standard_hours_per_day)}}h, Ram:${{hoursText(p.ramadan_hours_per_day)}}h)`;
          return `<option value="${{escapeHtml(value)}}">${{escapeHtml(label)}}</option>`;
        }})
        .join("");
    }}

    async function loadCapacityProfiles() {{
      if (!window.location.protocol.startsWith("http")) {{
        backendAvailable = false;
        capacityProfiles = embeddedCapacityProfiles.slice();
        renderCapacityProfiles();
        return;
      }}
      try {{
        const response = await fetch("/api/capacity/profiles");
        const data = await response.json();
        if (!response.ok) {{
          throw new Error(data.error || "Failed to load saved profiles.");
        }}
        backendAvailable = true;
        capacityProfiles = Array.isArray(data.profiles) ? data.profiles : [];
      }} catch (err) {{
        backendAvailable = false;
        capacityProfiles = embeddedCapacityProfiles.slice();
      }}
      renderCapacityProfiles();
    }}

    function effectiveHoursPerWorkday(settings) {{
      const from = parseIsoDate(settings.from_date);
      const to = parseIsoDate(settings.to_date);
      if (!from || !to || to < from) {{
        return Number(settings.standard_hours_per_day || 8);
      }}
      const holidaySet = new Set(Array.isArray(settings.holiday_dates) ? settings.holiday_dates : []);
      const ramadanStart = settings.ramadan_start_date ? parseIsoDate(settings.ramadan_start_date) : null;
      const ramadanEnd = settings.ramadan_end_date ? parseIsoDate(settings.ramadan_end_date) : null;
      const workdayHours = [];
      const cursor = new Date(from);
      while (cursor <= to) {{
        const iso = cursor.toISOString().slice(0, 10);
        const isWeekday = cursor.getDay() >= 1 && cursor.getDay() <= 5;
        if (isWeekday && !holidaySet.has(iso)) {{
          const inRamadan = ramadanStart && ramadanEnd && cursor >= ramadanStart && cursor <= ramadanEnd;
          workdayHours.push(inRamadan ? Number(settings.ramadan_hours_per_day || 6.5) : Number(settings.standard_hours_per_day || 8));
        }}
        cursor.setDate(cursor.getDate() + 1);
      }}
      if (!workdayHours.length) {{
        return Number(settings.standard_hours_per_day || 8);
      }}
      const sum = workdayHours.reduce((acc, value) => acc + Number(value || 0), 0);
      return sum / workdayHours.length;
    }}

    function hoursPerDayForIso(isoDay, settings) {{
      const day = parseIsoDate(isoDay);
      if (!day) return Number(settings.standard_hours_per_day || 8);
      const ramadanStart = settings.ramadan_start_date ? parseIsoDate(settings.ramadan_start_date) : null;
      const ramadanEnd = settings.ramadan_end_date ? parseIsoDate(settings.ramadan_end_date) : null;
      const inRamadan = ramadanStart && ramadanEnd && day >= ramadanStart && day <= ramadanEnd;
      return inRamadan ? Number(settings.ramadan_hours_per_day || 6.5) : Number(settings.standard_hours_per_day || 8);
    }}

    function dailyHoursToDaysMap(dailyMap, settings) {{
      let totalDays = 0;
      for (const [isoDay, hours] of Object.entries(dailyMap || {{}})) {{
        const dayHours = Number(hoursPerDayForIso(isoDay, settings) || 0);
        if (dayHours <= 0) continue;
        totalDays += Number(hours || 0) / dayHours;
      }}
      return Number(totalDays.toFixed(2));
    }}

    function buildLeaveMetricsFromEmbedded(settings) {{
      const from = parseIsoDate(settings.from_date);
      const to = parseIsoDate(settings.to_date);
      if (!from || !to || to < from) {{
        return {{
          planned_taken_hours: 0,
          unplanned_taken_hours: 0,
          planned_not_taken_hours: 0,
          taken_hours: 0,
          not_yet_taken_hours: 0,
          taken_days: 0,
          not_yet_taken_days: 0,
          source: "embedded_unavailable",
        }};
      }}

      const plannedTakenByDay = {{}};
      const unplannedTakenByDay = {{}};
      const plannedNotTakenByDay = {{}};
      for (const row of leaveDailyRows) {{
        const isoDay = String(row && row.period_day || "");
        const day = parseIsoDate(isoDay);
        if (!day || day < from || day > to) continue;
        plannedTakenByDay[isoDay] = Number((plannedTakenByDay[isoDay] || 0) + Number(row && row.planned_taken_hours || 0));
        unplannedTakenByDay[isoDay] = Number((unplannedTakenByDay[isoDay] || 0) + Number(row && row.unplanned_taken_hours || 0));
        plannedNotTakenByDay[isoDay] = Number((plannedNotTakenByDay[isoDay] || 0) + Number(row && row.planned_not_taken_hours || 0));
      }}

      const plannedTaken = Number(Object.values(plannedTakenByDay).reduce((a, b) => a + Number(b || 0), 0).toFixed(2));
      const unplannedTaken = Number(Object.values(unplannedTakenByDay).reduce((a, b) => a + Number(b || 0), 0).toFixed(2));
      const plannedNotTaken = Number(Object.values(plannedNotTakenByDay).reduce((a, b) => a + Number(b || 0), 0).toFixed(2));
      const taken = Number((plannedTaken + unplannedTaken).toFixed(2));

      const takenByDay = {{ ...plannedTakenByDay }};
      for (const [isoDay, hours] of Object.entries(unplannedTakenByDay)) {{
        takenByDay[isoDay] = Number((takenByDay[isoDay] || 0) + Number(hours || 0));
      }}
      return {{
        planned_taken_hours: plannedTaken,
        unplanned_taken_hours: unplannedTaken,
        planned_not_taken_hours: plannedNotTaken,
        taken_hours: taken,
        not_yet_taken_hours: plannedNotTaken,
        taken_days: dailyHoursToDaysMap(takenByDay, settings),
        not_yet_taken_days: dailyHoursToDaysMap(plannedNotTakenByDay, settings),
        source: leaveDailyRows.length ? "embedded_rlt_leave_report.xlsx" : "embedded_unavailable",
      }};
    }}

    function calculateCapacityClient(input) {{
      const settings = {{
        ...input,
        holiday_dates: Array.isArray(input.holiday_dates) ? input.holiday_dates : [],
      }};
      const from = parseIsoDate(settings.from_date);
      const to = parseIsoDate(settings.to_date);
      if (!from || !to || to < from) {{
        throw new Error("Invalid report date range.");
      }}
      if (settings.employee_count < 0) throw new Error("Employees must be >= 0.");
      if (settings.standard_hours_per_day <= 0) throw new Error("Standard hours/day must be > 0.");
      if (settings.ramadan_hours_per_day <= 0) throw new Error("Ramadan hours/day must be > 0.");
      if (!!settings.ramadan_start_date !== !!settings.ramadan_end_date) {{
        throw new Error("Provide both Ramadan start and Ramadan end, or keep both empty.");
      }}

      const holidaySet = new Set(settings.holiday_dates);
      const ramadanStart = settings.ramadan_start_date ? parseIsoDate(settings.ramadan_start_date) : null;
      const ramadanEnd = settings.ramadan_end_date ? parseIsoDate(settings.ramadan_end_date) : null;
      if (ramadanStart && ramadanEnd && ramadanEnd < ramadanStart) {{
        throw new Error("Ramadan end must be on/after Ramadan start.");
      }}

      let totalWeekdays = 0;
      let holidayWeekdays = 0;
      let ramadanWeekdays = 0;
      let nonRamadanWeekdays = 0;
      const cursor = new Date(from);
      while (cursor <= to) {{
        const iso = cursor.toISOString().slice(0, 10);
        const isWeekday = cursor.getDay() >= 1 && cursor.getDay() <= 5;
        if (isWeekday) {{
          totalWeekdays += 1;
          if (holidaySet.has(iso)) {{
            holidayWeekdays += 1;
          }} else {{
            const inRamadan = ramadanStart && ramadanEnd && cursor >= ramadanStart && cursor <= ramadanEnd;
            if (inRamadan) ramadanWeekdays += 1;
            else nonRamadanWeekdays += 1;
          }}
        }}
        cursor.setDate(cursor.getDate() + 1);
      }}
      const available = settings.employee_count * (
        nonRamadanWeekdays * settings.standard_hours_per_day +
        ramadanWeekdays * settings.ramadan_hours_per_day
      );
      const leaveMetrics = buildLeaveMetricsFromEmbedded(settings);
      const remaining = Number((available - Number(leaveMetrics.taken_hours || 0) - Number(leaveMetrics.not_yet_taken_hours || 0)).toFixed(2));
      return {{
        settings,
        metrics: {{
          total_weekdays: totalWeekdays,
          holiday_weekdays: holidayWeekdays,
          ramadan_weekdays: ramadanWeekdays,
          non_ramadan_weekdays: nonRamadanWeekdays,
          available_capacity_hours: Number(available.toFixed(2)),
        }},
        leave_metrics: {{
          ...leaveMetrics,
          remaining_balance_hours: remaining,
          remaining_balance_days: Number((remaining / Math.max(effectiveHoursPerWorkday(settings), 0.1)).toFixed(2)),
        }},
      }};
    }}

    function updateCapacityKpis(capacityResult) {{
      const available = Number(capacityResult?.metrics?.available_capacity_hours || 0);
      const util = available > 0 ? (currentProjectActualHours / available) * 100 : 0;
      kpiCapacityEl.textContent = `${{hoursText(available)}}h`;
      kpiUtilizationEl.textContent = `${{util.toFixed(1)}}%`;
      headerKpiUtilizationEl.textContent = `${{util.toFixed(1)}}%`;

      kpiUtilCardEl.classList.remove("good", "warn", "bad");
      if (util < 70) kpiUtilCardEl.classList.add("good");
      else if (util <= 100) kpiUtilCardEl.classList.add("warn");
      else kpiUtilCardEl.classList.add("bad");

      const s = capacityResult.settings || {{}};
      const range = `${{s.from_date || "-"}} to ${{s.to_date || "-"}}`;
      capacityAssumptionEl.textContent =
        `Employees ${{s.employee_count ?? 0}} | Std ${{hoursText(s.standard_hours_per_day || 8)}}h | Ramadan ${{hoursText(s.ramadan_hours_per_day || 6.5)}}h | Range ${{range}}`;

      const leave = capacityResult.leave_metrics || {{}};
      const plannedLeavesTakenHours = Number(leave.planned_taken_hours || 0);
      const plannedLeavesNotYetHours = Number(leave.planned_not_taken_hours || leave.not_yet_taken_hours || 0);
      const unplannedLeavesTakenHours = Number(leave.unplanned_taken_hours || 0);
      const totalLeavesTakenHours = plannedLeavesTakenHours + unplannedLeavesTakenHours;
      const leaveTotalHours = plannedLeavesTakenHours + plannedLeavesNotYetHours + unplannedLeavesTakenHours;
      const leaveSubtasksPlanned = getLeaveSubtaskPlannedInRange(s.from_date || "", s.to_date || "");
      const rltLeavesPlanned = leaveSubtasksPlanned.uses_leave_subtasks
        ? leaveSubtasksPlanned
        : getRltLeavesPlannedInRange(s.from_date || "", s.to_date || "");
      const totalLeavesPlannedHours = Number(rltLeavesPlanned.hours || 0);
      const totalCapacityPlannedLeavesAdjustedHours = available - totalLeavesPlannedHours;
      const remainingCapacityHours = available - leaveTotalHours;
      const subtractionResultHours = available - currentProjectActualHours - leaveTotalHours;
      const remainingHours = Number(leave.remaining_balance_hours || 0);
      kpiLoggedEl.textContent = `${{hoursText(currentProjectActualHours)}}h`;
      kpiPlannedLeavesTakenEl.textContent = `${{hoursText(plannedLeavesTakenHours)}}h`;
      kpiPlannedLeavesNotYetEl.textContent = `${{hoursText(plannedLeavesNotYetHours)}}h`;
      kpiUnplannedLeavesTakenEl.textContent = `${{hoursText(unplannedLeavesTakenHours)}}h`;
      kpiLeaveTotalEl.textContent = `${{hoursText(leaveTotalHours)}}h`;
      kpiLeaveTotalRefEl.textContent = `${{hoursText(leaveTotalHours)}}h`;
      kpiGapEl.textContent = `${{hoursText(subtractionResultHours)}}h`;
      headerKpiGapEl.textContent = `${{hoursText(subtractionResultHours)}}h`;
      kpiGapCardEl.classList.remove("good", "bad");
      kpiGapCardEl.classList.add(subtractionResultHours >= 0 ? "good" : "bad");

      const totalEmployees = Number(s.employee_count ?? 0);
      const totalCapacity = Number(available || 0);
      const plannedMinusLeaves = Number(currentProjectPlannedHours || 0);
      const actualHours = Number(currentProjectActualHours || 0);
      const delta = plannedMinusLeaves - actualHours;
      const capacityGapHours = totalCapacity - plannedMinusLeaves - totalLeavesPlannedHours;

      summaryTotalCapacityEl.textContent = `${{hoursText(totalCapacity)}}h`;
      summaryLeaveHoursEl.textContent = `${{hoursText(leaveTotalHours)}}h`;
      summaryRemainingCapacityEl.textContent = `${{hoursText(remainingCapacityHours)}}h`;
      summaryTotalEmployeesEl.textContent = `${{Math.round(totalEmployees)}}`;
      summaryPlannedHoursEl.textContent = `${{hoursText(plannedMinusLeaves)}}h`;
      summaryActualHoursEl.textContent = `${{hoursText(actualHours)}}h`;
      summaryDeltaHoursEl.textContent = `${{hoursText(delta)}}h`;
      headerKpiActualEl.textContent = `${{hoursText(actualHours)}}h`;
      headerKpiDeltaEl.textContent = `${{hoursText(delta)}}h`;
      summaryDeltaCardEl.classList.remove("good", "bad");
      summaryDeltaCardEl.classList.add(delta >= 0 ? "good" : "bad");
      capacityRiskBadgeEl.classList.remove("warn", "risk");
      if (util > 100 || subtractionResultHours < 0) {{
        capacityRiskBadgeEl.classList.add("risk");
        capacityRiskBadgeEl.textContent = "Capacity At Risk";
      }} else if (util > 85) {{
        capacityRiskBadgeEl.classList.add("warn");
        capacityRiskBadgeEl.textContent = "Capacity Tight";
      }} else {{
        capacityRiskBadgeEl.textContent = "Capacity Healthy";
      }}

      if (nestedScoreTotalCapacityEl) {{
        nestedScoreTotalCapacityEl.textContent = `${{hoursText(totalCapacity)}}h`;
        nestedScoreTotalPlannedEl.textContent = `${{hoursText(plannedMinusLeaves)}}h`;
        nestedScoreTotalLoggedEl.textContent = `${{hoursText(actualHours)}}h`;
        nestedScoreDeltaEl.textContent = `${{hoursText(delta)}}h`;
        nestedScoreTotalLeavesTakenEl.textContent = `${{hoursText(totalLeavesTakenHours)}}h`;
        nestedScoreTotalLeavesPlannedEl.textContent = `${{hoursText(totalLeavesPlannedHours)}}h`;
        nestedScoreTotalCapacityPlannedLeavesAdjustedEl.textContent = `${{hoursText(totalCapacityPlannedLeavesAdjustedHours)}}h`;
        nestedScoreCapacityGapEl.textContent = `${{hoursText(capacityGapHours)}}h`;

        const scoreRangeFrom = s.from_date || "-";
        const scoreRangeTo = s.to_date || "-";
        nestedScoreTotalCapacityTipEl.textContent =
          `Formula: Total Capacity = Capacity Profile Capacity for selected date range.\\nValues:\\nDate Range = ${{scoreRangeFrom}} to ${{scoreRangeTo}}\\nTotal Capacity = ${{hoursText(totalCapacity)}}h`;
        nestedScoreTotalPlannedTipEl.textContent =
          `Formula: Total Planned Projects = Sum(Original Estimates for Epic/Story/Subtask items where planned start OR end is in selected range), excluding RLT (RnD Leave Tracker).\\nValues:\\nDate Range = ${{scoreRangeFrom}} to ${{scoreRangeTo}}\\nTotal Planned Projects = ${{hoursText(plannedMinusLeaves)}}h`;
        nestedScoreTotalLoggedTipEl.textContent =
          `Formula: Total Actual Project Hours = Sum(Project Actual Hours), excluding RLT (RnD Leave Tracker).\\nValues:\\nTotal Actual Project Hours = ${{hoursText(actualHours)}}h`;
        nestedScoreDeltaTipEl.textContent =
          `Formula: Plan Gap = Total Planned Projects - Total Actual Project Hours.\\nValues:\\nTotal Planned Projects = ${{hoursText(plannedMinusLeaves)}}h\\nTotal Actual Project Hours = ${{hoursText(actualHours)}}h\\nPlan Gap = ${{hoursText(delta)}}h`;
        nestedScoreTotalLeavesTakenTipEl.textContent =
          `Formula: Total Leaves Taken = Sum(Logged Hours for RLT RnD Leave Tracker leave work).\\nValues:\\nPlanned Leaves Taken = ${{hoursText(plannedLeavesTakenHours)}}h\\nUnplanned Leaves Taken = ${{hoursText(unplannedLeavesTakenHours)}}h\\nTotal Leaves Taken = ${{hoursText(totalLeavesTakenHours)}}h`;
        nestedScoreTotalLeavesPlannedTipEl.textContent =
          `Formula: Total Leaves Planned = Sum(Original Estimates) from leave workbook planned subtasks overlapping selected date range.\\nValues:\\nDate Range = ${{scoreRangeFrom}} to ${{scoreRangeTo}}\\nPlanned Leave Issues Count = ${{Math.round(Number(rltLeavesPlanned.issue_count || 0))}}\\nTotal Leaves Planned = ${{hoursText(totalLeavesPlannedHours)}}h`;
        nestedScoreTotalCapacityPlannedLeavesAdjustedTipEl.textContent =
          `Formula: Total Capacity (Planned Leaves Adjusted) = Total Capacity (Hours) - Total Leaves Planned.\\nValues:\\nTotal Capacity = ${{hoursText(totalCapacity)}}h\\nTotal Leaves Planned = ${{hoursText(totalLeavesPlannedHours)}}h\\nTotal Capacity (Planned Leaves Adjusted) = ${{hoursText(totalCapacityPlannedLeavesAdjustedHours)}}h`;
        nestedScoreCapacityGapTipEl.textContent =
          `Formula: Capacity available for more work = Total Capacity (Hours) - Total Planned Projects (Hours) - RLT RnD Leave Tracker Original Estimates.\\nValues:\\nTotal Capacity = ${{hoursText(totalCapacity)}}h\\nTotal Planned Projects = ${{hoursText(plannedMinusLeaves)}}h\\nRLT RnD Leave Tracker Original Estimates = ${{hoursText(totalLeavesPlannedHours)}}h\\nCapacity available for more work = ${{hoursText(capacityGapHours)}}h`;
      }}

      capacitySummaryCardsEl.classList.add("visible");
    }}

    async function loadCapacityFromBackend() {{
      const fromDate = fromDateEl.value || defaults.from || "";
      const toDate = toDateEl.value || defaults.to || "";
      if (!fromDate || !toDate) return;
      if (!backendAvailable && window.location.protocol.startsWith("http")) {{
        await loadCapacityProfiles();
      }}
      if (!backendAvailable) {{
        const local = calculateCapacityClient({{
          from_date: fromDate,
          to_date: toDate,
          employee_count: 0,
          standard_hours_per_day: 8,
          ramadan_start_date: "",
          ramadan_end_date: "",
          ramadan_hours_per_day: 6.5,
          holiday_dates: [],
        }});
        setCapacityForm(local.settings);
        activeProfileLabel = "Project Defaults";
        updateActiveProfileIndicator();
        updateCapacityKpis(local);
        await loadCapacityProfiles();
        capacityStatusEl.textContent = "Static mode: values are local only.";
        return;
      }}
      const response = await fetch(`/api/capacity?from=${{encodeURIComponent(fromDate)}}&to=${{encodeURIComponent(toDate)}}`);
      if (!response.ok) {{
        const msg = await response.text();
        throw new Error(msg || "Failed to load capacity settings.");
      }}
      const data = await response.json();
      setCapacityForm(data.settings || {{}});
      activeProfileLabel = `Range ${{data?.settings?.from_date || fromDate}} to ${{data?.settings?.to_date || toDate}}`;
      updateActiveProfileIndicator();
      updateCapacityKpis(data);
      capacityStatusEl.textContent = "Loaded saved capacity settings for selected range.";
      try {{
        await loadCapacityProfiles();
      }} catch (profileErr) {{
        capacityStatusEl.textContent = `Loaded capacity, but profile list failed: ${{profileErr.message || profileErr}}`;
      }}
    }}

    async function applyFilters() {{
      const fromDate = parseIsoDate(fromDateEl.value || defaults.from);
      const toDate = parseIsoDate(toDateEl.value || defaults.to);
      const selected = selectedProjects();
      const actualHoursMode = actualHoursModeEl ? String(actualHoursModeEl.value || ACTUAL_MODE_DEFAULT) : ACTUAL_MODE_DEFAULT;
      const plannedHoursInRange = plannedWorkItems.length
        ? plannedWorkHoursInRange(fromDateEl.value || defaults.from, toDateEl.value || defaults.to, selected)
        : projectPlannedHoursAllEpics;
      localStorage.setItem(ACTUAL_MODE_STORAGE_KEY, actualHoursMode);

      let day;
      let week;
      let month;
      let totalHours = 0;
      let projectActualHours = 0;
      let filtered = [];

      if (actualHoursMode === "planned_dates" && String(window.location.protocol || "").startsWith("http")) {{
        try {{
          const payloadData = await fetchActualAggregate(
            fromDateEl.value || defaults.from,
            toDateEl.value || defaults.to,
            actualHoursMode,
            selected,
          );
          const byPeriod = payloadData && payloadData.assignee_hours_by_period && typeof payloadData.assignee_hours_by_period === "object"
            ? payloadData.assignee_hours_by_period
            : {{}};
          day = buildCrosstabFromPeriodMap(byPeriod.day, "day");
          week = buildCrosstabFromPeriodMap(byPeriod.week, "week");
          month = buildCrosstabFromPeriodMap(byPeriod.month, "month");
          const projectHours = payloadData && payloadData.project_hours_by_key && typeof payloadData.project_hours_by_key === "object"
            ? payloadData.project_hours_by_key
            : {{}};
          totalHours = Number(Object.values(projectHours).reduce((acc, value) => acc + Number(value || 0), 0).toFixed(2));
          projectActualHours = Number(Object.entries(projectHours).reduce((acc, entry) => {{
            const key = String(entry[0] || "").trim().toUpperCase();
            if (key === "RLT") {{
              return acc;
            }}
            return acc + Number(entry[1] || 0);
          }}, 0).toFixed(2));
        }} catch (err) {{
          capacityStatusEl.textContent = `Actual-hours mode fetch error: ${{err.message || err}}`;
        }}
      }}

      if (!day || !week || !month) {{
        filtered = rows.filter((row) => {{
          const rowDate = parseIsoDate(row.worklog_date);
          if (!rowDate || !selected.has(row.project_key)) return false;
          if (fromDate && rowDate < fromDate) return false;
          if (toDate && rowDate > toDate) return false;
          return true;
        }});
        day = buildCrosstab(filtered, "day");
        week = buildCrosstab(filtered, "week");
        month = buildCrosstab(filtered, "month");
        totalHours = filtered.reduce((acc, row) => acc + Number(row.hours_logged || 0), 0);
        projectActualHours = filtered.reduce((acc, row) => {{
          const key = String(row.project_key || "").trim().toUpperCase();
          if (key === "RLT") {{
            return acc;
          }}
          return acc + Number(row.hours_logged || 0);
        }}, 0);
      }}
      const selectedCount = selected.size;
      const projectCount = projects.length;
      const rangeText = `${{fromDateEl.value || defaults.from || "-"}} to ${{toDateEl.value || defaults.to || "-"}}`;

      summaryRangeEl.textContent = `Range: ${{rangeText}}`;
      summaryProjectsEl.textContent = `Projects: ${{selectedCount}} / ${{projectCount}}`;
      summaryTotalEl.textContent = `Hours: ${{hoursText(totalHours)}}`;
      currentProjectActualHours = Number(projectActualHours.toFixed(2));
      currentProjectPlannedHours = Number(plannedHoursInRange || 0);

      renderCrosstab(dayTableWrapEl, day);
      renderCrosstab(weekTableWrapEl, week);
      renderCrosstab(monthTableWrapEl, month);

      try {{
        await loadCapacityFromBackend();
      }} catch (err) {{
        capacityStatusEl.textContent = `Capacity load error: ${{err.message || err}}`;
      }}
    }}

    async function resetFilters() {{
      fromDateEl.value = defaults.from;
      toDateEl.value = defaults.to;
      selectAllProjects();
      activeProfileLabel = "Project Defaults";
      updateActiveProfileIndicator();
      await applyFilters();
    }}

    generatedAtEl.textContent = payload.generated_at || "-";
    assigneeCountHelpEl.textContent = String(assigneeCount);
    if (actualHoursModeEl) {{
      const storedActualMode = localStorage.getItem(ACTUAL_MODE_STORAGE_KEY);
      actualHoursModeEl.value = (storedActualMode === "planned_dates" || storedActualMode === "log_date")
        ? storedActualMode
        : ACTUAL_MODE_DEFAULT;
    }}
    updateActiveProfileIndicator();
    fillProjects();
    renderHolidayList();
    resetFilters();
    setActiveTab("month");
    applyEl.addEventListener("click", () => {{ applyFilters(); }});
    if (actualHoursModeEl) {{
      actualHoursModeEl.addEventListener("change", () => {{
        applyFilters();
      }});
    }}
    resetEl.addEventListener("click", () => {{ resetFilters(); }});
    tabButtons.forEach((btn) => {{
      btn.addEventListener("click", () => {{
        setActiveTab(btn.dataset.tab || "month");
      }});
    }});
    selectAllProjectsEl.addEventListener("click", () => {{
      selectAllProjects();
      applyFilters();
    }});
    clearProjectsEl.addEventListener("click", () => {{
      Array.from(projectSelectEl.options).forEach((opt) => {{ opt.selected = false; }});
      applyFilters();
    }});
    useAssigneeCountEl.addEventListener("click", () => {{
      capacityEmployeesEl.value = String(assigneeCount);
      try {{
        const result = calculateCapacityClient(capacityPayloadFromForm());
        updateCapacityKpis(result);
      }} catch (err) {{
        capacityStatusEl.textContent = `Validation error: ${{err.message || err}}`;
      }}
    }});
    holidayAddEl.addEventListener("click", () => {{
      addHolidayFromPicker();
    }});
    holidayDatePickerEl.addEventListener("change", () => {{
      addHolidayFromPicker();
    }});
    holidayClearEl.addEventListener("click", () => {{
      selectedHolidayDates = [];
      renderHolidayList();
    }});
    capacityRecalcEl.addEventListener("click", () => {{
      (async () => {{
        try {{
          const result = await recalcCapacityFromCurrentForm();
          updateCapacityKpis(result);
          capacityStatusEl.textContent = "Capacity recalculated.";
        }} catch (err) {{
          capacityStatusEl.textContent = `Validation error: ${{err.message || err}}`;
        }}
      }})();
    }});
    capacityProfileRefreshEl.addEventListener("click", () => {{
      (async () => {{
        try {{
          await loadCapacityProfiles();
          capacityStatusEl.textContent = "Saved profiles refreshed.";
        }} catch (err) {{
          capacityStatusEl.textContent = `Profile refresh error: ${{err.message || err}}`;
        }}
      }})();
    }});
    capacityProfileApplyEl.addEventListener("click", () => {{
      (async () => {{
        try {{
          const selected = String(capacityProfileSelectEl.value || "").trim();
          if (!selected || !selected.includes("|")) {{
            capacityStatusEl.textContent = "Select a saved profile first.";
            return;
          }}
          const [fromSaved, toSaved] = selected.split("|");
          activeProfileLabel = `${{fromSaved}} to ${{toSaved}}`;
          updateActiveProfileIndicator();
          if (backendAvailable) {{
            const response = await fetch(`/api/capacity?from=${{encodeURIComponent(fromSaved)}}&to=${{encodeURIComponent(toSaved)}}`);
            const data = await response.json();
            if (!response.ok) {{
              throw new Error(data.error || "Failed to load selected profile.");
            }}
            setCapacityFormFromProfile(data.settings || {{}});
          }} else {{
            const fallbackProfile = (capacityProfiles || []).find((p) =>
              String(p && p.from_date || "") === fromSaved && String(p && p.to_date || "") === toSaved
            );
            if (!fallbackProfile) {{
              throw new Error("Selected profile is not available in current report payload.");
            }}
            setCapacityFormFromProfile(fallbackProfile);
          }}
          const evaluated = await recalcCapacityFromCurrentForm();
          updateCapacityKpis(evaluated);
          capacityStatusEl.textContent = backendAvailable
            ? "Profile applied to current selected range. Save to persist."
            : "Profile applied in static mode. Save requires API backend.";
        }} catch (err) {{
          capacityStatusEl.textContent = `Profile apply error: ${{err.message || err}}`;
        }}
      }})();
    }});
    capacitySaveEl.addEventListener("click", async () => {{
      try {{
        const payloadToSave = capacityPayloadFromForm();
        if (!backendAvailable) {{
          const local = calculateCapacityClient(payloadToSave);
          updateCapacityKpis(local);
          capacityStatusEl.textContent = "Static mode: cannot persist to SQLite. Run with --server.";
          return;
        }}
        const response = await fetch("/api/capacity", {{
          method: "POST",
          headers: {{ "Content-Type": "application/json" }},
          body: JSON.stringify(payloadToSave),
        }});
        const data = await response.json();
        if (!response.ok) {{
          throw new Error(data.error || "Failed to save capacity settings.");
        }}
        setCapacityForm(data.settings || {{}});
        activeProfileLabel = `Range ${{data?.settings?.from_date || payloadToSave.from_date}} to ${{data?.settings?.to_date || payloadToSave.to_date}}`;
        updateActiveProfileIndicator();
        updateCapacityKpis(data);
        capacityStatusEl.textContent = "Capacity settings saved.";
        await loadCapacityProfiles();
      }} catch (err) {{
        capacityStatusEl.textContent = `Save error: ${{err.message || err}}`;
      }}
    }});
    openCapacitySettingsEl.addEventListener("click", () => {{
      openSettingsDrawer();
    }});
    closeCapacitySettingsEl.addEventListener("click", () => {{
      closeSettingsDrawer();
    }});
    settingsDrawerOverlayEl.addEventListener("click", () => {{
      closeSettingsDrawer();
    }});
    document.addEventListener("keydown", (event) => {{
      if (event.key === "Escape" && settingsDrawerEl.classList.contains("open")) {{
        closeSettingsDrawer();
        return;
      }}
      trapDrawerFocus(event);
    }});
  </script>
<script src="shared-nav.js"></script>
</body>
</html>"""


def _resolve_runtime_paths(base_dir: Path) -> dict:
    base_dir = Path(__file__).resolve().parent
    input_name = os.getenv("JIRA_WORKLOG_XLSX_PATH", DEFAULT_WORKLOG_INPUT_XLSX).strip() or DEFAULT_WORKLOG_INPUT_XLSX
    work_items_name = os.getenv("JIRA_EXPORT_XLSX_PATH", DEFAULT_WORK_ITEMS_INPUT_XLSX).strip() or DEFAULT_WORK_ITEMS_INPUT_XLSX
    summary_name = os.getenv("JIRA_ASSIGNEE_HOURS_XLSX_PATH", DEFAULT_SUMMARY_OUTPUT_XLSX).strip() or DEFAULT_SUMMARY_OUTPUT_XLSX
    html_name = os.getenv("JIRA_ASSIGNEE_HOURS_HTML_PATH", DEFAULT_HTML_OUTPUT).strip() or DEFAULT_HTML_OUTPUT
    db_name = os.getenv("JIRA_ASSIGNEE_HOURS_CAPACITY_DB_PATH", DEFAULT_CAPACITY_DB).strip() or DEFAULT_CAPACITY_DB
    leave_name = os.getenv("JIRA_LEAVE_REPORT_XLSX_PATH", DEFAULT_LEAVE_REPORT_INPUT_XLSX).strip() or DEFAULT_LEAVE_REPORT_INPUT_XLSX

    input_path = _resolve_path(input_name, base_dir)
    work_items_path = _resolve_path(work_items_name, base_dir)
    summary_path = _resolve_path(summary_name, base_dir)
    html_path = _resolve_path(html_name, base_dir)
    db_path = _resolve_path(db_name, base_dir)
    leave_report_path = _resolve_path(leave_name, base_dir)
    return {
        "base_dir": base_dir,
        "input_path": input_path,
        "work_items_path": work_items_path,
        "summary_path": summary_path,
        "html_path": html_path,
        "db_path": db_path,
        "leave_report_path": leave_report_path,
    }


def _generate_outputs(paths: dict) -> dict:
    rows = _load_worklog_rows(paths["input_path"])
    _write_summary_xlsx(rows, paths["summary_path"])
    summary_rows = _read_summary_xlsx(paths["summary_path"])
    capacity_profiles = _list_capacity_profiles(paths["db_path"])
    leave_daily_rows = _load_leave_daily_rows(paths["leave_report_path"])
    leave_subtask_rows = _load_leave_subtask_rows(paths["leave_report_path"])
    project_planned_hours = _load_project_planned_hours_from_work_items(paths["work_items_path"])
    rlt_leaves_planned_rows = _load_rlt_leaves_planned_rows_from_work_items(paths["work_items_path"])
    planned_work_items = _load_planned_work_items_from_work_items(paths["work_items_path"])
    payload = _build_payload(
        summary_rows,
        capacity_profiles=capacity_profiles,
        leave_daily_rows=leave_daily_rows,
        leave_subtask_rows=leave_subtask_rows,
        project_planned_hours=project_planned_hours,
        rlt_leaves_planned_rows=rlt_leaves_planned_rows,
        planned_work_items=planned_work_items,
    )
    html = _build_html(payload)
    paths["html_path"].write_text(html, encoding="utf-8")
    return {
        "summary_rows": summary_rows,
        "payload": payload,
        "html_path": paths["html_path"],
        "summary_path": paths["summary_path"],
    }


def create_server_app(paths: dict):
    from flask import Flask, jsonify, request

    app = Flask(__name__)
    _init_capacity_db(paths["db_path"])
    outputs = _generate_outputs(paths)

    @app.route("/")
    def index():
        html_path = outputs["html_path"]
        if html_path.exists():
            return html_path.read_text(encoding="utf-8")
        refreshed = _generate_outputs(paths)
        return refreshed["html_path"].read_text(encoding="utf-8")

    @app.route("/api/capacity", methods=["GET"])
    def get_capacity():
        try:
            from_date = _to_text(request.args.get("from"))
            to_date = _to_text(request.args.get("to"))
            if not from_date or not to_date:
                return jsonify({"error": "Query params 'from' and 'to' are required."}), 400
            settings = _load_capacity_settings(paths["db_path"], from_date, to_date)
            result = calculate_capacity_metrics(settings)
            leave = _load_leave_metrics(
                paths["leave_report_path"],
                result["settings"]["from_date"],
                result["settings"]["to_date"],
                result["settings"],
            )
            remaining = round(
                result["metrics"]["available_capacity_hours"] - leave["taken_hours"] - leave["not_yet_taken_hours"],
                2,
            )
            result["leave_metrics"] = {
                **leave,
                "remaining_balance_hours": remaining,
                "remaining_balance_days": _hours_to_days_over_range(remaining, result["settings"]),
            }
            return jsonify(result)
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400

    @app.route("/api/capacity", methods=["POST"])
    def save_capacity():
        try:
            payload = request.get_json(silent=True) or {}
            saved = _save_capacity_settings(paths["db_path"], payload)
            result = calculate_capacity_metrics(saved)
            leave = _load_leave_metrics(
                paths["leave_report_path"],
                result["settings"]["from_date"],
                result["settings"]["to_date"],
                result["settings"],
            )
            remaining = round(
                result["metrics"]["available_capacity_hours"] - leave["taken_hours"] - leave["not_yet_taken_hours"],
                2,
            )
            result["leave_metrics"] = {
                **leave,
                "remaining_balance_hours": remaining,
                "remaining_balance_days": _hours_to_days_over_range(remaining, result["settings"]),
            }
            return jsonify(result)
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400

    @app.route("/api/capacity", methods=["DELETE"])
    def delete_capacity():
        try:
            from_date = _to_text(request.args.get("from"))
            to_date = _to_text(request.args.get("to"))
            if not from_date or not to_date:
                payload = request.get_json(silent=True) or {}
                from_date = _to_text(payload.get("from_date"))
                to_date = _to_text(payload.get("to_date"))
            if not from_date or not to_date:
                return jsonify({"error": "Range 'from/to' or 'from_date/to_date' is required."}), 400
            deleted = _delete_capacity_settings(paths["db_path"], from_date, to_date)
            return jsonify(
                {
                    "deleted": bool(deleted),
                    "from_date": from_date,
                    "to_date": to_date,
                }
            )
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400

    @app.route("/api/capacity/calculate", methods=["POST"])
    def calculate_capacity():
        try:
            payload = request.get_json(silent=True) or {}
            result = calculate_capacity_metrics(payload)
            leave = _load_leave_metrics(
                paths["leave_report_path"],
                result["settings"]["from_date"],
                result["settings"]["to_date"],
                result["settings"],
            )
            remaining = round(
                result["metrics"]["available_capacity_hours"] - leave["taken_hours"] - leave["not_yet_taken_hours"],
                2,
            )
            result["leave_metrics"] = {
                **leave,
                "remaining_balance_hours": remaining,
                "remaining_balance_days": _hours_to_days_over_range(remaining, result["settings"]),
            }
            return jsonify(result)
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400

    @app.route("/api/capacity/profiles", methods=["GET"])
    def list_capacity_profiles():
        return jsonify({"profiles": _list_capacity_profiles(paths["db_path"])})

    @app.route("/api/report/refresh", methods=["POST"])
    def refresh_report():
        refreshed = _generate_outputs(paths)
        return jsonify(
            {
                "rows": len(refreshed["summary_rows"]),
                "summary_path": str(refreshed["summary_path"]),
                "html_path": str(refreshed["html_path"]),
            }
        )

    return app


def run_server(paths: dict, port: int) -> None:
    app = create_server_app(paths)
    print(f"Assignee-hours server: http://localhost:{port}")
    print(f"Using capacity DB: {paths['db_path']}")
    app.run(host="0.0.0.0", port=port)


def main(server_mode: bool = False, port: int = 5000) -> None:
    paths = _resolve_runtime_paths(Path(__file__).resolve().parent)
    _init_capacity_db(paths["db_path"])
    if server_mode:
        run_server(paths, port)
        return

    outputs = _generate_outputs(paths)

    print(f"Assignee-hours summary rows: {len(outputs['summary_rows'])}")
    print(f"Wrote summary workbook: {outputs['summary_path']}")
    print(f"Wrote HTML report: {outputs['html_path']}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Generate assignee-hours report or run API server.")
    parser.add_argument("--server", action="store_true", help="Run Flask server with SQLite-backed capacity APIs.")
    parser.add_argument("--port", type=int, default=int(os.getenv("PORT", "5000")), help="Server port for --server mode.")
    args = parser.parse_args()
    main(server_mode=args.server, port=args.port)


