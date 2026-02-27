"""
Generate a phase-owner RMI gantt chart HTML report from nested view.xlsx.
"""
from __future__ import annotations

import json
import os
from datetime import date, datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

REQUIRED_HEADERS = [
    "Aspect",
    "Man-days",
    "Planned Start Date",
    "Planned End Date",
]

DEFAULT_INPUT_XLSX = "nested view.xlsx"
DEFAULT_OUTPUT_HTML = "phase_rmi_gantt_report.html"
FIXED_PHASE_NAMES = [
    "Research/URS",
    "DDS",
    "Development",
    "SQA",
    "User Manual",
    "Production",
]
PHASE_NAME_LOOKUP = {name.strip().lower(): name for name in FIXED_PHASE_NAMES}


def _resolve_path(value: str, base_dir: Path) -> Path:
    path = Path(value)
    if path.is_absolute():
        return path
    return base_dir / path


def _to_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _to_number_or_zero(value) -> float:
    if value in (None, ""):
        return 0.0
    try:
        return round(float(value), 2)
    except (TypeError, ValueError):
        return 0.0


def _parse_to_iso_date(value) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()

    text = _to_text(value)
    if not text:
        return ""

    formats = (
        "%Y-%m-%d",
        "%Y/%m/%d",
        "%d-%b-%Y",
        "%d-%B-%Y",
        "%m/%d/%Y",
        "%d/%m/%Y",
    )
    for fmt in formats:
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue

    try:
        parsed = datetime.fromisoformat(text.replace("Z", "+00:00"))
        return parsed.date().isoformat()
    except ValueError:
        return ""


def _row_type_from_level(level: int) -> str:
    mapping = {
        1: "project",
        2: "product",
        3: "rmi",
        4: "story",
        5: "subtask",
        6: "assignee",
    }
    return mapping.get(level, "unknown")


def _project_key_from_aspect(aspect: str) -> str:
    text = _to_text(aspect)
    if " - " in text:
        return text.split(" - ", 1)[0].strip()
    return text


def _load_phase_rmi_records(input_path: Path) -> list[dict[str, object]]:
    if not input_path.exists():
        raise FileNotFoundError(f"Nested view workbook not found: {input_path}")

    wb = load_workbook(input_path, read_only=False, data_only=True)
    ws = wb["NestedView"] if "NestedView" in wb.sheetnames else wb.active

    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header:
        wb.close()
        raise ValueError("Nested view workbook has no header row.")

    found_headers = [_to_text(cell) for cell in header]
    header_index = {h.lower(): i for i, h in enumerate(found_headers)}
    missing = [h for h in REQUIRED_HEADERS if h.lower() not in header_index]
    if missing:
        wb.close()
        raise ValueError(
            "Nested view workbook headers are missing required columns. "
            f"Required: {REQUIRED_HEADERS}, missing: {missing}, found: {found_headers}"
        )

    idx_aspect = header_index["aspect"]
    idx_man_days = header_index["man-days"]
    idx_planned_start = header_index["planned start date"]
    idx_planned_end = header_index["planned end date"]

    rows: list[dict[str, object]] = []
    stack: dict[int, dict[str, object]] = {}

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        level = int(getattr(ws.row_dimensions[row_idx], "outlineLevel", 0) or 0)
        if level <= 0:
            level = 1

        for key in list(stack):
            if key >= level:
                del stack[key]

        row_type = _row_type_from_level(level)
        aspect = _to_text(row[idx_aspect] if len(row) > idx_aspect else "")
        man_days = _to_number_or_zero(row[idx_man_days] if len(row) > idx_man_days else "")
        planned_start = _parse_to_iso_date(row[idx_planned_start] if len(row) > idx_planned_start else "")
        planned_end = _parse_to_iso_date(row[idx_planned_end] if len(row) > idx_planned_end else "")

        current = {
            "level": level,
            "row_type": row_type,
            "aspect": aspect,
            "man_days": man_days,
            "planned_start": planned_start,
            "planned_end": planned_end,
        }
        stack[level] = current

        if row_type == "story":
            rmi_node = stack.get(3)
            project_node = stack.get(1)
            if not rmi_node:
                continue
            phase_name = PHASE_NAME_LOOKUP.get(aspect.strip().lower())
            if not phase_name:
                continue
            if not planned_start or not planned_end:
                continue
            try:
                start_date = datetime.strptime(planned_start, "%Y-%m-%d").date()
                end_date = datetime.strptime(planned_end, "%Y-%m-%d").date()
            except ValueError:
                continue
            if end_date < start_date:
                continue

            rows.append(
                {
                    "phase_name": phase_name,
                    "rmi_name": _to_text(rmi_node.get("aspect")),
                    "man_days": man_days,
                    "planned_start": planned_start,
                    "planned_end": planned_end,
                    "project_key": _project_key_from_aspect(_to_text((project_node or {}).get("aspect", ""))),
                }
            )

    wb.close()
    return rows


def _aggregate_phase_rmi_rows(rows: list[dict[str, object]]) -> list[dict[str, object]]:
    grouped: dict[tuple[str, str], dict[str, object]] = {}
    for row in rows:
        phase_name = _to_text(row.get("phase_name"))
        rmi_name = _to_text(row.get("rmi_name"))
        if not phase_name or not rmi_name:
            continue
        key = (phase_name, rmi_name)
        current = grouped.get(key)
        if not current:
            grouped[key] = {
                "phase_name": phase_name,
                "rmi_name": rmi_name,
                "man_days": float(row.get("man_days") or 0.0),
                "planned_start": _to_text(row.get("planned_start")),
                "planned_end": _to_text(row.get("planned_end")),
                "project_key": _to_text(row.get("project_key")),
            }
            continue

        current["man_days"] = round(float(current["man_days"]) + float(row.get("man_days") or 0.0), 2)
        row_start = _to_text(row.get("planned_start"))
        row_end = _to_text(row.get("planned_end"))
        if row_start and (not current["planned_start"] or row_start < current["planned_start"]):
            current["planned_start"] = row_start
        if row_end and (not current["planned_end"] or row_end > current["planned_end"]):
            current["planned_end"] = row_end

    out = list(grouped.values())
    out.sort(key=lambda item: (_to_text(item.get("phase_name")).lower(), _to_text(item.get("rmi_name")).lower()))
    return out


def _build_html(data: dict[str, object]) -> str:
    payload = json.dumps(data, ensure_ascii=True)
    return f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Phase Owner RMI Gantt</title>
  <link rel="stylesheet" href="https://fonts.googleapis.com/css2?family=Material+Symbols+Outlined:opsz,wght,FILL,GRAD@20..48,500,0,0">
  <style>
    :root {{
      --bg: #eef3f7;
      --panel: #ffffff;
      --line: #d6e1ea;
      --text: #1f2937;
      --muted: #5f6f7f;
      --title: #0c4054;
      --head: #0f4c5c;
      --head-text: #ffffff;
      --phase-col: 260px;
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      color: var(--text);
      font-family: "Segoe UI", Tahoma, Verdana, sans-serif;
      background:
        radial-gradient(1000px 320px at 8% -10%, #dbeef7 0%, transparent 62%),
        linear-gradient(180deg, #f4f8fb, var(--bg));
    }}
    .page {{
      max-width: 1880px;
      margin: 0 auto;
      padding: 16px;
    }}
    .panel {{
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 12px;
      padding: 14px 16px;
      margin-bottom: 12px;
    }}
    .title {{
      margin: 0 0 6px;
      font-size: 1.25rem;
      color: var(--title);
    }}
    .meta {{
      margin: 0;
      color: var(--muted);
      font-size: 0.9rem;
    }}
    .controls {{
      margin-top: 10px;
      display: flex;
      flex-wrap: wrap;
      align-items: end;
      gap: 10px;
    }}
    .control {{
      display: flex;
      flex-direction: column;
      gap: 4px;
      min-width: 180px;
    }}
    .control label {{
      font-size: 0.8rem;
      color: #314756;
      font-weight: 600;
    }}
    .control input {{
      border: 1px solid #b9cad5;
      border-radius: 8px;
      padding: 7px 9px;
      font-size: 0.9rem;
      color: #102c3b;
      background: #fff;
    }}
    .control input:focus {{
      outline: none;
      border-color: #2a6274;
      box-shadow: 0 0 0 2px rgba(42, 98, 116, 0.16);
    }}
    .btn {{
      border: 1px solid #255f73;
      background: #0f4c5c;
      color: #fff;
      border-radius: 8px;
      padding: 7px 12px;
      font-size: 0.86rem;
      cursor: pointer;
      height: 34px;
    }}
    .btn.alt {{
      background: #fff;
      color: #255f73;
    }}
    .legend {{
      margin-top: 10px;
      display: flex;
      flex-wrap: wrap;
      gap: 8px;
      color: #506575;
      font-size: 0.78rem;
    }}
    .legend-pill {{
      display: inline-flex;
      align-items: center;
      gap: 6px;
      border: 1px solid #d8e4ee;
      border-radius: 999px;
      padding: 3px 8px;
      background: #f8fbff;
    }}
    .chip-dot {{
      width: 10px;
      height: 10px;
      border-radius: 999px;
      background: #93c5fd;
    }}
    .gantt-wrap {{
      background: #fff;
      border: 1px solid var(--line);
      border-radius: 12px;
      overflow: auto;
      max-height: 76vh;
    }}
    .gantt-root {{
      min-width: 860px;
    }}
    .grid {{
      min-width: max-content;
      border-collapse: separate;
      border-spacing: 0;
    }}
    .grid-row {{
      display: grid;
      grid-template-columns: var(--phase-col) 1fr;
      min-width: max-content;
    }}
    .cell {{
      border-top: 1px solid var(--line);
      background: #fff;
    }}
    .head .cell {{
      position: sticky;
      top: 0;
      z-index: 40;
      border-top: 0;
      background: var(--head);
      color: var(--head-text);
      font-weight: 700;
    }}
    .head .phase-cell {{
      padding: 10px 12px;
      border-right: 1px solid rgba(255,255,255,0.18);
    }}
    .head .timeline-cell {{
      padding: 0;
      z-index: 38;
    }}
    .phase-cell {{
      position: sticky;
      left: 0;
      z-index: 22;
      border-right: 1px solid var(--line);
      padding: 10px 12px;
      background: #fff;
      min-width: var(--phase-col);
      max-width: var(--phase-col);
    }}
    .phase-title {{
      font-size: 0.88rem;
      font-weight: 700;
      color: #0f3040;
      white-space: nowrap;
      overflow: hidden;
      text-overflow: ellipsis;
    }}
    .phase-sub {{
      margin-top: 3px;
      font-size: 0.74rem;
      color: #637a8a;
    }}
    .load-row {{
      margin-top: 7px;
      display: grid;
      grid-auto-flow: column;
      grid-auto-columns: 12px;
      gap: 3px;
      justify-content: start;
      align-items: center;
      overflow: hidden;
      max-width: 230px;
      min-height: 12px;
    }}
    .load-chip {{
      width: 12px;
      height: 10px;
      border-radius: 3px;
      border: 1px solid #d5e0e8;
      background: #eff4f8;
    }}
    .timeline-cell {{
      position: relative;
      overflow: hidden;
      min-height: 48px;
      background: #fff;
    }}
    .lane-timeline {{
      position: relative;
      padding: 8px 8px 10px;
    }}
    .timeline-header {{
      position: relative;
      min-height: 58px;
      background: #f8fbff;
      border-bottom: 1px solid #d4e0e9;
    }}
    .month-block {{
      position: absolute;
      top: 0;
      height: 22px;
      border-right: 1px solid #d0dce5;
      background: rgba(210, 228, 239, 0.35);
      color: #334f60;
      padding: 3px 6px;
      font-size: 0.72rem;
      white-space: nowrap;
      overflow: hidden;
      text-overflow: ellipsis;
    }}
    .week-block {{
      position: absolute;
      top: 24px;
      height: 32px;
      border-right: 1px solid #d8e3eb;
      padding: 4px 6px;
      font-size: 0.72rem;
      color: #3b5565;
      white-space: nowrap;
      overflow: hidden;
      text-overflow: ellipsis;
      background: rgba(255,255,255,0.62);
    }}
    .week-block.current-week {{
      background: rgba(251, 191, 36, 0.28);
      border: 1px solid rgba(180, 83, 9, 0.45);
      color: #7c2d12;
      font-weight: 700;
      z-index: 2;
    }}
    .grid-line {{
      position: absolute;
      top: 0;
      bottom: 0;
      width: 1px;
      background: #e8eef3;
      pointer-events: none;
    }}
    .week-line {{
      position: absolute;
      top: 0;
      bottom: 0;
      width: 1px;
      background: #d8e2ea;
      pointer-events: none;
    }}
    .current-week-band {{
      position: absolute;
      top: 0;
      bottom: 0;
      background: rgba(251, 191, 36, 0.13);
      border-left: 1px solid rgba(180, 83, 9, 0.4);
      border-right: 1px solid rgba(180, 83, 9, 0.4);
      pointer-events: none;
      z-index: 1;
    }}
    .lane-empty {{
      padding: 10px;
      font-size: 0.8rem;
      color: #7b8c99;
    }}
    .card {{
      position: absolute;
      min-height: 54px;
      border-radius: 9px;
      border: 1px solid rgba(15, 76, 92, 0.32);
      background: linear-gradient(180deg, rgba(219, 238, 247, 0.82), rgba(233, 246, 253, 0.95));
      box-shadow: 0 2px 8px rgba(16, 43, 58, 0.08);
      padding: 6px 7px;
      overflow: hidden;
    }}
    .card-title {{
      font-size: 0.76rem;
      color: #103547;
      font-weight: 700;
      white-space: nowrap;
      overflow: hidden;
      text-overflow: ellipsis;
      margin-bottom: 4px;
    }}
    .card-meta {{
      font-size: 0.68rem;
      color: #375567;
      line-height: 1.32;
      white-space: nowrap;
      overflow: hidden;
      text-overflow: ellipsis;
    }}
    .empty {{
      padding: 16px;
      color: #607282;
      font-size: 0.9rem;
    }}
  </style>
  <link rel="stylesheet" href="shared-nav.css">
</head>
<body>
  <div class="page">
    <section class="panel">
      <h1 class="title">Phase Owner RMI Gantt</h1>
      <p class="meta">Generated: <span id="generated-at"></span> | Source: <span id="source-file"></span> | RMIs in view: <span id="visible-count"></span></p>
      <div class="controls">
        <div class="control">
          <label for="from-date">From</label>
          <input type="date" id="from-date">
        </div>
        <div class="control">
          <label for="to-date">To</label>
          <input type="date" id="to-date">
        </div>
        <button class="btn alt" type="button" id="apply-range">Apply</button>
        <button class="btn" type="button" id="reset-range">Reset</button>
      </div>
      <div class="legend">
        <span class="legend-pill"><span class="chip-dot"></span>Mini cards show planned phase workload by RMI</span>
        <span class="legend-pill">Sticky phase lanes + sticky weekly header + horizontal scroll</span>
      </div>
    </section>
    <section class="gantt-wrap" id="gantt-wrap">
      <div id="gantt-root" class="gantt-root"></div>
    </section>
  </div>
  <script>
    const reportData = {payload};
    const allItems = Array.isArray(reportData.items) ? reportData.items : [];
    const phaseNames = Array.isArray(reportData.phase_names) ? reportData.phase_names : [];

    const generatedNode = document.getElementById("generated-at");
    const sourceNode = document.getElementById("source-file");
    const visibleCountNode = document.getElementById("visible-count");
    const fromInput = document.getElementById("from-date");
    const toInput = document.getElementById("to-date");
    const applyButton = document.getElementById("apply-range");
    const resetButton = document.getElementById("reset-range");
    const ganttRoot = document.getElementById("gantt-root");

    const DAY_MS = 86400000;
    const DAY_PX = 16;
    const CARD_HEIGHT = 58;
    const CARD_GAP = 8;
    const TIMELINE_SIDE_PAD = 8;

    generatedNode.textContent = reportData.generated_at || "-";
    sourceNode.textContent = reportData.source_file || "-";

    function parseIso(iso) {{
      if (!iso) return null;
      const d = new Date(`${{iso}}T00:00:00`);
      if (Number.isNaN(d.getTime())) return null;
      return d;
    }}

    function isoFromDate(d) {{
      const y = d.getFullYear();
      const m = String(d.getMonth() + 1).padStart(2, "0");
      const day = String(d.getDate()).padStart(2, "0");
      return `${{y}}-${{m}}-${{day}}`;
    }}

    function addDays(d, days) {{
      return new Date(d.getTime() + (days * DAY_MS));
    }}

    function startOfMonth(d) {{
      return new Date(d.getFullYear(), d.getMonth(), 1);
    }}

    function endOfMonth(d) {{
      return new Date(d.getFullYear(), d.getMonth() + 1, 0);
    }}

    function defaultRange() {{
      const today = new Date();
      const currentMonthStart = startOfMonth(today);
      const prevMonthStart = startOfMonth(new Date(today.getFullYear(), today.getMonth() - 1, 1));
      const nextMonthEnd = endOfMonth(new Date(today.getFullYear(), today.getMonth() + 1, 1));
      return {{ from: prevMonthStart, to: nextMonthEnd, currentMonthStart }};
    }}

    function overlap(aStart, aEnd, bStart, bEnd) {{
      return aStart <= bEnd && aEnd >= bStart;
    }}

    function daysBetweenInclusive(startDate, endDate) {{
      return Math.max(1, Math.floor((endDate.getTime() - startDate.getTime()) / DAY_MS) + 1);
    }}

    function pctFromDays(days, totalDays) {{
      if (totalDays <= 0) return 0;
      return (days / totalDays) * 100;
    }}

    function formatDate(iso) {{
      const d = parseIso(iso);
      if (!d) return "-";
      return d.toLocaleDateString(undefined, {{ day: "2-digit", month: "short", year: "numeric" }});
    }}

    function formatShortDate(d) {{
      return d.toLocaleDateString(undefined, {{ day: "2-digit", month: "short" }});
    }}

    function monthLabel(d) {{
      return d.toLocaleDateString(undefined, {{ month: "short", year: "numeric" }});
    }}

    function weekStartFor(dateObj) {{
      const out = new Date(dateObj.getTime());
      const day = out.getDay();
      const shift = day === 0 ? 6 : day - 1;
      out.setDate(out.getDate() - shift);
      out.setHours(0, 0, 0, 0);
      return out;
    }}

    function buildWeeks(rangeFrom, rangeTo) {{
      const today = new Date();
      today.setHours(0, 0, 0, 0);
      const weeks = [];
      let cursor = weekStartFor(rangeFrom);
      while (cursor <= rangeTo) {{
        const start = new Date(cursor.getTime());
        const end = addDays(start, 6);
        const clippedStart = start < rangeFrom ? rangeFrom : start;
        const clippedEnd = end > rangeTo ? rangeTo : end;
        const totalDays = daysBetweenInclusive(rangeFrom, rangeTo);
        const leftDays = Math.floor((clippedStart.getTime() - rangeFrom.getTime()) / DAY_MS);
        const spanDays = daysBetweenInclusive(clippedStart, clippedEnd);
        weeks.push({{
          start,
          end,
          clippedStart,
          clippedEnd,
          isCurrent: today >= start && today <= end,
          leftPct: pctFromDays(leftDays, totalDays),
          widthPct: pctFromDays(spanDays, totalDays),
          label: `Wk of ${{formatShortDate(start)}}`,
        }});
        cursor = addDays(cursor, 7);
      }}
      return weeks;
    }}

    function buildMonths(rangeFrom, rangeTo) {{
      const months = [];
      let cursor = startOfMonth(rangeFrom);
      const totalDays = daysBetweenInclusive(rangeFrom, rangeTo);
      while (cursor <= rangeTo) {{
        const monthStart = new Date(cursor.getTime());
        const monthEnd = endOfMonth(cursor);
        const clippedStart = monthStart < rangeFrom ? rangeFrom : monthStart;
        const clippedEnd = monthEnd > rangeTo ? rangeTo : monthEnd;
        const leftDays = Math.floor((clippedStart.getTime() - rangeFrom.getTime()) / DAY_MS);
        const spanDays = daysBetweenInclusive(clippedStart, clippedEnd);
        months.push({{
          label: monthLabel(monthStart),
          leftPct: pctFromDays(leftDays, totalDays),
          widthPct: pctFromDays(spanDays, totalDays),
        }});
        cursor = new Date(cursor.getFullYear(), cursor.getMonth() + 1, 1);
      }}
      return months;
    }}

    function parseRangeFromInputs() {{
      const preset = defaultRange();
      const from = parseIso(fromInput.value) || preset.from;
      const to = parseIso(toInput.value) || preset.to;
      if (to < from) {{
        return {{ from: preset.from, to: preset.to }};
      }}
      return {{ from, to }};
    }}

    function injectInputs(from, to) {{
      fromInput.value = isoFromDate(from);
      toInput.value = isoFromDate(to);
    }}

    function filteredItems(rangeFrom, rangeTo) {{
      return allItems.filter((item) => {{
        const start = parseIso(item.planned_start);
        const end = parseIso(item.planned_end);
        if (!start || !end || end < start) return false;
        return overlap(start, end, rangeFrom, rangeTo);
      }});
    }}

    function stackCards(items) {{
      const sorted = [...items].sort((a, b) => {{
        if (a.planned_start !== b.planned_start) return a.planned_start < b.planned_start ? -1 : 1;
        if (a.planned_end !== b.planned_end) return a.planned_end < b.planned_end ? -1 : 1;
        return (a.rmi_name || "").localeCompare(b.rmi_name || "");
      }});
      const trackEnds = [];
      return sorted.map((item) => {{
        const start = parseIso(item.planned_start);
        const end = parseIso(item.planned_end);
        let track = 0;
        for (; track < trackEnds.length; track++) {{
          if (trackEnds[track] < start) break;
        }}
        if (track === trackEnds.length) {{
          trackEnds.push(end);
        }} else {{
          trackEnds[track] = end;
        }}
        return {{ ...item, _track: track }};
      }});
    }}

    function laneLoadChips(laneItems, weeks) {{
      if (!weeks.length) return "";
      const weekLoads = weeks.map((week) => {{
        let total = 0;
        for (const item of laneItems) {{
          const s = parseIso(item.planned_start);
          const e = parseIso(item.planned_end);
          if (!s || !e) continue;
          if (overlap(s, e, week.clippedStart, week.clippedEnd)) {{
            total += Number(item.man_days || 0);
          }}
        }}
        return total;
      }});
      const maxLoad = Math.max(...weekLoads, 0);
      return weekLoads.map((load, i) => {{
        const pct = maxLoad > 0 ? (load / maxLoad) : 0;
        const bg = pct <= 0
          ? "#eff4f8"
          : pct < 0.35
            ? "#dbeafe"
            : pct < 0.7
              ? "#93c5fd"
              : "#2563eb";
        const title = `${{weeks[i].label}}: ${{load.toFixed(2).replace(/\\.00$/, "")}} md`;
        return `<span class="load-chip" style="background:${{bg}};" title="${{title}}"></span>`;
      }}).join("");
    }}

    function render() {{
      const range = parseRangeFromInputs();
      injectInputs(range.from, range.to);

      const visible = filteredItems(range.from, range.to);
      visibleCountNode.textContent = String(visible.length);

      const totalDays = daysBetweenInclusive(range.from, range.to);
      const timelineWidth = Math.max(880, Math.floor(totalDays * DAY_PX));
      const months = buildMonths(range.from, range.to);
      const weeks = buildWeeks(range.from, range.to);

      if (!phaseNames.length) {{
        ganttRoot.innerHTML = '<div class="empty">No phase rows found in nested view workbook.</div>';
        return;
      }}

      const monthBlocks = months.map((m) =>
        `<div class="month-block" style="left:${{m.leftPct}}%; width:${{m.widthPct}}%;">${{m.label}}</div>`
      ).join("");
      const weekBlocks = weeks.map((w) =>
        `<div class="week-block${{w.isCurrent ? " current-week" : ""}}" style="left:${{w.leftPct}}%; width:${{w.widthPct}}%;" title="${{w.label}}">${{w.label}}</div>`
      ).join("");
      const weekLines = weeks.map((w) => `<span class="week-line" style="left:${{w.leftPct}}%;"></span>`).join("");
      const currentWeek = weeks.find((w) => w.isCurrent);
      const currentBandHtml = currentWeek
        ? `<span class="current-week-band" style="left:${{currentWeek.leftPct}}%; width:${{currentWeek.widthPct}}%;"></span>`
        : "";

      const laneRows = phaseNames.map((phaseName) => {{
        const laneItems = visible.filter((item) => item.phase_name === phaseName);
        const stacked = stackCards(laneItems);
        const maxTrack = stacked.reduce((acc, item) => Math.max(acc, Number(item._track || 0)), -1);
        const rowCount = Math.max(1, maxTrack + 1);
        const laneHeight = Math.max(48, (rowCount * CARD_HEIGHT) + ((rowCount - 1) * CARD_GAP) + 20);
        const chips = laneLoadChips(laneItems, weeks);

        let cardsHtml = "";
        for (const item of stacked) {{
          const start = parseIso(item.planned_start);
          const end = parseIso(item.planned_end);
          if (!start || !end) continue;

          const clippedStart = start < range.from ? range.from : start;
          const clippedEnd = end > range.to ? range.to : end;
          const leftDays = Math.floor((clippedStart.getTime() - range.from.getTime()) / DAY_MS);
          const spanDays = daysBetweenInclusive(clippedStart, clippedEnd);
          const leftPx = Math.floor(leftDays * DAY_PX) + TIMELINE_SIDE_PAD;
          const widthPx = Math.max(64, Math.floor(spanDays * DAY_PX) - 4);
          const topPx = 8 + (Number(item._track || 0) * (CARD_HEIGHT + CARD_GAP));
          const title = [
            `Phase: ${{item.phase_name}}`,
            `RMI: ${{item.rmi_name}}`,
            `Man Days: ${{String(item.man_days)}}`,
            `Planned Start: ${{item.planned_start}}`,
            `Planned End: ${{item.planned_end}}`,
            `Project: ${{item.project_key || "-"}}`,
          ].join("\\n");

          cardsHtml += `
            <article class="card" style="left:${{leftPx}}px; top:${{topPx}}px; width:${{widthPx}}px;" title="${{title}}">
              <div class="card-title">${{item.rmi_name || "-"}}</div>
              <div class="card-meta">Man Days: ${{String(item.man_days)}}</div>
              <div class="card-meta">Start: ${{item.planned_start || "-"}}</div>
              <div class="card-meta">End: ${{item.planned_end || "-"}}</div>
            </article>
          `;
        }}

        const lineEveryDays = 7;
        const lines = [];
        for (let day = 0; day <= totalDays; day += lineEveryDays) {{
          const leftPx = Math.floor(day * DAY_PX) + TIMELINE_SIDE_PAD;
          lines.push(`<span class="grid-line" style="left:${{leftPx}}px;"></span>`);
        }}
        const linesHtml = lines.join("");

        const contentHtml = stacked.length
          ? `<div class="lane-timeline" style="width:${{timelineWidth}}px; min-height:${{laneHeight}}px;">${{linesHtml}}${{weekLines}}${{currentBandHtml}}${{cardsHtml}}</div>`
          : `<div class="lane-timeline" style="width:${{timelineWidth}}px; min-height:48px;">${{linesHtml}}${{weekLines}}${{currentBandHtml}}<div class="lane-empty">No RMIs in selected range.</div></div>`;

        return `
          <div class="grid-row">
            <div class="cell phase-cell">
              <div class="phase-title" title="${{phaseName}}">${{phaseName}}</div>
              <div class="phase-sub">${{laneItems.length}} RMI card(s)</div>
              <div class="load-row">${{chips}}</div>
            </div>
            <div class="cell timeline-cell">${{contentHtml}}</div>
          </div>
        `;
      }}).join("");

      ganttRoot.innerHTML = `
        <div class="grid" style="width: calc(var(--phase-col) + ${{timelineWidth}}px);">
          <div class="grid-row head">
            <div class="cell phase-cell">Phase</div>
            <div class="cell timeline-cell">
              <div class="timeline-header" style="width:${{timelineWidth}}px;">
                ${{monthBlocks}}
                ${{currentBandHtml}}
                ${{weekBlocks}}
              </div>
            </div>
          </div>
          ${{laneRows}}
        </div>
      `;
    }}

    const defaults = defaultRange();
    injectInputs(defaults.from, defaults.to);
    applyButton.addEventListener("click", render);
    resetButton.addEventListener("click", () => {{
      const d = defaultRange();
      injectInputs(d.from, d.to);
      render();
    }});
    render();
  </script>
<script src="shared-nav.js"></script>
</body>
</html>
"""


def main() -> None:
    base_dir = Path(__file__).resolve().parent
    input_name = os.getenv("JIRA_PHASE_GANTT_INPUT_XLSX_PATH", DEFAULT_INPUT_XLSX).strip() or DEFAULT_INPUT_XLSX
    output_name = os.getenv("JIRA_PHASE_GANTT_HTML_PATH", DEFAULT_OUTPUT_HTML).strip() or DEFAULT_OUTPUT_HTML

    input_path = _resolve_path(input_name, base_dir)
    output_path = _resolve_path(output_name, base_dir)

    phase_rows = _load_phase_rmi_records(input_path)
    aggregated = _aggregate_phase_rmi_rows(phase_rows)
    phase_names = FIXED_PHASE_NAMES

    payload = {
        "generated_at": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC"),
        "source_file": str(input_path),
        "phase_names": phase_names,
        "items": aggregated,
    }
    output_path.write_text(_build_html(payload), encoding="utf-8")

    print(f"Source workbook: {input_path}")
    print(f"Phase story rows loaded: {len(phase_rows)}")
    print(f"Aggregated phase-RMI rows: {len(aggregated)}")
    print(f"Report written: {output_path}")


if __name__ == "__main__":
    main()

