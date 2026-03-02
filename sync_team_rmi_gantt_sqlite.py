from __future__ import annotations

import json
import os
import re
import sqlite3
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from jira_client import BASE_URL

DEFAULT_WORK_ITEMS_XLSX = "1_jira_work_items_export.xlsx"
DEFAULT_CAPACITY_DB = "assignee_hours_capacity.db"
UNMAPPED_TEAM_NAME = "Unmapped Team"


def _resolve_path(value: str, base_dir: Path) -> Path:
    path = Path(value)
    return path if path.is_absolute() else base_dir / path


def _to_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _normalize_person_name(value: Any) -> str:
    text = _to_text(value)
    if not text:
        return ""
    return re.sub(r"\s+", " ", text).strip().casefold()


def _to_iso_date(value: Any) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = _to_text(value)
    if not text:
        return ""
    if len(text) >= 10:
        try:
            date.fromisoformat(text[:10])
            return text[:10]
        except ValueError:
            pass
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%d-%b-%Y", "%d-%B-%Y", "%m/%d/%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00")).date().isoformat()
    except ValueError:
        return ""


def _to_positive_float(value: Any) -> float:
    try:
        out = float(value or 0)
    except (TypeError, ValueError):
        return 0.0
    return out if out > 0 else 0.0


def _extract_project_key(issue_key: str) -> str:
    text = _to_text(issue_key).upper()
    if "-" not in text:
        return "UNKNOWN"
    left = text.split("-", 1)[0].strip()
    return left if left else "UNKNOWN"


def _fallback_epic_url(epic_key: str) -> str:
    key = _to_text(epic_key).upper()
    if not key:
        return ""
    base = _to_text(BASE_URL).rstrip("/")
    if not base:
        return ""
    return f"{base}/browse/{key}"


def _ensure_tables(conn: sqlite3.Connection) -> None:
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS team_rmi_gantt_items (
            team_name TEXT NOT NULL,
            epic_key TEXT NOT NULL,
            epic_name TEXT NOT NULL,
            epic_url TEXT NOT NULL,
            project_key TEXT NOT NULL,
            planned_start TEXT NOT NULL,
            planned_end TEXT NOT NULL,
            planned_hours REAL NOT NULL,
            planned_man_days REAL NOT NULL,
            story_count INTEGER NOT NULL,
            is_unmapped_team INTEGER NOT NULL DEFAULT 0,
            snapshot_utc TEXT NOT NULL,
            UNIQUE(team_name, epic_key, snapshot_utc)
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS team_rmi_gantt_snapshot_meta (
            id INTEGER PRIMARY KEY CHECK(id = 1),
            snapshot_utc TEXT NOT NULL,
            source_work_items_path TEXT NOT NULL,
            total_story_rows INTEGER NOT NULL,
            included_story_rows INTEGER NOT NULL,
            excluded_missing_epic INTEGER NOT NULL,
            excluded_missing_dates INTEGER NOT NULL,
            excluded_missing_estimate INTEGER NOT NULL
        )
        """
    )


def _load_team_map(db_path: Path) -> dict[str, str]:
    if not db_path.exists():
        return {}
    conn = sqlite3.connect(db_path)
    try:
        exists = conn.execute(
            "SELECT 1 FROM sqlite_master WHERE type='table' AND name='performance_teams'"
        ).fetchone()
        if not exists:
            return {}
        rows = conn.execute("SELECT team_name, assignees_json FROM performance_teams").fetchall()
    finally:
        conn.close()

    out: dict[str, str] = {}
    for team_name, assignees_json in rows:
        name = _to_text(team_name)
        if not name:
            continue
        try:
            assignees = json.loads(_to_text(assignees_json) or "[]")
        except json.JSONDecodeError:
            assignees = []
        if not isinstance(assignees, list):
            continue
        for assignee in assignees:
            key = _normalize_person_name(assignee)
            if key and key not in out:
                out[key] = name
    return out


def _load_work_items_for_team_gantt(work_items_path: Path) -> tuple[dict[str, dict[str, str]], list[dict[str, Any]]]:
    if not work_items_path.exists():
        raise FileNotFoundError(f"Work items workbook not found: {work_items_path}")
    wb = load_workbook(work_items_path, read_only=True, data_only=True)
    try:
        ws = wb.active
        header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header:
            return {}, []
        headers = [_to_text(h) for h in header]
        idx = {name: i for i, name in enumerate(headers)}
        required = ["issue_key", "jira_issue_type", "summary", "assignee", "parent_issue_key", "start_date", "end_date", "original_estimate_hours", "jira_url", "project_key"]
        missing = [name for name in required if name not in idx]
        if missing:
            raise ValueError(f"Work items workbook missing required columns: {missing}")

        epics: dict[str, dict[str, str]] = {}
        stories: list[dict[str, Any]] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            issue_key = _to_text(row[idx["issue_key"]]).upper()
            issue_type = _to_text(row[idx["jira_issue_type"]]).lower()
            if not issue_key:
                continue

            if "epic" in issue_type:
                epic_summary = _to_text(row[idx["summary"]])
                epic_url = _to_text(row[idx["jira_url"]])
                project_key = _to_text(row[idx["project_key"]]).upper() or _extract_project_key(issue_key)
                epics[issue_key] = {
                    "epic_name": epic_summary or issue_key,
                    "epic_url": epic_url or _fallback_epic_url(issue_key),
                    "project_key": project_key or "UNKNOWN",
                }
                continue

            if "story" not in issue_type:
                continue
            stories.append(
                {
                    "issue_key": issue_key,
                    "assignee": _to_text(row[idx["assignee"]]),
                    "epic_key": _to_text(row[idx["parent_issue_key"]]).upper(),
                    "planned_start": _to_iso_date(row[idx["start_date"]]),
                    "planned_end": _to_iso_date(row[idx["end_date"]]),
                    "planned_hours": _to_positive_float(row[idx["original_estimate_hours"]]),
                    "project_key": _to_text(row[idx["project_key"]]).upper() or _extract_project_key(issue_key),
                }
            )
        return epics, stories
    finally:
        wb.close()


def build_team_rmi_gantt_snapshot(work_items_path: Path, db_path: Path) -> dict[str, Any]:
    team_map = _load_team_map(db_path)
    epics, stories = _load_work_items_for_team_gantt(work_items_path)

    total_story_rows = 0
    included_story_rows = 0
    excluded_missing_epic = 0
    excluded_missing_dates = 0
    excluded_missing_estimate = 0

    grouped: dict[tuple[str, str], dict[str, Any]] = {}
    for story in stories:
        total_story_rows += 1
        epic_key = _to_text(story.get("epic_key")).upper()
        if not epic_key:
            excluded_missing_epic += 1
            continue
        planned_start = _to_text(story.get("planned_start"))
        planned_end = _to_text(story.get("planned_end"))
        if not planned_start or not planned_end or planned_end < planned_start:
            excluded_missing_dates += 1
            continue
        planned_hours = _to_positive_float(story.get("planned_hours"))
        if planned_hours <= 0:
            excluded_missing_estimate += 1
            continue

        assignee_key = _normalize_person_name(story.get("assignee"))
        team_name = team_map.get(assignee_key) or UNMAPPED_TEAM_NAME
        is_unmapped_team = 1 if team_name == UNMAPPED_TEAM_NAME else 0

        epic = epics.get(epic_key, {})
        epic_name = _to_text(epic.get("epic_name")) or epic_key
        epic_url = _to_text(epic.get("epic_url")) or _fallback_epic_url(epic_key)
        project_key = _to_text(epic.get("project_key")) or _to_text(story.get("project_key")) or _extract_project_key(epic_key)

        key = (team_name, epic_key)
        current = grouped.get(key)
        if not current:
            grouped[key] = {
                "team_name": team_name,
                "epic_key": epic_key,
                "epic_name": epic_name,
                "epic_url": epic_url,
                "project_key": project_key,
                "planned_start": planned_start,
                "planned_end": planned_end,
                "planned_hours": planned_hours,
                "story_count": 1,
                "is_unmapped_team": is_unmapped_team,
            }
        else:
            current["planned_hours"] = round(float(current["planned_hours"]) + planned_hours, 4)
            current["story_count"] = int(current["story_count"]) + 1
            if planned_start < _to_text(current.get("planned_start")):
                current["planned_start"] = planned_start
            if planned_end > _to_text(current.get("planned_end")):
                current["planned_end"] = planned_end
            if not _to_text(current.get("epic_url")) and epic_url:
                current["epic_url"] = epic_url
            if not _to_text(current.get("epic_name")) and epic_name:
                current["epic_name"] = epic_name

        included_story_rows += 1

    snapshot_utc = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
    items: list[dict[str, Any]] = []
    for item in grouped.values():
        planned_hours = round(float(item["planned_hours"]), 2)
        planned_man_days = round(planned_hours / 8.0, 2)
        items.append(
            {
                "team_name": _to_text(item["team_name"]),
                "epic_key": _to_text(item["epic_key"]),
                "epic_name": _to_text(item["epic_name"]),
                "epic_url": _to_text(item["epic_url"]),
                "project_key": _to_text(item["project_key"]),
                "planned_start": _to_text(item["planned_start"]),
                "planned_end": _to_text(item["planned_end"]),
                "planned_hours": planned_hours,
                "planned_man_days": planned_man_days,
                "story_count": int(item["story_count"]),
                "is_unmapped_team": int(item["is_unmapped_team"]),
                "snapshot_utc": snapshot_utc,
            }
        )
    items.sort(key=lambda x: (_to_text(x["team_name"]).lower(), _to_text(x["epic_name"]).lower(), _to_text(x["epic_key"]).lower()))

    return {
        "snapshot_utc": snapshot_utc,
        "source_work_items_path": str(work_items_path),
        "total_story_rows": total_story_rows,
        "included_story_rows": included_story_rows,
        "excluded_missing_epic": excluded_missing_epic,
        "excluded_missing_dates": excluded_missing_dates,
        "excluded_missing_estimate": excluded_missing_estimate,
        "items": items,
    }


def write_team_rmi_gantt_snapshot(db_path: Path, snapshot: dict[str, Any]) -> None:
    db_path.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(db_path)
    try:
        _ensure_tables(conn)
        with conn:
            conn.execute("DELETE FROM team_rmi_gantt_items")
            conn.execute("DELETE FROM team_rmi_gantt_snapshot_meta WHERE id = 1")
            for item in snapshot.get("items", []):
                conn.execute(
                    """
                    INSERT INTO team_rmi_gantt_items (
                        team_name, epic_key, epic_name, epic_url, project_key,
                        planned_start, planned_end, planned_hours, planned_man_days, story_count,
                        is_unmapped_team, snapshot_utc
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        _to_text(item.get("team_name")),
                        _to_text(item.get("epic_key")),
                        _to_text(item.get("epic_name")),
                        _to_text(item.get("epic_url")),
                        _to_text(item.get("project_key")),
                        _to_text(item.get("planned_start")),
                        _to_text(item.get("planned_end")),
                        float(item.get("planned_hours") or 0.0),
                        float(item.get("planned_man_days") or 0.0),
                        int(item.get("story_count") or 0),
                        int(item.get("is_unmapped_team") or 0),
                        _to_text(item.get("snapshot_utc")),
                    ),
                )
            conn.execute(
                """
                INSERT INTO team_rmi_gantt_snapshot_meta (
                    id, snapshot_utc, source_work_items_path, total_story_rows, included_story_rows,
                    excluded_missing_epic, excluded_missing_dates, excluded_missing_estimate
                ) VALUES (1, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    _to_text(snapshot.get("snapshot_utc")),
                    _to_text(snapshot.get("source_work_items_path")),
                    int(snapshot.get("total_story_rows") or 0),
                    int(snapshot.get("included_story_rows") or 0),
                    int(snapshot.get("excluded_missing_epic") or 0),
                    int(snapshot.get("excluded_missing_dates") or 0),
                    int(snapshot.get("excluded_missing_estimate") or 0),
                ),
            )
    finally:
        conn.close()


def sync_team_rmi_gantt_to_sqlite(work_items_path: Path, db_path: Path) -> dict[str, Any]:
    snapshot = build_team_rmi_gantt_snapshot(work_items_path, db_path)
    write_team_rmi_gantt_snapshot(db_path, snapshot)
    return snapshot


def main() -> None:
    base_dir = Path(__file__).resolve().parent
    work_items_name = os.getenv("JIRA_EXPORT_XLSX_PATH", DEFAULT_WORK_ITEMS_XLSX).strip() or DEFAULT_WORK_ITEMS_XLSX
    db_name = os.getenv("JIRA_ASSIGNEE_HOURS_CAPACITY_DB_PATH", DEFAULT_CAPACITY_DB).strip() or DEFAULT_CAPACITY_DB
    work_items_path = _resolve_path(work_items_name, base_dir)
    db_path = _resolve_path(db_name, base_dir)

    snapshot = sync_team_rmi_gantt_to_sqlite(work_items_path, db_path)
    print(f"Source workbook: {work_items_path}")
    print(f"Capacity DB: {db_path}")
    print(f"Total story rows: {snapshot['total_story_rows']}")
    print(f"Included story rows: {snapshot['included_story_rows']}")
    print(f"Excluded (missing epic): {snapshot['excluded_missing_epic']}")
    print(f"Excluded (missing/invalid dates): {snapshot['excluded_missing_dates']}")
    print(f"Excluded (missing estimate): {snapshot['excluded_missing_estimate']}")
    print(f"Aggregated team-epic rows: {len(snapshot.get('items', []))}")


if __name__ == "__main__":
    main()
