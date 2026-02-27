"""
Transform IPP Meeting workbook phase planned-range fields into normalized columns.
"""
from __future__ import annotations

import argparse
import json
import os
import re
from datetime import date, datetime, timedelta, timezone
from pathlib import Path

from openpyxl import Workbook, load_workbook

DEFAULT_INPUT_XLSX = (
    r"C:\Users\hmalik\OneDrive - Octopus Digital\ALL DOCS\IPP Meeting\all ipp meetings.xlsx"
)
DEFAULT_FIXED_OUTPUT_NAME = "ipp_phase_breakdown.xlsx"
DEFAULT_OUTPUT_MODE = "timestamped"
DEFAULT_DATA_SHEET_NAME = "IPP Phase Breakdown"
DEFAULT_COMPUTED_SHEET_NAME = "IPP Dashboard Computed"
DEFAULT_METADATA_SHEET_NAME = "Metadata"
COMPUTATION_RULES_VERSION = "v2"

PHASE_COLUMNS = [
    ("Research/URS", "Research/URS - Planned Range"),
    ("DDS", "DDS - Planned Range"),
    ("Development", "Development - Planned Range"),
    ("SQA", "SQA - Planned Range"),
    ("User Manual", "User Manual - Planned Range"),
    ("Production", "Production - Planned Range"),
]

DATE_TOKEN_RE = re.compile(r"\b\d{4}-\d{2}-\d{2}\b|\b\d{1,2}[-/]\w{3}[-/]\d{2,4}\b", re.IGNORECASE)
ISSUE_KEY_RE = re.compile(r"\b([A-Za-z][A-Za-z0-9]+-\d+)\b")
MANDAYS_RE = re.compile(r"([0-9]+(?:\.[0-9]+)?)\s*man\s*-?\s*days?\b", re.IGNORECASE)
LATEST_SHEET_RE = re.compile(r"^\s*latest\s*-\s*(\d{1,2}\s+[A-Za-z]{3}\s+\d{4})\s*$", re.IGNORECASE)
DATE_SHEET_RE = re.compile(r"^\s*(\d{1,2}\s+[A-Za-z]{3}\s+\d{4})\s*$", re.IGNORECASE)

REQUIRED_BASE_HEADERS = [
    "Product",
    "Epic/RMI",
    "Jira Task ID",
    "Planned Start Date",
    "Planned End Date",
    "Actual Date (Production Date)",
]

ROADMAP_SIDE_PADDING_DAYS = 45
ROADMAP_TICK_DAYS = 15
MINI_SIDE_PADDING_DAYS = 10
MINI_TICK_DAYS = 7
MINI_SCROLL_THRESHOLD_DAYS = 7
MINI_TIMELINE_BASE_WIDTH = 520
MINI_TIMELINE_PX_PER_DAY = 24
MIN_BAR_THICKNESS = 8.0
MAX_BAR_THICKNESS = 22.0
MID_BAR_THICKNESS = 15.0
TRACK_VERTICAL_PADDING = 4.0
TRACK_BASE_HEIGHT = MAX_BAR_THICKNESS + (TRACK_VERTICAL_PADDING * 2.0)
SMALL_MIN_WIDTH_PCT = 0.8


def _normalize_header(text: str) -> str:
    return " ".join(str(text or "").strip().lower().split())


def _parse_sheet_date(text: str) -> date | None:
    cleaned = str(text or "").strip()
    for fmt in ("%d %b %Y", "%d %B %Y"):
        try:
            return datetime.strptime(cleaned, fmt).date()
        except ValueError:
            continue
    return None


def _select_source_sheet(sheet_names: list[str], explicit_sheet: str = "") -> str:
    if explicit_sheet:
        if explicit_sheet not in sheet_names:
            raise ValueError(f"Source sheet not found: {explicit_sheet}")
        return explicit_sheet

    candidates: list[tuple[date, str, int]] = []
    for name in sheet_names:
        latest_match = LATEST_SHEET_RE.match(name)
        if latest_match:
            parsed = _parse_sheet_date(latest_match.group(1))
            if parsed:
                candidates.append((parsed, name, 2))
                continue

        plain_match = DATE_SHEET_RE.match(name)
        if plain_match:
            parsed = _parse_sheet_date(plain_match.group(1))
            if parsed:
                candidates.append((parsed, name, 1))

    if candidates:
        candidates.sort(key=lambda item: (item[0], item[2]))
        return candidates[-1][1]
    return sheet_names[0]


def _as_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def _parse_date_token(token: str) -> str:
    text = (token or "").strip()
    if not text:
        return ""
    for fmt in ("%Y-%m-%d", "%d-%b-%Y", "%d/%b/%Y", "%d-%b-%y", "%d/%b/%y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    return ""


def _parse_iso_date(text: str) -> date | None:
    value = _as_text(text)
    if not value:
        return None
    for fmt in ("%Y-%m-%d", "%d-%b-%Y", "%d/%b/%Y", "%d-%b-%y", "%d/%b/%y"):
        try:
            return datetime.strptime(value, fmt).date()
        except ValueError:
            continue
    return None


def _normalize_iso_date(value) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = _as_text(value)
    if not text:
        return ""
    direct = _parse_iso_date(text)
    if direct:
        return direct.isoformat()
    for token in DATE_TOKEN_RE.findall(text):
        parsed = _parse_date_token(token)
        if parsed:
            return parsed
    return ""


def _extract_issue_key(value: str) -> str:
    match = ISSUE_KEY_RE.search(str(value or ""))
    return match.group(1).upper() if match else ""


def _jira_base_url() -> str:
    site = os.getenv("JIRA_SITE", "octopusdtlsupport").strip() or "octopusdtlsupport"
    return f"https://{site}.atlassian.net"


def _normalize_jira_link(value) -> str:
    text = _as_text(value)
    if not text:
        return ""
    if text.lower().startswith("http://") or text.lower().startswith("https://"):
        return text
    key = _extract_issue_key(text)
    if key:
        return f"{_jira_base_url()}/browse/{key}"
    return text


def _to_number(value) -> float | None:
    text = _as_text(value)
    if not text:
        return None
    try:
        out = float(text)
    except ValueError:
        return None
    return out if out == out else None


def _display_date(value: date | None) -> str:
    if value is None:
        return "-"
    return value.strftime("%d-%b-%Y")


def _clamp_percent(value: float) -> float:
    if value != value:
        return 0.0
    return max(0.0, min(100.0, value))


def _build_week_ticks(axis_start: date, axis_end: date, step_days: int) -> list[dict[str, object]]:
    span_days = max(1, (axis_end - axis_start).days + 1)
    ticks: list[dict[str, object]] = []
    cursor = axis_start
    while cursor <= axis_end:
        left_pct = _clamp_percent(((cursor - axis_start).days / max(1, span_days - 1)) * 100.0)
        ticks.append({"iso": cursor.isoformat(), "label": _display_date(cursor), "left_pct": round(left_pct, 4)})
        cursor = cursor + timedelta(days=step_days)
    return ticks


def _set_warning(result: dict[str, str], warnings: list[str]) -> dict[str, str]:
    if warnings:
        deduped = []
        seen = set()
        for item in warnings:
            if item not in seen:
                seen.add(item)
                deduped.append(item)
        result["warning"] = ", ".join(deduped)
    return result


def _parse_phase_cell(raw_value) -> dict[str, str]:
    raw = _as_text(raw_value)
    lowered = raw.lower()
    result = {"start": "", "end": "", "mandays": "", "raw": raw, "state": "no_entry", "warning": ""}
    if not raw:
        return result

    if "skip" in lowered:
        result["state"] = "skipped"
        return result

    has_not_planned = "not planned" in lowered
    date_tokens = DATE_TOKEN_RE.findall(raw)
    parsed_dates = [_parse_date_token(token) for token in date_tokens]
    parsed_dates = [item for item in parsed_dates if item]

    if len(parsed_dates) >= 2:
        result["start"] = parsed_dates[0]
        result["end"] = parsed_dates[1]
    elif len(parsed_dates) == 1:
        result["start"] = parsed_dates[0]

    manday_match = MANDAYS_RE.search(raw)
    if manday_match:
        result["mandays"] = manday_match.group(1)

    warnings: list[str] = []
    if has_not_planned:
        result["state"] = "not_planned"
        if date_tokens and len(parsed_dates) < 2:
            warnings.append("missing_or_invalid_date_range")
        return _set_warning(result, warnings)

    result["state"] = "planned"
    if len(parsed_dates) < 2:
        warnings.append("missing_or_invalid_date_range")
    if not result["mandays"]:
        warnings.append("missing_or_non_numeric_mandays")
    if result["start"] and result["end"] and result["start"] > result["end"]:
        warnings.append("start_after_end")
    if warnings:
        result["state"] = "invalid"
    return _set_warning(result, warnings)


def _row_has_payload(base_values: dict[str, str], phase_payloads: list[dict[str, str]]) -> bool:
    if any(str(value or "").strip() for value in base_values.values()):
        return True
    for payload in phase_payloads:
        if any(str(payload.get(key, "")).strip() for key in ("raw", "start", "end", "mandays", "warning")):
            return True
    return False


def _build_output_headers() -> list[str]:
    headers = [
        "Source Sheet", "Row Number", "Product", "Epic/RMI", "Epic/RMI Jira Link",
        "Epic Planned Start Date", "Epic Planned End Date", "Epic Actual Date (Production Date)", "Remarks",
    ]
    for phase_name, _ in PHASE_COLUMNS:
        headers.extend([
            f"{phase_name} Planned Start Date", f"{phase_name} Planned End Date", f"{phase_name} Planned Man-days",
            f"{phase_name} Raw Planned Range", f"{phase_name} Planning State", f"{phase_name} Parse Warning",
        ])
    headers.extend([
        "Any Phase Parse Warning", "Any No Entry", "Any Not Planned", "Any Skipped",
        "Computed Total Phase Man-days", "Computed Has Valid Epic Plan",
        "Computed Epic Planned Start ISO", "Computed Epic Planned End ISO", "Computed Epic Actual ISO",
    ])
    return headers


def _build_computed_headers() -> list[str]:
    return [
        "Source Sheet", "Row Number", "Product", "Epic/RMI", "Epic/RMI Jira Link",
        "Epic Planned Start Date", "Epic Planned End Date", "Epic Actual Date (Production Date)", "Remarks",
        "Computed Total Phase Man-days", "Computed Roadmap Valid",
        "Computed Roadmap Axis Start ISO", "Computed Roadmap Axis End ISO", "Computed Roadmap Axis Span Days",
        "Computed Roadmap Today In Range", "Computed Roadmap Today Left Pct",
        "Computed Roadmap Bar Left Pct", "Computed Roadmap Bar Width Pct", "Computed Roadmap Actual Left Pct",
        "Computed Roadmap Week Ticks JSON", "Computed MiniGantt Has Dated Phases",
        "Computed MiniGantt Axis Start ISO", "Computed MiniGantt Axis End ISO", "Computed MiniGantt Axis Span Days",
        "Computed MiniGantt Timeline Width Px", "Computed MiniGantt Scroll Enabled", "Computed MiniGantt Week Ticks JSON",
        "Computed MiniGantt Today In Range", "Computed MiniGantt Today Left Pct", "Computed Phase Geometry JSON",
    ]

def _resolve_input_headers(header_row: list[str]) -> dict[str, int]:
    normalized_to_index = {_normalize_header(value): idx for idx, value in enumerate(header_row)}
    missing = []
    resolved: dict[str, int] = {}

    for header_name in REQUIRED_BASE_HEADERS + [item[1] for item in PHASE_COLUMNS]:
        index = normalized_to_index.get(_normalize_header(header_name))
        if index is None:
            missing.append(header_name)
        else:
            resolved[header_name] = index

    remarks_index = normalized_to_index.get(_normalize_header("Remarks"))
    if remarks_index is not None:
        resolved["Remarks"] = remarks_index

    if missing:
        raise ValueError("Missing required columns: " + ", ".join(missing))
    return resolved


def _compute_roadmap_axis(records: list[dict[str, object]]) -> dict[str, object]:
    dated = [
        r for r in records
        if r.get("computed_has_valid_epic_plan") == "Yes"
        and isinstance(r.get("epic_start_date"), date)
        and isinstance(r.get("epic_end_date"), date)
    ]
    if not dated:
        return {
            "has_axis": False,
            "axis_start": None,
            "axis_end": None,
            "axis_span_days": 0,
            "today_in_range": False,
            "today_left_pct": "",
            "week_ticks": [],
        }

    min_start = min(r["epic_start_date"] for r in dated)
    max_end = max(r["epic_end_date"] for r in dated)
    axis_start = min_start - timedelta(days=ROADMAP_SIDE_PADDING_DAYS)
    axis_end = max_end + timedelta(days=ROADMAP_SIDE_PADDING_DAYS)
    axis_span_days = max(1, (axis_end - axis_start).days + 1)

    today = datetime.now(timezone.utc).date()
    today_in_range = axis_start <= today <= axis_end
    today_left_pct = ""
    if today_in_range:
        today_left_pct = round(_clamp_percent(((today - axis_start).days / max(1, axis_span_days - 1)) * 100.0), 4)

    return {
        "has_axis": True,
        "axis_start": axis_start,
        "axis_end": axis_end,
        "axis_span_days": axis_span_days,
        "today_in_range": today_in_range,
        "today_left_pct": today_left_pct,
        "week_ticks": _build_week_ticks(axis_start, axis_end, ROADMAP_TICK_DAYS),
    }


def _compute_phase_geometry_for_record(record: dict[str, object], global_max_mandays: float) -> dict[str, object]:
    phases_by_name = {p["name"]: p for p in record["phases"]}
    drawable = [
        p for p in record["phases"]
        if isinstance(p.get("start_date"), date)
        and isinstance(p.get("end_date"), date)
        and p["start_date"] <= p["end_date"]
    ]

    if not drawable:
        empty_phases = {}
        for phase_name, _ in PHASE_COLUMNS:
            p = phases_by_name[phase_name]
            empty_phases[phase_name] = {
                "state": p["state"], "state_label": p["state_label"], "warning": p["warning"],
                "start_iso": p["start_iso"], "end_iso": p["end_iso"],
                "mandays_text": p["mandays_text"], "mandays_num": p["mandays_num"], "raw": p["raw"],
                "valid": False, "bar_left_pct": "", "bar_width_pct": "", "bar_thickness_px": "",
                "bar_top_offset_px": "", "start_label": "-", "end_label": "-",
                "bar_label": f"{p['mandays_text']} md" if p["mandays_text"] else "-", "show_no_bar": True,
            }
        return {
            "has_dated_phases": False,
            "axis_start": None,
            "axis_end": None,
            "axis_span_days": 0,
            "timeline_width_px": MINI_TIMELINE_BASE_WIDTH,
            "scroll_enabled": False,
            "week_ticks": [],
            "today_in_range": False,
            "today_left_pct": "",
            "phases": empty_phases,
        }

    min_start = min(p["start_date"] for p in drawable)
    max_end = max(p["end_date"] for p in drawable)
    raw_span_days = max(1, (max_end - min_start).days + 1)
    axis_start = min_start - timedelta(days=MINI_SIDE_PADDING_DAYS)
    axis_end = max_end + timedelta(days=MINI_SIDE_PADDING_DAYS)
    axis_span_days = max(1, (axis_end - axis_start).days + 1)

    scroll_enabled = raw_span_days > MINI_SCROLL_THRESHOLD_DAYS
    timeline_width_px = max(MINI_TIMELINE_BASE_WIDTH, axis_span_days * MINI_TIMELINE_PX_PER_DAY) if scroll_enabled else MINI_TIMELINE_BASE_WIDTH

    today = datetime.now(timezone.utc).date()
    today_in_range = axis_start <= today <= axis_end
    today_left_pct = ""
    if today_in_range:
        today_left_pct = round(_clamp_percent(((today - axis_start).days / max(1, axis_span_days - 1)) * 100.0), 4)

    out_phases: dict[str, dict[str, object]] = {}
    for phase_name, _ in PHASE_COLUMNS:
        p = phases_by_name[phase_name]
        is_valid = (
            isinstance(p.get("start_date"), date)
            and isinstance(p.get("end_date"), date)
            and p["start_date"] <= p["end_date"]
        )
        phase_payload = {
            "state": p["state"], "state_label": p["state_label"], "warning": p["warning"],
            "start_iso": p["start_iso"], "end_iso": p["end_iso"],
            "mandays_text": p["mandays_text"], "mandays_num": p["mandays_num"], "raw": p["raw"],
            "valid": is_valid, "bar_left_pct": "", "bar_width_pct": "", "bar_thickness_px": "", "bar_top_offset_px": "",
            "start_label": "-", "end_label": "-", "bar_label": f"{p['mandays_text']} md" if p["mandays_text"] else "-",
            "show_no_bar": not is_valid,
        }

        if is_valid:
            left_pct = _clamp_percent(((p["start_date"] - axis_start).days / max(1, axis_span_days - 1)) * 100.0)
            right_pct = _clamp_percent(((p["end_date"] - axis_start).days / max(1, axis_span_days - 1)) * 100.0)
            width_pct = max(SMALL_MIN_WIDTH_PCT, right_pct - left_pct)

            md_num = p["mandays_num"]
            if md_num is None:
                thickness = MIN_BAR_THICKNESS
            elif global_max_mandays <= 0:
                thickness = MID_BAR_THICKNESS
            else:
                ratio = min(1.0, max(0.0, md_num / global_max_mandays))
                thickness = MIN_BAR_THICKNESS + (ratio * (MAX_BAR_THICKNESS - MIN_BAR_THICKNESS))
            top_offset = max(0.0, (TRACK_BASE_HEIGHT - thickness) / 2.0)

            phase_payload.update({
                "bar_left_pct": round(left_pct, 4), "bar_width_pct": round(width_pct, 4),
                "bar_thickness_px": round(thickness, 4), "bar_top_offset_px": round(top_offset, 4),
                "start_label": _display_date(p["start_date"]), "end_label": _display_date(p["end_date"]), "show_no_bar": False,
            })

        out_phases[phase_name] = phase_payload

    return {
        "has_dated_phases": True,
        "axis_start": axis_start,
        "axis_end": axis_end,
        "axis_span_days": axis_span_days,
        "timeline_width_px": timeline_width_px,
        "scroll_enabled": scroll_enabled,
        "week_ticks": _build_week_ticks(axis_start, axis_end, MINI_TICK_DAYS),
        "today_in_range": today_in_range,
        "today_left_pct": today_left_pct,
        "phases": out_phases,
    }

def _build_output_rows(workbook_path: Path, source_sheet: str) -> tuple[list[str], list[list[object]], list[str], list[list[object]], float]:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    ws = wb[source_sheet]

    header_values = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_values:
        wb.close()
        raise ValueError(f"Source sheet has no header row: {source_sheet}")

    source_headers = [_as_text(value) for value in header_values]
    header_idx = _resolve_input_headers(source_headers)
    output_headers = _build_output_headers()
    computed_headers = _build_computed_headers()

    records: list[dict[str, object]] = []

    for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        def get_col(name: str):
            index = header_idx.get(name)
            if index is None:
                return ""
            return row[index] if index < len(row) else ""

        epic_planned_start_text = _as_text(get_col("Planned Start Date"))
        epic_planned_end_text = _as_text(get_col("Planned End Date"))
        epic_actual_text = _as_text(get_col("Actual Date (Production Date)"))

        epic_start_iso = _normalize_iso_date(get_col("Planned Start Date"))
        epic_end_iso = _normalize_iso_date(get_col("Planned End Date"))
        epic_actual_iso = _normalize_iso_date(get_col("Actual Date (Production Date)"))

        base = {
            "Product": _as_text(get_col("Product")),
            "Epic/RMI": _as_text(get_col("Epic/RMI")),
            "Epic/RMI Jira Link": _normalize_jira_link(get_col("Jira Task ID")),
            "Epic Planned Start Date": epic_planned_start_text,
            "Epic Planned End Date": epic_planned_end_text,
            "Epic Actual Date (Production Date)": epic_actual_text,
            "Remarks": _as_text(get_col("Remarks")),
        }

        parsed_phases = []
        phases_for_compute = []
        for phase_name, input_header in PHASE_COLUMNS:
            parsed = _parse_phase_cell(get_col(input_header))
            parsed_phases.append(parsed)
            phases_for_compute.append(
                {
                    "name": phase_name,
                    "state": parsed["state"],
                    "state_label": parsed["state"].replace("_", " "),
                    "warning": parsed["warning"],
                    "start_iso": parsed["start"],
                    "end_iso": parsed["end"],
                    "start_date": _parse_iso_date(parsed["start"]),
                    "end_date": _parse_iso_date(parsed["end"]),
                    "mandays_text": parsed["mandays"],
                    "mandays_num": _to_number(parsed["mandays"]),
                    "raw": parsed["raw"],
                }
            )

        if not _row_has_payload(base, parsed_phases):
            continue

        total_mandays = sum((p["mandays_num"] or 0.0) for p in phases_for_compute)

        epic_start_date = _parse_iso_date(epic_start_iso)
        epic_end_date = _parse_iso_date(epic_end_iso)
        epic_actual_date = _parse_iso_date(epic_actual_iso)
        has_valid_epic_plan = bool(epic_start_date and epic_end_date and epic_start_date <= epic_end_date)

        any_warning = False
        any_no_entry = False
        any_not_planned = False
        any_skipped = False

        out_row: list[object] = [
            source_sheet,
            row_number,
            base["Product"],
            base["Epic/RMI"],
            base["Epic/RMI Jira Link"],
            base["Epic Planned Start Date"],
            base["Epic Planned End Date"],
            base["Epic Actual Date (Production Date)"],
            base["Remarks"],
        ]

        for payload in parsed_phases:
            out_row.extend(
                [
                    payload["start"], payload["end"], payload["mandays"], payload["raw"], payload["state"], payload["warning"],
                ]
            )
            any_warning = any_warning or bool(payload["warning"])
            any_no_entry = any_no_entry or payload["state"] == "no_entry"
            any_not_planned = any_not_planned or payload["state"] == "not_planned"
            any_skipped = any_skipped or payload["state"] == "skipped"

        out_row.extend(
            [
                "Yes" if any_warning else "No",
                "Yes" if any_no_entry else "No",
                "Yes" if any_not_planned else "No",
                "Yes" if any_skipped else "No",
                round(total_mandays, 4),
                "Yes" if has_valid_epic_plan else "No",
                epic_start_iso,
                epic_end_iso,
                epic_actual_iso,
            ]
        )

        records.append(
            {
                "source_sheet": source_sheet,
                "row_number": row_number,
                "base": base,
                "phases": phases_for_compute,
                "out_row": out_row,
                "total_mandays": total_mandays,
                "computed_has_valid_epic_plan": "Yes" if has_valid_epic_plan else "No",
                "epic_start_date": epic_start_date,
                "epic_end_date": epic_end_date,
                "epic_actual_date": epic_actual_date,
            }
        )

    wb.close()

    global_max_mandays = max((p["mandays_num"] or 0.0 for r in records for p in r["phases"]), default=0.0)
    roadmap_axis = _compute_roadmap_axis(records)

    data_rows: list[list[object]] = [r["out_row"] for r in records]
    computed_rows: list[list[object]] = []

    for r in records:
        if roadmap_axis["has_axis"]:
            axis_start = roadmap_axis["axis_start"]
            axis_span_days = roadmap_axis["axis_span_days"]
            bar_left = ""
            bar_width = ""
            if r["computed_has_valid_epic_plan"] == "Yes":
                left = _clamp_percent(((r["epic_start_date"] - axis_start).days / max(1, axis_span_days - 1)) * 100.0)
                right = _clamp_percent(((r["epic_end_date"] - axis_start).days / max(1, axis_span_days - 1)) * 100.0)
                bar_left = round(left, 4)
                bar_width = round(max(SMALL_MIN_WIDTH_PCT, right - left), 4)

            actual_left = ""
            if isinstance(r.get("epic_actual_date"), date):
                actual_pct = _clamp_percent(((r["epic_actual_date"] - axis_start).days / max(1, axis_span_days - 1)) * 100.0)
                actual_left = round(actual_pct, 4)
        else:
            bar_left = ""
            bar_width = ""
            actual_left = ""

        mini = _compute_phase_geometry_for_record(r, global_max_mandays)

        computed_rows.append(
            [
                r["source_sheet"], r["row_number"], r["base"]["Product"], r["base"]["Epic/RMI"], r["base"]["Epic/RMI Jira Link"],
                r["base"]["Epic Planned Start Date"], r["base"]["Epic Planned End Date"], r["base"]["Epic Actual Date (Production Date)"], r["base"]["Remarks"],
                round(r["total_mandays"], 4),
                r["computed_has_valid_epic_plan"],
                roadmap_axis["axis_start"].isoformat() if roadmap_axis["has_axis"] else "",
                roadmap_axis["axis_end"].isoformat() if roadmap_axis["has_axis"] else "",
                roadmap_axis["axis_span_days"] if roadmap_axis["has_axis"] else 0,
                "Yes" if roadmap_axis.get("today_in_range") else "No",
                roadmap_axis.get("today_left_pct", ""),
                bar_left,
                bar_width,
                actual_left,
                json.dumps(roadmap_axis.get("week_ticks", []), ensure_ascii=False),
                "Yes" if mini["has_dated_phases"] else "No",
                mini["axis_start"].isoformat() if isinstance(mini["axis_start"], date) else "",
                mini["axis_end"].isoformat() if isinstance(mini["axis_end"], date) else "",
                mini["axis_span_days"],
                mini["timeline_width_px"],
                "Yes" if mini["scroll_enabled"] else "No",
                json.dumps(mini["week_ticks"], ensure_ascii=False),
                "Yes" if mini["today_in_range"] else "No",
                mini["today_left_pct"],
                json.dumps(mini["phases"], ensure_ascii=False),
            ]
        )

    return output_headers, data_rows, computed_headers, computed_rows, global_max_mandays


def _write_workbook(
    output_path: Path,
    headers: list[str],
    rows: list[list[object]],
    computed_headers: list[str],
    computed_rows: list[list[object]],
    source_workbook: Path,
    source_sheet: str,
    output_mode: str,
    global_max_mandays: float,
) -> None:
    wb = Workbook()
    ws_data = wb.active
    ws_data.title = DEFAULT_DATA_SHEET_NAME
    ws_data.append(headers)
    for row in rows:
        ws_data.append(row)

    ws_computed = wb.create_sheet(DEFAULT_COMPUTED_SHEET_NAME)
    ws_computed.append(computed_headers)
    for row in computed_rows:
        ws_computed.append(row)

    ws_meta = wb.create_sheet(DEFAULT_METADATA_SHEET_NAME)
    generated_at = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
    ws_meta.append(["Generated At", generated_at])
    ws_meta.append(["Source Workbook", str(source_workbook)])
    ws_meta.append(["Source Sheet", source_sheet])
    ws_meta.append(["Output Mode", output_mode])
    ws_meta.append(["Data Row Count", len(rows)])
    ws_meta.append(["Phase Count", len(PHASE_COLUMNS)])
    ws_meta.append(["Computed Sheet Name", DEFAULT_COMPUTED_SHEET_NAME])
    ws_meta.append(["Global Max Man-days Benchmark", round(global_max_mandays, 4)])
    ws_meta.append(["Computation Timestamp UTC", generated_at])
    ws_meta.append(["Computation Rules Version", COMPUTATION_RULES_VERSION])

    wb.save(output_path)

def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Export normalized IPP phase breakdown workbook.")
    parser.add_argument("--input-xlsx", default="", help="Input workbook path override.")
    parser.add_argument("--sheet", default="", help="Exact source sheet name override.")
    parser.add_argument("--output-dir", default="", help="Output directory override.")
    parser.add_argument(
        "--output-mode",
        default="",
        choices=["timestamped", "fixed", "both"],
        help="Output write mode.",
    )
    parser.add_argument(
        "--fixed-output-name",
        default="",
        help=f"Fixed output file name (default: {DEFAULT_FIXED_OUTPUT_NAME}).",
    )
    return parser.parse_args()


def main() -> None:
    args = _parse_args()
    base_dir = Path(__file__).resolve().parent

    input_xlsx = (
        args.input_xlsx.strip()
        or os.getenv("IPP_MEETING_XLSX_PATH", "").strip()
        or DEFAULT_INPUT_XLSX
    )
    input_path = Path(input_xlsx)
    if not input_path.is_absolute():
        input_path = base_dir / input_path
    if not input_path.exists():
        raise FileNotFoundError(f"Input workbook not found: {input_path}")

    output_dir_value = (
        args.output_dir.strip()
        or os.getenv("IPP_PHASE_OUTPUT_DIR", "").strip()
        or str(base_dir)
    )
    output_dir = Path(output_dir_value)
    if not output_dir.is_absolute():
        output_dir = base_dir / output_dir
    output_dir.mkdir(parents=True, exist_ok=True)

    output_mode = (
        args.output_mode.strip()
        or os.getenv("IPP_PHASE_OUTPUT_MODE", "").strip().lower()
        or DEFAULT_OUTPUT_MODE
    )
    if output_mode not in {"timestamped", "fixed", "both"}:
        raise ValueError("IPP phase output mode must be one of: timestamped, fixed, both")

    fixed_output_name = (
        args.fixed_output_name.strip()
        or os.getenv("IPP_PHASE_OUTPUT_FIXED_NAME", "").strip()
        or DEFAULT_FIXED_OUTPUT_NAME
    )

    requested_sheet = args.sheet.strip() or os.getenv("IPP_PHASE_SOURCE_SHEET", "").strip()

    wb = load_workbook(input_path, read_only=True, data_only=True)
    source_sheet = _select_source_sheet(list(wb.sheetnames), requested_sheet)
    wb.close()

    headers, rows, computed_headers, computed_rows, global_max_mandays = _build_output_rows(input_path, source_sheet)

    stamp = datetime.now().strftime("%Y-%m-%d")
    timestamped_name = f"ipp_phase_breakdown_{stamp}.xlsx"
    writes: list[Path] = []

    if output_mode in {"timestamped", "both"}:
        timestamped_path = output_dir / timestamped_name
        _write_workbook(
            timestamped_path,
            headers,
            rows,
            computed_headers,
            computed_rows,
            input_path,
            source_sheet,
            output_mode,
            global_max_mandays,
        )
        writes.append(timestamped_path)

    if output_mode in {"fixed", "both"}:
        fixed_path = output_dir / fixed_output_name
        _write_workbook(
            fixed_path,
            headers,
            rows,
            computed_headers,
            computed_rows,
            input_path,
            source_sheet,
            output_mode,
            global_max_mandays,
        )
        writes.append(fixed_path)

    print(f"Source workbook: {input_path}")
    print(f"Source sheet: {source_sheet}")
    print(f"Rows exported: {len(rows)}")
    print(f"Computed rows exported: {len(computed_rows)}")
    print(f"Global max man-days benchmark: {round(global_max_mandays, 4)}")
    for path in writes:
        print(f"Wrote: {path}")


if __name__ == "__main__":
    main()
