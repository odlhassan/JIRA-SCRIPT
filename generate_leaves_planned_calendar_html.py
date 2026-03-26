"""
Generate a planned leaves calendar HTML with date drilldown details.
"""
from __future__ import annotations

import json
import os
from collections import defaultdict
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from canonical_report_data import build_rlt_leave_snapshot, resolve_canonical_run_id

DEFAULT_LEAVE_REPORT_INPUT_XLSX = "rlt_leave_report.xlsx"
DEFAULT_OUTPUT_HTML = "leaves_planned_calendar.html"
EPSILON = 0.01


def _resolve_path(value: str, base_dir: Path) -> Path:
    path = Path(value)
    return path if path.is_absolute() else base_dir / path


def _to_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _to_float(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _parse_iso_day(value: Any) -> str:
    text = _to_text(value)
    if not text:
        return ""
    try:
        return date.fromisoformat(text).isoformat()
    except ValueError:
        return ""


def _month_start(day_value: date) -> date:
    return date(day_value.year, day_value.month, 1)


def _jira_base_url() -> str:
    site = _to_text(os.getenv("JIRA_SITE")) or "octopusdtlsupport"
    return f"https://{site}.atlassian.net"


def _load_daily(wb) -> tuple[dict[str, int], dict[str, float], dict[tuple[str, str], float], int, list[str]]:
    warnings: list[str] = []
    if "Daily_Assignee" not in wb.sheetnames:
        warnings.append("Daily_Assignee sheet missing; calendar and drilldown unavailable.")
        return {}, {}, {}, 0, warnings
    ws = wb["Daily_Assignee"]
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header:
        warnings.append("Daily_Assignee header missing; calendar and drilldown unavailable.")
        return {}, {}, {}, 0, warnings
    idx = {str(h).strip(): i for i, h in enumerate(header)}
    need = ["assignee", "period_day", "planned_taken_hours", "planned_not_taken_hours"]
    if any(name not in idx for name in need):
        warnings.append("Daily_Assignee required columns missing; calendar and drilldown unavailable.")
        return {}, {}, {}, 0, warnings

    counts_by_date: dict[str, int] = {}
    planned_hours_by_date: dict[str, float] = {}
    planned_by_day_assignee: dict[tuple[str, str], float] = defaultdict(float)
    skipped = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        iso = _parse_iso_day(row[idx["period_day"]])
        if not iso:
            skipped += 1
            continue
        assignee = _to_text(row[idx["assignee"]]) or "Unassigned"
        planned = round(_to_float(row[idx["planned_taken_hours"]]) + _to_float(row[idx["planned_not_taken_hours"]]), 2)
        if planned <= 0:
            continue
        counts_by_date[iso] = counts_by_date.get(iso, 0) + 1
        planned_hours_by_date[iso] = round(planned_hours_by_date.get(iso, 0.0) + planned, 2)
        planned_by_day_assignee[(iso, assignee)] = round(planned_by_day_assignee[(iso, assignee)] + planned, 2)
    return counts_by_date, planned_hours_by_date, planned_by_day_assignee, skipped, warnings


def _load_subtasks(wb) -> tuple[list[dict[str, Any]], list[str]]:
    warnings: list[str] = []
    sheet_name = ""
    if "Subtasks_Distributed" in wb.sheetnames:
        sheet_name = "Subtasks_Distributed"
    elif "Raw_Subtasks" in wb.sheetnames:
        sheet_name = "Raw_Subtasks"
    else:
        warnings.append("Subtasks_Distributed/Raw_Subtasks sheet missing; issue-level drilldown limited.")
        return [], warnings
    ws = wb[sheet_name]
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header:
        warnings.append(f"{sheet_name} header missing; issue-level drilldown limited.")
        return [], warnings
    idx = {str(h).strip(): i for i, h in enumerate(header)}
    need = ["issue_key", "assignee", "leave_classification", "original_estimate_hours", "planned_date_for_bucket", "total_worklog_hours", "no_entry_flag", "start_date", "due_date"]
    if any(name not in idx for name in need):
        warnings.append(f"{sheet_name} required columns missing; issue-level drilldown limited.")
        return [], warnings

    out: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        key = _to_text(row[idx["issue_key"]]).upper()
        if not key:
            continue
        out.append(
            {
            "issue_key": key,
            "assignee": _to_text(row[idx["assignee"]]) or "Unassigned",
            "leave_classification": _to_text(row[idx["leave_classification"]]),
            "original_estimate_hours": _to_float(row[idx["original_estimate_hours"]]),
            "planned_date_for_bucket": _parse_iso_day(row[idx["planned_date_for_bucket"]]),
            "total_worklog_hours": _to_float(row[idx["total_worklog_hours"]]),
            "no_entry_flag": _to_text(row[idx["no_entry_flag"]]),
            "start_date": _parse_iso_day(row[idx["start_date"]]),
            "due_date": _parse_iso_day(row[idx["due_date"]]),
        }
        )
    return out, warnings


def _load_worklogs(wb) -> tuple[list[dict[str, Any]], list[str]]:
    warnings: list[str] = []
    if "Worklogs_Normalized" not in wb.sheetnames:
        warnings.append("Worklogs_Normalized sheet missing; planned taken issue-level drilldown limited.")
        return [], warnings
    ws = wb["Worklogs_Normalized"]
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header:
        warnings.append("Worklogs_Normalized header missing; planned taken issue-level drilldown limited.")
        return [], warnings
    idx = {str(h).strip(): i for i, h in enumerate(header)}
    need = ["issue_key", "started_date", "hours_logged"]
    if any(name not in idx for name in need):
        warnings.append("Worklogs_Normalized required columns missing; planned taken issue-level drilldown limited.")
        return [], warnings

    rows: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        key = _to_text(row[idx["issue_key"]]).upper()
        iso = _parse_iso_day(row[idx["started_date"]])
        hrs = _to_float(row[idx["hours_logged"]])
        if key and iso and hrs > 0:
            rows.append({"issue_key": key, "started_date": iso, "hours_logged": hrs})
    return rows, warnings


def _append_detail(
    details_raw: dict[str, dict[tuple[str, str, str], dict[str, Any]]],
    mapped: dict[tuple[str, str], float],
    *,
    iso: str,
    assignee: str,
    issue_key: str,
    estimate: float | None,
    hours: float,
    source: str,
    jira_base_url: str,
) -> None:
    if not iso or hours <= 0:
        return
    person = assignee or "Unassigned"
    key = (issue_key or "").upper()
    row_key = (person, key, source)
    bucket = details_raw.setdefault(iso, {})
    if row_key not in bucket:
        bucket[row_key] = {
            "assignee": person,
            "issue_key": key,
            "jira_url": f"{jira_base_url}/browse/{key}" if key else "",
            "original_estimate_hours": round(estimate, 2) if estimate and estimate > 0 else None,
            "hours_on_date": 0.0,
            "source": source,
        }
    bucket[row_key]["hours_on_date"] = round(bucket[row_key]["hours_on_date"] + hours, 2)
    mapped[(iso, person)] = round(mapped[(iso, person)] + hours, 2)


def _build_details(
    wb,
    planned_by_day_assignee: dict[tuple[str, str], float],
) -> tuple[dict[str, list[dict[str, Any]]], dict[str, list[dict[str, Any]]], list[str]]:
    warnings: list[str] = []
    jira_base_url = _jira_base_url()
    subtasks, subtask_warnings = _load_subtasks(wb)
    worklogs, worklog_warnings = _load_worklogs(wb)
    warnings.extend(subtask_warnings)
    warnings.extend(worklog_warnings)

    details_raw: dict[str, dict[tuple[str, str, str], dict[str, Any]]] = {}
    mapped: dict[tuple[str, str], float] = defaultdict(float)
    subtasks_by_issue: dict[str, list[dict[str, Any]]] = defaultdict(list)
    subtasks_by_issue_day: dict[tuple[str, str], dict[str, Any]] = {}
    for subtask in subtasks:
        key = _to_text(subtask.get("issue_key")).upper()
        if not key:
            continue
        subtasks_by_issue[key].append(subtask)
        day = _to_text(subtask.get("planned_date_for_bucket")) or _to_text(subtask.get("start_date")) or _to_text(subtask.get("due_date"))
        if day:
            subtasks_by_issue_day[(key, day)] = subtask

    for log in worklogs:
        issue_key = _to_text(log.get("issue_key")).upper()
        log_day = _to_text(log.get("started_date"))
        subtask = subtasks_by_issue_day.get((issue_key, log_day))
        if not subtask and issue_key in subtasks_by_issue:
            # Fallback for legacy rows where issue/date mapping is not available.
            subtask = subtasks_by_issue[issue_key][0]
        if not subtask:
            continue
        if _to_text(subtask.get("leave_classification")) != "Planned":
            continue
        estimate = _to_float(subtask.get("original_estimate_hours"))
        # Logged-date fallback: even with no estimate and no start/due, keep worklog started_date.
        _append_detail(
            details_raw,
            mapped,
            iso=log_day,
            assignee=_to_text(subtask.get("assignee")) or "Unassigned",
            issue_key=issue_key,
            estimate=estimate if estimate > 0 else None,
            hours=_to_float(log["hours_logged"]),
            source="planned_taken_worklog",
            jira_base_url=jira_base_url,
        )

    for subtask in subtasks:
        if _to_text(subtask.get("leave_classification")) != "Planned":
            continue
        if _to_float(subtask.get("total_worklog_hours")) > 0:
            continue
        if _to_text(subtask.get("no_entry_flag")) == "Yes":
            continue
        iso = _to_text(subtask.get("planned_date_for_bucket"))
        estimate = _to_float(subtask.get("original_estimate_hours"))
        if not iso or estimate <= 0:
            continue
        _append_detail(
            details_raw,
            mapped,
            iso=iso,
            assignee=_to_text(subtask.get("assignee")) or "Unassigned",
            issue_key=_to_text(subtask.get("issue_key")),
            estimate=estimate,
            hours=estimate,
            source="planned_not_taken_bucket",
            jira_base_url=jira_base_url,
        )

    details_by_date: dict[str, list[dict[str, Any]]] = {}
    for iso, rows in details_raw.items():
        row_list = list(rows.values())
        row_list.sort(key=lambda x: (_to_text(x.get("assignee")).lower(), _to_text(x.get("issue_key")).lower(), _to_text(x.get("source"))))
        details_by_date[iso] = row_list

    unmatched_by_date: dict[str, list[dict[str, Any]]] = {}
    temp_unmatched: dict[str, dict[str, dict[str, Any]]] = {}
    for (iso, assignee), daily_hours in planned_by_day_assignee.items():
        residual = round(daily_hours - _to_float(mapped.get((iso, assignee))), 2)
        if residual <= EPSILON:
            continue
        temp_unmatched.setdefault(iso, {})[assignee] = {
            "assignee": assignee,
            "hours_on_date": residual,
            "reason": "No issue-level mapping for this day",
        }
    for iso, rows in temp_unmatched.items():
        row_list = list(rows.values())
        row_list.sort(key=lambda x: _to_text(x.get("assignee")).lower())
        unmatched_by_date[iso] = row_list

    return details_by_date, unmatched_by_date, warnings


def _load_calendar_data(
    leave_report_path: Path,
) -> tuple[dict[str, int], dict[str, float], int, dict[str, list[dict[str, Any]]], dict[str, list[dict[str, Any]]], list[str]]:
    if not leave_report_path.exists():
        raise FileNotFoundError(f"Leave report workbook not found: {leave_report_path}")
    wb = load_workbook(leave_report_path, read_only=True, data_only=True)
    try:
        counts, planned_hours, planned_by_day_assignee, skipped, warnings = _load_daily(wb)
        details_by_date, unmatched_by_date, details_warnings = _build_details(wb, planned_by_day_assignee)
        warnings.extend(details_warnings)
        return counts, planned_hours, skipped, details_by_date, unmatched_by_date, sorted(set(warnings))
    finally:
        wb.close()


def _load_calendar_data_from_canonical(
    db_path: Path,
    run_id: str = "",
) -> tuple[dict[str, int], dict[str, float], int, dict[str, list[dict[str, Any]]], dict[str, list[dict[str, Any]]], list[str]]:
    effective_run_id = resolve_canonical_run_id(db_path, run_id)
    if not effective_run_id:
        return {}, {}, 0, {}, {}, []
    snapshot = build_rlt_leave_snapshot(db_path, effective_run_id)
    daily_rows = list(snapshot.get("daily") or [])
    counts_by_date: dict[str, int] = {}
    planned_hours_by_date: dict[str, float] = {}
    planned_by_day_assignee: dict[tuple[str, str], float] = defaultdict(float)
    for row in daily_rows:
        iso = _parse_iso_day(row.get("period_day"))
        if not iso:
            continue
        assignee = _to_text(row.get("assignee")) or "Unassigned"
        planned = round(_to_float(row.get("planned_taken_hours")) + _to_float(row.get("planned_not_taken_hours")), 2)
        if planned <= 0:
            continue
        counts_by_date[iso] = counts_by_date.get(iso, 0) + 1
        planned_hours_by_date[iso] = round(planned_hours_by_date.get(iso, 0.0) + planned, 2)
        planned_by_day_assignee[(iso, assignee)] = round(planned_by_day_assignee[(iso, assignee)] + planned, 2)

    subtasks = list(snapshot.get("distributed_subtasks") or snapshot.get("raw_subtasks") or [])
    worklogs = list(snapshot.get("worklogs_normalized") or [])
    details_raw: dict[str, dict[tuple[str, str, str], dict[str, Any]]] = {}
    mapped: dict[tuple[str, str], float] = defaultdict(float)
    jira_base_url = _jira_base_url()
    subtasks_by_issue_day: dict[tuple[str, str], dict[str, Any]] = {}
    subtasks_by_issue: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for subtask in subtasks:
        key = _to_text(subtask.get("issue_key")).upper()
        if not key:
            continue
        subtasks_by_issue[key].append(subtask)
        day = _to_text(subtask.get("planned_date_for_bucket")) or _to_text(subtask.get("start_date")) or _to_text(subtask.get("due_date"))
        if day:
            subtasks_by_issue_day[(key, day)] = subtask

    for log in worklogs:
        issue_key = _to_text(log.get("issue_key")).upper()
        log_day = _to_text(log.get("started_date"))
        subtask = subtasks_by_issue_day.get((issue_key, log_day))
        if not subtask and issue_key in subtasks_by_issue:
            subtask = subtasks_by_issue[issue_key][0]
        if not subtask or _to_text(subtask.get("leave_classification")) != "Planned":
            continue
        _append_detail(
            details_raw,
            mapped,
            iso=log_day,
            assignee=_to_text(subtask.get("assignee")) or "Unassigned",
            issue_key=issue_key,
            estimate=_to_float(subtask.get("original_estimate_hours")) or None,
            hours=_to_float(log.get("hours_logged")),
            source="planned_taken_worklog",
            jira_base_url=jira_base_url,
        )

    for subtask in subtasks:
        if _to_text(subtask.get("leave_classification")) != "Planned":
            continue
        if _to_float(subtask.get("total_worklog_hours")) > 0:
            continue
        if _to_text(subtask.get("no_entry_flag")) == "Yes":
            continue
        iso = _to_text(subtask.get("planned_date_for_bucket"))
        estimate = _to_float(subtask.get("original_estimate_hours"))
        if not iso or estimate <= 0:
            continue
        _append_detail(
            details_raw,
            mapped,
            iso=iso,
            assignee=_to_text(subtask.get("assignee")) or "Unassigned",
            issue_key=_to_text(subtask.get("issue_key")),
            estimate=estimate,
            hours=estimate,
            source="planned_not_taken_bucket",
            jira_base_url=jira_base_url,
        )

    details_by_date = {
        iso: sorted(rows.values(), key=lambda x: (_to_text(x.get("assignee")).lower(), _to_text(x.get("issue_key")).lower(), _to_text(x.get("source"))))
        for iso, rows in details_raw.items()
    }
    unmatched_by_date: dict[str, list[dict[str, Any]]] = {}
    for (iso, assignee), daily_hours in planned_by_day_assignee.items():
        residual = round(daily_hours - _to_float(mapped.get((iso, assignee))), 2)
        if residual > EPSILON:
            unmatched_by_date.setdefault(iso, []).append(
                {"assignee": assignee, "hours_on_date": residual, "reason": "No issue-level mapping for this day"}
            )
    for iso in list(unmatched_by_date):
        unmatched_by_date[iso] = sorted(unmatched_by_date[iso], key=lambda row: _to_text(row.get("assignee")).lower())
    return counts_by_date, planned_hours_by_date, 0, details_by_date, unmatched_by_date, []


def _resolve_date_range(counts: dict[str, int]) -> tuple[str, str]:
    min_day: date | None = None
    max_day: date | None = None
    for iso in counts:
        try:
            day = date.fromisoformat(iso)
        except ValueError:
            continue
        if min_day is None or day < min_day:
            min_day = day
        if max_day is None or day > max_day:
            max_day = day
    if min_day is None or max_day is None:
        today = date.today()
        min_day = _month_start(today)
        max_day = today
    return min_day.isoformat(), max_day.isoformat()


def _build_html(payload: dict[str, Any]) -> str:
    data = json.dumps(payload, ensure_ascii=True)
    template = """<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Planned Leaves Calendar</title>
<style>
body{margin:0;font-family:Segoe UI,Tahoma,sans-serif;background:#f4f7fb;color:#0f172a}.page{max-width:1400px;margin:0 auto;padding:16px}.card{background:#fff;border:1px solid #d7e0ea;border-radius:12px;padding:14px;margin-bottom:12px}.controls{display:grid;gap:10px;grid-template-columns:repeat(auto-fit,minmax(210px,1fr));align-items:end}label{display:block;font-size:.75rem;text-transform:uppercase;font-weight:700;color:#475569;margin-bottom:6px}input[type=month]{width:100%;border:1px solid #b8c7d5;border-radius:8px;padding:8px}.btn{border:1px solid #255f73;background:#0f4c5c;color:#fff;border-radius:8px;padding:8px 12px;cursor:pointer}.btn.alt{background:#eff6ff;border-color:#bfdbfe;color:#1d4ed8}.btn-row{display:flex;gap:8px;flex-wrap:wrap}.toggle-chip{border:1px solid #bfdbfe;background:#eff6ff;color:#1d4ed8;border-radius:999px;padding:6px 10px;cursor:pointer;font-weight:700}.toggle-chip.active{border-color:#1e40af;background:#1d4ed8;color:#fff}.months-grid{display:grid;gap:10px;grid-template-columns:repeat(auto-fill,minmax(320px,1fr))}.month-card{border:1px solid #d7e0ea;border-radius:10px;background:#fbfdff;overflow:hidden}.month-head{background:#eff6ff;color:#1d4ed8;font-weight:700;padding:8px 10px}table{width:100%;border-collapse:collapse;table-layout:fixed}th{font-size:.72rem;background:#f8fafc;border-bottom:1px solid #e2e8f0;padding:6px 0}td{height:38px;border-top:1px solid #eef2f7;border-right:1px solid #eef2f7;text-align:center;position:relative}tr td:last-child{border-right:none}.day-empty{background:#f8fafc}.day-has-date{cursor:pointer}.day-has-date:focus-visible{outline:2px solid #1d4ed8;outline-offset:-2px}.day-selected{box-shadow:inset 0 0 0 2px #1d4ed8}.day-planned{font-weight:700;color:#7f1d1d}.day-count{position:absolute;right:4px;bottom:3px;font-size:.62rem;font-weight:700}.details-empty{color:#475569}.details-table-wrap{overflow:auto;border:1px solid #dbe4ec;border-radius:10px;margin-top:8px}.details-table{width:100%;border-collapse:collapse;table-layout:auto;font-size:.82rem}.details-table th,.details-table td{border-bottom:1px solid #e8eef4;text-align:left;padding:8px 10px;white-space:nowrap}.source-chip{display:inline-block;border:1px solid #bfdbfe;background:#eff6ff;color:#1e40af;border-radius:999px;padding:2px 8px;font-size:.7rem;font-weight:700}
</style>
<link rel="stylesheet" href="shared-nav.css">
</head>
<body>
<div class="page">
<section class="card">
<h1>Planned Leaves Calendar</h1>
<p>Generated: <span id="generated-at"></span> | Source: <span id="source-file"></span></p>
<div class="controls">
<div><label for="from-month">From Month</label><input id="from-month" type="month"></div>
<div><label for="to-month">To Month</label><input id="to-month" type="month"></div>
<div><label>Color Intensity By</label><button class="toggle-chip active" id="mode-count" type="button">Planned Leaves Count</button> <button class="toggle-chip" id="mode-hours" type="button">Planned Leave Hours</button></div>
<div class="btn-row"><button class="btn" id="apply" type="button">Apply</button><button class="btn alt" id="reset" type="button">Reset</button></div>
</div>
<div id="status"></div>
</section>
<section class="months-grid" id="months-grid"></section>
<section class="card">
<h2>Date Details</h2>
<div id="details-meta">Select a date to view assignees, Jira links, and original estimates.</div>
<div id="details-content" class="details-empty">No date selected.</div>
</section>
</div>
<script>
const payload = __PAYLOAD__;
const countsByDate = payload.counts_by_date || {};
const plannedHoursByDate = payload.planned_hours_by_date || {};
const detailsByDate = payload.details_by_date || {};
const unmatchedByDate = payload.unmatched_assignee_hours_by_date || {};
const warnings = Array.isArray(payload.warnings) ? payload.warnings : [];
const maxCount = Number(payload.max_count || 0);
const maxPlannedHours = Number(payload.max_planned_hours || 0);
let heatMode = "count";
let selectedIsoDate = "";
const statusEl = document.getElementById("status");
const monthsGridEl = document.getElementById("months-grid");
const detailsMetaEl = document.getElementById("details-meta");
const detailsContentEl = document.getElementById("details-content");
document.getElementById("generated-at").textContent = payload.generated_at || "-";
document.getElementById("source-file").textContent = payload.source_file || "-";
if (payload.details_meta && payload.details_meta.jira_base_url) detailsMetaEl.textContent = "Select a date to view assignees, Jira links, and original estimates. Jira base: " + payload.details_meta.jira_base_url;
const parseIsoDate = (v) => { const d = new Date(v + "T00:00:00"); return Number.isNaN(d.getTime()) ? null : d; };
const monthStart = (d) => new Date(d.getFullYear(), d.getMonth(), 1);
const monthKey = (d) => d.getFullYear() + "-" + String(d.getMonth() + 1).padStart(2, "0");
const isoDate = (d) => d.getFullYear() + "-" + String(d.getMonth() + 1).padStart(2, "0") + "-" + String(d.getDate()).padStart(2, "0");
const monthLabel = (d) => d.toLocaleDateString(undefined, { month: "long", year: "numeric" });
const hoursText = (v) => Number(v || 0).toFixed(2).replace(/\\.00$/, "") + "h";
const sourceLabel = (s) => s === "planned_not_taken_bucket" ? "Planned Not Taken" : "Planned Taken";
const heatColor = (v,m) => { if (!v || v <= 0 || !m) return "#f8fafc"; const r = v / m; if (r <= .2) return "#fee2e2"; if (r <= .4) return "#fecaca"; if (r <= .6) return "#fca5a5"; if (r <= .8) return "#f87171"; return "#dc2626"; };
function clearDetails(msg){ detailsContentEl.className="details-empty"; detailsContentEl.textContent=msg || "No date selected."; }
function renderDetails(iso){ selectedIsoDate = iso || ""; const rows = detailsByDate[selectedIsoDate] || []; const unmatched = unmatchedByDate[selectedIsoDate] || []; const total = Number(plannedHoursByDate[selectedIsoDate] || 0); if(!selectedIsoDate) return clearDetails("No date selected."); if(total<=0 && rows.length===0 && unmatched.length===0) return clearDetails("No planned leave details for " + selectedIsoDate + "."); const people = new Set(); rows.forEach(r=>people.add(String(r.assignee || "Unassigned"))); unmatched.forEach(r=>people.add(String(r.assignee || "Unassigned"))); detailsContentEl.className=""; detailsContentEl.innerHTML = "<div>" + selectedIsoDate + " | Planned: " + hoursText(total) + " | Assignees: " + people.size + "</div>"; if(rows.length){ let html = "<div class='details-table-wrap'><table class='details-table'><thead><tr><th>Assignee</th><th>Jira</th><th>Original Estimate</th><th>Hours On Selected Date</th><th>Source</th></tr></thead><tbody>"; rows.forEach(r=>{ const key = String(r.issue_key || ""); const url = String(r.jira_url || ""); const jira = (key && url) ? "<a target='_blank' rel='noopener noreferrer' href='" + url + "'>" + key + "</a>" : "No Entry"; const est = r.original_estimate_hours == null ? "No Entry" : hoursText(r.original_estimate_hours); html += "<tr><td>" + String(r.assignee || "Unassigned") + "</td><td>" + jira + "</td><td>" + est + "</td><td>" + hoursText(r.hours_on_date) + "</td><td><span class='source-chip'>" + sourceLabel(String(r.source || "")) + "</span></td></tr>"; }); html += "</tbody></table></div>"; detailsContentEl.innerHTML += html; } if(unmatched.length){ let html = "<div style='margin-top:10px;font-weight:700'>Unmapped hours</div><div class='details-table-wrap'><table class='details-table'><thead><tr><th>Assignee</th><th>Hours On Selected Date</th><th>Reason</th></tr></thead><tbody>"; unmatched.forEach(r=>{ html += "<tr><td>" + String(r.assignee || "Unassigned") + "</td><td>" + hoursText(r.hours_on_date) + "</td><td>" + String(r.reason || "No issue-level mapping for this day") + "</td></tr>"; }); html += "</tbody></table></div>"; detailsContentEl.innerHTML += html; } }
function selectDate(iso){ selectedIsoDate = iso || ""; const old = monthsGridEl.querySelector("td.day-selected"); if(old) old.classList.remove("day-selected"); const cell = monthsGridEl.querySelector("td[data-iso='" + selectedIsoDate + "']"); if(cell) cell.classList.add("day-selected"); renderDetails(selectedIsoDate); }
function renderRange(fromM,toM){ monthsGridEl.innerHTML=""; if(!fromM || !toM || toM < fromM){ statusEl.textContent="Invalid month range."; clearDetails("No date selected."); return; } let months=0,high=0; const c=new Date(fromM); while(c<=toM){ months += 1; const card=document.createElement("article"); card.className="month-card"; card.innerHTML="<div class='month-head'>" + monthLabel(c) + "</div>"; const t=document.createElement("table"); t.innerHTML="<thead><tr><th>Mon</th><th>Tue</th><th>Wed</th><th>Thu</th><th>Fri</th><th>Sat</th><th>Sun</th></tr></thead>"; const b=document.createElement("tbody"); const first=new Date(c.getFullYear(), c.getMonth(), 1), last=new Date(c.getFullYear(), c.getMonth()+1, 0), firstW=(first.getDay()+6)%7; let day=1; const totalDays=last.getDate(); for(let w=0; w<6; w+=1){ const tr=document.createElement("tr"); for(let col=0; col<7; col+=1){ const td=document.createElement("td"); if((w===0 && col<firstW) || day>totalDays){ td.className="day-empty"; } else { const cur=new Date(c.getFullYear(), c.getMonth(), day); const iso=isoDate(cur); const count=Number(countsByDate[iso] || 0); const hours=Number(plannedHoursByDate[iso] || 0); td.textContent=String(day); td.classList.add("day-has-date"); td.dataset.iso=iso; td.tabIndex=0; td.setAttribute("role","button"); td.addEventListener("click", ()=>selectDate(iso)); td.addEventListener("keydown",(e)=>{ if(e.key==="Enter" || e.key===" "){ e.preventDefault(); selectDate(iso);} }); if(count>0){ high += 1; td.classList.add("day-planned"); const metric = heatMode==="hours" ? hours : count; const metricMax = heatMode==="hours" ? maxPlannedHours : maxCount; td.style.background = heatColor(metric, metricMax); const ce=document.createElement("span"); ce.className="day-count"; ce.textContent = heatMode==="hours" ? hoursText(hours) : String(count); td.appendChild(ce); } if(iso===selectedIsoDate) td.classList.add("day-selected"); day += 1; } tr.appendChild(td); } b.appendChild(tr); if(day>totalDays) break; } t.appendChild(b); card.appendChild(t); monthsGridEl.appendChild(card); c.setMonth(c.getMonth()+1,1); } const modeText = heatMode==="hours" ? "Mode: Planned Leave Hours | Darkest cell: " + hoursText(maxPlannedHours) : "Mode: Planned Leaves Count | Darkest cell: " + String(maxCount); const warnText = warnings.length ? " | Warning: " + warnings.join(" | ") : ""; statusEl.textContent = "Months: " + months + " | Highlighted dates: " + high + " | " + modeText + warnText; if(selectedIsoDate){ const exists = monthsGridEl.querySelector("td[data-iso='" + selectedIsoDate + "']"); if(exists) renderDetails(selectedIsoDate); else { selectedIsoDate = ""; clearDetails("Selected date is outside the current month range."); } } }
function applyRange(){ const fr=document.getElementById("from-month").value; const tr=document.getElementById("to-month").value; if(!fr || !tr){ statusEl.textContent="Select both From and To month."; return; } renderRange(monthStart(parseIsoDate(fr+"-01")), monthStart(parseIsoDate(tr+"-01"))); }
const min = parseIsoDate(payload.min_date || ""); const max = parseIsoDate(payload.max_date || ""); const defaultFrom = min ? monthKey(monthStart(min)) : monthKey(monthStart(new Date())); const defaultTo = max ? monthKey(monthStart(max)) : defaultFrom;
document.getElementById("from-month").value = defaultFrom; document.getElementById("to-month").value = defaultTo; clearDetails("No date selected."); renderRange(monthStart(parseIsoDate(defaultFrom + "-01")), monthStart(parseIsoDate(defaultTo + "-01")));
document.getElementById("apply").addEventListener("click", applyRange);
document.getElementById("mode-count").addEventListener("click", () => { heatMode = "count"; document.getElementById("mode-count").classList.add("active"); document.getElementById("mode-hours").classList.remove("active"); applyRange(); });
document.getElementById("mode-hours").addEventListener("click", () => { heatMode = "hours"; document.getElementById("mode-hours").classList.add("active"); document.getElementById("mode-count").classList.remove("active"); applyRange(); });
document.getElementById("reset").addEventListener("click", () => { document.getElementById("from-month").value = defaultFrom; document.getElementById("to-month").value = defaultTo; heatMode = "count"; document.getElementById("mode-count").classList.add("active"); document.getElementById("mode-hours").classList.remove("active"); selectedIsoDate = ""; clearDetails("No date selected."); renderRange(monthStart(parseIsoDate(defaultFrom + "-01")), monthStart(parseIsoDate(defaultTo + "-01"))); });
</script>
<script src="shared-nav.js"></script>
</body>
</html>"""
    return template.replace("__PAYLOAD__", data)


def main() -> None:
    base_dir = Path(__file__).resolve().parent
    leave_report_name = os.getenv("JIRA_LEAVE_REPORT_XLSX_PATH", DEFAULT_LEAVE_REPORT_INPUT_XLSX).strip() or DEFAULT_LEAVE_REPORT_INPUT_XLSX
    output_name = os.getenv("JIRA_LEAVES_CALENDAR_HTML_PATH", DEFAULT_OUTPUT_HTML).strip() or DEFAULT_OUTPUT_HTML
    db_name = os.getenv("JIRA_ASSIGNEE_HOURS_CAPACITY_DB_PATH", "assignee_hours_capacity.db").strip() or "assignee_hours_capacity.db"
    canonical_run_id = os.getenv("JIRA_CANONICAL_RUN_ID", "").strip()

    leave_report_path = _resolve_path(leave_report_name, base_dir)
    output_path = _resolve_path(output_name, base_dir)
    db_path = _resolve_path(db_name, base_dir)
    counts_by_date, planned_hours_by_date, skipped_invalid_dates, details_by_date, unmatched_by_date, warnings = _load_calendar_data_from_canonical(db_path, canonical_run_id)
    source_file = "canonical_db"
    if not counts_by_date and not details_by_date:
        counts_by_date, planned_hours_by_date, skipped_invalid_dates, details_by_date, unmatched_by_date, warnings = _load_calendar_data(leave_report_path)
        source_file = str(leave_report_path)
    min_date, max_date = _resolve_date_range(counts_by_date)
    max_count = max(counts_by_date.values()) if counts_by_date else 0
    max_planned_hours = max(planned_hours_by_date.values()) if planned_hours_by_date else 0

    payload = {
        "generated_at": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC"),
        "source_file": source_file,
        "days_with_planned_leaves": len(counts_by_date),
        "skipped_invalid_dates": skipped_invalid_dates,
        "min_date": min_date,
        "max_date": max_date,
        "max_count": int(max_count),
        "max_planned_hours": round(float(max_planned_hours), 2),
        "counts_by_date": counts_by_date,
        "planned_hours_by_date": planned_hours_by_date,
        "details_by_date": details_by_date,
        "unmatched_assignee_hours_by_date": unmatched_by_date,
        "details_meta": {"jira_base_url": _jira_base_url()},
        "warnings": warnings,
    }
    output_path.write_text(_build_html(payload), encoding="utf-8")

    print(f"Source data: {source_file}")
    print(f"Days with planned leaves: {len(counts_by_date)}")
    print(f"Skipped (invalid dates): {skipped_invalid_dates}")
    print(f"Calendar date range: {min_date} to {max_date}")
    print(f"Detail dates with issue mapping: {len(details_by_date)}")
    print(f"Detail dates with unmatched hours: {len(unmatched_by_date)}")
    print(f"HTML report written: {output_path}")


if __name__ == "__main__":
    main()
