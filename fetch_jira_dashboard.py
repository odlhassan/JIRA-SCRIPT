"""
Jira Deadline Dashboard Generator.

Canonical generation path:
- `dashboard_template.html` is the source of UI for generated `dashboard.html`.
- `run_all.py` and this script regenerate `dashboard.html` from the template.

Important:
- Manual edits to `dashboard.html` are overwritten on regeneration.
"""
import json
import os
import re
import sqlite3
import webbrowser
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

from dotenv import load_dotenv

from jira_client import (
    BASE_URL,
    build_dashboard_data,
    extract_jira_key_from_url,
    fetch_stories_for_epics,
    get_issue,
    get_session,
    get_board_id,
)
from managed_projects_registry import deterministic_color_for_project_key, list_managed_projects

load_dotenv()

IPP_XLSX_PATH = os.getenv(
    "IPP_XLSX_PATH",
    r"C:\Users\hmalik\OneDrive - Octopus Digital\ALL DOCS\IPP Meeting",
)
DEFAULT_DASHBOARD_RISK_SETTINGS = {
    "indicator_points": {
        "subtask_linear_lag": 3,
        "due_crossed_unresolved": 3,
        "subtask_late_actual_start": 1,
        "start_passed_not_in_progress": 1,
        "inherited_child_risk": 3,
    },
    "thresholds": {
        "can_be_min": 1,
        "medium_min": 2,
        "high_min": 4,
        "at_risk_min": 2,
    },
    "labels": {
        "low": "Low",
        "can_be": "Can Be",
        "medium": "Medium Risk",
        "high": "Highly At Risk",
    },
}


def _extract_dates_from_planned_cell(text):
    """Extract dates from Planned/Actual Date cell. Supports DD-Mon-YYYY and DD-Mon (e.g. 19-Jan-2026 - 27-Jan-2026)."""
    dates = []
    now = datetime.now()
    months = {
        "Jan": 1, "Feb": 2, "Mar": 3, "Apr": 4, "May": 5, "Jun": 6,
        "Jul": 7, "Aug": 8, "Sep": 9, "Oct": 10, "Nov": 11, "Dec": 12,
    }
    # Match DD-Mon-YYYY or DD-Mon (hyphen, en-dash, or space)
    for match in re.finditer(r"\b(\d{1,2})[-–\s]([A-Za-z]{3})(?:[-–\s](\d{2,4}))?\b", text, re.I):
        day = int(match.group(1))
        mon_str = match.group(2).capitalize()[:3]
        year_str = match.group(3)
        month = months.get(mon_str)
        if not month:
            continue
        if year_str:
            y = int(year_str)
            year = y if y > 99 else 2000 + y
        else:
            year = now.year
            if month < now.month:
                year = now.year - 1
            elif month > now.month + 1:
                year = now.year - 1
        try:
            dates.append(datetime(year, month, day).date())
        except ValueError:
            pass
    return dates


def get_omniconnect_from_xlsx():
    """
    Extract OmniConnect data from the IPP Meeting Excel file.
    XLSX structure: Product, Epic/RMI, Jira Task ID, Planned Date, Actual Date (Production Date), Sprint, Work Status, RMI-on-track-or-not Status.
    Note: Epic/RMI status is sourced from Jira, not from the Excel "Work Status" column.
    Returns (min_str, max_str, dates_iso, production_dates_iso, epic_actual_rows, jira_keys_by_actual, excel_jira_keys, unique_epics, epic_details).
    """
    path = IPP_XLSX_PATH.strip()
    if not path:
        return None, None, [], [], [], {}, set(), [], {}
    if os.path.isdir(path):
        xlsx_files = [
            os.path.join(path, f)
            for f in os.listdir(path)
            if f.lower().endswith(".xlsx")
        ]
        if not xlsx_files:
            return None, None, [], [], [], {}, set(), [], {}
        path = max(xlsx_files, key=os.path.getmtime)
    if not os.path.isfile(path):
        return None, None, [], [], [], {}, set(), [], {}

    try:
        import openpyxl
        import shutil

        local_path = path
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception:
            script_dir = os.path.dirname(os.path.abspath(__file__))
            local_path = os.path.join(script_dir, "_ipp_temp.xlsx")
            shutil.copy2(path, local_path)
            wb = openpyxl.load_workbook(local_path, read_only=True, data_only=True)

        planned_dates = []
        actual_dates = []
        epic_actual_rows = []
        jira_keys_by_actual = defaultdict(list)
        excel_jira_keys = set()  # All Jira keys mentioned in Excel for OmniConnect
        unique_epics = set()  # All unique Epic/RMI names
        epic_details = {}  # Dictionary mapping epic name to {work_status_from_jira, production_date}

        sheet = None
        for sn in wb.sheetnames:
            if "Latest" in sn or "latest" in sn:
                sheet = wb[sn]
                break
        if sheet is None:
            sheet = wb[wb.sheetnames[0]]

        headers = [sheet.cell(1, c).value for c in range(1, 9)]
        product_col = next((i for i, h in enumerate(headers) if h and "Product" in str(h)), 0)
        epic_col = next((i for i, h in enumerate(headers) if h and ("Epic" in str(h) or "RMI" in str(h))), 1)
        jira_col = next((i for i, h in enumerate(headers) if h and "Jira" in str(h)), 2)
        planned_col = next((i for i, h in enumerate(headers) if h and "Planned" in str(h) and "Date" in str(h)), 3)
        actual_col = next((i for i, h in enumerate(headers) if h and "Actual" in str(h) and "Date" in str(h)), 4)
        epic_link_col = next((i for i, h in enumerate(headers) if h and "Epic" in str(h) and "Link" in str(h)), None)

        for r in range(2, sheet.max_row + 1):
            product = sheet.cell(r, product_col + 1).value
            if not product or str(product).strip() != "OmniConnect":
                continue

            epic_name = sheet.cell(r, epic_col + 1).value
            epic_name_str = str(epic_name).strip() if epic_name else None
            if epic_name_str:
                unique_epics.add(epic_name_str)
            jira_val = sheet.cell(r, jira_col + 1).value
            jira_key = extract_jira_key_from_url(jira_val) if jira_val else None

            planned_val = sheet.cell(r, planned_col + 1).value
            planned_dates_for_row = []
            if planned_val:
                row_planned_dates = _extract_dates_from_planned_cell(str(planned_val))
                planned_dates.extend(row_planned_dates)
                planned_dates_for_row = [d.isoformat() for d in row_planned_dates]

            # Extract Epic link - use Jira Task ID column as the primary source
            epic_link = None
            if jira_val:
                jira_val_str = str(jira_val).strip()
                # If it's a URL, use it directly as the Epic Link
                if jira_val_str.startswith('http'):
                    epic_link = jira_val_str
                else:
                    # Try to extract Jira key and build URL
                    key = extract_jira_key_from_url(jira_val_str)
                    if key:
                        epic_link = f"{BASE_URL}/browse/{key}"
                    elif re.match(r'^[A-Za-z0-9]+-\d+$', jira_val_str):
                        # It's just a Jira key
                        epic_link = f"{BASE_URL}/browse/{jira_val_str}"
            
            # If there's a separate Epic Link column, prefer that (but only if Jira Task ID didn't provide a link)
            if not epic_link and epic_link_col is not None:
                epic_link_val = sheet.cell(r, epic_link_col + 1).value
                if epic_link_val:
                    epic_link_str = str(epic_link_val).strip()
                    # If it's a URL, keep it as is; if it's just a key, convert to URL
                    if epic_link_str.startswith('http'):
                        epic_link = epic_link_str
                    else:
                        # Try to extract Jira key from the string (might be a URL or just text with key)
                        key = extract_jira_key_from_url(epic_link_str)
                        if key:
                            epic_link = f"{BASE_URL}/browse/{key}"
                        elif re.match(r'^[A-Za-z0-9]+-\d+$', epic_link_str):
                            # It's just a Jira key
                            epic_link = f"{BASE_URL}/browse/{epic_link_str}"

            # Extract production date
            act_date = None
            actual_val = sheet.cell(r, actual_col + 1).value
            if actual_val:
                if hasattr(actual_val, "date"):
                    act_date = actual_val.date()
                else:
                    parsed = _extract_dates_from_planned_cell(str(actual_val))
                    act_date = parsed[0] if parsed else None
                
                if act_date:
                    actual_dates.append(act_date)
                    act_iso = act_date.isoformat()
                    epic_actual_rows.append((epic_name, act_iso))
                    if jira_key:
                        jira_keys_by_actual[act_iso].append(jira_key)

            # Store epic details (use latest production date if multiple rows exist)
            if epic_name_str:
                if epic_name_str not in epic_details:
                    epic_details[epic_name_str] = {
                        "work_status": "",
                        "production_date": act_date,
                        "epic_link": epic_link,
                        "planned_dates": planned_dates_for_row.copy() if planned_dates_for_row else []
                    }
                else:
                    # Update production date if this one is later
                    if act_date:
                        existing_date = epic_details[epic_name_str]["production_date"]
                        if not existing_date or (act_date > existing_date):
                            epic_details[epic_name_str]["production_date"] = act_date
                    # Update epic link if current is empty
                    if epic_link and not epic_details[epic_name_str].get("epic_link"):
                        epic_details[epic_name_str]["epic_link"] = epic_link
                    # Merge planned dates (avoid duplicates)
                    if planned_dates_for_row:
                        existing_planned = epic_details[epic_name_str].get("planned_dates", [])
                        for pd in planned_dates_for_row:
                            if pd not in existing_planned:
                                existing_planned.append(pd)
                        epic_details[epic_name_str]["planned_dates"] = sorted(existing_planned)

            if jira_key:
                excel_jira_keys.add(jira_key)

        if not planned_dates:
            wb.close()
            return None, None, [], [], [], {}, set(), sorted(unique_epics), epic_details

        min_d = min(planned_dates)
        max_d = max(planned_dates)
        dates_iso = sorted({d.isoformat() for d in planned_dates})
        production_dates_iso = sorted({d.isoformat() for d in actual_dates}) if actual_dates else dates_iso

        wb.close()

        if local_path != path and os.path.isfile(local_path):
            try:
                os.unlink(local_path)
            except OSError:
                pass

        # Convert date objects to formatted strings in epic_details
        formatted_epic_details = {}
        for epic_name, details in epic_details.items():
            formatted_epic_details[epic_name] = {
                "work_status": details["work_status"],
                "production_date": details["production_date"].strftime("%d-%b-%Y") if details["production_date"] else "",
                "epic_link": details.get("epic_link", ""),
                "planned_dates": details.get("planned_dates", [])
            }

        return (
            min_d.strftime("%d-%b-%Y"),
            max_d.strftime("%d-%b-%Y"),
            dates_iso,
            production_dates_iso,
            epic_actual_rows,
            dict(jira_keys_by_actual),
            excel_jira_keys,
            sorted(unique_epics),
            formatted_epic_details,
        )
    except Exception as e:
        print(f"Warning: Could not read IPP Excel: {e}")
        return None, None, [], [], [], {}, set(), [], {}


def generate_html(data):
    """Legacy inline-HTML generator (deprecated; template path is canonical)."""
    from datetime import datetime, timezone
    data = dict(data)
    data["generated_at"] = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    data["epics"] = {}  # Cards removed for now; API still returns full data for Refresh

    json_data = json.dumps(data, default=str)

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Jira Deadline Dashboard - O2</title>
    <link href="https://fonts.googleapis.com/icon?family=Material+Icons" rel="stylesheet">
    <style>
        * {{ box-sizing: border-box; }}
        body {{
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
            max-width: 900px;
            margin: 0 auto;
            padding: 1.5rem;
            background: #f5f5f5;
            color: #333;
        }}
        h1 {{
            margin: 0 0 0.5rem 0;
            font-size: 1.5rem;
        }}
        .meta {{
            color: #666;
            font-size: 0.875rem;
            margin-bottom: 1rem;
        }}
        .header-calendar {{
            background: #fff;
            border-radius: 8px;
            padding: 1rem;
            margin-bottom: 1.5rem;
            box-shadow: 0 1px 3px rgba(0,0,0,0.1);
            overflow-x: auto;
        }}
        .header-calendar h3 {{
            margin: 0 0 0.75rem 0;
            font-size: 0.85rem;
            color: #0747a6;
            font-weight: 600;
        }}
        .calendar-months {{
            display: flex;
            flex-wrap: wrap;
            gap: 1.5rem;
        }}
        .calendar-month {{
            min-width: 160px;
        }}
        .calendar-month-title {{
            font-size: 0.75rem;
            font-weight: 600;
            color: #505f79;
            margin-bottom: 0.25rem;
            text-align: center;
        }}
        .calendar-dow {{
            display: grid;
            grid-template-columns: repeat(7, 1fr);
            font-size: 0.6rem;
            color: #97a0af;
            margin-bottom: 0.15rem;
        }}
        .calendar-dow span {{
            text-align: center;
        }}
        .calendar-days {{
            display: grid;
            grid-template-columns: repeat(7, 1fr);
            gap: 2px;
            font-size: 0.7rem;
        }}
        .calendar-day {{
            aspect-ratio: 1;
            display: flex;
            align-items: center;
            justify-content: center;
            border-radius: 4px;
            min-width: 18px;
            min-height: 18px;
            color: #505f79;
        }}
        .calendar-day.other-month {{
            color: #c1c7d0;
        }}
        .calendar-day.today {{
            background: #0747a6;
            color: #fff;
            font-weight: 600;
        }}
        .calendar-day.nearest-production {{
            background: #ff5630;
            color: #fff;
            font-weight: 600;
        }}
        .calendar-legend {{
            margin-top: 0.75rem;
            font-size: 0.7rem;
            color: #505f79;
        }}
        .calendar-legend .leg-today {{ color: #0747a6; font-weight: 600; }}
        .calendar-legend .leg-production {{ color: #ff5630; font-weight: 600; margin-left: 0.5rem; }}
        .epics-section {{
            background: #fff;
            border-radius: 8px;
            padding: 1rem;
            margin-bottom: 1.5rem;
            box-shadow: 0 1px 3px rgba(0,0,0,0.1);
        }}
        .epics-section h3 {{
            margin: 0;
            font-size: 0.85rem;
            color: #0747a6;
            font-weight: 600;
        }}
        .epics-header {{
            display: flex;
            align-items: flex-end;
            justify-content: space-between;
            gap: 1rem;
            margin-bottom: 0.5rem;
        }}
        .epics-search {{
            position: relative;
            min-width: 0;
            flex: 0 0 260px;
        }}
        .epics-search-input {{
            width: 100%;
            padding: 0.35rem 1.75rem 0.35rem 1.75rem;
            border-radius: 16px;
            border: 1px solid #dfe1e6;
            font-size: 0.8rem;
            color: #172b4d;
            background: #fafbfc;
            outline: none;
            transition: border-color 0.2s, box-shadow 0.2s, background 0.2s;
        }}
        .epics-search-input::placeholder {{
            color: #97a0af;
        }}
        .epics-search-input:focus {{
            border-color: #0747a6;
            box-shadow: 0 0 0 2px rgba(7,71,166,0.2);
            background: #fff;
        }}
        .epics-search-icon {{
            position: absolute;
            left: 0.5rem;
            top: 50%;
            transform: translateY(-50%);
            font-size: 1rem;
            color: #97a0af;
        }}
        .epics-search-clear {{
            position: absolute;
            right: 0.4rem;
            top: 50%;
            transform: translateY(-50%);
            border: none;
            background: transparent;
            color: #97a0af;
            cursor: pointer;
            font-size: 0.9rem;
            padding: 0;
            line-height: 1;
        }}
        .epics-search-clear:hover {{
            color: #172b4d;
        }}
        .epics-search-hint {{
            font-size: 0.7rem;
            color: #97a0af;
            text-align: right;
            margin-bottom: 0.75rem;
        }}
        .epics-table {{
            width: 100%;
            border-collapse: collapse;
            font-size: 0.875rem;
        }}
        .epics-table th {{
            background: #f4f5f7;
            padding: 0.75rem;
            text-align: left;
            font-weight: 600;
            color: #505f79;
            border-bottom: 2px solid #dfe1e6;
            position: relative;
        }}
        .epics-table th.sortable {{
            cursor: pointer;
            user-select: none;
            padding-right: 1.5rem;
        }}
        .epics-table th.sortable:hover {{
            background: #e4e6ea;
        }}
        .epics-table th.sortable::after {{
            content: ' ↕';
            position: absolute;
            right: 0.5rem;
            opacity: 0.5;
            font-size: 0.75rem;
        }}
        .epics-table th.sortable.asc::after {{
            content: ' ↑';
            opacity: 1;
        }}
        .epics-table th.sortable.desc::after {{
            content: ' ↓';
            opacity: 1;
        }}
        .epics-table td {{
            padding: 0.75rem;
            color: #505f79;
            border-bottom: 1px solid #dfe1e6;
        }}
        .epics-table td.nearest-date {{
            color: #ff5630;
            font-weight: 600;
        }}
        .epics-table td.status-in-progress {{
            color: #00875a;
            font-weight: 600;
        }}
        .epics-table td.status-done {{
            color: #97a0af;
        }}
        .epics-table td.status-on-hold {{
            color: #ff8c00;
            font-weight: 600;
        }}
        .epics-table tr:last-child td {{
            border-bottom: none;
        }}
        .epics-table tr:hover {{
            background: #f9fafb;
        }}
        .epics-empty {{
            color: #97a0af;
            font-size: 0.875rem;
            font-style: italic;
        }}
        .epic-link-button {{
            display: inline-flex;
            align-items: center;
            gap: 0.25rem;
            padding: 0.35rem 0.65rem;
            margin-left: 0.5rem;
            font-size: 0.75rem;
            border: 1px solid #dfe1e6;
            border-radius: 4px;
            background: #fff;
            color: #0747a6;
            text-decoration: none;
            cursor: pointer;
            transition: all 0.2s;
        }}
        .epic-link-button:hover {{
            background: #f4f5f7;
            border-color: #0747a6;
        }}
        .epic-link-button:active {{
            background: #e4e6ea;
        }}
        .epic-link-button.no-link {{
            color: #97a0af;
            cursor: not-allowed;
            background: #f4f5f7;
        }}
        .epic-link-button.no-link:hover {{
            border-color: #dfe1e6;
            background: #f4f5f7;
        }}
        .epic-link-button .material-icons {{
            font-size: 0.9em;
        }}
        .epic-name-cell {{
            display: flex;
            align-items: center;
            flex-wrap: wrap;
            gap: 0.5rem;
        }}
        .material-icons {{
            font-family: 'Material Icons';
            font-weight: normal;
            font-style: normal;
            font-size: 1em;
            line-height: 1;
            letter-spacing: normal;
            text-transform: none;
            display: inline-block;
            white-space: nowrap;
            word-wrap: normal;
            direction: ltr;
            vertical-align: middle;
        }}
        .icon-text {{
            display: inline-flex;
            align-items: center;
            gap: 0.35rem;
        }}
        .accordion-container {{
            display: flex;
            flex-direction: column;
            gap: 0.5rem;
        }}
        .accordion-item {{
            background: #fff;
            border: 1px solid #dfe1e6;
            border-radius: 4px;
            overflow: hidden;
        }}
        .accordion-header {{
            display: grid;
            grid-template-columns: auto 1fr auto auto auto;
            gap: 1rem;
            align-items: center;
            padding: 0.75rem;
            cursor: pointer;
            background: #f4f5f7;
            transition: background 0.2s;
            user-select: none;
        }}
        .accordion-header:hover {{
            background: #e4e6ea;
        }}
        .accordion-header.active {{
            background: #e3fcef;
            border-bottom: 1px solid #dfe1e6;
        }}
        .accordion-toggle {{
            display: flex;
            align-items: center;
            justify-content: center;
            width: 24px;
            height: 24px;
            color: #505f79;
        }}
        .accordion-content {{
            max-height: 0;
            overflow: hidden;
            transition: max-height 0.3s ease-out;
        }}
        .accordion-content.open {{
            max-height: 5000px;
            transition: max-height 0.5s ease-in;
        }}
        .accordion-body {{
            padding: 1rem;
            border-top: 1px solid #dfe1e6;
        }}
        .task-accordion {{
            margin-bottom: 0.5rem;
        }}
        .task-accordion:last-child {{
            margin-bottom: 0;
        }}
        .task-header {{
            display: flex;
            align-items: center;
            gap: 0.5rem;
            padding: 0.5rem;
            background: #f9fafb;
            border: 1px solid #dfe1e6;
            border-radius: 4px;
            cursor: pointer;
            transition: background 0.2s;
        }}
        .task-header:hover {{
            background: #f4f5f7;
        }}
        .task-header.active {{
            background: #e3fcef;
        }}
        .task-content {{
            max-height: 0;
            overflow: hidden;
            transition: max-height 0.3s ease-out;
        }}
        .task-content.open {{
            max-height: 2000px;
            transition: max-height 0.4s ease-in;
        }}
        .task-body {{
            padding: 0.75rem;
            background: #fff;
            border-left: 2px solid #0747a6;
        }}
        .subtask-list {{
            list-style: none;
            padding: 0;
            margin: 0.5rem 0 0 0;
        }}
        .subtask-item {{
            display: flex;
            align-items: center;
            gap: 0.5rem;
            padding: 0.5rem;
            border-bottom: 1px solid #f4f5f7;
        }}
        .subtask-item:last-child {{
            border-bottom: none;
        }}
        .task-link, .subtask-link {{
            color: #0747a6;
            text-decoration: none;
            font-weight: 500;
        }}
        .task-link:hover, .subtask-link:hover {{
            text-decoration: underline;
        }}
        .status-badge {{
            display: inline-block;
            padding: 0.25rem 0.5rem;
            border-radius: 3px;
            font-size: 0.75rem;
            font-weight: 600;
        }}
        .status-badge.in-progress {{
            background: #e3fcef;
            color: #00875a;
        }}
        .status-badge.done {{
            background: #f4f5f7;
            color: #97a0af;
        }}
        .status-badge.on-hold {{
            background: #fff4e5;
            color: #ff8c00;
        }}
        .status-badge.unknown {{
            background: #f4f5f7;
            color: #505f79;
        }}
    </style>
</head>
<body>
    <h1>Jira Deadline Dashboard</h1>
    <p class="meta">
        Board: O2 | Generated: <span id="generated-at"></span>
    </p>
    <div id="header-calendar" class="header-calendar"></div>
    <div id="epics-section" class="epics-section"></div>

    <script>
        let dashboardData = {json_data};

        (function() {{
            const d = document.getElementById('dashboard');
            if (d) d.remove();
        }})();

        let epicSortState = {{ column: null, direction: 'asc' }};
        let epicSearchQuery = '';

        function updateUI() {{
            try {{
                document.getElementById('generated-at').textContent = dashboardData.generated_at || 'N/A';
                renderCalendar();
                renderEpics();
            }} catch (error) {{
                console.error('Error in updateUI:', error);
            }}
        }}

        updateUI();

        function renderCalendar() {{
            const container = document.getElementById('header-calendar');
            const dates = dashboardData.planned_dates || [];
            const productionDates = dashboardData.production_dates || [];
            if (dates.length === 0) {{ container.style.display = 'none'; return; }}
            container.style.display = 'block';
            const minDate = dates[0];
            const maxDate = dates[dates.length - 1];
            const today = new Date();
            const todayISO = today.getFullYear() + '-' + String(today.getMonth() + 1).padStart(2,'0') + '-' + String(today.getDate()).padStart(2,'0');
            const futureOrToday = (productionDates.length > 0 ? productionDates : dates).filter(d => d >= todayISO);
            const nearestProduction = futureOrToday.length > 0 ? futureOrToday[0] : (productionDates.length > 0 ? productionDates[productionDates.length - 1] : maxDate);
            const DOW = ['Su','Mo','Tu','We','Th','Fr','Sa'];
            const months = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec'];

            function getMonthDates(year, month) {{
                const first = new Date(year, month, 1);
                const last = new Date(year, month + 1, 0);
                const startPad = first.getDay();
                const days = last.getDate();
                const arr = [];
                for (let i = 0; i < startPad; i++) arr.push(null);
                for (let d = 1; d <= days; d++) arr.push({{year, month, day: d}});
                return arr;
            }}

            function toISO(y, m, d) {{
                const ms = m + 1;
                return y + '-' + String(ms).padStart(2,'0') + '-' + String(d).padStart(2,'0');
            }}

            const minParts = minDate.split('-').map(Number);
            const maxParts = maxDate.split('-').map(Number);
            const minY = minParts[0], minM = minParts[1] - 1, minD = minParts[2];
            const maxY = maxParts[0], maxM = maxParts[1] - 1, maxD = maxParts[2];

            let html = '<h3>OmniConnect Planned Dates</h3><div class="calendar-months">';
            let y = minY, m = minM;
            while (y < maxY || (y === maxY && m <= maxM)) {{
                const monthDates = getMonthDates(y, m);
                const monthTitle = months[m] + ' ' + y;
                html += '<div class="calendar-month"><div class="calendar-month-title">' + monthTitle + '</div>';
                html += '<div class="calendar-dow">' + DOW.map(d => '<span>' + d + '</span>').join('') + '</div>';
                html += '<div class="calendar-days">';
                for (const cell of monthDates) {{
                    if (cell === null) {{
                        html += '<div class="calendar-day other-month"></div>';
                    }} else {{
                        const iso = toISO(cell.year, cell.month, cell.day);
                        const isOther = (cell.year === y && cell.month === m) ? false : true;
                        let cls = 'calendar-day' + (isOther ? ' other-month' : '');
                        if (iso === todayISO) cls += ' today';
                        else if (iso === nearestProduction) cls += ' nearest-production';
                        html += '<div class="' + cls + '" title="' + iso + '">' + cell.day + '</div>';
                    }}
                }}
                html += '</div></div>';
                m++;
                if (m > 11) {{ m = 0; y++; }}
            }}
            html += '</div></div>';
            html += '<div class="calendar-legend"><span class="leg-today">Today: ' + todayISO + '</span> <span class="leg-production">Nearest production: ' + nearestProduction + '</span></div>';
            container.innerHTML = html;
        }}

        function parseDate(dateStr) {{
            if (!dateStr || dateStr === '-') return null;
            const months = {{ 'Jan': 1, 'Feb': 2, 'Mar': 3, 'Apr': 4, 'May': 5, 'Jun': 6, 'Jul': 7, 'Aug': 8, 'Sep': 9, 'Oct': 10, 'Nov': 11, 'Dec': 12 }};
            const parts = dateStr.split('-');
            if (parts.length === 3) {{
                const day = parseInt(parts[0]);
                const month = months[parts[1]];
                const year = parseInt(parts[2]);
                if (month && day && year) {{
                    return new Date(year, month - 1, day);
                }}
            }}
            return null;
        }}

        function sortEpics(epics, epicDetails, column, direction) {{
            const sorted = [...epics];
            sorted.sort((a, b) => {{
                let aVal, bVal;
                
                if (column === 'name') {{
                    aVal = a.toLowerCase();
                    bVal = b.toLowerCase();
                    return direction === 'asc' ? aVal.localeCompare(bVal) : bVal.localeCompare(aVal);
                }} else if (column === 'status') {{
                    aVal = (epicDetails[a]?.work_status || '-').toLowerCase();
                    bVal = (epicDetails[b]?.work_status || '-').toLowerCase();
                    return direction === 'asc' ? aVal.localeCompare(bVal) : bVal.localeCompare(aVal);
                }} else if (column === 'date') {{
                    const aDate = parseDate(epicDetails[a]?.production_date);
                    const bDate = parseDate(epicDetails[b]?.production_date);
                    if (!aDate && !bDate) return 0;
                    if (!aDate) return 1;
                    if (!bDate) return -1;
                    return direction === 'asc' ? aDate - bDate : bDate - aDate;
                }}
                return 0;
            }});
            return sorted;
        }}

        window.handleSort = function(column) {{
            if (epicSortState.column === column) {{
                epicSortState.direction = epicSortState.direction === 'asc' ? 'desc' : 'asc';
            }} else {{
                epicSortState.column = column;
                epicSortState.direction = 'asc';
            }}
            renderEpics();
        }};

        function renderEpics() {{
            try {{
            const container = document.getElementById('epics-section');
            if (!container) {{
                console.error('epics-section container not found');
                return;
            }}
            const epicDetails = dashboardData.epic_details || {{}};
            const nearestProductionDate = dashboardData.nearest_production_date || '';
            const allEpics = dashboardData.unique_epics || [];

            if (allEpics.length === 0) {{
                container.innerHTML = '<h3>RMIs / Epics</h3><div class="epics-empty">No RMIs/Epics found in Excel file.</div>';
                return;
            }}

            // Filter by search query
            const query = (epicSearchQuery || '').trim().toLowerCase();
            let epics = allEpics;
            if (query) {{
                epics = allEpics.filter(epicName => {{
                    const details = epicDetails[epicName] || {{}};
                    const workStatus = (details.work_status || '').toLowerCase();
                    const stories = details.stories || [];

                    const inName = epicName.toLowerCase().includes(query);
                    const inStatus = workStatus.includes(query);
                    const inStories = stories.some(story => {{
                        const key = (story.key || '').toLowerCase();
                        const summary = (story.summary || '').toLowerCase();
                        return key.includes(query) || summary.includes(query);
                    }});

                    return inName || inStatus || inStories;
                }});
            }}

            // Sort if needed
            if (epicSortState.column) {{
                epics = sortEpics(epics, epicDetails, epicSortState.column, epicSortState.direction);
            }}

            const totalEpics = allEpics.length;
            const visibleCount = epics.length;

            // Convert nearest production date to display format for comparison
            let nearestDateDisplay = '';
            if (nearestProductionDate) {{
                const nearestDate = new Date(nearestProductionDate);
                const months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec'];
                nearestDateDisplay = String(nearestDate.getDate()).padStart(2, '0') + '-' + months[nearestDate.getMonth()] + '-' + nearestDate.getFullYear();
            }}

            // Build header and search UI
            let html = '<div class="epics-header">';
            if (visibleCount === totalEpics) {{
                html += '<h3>RMIs / Epics (' + totalEpics + ')</h3>';
            }} else {{
                html += '<h3>RMIs / Epics (' + visibleCount + ' of ' + totalEpics + ')</h3>';
            }}
            html += '<div class="epics-search">';
            html += '<span class="material-icons epics-search-icon">search</span>';
            html += '<input type="text" id="epics-search-input" class="epics-search-input" placeholder="Search epics, status, or tasks..." value="' + escapeHtml(epicSearchQuery) + '" oninput="handleEpicSearchChange(this.value)" />';
            if (query) {{
                html += '<button type="button" class="epics-search-clear" onclick="clearEpicSearch()">×</button>';
            }}
            html += '</div></div>';
            html += '<div class="epics-search-hint">Search matches epic name, work status, story key and summary.</div>';

            if (visibleCount === 0) {{
                html += '<div class="epics-empty">No RMIs/Epics match this search.</div>';
                container.innerHTML = html;
                return;
            }}

            html += '<div class="accordion-container">';
            
            function getStatusBadgeClass(status) {{
                if (!status) return 'unknown';
                const statusLower = status.toLowerCase();
                if (statusLower.includes('in progress') || statusLower === 'in progress') {{
                    return 'in-progress';
                }} else if (statusLower.includes('done') || statusLower === 'done' || statusLower === 'closed') {{
                    return 'done';
                }} else if (statusLower.includes('on hold') || statusLower === 'on hold' || statusLower === 'hold') {{
                    return 'on-hold';
                }}
                return 'unknown';
            }}
            
            for (let i = 0; i < epics.length; i++) {{
                const epicName = epics[i];
                const details = epicDetails[epicName] || {{}};
                const workStatus = details.work_status || '-';
                const productionDate = details.production_date || '-';
                const stories = details.stories || [];
                
                // Determine status class
                let statusClass = '';
                const statusLower = workStatus.toLowerCase();
                if (statusLower.includes('in progress') || statusLower === 'in progress') {{
                    statusClass = 'status-in-progress';
                }} else if (statusLower.includes('done') || statusLower === 'done') {{
                    statusClass = 'status-done';
                }} else if (statusLower.includes('on hold') || statusLower === 'on hold' || statusLower === 'hold') {{
                    statusClass = 'status-on-hold';
                }}
                
                // Check if this is the nearest production date
                const isNearestDate = productionDate === nearestDateDisplay;
                const dateClass = isNearestDate ? ' nearest-date' : '';
                const statusClassAttr = statusClass ? ' ' + statusClass : '';
                
                // Epic link button
                const epicLink = details.epic_link || '';
                let epicLinkButton = '';
                if (epicLink) {{
                    epicLinkButton = '<a href="' + escapeHtml(epicLink) + '" target="_blank" class="epic-link-button"><span class="material-icons">link</span> Epic Link</a>';
                }} else {{
                    epicLinkButton = '<span class="epic-link-button no-link"><span class="material-icons">link_off</span> No Epic Link</span>';
                }}
                
                // Build tasks HTML
                let tasksHtml = '';
                if (stories.length > 0) {{
                    for (let j = 0; j < stories.length; j++) {{
                        const story = stories[j];
                        const taskStatusClass = getStatusBadgeClass(story.status);
                        let subtasksHtml = '';
                        if (story.subtasks && story.subtasks.length > 0) {{
                            subtasksHtml = '<ul class="subtask-list">';
                            for (let k = 0; k < story.subtasks.length; k++) {{
                                const subtask = story.subtasks[k];
                                const subtaskStatusClass = getStatusBadgeClass(subtask.status);
                                subtasksHtml += '<li class="subtask-item">';
                                subtasksHtml += '<a href="' + escapeHtml(subtask.url) + '" target="_blank" class="subtask-link">' + escapeHtml(subtask.key) + ': ' + escapeHtml(subtask.summary) + '</a>';
                                subtasksHtml += '<span class="status-badge ' + subtaskStatusClass + '">' + escapeHtml(subtask.status) + '</span>';
                                subtasksHtml += '</li>';
                            }}
                            subtasksHtml += '</ul>';
                        }}
                        
                        tasksHtml += '<div class="task-accordion">';
                        tasksHtml += '<div class="task-header" onclick="toggleTask(' + i + ', ' + j + ')">';
                        tasksHtml += '<span class="material-icons task-toggle-icon">expand_more</span>';
                        tasksHtml += '<a href="' + escapeHtml(story.url) + '" target="_blank" class="task-link">' + escapeHtml(story.key) + ': ' + escapeHtml(story.summary) + '</a>';
                        tasksHtml += '<span class="status-badge ' + taskStatusClass + '">' + escapeHtml(story.status) + '</span>';
                        tasksHtml += '</div>';
                        tasksHtml += '<div class="task-content" id="task-content-' + i + '-' + j + '">';
                        tasksHtml += '<div class="task-body">';
                        if (subtasksHtml) {{
                            tasksHtml += '<strong>Subtasks:</strong>' + subtasksHtml;
                        }} else {{
                            tasksHtml += '<em>No subtasks</em>';
                        }}
                        tasksHtml += '</div></div></div>';
                    }}
                }} else {{
                    tasksHtml = '<p style="color: #97a0af; font-style: italic;">No tasks found</p>';
                }}
                
                html += '<div class="accordion-item">';
                html += '<div class="accordion-header" onclick="toggleAccordion(' + i + ')" id="accordion-header-' + i + '">';
                html += '<span class="accordion-toggle"><span class="material-icons accordion-icon">expand_more</span></span>';
                html += '<div class="epic-name-cell">' + escapeHtml(epicName) + epicLinkButton + '</div>';
                html += '<div' + statusClassAttr + '>' + escapeHtml(workStatus) + '</div>';
                html += '<div' + (dateClass ? ' class="' + dateClass.trim() + '"' : '') + '>' + escapeHtml(productionDate) + '</div>';
                html += '</div>';
                html += '<div class="accordion-content" id="accordion-content-' + i + '">';
                html += '<div class="accordion-body">';
                html += '<h4 style="margin-top: 0; margin-bottom: 1rem; color: #505f79;">Tasks & Subtasks</h4>';
                html += tasksHtml;
                html += '</div></div></div>';
            }}
            html += '</div>';
            container.innerHTML = html;
            }} catch (error) {{
                console.error('Error in renderEpics:', error);
                const container = document.getElementById('epics-section');
                if (container) {{
                    container.innerHTML = '<h3>RMIs / Epics</h3><div class="epics-empty">Error loading epics: ' + escapeHtml(error.message) + '</div>';
                }}
            }}
        }}

        window.handleEpicSearchChange = function(value) {{
            epicSearchQuery = value || '';
            renderEpics();
        }};

        window.clearEpicSearch = function() {{
            epicSearchQuery = '';
            const input = document.getElementById('epics-search-input');
            if (input) {{
                input.value = '';
            }}
            renderEpics();
        }};

        function escapeHtml(text) {{
            const div = document.createElement('div');
            div.textContent = text;
            return div.innerHTML;
        }}

        function getStatusBadgeClass(status) {{
            if (!status) return 'unknown';
            const statusLower = status.toLowerCase();
            if (statusLower.includes('in progress') || statusLower === 'in progress') {{
                return 'in-progress';
            }} else if (statusLower.includes('done') || statusLower === 'done' || statusLower === 'closed') {{
                return 'done';
            }} else if (statusLower.includes('on hold') || statusLower === 'on hold' || statusLower === 'hold') {{
                return 'on-hold';
            }}
            return 'unknown';
        }}

        let openAccordionIndex = null;

        function toggleAccordion(index) {{
            const content = document.getElementById('accordion-content-' + index);
            const header = document.getElementById('accordion-header-' + index);
            const icon = header.querySelector('.accordion-icon');
            
            // Close previously open accordion
            if (openAccordionIndex !== null && openAccordionIndex !== index) {{
                const prevContent = document.getElementById('accordion-content-' + openAccordionIndex);
                const prevHeader = document.getElementById('accordion-header-' + openAccordionIndex);
                const prevIcon = prevHeader.querySelector('.accordion-icon');
                prevContent.classList.remove('open');
                prevHeader.classList.remove('active');
                prevIcon.textContent = 'expand_more';
            }}
            
            // Toggle current accordion
            if (content.classList.contains('open')) {{
                content.classList.remove('open');
                header.classList.remove('active');
                icon.textContent = 'expand_more';
                openAccordionIndex = null;
            }} else {{
                content.classList.add('open');
                header.classList.add('active');
                icon.textContent = 'expand_less';
                openAccordionIndex = index;
            }}
        }}

        function toggleTask(epicIndex, taskIndex) {{
            const content = document.getElementById('task-content-' + epicIndex + '-' + taskIndex);
            const header = content.previousElementSibling;
            const icon = header.querySelector('.task-toggle-icon');
            
            if (content.classList.contains('open')) {{
                content.classList.remove('open');
                header.classList.remove('active');
                icon.textContent = 'expand_more';
            }} else {{
                content.classList.add('open');
                header.classList.add('active');
                icon.textContent = 'expand_less';
            }}
        }}

    </script>
</body>
</html>
"""
    return html


def fetch_dashboard_data():
    """Build hierarchical dashboard data from generated export workbooks."""
    from openpyxl import load_workbook
    from ipp_meeting_utils import load_ipp_planned_dates_by_key, normalize_issue_key

    script_dir = Path(__file__).resolve().parent

    def resolve_path(env_var: str, default_name: str) -> Path:
        value = os.getenv(env_var, default_name).strip() or default_name
        path = Path(value)
        if not path.is_absolute():
            path = script_dir / path
        return path

    def read_rows(path: Path):
        if not path.exists():
            print(f"Warning: export file not found: {path}")
            return []
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            wb.close()
            return []
        headers = [str(h).strip() if h is not None else "" for h in header_row]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append({headers[i]: row[i] for i in range(len(headers))})
        wb.close()
        return rows

    def as_text(value):
        return str(value or "").strip()

    def normalize_date_text(value):
        text = as_text(value)
        if not text:
            return ""
        m = re.match(r"^(\d{4}-\d{2}-\d{2})", text)
        if m:
            return m.group(1)
        try:
            return datetime.fromisoformat(text.replace("Z", "+00:00")).date().isoformat()
        except ValueError:
            pass
        for fmt in ("%d-%b-%Y", "%d-%B-%Y", "%m/%d/%Y", "%d/%m/%Y"):
            try:
                return datetime.strptime(text, fmt).date().isoformat()
            except ValueError:
                continue
        return text

    def choose_earlier_date(existing, candidate):
        existing_text = normalize_date_text(existing)
        candidate_text = normalize_date_text(candidate)
        if not existing_text:
            return candidate_text
        if not candidate_text:
            return existing_text
        try:
            return min(existing_text, candidate_text)
        except TypeError:
            return existing_text

    def choose_later_date(existing, candidate):
        existing_text = normalize_date_text(existing)
        candidate_text = normalize_date_text(candidate)
        if not existing_text:
            return candidate_text
        if not candidate_text:
            return existing_text
        try:
            return max(existing_text, candidate_text)
        except TypeError:
            return existing_text

    def as_yes_no(value):
        return "Yes" if as_text(value).lower() == "yes" else "No"

    def merge_yes_no(existing, incoming):
        return "Yes" if as_yes_no(existing) == "Yes" or as_yes_no(incoming) == "Yes" else "No"

    def as_float(value):
        try:
            return float(value or 0)
        except (TypeError, ValueError):
            return 0.0

    def normalize_dashboard_risk_settings(raw):
        defaults = json.loads(json.dumps(DEFAULT_DASHBOARD_RISK_SETTINGS))
        settings_in = raw if isinstance(raw, dict) else {}
        points_in = settings_in.get("indicator_points") if isinstance(settings_in.get("indicator_points"), dict) else {}
        thresholds_in = settings_in.get("thresholds") if isinstance(settings_in.get("thresholds"), dict) else {}
        labels_in = settings_in.get("labels") if isinstance(settings_in.get("labels"), dict) else {}

        def as_int(value, default_value):
            try:
                return max(0, int(value))
            except (TypeError, ValueError):
                return int(default_value)

        points = {}
        for key, default_value in (defaults.get("indicator_points") or {}).items():
            points[key] = as_int(points_in.get(key, default_value), default_value)

        thresholds = {
            "can_be_min": as_int(thresholds_in.get("can_be_min", (defaults.get("thresholds") or {}).get("can_be_min", 1)), 1),
            "medium_min": as_int(thresholds_in.get("medium_min", (defaults.get("thresholds") or {}).get("medium_min", 2)), 2),
            "high_min": as_int(thresholds_in.get("high_min", (defaults.get("thresholds") or {}).get("high_min", 4)), 4),
            "at_risk_min": as_int(thresholds_in.get("at_risk_min", (defaults.get("thresholds") or {}).get("at_risk_min", 2)), 2),
        }
        if thresholds["can_be_min"] > thresholds["medium_min"]:
            thresholds["medium_min"] = thresholds["can_be_min"]
        if thresholds["medium_min"] > thresholds["high_min"]:
            thresholds["high_min"] = thresholds["medium_min"]
        if thresholds["at_risk_min"] < thresholds["medium_min"]:
            thresholds["at_risk_min"] = thresholds["medium_min"]
        if thresholds["at_risk_min"] > thresholds["high_min"]:
            thresholds["at_risk_min"] = thresholds["high_min"]

        labels = {}
        for key, default_value in (defaults.get("labels") or {}).items():
            label = as_text(labels_in.get(key, default_value))
            labels[key] = label if label else default_value

        return {
            "indicator_points": points,
            "thresholds": thresholds,
            "labels": labels,
        }

    def load_dashboard_risk_settings(db_path: Path):
        if not db_path.exists():
            return normalize_dashboard_risk_settings({})
        conn = None
        try:
            conn = sqlite3.connect(db_path)
            conn.row_factory = sqlite3.Row
            exists = conn.execute(
                "SELECT 1 FROM sqlite_master WHERE type='table' AND name='dashboard_risk_settings'"
            ).fetchone()
            if not exists:
                return normalize_dashboard_risk_settings({})
            row = conn.execute(
                "SELECT settings_json FROM dashboard_risk_settings WHERE id=1"
            ).fetchone()
            raw_json = as_text(row["settings_json"]) if row else ""
            if not raw_json:
                return normalize_dashboard_risk_settings({})
            try:
                parsed = json.loads(raw_json)
            except Exception:
                parsed = {}
            return normalize_dashboard_risk_settings(parsed)
        except Exception:
            return normalize_dashboard_risk_settings({})
        finally:
            if conn is not None:
                conn.close()

    def is_resolved_status(value):
        status_text = as_text(value).lower().replace("-", " ").replace("_", " ").strip()
        if not status_text:
            return False
        return ("resolved" in status_text) or ("done" in status_text) or ("closed" in status_text)

    def parse_original_estimate_hours(value):
        text = as_text(value)
        if not text:
            return 0.0
        try:
            numeric = float(text)
            if numeric == numeric and numeric > 0:
                return numeric
        except (TypeError, ValueError):
            pass
        token_pattern = re.compile(r"(\d+(?:\.\d+)?)\s*([wdhm])", re.IGNORECASE)
        total_hours = 0.0
        matched = False
        for amount_text, unit_text in token_pattern.findall(text):
            matched = True
            try:
                amount = float(amount_text)
            except ValueError:
                continue
            unit = unit_text.lower()
            if unit == "w":
                total_hours += amount * 40.0
            elif unit == "d":
                total_hours += amount * 8.0
            elif unit == "h":
                total_hours += amount
            elif unit == "m":
                total_hours += amount / 60.0
        if matched and total_hours > 0:
            return total_hours
        return 0.0

    def parse_json_object(value):
        text = as_text(value)
        if not text:
            return {}
        try:
            parsed = json.loads(text)
        except Exception:
            return {}
        return parsed if isinstance(parsed, dict) else {}

    def parse_planner_hours_from_man_days(value):
        text = as_text(value)
        if not text:
            return None
        try:
            man_days = float(text)
        except (TypeError, ValueError):
            return None
        if man_days != man_days or man_days < 0:
            return None
        return round(man_days * 8.0, 4)

    def load_epics_planner_epic_plan_by_key(db_path: Path):
        if not db_path.exists():
            return {}
        try:
            conn = sqlite3.connect(db_path)
            conn.row_factory = sqlite3.Row
        except Exception:
            return {}
        try:
            exists = conn.execute(
                "SELECT 1 FROM sqlite_master WHERE type='table' AND name='epics_management'"
            ).fetchone()
            if not exists:
                return {}
            rows = conn.execute("SELECT epic_key, epic_plan_json FROM epics_management").fetchall()
        except Exception:
            return {}
        finally:
            conn.close()

        lookup = {}
        for row in rows:
            epic_key = normalize_issue_key(as_text(row["epic_key"]))
            if not epic_key:
                continue
            epic_plan = parse_json_object(row["epic_plan_json"])
            planner_start = normalize_date_text(epic_plan.get("start_date"))
            planner_end = normalize_date_text(epic_plan.get("due_date"))
            planner_hours = parse_planner_hours_from_man_days(epic_plan.get("man_days"))
            lookup[epic_key] = {
                "planner_start_date": planner_start,
                "planner_end_date": planner_end,
                "planner_planned_hours": planner_hours,
            }
        return lookup

    def load_epics_planner_story_dates_by_key(db_path: Path):
        if not db_path.exists():
            return {}
        try:
            conn = sqlite3.connect(db_path)
            conn.row_factory = sqlite3.Row
        except Exception:
            return {}
        try:
            exists = conn.execute(
                "SELECT 1 FROM sqlite_master WHERE type='table' AND name='epics_management_story_sync'"
            ).fetchone()
            if not exists:
                return {}
            columns = conn.execute("PRAGMA table_info(epics_management_story_sync)").fetchall()
            column_names = {as_text(col[1]) for col in columns}
            has_estimate_hours = "estimate_hours" in column_names
            if has_estimate_hours:
                rows = conn.execute(
                    "SELECT story_key, start_date, due_date, estimate_hours FROM epics_management_story_sync"
                ).fetchall()
            else:
                rows = conn.execute(
                    "SELECT story_key, start_date, due_date FROM epics_management_story_sync"
                ).fetchall()
        except Exception:
            return {}
        finally:
            conn.close()

        lookup = {}
        for row in rows:
            story_key = normalize_issue_key(as_text(row["story_key"]))
            if not story_key:
                continue
            start_date = normalize_date_text(row["start_date"])
            due_date = normalize_date_text(row["due_date"])
            estimate_hours = None
            try:
                raw_hours = row["estimate_hours"] if "estimate_hours" in row.keys() else None
                estimate_hours = float(raw_hours) if raw_hours is not None else None
            except Exception:
                estimate_hours = None
            if not start_date and not due_date and estimate_hours is None:
                continue
            lookup[story_key] = {
                "start_date": start_date,
                "due_date": due_date,
                "estimate_hours": estimate_hours,
            }
        return lookup

    def parse_iso_date(value):
        text = normalize_date_text(value)
        if not text:
            return None
        try:
            return datetime.fromisoformat(text).date()
        except ValueError:
            return None

    def expected_hours_to_date(planned_hours: float, start_text: str, end_text: str, as_of_date):
        if planned_hours <= 0:
            return 0.0
        start_day = parse_iso_date(start_text)
        end_day = parse_iso_date(end_text)
        if not start_day or not end_day or end_day < start_day:
            return 0.0
        if as_of_date < start_day:
            return 0.0
        total_days = (end_day - start_day).days + 1
        elapsed_days = (min(as_of_date, end_day) - start_day).days + 1
        ratio = max(0.0, min(float(elapsed_days) / float(total_days), 1.0))
        return round(planned_hours * ratio, 2)

    def is_in_progress_status(value):
        status_text = as_text(value).lower().replace("-", " ").replace("_", " ").strip()
        if not status_text:
            return False
        return "in progress" in status_text

    def risk_level_from_score(score: int, thresholds: dict):
        safe_score = max(0, int(score or 0))
        high_min = int(thresholds.get("high_min", 4))
        medium_min = int(thresholds.get("medium_min", 2))
        can_be_min = int(thresholds.get("can_be_min", 1))
        if safe_score >= high_min:
            return "high"
        if safe_score >= medium_min:
            return "medium"
        if safe_score >= can_be_min:
            return "can_be"
        return "low"

    def is_subtask_kind(item):
        issue_type = as_text(item.get("issue_type")).lower()
        return ("subtask" in issue_type) or ("sub-task" in issue_type)

    def compute_self_risk(item: dict, as_of_date, risk_settings: dict):
        points = risk_settings.get("indicator_points") if isinstance(risk_settings.get("indicator_points"), dict) else {}
        thresholds = risk_settings.get("thresholds") if isinstance(risk_settings.get("thresholds"), dict) else {}
        labels = risk_settings.get("labels") if isinstance(risk_settings.get("labels"), dict) else {}
        point_start_passed_not_in_progress = int(points.get("start_passed_not_in_progress", 1))
        point_due_crossed = int(points.get("due_crossed_unresolved", 3))
        point_subtask_late_start = int(points.get("subtask_late_actual_start", 1))
        point_subtask_linear_lag = int(points.get("subtask_linear_lag", 3))
        at_risk_min = int(thresholds.get("at_risk_min", 2))
        status_for_risk = as_text(item.get("jira_status")) or as_text(item.get("status"))
        planned_hours = parse_original_estimate_hours(item.get("original_estimate"))
        logged_hours = as_float(item.get("total_hours_logged"))
        planned_start_day = parse_iso_date(item.get("jira_start_date"))
        planned_end_day = parse_iso_date(item.get("jira_end_date"))
        actual_start_day = parse_iso_date(item.get("actual_start_date"))
        expected_hours = expected_hours_to_date(
            planned_hours,
            item.get("jira_start_date"),
            item.get("jira_end_date"),
            as_of_date,
        )

        item["planned_hours_numeric"] = round(planned_hours, 2)
        item["expected_hours_to_date"] = expected_hours
        item["risk_score_self"] = 0
        item["risk_score_rollup"] = 0
        item["risk_score_final"] = 0
        item["risk_level"] = "low"
        item["risk_level_label"] = labels.get("low", "Low")
        item["risk_reasons"] = []
        item["risk_inherited_from"] = ""
        item["is_at_risk_self"] = False
        item["is_at_risk"] = False

        if is_resolved_status(status_for_risk):
            return

        reasons: list[str] = []
        score = 0

        # Minor: planned start has passed and item is still not in progress.
        if planned_start_day and as_of_date > planned_start_day and not is_in_progress_status(status_for_risk):
            score += point_start_passed_not_in_progress
            reasons.append(f"+{point_start_passed_not_in_progress} Planned start has passed and item is not In Progress.")

        # Major: due date has crossed and item is unresolved.
        if planned_end_day and as_of_date > planned_end_day:
            score += point_due_crossed
            reasons.append(f"+{point_due_crossed} Planned end has passed and item is still unresolved.")

        if is_subtask_kind(item):
            # Minor: actual start is later than planned start.
            if planned_start_day and actual_start_day and actual_start_day > planned_start_day:
                score += point_subtask_late_start
                reasons.append(f"+{point_subtask_late_start} Actual start is later than planned start.")

            # Major: linear progression effort lag against expected effort to date.
            if expected_hours > 0 and logged_hours + 1e-6 < expected_hours:
                score += point_subtask_linear_lag
                reasons.append(
                    f"+{point_subtask_linear_lag} Logged hours ({round(logged_hours, 2)}) are below expected effort to date ({round(expected_hours, 2)})."
                )

        item["risk_score_self"] = int(score)
        item["risk_level"] = risk_level_from_score(score, thresholds)
        item["risk_level_label"] = labels.get(item["risk_level"], item["risk_level"].replace("_", " ").title())
        item["risk_reasons"] = reasons
        item["is_at_risk_self"] = score >= at_risk_min

    def project_from_key(issue_key: str):
        key = as_text(issue_key)
        return key.split("-", 1)[0] if "-" in key else ""

    def build_item(row, issue_type_name: str):
        issue_key = as_text(row.get("issue_key"))
        parent_key = as_text(row.get("parent_issue_key"))
        status = as_text(row.get("status"))
        hours = as_float(row.get("total_hours_logged"))
        item = {
            "issue_key": issue_key,
            "issue_type": issue_type_name,
            "summary": as_text(row.get("summary")),
            "assignee": as_text(row.get("assignee")),
            "project_key": as_text(row.get("project_key")) or project_from_key(issue_key),
            "jira_url": as_text(row.get("jira_url")) or (f"{BASE_URL}/browse/{issue_key}" if issue_key else ""),
            "parent_issue_key": parent_key,
            "epic_key": "",
            "story_key": "",
            "jira_start_date": normalize_date_text(
                row.get("start_date") or row.get("planned start date") or row.get("planned_start_date")
            ),
            "jira_end_date": normalize_date_text(
                row.get("end_date") or row.get("planned end date") or row.get("planned_end_date")
            ),
            "actual_start_date": as_text(row.get("actual_start_date")),
            "actual_end_date": as_text(row.get("actual_end_date")),
            "original_estimate": as_text(row.get("original_estimate")),
            "total_hours_logged": hours,
            "latest_ipp_meeting": as_yes_no(row.get("Latest IPP Meeting", "No")),
            "jira_ipp_rmi_dates_altered": as_yes_no(row.get("Jira IPP RMI Dates Altered", "No")),
            "ipp_actual_date": normalize_date_text(
                row.get("IPP Actual Date (Production Date)") or row.get("ipp_actual_date")
            ),
            "ipp_remarks": as_text(row.get("IPP Remarks") or row.get("ipp_remarks")),
            "ipp_actual_matches_jira_end_date": as_yes_no(
                row.get("IPP Actual Date Matches Jira End Date", "No")
            ),
            "ipp_planned_start_date": normalize_date_text(
                row.get("IPP Planned Start Date") or row.get("ipp_planned_start_date")
            ),
            "ipp_planned_end_date": normalize_date_text(
                row.get("IPP Planned End Date") or row.get("ipp_planned_end_date")
            ),
            "jira_status": status,
            "status": status,
            # Compatibility aliases for existing UI fields.
            "epic_key_legacy": "",
            "epic_status": "",
            "subtask_hours_logged_total": 0.0,
            # Epics Planner validation fields.
            "planner_has_entry": False,
            "planner_planned_start_date": "",
            "planner_planned_end_date": "",
            "planner_planned_hours": None,
            "planner_dates_match": "N/A",
            "planner_hours_match": "N/A",
            "planner_validation_status": "No Planner Entry",
            # Story planner sync fields (from epics_management_story_sync).
            "planner_story_start_date": "",
            "planner_story_end_date": "",
            "planner_story_planned_hours": None,
            # Risk-scoring fields.
            "risk_score_self": 0,
            "risk_score_rollup": 0,
            "risk_score_final": 0,
            "risk_level": "low",
            "risk_level_label": "Low",
            "risk_reasons": [],
            "risk_inherited_from": "",
        }
        return item

    rows1 = read_rows(resolve_path("JIRA_EXPORT_XLSX_PATH", "1_jira_work_items_export.xlsx"))
    rows2 = read_rows(resolve_path("JIRA_WORKLOG_XLSX_PATH", "2_jira_subtask_worklogs.xlsx"))
    rows3 = read_rows(resolve_path("JIRA_SUBTASK_ROLLUP_XLSX_PATH", "3_jira_subtask_worklog_rollup.xlsx"))
    planner_db_path = resolve_path("JIRA_ASSIGNEE_HOURS_CAPACITY_DB_PATH", "assignee_hours_capacity.db")
    risk_settings = load_dashboard_risk_settings(planner_db_path)
    planner_epic_plan_by_key = load_epics_planner_epic_plan_by_key(planner_db_path)
    planner_story_dates_by_key = load_epics_planner_story_dates_by_key(planner_db_path)
    ipp_planned_dates_by_key = load_ipp_planned_dates_by_key()

    def apply_ipp_planned_dates(item: dict, epic_key: str):
        normalized_epic_key = normalize_issue_key(epic_key)
        if not normalized_epic_key:
            return
        ipp_dates = ipp_planned_dates_by_key.get(normalized_epic_key, {})
        item["ipp_planned_start_date"] = normalize_date_text(
            ipp_dates.get("planned_start") or item.get("ipp_planned_start_date")
        )
        item["ipp_planned_end_date"] = normalize_date_text(
            ipp_dates.get("planned_end") or item.get("ipp_planned_end_date")
        )

    epics_map = {}
    stories_map = {}
    subtasks_map = {}
    bug_subtasks_map = {}

    for row in rows1:
        issue_type = as_text(row.get("jira_issue_type")).lower()
        issue_type_raw = as_text(row.get("jira_issue_type"))
        if not issue_type:
            continue
        if "epic" in issue_type:
            item = build_item(row, issue_type_raw or "Epic")
            item["epic_key"] = item["issue_key"]
            apply_ipp_planned_dates(item, item["issue_key"])
            item["epic_key_legacy"] = item["issue_key"]
            item["epic_status"] = item["status"]
            item["subtask_hours_logged_total"] = item["total_hours_logged"]
            epics_map[item["issue_key"]] = item
        elif issue_type == "story":
            item = build_item(row, "Story")
            item["epic_key"] = as_text(item["parent_issue_key"])
            apply_ipp_planned_dates(item, item["epic_key"])
            stories_map[item["issue_key"]] = item
        elif issue_type == "bug subtask":
            item = build_item(row, "Bug Subtask")
            item["story_key"] = as_text(item["parent_issue_key"])
            bug_subtasks_map[item["issue_key"]] = item
        elif issue_type in ("sub-task", "subtask"):
            item = build_item(row, issue_type_raw or "Sub-task")
            item["story_key"] = as_text(item["parent_issue_key"])
            subtasks_map[item["issue_key"]] = item

    for row in rows2:
        epic_key = as_text(row.get("parent_epic_id"))
        if not epic_key:
            continue
        epic = epics_map.get(epic_key)
        if not epic:
            continue
        epic["subtask_hours_logged_total"] = as_float(epic.get("subtask_hours_logged_total")) + as_float(row.get("hours_logged"))
        epic["latest_ipp_meeting"] = merge_yes_no(epic.get("latest_ipp_meeting", "No"), row.get("Latest IPP Meeting", "No"))
        epic["jira_ipp_rmi_dates_altered"] = merge_yes_no(epic.get("jira_ipp_rmi_dates_altered", "No"), row.get("Jira IPP RMI Dates Altered", "No"))
        epic["ipp_actual_matches_jira_end_date"] = merge_yes_no(
            epic.get("ipp_actual_matches_jira_end_date", "No"),
            row.get("IPP Actual Date Matches Jira End Date", "No"),
        )
        epic["ipp_actual_date"] = choose_later_date(
            epic.get("ipp_actual_date"),
            row.get("IPP Actual Date (Production Date)") or row.get("ipp_actual_date"),
        )
        if not as_text(epic.get("ipp_remarks")):
            epic["ipp_remarks"] = as_text(row.get("IPP Remarks") or row.get("ipp_remarks"))

    for row in rows3:
        epic_key = as_text(row.get("parent_epic_id"))
        epic = epics_map.get(epic_key)
        if not epic:
            continue
        planned_start_date = normalize_date_text(row.get("planned start date") or row.get("planned_start_date"))
        story_key = as_text(row.get("parent_story_id"))
        subtask_key = as_text(row.get("issue_id"))

        if planned_start_date:
            epic["jira_start_date"] = choose_earlier_date(epic.get("jira_start_date"), planned_start_date)
            story = stories_map.get(story_key)
            if story:
                story["jira_start_date"] = choose_earlier_date(story.get("jira_start_date"), planned_start_date)
            subtask = subtasks_map.get(subtask_key)
            if subtask and not as_text(subtask.get("jira_start_date")):
                subtask["jira_start_date"] = planned_start_date
            bug_subtask = bug_subtasks_map.get(subtask_key)
            if bug_subtask and not as_text(bug_subtask.get("jira_start_date")):
                bug_subtask["jira_start_date"] = planned_start_date

        epic["latest_ipp_meeting"] = merge_yes_no(epic.get("latest_ipp_meeting", "No"), row.get("Latest IPP Meeting", "No"))
        epic["jira_ipp_rmi_dates_altered"] = merge_yes_no(epic.get("jira_ipp_rmi_dates_altered", "No"), row.get("Jira IPP RMI Dates Altered", "No"))
        epic["ipp_actual_matches_jira_end_date"] = merge_yes_no(
            epic.get("ipp_actual_matches_jira_end_date", "No"),
            row.get("IPP Actual Date Matches Jira End Date", "No"),
        )
        epic["ipp_actual_date"] = choose_later_date(
            epic.get("ipp_actual_date"),
            row.get("IPP Actual Date (Production Date)") or row.get("ipp_actual_date"),
        )
        if not as_text(epic.get("ipp_remarks")):
            epic["ipp_remarks"] = as_text(row.get("IPP Remarks") or row.get("ipp_remarks"))

    for subtask in subtasks_map.values():
        story = stories_map.get(as_text(subtask.get("story_key")))
        subtask["epic_key"] = as_text(story.get("epic_key")) if story else ""
        apply_ipp_planned_dates(subtask, subtask["epic_key"])
    for bug_subtask in bug_subtasks_map.values():
        story = stories_map.get(as_text(bug_subtask.get("story_key")))
        bug_subtask["epic_key"] = as_text(story.get("epic_key")) if story else ""
        apply_ipp_planned_dates(bug_subtask, bug_subtask["epic_key"])

    # Story planned dates: prefer Epics Planner module story-sync dates when available.
    for story in stories_map.values():
        planner_story = planner_story_dates_by_key.get(normalize_issue_key(as_text(story.get("issue_key"))), {})
        if not planner_story:
            continue
        planner_start = normalize_date_text(planner_story.get("start_date"))
        planner_due = normalize_date_text(planner_story.get("due_date"))
        story["planner_story_start_date"] = planner_start
        story["planner_story_end_date"] = planner_due
        story["planner_story_planned_hours"] = planner_story.get("estimate_hours")

    today_utc = datetime.now(timezone.utc).date()
    for item in subtasks_map.values():
        compute_self_risk(item, today_utc, risk_settings)
    for item in bug_subtasks_map.values():
        compute_self_risk(item, today_utc, risk_settings)
    for item in stories_map.values():
        compute_self_risk(item, today_utc, risk_settings)
    for item in epics_map.values():
        compute_self_risk(item, today_utc, risk_settings)
        planner = planner_epic_plan_by_key.get(normalize_issue_key(as_text(item.get("issue_key"))), {})
        has_planner_entry = bool(planner)
        planner_start = normalize_date_text(planner.get("planner_start_date"))
        planner_end = normalize_date_text(planner.get("planner_end_date"))
        planner_hours = planner.get("planner_planned_hours")
        planner_complete = bool(planner_start and planner_end and planner_hours is not None)

        jira_start = normalize_date_text(item.get("jira_start_date"))
        jira_end = normalize_date_text(item.get("jira_end_date"))
        jira_hours = float(item.get("planned_hours_numeric") or 0.0)
        jira_hours_available = as_text(item.get("original_estimate")) != ""

        dates_match = "N/A"
        if planner_complete and jira_start and jira_end:
            dates_match = "Yes" if planner_start == jira_start and planner_end == jira_end else "No"

        hours_match = "N/A"
        if planner_complete and jira_hours_available:
            hours_match = "Yes" if abs(float(planner_hours) - jira_hours) < 0.01 else "No"

        if not has_planner_entry:
            planner_status = "No Planner Entry"
        elif not planner_complete:
            planner_status = "Incomplete"
        elif dates_match == "Yes" and hours_match == "Yes":
            planner_status = "Matched"
        else:
            planner_status = "Mismatch"

        item["planner_has_entry"] = has_planner_entry
        item["planner_planned_start_date"] = planner_start
        item["planner_planned_end_date"] = planner_end
        item["planner_planned_hours"] = planner_hours
        item["planner_dates_match"] = dates_match
        item["planner_hours_match"] = hours_match
        item["planner_validation_status"] = planner_status

    risk_points = risk_settings.get("indicator_points") if isinstance(risk_settings.get("indicator_points"), dict) else {}
    risk_thresholds = risk_settings.get("thresholds") if isinstance(risk_settings.get("thresholds"), dict) else {}
    risk_labels = risk_settings.get("labels") if isinstance(risk_settings.get("labels"), dict) else {}
    inherited_child_points = int(risk_points.get("inherited_child_risk", 3))
    at_risk_min_score = int(risk_thresholds.get("at_risk_min", 2))

    # Finalize subtask/bug-subtask risk directly from self score.
    for subtask in list(subtasks_map.values()) + list(bug_subtasks_map.values()):
        self_score = int(subtask.get("risk_score_self") or 0)
        subtask["risk_score_rollup"] = 0
        subtask["risk_score_final"] = self_score
        subtask["risk_level"] = risk_level_from_score(self_score, risk_thresholds)
        subtask["risk_level_label"] = risk_labels.get(subtask["risk_level"], subtask["risk_level"].replace("_", " ").title())
        subtask["is_at_risk"] = self_score >= at_risk_min_score

    # Story roll-up from subtasks.
    subtasks_by_story: dict[str, list[dict]] = defaultdict(list)
    for subtask in list(subtasks_map.values()) + list(bug_subtasks_map.values()):
        story_key = as_text(subtask.get("story_key"))
        if story_key:
            subtasks_by_story[story_key].append(subtask)

    for story in stories_map.values():
        status_for_risk = as_text(story.get("jira_status")) or as_text(story.get("status"))
        if is_resolved_status(status_for_risk):
            story["risk_score_rollup"] = 0
            story["risk_score_final"] = 0
            story["risk_level"] = "low"
            story["risk_level_label"] = risk_labels.get("low", "Low")
            story["risk_inherited_from"] = ""
            story["is_at_risk"] = False
            story["risk_reasons"] = []
            continue

        children = subtasks_by_story.get(as_text(story.get("issue_key")), [])
        strongest_child = None
        if children:
            strongest_child = max(children, key=lambda child: int(child.get("risk_score_final") or 0))
        rollup_score = int(strongest_child.get("risk_score_final") or 0) if strongest_child else 0

        score_self = int(story.get("risk_score_self") or 0)
        reasons = list(story.get("risk_reasons") or [])
        if strongest_child and bool(strongest_child.get("is_at_risk")):
            score_self += inherited_child_points
            child_key = as_text(strongest_child.get("issue_key"))
            child_level_label = as_text(strongest_child.get("risk_level_label") or strongest_child.get("risk_level") or "low")
            reasons.append(f"+{inherited_child_points} Child subtask risk inherited from {child_key} ({child_level_label}).")
            story["risk_inherited_from"] = child_key
        else:
            story["risk_inherited_from"] = ""

        final_score = max(score_self, rollup_score)
        if strongest_child and rollup_score > score_self:
            reasons.append(
                f"Roll-up used strongest child score: {rollup_score} from {as_text(strongest_child.get('issue_key'))}."
            )

        story["risk_score_self"] = score_self
        story["risk_score_rollup"] = rollup_score
        story["risk_score_final"] = final_score
        story["risk_reasons"] = reasons
        story["risk_level"] = risk_level_from_score(final_score, risk_thresholds)
        story["risk_level_label"] = risk_labels.get(story["risk_level"], story["risk_level"].replace("_", " ").title())
        story["is_at_risk"] = final_score >= at_risk_min_score

    # Epic roll-up from stories.
    stories_by_epic: dict[str, list[dict]] = defaultdict(list)
    for story in stories_map.values():
        epic_key = as_text(story.get("epic_key"))
        if epic_key:
            stories_by_epic[epic_key].append(story)

    for epic in epics_map.values():
        status_for_risk = as_text(epic.get("jira_status")) or as_text(epic.get("status"))
        if is_resolved_status(status_for_risk):
            epic["risk_score_rollup"] = 0
            epic["risk_score_final"] = 0
            epic["risk_level"] = "low"
            epic["risk_level_label"] = risk_labels.get("low", "Low")
            epic["risk_inherited_from"] = ""
            epic["is_at_risk"] = False
            epic["risk_reasons"] = []
            continue

        children = stories_by_epic.get(as_text(epic.get("issue_key")), [])
        strongest_child = None
        if children:
            strongest_child = max(children, key=lambda child: int(child.get("risk_score_final") or 0))
        rollup_score = int(strongest_child.get("risk_score_final") or 0) if strongest_child else 0

        score_self = int(epic.get("risk_score_self") or 0)
        reasons = list(epic.get("risk_reasons") or [])
        if strongest_child and bool(strongest_child.get("is_at_risk")):
            score_self += inherited_child_points
            child_key = as_text(strongest_child.get("issue_key"))
            child_level_label = as_text(strongest_child.get("risk_level_label") or strongest_child.get("risk_level") or "low")
            reasons.append(f"+{inherited_child_points} Child story risk inherited from {child_key} ({child_level_label}).")
            epic["risk_inherited_from"] = child_key
        else:
            epic["risk_inherited_from"] = ""

        final_score = max(score_self, rollup_score)
        if strongest_child and rollup_score > score_self:
            reasons.append(
                f"Roll-up used strongest child score: {rollup_score} from {as_text(strongest_child.get('issue_key'))}."
            )

        epic["risk_score_self"] = score_self
        epic["risk_score_rollup"] = rollup_score
        epic["risk_score_final"] = final_score
        epic["risk_reasons"] = reasons
        epic["risk_level"] = risk_level_from_score(final_score, risk_thresholds)
        epic["risk_level_label"] = risk_labels.get(epic["risk_level"], epic["risk_level"].replace("_", " ").title())
        epic["is_at_risk"] = final_score >= at_risk_min_score

    epics = sorted(epics_map.values(), key=lambda x: (x.get("project_key", ""), x.get("issue_key", "")))
    stories = sorted(stories_map.values(), key=lambda x: (x.get("project_key", ""), x.get("issue_key", "")))
    subtasks = sorted(subtasks_map.values(), key=lambda x: (x.get("project_key", ""), x.get("issue_key", "")))
    bug_subtasks = sorted(bug_subtasks_map.values(), key=lambda x: (x.get("project_key", ""), x.get("issue_key", "")))

    epic_keys = {as_text(e.get("issue_key")) for e in epics}
    story_keys = {as_text(s.get("issue_key")) for s in stories}
    orphan_stories = [s for s in stories if as_text(s.get("epic_key")) not in epic_keys]
    orphan_subtasks = [s for s in subtasks if as_text(s.get("story_key")) not in story_keys]
    orphan_bug_subtasks = [s for s in bug_subtasks if as_text(s.get("story_key")) not in story_keys]

    projects = sorted(
        {
            as_text(i.get("project_key"))
            for i in (epics + stories + subtasks + bug_subtasks)
            if as_text(i.get("project_key"))
        }
    )

    # Load project display names and colors from managed projects database
    project_names = {}
    project_colors = {}
    try:
        managed_projects = list_managed_projects(planner_db_path, include_inactive=False)
        for proj in managed_projects:
            key = proj.get("project_key")
            display = proj.get("display_name") or proj.get("project_name") or key
            color = proj.get("color_hex")
            if key:
                project_names[key] = display
                if color:
                    project_colors[key] = color
    except Exception:
        pass
    # Fallback: deterministic color for projects not in managed_projects
    for p in projects:
        if p not in project_colors:
            try:
                project_colors[p] = deterministic_color_for_project_key(p)
            except Exception:
                project_colors[p] = "#64748b"

    return {
        "epics": epics,
        "stories": stories,
        "subtasks": subtasks,
        "bug_subtasks": bug_subtasks,
        "orphans": {
            "stories": orphan_stories,
            "subtasks": orphan_subtasks,
            "bug_subtasks": orphan_bug_subtasks,
        },
        "risk_config": risk_settings,
        "projects": projects,
        "project_names": project_names,
        "project_colors": project_colors,
        "generated_at": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC"),
    }


def generate_exports_dashboard_html(data):
    """Generate dashboard HTML from template file with embedded data."""
    json_data = json.dumps(data, default=str)
    template_path = Path(__file__).resolve().parent / "dashboard_template.html"
    if not template_path.exists():
        raise FileNotFoundError(f"Dashboard template not found: {template_path}")
    template = template_path.read_text(encoding="utf-8")
    required_tokens = [
        "__DASHBOARD_DATA__",
        'id="calendar-panel"',
        'id="date-filter-from"',
        'id="date-filter-to"',
        'id="date-filter-reset"',
        "top-date-range-chip",
        "function renderGantt(",
        "let selectedEpicKey = ''",
        "function renderAll()",
        'id="bug-subtasks-list"',
        "gantt-axis-track",
        "gantt-date-label",
    ]
    missing = [token for token in required_tokens if token not in template]
    if missing:
        raise ValueError(
            "dashboard_template.html is missing required generator invariants: "
            + ", ".join(missing)
        )
    return template.replace("__DASHBOARD_DATA__", json_data)


def main():
    print("Building dashboard from generated export files...")
    try:
        data = fetch_dashboard_data()
        epics_count = len(data.get("epics", []))
        print(f"Found {epics_count} RMIs/Epics in export files")
        html = generate_exports_dashboard_html(data)
        out_path = os.path.join(os.path.dirname(__file__), "dashboard.html")
        with open(out_path, "w", encoding="utf-8") as f:
            f.write(html)
        print(f"Dashboard written to {out_path}")
        # Open dashboard in default browser on success
        webbrowser.open(Path(out_path).resolve().as_uri())
        print("Dashboard opened in your browser.")
    except Exception as e:
        print(f"Error generating dashboard: {e}")
        import traceback
        traceback.print_exc()


def run_server():
    from flask import Flask

    app = Flask(__name__)
    script_dir = os.path.dirname(os.path.abspath(__file__))

    @app.route("/")
    def index():
        html_path = os.path.join(script_dir, "dashboard.html")
        if os.path.isfile(html_path):
            with open(html_path, "r", encoding="utf-8") as f:
                return f.read()
        data = fetch_dashboard_data()
        return generate_exports_dashboard_html(data)

    port = int(os.getenv("PORT", 5000))
    print(f"Dashboard server: http://localhost:{port}")
    print("Run the script to regenerate dashboard.html for fresh data.")
    app.run(host="0.0.0.0", port=port)


if __name__ == "__main__":
    import sys
    if "--server" in sys.argv or "-s" in sys.argv:
        run_server()
    else:
        main()
