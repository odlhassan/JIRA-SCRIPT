"""
Generate an interactive HTML report from nested view.xlsx.
"""
from __future__ import annotations

import json
import os
import shutil
import sqlite3
from collections import defaultdict, deque
from datetime import date, datetime, timedelta, timezone
from pathlib import Path

from openpyxl import load_workbook
from generate_assignee_hours_report import (
    DEFAULT_LEAVE_REPORT_INPUT_XLSX,
    _load_leave_daily_rows,
    _load_leave_subtask_rows,
)

EXPECTED_HEADERS = [
    "Aspect",
    "Approved Days",
    "Approved Hours",
    "Planned Days",
    "Planned Hours",
    "Actual Hours",
    "Actual Days",
    "Planned Start Date",
    "Planned End Date",
]

DEFAULT_INPUT_XLSX = "nested view.xlsx"
DEFAULT_OUTPUT_HTML = "nested_view_report.html"
DEFAULT_WORK_ITEMS_XLSX = "1_jira_work_items_export.xlsx"
DEFAULT_CAPACITY_DB = "assignee_hours_capacity.db"
DEFAULT_EXPORTS_DB = "jira_exports.db"
REPORT_HTML_DIRNAME = "report_html"


def _resolve_path(value: str, base_dir: Path) -> Path:
    path = Path(value)
    if path.is_absolute():
        return path
    return base_dir / path


def _to_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _to_number_or_blank(value):
    if value in (None, ""):
        return ""
    try:
        return round(float(value), 2)
    except (TypeError, ValueError):
        return ""


def _subtract_numbers_or_blank(left, right):
    if left == "" or right == "":
        return ""
    try:
        return round(float(left) - float(right), 2)
    except (TypeError, ValueError):
        return ""


def _hours_to_days(hours: float | int | str) -> float:
    try:
        return round(float(hours or 0) / 8.0, 2)
    except (TypeError, ValueError):
        return 0.0


def _to_key_text(value) -> str:
    return _to_text(value).lower()


def _calculate_capacity_hours(settings: dict) -> float:
    from_date = _to_text(settings.get("from_date"))
    to_date = _to_text(settings.get("to_date"))
    if not from_date or not to_date:
        return 0.0
    try:
        from_value = date.fromisoformat(from_date)
        to_value = date.fromisoformat(to_date)
    except ValueError:
        return 0.0
    if to_value < from_value:
        return 0.0

    employee_count = int(settings.get("employee_count") or 0)
    standard_hours_per_day = float(settings.get("standard_hours_per_day") or 0)
    ramadan_hours_per_day = float(settings.get("ramadan_hours_per_day") or 0)
    if employee_count <= 0 or standard_hours_per_day <= 0 or ramadan_hours_per_day <= 0:
        return 0.0

    ramadan_start_date = _to_text(settings.get("ramadan_start_date"))
    ramadan_end_date = _to_text(settings.get("ramadan_end_date"))
    ramadan_start = None
    ramadan_end = None
    if ramadan_start_date and ramadan_end_date:
        try:
            ramadan_start = date.fromisoformat(ramadan_start_date)
            ramadan_end = date.fromisoformat(ramadan_end_date)
        except ValueError:
            ramadan_start = None
            ramadan_end = None

    holiday_raw = settings.get("holiday_dates", [])
    holiday_dates: set[date] = set()
    if isinstance(holiday_raw, list):
        for item in holiday_raw:
            text = _to_text(item)
            if not text:
                continue
            try:
                holiday_dates.add(date.fromisoformat(text))
            except ValueError:
                continue

    non_ramadan_weekdays = 0
    ramadan_weekdays = 0
    cursor = from_value
    while cursor <= to_value:
        if cursor.weekday() < 5 and cursor not in holiday_dates:
            in_ramadan = bool(ramadan_start and ramadan_end and ramadan_start <= cursor <= ramadan_end)
            if in_ramadan:
                ramadan_weekdays += 1
            else:
                non_ramadan_weekdays += 1
        cursor += timedelta(days=1)

    available_capacity_hours = employee_count * (
        non_ramadan_weekdays * standard_hours_per_day
        + ramadan_weekdays * ramadan_hours_per_day
    )
    return round(float(available_capacity_hours), 2)


def _load_capacity_profiles(db_path: Path) -> list[dict]:
    if not db_path.exists():
        return []
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    try:
        rows = conn.execute(
            """
            SELECT from_date, to_date, employee_count, standard_hours_per_day,
                   ramadan_start_date, ramadan_end_date, ramadan_hours_per_day,
                   holiday_dates_json, updated_at_utc
            FROM assignee_capacity_settings
            ORDER BY updated_at_utc DESC, from_date DESC, to_date DESC
            """
        ).fetchall()
    except sqlite3.Error:
        conn.close()
        return []
    finally:
        try:
            conn.close()
        except Exception:
            pass

    profiles: list[dict] = []
    for row in rows:
        holiday_dates: list[str] = []
        raw_holidays = _to_text(row["holiday_dates_json"])
        if raw_holidays:
            try:
                decoded = json.loads(raw_holidays)
                if isinstance(decoded, list):
                    holiday_dates = [_to_text(item) for item in decoded if _to_text(item)]
            except json.JSONDecodeError:
                holiday_dates = []
        profile = {
            "from_date": _to_text(row["from_date"]),
            "to_date": _to_text(row["to_date"]),
            "employee_count": int(row["employee_count"] or 0),
            "standard_hours_per_day": float(row["standard_hours_per_day"] or 0),
            "ramadan_start_date": _to_text(row["ramadan_start_date"]),
            "ramadan_end_date": _to_text(row["ramadan_end_date"]),
            "ramadan_hours_per_day": float(row["ramadan_hours_per_day"] or 0),
            "holiday_dates": holiday_dates,
            "updated_at_utc": _to_text(row["updated_at_utc"]),
        }
        profile["available_capacity_hours"] = _calculate_capacity_hours(profile)
        profiles.append(profile)
    return profiles


def _project_key_from_aspect(aspect: str) -> str:
    text = _to_text(aspect)
    if " - " in text:
        return text.split(" - ", 1)[0].strip().upper()
    return text.strip().upper()


def _project_key_and_name_from_aspect(aspect: str) -> tuple[str, str]:
    text = _to_text(aspect)
    if " - " in text:
        key, name = text.split(" - ", 1)
        return key.strip().upper(), name.strip()
    return text.strip().upper(), ""


def _jira_base_url() -> str:
    site = os.getenv("JIRA_SITE", "octopusdtlsupport").strip()
    return f"https://{site}.atlassian.net"


def _work_item_kind(jira_issue_type: str) -> str:
    value = _to_text(jira_issue_type).lower()
    if "epic" in value:
        return "epic"
    if "sub-task" in value or "subtask" in value or ("bug" in value and "sub" in value):
        return "subtask"
    if "story" in value or "task" in value:
        return "story"
    return "other"


def _load_work_items_link_index(work_items_path: Path) -> dict[str, object]:
    empty = {
        "epic_by_project_summary": defaultdict(deque),
        "story_by_project_parent_summary": defaultdict(deque),
        "story_by_project_summary": defaultdict(deque),
        "subtask_by_project_parent_summary": defaultdict(deque),
        "subtask_by_project_summary": defaultdict(deque),
    }
    if not work_items_path.exists():
        return empty

    wb = load_workbook(work_items_path, read_only=True, data_only=True)
    ws = wb.active
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header:
        wb.close()
        return empty

    headers = [_to_text(h) for h in header]
    idx = {name: i for i, name in enumerate(headers)}
    required = ["project_key", "issue_key", "jira_issue_type", "summary", "parent_issue_key", "jira_url"]
    if any(name not in idx for name in required):
        wb.close()
        return empty

    base_url = _jira_base_url()

    for row in ws.iter_rows(min_row=2, values_only=True):
        project_key = _to_text(row[idx["project_key"]]).upper()
        issue_key = _to_text(row[idx["issue_key"]]).upper()
        parent_key = _to_text(row[idx["parent_issue_key"]]).upper()
        summary_key = _to_key_text(row[idx["summary"]])
        issue_type = _to_text(row[idx["jira_issue_type"]])
        jira_url = _to_text(row[idx["jira_url"]]) or (f"{base_url}/browse/{issue_key}" if issue_key else "")
        if not project_key or not issue_key or not summary_key:
            continue

        payload = {"key": issue_key, "url": jira_url}
        kind = _work_item_kind(issue_type)
        if kind == "epic":
            empty["epic_by_project_summary"][(project_key, summary_key)].append(payload)
        elif kind == "story":
            empty["story_by_project_parent_summary"][(project_key, parent_key, summary_key)].append(payload)
            empty["story_by_project_summary"][(project_key, summary_key)].append(payload)
        elif kind == "subtask":
            empty["subtask_by_project_parent_summary"][(project_key, parent_key, summary_key)].append(payload)
            empty["subtask_by_project_summary"][(project_key, summary_key)].append(payload)

    wb.close()
    return empty


def _take_indexed_item(mapping, key):
    queue = mapping.get(key)
    if not queue:
        return None
    return queue.popleft()


def _attach_jira_links(rows: list[dict], work_items_path: Path) -> None:
    index = _load_work_items_link_index(work_items_path)

    current_project = ""
    current_epic_key = ""
    current_story_key = ""

    for row in rows:
        row["jira_key"] = ""
        row["jira_url"] = ""
        row_type = row.get("row_type")
        level = int(row.get("level", 0) or 0)

        if level <= 1:
            current_epic_key = ""
            current_story_key = ""
        elif level <= 3:
            current_story_key = ""

        if row_type == "project":
            current_project = _project_key_from_aspect(str(row.get("aspect", "")))
            continue
        if row_type == "product":
            continue

        summary_key = _to_key_text(row.get("aspect", ""))
        if not current_project or not summary_key:
            continue

        if row_type == "rmi":
            hit = _take_indexed_item(index["epic_by_project_summary"], (current_project, summary_key))
            if hit:
                row["jira_key"] = hit["key"]
                row["jira_url"] = hit["url"]
                current_epic_key = hit["key"]
            else:
                current_epic_key = ""
        elif row_type == "story":
            hit = _take_indexed_item(
                index["story_by_project_parent_summary"],
                (current_project, current_epic_key, summary_key),
            )
            if hit is None:
                hit = _take_indexed_item(index["story_by_project_summary"], (current_project, summary_key))
            if hit:
                row["jira_key"] = hit["key"]
                row["jira_url"] = hit["url"]
                current_story_key = hit["key"]
            else:
                current_story_key = ""
        elif row_type == "subtask":
            hit = _take_indexed_item(
                index["subtask_by_project_parent_summary"],
                (current_project, current_story_key, summary_key),
            )
            if hit is None:
                hit = _take_indexed_item(index["subtask_by_project_summary"], (current_project, summary_key))
            if hit:
                row["jira_key"] = hit["key"]
                row["jira_url"] = hit["url"]


def _assign_row_metadata(rows: list[dict], *, attach_links_from_workbook: bool, work_items_path: Path | None = None) -> list[dict]:
    stack: dict[int, int] = {}
    row_by_id: dict[int, dict] = {}
    next_id = 1
    current_project_key = ""
    current_project_name = ""

    for row_data in rows:
        level = int(row_data.get("level", 0) or 0)
        for key in list(stack):
            if key >= level:
                del stack[key]

        parent_id = stack.get(level - 1)
        row_id = next_id
        next_id += 1
        stack[level] = row_id

        row_type = _row_type_from_level(level)
        row_data["id"] = row_id
        row_data["parent_id"] = parent_id
        row_data["row_type"] = row_type
        row_data["type_label"] = ""
        row_data["is_missing_parent"] = False
        row_data["missing_parent_reason"] = ""
        row_data["has_defined_product_category"] = False
        row_data.setdefault("jira_key", "")
        row_data.setdefault("jira_url", "")

        if row_type == "project":
            current_project_key, current_project_name = _project_key_and_name_from_aspect(row_data["aspect"])
        row_data["project_key"] = current_project_key
        row_data["project_name"] = current_project_name
        approved_hours = row_data.get("approved_hours", row_data.get("man_hours", ""))
        approved_days = row_data.get("approved_days", row_data.get("man_days", ""))
        planned_hours = row_data.get("planned_hours", approved_hours)
        planned_days = row_data.get("planned_days", approved_days)
        row_data["approved_hours"] = approved_hours
        row_data["approved_days"] = approved_days
        row_data["planned_hours"] = planned_hours
        row_data["planned_days"] = planned_days
        # Keep legacy aliases so existing scorecards and formulas continue to work.
        row_data["man_hours"] = approved_hours
        row_data["man_days"] = approved_days
        row_data["delta_hours"] = _subtract_numbers_or_blank(approved_hours, row_data["actual_hours"])
        row_data["delta_days"] = _subtract_numbers_or_blank(approved_days, row_data["actual_days"])

        if row_data["row_type"] == "story":
            parent = row_by_id.get(parent_id) if parent_id else None
            parent_is_no_rmi = bool(parent and parent.get("row_type") == "rmi" and _to_text(parent.get("aspect")) == "No RMI")
            if parent_is_no_rmi:
                row_data["is_missing_parent"] = True
                row_data["missing_parent_reason"] = "missing_rmi_parent"
        elif row_data["row_type"] == "subtask":
            parent = row_by_id.get(parent_id) if parent_id else None
            parent_is_story = bool(parent and parent.get("row_type") == "story")
            if not parent_is_story:
                row_data["is_missing_parent"] = True
                row_data["missing_parent_reason"] = "missing_story_parent"
        elif row_data["row_type"] == "assignee":
            parent = row_by_id.get(parent_id) if parent_id else None
            if parent and parent.get("row_type") == "subtask" and parent.get("is_missing_parent"):
                row_data["is_missing_parent"] = True
                row_data["missing_parent_reason"] = parent.get("missing_parent_reason", "")

        if row_data["row_type"] == "product":
            row_data["has_defined_product_category"] = _is_defined_product_category(row_data["aspect"])
        else:
            product_parent = row_by_id.get(stack.get(2))
            if product_parent and product_parent.get("row_type") == "product":
                row_data["has_defined_product_category"] = bool(
                    product_parent.get("has_defined_product_category")
                )

        if row_data["row_type"] == "project":
            row_data["type_label"] = "Project"
        elif row_data["row_type"] == "product":
            row_data["type_label"] = "Category"
        elif row_data["row_type"] == "rmi":
            row_data["type_label"] = "Epic"
        elif row_data["row_type"] == "story":
            row_data["type_label"] = "Bug" if _detect_bug_label("story", row_data["aspect"]) else "Story"
        elif row_data["row_type"] == "subtask":
            row_data["type_label"] = "Bug" if _detect_bug_label("subtask", row_data["aspect"]) else "Subtask"
        else:
            row_data["type_label"] = row_data["row_type"].capitalize()

        row_by_id[row_id] = row_data

    if attach_links_from_workbook and work_items_path is not None:
        _attach_jira_links(rows, work_items_path)
    return rows


def _row_type_from_level(level: int) -> str:
    mapping = {
        1: "project",
        2: "product",
        3: "rmi",
        4: "story",
        5: "subtask",
        6: "assignee",
    }
    return mapping.get(level, "unknown")


def _detect_bug_label(row_type: str, aspect: str) -> bool:
    if row_type not in ("story", "subtask"):
        return False
    text = _to_text(aspect).lower()
    return "bug" in text


def _is_defined_product_category(value: str) -> bool:
    text = _to_text(value).strip().lower()
    if not text:
        return False
    return text not in {"uncategorized", "no product", "n/a", "na", "none"}


def _canonical_last_success_run_id(db_path: Path) -> str:
    if not db_path.exists():
        return ""
    conn = sqlite3.connect(db_path)
    try:
        row = conn.execute(
            "SELECT last_success_run_id FROM canonical_refresh_state WHERE id = 1"
        ).fetchone()
    except sqlite3.Error:
        return ""
    finally:
        conn.close()
    return _to_text(row[0] if row else "")


def _load_project_display_names(db_path: Path) -> dict[str, str]:
    if not db_path.exists():
        return {}
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    try:
        rows = conn.execute(
            """
            SELECT project_key, display_name, project_name
            FROM managed_projects
            WHERE is_active = 1
            """
        ).fetchall()
    except sqlite3.Error:
        return {}
    finally:
        conn.close()
    result: dict[str, str] = {}
    for row in rows:
        key = _to_text(row["project_key"]).upper()
        if not key:
            continue
        result[key] = _to_text(row["display_name"]) or _to_text(row["project_name"]) or key
    return result


def _load_epic_categories(db_path: Path) -> tuple[dict[str, str], dict[str, str]]:
    if not db_path.exists():
        return {}, {}
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    try:
        rows = conn.execute(
            """
            SELECT epic_key, product_category, project_name
            FROM epics_management
            """
        ).fetchall()
    except sqlite3.Error:
        return {}, {}
    finally:
        conn.close()
    categories: dict[str, str] = {}
    project_names: dict[str, str] = {}
    for row in rows:
        epic_key = _to_text(row["epic_key"]).upper()
        if epic_key:
            categories[epic_key] = _to_text(row["product_category"])
        project_name = _to_text(row["project_name"])
        project_key = epic_key.split("-", 1)[0] if "-" in epic_key else ""
        if project_key and project_name and project_key not in project_names:
            project_names[project_key] = project_name
    return categories, project_names


def _load_nested_node_categories(exports_db_path: Path) -> dict[str, str]:
    if not exports_db_path.exists() or not exports_db_path.is_file():
        return {}
    try:
        conn = sqlite3.connect(exports_db_path)
        rows = conn.execute(
            "SELECT key, product_category FROM nested_view_nodes"
        ).fetchall()
    except sqlite3.Error:
        return {}
    finally:
        try:
            conn.close()
        except Exception:
            pass
    result: dict[str, str] = {}
    for key, product_category in rows:
        normalized_key = _to_text(key).upper()
        if normalized_key and normalized_key not in result:
            result[normalized_key] = _to_text(product_category)
    return result


def _canonical_kind(issue_type: str) -> str:
    value = _to_text(issue_type).lower()
    if "epic" in value:
        return "epic"
    if "sub-task" in value or "subtask" in value or ("bug" in value and "sub" in value):
        return "subtask"
    if "story" in value or "task" in value:
        return "story"
    return "other"


def _load_nested_rows_from_canonical_db(db_path: Path, exports_db_path: Path | None = None, run_id: str = "") -> list[dict]:
    effective_run_id = _to_text(run_id) or _canonical_last_success_run_id(db_path)
    if not effective_run_id:
        raise ValueError("No successful canonical refresh found.")

    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    try:
        issue_rows = conn.execute(
            """
            SELECT issue_key, project_key, issue_type, summary, assignee,
                   start_date, due_date, original_estimate_hours, total_hours_logged,
                   parent_issue_key, story_key, epic_key
            FROM canonical_issues
            WHERE run_id = ?
            ORDER BY project_key ASC, issue_key ASC
            """,
            (effective_run_id,),
        ).fetchall()
        actual_rows = conn.execute(
            """
            SELECT issue_key, total_worklog_hours
            FROM canonical_issue_actuals
            WHERE run_id = ?
            """,
            (effective_run_id,),
        ).fetchall()
    finally:
        conn.close()

    if not issue_rows:
        raise ValueError(f"No canonical issues found for run_id={effective_run_id}.")

    project_display_names = _load_project_display_names(db_path)
    epic_categories, epic_project_names = _load_epic_categories(db_path)
    nested_node_categories = _load_nested_node_categories(exports_db_path or Path(""))

    actual_hours_by_key = {
        _to_text(row["issue_key"]).upper(): round(float(row["total_worklog_hours"] or 0), 2)
        for row in actual_rows
        if _to_text(row["issue_key"])
    }

    epics: dict[str, dict] = {}
    stories: dict[str, dict] = {}
    subtasks: dict[str, dict] = {}

    for row in issue_rows:
        issue_key = _to_text(row["issue_key"]).upper()
        if not issue_key:
            continue
        kind = _canonical_kind(row["issue_type"])
        if kind == "other":
            continue
        project_key = _to_text(row["project_key"]).upper() or "UNKNOWN"
        epic_key = _to_text(row["epic_key"]).upper()
        story_key = _to_text(row["story_key"]).upper()
        parent_key = _to_text(row["parent_issue_key"]).upper()
        issue_payload = {
            "key": issue_key,
            "project_key": project_key,
            "summary": _to_text(row["summary"]) or issue_key,
            "assignee": _to_text(row["assignee"]) or "Unassigned",
            "planned_start": _to_text(row["start_date"]),
            "planned_end": _to_text(row["due_date"]),
            "approved_hours": round(float(row["original_estimate_hours"] or 0), 2),
            "approved_days": _hours_to_days(row["original_estimate_hours"] or 0),
            "planned_hours": round(float(row["original_estimate_hours"] or 0), 2) if kind == "subtask" else 0.0,
            "planned_days": _hours_to_days(row["original_estimate_hours"] or 0) if kind == "subtask" else 0.0,
            "man_hours": round(float(row["original_estimate_hours"] or 0), 2),
            "man_days": _hours_to_days(row["original_estimate_hours"] or 0),
            "actual_hours": round(float(row["total_hours_logged"] or 0), 2),
            "actual_days": _hours_to_days(row["total_hours_logged"] or 0),
            "parent_key": parent_key,
            "epic_key": epic_key,
            "story_key": story_key,
        }
        if kind == "epic":
            issue_payload["product_category"] = (
                _to_text(epic_categories.get(issue_key))
                or _to_text(nested_node_categories.get(issue_key))
                or "Uncategorized"
            )
            epics[issue_key] = issue_payload
        elif kind == "story":
            issue_payload["epic_key"] = epic_key or parent_key
            stories[issue_key] = issue_payload
        elif kind == "subtask":
            issue_payload["story_key"] = story_key or parent_key
            issue_payload["actual_hours"] = actual_hours_by_key.get(issue_key, issue_payload["actual_hours"])
            issue_payload["actual_days"] = _hours_to_days(issue_payload["actual_hours"])
            subtasks[issue_key] = issue_payload

    for story in stories.values():
        epic = epics.get(_to_text(story.get("epic_key")).upper())
        story["product_category"] = (
            _to_text(nested_node_categories.get(_to_text(story.get("key")).upper()))
            or _to_text(epic_categories.get(_to_text(story.get("epic_key")).upper()))
            or _to_text(epic.get("product_category") if epic else "")
            or "Uncategorized"
        )
    for subtask in subtasks.values():
        story = stories.get(_to_text(subtask.get("story_key")).upper())
        epic = epics.get(_to_text(story.get("epic_key")).upper()) if story else None
        subtask["product_category"] = (
            _to_text(nested_node_categories.get(_to_text(subtask.get("key")).upper()))
            or _to_text(story.get("product_category") if story else "")
            or _to_text(epic.get("product_category") if epic else "")
            or "Uncategorized"
        )
        if subtask.get("project_key") == "UNKNOWN" and story:
            subtask["project_key"] = story.get("project_key", "UNKNOWN")

    stories_by_epic: dict[str, list[dict]] = defaultdict(list)
    for story in stories.values():
        grouped_epic_key = _to_text(story.get("epic_key")).upper() or f"__NO_RMI__::{story['project_key']}::{story['product_category']}"
        stories_by_epic[grouped_epic_key].append(story)

    subtasks_by_story: dict[str, list[dict]] = defaultdict(list)
    for subtask in subtasks.values():
        subtasks_by_story[_to_text(subtask.get("story_key")).upper()].append(subtask)

    projects_found: set[str] = set()
    products_by_project: dict[str, set[str]] = defaultdict(set)
    for collection in (epics.values(), stories.values(), subtasks.values()):
        for item in collection:
            project_key = _to_text(item.get("project_key")).upper()
            if not project_key:
                continue
            projects_found.add(project_key)
            products_by_project[project_key].add(_to_text(item.get("product_category")) or "Uncategorized")

    def _merge_date_bounds(values: list[str]) -> tuple[str, str]:
        clean = [value for value in (_to_text(item) for item in values) if value]
        if not clean:
            return "", ""
        return min(clean), max(clean)

    story_metrics: dict[str, dict[str, object]] = {}
    for story in stories.values():
        related_subtasks = subtasks_by_story.get(story["key"], [])
        fallback_start, _ = _merge_date_bounds([_to_text(subtask.get("planned_start")) for subtask in related_subtasks])
        _, fallback_end = _merge_date_bounds([_to_text(subtask.get("planned_end")) for subtask in related_subtasks])
        planned_start = _to_text(story.get("planned_start")) or fallback_start
        planned_end = _to_text(story.get("planned_end")) or fallback_end
        story_metrics[story["key"]] = {
            "approved_days": round(float(story.get("approved_days") or 0), 2),
            "approved_hours": round(float(story.get("approved_hours") or 0), 2),
            "planned_days": round(sum(float(item.get("planned_days") or 0) for item in related_subtasks), 2),
            "planned_hours": round(sum(float(item.get("planned_hours") or 0) for item in related_subtasks), 2),
            "actual_hours": round(float(story.get("actual_hours") or 0), 2),
            "planned_start": planned_start,
            "planned_end": planned_end,
        }

    def _aggregate_metrics(items: list[dict[str, object]]) -> dict[str, object]:
        approved_days = round(sum(float(item.get("approved_days", 0) or 0) for item in items), 2)
        approved_hours = round(sum(float(item.get("approved_hours", 0) or 0) for item in items), 2)
        planned_days = round(sum(float(item.get("planned_days", 0) or 0) for item in items), 2)
        planned_hours = round(sum(float(item.get("planned_hours", 0) or 0) for item in items), 2)
        actual_hours = round(sum(float(item.get("actual_hours", 0) or 0) for item in items), 2)
        planned_start, _ = _merge_date_bounds([_to_text(item.get("planned_start")) for item in items])
        _, planned_end = _merge_date_bounds([_to_text(item.get("planned_end")) for item in items])
        return {
            "approved_days": approved_days,
            "approved_hours": approved_hours,
            "planned_days": planned_days,
            "planned_hours": planned_hours,
            "actual_hours": actual_hours,
            "planned_start": planned_start,
            "planned_end": planned_end,
        }

    flat_rows: list[dict] = []
    for project_key in sorted(projects_found):
        product_values = sorted(products_by_project.get(project_key, {"Uncategorized"}))
        product_blocks: list[dict[str, object]] = []
        for product_category in product_values:
            product_label = _to_text(product_category) or "Uncategorized"
            epics_in_group = [
                epic for epic in epics.values()
                if epic["project_key"] == project_key and _to_text(epic.get("product_category")) == product_label
            ]
            orphan_epic_placeholder_key = f"__NO_RMI__::{project_key}::{product_label}"
            orphan_stories = stories_by_epic.get(orphan_epic_placeholder_key, [])
            if orphan_stories:
                epics_in_group.append(
                    {
                        "key": "",
                        "summary": "No RMI",
                        "project_key": project_key,
                        "product_category": product_label,
                        "approved_hours": 0.0,
                        "approved_days": 0.0,
                        "planned_hours": 0.0,
                        "planned_days": 0.0,
                        "man_hours": 0.0,
                        "man_days": 0.0,
                        "actual_hours": 0.0,
                        "planned_start": "",
                        "planned_end": "",
                    }
                )
            epics_in_group = sorted(epics_in_group, key=lambda item: (_to_text(item.get("summary")).lower(), _to_text(item.get("key"))))
            epic_metrics: dict[str, dict[str, object]] = {}
            for epic in epics_in_group:
                epic_story_key = _to_text(epic.get("key")) or orphan_epic_placeholder_key
                related_stories = stories_by_epic.get(epic_story_key, [])
                related_story_metrics = [story_metrics[story["key"]] for story in related_stories if story["key"] in story_metrics]
                fallback_start, _ = _merge_date_bounds([_to_text(metric.get("planned_start")) for metric in related_story_metrics])
                _, fallback_end = _merge_date_bounds([_to_text(metric.get("planned_end")) for metric in related_story_metrics])
                epic_metrics[_to_text(epic.get("key"))] = {
                    "approved_days": round(float(epic.get("approved_days") or 0), 2),
                    "approved_hours": round(float(epic.get("approved_hours") or 0), 2),
                    "planned_days": round(sum(float(metric.get("planned_days") or 0) for metric in related_story_metrics), 2),
                    "planned_hours": round(sum(float(metric.get("planned_hours") or 0) for metric in related_story_metrics), 2),
                    "actual_hours": round(float(epic.get("actual_hours") or 0), 2),
                    "planned_start": _to_text(epic.get("planned_start")) or fallback_start,
                    "planned_end": _to_text(epic.get("planned_end")) or fallback_end,
                }
            product_blocks.append(
                {
                    "label": product_label,
                    "epics": epics_in_group,
                    "epic_metrics": epic_metrics,
                    "metric": _aggregate_metrics(list(epic_metrics.values())),
                }
            )

        project_metric = _aggregate_metrics([block["metric"] for block in product_blocks])
        project_name = (
            project_display_names.get(project_key)
            or epic_project_names.get(project_key)
            or project_key
        )
        flat_rows.append(
            {
                "level": 1,
                "aspect": f"{project_key} - {project_name}",
                "approved_days": project_metric["approved_days"],
                "approved_hours": project_metric["approved_hours"],
                "planned_days": project_metric["planned_days"],
                "planned_hours": project_metric["planned_hours"],
                "man_days": project_metric["approved_days"],
                "man_hours": project_metric["approved_hours"],
                "actual_hours": project_metric["actual_hours"],
                "actual_days": _hours_to_days(project_metric["actual_hours"]),
                "planned_start": _to_text(project_metric["planned_start"]),
                "planned_end": _to_text(project_metric["planned_end"]),
            }
        )
        for block in product_blocks:
            flat_rows.append(
                {
                    "level": 2,
                    "aspect": _to_text(block["label"]),
                    "approved_days": block["metric"]["approved_days"],
                    "approved_hours": block["metric"]["approved_hours"],
                    "planned_days": block["metric"]["planned_days"],
                    "planned_hours": block["metric"]["planned_hours"],
                    "man_days": block["metric"]["approved_days"],
                    "man_hours": block["metric"]["approved_hours"],
                    "actual_hours": block["metric"]["actual_hours"],
                    "actual_days": _hours_to_days(block["metric"]["actual_hours"]),
                    "planned_start": _to_text(block["metric"]["planned_start"]),
                    "planned_end": _to_text(block["metric"]["planned_end"]),
                }
            )
            for epic in block["epics"]:
                epic_key = _to_text(epic.get("key")).upper()
                metrics = block["epic_metrics"].get(epic_key if epic_key else "", {})
                flat_rows.append(
                    {
                        "level": 3,
                        "aspect": _to_text(epic.get("summary")),
                        "approved_days": metrics.get("approved_days", 0.0),
                        "approved_hours": metrics.get("approved_hours", 0.0),
                        "planned_days": metrics.get("planned_days", 0.0),
                        "planned_hours": metrics.get("planned_hours", 0.0),
                        "man_days": metrics.get("approved_days", 0.0),
                        "man_hours": metrics.get("approved_hours", 0.0),
                        "actual_hours": metrics.get("actual_hours", 0.0),
                        "actual_days": _hours_to_days(metrics.get("actual_hours", 0.0)),
                        "planned_start": _to_text(metrics.get("planned_start")),
                        "planned_end": _to_text(metrics.get("planned_end")),
                        "jira_key": epic_key,
                        "jira_url": f"{_jira_base_url()}/browse/{epic_key}" if epic_key else "",
                    }
                )
                orphan_epic_placeholder_key = f"__NO_RMI__::{project_key}::{block['label']}"
                epic_story_key = epic_key or orphan_epic_placeholder_key
                related_stories = sorted(
                    stories_by_epic.get(epic_story_key, []),
                    key=lambda item: (_to_text(item.get("summary")).lower(), _to_text(item.get("key"))),
                )
                for story in related_stories:
                    story_key = _to_text(story.get("key")).upper()
                    story_metric = story_metrics.get(story_key, {})
                    flat_rows.append(
                        {
                            "level": 4,
                            "aspect": _to_text(story.get("summary")),
                            "approved_days": story_metric.get("approved_days", round(float(story.get("approved_days") or 0), 2)),
                            "approved_hours": story_metric.get("approved_hours", round(float(story.get("approved_hours") or 0), 2)),
                            "planned_days": story_metric.get("planned_days", 0.0),
                            "planned_hours": story_metric.get("planned_hours", 0.0),
                            "man_days": story_metric.get("approved_days", round(float(story.get("approved_days") or 0), 2)),
                            "man_hours": story_metric.get("approved_hours", round(float(story.get("approved_hours") or 0), 2)),
                            "actual_hours": story_metric.get("actual_hours", round(float(story.get("actual_hours") or 0), 2)),
                            "actual_days": _hours_to_days(story_metric.get("actual_hours", story.get("actual_hours", 0))),
                            "planned_start": _to_text(story_metric.get("planned_start")) or _to_text(story.get("planned_start")),
                            "planned_end": _to_text(story_metric.get("planned_end")) or _to_text(story.get("planned_end")),
                            "jira_key": story_key,
                            "jira_url": f"{_jira_base_url()}/browse/{story_key}" if story_key else "",
                        }
                    )
                    related_subtasks = sorted(
                        subtasks_by_story.get(story_key, []),
                        key=lambda item: (_to_text(item.get("summary")).lower(), _to_text(item.get("key"))),
                    )
                    for subtask in related_subtasks:
                        subtask_key = _to_text(subtask.get("key")).upper()
                        assignee_text = _to_text(subtask.get("assignee")) or "Unassigned"
                        flat_rows.append(
                            {
                                "level": 5,
                                "aspect": _to_text(subtask.get("summary")),
                                "approved_days": round(float(subtask.get("approved_days") or 0), 2),
                                "approved_hours": round(float(subtask.get("approved_hours") or 0), 2),
                                "planned_days": round(float(subtask.get("planned_days") or 0), 2),
                                "planned_hours": round(float(subtask.get("planned_hours") or 0), 2),
                                "man_days": round(float(subtask.get("approved_days") or 0), 2),
                                "man_hours": round(float(subtask.get("approved_hours") or 0), 2),
                                "actual_hours": round(float(subtask.get("actual_hours") or 0), 2),
                                "actual_days": _hours_to_days(subtask.get("actual_hours", 0)),
                                "planned_start": _to_text(subtask.get("planned_start")),
                                "planned_end": _to_text(subtask.get("planned_end")),
                                "jira_key": subtask_key,
                                "jira_url": f"{_jira_base_url()}/browse/{subtask_key}" if subtask_key else "",
                                "assignee": assignee_text,
                            }
                        )
                        flat_rows.append(
                            {
                                "level": 6,
                                "aspect": assignee_text,
                                "approved_days": round(float(subtask.get("approved_days") or 0), 2),
                                "approved_hours": round(float(subtask.get("approved_hours") or 0), 2),
                                "planned_days": round(float(subtask.get("planned_days") or 0), 2),
                                "planned_hours": round(float(subtask.get("planned_hours") or 0), 2),
                                "man_days": round(float(subtask.get("approved_days") or 0), 2),
                                "man_hours": round(float(subtask.get("approved_hours") or 0), 2),
                                "actual_hours": round(float(subtask.get("actual_hours") or 0), 2),
                                "actual_days": _hours_to_days(subtask.get("actual_hours", 0)),
                                "planned_start": _to_text(subtask.get("planned_start")),
                                "planned_end": _to_text(subtask.get("planned_end")),
                                "assignee": assignee_text,
                            }
                        )

    return _assign_row_metadata(flat_rows, attach_links_from_workbook=False)


def load_nested_view_tree_for_api(
    db_path: Path,
    exports_db_path: Path | None = None,
    run_id: str = "",
) -> list[dict]:
    """
    Load the full nested view tree from the canonical DB for API consumption.
    Returns the same row structure as used in the HTML report (with id, parent_id,
    row_type, type_label, etc.) so the table reflects database data.
    """
    return _load_nested_rows_from_canonical_db(
        db_path,
        exports_db_path=exports_db_path or db_path.parent / "jira_exports.db",
        run_id=run_id,
    )


def _load_nested_rows(input_path: Path) -> list[dict]:
    if not input_path.exists():
        raise FileNotFoundError(f"Nested view workbook not found: {input_path}")

    wb = load_workbook(input_path, read_only=False, data_only=True)
    ws = wb["NestedView"] if "NestedView" in wb.sheetnames else wb.active

    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header:
        wb.close()
        raise ValueError("Nested view workbook has no header row.")

    found_headers = [_to_text(cell) for cell in header]
    normalized_headers = [h.lower() for h in found_headers]
    has_new_headers = normalized_headers[: len(EXPECTED_HEADERS)] == [h.lower() for h in EXPECTED_HEADERS]
    old_headers = [
        "aspect",
        "man-days",
        "man-hours",
        "actual hours",
        "actual days",
        "planned start date",
        "planned end date",
    ]
    has_old_headers = normalized_headers[: len(old_headers)] == old_headers
    if not has_new_headers and not has_old_headers:
        wb.close()
        raise ValueError(
            "Nested view workbook headers do not match expected layout. "
            f"Expected first columns: {EXPECTED_HEADERS}, got: {found_headers[:len(EXPECTED_HEADERS)]}"
        )

    rows: list[dict] = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        cell_value = row[0] if len(row) > 0 else ""
        raw_aspect = "" if cell_value is None else str(cell_value)
        level = int(getattr(ws.row_dimensions[row_idx], "outlineLevel", 0) or 0)
        if level <= 0:
            leading_spaces = len(raw_aspect) - len(raw_aspect.lstrip(" "))
            # Some generated workbooks flatten Excel outline metadata and keep only
            # two-space indentation in the Aspect text. Recover the hierarchy from it.
            level = max(1, (leading_spaces // 2) + 1)
        aspect = _to_text(raw_aspect.lstrip(" "))

        if has_new_headers:
            approved_days = _to_number_or_blank(row[1] if len(row) > 1 else "")
            approved_hours = _to_number_or_blank(row[2] if len(row) > 2 else "")
            planned_days = _to_number_or_blank(row[3] if len(row) > 3 else "")
            planned_hours = _to_number_or_blank(row[4] if len(row) > 4 else "")
            actual_hours = _to_number_or_blank(row[5] if len(row) > 5 else "")
            actual_days = _to_number_or_blank(row[6] if len(row) > 6 else "")
            planned_start = _to_text(row[7] if len(row) > 7 else "")
            planned_end = _to_text(row[8] if len(row) > 8 else "")
        else:
            approved_days = _to_number_or_blank(row[1] if len(row) > 1 else "")
            approved_hours = _to_number_or_blank(row[2] if len(row) > 2 else "")
            planned_days = approved_days
            planned_hours = approved_hours
            actual_hours = _to_number_or_blank(row[3] if len(row) > 3 else "")
            actual_days = _to_number_or_blank(row[4] if len(row) > 4 else "")
            planned_start = _to_text(row[5] if len(row) > 5 else "")
            planned_end = _to_text(row[6] if len(row) > 6 else "")

        row_data = {
            "level": level,
            "aspect": aspect,
            "approved_days": approved_days,
            "approved_hours": approved_hours,
            "planned_days": planned_days,
            "planned_hours": planned_hours,
            "man_days": approved_days,
            "man_hours": approved_hours,
            "actual_hours": actual_hours,
            "actual_days": actual_days,
            "planned_start": planned_start,
            "planned_end": planned_end,
        }
        rows.append(row_data)

    wb.close()
    work_items_path = _resolve_path(os.getenv("JIRA_WORK_ITEMS_XLSX_PATH", DEFAULT_WORK_ITEMS_XLSX), input_path.parent)
    return _assign_row_metadata(rows, attach_links_from_workbook=True, work_items_path=work_items_path)


def _build_html(data: dict) -> str:
    payload = json.dumps(data, ensure_ascii=True)
    return f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Nested View Report</title>
  <link href="https://fonts.googleapis.com/icon?family=Material+Icons+Outlined" rel="stylesheet">
  <script>
    (function () {{
      const themeStorageKey = "rmi-nested-report-theme";
      const themeStored = localStorage.getItem(themeStorageKey);
      const theme = (themeStored === "dark" || themeStored === "light") ? themeStored : "light";
      const densityStorageKey = "rmi-nested-report-density";
      const densityStored = localStorage.getItem(densityStorageKey);
      const density = (densityStored === "compact" || densityStored === "cozy") ? densityStored : "cozy";
      document.documentElement.setAttribute("data-theme", theme);
      document.documentElement.setAttribute("data-density", density);
    }})();
  </script>
  <style>
    :root {{
      --bg: #f3f6f9;
      --panel: #ffffff;
      --text: #1f2937;
      --muted: #6b7280;
      --line: #dbe3ea;
      --head: #0f4c5c;
      --head-text: #ffffff;
      --accent: #ffb703;
      --kpi-capacity-bg: #fee2e2;
      --kpi-capacity-line: #fca5a5;
      --kpi-capacity-ink: #7f1d1d;
      --kpi-planned-bg: #dbeafe;
      --kpi-planned-line: #93c5fd;
      --kpi-planned-ink: #1e3a8a;
      --kpi-actual-bg: #f3e8ff;
      --kpi-actual-line: #d8b4fe;
      --kpi-actual-ink: #a855f7;
      --kpi-gap-bg: #e0f2fe;
      --kpi-gap-line: #7dd3fc;
      --kpi-gap-ink: #2563eb;
      --kpi-leaves-bg: #fecaca;
      --kpi-leaves-line: #f87171;
      --kpi-leaves-ink: #b91c1c;
      --kpi-capacity-gap-bg: #fef2f2;
      --kpi-capacity-gap-line: #fecaca;
      --kpi-capacity-gap-ink: #ef4444;
    }}
    :root {{ color-scheme: light; }}
    html[data-theme="dark"] {{ color-scheme: dark; }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      font-family: "Segoe UI", Tahoma, Verdana, sans-serif;
      color: var(--text);
      background:
        radial-gradient(1000px 280px at 10% -5%, #d7eef6 0%, transparent 60%),
        linear-gradient(180deg, #eef4f7, var(--bg));
    }}
    .top-date-range-wrap {{
      position: sticky;
      top: 0;
      z-index: 20;
      display: flex;
      justify-content: center;
      padding: 8px 12px 0;
      pointer-events: none;
    }}
    .top-date-range-chip {{
      pointer-events: auto;
      display: inline-flex;
      align-items: center;
      gap: 8px;
      flex-wrap: wrap;
      padding: 6px 12px;
      border-radius: 999px;
      border: 1px solid #7a4b24;
      background: #8b5e34;
      box-shadow: 0 6px 16px rgba(17, 24, 39, 0.14);
    }}
    .date-chip-segment {{
      display: inline-flex;
      align-items: center;
      color: #fff7ed;
      font-size: 0.72rem;
      font-weight: 700;
      line-height: 1;
      text-transform: uppercase;
      letter-spacing: 0.02em;
    }}
    .date-chip-input {{
      border: 1px solid #f3e8d8;
      border-radius: 999px;
      padding: 3px 8px;
      background: #fff7ed;
      color: #7c2d12;
      font-size: 0.74rem;
      font-weight: 700;
      min-height: 30px;
    }}
    .date-chip-input:focus {{
      outline: none;
      border-color: #fcd7aa;
      box-shadow: 0 0 0 2px rgba(255, 237, 213, 0.38);
    }}
    .date-chip-apply {{
      display: inline-flex;
      align-items: center;
      gap: 5px;
      border: 1px solid #0f766e;
      border-radius: 999px;
      background: #0d9488;
      color: #ffffff;
      font-size: 0.74rem;
      font-weight: 700;
      padding: 5px 14px 5px 10px;
      cursor: pointer;
      line-height: 1;
      transition: background 0.15s, box-shadow 0.15s;
    }}
    .date-chip-apply .material-icons-outlined {{ font-size: 15px; }}
    .date-chip-apply:hover {{ background: #0f766e; }}
    .date-chip-apply:focus {{
      outline: none;
      box-shadow: 0 0 0 2px rgba(13, 148, 136, 0.35);
    }}
    .date-chip-apply:disabled {{ opacity: 0.55; cursor: not-allowed; }}
    .date-chip-reset {{
      display: inline-flex;
      align-items: center;
      justify-content: center;
      border: 1px solid #d4b896;
      border-radius: 999px;
      background: #fff7ed;
      color: #92400e;
      width: 30px;
      height: 30px;
      padding: 0;
      cursor: pointer;
      line-height: 1;
      transition: background 0.15s;
    }}
    .date-chip-reset .material-icons-outlined {{ font-size: 16px; }}
    .date-chip-reset:hover {{ background: #ffedd5; }}
    .date-chip-reset:focus {{
      outline: none;
      box-shadow: 0 0 0 2px rgba(255, 237, 213, 0.38);
    }}
    .adv-filter-wrap {{ position: relative; display: inline-flex; }}
    .adv-filter-btn {{
      display: inline-flex;
      align-items: center;
      gap: 5px;
      border: 1px solid #6366f1;
      border-radius: 999px;
      background: #4f46e5;
      color: #ffffff;
      font-size: 0.72rem;
      font-weight: 700;
      padding: 5px 12px 5px 9px;
      cursor: pointer;
      line-height: 1;
      transition: background 0.15s, box-shadow 0.15s;
    }}
    .adv-filter-btn .material-icons-outlined {{ font-size: 15px; }}
    .adv-filter-btn:hover {{ background: #4338ca; }}
    .adv-filter-btn:focus {{
      outline: none;
      box-shadow: 0 0 0 2px rgba(99, 102, 241, 0.35);
    }}
    .adv-filter-btn[aria-expanded="true"] {{ background: #4338ca; }}
    .adv-filter-menu {{
      position: absolute;
      top: calc(100% + 8px);
      right: 0;
      min-width: 230px;
      background: #fffbf5;
      border: 1px solid #e0d4c8;
      border-radius: 12px;
      box-shadow: 0 12px 36px rgba(0, 0, 0, 0.14);
      padding: 6px 0;
      z-index: 100;
      animation: advFilterFadeIn 0.12s ease-out;
    }}
    @keyframes advFilterFadeIn {{
      from {{ opacity: 0; transform: translateY(-4px); }}
      to   {{ opacity: 1; transform: translateY(0); }}
    }}
    .adv-filter-group-label {{
      padding: 8px 14px 4px;
      font-size: 0.65rem;
      font-weight: 800;
      text-transform: uppercase;
      letter-spacing: 0.06em;
      color: #78716c;
    }}
    .adv-filter-item {{
      display: block;
      width: 100%;
      padding: 8px 14px;
      border: none;
      background: transparent;
      text-align: left;
      font-size: 0.8rem;
      font-weight: 600;
      color: #44403c;
      cursor: pointer;
      transition: background 0.1s, color 0.1s;
    }}
    .adv-filter-item:hover {{ background: #e0f2f1; color: #0f766e; }}
    .adv-filter-divider {{
      height: 1px;
      background: #e7e0d8;
      margin: 6px 10px;
    }}
    .adv-filter-radio {{
      display: flex;
      align-items: center;
      gap: 8px;
      padding: 7px 14px;
      font-size: 0.8rem;
      font-weight: 600;
      color: #44403c;
      cursor: pointer;
      transition: background 0.1s;
    }}
    .adv-filter-radio:hover {{ background: #ede9fe; }}
    .adv-filter-radio input[type="radio"] {{
      accent-color: #4f46e5;
      margin: 0;
      width: 14px;
      height: 14px;
    }}
    .adv-filter-radio input[type="radio"]:checked + span {{ color: #4f46e5; font-weight: 700; }}
    .date-chip-status {{
      font-size: 0.72rem;
      color: #fef3c7;
      font-weight: 600;
      margin-left: 2px;
      white-space: nowrap;
    }}
    .date-chip-toggle {{
      display: inline-flex;
      align-items: center;
      gap: 6px;
      border: 1px solid #f3e8d8;
      border-radius: 999px;
      background: #fff7ed;
      color: #7c2d12;
      font-size: 0.74rem;
      font-weight: 700;
      min-height: 30px;
      padding: 0 10px;
      cursor: pointer;
      user-select: none;
    }}
    .date-chip-toggle input[type="checkbox"] {{
      width: 14px;
      height: 14px;
      margin: 0;
      accent-color: #0f766e;
      cursor: pointer;
    }}
    .page {{
      max-width: 1500px;
      margin: 0 auto;
      padding: 18px;
    }}
    .header {{
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 12px;
      padding: 14px 16px;
      margin-bottom: 12px;
    }}
    .header.is-collapsed > :not(.header-top) {{
      display: none;
    }}
    .header-top {{
      display: flex;
      align-items: center;
      justify-content: space-between;
      gap: 10px;
      flex-wrap: wrap;
    }}
    .header-actions {{
      display: flex;
      align-items: center;
      gap: 8px;
      flex-wrap: wrap;
    }}
    .title {{
      margin: 0 0 6px;
      font-size: 1.25rem;
      font-weight: 700;
      color: #0b3142;
    }}
    .meta {{
      margin: 0;
      color: var(--muted);
      font-size: 0.9rem;
    }}
    .scorecards {{
      margin-top: 10px;
      display: grid;
      grid-template-columns: repeat(2, minmax(220px, 1fr));
      gap: 8px;
    }}
    .section-title {{
      margin: 12px 0 6px;
      font-size: 0.86rem;
      font-weight: 800;
      letter-spacing: 0.03em;
      text-transform: uppercase;
      color: #355564;
    }}
    .section-note {{
      margin: 0 0 8px;
      font-size: 0.78rem;
      color: #5a7480;
      line-height: 1.35;
    }}
    .capacity-profile-bar {{
      margin-top: 10px;
      border: 1px solid #d7e3ea;
      border-radius: 10px;
      background: #f8fbfd;
      padding: 8px 10px;
      display: flex;
      flex-wrap: wrap;
      gap: 8px;
      align-items: center;
    }}
    .capacity-profile-toggle {{
      white-space: nowrap;
    }}
    .capacity-profile-overlay {{
      position: fixed;
      inset: 0;
      background: rgba(15, 23, 42, 0.45);
      z-index: 70;
      opacity: 0;
      visibility: hidden;
      pointer-events: none;
      transition: opacity 180ms ease, visibility 180ms ease;
    }}
    .capacity-profile-overlay.is-open {{
      opacity: 1;
      visibility: visible;
      pointer-events: auto;
    }}
    .capacity-profile-drawer {{
      position: fixed;
      top: 0;
      right: 0;
      width: 50vw;
      min-width: 360px;
      max-width: 92vw;
      height: 100vh;
      overflow-y: auto;
      border-left: 1px solid #d7e3ea;
      background: #f8fbfd;
      box-shadow: -12px 0 30px rgba(15, 23, 42, 0.24);
      z-index: 80;
      padding: 14px;
      transform: translateX(102%);
      pointer-events: none;
      transition: transform 220ms ease;
    }}
    .capacity-profile-drawer.is-resizing {{
      transition: none;
      user-select: none;
      cursor: ew-resize;
    }}
    .capacity-profile-resize-handle {{
      position: absolute;
      top: 0;
      left: 0;
      width: 14px;
      height: 100%;
      cursor: ew-resize;
      touch-action: none;
      z-index: 2;
    }}
    .capacity-profile-resize-handle::before {{
      content: "";
      position: absolute;
      top: 50%;
      left: 4px;
      transform: translateY(-50%);
      width: 4px;
      height: 48px;
      border-radius: 999px;
      background: #b6c7d2;
      box-shadow: 0 0 0 1px rgba(18, 49, 63, 0.06);
    }}
    .capacity-profile-resize-handle:hover::before,
    .capacity-profile-drawer.is-resizing .capacity-profile-resize-handle::before {{
      background: #2a6274;
    }}
    .capacity-profile-drawer.is-open {{
      transform: translateX(0);
      pointer-events: auto;
    }}
    .capacity-profile-drawer-head {{
      display: flex;
      align-items: center;
      justify-content: space-between;
      gap: 8px;
      margin-bottom: 8px;
    }}
    .capacity-profile-drawer-title {{
      margin: 0;
      font-size: 1rem;
      color: #0b3142;
    }}
    .capacity-profile-close {{
      width: 30px;
      height: 30px;
      border-radius: 999px;
      border: 1px solid #b6c7d2;
      background: #fff;
      color: #12313f;
      cursor: pointer;
      display: inline-flex;
      align-items: center;
      justify-content: center;
      padding: 0;
    }}
    .capacity-profile-close .material-icons-outlined {{
      font-size: 18px;
      line-height: 1;
    }}
    .capacity-profile-close:hover {{
      background: #eff6ff;
    }}
    body.capacity-modal-open {{
      overflow: hidden;
    }}
    body.capacity-modal-resizing {{
      cursor: ew-resize;
      user-select: none;
    }}
    .capacity-profile-label {{
      font-size: 0.8rem;
      font-weight: 700;
      color: #264653;
      white-space: nowrap;
    }}
    .capacity-profile-select {{
      min-width: 280px;
      flex: 1 1 340px;
      border: 1px solid #b6c7d2;
      border-radius: 8px;
      padding: 6px 8px;
      font-size: 0.86rem;
      color: #12313f;
      background: #fff;
    }}
    .capacity-profile-select:focus {{
      outline: none;
      border-color: #2a6274;
      box-shadow: 0 0 0 2px rgba(42, 98, 116, 0.15);
    }}
    .capacity-profile-details {{
      flex: 1 1 100%;
      font-size: 0.78rem;
      color: #3b5562;
      line-height: 1.35;
    }}
    .capacity-expanded {{
      flex: 1 1 100%;
      border: 1px solid #d7e3ea;
      border-radius: 10px;
      background: #ffffff;
      padding: 10px;
    }}
    .capacity-expanded-head {{
      display: flex;
      justify-content: space-between;
      align-items: center;
      gap: 10px;
      flex-wrap: wrap;
      margin-bottom: 8px;
    }}
    .capacity-expanded-title {{
      font-size: 0.86rem;
      font-weight: 800;
      color: #12313f;
      letter-spacing: 0.02em;
      text-transform: uppercase;
    }}
    .capacity-expanded-sub {{
      font-size: 0.74rem;
      color: #5a7480;
    }}
    .capacity-expanded-grid {{
      display: grid;
      grid-template-columns: repeat(auto-fit, minmax(170px, 1fr));
      gap: 6px;
      margin-bottom: 8px;
    }}
    .capacity-chip {{
      border: 1px solid #d7e3ea;
      border-radius: 8px;
      background: #f8fbfd;
      padding: 6px 8px;
    }}
    .capacity-chip .k {{
      font-size: 0.66rem;
      color: #5a7480;
      text-transform: uppercase;
      letter-spacing: 0.04em;
    }}
    .capacity-chip .v {{
      font-size: 0.85rem;
      color: #12313f;
      font-weight: 700;
      margin-top: 2px;
    }}
    .capacity-legend {{
      display: flex;
      gap: 6px;
      flex-wrap: wrap;
      margin-bottom: 8px;
    }}
    .capacity-legend .pill {{
      font-size: 0.68rem;
      padding: 2px 7px;
      border-radius: 999px;
      border: 1px solid #d7e3ea;
      color: #355564;
      background: #f1f7fb;
    }}
    .capacity-calendar-wrap {{
      display: flex;
      gap: 8px;
      overflow-x: auto;
      padding-bottom: 6px;
      scroll-snap-type: x proximity;
      scrollbar-width: thin;
      scrollbar-color: #b6c7d2 #edf4f8;
    }}
    .capacity-calendar-wrap::-webkit-scrollbar {{
      height: 10px;
    }}
    .capacity-calendar-wrap::-webkit-scrollbar-track {{
      background: #edf4f8;
      border-radius: 999px;
    }}
    .capacity-calendar-wrap::-webkit-scrollbar-thumb {{
      background: #b6c7d2;
      border-radius: 999px;
      border: 2px solid #edf4f8;
    }}
    .capacity-calendar-wrap::-webkit-scrollbar-thumb:hover {{
      background: #8ca6b5;
    }}
    .capacity-month {{
      border: 1px solid #d7e3ea;
      border-radius: 10px;
      background: #f8fbfd;
      padding: 8px;
      flex: 0 0 calc((100% - 16px) / 3);
      min-width: 250px;
      scroll-snap-align: start;
    }}
    .capacity-month-head {{
      font-size: 0.8rem;
      font-weight: 800;
      color: #12313f;
      margin-bottom: 6px;
    }}
    .capacity-month-grid {{
      display: grid;
      grid-template-columns: repeat(7, minmax(0, 1fr));
      gap: 4px;
    }}
    .capacity-dow {{
      font-size: 0.62rem;
      color: #5a7480;
      text-transform: uppercase;
      text-align: center;
    }}
    .capacity-day {{
      min-height: 44px;
      border: 1px solid #d7e3ea;
      border-radius: 6px;
      background: #ffffff;
      padding: 3px 4px;
    }}
    .capacity-day.is-out {{
      opacity: 0.35;
    }}
    .capacity-day.is-weekend {{
      background: #eef3f6;
      border-color: #d0d9df;
    }}
    .capacity-day.is-ramadan {{
      border-color: #38bdf8;
      box-shadow: inset 0 0 0 1px rgba(56, 189, 248, 0.18);
    }}
    .capacity-day.is-holiday {{
      border-color: #f59e0b;
      box-shadow: inset 0 0 0 1px rgba(245, 158, 11, 0.18);
    }}
    .capacity-day.has-leave {{
      border-color: #94a3b8;
      box-shadow: inset 0 0 0 1px rgba(148, 163, 184, 0.22);
    }}
    .capacity-day.has-ramadan-leave {{
      border-color: #38bdf8;
      box-shadow: inset 0 0 0 1px rgba(56, 189, 248, 0.28);
    }}
    .capacity-day.is-today {{
      background: #fde68a;
      border-color: #f59e0b;
      box-shadow: inset 0 0 0 1px rgba(217, 119, 6, 0.32), 0 0 0 1px rgba(245, 158, 11, 0.18);
    }}
    .capacity-day.is-today .capacity-day-num {{
      color: #78350f;
      font-weight: 900;
    }}
    .capacity-day.is-today .capacity-day-tag {{
      color: #78350f;
      border-color: #d97706;
      background: #fef3c7;
    }}
    .capacity-day-num {{
      font-size: 0.72rem;
      font-weight: 700;
      color: #12313f;
      line-height: 1.1;
    }}
    .capacity-day-tags {{
      margin-top: 2px;
      display: flex;
      gap: 3px;
      flex-wrap: wrap;
    }}
    .capacity-day-tag {{
      font-size: 0.58rem;
      line-height: 1;
      padding: 2px 4px;
      border-radius: 999px;
      border: 1px solid #d7e3ea;
      background: #f1f7fb;
      color: #355564;
    }}
    .capacity-day-tag.r {{
      border-color: #38bdf8;
      background: #e0f2fe;
    }}
    .capacity-day-tag.h {{
      border-color: #f59e0b;
      background: #fff4df;
    }}
    .capacity-day-tag.l {{
      border-color: #94a3b8;
      background: #eef2f6;
    }}
    .capacity-day-tag.rl {{
      border-color: #38bdf8;
      background: #e0f2fe;
    }}
    .capacity-empty {{
      font-size: 0.76rem;
      color: #5a7480;
    }}
    .capacity-profile-editor {{
      flex: 1 1 100%;
      display: grid;
      grid-template-columns: repeat(3, minmax(180px, 1fr));
      gap: 8px;
      align-items: end;
    }}
    .capacity-profile-field {{
      display: flex;
      flex-direction: column;
      gap: 4px;
      min-width: 0;
    }}
    .capacity-profile-field label {{
      font-size: 0.72rem;
      font-weight: 700;
      color: #264653;
      text-transform: uppercase;
      letter-spacing: 0.02em;
    }}
    .capacity-profile-field input {{
      border: 1px solid #b6c7d2;
      border-radius: 8px;
      padding: 6px 8px;
      font-size: 0.84rem;
      color: #12313f;
      background: #fff;
      min-height: 32px;
    }}
    .capacity-profile-field input:focus {{
      outline: none;
      border-color: #2a6274;
      box-shadow: 0 0 0 2px rgba(42, 98, 116, 0.15);
    }}
    .capacity-profile-actions {{
      display: flex;
      gap: 8px;
      align-items: center;
      flex-wrap: wrap;
    }}
    .capacity-profile-status {{
      flex: 1 1 100%;
      font-size: 0.78rem;
      line-height: 1.35;
    }}
    .capacity-profile-status[data-variant="info"] {{
      color: #3b5562;
    }}
    .capacity-profile-status[data-variant="success"] {{
      color: #166534;
    }}
    .capacity-profile-status[data-variant="error"] {{
      color: #991b1b;
    }}
    @media (max-width: 1200px) {{
      .capacity-month {{
        flex-basis: calc((100% - 8px) / 2);
      }}
    }}
    @media (max-width: 760px) {{
      .capacity-profile-drawer {{
        width: min(92vw, 560px);
        min-width: 280px;
      }}
      .capacity-profile-resize-handle {{
        display: none;
      }}
      .capacity-month {{
        flex-basis: 100%;
        min-width: 0;
      }}
    }}
    .score-card {{
      border: 1px solid #d7e3ea;
      background: #f8fbfd;
      border-radius: 10px;
      padding: 8px 10px;
      min-height: 66px;
      position: relative;
    }}
    .score-card.is-expandable {{
      cursor: pointer;
      transition: box-shadow 0.14s ease, transform 0.14s ease;
    }}
    .score-card.is-expandable:hover {{
      box-shadow: 0 6px 20px rgba(15, 23, 42, 0.09);
      transform: translateY(-1px);
    }}
    .score-card.is-expanded {{
      box-shadow: 0 0 0 2px rgba(244, 63, 94, 0.28);
    }}
    .efficiency-inline-wrap {{
      margin-top: 0;
      display: grid;
      grid-template-columns: 1fr;
      gap: 8px;
    }}
    #efficiency-under-actual {{
      grid-column: 2;
      align-self: start;
    }}
    .efficiency-inline-item {{
      border-radius: 10px;
      border: 1px solid #93c5fd;
      background: #eff6ff;
      padding: 8px 10px;
      min-height: 66px;
    }}
    .efficiency-inline-label {{
      color: #355564;
      font-size: 0.68rem;
      font-weight: 700;
      text-transform: uppercase;
      letter-spacing: 0.01em;
      margin: 0 0 4px;
      display: inline-flex;
      align-items: center;
      gap: 6px;
    }}
    .efficiency-inline-value {{
      margin: 0;
      color: #1d4ed8;
      font-size: 1.15rem;
      font-weight: 800;
      line-height: 1.15;
    }}
    .efficiency-inline-formula {{
      margin: 4px 0 0;
      color: #475569;
      font-size: 0.66rem;
      line-height: 1.25;
      font-weight: 600;
    }}
    .score-formula-chip {{
      position: absolute;
      top: 6px;
      right: 8px;
      font-size: 0.6rem;
      line-height: 1;
      font-weight: 700;
      color: #475569;
      background: rgba(248, 250, 252, 0.88);
      border: 1px solid #cbd5e1;
      border-radius: 999px;
      padding: 2px 6px;
      white-space: nowrap;
      pointer-events: none;
      display: inline-flex;
      align-items: center;
      gap: 3px;
    }}
    .score-formula-part {{
      display: inline-flex;
      align-items: center;
      gap: 2px;
    }}
    .score-formula-part .material-icons-outlined {{
      font-size: 11px;
      line-height: 1;
    }}
    .score-formula-mult {{
      color: #64748b;
    }}
    #score-total-capacity-card,
    #score-total-leaves-card,
    #score-total-leaves-planned-card,
    #score-total-capacity-planned-leaves-adjusted-card {{
      background: #fff1f2;
      border-color: #fda4af;
    }}
    #score-total-planned-card,
    #score-total-logged-card {{
      background: #eff6ff;
      border-color: #93c5fd;
    }}
    .score-label {{
      color: #355564;
      font-size: 0.75rem;
      font-weight: 700;
      text-transform: uppercase;
      letter-spacing: 0.01em;
      margin: 0 0 5px;
      display: inline-flex;
      align-items: center;
      gap: 6px;
    }}
    .score-action-btn {{
      display: inline-flex;
      align-items: center;
      justify-content: center;
      width: 18px;
      height: 18px;
      border-radius: 999px;
      border: 1px solid #94a3b8;
      background: #f8fafc;
      color: #334155;
      cursor: pointer;
      padding: 0;
    }}
    .score-action-btn .material-icons-outlined {{
      font-size: 13px;
      line-height: 1;
    }}
    .score-action-btn:hover {{
      background: #e2e8f0;
    }}
    .score-action-btn:focus {{
      outline: none;
      border-color: #2a6274;
      box-shadow: 0 0 0 2px rgba(42, 98, 116, 0.15);
    }}
    .score-info {{
      position: relative;
      display: inline-flex;
      align-items: center;
      justify-content: center;
      width: 16px;
      height: 16px;
      border-radius: 999px;
      border: 1px solid #94a3b8;
      color: #334155;
      font-size: 11px;
      line-height: 1;
      font-weight: 700;
      background: #f8fafc;
      cursor: default;
      user-select: none;
      text-transform: none;
    }}
    .score-info-tip {{
      position: absolute;
      left: 50%;
      top: calc(100% + 8px);
      transform: translateX(-50%);
      min-width: 240px;
      max-width: 320px;
      padding: 8px 10px;
      border-radius: 8px;
      border: 1px solid #cbd5e1;
      background: #ffffff;
      color: #1f2937;
      font-size: 12px;
      line-height: 1.35;
      text-transform: none;
      letter-spacing: normal;
      box-shadow: 0 10px 24px rgba(15, 23, 42, 0.18);
      opacity: 0;
      visibility: hidden;
      pointer-events: none;
      z-index: 30;
      transition: opacity 0.14s ease;
      white-space: pre-line;
    }}
    .score-info:hover .score-info-tip,
    .score-info:focus .score-info-tip,
    .score-info:focus-visible .score-info-tip {{
      opacity: 1;
      visibility: visible;
    }}
    .score-value {{
      color: #0b3142;
      font-size: 1.15rem;
      font-weight: 800;
      line-height: 1.15;
      margin: 0;
    }}
    .score-capacity-meta {{
      margin: 4px 0 0;
      display: inline-flex;
      align-items: center;
      gap: 6px;
      font-size: 0.72rem;
      color: #355564;
      font-weight: 600;
    }}
    .score-capacity-meta-item {{
      display: inline-flex;
      align-items: center;
      gap: 3px;
      line-height: 1;
    }}
    .score-capacity-meta-item .material-icons-outlined {{
      font-size: 13px;
      line-height: 1;
    }}
    .score-capacity-meta-mult {{
      font-size: 0.75rem;
      color: #47606d;
    }}
    #score-total-capacity-card .score-value,
    #score-total-leaves-card .score-value,
    #score-total-leaves-planned-card .score-value,
    #score-total-capacity-planned-leaves-adjusted-card .score-value {{ color: #9f1239; }}
    #score-total-planned-card .score-value,
    #score-total-logged-card .score-value {{ color: #1d4ed8; }}
    .score-card.delta-pos .score-value {{
      color: #15803d;
    }}
    .score-card.delta-neg .score-value {{
      color: #b91c1c;
    }}
    .score-card.delta-zero .score-value {{
      color: #1f2937;
    }}
    .score-details-panel {{
      margin-top: 10px;
      border: 1px solid #fda4af;
      border-radius: 10px;
      background: #fff7f8;
      padding: 10px;
    }}
    .score-details-panel[hidden] {{
      display: none;
    }}
    .score-details-title {{
      margin: 0 0 8px;
      font-size: 0.9rem;
      font-weight: 700;
      color: #9f1239;
    }}
    .score-details-meta {{
      margin: 0 0 8px;
      font-size: 0.78rem;
      color: #6b7280;
    }}
    .score-details-table-wrap {{
      overflow-x: auto;
    }}
    .score-details-table {{
      width: 100%;
      border-collapse: collapse;
      min-width: 640px;
      font-size: 0.84rem;
    }}
    .score-details-table th,
    .score-details-table td {{
      border: 1px solid #fbcfe8;
      padding: 7px 8px;
      text-align: left;
      vertical-align: middle;
      white-space: nowrap;
    }}
    .score-details-table th {{
      background: #ffe4e6;
      color: #9f1239;
      font-size: 0.75rem;
      text-transform: uppercase;
      letter-spacing: 0.02em;
    }}
    .score-details-table td.hours {{
      text-align: right;
      font-variant-numeric: tabular-nums;
    }}
    .score-details-empty {{
      color: #64748b;
      font-style: italic;
      text-align: center;
    }}
    .leaves-link-btn {{
      display: inline-flex;
      align-items: center;
      justify-content: center;
      width: 24px;
      height: 24px;
      border-radius: 999px;
      border: 1px solid #fda4af;
      background: #fff1f2;
      color: #be123c;
      text-decoration: none;
    }}
    .leaves-link-btn:hover {{
      background: #ffe4e6;
    }}
    .leaves-link-btn .material-icons-outlined {{
      font-size: 15px;
      line-height: 1;
    }}
    .planned-project-row td {{
      background: #eff6ff;
      border-color: #bfdbfe;
      padding: 6px 8px;
    }}
    .planned-project-toggle {{
      border: 0;
      background: transparent;
      color: #1e3a8a;
      font-size: 0.88rem;
      font-weight: 700;
      cursor: pointer;
      padding: 8px 6px;
      min-height: 36px;
      min-width: 260px;
      display: inline-flex;
      align-items: center;
      gap: 6px;
      border-radius: 8px;
    }}
    .planned-project-toggle:hover {{
      background: rgba(59, 130, 246, 0.12);
    }}
    .planned-project-toggle:focus-visible {{
      outline: 2px solid #2563eb;
      outline-offset: 1px;
    }}
    .planned-project-toggle .caret {{
      width: 10px;
      text-align: center;
      color: #1d4ed8;
    }}
    .planned-epic-row td {{
      background: #ffffff;
    }}
    .planned-epic-row td.hours {{
      text-align: right;
      font-variant-numeric: tabular-nums;
    }}
    .toolbar {{
      display: flex;
      flex-wrap: wrap;
      gap: 8px;
      margin-top: 12px;
    }}
    .search-wrap {{
      display: flex;
      align-items: center;
      gap: 8px;
      flex: 1 1 420px;
      min-width: 260px;
    }}
    .search-input {{
      width: 100%;
      border: 1px solid #b6c7d2;
      border-radius: 8px;
      padding: 7px 10px;
      font-size: 0.9rem;
      color: #12313f;
      background: #fff;
    }}
    .search-input:focus {{
      outline: none;
      border-color: #2a6274;
      box-shadow: 0 0 0 2px rgba(42, 98, 116, 0.15);
    }}
    .search-meta {{
      color: #5f7481;
      font-size: 0.82rem;
      white-space: nowrap;
    }}
    .project-filter {{
      position: relative;
      flex: 0 1 290px;
      min-width: 220px;
    }}
    .project-filter-progress {{
      position: absolute;
      left: 8px;
      right: 8px;
      top: calc(100% - 3px);
      height: 3px;
      border-radius: 999px;
      background: rgba(148, 163, 184, 0.2);
      overflow: hidden;
      opacity: 0;
      pointer-events: none;
      transform: scaleY(0.8);
      transition: opacity 140ms ease;
      z-index: 24;
    }}
    .project-filter-progress.is-active {{
      opacity: 1;
    }}
    .project-filter-progress-bar {{
      width: 42%;
      height: 100%;
      border-radius: inherit;
      background: linear-gradient(90deg, #0f766e 0%, #22c55e 55%, #67e8f9 100%);
      animation: project-filter-progress-slide 1.1s linear infinite;
      transform: translateX(-110%);
    }}
    @keyframes project-filter-progress-slide {{
      0% {{ transform: translateX(-110%); }}
      100% {{ transform: translateX(240%); }}
    }}
    .team-filter-progress {{
      position: absolute;
      left: 8px;
      right: 8px;
      top: calc(100% - 3px);
      height: 3px;
      border-radius: 999px;
      background: rgba(148, 163, 184, 0.2);
      overflow: hidden;
      opacity: 0;
      pointer-events: none;
      transform: scaleY(0.8);
      transition: opacity 140ms ease;
      z-index: 24;
    }}
    .team-filter-progress.is-active {{
      opacity: 1;
    }}
    .team-filter-progress-bar {{
      width: 42%;
      height: 100%;
      border-radius: inherit;
      background: linear-gradient(90deg, #2563eb 0%, #22c55e 55%, #67e8f9 100%);
      animation: project-filter-progress-slide 1.1s linear infinite;
      transform: translateX(-110%);
    }}
    .project-filter-btn {{
      width: 100%;
      display: inline-flex;
      align-items: center;
      justify-content: space-between;
      gap: 8px;
      border: 1px solid #b6c7d2;
      background: #fff;
      color: #12313f;
      border-radius: 8px;
      padding: 7px 10px;
      cursor: pointer;
      font-size: 0.9rem;
      text-align: left;
    }}
    .project-filter-btn:focus {{
      outline: none;
      border-color: #2a6274;
      box-shadow: 0 0 0 2px rgba(42, 98, 116, 0.15);
    }}
    .project-filter-caret {{
      font-size: 0.85rem;
      color: #5f7481;
      flex: 0 0 auto;
    }}
    .project-filter-menu {{
      position: absolute;
      top: calc(100% + 6px);
      left: 0;
      width: 100%;
      background: #fff;
      border: 1px solid #d7e3ea;
      border-radius: 10px;
      box-shadow: 0 10px 26px rgba(2, 6, 23, 0.16);
      z-index: 20;
      padding: 8px;
      display: none;
    }}
    .project-filter.open .project-filter-menu {{
      display: block;
    }}
    .project-filter-actions {{
      display: flex;
      gap: 6px;
      margin-bottom: 8px;
    }}
    .project-filter-action {{
      border: 1px solid #d1d5db;
      background: #f8fafc;
      color: #1f2937;
      border-radius: 7px;
      padding: 4px 9px;
      font-size: 0.76rem;
      cursor: pointer;
    }}
    .project-filter-action:hover {{
      background: #eef2f7;
    }}
    .project-filter-options {{
      max-height: 220px;
      overflow: auto;
      border: 1px solid #e5e7eb;
      border-radius: 8px;
      padding: 5px;
      background: #fdfefe;
    }}
    .project-option {{
      display: flex;
      align-items: center;
      gap: 7px;
      font-size: 0.82rem;
      color: #1f2937;
      padding: 5px 6px;
      border-radius: 6px;
      cursor: pointer;
    }}
    .project-option:hover {{
      background: #eff6ff;
    }}
    .project-option input {{
      margin: 0;
      flex: 0 0 auto;
    }}
    .project-option-label {{
      display: inline-block;
      min-width: 0;
      overflow: hidden;
      text-overflow: ellipsis;
      white-space: nowrap;
    }}
    .view-options {{
      position: relative;
    }}
    .view-options-toggle {{
      display: inline-flex;
      align-items: center;
      gap: 8px;
      border: 1px solid #bfdbfe;
      background: #eff6ff;
      color: #1d4ed8;
      border-radius: 8px;
      padding: 6px 10px;
      cursor: pointer;
      font-size: .82rem;
      font-weight: 700;
    }}
    .view-options-menu {{
      position: absolute;
      right: 0;
      top: calc(100% + 6px);
      min-width: 220px;
      border: 1px solid #d7e3ea;
      border-radius: 10px;
      background: #fff;
      box-shadow: 0 10px 26px rgba(2, 6, 23, 0.16);
      padding: 8px;
      display: none;
      z-index: 25;
    }}
    .view-options.open .view-options-menu {{
      display: grid;
      gap: 6px;
    }}
    .view-options-menu .btn {{
      justify-content: flex-start;
      width: 100%;
    }}
    .legend {{
      margin-top: 10px;
      display: flex;
      flex-wrap: wrap;
      gap: 8px 10px;
      align-items: center;
    }}
    .legend-title {{
      font-size: 0.82rem;
      font-weight: 700;
      color: #244552;
      margin-right: 4px;
    }}
    .legend-item {{
      display: inline-flex;
      align-items: center;
      gap: 6px;
      font-size: 0.78rem;
      color: #264653;
      background: #f8fbfd;
      border: 1px solid #d7e3ea;
      border-radius: 999px;
      padding: 3px 8px;
    }}
    .legend-swatch {{
      width: 14px;
      height: 14px;
      border-radius: 4px;
      border: 1px solid rgba(0,0,0,.15);
      flex: 0 0 auto;
    }}
    .btn {{
      border: 1px solid #255f73;
      background: #0f4c5c;
      color: #fff;
      border-radius: 8px;
      padding: 7px 12px;
      cursor: pointer;
      font-size: 0.9rem;
    }}
    .btn .material-icons-outlined {{
      font-size: 0.92rem;
      vertical-align: -2px;
      margin-right: 0.2rem;
    }}
    #theme-toggle {{
      background: rgba(191, 219, 254, 0.45);
      border-color: #93c5fd;
      color: #1e3a8a;
    }}
    #theme-toggle .material-icons-outlined {{
      color: #1d4ed8;
    }}
    #theme-toggle:hover {{
      background: rgba(191, 219, 254, 0.62);
    }}
    html[data-theme="dark"] #theme-toggle {{
      background: rgba(254, 243, 199, 0.45);
      border-color: #fcd34d;
      color: #fff;
    }}
    html[data-theme="dark"] #theme-toggle .material-icons-outlined {{
      color: #facc15;
    }}
    html[data-theme="dark"] #theme-toggle:hover {{
      background: rgba(254, 243, 199, 0.62);
    }}
    .btn.alt {{
      background: #fff;
      color: #255f73;
    }}
    .btn.active {{
      background: #0f4c5c;
      color: #fff;
      border-color: #0f4c5c;
    }}
    .btn.alert {{
      border-color: #b91c1c;
      background: #fff1f2;
      color: #b91c1c;
    }}
    .btn.alert.active {{
      border-color: #991b1b;
      background: #b91c1c;
      color: #fff;
    }}
    .btn:disabled {{
      opacity: 0.55;
      cursor: not-allowed;
      filter: none;
    }}
    .btn:disabled:hover {{ filter: none; }}
    .btn:hover {{ filter: brightness(1.05); }}
    .table-wrap {{
      --aspect-col-width: 360px;
      --type-col-width: 120px;
      --assignee-col-width: 170px;
      --metric-col-width: 130px;
      --resource-col-width: 150px;
      --date-col-width: 150px;
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 12px;
      max-height: calc(100vh - 220px);
      overflow: auto;
      width: 100%;
    }}
    .table-section-title {{
      margin: 0 0 10px 0;
      font-size: 1.15rem;
      font-weight: 700;
      color: var(--head);
      letter-spacing: 0.02em;
    }}
    table {{
      width: max-content;
      border-collapse: separate;
      border-spacing: 0;
      min-width: 100%;
      table-layout: auto;
    }}
    thead th {{
      position: sticky;
      top: 0;
      z-index: 2;
      background: var(--head);
      color: var(--head-text);
      text-align: left;
      padding: 10px 10px;
      font-size: 0.9rem;
      white-space: nowrap;
      border-bottom: 1px solid #0a3946;
    }}
    tbody td {{
      border-top: 1px solid var(--line);
      padding: 8px 10px;
      font-size: 0.9rem;
      vertical-align: middle;
      background: #fff;
      min-width: var(--metric-col-width);
    }}
    tbody tr:nth-child(even) td {{ background: #fbfdff; }}
    tbody tr.sticky-parent-row > td {{
      position: sticky;
      top: var(--sticky-parent-top, 0px);
      z-index: 1;
      box-shadow: inset 0 -1px 0 rgba(15, 23, 42, 0.12);
    }}
    tbody tr.sticky-parent-row > td.col-aspect {{
      z-index: 2;
    }}
    thead th.col-aspect,
    tbody td.col-aspect {{
      width: var(--aspect-col-width);
      min-width: var(--aspect-col-width);
      max-width: var(--aspect-col-width);
      position: sticky;
      left: 0;
      z-index: 3;
      box-shadow: 2px 0 0 rgba(15, 76, 92, 0.14);
    }}
    thead th.col-aspect {{
      z-index: 4;
      position: sticky;
      padding-right: 18px;
    }}
    .col-resizer {{
      position: absolute;
      top: 0;
      right: -4px;
      width: 10px;
      height: 100%;
      cursor: col-resize;
      touch-action: none;
      z-index: 6;
    }}
    .col-resizer::before {{
      content: "";
      position: absolute;
      right: 4px;
      top: 20%;
      bottom: 20%;
      width: 2px;
      border-radius: 2px;
      background: rgba(12, 58, 72, 0.25);
    }}
    thead th.col-aspect:hover .col-resizer::before {{
      background: rgba(12, 58, 72, 0.45);
    }}
    body.resizing-col {{
      cursor: col-resize;
      user-select: none;
    }}
    .col-type {{
      width: var(--type-col-width);
      min-width: var(--type-col-width);
      white-space: nowrap;
    }}
    .col-assignee {{
      width: var(--assignee-col-width);
      min-width: var(--assignee-col-width);
    }}
    .col-resource {{
      width: var(--resource-col-width);
      min-width: var(--resource-col-width);
      white-space: nowrap;
    }}
    .col-date {{
      width: var(--date-col-width);
      min-width: var(--date-col-width);
      white-space: nowrap;
    }}
    .num {{ text-align: right; }}
    thead th:nth-child(4),
    tbody td:nth-child(4),
    thead th:nth-child(5),
    tbody td:nth-child(5),
    thead th:nth-child(6),
    tbody td:nth-child(6) {{
      width: var(--metric-col-width);
      min-width: var(--metric-col-width);
      white-space: nowrap;
    }}
    thead th:nth-child(7),
    tbody td:nth-child(7),
    thead th:nth-child(8),
    tbody td:nth-child(8),
    thead th:nth-child(9),
    tbody td:nth-child(9),
    thead th:nth-child(10),
    tbody td:nth-child(10),
    thead th:nth-child(11),
    tbody td:nth-child(11) {{
      width: var(--metric-col-width);
      min-width: var(--metric-col-width);
      white-space: nowrap;
    }}
    .aspect-cell {{
      display: flex;
      align-items: center;
      gap: 6px;
      min-height: 20px;
      min-width: 0;
    }}
    .toggle {{
      width: 18px;
      height: 18px;
      border: 1px solid #a5b7c3;
      border-radius: 5px;
      background: #fff;
      color: #294b5a;
      cursor: pointer;
      line-height: 1;
      font-size: 12px;
      padding: 0;
      flex: 0 0 auto;
    }}
    .toggle.placeholder {{
      visibility: hidden;
      cursor: default;
    }}
    .node-text {{
      display: block;
      flex: 1 1 auto;
      min-width: 0;
      overflow: hidden;
      text-overflow: ellipsis;
      white-space: nowrap;
    }}
    .assignee-chip {{
      display: inline-flex;
      align-items: center;
      padding: 2px 8px;
      border-radius: 999px;
      font-size: 0.72rem;
      font-weight: 600;
      border: 1px solid #d1d5db;
      background: #f3f4f6;
      color: #374151;
      white-space: nowrap;
      margin-left: 8px;
      flex: 0 0 auto;
    }}
    .jira-link {{
      display: inline-flex;
      align-items: center;
      justify-content: center;
      width: 24px;
      height: 24px;
      border-radius: 999px;
      border: 1px solid #93c5fd;
      background: #eff6ff;
      color: #1d4ed8;
      text-decoration: none;
      font-size: 0.8rem;
      margin-left: 6px;
      white-space: nowrap;
      flex: 0 0 auto;
    }}
    .jira-link .material-icons-outlined {{
      font-size: 15px;
      line-height: 1;
    }}
    .jira-link:hover {{
      background: #dbeafe;
    }}
    .lvl-1 .node-text {{ font-weight: 700; color: #0b3142; }}
    .lvl-2 .node-text {{ font-weight: 700; color: #264653; }}
    .lvl-3 .node-text {{ font-weight: 600; color: #2a6274; }}
    .lvl-4 .node-text {{ font-weight: 600; }}
    .lvl-5 .node-text {{ color: #374151; }}
    .lvl-6 .node-text {{ color: #4b5563; font-style: italic; }}
    tr.row-type-project td {{ background: #dbeeff; }}
    tr.row-type-product td {{ background: #f2f7a8; }}
    tr.row-type-rmi td {{ background: #eadfff; }}
    tr.row-type-story td {{ background: #dfeaff; }}
    tr.row-type-subtask td {{ background: #ecfde9; }}
    tr.row-type-assignee td {{ background: #f6fff4; }}
    tr.row-type-bug td {{ background: #ffeedd; }}
    .type-chip {{
      display: inline-flex;
      align-items: center;
      justify-content: flex-start;
      gap: 4px;
      padding: 2px 8px;
      border-radius: 999px;
      font-size: 0.74rem;
      font-weight: 700;
      border: 1px solid rgba(0,0,0,.15);
      line-height: 1.1;
      white-space: nowrap;
    }}
    .type-chip-code {{
      display: inline-flex;
      align-items: center;
      justify-content: center;
      min-width: 26px;
      padding: 0 4px;
      border-radius: 999px;
      border: 1px solid currentColor;
      font-size: 0.56rem;
      letter-spacing: 0.03em;
      font-weight: 800;
      line-height: 1.5;
    }}
    .type-chip-text {{
      font-weight: 700;
    }}
    .type-chip.project {{ background: #dbeeff; color: #0b3f75; border-color: #91bde7; }}
    .type-chip.product {{ background: #f2f7a8; color: #5f6200; border-color: #d0d66b; }}
    .type-chip.rmi {{ background: #eadfff; color: #4f3191; border-color: #b39be5; }}
    .type-chip.story {{ background: #dfeaff; color: #1f4f9e; border-color: #97b6ee; }}
    .type-chip.subtask {{ background: #ecfde9; color: #2d7b2a; border-color: #b0dbab; }}
    .type-chip.bug {{ background: #ffeedd; color: #8d4a12; border-color: #f0c8a5; }}
    .tree-lines {{
      background-repeat: no-repeat;
      background-size: 100% 100%;
    }}
    .pill {{
      display: inline-block;
      font-size: 0.75rem;
      line-height: 1;
      padding: 3px 6px;
      border-radius: 999px;
      border: 1px solid #ecdba3;
      background: #fff8df;
      color: #7a6100;
      margin-left: 8px;
    }}
    .danger-chip {{
      display: inline-flex;
      align-items: center;
      padding: 2px 6px;
      border-radius: 999px;
      border: 1px solid #fecaca;
      background: #fef2f2;
      color: #991b1b;
      font-size: 0.72rem;
      line-height: 1;
      font-weight: 700;
      margin-left: 8px;
      white-space: nowrap;
    }}
    html[data-density="compact"] thead th {{
      padding: 6px 10px;
    }}
    html[data-density="compact"] tbody td {{
      padding: 5px 10px;
    }}
    html[data-density="compact"] .toggle {{
      width: 15px;
      height: 15px;
      font-size: 11px;
    }}
    html[data-density="compact"] .type-chip {{
      padding: 1px 7px;
      line-height: 1;
    }}
    html[data-density="compact"] .assignee-chip {{
      padding: 1px 7px;
      line-height: 1;
    }}
    html[data-density="compact"] .jira-link {{
      width: 20px;
      height: 20px;
    }}
    html[data-density="compact"] .danger-chip {{
      padding: 1px 6px;
      line-height: 1;
    }}
    .delta-neg {{
      color: #b91c1c;
      font-weight: 700;
    }}
    .delta-pos {{
      color: #15803d;
      font-weight: 700;
    }}
    .delta-zero {{
      color: #ffffff;
      font-weight: 700;
    }}
    tr.row-danger td {{
      border-top-color: #f3b4b4;
    }}
    @media (max-width: 900px) {{
      .page {{ padding: 10px; }}
      .header {{ padding: 12px; }}
      .title {{ font-size: 1.05rem; }}
      tbody td, thead th {{ font-size: 0.82rem; }}
      .scorecards {{
        grid-template-columns: repeat(2, minmax(150px, 1fr));
      }}
      .efficiency-inline-wrap {{
        grid-template-columns: 1fr;
      }}
      #efficiency-under-actual {{
        grid-column: 1 / -1;
      }}
      .capacity-profile-select {{
        min-width: 100%;
      }}
      .capacity-profile-editor {{
        grid-template-columns: 1fr;
      }}
      .table-wrap {{
        width: 100%;
      }}
      .col-aspect {{ min-width: 300px; }}
      .project-filter {{ flex-basis: 100%; }}
      .header-top {{
        align-items: flex-start;
      }}
    }}
    html[data-theme="dark"] body {{
      color: #e5e7eb;
      background:
        radial-gradient(1000px 280px at 10% -5%, #102a43 0%, transparent 65%),
        linear-gradient(180deg, #0f172a, #0b1220);
    }}
    html[data-theme="dark"] .header,
    html[data-theme="dark"] .table-wrap {{
      background: #111827;
      border-color: #1f2937;
    }}
    html[data-theme="dark"] .table-section-title {{ color: #e5e7eb; }}
    html[data-theme="dark"] .title {{ color: #f3f4f6; }}
    html[data-theme="dark"] .meta,
    html[data-theme="dark"] .search-meta,
    html[data-theme="dark"] .legend-title {{ color: #9ca3af; }}
    html[data-theme="dark"] .score-card {{
      background: #0f172a;
      border-color: #334155;
    }}
    html[data-theme="dark"] .score-formula-chip {{
      color: #cbd5e1;
      background: rgba(15, 23, 42, 0.82);
      border-color: #475569;
    }}
    html[data-theme="dark"] .score-formula-mult {{
      color: #94a3b8;
    }}
    html[data-theme="dark"] #score-total-capacity-card,
    html[data-theme="dark"] #score-total-leaves-card,
    html[data-theme="dark"] #score-total-leaves-planned-card,
    html[data-theme="dark"] #score-total-capacity-planned-leaves-adjusted-card {{
      background: #4c0519;
      border-color: #fb7185;
    }}
    html[data-theme="dark"] #score-total-planned-card,
    html[data-theme="dark"] #score-total-logged-card {{
      background: #172554;
      border-color: #60a5fa;
    }}
    html[data-theme="dark"] .efficiency-inline-item {{
      background: rgba(2, 6, 23, 0.62);
      border-color: #60a5fa;
    }}
    html[data-theme="dark"] .efficiency-inline-label,
    html[data-theme="dark"] .efficiency-inline-value {{
      color: #bfdbfe;
    }}
    html[data-theme="dark"] .efficiency-inline-formula {{
      color: #cbd5e1;
    }}
    html[data-theme="dark"] .capacity-profile-bar {{
      background: #0f172a;
      border-color: #334155;
    }}
    html[data-theme="dark"] .capacity-profile-overlay {{
      background: rgba(2, 6, 23, 0.7);
    }}
    html[data-theme="dark"] .capacity-profile-drawer {{
      background: #0f172a;
      border-left-color: #334155;
    }}
    html[data-theme="dark"] .capacity-profile-resize-handle::before {{
      background: #475569;
      box-shadow: 0 0 0 1px rgba(148, 163, 184, 0.12);
    }}
    html[data-theme="dark"] .capacity-profile-resize-handle:hover::before,
    html[data-theme="dark"] .capacity-profile-drawer.is-resizing .capacity-profile-resize-handle::before {{
      background: #93c5fd;
    }}
    html[data-theme="dark"] .capacity-profile-drawer-title {{
      color: #f3f4f6;
    }}
    html[data-theme="dark"] .capacity-profile-close {{
      background: #111827;
      border-color: #334155;
      color: #e5e7eb;
    }}
    html[data-theme="dark"] .capacity-profile-close:hover {{
      background: #1f2937;
    }}
    html[data-theme="dark"] .capacity-profile-label {{
      color: #93c5fd;
    }}
    html[data-theme="dark"] .capacity-profile-select {{
      background: #111827;
      color: #e5e7eb;
      border-color: #334155;
    }}
    html[data-theme="dark"] .capacity-profile-details {{
      color: #cbd5e1;
    }}
    html[data-theme="dark"] .capacity-expanded,
    html[data-theme="dark"] .capacity-chip,
    html[data-theme="dark"] .capacity-month {{
      background: #111827;
      border-color: #334155;
    }}
    html[data-theme="dark"] .capacity-expanded-title,
    html[data-theme="dark"] .capacity-month-head,
    html[data-theme="dark"] .capacity-chip .v,
    html[data-theme="dark"] .capacity-day-num {{
      color: #f3f4f6;
    }}
    html[data-theme="dark"] .capacity-expanded-sub,
    html[data-theme="dark"] .capacity-chip .k,
    html[data-theme="dark"] .capacity-dow,
    html[data-theme="dark"] .capacity-empty {{
      color: #cbd5e1;
    }}
    html[data-theme="dark"] .capacity-legend .pill {{
      border-color: #334155;
      background: #172554;
      color: #bfdbfe;
    }}
    html[data-theme="dark"] .capacity-calendar-wrap {{
      scrollbar-color: #475569 #0f172a;
    }}
    html[data-theme="dark"] .capacity-calendar-wrap::-webkit-scrollbar-track {{
      background: #0f172a;
    }}
    html[data-theme="dark"] .capacity-calendar-wrap::-webkit-scrollbar-thumb {{
      background: #475569;
      border-color: #0f172a;
    }}
    html[data-theme="dark"] .capacity-calendar-wrap::-webkit-scrollbar-thumb:hover {{
      background: #64748b;
    }}
    html[data-theme="dark"] .capacity-day {{
      background: #0f172a;
      border-color: #334155;
    }}
    html[data-theme="dark"] .capacity-day.is-weekend {{
      background: #1f2937;
      border-color: #475569;
    }}
    html[data-theme="dark"] .capacity-day-tag {{
      background: #172554;
      border-color: #334155;
      color: #bfdbfe;
    }}
    html[data-theme="dark"] .capacity-day-tag.h {{
      background: #422006;
    }}
    html[data-theme="dark"] .capacity-day-tag.l {{
      background: #374151;
    }}
    html[data-theme="dark"] .capacity-profile-field label {{
      color: #93c5fd;
    }}
    html[data-theme="dark"] .capacity-profile-field input {{
      background: #111827;
      color: #e5e7eb;
      border-color: #334155;
    }}
    html[data-theme="dark"] .capacity-profile-status[data-variant="info"] {{
      color: #cbd5e1;
    }}
    html[data-theme="dark"] .capacity-profile-status[data-variant="success"] {{
      color: #86efac;
    }}
    html[data-theme="dark"] .capacity-profile-status[data-variant="error"] {{
      color: #fca5a5;
    }}
    html[data-theme="dark"] .score-label {{
      color: #93c5fd;
    }}
    html[data-theme="dark"] .section-title {{
      color: #93c5fd;
    }}
    html[data-theme="dark"] .section-note {{
      color: #cbd5e1;
    }}
    html[data-theme="dark"] .score-info {{
      border-color: #475569;
      color: #cbd5e1;
      background: #0f172a;
    }}
    html[data-theme="dark"] .score-action-btn {{
      border-color: #475569;
      color: #cbd5e1;
      background: #0f172a;
    }}
    html[data-theme="dark"] .score-action-btn:hover {{
      background: #1e293b;
    }}
    html[data-theme="dark"] .score-info-tip {{
      border-color: #334155;
      background: #111827;
      color: #e5e7eb;
    }}
    html[data-theme="dark"] .score-value {{
      color: #f3f4f6;
    }}
    html[data-theme="dark"] .score-capacity-meta {{
      color: #bfdbfe;
    }}
    html[data-theme="dark"] .score-capacity-meta-mult {{
      color: #93c5fd;
    }}
    html[data-theme="dark"] .score-card.delta-pos .score-value {{
      color: #86efac;
    }}
    html[data-theme="dark"] .score-card.delta-neg .score-value {{
      color: #fca5a5;
    }}
    html[data-theme="dark"] .score-card.delta-zero .score-value {{
      color: #e5e7eb;
    }}
    html[data-theme="dark"] .score-details-panel {{
      border-color: #fb7185;
      background: #3f0b1f;
    }}
    html[data-theme="dark"] .score-details-title {{
      color: #fda4af;
    }}
    html[data-theme="dark"] .score-details-meta {{
      color: #cbd5e1;
    }}
    html[data-theme="dark"] .score-details-table th,
    html[data-theme="dark"] .score-details-table td {{
      border-color: #9f1239;
    }}
    html[data-theme="dark"] .score-details-table th {{
      background: #6b102b;
      color: #fecdd3;
    }}
    html[data-theme="dark"] .score-details-empty {{
      color: #cbd5e1;
    }}
    html[data-theme="dark"] .leaves-link-btn {{
      border-color: #fb7185;
      background: #6b102b;
      color: #fecdd3;
    }}
    html[data-theme="dark"] .leaves-link-btn:hover {{
      background: #881337;
    }}
    html[data-theme="dark"] .planned-project-row td {{
      background: #172554;
      border-color: #1d4ed8;
    }}
    html[data-theme="dark"] .planned-project-toggle {{
      color: #bfdbfe;
    }}
    html[data-theme="dark"] .planned-project-toggle:hover {{
      background: rgba(147, 197, 253, 0.2);
    }}
    html[data-theme="dark"] .planned-project-toggle:focus-visible {{
      outline-color: #93c5fd;
    }}
    html[data-theme="dark"] .planned-project-toggle .caret {{
      color: #93c5fd;
    }}
    html[data-theme="dark"] .planned-epic-row td {{
      background: #111827;
    }}
    html[data-theme="dark"] .search-input {{
      background: #0f172a;
      color: #e5e7eb;
      border-color: #374151;
    }}
    html[data-theme="dark"] .project-filter-btn {{
      background: #0f172a;
      color: #e5e7eb;
      border-color: #374151;
    }}
    html[data-theme="dark"] .project-filter-progress {{
      background: rgba(71, 85, 105, 0.45);
    }}
    html[data-theme="dark"] .project-filter-progress-bar {{
      background: linear-gradient(90deg, #22c55e 0%, #38bdf8 60%, #f8fafc 100%);
    }}
    html[data-theme="dark"] .team-filter-progress {{
      background: rgba(71, 85, 105, 0.45);
    }}
    html[data-theme="dark"] .team-filter-progress-bar {{
      background: linear-gradient(90deg, #60a5fa 0%, #22c55e 60%, #f8fafc 100%);
    }}
    html[data-theme="dark"] .project-filter-caret {{
      color: #94a3b8;
    }}
    html[data-theme="dark"] .project-filter-menu {{
      background: #0f172a;
      border-color: #334155;
      box-shadow: 0 14px 32px rgba(2, 6, 23, 0.5);
    }}
    html[data-theme="dark"] .view-options-toggle {{
      background: #1e3a5f;
      border-color: #3b82f6;
      color: #dbeafe;
    }}
    html[data-theme="dark"] .view-options-menu {{
      background: #0f172a;
      border-color: #334155;
      box-shadow: 0 14px 32px rgba(2, 6, 23, 0.5);
    }}
    html[data-theme="dark"] .project-filter-action {{
      background: #111827;
      color: #d1d5db;
      border-color: #334155;
    }}
    html[data-theme="dark"] .project-filter-action:hover {{
      background: #1f2937;
    }}
    html[data-theme="dark"] .project-filter-options {{
      border-color: #334155;
      background: #111827;
    }}
    html[data-theme="dark"] .date-chip-apply {{
      background: #0f766e;
      border-color: #0d9488;
      color: #ccfbf1;
    }}
    html[data-theme="dark"] .date-chip-apply:hover {{ background: #115e59; }}
    html[data-theme="dark"] .date-chip-reset {{
      background: #1f2937;
      border-color: #475569;
      color: #d6d3d1;
    }}
    html[data-theme="dark"] .date-chip-reset:hover {{ background: #374151; }}
    html[data-theme="dark"] .adv-filter-btn {{
      background: #4338ca;
      border-color: #6366f1;
      color: #e0e7ff;
    }}
    html[data-theme="dark"] .adv-filter-btn:hover {{ background: #3730a3; }}
    html[data-theme="dark"] .adv-filter-menu {{
      background: #1e1b2e;
      border-color: #3f3a5a;
      box-shadow: 0 12px 36px rgba(0, 0, 0, 0.4);
    }}
    html[data-theme="dark"] .adv-filter-group-label {{ color: #a8a29e; }}
    html[data-theme="dark"] .adv-filter-item {{ color: #d6d3d1; }}
    html[data-theme="dark"] .adv-filter-item:hover {{ background: #164e63; color: #5eead4; }}
    html[data-theme="dark"] .adv-filter-divider {{ background: #3f3a5a; }}
    html[data-theme="dark"] .adv-filter-radio {{ color: #d6d3d1; }}
    html[data-theme="dark"] .adv-filter-radio:hover {{ background: #312e81; }}
    html[data-theme="dark"] .adv-filter-radio input[type="radio"]:checked + span {{ color: #a5b4fc; }}
    html[data-theme="dark"] .date-chip-status {{
      color: #e2e8f0;
    }}
    html[data-theme="dark"] .date-chip-toggle {{
      border-color: #475569;
      background: #1f2937;
      color: #e2e8f0;
    }}
    html[data-theme="dark"] .project-option {{
      color: #d1d5db;
    }}
    html[data-theme="dark"] .project-option:hover {{
      background: #1e293b;
    }}
    html[data-theme="dark"] .search-input:focus {{
      border-color: #60a5fa;
      box-shadow: 0 0 0 2px rgba(96, 165, 250, 0.25);
    }}
    html[data-theme="dark"] .btn.alt {{
      background: #0f172a;
      color: #cbd5e1;
      border-color: #374151;
    }}
    html[data-theme="dark"] .btn.alt.active {{
      background: #1e3a8a;
      border-color: #2563eb;
      color: #dbeafe;
    }}
    html[data-theme="dark"] .btn.alert {{
      background: #450a0a;
      border-color: #b91c1c;
      color: #fecaca;
    }}
    html[data-theme="dark"] .legend-item {{
      background: #0f172a;
      border-color: #334155;
      color: #cbd5e1;
    }}
    html[data-theme="dark"] thead th {{
      border-bottom-color: #334155;
    }}
    html[data-theme="dark"] tbody td {{
      background: #111827;
      border-top-color: #334155;
      color: #d1d5db;
    }}
    html[data-theme="dark"] tbody tr:nth-child(even) td {{ background: #0f172a; }}
    html[data-theme="dark"] .toggle {{
      background: #111827;
      border-color: #475569;
      color: #d1d5db;
    }}
    html[data-theme="dark"] .jira-link {{
      background: #172554;
      border-color: #1d4ed8;
      color: #bfdbfe;
    }}
    html[data-theme="dark"] .jira-link:hover {{ background: #1e3a8a; }}
    html[data-theme="dark"] .assignee-chip {{
      background: #1f2937;
      border-color: #475569;
      color: #d1d5db;
    }}
    html[data-theme="dark"] .danger-chip {{
      background: #450a0a;
      border-color: #b91c1c;
      color: #fecaca;
    }}
    html[data-theme="dark"] tr.row-type-project td {{ background: #102a43; }}
    html[data-theme="dark"] tr.row-type-product td {{ background: #3a330a; }}
    html[data-theme="dark"] tr.row-type-rmi td {{ background: #2e1f4f; }}
    html[data-theme="dark"] tr.row-type-story td {{ background: #1f2d4d; }}
    html[data-theme="dark"] tr.row-type-subtask td {{ background: #0f3a2c; }}
    html[data-theme="dark"] tr.row-type-assignee td {{ background: #1e293b; }}
    html[data-theme="dark"] tr.row-type-bug td {{ background: #3f2a1d; }}
    html[data-theme="dark"] .type-chip.project {{ background: #102a43; color: #93c5fd; border-color: #1d4ed8; }}
    html[data-theme="dark"] .type-chip.product {{ background: #3a330a; color: #fde68a; border-color: #ca8a04; }}
    html[data-theme="dark"] .type-chip.rmi {{ background: #2e1f4f; color: #d8b4fe; border-color: #9333ea; }}
    html[data-theme="dark"] .type-chip.story {{ background: #1f2d4d; color: #bfdbfe; border-color: #3b82f6; }}
    html[data-theme="dark"] .type-chip.subtask {{ background: #0f3a2c; color: #86efac; border-color: #16a34a; }}
    html[data-theme="dark"] .type-chip.bug {{ background: #3f2a1d; color: #fdba74; border-color: #ea580c; }}
    html[data-theme="dark"] .pill {{ background: #3a330a; border-color: #a16207; color: #fde68a; }}
    html[data-theme="dark"] .col-resizer::before {{
      background: rgba(148, 163, 184, 0.35);
    }}
    html[data-theme="dark"] thead th.col-aspect:hover .col-resizer::before {{
      background: rgba(148, 163, 184, 0.6);
    }}
  </style>
  <link rel="stylesheet" href="shared-nav.css">
</head>
<body>
  <div class="top-date-range-wrap">
    <div class="top-date-range-chip" aria-label="Date range filter">
      <span class="date-chip-segment">From</span>
      <input id="date-filter-from" class="date-chip-input" type="date" aria-label="From date">
      <span class="date-chip-segment">To</span>
      <input id="date-filter-to" class="date-chip-input" type="date" aria-label="To date">
      <button id="date-filter-apply" class="date-chip-apply" type="button" aria-label="Apply date range">
        <span class="material-icons-outlined" aria-hidden="true">check_circle</span>
        <span class="date-chip-apply-label">Apply</span>
      </button>
      <button id="date-filter-reset" class="date-chip-reset" type="button" aria-label="Reset date range" title="Reset">
        <span class="material-icons-outlined" aria-hidden="true">restart_alt</span>
      </button>
      <div class="adv-filter-wrap">
        <button id="adv-filter-toggle" class="adv-filter-btn" type="button" aria-expanded="false" aria-haspopup="true" aria-label="Advanced Filters">
          <span class="material-icons-outlined" aria-hidden="true">tune</span>
          <span>Advanced Filters</span>
        </button>
        <div class="adv-filter-menu" id="adv-filter-menu" role="menu" hidden>
          <div class="adv-filter-group-label">Date Presets</div>
          <button class="adv-filter-item" type="button" data-preset="last30" role="menuitem">Last 30 Days</button>
          <button class="adv-filter-item" type="button" data-preset="lastMonth" role="menuitem">Last Month</button>
          <button class="adv-filter-item" type="button" data-preset="currentMonth" role="menuitem">Current Month</button>
          <button class="adv-filter-item" type="button" data-preset="last90" role="menuitem">Last 90 Days</button>
          <button class="adv-filter-item" type="button" data-preset="lastQuarter" role="menuitem">Last Quarter</button>
          <button class="adv-filter-item" type="button" data-preset="currentQuarter" role="menuitem">Current Quarter</button>
        </div>
      </div>
      <span class="date-chip-segment">Planned Hours Source</span>
      <select id="planned-hours-source" class="date-chip-input" aria-label="Planned Hours Source">
        <option value="subtask_estimates">Subtask Estimates</option>
        <option value="subtask_logs">Subtask Logs</option>
        <option value="epic_estimates">Epic Estimates</option>
      </select>
      <label class="date-chip-toggle" for="extended-actual-hours-toggle" title="For subtasks whose planned start/end is in range, include total logged hours across all dates.">
        <input id="extended-actual-hours-toggle" type="checkbox" aria-label="Extended Actual Hours">
        <span>Extended Actual Hours</span>
      </label>
      <select id="actual-hours-mode" style="display:none" aria-hidden="true">
        <option value="log_date">By Log Date</option>
        <option value="planned_dates">By Planned Dates</option>
      </select>
      <span id="date-filter-status" class="date-chip-status" aria-live="polite"></span>
    </div>
  </div>
  <div class="page">
    <div class="capacity-profile-overlay" id="capacity-profile-overlay"></div>
    <section class="header">
      <div class="header-top">
        <h1 class="title">Nested View Report</h1>
        <div class="header-actions">
          <button
            class="btn alt capacity-profile-toggle"
            type="button"
            id="capacity-profile-toggle"
            aria-expanded="false"
            aria-controls="capacity-profile-drawer"
          >
            Show Capacity Profile
          </button>
          <button
            class="btn alt"
            type="button"
            id="header-toggle"
            aria-expanded="true"
            aria-controls="report-header-content"
          >
            Collapse Header
          </button>
        </div>
      </div>
      <p class="meta" id="report-header-content">Generated: <span id="generated-at"></span> | Rows: <span id="row-count"></span></p>
      <h2 class="section-title">Filters And Actions</h2>
      <p class="section-note">Search, project scoping, hierarchy controls, and view preferences.</p>
      <div class="toolbar">
        <div class="search-wrap">
          <input class="search-input" type="text" id="search-input" placeholder="Search any column (aspect, approved days, approved hours, planned days, planned hours, actual hours, actual days, delta hours, delta days, resource logged, planned dates)">
          <span class="search-meta" id="search-meta"></span>
        </div>
        <div class="project-filter" id="project-filter">
          <button
            class="project-filter-btn"
            type="button"
            id="project-filter-toggle"
            aria-haspopup="true"
            aria-expanded="false"
            aria-controls="project-filter-menu"
          >
            <span id="project-filter-summary">Projects: All</span>
            <span class="project-filter-caret" aria-hidden="true">v</span>
          </button>
          <div class="project-filter-progress" id="project-filter-progress" hidden aria-hidden="true">
            <div class="project-filter-progress-bar"></div>
          </div>
          <div class="project-filter-menu" id="project-filter-menu">
            <div class="project-filter-actions">
              <button class="project-filter-action" type="button" id="project-filter-select-all">Select all</button>
              <button class="project-filter-action" type="button" id="project-filter-clear-all">Clear all</button>
            </div>
            <div class="project-filter-options" id="project-filter-options"></div>
          </div>
        </div>
        <div class="project-filter" id="team-filter">
          <button
            class="project-filter-btn"
            type="button"
            id="team-filter-toggle"
            aria-haspopup="true"
            aria-expanded="false"
            aria-controls="team-filter-menu"
          >
            <span id="team-filter-summary">Teams: All</span>
            <span class="project-filter-caret" aria-hidden="true">v</span>
          </button>
          <div class="team-filter-progress" id="team-filter-progress" hidden aria-hidden="true">
            <div class="team-filter-progress-bar"></div>
          </div>
          <div class="project-filter-menu" id="team-filter-menu">
            <div class="project-filter-actions">
              <button class="project-filter-action" type="button" id="team-filter-select-all">Select all</button>
              <button class="project-filter-action" type="button" id="team-filter-clear-all">Clear all</button>
            </div>
            <div class="project-filter-options" id="team-filter-options"></div>
          </div>
        </div>
        <button class="btn" type="button" id="expand-all">Expand All</button>
        <button class="btn alt" type="button" id="collapse-all">Collapse To Projects</button>
        <button class="btn alt" type="button" id="collapse-epics">Collapse to Epics</button>
        <button class="btn alt" type="button" id="clear-search">Clear Search</button>
        <div class="view-options" id="view-options">
          <button
            class="view-options-toggle"
            type="button"
            id="view-options-toggle"
            aria-haspopup="true"
            aria-expanded="false"
            aria-controls="view-options-menu"
          >
            View Options
          </button>
          <div class="view-options-menu" id="view-options-menu">
            <button class="btn alt" type="button" id="theme-toggle"><span class="material-icons-outlined" aria-hidden="true">dark_mode</span>Dark mode</button>
            <button class="btn alt" type="button" id="toggle-density">Compact View</button>
            <button class="btn alt" type="button" id="toggle-no-entry">No Entry &lt;0&gt;</button>
            <button class="btn alt" type="button" id="toggle-product">Show Category</button>
          </div>
        </div>
      </div>
      <h2 class="section-title">Performance Overview</h2>
      <p class="section-note">KPIs react to the selected date range, project filter, and applied capacity profile.</p>
      <div class="scorecards">
        <article class="score-card" id="score-total-capacity-card">
          <span class="score-formula-chip" id="score-total-capacity-formula" aria-label="Employee count multiplied by business days multiplied by per day hours">
            <span class="score-formula-part">
              <span class="material-icons-outlined" aria-hidden="true">person</span>
              <span id="score-total-capacity-formula-employee">-</span>
            </span>
            <span class="score-formula-mult">x</span>
            <span class="score-formula-part">
              <span class="material-icons-outlined" aria-hidden="true">calendar_month</span>
              <span id="score-total-capacity-formula-days">-</span>
            </span>
            <span class="score-formula-mult">x</span>
            <span class="score-formula-part">
              <span class="material-icons-outlined" aria-hidden="true">hourglass_top</span>
              <span id="score-total-capacity-formula-hours">-</span>
            </span>
          </span>
          <p class="score-label">
            Total Capacity (Hours)
            <span class="score-info" tabindex="0" aria-label="Total Capacity information">
              i
              <span class="score-info-tip" id="score-total-capacity-tip">Formula: Total Capacity (Hours) = Employee Count x Available Business Days x Per Day Hours.
Values:
Selected Profile = None
Date Range = -
Employee Count = 0
Per Day Hours = 0h
Standard Hours/Day = 0h
Ramadan Hours/Day = 0h
Working Weekdays (Mon-Fri) = 0
Ramadan Weekdays (Mon-Fri, non-holiday) = 0
Non-Ramadan Weekdays (Mon-Fri, non-holiday) = 0
Holiday Weekdays in Range = 0
Capacity Profile Hours = N/A
Fallback Project Capacity = 0h
Total Capacity = 0h</span>
            </span>
            <button
              type="button"
              class="score-action-btn"
              id="score-capacity-profile-open"
              title="Load capacity profile"
              aria-label="Load capacity profile"
            >
              <span class="material-icons-outlined" aria-hidden="true">tune</span>
            </button>
          </p>
          <p class="score-value" id="score-total-capacity">0h</p>
        </article>
        <article class="score-card" id="score-total-planned-card">
          <p class="score-label">
            Total Planned Hours
            <span class="score-info" tabindex="0" aria-label="Total Planned Hours information">
              i
              <span class="score-info-tip" id="score-total-planned-tip">Formula: Total Planned Hours = Sum(Subtask Planned Hours) where subtask Start OR Due date is within selected range, excluding RLT (RnD Leave Tracker).
Values:
Included Subtasks Count = 0
Excluded (RLT) Subtasks Count = 0
Excluded (RLT) Subtasks Planned Sum = 0h
Total Planned Hours = 0h</span>
            </span>
          </p>
          <p class="score-value" id="score-total-planned">0h</p>
          <section class="score-details-panel" id="score-total-planned-details" hidden>
            <h3 class="score-details-title">Total Planned Hours - Subtask Breakdown</h3>
            <p class="score-details-meta" id="score-total-planned-details-meta">No rows.</p>
            <div class="score-details-table-wrap">
              <table class="score-details-table" aria-label="Total Planned Hours details table">
                <thead>
                  <tr>
                    <th>Project Name</th>
                    <th>Subtask Jira ID</th>
                    <th>Task Type</th>
                    <th>Subtask Jira Name</th>
                    <th>Subtask Start Date</th>
                    <th>Subtask Due Date</th>
                    <th>Planned Hours</th>
                  </tr>
                </thead>
                <tbody id="score-total-planned-details-body"></tbody>
              </table>
            </div>
          </section>
        </article>
        <article class="score-card" id="score-total-leaves-planned-card">
          <span class="score-formula-chip" id="score-total-leaves-planned-formula">Planned Taken + Planned Not Yet Taken</span>
          <p class="score-label">
            Total Leaves Planned
            <span class="score-info" tabindex="0" aria-label="Total Leaves Planned information">
              i
              <span class="score-info-tip" id="score-total-leaves-planned-tip">Formula: Total Leaves Planned = Planned Taken + Planned Not Yet Taken from day-bucketed leave rows in selected date range.
Values:
Capacity Profile Effect = None (independent of selected profile)
Planned Taken + Planned Not Yet Taken = 0h
Total Leaves Planned = 0h</span>
            </span>
          </p>
          <p class="score-value" id="score-total-leaves-planned">0h</p>
          <section class="score-details-panel" id="score-total-leaves-planned-details" hidden>
            <h3 class="score-details-title">Total Leaves Planned - Assignee Summary</h3>
            <p class="score-details-meta" id="score-total-leaves-planned-assignee-summary-meta">No rows.</p>
            <div class="score-details-table-wrap">
              <table class="score-details-table" aria-label="Total Leaves Planned assignee summary table">
                <thead>
                  <tr>
                    <th>Assignee Name</th>
                    <th>Planned Leaves (h)</th>
                  </tr>
                </thead>
                <tbody id="score-total-leaves-planned-assignee-summary-body"></tbody>
              </table>
            </div>
            <h3 class="score-details-title" style="margin-top:12px;">Total Leaves Planned - Subtask Breakdown</h3>
            <p class="score-details-meta" id="score-total-leaves-planned-details-meta">No rows.</p>
            <div class="score-details-table-wrap">
              <table class="score-details-table" aria-label="Total Leaves Planned details table">
                <thead>
                  <tr>
                    <th>Date</th>
                    <th>Assignee Name</th>
                    <th>Hours</th>
                    <th>Subtask Jira ID</th>
                    <th>Jira</th>
                  </tr>
                </thead>
                <tbody id="score-total-leaves-planned-details-body"></tbody>
              </table>
            </div>
          </section>
        </article>
        <article class="score-card" id="score-total-logged-card">
          <p class="score-label">
            Total Actual Hours
            <span class="score-info" tabindex="0" aria-label="Total Actual Hours information">
              i
              <span class="score-info-tip" id="score-total-logged-tip">Formula: Total Actual Hours = Sum(Project Actual Hours), excluding RLT (RnD Leave Tracker).
Values:
Included Projects Actual Sum = 0h
Excluded Projects Actual Sum = 0h
Total Actual Hours = 0h</span>
            </span>
          </p>
          <p class="score-value" id="score-total-logged">0h</p>
          <section class="score-details-panel" id="score-total-logged-details" hidden>
            <h3 class="score-details-title">Total Actual Hours - Subtask Breakdown</h3>
            <p class="score-details-meta">
              <label><input type="checkbox" id="score-total-logged-include-bugs"> Include Bugs</label>
            </p>
            <p class="score-details-meta" id="score-total-logged-details-meta">No rows.</p>
            <div class="score-details-table-wrap">
              <table class="score-details-table" aria-label="Total Actual Hours details table">
                <thead>
                  <tr>
                    <th>Project Name</th>
                    <th>Subtask Jira ID</th>
                    <th>Task Type</th>
                    <th>Subtask Jira Name</th>
                    <th>Subtask Start Date</th>
                    <th>Subtask Due Date</th>
                    <th>Actual Hours</th>
                  </tr>
                </thead>
                <tbody id="score-total-logged-details-body"></tbody>
              </table>
            </div>
          </section>
        </article>
        <article class="score-card" id="score-total-capacity-planned-leaves-adjusted-card">
          <span class="score-formula-chip" id="score-availability-formula">Total Capacity - Total Leaves Planned</span>
          <p class="score-label">
            Availability
            <span class="score-info" tabindex="0" aria-label="Availability information">
              i
              <span class="score-info-tip" id="score-total-capacity-planned-leaves-adjusted-tip">Formula: Availability = Total Capacity (Hours) - Total Leaves Planned.
Values:
Total Capacity = 0h
Total Leaves Planned = 0h
Availability = 0h</span>
            </span>
          </p>
          <p class="score-value" id="score-total-capacity-planned-leaves-adjusted">0h</p>
        </article>
        <div class="efficiency-inline-wrap" id="efficiency-under-actual">
          <div class="efficiency-inline-item">
            <p class="efficiency-inline-label">
              Planning Efficiency
              <span class="score-info" tabindex="0" aria-label="Planning Efficiency information">
                i
                <span class="score-info-tip" id="score-loading-efficiency-tip">Formula: Planning Efficiency = Total Planned Hours / Availability x 100.
Values:
Total Planned Hours = 0h
Availability = 0h
Planning Efficiency = 0%</span>
              </span>
            </p>
            <p class="efficiency-inline-value" id="score-loading-efficiency">0%</p>
            <p class="efficiency-inline-formula">Formula: Total Planned Hours / Availability x 100</p>
          </div>
          <!-- Delivery Efficiency scorecard commented out
          <div class="efficiency-inline-item">
            <p class="efficiency-inline-label">
              Delivery Efficiency
              <span class="score-info" tabindex="0" aria-label="Delivery Efficiency information">
                i
                <span class="score-info-tip" id="score-delivery-efficiency-tip">Formula: Delivery Efficiency = Total Planned Hours / Total Actual Hours x 100.
Values:
Total Planned Hours = 0h
Total Actual Hours = 0h
Delivery Efficiency = 0%</span>
              </span>
            </p>
            <p class="efficiency-inline-value" id="score-delivery-efficiency">0%</p>
            <p class="efficiency-inline-formula">Formula: Total Planned Hours / Total Actual Hours x 100</p>
          </div>
          -->
        </div>
        <!--
        <article class="score-card" id="score-total-leaves-card">
          <p class="score-label">
            Total Leaves Taken
            <span class="score-info" tabindex="0" aria-label="Total Leaves Taken information">
              i
              <span class="score-info-tip" id="score-total-leaves-tip">Formula: Total Leaves Taken = Sum(Logged Hours for RLT RnD Leave Tracker leave work).
Values:
Planned Leaves Taken = 0h
Unplanned Leaves Taken = 0h
Total Leaves Taken = 0h</span>
            </span>
          </p>
          <p class="score-value" id="score-total-leaves">0h</p>
        </article>
        </article>
        -->
      </div>
      <div class="capacity-profile-drawer" id="capacity-profile-drawer" role="dialog" aria-modal="true" aria-label="Capacity profile">
        <div class="capacity-profile-resize-handle" id="capacity-profile-resize-handle" role="separator" aria-orientation="vertical" aria-label="Resize capacity profile drawer"></div>
        <div class="capacity-profile-drawer-head">
          <h2 class="capacity-profile-drawer-title">Capacity Profile</h2>
          <button class="capacity-profile-close" type="button" id="capacity-profile-close" aria-label="Close capacity profile">
            <span class="material-icons-outlined" aria-hidden="true">close</span>
          </button>
        </div>
        <p class="section-note">Apply a saved profile to this page range. Capacity profile management is available on the dedicated settings page.</p>
        <div class="capacity-profile-bar">
          <label class="capacity-profile-label" for="capacity-profile-select">Saved Capacity Profiles</label>
          <select id="capacity-profile-select" class="capacity-profile-select"></select>
          <button class="btn alt" type="button" id="capacity-profile-apply">Apply</button>
          <button class="btn alt" type="button" id="capacity-profile-refresh">Refresh</button>
          <button class="btn alt" type="button" id="capacity-profile-reset">Use Project Totals</button>
          <a class="btn alt" href="/settings/capacity">Manage Capacity Profiles</a>
          <div class="capacity-profile-status" id="capacity-profile-status" data-variant="info"></div>
          <div class="capacity-profile-details" id="capacity-profile-details"></div>
          <div class="capacity-expanded" id="capacity-profile-expanded"></div>
        </div>
      </div>
    </section>
    <section class="table-wrap">
      <h2 class="table-section-title">Timeless Breakdown of Work</h2>
      <div class="legend">
        <span class="legend-title">Legend</span>
        <span class="legend-item"><span class="legend-swatch" style="background:#dbeeff;border-color:#91bde7"></span>Project</span>
        <span class="legend-item"><span class="legend-swatch" style="background:#f2f7a8;border-color:#d0d66b"></span>Category</span>
        <span class="legend-item"><span class="legend-swatch" style="background:#eadfff;border-color:#b39be5"></span>Epic</span>
        <span class="legend-item"><span class="legend-swatch" style="background:#dfeaff;border-color:#97b6ee"></span>Story</span>
        <span class="legend-item"><span class="legend-swatch" style="background:#ecfde9;border-color:#b0dbab"></span>Subtask</span>
        <span class="legend-item"><span class="danger-chip" style="margin-left:0">No entry</span>Missing value warning</span>
      </div>
      <table>
        <caption class="section-note" style="caption-side:top;text-align:left;padding:8px 10px;">
          Hierarchical work breakdown with approved, planned, and actual effort plus schedule context.
        </caption>
        <thead>
          <tr>
            <th class="col-aspect" title="Hierarchy node label (project/category/epic/story/subtask).">Aspect<span class="col-resizer" id="aspect-col-resizer" role="separator" aria-orientation="vertical" aria-label="Resize Aspect column"></span></th>
            <th class="col-type" title="Work item type with semantic color coding.">Type</th>
            <th class="col-assignee" title="Assignee name where available at this level.">Assignee</th>
            <th class="num" title="Approved effort in days.">Approved Days</th>
            <th class="num" title="Rolled-up subtask original estimates in days.">Planned Days</th>
            <th class="num" title="Actual logged effort converted to days.">Actual Days</th>
            <th class="num" title="Approved days minus actual days.">Delta Days</th>
            <th class="num" title="Approved effort in hours.">Approved Hours</th>
            <th class="num" title="Rolled-up subtask original estimates in hours.">Planned Hours</th>
            <th class="num" title="Actual logged effort in hours.">Actual Hours</th>
            <th class="num" title="Approved hours minus actual hours.">Delta Hours</th>
            <th class="col-resource" title="Distinct contributors who logged hours.">Resource Logged</th>
            <th class="col-date" title="Planned start date from source.">Planned Start Date</th>
            <th class="col-date" title="Planned end date from source.">Planned End Date</th>
          </tr>
        </thead>
        <tbody id="rows"></tbody>
      </table>
    </section>
  </div>
  <script>
    const reportData = {payload};
    const tbody = document.getElementById("rows");
    document.getElementById("generated-at").textContent = reportData.generated_at || "-";
    const rowCountNode = document.getElementById("row-count");

    let allRows = reportData.rows || [];
    const leaveDailyRows = Array.isArray(reportData.leave_daily_rows) ? reportData.leave_daily_rows : [];
    const leaveSubtaskRows = Array.isArray(reportData.leave_subtask_rows) ? reportData.leave_subtask_rows : [];
    const jiraBaseUrl = String(reportData.jira_base_url || "").trim().replace(/[/]+$/, "");
    const rowsById = new Map();
    const childrenByParent = new Map();
    const collapsed = new Set();
    const searchInput = document.getElementById("search-input");
    const clearSearchButton = document.getElementById("clear-search");
    const themeToggleButton = document.getElementById("theme-toggle");
    const densityToggleButton = document.getElementById("toggle-density");
    const toggleProductButton = document.getElementById("toggle-product");
    const toggleNoEntryButton = document.getElementById("toggle-no-entry");
    const collapseEpicsButton = document.getElementById("collapse-epics");
    const searchMeta = document.getElementById("search-meta");
    const dateFilterFromInput = document.getElementById("date-filter-from");
    const dateFilterToInput = document.getElementById("date-filter-to");
    const plannedHoursSourceSelect = document.getElementById("planned-hours-source");
    const extendedActualHoursToggle = document.getElementById("extended-actual-hours-toggle");
    const actualHoursModeSelect = document.getElementById("actual-hours-mode");
    const dateFilterApplyButton = document.getElementById("date-filter-apply");
    const dateFilterResetButton = document.getElementById("date-filter-reset");
    const dateFilterStatusNode = document.getElementById("date-filter-status");
    const projectFilterRoot = document.getElementById("project-filter");
    const projectFilterToggle = document.getElementById("project-filter-toggle");
    const projectFilterMenu = document.getElementById("project-filter-menu");
    const projectFilterOptions = document.getElementById("project-filter-options");
    const projectFilterSummary = document.getElementById("project-filter-summary");
    const projectFilterProgress = document.getElementById("project-filter-progress");
    const projectFilterSelectAll = document.getElementById("project-filter-select-all");
    const projectFilterClearAll = document.getElementById("project-filter-clear-all");
    const teamFilterRoot = document.getElementById("team-filter");
    const teamFilterToggle = document.getElementById("team-filter-toggle");
    const teamFilterMenu = document.getElementById("team-filter-menu");
    const teamFilterOptions = document.getElementById("team-filter-options");
    const teamFilterSummary = document.getElementById("team-filter-summary");
    const teamFilterProgress = document.getElementById("team-filter-progress");
    const teamFilterSelectAll = document.getElementById("team-filter-select-all");
    const teamFilterClearAll = document.getElementById("team-filter-clear-all");
    const viewOptionsRoot = document.getElementById("view-options");
    const viewOptionsToggle = document.getElementById("view-options-toggle");
    const viewOptionsMenu = document.getElementById("view-options-menu");
    const tableWrapEl = document.querySelector(".table-wrap");
    const tableHeadEl = document.querySelector("thead");
    const totalCapacityScoreNode = document.getElementById("score-total-capacity");
    const totalCapacityFormulaEmployeeNode = document.getElementById("score-total-capacity-formula-employee");
    const totalCapacityFormulaDaysNode = document.getElementById("score-total-capacity-formula-days");
    const totalCapacityFormulaHoursNode = document.getElementById("score-total-capacity-formula-hours");
    const totalPlannedScoreNode = document.getElementById("score-total-planned");
    const totalPlannedCardEl = document.getElementById("score-total-planned-card");
    const totalPlannedDetailsEl = document.getElementById("score-total-planned-details");
    const totalPlannedDetailsMetaEl = document.getElementById("score-total-planned-details-meta");
    const totalPlannedDetailsBodyEl = document.getElementById("score-total-planned-details-body");
    const totalLoggedScoreNode = document.getElementById("score-total-logged");
    const totalLoggedCardEl = document.getElementById("score-total-logged-card");
    const totalLoggedDetailsEl = document.getElementById("score-total-logged-details");
    const totalLoggedDetailsMetaEl = document.getElementById("score-total-logged-details-meta");
    const totalLoggedDetailsBodyEl = document.getElementById("score-total-logged-details-body");
    const totalLoggedIncludeBugsEl = document.getElementById("score-total-logged-include-bugs");
    const deltaScoreNode = document.getElementById("score-delta");
    const totalLeavesScoreNode = document.getElementById("score-total-leaves");
    const totalLeavesPlannedScoreNode = document.getElementById("score-total-leaves-planned");
    const totalLeavesPlannedCardEl = document.getElementById("score-total-leaves-planned-card");
    const totalLeavesPlannedDetailsEl = document.getElementById("score-total-leaves-planned-details");
    const totalLeavesPlannedDetailsMetaEl = document.getElementById("score-total-leaves-planned-details-meta");
    const totalLeavesPlannedDetailsBodyEl = document.getElementById("score-total-leaves-planned-details-body");
    const totalLeavesPlannedAssigneeSummaryMetaEl = document.getElementById("score-total-leaves-planned-assignee-summary-meta");
    const totalLeavesPlannedAssigneeSummaryBodyEl = document.getElementById("score-total-leaves-planned-assignee-summary-body");
    const totalCapacityPlannedLeavesAdjustedScoreNode = document.getElementById("score-total-capacity-planned-leaves-adjusted");
    const loadingEfficiencyScoreNode = document.getElementById("score-loading-efficiency");
    const deliveryEfficiencyScoreNode = document.getElementById("score-delivery-efficiency");
    const capacityGapScoreNode = document.getElementById("score-capacity-gap");
    const deltaScoreCard = document.getElementById("score-delta-card");
    const totalCapacityTipNode = document.getElementById("score-total-capacity-tip");
    const totalPlannedTipNode = document.getElementById("score-total-planned-tip");
    const totalLoggedTipNode = document.getElementById("score-total-logged-tip");
    const deltaTipNode = document.getElementById("score-delta-tip");
    const totalLeavesTipNode = document.getElementById("score-total-leaves-tip");
    const totalLeavesPlannedTipNode = document.getElementById("score-total-leaves-planned-tip");
    const totalCapacityPlannedLeavesAdjustedTipNode = document.getElementById("score-total-capacity-planned-leaves-adjusted-tip");
    const loadingEfficiencyTipNode = document.getElementById("score-loading-efficiency-tip");
    const deliveryEfficiencyTipNode = document.getElementById("score-delivery-efficiency-tip");
    const capacityGapTipNode = document.getElementById("score-capacity-gap-tip");
    const availabilityFormulaNode = document.getElementById("score-availability-formula");
    const scoreCapacityProfileOpenButton = document.getElementById("score-capacity-profile-open");
    const headerSectionEl = document.querySelector("section.header");
    const headerToggleButton = document.getElementById("header-toggle");
    const capacityProfileSelectEl = document.getElementById("capacity-profile-select");
    const capacityProfileApplyEl = document.getElementById("capacity-profile-apply");
    const capacityProfileRefreshEl = document.getElementById("capacity-profile-refresh");
    const capacityProfileResetEl = document.getElementById("capacity-profile-reset");
    const capacityProfileToggleEl = document.getElementById("capacity-profile-toggle");
    const capacityProfileDrawerEl = document.getElementById("capacity-profile-drawer");
    const capacityProfileOverlayEl = document.getElementById("capacity-profile-overlay");
    const capacityProfileCloseEl = document.getElementById("capacity-profile-close");
    const capacityProfileResizeHandleEl = document.getElementById("capacity-profile-resize-handle");
    const capacityProfileStatusEl = document.getElementById("capacity-profile-status");
    const capacityProfileDetailsEl = document.getElementById("capacity-profile-details");
    const capacityProfileExpandedEl = document.getElementById("capacity-profile-expanded");
    const DATE_FILTER_WORK_TYPES = new Set(["rmi"]);
    const LEAF_WORK_ROW_TYPES = new Set(["subtask"]);
    const ACTUAL_HOURS_MODE_STORAGE_KEY = "actual-hours-mode:nested-view";
    const PLANNED_HOURS_SOURCE_STORAGE_KEY = "planned-hours-source:nested-view";
    const DEFAULT_ACTUAL_HOURS_MODE = "log_date";
    const DEFAULT_PLANNED_HOURS_SOURCE = "subtask_estimates";
    const DEFAULT_DATE_FROM = "2026-01-01";
    const DEFAULT_DATE_TO = new Date().toISOString().slice(0, 10);
    let selectedDateFrom = DEFAULT_DATE_FROM;
    let selectedDateTo = DEFAULT_DATE_TO;
    let pendingDateFrom = DEFAULT_DATE_FROM;
    let pendingDateTo = DEFAULT_DATE_TO;
    let selectedActualHoursMode = DEFAULT_ACTUAL_HOURS_MODE;
    let pendingActualHoursMode = DEFAULT_ACTUAL_HOURS_MODE;
    let selectedPlannedHoursSource = DEFAULT_PLANNED_HOURS_SOURCE;
    let pendingPlannedHoursSource = DEFAULT_PLANNED_HOURS_SOURCE;
    let isApplyingDateRange = false;
    let activeSearchQuery = "";
    let showProductCategorization = false;
    let onlyNoEntry = false;
    const THEME_STORAGE_KEY = "rmi-nested-report-theme";
    const DENSITY_STORAGE_KEY = "rmi-nested-report-density";
    const HEADER_COLLAPSED_STORAGE_KEY = "rmi-nested-report-header-collapsed";
    const CAPACITY_DRAWER_WIDTH_STORAGE_KEY = "nested-view:capacity-drawer-width";
    const ASPECT_COL_WIDTH_STORAGE_KEY = "rmi-nested-report-aspect-width";
    const MIN_ASPECT_COL_WIDTH = 240;
    const MAX_ASPECT_COL_WIDTH = 900;
    const DEFAULT_CAPACITY_DRAWER_WIDTH_VW = 50;
    const MIN_CAPACITY_DRAWER_WIDTH_VW = 35;
    const MAX_CAPACITY_DRAWER_WIDTH_VW = 85;
    const WORK_NO_ENTRY_TYPES = new Set(["rmi", "story", "subtask"]);
    const allProjects = [];
    const selectedProjectKeys = new Set();
    let scorecardSourceRows = [];
    let capacityProfiles = Array.isArray(reportData.capacity_profiles) ? reportData.capacity_profiles.slice() : [];
    let appliedCapacityProfile = null;
    let appliedCapacityProfileKey = "";
    const CAPACITY_PROFILES_ENDPOINT = "/api/capacity/profiles";
    const hasCapacityApi = window.location.protocol !== "file:";
    const MANAGED_FIELDS_ENDPOINT = "/api/manage-fields?include_inactive=0";
    const hasManagedFieldsApi = window.location.protocol !== "file:";
    const ACTUAL_HOURS_AGGREGATE_ENDPOINT = "/api/actual-hours/aggregate";
    const NESTED_ACTUALS_ENDPOINT = "/api/nested-view/actual-hours";
    const SCOPED_SUBTASKS_ENDPOINT = "/api/scoped-subtasks";
    const APPROVED_PLANNED_SUMMARY_ENDPOINT = "/api/approved-vs-planned-hours/summary";
    const APPROVED_PLANNED_DETAILS_ENDPOINT = "/api/approved-vs-planned-hours/details";
    const PERFORMANCE_TEAMS_ENDPOINT = "/api/performance/teams";
    const hasNestedActualsApi = window.location.protocol !== "file:";
    const hasNestedTreeApi = window.location.protocol !== "file:";
    const hasPlannedParityApi = window.location.protocol !== "file:";
    const hasTeamsApi = window.location.protocol !== "file:";
    const originalMetricsById = new Map();
    let managedFieldsByKey = new Map();
    const allTeams = [];
    const selectedTeamNames = new Set();
    const selectedTeamAssignees = new Set();
    let totalPlannedDetailsOpen = false;
    let totalPlannedDetailsGroups = [];
    const collapsedTotalPlannedProjectKeys = new Set();
    let totalActualDetailsOpen = false;
    let totalActualDetailsGroups = [];
    const collapsedTotalActualProjectKeys = new Set();
    let totalActualIncludeBugs = false;
    let totalLeavesPlannedDetailsOpen = false;
    let totalLeavesPlannedDetailsRows = [];
    let subtaskLogHoursByIssue = {{}};
    const plannedHoursInRangeParityCache = new Map();
    const plannedHoursDetailsParityCache = new Map();
    let scorecardUpdateVersion = 0;
    let projectFilterLoadingVersion = 0;
    let teamFilterLoadingVersion = 0;
    let capacityDrawerPointerId = null;

    function toFiniteNumber(value, fallback = 0) {{
      const n = Number(value);
      return Number.isFinite(n) ? n : fallback;
    }}

    function roundHours(value) {{
      return Math.round(toFiniteNumber(value, 0) * 100) / 100;
    }}

    function normalizePlannedHoursSource(value) {{
      const raw = String(value || "").trim().toLowerCase();
      if (raw === "subtask_logs") return "subtask_logs";
      if (raw === "epic_estimates") return "epic_estimates";
      return "subtask_estimates";
    }}

    function resolveScopedSubtaskBasis(plannedHoursSource) {{
      return normalizePlannedHoursSource(plannedHoursSource || selectedPlannedHoursSource || DEFAULT_PLANNED_HOURS_SOURCE) === "subtask_logs"
        ? "log_date"
        : "planned_dates";
    }}

    function buildPlannedHoursInRangeParityKey() {{
      return buildPlannedHoursInRangeParityKeyFor({{
        dateFrom: selectedDateFrom,
        dateTo: selectedDateTo,
        plannedHoursSource: selectedPlannedHoursSource,
      }});
    }}

    function buildPlannedHoursInRangeParityKeyFor(selection) {{
      const snapshot = selection && typeof selection === "object" ? selection : {{}};
      const projectScope = Array.from(selectedProjectKeys).sort().join(",");
      const assigneeScope = (!isAllTeamsSelected() && selectedTeamAssignees.size)
        ? Array.from(selectedTeamAssignees).sort().join(",")
        : "";
      return [
        String(snapshot.dateFrom || selectedDateFrom || ""),
        String(snapshot.dateTo || selectedDateTo || ""),
        String(snapshot.plannedHoursSource || selectedPlannedHoursSource || ""),
        projectScope,
        assigneeScope,
      ].join("|");
    }}

    async function fetchPlannedHoursInRangeParityTotal(selection) {{
      const snapshot = selection && typeof selection === "object" ? selection : {{}};
      const scorecardDateFrom = String(snapshot.dateFrom || selectedDateFrom || "");
      const scorecardDateTo = String(snapshot.dateTo || selectedDateTo || "");
      const scorecardPlannedHoursSource = normalizePlannedHoursSource(
        snapshot.plannedHoursSource || selectedPlannedHoursSource
      );
      if (!hasPlannedParityApi) {{
        return null;
      }}
      if (scorecardPlannedHoursSource !== "subtask_estimates") {{
        return null;
      }}
      if (selectedProjectKeys.size === 0) {{
        return 0;
      }}
      if (!isAllTeamsSelected() && selectedTeamAssignees.size === 0) {{
        return 0;
      }}
      const parityKey = buildPlannedHoursInRangeParityKeyFor({{
        dateFrom: scorecardDateFrom,
        dateTo: scorecardDateTo,
        plannedHoursSource: scorecardPlannedHoursSource,
      }});
      if (plannedHoursInRangeParityCache.has(parityKey)) {{
        return plannedHoursInRangeParityCache.get(parityKey);
      }}
      const params = new URLSearchParams();
      params.set("from", scorecardDateFrom);
      params.set("to", scorecardDateTo);
      params.set("mode", "planned_dates");
      params.set("plan_source", "jira_estimates");
      params.set("projects", Array.from(selectedProjectKeys).sort().join(","));
      if (!isAllTeamsSelected() && selectedTeamAssignees.size) {{
        params.set("assignees", Array.from(selectedTeamAssignees).sort().join(","));
      }}
      const response = await fetch(
        APPROVED_PLANNED_SUMMARY_ENDPOINT + "?" + params.toString(),
        {{
          method: "GET",
        }}
      );
      const payload = await response.json().catch(() => ({{}}));
      if (!response.ok || !payload || payload.ok === false) {{
        throw new Error(String(payload && payload.error || "Failed to fetch planned-hours parity total."));
      }}
      const rows = Array.isArray(payload && payload.rows) ? payload.rows : [];
      const total = roundHours(rows.reduce((sum, item) => (
        sum + toFiniteNumber(item && item.dispensed_in_range_hours, 0)
      ), 0));
      plannedHoursInRangeParityCache.set(parityKey, total);
      return total;
    }}

    async function fetchPlannedHoursParityDetailsGroups(selection) {{
      const snapshot = selection && typeof selection === "object" ? selection : {{}};
      const scorecardDateFrom = String(snapshot.dateFrom || selectedDateFrom || "");
      const scorecardDateTo = String(snapshot.dateTo || selectedDateTo || "");
      const scorecardPlannedHoursSource = normalizePlannedHoursSource(
        snapshot.plannedHoursSource || selectedPlannedHoursSource
      );
      if (!hasPlannedParityApi) {{
        return null;
      }}
      if (scorecardPlannedHoursSource !== "subtask_estimates") {{
        return null;
      }}
      if (selectedProjectKeys.size === 0) {{
        return [];
      }}
      if (!isAllTeamsSelected() && selectedTeamAssignees.size === 0) {{
        return [];
      }}
      const parityKey = buildPlannedHoursInRangeParityKeyFor({{
        dateFrom: scorecardDateFrom,
        dateTo: scorecardDateTo,
        plannedHoursSource: scorecardPlannedHoursSource,
      }});
      if (plannedHoursDetailsParityCache.has(parityKey)) {{
        return plannedHoursDetailsParityCache.get(parityKey);
      }}
      const projectKeys = Array.from(selectedProjectKeys).sort();
      const projectScope = projectKeys.join(",");
      const groups = await Promise.all(projectKeys.map(async (projectKey) => {{
        const params = new URLSearchParams();
        params.set("from", scorecardDateFrom);
        params.set("to", scorecardDateTo);
        params.set("mode", "planned_dates");
        params.set("plan_source", "jira_estimates");
        params.set("projects", projectScope);
        params.set("project_key", projectKey);
        if (!isAllTeamsSelected() && selectedTeamAssignees.size) {{
          params.set("assignees", Array.from(selectedTeamAssignees).sort().join(","));
        }}
        const response = await fetch(
          APPROVED_PLANNED_DETAILS_ENDPOINT + "?" + params.toString(),
          {{
            method: "GET",
          }}
        );
        const payload = await response.json().catch(() => ({{}}));
        if (!response.ok || !payload || payload.ok === false) {{
          throw new Error(String(payload && payload.error || ("Failed to fetch planned-hours details for " + projectKey + ".")));
        }}
        const epics = Array.isArray(payload && payload.epics) ? payload.epics : [];
        return {{
          project_key: String(payload && payload.project_key || projectKey || "").trim().toUpperCase(),
          project_name: String(payload && payload.project_name || projectKey || "").trim(),
          epics: epics.map((epic) => ({{
            epic_jira_id: String(epic && epic.issue_key || "").trim().toUpperCase(),
            epic_jira_name: String(epic && epic.summary || "").trim(),
            epic_start_date: String(epic && epic.planned_start || "").trim(),
            epic_due_date: String(epic && epic.planned_due || "").trim(),
            planned_hours: roundHours(epic && epic.dispensed_in_range_hours),
          }})).filter((epic) => (
            epic.epic_jira_id || epic.epic_jira_name || toFiniteNumber(epic.planned_hours, 0) > 0
          )),
        }};
      }}));
      plannedHoursDetailsParityCache.set(parityKey, groups);
      return groups;
    }}

    function extractSubtaskHoursMap(payload) {{
      const source = payload && payload.subtask_hours_by_issue && typeof payload.subtask_hours_by_issue === "object"
        ? payload.subtask_hours_by_issue
        : {{}};
      const out = {{}};
      for (const [key, value] of Object.entries(source)) {{
        const issueKey = String(key || "").trim().toUpperCase();
        if (!issueKey) continue;
        out[issueKey] = toFiniteNumber(value, 0);
      }}
      return out;
    }}

    function toHoursToDays(hoursValue) {{
      return roundHours(toFiniteNumber(hoursValue, 0) / 8);
    }}

    function setDateFilterStatus(text) {{
      if (!dateFilterStatusNode) {{
        return;
      }}
      dateFilterStatusNode.textContent = String(text || "");
    }}

    function setDateApplyBusy(busy) {{
      isApplyingDateRange = !!busy;
      updateDateRangeApplyState();
      var applyLabel = dateFilterApplyButton && dateFilterApplyButton.querySelector(".date-chip-apply-label");
      if (applyLabel) {{
        applyLabel.textContent = busy ? "Applying\u2026" : "Apply";
      }}
    }}

    function formatHours(value) {{
      const n = toFiniteNumber(value, 0);
      const rounded = Math.round(n * 100) / 100;
      const text = Number.isInteger(rounded)
        ? String(rounded)
        : rounded.toFixed(2).replace(/0+$/, "").replace(/\\.$/, "");
      return text + "h";
    }}

    function formatHoursPlain(value) {{
      const n = toFiniteNumber(value, 0);
      const rounded = Math.round(n * 100) / 100;
      if (Number.isInteger(rounded)) {{
        return String(rounded);
      }}
      return rounded.toFixed(2).replace(/0+$/, "").replace(/\\.$/, "");
    }}

    function formatPercent(value) {{
      const n = toFiniteNumber(value, 0);
      const rounded = Math.round(n * 100) / 100;
      if (Number.isInteger(rounded)) {{
        return String(rounded) + "%";
      }}
      return rounded.toFixed(2).replace(/0+$/, "").replace(/\\.$/, "") + "%";
    }}

    function setTotalPlannedDetailsOpen(open) {{
      totalPlannedDetailsOpen = !!open;
      if (totalPlannedDetailsEl) {{
        totalPlannedDetailsEl.hidden = !totalPlannedDetailsOpen;
      }}
      if (totalPlannedCardEl) {{
        totalPlannedCardEl.classList.toggle("is-expanded", totalPlannedDetailsOpen);
        totalPlannedCardEl.setAttribute("aria-expanded", totalPlannedDetailsOpen ? "true" : "false");
      }}
    }}

    function renderTotalPlannedDetails() {{
      if (!totalPlannedDetailsBodyEl || !totalPlannedDetailsMetaEl) {{
        return;
      }}
      totalPlannedDetailsBodyEl.innerHTML = "";
      const groups = Array.isArray(totalPlannedDetailsGroups) ? totalPlannedDetailsGroups : [];
      let totalSubtasks = 0;
      let totalHours = 0;
      for (const group of groups) {{
        const subtasks = Array.isArray(group && group.subtasks) ? group.subtasks : [];
        totalSubtasks += subtasks.length;
        for (const subtask of subtasks) {{
          totalHours += toFiniteNumber(subtask && subtask.planned_hours, 0);
        }}
      }}
      totalPlannedDetailsMetaEl.textContent =
        "Projects: " + String(groups.length) + " | Subtasks: " + String(totalSubtasks) + " | Planned Hours: " + formatHours(totalHours);
      if (!groups.length) {{
        const tr = document.createElement("tr");
        const td = document.createElement("td");
        td.colSpan = 7;
        td.className = "score-details-empty";
        td.textContent = "No planned subtask rows found for current date/team filters.";
        tr.appendChild(td);
        totalPlannedDetailsBodyEl.appendChild(tr);
        return;
      }}
      for (const group of groups) {{
        const projectKey = String(group && group.project_key || "");
        const projectName = String(group && group.project_name || projectKey || "Unknown Project");
        const subtasks = Array.isArray(group && group.subtasks) ? group.subtasks : [];
        const groupHours = subtasks.reduce((sum, subtask) => sum + toFiniteNumber(subtask && subtask.planned_hours, 0), 0);
        const projectTr = document.createElement("tr");
        projectTr.className = "planned-project-row";
        const projectTd = document.createElement("td");
        projectTd.colSpan = 7;
        const projectBtn = document.createElement("button");
        projectBtn.type = "button";
        projectBtn.className = "planned-project-toggle";
        projectBtn.setAttribute("data-project-key", projectKey);
        const collapsed = collapsedTotalPlannedProjectKeys.has(projectKey);
        projectBtn.innerHTML =
          '<span class="caret">' + (collapsed ? "+" : "-") + '</span>'
          + '<span>' + projectName + "</span>"
          + '<span>(' + String(subtasks.length) + " subtasks, " + formatHours(groupHours) + ")</span>";
        projectTd.appendChild(projectBtn);
        projectTr.appendChild(projectTd);
        totalPlannedDetailsBodyEl.appendChild(projectTr);

        if (!collapsed) {{
          for (const subtask of subtasks) {{
            const tr = document.createElement("tr");
            tr.className = "planned-epic-row";
            const tdProject = document.createElement("td");
            tdProject.textContent = "";
            const tdSubtaskKey = document.createElement("td");
            tdSubtaskKey.textContent = String(subtask && subtask.subtask_jira_id || "");
            const tdTaskType = document.createElement("td");
            tdTaskType.textContent = String(subtask && subtask.task_type || "Subtask");
            const tdSubtaskName = document.createElement("td");
            tdSubtaskName.textContent = String(subtask && subtask.subtask_name || "");
            const tdStart = document.createElement("td");
            tdStart.textContent = String(subtask && subtask.subtask_start_date || "");
            const tdDue = document.createElement("td");
            tdDue.textContent = String(subtask && subtask.subtask_due_date || "");
            const tdHours = document.createElement("td");
            tdHours.className = "hours";
            tdHours.textContent = formatHoursPlain(subtask && subtask.planned_hours);
            tr.appendChild(tdProject);
            tr.appendChild(tdSubtaskKey);
            tr.appendChild(tdTaskType);
            tr.appendChild(tdSubtaskName);
            tr.appendChild(tdStart);
            tr.appendChild(tdDue);
            tr.appendChild(tdHours);
            totalPlannedDetailsBodyEl.appendChild(tr);
          }}
        }}
      }}
      for (const btn of Array.from(totalPlannedDetailsBodyEl.querySelectorAll(".planned-project-toggle"))) {{
        btn.addEventListener("click", (event) => {{
          event.preventDefault();
          event.stopPropagation();
          const key = String(btn.getAttribute("data-project-key") || "");
          if (!key) {{
            return;
          }}
          if (collapsedTotalPlannedProjectKeys.has(key)) {{
            collapsedTotalPlannedProjectKeys.delete(key);
          }} else {{
            collapsedTotalPlannedProjectKeys.add(key);
          }}
          renderTotalPlannedDetails();
        }});
      }}
    }}

    function setTotalLeavesPlannedDetailsOpen(open) {{
      totalLeavesPlannedDetailsOpen = !!open;
      if (totalLeavesPlannedDetailsEl) {{
        totalLeavesPlannedDetailsEl.hidden = !totalLeavesPlannedDetailsOpen;
      }}
      if (totalLeavesPlannedCardEl) {{
        totalLeavesPlannedCardEl.classList.toggle("is-expanded", totalLeavesPlannedDetailsOpen);
        totalLeavesPlannedCardEl.setAttribute("aria-expanded", totalLeavesPlannedDetailsOpen ? "true" : "false");
      }}
    }}

    function setTotalActualDetailsOpen(open) {{
      totalActualDetailsOpen = !!open;
      if (totalLoggedDetailsEl) {{
        totalLoggedDetailsEl.hidden = !totalActualDetailsOpen;
      }}
      if (totalLoggedCardEl) {{
        totalLoggedCardEl.classList.toggle("is-expanded", totalActualDetailsOpen);
        totalLoggedCardEl.setAttribute("aria-expanded", totalActualDetailsOpen ? "true" : "false");
      }}
    }}

    function renderTotalActualDetails() {{
      if (!totalLoggedDetailsBodyEl || !totalLoggedDetailsMetaEl) {{
        return;
      }}
      totalLoggedDetailsBodyEl.innerHTML = "";
      const groups = Array.isArray(totalActualDetailsGroups) ? totalActualDetailsGroups : [];
      let totalSubtasks = 0;
      let totalHours = 0;
      for (const group of groups) {{
        const subtasks = Array.isArray(group && group.subtasks) ? group.subtasks : [];
        totalSubtasks += subtasks.length;
        for (const subtask of subtasks) {{
          totalHours += toFiniteNumber(subtask && subtask.actual_hours, 0);
        }}
      }}
      totalLoggedDetailsMetaEl.textContent =
        "Projects: " + String(groups.length) + " | Subtasks: " + String(totalSubtasks) + " | Actual Hours: " + formatHours(totalHours);
      if (!groups.length) {{
        const tr = document.createElement("tr");
        const td = document.createElement("td");
        td.colSpan = 7;
        td.className = "score-details-empty";
        td.textContent = "No actual subtask rows found for current date/team filters.";
        tr.appendChild(td);
        totalLoggedDetailsBodyEl.appendChild(tr);
        return;
      }}
      for (const group of groups) {{
        const projectKey = String(group && group.project_key || "");
        const projectName = String(group && group.project_name || projectKey || "Unknown Project");
        const subtasks = Array.isArray(group && group.subtasks) ? group.subtasks : [];
        const groupHours = subtasks.reduce((sum, subtask) => sum + toFiniteNumber(subtask && subtask.actual_hours, 0), 0);
        const projectTr = document.createElement("tr");
        projectTr.className = "planned-project-row";
        const projectTd = document.createElement("td");
        projectTd.colSpan = 7;
        const projectBtn = document.createElement("button");
        projectBtn.type = "button";
        projectBtn.className = "planned-project-toggle";
        projectBtn.setAttribute("data-project-key", projectKey);
        const collapsed = collapsedTotalActualProjectKeys.has(projectKey);
        projectBtn.innerHTML =
          '<span class="caret">' + (collapsed ? "+" : "-") + '</span>'
          + '<span>' + projectName + "</span>"
          + '<span>(' + String(subtasks.length) + " subtasks, " + formatHours(groupHours) + ")</span>";
        projectTd.appendChild(projectBtn);
        projectTr.appendChild(projectTd);
        totalLoggedDetailsBodyEl.appendChild(projectTr);

        if (!collapsed) {{
          for (const subtask of subtasks) {{
            const tr = document.createElement("tr");
            tr.className = "planned-epic-row";
            const tdProject = document.createElement("td");
            tdProject.textContent = "";
            const tdSubtaskKey = document.createElement("td");
            tdSubtaskKey.textContent = String(subtask && subtask.subtask_jira_id || "");
            const tdTaskType = document.createElement("td");
            tdTaskType.textContent = String(subtask && subtask.task_type || "Subtask");
            const tdSubtaskName = document.createElement("td");
            tdSubtaskName.textContent = String(subtask && subtask.subtask_name || "");
            const tdStart = document.createElement("td");
            tdStart.textContent = String(subtask && subtask.subtask_start_date || "");
            const tdDue = document.createElement("td");
            tdDue.textContent = String(subtask && subtask.subtask_due_date || "");
            const tdHours = document.createElement("td");
            tdHours.className = "hours";
            tdHours.textContent = formatHoursPlain(subtask && subtask.actual_hours);
            tr.appendChild(tdProject);
            tr.appendChild(tdSubtaskKey);
            tr.appendChild(tdTaskType);
            tr.appendChild(tdSubtaskName);
            tr.appendChild(tdStart);
            tr.appendChild(tdDue);
            tr.appendChild(tdHours);
            totalLoggedDetailsBodyEl.appendChild(tr);
          }}
        }}
      }}
      for (const btn of Array.from(totalLoggedDetailsBodyEl.querySelectorAll(".planned-project-toggle"))) {{
        btn.addEventListener("click", (event) => {{
          event.preventDefault();
          event.stopPropagation();
          const key = String(btn.getAttribute("data-project-key") || "");
          if (!key) {{
            return;
          }}
          if (collapsedTotalActualProjectKeys.has(key)) {{
            collapsedTotalActualProjectKeys.delete(key);
          }} else {{
            collapsedTotalActualProjectKeys.add(key);
          }}
          renderTotalActualDetails();
        }});
      }}
    }}

    function renderTotalLeavesPlannedDetails() {{
      if (!totalLeavesPlannedDetailsBodyEl || !totalLeavesPlannedDetailsMetaEl) {{
        return;
      }}
      totalLeavesPlannedDetailsBodyEl.innerHTML = "";
      if (totalLeavesPlannedAssigneeSummaryBodyEl) {{
        totalLeavesPlannedAssigneeSummaryBodyEl.innerHTML = "";
      }}
      const rows = Array.isArray(totalLeavesPlannedDetailsRows) ? totalLeavesPlannedDetailsRows : [];
      const totalHours = rows.reduce((sum, row) => sum + toFiniteNumber(row && row.hours, 0), 0);
      totalLeavesPlannedDetailsMetaEl.textContent =
        "Rows: " + String(rows.length) + " | Total Hours: " + formatHours(totalHours);
      const assigneeSummaryMap = new Map();
      for (const row of rows) {{
        const assigneeName = String(row && row.assignee_name || "").trim() || "Unassigned";
        const rowHours = toFiniteNumber(row && row.hours, 0);
        assigneeSummaryMap.set(assigneeName, toFiniteNumber(assigneeSummaryMap.get(assigneeName), 0) + rowHours);
      }}
      const assigneeSummaryRows = Array.from(assigneeSummaryMap.entries())
        .map(([assignee_name, hours]) => ({{ assignee_name, hours }}))
        .sort((left, right) => {{
          const diff = toFiniteNumber(right && right.hours, 0) - toFiniteNumber(left && left.hours, 0);
          if (Math.abs(diff) > 1e-9) {{
            return diff;
          }}
          return String(left && left.assignee_name || "").localeCompare(String(right && right.assignee_name || ""));
        }});
      if (totalLeavesPlannedAssigneeSummaryMetaEl) {{
        totalLeavesPlannedAssigneeSummaryMetaEl.textContent =
          "Assignees: " + String(assigneeSummaryRows.length) + " | Total Hours: " + formatHours(totalHours);
      }}
      if (!rows.length) {{
        const tr = document.createElement("tr");
        const td = document.createElement("td");
        td.colSpan = 5;
        td.className = "score-details-empty";
        td.textContent = "No planned leave rows found for current date/team filters.";
        tr.appendChild(td);
        totalLeavesPlannedDetailsBodyEl.appendChild(tr);
      }} else {{
        for (const row of rows) {{
          const tr = document.createElement("tr");
          const tdDate = document.createElement("td");
          tdDate.textContent = String(row && row.date || "");
          const tdAssignee = document.createElement("td");
          tdAssignee.textContent = String(row && row.assignee_name || "");
          const tdHours = document.createElement("td");
          tdHours.className = "hours";
          tdHours.textContent = formatHoursPlain(row && row.hours);
        const tdJiraId = document.createElement("td");
        tdJiraId.textContent = String(row && row.jira_key || "");
        const tdLink = document.createElement("td");
        const jiraUrl = String(row && row.jira_url || "");
        if (jiraUrl) {{
          const urls = jiraUrl.split(",").map((part) => String(part || "").trim()).filter(Boolean);
          if (urls.length > 1) {{
            tdLink.style.whiteSpace = "normal";
          }}
          urls.forEach((urlValue, index) => {{
            const link = document.createElement("a");
            link.className = "leaves-link-btn";
            link.href = urlValue;
            link.target = "_blank";
            link.rel = "noopener noreferrer";
            link.title = "Open Jira task";
            link.setAttribute("aria-label", "Open Jira task");
            link.innerHTML = '<span class="material-icons-outlined" aria-hidden="true">open_in_new</span>';
            tdLink.appendChild(link);
            if (index < urls.length - 1) {{
              tdLink.appendChild(document.createTextNode(" "));
            }}
          }});
        }} else {{
          tdLink.textContent = "-";
        }}
          tr.appendChild(tdDate);
          tr.appendChild(tdAssignee);
          tr.appendChild(tdHours);
          tr.appendChild(tdJiraId);
          tr.appendChild(tdLink);
          totalLeavesPlannedDetailsBodyEl.appendChild(tr);
        }}
      }}
      if (!totalLeavesPlannedAssigneeSummaryBodyEl) {{
        return;
      }}
      if (!assigneeSummaryRows.length) {{
        const tr = document.createElement("tr");
        const td = document.createElement("td");
        td.colSpan = 2;
        td.className = "score-details-empty";
        td.textContent = "No assignee summary rows for current date/team filters.";
        tr.appendChild(td);
        totalLeavesPlannedAssigneeSummaryBodyEl.appendChild(tr);
        return;
      }}
      for (const row of assigneeSummaryRows) {{
        const tr = document.createElement("tr");
        const tdAssignee = document.createElement("td");
        tdAssignee.textContent = String(row && row.assignee_name || "");
        const tdHours = document.createElement("td");
        tdHours.className = "hours";
        tdHours.textContent = formatHoursPlain(row && row.hours);
        tr.appendChild(tdAssignee);
        tr.appendChild(tdHours);
        totalLeavesPlannedAssigneeSummaryBodyEl.appendChild(tr);
      }}
    }}

    function capacityProfileKey(profile) {{
      const fromDate = String(profile && profile.from_date || "");
      const toDate = String(profile && profile.to_date || "");
      return fromDate + "|" + toDate;
    }}

    function setCapacityStatus(text, variant) {{
      if (!capacityProfileStatusEl) {{
        return;
      }}
      capacityProfileStatusEl.textContent = String(text || "");
      capacityProfileStatusEl.setAttribute("data-variant", variant || "info");
    }}

    function findCapacityProfileByKey(key) {{
      const wanted = String(key || "");
      if (!wanted) {{
        return null;
      }}
      for (const profile of capacityProfiles) {{
        if (capacityProfileKey(profile) === wanted) {{
          return profile;
        }}
      }}
      return null;
    }}

    function renderCapacityProfileOptions() {{
      if (!capacityProfileSelectEl) {{
        return;
      }}
      const previousSelectedKey = String(capacityProfileSelectEl.value || "");
      if (!capacityProfiles.length) {{
        capacityProfileSelectEl.innerHTML = '<option value="">No saved profiles found</option>';
        capacityProfileSelectEl.disabled = true;
        if (capacityProfileApplyEl) {{
          capacityProfileApplyEl.disabled = true;
        }}
        if (capacityProfileDetailsEl) {{
          capacityProfileDetailsEl.textContent = "No saved capacity profile found in assignee_hours_capacity.db.";
        }}
        return;
      }}
      capacityProfileSelectEl.disabled = false;
      const options = ['<option value="">Select a saved profile</option>'];
      for (const profile of capacityProfiles) {{
        const key = capacityProfileKey(profile);
        const fromDate = String(profile && profile.from_date || "");
        const toDate = String(profile && profile.to_date || "");
        const employees = toFiniteNumber(profile && profile.employee_count, 0);
        const standardHours = toFiniteNumber(profile && profile.standard_hours_per_day, 0);
        const ramadanHours = toFiniteNumber(profile && profile.ramadan_hours_per_day, 0);
        const label = fromDate + " to " + toDate + " (Emp:" + String(employees) + ", Std:" + formatHours(standardHours) + ", Ram:" + formatHours(ramadanHours) + ")";
        options.push('<option value="' + key + '">' + label + '</option>');
      }}
      capacityProfileSelectEl.innerHTML = options.join("");
      if (appliedCapacityProfileKey) {{
        capacityProfileSelectEl.value = appliedCapacityProfileKey;
      }} else if (previousSelectedKey && findCapacityProfileByKey(previousSelectedKey)) {{
        capacityProfileSelectEl.value = previousSelectedKey;
      }}
      if (capacityProfileApplyEl) {{
        capacityProfileApplyEl.disabled = false;
      }}
    }}

    function getSelectedCapacityProfile() {{
      if (!capacityProfileSelectEl) {{
        return null;
      }}
      const selectedKey = String(capacityProfileSelectEl.value || "");
      return findCapacityProfileByKey(selectedKey);
    }}

    function renderCapacityProfileExpanded(profile) {{
      if (!capacityProfileExpandedEl) {{
        return;
      }}
      const bounds = getDateFilterBounds();
      const fromDate = bounds && bounds.start ? toIsoDate(bounds.start) : "";
      const toDate = bounds && bounds.end ? toIsoDate(bounds.end) : "";
      if (!profile) {{
        capacityProfileExpandedEl.innerHTML = '<div class="capacity-empty">Select a saved profile to preview its calendar for the active nested-view date range.</div>';
        return;
      }}
      if (!fromDate || !toDate || toDate < fromDate) {{
        capacityProfileExpandedEl.innerHTML = '<div class="capacity-empty">Select a valid date range to view the capacity calendar.</div>';
        return;
      }}
      const breakdown = computeCapacityBreakdownForRange(profile, bounds);
      const employees = Math.max(0, Math.round(toFiniteNumber(profile.employee_count, 0)));
      const standardHours = toFiniteNumber(profile.standard_hours_per_day, 0);
      const ramadanHours = toFiniteNumber(profile.ramadan_hours_per_day, standardHours);
      const ramadanStart = asText(profile.ramadan_start_date);
      const ramadanEnd = asText(profile.ramadan_end_date);
      const holidayDates = Array.isArray(profile.holiday_dates)
        ? profile.holiday_dates.map((item) => asText(item)).filter(Boolean).sort()
        : [];
      const holidaySet = new Set(holidayDates);
      const leaveByDay = new Map();
      for (const row of leaveDailyRows) {{
        const day = asText(row && row.period_day);
        if (!day || day < fromDate || day > toDate) {{
          continue;
        }}
        const planned = toFiniteNumber(row && row.planned_taken_hours, 0);
        const unplanned = toFiniteNumber(row && row.unplanned_taken_hours, 0);
        if (planned <= 0 && unplanned <= 0) {{
          continue;
        }}
        const current = leaveByDay.get(day) || {{ planned: 0, unplanned: 0 }};
        current.planned += planned;
        current.unplanned += unplanned;
        leaveByDay.set(day, current);
      }}
      let totalPlannedLeave = 0;
      let totalUnplannedLeave = 0;
      leaveByDay.forEach((item) => {{
        totalPlannedLeave += toFiniteNumber(item && item.planned, 0);
        totalUnplannedLeave += toFiniteNumber(item && item.unplanned, 0);
      }});
      const cursor = parseDateValue(fromDate);
      const end = parseDateValue(toDate);
      if (!cursor || !end) {{
        capacityProfileExpandedEl.innerHTML = '<div class="capacity-empty">Date range is invalid for calendar rendering.</div>';
        return;
      }}
      const todayIso = toIsoDate(new Date());
      const monthCards = [];
      cursor.setDate(1);
      while (cursor <= end) {{
        const monthStart = new Date(cursor.getFullYear(), cursor.getMonth(), 1);
        const monthEnd = new Date(cursor.getFullYear(), cursor.getMonth() + 1, 0);
        const firstDayOffset = (monthStart.getDay() + 6) % 7;
        const lastDayOffset = (monthEnd.getDay() + 6) % 7;
        const gridStart = new Date(monthStart);
        gridStart.setDate(gridStart.getDate() - firstDayOffset);
        const gridEnd = new Date(monthEnd);
        gridEnd.setDate(gridEnd.getDate() + (6 - lastDayOffset));
        const cells = [
          '<div class="capacity-dow">Mon</div>',
          '<div class="capacity-dow">Tue</div>',
          '<div class="capacity-dow">Wed</div>',
          '<div class="capacity-dow">Thu</div>',
          '<div class="capacity-dow">Fri</div>',
          '<div class="capacity-dow">Sat</div>',
          '<div class="capacity-dow">Sun</div>',
        ];
        for (let day = new Date(gridStart); day <= gridEnd; day.setDate(day.getDate() + 1)) {{
          const iso = toIsoDate(day);
          const isOutsideMonth = day.getMonth() !== monthStart.getMonth();
          const isWeekend = day.getDay() === 0 || day.getDay() === 6;
          const isRamadan = !!(ramadanStart && ramadanEnd && iso >= ramadanStart && iso <= ramadanEnd);
          const isHoliday = holidaySet.has(iso);
          const isToday = iso === todayIso;
          const leave = leaveByDay.get(iso) || null;
          const classes = ["capacity-day"];
          if (isOutsideMonth) {{
            classes.push("is-out");
          }}
          if (isWeekend) {{
            classes.push("is-weekend");
          }}
          if (isRamadan) {{
            classes.push("is-ramadan");
          }}
          if (isHoliday) {{
            classes.push("is-holiday");
          }}
          if (isToday) {{
            classes.push("is-today");
          }}
          if (leave && isRamadan) {{
            classes.push("has-ramadan-leave");
          }} else if (leave) {{
            classes.push("has-leave");
          }}
          const tags = [];
          if (isRamadan) {{
            tags.push('<span class="capacity-day-tag r">R</span>');
          }}
          if (isHoliday) {{
            tags.push('<span class="capacity-day-tag h">H</span>');
          }}
          if (leave) {{
            const rawLeaveHours = toFiniteNumber(leave.planned, 0) + toFiniteNumber(leave.unplanned, 0);
            const leaveHours = rawLeaveHours > 0 ? (isRamadan ? ramadanHours : standardHours) : 0;
            tags.push(
              isRamadan
                ? '<span class="capacity-day-tag rl">RL ' + escapeHtml(leaveHours.toFixed(1)) + 'h</span>'
                : '<span class="capacity-day-tag l">L ' + escapeHtml(leaveHours.toFixed(1)) + 'h</span>'
            );
          }}
          cells.push(
            '<div class="' + classes.join(" ") + '"><div class="capacity-day-num">' + String(day.getDate()) + '</div><div class="capacity-day-tags">' + tags.join("") + '</div></div>'
          );
        }}
        const monthLabel = monthStart.toLocaleDateString(undefined, {{
          year: "numeric",
          month: "long",
        }});
        const isCurrentMonth = monthStart.getFullYear() === new Date().getFullYear()
          && monthStart.getMonth() === new Date().getMonth();
        monthCards.push(
          '<div class="capacity-month"' + (isCurrentMonth ? ' data-current-month="1"' : "") + '><div class="capacity-month-head">'
          + escapeHtml(monthLabel)
          + '</div><div class="capacity-month-grid">'
          + cells.join("")
          + '</div></div>'
        );
        cursor.setMonth(cursor.getMonth() + 1, 1);
      }}
      const isAppliedProfile = !!(
        appliedCapacityProfile
        && capacityProfileKey(appliedCapacityProfile) === capacityProfileKey(profile)
      );
      const modeLabel = isAppliedProfile ? "Applied profile" : "Preview profile";
      const holidayText = holidayDates.length
        ? holidayDates.map((value) => formatCapacityDate(value)).join(", ")
        : "None";
      const ramadanText = ramadanStart && ramadanEnd
        ? formatCapacityDate(ramadanStart) + " to " + formatCapacityDate(ramadanEnd)
        : "Not set";
      const perAssigneeCapacity = employees > 0
        ? (toFiniteNumber(breakdown.profileCapacityHours, 0) / employees)
        : 0;
      capacityProfileExpandedEl.innerHTML = ''
        + '<div class="capacity-expanded-head">'
        + '<div class="capacity-expanded-title">Capacity Profile Calendar</div>'
        + '<div class="capacity-expanded-sub">' + escapeHtml(modeLabel) + ' | Active range: '
        + escapeHtml(formatCapacityDate(fromDate)) + ' to ' + escapeHtml(formatCapacityDate(toDate)) + '</div>'
        + '</div>'
        + '<div class="capacity-expanded-grid">'
        + '<div class="capacity-chip"><div class="k">Profile Range</div><div class="v">' + escapeHtml(formatCapacityDate(profile.from_date)) + ' to ' + escapeHtml(formatCapacityDate(profile.to_date)) + '</div></div>'
        + '<div class="capacity-chip"><div class="k">Employee Count</div><div class="v">' + String(employees) + '</div></div>'
        + '<div class="capacity-chip"><div class="k">Standard Hours/Day</div><div class="v">' + escapeHtml(standardHours.toFixed(2)) + 'h</div></div>'
        + '<div class="capacity-chip"><div class="k">Ramadan Hours/Day</div><div class="v">' + escapeHtml(ramadanHours.toFixed(2)) + 'h</div></div>'
        + '<div class="capacity-chip"><div class="k">Ramadan Range</div><div class="v">' + escapeHtml(ramadanText) + '</div></div>'
        + '<div class="capacity-chip"><div class="k">Business Days</div><div class="v">' + String(Math.round(toFiniteNumber(breakdown.weekdayCount, 0))) + 'd</div></div>'
        + '<div class="capacity-chip"><div class="k">Holiday Weekdays</div><div class="v">' + String(Math.round(toFiniteNumber(breakdown.holidayWeekdayCount, 0))) + '</div></div>'
        + '<div class="capacity-chip"><div class="k">Range Capacity</div><div class="v">' + escapeHtml(formatHours(breakdown.profileCapacityHours)) + '</div></div>'
        + '<div class="capacity-chip"><div class="k">Per Assignee Capacity</div><div class="v">' + escapeHtml(formatHours(perAssigneeCapacity)) + '</div></div>'
        + '<div class="capacity-chip"><div class="k">Planned Leave (Range)</div><div class="v">' + escapeHtml(formatHours(totalPlannedLeave)) + '</div></div>'
        + '<div class="capacity-chip"><div class="k">Unplanned Leave (Range)</div><div class="v">' + escapeHtml(formatHours(totalUnplannedLeave)) + '</div></div>'
        + '</div>'
        + '<div class="capacity-legend">'
        + '<span class="pill">R = Ramadan</span>'
        + '<span class="pill">H = Holiday</span>'
        + '<span class="pill">L = Leave hours</span>'
        + '<span class="pill">RL = Ramadan leave hours</span>'
        + '</div>'
        + '<div class="capacity-expanded-sub" style="margin-bottom:8px;">Holiday dates: ' + escapeHtml(holidayText) + '</div>'
        + '<div class="capacity-calendar-wrap">' + monthCards.join("") + '</div>';
      const calendarWrap = capacityProfileExpandedEl.querySelector(".capacity-calendar-wrap");
      const currentMonthCard = calendarWrap ? calendarWrap.querySelector('[data-current-month="1"]') : null;
      if (calendarWrap && currentMonthCard) {{
        calendarWrap.scrollLeft = Math.max(0, currentMonthCard.offsetLeft - 8);
      }}
    }}

    function renderCapacityProfileDetails() {{
      if (!capacityProfileDetailsEl) {{
        return;
      }}
      const profile = getSelectedCapacityProfile();
      if (!profile) {{
        capacityProfileDetailsEl.textContent = "Select a saved profile and click Apply to use its calculated capacity.";
        renderCapacityProfileExpanded(null);
        return;
      }}
      const holidays = Array.isArray(profile.holiday_dates) ? profile.holiday_dates.length : 0;
      const fromDate = String(profile.from_date || "-");
      const toDate = String(profile.to_date || "-");
      const emp = String(toFiniteNumber(profile.employee_count, 0));
      const std = formatHours(profile.standard_hours_per_day || 0);
      const ramadan = formatHours(profile.ramadan_hours_per_day || 0);
      const ramadanRange = (profile.ramadan_start_date && profile.ramadan_end_date)
        ? (String(profile.ramadan_start_date) + " to " + String(profile.ramadan_end_date))
        : "Not set";
      const cap = formatHours(profile.available_capacity_hours || 0);
      const updated = String(profile.updated_at_utc || "-");
      capacityProfileDetailsEl.textContent =
        "Range: " + fromDate + " to " + toDate +
        " | Employees: " + emp +
        " | Std/Day: " + std +
        " | Ramadan/Day: " + ramadan +
        " | Ramadan Range: " + ramadanRange +
        " | Holidays: " + String(holidays) +
        " | Calculated Capacity: " + cap +
        " | Updated: " + updated;
      if (appliedCapacityProfile && appliedCapacityProfileKey === capacityProfileKey(profile)) {{
        const bounds = getDateFilterBounds();
        const dynamicCapacity = computeCapacityHoursForRange(profile, bounds);
        const rangeStart = bounds && bounds.start ? toIsoDate(bounds.start) : "-";
        const rangeEnd = bounds && bounds.end ? toIsoDate(bounds.end) : "-";
        capacityProfileDetailsEl.textContent +=
          " | Applied to selected range: " + rangeStart + " to " + rangeEnd +
          " | Dynamic Capacity: " + formatHours(dynamicCapacity);
      }}
      renderCapacityProfileExpanded(profile);
    }}

    function applyProfilesPayload(nextProfiles) {{
      capacityProfiles = Array.isArray(nextProfiles) ? nextProfiles.slice() : [];
      if (appliedCapacityProfileKey) {{
        appliedCapacityProfile = findCapacityProfileByKey(appliedCapacityProfileKey);
        if (!appliedCapacityProfile) {{
          appliedCapacityProfileKey = "";
        }}
      }}
      if (!appliedCapacityProfileKey && capacityProfiles.length) {{
        appliedCapacityProfile = capacityProfiles[0];
        appliedCapacityProfileKey = capacityProfileKey(appliedCapacityProfile);
      }}
      renderCapacityProfileOptions();
      renderCapacityProfileDetails();
      const requestVersion = ++scorecardUpdateVersion;
      if (projectFilterProgress && !projectFilterProgress.hidden) {{
        setProjectFilterLoading(true, requestVersion);
      }}
      if (teamFilterProgress && !teamFilterProgress.hidden) {{
        setTeamFilterLoading(true, requestVersion);
      }}
      void updateScoreCards(scorecardSourceRows, requestVersion, buildScorecardSelectionSnapshot());
    }}

    async function refreshCapacityProfilesFromApi() {{
      if (!hasCapacityApi) {{
        setCapacityStatus("Profile save/delete needs server mode (not file://).", "info");
        return false;
      }}
      const response = await fetch(CAPACITY_PROFILES_ENDPOINT, {{ method: "GET" }});
      if (!response.ok) {{
        throw new Error("Failed to load profiles (" + String(response.status) + ").");
      }}
      const payload = await response.json();
      applyProfilesPayload(payload && payload.profiles);
      return true;
    }}

    function applyManagedFieldsPayload(payload) {{
      const next = new Map();
      const fields = Array.isArray(payload && payload.fields) ? payload.fields : [];
      for (const field of fields) {{
        const key = String(field && field.field_key || "").trim().toLowerCase();
        if (!key || !field || !field.is_active) {{
          continue;
        }}
        next.set(key, field);
      }}
      managedFieldsByKey = next;
      const requestVersion = ++scorecardUpdateVersion;
      if (projectFilterProgress && !projectFilterProgress.hidden) {{
        setProjectFilterLoading(true, requestVersion);
      }}
      if (teamFilterProgress && !teamFilterProgress.hidden) {{
        setTeamFilterLoading(true, requestVersion);
      }}
      void updateScoreCards(scorecardSourceRows, requestVersion, buildScorecardSelectionSnapshot());
    }}

    async function refreshManagedFieldsFromApi() {{
      if (!hasManagedFieldsApi) {{
        return false;
      }}
      const response = await fetch(MANAGED_FIELDS_ENDPOINT, {{ method: "GET" }});
      if (!response.ok) {{
        throw new Error("Failed to load managed fields (" + String(response.status) + ").");
      }}
      const payload = await response.json();
      applyManagedFieldsPayload(payload);
      return true;
    }}

    function evaluateExpressionWithContext(expression, context) {{
      const text = String(expression || "").trim();
      if (!text) {{
        return {{ ok: false, value: NaN, error: "Empty expression." }};
      }}
      const tokens = [];
      let idx = 0;
      while (idx < text.length) {{
        const ch = text[idx];
        if (/\\s/.test(ch)) {{
          idx += 1;
          continue;
        }}
        if ("+-*/".includes(ch)) {{
          tokens.push({{ t: "op", v: ch, p: idx }});
          idx += 1;
          continue;
        }}
        if (ch === "(") {{
          tokens.push({{ t: "lparen", v: ch, p: idx }});
          idx += 1;
          continue;
        }}
        if (ch === ")") {{
          tokens.push({{ t: "rparen", v: ch, p: idx }});
          idx += 1;
          continue;
        }}
        if (ch === ",") {{
          tokens.push({{ t: "comma", v: ch, p: idx }});
          idx += 1;
          continue;
        }}
        if (/[A-Za-z_]/.test(ch)) {{
          const start = idx;
          idx += 1;
          while (idx < text.length && /[A-Za-z0-9_]/.test(text[idx])) {{
            idx += 1;
          }}
          tokens.push({{ t: "ident", v: text.slice(start, idx), p: start }});
          continue;
        }}
        return {{ ok: false, value: NaN, error: "Invalid character at position " + String(idx + 1) + "." }};
      }}
      tokens.push({{ t: "eof", v: "", p: text.length }});
      let pos = 0;
      function peek() {{
        return tokens[pos];
      }}
      function consume(expected) {{
        const token = tokens[pos];
        if (expected && token.t !== expected) {{
          throw new Error("Expected " + expected + " at position " + String(token.p + 1) + ".");
        }}
        pos += 1;
        return token;
      }}
      function callFunction(name, arg) {{
        const fn = String(name || "").toLowerCase();
        const safe = toFiniteNumber(arg, 0);
        if (fn === "sum" || fn === "min" || fn === "max" || fn === "average") {{
          return safe;
        }}
        if (fn === "count") {{
          return safe !== 0 ? 1 : 0;
        }}
        throw new Error("Unknown function '" + String(name || "") + "'.");
      }}
      function parseExpression() {{
        let value = parseTerm();
        while (peek().t === "op" && (peek().v === "+" || peek().v === "-")) {{
          const op = consume("op").v;
          const rhs = parseTerm();
          value = op === "+" ? (value + rhs) : (value - rhs);
        }}
        return value;
      }}
      function parseTerm() {{
        let value = parseFactor();
        while (peek().t === "op" && (peek().v === "*" || peek().v === "/")) {{
          const op = consume("op").v;
          const rhs = parseFactor();
          if (op === "*") {{
            value *= rhs;
          }} else {{
            value = rhs === 0 ? 0 : (value / rhs);
          }}
        }}
        return value;
      }}
      function parseFactor() {{
        const token = peek();
        if (token.t === "ident") {{
          const ident = consume("ident");
          const key = String(ident.v || "").toLowerCase();
          if (peek().t === "lparen") {{
            consume("lparen");
            const arg = parseExpression();
            if (peek().t === "comma") {{
              throw new Error("Function '" + ident.v + "' accepts one argument at position " + String(peek().p + 1) + ".");
            }}
            consume("rparen");
            return callFunction(ident.v, arg);
          }}
          if (!Object.prototype.hasOwnProperty.call(context, key)) {{
            throw new Error("Unknown identifier '" + ident.v + "'.");
          }}
          return toFiniteNumber(context[key], 0);
        }}
        if (token.t === "lparen") {{
          consume("lparen");
          const value = parseExpression();
          consume("rparen");
          return value;
        }}
        throw new Error("Unexpected token at position " + String(token.p + 1) + ".");
      }}
      try {{
        const value = parseExpression();
        if (peek().t !== "eof") {{
          throw new Error("Unexpected token at position " + String(peek().p + 1) + ".");
        }}
        return {{ ok: true, value: roundHours(value), error: "" }};
      }} catch (error) {{
        return {{ ok: false, value: NaN, error: String(error && error.message || error || "Expression parse error.") }};
      }}
    }}

    function managedFieldFormulaText(fieldKey, fallbackFormulaText) {{
      const key = String(fieldKey || "").trim().toLowerCase();
      if (!managedFieldsByKey.has(key)) {{
        return fallbackFormulaText;
      }}
      const item = managedFieldsByKey.get(key);
      const expression = String(item && item.formula_expression || "").trim();
      return expression || fallbackFormulaText;
    }}

    function evaluateManagedField(fieldKey, fallbackValue, context) {{
      const key = String(fieldKey || "").trim().toLowerCase();
      if (!managedFieldsByKey.has(key)) {{
        return {{ value: fallbackValue, usedManagedField: false, formulaText: "", error: "" }};
      }}
      const item = managedFieldsByKey.get(key);
      const expression = String(item && item.formula_expression || "").trim();
      if (!expression) {{
        return {{ value: fallbackValue, usedManagedField: false, formulaText: "", error: "" }};
      }}
      const evaluated = evaluateExpressionWithContext(expression, context || {{}});
      if (!evaluated.ok) {{
        return {{ value: fallbackValue, usedManagedField: false, formulaText: expression, error: evaluated.error }};
      }}
      return {{ value: evaluated.value, usedManagedField: true, formulaText: expression, error: "" }};
    }}

    function toIsoDate(dateValue) {{
      if (!(dateValue instanceof Date) || Number.isNaN(dateValue.getTime())) {{
        return "";
      }}
      const y = dateValue.getFullYear();
      const m = String(dateValue.getMonth() + 1).padStart(2, "0");
      const d = String(dateValue.getDate()).padStart(2, "0");
      return String(y) + "-" + m + "-" + d;
    }}

    function computeCapacityBreakdownForRange(profile, bounds) {{
      const emptyBreakdown = {{
        profileCapacityHours: 0,
        employees: 0,
        standardHours: 0,
        ramadanHours: 0,
        weekdayCount: 0,
        ramadanWeekdayCount: 0,
        regularWeekdayCount: 0,
        holidayWeekdayCount: 0,
      }};
      if (!profile || !bounds || !bounds.start || !bounds.end) {{
        return emptyBreakdown;
      }}
      const employees = toFiniteNumber(profile.employee_count, 0);
      const standardHours = toFiniteNumber(profile.standard_hours_per_day, 0);
      const ramadanHours = toFiniteNumber(profile.ramadan_hours_per_day, 0);
      if (employees <= 0 || standardHours <= 0 || ramadanHours <= 0) {{
        return {{
          ...emptyBreakdown,
          employees,
          standardHours,
          ramadanHours,
        }};
      }}
      const start = new Date(bounds.start.getFullYear(), bounds.start.getMonth(), bounds.start.getDate());
      const end = new Date(bounds.end.getFullYear(), bounds.end.getMonth(), bounds.end.getDate());
      const ramadanStart = parseDateValue(profile.ramadan_start_date);
      const ramadanEnd = parseDateValue(profile.ramadan_end_date);
      const holidaySet = new Set();
      if (Array.isArray(profile.holiday_dates)) {{
        for (const holiday of profile.holiday_dates) {{
          const holidayDate = parseDateValue(holiday);
          if (holidayDate) {{
            holidaySet.add(toIsoDate(holidayDate));
          }}
        }}
      }}

      let capacity = 0;
      let weekdayCount = 0;
      let ramadanWeekdayCount = 0;
      let regularWeekdayCount = 0;
      let holidayWeekdayCount = 0;
      const cursor = new Date(start);
      while (cursor <= end) {{
        const day = cursor.getDay();
        if (day >= 1 && day <= 5) {{
          const iso = toIsoDate(cursor);
          if (holidaySet.has(iso)) {{
            holidayWeekdayCount += 1;
          }} else {{
            weekdayCount += 1;
            const inRamadan = ramadanStart && ramadanEnd && cursor >= ramadanStart && cursor <= ramadanEnd;
            const perDay = inRamadan ? ramadanHours : standardHours;
            capacity += (employees * perDay);
            if (inRamadan) {{
              ramadanWeekdayCount += 1;
            }} else {{
              regularWeekdayCount += 1;
            }}
          }}
        }}
        cursor.setDate(cursor.getDate() + 1);
      }}
      return {{
        profileCapacityHours: Math.round(capacity * 100) / 100,
        employees,
        standardHours,
        ramadanHours,
        weekdayCount,
        ramadanWeekdayCount,
        regularWeekdayCount,
        holidayWeekdayCount,
      }};
    }}

    function computeCapacityHoursForRange(profile, bounds) {{
      const breakdown = computeCapacityBreakdownForRange(profile, bounds);
      return toFiniteNumber(breakdown.profileCapacityHours, 0);
    }}

    async function updateScoreCards(sourceRows, requestVersion, selection) {{
      const activeRequestVersion = Number.isFinite(Number(requestVersion))
        ? Number(requestVersion)
        : ++scorecardUpdateVersion;
      const scorecardSelection = selection && typeof selection === "object"
        ? selection
        : buildScorecardSelectionSnapshot();
      const scorecardDateFrom = String(scorecardSelection.dateFrom || selectedDateFrom || DEFAULT_DATE_FROM);
      const scorecardDateTo = String(scorecardSelection.dateTo || selectedDateTo || DEFAULT_DATE_TO);
      const scorecardActualHoursMode = String(
        scorecardSelection.actualHoursMode || selectedActualHoursMode || DEFAULT_ACTUAL_HOURS_MODE
      );
      const scorecardPlannedHoursSource = normalizePlannedHoursSource(
        scorecardSelection.plannedHoursSource || selectedPlannedHoursSource || DEFAULT_PLANNED_HOURS_SOURCE
      );
      function isStaleScorecardUpdate() {{
        return activeRequestVersion !== scorecardUpdateVersion;
      }}
      const rows = Array.isArray(sourceRows) ? sourceRows : [];
      const rowsById = new Map();
      const rowsByIdText = new Map();
      for (const row of rows) {{
        rowsById.set(row.id, row);
        rowsByIdText.set(String(row && row.id || ""), row);
      }}
      function resolveProjectContext(row) {{
        let current = row || null;
        let fallbackKey = String(row && row.project_key || "").trim();
        let fallbackName = String(row && row.project_name || "").trim();
        while (current) {{
          if (String(current && current.row_type || "") === "project") {{
            return {{
              key: String(current && current.project_key || "").trim(),
              name: String(current && current.project_name || "").trim(),
            }};
          }}
          if (!fallbackKey) {{
            fallbackKey = String(current && current.project_key || "").trim();
          }}
          if (!fallbackName) {{
            fallbackName = String(current && current.project_name || "").trim();
          }}
          const parentId = current && current.parent_id;
          current = parentId ? (rowsById.get(parentId) || null) : null;
        }}
        return {{ key: fallbackKey, name: fallbackName }};
      }}
      function resolveEpicContext(row) {{
        let current = row || null;
        while (current) {{
          if (String(current && current.row_type || "") === "rmi") {{
            return {{
              id: String(current && current.id || ""),
              jira_key: String(current && current.jira_key || "").trim().toUpperCase(),
              name: String(current && current.aspect || "").trim(),
            }};
          }}
          const parentId = current && current.parent_id;
          current = parentId ? (rowsById.get(parentId) || null) : null;
        }}
        return {{ id: "", jira_key: "", name: "" }};
      }}
      function isLeaveProject(row) {{
        const projectCtx = resolveProjectContext(row);
        const projectKey = String(projectCtx && projectCtx.key || "").trim().toUpperCase();
        const projectName = String(projectCtx && projectCtx.name || "").trim().toUpperCase();
        if (projectKey === "RLT") {{
          return true;
        }}
        return projectName === "RND LEAVE TRACKER" || projectName === "RLT RND LEAVE TRACKER";
      }}
      function isExcludedPlannedProject(row) {{
        return isLeaveProject(row);
      }}
      function subtaskPlannedInRange(row, bounds) {{
        if (String(row && row.row_type || "") !== "subtask") {{
          return false;
        }}
        const plannedStart = parseDateValue(row && (row.planned_start || row.start_date));
        const plannedDue = parseDateValue(row && (row.planned_end || row.planned_due || row.due_date));
        return isDateWithinBounds(plannedStart, bounds) || isDateWithinBounds(plannedDue, bounds);
      }}
      function subtaskLoggedInRange(row) {{
        if (String(row && row.row_type || "") !== "subtask") {{
          return false;
        }}
        const jiraKey = String(row && row.jira_key || "").trim().toUpperCase();
        if (!jiraKey) {{
          return false;
        }}
        if (Object.prototype.hasOwnProperty.call(subtaskLogHoursByIssue, jiraKey)) {{
          return toFiniteNumber(subtaskLogHoursByIssue[jiraKey], 0) > 0;
        }}
        if (scorecardActualHoursMode === "log_date") {{
          return toFiniteNumber(row && row.actual_hours, 0) > 0;
        }}
        return false;
      }}
      function isBugSubtaskRow(row) {{
        const rowType = String(row && row.row_type || "");
        const typeLabel = String(row && row.type_label || "").toLowerCase();
        return rowType === "subtask" && typeLabel === "bug";
      }}
      function isBugSubtaskIssueType(issueType) {{
        const low = String(issueType || "").toLowerCase();
        return low.includes("bug") && (low.includes("subtask") || low.includes("sub-task") || low.includes("task"));
      }}
      function epicPlannedInRangeFromSubtask(row, bounds) {{
        const epicCtx = resolveEpicContext(row);
        const epicRow = rowsByIdText.get(String(epicCtx && epicCtx.id || ""));
        if (!epicRow) {{
          return false;
        }}
        const epicStart = parseDateValue(epicRow && epicRow.planned_start);
        const epicDue = parseDateValue(epicRow && epicRow.planned_end);
        return isDateWithinBounds(epicStart, bounds) || isDateWithinBounds(epicDue, bounds);
      }}
      function subtaskMatchesPlannedHoursSource(row, bounds) {{
        const source = scorecardPlannedHoursSource;
        let match = false;
        if (source === "epic_estimates") {{
          match = epicPlannedInRangeFromSubtask(row, bounds);
        }} else if (source === "subtask_logs") {{
          match = subtaskLoggedInRange(row);
        }} else {{
          match = subtaskPlannedInRange(row, bounds);
        }}
        if (!match) {{
          return false;
        }}
        if (isBugSubtaskRow(row)) {{
          return false;
        }}
        return true;
      }}
      function subtaskMatchesActualHoursMode(row, bounds) {{
        if (String(row && row.row_type || "") !== "subtask") {{
          return false;
        }}
        if (scorecardActualHoursMode === "planned_dates") {{
          if (scorecardPlannedHoursSource === "subtask_logs") {{
            return subtaskLoggedInRange(row);
          }}
          return subtaskPlannedInRange(row, bounds);
        }}
        return subtaskLoggedInRange(row);
      }}
      function subtaskPlannedHours(row) {{
        const manHours = Number(row && row.man_hours);
        const manDays = Number(row && row.man_days);
        if (Number.isFinite(manHours)) {{
          return manHours;
        }}
        if (Number.isFinite(manDays)) {{
          return manDays * 8;
        }}
        return 0;
      }}
      function rowMatchesSelectedTeams(row) {{
        if (isAllTeamsSelected()) {{
          return true;
        }}
        if (!selectedTeamNames.size) {{
          return false;
        }}
        const assigneeText = String(row && row.assignee_name || row && row.assignee || "").trim();
        if (!assigneeText) {{
          return false;
        }}
        const assignees = assigneeText.split(",").map((name) => normalizeAssigneeName(name)).filter(Boolean);
        if (!assignees.length) {{
          return false;
        }}
        for (const assignee of assignees) {{
          if (selectedTeamAssignees.has(assignee)) {{
            return true;
          }}
        }}
        return false;
      }}
      function leaveSubtaskMatchesSelectedTeams(row) {{
        if (isAllTeamsSelected()) {{
          return true;
        }}
        if (!selectedTeamNames.size) {{
          return false;
        }}
        const assignee = normalizeAssigneeName(String(row && row.assignee || "").trim());
        return !!(assignee && selectedTeamAssignees.has(assignee));
      }}
      function leaveSubtaskOverlapsBounds(row, bounds) {{
        const start = parseDateValue(row && row.start_date);
        const due = parseDateValue(row && row.due_date);
        if (start && due && bounds && bounds.start && bounds.end) {{
          return start.getTime() <= bounds.end.getTime() && due.getTime() >= bounds.start.getTime();
        }}
        return isDateWithinBounds(start, bounds) || isDateWithinBounds(due, bounds);
      }}
      function leaveJiraUrl(row) {{
        const directUrl = String(row && row.jira_url || "").trim();
        if (directUrl) {{
          return directUrl;
        }}
        const issueKey = String(row && row.issue_key || "").trim().toUpperCase();
        if (!issueKey || !jiraBaseUrl) {{
          return "";
        }}
        return jiraBaseUrl + "/browse/" + issueKey;
      }}
      let totalCapacityHours = 0;
      let totalPlannedHours = 0;
      let excludedPlannedHours = 0;
      let includedSubtaskCount = 0;
      let excludedSubtaskCount = 0;
      const plannedSubtasksByProject = new Map();
      const fallbackLeavesDetailsRowsNext = [];
      let totalActualProjectHoursFromProjects = 0;
      let totalActualProjectHoursFromFilteredSubtasks = 0;
      let filteredSubtaskActualCount = 0;
      let excludedActualHours = 0;
      let plannedLeavesTaken = 0;
      let plannedLeavesNotTakenYet = 0;
      let unplannedLeavesTaken = 0;
      let totalPlannedDetailsGroupsNext = [];
      let totalActualDetailsGroupsNext = [];
      let scopedSummaryPayload = null;
      const dateBounds = getDateFilterBoundsFor(scorecardDateFrom, scorecardDateTo);
      const actualSubtasksByProject = new Map();
      for (const row of rows) {{
        if (String(row && row.row_type || "") !== "project") {{
          continue;
        }}
        const projectManDays = Number(row && row.man_days);
        const projectManHours = Number(row && row.man_hours);
        const plannedHours = Number.isFinite(projectManHours)
          ? projectManHours
          : (Number.isFinite(projectManDays) ? (projectManDays * 8) : 0);
        const capacityHours = Number.isFinite(projectManDays)
          ? (projectManDays * 8)
          : plannedHours;
        const actualHours = Number(row && row.actual_hours);
        totalCapacityHours += capacityHours;
        if (!isLeaveProject(row)) {{
          if (Number.isFinite(actualHours)) {{
            totalActualProjectHoursFromProjects += actualHours;
          }}
        }} else {{
          const actualLeaveHours = Number.isFinite(actualHours) ? actualHours : 0;
          excludedActualHours += actualLeaveHours;
          const cappedPlannedTaken = Math.max(0, Math.min(actualLeaveHours, plannedHours));
          plannedLeavesTaken += cappedPlannedTaken;
          plannedLeavesNotTakenYet += Math.max(0, plannedHours - cappedPlannedTaken);
          unplannedLeavesTaken += Math.max(0, actualLeaveHours - plannedHours);
        }}
      }}
      for (const row of rows) {{
        if (!subtaskMatchesPlannedHoursSource(row, dateBounds)) {{
          continue;
        }}
        if (!rowMatchesSelectedTeams(row)) {{
          continue;
        }}
        if (isBugSubtaskRow(row)) {{
          continue;
        }}
        const plannedHours = subtaskPlannedHours(row);
        if (isExcludedPlannedProject(row)) {{
          excludedPlannedHours += plannedHours;
          excludedSubtaskCount += 1;
          const plannedStartDate = parseDateValue(row && (row.planned_start || row.start_date));
          fallbackLeavesDetailsRowsNext.push({{
            date: plannedStartDate ? toIsoDate(plannedStartDate) : "",
            assignee_name: String(row && row.assignee_name || row && row.assignee || "").trim(),
            hours: plannedHours,
            jira_key: String(row && row.jira_key || "").trim(),
            jira_url: String(row && row.jira_url || "").trim(),
          }});
        }} else {{
          totalPlannedHours += plannedHours;
          includedSubtaskCount += 1;
          const projectCtx = resolveProjectContext(row);
          const projectKey = String(projectCtx && projectCtx.key || "").trim().toUpperCase();
          const projectName = String(projectCtx && projectCtx.name || "").trim() || projectKey || "Unknown Project";
          if (!plannedSubtasksByProject.has(projectKey)) {{
            plannedSubtasksByProject.set(projectKey, {{
              project_key: projectKey,
              project_name: projectName,
              subtasks: [],
            }});
          }}
          const plannedStart = parseDateValue(row && (row.planned_start || row.start_date));
          const plannedDue = parseDateValue(row && (row.planned_end || row.planned_due || row.due_date));
          plannedSubtasksByProject.get(projectKey).subtasks.push({{
            subtask_jira_id: String(row && row.jira_key || "").trim().toUpperCase(),
            subtask_name: String(row && row.aspect || "").trim(),
            task_type: "Subtask",
            subtask_start_date: plannedStart ? toIsoDate(plannedStart) : "",
            subtask_due_date: plannedDue ? toIsoDate(plannedDue) : "",
            planned_hours: roundHours(plannedHours),
          }});
        }}
      }}
      for (const row of rows) {{
        if (!subtaskMatchesActualHoursMode(row, dateBounds)) {{
          continue;
        }}
        if (!rowMatchesSelectedTeams(row)) {{
          continue;
        }}
        if (isExcludedPlannedProject(row)) {{
          continue;
        }}
        if (!totalActualIncludeBugs && isBugSubtaskRow(row)) {{
          continue;
        }}
        const subtaskActualHours = Number(row && row.actual_hours);
        if (Number.isFinite(subtaskActualHours)) {{
          totalActualProjectHoursFromFilteredSubtasks += subtaskActualHours;
          filteredSubtaskActualCount += 1;
          const projectCtx = resolveProjectContext(row);
          const projectKey = String(projectCtx && projectCtx.key || "").trim().toUpperCase();
          const projectName = String(projectCtx && projectCtx.name || "").trim() || projectKey || "Unknown Project";
          if (!actualSubtasksByProject.has(projectKey)) {{
            actualSubtasksByProject.set(projectKey, {{
              project_key: projectKey,
              project_name: projectName,
              subtasks: [],
            }});
          }}
          const plannedStart = parseDateValue(row && (row.planned_start || row.start_date));
          const plannedDue = parseDateValue(row && (row.planned_end || row.planned_due || row.due_date));
          actualSubtasksByProject.get(projectKey).subtasks.push({{
            subtask_jira_id: String(row && row.jira_key || "").trim().toUpperCase(),
            subtask_name: String(row && row.aspect || "").trim(),
            task_type: isBugSubtaskRow(row) ? "Bug Subtask" : "Subtask",
            subtask_start_date: plannedStart ? toIsoDate(plannedStart) : "",
            subtask_due_date: plannedDue ? toIsoDate(plannedDue) : "",
            actual_hours: roundHours(subtaskActualHours),
          }});
        }}
      }}
      if (scorecardPlannedHoursSource === "subtask_estimates") {{
        try {{
          scopedSummaryPayload = await fetchScopedSubtasksSummary(scorecardSelection);
          if (isStaleScorecardUpdate()) {{
            return;
          }}
          const scopedRows = Array.isArray(scopedSummaryPayload && scopedSummaryPayload.rows) ? scopedSummaryPayload.rows : [];
          if (scopedSummaryPayload && typeof scopedSummaryPayload === "object") {{
            const scopedPlannedSubtasks = scopedRows.filter((row) => !isBugSubtaskIssueType(row && row.issue_type));
            totalPlannedHours = roundHours(scopedPlannedSubtasks.reduce((sum, row) => (
              sum + toFiniteNumber(row && row.original_estimate_hours, 0)
            ), 0));
            const scopedActualSubtasks = scopedRows.filter((row) => (
              toFiniteNumber(row && row.logged_hours, 0) > 0
              && (totalActualIncludeBugs || !isBugSubtaskIssueType(row && row.issue_type))
            ));
            totalActualProjectHoursFromFilteredSubtasks = roundHours(scopedActualSubtasks.reduce((sum, row) => (
              sum + toFiniteNumber(row && row.logged_hours, 0)
            ), 0));
            filteredSubtaskActualCount = scopedActualSubtasks.length;
          }}
          if (Array.isArray(scopedRows) && scopedRows.length) {{
            const scopedSubtasksByProject = new Map();
            const scopedActualSubtasksByProject = new Map();
            for (const row of scopedRows) {{
              const issueKey = String(row && row.issue_key || "").trim().toUpperCase();
              const projectKey = String(row && row.project_key || "").trim().toUpperCase();
              const projectName = projectKey;
              const issueMeta = allRows.find((item) => String(item && item.jira_key || "").trim().toUpperCase() === issueKey) || null;
              const projectCtx = issueMeta ? resolveProjectContext(issueMeta) : null;
              const resolvedProjectName = String(projectCtx && projectCtx.name || "").trim() || projectName || projectKey || "Unknown Project";
              const plannedStart = parseDateValue(row && row.start_date);
              const plannedDue = parseDateValue(row && row.due_date);
              const loggedHours = roundHours(toFiniteNumber(row && row.logged_hours, 0));
              const isBugRow = isBugSubtaskIssueType(row && row.issue_type);
              if (loggedHours > 0) {{
                if (!totalActualIncludeBugs && isBugRow) {{
                  // Skip bug subtasks from Actual Hours scorecard when Include Bugs is OFF.
                }} else {{
                if (!scopedActualSubtasksByProject.has(projectKey)) {{
                  scopedActualSubtasksByProject.set(projectKey, {{
                    project_key: projectKey,
                    project_name: resolvedProjectName,
                    subtasks: [],
                  }});
                }}
                scopedActualSubtasksByProject.get(projectKey).subtasks.push({{
                  subtask_jira_id: issueKey,
                  subtask_name: String(issueMeta && issueMeta.aspect || "").trim(),
                  task_type: isBugRow ? "Bug Subtask" : "Subtask",
                  subtask_start_date: plannedStart ? toIsoDate(plannedStart) : "",
                  subtask_due_date: plannedDue ? toIsoDate(plannedDue) : "",
                  actual_hours: loggedHours,
                }});
                }}
              }}
              if (isBugSubtaskIssueType(row && row.issue_type)) {{
                continue;
              }}
              if (!scopedSubtasksByProject.has(projectKey)) {{
                scopedSubtasksByProject.set(projectKey, {{
                  project_key: projectKey,
                  project_name: resolvedProjectName,
                  subtasks: [],
                }});
              }}
              scopedSubtasksByProject.get(projectKey).subtasks.push({{
                subtask_jira_id: issueKey,
                subtask_name: String(issueMeta && issueMeta.aspect || "").trim(),
                task_type: "Subtask",
                subtask_start_date: plannedStart ? toIsoDate(plannedStart) : "",
                subtask_due_date: plannedDue ? toIsoDate(plannedDue) : "",
                planned_hours: roundHours(toFiniteNumber(row && row.original_estimate_hours, 0)),
              }});
            }}
            totalPlannedDetailsGroupsNext.length = 0;
            for (const group of Array.from(scopedSubtasksByProject.values())) {{
              totalPlannedDetailsGroupsNext.push(group);
            }}
            totalActualDetailsGroupsNext.length = 0;
            for (const group of Array.from(scopedActualSubtasksByProject.values())) {{
              totalActualDetailsGroupsNext.push(group);
            }}
          }}
        }} catch (error) {{
          if (isStaleScorecardUpdate()) {{
            return;
          }}
          console.warn("Failed to align Nested View scorecards with shared scoped subtasks:", error);
        }}
      }}
      if (isStaleScorecardUpdate()) {{
        return;
      }}
      const extendedActualsEnabled = scorecardActualHoursMode === "planned_dates";
      const totalActualProjectHours = (filteredSubtaskActualCount > 0 || hasNestedActualsApi)
        ? totalActualProjectHoursFromFilteredSubtasks
        : totalActualProjectHoursFromProjects;
      const groupedPlannedSubtasks = new Map(plannedSubtasksByProject);
      const groupedActualSubtasks = new Map(actualSubtasksByProject);
      totalPlannedDetailsGroupsNext = totalPlannedDetailsGroupsNext.length
        ? totalPlannedDetailsGroupsNext
        : Array.from(groupedPlannedSubtasks.values());
      totalActualDetailsGroupsNext = totalActualDetailsGroupsNext.length
        ? totalActualDetailsGroupsNext
        : Array.from(groupedActualSubtasks.values());
      totalPlannedDetailsGroupsNext.sort((left, right) =>
        String(left && left.project_name || "").localeCompare(String(right && right.project_name || ""))
      );
      totalActualDetailsGroupsNext.sort((left, right) =>
        String(left && left.project_name || "").localeCompare(String(right && right.project_name || ""))
      );
      for (const group of totalPlannedDetailsGroupsNext) {{
        const subtasks = Array.isArray(group && group.subtasks) ? group.subtasks : [];
        subtasks.sort((left, right) => {{
          const diff = toFiniteNumber(right && right.planned_hours, 0) - toFiniteNumber(left && left.planned_hours, 0);
          if (Math.abs(diff) > 1e-9) {{
            return diff;
          }}
          return String(left && left.subtask_jira_id || "").localeCompare(String(right && right.subtask_jira_id || ""));
        }});
      }}
      for (const group of totalActualDetailsGroupsNext) {{
        const subtasks = Array.isArray(group && group.subtasks) ? group.subtasks : [];
        subtasks.sort((left, right) => {{
          const diff = toFiniteNumber(right && right.actual_hours, 0) - toFiniteNumber(left && left.actual_hours, 0);
          if (Math.abs(diff) > 1e-9) {{
            return diff;
          }}
          return String(left && left.subtask_jira_id || "").localeCompare(String(right && right.subtask_jira_id || ""));
        }});
      }}
      const validProjectKeys = new Set(totalPlannedDetailsGroupsNext.map((group) => String(group && group.project_key || "")));
      for (const key of Array.from(collapsedTotalPlannedProjectKeys)) {{
        if (!validProjectKeys.has(key)) {{
          collapsedTotalPlannedProjectKeys.delete(key);
        }}
      }}
      for (const key of validProjectKeys) {{
        collapsedTotalPlannedProjectKeys.add(key);
      }}
      const validActualProjectKeys = new Set(totalActualDetailsGroupsNext.map((group) => String(group && group.project_key || "")));
      for (const key of Array.from(collapsedTotalActualProjectKeys)) {{
        if (!validActualProjectKeys.has(key)) {{
          collapsedTotalActualProjectKeys.delete(key);
        }}
      }}
      for (const key of validActualProjectKeys) {{
        collapsedTotalActualProjectKeys.add(key);
      }}
      totalPlannedDetailsGroups = totalPlannedDetailsGroupsNext;
      totalActualDetailsGroups = totalActualDetailsGroupsNext;
      renderTotalPlannedDetails();
      renderTotalActualDetails();
      function computeEmbeddedLeaveMetricsForRange(bounds) {{
        const empty = {{
          hasData: false,
          plannedTakenHours: 0,
          plannedNotTakenHours: 0,
          unplannedTakenHours: 0,
        }};
        if (!bounds || !bounds.start || !bounds.end || !leaveDailyRows.length) {{
          return empty;
        }}
        let plannedTakenHours = 0;
        let plannedNotTakenHours = 0;
        let unplannedTakenHours = 0;
        let hasData = false;
        for (const row of leaveDailyRows) {{
          const day = parseDateValue(row && row.period_day);
          if (!day || !isDateWithinBounds(day, bounds)) {{
            continue;
          }}
          hasData = true;
          plannedTakenHours += toFiniteNumber(row && row.planned_taken_hours, 0);
          plannedNotTakenHours += toFiniteNumber(row && row.planned_not_taken_hours, 0);
          unplannedTakenHours += toFiniteNumber(row && row.unplanned_taken_hours, 0);
        }}
        return {{
          hasData,
          plannedTakenHours,
          plannedNotTakenHours,
          unplannedTakenHours,
        }};
      }}
      const embeddedLeaveMetrics = computeEmbeddedLeaveMetricsForRange(dateBounds);
      const dailyPlannedDetailsRowsNext = [];
      if (dateBounds && dateBounds.start && dateBounds.end && leaveDailyRows.length) {{
        const byDayAssignee = new Map();
        for (const row of leaveDailyRows) {{
          const day = parseDateValue(row && row.period_day);
          if (!day || !isDateWithinBounds(day, dateBounds)) {{
            continue;
          }}
          const plannedTaken = toFiniteNumber(row && row.planned_taken_hours, 0);
          const plannedNotTaken = toFiniteNumber(row && row.planned_not_taken_hours, 0);
          const plannedTotal = plannedTaken + plannedNotTaken;
          if (plannedTotal <= 0) {{
            continue;
          }}
          const isoDay = toIsoDate(day);
          const assigneeName = String(row && row.assignee || "").trim();
          const groupKey = isoDay + "|" + assigneeName.toLowerCase();
          const jiraTaskIds = String(row && row.jira_task_ids || "").trim();
          const jiraTaskLinks = String(row && row.jira_task_links || "").trim();
          const existing = byDayAssignee.get(groupKey) || {{
            date: isoDay,
            assignee_name: assigneeName,
            hours: 0,
            jira_key: "",
            jira_url: "",
            _jiraKeySet: new Set(),
            _jiraUrlSet: new Set(),
          }};
          existing.hours += plannedTotal;
          if (jiraTaskIds) {{
            for (const key of jiraTaskIds.split(",")) {{
              const keyText = String(key || "").trim().toUpperCase();
              if (keyText) {{
                existing._jiraKeySet.add(keyText);
              }}
            }}
          }}
          if (jiraTaskLinks) {{
            for (const linkTextRaw of jiraTaskLinks.split(",")) {{
              const linkText = String(linkTextRaw || "").trim();
              if (linkText) {{
                existing._jiraUrlSet.add(linkText);
              }}
            }}
          }}
          existing.jira_key = Array.from(existing._jiraKeySet).sort().join(", ");
          existing.jira_url = Array.from(existing._jiraUrlSet).sort().join(", ");
          byDayAssignee.set(groupKey, existing);
        }}
        for (const item of byDayAssignee.values()) {{
          delete item._jiraKeySet;
          delete item._jiraUrlSet;
          dailyPlannedDetailsRowsNext.push(item);
        }}
      }}
      const totalLeavesPlannedDetailsRowsNext = [];
      for (const leaveRow of leaveSubtaskRows) {{
        if (!leaveSubtaskOverlapsBounds(leaveRow, dateBounds)) {{
          continue;
        }}
        if (!leaveSubtaskMatchesSelectedTeams(leaveRow)) {{
          continue;
        }}
        const issueKey = String(leaveRow && leaveRow.issue_key || "").trim().toUpperCase();
        totalLeavesPlannedDetailsRowsNext.push({{
          date: String(leaveRow && leaveRow.start_date || "").trim(),
          assignee_name: String(leaveRow && leaveRow.assignee || "").trim(),
          hours: toFiniteNumber(leaveRow && leaveRow.original_estimate_hours, 0),
          jira_key: issueKey,
          jira_url: leaveJiraUrl(leaveRow),
        }});
      }}
      const detailsRowsForRender = dailyPlannedDetailsRowsNext.length
        ? dailyPlannedDetailsRowsNext
        : (totalLeavesPlannedDetailsRowsNext.length ? totalLeavesPlannedDetailsRowsNext : fallbackLeavesDetailsRowsNext);
      detailsRowsForRender.sort((left, right) => {{
        const lDate = String(left && left.date || "");
        const rDate = String(right && right.date || "");
        if (lDate && rDate && lDate !== rDate) {{
          return lDate.localeCompare(rDate);
        }}
        if (lDate && !rDate) {{
          return -1;
        }}
        if (!lDate && rDate) {{
          return 1;
        }}
        const lAssignee = String(left && left.assignee_name || "");
        const rAssignee = String(right && right.assignee_name || "");
        if (lAssignee !== rAssignee) {{
          return lAssignee.localeCompare(rAssignee);
        }}
        const lJira = String(left && left.jira_key || "");
        const rJira = String(right && right.jira_key || "");
        return lJira.localeCompare(rJira);
      }});
      totalLeavesPlannedDetailsRows = detailsRowsForRender;
      renderTotalLeavesPlannedDetails();
      if (embeddedLeaveMetrics.hasData) {{
        plannedLeavesTaken = embeddedLeaveMetrics.plannedTakenHours;
        plannedLeavesNotTakenYet = embeddedLeaveMetrics.plannedNotTakenHours;
        unplannedLeavesTaken = embeddedLeaveMetrics.unplannedTakenHours;
      }}
      const totalLeavesTakenHours = plannedLeavesTaken + unplannedLeavesTaken;
      const deltaHours = totalPlannedHours - totalActualProjectHours;
      const capacityBreakdown = appliedCapacityProfile
        ? computeCapacityBreakdownForRange(appliedCapacityProfile, dateBounds)
        : null;
      const profileCapacityHours = appliedCapacityProfile
        ? toFiniteNumber(capacityBreakdown && capacityBreakdown.profileCapacityHours, 0)
        : NaN;
      const totalCapacityHoursValue = Number.isFinite(profileCapacityHours)
        ? profileCapacityHours
        : totalCapacityHours;
      const capacityEmployeeCount = appliedCapacityProfile
        ? Math.round(toFiniteNumber(capacityBreakdown && capacityBreakdown.employees, 0))
        : null;
      const capacityBusinessDays = appliedCapacityProfile
        ? Math.round(toFiniteNumber(capacityBreakdown && capacityBreakdown.weekdayCount, 0))
        : null;
      const capacityPerDayHours = appliedCapacityProfile
        && Number.isFinite(profileCapacityHours)
        && Number(capacityEmployeeCount) > 0
        && Number(capacityBusinessDays) > 0
        ? (profileCapacityHours / (Number(capacityEmployeeCount) * Number(capacityBusinessDays)))
        : null;
      const totalPlannedLeavesHours = plannedLeavesTaken + plannedLeavesNotTakenYet;
      const totalCapacityPlannedLeavesAdjustedHoursDefault = totalCapacityHoursValue - totalPlannedLeavesHours;
      const capacityGapHoursDefault = totalCapacityHoursValue - totalPlannedHours - totalPlannedLeavesHours;
      const hoursRequiredToCompleteProjectsDefault = deltaHours;
      const scorecardFormulaContext = {{
        "capacity": totalCapacityHoursValue,
        "planned_hours": totalPlannedHours,
        "actual_hours": totalActualProjectHours,
        "planned_leaves": totalPlannedLeavesHours,
      }};
      const availabilityEval = evaluateManagedField("availability", totalCapacityPlannedLeavesAdjustedHoursDefault, scorecardFormulaContext);
      const capacityMoreWorkEval = evaluateManagedField("capacity_available_for_more_work", capacityGapHoursDefault, scorecardFormulaContext);
      const hoursRequiredEval = evaluateManagedField("hours_required_to_complete_projects", hoursRequiredToCompleteProjectsDefault, scorecardFormulaContext);
      const totalCapacityPlannedLeavesAdjustedHours = availabilityEval.value;
      const capacityGapHours = capacityMoreWorkEval.value;
      const hoursRequiredToCompleteProjectsHours = hoursRequiredEval.value;
      const loadingEfficiencyPercent = totalCapacityPlannedLeavesAdjustedHours === 0
        ? 0
        : ((totalPlannedHours / totalCapacityPlannedLeavesAdjustedHours) * 100);
      const deliveryEfficiencyPercent = totalActualProjectHours === 0
        ? 0
        : ((totalPlannedHours / totalActualProjectHours) * 100);
      totalCapacityScoreNode.textContent = formatHours(totalCapacityHoursValue);
      totalPlannedScoreNode.textContent = formatHours(totalPlannedHours);
      totalLoggedScoreNode.textContent = formatHours(totalActualProjectHours);
      if (deltaScoreNode) {{
        deltaScoreNode.textContent = formatHours(hoursRequiredToCompleteProjectsHours);
      }}
      if (totalLeavesScoreNode) {{
        totalLeavesScoreNode.textContent = formatHours(totalLeavesTakenHours);
      }}
      totalLeavesPlannedScoreNode.textContent = formatHours(totalPlannedLeavesHours);
      totalCapacityPlannedLeavesAdjustedScoreNode.textContent = formatHours(totalCapacityPlannedLeavesAdjustedHours);
      if (loadingEfficiencyScoreNode) {{
        loadingEfficiencyScoreNode.textContent = formatPercent(loadingEfficiencyPercent);
      }}
      if (deliveryEfficiencyScoreNode) {{
        deliveryEfficiencyScoreNode.textContent = formatPercent(deliveryEfficiencyPercent);
      }}
      if (capacityGapScoreNode) {{
        capacityGapScoreNode.textContent = formatHours(capacityGapHours);
      }}
      if (availabilityFormulaNode) {{
        availabilityFormulaNode.textContent = managedFieldFormulaText("availability", "Total Capacity - Total Leaves Planned");
      }}
      if (totalCapacityFormulaEmployeeNode) {{
        totalCapacityFormulaEmployeeNode.textContent = capacityEmployeeCount === null ? "-" : String(capacityEmployeeCount);
      }}
      if (totalCapacityFormulaDaysNode) {{
        totalCapacityFormulaDaysNode.textContent = capacityBusinessDays === null ? "-" : String(capacityBusinessDays);
      }}
      if (totalCapacityFormulaHoursNode) {{
        totalCapacityFormulaHoursNode.textContent = capacityPerDayHours === null ? "-" : formatHours(capacityPerDayHours);
      }}
      if (totalCapacityTipNode) {{
        const selectedProfileLabel = appliedCapacityProfile
          ? (String(appliedCapacityProfile.name || "").trim() || (String(appliedCapacityProfile.from_date || "") + " to " + String(appliedCapacityProfile.to_date || "")))
          : "None";
        const fromDate = dateBounds && dateBounds.start ? toIsoDate(dateBounds.start) : scorecardDateFrom;
        const toDate = dateBounds && dateBounds.end ? toIsoDate(dateBounds.end) : scorecardDateTo;
        const profileEmployees = toFiniteNumber(capacityBreakdown && capacityBreakdown.employees, 0);
        const profileWeekdays = toFiniteNumber(capacityBreakdown && capacityBreakdown.weekdayCount, 0);
        const perDayHours = Number.isFinite(profileCapacityHours) && profileEmployees > 0 && profileWeekdays > 0
          ? (profileCapacityHours / (profileEmployees * profileWeekdays))
          : 0;
        totalCapacityTipNode.textContent =
          "Formula: Total Capacity (Hours) = Employee Count x Available Business Days x Per Day Hours.\\n"
          + "Values:\\n"
          + "Selected Profile = " + selectedProfileLabel + "\\n"
          + "Date Range = " + fromDate + " to " + toDate + "\\n"
          + "Employee Count = " + String(profileEmployees) + "\\n"
          + "Per Day Hours = " + formatHours(perDayHours) + "\\n"
          + "Standard Hours/Day = " + formatHours(toFiniteNumber(capacityBreakdown && capacityBreakdown.standardHours, 0)) + "\\n"
          + "Ramadan Hours/Day = " + formatHours(toFiniteNumber(capacityBreakdown && capacityBreakdown.ramadanHours, 0)) + "\\n"
          + "Available Business Days (Mon-Fri, non-holiday) = " + String(profileWeekdays) + "\\n"
          + "Ramadan Weekdays (Mon-Fri, non-holiday) = " + String(toFiniteNumber(capacityBreakdown && capacityBreakdown.ramadanWeekdayCount, 0)) + "\\n"
          + "Non-Ramadan Weekdays (Mon-Fri, non-holiday) = " + String(toFiniteNumber(capacityBreakdown && capacityBreakdown.regularWeekdayCount, 0)) + "\\n"
          + "Holiday Weekdays in Range = " + String(toFiniteNumber(capacityBreakdown && capacityBreakdown.holidayWeekdayCount, 0)) + "\\n"
          + "Capacity Profile Hours = " + (Number.isFinite(profileCapacityHours) ? formatHours(profileCapacityHours) : "N/A") + "\\n"
          + "Fallback Project Capacity = " + formatHours(totalCapacityHours) + "\\n"
          + "Total Capacity = " + formatHours(totalCapacityHoursValue);
      }}
      if (totalPlannedTipNode) {{
        totalPlannedTipNode.textContent =
          "Formula: Total Planned Hours = Sum(Subtask Planned Hours) where subtask Start OR Due date is within selected range, excluding RLT (RnD Leave Tracker).\\n"
          + "Values:\\n"
          + "Included Subtasks Count = " + String(includedSubtaskCount) + "\\n"
          + "Excluded (RLT) Subtasks Count = " + String(excludedSubtaskCount) + "\\n"
          + "Excluded (RLT) Subtasks Planned Sum = " + formatHours(excludedPlannedHours) + "\\n"
          + "Total Planned Hours = " + formatHours(totalPlannedHours) + "\\n"
          + "\\nSQL example (subtasks only; adjust dates to match this report):\\n"
          + "WITH latest_run AS (\\n"
          + "  SELECT run_id\\n"
          + "  FROM canonical_refresh_runs\\n"
          + "  WHERE status = 'success'\\n"
          + "  ORDER BY updated_at_utc DESC\\n"
          + "  LIMIT 1\\n"
          + "), scoped_subtasks AS (\\n"
          + "  SELECT\\n"
          + "    ci.issue_key,\\n"
          + "    ci.issue_type,\\n"
          + "    COALESCE(ci.original_estimate_hours, 0) AS planned_hours\\n"
          + "  FROM canonical_issues ci\\n"
          + "  WHERE ci.run_id = (SELECT run_id FROM latest_run)\\n"
          + "    AND ci.project_key <> 'RLT'\\n"
          + "    AND (\\n"
          + "      LOWER(ci.issue_type) LIKE '%sub-task%'\\n"
          + "      OR LOWER(ci.issue_type) LIKE '%subtask%'\\n"
          + "    )\\n"
          + "    AND (\\n"
          + "      (ci.start_date >= '2026-02-01' AND ci.start_date <= '2026-02-28')\\n"
          + "      OR\\n"
          + "      (ci.due_date   >= '2026-02-01' AND ci.due_date   <= '2026-02-28')\\n"
          + "    )\\n"
          + "), worklog_totals AS (\\n"
          + "  SELECT\\n"
          + "    cw.issue_key,\\n"
          + "    ROUND(SUM(COALESCE(cw.hours_logged, 0)), 2) AS logged_hours\\n"
          + "  FROM canonical_worklogs cw\\n"
          + "  WHERE cw.run_id = (SELECT run_id FROM latest_run)\\n"
          + "    AND cw.started_date >= '2026-02-01'\\n"
          + "    AND cw.started_date <= '2026-02-28'\\n"
          + "  GROUP BY cw.issue_key\\n"
          + ")\\n"
          + "SELECT\\n"
          + "  ss.issue_key                          AS subtask_id,\\n"
          + "  ss.issue_type                         AS work_item_type,\\n"
          + "  ROUND(ss.planned_hours, 2)            AS original_estimate_hours,\\n"
          + "  COALESCE(wt.logged_hours, 0)          AS logged_hours\\n"
          + "FROM scoped_subtasks ss\\n"
          + "LEFT JOIN worklog_totals wt\\n"
          + "  ON wt.issue_key = ss.issue_key\\n"
          + "ORDER BY ss.issue_key;\\n";
      }}
      if (totalLoggedTipNode) {{
        if (extendedActualsEnabled) {{
          totalLoggedTipNode.textContent =
            "Formula: Total Actual Hours = Sum(All Logged Hours for subtasks whose planned Start OR Due date is within selected range), excluding RLT (RnD Leave Tracker).\\n"
            + "Values:\\n"
            + "Included Subtasks Actual Sum = " + formatHours(totalActualProjectHoursFromFilteredSubtasks) + "\\n"
            + "Included Subtasks Count = " + String(filteredSubtaskActualCount) + "\\n"
            + "Excluded Projects Actual Sum = " + formatHours(excludedActualHours) + "\\n"
            + "Total Actual Hours = " + formatHours(totalActualProjectHours);
        }} else {{
          totalLoggedTipNode.textContent =
            "Formula: Total Actual Hours = Sum(Logged Hours in selected date range for subtasks with worklog dates in selected range), excluding RLT (RnD Leave Tracker).\\n"
            + "Values:\\n"
            + "Included Subtasks Actual Sum = " + formatHours(totalActualProjectHoursFromFilteredSubtasks) + "\\n"
            + "Included Subtasks Count = " + String(filteredSubtaskActualCount) + "\\n"
            + "Excluded Projects Actual Sum = " + formatHours(excludedActualHours) + "\\n"
            + "Total Actual Hours = " + formatHours(totalActualProjectHours);
        }}
      }}
      if (deltaTipNode) {{
        const deltaFormulaText = managedFieldFormulaText(
          "hours_required_to_complete_projects",
          "Total Planned Hours - Total Actual Hours"
        );
        deltaTipNode.textContent =
          "Formula: Hours Required To Complete Projects = " + deltaFormulaText + ".\\n"
          + "Values:\\n"
          + "Total Planned Hours = " + formatHours(totalPlannedHours) + "\\n"
          + "Total Actual Hours = " + formatHours(totalActualProjectHours) + "\\n"
          + "Hours Required To Complete Projects = " + formatHours(hoursRequiredToCompleteProjectsHours);
      }}
      if (totalLeavesTipNode) {{
        totalLeavesTipNode.textContent =
          "Formula: Total Leaves Taken = Sum(Logged Hours for RLT RnD Leave Tracker leave work).\\n"
          + "Values:\\n"
          + "Planned Leaves Taken = " + formatHours(plannedLeavesTaken) + "\\n"
          + "Unplanned Leaves Taken = " + formatHours(unplannedLeavesTaken) + "\\n"
          + "Total Leaves Taken = " + formatHours(totalLeavesTakenHours);
      }}
      if (totalLeavesPlannedTipNode) {{
        const scoreRangeFrom = dateBounds && dateBounds.start ? toIsoDate(dateBounds.start) : "-";
        const scoreRangeTo = dateBounds && dateBounds.end ? toIsoDate(dateBounds.end) : "-";
        totalLeavesPlannedTipNode.textContent =
          "Formula: Total Leaves Planned = Planned Taken + Planned Not Yet Taken from day-bucketed leave rows within selected date range.\\n"
          + "Values:\\n"
          + "Capacity Profile Effect = None (independent of selected profile)\\n"
          + "Date Range = " + scoreRangeFrom + " to " + scoreRangeTo + "\\n"
          + "Planned Taken = " + formatHours(plannedLeavesTaken) + "\\n"
          + "Planned Not Yet Taken = " + formatHours(plannedLeavesNotTakenYet) + "\\n"
          + "Total Leaves Planned = " + formatHours(totalPlannedLeavesHours);
      }}
      if (totalCapacityPlannedLeavesAdjustedTipNode) {{
        const availabilityFormulaText = managedFieldFormulaText("availability", "capacity-planned_leaves");
        totalCapacityPlannedLeavesAdjustedTipNode.textContent =
          "Formula: Availability = " + availabilityFormulaText + ".\\n"
          + "Values:\\n"
          + "Total Capacity = " + formatHours(totalCapacityHoursValue) + "\\n"
          + "Total Leaves Planned = " + formatHours(totalPlannedLeavesHours) + "\\n"
          + "Availability = " + formatHours(totalCapacityPlannedLeavesAdjustedHours);
      }}
      if (loadingEfficiencyTipNode) {{
        loadingEfficiencyTipNode.textContent =
          "Formula: Planning Efficiency = Total Planned Hours / Availability x 100.\\n"
          + "Values:\\n"
          + "Total Planned Hours = " + formatHours(totalPlannedHours) + "\\n"
          + "Availability = " + formatHours(totalCapacityPlannedLeavesAdjustedHours) + "\\n"
          + "Planning Efficiency = " + formatPercent(loadingEfficiencyPercent);
      }}
      if (deliveryEfficiencyTipNode) {{
        deliveryEfficiencyTipNode.textContent =
          "Formula: Delivery Efficiency = Total Planned Hours / Total Actual Hours x 100.\\n"
          + "Values:\\n"
          + "Total Planned Hours = " + formatHours(totalPlannedHours) + "\\n"
          + "Total Actual Hours = " + formatHours(totalActualProjectHours) + "\\n"
          + "Delivery Efficiency = " + formatPercent(deliveryEfficiencyPercent);
      }}
      if (capacityGapTipNode) {{
        const capacityGapFormulaText = managedFieldFormulaText(
          "capacity_available_for_more_work",
          "capacity-planned_hours-planned_leaves"
        );
        capacityGapTipNode.textContent =
          "Formula: Capacity available for more work = " + capacityGapFormulaText + ".\\n"
          + "Values:\\n"
          + "Total Capacity = " + formatHours(totalCapacityHoursValue) + "\\n"
          + "Total Planned Hours = " + formatHours(totalPlannedHours) + "\\n"
          + "Total Leaves Planned = " + formatHours(totalPlannedLeavesHours) + "\\n"
          + "Capacity available for more work = " + formatHours(capacityGapHours);
      }}
      if (deltaScoreCard) {{
        deltaScoreCard.classList.remove("delta-pos", "delta-neg", "delta-zero");
        if (hoursRequiredToCompleteProjectsHours > 0) {{
          deltaScoreCard.classList.add("delta-pos");
        }} else if (hoursRequiredToCompleteProjectsHours < 0) {{
          deltaScoreCard.classList.add("delta-neg");
        }} else {{
          deltaScoreCard.classList.add("delta-zero");
        }}
      }}
      if (activeRequestVersion === projectFilterLoadingVersion) {{
        setProjectFilterLoading(false);
      }}
      if (activeRequestVersion === teamFilterLoadingVersion) {{
        setTeamFilterLoading(false);
      }}
    }}

    function applySelectedCapacityProfile() {{
      const profile = getSelectedCapacityProfile();
      if (!profile) {{
        renderCapacityProfileDetails();
        return;
      }}
      appliedCapacityProfile = profile;
      appliedCapacityProfileKey = capacityProfileKey(profile);
      const requestVersion = ++scorecardUpdateVersion;
      if (projectFilterProgress && !projectFilterProgress.hidden) {{
        setProjectFilterLoading(true, requestVersion);
      }}
      if (teamFilterProgress && !teamFilterProgress.hidden) {{
        setTeamFilterLoading(true, requestVersion);
      }}
      void updateScoreCards(scorecardSourceRows, requestVersion, buildScorecardSelectionSnapshot());
      renderCapacityProfileDetails();
    }}

    function resetCapacityProfileOverride() {{
      appliedCapacityProfile = null;
      appliedCapacityProfileKey = "";
      const requestVersion = ++scorecardUpdateVersion;
      if (projectFilterProgress && !projectFilterProgress.hidden) {{
        setProjectFilterLoading(true, requestVersion);
      }}
      if (teamFilterProgress && !teamFilterProgress.hidden) {{
        setTeamFilterLoading(true, requestVersion);
      }}
      void updateScoreCards(scorecardSourceRows, requestVersion, buildScorecardSelectionSnapshot());
      renderCapacityProfileDetails();
      if (capacityProfileDetailsEl) {{
        capacityProfileDetailsEl.textContent += " | Using project totals.";
      }}
    }}

    function setCapacityProfileDrawerOpen(isOpen) {{
      const open = Boolean(isOpen);
      if (!open) {{
        stopCapacityDrawerResize();
      }}
      if (capacityProfileDrawerEl) {{
        capacityProfileDrawerEl.classList.toggle("is-open", open);
      }}
      if (capacityProfileOverlayEl) {{
        capacityProfileOverlayEl.classList.toggle("is-open", open);
      }}
      document.body.classList.toggle("capacity-modal-open", open);
      if (capacityProfileToggleEl) {{
        capacityProfileToggleEl.setAttribute("aria-expanded", open ? "true" : "false");
        capacityProfileToggleEl.textContent = open ? "Hide Capacity Profile" : "Show Capacity Profile";
      }}
    }}

    function setHeaderCollapsed(isCollapsed) {{
      const collapsedState = Boolean(isCollapsed);
      if (headerSectionEl) {{
        headerSectionEl.classList.toggle("is-collapsed", collapsedState);
      }}
      if (headerToggleButton) {{
        headerToggleButton.setAttribute("aria-expanded", collapsedState ? "false" : "true");
        headerToggleButton.textContent = collapsedState ? "Expand Header" : "Collapse Header";
      }}
      localStorage.setItem(HEADER_COLLAPSED_STORAGE_KEY, collapsedState ? "1" : "0");
    }}

    function clampAspectWidth(value) {{
      return Math.min(MAX_ASPECT_COL_WIDTH, Math.max(MIN_ASPECT_COL_WIDTH, value));
    }}

    function clampCapacityDrawerWidthVw(value) {{
      const numeric = Number(value);
      if (!Number.isFinite(numeric)) {{
        return DEFAULT_CAPACITY_DRAWER_WIDTH_VW;
      }}
      return Math.min(MAX_CAPACITY_DRAWER_WIDTH_VW, Math.max(MIN_CAPACITY_DRAWER_WIDTH_VW, numeric));
    }}

    function setCapacityDrawerWidth(widthVw) {{
      const nextWidth = clampCapacityDrawerWidthVw(widthVw);
      document.documentElement.style.setProperty("--capacity-profile-drawer-width", nextWidth.toFixed(2) + "vw");
      try {{
        localStorage.setItem(CAPACITY_DRAWER_WIDTH_STORAGE_KEY, String(nextWidth));
      }} catch (_error) {{
      }}
      if (capacityProfileDrawerEl) {{
        capacityProfileDrawerEl.style.width = nextWidth.toFixed(2) + "vw";
      }}
      return nextWidth;
    }}

    function initializeCapacityDrawerWidth() {{
      let initialWidth = DEFAULT_CAPACITY_DRAWER_WIDTH_VW;
      try {{
        initialWidth = clampCapacityDrawerWidthVw(localStorage.getItem(CAPACITY_DRAWER_WIDTH_STORAGE_KEY));
      }} catch (_error) {{
        initialWidth = DEFAULT_CAPACITY_DRAWER_WIDTH_VW;
      }}
      setCapacityDrawerWidth(initialWidth);
    }}

    function stopCapacityDrawerResize() {{
      if (capacityProfileDrawerEl) {{
        capacityProfileDrawerEl.classList.remove("is-resizing");
      }}
      document.body.classList.remove("capacity-modal-resizing");
      capacityDrawerPointerId = null;
    }}

    function handleCapacityDrawerResizeClientX(clientX) {{
      const viewportWidth = Math.max(window.innerWidth || 0, 1);
      const desiredWidthPx = viewportWidth - Number(clientX || 0);
      const desiredWidthVw = (desiredWidthPx / viewportWidth) * 100;
      setCapacityDrawerWidth(desiredWidthVw);
    }}

    function startCapacityDrawerResize(event) {{
      if (!capacityProfileDrawerEl || !capacityProfileResizeHandleEl) {{
        return;
      }}
      if (window.innerWidth <= 760) {{
        return;
      }}
      event.preventDefault();
      capacityDrawerPointerId = typeof event.pointerId === "number" ? event.pointerId : null;
      capacityProfileDrawerEl.classList.add("is-resizing");
      document.body.classList.add("capacity-modal-resizing");
      handleCapacityDrawerResizeClientX(event.clientX);
      if (capacityDrawerPointerId !== null && typeof capacityProfileResizeHandleEl.setPointerCapture === "function") {{
        capacityProfileResizeHandleEl.setPointerCapture(capacityDrawerPointerId);
      }}
    }}

    function setAspectColumnWidth(width) {{
      const next = clampAspectWidth(Math.round(Number(width) || MIN_ASPECT_COL_WIDTH));
      document.documentElement.style.setProperty("--aspect-col-width", String(next) + "px");
      localStorage.setItem(ASPECT_COL_WIDTH_STORAGE_KEY, String(next));
    }}

    function initializeAspectColumnResize() {{
      const resizer = document.getElementById("aspect-col-resizer");
      const aspectHeader = document.querySelector("thead th.col-aspect");
      if (!resizer || !aspectHeader) {{
        return;
      }}

      const storedWidth = Number.parseInt(localStorage.getItem(ASPECT_COL_WIDTH_STORAGE_KEY) || "", 10);
      if (Number.isFinite(storedWidth)) {{
        setAspectColumnWidth(storedWidth);
      }}

      let resizing = false;
      let startX = 0;
      let startWidth = 0;

      function stopResize() {{
        if (!resizing) {{
          return;
        }}
        resizing = false;
        document.body.classList.remove("resizing-col");
        window.removeEventListener("pointermove", onResizeMove);
        window.removeEventListener("pointerup", stopResize);
      }}

      function onResizeMove(event) {{
        if (!resizing) {{
          return;
        }}
        const delta = Number(event.clientX || 0) - startX;
        setAspectColumnWidth(startWidth + delta);
      }}

      resizer.addEventListener("pointerdown", (event) => {{
        event.preventDefault();
        resizing = true;
        startX = Number(event.clientX || 0);
        startWidth = aspectHeader.getBoundingClientRect().width || MIN_ASPECT_COL_WIDTH;
        document.body.classList.add("resizing-col");
        window.addEventListener("pointermove", onResizeMove);
        window.addEventListener("pointerup", stopResize);
      }});
    }}

    function setTheme(theme) {{
      const nextTheme = theme === "dark" ? "dark" : "light";
      document.documentElement.setAttribute("data-theme", nextTheme);
      localStorage.setItem(THEME_STORAGE_KEY, nextTheme);
      updateThemeToggleLabel(nextTheme);
    }}

    function getInitialTheme() {{
      const stored = localStorage.getItem(THEME_STORAGE_KEY);
      if (stored === "dark" || stored === "light") {{
        return stored;
      }}
      return "light";
    }}

    function updateThemeToggleLabel(theme) {{
      if (!themeToggleButton) {{
        return;
      }}
      themeToggleButton.innerHTML = theme === "dark"
        ? '<span class="material-icons-outlined" aria-hidden="true">light_mode</span>Light mode'
        : '<span class="material-icons-outlined" aria-hidden="true">dark_mode</span>Dark mode';
    }}

    function initializeThemeToggle() {{
      const initialTheme = getInitialTheme();
      document.documentElement.setAttribute("data-theme", initialTheme);
      updateThemeToggleLabel(initialTheme);
      if (!themeToggleButton) {{
        return;
      }}
      themeToggleButton.addEventListener("click", () => {{
        const currentTheme = document.documentElement.getAttribute("data-theme") === "dark" ? "dark" : "light";
        setTheme(currentTheme === "dark" ? "light" : "dark");
        closeViewOptionsMenu();
      }});
    }}

    function getInitialDensity() {{
      const stored = localStorage.getItem(DENSITY_STORAGE_KEY);
      if (stored === "compact" || stored === "cozy") {{
        return stored;
      }}
      return "cozy";
    }}

    function updateDensityToggleLabel(density) {{
      if (!densityToggleButton) {{
        return;
      }}
      densityToggleButton.textContent = density === "compact"
        ? "Cozy View"
        : "Compact View";
    }}

    function setDensity(density) {{
      const nextDensity = density === "compact" ? "compact" : "cozy";
      document.documentElement.setAttribute("data-density", nextDensity);
      localStorage.setItem(DENSITY_STORAGE_KEY, nextDensity);
      updateDensityToggleLabel(nextDensity);
    }}

    function initializeDensityToggle() {{
      const initialDensity = getInitialDensity();
      setDensity(initialDensity);
      if (!densityToggleButton) {{
        return;
      }}
      densityToggleButton.addEventListener("click", () => {{
        const currentDensity = document.documentElement.getAttribute("data-density") === "compact"
          ? "compact"
          : "cozy";
        setDensity(currentDensity === "compact" ? "cozy" : "compact");
        closeViewOptionsMenu();
      }});
    }}

    function projectLabelFor(project) {{
      const key = String(project && project.key || "").trim();
      const name = String(project && project.name || "").trim();
      return name || key;
    }}

    function updateProjectFilterSummary() {{
      if (!projectFilterSummary) {{
        return;
      }}
      const total = allProjects.length;
      const selected = selectedProjectKeys.size;
      if (selected === total) {{
        projectFilterSummary.textContent = "Projects: All";
        return;
      }}
      if (selected === 0) {{
        projectFilterSummary.textContent = "Projects: None";
        return;
      }}
      projectFilterSummary.textContent = "Projects: " + String(selected) + " selected";
    }}

    function setProjectFilterLoading(isLoading, requestVersion) {{
      if (Number.isFinite(Number(requestVersion))) {{
        projectFilterLoadingVersion = Number(requestVersion);
      }} else if (!isLoading) {{
        projectFilterLoadingVersion = 0;
      }}
      if (!projectFilterProgress) {{
        return;
      }}
      const active = !!isLoading;
      projectFilterProgress.hidden = !active;
      projectFilterProgress.classList.toggle("is-active", active);
    }}

    function setTeamFilterLoading(isLoading, requestVersion) {{
      if (Number.isFinite(Number(requestVersion))) {{
        teamFilterLoadingVersion = Number(requestVersion);
      }} else if (!isLoading) {{
        teamFilterLoadingVersion = 0;
      }}
      if (!teamFilterProgress) {{
        return;
      }}
      const active = !!isLoading;
      teamFilterProgress.hidden = !active;
      teamFilterProgress.classList.toggle("is-active", active);
    }}

    function closeProjectFilterMenu() {{
      if (!projectFilterRoot || !projectFilterToggle) {{
        return;
      }}
      projectFilterRoot.classList.remove("open");
      projectFilterToggle.setAttribute("aria-expanded", "false");
    }}

    function openProjectFilterMenu() {{
      if (!projectFilterRoot || !projectFilterToggle) {{
        return;
      }}
      projectFilterRoot.classList.add("open");
      projectFilterToggle.setAttribute("aria-expanded", "true");
    }}

    function toggleProjectFilterMenu() {{
      if (!projectFilterRoot) {{
        return;
      }}
      const isOpen = projectFilterRoot.classList.contains("open");
      if (isOpen) {{
        closeProjectFilterMenu();
      }} else {{
        openProjectFilterMenu();
      }}
    }}

    function closeViewOptionsMenu() {{
      if (!viewOptionsRoot || !viewOptionsToggle) {{
        return;
      }}
      viewOptionsRoot.classList.remove("open");
      viewOptionsToggle.setAttribute("aria-expanded", "false");
    }}

    function openViewOptionsMenu() {{
      if (!viewOptionsRoot || !viewOptionsToggle) {{
        return;
      }}
      viewOptionsRoot.classList.add("open");
      viewOptionsToggle.setAttribute("aria-expanded", "true");
    }}

    function toggleViewOptionsMenu() {{
      if (!viewOptionsRoot) {{
        return;
      }}
      const isOpen = viewOptionsRoot.classList.contains("open");
      if (isOpen) {{
        closeViewOptionsMenu();
      }} else {{
        openViewOptionsMenu();
      }}
    }}

    function renderProjectFilterOptions() {{
      if (!projectFilterOptions) {{
        return;
      }}
      projectFilterOptions.innerHTML = "";
      for (const project of allProjects) {{
        const label = document.createElement("label");
        label.className = "project-option";

        const checkbox = document.createElement("input");
        checkbox.type = "checkbox";
        checkbox.checked = selectedProjectKeys.has(project.key);
        checkbox.dataset.projectKey = project.key;
        checkbox.setAttribute("aria-label", project.label);
        checkbox.addEventListener("change", () => {{
          if (checkbox.checked) {{
            selectedProjectKeys.add(project.key);
          }} else {{
            selectedProjectKeys.delete(project.key);
          }}
          updateProjectFilterSummary();
          rerender(true, {{ projectFilterLoading: true }});
        }});

        const text = document.createElement("span");
        text.className = "project-option-label";
        text.textContent = project.label;

        label.appendChild(checkbox);
        label.appendChild(text);
        projectFilterOptions.appendChild(label);
      }}
    }}

    function refreshProjectFilterOptions(options) {{
      const config = options && typeof options === "object" ? options : {{}};
      const preserveSelection = config.preserveSelection !== false;
      const byKey = new Map();
      for (const row of allRows) {{
        const rowType = String(row && row.row_type || "");
        const key = String(row && row.project_key || "").trim();
        if (rowType !== "project" || !key) {{
          continue;
        }}
        if (!byKey.has(key)) {{
          byKey.set(key, {{
            key,
            name: String(row && row.project_name || "").trim(),
          }});
        }}
      }}
      allProjects.length = 0;
      for (const entry of byKey.values()) {{
        allProjects.push({{
          key: entry.key,
          name: entry.name,
          label: projectLabelFor(entry),
        }});
      }}
      allProjects.sort((left, right) => left.label.localeCompare(right.label));
      const nextSelectedKeys = new Set();
      if (preserveSelection && selectedProjectKeys.size) {{
        for (const key of selectedProjectKeys) {{
          if (allProjects.some((project) => project.key === key)) {{
            nextSelectedKeys.add(key);
          }}
        }}
      }}
      if (!nextSelectedKeys.size) {{
        for (const project of allProjects) {{
          nextSelectedKeys.add(project.key);
        }}
      }} else {{
        for (const project of allProjects) {{
          if (!selectedProjectKeys.has(project.key)) {{
            nextSelectedKeys.add(project.key);
          }}
        }}
      }}
      selectedProjectKeys.clear();
      for (const key of nextSelectedKeys) {{
        selectedProjectKeys.add(key);
      }}
      renderProjectFilterOptions();
      updateProjectFilterSummary();
    }}

    function initializeProjectFilter() {{
      refreshProjectFilterOptions({{ preserveSelection: false }});
      if (projectFilterToggle) {{
        projectFilterToggle.addEventListener("click", () => {{
          toggleProjectFilterMenu();
        }});
        projectFilterToggle.addEventListener("keydown", (event) => {{
          if (event.key === "Enter" || event.key === " ") {{
            event.preventDefault();
            toggleProjectFilterMenu();
          }}
          if (event.key === "Escape") {{
            closeProjectFilterMenu();
          }}
        }});
      }}
      if (projectFilterSelectAll) {{
        projectFilterSelectAll.addEventListener("click", () => {{
          selectedProjectKeys.clear();
          for (const project of allProjects) {{
            selectedProjectKeys.add(project.key);
          }}
          renderProjectFilterOptions();
          updateProjectFilterSummary();
          rerender(true, {{ projectFilterLoading: true }});
        }});
      }}
      if (projectFilterClearAll) {{
        projectFilterClearAll.addEventListener("click", () => {{
          selectedProjectKeys.clear();
          renderProjectFilterOptions();
          updateProjectFilterSummary();
          rerender(true, {{ projectFilterLoading: true }});
        }});
      }}
      document.addEventListener("click", (event) => {{
        if (!projectFilterRoot) {{
          return;
        }}
        if (!projectFilterRoot.contains(event.target)) {{
          closeProjectFilterMenu();
        }}
        if (teamFilterRoot && !teamFilterRoot.contains(event.target)) {{
          closeTeamFilterMenu();
        }}
        if (viewOptionsRoot && !viewOptionsRoot.contains(event.target)) {{
          closeViewOptionsMenu();
        }}
      }});
      document.addEventListener("keydown", (event) => {{
        if (event.key === "Escape") {{
          closeProjectFilterMenu();
          closeTeamFilterMenu();
          closeViewOptionsMenu();
        }}
      }});
    }}

    function normalizeAssigneeName(value) {{
      return String(value || "").trim().toLowerCase();
    }}

    function isAllTeamsSelected() {{
      return !allTeams.length || selectedTeamNames.size === allTeams.length;
    }}

    function recomputeSelectedTeamAssignees() {{
      selectedTeamAssignees.clear();
      if (isAllTeamsSelected()) {{
        return;
      }}
      for (const team of allTeams) {{
        if (!selectedTeamNames.has(team.team_name)) {{
          continue;
        }}
        const members = Array.isArray(team.assignees) ? team.assignees : [];
        for (const member of members) {{
          const normalized = normalizeAssigneeName(member);
          if (normalized) {{
            selectedTeamAssignees.add(normalized);
          }}
        }}
      }}
    }}

    function updateTeamFilterSummary() {{
      if (!teamFilterSummary) {{
        return;
      }}
      const total = allTeams.length;
      const selected = selectedTeamNames.size;
      if (!total) {{
        teamFilterSummary.textContent = "Teams: None";
        return;
      }}
      if (selected === total) {{
        teamFilterSummary.textContent = "Teams: All";
        return;
      }}
      if (selected === 0) {{
        teamFilterSummary.textContent = "Teams: None";
        return;
      }}
      teamFilterSummary.textContent = "Teams: " + String(selected) + " selected";
    }}

    function closeTeamFilterMenu() {{
      if (!teamFilterRoot || !teamFilterToggle) {{
        return;
      }}
      teamFilterRoot.classList.remove("open");
      teamFilterToggle.setAttribute("aria-expanded", "false");
    }}

    function openTeamFilterMenu() {{
      if (!teamFilterRoot || !teamFilterToggle) {{
        return;
      }}
      teamFilterRoot.classList.add("open");
      teamFilterToggle.setAttribute("aria-expanded", "true");
    }}

    function toggleTeamFilterMenu() {{
      if (!teamFilterRoot) {{
        return;
      }}
      const isOpen = teamFilterRoot.classList.contains("open");
      if (isOpen) {{
        closeTeamFilterMenu();
      }} else {{
        openTeamFilterMenu();
      }}
    }}

    async function applyTeamFilterSelection() {{
      recomputeSelectedTeamAssignees();
      updateTeamFilterSummary();
      if (!hasNestedActualsApi) {{
        applyOriginalMetricsToRows();
        rerender(true, {{ teamFilterLoading: true }});
        return;
      }}
      setDateApplyBusy(true);
      try {{
        setTeamFilterLoading(true);
        const payload = await fetchActualHoursForDateRange(
          selectedDateFrom,
          selectedDateTo,
          selectedActualHoursMode,
          selectedTeamAssignees,
          selectedPlannedHoursSource
        );
        applyFetchedActualHours(payload);
        rerender(true, {{ teamFilterLoading: true }});
      }} catch (error) {{
        setTeamFilterLoading(false);
        setDateFilterStatus(String(error && error.message || error || "Failed to apply team filter."));
      }} finally {{
        setDateApplyBusy(false);
      }}
      updateDateRangeApplyState();
    }}

    function renderTeamFilterOptions() {{
      if (!teamFilterOptions) {{
        return;
      }}
      teamFilterOptions.innerHTML = "";
      if (!allTeams.length) {{
        const empty = document.createElement("div");
        empty.className = "project-option-label";
        empty.textContent = "No teams found in Performance Settings.";
        teamFilterOptions.appendChild(empty);
        return;
      }}
      for (const team of allTeams) {{
        const label = document.createElement("label");
        label.className = "project-option";

        const checkbox = document.createElement("input");
        checkbox.type = "checkbox";
        checkbox.checked = selectedTeamNames.has(team.team_name);
        checkbox.dataset.teamName = team.team_name;
        checkbox.setAttribute("aria-label", team.team_name);
        checkbox.addEventListener("change", () => {{
          if (checkbox.checked) {{
            selectedTeamNames.add(team.team_name);
          }} else {{
            selectedTeamNames.delete(team.team_name);
          }}
          applyTeamFilterSelection();
        }});

        const text = document.createElement("span");
        text.className = "project-option-label";
        const leader = String(team.team_leader || "").trim();
        text.textContent = leader
          ? (team.team_name + " (Lead: " + leader + ")")
          : team.team_name;

        label.appendChild(checkbox);
        label.appendChild(text);
        teamFilterOptions.appendChild(label);
      }}
    }}

    async function initializeTeamFilter() {{
      allTeams.length = 0;
      selectedTeamNames.clear();
      selectedTeamAssignees.clear();
      if (!hasTeamsApi) {{
        updateTeamFilterSummary();
        return;
      }}
      let teamsPayload = [];
      try {{
        const response = await fetch(PERFORMANCE_TEAMS_ENDPOINT, {{ method: "GET" }});
        const payload = await response.json().catch(() => ({{}}));
        if (!response.ok) {{
          throw new Error(String(payload && payload.error || "Failed to load teams."));
        }}
        teamsPayload = Array.isArray(payload && payload.teams) ? payload.teams : [];
      }} catch (error) {{
        console.warn("Failed to load teams:", error);
      }}
      for (const item of teamsPayload) {{
        const teamName = String(item && item.team_name || "").trim();
        if (!teamName) {{
          continue;
        }}
        const members = Array.isArray(item && item.assignees) ? item.assignees : [];
        const assignees = Array.from(new Set(
          members
            .map((entry) => String(entry || "").trim())
            .filter(Boolean)
        ));
        allTeams.push({{
          team_name: teamName,
          team_leader: String(item && item.team_leader || "").trim(),
          assignees,
        }});
      }}
      allTeams.sort((left, right) => left.team_name.localeCompare(right.team_name));
      for (const team of allTeams) {{
        selectedTeamNames.add(team.team_name);
      }}
      renderTeamFilterOptions();
      updateTeamFilterSummary();

      if (teamFilterToggle) {{
        teamFilterToggle.addEventListener("click", () => {{
          toggleTeamFilterMenu();
        }});
        teamFilterToggle.addEventListener("keydown", (event) => {{
          if (event.key === "Enter" || event.key === " ") {{
            event.preventDefault();
            toggleTeamFilterMenu();
          }}
          if (event.key === "Escape") {{
            closeTeamFilterMenu();
          }}
        }});
      }}
      if (teamFilterSelectAll) {{
        teamFilterSelectAll.addEventListener("click", () => {{
          selectedTeamNames.clear();
          for (const team of allTeams) {{
            selectedTeamNames.add(team.team_name);
          }}
          renderTeamFilterOptions();
          applyTeamFilterSelection();
        }});
      }}
      if (teamFilterClearAll) {{
        teamFilterClearAll.addEventListener("click", () => {{
          selectedTeamNames.clear();
          renderTeamFilterOptions();
          applyTeamFilterSelection();
        }});
      }}
    }}

    function addChild(parentId, childId) {{
      if (!childrenByParent.has(parentId)) {{
        childrenByParent.set(parentId, []);
      }}
      childrenByParent.get(parentId).push(childId);
    }}

    function hasChildren(id) {{
      const children = childrenByParent.get(id) || [];
      return children.length > 0;
    }}

    function clearTreeState() {{
      rowsById.clear();
      childrenByParent.clear();
    }}

    function findAncestorByType(row, targetType, byId) {{
      let current = row;
      while (current && current.parent_id) {{
        const parent = byId.get(current.parent_id);
        if (!parent) {{
          return null;
        }}
        if (parent.row_type === targetType) {{
          return parent;
        }}
        current = parent;
      }}
      return null;
    }}

    function asText(value) {{
      return String(value || "").trim();
    }}

    function escapeHtml(value) {{
      return String(value || "")
        .replace(/&/g, "&amp;")
        .replace(/</g, "&lt;")
        .replace(/>/g, "&gt;")
        .replace(/"/g, "&quot;")
        .replace(/'/g, "&#39;");
    }}

    function parseDateValue(value) {{
      const text = asText(value);
      if (!text) {{
        return null;
      }}
      const exact = text.match(/^(\\d{{4}})-(\\d{{2}})-(\\d{{2}})$/);
      if (exact) {{
        return new Date(Number(exact[1]), Number(exact[2]) - 1, Number(exact[3]));
      }}
      const iso = text.match(/^(\\d{{4}})-(\\d{{2}})-(\\d{{2}})T/);
      if (iso) {{
        return new Date(Number(iso[1]), Number(iso[2]) - 1, Number(iso[3]));
      }}
      const d = new Date(text);
      if (Number.isNaN(d.getTime())) {{
        return null;
      }}
      return new Date(d.getFullYear(), d.getMonth(), d.getDate());
    }}

    function formatCapacityDate(value) {{
      const parsed = parseDateValue(value);
      if (!parsed) {{
        return asText(value);
      }}
      return parsed.toLocaleDateString(undefined, {{
        year: "numeric",
        month: "short",
        day: "numeric",
      }});
    }}

    function parseMonthValue(value) {{
      const text = asText(value);
      const m = text.match(/^(\\d{{4}})-(\\d{{2}})$/);
      if (!m) {{
        return null;
      }}
      const year = Number(m[1]);
      const monthIndex = Number(m[2]) - 1;
      if (!Number.isFinite(year) || !Number.isFinite(monthIndex) || monthIndex < 0 || monthIndex > 11) {{
        return null;
      }}
      return {{ year, monthIndex }};
    }}

    function monthStart(value) {{
      const parsed = parseMonthValue(value);
      if (!parsed) {{
        return null;
      }}
      return new Date(parsed.year, parsed.monthIndex, 1);
    }}

    function monthEnd(value) {{
      const parsed = parseMonthValue(value);
      if (!parsed) {{
        return null;
      }}
      return new Date(parsed.year, parsed.monthIndex + 1, 0);
    }}

    function parseFilterDate(value) {{
      const text = asText(value);
      const match = text.match(/^(\\d{{4}})-(\\d{{2}})-(\\d{{2}})$/);
      if (!match) {{
        return null;
      }}
      const year = Number(match[1]);
      const month = Number(match[2]) - 1;
      const day = Number(match[3]);
      if (!Number.isFinite(year) || !Number.isFinite(month) || !Number.isFinite(day)) {{
        return null;
      }}
      return new Date(year, month, day);
    }}

    function syncExtendedActualHoursToggle() {{
      if (!extendedActualHoursToggle) {{
        return;
      }}
      extendedActualHoursToggle.checked = pendingActualHoursMode === "planned_dates";
    }}

    function normalizePendingDateRange() {{
      if (pendingDateFrom > pendingDateTo) {{
        const tmp = pendingDateFrom;
        pendingDateFrom = pendingDateTo;
        pendingDateTo = tmp;
      }}
      if (pendingActualHoursMode !== "log_date" && pendingActualHoursMode !== "planned_dates") {{
        pendingActualHoursMode = DEFAULT_ACTUAL_HOURS_MODE;
      }}
      pendingPlannedHoursSource = normalizePlannedHoursSource(pendingPlannedHoursSource);
      if (dateFilterFromInput) {{
        dateFilterFromInput.value = pendingDateFrom;
      }}
      if (dateFilterToInput) {{
        dateFilterToInput.value = pendingDateTo;
      }}
      if (plannedHoursSourceSelect) {{
        plannedHoursSourceSelect.value = pendingPlannedHoursSource;
      }}
      if (actualHoursModeSelect) {{
        actualHoursModeSelect.value = pendingActualHoursMode;
      }}
      syncExtendedActualHoursToggle();
    }}

    function hasPendingDateChange() {{
      normalizePendingDateRange();
      return pendingDateFrom !== selectedDateFrom
        || pendingDateTo !== selectedDateTo
        || pendingActualHoursMode !== selectedActualHoursMode
        || pendingPlannedHoursSource !== selectedPlannedHoursSource;
    }}

    function updateDateRangeApplyState() {{
      if (!dateFilterApplyButton) {{
        return;
      }}
      const fromDate = parseFilterDate(pendingDateFrom);
      const toDate = parseFilterDate(pendingDateTo);
      const isValid = !!(fromDate && toDate);
      const dirty = hasPendingDateChange();
      dateFilterApplyButton.disabled = isApplyingDateRange || !isValid;
      if (isApplyingDateRange) {{
        setDateFilterStatus("Recomputing actual hours for selected range and mode...");
      }} else if (!isValid) {{
        setDateFilterStatus("Select a valid date range.");
      }} else if (dirty) {{
        setDateFilterStatus("Date range, mode, or planned hours source changed. Click apply.");
      }} else {{
        setDateFilterStatus("");
      }}
    }}

    function getDateFilterBounds() {{
      if (selectedDateFrom > selectedDateTo) {{
        const tmp = selectedDateFrom;
        selectedDateFrom = selectedDateTo;
        selectedDateTo = tmp;
      }}
      return getDateFilterBoundsFor(selectedDateFrom, selectedDateTo);
    }}

    function getDateFilterBoundsFor(fromValue, toValue) {{
      let normalizedFrom = String(fromValue || DEFAULT_DATE_FROM);
      let normalizedTo = String(toValue || DEFAULT_DATE_TO);
      if (normalizedFrom > normalizedTo) {{
        const tmp = normalizedFrom;
        normalizedFrom = normalizedTo;
        normalizedTo = tmp;
      }}
      return {{
        start: parseFilterDate(normalizedFrom) || parseFilterDate(DEFAULT_DATE_FROM),
        end: parseFilterDate(normalizedTo) || parseFilterDate(DEFAULT_DATE_TO),
      }};
    }}

    function normalizeActualHoursMode(value) {{
      return String(value || "").trim().toLowerCase() === "planned_dates"
        ? "planned_dates"
        : "log_date";
    }}

    function buildScorecardSelectionSnapshot(overrides) {{
      const snapshot = overrides && typeof overrides === "object" ? overrides : {{}};
      return {{
        dateFrom: String(snapshot.dateFrom || selectedDateFrom || DEFAULT_DATE_FROM),
        dateTo: String(snapshot.dateTo || selectedDateTo || DEFAULT_DATE_TO),
        actualHoursMode: normalizeActualHoursMode(snapshot.actualHoursMode || selectedActualHoursMode || DEFAULT_ACTUAL_HOURS_MODE),
        plannedHoursSource: normalizePlannedHoursSource(
          snapshot.plannedHoursSource || selectedPlannedHoursSource || DEFAULT_PLANNED_HOURS_SOURCE
        ),
      }};
    }}

    function captureOriginalMetrics() {{
      originalMetricsById.clear();
      for (const row of allRows) {{
        originalMetricsById.set(row.id, {{
          actual_hours: row.actual_hours,
          actual_days: row.actual_days,
          delta_hours: row.delta_hours,
          delta_days: row.delta_days,
        }});
      }}
    }}

    function assignComputedMetrics(row, nextHours) {{
      const hasPlannedHours = Number.isFinite(Number(row && row.man_hours));
      const hasPlannedDays = Number.isFinite(Number(row && row.man_days));
      const actualHours = roundHours(nextHours);
      row.actual_hours = actualHours;
      row.actual_days = toHoursToDays(actualHours);
      row.delta_hours = hasPlannedHours ? roundHours(Number(row.man_hours) - actualHours) : "";
      row.delta_days = hasPlannedDays ? roundHours(Number(row.man_days) - row.actual_days) : "";
    }}

    function applyOriginalMetricsToRows() {{
      for (const row of allRows) {{
        const original = originalMetricsById.get(row.id);
        if (!original) {{
          continue;
        }}
        row.actual_hours = original.actual_hours;
        row.actual_days = original.actual_days;
        row.delta_hours = original.delta_hours;
        row.delta_days = original.delta_days;
      }}
    }}

    function resolveRowIssueKey(row) {{
      const explicit = String(row && row.jira_key || "").trim().toUpperCase();
      if (explicit) {{
        return explicit;
      }}
      const aspect = String(row && row.aspect || "").trim().toUpperCase();
      const match = aspect.match(/([A-Z][A-Z0-9]+-\\d+)/);
      return match ? String(match[1] || "").trim().toUpperCase() : "";
    }}

    function applyFetchedActualHours(payload) {{
      const subtaskHours = payload && payload.subtask_hours_by_issue && typeof payload.subtask_hours_by_issue === "object"
        ? payload.subtask_hours_by_issue
        : {{}};
      const childrenByParentId = new Map();
      const byId = new Map();
      for (const row of allRows) {{
        byId.set(row.id, row);
        if (!row.parent_id) {{
          continue;
        }}
        if (!childrenByParentId.has(row.parent_id)) {{
          childrenByParentId.set(row.parent_id, []);
        }}
        childrenByParentId.get(row.parent_id).push(row.id);
      }}

      function getChildRows(parentId) {{
        const childIds = childrenByParentId.get(parentId) || [];
        return childIds
          .map((childId) => byId.get(childId))
          .filter(Boolean);
      }}

      function hasLeafWorkChildren(row) {{
        const childRows = getChildRows(row && row.id);
        return childRows.some((child) => LEAF_WORK_ROW_TYPES.has(String(child && child.row_type || "")));
      }}

      for (const row of allRows) {{
        const rowType = String(row && row.row_type || "");
        if (!LEAF_WORK_ROW_TYPES.has(rowType)) {{
          continue;
        }}
        const jiraKey = resolveRowIssueKey(row);
        const nextHours = jiraKey ? toFiniteNumber(subtaskHours[jiraKey], 0) : 0;
        assignComputedMetrics(row, nextHours);
      }}

      function sumWorkHoursDescendants(parentId) {{
        const childIds = childrenByParentId.get(parentId) || [];
        let total = 0;
        for (const childId of childIds) {{
          const child = byId.get(childId);
          if (!child) {{
            continue;
          }}
          total += toFiniteNumber(child.actual_hours, 0);
        }}
        return total;
      }}

      for (const rowType of ["story", "rmi", "product", "project"]) {{
        for (const row of allRows) {{
          if (row.row_type !== rowType) {{
            continue;
          }}
          const rolledHours = sumWorkHoursDescendants(row.id);
          assignComputedMetrics(row, rolledHours);
        }}
      }}
    }}

    async function fetchActualHoursForDateRange(fromDate, toDate, mode, selectedAssigneesSet, plannedHoursSource) {{
      const fromParam = encodeURIComponent(String(fromDate || ""));
      const toParam = encodeURIComponent(String(toDate || ""));
      const modeParam = encodeURIComponent(String(mode || DEFAULT_ACTUAL_HOURS_MODE));
      const scopeBasis = resolveScopedSubtaskBasis(plannedHoursSource);
      const scopeBasisParam = encodeURIComponent(scopeBasis);
      const projectsParam = selectedProjectKeys.size
        ? ("&projects=" + encodeURIComponent(Array.from(selectedProjectKeys).sort().join(",")))
        : "";
      const assignees = Array.isArray(selectedAssigneesSet)
        ? selectedAssigneesSet
        : Array.from(selectedAssigneesSet || []);
      const assigneesParam = assignees.length
        ? ("&assignees=" + encodeURIComponent(assignees.join(",")))
        : "";
      const response = await fetch(
        NESTED_ACTUALS_ENDPOINT
        + "?from=" + fromParam
        + "&to=" + toParam
        + "&mode=" + modeParam
        + "&scope_basis=" + scopeBasisParam
        + projectsParam
        + assigneesParam
        + "&report=nested_view",
        {{
        method: "GET",
      }});
      const payload = await response.json().catch(() => ({{}}));
      if (!response.ok || !payload || payload.ok === false) {{
        throw new Error(String(payload && payload.error || "Failed to fetch actual hours for selected date range."));
      }}
      return payload;
    }}

    async function fetchScopedSubtasksSummary(selection) {{
      const snapshot = selection && typeof selection === "object" ? selection : buildScorecardSelectionSnapshot();
      const scorecardDateFrom = String(snapshot.dateFrom || selectedDateFrom || DEFAULT_DATE_FROM);
      const scorecardDateTo = String(snapshot.dateTo || selectedDateTo || DEFAULT_DATE_TO);
      const scorecardActualHoursMode = String(
        snapshot.actualHoursMode || selectedActualHoursMode || DEFAULT_ACTUAL_HOURS_MODE
      );
      const params = new URLSearchParams();
      params.set("from", scorecardDateFrom);
      params.set("to", scorecardDateTo);
      params.set("mode", scorecardActualHoursMode === "planned_dates" ? "extended" : "log_date");
      params.set("projects", Array.from(selectedProjectKeys).sort().join(","));
      if (!isAllTeamsSelected() && selectedTeamAssignees.size) {{
        params.set("assignees", Array.from(selectedTeamAssignees).sort().join(","));
      }}
      const response = await fetch(SCOPED_SUBTASKS_ENDPOINT + "?" + params.toString(), {{ method: "GET" }});
      const payload = await response.json().catch(() => ({{}}));
      if (!response.ok || !payload || payload.ok === false) {{
        throw new Error(String(payload && payload.error || "Failed to fetch scoped subtasks."));
      }}
      return payload;
    }}

    async function applyPendingDateRange() {{
      pendingDateFrom = dateFilterFromInput
        ? (dateFilterFromInput.value || DEFAULT_DATE_FROM)
        : pendingDateFrom;
      pendingDateTo = dateFilterToInput
        ? (dateFilterToInput.value || DEFAULT_DATE_TO)
        : pendingDateTo;
      pendingActualHoursMode = actualHoursModeSelect
        ? String(actualHoursModeSelect.value || DEFAULT_ACTUAL_HOURS_MODE)
        : pendingActualHoursMode;
      pendingPlannedHoursSource = plannedHoursSourceSelect
        ? normalizePlannedHoursSource(plannedHoursSourceSelect.value || DEFAULT_PLANNED_HOURS_SOURCE)
        : pendingPlannedHoursSource;
      if (extendedActualHoursToggle) {{
        pendingActualHoursMode = extendedActualHoursToggle.checked ? "planned_dates" : "log_date";
        if (actualHoursModeSelect) {{
          actualHoursModeSelect.value = pendingActualHoursMode;
        }}
      }}
      normalizePendingDateRange();
      const nextFrom = pendingDateFrom;
      const nextTo = pendingDateTo;
      const nextMode = normalizeActualHoursMode(pendingActualHoursMode);
      const nextPlannedHoursSource = normalizePlannedHoursSource(pendingPlannedHoursSource);
      if (!parseFilterDate(nextFrom) || !parseFilterDate(nextTo)) {{
        updateDateRangeApplyState();
        return;
      }}
      setDateApplyBusy(true);
      try {{
        if (hasNestedActualsApi) {{
          const payload = await fetchActualHoursForDateRange(
            nextFrom,
            nextTo,
            nextMode,
            selectedTeamAssignees,
            nextPlannedHoursSource
          );
          applyFetchedActualHours(payload);
          const logPayload = nextMode === "log_date"
            ? payload
            : await fetchActualHoursForDateRange(
                nextFrom,
                nextTo,
                "log_date",
                selectedTeamAssignees,
                nextPlannedHoursSource
              );
          subtaskLogHoursByIssue = extractSubtaskHoursMap(logPayload);
        }} else {{
          applyOriginalMetricsToRows();
        }}
        selectedDateFrom = nextFrom;
        selectedDateTo = nextTo;
        selectedActualHoursMode = nextMode;
        selectedPlannedHoursSource = nextPlannedHoursSource;
        localStorage.setItem(ACTUAL_HOURS_MODE_STORAGE_KEY, selectedActualHoursMode);
        localStorage.setItem(PLANNED_HOURS_SOURCE_STORAGE_KEY, selectedPlannedHoursSource);
        rerender(true, {{
          selection: {{
            dateFrom: nextFrom,
            dateTo: nextTo,
            actualHoursMode: nextMode,
            plannedHoursSource: nextPlannedHoursSource,
          }},
        }});
      }} catch (error) {{
        if (hasNestedActualsApi) {{
          applyOriginalMetricsToRows();
        }}
        selectedDateFrom = nextFrom;
        selectedDateTo = nextTo;
        selectedActualHoursMode = nextMode;
        selectedPlannedHoursSource = nextPlannedHoursSource;
        rerender(true, {{
          selection: {{
            dateFrom: nextFrom,
            dateTo: nextTo,
            actualHoursMode: nextMode,
            plannedHoursSource: nextPlannedHoursSource,
          }},
        }});
        setDateFilterStatus(String(error && error.message || error || "Actual hours API unavailable; showing client-side filtered view."));
      }} finally {{
        setDateApplyBusy(false);
      }}
      updateDateRangeApplyState();
    }}

    function isDateWithinBounds(dateValue, bounds) {{
      if (!dateValue || !bounds || !bounds.start || !bounds.end) {{
        return false;
      }}
      const value = dateValue.getTime();
      return value >= bounds.start.getTime() && value <= bounds.end.getTime();
    }}

    function matchesDateFilter(row, selection) {{
      const activeSelection = buildScorecardSelectionSnapshot(selection);
      const rowType = String(row && row.row_type || "");
      const source = activeSelection.plannedHoursSource;
      const bounds = getDateFilterBoundsFor(activeSelection.dateFrom, activeSelection.dateTo);
      if (source === "epic_estimates") {{
        if (!DATE_FILTER_WORK_TYPES.has(rowType)) {{
          return false;
        }}
        const plannedStart = parseDateValue(row && row.planned_start);
        const plannedEnd = parseDateValue(row && row.planned_end);
        if (!plannedStart && !plannedEnd) {{
          return false;
        }}
        return isDateWithinBounds(plannedStart, bounds) || isDateWithinBounds(plannedEnd, bounds);
      }}
      if (rowType !== "subtask") {{
        return false;
      }}
      if (source === "subtask_logs") {{
        const jiraKey = String(row && row.jira_key || "").trim().toUpperCase();
        if (!jiraKey) {{
          return false;
        }}
        if (Object.prototype.hasOwnProperty.call(subtaskLogHoursByIssue, jiraKey)) {{
          return toFiniteNumber(subtaskLogHoursByIssue[jiraKey], 0) > 0;
        }}
        if (activeSelection.actualHoursMode === "log_date") {{
          return toFiniteNumber(row && row.actual_hours, 0) > 0;
        }}
        return false;
      }}
      const plannedStart = parseDateValue(row && (row.planned_start || row.start_date));
      const plannedEnd = parseDateValue(row && (row.planned_end || row.planned_due || row.due_date));
      if (!plannedStart && !plannedEnd) {{
        return false;
      }}
      return isDateWithinBounds(plannedStart, bounds) || isDateWithinBounds(plannedEnd, bounds);
    }}

    function applyDateFilterWithAncestors(rows, includeDescendants, selection) {{
      if (!rows.length) {{
        return rows;
      }}
      const keepDescendants = includeDescendants !== false;
      const byId = new Map();
      const childrenByParent = new Map();
      for (const row of rows) {{
        byId.set(row.id, row);
        if (!row.parent_id) {{
          continue;
        }}
        if (!childrenByParent.has(row.parent_id)) {{
          childrenByParent.set(row.parent_id, []);
        }}
        childrenByParent.get(row.parent_id).push(row.id);
      }}
      const visibleIds = new Set();
      function addDescendants(parentId) {{
        const childIds = childrenByParent.get(parentId) || [];
        for (const childId of childIds) {{
          if (visibleIds.has(childId)) {{
            continue;
          }}
          visibleIds.add(childId);
          addDescendants(childId);
        }}
      }}
      for (const row of rows) {{
        if (!matchesDateFilter(row, selection)) {{
          continue;
        }}
        visibleIds.add(row.id);
        let current = row;
        while (current && current.parent_id) {{
          visibleIds.add(current.parent_id);
          current = byId.get(current.parent_id) || null;
        }}
        if (keepDescendants) {{
          addDescendants(row.id);
        }}
      }}
      return rows.filter((row) => visibleIds.has(row.id));
    }}

    function applyEpicPlannedTotalsToAncestors(rows) {{
      if (!rows.length) {{
        return;
      }}
      const byId = new Map();
      const childrenByParent = new Map();
      for (const row of rows) {{
        byId.set(row.id, row);
        if (!row.parent_id) {{
          continue;
        }}
        if (!childrenByParent.has(row.parent_id)) {{
          childrenByParent.set(row.parent_id, []);
        }}
        childrenByParent.get(row.parent_id).push(row.id);
      }}
      function sumApprovedEpicHours(parentId) {{
        const childIds = childrenByParent.get(parentId) || [];
        let total = 0;
        for (const childId of childIds) {{
          const child = byId.get(childId);
          if (!child) {{
            continue;
          }}
          if (child.row_type === "rmi") {{
            total += toFiniteNumber(child.man_hours, 0);
          }}
          total += sumApprovedEpicHours(childId);
        }}
        return total;
      }}
      function sumPlannedLeafHours(parentId) {{
        const childIds = childrenByParent.get(parentId) || [];
        let total = 0;
        for (const childId of childIds) {{
          const child = byId.get(childId);
          if (!child) {{
            continue;
          }}
          if (child.row_type === "subtask") {{
            total += toFiniteNumber(child.planned_hours, 0);
            continue;
          }}
          total += sumPlannedLeafHours(childId);
        }}
        return total;
      }}
      for (const row of rows) {{
        if (row.row_type === "project" || row.row_type === "product") {{
          const epicHours = roundHours(sumApprovedEpicHours(row.id));
          row.man_hours = epicHours;
          row.man_days = toHoursToDays(epicHours);
          row.approved_hours = epicHours;
          row.approved_days = toHoursToDays(epicHours);
        }}
        if (row.row_type !== "story" && row.row_type !== "rmi" && row.row_type !== "product" && row.row_type !== "project") {{
          continue;
        }}
        const plannedHours = roundHours(sumPlannedLeafHours(row.id));
        row.planned_hours = plannedHours;
        row.planned_days = toHoursToDays(plannedHours);
        const actualHours = Number(row.actual_hours);
        const hasActual = Number.isFinite(actualHours);
        const approvedHours = Number(row.approved_hours);
        const approvedDays = Number(row.approved_days);
        row.delta_hours = hasActual && Number.isFinite(approvedHours) ? roundHours(approvedHours - actualHours) : "";
        row.delta_days = hasActual && Number.isFinite(approvedDays) ? roundHours(approvedDays - row.actual_days) : "";
      }}
    }}

    function buildDisplayRows(selection) {{
      const activeSelection = buildScorecardSelectionSnapshot(selection);
      const byId = new Map();
      for (const row of allRows) {{
        byId.set(row.id, row);
      }}

      const assigneeBySubtaskId = new Map();
      for (const row of allRows) {{
        if (row.row_type !== "assignee" || !row.parent_id) {{
          continue;
        }}
        const name = String(row.aspect || "").trim();
        if (!name) {{
          continue;
        }}
        if (!assigneeBySubtaskId.has(row.parent_id)) {{
          assigneeBySubtaskId.set(row.parent_id, []);
        }}
        assigneeBySubtaskId.get(row.parent_id).push(name);
      }}
      for (const [subtaskId, names] of assigneeBySubtaskId.entries()) {{
        const unique = Array.from(new Set(names.map((name) => String(name || "").trim()).filter(Boolean)));
        assigneeBySubtaskId.set(subtaskId, unique);
      }}

      const assigneesByStoryId = new Map();
      const assigneesByEpicId = new Map();
      for (const row of allRows) {{
        if (row.row_type !== "subtask") {{
          continue;
        }}
        const subtaskAssignees = assigneeBySubtaskId.get(row.id) || [];
        if (!subtaskAssignees.length) {{
          continue;
        }}
        const storyParent = findAncestorByType(row, "story", byId);
        if (storyParent) {{
          if (!assigneesByStoryId.has(storyParent.id)) {{
            assigneesByStoryId.set(storyParent.id, new Set());
          }}
          const storySet = assigneesByStoryId.get(storyParent.id);
          for (const name of subtaskAssignees) {{
            storySet.add(name);
          }}
        }}
        const epicParent = findAncestorByType(row, "rmi", byId);
        if (epicParent) {{
          if (!assigneesByEpicId.has(epicParent.id)) {{
            assigneesByEpicId.set(epicParent.id, new Set());
          }}
          const epicSet = assigneesByEpicId.get(epicParent.id);
          for (const name of subtaskAssignees) {{
            epicSet.add(name);
          }}
        }}
      }}

      const filtered = [];
      for (const row of allRows) {{
        const itemProjectKey = String(row && row.project_key || "").trim();
        if (!selectedProjectKeys.has(itemProjectKey)) {{
          continue;
        }}
        if (row.row_type === "assignee") {{
          continue;
        }}

        if (!showProductCategorization && row.row_type === "product") {{
          continue;
        }}

        const item = Object.assign({{}}, row);
        if (item.row_type === "subtask") {{
          const assignees = assigneeBySubtaskId.get(item.id) || [];
          item.assignee_name = assignees.join(", ");
        }} else if (item.row_type === "story") {{
          const assignees = Array.from(assigneesByStoryId.get(item.id) || []);
          assignees.sort((left, right) => left.localeCompare(right));
          item.assignee_name = assignees.join(", ");
        }} else if (item.row_type === "rmi") {{
          const assignees = Array.from(assigneesByEpicId.get(item.id) || []);
          assignees.sort((left, right) => left.localeCompare(right));
          item.assignee_name = assignees.join(", ");
        }}

        if (!showProductCategorization) {{
          if (item.row_type === "project") {{
            item.level = 1;
            item.parent_id = null;
          }} else if (item.row_type === "rmi") {{
            const projectParent = findAncestorByType(row, "project", byId);
            item.level = 2;
            item.parent_id = projectParent ? projectParent.id : null;
          }} else if (item.row_type === "story") {{
            const rmiParent = findAncestorByType(row, "rmi", byId);
            const projectParent = findAncestorByType(row, "project", byId);
            item.level = 3;
            item.parent_id = rmiParent ? rmiParent.id : (projectParent ? projectParent.id : null);
          }} else if (item.row_type === "subtask") {{
            const storyParent = findAncestorByType(row, "story", byId);
            const rmiParent = findAncestorByType(row, "rmi", byId);
            item.level = 4;
            item.parent_id = storyParent ? storyParent.id : (rmiParent ? rmiParent.id : null);
          }}
        }}

        if (!item.type_label) {{
          if (item.row_type === "project") {{
            item.type_label = "Project";
          }} else if (item.row_type === "rmi") {{
            item.type_label = "Epic";
          }} else if (item.row_type === "story") {{
            item.type_label = "Story";
          }} else if (item.row_type === "subtask") {{
            item.type_label = "Subtask";
          }}
        }}
        filtered.push(item);
      }}
      const filteredByDate = applyDateFilterWithAncestors(filtered, true, activeSelection);
      const scorecardFilteredByDate = applyDateFilterWithAncestors(
        filtered.map((row) => Object.assign({{}}, row)),
        false,
        activeSelection
      );
      applyEpicPlannedTotalsToAncestors(filteredByDate);
      applyEpicPlannedTotalsToAncestors(scorecardFilteredByDate);
      scorecardSourceRows = scorecardFilteredByDate;
      if (!onlyNoEntry) {{
        return filteredByDate;
      }}

      const noEntryRows = filteredByDate
        .filter((row) => hasWorkNoEntry(row))
        .map((row) => {{
          const item = Object.assign({{}}, row);
          item.level = 1;
          item.parent_id = null;
          return item;
        }});

      noEntryRows.sort((a, b) => {{
        const diff = noEntryCount(b) - noEntryCount(a);
        if (diff !== 0) {{
          return diff;
        }}
        const aType = String(a.type_label || "");
        const bType = String(b.type_label || "");
        if (aType !== bType) {{
          return aType.localeCompare(bType);
        }}
        return String(a.aspect || "").localeCompare(String(b.aspect || ""));
      }});

      return noEntryRows;
    }}

    function toNumberText(value) {{
      if (value === null || value === undefined || value === "") {{
        return "";
      }}
      const n = Number(value);
      if (Number.isFinite(n) && n === 0) {{
        return "No entry";
      }}
      return String(n.toFixed(2)).replace(/\\.00$/, "");
    }}

    function toMetricHtml(value) {{
      const text = toNumberText(value);
      if (text === "No entry") {{
        return '<span class="danger-chip">No entry</span>';
      }}
      return text;
    }}

    function toDeltaText(value) {{
      if (value === null || value === undefined || value === "") {{
        return "No entry";
      }}
      const n = Number(value);
      if (!Number.isFinite(n)) {{
        return "No entry";
      }}
      if (n === 0) {{
        return "0";
      }}
      return String(n.toFixed(2)).replace(/\\.00$/, "");
    }}

    function toDeltaHtml(value) {{
      const text = toDeltaText(value);
      if (text === "No entry") {{
        return '<span class="danger-chip">No entry</span>';
      }}
      const n = Number(value);
      if (!Number.isFinite(n)) {{
        return '<span class="danger-chip">No entry</span>';
      }}
      if (n < 0) {{
        return '<span class="delta-neg">' + text + '</span>';
      }}
      if (n > 0) {{
        return '<span class="delta-pos">' + text + '</span>';
      }}
      return '<span class="delta-zero">' + text + '</span>';
    }}

    function resourceLoggedText(row) {{
      const rowType = String(row && row.row_type || "");
      if (!WORK_NO_ENTRY_TYPES.has(rowType)) {{
        return "";
      }}
      const hours = Number(row && row.actual_hours);
      if (!Number.isFinite(hours)) {{
        return "No";
      }}
      return hours > 0 ? "Yes" : "No";
    }}

    function typeClassFromLabel(typeLabel) {{
      if (typeLabel === "Project") return "project";
      if (typeLabel === "Category") return "product";
      if (typeLabel === "Epic") return "rmi";
      if (typeLabel === "Story") return "story";
      if (typeLabel === "Subtask") return "subtask";
      if (typeLabel === "Assignee") return "assignee";
      if (typeLabel === "Bug") return "bug";
      return "";
    }}

    const TREE_LINE_COLOR = {{
      project: "#69aef5",
      product: "#c7cf43",
      rmi: "#9b7be5",
      story: "#5d88d8",
      subtask: "#9ad586",
      bug: "#e7b184",
      assignee: "#b7dfad",
    }};

    function rowColorClass(row) {{
      const labelClass = typeClassFromLabel(String(row.type_label || ""));
      if (labelClass === "bug") {{
        return "bug";
      }}
      if (labelClass) {{
        return labelClass;
      }}
      if (row.row_type === "project") return "project";
      if (row.row_type === "product") return "product";
      if (row.row_type === "rmi") return "rmi";
      if (row.row_type === "story") return "story";
      if (row.row_type === "subtask") return "subtask";
      if (row.row_type === "assignee") return "assignee";
      return "";
    }}

    function buildLineageClasses(row, byId) {{
      const chain = [];
      let current = row;
      while (current) {{
        const cls = rowColorClass(current);
        if (cls) {{
          if (cls === "bug" && current.row_type === "subtask") {{
            chain.push("subtask");
          }} else {{
            chain.push(cls);
          }}
        }}
        if (!current.parent_id) {{
          break;
        }}
        current = byId.get(current.parent_id) || null;
      }}
      return chain.reverse();
    }}

    function buildTreeLinesBackground(row, byId) {{
      const classes = buildLineageClasses(row, byId);
      if (!classes.length) {{
        return "";
      }}
      const gradients = [];
      const startX = 2;
      const step = 18;
      const width = 4;
      for (let i = 0; i < classes.length; i += 1) {{
        const color = TREE_LINE_COLOR[classes[i]] || "#cbd5e1";
        const x1 = startX + (i * step);
        const x2 = x1 + width;
        gradients.push("linear-gradient(to right, transparent " + x1 + "px, " + color + " " + x1 + "px, " + color + " " + x2 + "px, transparent " + x2 + "px)");
      }}
      return gradients.join(",");
    }}

    function typeChipHtml(typeLabel) {{
      const cls = typeClassFromLabel(typeLabel);
      if (!typeLabel) return "";
      if (!cls) return typeLabel;
      const codeByType = {{
        project: "PRJ",
        product: "CAT",
        rmi: "EPC",
        story: "STY",
        subtask: "TSK",
        bug: "BUG",
      }};
      const code = codeByType[cls] || "ROW";
      return '<span class="type-chip ' + cls + '"><span class="type-chip-code">' + code + '</span><span class="type-chip-text">' + typeLabel + "</span></span>";
    }}

    function isZeroMetric(value) {{
      if (value === null || value === undefined || value === "") {{
        return false;
      }}
      const n = Number(value);
      return Number.isFinite(n) && n === 0;
    }}

    function noEntryCount(row) {{
      let count = 0;
      if (isZeroMetric(row.man_days)) count += 1;
      if (isZeroMetric(row.man_hours)) count += 1;
      if (isZeroMetric(row.actual_hours)) count += 1;
      return count;
    }}

    function hasWorkNoEntry(row) {{
      return WORK_NO_ENTRY_TYPES.has(String(row.row_type || "")) && noEntryCount(row) > 0;
    }}

    function updateNoEntryToggle(filteredRows) {{
      const totalWorkNoEntryCount = (filteredRows || []).filter((row) => hasWorkNoEntry(row)).length;
      const hasNoEntry = totalWorkNoEntryCount > 0;
      toggleNoEntryButton.textContent = "No Entry <" + String(totalWorkNoEntryCount) + ">";
      toggleNoEntryButton.disabled = !hasNoEntry;
      toggleNoEntryButton.classList.toggle("alert", hasNoEntry);
      if (!hasNoEntry) {{
        onlyNoEntry = false;
      }}
      toggleNoEntryButton.classList.toggle("active", hasNoEntry && onlyNoEntry);
    }}

    function hideDescendants(id) {{
      const children = childrenByParent.get(id) || [];
      for (const childId of children) {{
        const child = rowsById.get(childId);
        if (!child) continue;
        child.el.style.display = "none";
        hideDescendants(childId);
      }}
    }}

    function showChildren(id) {{
      const children = childrenByParent.get(id) || [];
      for (const childId of children) {{
        const child = rowsById.get(childId);
        if (!child) continue;
        child.el.style.display = "";
        if (collapsed.has(childId)) {{
          hideDescendants(childId);
        }} else {{
          showChildren(childId);
        }}
      }}
    }}

    function updateToggleIcon(id) {{
      const row = rowsById.get(id);
      if (!row || !row.toggle) return;
      row.toggle.textContent = collapsed.has(id) ? "+" : "-";
    }}

    function getSearchBlob(row) {{
      const warningText = row.missing_parent_reason === "missing_rmi_parent"
        ? "missing parent missing rmi parent orphan"
        : (row.missing_parent_reason === "missing_story_parent"
            ? "missing parent missing story parent orphan"
            : "");
      const noEntryText = ((row.row_type === "rmi" || row.row_type === "story" || row.row_type === "subtask")
        && (isZeroMetric(row.man_days) || isZeroMetric(row.man_hours) || isZeroMetric(row.actual_hours)))
        ? "no entry missing metrics zero approved days approved hours actual hours"
        : "";
      return [
        row.aspect || "",
        row.type_label || "",
        row.project_key || "",
        row.project_name || "",
        row.assignee_name || "",
        row.jira_key || "",
        row.jira_url || "",
        row.approved_days === "" ? "" : String(row.approved_days),
        row.approved_hours === "" ? "" : String(row.approved_hours),
        row.planned_days === "" ? "" : String(row.planned_days),
        row.planned_hours === "" ? "" : String(row.planned_hours),
        row.actual_hours === "" ? "" : String(row.actual_hours),
        row.actual_days === "" ? "" : String(row.actual_days),
        row.delta_hours === "" ? "" : String(row.delta_hours),
        row.delta_days === "" ? "" : String(row.delta_days),
        resourceLoggedText(row),
        row.planned_start || "",
        row.planned_end || "",
        row.level === undefined || row.level === null ? "" : String(row.level),
        row.row_type || "",
        warningText,
        noEntryText,
      ].join(" ").toLowerCase();
    }}

    function collectAncestors(id, visible) {{
      let current = rowsById.get(id);
      while (current && current.parentId) {{
        visible.add(current.parentId);
        current = rowsById.get(current.parentId);
      }}
    }}

    function applyVisibility() {{
      const query = (activeSearchQuery || "").trim().toLowerCase();
      if (!query) {{
        for (const row of rowsById.values()) {{
          row.el.style.display = "";
        }}
        for (const id of collapsed) {{
          hideDescendants(id);
        }}
        searchMeta.textContent = "";
        updateStickyParentRows();
        return;
      }}

      const visible = new Set();
      let matchedCount = 0;
      for (const [id, row] of rowsById.entries()) {{
        if ((row.searchBlob || "").includes(query)) {{
          matchedCount += 1;
          visible.add(id);
          collectAncestors(id, visible);
        }}
      }}

      for (const [id, row] of rowsById.entries()) {{
        row.el.style.display = visible.has(id) ? "" : "none";
      }}
      searchMeta.textContent = matchedCount + " match" + (matchedCount === 1 ? "" : "es");
      updateStickyParentRows();
    }}

    function clearStickyParentRows() {{
      for (const row of rowsById.values()) {{
        if (!row || !row.el) {{
          continue;
        }}
        row.el.classList.remove("sticky-parent-row");
        row.el.style.removeProperty("--sticky-parent-top");
      }}
    }}

    function updateStickyParentRows() {{
      clearStickyParentRows();
      if (!tableWrapEl || !tableHeadEl) {{
        return;
      }}
      const wrapRect = tableWrapEl.getBoundingClientRect();
      const headerHeight = tableHeadEl.getBoundingClientRect().height || 0;
      const minVisibleTop = wrapRect.top + headerHeight + 1;
      let firstVisibleEntry = null;
      const renderedRows = tbody.querySelectorAll("tr");
      for (const tr of renderedRows) {{
        if (!tr || tr.style.display === "none") {{
          continue;
        }}
        if (tr.classList.contains("sticky-parent-row")) {{
          continue;
        }}
        const rect = tr.getBoundingClientRect();
        if (rect.bottom <= minVisibleTop) {{
          continue;
        }}
        if (rect.top >= wrapRect.bottom) {{
          break;
        }}
        const id = tr.dataset && tr.dataset.id ? tr.dataset.id : "";
        firstVisibleEntry = rowsById.get(id) || null;
        if (firstVisibleEntry) {{
          break;
        }}
      }}
      if (!firstVisibleEntry) {{
        return;
      }}
      const activeAncestors = [];
      let current = firstVisibleEntry;
      while (current && current.parentId) {{
        const parent = rowsById.get(current.parentId);
        if (!parent || !parent.el || parent.el.style.display === "none") {{
          break;
        }}
        activeAncestors.unshift(parent);
        current = parent;
      }}
      let topOffset = headerHeight;
      for (const ancestor of activeAncestors) {{
        ancestor.el.classList.add("sticky-parent-row");
        ancestor.el.style.setProperty("--sticky-parent-top", String(topOffset) + "px");
        const rowHeight = ancestor.el.getBoundingClientRect().height || 32;
        topOffset += rowHeight;
      }}
    }}

    function toggleRow(id) {{
      if (!hasChildren(id)) return;
      if (collapsed.has(id)) {{
        collapsed.delete(id);
        showChildren(id);
      }} else {{
        collapsed.add(id);
        hideDescendants(id);
      }}
      updateToggleIcon(id);
      applyVisibility();
    }}

    function collapseToProjects() {{
      if (showProductCategorization) {{
        showProductCategorization = false;
        toggleProductButton.textContent = "Show Category";
        rerender(true);
      }}
      collapsed.clear();
      for (const [id, row] of rowsById.entries()) {{
        if ((row.level || 0) === 1 && hasChildren(id)) {{
          collapsed.add(id);
        }}
      }}
      for (const row of rowsById.values()) {{
        row.el.style.display = "";
      }}
      for (const id of collapsed) {{
        hideDescendants(id);
      }}
      for (const id of rowsById.keys()) {{
        updateToggleIcon(id);
      }}
      applyVisibility();
    }}

    function collapseToEpics() {{
      collapsed.clear();
      for (const [id, row] of rowsById.entries()) {{
        if (row.rowType === "rmi" && hasChildren(id)) {{
          collapsed.add(id);
        }}
      }}
      for (const row of rowsById.values()) {{
        row.el.style.display = "";
      }}
      for (const id of collapsed) {{
        hideDescendants(id);
      }}
      for (const id of rowsById.keys()) {{
        updateToggleIcon(id);
      }}
      applyVisibility();
    }}

    function expandAll() {{
      collapsed.clear();
      for (const row of rowsById.values()) {{
        row.el.style.display = "";
      }}
      for (const id of rowsById.keys()) {{
        updateToggleIcon(id);
      }}
      applyVisibility();
    }}

    function renderRows(sourceRows) {{
      const sourceById = new Map();
      for (const row of sourceRows) {{
        sourceById.set(row.id, row);
        if (row.parent_id) {{
          addChild(row.parent_id, row.id);
        }}
      }}

      for (const row of sourceRows) {{
        const tr = document.createElement("tr");
        tr.className = "lvl-" + String(row.level || 1);
        const typeLabel = String(row.type_label || "");
        const typeClass = rowColorClass(row);
        if (typeClass) {{
          tr.classList.add("row-type-" + typeClass);
        }}
        const noEntryWarning = ((row.row_type === "rmi" || row.row_type === "story" || row.row_type === "subtask")
          && (isZeroMetric(row.man_days) || isZeroMetric(row.man_hours) || isZeroMetric(row.actual_hours)));
        const showTypeNoEntryChip = noEntryWarning && !row.has_defined_product_category;
        if (row.is_missing_parent || noEntryWarning) {{
          tr.classList.add("row-danger");
        }}
        tr.dataset.id = String(row.id);
        tr.dataset.parentId = row.parent_id ? String(row.parent_id) : "";
        tr.dataset.level = String(row.level || 1);

        const tdAspect = document.createElement("td");
        tdAspect.className = "col-aspect";
        tdAspect.classList.add("tree-lines");
        const treeLinesBackground = buildTreeLinesBackground(row, sourceById);
        if (treeLinesBackground) {{
          tdAspect.style.backgroundImage = treeLinesBackground;
        }}
        const wrap = document.createElement("div");
        wrap.className = "aspect-cell";
        wrap.style.paddingLeft = (Math.max(0, (Number(row.level || 1) - 1)) * 18) + "px";

        const btn = document.createElement("button");
        btn.type = "button";
        btn.className = "toggle";
        if (hasChildren(row.id)) {{
          btn.textContent = "-";
          btn.addEventListener("click", () => toggleRow(row.id));
        }} else {{
          btn.textContent = ".";
          btn.classList.add("placeholder");
          btn.disabled = true;
        }}

        const text = document.createElement("span");
        text.className = "node-text";
        text.textContent = (row.row_type === "project" && (row.project_name || "").trim()) ? String(row.project_name || "").trim() : (row.aspect || "");

        wrap.appendChild(btn);
        wrap.appendChild(text);
        if ((row.row_type === "rmi" || row.row_type === "story" || row.row_type === "subtask") && row.jira_url) {{
          const jiraLink = document.createElement("a");
          jiraLink.className = "jira-link";
          jiraLink.href = row.jira_url;
          jiraLink.target = "_blank";
          jiraLink.rel = "noopener noreferrer";
          jiraLink.title = "Open JIRA";
          jiraLink.setAttribute("aria-label", "Open JIRA");
          jiraLink.innerHTML = '<span class="material-icons-outlined" aria-hidden="true">open_in_new</span>';
          wrap.appendChild(jiraLink);
        }}
        tdAspect.appendChild(wrap);

        const tdType = document.createElement("td");
        tdType.className = "col-type";
        let typeHtml = typeChipHtml(typeLabel);
        if (row.is_missing_parent) {{
          let warningText = "Missing Parent";
          if (row.missing_parent_reason === "missing_rmi_parent") {{
            warningText = "Missing RMI Parent";
          }} else if (row.missing_parent_reason === "missing_story_parent") {{
            warningText = "Missing Story Parent";
          }}
          typeHtml += ' <span class="danger-chip">' + warningText + '</span>';
        }}
        if (showTypeNoEntryChip) {{
          typeHtml += ' <span class="danger-chip">No entry</span>';
        }}
        tdType.innerHTML = typeHtml;

        const tdAssignee = document.createElement("td");
        tdAssignee.textContent = row.assignee_name || "";

        const tdApprovedDays = document.createElement("td");
        tdApprovedDays.className = "num";
        tdApprovedDays.innerHTML = toMetricHtml(row.approved_days);

        const tdPlannedDays = document.createElement("td");
        tdPlannedDays.className = "num";
        tdPlannedDays.innerHTML = toMetricHtml(row.planned_days);

        const tdActualDays = document.createElement("td");
        tdActualDays.className = "num";
        tdActualDays.innerHTML = toMetricHtml(row.actual_days);

        const tdDeltaDays = document.createElement("td");
        tdDeltaDays.className = "num";
        tdDeltaDays.innerHTML = toDeltaHtml(row.delta_days);

        const tdApprovedHours = document.createElement("td");
        tdApprovedHours.className = "num";
        tdApprovedHours.innerHTML = toMetricHtml(row.approved_hours);

        const tdPlannedHours = document.createElement("td");
        tdPlannedHours.className = "num";
        tdPlannedHours.innerHTML = toMetricHtml(row.planned_hours);

        const tdActual = document.createElement("td");
        tdActual.className = "num";
        tdActual.innerHTML = toMetricHtml(row.actual_hours);

        const tdDeltaHours = document.createElement("td");
        tdDeltaHours.className = "num";
        tdDeltaHours.innerHTML = toDeltaHtml(row.delta_hours);

        const tdResourceLogged = document.createElement("td");
        tdResourceLogged.className = "col-resource";
        tdResourceLogged.textContent = resourceLoggedText(row);

        const tdStart = document.createElement("td");
        tdStart.className = "col-date";
        tdStart.textContent = row.planned_start || "";

        const tdEnd = document.createElement("td");
        tdEnd.className = "col-date";
        tdEnd.textContent = row.planned_end || "";

        tr.appendChild(tdAspect);
        tr.appendChild(tdType);
        tr.appendChild(tdAssignee);
        tr.appendChild(tdApprovedDays);
        tr.appendChild(tdPlannedDays);
        tr.appendChild(tdActualDays);
        tr.appendChild(tdDeltaDays);
        tr.appendChild(tdApprovedHours);
        tr.appendChild(tdPlannedHours);
        tr.appendChild(tdActual);
        tr.appendChild(tdDeltaHours);
        tr.appendChild(tdResourceLogged);
        tr.appendChild(tdStart);
        tr.appendChild(tdEnd);
        tbody.appendChild(tr);

        rowsById.set(row.id, {{
          id: row.id,
          level: Number(row.level || 1),
          rowType: String(row.row_type || ""),
          parentId: row.parent_id || null,
          el: tr,
          toggle: hasChildren(row.id) ? btn : null,
          searchBlob: getSearchBlob(row),
        }});
      }}
    }}

    function rerender(resetCollapsed, options) {{
      const nextOptions = options && typeof options === "object" ? options : {{}};
      tbody.innerHTML = "";
      clearTreeState();
      if (resetCollapsed) {{
        collapsed.clear();
      }}
      const renderSelection = buildScorecardSelectionSnapshot(nextOptions.selection);
      const displayRows = buildDisplayRows(renderSelection);
      rowCountNode.textContent = String(displayRows.length);
      updateNoEntryToggle(displayRows);
      const requestVersion = ++scorecardUpdateVersion;
      if (nextOptions.projectFilterLoading || (projectFilterProgress && !projectFilterProgress.hidden)) {{
        setProjectFilterLoading(true, requestVersion);
      }}
      if (nextOptions.teamFilterLoading || (teamFilterProgress && !teamFilterProgress.hidden)) {{
        setTeamFilterLoading(true, requestVersion);
      }}
      renderCapacityProfileDetails();
      void updateScoreCards(scorecardSourceRows, requestVersion, renderSelection);
      renderRows(displayRows);
      applyVisibility();
    }}

    initializeThemeToggle();
    initializeDensityToggle();
    initializeAspectColumnResize();
    initializeCapacityDrawerWidth();
    initializeProjectFilter();
    initializeTeamFilter().catch((error) => {{
      console.warn("Failed to initialize team filter:", error);
    }});
    if (!hasCapacityApi) {{
      setCapacityStatus("Read-only in static mode. Open via server to load latest saved profiles.", "info");
    }} else {{
      setCapacityStatus("First available profile is auto-applied on load.", "info");
    }}
    applyProfilesPayload(capacityProfiles);
    if (hasManagedFieldsApi) {{
      refreshManagedFieldsFromApi().catch((error) => {{
        console.warn("Failed to load managed fields:", error);
      }});
    }}
    captureOriginalMetrics();
    const storedActualHoursMode = localStorage.getItem(ACTUAL_HOURS_MODE_STORAGE_KEY);
    if (storedActualHoursMode === "log_date" || storedActualHoursMode === "planned_dates") {{
      pendingActualHoursMode = storedActualHoursMode;
    }}
    const storedPlannedHoursSource = localStorage.getItem(PLANNED_HOURS_SOURCE_STORAGE_KEY);
    pendingPlannedHoursSource = normalizePlannedHoursSource(storedPlannedHoursSource || DEFAULT_PLANNED_HOURS_SOURCE);
    normalizePendingDateRange();
    selectedDateFrom = pendingDateFrom;
    selectedDateTo = pendingDateTo;
    selectedActualHoursMode = pendingActualHoursMode;
    selectedPlannedHoursSource = pendingPlannedHoursSource;
    if (!hasNestedActualsApi) {{
      setDateFilterStatus("Date apply works without live recompute in file mode.");
    }}
    updateDateRangeApplyState();
    rerender(true);
    if (hasNestedTreeApi) {{
      fetch("/api/nested-view/tree", {{ method: "GET", headers: {{ Accept: "application/json" }} }})
        .then((r) => r.json())
        .then((data) => {{
          if (data && data.ok === true && Array.isArray(data.rows)) {{
            reportData.rows = data.rows;
            reportData.generated_at = data.generated_at || reportData.generated_at;
            reportData.source_file = data.source_file || reportData.source_file;
            allRows = data.rows;
            refreshProjectFilterOptions({{ preserveSelection: true }});
            if (document.getElementById("generated-at")) {{
              document.getElementById("generated-at").textContent = reportData.generated_at || "-";
            }}
            rerender(true);
          }}
        }})
        .catch((err) => {{ console.warn("Nested view tree from API:", err && err.message || err); }});
    }}
    if (tableWrapEl) {{
      tableWrapEl.addEventListener("scroll", () => {{
        updateStickyParentRows();
      }}, {{ passive: true }});
    }}
    window.addEventListener("resize", () => {{
      updateStickyParentRows();
    }});
    if (capacityProfileSelectEl) {{
      capacityProfileSelectEl.addEventListener("change", () => {{
        renderCapacityProfileDetails();
      }});
    }}
    if (capacityProfileApplyEl) {{
      capacityProfileApplyEl.addEventListener("click", () => {{
        applySelectedCapacityProfile();
      }});
    }}
    if (capacityProfileRefreshEl) {{
      capacityProfileRefreshEl.addEventListener("click", async () => {{
        try {{
          const refreshed = await refreshCapacityProfilesFromApi();
          if (refreshed) {{
            setCapacityStatus("Profiles refreshed.", "success");
          }}
        }} catch (error) {{
          setCapacityStatus(String(error && error.message || error || "Failed to refresh profiles."), "error");
        }}
      }});
    }}
    if (capacityProfileResetEl) {{
      capacityProfileResetEl.addEventListener("click", () => {{
        resetCapacityProfileOverride();
      }});
    }}
    const initialHeaderCollapsed = localStorage.getItem(HEADER_COLLAPSED_STORAGE_KEY) === "1";
    setHeaderCollapsed(initialHeaderCollapsed);
    if (headerToggleButton) {{
      headerToggleButton.addEventListener("click", () => {{
        const currentlyCollapsed = headerSectionEl ? headerSectionEl.classList.contains("is-collapsed") : false;
        setHeaderCollapsed(!currentlyCollapsed);
      }});
    }}
    setCapacityProfileDrawerOpen(false);
    if (totalPlannedCardEl) {{
      totalPlannedCardEl.classList.add("is-expandable");
      totalPlannedCardEl.setAttribute("role", "button");
      totalPlannedCardEl.setAttribute("tabindex", "0");
      totalPlannedCardEl.setAttribute("aria-controls", "score-total-planned-details");
      totalPlannedCardEl.setAttribute("aria-expanded", "false");
      totalPlannedCardEl.addEventListener("click", (event) => {{
        const target = event.target instanceof Element ? event.target : null;
        if (target && (target.closest(".score-info") || target.closest(".score-action-btn") || target.closest("a") || target.closest(".score-details-panel"))) {{
          return;
        }}
        setTotalPlannedDetailsOpen(!totalPlannedDetailsOpen);
      }});
      totalPlannedCardEl.addEventListener("keydown", (event) => {{
        if (event.key === "Enter" || event.key === " ") {{
          event.preventDefault();
          setTotalPlannedDetailsOpen(!totalPlannedDetailsOpen);
        }}
      }});
    }}
    if (totalLoggedCardEl) {{
      totalLoggedCardEl.classList.add("is-expandable");
      totalLoggedCardEl.setAttribute("role", "button");
      totalLoggedCardEl.setAttribute("tabindex", "0");
      totalLoggedCardEl.setAttribute("aria-controls", "score-total-logged-details");
      totalLoggedCardEl.setAttribute("aria-expanded", "false");
      totalLoggedCardEl.addEventListener("click", (event) => {{
        const target = event.target instanceof Element ? event.target : null;
        if (target && (target.closest(".score-info") || target.closest(".score-action-btn") || target.closest("a") || target.closest(".score-details-panel") || target.closest("#score-total-logged-include-bugs"))) {{
          return;
        }}
        setTotalActualDetailsOpen(!totalActualDetailsOpen);
      }});
      totalLoggedCardEl.addEventListener("keydown", (event) => {{
        if (event.key === "Enter" || event.key === " ") {{
          event.preventDefault();
          setTotalActualDetailsOpen(!totalActualDetailsOpen);
        }}
      }});
    }}
    if (totalLoggedIncludeBugsEl) {{
      totalLoggedIncludeBugsEl.checked = !!totalActualIncludeBugs;
      totalLoggedIncludeBugsEl.addEventListener("change", () => {{
        totalActualIncludeBugs = !!totalLoggedIncludeBugsEl.checked;
        const requestVersion = ++scorecardUpdateVersion;
        if (projectFilterProgress && !projectFilterProgress.hidden) {{
          setProjectFilterLoading(true, requestVersion);
        }}
        if (teamFilterProgress && !teamFilterProgress.hidden) {{
          setTeamFilterLoading(true, requestVersion);
        }}
        void updateScoreCards(scorecardSourceRows, requestVersion, buildScorecardSelectionSnapshot());
      }});
    }}
    if (totalLeavesPlannedCardEl) {{
      totalLeavesPlannedCardEl.classList.add("is-expandable");
      totalLeavesPlannedCardEl.setAttribute("role", "button");
      totalLeavesPlannedCardEl.setAttribute("tabindex", "0");
      totalLeavesPlannedCardEl.setAttribute("aria-controls", "score-total-leaves-planned-details");
      totalLeavesPlannedCardEl.setAttribute("aria-expanded", "false");
      totalLeavesPlannedCardEl.addEventListener("click", (event) => {{
        const target = event.target instanceof Element ? event.target : null;
        if (target && (target.closest(".score-info") || target.closest(".score-action-btn") || target.closest("a") || target.closest(".score-details-panel"))) {{
          return;
        }}
        setTotalLeavesPlannedDetailsOpen(!totalLeavesPlannedDetailsOpen);
      }});
      totalLeavesPlannedCardEl.addEventListener("keydown", (event) => {{
        if (event.key === "Enter" || event.key === " ") {{
          event.preventDefault();
          setTotalLeavesPlannedDetailsOpen(!totalLeavesPlannedDetailsOpen);
        }}
      }});
    }}
    if (capacityProfileToggleEl) {{
      capacityProfileToggleEl.addEventListener("click", () => {{
        const expanded = capacityProfileToggleEl.getAttribute("aria-expanded") === "true";
        setCapacityProfileDrawerOpen(!expanded);
      }});
    }}
    if (capacityProfileOverlayEl) {{
      capacityProfileOverlayEl.addEventListener("click", () => {{
        setCapacityProfileDrawerOpen(false);
      }});
    }}
    if (capacityProfileCloseEl) {{
      capacityProfileCloseEl.addEventListener("click", () => {{
        setCapacityProfileDrawerOpen(false);
      }});
    }}
    if (capacityProfileResizeHandleEl) {{
      capacityProfileResizeHandleEl.addEventListener("pointerdown", (event) => {{
        startCapacityDrawerResize(event);
      }});
      capacityProfileResizeHandleEl.addEventListener("pointermove", (event) => {{
        if (capacityDrawerPointerId === null || event.pointerId !== capacityDrawerPointerId) {{
          return;
        }}
        event.preventDefault();
        handleCapacityDrawerResizeClientX(event.clientX);
      }});
      capacityProfileResizeHandleEl.addEventListener("pointerup", (event) => {{
        if (capacityDrawerPointerId !== null && event.pointerId === capacityDrawerPointerId) {{
          stopCapacityDrawerResize();
        }}
      }});
      capacityProfileResizeHandleEl.addEventListener("pointercancel", (event) => {{
        if (capacityDrawerPointerId !== null && event.pointerId === capacityDrawerPointerId) {{
          stopCapacityDrawerResize();
        }}
      }});
    }}
    if (scoreCapacityProfileOpenButton) {{
      scoreCapacityProfileOpenButton.addEventListener("click", () => {{
        setHeaderCollapsed(false);
        setCapacityProfileDrawerOpen(true);
        if (capacityProfileSelectEl) {{
          capacityProfileSelectEl.focus();
        }}
      }});
    }}
    document.addEventListener("keydown", (event) => {{
      if (event.key === "Escape" && capacityProfileDrawerEl && capacityProfileDrawerEl.classList.contains("is-open")) {{
        setCapacityProfileDrawerOpen(false);
      }}
    }});
    window.addEventListener("pointerup", () => {{
      if (capacityDrawerPointerId !== null) {{
        stopCapacityDrawerResize();
      }}
    }});
    window.addEventListener("resize", () => {{
      if (window.innerWidth <= 760) {{
        stopCapacityDrawerResize();
      }}
    }});
    if (viewOptionsToggle) {{
      viewOptionsToggle.addEventListener("click", () => {{
        toggleViewOptionsMenu();
      }});
      viewOptionsToggle.addEventListener("keydown", (event) => {{
        if (event.key === "Enter" || event.key === " ") {{
          event.preventDefault();
          toggleViewOptionsMenu();
        }}
        if (event.key === "Escape") {{
          closeViewOptionsMenu();
        }}
      }});
    }}
    searchInput.addEventListener("input", () => {{
      activeSearchQuery = searchInput.value || "";
      applyVisibility();
    }});
    clearSearchButton.addEventListener("click", () => {{
      searchInput.value = "";
      activeSearchQuery = "";
      applyVisibility();
      searchInput.focus();
    }});
    toggleProductButton.addEventListener("click", () => {{
      showProductCategorization = !showProductCategorization;
      toggleProductButton.textContent = showProductCategorization
        ? "Hide Category"
        : "Show Category";
      rerender(true);
      closeViewOptionsMenu();
    }});
    toggleNoEntryButton.addEventListener("click", () => {{
      if (toggleNoEntryButton.disabled) {{
        return;
      }}
      onlyNoEntry = !onlyNoEntry;
      rerender(true);
      closeViewOptionsMenu();
    }});
    if (dateFilterFromInput) {{
      dateFilterFromInput.value = pendingDateFrom;
      dateFilterFromInput.addEventListener("change", () => {{
        pendingDateFrom = dateFilterFromInput.value || DEFAULT_DATE_FROM;
        normalizePendingDateRange();
        updateDateRangeApplyState();
      }});
    }}
    if (dateFilterToInput) {{
      dateFilterToInput.value = pendingDateTo;
      dateFilterToInput.addEventListener("change", () => {{
        pendingDateTo = dateFilterToInput.value || DEFAULT_DATE_TO;
        normalizePendingDateRange();
        updateDateRangeApplyState();
      }});
    }}
    if (actualHoursModeSelect) {{
      actualHoursModeSelect.value = pendingActualHoursMode;
      actualHoursModeSelect.addEventListener("change", () => {{
        const nextValue = String(actualHoursModeSelect.value || DEFAULT_ACTUAL_HOURS_MODE);
        pendingActualHoursMode = (nextValue === "planned_dates") ? "planned_dates" : "log_date";
        normalizePendingDateRange();
        updateDateRangeApplyState();
      }});
    }}
    if (extendedActualHoursToggle) {{
      extendedActualHoursToggle.checked = pendingActualHoursMode === "planned_dates";
      extendedActualHoursToggle.addEventListener("change", () => {{
        pendingActualHoursMode = extendedActualHoursToggle.checked ? "planned_dates" : "log_date";
        if (actualHoursModeSelect) {{
          actualHoursModeSelect.value = pendingActualHoursMode;
        }}
        normalizePendingDateRange();
        updateDateRangeApplyState();
        applyPendingDateRange();
      }});
    }}
    if (plannedHoursSourceSelect) {{
      plannedHoursSourceSelect.value = pendingPlannedHoursSource;
      plannedHoursSourceSelect.addEventListener("change", () => {{
        pendingPlannedHoursSource = normalizePlannedHoursSource(plannedHoursSourceSelect.value || DEFAULT_PLANNED_HOURS_SOURCE);
        normalizePendingDateRange();
        updateDateRangeApplyState();
      }});
    }}
    if (dateFilterApplyButton) {{
      dateFilterApplyButton.addEventListener("click", () => {{
        applyPendingDateRange();
      }});
    }}
    if (dateFilterResetButton) {{
      dateFilterResetButton.addEventListener("click", () => {{
        pendingDateFrom = DEFAULT_DATE_FROM;
        pendingDateTo = DEFAULT_DATE_TO;
        if (dateFilterFromInput) dateFilterFromInput.value = DEFAULT_DATE_FROM;
        if (dateFilterToInput) dateFilterToInput.value = DEFAULT_DATE_TO;
        normalizePendingDateRange();
        updateDateRangeApplyState();
      }});
    }}
    document.getElementById("expand-all").addEventListener("click", expandAll);
    document.getElementById("collapse-all").addEventListener("click", collapseToProjects);
    if (collapseEpicsButton) {{
      collapseEpicsButton.addEventListener("click", collapseToEpics);
    }}

    (function initAdvancedFilters() {{
      var advToggle = document.getElementById("adv-filter-toggle");
      var advMenu = document.getElementById("adv-filter-menu");
      if (!advToggle || !advMenu) return;

      advToggle.addEventListener("click", function (e) {{
        e.stopPropagation();
        var open = advToggle.getAttribute("aria-expanded") === "true";
        advToggle.setAttribute("aria-expanded", String(!open));
        advMenu.hidden = open;
      }});
      document.addEventListener("click", function (e) {{
        if (!advMenu.contains(e.target) && e.target !== advToggle && !advToggle.contains(e.target)) {{
          advToggle.setAttribute("aria-expanded", "false");
          advMenu.hidden = true;
        }}
      }});

      function computePreset(key) {{
        var today = new Date();
        var iso = function (d) {{ return d.toISOString().slice(0, 10); }};
        if (key === "last30") {{
          var f = new Date(today); f.setDate(f.getDate() - 30);
          return {{ from: iso(f), to: iso(today) }};
        }}
        if (key === "lastMonth") {{
          var f2 = new Date(today.getFullYear(), today.getMonth() - 1, 1);
          var t2 = new Date(today.getFullYear(), today.getMonth(), 0);
          return {{ from: iso(f2), to: iso(t2) }};
        }}
        if (key === "currentMonth") {{
          var fcm = new Date(today.getFullYear(), today.getMonth(), 1);
          var tcm = new Date(today.getFullYear(), today.getMonth() + 1, 0);
          return {{ from: iso(fcm), to: iso(tcm) }};
        }}
        if (key === "last90") {{
          var f3 = new Date(today); f3.setDate(f3.getDate() - 90);
          return {{ from: iso(f3), to: iso(today) }};
        }}
        if (key === "lastQuarter") {{
          var q = Math.floor(today.getMonth() / 3);
          var pq = q === 0 ? 3 : q - 1;
          var yr = q === 0 ? today.getFullYear() - 1 : today.getFullYear();
          return {{ from: iso(new Date(yr, pq * 3, 1)), to: iso(new Date(yr, pq * 3 + 3, 0)) }};
        }}
        if (key === "currentQuarter") {{
          var cq = Math.floor(today.getMonth() / 3);
          return {{ from: iso(new Date(today.getFullYear(), cq * 3, 1)), to: iso(new Date(today.getFullYear(), cq * 3 + 3, 0)) }};
        }}
        return null;
      }}

      document.querySelectorAll(".adv-filter-item[data-preset]").forEach(function (btn) {{
        btn.addEventListener("click", function () {{
          var preset = computePreset(btn.dataset.preset);
          if (!preset) return;
          pendingDateFrom = preset.from;
          pendingDateTo = preset.to;
          if (dateFilterFromInput) dateFilterFromInput.value = preset.from;
          if (dateFilterToInput) dateFilterToInput.value = preset.to;
          normalizePendingDateRange();
          updateDateRangeApplyState();
          advToggle.setAttribute("aria-expanded", "false");
          advMenu.hidden = true;
        }});
      }});

    }})();
  </script>
<script src="shared-nav.js"></script>
</body>
</html>
"""


def main() -> None:
    base_dir = Path(__file__).resolve().parent
    input_name = os.getenv("JIRA_NESTED_VIEW_XLSX_PATH", DEFAULT_INPUT_XLSX).strip() or DEFAULT_INPUT_XLSX
    output_name = os.getenv("JIRA_NESTED_VIEW_HTML_PATH", DEFAULT_OUTPUT_HTML).strip() or DEFAULT_OUTPUT_HTML
    capacity_db_name = os.getenv("JIRA_ASSIGNEE_CAPACITY_DB_PATH", DEFAULT_CAPACITY_DB).strip() or DEFAULT_CAPACITY_DB
    exports_db_name = os.getenv("JIRA_EXPORTS_DB_PATH", DEFAULT_EXPORTS_DB).strip() or DEFAULT_EXPORTS_DB
    leave_name = os.getenv("JIRA_LEAVE_REPORT_XLSX_PATH", DEFAULT_LEAVE_REPORT_INPUT_XLSX).strip() or DEFAULT_LEAVE_REPORT_INPUT_XLSX

    input_path = _resolve_path(input_name, base_dir)
    output_path = _resolve_path(output_name, base_dir)
    capacity_db_path = _resolve_path(capacity_db_name, base_dir)
    exports_db_path = _resolve_path(exports_db_name, base_dir)
    leave_path = _resolve_path(leave_name, base_dir)
    canonical_run_id = _to_text(os.getenv("JIRA_CANONICAL_RUN_ID"))

    source_file = str(input_path)
    try:
        rows = _load_nested_rows_from_canonical_db(
            capacity_db_path,
            exports_db_path=exports_db_path,
            run_id=canonical_run_id,
        )
        source_file = "canonical_db"
    except Exception:
        rows = _load_nested_rows(input_path)
    capacity_profiles = _load_capacity_profiles(capacity_db_path)
    leave_daily_rows = _load_leave_daily_rows(leave_path)
    leave_subtask_rows = _load_leave_subtask_rows(leave_path)
    data = {
        "generated_at": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC"),
        "source_file": source_file,
        "rows": rows,
        "capacity_profiles": capacity_profiles,
        "leave_daily_rows": leave_daily_rows,
        "leave_subtask_rows": leave_subtask_rows,
        "jira_base_url": _jira_base_url(),
    }
    html = _build_html(data)
    output_path.write_text(html, encoding="utf-8")
    served_output_path = base_dir / REPORT_HTML_DIRNAME / output_path.name
    if served_output_path.resolve() != output_path.resolve():
        served_output_path.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(output_path, served_output_path)

    print(f"Source data: {source_file}")
    print(f"Rows loaded: {len(rows)}")
    print(f"Capacity profiles loaded: {len(capacity_profiles)}")
    print(f"Leave daily rows loaded: {len(leave_daily_rows)}")
    print(f"Leave subtask rows loaded: {len(leave_subtask_rows)}")
    print(f"HTML report written: {output_path}")
    if served_output_path.resolve() != output_path.resolve():
        print(f"Served HTML synced: {served_output_path}")


if __name__ == "__main__":
    main()
