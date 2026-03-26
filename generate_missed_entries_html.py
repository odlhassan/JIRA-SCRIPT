"""
Generate an interactive missed-entries report from 1_jira_work_items_export.xlsx.
"""
from __future__ import annotations

import json
import os
import re
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

from canonical_report_data import load_canonical_issues, resolve_canonical_run_id

DEFAULT_INPUT_XLSX = "1_jira_work_items_export.xlsx"
DEFAULT_OUTPUT_HTML = "missed_entries.html"


def _resolve_path(value: str, base_dir: Path) -> Path:
    path = Path(value)
    if path.is_absolute():
        return path
    return base_dir / path


def _to_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize_date_text(value) -> str:
    text = _to_text(value)
    if not text:
        return ""
    iso = re.match(r"^(\d{4}-\d{2}-\d{2})", text)
    if iso:
        return iso.group(1)
    for fmt in ("%d-%b-%Y", "%d-%B-%Y", "%m/%d/%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    return ""


def _pick_first(row: dict[str, object], names: list[str]) -> str:
    for name in names:
        if name in row:
            value = _to_text(row.get(name))
            if value:
                return value
    return ""


def _to_float(value: object) -> float:
    text = _to_text(value)
    if not text:
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def _default_month_window() -> tuple[str, str]:
    now = datetime.now(timezone.utc)
    current = now.strftime("%Y-%m")
    year = now.year
    month = now.month - 1
    if month == 0:
        month = 12
        year -= 1
    previous = f"{year:04d}-{month:02d}"
    return previous, current


def _load_rows(input_path: Path) -> tuple[list[dict], str, str]:
    if not input_path.exists():
        raise FileNotFoundError(f"Work items workbook not found: {input_path}")

    wb = load_workbook(input_path, read_only=True, data_only=True)
    ws = wb.active
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header:
        wb.close()
        raise ValueError("Work items workbook has no header row.")

    headers = [_to_text(h) for h in header]
    lower_headers = [h.lower() for h in headers]
    date_values: list[str] = []
    rows: list[dict] = []
    jira_site = os.getenv("JIRA_SITE", "octopusdtlsupport").strip() or "octopusdtlsupport"
    base_url = f"https://{jira_site}.atlassian.net"

    for raw in ws.iter_rows(min_row=2, values_only=True):
        source = {lower_headers[i]: raw[i] for i in range(len(lower_headers))}
        issue_key = _pick_first(source, ["issue_key"])
        if not issue_key:
            continue

        issue_type = _pick_first(source, ["jira_issue_type", "issue_type"]) or "Unknown"
        assignee = _pick_first(source, ["assignee"])
        summary = _pick_first(source, ["summary"])
        jira_start_date = _normalize_date_text(
            _pick_first(source, ["start_date", "planned start date", "planned_start_date"])
        )
        jira_due_date = _normalize_date_text(
            _pick_first(source, ["end_date", "planned end date", "planned_end_date", "duedate"])
        )
        original_estimate = _pick_first(source, ["original_estimate"])
        total_hours_logged = _to_float(
            _pick_first(source, ["total_hours_logged", "total hours_logged", "hours_logged"])
        )
        jira_url = _pick_first(source, ["jira_url"])
        if not jira_url:
            jira_url = f"{base_url}/browse/{issue_key}"

        if jira_start_date:
            date_values.append(jira_start_date)
        if jira_due_date:
            date_values.append(jira_due_date)

        rows.append(
            {
                "issue_key": issue_key,
                "issue_type": issue_type,
                "assignee": assignee,
                "summary": summary,
                "jira_start_date": jira_start_date,
                "jira_due_date": jira_due_date,
                "original_estimate": original_estimate,
                "resource_logged_hours": "Yes" if total_hours_logged > 0 else "No",
                "jira_url": jira_url,
            }
        )

    wb.close()

    default_from, default_to = _default_month_window()

    return rows, default_from, default_to


def _load_rows_from_canonical_db(db_path: Path, run_id: str = "") -> tuple[list[dict], str, str]:
    effective_run_id = resolve_canonical_run_id(db_path, run_id)
    if not effective_run_id:
        return [], *_default_month_window()

    rows: list[dict] = []
    jira_site = os.getenv("JIRA_SITE", "octopusdtlsupport").strip() or "octopusdtlsupport"
    base_url = f"https://{jira_site}.atlassian.net"
    for source in load_canonical_issues(db_path, effective_run_id):
        issue_key = _to_text(source.get("issue_key")).upper()
        if not issue_key:
            continue
        total_hours_logged = _to_float(source.get("total_hours_logged"))
        rows.append(
            {
                "issue_key": issue_key,
                "issue_type": _to_text(source.get("issue_type")) or "Unknown",
                "assignee": _to_text(source.get("assignee")),
                "summary": _to_text(source.get("summary")),
                "jira_start_date": _normalize_date_text(source.get("start_date")),
                "jira_due_date": _normalize_date_text(source.get("due_date")),
                "original_estimate": round(_to_float(source.get("original_estimate_hours")), 2),
                "resource_logged_hours": "Yes" if total_hours_logged > 0 else "No",
                "jira_url": f"{base_url}/browse/{issue_key}",
            }
        )
    return rows, *_default_month_window()


def _build_html(payload: dict) -> str:
    data = json.dumps(payload, ensure_ascii=True)
    return f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Missed Entries Report</title>
  <link rel="stylesheet" href="https://fonts.googleapis.com/css2?family=Material+Symbols+Outlined:opsz,wght,FILL,GRAD@20..48,500,0,0">
  <style>
    :root {{
      --bg: #f3f6f9;
      --panel: #ffffff;
      --text: #1f2937;
      --muted: #6b7280;
      --line: #dbe3ea;
      --head: #0f4c5c;
      --head-text: #ffffff;
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      font-family: "Segoe UI", Tahoma, Verdana, sans-serif;
      color: var(--text);
      background:
        radial-gradient(1000px 280px at 10% -5%, #d7eef6 0%, transparent 60%),
        linear-gradient(180deg, #eef4f7, var(--bg));
    }}
    .page {{
      max-width: 1500px;
      margin: 0 auto;
      padding: 16px;
    }}
    .card {{
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 12px;
      padding: 14px;
      margin-bottom: 12px;
    }}
    .title {{
      margin: 0 0 6px;
      font-size: 1.2rem;
      font-weight: 700;
      color: #0b3142;
    }}
    .page-head {{
      display: flex;
      align-items: center;
      justify-content: space-between;
      gap: 10px;
      flex-wrap: wrap;
      margin-bottom: 6px;
    }}
    .meta {{
      margin: 0;
      color: var(--muted);
      font-size: 0.9rem;
    }}
    .controls {{
      display: grid;
      gap: 10px;
      grid-template-columns: repeat(auto-fit, minmax(260px, 1fr));
      margin-top: 12px;
    }}
    .control-block {{
      border: 1px solid var(--line);
      border-radius: 10px;
      padding: 10px;
      background: #f8fbfd;
    }}
    .control-label {{
      font-size: 0.8rem;
      font-weight: 700;
      color: #334155;
      text-transform: uppercase;
      letter-spacing: 0.02em;
      margin-bottom: 8px;
      display: block;
    }}
    .row {{
      display: flex;
      flex-wrap: wrap;
      gap: 8px;
      align-items: center;
    }}
    input[type="month"] {{
      border: 1px solid #b6c7d2;
      border-radius: 8px;
      padding: 6px 8px;
      font-size: 0.86rem;
      color: #12313f;
      background: #fff;
    }}
    .btn {{
      border: 1px solid #255f73;
      background: #0f4c5c;
      color: #fff;
      border-radius: 8px;
      padding: 6px 10px;
      cursor: pointer;
      font-size: 0.85rem;
      font-weight: 700;
    }}
    .field-option {{
      display: flex;
      align-items: center;
      gap: 6px;
      margin: 4px 0;
      font-size: 0.86rem;
    }}
    .summary-head {{
      display: flex;
      justify-content: space-between;
      align-items: center;
      gap: 10px;
      margin-bottom: 8px;
      flex-wrap: wrap;
    }}
    .chip {{
      display: inline-flex;
      align-items: center;
      border: 1px solid #fecaca;
      background: #fef2f2;
      color: #991b1b;
      border-radius: 999px;
      padding: 3px 9px;
      font-size: 0.8rem;
      font-weight: 700;
    }}
    table {{
      width: 100%;
      border-collapse: collapse;
      table-layout: fixed;
    }}
    .table-wrap {{
      overflow-x: auto;
    }}
    th {{
      background: var(--head);
      color: var(--head-text);
      text-align: left;
      padding: 8px 10px;
      font-size: 0.84rem;
      white-space: nowrap;
    }}
    td {{
      border-top: 1px solid var(--line);
      padding: 8px 10px;
      font-size: 0.86rem;
      vertical-align: top;
      background: #fff;
      word-break: break-word;
    }}
    tr:nth-child(even) td {{
      background: #fbfdff;
    }}
    .num {{
      text-align: right;
      white-space: nowrap;
    }}
    .empty {{
      color: var(--muted);
      font-style: italic;
    }}
    .jira-link {{
      color: #1d4ed8;
      font-weight: 700;
      text-decoration: none;
    }}
    .jira-link:hover {{
      text-decoration: underline;
    }}
    .missed-field {{
      display: inline-block;
      border: 1px solid #fecaca;
      border-radius: 999px;
      background: #fef2f2;
      color: #991b1b;
      font-size: 0.74rem;
      font-weight: 700;
      padding: 2px 8px;
      margin: 1px 4px 1px 0;
      white-space: nowrap;
    }}
    .assignee-toggle {{
      border: none;
      background: transparent;
      color: #1e3a8a;
      font-weight: 700;
      cursor: pointer;
      padding: 0;
      text-align: left;
    }}
    .assignee-toggle:hover {{
      text-decoration: underline;
    }}
    .assignee-detail-row {{
      display: none;
    }}
    .assignee-detail-row.open {{
      display: table-row;
    }}
    .assignee-detail-cell {{
      background: #f8fbfd !important;
      border-top: none;
      padding-top: 4px;
      padding-bottom: 12px;
    }}
    .assignee-jira-list {{
      margin: 0;
      padding-left: 18px;
    }}
    .assignee-jira-item {{
      margin: 5px 0;
      font-size: 0.84rem;
      color: #334155;
    }}
    .detail-workitem-table {{
      width: 100%;
      border-collapse: collapse;
      table-layout: fixed;
      margin-top: 4px;
    }}
    .detail-workitem-table th {{
      background: #e2ecf2;
      color: #12313f;
      font-size: 0.79rem;
      border-bottom: 1px solid #c8d9e3;
    }}
    .detail-workitem-table td {{
      background: #ffffff !important;
      border-top: 1px solid #d9e5ec;
      font-size: 0.82rem;
      vertical-align: middle;
    }}
    .status-chip {{
      display: inline-block;
      border-radius: 999px;
      padding: 2px 9px;
      font-size: 0.74rem;
      font-weight: 700;
      border: 1px solid transparent;
      white-space: nowrap;
    }}
    .status-yes {{
      color: #166534;
      background: #dcfce7;
      border-color: #86efac;
    }}
    .status-no {{
      color: #991b1b;
      background: #fee2e2;
      border-color: #fca5a5;
    }}
  </style>
  <link rel="stylesheet" href="shared-nav.css">
</head>
<body>
  <div class="page">
    <section class="card">
      <div class="page-head">
        <h1 class="title" style="margin:0">Missed Entries Report</h1>
        <div class="row">
          <button id="export-excel" class="btn" type="button">Export Excel</button>
          <button id="refresh-report" class="btn" type="button">Refresh Report</button>
        </div>
      </div>
      <p class="meta">
        Generated: <span id="generated-at"></span> |
        Source: <span id="source-file"></span> |
        Rows Loaded: <span id="rows-loaded"></span>
      </p>
      <div class="controls">
        <div class="control-block">
          <span class="control-label">Date Range (Month)</span>
          <div class="row">
            <label for="date-from">From</label>
            <input id="date-from" type="month">
            <label for="date-to">To</label>
            <input id="date-to" type="month">
            <button id="date-reset" class="btn" type="button">Reset</button>
          </div>
        </div>
        <div class="control-block">
          <span class="control-label">Missing Fields (Any Selected)</span>
          <label class="field-option"><input type="checkbox" class="field-checkbox" value="jira_start_date" checked> Jira Planned Start Date</label>
          <label class="field-option"><input type="checkbox" class="field-checkbox" value="jira_due_date" checked> Jira Planned Due Date</label>
          <label class="field-option"><input type="checkbox" class="field-checkbox" value="original_estimate" checked> Original Estimates</label>
        </div>
      </div>
    </section>

    <section class="card">
      <div class="summary-head">
        <h2 class="title" style="font-size:1rem;margin:0">Missed Entries by Assignee</h2>
        <span class="chip">Total Missed Entries: <span id="total-missed-count" style="margin-left:5px">0</span></span>
      </div>
      <div class="table-wrap">
        <table>
          <thead>
            <tr>
              <th>Assignee</th>
              <th class="num">Total Missed</th>
              <th class="num">Planned Start Missing</th>
              <th class="num">Planned Due Missing</th>
              <th class="num">Original Estimate Missing</th>
            </tr>
          </thead>
          <tbody id="summary-rows"></tbody>
        </table>
      </div>
    </section>

  </div>

  <script>
    const reportData = {data};
    const rows = reportData.rows || [];
    const fieldLabels = {{
      jira_start_date: "Jira Planned Start Date",
      jira_due_date: "Jira Planned Due Date",
      original_estimate: "Original Estimates",
    }};

    const generatedAtNode = document.getElementById("generated-at");
    const sourceFileNode = document.getElementById("source-file");
    const rowsLoadedNode = document.getElementById("rows-loaded");
    const refreshReportButton = document.getElementById("refresh-report");
    const exportExcelButton = document.getElementById("export-excel");
    const dateFromInput = document.getElementById("date-from");
    const dateToInput = document.getElementById("date-to");
    const dateResetButton = document.getElementById("date-reset");
    const totalMissedCountNode = document.getElementById("total-missed-count");
    const summaryRowsNode = document.getElementById("summary-rows");
    const fieldCheckboxes = Array.from(document.querySelectorAll(".field-checkbox"));

    const DEFAULT_DATE_FROM_MONTH = reportData.default_date_from || "2026-01";
    const DEFAULT_DATE_TO_MONTH = reportData.default_date_to || "2026-02";
    let selectedDateFromMonth = DEFAULT_DATE_FROM_MONTH;
    let selectedDateToMonth = DEFAULT_DATE_TO_MONTH;

    generatedAtNode.textContent = reportData.generated_at || "-";
    sourceFileNode.textContent = reportData.source_file || "-";
    rowsLoadedNode.textContent = String(rows.length);

    function asText(value) {{
      return String(value || "").trim();
    }}

    function parseDateValue(value) {{
      const text = asText(value);
      if (!text) return null;
      const m = text.match(/^(\\d{{4}})-(\\d{{2}})-(\\d{{2}})$/);
      if (!m) return null;
      return new Date(Number(m[1]), Number(m[2]) - 1, Number(m[3]));
    }}

    function parseMonthValue(value) {{
      const text = asText(value);
      const m = text.match(/^(\\d{{4}})-(\\d{{2}})$/);
      if (!m) return null;
      const year = Number(m[1]);
      const monthIndex = Number(m[2]) - 1;
      if (!Number.isFinite(year) || !Number.isFinite(monthIndex) || monthIndex < 0 || monthIndex > 11) {{
        return null;
      }}
      return {{ year, monthIndex }};
    }}

    function monthStart(value) {{
      const parsed = parseMonthValue(value);
      if (!parsed) return null;
      return new Date(parsed.year, parsed.monthIndex, 1);
    }}

    function monthEnd(value) {{
      const parsed = parseMonthValue(value);
      if (!parsed) return null;
      return new Date(parsed.year, parsed.monthIndex + 1, 0);
    }}

    function normalizeMonths() {{
      if (selectedDateFromMonth > selectedDateToMonth) {{
        const temp = selectedDateFromMonth;
        selectedDateFromMonth = selectedDateToMonth;
        selectedDateToMonth = temp;
      }}
      dateFromInput.value = selectedDateFromMonth;
      dateToInput.value = selectedDateToMonth;
    }}

    function rangesOverlap(startA, endA, startB, endB) {{
      if (!startA || !endA || !startB || !endB) return false;
      return startA.getTime() <= endB.getTime() && startB.getTime() <= endA.getTime();
    }}

    function matchesDateRange(row) {{
      const start = parseDateValue(row.jira_start_date);
      const due = parseDateValue(row.jira_due_date);
      if (!start && !due) {{
        return false;
      }}
      const rowStart = start || due;
      const rowEnd = due || start;
      const filterStart = monthStart(selectedDateFromMonth) || monthStart(DEFAULT_DATE_FROM_MONTH);
      const filterEnd = monthEnd(selectedDateToMonth) || monthEnd(DEFAULT_DATE_TO_MONTH);
      return rangesOverlap(filterStart, filterEnd, rowStart, rowEnd);
    }}

    function isMissing(value) {{
      return asText(value) === "";
    }}

    function selectedFieldKeys() {{
      return fieldCheckboxes.filter((node) => node.checked).map((node) => node.value);
    }}

    function missingFields(row, fields) {{
      const misses = [];
      for (const field of fields) {{
        if (isMissing(row[field])) {{
          misses.push(field);
        }}
      }}
      return misses;
    }}

    function normalizeAssignee(value) {{
      const text = asText(value);
      if (!text || text.toLowerCase() === "unassigned") {{
        return "Unassigned";
      }}
      return text;
    }}

    function buildFilteredRows() {{
      const selectedFields = selectedFieldKeys();
      if (!selectedFields.length) {{
        return [];
      }}
      const filtered = [];
      for (const row of rows) {{
        if (!matchesDateRange(row)) {{
          continue;
        }}
        const misses = missingFields(row, selectedFields);
        if (!misses.length) {{
          continue;
        }}
        const item = Object.assign({{}}, row);
        item._misses = misses;
        filtered.push(item);
      }}
      return filtered;
    }}

    function renderSummary(items) {{
      const byAssignee = new Map();
      for (const row of items) {{
        const assignee = normalizeAssignee(row.assignee);
        if (!byAssignee.has(assignee)) {{
          byAssignee.set(assignee, {{
            assignee,
            total: 0,
            jira_start_date: 0,
            jira_due_date: 0,
            original_estimate: 0,
            rows: [],
          }});
        }}
        const entry = byAssignee.get(assignee);
        entry.total += 1;
        for (const field of row._misses || []) {{
          entry[field] += 1;
        }}
        entry.rows.push(row);
      }}

      const rowsOut = Array.from(byAssignee.values()).sort((a, b) => {{
        const byTotal = b.total - a.total;
        if (byTotal !== 0) return byTotal;
        return a.assignee.localeCompare(b.assignee);
      }});

      summaryRowsNode.innerHTML = "";
      totalMissedCountNode.textContent = String(items.length);
      if (!rowsOut.length) {{
        const tr = document.createElement("tr");
        const td = document.createElement("td");
        td.colSpan = 5;
        td.className = "empty";
        td.textContent = "No missed entries for current filters.";
        tr.appendChild(td);
        summaryRowsNode.appendChild(tr);
        return;
      }}

      for (const entry of rowsOut) {{
        const assigneeId = "assignee-" + entry.assignee.toLowerCase().replace(/[^a-z0-9]+/g, "-");
        const tr = document.createElement("tr");
        tr.innerHTML = `
          <td><button type="button" class="assignee-toggle" data-target="${{assigneeId}}">${{entry.assignee}}</button></td>
          <td class="num">${{entry.total}}</td>
          <td class="num">${{entry.jira_start_date}}</td>
          <td class="num">${{entry.jira_due_date}}</td>
          <td class="num">${{entry.original_estimate}}</td>
        `;
        summaryRowsNode.appendChild(tr);

        const detailRow = document.createElement("tr");
        detailRow.className = "assignee-detail-row";
        detailRow.id = assigneeId;
        const itemsHtml = entry.rows
          .slice()
          .sort((a, b) => asText(a.issue_key).localeCompare(asText(b.issue_key)))
          .map((row) => {{
            const link = asText(row.jira_url);
            const misses = (row._misses || [])
              .map((field) => `<span class="missed-field">${{fieldLabels[field] || field}}</span>`)
              .join("");
            const issue = asText(row.issue_key);
            const issueHtml = link
              ? `<a class="jira-link" href="${{link}}" target="_blank" rel="noopener noreferrer">${{issue}}</a>`
              : issue;
            const hoursLogged = asText(row.resource_logged_hours).toLowerCase() === "yes";
            const hoursChip = hoursLogged
              ? '<span class="status-chip status-yes">Yes</span>'
              : '<span class="status-chip status-no">No</span>';
            return `
              <tr>
                <td>${{issueHtml}}</td>
                <td>${{asText(row.issue_type) || "-"}}</td>
                <td>${{misses || '<span class="empty">None</span>'}}</td>
                <td>${{hoursChip}}</td>
              </tr>
            `;
          }})
          .join("");
        detailRow.innerHTML = `
          <td colspan="5" class="assignee-detail-cell">
            <table class="detail-workitem-table">
              <thead>
                <tr>
                  <th>Work Item</th>
                  <th>Issue Type</th>
                  <th>Missing Fields</th>
                  <th>Resource Logged Hours</th>
                </tr>
              </thead>
              <tbody>
                ${{itemsHtml || '<tr><td colspan="4" class="assignee-jira-item empty">No Jira rows.</td></tr>'}}
              </tbody>
            </table>
          </td>
        `;
        summaryRowsNode.appendChild(detailRow);
      }}

      summaryRowsNode.querySelectorAll(".assignee-toggle").forEach((button) => {{
        button.addEventListener("click", () => {{
          const targetId = button.getAttribute("data-target");
          const row = document.getElementById(targetId);
          if (!row) return;
          row.classList.toggle("open");
        }});
      }});
    }}

    function buildSummaryRows(items) {{
      const byAssignee = new Map();
      for (const row of items) {{
        const assignee = normalizeAssignee(row.assignee);
        if (!byAssignee.has(assignee)) {{
          byAssignee.set(assignee, {{
            assignee,
            total: 0,
            jira_start_date: 0,
            jira_due_date: 0,
            original_estimate: 0,
          }});
        }}
        const entry = byAssignee.get(assignee);
        entry.total += 1;
        for (const field of row._misses || []) {{
          entry[field] += 1;
        }}
      }}
      return Array.from(byAssignee.values()).sort((a, b) => {{
        const byTotal = b.total - a.total;
        if (byTotal !== 0) return byTotal;
        return a.assignee.localeCompare(b.assignee);
      }});
    }}

    function exportToExcel() {{
      const items = buildFilteredRows();
      if (!items.length) {{
        alert("No missed entries for current filters.");
        return;
      }}
      if (!window.XLSX) {{
        alert("Excel export library failed to load. Refresh and try again.");
        return;
      }}

      const detailsRows = items
        .slice()
        .sort((a, b) => normalizeAssignee(a.assignee).localeCompare(normalizeAssignee(b.assignee)) || asText(a.issue_key).localeCompare(asText(b.issue_key)))
        .map((row) => {{
          const misses = (row._misses || []).map((field) => fieldLabels[field] || field).join(", ");
          return {{
            "Assignee": normalizeAssignee(row.assignee),
            "Work Item": asText(row.issue_key),
            "Issue Type": asText(row.issue_type),
            "Missing Fields": misses,
            "Resource Logged Hours": asText(row.resource_logged_hours),
            "Jira Planned Start Date": asText(row.jira_start_date),
            "Jira Planned Due Date": asText(row.jira_due_date),
            "Original Estimates": asText(row.original_estimate),
            "Summary": asText(row.summary),
            "Jira URL": asText(row.jira_url),
          }};
        }});

      const summaryRows = buildSummaryRows(items).map((entry) => ({{
        "Assignee": entry.assignee,
        "Total Missed": entry.total,
        "Planned Start Missing": entry.jira_start_date,
        "Planned Due Missing": entry.jira_due_date,
        "Original Estimate Missing": entry.original_estimate,
      }}));

      const wb = XLSX.utils.book_new();
      const wsSummary = XLSX.utils.json_to_sheet(summaryRows);
      const wsDetails = XLSX.utils.json_to_sheet(detailsRows);
      XLSX.utils.book_append_sheet(wb, wsSummary, "Summary");
      XLSX.utils.book_append_sheet(wb, wsDetails, "Details");

      const stamp = new Date().toISOString().replace(/[-:]/g, "").replace(/\\.\\d+Z$/, "Z");
      const fileName = `missed_entries_export_${{stamp}}.xlsx`;
      XLSX.writeFile(wb, fileName);
    }}

    function rerender() {{
      normalizeMonths();
      const filtered = buildFilteredRows();
      renderSummary(filtered);
    }}

    dateFromInput.value = selectedDateFromMonth;
    dateToInput.value = selectedDateToMonth;

    dateFromInput.addEventListener("change", () => {{
      selectedDateFromMonth = dateFromInput.value || DEFAULT_DATE_FROM_MONTH;
      rerender();
    }});
    dateToInput.addEventListener("change", () => {{
      selectedDateToMonth = dateToInput.value || DEFAULT_DATE_TO_MONTH;
      rerender();
    }});
    dateResetButton.addEventListener("click", () => {{
      selectedDateFromMonth = DEFAULT_DATE_FROM_MONTH;
      selectedDateToMonth = DEFAULT_DATE_TO_MONTH;
      rerender();
    }});
    refreshReportButton.addEventListener("click", () => {{
      rerender();
    }});
    fieldCheckboxes.forEach((node) => {{
      node.addEventListener("change", () => {{
        rerender();
      }});
    }});
    exportExcelButton.addEventListener("click", () => {{
      exportToExcel();
    }});

    rerender();
  </script>
<script src="https://cdn.jsdelivr.net/npm/xlsx@0.18.5/dist/xlsx.full.min.js"></script>
<script src="shared-nav.js"></script>
</body>
</html>
"""


def main() -> None:
    base_dir = Path(__file__).resolve().parent
    input_name = os.getenv("JIRA_EXPORT_XLSX_PATH", DEFAULT_INPUT_XLSX).strip() or DEFAULT_INPUT_XLSX
    output_name = os.getenv("JIRA_MISSED_ENTRIES_HTML_PATH", DEFAULT_OUTPUT_HTML).strip() or DEFAULT_OUTPUT_HTML
    db_name = os.getenv("JIRA_ASSIGNEE_HOURS_CAPACITY_DB_PATH", "assignee_hours_capacity.db").strip() or "assignee_hours_capacity.db"
    canonical_run_id = os.getenv("JIRA_CANONICAL_RUN_ID", "").strip()

    input_path = _resolve_path(input_name, base_dir)
    output_path = _resolve_path(output_name, base_dir)
    db_path = _resolve_path(db_name, base_dir)

    rows, default_from, default_to = _load_rows_from_canonical_db(db_path, canonical_run_id)
    source_file = "canonical_db"
    if not rows:
        rows, default_from, default_to = _load_rows(input_path)
        source_file = str(input_path)
    payload = {
        "generated_at": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC"),
        "source_file": source_file,
        "rows": rows,
        "default_date_from": default_from,
        "default_date_to": default_to,
    }
    html = _build_html(payload)
    output_path.write_text(html, encoding="utf-8")

    print(f"Source data: {source_file}")
    print(f"Rows loaded: {len(rows)}")
    print(f"HTML report written: {output_path}")


if __name__ == "__main__":
    main()

