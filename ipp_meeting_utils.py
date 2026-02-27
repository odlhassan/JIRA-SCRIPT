"""
Utilities for checking whether Jira issues appear in the latest IPP Meeting workbook.
"""
from __future__ import annotations

import os
import re
import shutil
import tempfile
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

DEFAULT_IPP_MEETING_XLSX_PATH = (
    r"C:\Users\hmalik\OneDrive - Octopus Digital\ALL DOCS\IPP Meeting\all ipp meetings.xlsx"
)

_ISSUE_KEY_PATTERN = re.compile(r"\b([A-Za-z][A-Za-z0-9]+-\d+)\b")
_DATE_PATTERN = re.compile(r"\b(\d{1,2})[-/\s]([A-Za-z]{3})[-/\s](\d{2,4})\b")
_MONTH_MAP = {
    "jan": 1,
    "feb": 2,
    "mar": 3,
    "apr": 4,
    "may": 5,
    "jun": 6,
    "jul": 7,
    "aug": 8,
    "sep": 9,
    "oct": 10,
    "nov": 11,
    "dec": 12,
}
_DEFAULT_PROJECT_KEYS = ["DIGITALLOG", "FF", "O2", "ODL", "MN"]


def get_ipp_meeting_xlsx_path() -> Path:
    path_value = os.getenv("IPP_MEETING_XLSX_PATH", DEFAULT_IPP_MEETING_XLSX_PATH).strip()
    return Path(path_value)


def normalize_issue_key(value: str) -> str:
    if not value:
        return ""
    match = _ISSUE_KEY_PATTERN.search(str(value))
    if not match:
        return ""
    return match.group(1).upper()


def load_ipp_issue_keys(path: Path | None = None) -> set[str]:
    workbook_path = path or get_ipp_meeting_xlsx_path()
    if not workbook_path.exists():
        print(f"Warning: IPP Meeting file not found: {workbook_path}")
        return set()
    temp_copy_path = None
    try:
        wb = load_workbook(workbook_path, read_only=True, data_only=True)
    except Exception:
        temp_dir = Path(tempfile.gettempdir())
        temp_copy_path = temp_dir / f"ipp_meeting_copy_{os.getpid()}.xlsx"
        shutil.copy2(workbook_path, temp_copy_path)
        wb = load_workbook(temp_copy_path, read_only=True, data_only=True)
    issue_keys: set[str] = set()
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(values_only=True):
                for cell_value in row:
                    if cell_value is None:
                        continue
                    key = normalize_issue_key(str(cell_value))
                    if key:
                        issue_keys.add(key)
    finally:
        wb.close()
        if temp_copy_path and temp_copy_path.exists():
            try:
                temp_copy_path.unlink()
            except OSError:
                pass
    return issue_keys


def _parse_date_value(value) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None

    formats = ("%Y-%m-%d", "%d-%b-%Y", "%d/%m/%Y", "%m/%d/%Y")
    for fmt in formats:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue

    match = _DATE_PATTERN.search(text)
    if not match:
        return None
    day = int(match.group(1))
    month_text = match.group(2).strip().lower()[:3]
    year = int(match.group(3))
    if year < 100:
        year += 2000
    month = _MONTH_MAP.get(month_text)
    if not month:
        return None
    try:
        return date(year, month, day)
    except ValueError:
        return None


def _extract_date_range(value) -> tuple[str, str]:
    if value is None:
        return "", ""
    if isinstance(value, (datetime, date)):
        parsed = _parse_date_value(value)
        iso = parsed.isoformat() if parsed else ""
        return iso, iso

    text = str(value).strip()
    if not text:
        return "", ""

    parsed_dates: list[date] = []
    for match in _DATE_PATTERN.finditer(text):
        parsed = _parse_date_value(match.group(0))
        if parsed:
            parsed_dates.append(parsed)

    if not parsed_dates:
        parsed = _parse_date_value(text)
        if parsed:
            iso = parsed.isoformat()
            return iso, iso
        return "", ""

    parsed_dates.sort()
    return parsed_dates[0].isoformat(), parsed_dates[-1].isoformat()


def load_ipp_planned_dates_by_key(path: Path | None = None) -> dict[str, dict[str, str]]:
    workbook_path = path or get_ipp_meeting_xlsx_path()
    if not workbook_path.exists():
        print(f"Warning: IPP Meeting file not found: {workbook_path}")
        return {}
    temp_copy_path = None
    try:
        wb = load_workbook(workbook_path, read_only=True, data_only=True)
    except Exception:
        temp_dir = Path(tempfile.gettempdir())
        temp_copy_path = temp_dir / f"ipp_meeting_copy_{os.getpid()}.xlsx"
        shutil.copy2(workbook_path, temp_copy_path)
        wb = load_workbook(temp_copy_path, read_only=True, data_only=True)
    result: dict[str, dict[str, str]] = {}
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if not header_row:
                continue
            headers = [str(h).strip() if h is not None else "" for h in header_row]

            jira_idx = None
            epic_link_idx = None
            planned_idx = None
            planned_start_idx = None
            planned_end_idx = None
            for idx, header in enumerate(headers):
                lower = header.lower()
                if jira_idx is None and "jira" in lower:
                    jira_idx = idx
                if epic_link_idx is None and "epic" in lower and "link" in lower:
                    epic_link_idx = idx
                if planned_start_idx is None and "planned" in lower and "start" in lower and "date" in lower:
                    planned_start_idx = idx
                if planned_end_idx is None and "planned" in lower and "end" in lower and "date" in lower:
                    planned_end_idx = idx
                if planned_idx is None and "planned" in lower and "date" in lower:
                    planned_idx = idx
            if planned_idx is None and planned_start_idx is None and planned_end_idx is None:
                continue

            for row in ws.iter_rows(min_row=2, values_only=True):
                key_candidates = []
                if epic_link_idx is not None and epic_link_idx < len(row):
                    key_candidates.append(row[epic_link_idx])
                if jira_idx is not None and jira_idx < len(row):
                    key_candidates.append(row[jira_idx])

                issue_key = ""
                for candidate in key_candidates:
                    issue_key = normalize_issue_key(str(candidate) if candidate is not None else "")
                    if issue_key:
                        break
                if not issue_key:
                    continue

                # Preferred path: separate Planned Start Date / Planned End Date columns.
                if planned_start_idx is not None and planned_end_idx is not None:
                    start_value = row[planned_start_idx] if planned_start_idx < len(row) else None
                    end_value = row[planned_end_idx] if planned_end_idx < len(row) else None
                    start_date = _parse_date_value(start_value)
                    end_date = _parse_date_value(end_value)
                    planned_start = start_date.isoformat() if start_date else ""
                    planned_end = end_date.isoformat() if end_date else ""
                else:
                    planned_value = row[planned_idx] if planned_idx is not None and planned_idx < len(row) else None
                    planned_start, planned_end = _extract_date_range(planned_value)
                if not planned_start and not planned_end:
                    continue

                existing = result.get(issue_key)
                if existing is None:
                    result[issue_key] = {
                        "planned_start": planned_start,
                        "planned_end": planned_end,
                    }
                    continue

                starts = [value for value in [existing.get("planned_start", ""), planned_start] if value]
                ends = [value for value in [existing.get("planned_end", ""), planned_end] if value]
                result[issue_key] = {
                    "planned_start": min(starts) if starts else "",
                    "planned_end": max(ends) if ends else "",
                }
    finally:
        wb.close()
        if temp_copy_path and temp_copy_path.exists():
            try:
                temp_copy_path.unlink()
            except OSError:
                pass
    return result


def load_ipp_actual_and_remarks_by_key(path: Path | None = None) -> dict[str, dict[str, str]]:
    workbook_path = path or get_ipp_meeting_xlsx_path()
    if not workbook_path.exists():
        print(f"Warning: IPP Meeting file not found: {workbook_path}")
        return {}
    temp_copy_path = None
    try:
        wb = load_workbook(workbook_path, read_only=True, data_only=True)
    except Exception:
        temp_dir = Path(tempfile.gettempdir())
        temp_copy_path = temp_dir / f"ipp_meeting_copy_{os.getpid()}.xlsx"
        shutil.copy2(workbook_path, temp_copy_path)
        wb = load_workbook(temp_copy_path, read_only=True, data_only=True)

    result: dict[str, dict[str, str]] = {}
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if not header_row:
                continue
            headers = [str(h).strip() if h is not None else "" for h in header_row]

            jira_idx = None
            epic_link_idx = None
            actual_date_idx = None
            remarks_idx = None
            for idx, header in enumerate(headers):
                lower = header.lower()
                if jira_idx is None and "jira" in lower:
                    jira_idx = idx
                if epic_link_idx is None and "epic" in lower and "link" in lower:
                    epic_link_idx = idx
                if actual_date_idx is None and "actual" in lower and "date" in lower:
                    actual_date_idx = idx
                if remarks_idx is None and "remarks" in lower:
                    remarks_idx = idx
            if actual_date_idx is None and remarks_idx is None:
                continue

            for row in ws.iter_rows(min_row=2, values_only=True):
                key_candidates = []
                if epic_link_idx is not None and epic_link_idx < len(row):
                    key_candidates.append(row[epic_link_idx])
                if jira_idx is not None and jira_idx < len(row):
                    key_candidates.append(row[jira_idx])

                issue_key = ""
                for candidate in key_candidates:
                    issue_key = normalize_issue_key(str(candidate) if candidate is not None else "")
                    if issue_key:
                        break
                if not issue_key:
                    continue

                actual_value = row[actual_date_idx] if actual_date_idx is not None and actual_date_idx < len(row) else None
                parsed_actual = _parse_date_value(actual_value)
                actual_date_iso = parsed_actual.isoformat() if parsed_actual else ""
                remarks_value = row[remarks_idx] if remarks_idx is not None and remarks_idx < len(row) else ""
                remarks_text = str(remarks_value).strip() if remarks_value is not None else ""
                if not actual_date_iso and not remarks_text:
                    continue

                existing = result.get(issue_key)
                if existing is None:
                    result[issue_key] = {
                        "ipp_actual_date": actual_date_iso,
                        "ipp_remarks": remarks_text,
                    }
                    continue

                existing_actual = existing.get("ipp_actual_date", "")
                if actual_date_iso:
                    if not existing_actual or actual_date_iso > existing_actual:
                        result[issue_key] = {
                            "ipp_actual_date": actual_date_iso,
                            "ipp_remarks": remarks_text,
                        }
    finally:
        wb.close()
        if temp_copy_path and temp_copy_path.exists():
            try:
                temp_copy_path.unlink()
            except OSError:
                pass
    return result


def yes_no_dates_altered(
    epic_key: str,
    ipp_planned_dates: dict[str, dict[str, str]],
    jira_epic_dates: dict[str, dict[str, str]],
) -> str:
    normalized = normalize_issue_key(epic_key)
    if not normalized:
        return "No"

    ipp_dates = ipp_planned_dates.get(normalized)
    jira_dates = jira_epic_dates.get(normalized)
    if not ipp_dates or not jira_dates:
        return "No"

    ipp_start = ipp_dates.get("planned_start", "")
    ipp_end = ipp_dates.get("planned_end", "")
    jira_start = jira_dates.get("planned_start", "")
    jira_end = jira_dates.get("planned_end", "")
    return "Yes" if ipp_start != jira_start or ipp_end != jira_end else "No"


def yes_no_ipp_actual_matches_jira_end(
    epic_key: str,
    ipp_actual_by_key: dict[str, dict[str, str]],
    jira_epic_dates: dict[str, dict[str, str]],
) -> str:
    normalized = normalize_issue_key(epic_key)
    if not normalized:
        return "No"

    ipp_data = ipp_actual_by_key.get(normalized)
    jira_dates = jira_epic_dates.get(normalized)
    if not ipp_data or not jira_dates:
        return "No"

    ipp_actual = _parse_date_value(ipp_data.get("ipp_actual_date", ""))
    jira_end = _parse_date_value(jira_dates.get("planned_end", ""))
    if not ipp_actual or not jira_end:
        return "No"
    return "Yes" if ipp_actual.isoformat() == jira_end.isoformat() else "No"


def _iter_issue_search_pages(session, base_url: str, jql: str, fields: list[str]):
    url = f"{base_url}/rest/api/3/search/jql"
    next_page_token = None
    while True:
        payload = {"jql": jql, "maxResults": 100, "fields": fields}
        if next_page_token:
            payload["nextPageToken"] = next_page_token
        response = session.post(url, json=payload)
        response.raise_for_status()
        data = response.json()
        for issue in data.get("issues", []):
            yield issue
        next_page_token = data.get("nextPageToken")
        if not next_page_token:
            break


def _get_project_keys_for_sampling(project_keys: list[str] | None) -> list[str]:
    if project_keys:
        values = [str(key).strip() for key in project_keys if str(key).strip()]
        if values:
            return values
    raw = os.getenv("JIRA_PROJECT_KEYS", "").strip()
    if raw:
        values = [key.strip() for key in raw.split(",") if key.strip()]
        if values:
            return values
    return list(_DEFAULT_PROJECT_KEYS)


def _build_sampling_jql(project_keys: list[str]) -> str:
    keys_str = ", ".join(project_keys)
    return (
        f'project in ({keys_str}) AND issuetype in ("Epic", "Story", "Task", '
        f'"Sub-task", "Subtask", "Bug Task", "Bug Subtask")'
    )


def _field_has_value(value) -> bool:
    if value is None:
        return False
    if isinstance(value, str):
        return bool(value.strip())
    if isinstance(value, list):
        return any(_field_has_value(item) for item in value)
    return True


def resolve_jira_start_date_field_id(
    session,
    base_url: str,
    project_keys: list[str] | None = None,
) -> str:
    url = f"{base_url}/rest/api/3/field"
    response = session.get(url)
    response.raise_for_status()
    all_fields = response.json()

    candidates = []
    for field in all_fields:
        field_id = str(field.get("id") or "").strip()
        if not field_id:
            continue
        name = (field.get("name") or "").strip().lower()
        if "start date" in name:
            candidates.append({"id": field_id, "name": name})

    if not candidates:
        return ""
    if len(candidates) == 1:
        return candidates[0]["id"]

    counts_by_field_id: dict[str, int] = {item["id"]: 0 for item in candidates}
    sampling_keys = _get_project_keys_for_sampling(project_keys)
    sampling_jql = _build_sampling_jql(sampling_keys)
    sampling_fields = ["issuetype"] + [item["id"] for item in candidates]

    try:
        for issue in _iter_issue_search_pages(session, base_url, sampling_jql, sampling_fields):
            issue_fields = issue.get("fields", {}) or {}
            for candidate in candidates:
                field_id = candidate["id"]
                if _field_has_value(issue_fields.get(field_id)):
                    counts_by_field_id[field_id] += 1
    except Exception:
        counts_by_field_id = {}

    if counts_by_field_id:
        max_count = max(counts_by_field_id.values())
        winners = [field_id for field_id, count in counts_by_field_id.items() if count == max_count]
        if len(winners) == 1:
            return winners[0]
        winners_set = set(winners)
    else:
        winners_set = {item["id"] for item in candidates}

    for field in candidates:
        if field["name"] == "start date" and field["id"] in winners_set:
            return field["id"]

    for field in all_fields:
        if str(field.get("id") or "").strip() not in winners_set:
            continue
        name = (field.get("name") or "").strip().lower()
        if "start date" in name:
            return field.get("id", "")
    return ""


def resolve_jira_end_date_field_id(
    session,
    base_url: str,
    project_keys: list[str] | None = None,
) -> str:
    url = f"{base_url}/rest/api/3/field"
    response = session.get(url)
    response.raise_for_status()
    all_fields = response.json()

    candidates = []
    for field in all_fields:
        field_id = str(field.get("id") or "").strip()
        if not field_id:
            continue
        name = (field.get("name") or "").strip().lower()
        if "date" not in name:
            continue
        if "due date" in name or "end date" in name:
            candidates.append({"id": field_id, "name": name})

    if not candidates:
        return ""
    if len(candidates) == 1:
        return candidates[0]["id"]

    counts_by_field_id: dict[str, int] = {item["id"]: 0 for item in candidates}
    sampling_keys = _get_project_keys_for_sampling(project_keys)
    sampling_jql = _build_sampling_jql(sampling_keys)
    sampling_fields = ["issuetype"] + [item["id"] for item in candidates]

    try:
        for issue in _iter_issue_search_pages(session, base_url, sampling_jql, sampling_fields):
            issue_fields = issue.get("fields", {}) or {}
            for candidate in candidates:
                field_id = candidate["id"]
                if _field_has_value(issue_fields.get(field_id)):
                    counts_by_field_id[field_id] += 1
    except Exception:
        counts_by_field_id = {}

    if counts_by_field_id:
        max_count = max(counts_by_field_id.values())
        winners = [field_id for field_id, count in counts_by_field_id.items() if count == max_count]
        if len(winners) == 1:
            return winners[0]
        winners_set = set(winners)
    else:
        winners_set = {item["id"] for item in candidates}

    preferred_exact = [
        "planned due date",
        "planned end date",
        "due date",
        "end date",
    ]
    for name in preferred_exact:
        for field in candidates:
            if field["id"] in winners_set and field["name"] == name:
                return field["id"]

    preferred_contains = [
        "planned due date",
        "planned end date",
        "due date",
        "end date",
    ]
    for text in preferred_contains:
        for field in candidates:
            if field["id"] in winners_set and text in field["name"]:
                return field["id"]

    for field in candidates:
        if field["id"] in winners_set:
            return field["id"]
    return candidates[0]["id"]


def resolve_jira_end_date_field_ids(
    session,
    base_url: str,
    project_keys: list[str] | None = None,
) -> list[str]:
    url = f"{base_url}/rest/api/3/field"
    response = session.get(url)
    response.raise_for_status()
    all_fields = response.json()

    candidates: list[dict[str, str]] = []
    seen_ids: set[str] = set()
    for field in all_fields:
        field_id = str(field.get("id") or "").strip()
        if not field_id or field_id in seen_ids:
            continue
        name = (field.get("name") or "").strip().lower()
        if "date" not in name:
            continue
        if "due date" in name or "end date" in name:
            candidates.append({"id": field_id, "name": name})
            seen_ids.add(field_id)

    primary = resolve_jira_end_date_field_id(session, base_url, project_keys=project_keys)
    ordered_ids: list[str] = []
    if primary:
        ordered_ids.append(primary)

    for candidate in candidates:
        field_id = candidate["id"]
        if field_id in ordered_ids:
            continue
        ordered_ids.append(field_id)

    # Keep system duedate as final fallback if available.
    if "duedate" not in ordered_ids:
        ordered_ids.append("duedate")

    custom_ids = [field_id for field_id in ordered_ids if field_id.startswith("customfield_")]
    system_ids = [field_id for field_id in ordered_ids if not field_id.startswith("customfield_")]
    if primary and primary in custom_ids:
        return [primary] + [field_id for field_id in custom_ids if field_id != primary] + system_ids
    if primary and primary in system_ids:
        return [primary] + custom_ids + [field_id for field_id in system_ids if field_id != primary]
    return custom_ids + system_ids


def _first_non_empty_field_value(issue_fields: dict, field_ids: list[str]) -> str:
    for field_id in field_ids:
        value = issue_fields.get(field_id)
        if value is None:
            continue
        if isinstance(value, str):
            text = value.strip()
            if text:
                return text
            continue
        if isinstance(value, (list, dict)):
            if value:
                return str(value)
            continue
        return str(value)
    return ""


def fetch_jira_issue_planned_dates(
    session,
    base_url: str,
    issue_keys: set[str],
    start_date_field_id: str | None = None,
    end_date_field_id: str | None = None,
    end_date_field_ids: list[str] | None = None,
    project_keys: list[str] | None = None,
) -> dict[str, dict[str, str]]:
    normalized_keys = {normalize_issue_key(key) for key in issue_keys}
    normalized_keys = {key for key in normalized_keys if key}
    if not normalized_keys:
        return {}

    start_field_id = start_date_field_id or resolve_jira_start_date_field_id(
        session, base_url, project_keys=project_keys
    )
    if end_date_field_ids:
        end_field_ids = [field_id for field_id in end_date_field_ids if str(field_id).strip()]
    else:
        resolved_end_field_id = end_date_field_id or resolve_jira_end_date_field_id(
            session, base_url, project_keys=project_keys
        )
        if resolved_end_field_id:
            end_field_ids = [resolved_end_field_id]
        else:
            end_field_ids = resolve_jira_end_date_field_ids(session, base_url, project_keys=project_keys)
    end_field_ids = [field_id for field_id in end_field_ids if field_id]
    if "duedate" not in end_field_ids:
        end_field_ids.append("duedate")
    fields = ["issuetype"]
    for field_id in end_field_ids:
        if field_id not in fields:
            fields.append(field_id)
    if start_field_id and start_field_id not in fields:
        fields.append(start_field_id)

    url = f"{base_url}/rest/api/3/search/jql"
    results: dict[str, dict[str, str]] = {}
    key_list = sorted(normalized_keys)
    for offset in range(0, len(key_list), 500):
        chunk = key_list[offset : offset + 500]
        keys_clause = ", ".join(f'"{key}"' for key in chunk)
        jql = f"key in ({keys_clause})"

        next_page_token = None
        while True:
            payload = {"jql": jql, "maxResults": 100, "fields": fields}
            if next_page_token:
                payload["nextPageToken"] = next_page_token
            response = session.post(url, json=payload)
            response.raise_for_status()
            data = response.json()

            for issue in data.get("issues", []):
                key = normalize_issue_key(issue.get("key", ""))
                issue_fields = issue.get("fields", {}) or {}
                issue_type = ((issue_fields.get("issuetype") or {}).get("name") or "").strip().lower()
                if "epic" not in issue_type:
                    continue
                planned_start = issue_fields.get(start_field_id, "") if start_field_id else ""
                planned_end = _first_non_empty_field_value(issue_fields, end_field_ids)
                results[key] = {
                    "planned_start": str(planned_start or ""),
                    "planned_end": str(planned_end or ""),
                }

            next_page_token = data.get("nextPageToken")
            if not next_page_token:
                break
    return results


def yes_no_in_ipp(issue_key: str, ipp_issue_keys: set[str]) -> str:
    normalized = normalize_issue_key(issue_key)
    return "Yes" if normalized and normalized in ipp_issue_keys else "No"
