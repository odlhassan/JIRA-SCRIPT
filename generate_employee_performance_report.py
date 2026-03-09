from __future__ import annotations

import argparse
import json
import os
import re
import sqlite3
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from generate_assignee_hours_report import _list_capacity_profiles
from manage_fields_registry import load_manage_fields
from report_entity_registry import load_report_entities

DEFAULT_WORKLOG_INPUT_XLSX = "2_jira_subtask_worklogs.xlsx"
DEFAULT_WORK_ITEMS_INPUT_XLSX = "1_jira_work_items_export.xlsx"
DEFAULT_LEAVE_REPORT_INPUT_XLSX = "rlt_leave_report.xlsx"
DEFAULT_HTML_OUTPUT = "employee_performance_report.html"
DEFAULT_CAPACITY_DB = "assignee_hours_capacity.db"
LEAVE_HOURS_PER_DAY = 8.0

DEFAULT_PERFORMANCE_SETTINGS: dict[str, float] = {
    "base_score": 100.0,
    "min_score": 0.0,
    "max_score": 100.0,
    "points_per_bug_hour": 0.5,
    "points_per_bug_late_hour": 1.5,
    "points_per_unplanned_leave_hour": 0.75,
    "points_per_subtask_late_hour": 1.0,
    "points_per_estimate_overrun_hour": 1.25,
    "points_per_missed_due_date": 2.0,
    "overloaded_penalty_enabled": 0.0,
    "planning_realism_enabled": 0.0,
    "overloaded_penalty_threshold_pct": 10.0,
}


def _to_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _to_float(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _resolve_path(value: str, base_dir: Path) -> Path:
    path = Path(value)
    return path if path.is_absolute() else base_dir / path


def _parse_iso_date(value: Any) -> str:
    text = _to_text(value)
    if not text:
        return ""
    if len(text) >= 10:
        try:
            date.fromisoformat(text[:10])
            return text[:10]
        except ValueError:
            pass
    for fmt in ("%Y-%m-%dT%H:%M:%S.%f%z", "%Y-%m-%dT%H:%M:%S%z", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    return ""


def _derive_actual_completion(
    planned_due_date: Any,
    last_logged_date: Any,
    resolved_stable_since_date: Any,
) -> dict[str, str]:
    due_date = _parse_iso_date(planned_due_date)
    last_log = _parse_iso_date(last_logged_date)
    resolved_date = _parse_iso_date(resolved_stable_since_date)
    if last_log and resolved_date:
        if last_log >= resolved_date:
            actual_complete_date = last_log
            actual_complete_source = "last_logged_date"
        else:
            actual_complete_date = resolved_date
            actual_complete_source = "resolved_stable_since_date"
    elif last_log:
        actual_complete_date = last_log
        actual_complete_source = "last_logged_date"
    elif resolved_date:
        actual_complete_date = resolved_date
        actual_complete_source = "resolved_stable_since_date"
    else:
        actual_complete_date = ""
        actual_complete_source = "none"

    if not due_date:
        completion_bucket = "no_due_date"
    elif not actual_complete_date:
        completion_bucket = "not_completed"
    elif actual_complete_date < due_date:
        completion_bucket = "before_due"
    elif actual_complete_date == due_date:
        completion_bucket = "on_due"
    else:
        completion_bucket = "after_due"

    return {
        "planned_due_date": due_date,
        "last_logged_date": last_log,
        "resolved_stable_since_date": resolved_date,
        "actual_complete_date": actual_complete_date,
        "actual_complete_source": actual_complete_source,
        "completion_bucket": completion_bucket,
    }


def _normalize_header_key(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "_", _to_text(value).strip().lower()).strip("_")


def _extract_project_key(issue_id: str) -> str:
    text = _to_text(issue_id).upper()
    if not text or "-" not in text:
        return "UNKNOWN"
    project = text.split("-", 1)[0].strip()
    return project if re.match(r"^[A-Z0-9]+$", project) else "UNKNOWN"


def _is_bug_type(text: str) -> bool:
    low = _to_text(text).lower()
    return "bug" in low and ("subtask" in low or "sub-task" in low or "task" in low)


def _normalize_performance_settings(payload: dict, require_all_fields: bool = True) -> dict[str, float]:
    source = payload or {}
    out: dict[str, float] = {}
    for key, default in DEFAULT_PERFORMANCE_SETTINGS.items():
        if key not in source and require_all_fields:
            raise ValueError(f"Missing field: {key}")
        value = _to_float(source.get(key, default))
        if key.startswith("points_per_") and value < 0:
            raise ValueError(f"{key} must be >= 0")
        if key == "overloaded_penalty_enabled":
            value = 1.0 if value > 0 else 0.0
        if key == "planning_realism_enabled":
            value = 1.0 if value > 0 else 0.0
        if key == "overloaded_penalty_threshold_pct":
            if value < 0 or value > 100:
                raise ValueError("overloaded_penalty_threshold_pct must be between 0 and 100")
        out[key] = round(value, 4)
    if out["min_score"] > out["base_score"] or out["base_score"] > out["max_score"]:
        raise ValueError("Score bounds must satisfy: max_score >= base_score >= min_score")
    return out


def _init_performance_settings_db(db_path: Path) -> None:
    db_path.parent.mkdir(parents=True, exist_ok=True)
    with sqlite3.connect(db_path) as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS performance_point_settings (
                id INTEGER PRIMARY KEY CHECK(id = 1),
                base_score REAL NOT NULL,
                min_score REAL NOT NULL,
                max_score REAL NOT NULL,
                points_per_bug_hour REAL NOT NULL,
                points_per_bug_late_hour REAL NOT NULL,
                points_per_unplanned_leave_hour REAL NOT NULL,
                points_per_subtask_late_hour REAL NOT NULL,
                points_per_estimate_overrun_hour REAL NOT NULL,
                points_per_missed_due_date REAL NOT NULL,
                overloaded_penalty_enabled INTEGER NOT NULL DEFAULT 0,
                planning_realism_enabled INTEGER NOT NULL DEFAULT 0,
                overloaded_penalty_threshold_pct REAL NOT NULL DEFAULT 10.0,
                updated_at TEXT NOT NULL
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS performance_teams (
                team_name TEXT PRIMARY KEY,
                team_leader TEXT NOT NULL DEFAULT '',
                assignees_json TEXT NOT NULL,
                updated_at TEXT NOT NULL
            )
            """
        )
        team_cols = [str(row[1]).lower() for row in conn.execute("PRAGMA table_info(performance_teams)").fetchall()]
        if "team_leader" not in team_cols:
            conn.execute("ALTER TABLE performance_teams ADD COLUMN team_leader TEXT NOT NULL DEFAULT ''")
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS simple_scoring_subtasks (
                issue_key TEXT PRIMARY KEY,
                assignee TEXT NOT NULL,
                original_estimate_hours REAL NOT NULL DEFAULT 0,
                actual_hours_logged REAL NOT NULL DEFAULT 0,
                overrun_hours REAL NOT NULL DEFAULT 0,
                estimate_status TEXT NOT NULL DEFAULT 'unknown',
                due_date TEXT,
                planned_due_date TEXT,
                last_logged_date TEXT,
                effective_completion_date TEXT,
                actual_complete_date TEXT,
                actual_complete_source TEXT,
                due_completion_status TEXT NOT NULL DEFAULT 'unknown',
                is_commitment INTEGER NOT NULL DEFAULT 0,
                status TEXT,
                updated_at TEXT NOT NULL
            )
            """
        )
        simple_scoring_cols = [str(row[1]).lower() for row in conn.execute("PRAGMA table_info(simple_scoring_subtasks)").fetchall()]
        if "planned_due_date" not in simple_scoring_cols:
            conn.execute("ALTER TABLE simple_scoring_subtasks ADD COLUMN planned_due_date TEXT")
        if "last_logged_date" not in simple_scoring_cols:
            conn.execute("ALTER TABLE simple_scoring_subtasks ADD COLUMN last_logged_date TEXT")
        if "actual_complete_date" not in simple_scoring_cols:
            conn.execute("ALTER TABLE simple_scoring_subtasks ADD COLUMN actual_complete_date TEXT")
        if "actual_complete_source" not in simple_scoring_cols:
            conn.execute("ALTER TABLE simple_scoring_subtasks ADD COLUMN actual_complete_source TEXT")
        settings_cols = [str(row[1]).lower() for row in conn.execute("PRAGMA table_info(performance_point_settings)").fetchall()]
        if "points_per_missed_due_date" not in settings_cols:
            conn.execute("ALTER TABLE performance_point_settings ADD COLUMN points_per_missed_due_date REAL NOT NULL DEFAULT 2.0")
        if "overloaded_penalty_enabled" not in settings_cols:
            conn.execute("ALTER TABLE performance_point_settings ADD COLUMN overloaded_penalty_enabled INTEGER NOT NULL DEFAULT 0")
        if "planning_realism_enabled" not in settings_cols:
            conn.execute("ALTER TABLE performance_point_settings ADD COLUMN planning_realism_enabled INTEGER NOT NULL DEFAULT 0")
        if "overloaded_penalty_threshold_pct" not in settings_cols:
            conn.execute("ALTER TABLE performance_point_settings ADD COLUMN overloaded_penalty_threshold_pct REAL NOT NULL DEFAULT 10.0")
        row = conn.execute("SELECT id FROM performance_point_settings WHERE id = 1").fetchone()
        if not row:
            defaults = _normalize_performance_settings(DEFAULT_PERFORMANCE_SETTINGS, require_all_fields=True)
            conn.execute(
                """
                INSERT INTO performance_point_settings (
                    id, base_score, min_score, max_score,
                    points_per_bug_hour, points_per_bug_late_hour, points_per_unplanned_leave_hour,
                    points_per_subtask_late_hour, points_per_estimate_overrun_hour, points_per_missed_due_date,
                    overloaded_penalty_enabled, planning_realism_enabled, overloaded_penalty_threshold_pct, updated_at
                ) VALUES (1, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    defaults["base_score"],
                    defaults["min_score"],
                    defaults["max_score"],
                    defaults["points_per_bug_hour"],
                    defaults["points_per_bug_late_hour"],
                    defaults["points_per_unplanned_leave_hour"],
                    defaults["points_per_subtask_late_hour"],
                    defaults["points_per_estimate_overrun_hour"],
                    defaults["points_per_missed_due_date"],
                    int(defaults["overloaded_penalty_enabled"]),
                    int(defaults["planning_realism_enabled"]),
                    defaults["overloaded_penalty_threshold_pct"],
                    datetime.now(timezone.utc).isoformat(),
                ),
            )
        conn.commit()


def _load_performance_settings(db_path: Path) -> dict[str, float]:
    _init_performance_settings_db(db_path)
    with sqlite3.connect(db_path) as conn:
        row = conn.execute(
            """
            SELECT base_score, min_score, max_score, points_per_bug_hour, points_per_bug_late_hour,
                   points_per_unplanned_leave_hour, points_per_subtask_late_hour, points_per_estimate_overrun_hour, points_per_missed_due_date,
                   overloaded_penalty_enabled, planning_realism_enabled, overloaded_penalty_threshold_pct
            FROM performance_point_settings WHERE id = 1
            """
        ).fetchone()
    if not row:
        return dict(DEFAULT_PERFORMANCE_SETTINGS)
    return _normalize_performance_settings(
        {
            "base_score": row[0],
            "min_score": row[1],
            "max_score": row[2],
            "points_per_bug_hour": row[3],
            "points_per_bug_late_hour": row[4],
            "points_per_unplanned_leave_hour": row[5],
            "points_per_subtask_late_hour": row[6],
            "points_per_estimate_overrun_hour": row[7],
            "points_per_missed_due_date": row[8],
            "overloaded_penalty_enabled": row[9],
            "planning_realism_enabled": row[10],
            "overloaded_penalty_threshold_pct": row[11],
        },
        require_all_fields=True,
    )


def _save_performance_settings(db_path: Path, payload: dict) -> dict[str, float]:
    _init_performance_settings_db(db_path)
    normalized = _normalize_performance_settings(payload, require_all_fields=True)
    with sqlite3.connect(db_path) as conn:
        conn.execute(
            """
            UPDATE performance_point_settings
            SET base_score=?, min_score=?, max_score=?, points_per_bug_hour=?, points_per_bug_late_hour=?,
                points_per_unplanned_leave_hour=?, points_per_subtask_late_hour=?, points_per_estimate_overrun_hour=?, points_per_missed_due_date=?,
                overloaded_penalty_enabled=?, planning_realism_enabled=?, overloaded_penalty_threshold_pct=?, updated_at=?
            WHERE id=1
            """,
            (
                normalized["base_score"],
                normalized["min_score"],
                normalized["max_score"],
                normalized["points_per_bug_hour"],
                normalized["points_per_bug_late_hour"],
                normalized["points_per_unplanned_leave_hour"],
                normalized["points_per_subtask_late_hour"],
                normalized["points_per_estimate_overrun_hour"],
                normalized["points_per_missed_due_date"],
                int(normalized["overloaded_penalty_enabled"]),
                int(normalized["planning_realism_enabled"]),
                normalized["overloaded_penalty_threshold_pct"],
                datetime.now(timezone.utc).isoformat(),
            ),
        )
        conn.commit()
    return normalized


def _normalize_team_name(value: Any) -> str:
    name = _to_text(value)
    if not name:
        raise ValueError("Team name is required.")
    if len(name) > 80:
        raise ValueError("Team name is too long (max 80 chars).")
    return name


def _normalize_assignees(values: list[Any]) -> list[str]:
    out: list[str] = []
    seen: set[str] = set()
    for raw in values or []:
        value = _to_text(raw)
        if not value:
            continue
        key = value.casefold()
        if key in seen:
            continue
        seen.add(key)
        out.append(value)
    if not out:
        raise ValueError("Select at least one assignee.")
    out.sort(key=lambda s: s.casefold())
    return out


def _normalize_team_leader(value: Any, assignees: list[str]) -> str:
    leader = _to_text(value)
    if not leader:
        raise ValueError("Team leader is required.")
    if leader.casefold() not in {a.casefold() for a in assignees}:
        raise ValueError("Team leader must be one of selected assignees.")
    for a in assignees:
        if a.casefold() == leader.casefold():
            return a
    return leader


def _list_performance_teams(db_path: Path) -> list[dict]:
    _init_performance_settings_db(db_path)
    with sqlite3.connect(db_path) as conn:
        rows = conn.execute(
            """
            SELECT team_name, team_leader, assignees_json, updated_at
            FROM performance_teams
            ORDER BY lower(team_name) ASC
            """
        ).fetchall()
    out: list[dict] = []
    for row in rows:
        try:
            assignees = json.loads(_to_text(row[2]))
        except json.JSONDecodeError:
            assignees = []
        out.append(
            {
                "team_name": _to_text(row[0]),
                "team_leader": _to_text(row[1]),
                "assignees": [str(a) for a in (assignees or []) if _to_text(a)],
                "updated_at": _to_text(row[3]),
            }
        )
    return out


def _save_performance_team(db_path: Path, team_name: Any, assignees: list[Any], team_leader: Any) -> dict:
    _init_performance_settings_db(db_path)
    normalized_name = _normalize_team_name(team_name)
    normalized_assignees = _normalize_assignees(assignees)
    normalized_leader = _normalize_team_leader(team_leader, normalized_assignees)
    updated_at = datetime.now(timezone.utc).isoformat()
    with sqlite3.connect(db_path) as conn:
        conn.execute(
            """
            INSERT INTO performance_teams (team_name, team_leader, assignees_json, updated_at)
            VALUES (?, ?, ?, ?)
            ON CONFLICT(team_name) DO UPDATE SET
              team_leader=excluded.team_leader,
              assignees_json=excluded.assignees_json,
              updated_at=excluded.updated_at
            """,
            (normalized_name, normalized_leader, json.dumps(normalized_assignees, ensure_ascii=True), updated_at),
        )
        conn.commit()
    return {
        "team_name": normalized_name,
        "team_leader": normalized_leader,
        "assignees": normalized_assignees,
        "updated_at": updated_at,
    }


def _update_performance_team(
    db_path: Path,
    existing_team_name: Any,
    team_name: Any,
    assignees: list[Any],
    team_leader: Any,
) -> dict:
    _init_performance_settings_db(db_path)
    normalized_existing_name = _normalize_team_name(existing_team_name)
    normalized_name = _normalize_team_name(team_name)
    normalized_assignees = _normalize_assignees(assignees)
    normalized_leader = _normalize_team_leader(team_leader, normalized_assignees)
    updated_at = datetime.now(timezone.utc).isoformat()
    with sqlite3.connect(db_path) as conn:
        current = conn.execute(
            "SELECT team_name FROM performance_teams WHERE team_name = ?",
            (normalized_existing_name,),
        ).fetchone()
        if not current:
            raise ValueError("Team not found.")
        if normalized_existing_name.casefold() != normalized_name.casefold():
            conflict = conn.execute(
                "SELECT team_name FROM performance_teams WHERE team_name = ?",
                (normalized_name,),
            ).fetchone()
            if conflict:
                raise ValueError("A team with this name already exists.")
        conn.execute(
            """
            UPDATE performance_teams
            SET team_name = ?, team_leader = ?, assignees_json = ?, updated_at = ?
            WHERE team_name = ?
            """,
            (
                normalized_name,
                normalized_leader,
                json.dumps(normalized_assignees, ensure_ascii=True),
                updated_at,
                normalized_existing_name,
            ),
        )
        conn.commit()
    return {
        "team_name": normalized_name,
        "team_leader": normalized_leader,
        "assignees": normalized_assignees,
        "updated_at": updated_at,
    }


def _delete_performance_team(db_path: Path, team_name: Any) -> bool:
    _init_performance_settings_db(db_path)
    normalized_name = _normalize_team_name(team_name)
    with sqlite3.connect(db_path) as conn:
        cur = conn.execute("DELETE FROM performance_teams WHERE team_name = ?", (normalized_name,))
        conn.commit()
        return cur.rowcount > 0


def _load_work_items(path: Path) -> dict[str, dict]:
    if not path.exists():
        return {}
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header:
            return {}
        headers = [_to_text(h) for h in header]
        idx = {name: i for i, name in enumerate(headers)}
        if "issue_key" not in idx:
            return {}
        issue_type_col = "jira_issue_type" if "jira_issue_type" in idx else ("work_item_type" if "work_item_type" in idx else "")
        out: dict[str, dict] = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            issue_key = _to_text(row[idx["issue_key"]]).upper()
            if not issue_key:
                continue
            out[issue_key] = {
                "issue_key": issue_key,
                "project_key": _to_text(row[idx.get("project_key", -1)]).upper() if "project_key" in idx else _extract_project_key(issue_key),
                "issue_type": _to_text(row[idx[issue_type_col]]) if issue_type_col else "",
                "fix_type": _to_text(row[idx["fix_type"]]).lower() if "fix_type" in idx else "",
                "summary": _to_text(row[idx["summary"]]) if "summary" in idx else "",
                "status": _to_text(row[idx["status"]]) if "status" in idx else "",
                "assignee": _to_text(row[idx["assignee"]]) if "assignee" in idx else "",
                "start_date": _parse_iso_date(row[idx["start_date"]]) if "start_date" in idx else "",
                "due_date": _parse_iso_date(row[idx["end_date"]]) if "end_date" in idx else "",
                "resolved_stable_since_date": _parse_iso_date(row[idx["resolved_stable_since_date"]]) if "resolved_stable_since_date" in idx else "",
                "original_estimate_hours": round(_to_float(row[idx["original_estimate_hours"]]), 2) if "original_estimate_hours" in idx else 0.0,
                "parent_issue_key": _to_text(row[idx["parent_issue_key"]]).upper() if "parent_issue_key" in idx else "",
            }
        return out
    finally:
        wb.close()


def _load_worklogs(path: Path, work_items: dict[str, dict]) -> list[dict]:
    if not path.exists():
        return []
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header:
            return []
        headers = [_to_text(h) for h in header]
        idx = {name: i for i, name in enumerate(headers)}
        required = {"issue_id", "issue_assignee", "worklog_started", "hours_logged"}
        if not required.issubset(set(idx)):
            return []
        out: list[dict] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            issue_id = _to_text(row[idx["issue_id"]]).upper()
            if not issue_id:
                continue
            worklog_date = _parse_iso_date(row[idx["worklog_started"]])
            hours = round(_to_float(row[idx["hours_logged"]]), 2)
            if not worklog_date or hours <= 0:
                continue
            assignee = _to_text(row[idx["issue_assignee"]]) or "Unassigned"
            item = work_items.get(issue_id, {})
            parent_story_id = _to_text(row[idx["parent_story_id"]]).upper() if "parent_story_id" in idx else ""
            parent_story_id = parent_story_id or _to_text(item.get("parent_issue_key")).upper()
            story_item = work_items.get(parent_story_id, {}) if parent_story_id else {}
            raw_type = _to_text(row[idx["issue_type"]]) if "issue_type" in idx else ""
            type_hint = raw_type or _to_text(item.get("issue_type"))
            out.append(
                {
                    "issue_id": issue_id,
                    "issue_assignee": assignee,
                    "worklog_date": worklog_date,
                    "hours_logged": hours,
                    "project_key": _to_text(item.get("project_key")) or _extract_project_key(issue_id),
                    "is_bug": _is_bug_type(type_hint),
                    "fix_type": _to_text(item.get("fix_type")).lower(),
                    "item_summary": _to_text(item.get("summary")),
                    "item_status": _to_text(item.get("status")),
                    "item_issue_type": _to_text(item.get("issue_type")),
                    "item_assignee": _to_text(item.get("assignee")),
                    "item_parent_issue_key": _to_text(item.get("parent_issue_key")).upper(),
                    "item_start_date": _to_text(item.get("start_date")),
                    "item_due_date": _to_text(item.get("due_date")),
                    "item_resolved_stable_since_date": _to_text(item.get("resolved_stable_since_date")),
                    "story_due_date": _to_text(story_item.get("due_date")),
                    "original_estimate_hours": round(_to_float(item.get("original_estimate_hours")), 2),
                }
            )
        return out
    finally:
        wb.close()


def _load_unplanned_leave_rows(path: Path) -> list[dict]:
    if not path.exists():
        return []
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        preferred_sheets = ["Daily_Assignee", "Leaves", "Leave_Daily", "Daily_Leaves"]
        sheet_names = [name for name in preferred_sheets if name in wb.sheetnames]
        sheet_names.extend([name for name in wb.sheetnames if name not in sheet_names])
        for sheet_name in sheet_names:
            ws = wb[sheet_name]
            header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if not header:
                continue
            header_keys = [_normalize_header_key(h) for h in header]
            index_by_header = {key: i for i, key in enumerate(header_keys) if key}

            def idx(*aliases: str) -> int | None:
                for alias in aliases:
                    key = _normalize_header_key(alias)
                    if key in index_by_header:
                        return index_by_header[key]
                return None

            assignee_idx = idx("assignee", "employee", "resource_name")
            day_idx = idx("period_day", "day", "date", "leave_date")
            planned_idx = idx("planned_taken_hours", "planned_leave_hours", "planned_hours")
            unplanned_idx = idx("unplanned_taken_hours", "unplanned_leave_hours", "unplanned_hours")
            leave_type_idx = idx("leave_type", "type")
            leave_hours_idx = idx("leave_hours", "hours", "taken_hours")
            leave_days_idx = idx("leave_days", "days")

            has_wide_schema = planned_idx is not None and unplanned_idx is not None
            has_long_schema = leave_type_idx is not None and (leave_hours_idx is not None or leave_days_idx is not None)
            if assignee_idx is None or day_idx is None or (not has_wide_schema and not has_long_schema):
                continue

            out: list[dict] = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                iso_day = _parse_iso_date(row[day_idx])
                if not iso_day:
                    continue
                assignee = _to_text(row[assignee_idx]) or "Unassigned"
                planned_hours = 0.0
                unplanned_hours = 0.0
                if has_wide_schema:
                    planned_hours = round(_to_float(row[planned_idx]), 2)
                    unplanned_hours = round(_to_float(row[unplanned_idx]), 2)
                else:
                    leave_type = _to_text(row[leave_type_idx]).lower()
                    hours = _to_float(row[leave_hours_idx]) if leave_hours_idx is not None else 0.0
                    if leave_days_idx is not None:
                        hours = hours or (_to_float(row[leave_days_idx]) * LEAVE_HOURS_PER_DAY)
                    hours = round(hours, 2)
                    if "planned" in leave_type and "unplanned" not in leave_type:
                        planned_hours = hours
                    elif "unplanned" in leave_type:
                        unplanned_hours = hours
                    else:
                        continue
                out.append(
                    {
                        "assignee": assignee,
                        "period_day": iso_day,
                        "unplanned_taken_hours": unplanned_hours,
                        "planned_taken_hours": planned_hours,
                    }
                )
            return out
        return []
    finally:
        wb.close()


def _load_leave_issue_keys(path: Path) -> list[str]:
    if not path.exists():
        return []
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        preferred_sheets = ["Raw_Subtasks", "Subtasks_Distributed"]
        for sheet_name in preferred_sheets:
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if not header:
                continue
            header_keys = [_normalize_header_key(h) for h in header]
            index_by_header = {key: i for i, key in enumerate(header_keys) if key}
            key_idx = index_by_header.get("issue_key")
            if key_idx is None:
                continue
            keys: set[str] = set()
            for row in ws.iter_rows(min_row=2, values_only=True):
                issue_key = _to_text(row[key_idx]).upper()
                if issue_key:
                    keys.add(issue_key)
            return sorted(keys)
        return []
    finally:
        wb.close()


def _resolve_epf_run_id(db_path: Path, requested_run_id: str) -> str:
    run_id = _to_text(requested_run_id)
    if run_id:
        return run_id
    if not db_path.exists():
        return ""
    with sqlite3.connect(db_path) as conn:
        row = conn.execute(
            "SELECT active_run_id FROM epf_refresh_state WHERE id = 1"
        ).fetchone()
    return _to_text(row[0] if row else "")


def _load_work_items_from_epf_db(db_path: Path, run_id: str) -> dict[str, dict]:
    if not db_path.exists() or not run_id:
        return {}
    with sqlite3.connect(db_path) as conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute(
            """
            SELECT issue_key, project_key, issue_type, fix_type, summary, status, assignee,
                   start_date, due_date, resolved_stable_since_date, original_estimate_hours, parent_issue_key
            FROM epf_work_items
            WHERE run_id = ?
            """,
            (run_id,),
        ).fetchall()
    out: dict[str, dict] = {}
    for row in rows:
        issue_key = _to_text(row["issue_key"]).upper()
        if not issue_key:
            continue
        out[issue_key] = {
            "issue_key": issue_key,
            "project_key": _to_text(row["project_key"]).upper() or _extract_project_key(issue_key),
            "issue_type": _to_text(row["issue_type"]),
            "fix_type": _to_text(row["fix_type"]).lower(),
            "summary": _to_text(row["summary"]),
            "status": _to_text(row["status"]),
            "assignee": _to_text(row["assignee"]),
            "start_date": _parse_iso_date(row["start_date"]),
            "due_date": _parse_iso_date(row["due_date"]),
            "resolved_stable_since_date": _parse_iso_date(row["resolved_stable_since_date"]),
            "original_estimate_hours": round(_to_float(row["original_estimate_hours"]), 2),
            "parent_issue_key": _to_text(row["parent_issue_key"]).upper(),
        }
    return out


def _load_worklogs_from_epf_db(db_path: Path, run_id: str, work_items: dict[str, dict]) -> list[dict]:
    if not db_path.exists() or not run_id:
        return []
    with sqlite3.connect(db_path) as conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute(
            """
            SELECT issue_id, issue_assignee, worklog_date, hours_logged, project_key, is_bug, fix_type,
                   item_summary, item_status, item_issue_type, item_assignee, item_parent_issue_key,
                   item_start_date, item_due_date, item_resolved_stable_since_date, story_due_date,
                   original_estimate_hours
            FROM epf_worklogs
            WHERE run_id = ?
            """,
            (run_id,),
        ).fetchall()
    out: list[dict] = []
    for row in rows:
        issue_id = _to_text(row["issue_id"]).upper()
        if not issue_id:
            continue
        issue_assignee = _to_text(row["issue_assignee"]) or "Unassigned"
        worklog_date = _parse_iso_date(row["worklog_date"])
        hours_logged = round(_to_float(row["hours_logged"]), 2)
        if not worklog_date or hours_logged <= 0:
            continue
        item = work_items.get(issue_id, {})
        issue_type = _to_text(row["item_issue_type"]) or _to_text(item.get("issue_type"))
        out.append(
            {
                "issue_id": issue_id,
                "issue_assignee": issue_assignee,
                "worklog_date": worklog_date,
                "hours_logged": hours_logged,
                "project_key": _to_text(row["project_key"]).upper()
                or _to_text(item.get("project_key"))
                or _extract_project_key(issue_id),
                "is_bug": bool(int(_to_float(row["is_bug"]))) if _to_text(row["is_bug"]) else _is_bug_type(issue_type),
                "fix_type": _to_text(row["fix_type"]).lower(),
                "item_summary": _to_text(row["item_summary"]) or _to_text(item.get("summary")),
                "item_status": _to_text(row["item_status"]) or _to_text(item.get("status")),
                "item_issue_type": issue_type,
                "item_assignee": _to_text(row["item_assignee"]) or _to_text(item.get("assignee")),
                "item_parent_issue_key": _to_text(row["item_parent_issue_key"]).upper() or _to_text(item.get("parent_issue_key")).upper(),
                "item_start_date": _parse_iso_date(row["item_start_date"]) or _to_text(item.get("start_date")),
                "item_due_date": _parse_iso_date(row["item_due_date"]) or _to_text(item.get("due_date")),
                "item_resolved_stable_since_date": _parse_iso_date(row["item_resolved_stable_since_date"]) or _to_text(item.get("resolved_stable_since_date")),
                "story_due_date": _parse_iso_date(row["story_due_date"]),
                "original_estimate_hours": round(_to_float(row["original_estimate_hours"]), 2),
            }
        )
    return out


def _load_unplanned_leave_rows_from_epf_db(db_path: Path, run_id: str) -> list[dict]:
    if not db_path.exists() or not run_id:
        return []
    with sqlite3.connect(db_path) as conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute(
            """
            SELECT assignee, period_day, unplanned_taken_hours, planned_taken_hours
            FROM epf_leave_rows
            WHERE run_id = ?
            """,
            (run_id,),
        ).fetchall()
    out: list[dict] = []
    for row in rows:
        period_day = _parse_iso_date(row["period_day"])
        if not period_day:
            continue
        out.append(
            {
                "assignee": _to_text(row["assignee"]) or "Unassigned",
                "period_day": period_day,
                "unplanned_taken_hours": round(_to_float(row["unplanned_taken_hours"]), 2),
                "planned_taken_hours": round(_to_float(row["planned_taken_hours"]), 2),
            }
        )
    return out


def _load_leave_issue_keys_from_epf_db(db_path: Path, run_id: str) -> list[str]:
    if not db_path.exists() or not run_id:
        return []
    with sqlite3.connect(db_path) as conn:
        rows = conn.execute(
            "SELECT issue_key FROM epf_leave_issue_keys WHERE run_id = ? ORDER BY issue_key",
            (run_id,),
        ).fetchall()
    return sorted({_to_text(r[0]).upper() for r in rows if _to_text(r[0])})


def _default_range(rows: list[dict]) -> tuple[str, str]:
    today = datetime.now(timezone.utc).date()
    month_start = date(today.year, today.month, 1)
    next_month_start = date(today.year + (1 if today.month == 12 else 0), 1 if today.month == 12 else today.month + 1, 1)
    month_end = next_month_start - timedelta(days=1)
    return month_start.isoformat(), month_end.isoformat()


def _precompute_simple_scoring(
    db_path: Path,
    work_items: dict[str, dict],
    worklogs: list[dict],
) -> list[dict]:
    """Precompute per-subtask simple scoring metrics and persist to SQLite."""
    hours_by_issue: dict[str, float] = {}
    last_log_by_issue: dict[str, str] = {}
    for wl in worklogs:
        key = _to_text(wl.get("issue_id")).upper()
        if not key:
            continue
        hrs = _to_float(wl.get("hours_logged"))
        hours_by_issue[key] = hours_by_issue.get(key, 0.0) + hrs
        log_date = _to_text(wl.get("worklog_date"))
        if log_date and (not last_log_by_issue.get(key) or log_date > last_log_by_issue[key]):
            last_log_by_issue[key] = log_date

    rows: list[dict] = []
    now_iso = datetime.now(timezone.utc).isoformat()
    for key, wi in work_items.items():
        issue_type = _to_text(wi.get("issue_type")).lower()
        if "subtask" not in issue_type and "sub-task" not in issue_type and "bug" not in issue_type:
            continue
        assignee = _to_text(wi.get("assignee"))
        if not assignee:
            continue
        estimate = round(_to_float(wi.get("original_estimate_hours")), 2)
        actual = round(hours_by_issue.get(key, 0.0), 2)
        overrun = round(max(0.0, actual - estimate), 2) if estimate > 0 else 0.0
        if estimate <= 0:
            est_status = "no_estimate"
        elif actual <= estimate:
            est_status = "within_estimate"
        else:
            est_status = "over_estimate"

        completion_info = _derive_actual_completion(
            wi.get("due_date"),
            last_log_by_issue.get(key, ""),
            wi.get("resolved_stable_since_date"),
        )
        due_date = completion_info["planned_due_date"]
        actual_complete_date = completion_info["actual_complete_date"]
        actual_complete_source = completion_info["actual_complete_source"]
        last_log = completion_info["last_logged_date"]

        if not due_date:
            due_status = "no_due_date"
        elif not actual_complete_date:
            due_status = "not_completed"
        elif actual_complete_date <= due_date:
            due_status = "on_time"
        else:
            due_status = "late"

        is_commitment = 1 if (est_status == "over_estimate" and due_status == "on_time") else 0

        rows.append({
            "issue_key": key,
            "assignee": assignee,
            "original_estimate_hours": estimate,
            "actual_hours_logged": actual,
            "overrun_hours": overrun,
            "estimate_status": est_status,
            "due_date": due_date,
            "planned_due_date": due_date,
            "last_logged_date": last_log,
            "effective_completion_date": actual_complete_date,
            "actual_complete_date": actual_complete_date,
            "actual_complete_source": actual_complete_source,
            "due_completion_status": due_status,
            "is_commitment": is_commitment,
            "status": _to_text(wi.get("status")),
            "updated_at": now_iso,
        })

    _init_performance_settings_db(db_path)
    with sqlite3.connect(db_path) as conn:
        conn.execute("DELETE FROM simple_scoring_subtasks")
        for r in rows:
            conn.execute(
                """INSERT OR REPLACE INTO simple_scoring_subtasks
                   (issue_key, assignee, original_estimate_hours, actual_hours_logged,
                    overrun_hours, estimate_status, due_date, planned_due_date, last_logged_date,
                    effective_completion_date, actual_complete_date, actual_complete_source,
                    due_completion_status, is_commitment, status, updated_at)
                   VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (
                    r["issue_key"], r["assignee"], r["original_estimate_hours"],
                    r["actual_hours_logged"], r["overrun_hours"], r["estimate_status"],
                    r["due_date"], r["planned_due_date"], r["last_logged_date"],
                    r["effective_completion_date"], r["actual_complete_date"], r["actual_complete_source"],
                    r["due_completion_status"], r["is_commitment"], r["status"], r["updated_at"],
                ),
            )
        conn.commit()
    return rows


def _load_simple_scoring(db_path: Path) -> list[dict]:
    _init_performance_settings_db(db_path)
    with sqlite3.connect(db_path) as conn:
        conn.row_factory = sqlite3.Row
        return [dict(r) for r in conn.execute("SELECT * FROM simple_scoring_subtasks").fetchall()]


def _build_payload(
    worklogs: list[dict],
    work_items: list[dict],
    leave_rows: list[dict],
    settings: dict[str, float],
    teams: list[dict],
    entities_catalog: list[dict],
    managed_fields: list[dict],
    capacity_profiles: list[dict],
    leave_issue_keys: list[str] | None = None,
    simple_scoring: list[dict] | None = None,
) -> dict:
    jira_browse_base = _to_text(os.getenv("JIRA_BROWSE_BASE"))
    if not jira_browse_base:
        jira_site = _to_text(os.getenv("JIRA_SITE")) or "octopusdtlsupport"
        if re.match(r"^https?://", jira_site, re.IGNORECASE):
            jira_browse_base = f"{jira_site.rstrip('/')}/browse"
        else:
            jira_browse_base = f"https://{jira_site}.atlassian.net/browse"
    projects = sorted(
        {
            *({_to_text(r.get("project_key")) or "UNKNOWN" for r in worklogs}),
            *({_to_text(r.get("project_key")) or "UNKNOWN" for r in work_items}),
        }
    )
    default_from, default_to = _default_range(worklogs)
    return {
        "worklogs": worklogs,
        "work_items": work_items,
        "leave_rows": leave_rows,
        "leave_issue_keys": leave_issue_keys or [],
        "teams": teams or [],
        "projects": projects,
        "default_from": default_from,
        "default_to": default_to,
        "leave_hours_per_day": LEAVE_HOURS_PER_DAY,
        "settings": settings,
        "entities_catalog": entities_catalog or [],
        "managed_fields": managed_fields or [],
        "capacity_profiles": capacity_profiles or [],
        "simple_scoring": simple_scoring or [],
        "jira_browse_base": jira_browse_base.rstrip("/"),
        "generated_at": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC"),
    }


def _build_html(payload: dict) -> str:
    data = json.dumps(payload, ensure_ascii=True)
    return f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Employee Performance Dashboard</title>
  <style>
    :root {{ --bg:#0a1325; --panel:#121f39; --line:#2a3f65; --ink:#dce8ff; --muted:#9db1d8; --bad:#fb7185; --good:#34d399; }}
    * {{ box-sizing:border-box; }} body {{ margin:0; font-family:"Trebuchet MS","Segoe UI",sans-serif; color:var(--ink); background:radial-gradient(circle at 5% 0%,#1a3356 0%,#0a1325 60%); }}
    .top-date-range-wrap {{ position:sticky; top:0; z-index:25; display:flex; justify-content:center; padding:10px 12px 0; }}
    .top-date-range-chip {{ display:inline-flex; align-items:center; gap:8px; flex-wrap:wrap; padding:7px 12px; border:1px solid #3f5f93; border-radius:999px; background:#0f2342; box-shadow:0 8px 18px rgba(2, 8, 23, .35); }}
    .date-chip-segment {{ color:#cfe0ff; font-size:.72rem; font-weight:800; text-transform:uppercase; letter-spacing:.03em; }}
    .date-chip-input {{ border:1px solid #3a5c91; border-radius:999px; background:#0d1830; color:var(--ink); padding:6px 10px; min-height:32px; font-weight:700; }}
    .date-chip-input:focus {{ outline:none; border-color:#7cb2ff; box-shadow:0 0 0 2px rgba(124,178,255,.25); }}
    .date-chip-select {{ border:1px solid #3a5c91; border-radius:999px; background:#0d1830; color:var(--ink); padding:6px 10px; min-height:32px; font-weight:700; min-width:220px; }}
    .date-chip-select:focus {{ outline:none; border-color:#7cb2ff; box-shadow:0 0 0 2px rgba(124,178,255,.25); }}
    .date-chip-control {{ display:inline-flex; align-items:center; gap:6px; }}
    .date-chip-label {{ color:#cfe0ff; font-size:.72rem; font-weight:800; text-transform:uppercase; letter-spacing:.03em; white-space:nowrap; }}
    .date-chip-control .date-chip-select {{ min-width:200px; }}
    .adv-filter-wrap {{ position:relative; display:inline-flex; }}
    .adv-filter-btn {{ border:1px solid #4f46e5; border-radius:999px; background:#4338ca; color:#eef2ff; font-size:.74rem; font-weight:700; padding:6px 12px; cursor:pointer; }}
    .adv-filter-btn:hover {{ background:#3730a3; }}
    .adv-filter-btn:focus {{ outline:none; box-shadow:0 0 0 2px rgba(99,102,241,.35); }}
    .adv-filter-menu {{ position:absolute; top:calc(100% + 6px); right:0; min-width:200px; padding:6px; border:1px solid #314d7a; border-radius:10px; background:#0f1b32; box-shadow:0 10px 20px rgba(2,8,23,.4); z-index:45; }}
    .adv-filter-group-label {{ padding:4px 8px; color:#9db1d8; font-size:.66rem; text-transform:uppercase; font-weight:800; letter-spacing:.04em; }}
    .adv-filter-item {{ width:100%; border:0; background:transparent; color:#dce8ff; text-align:left; padding:7px 8px; border-radius:8px; font-size:.76rem; cursor:pointer; }}
    .adv-filter-item:hover {{ background:#17325a; }}
    .date-chip-status {{ color:#c7d7f3; font-size:.72rem; font-weight:700; }}
    .header-expand-fab {{ position:fixed; top:12px; right:12px; z-index:60; display:none; width:38px; height:38px; border:1px solid #4a6ea9; border-radius:999px; background:#1b325a; color:#eef4ff; font-size:1.05rem; font-weight:800; cursor:pointer; line-height:1; }}
    .header-expand-fab:hover {{ background:#244172; }}
    body.header-collapsed .header-expand-fab {{ display:inline-flex; align-items:center; justify-content:center; }}
    .wrap {{ max-width:1800px; margin:0 auto; padding:14px; }}
    .hero {{ background:#101e37; border:1px solid #34507e; border-radius:12px; padding:12px; }}
    .hero.is-collapsed > :not(.hero-top) {{ display:none; }}
    .hero-top {{ display:flex; align-items:flex-start; justify-content:space-between; gap:8px; flex-wrap:wrap; }}
    .hero-actions {{ display:flex; align-items:center; gap:8px; flex-wrap:wrap; }}
    .meta {{ color:#c4d4ef; font-size:.8rem; margin-top:4px; }}
    .toolbar {{ display:grid; gap:8px; grid-template-columns:minmax(180px,1fr) minmax(210px,1fr) minmax(210px,1fr) auto; margin-top:8px; }}
    .shortcut-bar {{ display:flex; gap:8px; flex-wrap:wrap; margin-top:8px; }}
    .shortcut-btn {{ border:1px solid #3f5f93; background:#0f2342; color:#dce8ff; border-radius:999px; font-size:.74rem; padding:5px 10px; cursor:pointer; }}
    .shortcut-btn:hover {{ background:#17325a; }}
    .f label {{ display:block; font-size:.7rem; color:var(--muted); margin-bottom:3px; text-transform:uppercase; font-weight:700; }} .f input,.f select {{ width:100%; border:1px solid #3a5c91; border-radius:8px; background:#0d1830; color:var(--ink); padding:7px; }}
    #projects option {{ color:#0f172a; background:#ffffff; }}
    .btn {{ border:1px solid #4a6ea9; background:#1b325a; color:#eef4ff; border-radius:8px; font-weight:700; padding:7px 10px; cursor:pointer; text-decoration:none; display:inline-flex; align-items:center; }}
    .report-refresh-wrap {{ display:flex; align-items:center; gap:8px; flex-wrap:wrap; }}
    .report-refresh-status {{ font-size:.72rem; color:#9db1d8; min-height:18px; }}
    .report-refresh-details {{ width:100%; font-size:.68rem; color:#c7d7f3; min-height:16px; }}
    .guide {{ margin-top:8px; border:1px solid #2f517e; border-radius:10px; background:#0f2140; padding:8px; }}
    .guide h2 {{ margin:0; font-size:.86rem; }}
    .guide p {{ margin:5px 0 0; font-size:.77rem; color:#c7d7f3; line-height:1.35; }}
    .hero-scorecard {{ margin-top:10px; border:1px solid #5b8bd1; border-radius:16px; background:linear-gradient(135deg,#15345f 0%, #102546 52%, #0d1b32 100%); padding:14px 16px; display:flex; justify-content:space-between; align-items:flex-end; gap:14px; box-shadow:0 14px 30px rgba(2,8,23,.28); }}
    .hero-scorecard-copy {{ min-width:0; }}
    .hero-scorecard-kicker {{ font-size:.7rem; font-weight:900; text-transform:uppercase; letter-spacing:.08em; color:#93c5fd; }}
    .hero-scorecard-title {{ margin-top:4px; font-size:1rem; font-weight:900; color:#eff6ff; }}
    .hero-scorecard-sub {{ margin-top:4px; font-size:.76rem; color:#c7d7f3; line-height:1.35; }}
    .hero-scorecard-metric {{ text-align:right; flex:0 0 auto; }}
    .hero-scorecard-value {{ font-size:clamp(2rem,3.6vw,3rem); font-weight:900; line-height:.95; color:#f8fbff; }}
    .hero-scorecard-mode {{ margin-top:6px; display:inline-flex; align-items:center; gap:6px; border:1px solid #4f77b3; border-radius:999px; padding:4px 10px; background:rgba(15,35,66,.55); color:#dbeafe; font-size:.72rem; font-weight:800; }}
    .hero-scorecard-mode .material-symbols-outlined {{ font-size:15px; }}
    .hero-scorecard-meta {{ margin-top:6px; font-size:.72rem; color:#9db1d8; }}
    .discover {{ display:flex; gap:8px; flex-wrap:wrap; margin-top:8px; }}
    .discover .pill {{ border:1px solid #365c8d; border-radius:999px; padding:4px 10px; font-size:.74rem; color:#dce8ff; background:#132949; }}
    .leader-controls {{ display:flex; gap:8px; flex-wrap:wrap; align-items:end; padding:8px 10px 0; }}
    .leader-controls .f {{ min-width:160px; }}
    .leader-actions-wrap {{ margin-left:auto; display:flex; align-items:center; gap:8px; }}
    .leader-actions-menu-wrap {{ position:relative; }}
    .leader-icon-btn {{ width:34px; height:34px; border:1px solid #3f5f93; border-radius:8px; background:#102949; color:#dce8ff; display:inline-flex; align-items:center; justify-content:center; cursor:pointer; }}
    .leader-icon-btn:hover {{ background:#17325a; border-color:#7cb2ff; }}
    .leader-icon-btn .material-symbols-outlined {{ font-size:18px; }}
    .leader-actions-menu {{ position:absolute; top:calc(100% + 6px); right:0; min-width:210px; padding:6px; border:1px solid #314d7a; border-radius:10px; background:#0f1b32; box-shadow:0 10px 20px rgba(2,8,23,.4); z-index:45; }}
    .leader-actions-item {{ width:100%; border:0; background:transparent; color:#dce8ff; text-align:left; padding:7px 8px; border-radius:8px; font-size:.76rem; cursor:pointer; display:flex; align-items:center; gap:6px; }}
    .leader-actions-item:hover {{ background:#17325a; }}
    .leader-actions-item .material-symbols-outlined {{ font-size:16px; }}
    .leader-action-status {{ padding:0 10px 8px; min-height:16px; }}
    .section-head {{ margin:10px 0 4px; font-size:.8rem; color:var(--muted); text-transform:uppercase; letter-spacing:.04em; font-weight:800; }}
    .section-head.collapse-toggle {{ cursor:pointer; user-select:none; }}
    .section-head.collapse-toggle .hint {{ font-size:.68rem; color:#7fa3d6; margin-left:8px; text-transform:none; letter-spacing:0; }}
    .is-collapsed {{ display:none; }}
    .kpis {{ display:grid; gap:8px; grid-template-columns:repeat(4,minmax(0,1fr)); margin-top:8px; }} .kpi {{ border:1px solid var(--line); border-radius:10px; background:var(--panel); padding:8px; }} .kpi .k {{ font-size:.72rem; color:var(--muted); text-transform:uppercase; }} .kpi .v {{ margin-top:4px; font-size:1.1rem; font-weight:800; }}
    .kpi.actionable {{ cursor:pointer; transition:background .15s ease, border-color .15s ease; }}
    .kpi.actionable:hover {{ background:#173158; border-color:#5f88c0; }}
    .kpi .kpi-note {{ margin-top:3px; font-size:.66rem; color:#9db1d8; }}
    .rmi-list-wrap {{ margin-top:8px; }}
    .planning-kpis .kpi {{ border-color:#315a86; background:#10223f; }}
    .top3-wrap {{ display:grid; gap:8px; grid-template-columns:repeat(2,minmax(0,1fr)); margin-top:8px; }}
    .top3-card {{ border:1px solid var(--line); border-radius:10px; background:var(--panel); padding:8px; }}
    .top3-title {{ font-size:.74rem; color:var(--muted); text-transform:uppercase; font-weight:700; margin-bottom:5px; }}
    .top3-item {{ display:flex; justify-content:space-between; gap:8px; padding:3px 0; font-size:.82rem; }}
    .top3-item .nm {{ color:var(--ink); font-weight:700; white-space:nowrap; overflow:hidden; text-overflow:ellipsis; }}
    .top3-item .sc {{ font-weight:800; }}
    .top3-item.high .sc {{ color:var(--good); }}
    .top3-item.low .sc {{ color:var(--bad); }}
    .teams-wrap {{ margin-top:8px; border:1px solid var(--line); border-radius:10px; background:var(--panel); padding:8px; }}
    .teams-wrap.is-collapsed .team-score-layout {{ display:none; }}
    .teams-title {{ font-size:.74rem; color:var(--muted); text-transform:uppercase; font-weight:700; margin-bottom:6px; }}
    .team-score-layout {{ display:grid; gap:8px; grid-template-columns:7fr 3fr; align-items:stretch; }}
    .team-card {{ border:1px solid #314d7a; border-radius:8px; background:#111e37; padding:8px; cursor:pointer; transition:box-shadow .12s ease, border-color .12s ease, transform .12s ease; }}
    .team-card:hover {{ border-color:#5f86bf; transform:translateY(-1px); }}
    .team-card.sel {{ border-color:#6ea8ff; box-shadow:inset 0 0 0 1px #6ea8ff; }}
    .team-head {{ display:flex; justify-content:space-between; gap:8px; align-items:flex-start; }}
    .team-name {{ font-size:.86rem; font-weight:800; }}
    .team-sub {{ font-size:.74rem; color:var(--muted); margin-top:2px; }}
    .team-score {{ font-size:1.1rem; font-weight:900; }}
    .team-metrics {{ margin-top:6px; display:grid; gap:3px; font-size:.76rem; }}
    .team-chart-shell {{ border:1px solid #2b446e; border-radius:10px; background:#0f1b32; padding:8px; }}
    .team-chart-svg {{ width:100%; height:220px; display:block; border:1px solid #223a61; border-radius:8px; background:#0d172b; }}
    .team-detail-shell {{ border:1px solid #2b446e; border-radius:10px; background:#0f1b32; padding:8px; }}
    .team-detail-body {{ margin-top:6px; }}
    .arena {{ display:grid; gap:10px; grid-template-columns:minmax(320px,38%) minmax(0,62%); margin-top:10px; align-items:stretch; }} .panel {{ border:1px solid var(--line); border-radius:12px; background:var(--panel); overflow:hidden; }} .panel h2 {{ margin:0; padding:9px 10px; font-size:.9rem; border-bottom:1px solid var(--line); }}
    .panel-head {{ display:flex; justify-content:space-between; align-items:center; gap:10px; padding:9px 10px; border-bottom:1px solid var(--line); }}
    .panel-head-title {{ margin:0; padding:0; border-bottom:0; font-size:.9rem; }}
    .panel-inline-toggle {{ display:inline-flex; align-items:center; gap:6px; font-size:.72rem; color:#c7d7f3; font-weight:700; user-select:none; }}
    .panel-inline-toggle input {{ margin:0; accent-color:#60a5fa; }}
    .arena > .panel {{ display:flex; flex-direction:column; min-height:72vh; max-height:72vh; }}
    .leaderboard {{ overflow:auto; flex:1; min-height:0; max-height:none; }} .row {{ display:grid; gap:6px; grid-template-columns:24px 1fr auto; align-items:center; padding:8px 10px; border-bottom:1px solid #243b61; cursor:pointer; }} .row.sel {{ background:#193766; box-shadow:inset 0 0 0 1px #5d89cf; }} .rank {{ color:#5eead4; font-weight:800; }} .sub {{ color:var(--muted); font-size:.72rem; }} .score {{ border:1px solid #3f5f93; border-radius:999px; padding:2px 8px; font-weight:800; display:inline-flex; gap:4px; align-items:center; }}
    .leader-metrics {{ display:flex; gap:6px; flex-wrap:wrap; margin-top:2px; }}
    .metric-chip {{ display:inline-flex; align-items:center; gap:4px; border:1px solid #36598a; border-radius:999px; padding:2px 8px; background:#112546; }}
    .metric-chip .metric-value {{ color:#e2ecff; font-weight:900; font-size:.74rem; }}
    .metric-chip .metric-value.warn {{ color:#f59e0b; }}
    .metric-chip .material-symbols-outlined {{ font-size:15px; color:#93c5fd; font-variation-settings:"FILL" 1, "wght" 500, "GRAD" 0, "opsz" 20; }}
    .assignee-refresh-btn {{ width:26px; height:26px; border:1px solid #36598a; border-radius:999px; background:#112546; color:#93c5fd; display:inline-flex; align-items:center; justify-content:center; cursor:pointer; }}
    .assignee-refresh-btn:hover {{ border-color:#7cb2ff; color:#dbeafe; background:#17325a; }}
    .assignee-refresh-btn:disabled {{ opacity:.55; cursor:progress; border-color:#476b9f; color:#a8c8ff; background:#17325a; }}
    .assignee-refresh-btn .material-symbols-outlined {{ font-size:16px; }}
    .row-refresh {{ grid-column:1 / -1; margin-top:2px; padding:6px 8px; border-radius:10px; border:1px solid #29476f; background:#0f1f3a; }}
    .row-refresh-head {{ display:flex; justify-content:space-between; gap:8px; align-items:center; font-size:.7rem; color:#dbeafe; }}
    .row-refresh-status {{ font-weight:800; }}
    .row-refresh-pct {{ color:#93c5fd; font-weight:800; }}
    .row-refresh-track {{ margin-top:6px; width:100%; height:7px; border-radius:999px; background:#09162c; border:1px solid #223a61; overflow:hidden; }}
    .row-refresh-fill {{ height:100%; background:linear-gradient(90deg,#38bdf8,#60a5fa); transition:width .25s ease; }}
    .row-refresh-sub {{ margin-top:4px; font-size:.67rem; color:#9db1d8; white-space:nowrap; overflow:hidden; text-overflow:ellipsis; }}
    .row-refresh.row-refresh-err {{ border-color:#7f1d1d; background:#2c1016; }}
    .row-refresh.row-refresh-err .row-refresh-fill {{ background:linear-gradient(90deg,#f97316,#ef4444); }}
    .row-refresh.row-refresh-ok {{ border-color:#166534; background:#0f2418; }}
    .row-refresh.row-refresh-ok .row-refresh-fill {{ background:linear-gradient(90deg,#22c55e,#86efac); }}
    .detail {{ padding:10px; overflow-y:auto; flex:1; min-height:0; }} .score-arena {{ margin-top:10px; }} .score-drill {{ padding:10px; }} .card {{ border:1px solid #314e7f; border-radius:10px; background:#12213d; padding:10px; }} .big {{ font-size:2rem; font-weight:900; line-height:1; }}
    .tabs {{ display:flex; gap:6px; flex-wrap:wrap; margin-top:8px; }}
    .tab-btn {{ border:1px solid #3b5f91; background:#12284b; color:#e6efff; border-radius:999px; padding:4px 10px; font-size:.74rem; cursor:pointer; }}
    .tab-btn.active {{ border-color:#7cb2ff; box-shadow:inset 0 0 0 1px #7cb2ff; background:#173866; }}
    .tabs[data-tab-group="scoring"] {{ display:grid; grid-template-columns:repeat(auto-fit,minmax(220px,1fr)); gap:10px; }}
    .tabs[data-tab-group="scoring"] .tab-btn {{ min-height:66px; border-radius:12px; padding:9px 12px; display:flex; flex-direction:column; align-items:flex-start; justify-content:center; gap:4px; font-size:.79rem; transition:border-color .15s ease, box-shadow .15s ease, background .15s ease; }}
    .tabs[data-tab-group="scoring"] .tab-btn.active {{ border-color:#93c5fd; box-shadow:inset 0 0 0 1px #93c5fd; background:#173866; }}
    .tab-kicker {{ font-size:.62rem; line-height:1; color:#9bb5df; text-transform:uppercase; letter-spacing:.06em; }}
    .tab-main-row {{ width:100%; display:flex; justify-content:space-between; align-items:center; gap:10px; }}
    .tab-title {{ font-size:.86rem; font-weight:800; color:#eff6ff; }}
    .tab-score {{ font-size:1rem; font-weight:900; color:#e0f2fe; }}
    .tab-btn.tab-btn-beta {{ border-color:#7c6a1c; background:linear-gradient(180deg,#2f2a15 0%,#251f10 100%); }}
    .tab-btn.tab-btn-beta .tab-score {{ color:#fde68a; }}
    .tab-btn.tab-btn-beta.active {{ border-color:#facc15; box-shadow:inset 0 0 0 1px #facc15; background:linear-gradient(180deg,#43380d 0%,#362b08 100%); }}
    .tab-beta {{ display:inline-flex; align-items:center; border:1px solid #facc15; color:#facc15; border-radius:999px; padding:1px 7px; margin-left:6px; font-size:.64rem; font-weight:800; text-transform:uppercase; vertical-align:middle; }}
    .assignment-scorecards {{ display:grid; gap:8px; grid-template-columns:repeat(4,minmax(140px,1fr)); margin-top:10px; }}
    .assignment-card {{ border:1px solid #375f95; border-radius:10px; background:#12284b; padding:8px 10px; }}
    .assignment-card .k {{ font-size:.66rem; text-transform:uppercase; letter-spacing:.05em; color:#cfe0ff; font-weight:800; }}
    .assignment-card .v {{ margin-top:3px; font-size:1.05rem; font-weight:900; color:#eef4ff; }}
    .assignment-card.epic {{ border-color:#c084fc; background:rgba(192,132,252,.14); }}
    .assignment-card.story {{ border-color:#60a5fa; background:rgba(96,165,250,.14); }}
    .assignment-card.subtask {{ border-color:#4ade80; background:rgba(74,222,128,.14); }}
    .assignment-card.bug {{ border-color:#f87171; background:rgba(248,113,113,.14); }}
    @media (max-width: 980px) {{ .assignment-scorecards {{ grid-template-columns:repeat(2,minmax(140px,1fr)); }} }}
    .tab-pane {{ display:none; }}
    .tab-pane.active {{ display:block; }}
    .grid2 {{ display:grid; gap:8px; grid-template-columns:repeat(2,minmax(0,1fr)); margin-top:8px; }} .mini {{ border:1px solid #2b446e; border-radius:10px; background:#0f1b32; padding:8px; }} .mini h3 {{ margin:0 0 6px; font-size:.8rem; }} .mini .l {{ display:flex; justify-content:space-between; font-size:.78rem; padding:2px 0; }}
    .mini .l.actionable {{ cursor:pointer; border-radius:6px; transition:background .15s ease; }}
    .mini .l.actionable:hover {{ background:#173158; }}
    .metric-link-btn {{ border:1px solid #36598a; border-radius:999px; background:#12284b; color:#dce8ff; font-size:.72rem; padding:2px 8px; cursor:pointer; }}
    .metric-link-btn:hover {{ background:#1a3a67; border-color:#5f88c0; }}
    .jira-link-icon {{ display:inline-flex; align-items:center; justify-content:center; width:22px; height:22px; border:1px solid #3a5c91; border-radius:6px; color:#cfe0ff; background:#0f2342; text-decoration:none; }}
    .jira-link-icon:hover {{ border-color:#7cb2ff; color:#ffffff; background:#17325a; }}
    .jira-link-icon .material-symbols-outlined {{ font-size:15px; }}
    .jira-link-disabled {{ color:#6e85ac; font-size:.72rem; }}
    .ts-card {{ margin-top:8px; border:1px solid #2b446e; border-radius:10px; background:#0f1b32; padding:8px; }}
    .ts-svg {{ width:100%; height:220px; display:block; border:1px solid #223a61; border-radius:8px; background:#0d172b; }}
    .kpi-charts {{ display:grid; gap:8px; grid-template-columns:repeat(2,minmax(0,1fr)); margin-top:8px; }}
    .mini-chart {{ border:1px solid #2b446e; border-radius:10px; background:#0f1b32; padding:8px; }}
    .mini-svg {{ width:100%; height:160px; display:block; border:1px solid #223a61; border-radius:8px; background:#0d172b; }}
    .tbl-wrap {{ margin-top:8px; border:1px solid #2b446e; border-radius:10px; background:#0f1b32; overflow:auto; }}
    .tbl-title {{ margin:0; padding:8px; font-size:.82rem; border-bottom:1px solid #21385e; }}
    .tbl {{ width:100%; border-collapse:collapse; min-width:720px; }}
    .tbl th,.tbl td {{ border-bottom:1px solid #21385e; padding:6px 8px; font-size:.76rem; vertical-align:top; }}
    .tbl th {{ color:#9db1d8; text-transform:uppercase; font-size:.68rem; letter-spacing:.04em; position:sticky; top:0; background:#10223f; }}
    .tbl tr.exec-negative-subtask td {{ background:rgba(244,63,94,.12); }}
    .tbl tr.exec-negative-subtask .issue-id,.tbl tr.exec-negative-subtask .issue-title {{ color:#fecdd3; }}
    .tbl tr.due-missed td {{ background:rgba(249,115,22,.12); }}
    .tbl tr.penalized-row td, .ss-tbl tr.penalized-row td {{ background:rgba(190,24,93,.12); }}
    .tbl tr.penalized-row td:first-child, .ss-tbl tr.penalized-row td:first-child {{ box-shadow:inset 3px 0 0 #fb7185; }}
    .tbl tr.penalized-row .issue-id, .tbl tr.penalized-row .issue-title, .ss-tbl tr.penalized-row .issue-id, .ss-tbl tr.penalized-row .issue-title {{ color:#fecdd3; }}
    .due-bucket {{ display:inline-flex; align-items:center; border-radius:999px; padding:2px 8px; border:1px solid transparent; font-weight:700; }}
    .due-before {{ background:rgba(34,197,94,.18); border-color:#22c55e; color:#bbf7d0; }}
    .due-on {{ background:rgba(56,189,248,.18); border-color:#38bdf8; color:#bae6fd; }}
    .due-after {{ background:rgba(249,115,22,.2); border-color:#f97316; color:#fed7aa; }}
    .exec-accordion {{ padding:8px; display:grid; gap:8px; }}
    .exec-epic {{ border:1px solid #2f4f7f; border-radius:10px; background:#10203c; overflow:hidden; }}
    .exec-epic > summary {{ list-style:none; cursor:pointer; display:flex; justify-content:space-between; gap:8px; align-items:flex-start; padding:8px; }}
    .exec-epic > summary::-webkit-details-marker {{ display:none; }}
    .exec-epic-left {{ min-width:0; }}
    .exec-epic-metrics {{ display:flex; gap:6px; flex-wrap:wrap; justify-content:flex-end; }}
    .exec-epic-body {{ border-top:1px solid #29476f; padding:8px; display:grid; gap:8px; }}
    .exec-story-block {{ border:1px solid #29476f; border-radius:10px; background:#0e1c35; overflow:hidden; }}
    .exec-story-head {{ display:flex; justify-content:space-between; gap:8px; align-items:flex-start; padding:8px; border-bottom:1px solid #21385e; }}
    .exec-story-metrics {{ display:flex; gap:6px; flex-wrap:wrap; justify-content:flex-end; }}
    .exec-subtask-table {{ min-width:0; border-top:0; }}
    .exec-neg-pill {{ border-color:#fb7185; color:#fecdd3; background:#3b1020; }}
    .focus-pulse {{ animation:focusPulse 1.2s ease; }}
    @keyframes focusPulse {{ 0% {{ box-shadow:0 0 0 0 rgba(56,189,248,.65); }} 100% {{ box-shadow:0 0 0 12px rgba(56,189,248,0); }} }}
    .issue-id {{ font-size:.67rem; color:#8fb1e8; font-family:Consolas,monospace; }}
    .issue-title {{ font-size:.83rem; color:#e2ebff; line-height:1.25; }}
    .tree-wrap {{ border:1px solid #2b446e; border-radius:10px; background:#0f1b32; padding:8px; margin-top:8px; }}
    .tree-node {{ margin:4px 0; }}
    .tree-node details {{ border:1px solid #223a61; border-radius:8px; background:#112241; padding:6px; }}
    .tree-node summary {{ cursor:pointer; list-style:none; display:flex; justify-content:space-between; gap:8px; align-items:flex-start; }}
    .tree-node summary::-webkit-details-marker {{ display:none; }}
    .tree-left {{ min-width:0; }}
    .tree-metrics {{ display:flex; gap:6px; flex-wrap:wrap; justify-content:flex-end; }}
    .metric-pill {{ border:1px solid #375f95; border-radius:999px; padding:2px 8px; font-size:.68rem; color:#d7e7ff; background:#17325a; white-space:nowrap; }}
    .issue-type-pill {{ display:inline-flex; align-items:center; justify-content:center; min-width:30px; padding:2px 7px; }}
    .issue-type-pill .material-symbols-outlined {{ font-size:18px; line-height:1; font-variation-settings:"FILL" 1, "wght" 500, "GRAD" 0, "opsz" 24; }}
    .issue-type-pill.issue-epic .material-symbols-outlined {{ color:#c084fc; }}
    .issue-type-pill.issue-story .material-symbols-outlined {{ color:#60a5fa; }}
    .issue-type-pill.issue-subtask .material-symbols-outlined {{ color:#4ade80; }}
    .issue-type-pill.issue-bug-subtask .material-symbols-outlined {{ color:#f87171; }}
    .subtask-type-icon {{ display:inline-flex; align-items:center; gap:6px; font-size:12px; font-weight:700; }}
    .subtask-type-icon .material-symbols-outlined {{ font-size:16px; line-height:1; font-variation-settings:"FILL" 1, "wght" 500, "GRAD" 0, "opsz" 20; }}
    .subtask-type-icon.issue-subtask {{ color:#4ade80; }}
    .subtask-type-icon.issue-bug-subtask {{ color:#f87171; }}
    .issue-kind-inline {{ margin-left:6px; font-size:.68rem; font-weight:800; color:#fecdd3; }}
    .tree-children {{ margin-left:14px; margin-top:6px; border-left:1px dashed #2e4f7d; padding-left:8px; }}
    .exec-metrics {{ display:grid; gap:10px; }}
    .exec-metric {{ border:1px solid #29486f; border-radius:8px; padding:8px; background:#10223f; }}
    .exec-metric.actionable {{ cursor:pointer; }}
    .exec-metric.actionable:hover {{ border-color:#5f88c0; background:#12284b; }}
    .exec-m-head {{ display:flex; justify-content:space-between; gap:8px; align-items:center; }}
    .exec-m-name {{ font-size:.86rem; font-weight:800; color:#e6f0ff; }}
    .exec-m-value {{ font-size:.88rem; font-weight:900; color:#7dd3fc; }}
    .exec-m-meaning {{ margin-top:2px; font-size:.74rem; color:#b7c9e8; line-height:1.3; }}
    .exec-bar-track {{ margin-top:6px; width:100%; height:12px; border-radius:999px; border:1px solid #2f4e7d; background:#0a1a33; overflow:hidden; }}
    .exec-bar-fill {{ height:100%; background:linear-gradient(90deg,#38bdf8,#60a5fa); }}
    .exec-scale-note {{ font-size:.72rem; color:#93acd2; margin-top:2px; }}
    .availability-breakdown {{ margin-top:8px; border-top:1px dashed #2f4e7d; padding-top:8px; display:grid; gap:4px; }}
    .availability-line {{ display:flex; justify-content:space-between; gap:8px; font-size:.75rem; }}
    .availability-name {{ color:#dce8ff; }}
    .availability-num {{ color:#f8fafc; font-weight:800; font-family:Consolas,monospace; }}
    .availability-note {{ color:#93acd2; font-size:.72rem; }}
    .neg {{ color:var(--bad); font-weight:700; }} .feed {{ margin-top:8px; border:1px solid #2b446e; border-radius:10px; max-height:230px; overflow:auto; background:#0e182d; }} .feed .i {{ padding:7px 8px; border-bottom:1px solid #21385e; font-size:.76rem; }} .empty {{ color:var(--muted); font-style:italic; }}
    .scoring-section {{ margin-top:10px; border:1px solid #2b446e; border-radius:10px; background:#0f1b32; padding:10px; }}
    .score-drawer-overlay {{ position:fixed; inset:0; background:rgba(2,8,23,.56); opacity:0; pointer-events:none; transition:opacity .18s ease; z-index:95; }}
    .score-drawer-overlay.open {{ opacity:1; pointer-events:auto; }}
    .score-drawer {{ position:fixed; top:0; right:0; width:40vw; max-width:96vw; height:100vh; background:#0d1830; border-left:1px solid #315b92; box-shadow:-20px 0 40px rgba(2,8,23,.45); transform:translateX(100%); transition:transform .2s ease; z-index:96; display:flex; flex-direction:column; overflow-y:scroll; }}
    .score-drawer.open {{ transform:translateX(0); }}
    .score-drawer-head {{ padding:14px; border-bottom:1px solid #223a61; display:flex; align-items:flex-start; justify-content:space-between; gap:10px; background:#10223f; }}
    .score-drawer-title {{ margin:0; font-size:1rem; font-weight:900; color:#eef4ff; }}
    .score-drawer-subtitle {{ margin:4px 0 0; color:#9db1d8; font-size:.76rem; line-height:1.35; }}
    .score-drawer-body {{ padding:14px; overflow:auto; display:grid; gap:12px; }}
    .score-drawer-close {{ border:1px solid #4a6ea9; background:#1b325a; color:#eef4ff; border-radius:8px; font-weight:700; padding:7px 10px; cursor:pointer; }}
    .score-drawer-section {{ border:1px solid #223a61; border-radius:12px; background:#0f1b32; overflow:hidden; display:flex; flex-direction:column; min-height:0; }}
    .score-drawer-section-head {{ padding:10px 12px; border-bottom:1px solid #223a61; display:flex; justify-content:space-between; gap:8px; align-items:center; background:#10223f; }}
    .score-drawer-section-title {{ margin:0; font-size:.82rem; text-transform:uppercase; letter-spacing:.05em; color:#cfe0ff; }}
    .score-drawer-section-note {{ color:#93acd2; font-size:.72rem; }}
    .score-drawer-section-content {{ min-height:0; overflow-y:auto; scrollbar-width:thin; scrollbar-color:#4f7fb8 #0b213e; }}
    .score-drawer-section-content::-webkit-scrollbar {{ width:10px; }}
    .score-drawer-section-content::-webkit-scrollbar-track {{ background:#0b213e; border-radius:999px; }}
    .score-drawer-section-content::-webkit-scrollbar-thumb {{ background:#4f7fb8; border-radius:999px; border:2px solid #0b213e; }}
    .score-drawer-section-content::-webkit-scrollbar-thumb:hover {{ background:#6fa0de; }}
    .score-drawer-section-rules .score-drawer-section-content {{ max-height:min(40vh, 420px); }}
    .score-drawer-section-penalties .score-drawer-section-content {{ max-height:min(42vh, 460px); }}
    .score-rule-list {{ display:grid; gap:8px; padding:12px; }}
    .score-rule-card {{ border:1px solid #29486f; border-radius:10px; background:#10223f; padding:10px; display:grid; gap:4px; }}
    .score-rule-top {{ display:flex; justify-content:space-between; gap:8px; align-items:flex-start; }}
    .score-rule-name {{ font-size:.82rem; font-weight:800; color:#eef4ff; }}
    .score-rule-impact {{ font-size:.82rem; font-weight:900; color:#7dd3fc; white-space:nowrap; }}
    .score-rule-impact.neg {{ color:#fda4af; }}
    .score-rule-impact.pos {{ color:#86efac; }}
    .score-rule-desc {{ color:#c7d7f3; font-size:.74rem; line-height:1.35; }}
    .score-rule-meta {{ color:#93acd2; font-size:.71rem; line-height:1.35; }}
    .score-subtask-table-wrap {{ padding:12px; }}
    .score-drawer-empty {{ color:#9db1d8; font-size:.78rem; font-style:italic; padding:12px; }}
    body.score-drawer-open {{ overflow:hidden; }}
    .scoring-section-head {{ display:flex; justify-content:space-between; align-items:center; margin-bottom:8px; }}
    .scoring-section-title {{ font-size:.88rem; font-weight:800; color:#e6f0ff; }}
    .scoring-section-title .beta-tag {{ font-size:.66rem; font-weight:700; color:#facc15; border:1px solid #facc15; border-radius:999px; padding:1px 7px; margin-left:6px; vertical-align:middle; text-transform:uppercase; }}
    .ss-toggle {{ display:flex; align-items:center; gap:6px; }}
    .ss-toggle-label {{ font-size:.68rem; color:#7fa3d6; user-select:none; }}
    .ss-switch {{ position:relative; width:32px; height:18px; }}
    .ss-switch input {{ opacity:0; width:0; height:0; }}
    .ss-switch .ss-slider {{ position:absolute; inset:0; background:#1e3560; border:1px solid #3a5c91; border-radius:999px; cursor:pointer; transition:background .2s; }}
    .ss-switch .ss-slider::before {{ content:""; position:absolute; left:2px; top:2px; width:12px; height:12px; background:#7fa3d6; border-radius:50%; transition:transform .2s; }}
    .ss-switch input:checked + .ss-slider {{ background:#1b4a7a; border-color:#5f88c0; }}
    .ss-switch input:checked + .ss-slider::before {{ transform:translateX(14px); background:#93c5fd; }}
    .ss-big-score {{ font-size:1.6rem; font-weight:900; line-height:1; margin:4px 0; }}
    .summary-score-trigger {{ display:inline-flex; align-items:center; gap:6px; border:0; background:none; color:inherit; padding:0; cursor:pointer; font:inherit; }}
    .summary-score-trigger .material-symbols-outlined {{ font-size:16px; color:#93c5fd; }}
    .summary-score-trigger:hover .material-symbols-outlined, .summary-score-trigger:hover .score-label-text {{ color:#ffffff; }}
    .score-label-text {{ text-decoration:underline; text-decoration-style:dotted; text-underline-offset:3px; }}
    .ss-summary-row {{ display:flex; gap:8px; flex-wrap:wrap; margin-top:6px; }}
    .ss-chip {{ border:1px solid #365c8d; border-radius:999px; padding:3px 9px; font-size:.72rem; color:#dce8ff; background:#132949; }}
    .ss-tbl {{ width:100%; border-collapse:collapse; margin-top:8px; font-size:.74rem; }}
    .ss-tbl th {{ color:#9db1d8; text-transform:uppercase; font-size:.66rem; letter-spacing:.04em; padding:5px 6px; border-bottom:1px solid #21385e; text-align:left; position:sticky; top:0; background:#10223f; }}
    .ss-tbl td {{ padding:5px 6px; border-bottom:1px solid #1b2f52; vertical-align:top; }}
    .ramadan-chip {{ display:inline-flex; align-items:center; margin-left:6px; border:1px solid #7c3aed; border-radius:999px; padding:1px 7px; font-size:.66rem; font-weight:800; color:#e9d5ff; background:rgba(124,58,237,.22); }}
    .ss-tbl .ss-row-within td {{ background:rgba(34,197,94,.10); }}
    .ss-tbl .ss-row-commitment td {{ background:rgba(129,140,248,.12); }}
    .ss-tbl .ss-row-over td {{ background:rgba(249,115,22,.10); }}
    .ss-tbl .ss-row-over-late td {{ background:rgba(251,113,133,.10); }}
    .ss-status-pill {{ display:inline-block; border-radius:999px; padding:2px 8px; font-size:.68rem; font-weight:700; border:1px solid transparent; }}
    .ss-pill-within {{ background:rgba(34,197,94,.18); border-color:#22c55e; color:#bbf7d0; }}
    .ss-pill-commitment {{ background:rgba(129,140,248,.18); border-color:#818cf8; color:#c7d2fe; }}
    .ss-pill-over {{ background:rgba(249,115,22,.2); border-color:#f97316; color:#fed7aa; }}
    .ss-pill-late {{ background:rgba(251,113,133,.18); border-color:#fb7185; color:#fecdd3; }}
    .ss-pill-noest {{ background:rgba(148,163,184,.15); border-color:#94a3b8; color:#cbd5e1; }}
    .ss-donut-wrap {{ display:flex; gap:12px; align-items:center; margin-top:8px; }}
    .ss-legend {{ font-size:.72rem; display:grid; gap:3px; }}
    .ss-legend-dot {{ display:inline-block; width:10px; height:10px; border-radius:50%; margin-right:4px; vertical-align:middle; }}
    .formula-guide {{ margin-top:10px; border:1px solid #315b92; border-radius:12px; background:linear-gradient(180deg,#0f2746 0%, #0b1d37 100%); padding:12px; color:#e8f1ff; }}
    .formula-head {{ display:flex; align-items:center; justify-content:space-between; gap:10px; margin-bottom:10px; }}
    .formula-title {{ margin:0; font-size:clamp(1rem,1.2vw,1.14rem); font-weight:900; letter-spacing:.01em; }}
    .formula-head-actions {{ display:flex; align-items:center; gap:8px; }}
    .formula-toggle-btn {{ border:1px solid #4d79b3; background:#16365f; color:#eaf3ff; border-radius:999px; padding:4px 10px; display:inline-flex; align-items:center; gap:4px; cursor:pointer; font-size:.8rem; font-weight:700; }}
    .formula-toggle-btn:hover {{ background:#1c4477; border-color:#6ea8ff; }}
    .formula-toggle-btn .material-symbols-outlined {{ font-size:16px; }}
    .formula-score-pill {{ border:1px solid #6ea8ff; background:#173866; color:#eaf3ff; border-radius:999px; padding:4px 10px; font-weight:900; font-size:clamp(.95rem,1.1vw,1.02rem); }}
    .formula-layout {{ display:grid; gap:10px; grid-template-columns:1.25fr .85fr; }}
    .formula-layout[hidden] {{ display:none !important; }}
    .formula-steps {{ border:1px solid #2f5588; border-radius:10px; background:#112b4d; padding:10px; }}
    .formula-step {{ margin-bottom:10px; }}
    .formula-step:last-child {{ margin-bottom:0; }}
    .formula-step.step-gap {{ margin-bottom:18px; padding-bottom:10px; border-bottom:1px dashed #3f6aa3; }}
    .formula-kicker {{ font-size:.78rem; text-transform:uppercase; letter-spacing:.06em; color:#9ec2f5; margin-bottom:4px; font-weight:800; }}
    .formula-eq {{ margin:0; font-family:"Consolas","Lucida Console","Courier New",monospace; font-size:clamp(1rem,1.15vw,1.08rem); font-weight:800; color:#f4f8ff; line-height:1.45; word-break:break-word; }}
    .formula-applied {{ margin-top:6px; font-size:clamp(.92rem,1vw,.98rem); color:#d4e5ff; }}
    .formula-note {{ margin-top:8px; font-size:clamp(.86rem,.95vw,.92rem); color:#b6cdef; }}
    .formula-safeguards {{ margin-top:10px; border:1px solid #315b92; border-radius:8px; background:#0d223e; padding:8px 10px; }}
    .formula-safeguards-title {{ font-size:.78rem; text-transform:uppercase; letter-spacing:.06em; color:#9ec2f5; font-weight:800; margin-bottom:4px; }}
    .formula-safeguard-item {{ font-size:clamp(.86rem,.95vw,.92rem); color:#d4e5ff; line-height:1.45; margin:2px 0; }}
    .formula-metrics {{ border:1px solid #2f5588; border-radius:10px; background:#102744; padding:10px; display:grid; gap:6px; align-content:start; }}
    .formula-row {{ display:flex; justify-content:space-between; align-items:center; gap:10px; font-size:clamp(.94rem,1vw,1rem); }}
    .formula-row span:first-child {{ color:#cfe1fb; }}
    .formula-row span:last-child {{ font-weight:800; color:#ffffff; }}
    .formula-row.final {{ margin-top:4px; padding-top:8px; border-top:1px dashed #3f6aa3; font-size:clamp(1rem,1.12vw,1.08rem); }}
    .formula-mini-help {{ margin-top:4px; font-size:.8rem; color:#9fbbe3; }}
    @media (max-width:900px) {{ .formula-layout {{ grid-template-columns:1fr; }} }}
    .score-label {{ font-size:.66rem; color:var(--muted); text-transform:uppercase; letter-spacing:.04em; }}
    .capacity-expanded {{ margin-top:10px; border:1px solid #36598a; border-radius:10px; background:#0d2543; padding:10px; }}
    .capacity-expanded-head {{ display:flex; justify-content:space-between; align-items:center; gap:10px; flex-wrap:wrap; margin-bottom:8px; }}
    .capacity-expanded-title {{ font-size:.86rem; font-weight:800; color:#e2ecff; letter-spacing:.02em; text-transform:uppercase; }}
    .capacity-expanded-sub {{ font-size:.74rem; color:#b7caea; }}
    .capacity-expanded-grid {{ display:grid; grid-template-columns:repeat(auto-fit, minmax(180px, 1fr)); gap:6px; margin-bottom:8px; }}
    .capacity-chip {{ border:1px solid #294b78; border-radius:8px; background:#102b4d; padding:6px 8px; }}
    .capacity-chip .k {{ font-size:.66rem; color:#9bb7df; text-transform:uppercase; letter-spacing:.04em; }}
    .capacity-chip .v {{ font-size:.85rem; color:#edf3ff; font-weight:700; margin-top:2px; }}
    .capacity-legend {{ display:flex; gap:6px; flex-wrap:wrap; margin-bottom:8px; }}
    .capacity-legend .pill {{ font-size:.68rem; padding:2px 7px; border-radius:999px; border:1px solid #355d92; color:#d9e7ff; background:#0f2a4b; }}
    .capacity-calendar-wrap {{ display:flex; gap:8px; overflow-x:auto; padding-bottom:6px; scroll-snap-type:x proximity; scrollbar-width:thin; scrollbar-color:#4f7fb8 #0b213e; }}
    .capacity-calendar-wrap::-webkit-scrollbar {{ height:10px; }}
    .capacity-calendar-wrap::-webkit-scrollbar-track {{ background:#0b213e; border-radius:999px; }}
    .capacity-calendar-wrap::-webkit-scrollbar-thumb {{ background:#4f7fb8; border-radius:999px; border:2px solid #0b213e; }}
    .capacity-calendar-wrap::-webkit-scrollbar-thumb:hover {{ background:#6fa0de; }}
    .capacity-month {{ border:1px solid #2f5486; border-radius:10px; background:#0f2b4d; padding:8px; flex:0 0 calc((100% - 16px)/3); min-width:280px; scroll-snap-align:start; }}
    .capacity-month-head {{ font-size:.8rem; font-weight:800; color:#dbe8ff; margin-bottom:6px; }}
    .capacity-month-grid {{ display:grid; grid-template-columns:repeat(7, minmax(0, 1fr)); gap:4px; }}
    .capacity-dow {{ font-size:.62rem; color:#8ea9cf; text-transform:uppercase; text-align:center; }}
    .capacity-day {{ min-height:44px; border:1px solid #22466f; border-radius:6px; background:#0c2340; padding:3px 4px; }}
    .capacity-day.is-out {{ opacity:.35; }}
    .capacity-day.is-weekend {{ background:#616161; border-color:#8b8b8b; }}
    .capacity-day.is-ramadan {{ border-color:#0ea5e9; box-shadow:inset 0 0 0 1px rgba(14,165,233,.25); }}
    .capacity-day.is-holiday {{ border-color:#f59e0b; box-shadow:inset 0 0 0 1px rgba(245,158,11,.25); }}
    .capacity-day.has-leave {{ border-color:#6b7280; box-shadow:inset 0 0 0 1px rgba(148,163,184,.25); }}
    .capacity-day.has-ramadan-leave {{ border-color:#0ea5e9; box-shadow:inset 0 0 0 1px rgba(14,165,233,.35); }}
    .capacity-day.is-today {{ background:#fde047; border-color:#facc15; box-shadow:inset 0 0 0 1px rgba(234,179,8,.55), 0 0 0 1px rgba(250,204,21,.35); }}
    .capacity-day.is-today .capacity-day-num {{ color:#1f2937; font-weight:900; }}
    .capacity-day.is-today .capacity-day-tag {{ color:#0f172a; border-color:#b45309; background:#fcd34d; }}
    .capacity-day-num {{ font-size:.72rem; font-weight:700; color:#dce9ff; line-height:1.1; }}
    .capacity-day-tags {{ margin-top:2px; display:flex; gap:3px; flex-wrap:wrap; }}
    .capacity-day-tag {{ font-size:.58rem; line-height:1; padding:2px 4px; border-radius:999px; border:1px solid #355c8f; background:#173761; color:#d9e8ff; }}
    .capacity-day-tag.r {{ border-color:#0ea5e9; background:#0c3a58; }}
    .capacity-day-tag.h {{ border-color:#f59e0b; background:#4a2a00; }}
    .capacity-day-tag.l {{ border-color:#64748b; background:#334155; }}
    .capacity-day-tag.rl {{ border-color:#0ea5e9; background:#0c3a58; }}
    .capacity-empty {{ font-size:.76rem; color:#9fb7db; }}
    @media (max-width:1200px) {{ .capacity-month{{ flex-basis:calc((100% - 8px)/2); }} }}
    @media (max-width:760px) {{ .capacity-month{{ flex-basis:100%; }} }}
    @media (max-width:1200px) {{ .toolbar{{grid-template-columns:1fr 1fr;}} .kpis{{grid-template-columns:1fr 1fr;}} .top3-wrap{{grid-template-columns:1fr;}} .team-score-layout{{grid-template-columns:1fr;}} .arena{{grid-template-columns:1fr;}} .arena > .panel {{ min-height:auto; max-height:none; }} .leaderboard {{ max-height:56vh; }} .detail {{ max-height:56vh; }} }}
  </style>
  <link rel="stylesheet" href="shared-nav.css">
  <link rel="stylesheet" href="https://fonts.googleapis.com/css2?family=Material+Symbols+Outlined:FILL,wght,GRAD,opsz@1,500,0,20">
</head>
<body>
<div class="top-date-range-wrap">
  <div class="top-date-range-chip" aria-label="Date range filter">
    <span class="date-chip-segment">From</span>
    <input id="from" class="date-chip-input" type="date" aria-label="From date">
    <span class="date-chip-segment">To</span>
    <input id="to" class="date-chip-input" type="date" aria-label="To date">
    <span class="date-chip-segment">Capacity Profile</span>
    <select id="capacity-profile-top" class="date-chip-select" aria-label="Capacity profile (top)"></select>
    <button id="apply" class="btn" type="button">Apply Filters</button>
    <button id="reset" class="btn" type="button">Reset</button>
    <div class="adv-filter-wrap">
      <button id="adv-filter-toggle" class="adv-filter-btn" type="button" aria-expanded="false" aria-haspopup="true" aria-controls="adv-filter-menu">Advanced Filters</button>
      <div class="adv-filter-menu" id="adv-filter-menu" role="menu" hidden>
        <div class="adv-filter-group-label">Date Presets</div>
        <button class="adv-filter-item" type="button" data-preset="last30" role="menuitem">Last 30 Days</button>
        <button class="adv-filter-item" type="button" data-preset="lastMonth" role="menuitem">Last Month</button>
        <button class="adv-filter-item" type="button" data-preset="currentMonth" role="menuitem">Current Month</button>
        <button class="adv-filter-item" type="button" data-preset="last90" role="menuitem">Last 90 Days</button>
        <button class="adv-filter-item" type="button" data-preset="lastQuarter" role="menuitem">Last Quarter</button>
        <button class="adv-filter-item" type="button" data-preset="currentQuarter" role="menuitem">Current Quarter</button>
      </div>
    </div>
    <div class="date-chip-control">
      <span class="date-chip-label">Simple Overrun Basis</span>
      <select id="simple-overrun-mode" class="date-chip-select" aria-label="Simple overrun basis">
        <option value="subtasks" selected>Overrun Subtask Hours</option>
        <option value="total">Total Overrun Hours</option>
      </select>
    </div>
    <span id="date-filter-status" class="date-chip-status" aria-live="polite"></span>
  </div>
</div>
<button id="header-expand-fab" class="header-expand-fab" type="button" aria-label="Expand header" title="Expand header">&#9776;</button>
<div id="score-detail-drawer-overlay" class="score-drawer-overlay" aria-hidden="true"></div>
<aside id="score-detail-drawer" class="score-drawer" role="dialog" aria-modal="true" aria-hidden="true" aria-labelledby="score-detail-drawer-title">
  <div class="score-drawer-head">
    <div>
      <h2 id="score-detail-drawer-title" class="score-drawer-title">Simple Score Details</h2>
      <p id="score-detail-drawer-subtitle" class="score-drawer-subtitle"></p>
    </div>
    <button id="score-detail-drawer-close" class="score-drawer-close" type="button">Close</button>
  </div>
  <div id="score-detail-drawer-body" class="score-drawer-body"></div>
</aside>
<div class="wrap">
  <section class="hero" id="performance-header">
    <div class="hero-top">
      <div>
        <h1>Employee Performance Dashboard</h1>
        <div class="meta" id="meta"></div>
      </div>
      <div class="hero-actions">
        <div class="report-refresh-wrap">
          <button id="employee-refresh-btn" class="btn" type="button">Refresh Report</button>
          <button id="employee-refresh-cancel-btn" class="btn" type="button">Cancel Run</button>
          <span id="employee-refresh-status" class="report-refresh-status" aria-live="polite"></span>
          <div id="employee-refresh-details" class="report-refresh-details" aria-live="polite"></div>
        </div>
        <button id="header-toggle" class="btn" type="button" aria-expanded="true" aria-controls="performance-header">Collapse Header</button>
      </div>
    </div>
    <section class="guide">
      <h2>Executive View Guide</h2>
      <p>Start with Planning & Start-Adherence KPIs to check workload realism, then use Performance Score KPIs for risk posture. In leaderboard, sort by the lens you want and click a person for full diagnostic detail.</p>
      <div class="discover" id="discover-insights"></div>
    </section>
    <section class="hero-scorecard" aria-label="Average performance percentage">
      <div class="hero-scorecard-copy">
        <div class="hero-scorecard-kicker">Average Performance</div>
        <div class="hero-scorecard-title">Current filtered team performance percentage</div>
        <div class="hero-scorecard-sub">Updates from the active date range, project selection, assignee search, and scoring mode.</div>
      </div>
      <div class="hero-scorecard-metric">
        <div class="hero-scorecard-value" id="header-average-performance-value">N/A</div>
        <div class="hero-scorecard-mode" id="header-average-performance-mode"><span class="material-symbols-outlined" aria-hidden="true">monitoring</span><span>Simple Scoring</span></div>
        <div class="hero-scorecard-meta" id="header-average-performance-meta">Eligible assignees: 0</div>
      </div>
    </section>
    <div class="toolbar">
      <div class="f"><label for="projects">Project</label><select id="projects" multiple size="1"></select></div>
      <div class="f"><label for="capacity-profile">Capacity Profile</label><select id="capacity-profile"></select></div>
      <div class="f"><label for="search">Search Assignee</label><input id="search" type="text"></div>
      <a href="/settings/performance" class="btn">Performance Settings</a>
    </div>
    <div class="shortcut-bar">
      <button id="shortcut-current-month" class="shortcut-btn" type="button">Current Month</button>
      <button id="shortcut-previous-month" class="shortcut-btn" type="button">Previous Month</button>
      <button id="shortcut-last-30-days" class="shortcut-btn" type="button">Last 30 Days</button>
      <button id="shortcut-quarter-to-date" class="shortcut-btn" type="button">Quarter To Date</button>
      <button id="shortcut-reset" class="shortcut-btn" type="button">Reset</button>
    </div>
    <div class="meta" id="capacity-profile-meta"></div>
    <div id="capacity-profile-expanded" class="capacity-expanded"></div>
    <div class="section-head collapse-toggle" id="toggle-performance-kpis">Performance Score KPIs<span class="hint">click to expand</span></div>
    <div id="performance-kpis-wrap" class="is-collapsed">
    <div class="kpis" id="performance-kpis">
      <div class="kpi"><div class="k">Team Avg Score</div><div class="v" id="kpi-avg">0</div></div>
      <div class="kpi"><div class="k">Top Performer</div><div class="v" id="kpi-top">-</div></div>
      <div class="kpi"><div class="k">At Risk (&lt;60)</div><div class="v" id="kpi-risk">0</div></div>
      <div class="kpi"><div class="k">Total Penalty</div><div class="v" id="kpi-pen">0</div></div>
      <div class="kpi"><div class="k">Total Rework Hours</div><div class="v" id="kpi-rework">0h</div></div>
    </div>
    </div>
    <div class="top3-wrap">
      <div class="top3-card">
        <div class="top3-title">Top 3 High Performing</div>
        <div id="top3-high"></div>
      </div>
      <div class="top3-card">
        <div class="top3-title">Top 3 Lowest Performing</div>
        <div id="top3-low"></div>
      </div>
    </div>
    <div class="teams-wrap is-collapsed" id="team-performance-section">
      <div class="teams-title collapse-toggle" id="toggle-team-performance">Team Performance <span class="hint">click to expand</span></div>
      <div class="team-score-layout">
        <div class="team-chart-shell">
          <div class="teams-title">Team Score Chart</div>
          <div id="team-performance-chart"></div>
        </div>
        <div class="team-detail-shell">
          <div class="teams-title">Team Performance Card</div>
          <div id="selected-team-performance" class="team-detail-body"><div class="empty">Select a team from chart.</div></div>
        </div>
      </div>
    </div>
  </section>
  <section class="arena">
    <article class="panel"><h2 id="leaderboard-title">Leaderboard</h2><div class="leader-controls"><div class="f"><label for="leader-scoring-mode">Scoring Mode</label><select id="leader-scoring-mode"><option value="simple" selected>Simple Scoring</option><option value="advanced">Advanced Scoring</option></select></div><div class="f"><label for="leader-sort">Sort By</label><select id="leader-sort"><option value="rmis">RMIs In Range</option><option value="score" selected>Performance Score</option><option value="missed">Missed Start Ratio</option><option value="capacity_gap">Capacity Gap (Cap - Planned)</option><option value="available_more_work">Available for more work</option></select></div><div class="f"><label for="leader-sort-direction">Sort Direction</label><select id="leader-sort-direction"><option value="desc" selected>Descending</option><option value="asc">Ascending</option></select></div><div class="f"><label for="filter-risk">At-Risk View</label><select id="filter-risk"><option value="all" selected>All Assignees</option><option value="risk">Only At-Risk (&lt;60)</option></select></div><div class="f"><label for="filter-missed">Start Discipline</label><select id="filter-missed"><option value="all" selected>All</option><option value="missed">Only Missed Starts</option></select></div><div class="f"><label for="leader-search">Leaderboard Search</label><input id="leader-search" type="text" placeholder="Search assignee"></div><div class="leader-actions-wrap"><div class="leader-actions-menu-wrap"><button id="leader-actions-toggle" class="leader-icon-btn" type="button" aria-label="Leaderboard actions" aria-expanded="false"><span class="material-symbols-outlined">settings</span></button><div id="leader-actions-menu" class="leader-actions-menu" hidden><button type="button" class="leader-actions-item" data-action="copy-gap-people"><span class="material-symbols-outlined">content_copy</span><span>Copy Gap People</span></button></div></div></div></div><div id="leaderboard-filter" class="sub" style="padding:0 10px 8px;"></div><div id="leaderboard-action-status" class="sub leader-action-status"></div><div id="leaderboard" class="leaderboard"></div></article>
    <article class="panel"><div class="panel-head"><h2 class="panel-head-title">Assignee Drilldown</h2><label class="panel-inline-toggle" for="assignee-overloaded-penalty-toggle"><input id="assignee-overloaded-penalty-toggle" type="checkbox"><span>Overloaded Penalty</span></label><label class="panel-inline-toggle" for="assignee-planning-realism-toggle"><input id="assignee-planning-realism-toggle" type="checkbox"><span>Overload Capping/ Planning Realism</span></label><label class="panel-inline-toggle" for="assignee-extended-actuals-toggle"><input id="assignee-extended-actuals-toggle" type="checkbox"><span>Extended Actuals</span></label></div><div id="detail" class="detail"><div class="empty">Select an assignee.</div></div></article>
  </section>
  <section class="score-arena">
    <article class="panel"><h2>Score Drilldown</h2><div id="score-drilldown" class="score-drill"><div class="empty">Select an assignee.</div></div></article>
  </section>
</div>
<script>
const payload = {data};
const worklogs = Array.isArray(payload.worklogs) ? payload.worklogs : [];
const workItems = Array.isArray(payload.work_items) ? payload.work_items : [];
const leaveRows = Array.isArray(payload.leave_rows) ? payload.leave_rows : [];
const leaveIssueKeySet = new Set((Array.isArray(payload.leave_issue_keys) ? payload.leave_issue_keys : []).map((v) => String(v || "").toUpperCase()).filter(Boolean));
const teams = Array.isArray(payload.teams) ? payload.teams : [];
const projects = Array.isArray(payload.projects) ? payload.projects : [];
const entitiesCatalog = Array.isArray(payload.entities_catalog) ? payload.entities_catalog : [];
const managedFields = Array.isArray(payload.managed_fields) ? payload.managed_fields : [];
const capacityProfiles = Array.isArray(payload.capacity_profiles) ? payload.capacity_profiles : [];
const simpleScoringRaw = Array.isArray(payload.simple_scoring) ? payload.simple_scoring : [];
const jiraBrowseBase = String(payload.jira_browse_base || "").replace(/\\/+$/, "");
const simpleScoringByKey = new Map(simpleScoringRaw.map((r) => [String(r && r.issue_key || "").toUpperCase(), r]));
const capacityProfileSelectEl = document.getElementById("capacity-profile");
const capacityProfileTopSelectEl = document.getElementById("capacity-profile-top");
const capacityProfileMetaEl = document.getElementById("capacity-profile-meta");
const capacityProfileExpandedEl = document.getElementById("capacity-profile-expanded");
const headerSectionEl = document.getElementById("performance-header");
const headerToggleButton = document.getElementById("header-toggle");
const headerExpandFabButton = document.getElementById("header-expand-fab");
const advFilterToggleButton = document.getElementById("adv-filter-toggle");
const advFilterMenu = document.getElementById("adv-filter-menu");
const dateFilterStatusNode = document.getElementById("date-filter-status");
const employeeRefreshBtn = document.getElementById("employee-refresh-btn");
const employeeRefreshCancelBtn = document.getElementById("employee-refresh-cancel-btn");
const employeeRefreshStatus = document.getElementById("employee-refresh-status");
const employeeRefreshDetails = document.getElementById("employee-refresh-details");
const leaderActionsToggle = document.getElementById("leader-actions-toggle");
const leaderActionsMenu = document.getElementById("leader-actions-menu");
const leaderboardActionStatusEl = document.getElementById("leaderboard-action-status");
const simpleOverrunModeEl = document.getElementById("simple-overrun-mode");
const assigneeExtendedActualsToggleEl = document.getElementById("assignee-extended-actuals-toggle");
const assigneeOverloadedPenaltyToggleEl = document.getElementById("assignee-overloaded-penalty-toggle");
const assigneePlanningRealismToggleEl = document.getElementById("assignee-planning-realism-toggle");
const scoreDrawerEl = document.getElementById("score-detail-drawer");
const scoreDrawerOverlayEl = document.getElementById("score-detail-drawer-overlay");
const scoreDrawerCloseEl = document.getElementById("score-detail-drawer-close");
const scoreDrawerTitleEl = document.getElementById("score-detail-drawer-title");
const scoreDrawerSubtitleEl = document.getElementById("score-detail-drawer-subtitle");
const scoreDrawerBodyEl = document.getElementById("score-detail-drawer-body");
let employeeRefreshPollHandle = null;
let employeeRefreshInlineRun = null;
let employeeRefreshInlineAssignee = "";
const HEADER_COLLAPSED_STORAGE_KEY = "employee-performance-header-collapsed";
const workItemsByKey = new Map(workItems.map((row) => [String(row && row.issue_key || "").toUpperCase(), row || {{}}]));
const defaultFrom = payload.default_from || "";
const defaultTo = payload.default_to || "";
const leaveHoursPerDay = n(payload.leave_hours_per_day) > 0 ? n(payload.leave_hours_per_day) : 8;
const settings = payload.settings || {{}};
let selectedName = "";
let selectedTeam = "";
let availabilityBreakdownForAssignee = "";
let plannedHoursBreakdownForAssignee = "";
let actualHoursBreakdownForAssignee = "";
let extendedActualsEnabled = false;
let rmiListForAssignee = "";
let dueCompletionEnabled = false;
let overloadedPenaltyEnabled = n(settings.overloaded_penalty_enabled) > 0;
let planningRealismEnabled = n(settings.planning_realism_enabled) > 0;
let overloadedPenaltyThresholdPct = clamp(n(settings.overloaded_penalty_threshold_pct), 0, 100);
let simpleOverrunMode = "subtasks";
let activeScoringTab = "simple";
let simpleFormulaGuideExpanded = true;
let advancedFormulaGuideExpanded = true;
let lastLeaderboardViewItems = [];
let leaderboardActionStatusTimer = null;
let scoreDrawerAssignee = "";
let performanceSettingsReady = false;
function n(v) {{ const x = Number(v); return Number.isFinite(x) ? x : 0; }}
function e(t) {{ return String(t ?? "").replace(/&/g, "&amp;").replace(/</g, "&lt;").replace(/>/g, "&gt;"); }}
function jiraIssueUrl(issueKey) {{
  const key = String(issueKey || "").trim().toUpperCase();
  if (!key) return "";
  if (!jiraBrowseBase) return "";
  return `${{jiraBrowseBase}}/${{encodeURIComponent(key)}}`;
}}
function scorePctText(v) {{ return `${{n(v).toFixed(1)}}%`; }}
function hoursText(v) {{ return `${{n(v).toFixed(1)}}h`; }}
function simpleOverrunLabel() {{ return simpleOverrunMode === "total" ? "Total Overrun Hours" : "Overrun Subtask Hours"; }}
function simpleOverrunShortLabel() {{ return simpleOverrunMode === "total" ? "Total Overrun" : "Overrun Subtask Hours"; }}
function actualCompletionSourceText(source) {{
  const src = String(source || "");
  if (src === "last_logged_date") return "Last Logged Date";
  if (src === "resolved_stable_since_date") return "Status Resolved Date";
  return "Not completed";
}}
function actualCompletionReason(row) {{
  const source = String(row?.actual_complete_source || "");
  if (source === "last_logged_date") return "Actual complete date came from last logged date after resolved date.";
  if (source === "resolved_stable_since_date") return "Actual complete date came from resolved date.";
  return "No completion date is available yet.";
}}
function deriveActualCompletion(plannedDueDate, lastLoggedDate, resolvedStableDate) {{
  const dueDate = String(plannedDueDate || "");
  const lastLog = String(lastLoggedDate || "");
  const resolvedDate = String(resolvedStableDate || "");
  let actualCompleteDate = "";
  let actualCompleteSource = "none";
  if (lastLog && resolvedDate) {{
    if (lastLog >= resolvedDate) {{
      actualCompleteDate = lastLog;
      actualCompleteSource = "last_logged_date";
    }} else {{
      actualCompleteDate = resolvedDate;
      actualCompleteSource = "resolved_stable_since_date";
    }}
  }} else if (lastLog) {{
    actualCompleteDate = lastLog;
    actualCompleteSource = "last_logged_date";
  }} else if (resolvedDate) {{
    actualCompleteDate = resolvedDate;
    actualCompleteSource = "resolved_stable_since_date";
  }}
  let completionBucket = "Not completed";
  if (!dueDate) completionBucket = "No due date";
  else if (!actualCompleteDate) completionBucket = "Not completed";
  else if (actualCompleteDate < dueDate) completionBucket = "Before due";
  else if (actualCompleteDate === dueDate) completionBucket = "On due";
  else completionBucket = "After due";
  return {{
    planned_due_date: dueDate,
    last_logged_date: lastLog,
    resolved_stable_since_date: resolvedDate,
    actual_complete_date: actualCompleteDate,
    actual_complete_source: actualCompleteSource,
    completion_bucket: completionBucket,
    is_penalized_for_due: !!(dueDate && actualCompleteDate && actualCompleteDate > dueDate),
  }};
}}
function openScoreDrawerForAssignee(item) {{
  if (!item || !scoreDrawerEl || !scoreDrawerOverlayEl || !scoreDrawerBodyEl) return;
  const eligibleForScore = isScoreEligible(item);
  const dueMode = dueCompletionEnabled;
  const totalEstimate = n(item.ss_total_estimate);
  const totalActual = n(item.ss_total_actual);
  const totalOverrun = n(item.ss_total_overrun);
  const commitmentForgiven = dueMode ? n(item.ss_commitment_overrun) : 0;
  const lateCompletionPenalty = dueMode ? n(item.ss_due_penalty_estimate) : 0;
  const appliedOverrun = dueMode ? Math.max(0, totalOverrun - commitmentForgiven) : totalOverrun;
  const overloadedApplied = n(item.simple_score_overloaded_applied) > 0;
  const planningRealismApplied = overloadedApplied && planningRealismEnabled;
  const baseSimpleScore = dueMode ? n(item.simple_score_due) : n(item.simple_score_raw);
  const finalSimpleScore = eligibleForScore
    ? n(item.simple_score)
    : NaN;
  const penalizedRows = (Array.isArray(item.ss_subtask_details) ? item.ss_subtask_details : [])
    .map((row) => {{
      const rawOverrun = n(row.overrun);
      const appliedOverrunHours = dueMode && n(row.is_commitment) ? 0 : rawOverrun;
      const appliedLateHours = dueMode && String(row.due_completion_status || "") === "late" ? n(row.estimate) : 0;
      const applied = appliedOverrunHours + appliedLateHours;
      return {{
        ...row,
        applied_overrun_component_hours: appliedOverrunHours,
        applied_due_component_hours: appliedLateHours,
        applied_overrun_hours: applied,
        contribution_pct: totalEstimate > 0 ? (applied / totalEstimate) * 100 : 0,
      }};
    }})
    .filter((row) => n(row.applied_overrun_hours) > 0)
    .sort((a, b) => n(b.contribution_pct) - n(a.contribution_pct) || String(a.issue_key || "").localeCompare(String(b.issue_key || "")));
  const ruleCards = [
    {{
      name: "Planned Hours Baseline",
      impact: `+${{scorePctText(100)}}`,
      tone: "pos",
      desc: `Simple score starts from the planned estimate baseline and then deducts applied ${{simpleOverrunShortLabel().toLowerCase()}}.`,
      meta: `Planned total: ${{hoursText(totalEstimate)}}`,
    }},
    {{
      name: dueMode ? `Applied ${{simpleOverrunShortLabel()}} Rule` : simpleOverrunShortLabel(),
      impact: `-${{scorePctText(totalEstimate > 0 ? (appliedOverrun / totalEstimate) * 100 : 0)}}`,
      tone: penalizedRows.length ? "neg" : "",
      desc: dueMode
        ? (simpleOverrunMode === "total"
            ? "Only positive total overrun hours reduce the score. Subtask overruns are ignored when overall total actual stays within planned hours."
            : "Only positive overrun subtask hours reduce the score. Overruns from on-time completions are forgiven in due-completion mode.")
        : (simpleOverrunMode === "total"
            ? "Only positive total overrun hours reduce the score. Subtask overruns are ignored when overall total actual stays within planned hours."
            : "Only positive overrun subtask hours reduce the score. Underruns do not offset overruns."),
      meta: dueMode
        ? `Actual: ${{hoursText(totalActual)}} | ${{simpleOverrunLabel()}}: ${{hoursText(totalOverrun)}}${{simpleOverrunMode === "total" ? "" : ` | Commitment forgiven: ${{hoursText(commitmentForgiven)}}`}} | Applied overrun: ${{hoursText(appliedOverrun)}}`
        : `Actual: ${{hoursText(totalActual)}} | ${{simpleOverrunLabel()}}: ${{hoursText(appliedOverrun)}}`,
    }},
    {{
      name: "Late Completion Rule",
      impact: dueMode ? `-${{scorePctText(totalEstimate > 0 ? (lateCompletionPenalty / totalEstimate) * 100 : 0)}}` : "Off",
      tone: dueMode && lateCompletionPenalty > 0 ? "neg" : "",
      desc: dueMode
        ? "Late-completed subtasks add their original estimate as penalty input, using the same due-completion miss data shown in assignee drilldown."
        : "Enable due-completion mode to apply late completion penalties.",
      meta: dueMode
        ? `Late items: ${{n(item.ss_due_penalty_count).toFixed(0)}} | Estimate penalty: ${{hoursText(lateCompletionPenalty)}}`
        : "Due-completion mode is off.",
    }},
    {{
      name: "Base Simple Score",
      impact: Number.isFinite(baseSimpleScore) ? scorePctText(baseSimpleScore) : "N/A",
      tone: Number.isFinite(baseSimpleScore) ? "pos" : "",
      desc: dueMode
        ? "Base score after due-completion adjustment and before overload handling."
        : "Base score after overrun deduction and before overload handling.",
      meta: totalEstimate > 0
        ? `Formula: 100 x (1 - Applied Penalty Hours / Planned) = ${{scorePctText(baseSimpleScore)}}`
        : "Planned hours are 0, so simple scoring is not eligible.",
    }},
    {{
      name: "Overloaded Penalty",
      impact: overloadedPenaltyEnabled
        ? (overloadedApplied
            ? (planningRealismApplied
                ? `Capped to ${{scorePctText(n(item.simple_score_overloaded))}}`
                : `Deducted ${{scorePctText(n(item.simple_score_overloaded_penalty_pct))}}`)
            : "Not applied")
        : "Disabled",
      tone: overloadedApplied ? "neg" : "",
      desc: planningRealismEnabled
        ? "When overload applies, Overload Capping/ Planning Realism caps the final simple score to the overload score."
        : "When overload applies, the overload gap is deducted from the base simple score like the other penalty terms.",
      meta: overloadedPenaltyEnabled
        ? `Overload Capping/ Planning Realism: ${{planningRealismEnabled ? "On" : "Off"}} | Threshold: ${{n(overloadedPenaltyThresholdPct).toFixed(1)}}% | Max planned before overload: ${{hoursText(n(item.employee_capacity_hours) * (1 + overloadedPenaltyThresholdPct / 100))}} | Capacity: ${{hoursText(item.employee_capacity_hours)}} | Capacity/Planned: ${{scorePctText(item.simple_score_overloaded)}} | Overload penalty: ${{scorePctText(n(item.simple_score_overloaded_penalty_pct))}}`
        : "Overloaded penalty is turned off.",
    }},
  ];
  const ruleHtml = ruleCards.map((rule) => `<article class="score-rule-card"><div class="score-rule-top"><div class="score-rule-name">${{e(rule.name)}}</div><span class="score-rule-impact${{rule.tone ? ` ${{rule.tone}}` : ""}}">${{e(rule.impact)}}</span></div><div class="score-rule-desc">${{e(rule.desc)}}</div><div class="score-rule-meta">${{e(rule.meta)}}</div></article>`).join("");
  const penalizedRowsHtml = penalizedRows.length
    ? penalizedRows.map((row) => {{
      const issueKey = String(row.issue_key || "").toUpperCase();
      const issueUrl = String(row.jira_url || jiraIssueUrl(issueKey) || "");
      const linkCell = issueUrl
        ? `<a class="jira-link-icon" href="${{e(issueUrl)}}" target="_blank" rel="noopener noreferrer" title="Open in Jira"><span class="material-symbols-outlined">open_in_new</span></a>`
        : `<span class="jira-link-disabled">-</span>`;
      const penaltyBreakdown = dueMode
        ? `Overrun: ${{hoursText(row.applied_overrun_component_hours)}} | Late estimate: ${{hoursText(row.applied_due_component_hours)}}`
        : `Overrun: ${{hoursText(row.applied_overrun_component_hours)}}`;
      return `<tr class="${{row.is_penalized_for_due ? "penalized-row" : ""}}"><td class="issue-id">${{e(issueKey || "-")}}</td><td class="issue-title">${{e(row.summary || "-")}}</td><td>${{e(row.project_name || "-")}}</td><td>${{e(row.epic_name || row.epic_key || "-")}}</td><td>${{e(row.planned_due_date || row.due_date || "-")}}</td><td>${{e(row.last_logged_date || "-")}}</td><td>${{e(row.actual_complete_date || row.effective_completion_date || "-")}}<div class="sub">${{e(actualCompletionSourceText(row.actual_complete_source))}}</div></td><td>${{e(row.penalty_reason || "-")}}<div class="sub">${{e(actualCompletionReason(row))}}</div></td><td>${{hoursText(row.applied_overrun_hours)}}<div class="sub">${{e(penaltyBreakdown)}}</div></td><td>${{scorePctText(row.contribution_pct)}}</td><td>${{linkCell}}</td></tr>`;
    }}).join("")
    : "";
  const penalizedSectionHtml = penalizedRows.length
    ? `<div class="score-subtask-table-wrap"><table class="ss-tbl"><thead><tr><th>Subtask</th><th>Name</th><th>Project</th><th>Epic</th><th>Planned Due Date</th><th>Last Logged Date</th><th>Actual Complete Date</th><th>Penalty Reason</th><th>Applied Overrun</th><th>Score Impact</th><th>Jira</th></tr></thead><tbody>${{penalizedRowsHtml}}</tbody></table></div>`
    : `<div class="score-drawer-empty">${{dueMode ? "No subtasks are penalizing simple score in due-completion mode." : "No subtasks are penalizing simple score in the current scope."}}</div>`;
  if (scoreDrawerTitleEl) scoreDrawerTitleEl.textContent = `Simple Score Details${{item.assignee ? ` - ${{String(item.assignee)}}` : ""}}`;
  if (scoreDrawerSubtitleEl) {{
    scoreDrawerSubtitleEl.textContent = eligibleForScore
      ? `Final Simple Score: ${{scorePctText(finalSimpleScore)}} | Mode: ${{dueMode ? "Due Completion ON" : "Due Completion OFF"}}`
      : `Simple scoring is not available because Planned Hours Assigned is ${{hoursText(item.planned_hours_assigned)}}.`;
  }}
  scoreDrawerBodyEl.innerHTML = `<section class="score-drawer-section score-drawer-section-rules"><div class="score-drawer-section-head"><h3 class="score-drawer-section-title">Rules</h3><div class="score-drawer-section-note">Exact formula inputs for this assignee</div></div><div class="score-drawer-section-content"><div class="score-rule-list">${{ruleHtml}}</div></div></section><section class="score-drawer-section score-drawer-section-penalties"><div class="score-drawer-section-head"><h3 class="score-drawer-section-title">Penalized Subtasks</h3><div class="score-drawer-section-note">${{penalizedRows.length ? `${{penalizedRows.length}} item(s) reducing the current simple score` : "No current penalties"}}</div></div><div class="score-drawer-section-content">${{penalizedSectionHtml}}</div></section>`;
  scoreDrawerAssignee = String(item.assignee || "");
  scoreDrawerOverlayEl.classList.add("open");
  scoreDrawerEl.classList.add("open");
  scoreDrawerEl.setAttribute("aria-hidden", "false");
  scoreDrawerOverlayEl.setAttribute("aria-hidden", "false");
  document.body.classList.add("score-drawer-open");
  if (scoreDrawerCloseEl) scoreDrawerCloseEl.focus();
}}
function closeScoreDrawer() {{
  if (!scoreDrawerEl || !scoreDrawerOverlayEl) return;
  scoreDrawerOverlayEl.classList.remove("open");
  scoreDrawerEl.classList.remove("open");
  scoreDrawerEl.setAttribute("aria-hidden", "true");
  scoreDrawerOverlayEl.setAttribute("aria-hidden", "true");
  document.body.classList.remove("score-drawer-open");
}}
function clamp(v, minv, maxv) {{ return Math.max(minv, Math.min(maxv, v)); }}
function applyPerformanceSettings(nextSettings) {{
  if (!nextSettings || typeof nextSettings !== "object") return;
  Object.assign(settings, nextSettings);
  overloadedPenaltyEnabled = n(settings.overloaded_penalty_enabled) > 0;
  planningRealismEnabled = n(settings.planning_realism_enabled) > 0;
  overloadedPenaltyThresholdPct = clamp(n(settings.overloaded_penalty_threshold_pct), 0, 100);
  if (assigneeOverloadedPenaltyToggleEl) {{
    assigneeOverloadedPenaltyToggleEl.checked = overloadedPenaltyEnabled;
  }}
  if (assigneePlanningRealismToggleEl) {{
    assigneePlanningRealismToggleEl.checked = planningRealismEnabled;
  }}
}}
function syncSimpleOverrunMode(nextMode) {{
  simpleOverrunMode = String(nextMode || "subtasks") === "total" ? "total" : "subtasks";
  if (simpleOverrunModeEl) simpleOverrunModeEl.value = simpleOverrunMode;
}}
async function hydratePerformanceSettings() {{
  if (typeof window === "undefined" || !window.location || window.location.protocol === "file:") return false;
  try {{
    const response = await fetch("/api/performance/settings");
    const body = await response.json().catch(() => ({{}}));
    const latestSettings = body && body.settings && typeof body.settings === "object" ? body.settings : null;
    if (!response.ok || !latestSettings) return false;
    applyPerformanceSettings(latestSettings);
    return true;
  }} catch (_error) {{
    return false;
  }}
}}
function inRange(day, from, to) {{ if (!day) return false; if (from && day < from) return false; if (to && day > to) return false; return true; }}
function setEmployeeRefreshStatus(text, tone) {{
  if (!employeeRefreshStatus) return;
  employeeRefreshStatus.textContent = String(text || "");
  if (tone === "ok") employeeRefreshStatus.style.color = "#86efac";
  else if (tone === "err") employeeRefreshStatus.style.color = "#fca5a5";
  else employeeRefreshStatus.style.color = "#9db1d8";
}}
function setEmployeeRefreshDetails(text, tone) {{
  if (!employeeRefreshDetails) return;
  employeeRefreshDetails.textContent = String(text || "");
  if (tone === "ok") employeeRefreshDetails.style.color = "#86efac";
  else if (tone === "err") employeeRefreshDetails.style.color = "#fca5a5";
  else employeeRefreshDetails.style.color = "#c7d7f3";
}}
function setEmployeeRefreshInlineState(run, fallbackAssignee, rerender = true) {{
  const normalizedRun = run && typeof run === "object" ? {{ ...run }} : null;
  const assignee = String((normalizedRun && normalizedRun.assignee) || fallbackAssignee || employeeRefreshInlineAssignee || "").trim();
  if (normalizedRun && assignee) {{
    employeeRefreshInlineRun = {{ ...normalizedRun, assignee }};
    employeeRefreshInlineAssignee = assignee;
  }} else {{
    employeeRefreshInlineRun = null;
    employeeRefreshInlineAssignee = "";
  }}
  if (rerender) render(compute());
}}
function getEmployeeRefreshInlineState(assigneeName) {{
  const run = employeeRefreshInlineRun && typeof employeeRefreshInlineRun === "object" ? employeeRefreshInlineRun : null;
  if (!run) return null;
  const rowAssignee = String(assigneeName || "").trim().toLowerCase();
  const targetAssignee = String((run.assignee || employeeRefreshInlineAssignee || "")).trim().toLowerCase();
  if (!rowAssignee || !targetAssignee || rowAssignee !== targetAssignee) return null;
  const status = String(run.status || "").toLowerCase();
  const progress = Math.max(0, Math.min(100, Math.round(Number(run.progress) || 0)));
  const tone = status === "success" ? "ok" : (status === "failed" || status === "canceled" ? "err" : "");
  return {{
    status,
    tone,
    progress,
    statusText: refreshStatusText(run),
    detailText: refreshDetailsText(run),
    isBusy: status === "running" || status === "cancel_requested",
  }};
}}
function selectedProjects() {{ return new Set(Array.from(document.getElementById("projects").selectedOptions).map(o => o.value)); }}
function addDayPenalty(rec, day, points) {{
  const d = String(day || "");
  const p = n(points);
  if (!d || p <= 0) return;
  rec.daily_penalty_by_day[d] = n(rec.daily_penalty_by_day[d]) + p;
}}
function dateRangeDays(from, to) {{
  const out = [];
  if (!from || !to || to < from) return out;
  let cur = new Date(from + "T00:00:00");
  const end = new Date(to + "T00:00:00");
  while (cur <= end) {{
    const y = cur.getFullYear();
    const m = String(cur.getMonth() + 1).padStart(2, "0");
    const d = String(cur.getDate()).padStart(2, "0");
    out.push(`${{y}}-${{m}}-${{d}}`);
    cur.setDate(cur.getDate() + 1);
  }}
  return out;
}}
function ensure(map, name) {{
  if (!map.has(name)) map.set(name, {{assignee:name, bug_hours:0, bug_late_hours:0, subtask_late_hours:0, estimate_overrun_hours:0, rework_hours:0, unplanned_leave_hours:0, planned_leave_hours:0, unplanned_leave_count:0, planned_leave_count:0, unplanned_leave_days:0, planned_leave_days:0, missing_story_due_count:0, missing_due_count:0, missing_estimate_issue_count:0, total_hours:0, planned_hours_assigned:0, base_capacity_hours:0, employee_capacity_hours:0, assigned_counts:{{epic:0,story:0,subtask:0}}, total_assigned_count:0, due_dated_assigned_count:0, missed_start_count:0, missed_start_ratio:0, missed_due_date_count:0, missed_due_date_ratio:0, active_rmi_count:0, active_rmi_keys:[], assigned_hierarchy:[], missed_start_items:[], due_compliance_items:[], start_day_activity:[], last_log_by_issue:{{}}, issue_logged_hours_by_issue:{{}}, subtask_late_by_issue:{{}}, entity_values:{{}}, managed_values:{{}}, managed_scope:{{}}, feed:[], daily_penalty_by_day:{{}}, daily_series:[], ss_total_estimate:0, ss_total_actual:0, ss_total_overrun:0, ss_commitment_overrun:0, ss_due_penalty_estimate:0, ss_due_penalty_count:0, ss_commitment_count:0, ss_within_count:0, ss_over_count:0, ss_no_estimate_count:0, ss_on_time_count:0, ss_late_count:0, ss_subtask_details:[], simple_score_overrun_total:0, simple_score_overrun_active:0, simple_score_overloaded:100, simple_score_overloaded_penalty_pct:0, simple_score_overloaded_applied:0}});
  return map.get(name);
}}
function normalizeType(t) {{
  const low = String(t || "").toLowerCase();
  if (low.includes("epic")) return "epic";
  if (low.includes("story")) return "story";
  return "subtask";
}}
function normalizeHierarchyType(t) {{
  const low = String(t || "").toLowerCase();
  if (low.includes("epic")) return "epic";
  if (low.includes("story")) return "story";
  if (low.includes("bug") && (low.includes("subtask") || low.includes("sub-task") || low.includes("task"))) return "bug_subtask";
  return "subtask";
}}
function issueTypeLabel(t) {{
  return String(t || "").trim().toLowerCase();
}}
function _isoDateLocal(dt) {{
  const y = dt.getFullYear();
  const m = String(dt.getMonth() + 1).padStart(2, "0");
  const d = String(dt.getDate()).padStart(2, "0");
  return `${{y}}-${{m}}-${{d}}`;
}}
function capacityProfileLabel(profile) {{
  const fromDate = String(profile?.from_date || "");
  const toDate = String(profile?.to_date || "");
  const std = n(profile?.standard_hours_per_day) > 0 ? n(profile.standard_hours_per_day) : 8;
  const ramadan = n(profile?.ramadan_hours_per_day) > 0 ? n(profile.ramadan_hours_per_day) : std;
  return `${{fromDate}} -> ${{toDate}} | Std ${{std.toFixed(1)}}h | Ramadan ${{ramadan.toFixed(1)}}h`;
}}
function getActiveCapacityProfileSelection() {{
  if (capacityProfileTopSelectEl) return String(capacityProfileTopSelectEl.value || "auto");
  if (capacityProfileSelectEl) return String(capacityProfileSelectEl.value || "auto");
  return "auto";
}}
function syncCapacityProfileSelection(selectedValue, source) {{
  const normalized = String(selectedValue || "auto");
  if (capacityProfileSelectEl && source !== "header") capacityProfileSelectEl.value = normalized;
  if (capacityProfileTopSelectEl && source !== "top") capacityProfileTopSelectEl.value = normalized;
}}
function refreshCapacityProfileOptions() {{
  if (!capacityProfileSelectEl && !capacityProfileTopSelectEl) return;
  const options = [`<option value="auto">Auto (Match selected date range)</option>`];
  capacityProfiles.forEach((profile, idx) => {{
    options.push(`<option value="${{idx}}">${{e(capacityProfileLabel(profile))}}</option>`);
  }});
  if (capacityProfileSelectEl) {{
    capacityProfileSelectEl.innerHTML = options.join("");
    capacityProfileSelectEl.value = "auto";
  }}
  if (capacityProfileTopSelectEl) {{
    capacityProfileTopSelectEl.innerHTML = options.join("");
    capacityProfileTopSelectEl.value = "auto";
  }}
}}
function resolveActiveCapacityProfile(fromIso, toIso) {{
  const selected = getActiveCapacityProfileSelection();
  if (selected === "auto") {{
    return capacityProfiles.find((p) => String(p.from_date || "") === String(fromIso || "") && String(p.to_date || "") === String(toIso || "")) || null;
  }}
  const idx = Number(selected);
  if (Number.isInteger(idx) && idx >= 0 && idx < capacityProfiles.length) return capacityProfiles[idx];
  return null;
}}
function updateCapacityProfileMeta(fromIso, toIso, profile) {{
  if (!capacityProfileMetaEl) return;
  const selected = getActiveCapacityProfileSelection();
  if (!profile) {{
    capacityProfileMetaEl.textContent = "Capacity profile: Default weekdays (8h/day)";
  }} else {{
    const mode = selected === "auto" ? "Auto profile" : "Applied profile";
    capacityProfileMetaEl.textContent = `${{mode}}: ${{capacityProfileLabel(profile)}} | Active range: ${{String(fromIso || "")}} -> ${{String(toIso || "")}}`;
  }}
  renderCapacityProfileExpanded(fromIso, toIso, profile || null);
}}
function renderCapacityProfileExpanded(fromIso, toIso, profile) {{
  if (!capacityProfileExpandedEl) return;
  const from = String(fromIso || "");
  const to = String(toIso || "");
  if (!from || !to || to < from) {{
    capacityProfileExpandedEl.innerHTML = '<div class="capacity-empty">Select a valid date range to view full capacity settings and calendar.</div>';
    return;
  }}
  const selected = getActiveCapacityProfileSelection();
  const mode = selected === "auto" ? "Auto" : "Manual";
  const std = n(profile?.standard_hours_per_day) > 0 ? n(profile.standard_hours_per_day) : 8;
  const ram = n(profile?.ramadan_hours_per_day) > 0 ? n(profile.ramadan_hours_per_day) : std;
  const ramadanStart = String(profile?.ramadan_start_date || "");
  const ramadanEnd = String(profile?.ramadan_end_date || "");
  const holidays = Array.isArray(profile?.holiday_dates) ? profile.holiday_dates.map((d) => String(d || "")).filter(Boolean).sort() : [];
  const holidaySet = new Set(holidays);
  const employeeCount = Math.max(0, Math.round(n(profile?.employee_count)));
  const businessDays = computeBusinessDays(from, to, profile || null);
  const perAssigneeHours = computePerAssigneeCapacity(from, to, profile || null);
  const leaveByDay = new Map();
  for (const row of leaveRows) {{
    const day = String(row && row.period_day || "");
    if (!inRange(day, from, to)) continue;
    const planned = n(row && row.planned_taken_hours);
    const unplanned = n(row && row.unplanned_taken_hours);
    if (planned <= 0 && unplanned <= 0) continue;
    const rec = leaveByDay.get(day) || {{ planned: 0, unplanned: 0, count: 0 }};
    rec.planned += planned;
    rec.unplanned += unplanned;
    rec.count += 1;
    leaveByDay.set(day, rec);
  }}
  let totalPlannedLeave = 0;
  let totalUnplannedLeave = 0;
  leaveByDay.forEach((v) => {{
    totalPlannedLeave += n(v.planned);
    totalUnplannedLeave += n(v.unplanned);
  }});
  const monthCards = [];
  const cursor = new Date(from + "T00:00:00");
  const end = new Date(to + "T00:00:00");
  const todayIso = _isoDateLocal(new Date());
  if (!Number.isFinite(cursor.getTime()) || !Number.isFinite(end.getTime())) {{
    capacityProfileExpandedEl.innerHTML = '<div class="capacity-empty">Date range is invalid for calendar rendering.</div>';
    return;
  }}
  cursor.setDate(1);
  while (cursor <= end) {{
    const mStart = new Date(cursor.getFullYear(), cursor.getMonth(), 1);
    const mEnd = new Date(cursor.getFullYear(), cursor.getMonth() + 1, 0);
    const firstDow = (mStart.getDay() + 6) % 7;
    const lastDow = (mEnd.getDay() + 6) % 7;
    const gridStart = new Date(mStart);
    gridStart.setDate(gridStart.getDate() - firstDow);
    const gridEnd = new Date(mEnd);
    gridEnd.setDate(gridEnd.getDate() + (6 - lastDow));
    const isCurrentMonth = mStart.getFullYear() === new Date().getFullYear() && mStart.getMonth() === new Date().getMonth();
    const cells = ['<div class="capacity-dow">Mon</div><div class="capacity-dow">Tue</div><div class="capacity-dow">Wed</div><div class="capacity-dow">Thu</div><div class="capacity-dow">Fri</div><div class="capacity-dow">Sat</div><div class="capacity-dow">Sun</div>'];
    for (let d = new Date(gridStart); d <= gridEnd; d.setDate(d.getDate() + 1)) {{
      const iso = _isoDateLocal(d);
      const isOut = d.getMonth() !== mStart.getMonth();
      const isWeekend = d.getDay() === 0 || d.getDay() === 6;
      const isRamadan = ramadanStart && ramadanEnd && iso >= ramadanStart && iso <= ramadanEnd;
      const isHoliday = holidaySet.has(iso);
      const isToday = iso === todayIso;
      const leave = leaveByDay.get(iso) || null;
      const classes = ["capacity-day"];
      if (isOut) classes.push("is-out");
      if (isWeekend) classes.push("is-weekend");
      if (isRamadan) classes.push("is-ramadan");
      if (isHoliday) classes.push("is-holiday");
      if (isToday) classes.push("is-today");
      if (leave && isRamadan) classes.push("has-ramadan-leave");
      else if (leave) classes.push("has-leave");
      const tags = [];
      if (isRamadan) tags.push('<span class="capacity-day-tag r">R</span>');
      if (isHoliday) tags.push('<span class="capacity-day-tag h">H</span>');
      if (leave) {{
        const rawLeaveHours = n(leave.planned) + n(leave.unplanned);
        const effectiveLeaveHours = isRamadan ? ram : std;
        const leaveHours = (rawLeaveHours > 0 ? effectiveLeaveHours : 0).toFixed(1);
        tags.push(isRamadan
          ? `<span class="capacity-day-tag rl">RL ${{leaveHours}}h</span>`
          : `<span class="capacity-day-tag l">L ${{leaveHours}}h</span>`);
      }}
      cells.push(`<div class="${{classes.join(" ")}}"><div class="capacity-day-num">${{String(d.getDate())}}</div><div class="capacity-day-tags">${{tags.join("")}}</div></div>`);
    }}
    const label = mStart.toLocaleString(undefined, {{ month: "long", year: "numeric" }});
    monthCards.push(`<div class="capacity-month"${{isCurrentMonth ? ' data-current-month="1"' : ""}}><div class="capacity-month-head">${{e(label)}}</div><div class="capacity-month-grid">${{cells.join("")}}</div></div>`);
    cursor.setMonth(cursor.getMonth() + 1, 1);
  }}
  const holidaysText = holidays.length ? e(holidays.join(", ")) : "None";
  const ramadanText = ramadanStart && ramadanEnd ? `${{e(ramadanStart)}} -> ${{e(ramadanEnd)}}` : "Not configured";
  const scopeLabel = profile ? `${{e(String(profile.from_date || ""))}} -> ${{e(String(profile.to_date || ""))}}` : "Default weekdays profile";
  capacityProfileExpandedEl.innerHTML = `
    <div class="capacity-expanded-head">
      <div class="capacity-expanded-title">Capacity Profile Expanded Settings</div>
      <div class="capacity-expanded-sub">${{mode}} selection | Active filter: ${{e(from)}} -> ${{e(to)}}</div>
    </div>
    <div class="capacity-expanded-grid">
      <div class="capacity-chip"><div class="k">Profile Range</div><div class="v">${{scopeLabel}}</div></div>
      <div class="capacity-chip"><div class="k">Employee Count</div><div class="v">${{employeeCount}}</div></div>
      <div class="capacity-chip"><div class="k">Standard Hours/Day</div><div class="v">${{std.toFixed(2)}}h</div></div>
      <div class="capacity-chip"><div class="k">Ramadan Hours/Day</div><div class="v">${{ram.toFixed(2)}}h</div></div>
      <div class="capacity-chip"><div class="k">Ramadan Range</div><div class="v">${{ramadanText}}</div></div>
      <div class="capacity-chip"><div class="k">Holidays</div><div class="v">${{holidays.length}}</div></div>
      <div class="capacity-chip"><div class="k">Business Days</div><div class="v">${{businessDays.toFixed(0)}}d</div></div>
      <div class="capacity-chip"><div class="k">Per Assignee Capacity</div><div class="v">${{perAssigneeHours.toFixed(2)}}h</div></div>
      <div class="capacity-chip"><div class="k">Planned Leave (Range)</div><div class="v">${{totalPlannedLeave.toFixed(2)}}h</div></div>
      <div class="capacity-chip"><div class="k">Unplanned Leave (Range)</div><div class="v">${{totalUnplannedLeave.toFixed(2)}}h</div></div>
    </div>
    <div class="capacity-legend">
      <span class="pill">R = Ramadan</span>
      <span class="pill">H = Holiday</span>
      <span class="pill">L = Leave hours</span>
      <span class="pill">RL = Ramadan leave hours</span>
    </div>
    <div class="capacity-expanded-sub" style="margin-bottom:8px;">Holiday dates: ${{holidaysText}}</div>
    <div class="capacity-calendar-wrap">${{monthCards.join("")}}</div>
  `;
  const calWrap = capacityProfileExpandedEl.querySelector(".capacity-calendar-wrap");
  const currentMonthCard = calWrap ? calWrap.querySelector('[data-current-month="1"]') : null;
  if (calWrap && currentMonthCard) {{
    const target = Math.max(0, currentMonthCard.offsetLeft - 8);
    calWrap.scrollLeft = target;
  }}
}}
function computePerAssigneeCapacity(fromIso, toIso, profile) {{
  const fromDate = String(fromIso || "");
  const toDate = String(toIso || "");
  if (!fromDate || !toDate || toDate < fromDate) return 0;
  const start = new Date(fromDate + "T00:00:00");
  const end = new Date(toDate + "T00:00:00");
  if (!Number.isFinite(start.getTime()) || !Number.isFinite(end.getTime())) return 0;
  const standardHours = n(profile?.standard_hours_per_day) > 0 ? n(profile.standard_hours_per_day) : 8;
  const ramadanHours = n(profile?.ramadan_hours_per_day) > 0 ? n(profile.ramadan_hours_per_day) : standardHours;
  const ramadanStart = String(profile?.ramadan_start_date || "");
  const ramadanEnd = String(profile?.ramadan_end_date || "");
  const holidaySet = new Set((Array.isArray(profile?.holiday_dates) ? profile.holiday_dates : []).map((day) => String(day || "")));
  let total = 0;
  for (let d = new Date(start); d <= end; d.setDate(d.getDate() + 1)) {{
    const dow = d.getDay();
    if (dow === 0 || dow === 6) continue;
    const dayIso = _isoDateLocal(d);
    if (holidaySet.has(dayIso)) continue;
    const inRamadan = ramadanStart && ramadanEnd && dayIso >= ramadanStart && dayIso <= ramadanEnd;
    total += inRamadan ? ramadanHours : standardHours;
  }}
  return total;
}}
function computeBusinessDays(fromIso, toIso, profile) {{
  const fromDate = String(fromIso || "");
  const toDate = String(toIso || "");
  if (!fromDate || !toDate || toDate < fromDate) return 0;
  const start = new Date(fromDate + "T00:00:00");
  const end = new Date(toDate + "T00:00:00");
  if (!Number.isFinite(start.getTime()) || !Number.isFinite(end.getTime())) return 0;
  const holidaySet = new Set((Array.isArray(profile?.holiday_dates) ? profile.holiday_dates : []).map((day) => String(day || "")));
  let days = 0;
  for (let d = new Date(start); d <= end; d.setDate(d.getDate() + 1)) {{
    const dow = d.getDay();
    if (dow === 0 || dow === 6) continue;
    const dayIso = _isoDateLocal(d);
    if (holidaySet.has(dayIso)) continue;
    days += 1;
  }}
  return days;
}}
function resolveParentEpicKey(wi) {{
  const parentKey = String(wi && wi.parent_issue_key || "").toUpperCase();
  if (!parentKey) return "";
  const parentRow = workItemsByKey.get(parentKey) || null;
  if (!parentRow) return "";
  const parentType = issueTypeLabel(parentRow.jira_issue_type || parentRow.issue_type || parentRow.work_item_type);
  if (parentType.includes("epic")) return parentKey;
  return String(parentRow.parent_issue_key || "").toUpperCase();
}}
function resolveEpicKeyFromWorklogRow(row) {{
  const issueKey = String(row && row.issue_id || "").toUpperCase();
  const issueRow = workItemsByKey.get(issueKey) || null;
  if (issueRow) {{
    return resolveParentEpicKey(issueRow);
  }}
  const parentKey = String(row && row.item_parent_issue_key || "").toUpperCase();
  if (!parentKey) return "";
  const parentRow = workItemsByKey.get(parentKey) || null;
  if (!parentRow) return "";
  const parentType = issueTypeLabel(parentRow.jira_issue_type || parentRow.issue_type || parentRow.work_item_type);
  if (parentType.includes("epic")) return parentKey;
  return String(parentRow.parent_issue_key || "").toUpperCase();
}}
function setDateFilterRange(from, to) {{
  document.getElementById("from").value = String(from || "");
  document.getElementById("to").value = String(to || "");
}}
function setDateFilterStatus(text) {{
  if (!dateFilterStatusNode) return;
  dateFilterStatusNode.textContent = String(text || "");
}}
function setAdvancedFilterMenuOpen(open) {{
  if (!advFilterToggleButton || !advFilterMenu) return;
  const isOpen = Boolean(open);
  advFilterToggleButton.setAttribute("aria-expanded", isOpen ? "true" : "false");
  if (isOpen) {{
    advFilterMenu.removeAttribute("hidden");
  }} else {{
    advFilterMenu.setAttribute("hidden", "");
  }}
}}
function applyDateShortcut(kind) {{
  const now = new Date();
  const today = _isoDateLocal(now);
  if (kind === "current_month") {{
    const first = new Date(now.getFullYear(), now.getMonth(), 1);
    const last = new Date(now.getFullYear(), now.getMonth() + 1, 0);
    setDateFilterRange(_isoDateLocal(first), _isoDateLocal(last));
    return;
  }}
  if (kind === "previous_month") {{
    const first = new Date(now.getFullYear(), now.getMonth() - 1, 1);
    const last = new Date(now.getFullYear(), now.getMonth(), 0);
    setDateFilterRange(_isoDateLocal(first), _isoDateLocal(last));
    return;
  }}
  if (kind === "last_30_days") {{
    const start = new Date(now.getFullYear(), now.getMonth(), now.getDate() - 29);
    setDateFilterRange(_isoDateLocal(start), today);
    return;
  }}
  if (kind === "last_90_days") {{
    const start = new Date(now.getFullYear(), now.getMonth(), now.getDate() - 89);
    setDateFilterRange(_isoDateLocal(start), today);
    return;
  }}
  if (kind === "last_quarter") {{
    const qStartMonth = Math.floor(now.getMonth() / 3) * 3;
    const start = new Date(now.getFullYear(), qStartMonth - 3, 1);
    const end = new Date(now.getFullYear(), qStartMonth, 0);
    setDateFilterRange(_isoDateLocal(start), _isoDateLocal(end));
    return;
  }}
  if (kind === "current_quarter") {{
    const qStartMonth = Math.floor(now.getMonth() / 3) * 3;
    const start = new Date(now.getFullYear(), qStartMonth, 1);
    const end = new Date(now.getFullYear(), qStartMonth + 3, 0);
    setDateFilterRange(_isoDateLocal(start), _isoDateLocal(end));
    return;
  }}
  if (kind === "quarter_to_date") {{
    const qStartMonth = Math.floor(now.getMonth() / 3) * 3;
    const qStart = new Date(now.getFullYear(), qStartMonth, 1);
    setDateFilterRange(_isoDateLocal(qStart), today);
    return;
  }}
  setDateFilterRange(defaultFrom, defaultTo);
}}
function isSubtaskPerformanceType(t) {{
  const label = issueTypeLabel(t);
  if (!label) return false;
  if (label.includes("sub-task") || label.includes("subtask")) return true;
  return label.includes("bug") && (label.includes("sub-task") || label.includes("subtask"));
}}
function isBugIssueType(t) {{
  const label = issueTypeLabel(t);
  return !!label && label.includes("bug");
}}
function isLeaveIssueKey(issueKey) {{
  const key = String(issueKey || "").toUpperCase();
  return !!key && leaveIssueKeySet.has(key);
}}
function buildCapacityMap(from, to, assigneeNames, activeProfile) {{
  const out = new Map();
  const names = Array.from(assigneeNames || []);
  const fromIso = String(from || "");
  const toIso = String(to || "");
  const perAssignee = computePerAssigneeCapacity(fromIso, toIso, activeProfile || null);
  names.forEach((name) => out.set(String(name || ""), perAssignee));
  return out;
}}
function evalFormula(expr, scope) {{
  const text = String(expr || "").trim();
  if (!text) return 0;
  const src = text
    .replace(/\\baverage\\(/gi, "AVG(")
    .replace(/\\bsum\\(/gi, "SUM(")
    .replace(/\\bcount\\(/gi, "COUNT(")
    .replace(/\\bmin\\(/gi, "MIN(")
    .replace(/\\bmax\\(/gi, "MAX(");
  const api = {{
    SUM:(v)=>n(v),
    COUNT:(v)=>n(v)>0?1:0,
    MIN:(v)=>n(v),
    MAX:(v)=>n(v),
    AVG:(v)=>n(v)
  }};
  const replaced = src.replace(/[A-Za-z_][A-Za-z0-9_]*/g, (token) => {{
    if (token in api) return token;
    return String(n(scope[token.toLowerCase()]));
  }});
  if (!/^[0-9+\\-*/().,\\sA-Z]+$/.test(replaced)) return 0;
  try {{
    const fn = Function("SUM","COUNT","MIN","MAX","AVG", `return (${{replaced}});`);
    return n(fn(api.SUM, api.COUNT, api.MIN, api.MAX, api.AVG));
  }} catch (_err) {{
    return 0;
  }}
}}
function compute() {{
  const from = document.getElementById("from").value || defaultFrom;
  const to = document.getElementById("to").value || defaultTo;
  const activeCapacityProfile = resolveActiveCapacityProfile(from, to);
  updateCapacityProfileMeta(from, to, activeCapacityProfile);
  const s = String(document.getElementById("search").value || "").trim().toLowerCase();
  const pset = selectedProjects();
  const useP = pset.size > 0;
  const scopedWorklogs = worklogs.filter((r) => {{
    if (useP && !pset.has(String(r.project_key || "UNKNOWN"))) return false;
    if (s && !String(r.issue_assignee || "").toLowerCase().includes(s)) return false;
    const issueType = String(r.item_issue_type || r.issue_type || "");
    return isSubtaskPerformanceType(issueType);
  }});
  const logs = scopedWorklogs.filter((r) => inRange(String(r.worklog_date || ""), from, to));
  const allLoggedHoursByAssigneeIssue = new Map();
  for (const wl of scopedWorklogs) {{
    const assignee = String(wl.issue_assignee || "Unassigned");
    const issueKey = String(wl.issue_id || "").toUpperCase();
    if (!issueKey) continue;
    const compound = `${{assignee}}\\u0000${{issueKey}}`;
    allLoggedHoursByAssigneeIssue.set(compound, n(allLoggedHoursByAssigneeIssue.get(compound)) + n(wl.hours_logged));
  }}
  const assignedItems = workItems.filter((r) => {{
    const assignee = String(r.assignee || "");
    if (!assignee) return false;
    const project = String(r.project_key || "UNKNOWN");
    if (useP && !pset.has(project)) return false;
    if (s && !assignee.toLowerCase().includes(s)) return false;
    const issueType = String(r.jira_issue_type || r.issue_type || r.work_item_type || "");
    if (!isSubtaskPerformanceType(issueType)) return false;
    const start = String(r.start_date || "");
    const due = String(r.due_date || "");
    if (!start && !due) return false;
    return inRange(start, from, to) || inRange(due, from, to);
  }});
  const assignedItemsWork = assignedItems.filter((r) => !isLeaveIssueKey(String(r.issue_key || "")));
  const leaves = leaveRows.filter((r) => inRange(String(r.period_day || ""), from, to));
  const byA = new Map();
  const epicKeysByAssignee = new Map();
  function addAssigneeEpic(name, epicKey) {{
    const assignee = String(name || "Unassigned");
    const epic = String(epicKey || "").toUpperCase();
    if (!epic) return;
    if (!epicKeysByAssignee.has(assignee)) epicKeysByAssignee.set(assignee, new Set());
    epicKeysByAssignee.get(assignee).add(epic);
  }}
  const issueAgg = new Map();
  for (const wi of assignedItemsWork) ensure(byA, String(wi.assignee || "Unassigned"));
  for (const l of logs) {{
    const a = ensure(byA, String(l.issue_assignee || "Unassigned"));
    addAssigneeEpic(a.assignee, resolveEpicKeyFromWorklogRow(l));
    const hrs = n(l.hours_logged);
    a.total_hours += hrs;
    const issueKey = String(l.issue_id || "");
    const lastLogged = String(a.last_log_by_issue[issueKey] || "");
    if (!lastLogged || String(l.worklog_date || "") > lastLogged) a.last_log_by_issue[issueKey] = String(l.worklog_date || "");
    a.issue_logged_hours_by_issue[issueKey] = n(a.issue_logged_hours_by_issue[issueKey]) + hrs;
    const issue = String(l.issue_id || "");
    if (!issueAgg.has(issue)) issueAgg.set(issue, {{ total:0, estimate:n(l.original_estimate_hours), shares:new Map(), dayShares:new Map() }});
    const agg = issueAgg.get(issue);
    agg.total += hrs;
    agg.shares.set(a.assignee, n(agg.shares.get(a.assignee)) + hrs);
    if (!agg.dayShares.has(a.assignee)) agg.dayShares.set(a.assignee, new Map());
    const byDay = agg.dayShares.get(a.assignee);
    const dayKey = String(l.worklog_date || "");
    byDay.set(dayKey, n(byDay.get(dayKey)) + hrs);
    if (l.is_bug) {{
      a.bug_hours += hrs;
      addDayPenalty(a, dayKey, hrs * n(settings.points_per_bug_hour));
      const due = String(l.story_due_date || "");
      if (due && String(l.worklog_date || "") > due) {{
        a.bug_late_hours += hrs;
        addDayPenalty(a, dayKey, hrs * n(settings.points_per_bug_late_hour));
      }}
      else if (!due) a.missing_story_due_count += 1;
    }} else {{
      const due = String(l.item_due_date || "");
      if (due && String(l.worklog_date || "") > due) {{
        a.subtask_late_hours += hrs;
        a.subtask_late_by_issue[issueKey] = n(a.subtask_late_by_issue[issueKey]) + hrs;
        addDayPenalty(a, dayKey, hrs * n(settings.points_per_subtask_late_hour));
      }}
      else if (!due) a.missing_due_count += 1;
    }}
    if (String(l.fix_type || "").toLowerCase() === "rework") {{
      a.rework_hours += hrs;
    }}
  }}
  for (const agg of issueAgg.values()) {{
    if (n(agg.estimate) <= 0) {{
      for (const name of agg.shares.keys()) ensure(byA, name).missing_estimate_issue_count += 1;
      continue;
    }}
    const overrun = Math.max(0, n(agg.total) - n(agg.estimate));
    if (overrun <= 0 || n(agg.total) <= 0) continue;
    for (const [name, hrs] of agg.shares.entries()) {{
      const rec = ensure(byA, name);
      const assigneeOverrun = overrun * (n(hrs) / n(agg.total));
      rec.estimate_overrun_hours += assigneeOverrun;
      const dayMap = agg.dayShares.get(name) || new Map();
      if (n(hrs) > 0) {{
        for (const [day, dayHours] of dayMap.entries()) {{
          const dayShare = n(dayHours) / n(hrs);
          const dayOverrun = assigneeOverrun * dayShare;
          addDayPenalty(rec, day, dayOverrun * n(settings.points_per_estimate_overrun_hour));
        }}
      }}
    }}
  }}
  for (const r of leaves) {{
    const a = ensure(byA, String(r.assignee || "Unassigned"));
    const unplannedHours = n(r.unplanned_taken_hours);
    const plannedHours = n(r.planned_taken_hours);
    a.unplanned_leave_hours += unplannedHours;
    a.planned_leave_hours += plannedHours;
    if (unplannedHours > 0) {{
      a.unplanned_leave_count += 1;
      a.unplanned_leave_days += unplannedHours / leaveHoursPerDay;
    }}
    if (plannedHours > 0) {{
      a.planned_leave_count += 1;
      a.planned_leave_days += plannedHours / leaveHoursPerDay;
    }}
    addDayPenalty(a, String(r.period_day || ""), unplannedHours * n(settings.points_per_unplanned_leave_hour));
  }}
  const assigneeNames = new Set(Array.from(byA.keys()));
  const capacityByAssignee = buildCapacityMap(from, to, assigneeNames, activeCapacityProfile);
  for (const wi of assignedItemsWork) {{
    const a = ensure(byA, String(wi.assignee || "Unassigned"));
    addAssigneeEpic(a.assignee, resolveParentEpicKey(wi));
    const issueKey = String(wi.issue_key || "");
    const startDate = String(wi.start_date || "");
    const dueDate = String(wi.due_date || "");
    const rawIssueType = wi.issue_type || wi.work_item_type || wi.jira_issue_type;
    const issueType = normalizeType(rawIssueType);
    const hierarchyIssueType = normalizeHierarchyType(rawIssueType);
    const originalEstimateHours = n(wi.original_estimate_hours);
    const includePlannedHours = !isBugIssueType(rawIssueType) && originalEstimateHours > 0;
    if (includePlannedHours) a.planned_hours_assigned += originalEstimateHours;
    a.assigned_counts[issueType] = n(a.assigned_counts[issueType]) + 1;
    a.total_assigned_count += 1;
    if (dueDate) a.due_dated_assigned_count += 1;
    const lastLogDate = String(a.last_log_by_issue[issueKey] || "");
    const completionMeta = deriveActualCompletion(dueDate, lastLogDate, String(wi.resolved_stable_since_date || ""));
    a.assigned_hierarchy.push({{
      issue_key: issueKey,
      summary: String(wi.summary || ""),
      issue_type: issueType,
      hierarchy_type: hierarchyIssueType,
      parent_issue_key: String(wi.parent_issue_key || ""),
      parent_epic_key: resolveParentEpicKey(wi),
      planned_due_date: dueDate,
      due_date: dueDate,
      start_date: startDate,
      original_estimate_hours: n(wi.original_estimate_hours),
      actual_hours: n(a.issue_logged_hours_by_issue[issueKey]),
      negative_hours: n(a.subtask_late_by_issue[issueKey]),
      last_logged_date: completionMeta.last_logged_date,
      actual_complete_date: completionMeta.actual_complete_date,
      actual_complete_source: completionMeta.actual_complete_source,
      completion_date: lastLogDate,
      resolved_stable_since_date: String(wi.resolved_stable_since_date || ""),
      is_penalized_for_due: completionMeta.is_penalized_for_due,
      status: String(wi.status || "")
    }});
    let missedDueDate = false;
    let dueStatus = completionMeta.completion_bucket;
    if (completionMeta.actual_complete_date && dueDate && completionMeta.actual_complete_date > dueDate) {{
      missedDueDate = true;
    }} else if (!completionMeta.actual_complete_date && dueDate && to >= dueDate) {{
      missedDueDate = true;
    }}
    if (missedDueDate) {{
      a.missed_due_date_count += 1;
      addDayPenalty(a, dueDate || completionMeta.actual_complete_date || to, n(settings.points_per_missed_due_date));
    }}
    a.due_compliance_items.push({{
      issue_key: issueKey,
      summary: String(wi.summary || ""),
      planned_due_date: dueDate,
      due_date: dueDate,
      last_logged_date: completionMeta.last_logged_date,
      actual_complete_date: completionMeta.actual_complete_date,
      actual_complete_source: completionMeta.actual_complete_source,
      completion_date: lastLogDate,
      resolved_stable_since_date: String(wi.resolved_stable_since_date || ""),
      status_bucket: dueStatus,
      is_missed_due_date: missedDueDate
    }});
    const ssRow = simpleScoringByKey.get(issueKey.toUpperCase());
    if (ssRow) {{
      const epicKey = resolveParentEpicKey(wi);
      const epicRow = workItemsByKey.get(String(epicKey || "").toUpperCase()) || {{}};
      const ssEst = n(ssRow.original_estimate_hours);
      const ssAct = n(ssRow.actual_hours_logged);
      const ssOver = n(ssRow.overrun_hours);
      const ssEstStatus = String(ssRow.estimate_status || "");
      const ssDueStatus = String(ssRow.due_completion_status || "");
      const ssCommit = n(ssRow.is_commitment);
      const penaltyReason = ssEstStatus === "no_estimate"
        ? "No estimate; this subtask does not affect simple score."
        : (ssCommit
          ? "Over estimate but completed on time, so due-completion mode forgives its overrun."
          : (ssDueStatus === "late" && ssEstStatus === "over_estimate"
            ? "Late completion: original estimate is penalized in due-completion mode, and its overrun also reduces score."
            : (ssDueStatus === "late"
              ? "Late completion: original estimate is penalized in due-completion mode."
              : (ssEstStatus === "over_estimate"
                ? "Over estimate; its overrun reduces simple score."
                : "Within estimate; no simple-score penalty."))));
      const issueUrl = jiraIssueUrl(issueKey);
      a.ss_total_estimate += ssEst;
      a.ss_total_actual += ssAct;
      a.ss_total_overrun += ssOver;
      if (ssCommit) {{ a.ss_commitment_overrun += ssOver; a.ss_commitment_count += 1; }}
      if (ssDueStatus === "late" && ssEst > 0) {{
        a.ss_due_penalty_estimate += ssEst;
        a.ss_due_penalty_count += 1;
      }}
      if (ssEstStatus === "within_estimate") a.ss_within_count += 1;
      else if (ssEstStatus === "over_estimate") a.ss_over_count += 1;
      else a.ss_no_estimate_count += 1;
      if (ssDueStatus === "on_time") a.ss_on_time_count += 1;
      else if (ssDueStatus === "late") a.ss_late_count += 1;
      a.ss_subtask_details.push({{
        issue_key: issueKey, summary: String(wi.summary || ""), estimate: ssEst, actual: ssAct,
        overrun: ssOver, estimate_status: ssEstStatus, planned_due_date: dueDate, due_date: dueDate,
        last_logged_date: completionMeta.last_logged_date,
        effective_completion_date: String(ssRow.effective_completion_date || ""),
        actual_complete_date: String(ssRow.actual_complete_date || completionMeta.actual_complete_date || ""),
        actual_complete_source: String(ssRow.actual_complete_source || completionMeta.actual_complete_source || "none"),
        due_completion_status: ssDueStatus, is_commitment: ssCommit, status: String(ssRow.status || ""),
        is_penalized_for_due: ssDueStatus === "late",
        project_key: String(wi.project_key || ""),
        project_name: String(wi.project_name || wi.project_key || ""),
        epic_key: String(epicKey || ""),
        epic_name: String(epicRow.summary || epicRow.item_summary || epicKey || ""),
        jira_url: issueUrl,
        penalty_reason: penaltyReason
      }});
    }}
    if (startDate && lastLogDate !== startDate) {{
      const loggedOnStart = logs.some((x) => String(x.issue_assignee || "") === a.assignee && String(x.issue_id || "") === issueKey && String(x.worklog_date || "") === startDate);
      if (!loggedOnStart) {{
        a.missed_start_count += 1;
        const otherWork = logs
          .filter((x) => String(x.issue_assignee || "") === a.assignee && String(x.worklog_date || "") === startDate && String(x.issue_id || "") !== issueKey)
          .map((x) => ({{issue_id:String(x.issue_id || ""), summary:String(x.item_summary || ""), hours:n(x.hours_logged)}}));
        const leaveOnDay = leaves
          .filter((x) => String(x.assignee || "") === a.assignee && String(x.period_day || "") === startDate)
          .reduce((acc, x) => acc + n(x.planned_taken_hours) + n(x.unplanned_taken_hours), 0);
        const idle = !otherWork.length && leaveOnDay <= 0;
        const entry = {{
          issue_key: issueKey,
          summary: String(wi.summary || ""),
          start_date: startDate,
          other_work: otherWork,
          leave_hours: leaveOnDay,
          idle: idle
        }};
        a.missed_start_items.push(entry);
        a.start_day_activity.push(entry);
      }}
    }}
  }}
  for (const it of byA.values()) {{
    const statsByIssue = {{}};
    const seenIssueKeys = new Set();
    let statsTotal = 0;
    for (const row of (Array.isArray(it.assigned_hierarchy) ? it.assigned_hierarchy : [])) {{
      const issueKey = String(row?.issue_key || "").toUpperCase();
      if (!issueKey || seenIssueKeys.has(issueKey)) continue;
      seenIssueKeys.add(issueKey);
      const compound = `${{String(it.assignee || "Unassigned")}}\\u0000${{issueKey}}`;
      const hours = extendedActualsEnabled
        ? n(allLoggedHoursByAssigneeIssue.get(compound))
        : n(it.issue_logged_hours_by_issue[issueKey]);
      if (hours > 0) {{
        statsByIssue[issueKey] = hours;
        statsTotal += hours;
      }}
    }}
    it.issue_logged_hours_stats_by_issue = statsByIssue;
    it.actual_hours_stats_total = statsTotal;
    for (const row of (Array.isArray(it.assigned_hierarchy) ? it.assigned_hierarchy : [])) {{
      const issueKey = String(row?.issue_key || "").toUpperCase();
      row.actual_hours = n(statsByIssue[issueKey]);
    }}
    let liveTotalEstimate = 0;
    let liveTotalActual = 0;
    let liveTotalOverrun = 0;
    let liveCommitmentOverrun = 0;
    let liveDuePenaltyEstimate = 0;
    let liveDuePenaltyCount = 0;
    let liveCommitmentCount = 0;
    let liveWithinCount = 0;
    let liveOverCount = 0;
    let liveNoEstimateCount = 0;
    let liveOnTimeCount = 0;
    let liveLateCount = 0;
    for (const ssRow of (Array.isArray(it.ss_subtask_details) ? it.ss_subtask_details : [])) {{
      const issueKey = String(ssRow?.issue_key || "").toUpperCase();
      const estimate = n(ssRow?.estimate);
      const actual = n(statsByIssue[issueKey]);
      const overrun = estimate > 0 ? Math.max(0, actual - estimate) : 0;
      const dueStatus = String(ssRow?.due_completion_status || "");
      let estimateStatus = "no_estimate";
      if (estimate > 0) estimateStatus = actual <= estimate ? "within_estimate" : "over_estimate";
      const isCommitment = estimateStatus === "over_estimate" && dueStatus === "on_time" ? 1 : 0;
      const penaltyReason = estimateStatus === "no_estimate"
        ? "No estimate; this subtask does not affect simple score."
        : (isCommitment
          ? "Over estimate but completed on time, so due-completion mode forgives its overrun."
          : (dueStatus === "late" && estimateStatus === "over_estimate"
            ? "Late completion: original estimate is penalized in due-completion mode, and its overrun also reduces score."
            : (dueStatus === "late"
              ? "Late completion: original estimate is penalized in due-completion mode."
              : (estimateStatus === "over_estimate"
                ? "Over estimate; its overrun reduces simple score."
                : "Within estimate; no simple-score penalty."))));
      ssRow.actual = actual;
      ssRow.overrun = overrun;
      ssRow.estimate_status = estimateStatus;
      ssRow.is_commitment = isCommitment;
      ssRow.penalty_reason = penaltyReason;
      liveTotalEstimate += estimate;
      liveTotalActual += actual;
      liveTotalOverrun += overrun;
      if (isCommitment) {{
        liveCommitmentOverrun += overrun;
        liveCommitmentCount += 1;
      }}
      if (dueStatus === "late" && estimate > 0) {{
        liveDuePenaltyEstimate += estimate;
        liveDuePenaltyCount += 1;
      }}
      if (estimateStatus === "within_estimate") liveWithinCount += 1;
      else if (estimateStatus === "over_estimate") liveOverCount += 1;
      else liveNoEstimateCount += 1;
      if (dueStatus === "on_time") liveOnTimeCount += 1;
      else if (dueStatus === "late") liveLateCount += 1;
    }}
    it.ss_total_estimate = liveTotalEstimate;
    it.ss_total_actual = liveTotalActual;
    it.ss_total_overrun = liveTotalOverrun;
    it.ss_commitment_overrun = liveCommitmentOverrun;
    it.ss_due_penalty_estimate = liveDuePenaltyEstimate;
    it.ss_due_penalty_count = liveDuePenaltyCount;
    it.ss_commitment_count = liveCommitmentCount;
    it.ss_within_count = liveWithinCount;
    it.ss_over_count = liveOverCount;
    it.ss_no_estimate_count = liveNoEstimateCount;
    it.ss_on_time_count = liveOnTimeCount;
    it.ss_late_count = liveLateCount;
  }}
  const items = Array.from(byA.values());
  for (const it of items) {{
    const activeRmiSet = epicKeysByAssignee.get(String(it.assignee || "")) || new Set();
    it.active_rmi_count = activeRmiSet.size;
    it.active_rmi_keys = Array.from(activeRmiSet).sort((a, b) => String(a || "").localeCompare(String(b || "")));
  }}
  const seriesDays = dateRangeDays(from, to);
  for (const it of items) {{
    it.penalties = {{ bug: it.bug_hours * n(settings.points_per_bug_hour), bug_late: it.bug_late_hours * n(settings.points_per_bug_late_hour), leave: it.unplanned_leave_hours * n(settings.points_per_unplanned_leave_hour), subtask_late: it.subtask_late_hours * n(settings.points_per_subtask_late_hour), estimate: it.estimate_overrun_hours * n(settings.points_per_estimate_overrun_hour), missed_due_date: it.missed_due_date_count * n(settings.points_per_missed_due_date) }};
    it.rework_ratio_pct = n(it.total_hours) > 0 ? (it.rework_hours / n(it.total_hours)) * 100 : 0;
    it.total_penalty = Object.values(it.penalties).reduce((a,b)=>a+n(b),0);
    it.raw_score = n(settings.base_score) - it.total_penalty;
    it.final_score = clamp(it.raw_score, n(settings.min_score), n(settings.max_score));
    it.missed_due_date_ratio = n(it.due_dated_assigned_count) > 0 ? (n(it.missed_due_date_count) / n(it.due_dated_assigned_count)) * 100 : 0;
    if (it.penalties.bug) it.feed.push({{label:"Bug hours", points:it.penalties.bug, hours:it.bug_hours}});
    if (it.penalties.bug_late) it.feed.push({{label:"Bug hours after story due", points:it.penalties.bug_late, hours:it.bug_late_hours}});
    if (it.penalties.subtask_late) it.feed.push({{label:"Subtask due overrun", points:it.penalties.subtask_late, hours:it.subtask_late_hours}});
    if (it.penalties.estimate) it.feed.push({{label:"Estimate overrun", points:it.penalties.estimate, hours:it.estimate_overrun_hours}});
    if (it.penalties.leave) it.feed.push({{label:"Unplanned leave", points:it.penalties.leave, hours:it.unplanned_leave_hours}});
    if (it.penalties.missed_due_date) it.feed.push({{label:"Missed due dates", points:it.penalties.missed_due_date, hours:it.missed_due_date_count}});
    it.feed.sort((a,b)=>n(b.points)-n(a.points));
    it.base_capacity_hours = Math.max(0, n(capacityByAssignee.get(it.assignee)));
    it.employee_capacity_hours = Math.max(0, n(it.base_capacity_hours) - n(it.planned_leave_hours) - n(it.unplanned_leave_hours));
    it.missed_start_ratio = n(it.total_assigned_count) > 0 ? (n(it.missed_start_count) / n(it.total_assigned_count)) * 100 : 0;
    it.entity_values = {{
      capacity: n(it.base_capacity_hours),
      planned_hours: n(it.planned_hours_assigned),
      actual_hours: n(it.total_hours),
      planned_leaves: n(it.planned_leave_hours),
      unplanned_leaves: n(it.unplanned_leave_hours),
      planned_dates: n(it.total_assigned_count),
      status: n(it.final_score),
      activity: n(it.total_hours)
    }};
    const scope = {{}};
    for (const e of entitiesCatalog) {{
      const key = String(e.entity_key || "").toLowerCase();
      if (!key) continue;
      scope[key] = n(it.entity_values[key]);
    }}
    for (const m of managedFields) {{
      const key = String(m.field_key || "").toLowerCase();
      if (!key) continue;
      const formula = String(m.formula_expression || "");
      const value = evalFormula(formula, scope);
      it.managed_values[key] = value;
      scope[key] = value;
    }}
    it.managed_scope = {{...scope}};
    const days = seriesDays.length ? seriesDays : Object.keys(it.daily_penalty_by_day || {{}}).sort();
    let cumulative = 0;
    it.daily_series = days.map((d) => {{
      cumulative += n(it.daily_penalty_by_day[d]);
      return {{
        day: d,
        penalty: n(it.daily_penalty_by_day[d]),
        score: clamp(n(settings.base_score) - cumulative, n(settings.min_score), n(settings.max_score))
      }};
    }});
  }}
  for (const it of items) {{
    const totalEst = n(it.ss_total_estimate);
    const totalActual = n(it.ss_total_actual);
    const totalOverSubtasks = n(it.ss_total_overrun);
    const totalOverTotal = totalEst > 0 ? Math.max(0, totalActual - totalEst) : 0;
    const totalOver = simpleOverrunMode === "total" ? totalOverTotal : totalOverSubtasks;
    const commitOver = simpleOverrunMode === "total" ? 0 : n(it.ss_commitment_overrun);
    const latePenaltyEstimate = n(it.ss_due_penalty_estimate);
    it.simple_score_overrun_total = totalOverTotal;
    it.simple_score_overrun_active = totalOver;
    it.score_eligible = n(it.planned_hours_assigned) > 0 ? 1 : 0;
    it.simple_score_raw = totalEst > 0 ? clamp(100 * (1 - totalOver / totalEst), 0, 100) : 0;
    const adjOver = Math.max(0, totalOver - commitOver);
    const dueAdjustedPenaltyHours = adjOver + latePenaltyEstimate;
    it.simple_score_due = totalEst > 0 ? clamp(100 * (1 - dueAdjustedPenaltyHours / totalEst), 0, 100) : 0;
    const effectiveCapacity = Math.max(0, n(it.employee_capacity_hours));
    it.simple_score_overloaded = totalEst > 0 ? clamp((effectiveCapacity / totalEst) * 100, 0, 100) : 0;
    const maxPlannedBeforeCap = effectiveCapacity * (1 + overloadedPenaltyThresholdPct / 100);
    const overloadedApplies = overloadedPenaltyEnabled && totalEst > 0 && effectiveCapacity > 0 && totalEst > maxPlannedBeforeCap;
    it.simple_score_overloaded_applied = overloadedApplies ? 1 : 0;
    it.simple_score_overloaded_penalty_pct = overloadedApplies ? clamp(100 - it.simple_score_overloaded, 0, 100) : 0;
    const baseSimpleScore = dueCompletionEnabled ? it.simple_score_due : it.simple_score_raw;
    it.simple_score = it.score_eligible
      ? (overloadedApplies
          ? (planningRealismEnabled
              ? it.simple_score_overloaded
              : clamp(baseSimpleScore - it.simple_score_overloaded_penalty_pct, 0, 100))
          : baseSimpleScore)
      : NaN;
  }}
  items.sort((a,b)=>n(b.final_score)-n(a.final_score) || a.assignee.localeCompare(b.assignee));
  return items;
}}
function renderSeriesSvg(series) {{
  const data = Array.isArray(series) ? series : [];
  if (!data.length) return '<div class="empty">No daily performance series in selected range.</div>';
  const w = 920;
  const h = 220;
  const padL = 42, padR = 12, padT = 12, padB = 28;
  const minS = n(settings.min_score), maxS = n(settings.max_score);
  const plotW = Math.max(1, w - padL - padR), plotH = Math.max(1, h - padT - padB);
  function x(i) {{ return padL + ((data.length <= 1 ? 0 : i / (data.length - 1)) * plotW); }}
  function y(score) {{
    const clamped = clamp(n(score), minS, maxS);
    const ratio = (maxS - minS) > 0 ? ((clamped - minS) / (maxS - minS)) : 0;
    return padT + (1 - ratio) * plotH;
  }}
  const pts = data.map((d, i) => `${{x(i).toFixed(2)}},${{y(d.score).toFixed(2)}}`).join(" ");
  const ticks = [minS, minS + (maxS - minS) * 0.5, maxS];
  const yTicks = ticks.map((v) => {{
    const yy = y(v).toFixed(2);
    return `<line x1="${{padL}}" y1="${{yy}}" x2="${{w - padR}}" y2="${{yy}}" stroke="#27406a" stroke-width="1"></line><text x="${{padL - 6}}" y="${{Number(yy) + 4}}" fill="#9db1d8" font-size="10" text-anchor="end">${{n(v).toFixed(0)}}</text>`;
  }}).join("");
  const first = data[0], last = data[data.length - 1];
  const xLabels = `<text x="${{padL}}" y="${{h - 8}}" fill="#9db1d8" font-size="10">${{e(first.day)}}</text><text x="${{w - padR}}" y="${{h - 8}}" fill="#9db1d8" font-size="10" text-anchor="end">${{e(last.day)}}</text>`;
  return `<svg class="ts-svg" viewBox="0 0 ${{w}} ${{h}}" preserveAspectRatio="none"><rect x="0" y="0" width="${{w}}" height="${{h}}" fill="#0d172b"></rect>${{yTicks}}<polyline fill="none" stroke="#60a5fa" stroke-width="2.5" points="${{pts}}"></polyline>${{xLabels}}</svg>`;
}}
function renderTeamChartSvg(rows, selectedTeamName) {{
  const data = Array.isArray(rows) ? rows.slice(0, 12) : [];
  if (!data.length) return '<div class="empty">No team data to chart.</div>';
  const w = 920;
  const barH = 18;
  const gap = 10;
  const h = 18 + (data.length * (barH + gap)) + 16;
  const padL = 180, padR = 24, padT = 14, padB = 12;
  const minS = n(settings.min_score), maxS = n(settings.max_score);
  const plotW = Math.max(1, w - padL - padR);
  function x(score) {{
    const clamped = clamp(n(score), minS, maxS);
    const ratio = (maxS - minS) > 0 ? ((clamped - minS) / (maxS - minS)) : 0;
    return padL + (ratio * plotW);
  }}
  const grid = [minS, minS + (maxS - minS) * 0.5, maxS].map((v) => {{
    const xx = x(v).toFixed(2);
    return `<line x1="${{xx}}" y1="${{padT - 2}}" x2="${{xx}}" y2="${{h - padB}}" stroke="#27406a" stroke-width="1"></line><text x="${{xx}}" y="${{padT - 4}}" fill="#9db1d8" font-size="10" text-anchor="middle">${{n(v).toFixed(0)}}</text>`;
  }}).join("");
  const bars = data.map((tr, i) => {{
    const y = padT + i * (barH + gap);
    const teamScore = n(tr.avg_score);
    const hasTeamScore = Number.isFinite(teamScore);
    const xx = hasTeamScore ? x(teamScore) : padL;
    const width = hasTeamScore ? Math.max(2, xx - padL) : 2;
    const selected = String(tr.team_name || "") === String(selectedTeamName || "");
    const fill = hasTeamScore ? (selected ? "#10b981" : "#22c55e") : "#475569";
    const stroke = selected ? "#93c5fd" : "none";
    const strokeW = selected ? "1.2" : "0";
    return `<g><text x="${{padL - 8}}" y="${{y + 12}}" fill="#dce8ff" font-size="11" text-anchor="end">${{e(tr.team_name)}}</text><rect x="${{padL}}" y="${{y}}" width="${{plotW}}" height="${{barH}}" rx="6" ry="6" fill="#122746"></rect><rect x="${{padL}}" y="${{y}}" width="${{width.toFixed(2)}}" height="${{barH}}" rx="6" ry="6" fill="${{fill}}" stroke="${{stroke}}" stroke-width="${{strokeW}}"></rect><rect class="team-bar-hit" data-team-name="${{e(tr.team_name)}}" x="${{padL}}" y="${{y}}" width="${{plotW}}" height="${{barH}}" rx="6" ry="6" fill="transparent" style="cursor:pointer;"></rect><text x="${{(xx + 6).toFixed(2)}}" y="${{y + 12}}" fill="#dce8ff" font-size="11">${{hasTeamScore ? teamScore.toFixed(1) : "N/A"}}</text></g>`;
  }}).join("");
  return `<svg class="team-chart-svg" viewBox="0 0 ${{w}} ${{h}}" preserveAspectRatio="none"><rect x="0" y="0" width="${{w}}" height="${{h}}" fill="#0d172b"></rect>${{grid}}${{bars}}</svg>`;
}}
function renderSimpleList(rows, formatter) {{
  const data = Array.isArray(rows) ? rows : [];
  if (!data.length) return '<div class="empty">No items.</div>';
  return data.map(formatter).join("");
}}
function renderAssignedMixChart(item) {{
  const vals = [n(item?.assigned_counts?.epic), n(item?.assigned_counts?.story), n(item?.assigned_counts?.subtask)];
  const labels = ["Epic", "Story", "Subtask"];
  const colors = ["#38bdf8", "#22c55e", "#f59e0b"];
  const maxv = Math.max(1, ...vals);
  const w = 520, h = 160, padL = 85, padR = 10, padT = 10, padB = 16;
  const rowH = 34;
  const bars = vals.map((v, i) => {{
    const y = padT + i * rowH;
    const ww = ((w - padL - padR) * (v / maxv));
    return `<g><text x="${{padL - 8}}" y="${{y + 15}}" fill="#dce8ff" font-size="11" text-anchor="end">${{labels[i]}}</text><rect x="${{padL}}" y="${{y}}" width="${{(w-padL-padR)}}" height="18" fill="#132846" rx="6"></rect><rect x="${{padL}}" y="${{y}}" width="${{ww.toFixed(2)}}" height="18" fill="${{colors[i]}}" rx="6"></rect><text x="${{(padL + ww + 6).toFixed(2)}}" y="${{y + 14}}" fill="#cfe3ff" font-size="11">${{v.toFixed(0)}}</text></g>`;
  }}).join("");
  return `<svg class="mini-svg" viewBox="0 0 ${{w}} ${{h}}" preserveAspectRatio="none"><rect x="0" y="0" width="${{w}}" height="${{h}}" fill="#0d172b"></rect>${{bars}}</svg>`;
}}
function renderDueComplianceChart(rows) {{
  const data = Array.isArray(rows) ? rows : [];
  const buckets = {{
    "Before due": 0,
    "On due": 0,
    "After due": 0,
    "Not completed": 0
  }};
  data.forEach((r) => {{
    const k = String(r?.status_bucket || "Not completed");
    if (!(k in buckets)) buckets["Not completed"] += 1;
    else buckets[k] += 1;
  }});
  const vals = [buckets["Before due"], buckets["On due"], buckets["After due"], buckets["Not completed"]];
  const labels = ["Before", "On", "After", "Open"];
  const colors = ["#22c55e", "#38bdf8", "#f43f5e", "#94a3b8"];
  const total = Math.max(1, vals.reduce((a, b) => a + b, 0));
  const w = 520, h = 160, cx = 115, cy = 80, r = 52, ir = 28;
  let angle = -Math.PI / 2;
  const segs = vals.map((v, i) => {{
    const frac = v / total;
    const next = angle + frac * Math.PI * 2;
    const large = frac > 0.5 ? 1 : 0;
    const x1 = cx + Math.cos(angle) * r, y1 = cy + Math.sin(angle) * r;
    const x2 = cx + Math.cos(next) * r, y2 = cy + Math.sin(next) * r;
    const x3 = cx + Math.cos(next) * ir, y3 = cy + Math.sin(next) * ir;
    const x4 = cx + Math.cos(angle) * ir, y4 = cy + Math.sin(angle) * ir;
    const d = `M ${{x1}} ${{y1}} A ${{r}} ${{r}} 0 ${{large}} 1 ${{x2}} ${{y2}} L ${{x3}} ${{y3}} A ${{ir}} ${{ir}} 0 ${{large}} 0 ${{x4}} ${{y4}} Z`;
    angle = next;
    return `<path d="${{d}}" fill="${{colors[i]}}"></path>`;
  }}).join("");
  const legend = vals.map((v, i) => `<text x="210" y="${{26 + i*24}}" fill="#dce8ff" font-size="11">${{labels[i]}}: ${{v}}</text><rect x="188" y="${{18 + i*24}}" width="14" height="14" rx="3" fill="${{colors[i]}}"></rect>`).join("");
  return `<svg class="mini-svg" viewBox="0 0 ${{w}} ${{h}}" preserveAspectRatio="none"><rect x="0" y="0" width="${{w}}" height="${{h}}" fill="#0d172b"></rect>${{segs}}<text x="${{cx}}" y="${{cy + 4}}" fill="#cfe3ff" font-size="12" text-anchor="middle">${{total}}</text>${{legend}}</svg>`;
}}
function issueTypeMeta(rawType) {{
  const type = String(rawType || "").toLowerCase();
  if (type === "epic") return {{ label: "Epic", icon: "bolt" }};
  if (type === "story") return {{ label: "Story", icon: "person" }};
  if (type === "bug_subtask") return {{ label: "Bug Subtask", icon: "bug_report" }};
  return {{ label: "Subtask", icon: "construction" }};
}}
function issueTypePill(rawType) {{
  const meta = issueTypeMeta(rawType);
  const low = String(rawType || "").toLowerCase();
  const tone = low === "epic" ? "issue-epic" : (low === "story" ? "issue-story" : (low === "bug_subtask" ? "issue-bug-subtask" : "issue-subtask"));
  return `<span class="metric-pill issue-type-pill ${{tone}}" title="${{e(meta.label)}}" aria-label="${{e(meta.label)}}"><span class="material-symbols-outlined" aria-hidden="true">${{e(meta.icon)}}</span></span>`;
}}
function subtaskTypeIcon(rawType) {{
  const low = String(rawType || "").toLowerCase();
  const isBug = low === "bug_subtask";
  const icon = isBug ? "bug_report" : "construction";
  const label = isBug ? "Bug Subtask" : "Subtask";
  const tone = isBug ? "issue-bug-subtask" : "issue-subtask";
  return `<span class="subtask-type-icon ${{tone}}" title="${{e(label)}}" aria-label="${{e(label)}}"><span class="material-symbols-outlined" aria-hidden="true">${{e(icon)}}</span><span>${{e(label)}}</span></span>`;
}}
function renderHierarchyTable(rows, dueRows, missedRows) {{
  const data = Array.isArray(rows) ? rows : [];
  if (!data.length) return '<div class="empty" style="padding:8px;">No assigned items.</div>';
  const dueMap = new Map((Array.isArray(dueRows) ? dueRows : []).map((r) => [String(r.issue_key || ""), String(r.status_bucket || "")]));
  const missedSet = new Set((Array.isArray(missedRows) ? missedRows : []).map((r) => String(r.issue_key || "")));
  const byKey = new Map(data.map((r) => [String(r.issue_key || ""), r]));
  const childMap = new Map();
  for (const r of data) {{
    const p = String(r.parent_issue_key || "");
    if (!childMap.has(p)) childMap.set(p, []);
    childMap.get(p).push(r);
  }}
  function typeRank(t) {{
    const x = String(t || "");
    if (x === "epic") return 0;
    if (x === "story") return 1;
    if (x === "subtask") return 2;
    return 3;
  }}
  function nodeType(node) {{
    return String(node?.hierarchy_type || node?.issue_type || "").toLowerCase();
  }}
  function sorted(list) {{
    return (list || []).slice().sort((a,b) => typeRank(nodeType(a)) - typeRank(nodeType(b)) || String(a.issue_key).localeCompare(String(b.issue_key)));
  }}
  function renderNode(node) {{
    const key = String(node.issue_key || "");
    const nodeKind = nodeType(node);
    const kids = sorted(childMap.get(key) || []);
    const dueBucket = dueMap.get(key) || "-";
    const missed = missedSet.has(key) ? "Yes" : "No";
    const openAttr = nodeKind === "epic" ? " open" : "";
    const isBugSubtask = nodeKind === "bug_subtask";
    const actualCompleteText = nodeKind === "subtask" || nodeKind === "bug_subtask"
      ? `Actual Complete: ${{e(node.actual_complete_date || node.completion_date || "-")}}`
      : `Done: ${{e(node.completion_date || "-")}}`;
    const summary = `<summary><div class="tree-left"><div class="issue-id">${{e(key)}}</div><div class="issue-title">${{e(node.summary || "")}}${{isBugSubtask ? '<span class="issue-kind-inline">Bug Subtask</span>' : ""}}</div></div><div class="tree-metrics">${{issueTypePill(nodeKind)}}<span class="metric-pill">Est: ${{n(node.original_estimate_hours).toFixed(1)}}h</span><span class="metric-pill">Start: ${{e(node.start_date || "-")}}</span><span class="metric-pill">Planned Due: ${{e(node.planned_due_date || node.due_date || "-")}}</span><span class="metric-pill">${{actualCompleteText}}</span><span class="metric-pill">Due Status: ${{e(dueBucket)}}</span><span class="metric-pill">Missed Start: ${{missed}}</span></div></summary>`;
    if (!kids.length) return `<div class="tree-node"><details${{openAttr}}>${{summary}}</details></div>`;
    return `<div class="tree-node"><details${{openAttr}}>${{summary}}<div class="tree-children">${{kids.map(renderNode).join("")}}</div></details></div>`;
  }}
  const roots = sorted(data.filter((r) => {{
    const p = String(r.parent_issue_key || "");
    return !p || !byKey.has(p);
  }}));
  return `<div class="tree-wrap">${{roots.map(renderNode).join("")}}</div>`;
}}
function renderDueTable(rows) {{
  const data = Array.isArray(rows) ? rows : [];
  if (!data.length) return '<div class="empty" style="padding:8px;">No logged items.</div>';
  function dueBucketClass(bucket) {{
    const b = String(bucket || "");
    if (b === "Before due") return "due-bucket due-before";
    if (b === "On due") return "due-bucket due-on";
    if (b === "After due") return "due-bucket due-after";
    return "due-bucket";
  }}
  return `<table class="tbl"><thead><tr><th>Issue</th><th>Due</th><th>Last Logged Date</th><th>Actual Completed Date</th><th>Stable Resolved</th><th>Bucket</th></tr></thead><tbody>${{data.map((r)=>`<tr class="${{r.is_missed_due_date ? "due-missed penalized-row" : ""}}"><td><div class="issue-id">${{e(r.issue_key)}}</div><div class="issue-title">${{e(r.summary)}}</div></td><td>${{e(r.planned_due_date || r.due_date || "-")}}</td><td>${{e(r.last_logged_date || "-")}}</td><td>${{e(r.actual_complete_date || r.completion_date || "-")}}<div class="sub">${{e(actualCompletionSourceText(r.actual_complete_source))}}</div></td><td>${{e(r.resolved_stable_since_date || "-")}}</td><td><span class="${{dueBucketClass(r.status_bucket)}}">${{e(r.status_bucket)}}</span></td></tr>`).join("")}}</tbody></table>`;
}}
function renderMissedTable(rows) {{
  const data = Array.isArray(rows) ? rows : [];
  if (!data.length) return '<div class="empty" style="padding:8px;">No missed starts.</div>';
  return `<table class="tbl"><thead><tr><th>Issue</th><th>Planned Start</th><th>Other Work</th><th>Leave (h)</th><th>State</th></tr></thead><tbody>${{data.map((r)=>{{ const other=(r.other_work||[]).map((x)=>`${{e(x.issue_id)}} (${{n(x.hours).toFixed(1)}}h)`).join(", "); const state=r.idle ? "Idle" : (n(r.leave_hours)>0 ? "On Leave" : "Working on other items"); return `<tr><td><div class="issue-id">${{e(r.issue_key)}}</div><div class="issue-title">${{e(r.summary)}}</div></td><td>${{e(r.start_date || "-")}}</td><td>${{other || "-"}}</td><td>${{n(r.leave_hours).toFixed(1)}}</td><td>${{state}}</td></tr>`; }}).join("")}}</tbody></table>`;
}}
function renderExecutionHierarchyTable(item) {{
  const nodes = Array.isArray(item?.assigned_hierarchy) ? item.assigned_hierarchy : [];
  const subtasks = nodes.filter((n0) => String(n0?.issue_type || "").toLowerCase() === "subtask");
  if (!subtasks.length) return '<div class="empty" style="padding:8px;">No subtasks in current scope.</div>';

  function wiByKey(key) {{
    return workItemsByKey.get(String(key || "").toUpperCase()) || null;
  }}
  function itemStart(x) {{ return String(x?.start_date || x?.item_start_date || ""); }}
  function itemDue(x) {{ return String(x?.due_date || x?.end_date || x?.item_due_date || ""); }}
  function itemSummary(x, fallback) {{ return String(x?.summary || fallback || ""); }}
  function maxDate(values) {{
    const vals = (Array.isArray(values) ? values : []).map((v) => String(v || "")).filter(Boolean).sort();
    return vals.length ? vals[vals.length - 1] : "";
  }}
  function minDate(values) {{
    const vals = (Array.isArray(values) ? values : []).map((v) => String(v || "")).filter(Boolean).sort();
    return vals.length ? vals[0] : "";
  }}

  const epicMap = new Map();
  for (const st of subtasks) {{
    const stKey = String(st.issue_key || "").toUpperCase();
    const storyKey = String(st.parent_issue_key || "").toUpperCase();
    const storyWi = wiByKey(storyKey);
    const epicKey = String(st.parent_epic_key || storyWi?.parent_issue_key || "").toUpperCase();
    const epicWi = wiByKey(epicKey);
    const storyActual = n(st.actual_hours);
    const storyCompletion = String(st.completion_date || "");

    if (!epicMap.has(epicKey || "NO_EPIC")) {{
      epicMap.set(epicKey || "NO_EPIC", {{
        epic_key: epicKey || "",
        epic_summary: itemSummary(epicWi, epicKey || "No Epic"),
        planned_start: itemStart(epicWi),
        planned_due: itemDue(epicWi),
        planned_hours: n(epicWi?.original_estimate_hours),
        actual_hours: 0,
        completion_dates: [],
        resolved_stable: String(epicWi?.resolved_stable_since_date || ""),
        stories: new Map(),
      }});
    }}
    const epicNode = epicMap.get(epicKey || "NO_EPIC");
    epicNode.actual_hours += storyActual;
    if (storyCompletion) epicNode.completion_dates.push(storyCompletion);

    if (!epicNode.stories.has(storyKey || "NO_STORY")) {{
      epicNode.stories.set(storyKey || "NO_STORY", {{
        story_key: storyKey || "",
        story_summary: itemSummary(storyWi, storyKey || "No Story"),
        planned_start: itemStart(storyWi),
        planned_due: itemDue(storyWi),
        planned_hours: n(storyWi?.original_estimate_hours),
        actual_hours: 0,
        completion_dates: [],
        resolved_stable: String(storyWi?.resolved_stable_since_date || ""),
        subtasks: new Map(),
      }});
    }}
    const storyNode = epicNode.stories.get(storyKey || "NO_STORY");
    storyNode.actual_hours += storyActual;
    if (storyCompletion) storyNode.completion_dates.push(storyCompletion);
    const stNodeKey = stKey || "NO_SUBTASK";
    if (!storyNode.subtasks.has(stNodeKey)) {{
      storyNode.subtasks.set(stNodeKey, {{
        issue_key: stKey,
        summary: String(st.summary || ""),
        planned_start: String(st.start_date || ""),
        planned_due: String(st.due_date || ""),
        planned_hours: n(st.original_estimate_hours),
        actual_hours: 0,
        negative_hours: 0,
        last_logged_date: String(st.last_logged_date || ""),
        actual_complete_date: String(st.actual_complete_date || st.completion_date || ""),
        actual_complete_source: String(st.actual_complete_source || "none"),
        completion_date: storyCompletion,
        resolved_stable: String(st.resolved_stable_since_date || ""),
        is_penalized_for_due: !!st.is_penalized_for_due,
        status: String(st.status || ""),
      }});
    }}
    const stNode = storyNode.subtasks.get(stNodeKey);
    stNode.actual_hours += n(st.actual_hours);
    stNode.negative_hours += n(st.negative_hours);
    if (storyCompletion && (!stNode.completion_date || storyCompletion > stNode.completion_date)) stNode.completion_date = storyCompletion;
  }}

  const epicNodes = Array.from(epicMap.values()).sort((a, b) => a.epic_summary.localeCompare(b.epic_summary) || a.epic_key.localeCompare(b.epic_key));
  return `<div class="exec-accordion">${{epicNodes.map((epic, epicIndex) => {{
    const storyNodes = Array.from(epic.stories.values()).sort((a, b) => a.story_summary.localeCompare(b.story_summary) || a.story_key.localeCompare(b.story_key));
    const epicPlannedHours = epic.planned_hours > 0 ? epic.planned_hours : storyNodes.reduce((acc, s0) => acc + n(s0.planned_hours), 0);
    const epicStart = epic.planned_start || minDate(storyNodes.map((s0) => s0.planned_start));
    const epicDue = epic.planned_due || maxDate(storyNodes.map((s0) => s0.planned_due));
    const epicDone = maxDate(epic.completion_dates);
    const epicNegativeCount = storyNodes.reduce((acc, story) => acc + Array.from(story.subtasks.values()).filter((st0) => n(st0.negative_hours) > 0).length, 0);
    const storyHtml = storyNodes.map((story) => {{
      const subRows = Array.from(story.subtasks.values()).sort((a, b) => a.summary.localeCompare(b.summary) || a.issue_key.localeCompare(b.issue_key));
      const storyPlannedHours = story.planned_hours > 0 ? story.planned_hours : subRows.reduce((acc, st0) => acc + n(st0.planned_hours), 0);
      const storyStart = story.planned_start || minDate(subRows.map((st0) => st0.planned_start));
      const storyDue = story.planned_due || maxDate(subRows.map((st0) => st0.planned_due));
      return `<section class="exec-story-block"><div class="exec-story-head"><div><span class="metric-pill">STORY</span><div class="issue-id">${{e(story.story_key || "-")}}</div><div class="issue-title">${{e(story.story_summary || "-")}}</div></div><div class="exec-story-metrics"><span class="metric-pill">Start: ${{e(storyStart || "-")}}</span><span class="metric-pill">Due: ${{e(storyDue || "-")}}</span><span class="metric-pill">Planned: ${{n(storyPlannedHours).toFixed(2)}}h</span><span class="metric-pill">Actual: ${{n(story.actual_hours).toFixed(2)}}h</span><span class="metric-pill">Done: ${{e(maxDate(story.completion_dates) || "-")}}</span></div></div><table class="tbl exec-subtask-table"><thead><tr><th>Subtask</th><th>Planned Start</th><th>Planned Due Date</th><th>Planned Hours</th><th>Actual Hours</th><th>Last Logged Date</th><th>Actual Complete Date</th><th>Status Resolved Date</th><th>Status</th></tr></thead><tbody>${{subRows.map((st) => {{
        const negHrs = n(st.negative_hours);
        const status = negHrs > 0 ? `Penalty hit: late by ${{negHrs.toFixed(2)}}h` : (st.status || "");
        const rowCls = (negHrs > 0 || st.is_penalized_for_due) ? "exec-negative-subtask penalized-row" : "";
        return `<tr class="${{rowCls}}"><td><div><span class="metric-pill">SUBTASK</span><div class="issue-id">${{e(st.issue_key || "-")}}</div><div class="issue-title">${{e(st.summary || "-")}}</div></div></td><td>${{e(st.planned_start || "-")}}</td><td>${{e(st.planned_due || "-")}}</td><td>${{n(st.planned_hours).toFixed(2)}}h</td><td>${{n(st.actual_hours).toFixed(2)}}h</td><td>${{e(st.last_logged_date || "-")}}</td><td>${{e(st.actual_complete_date || st.completion_date || "-")}}<div class="sub">${{e(actualCompletionSourceText(st.actual_complete_source))}}</div></td><td>${{e(st.resolved_stable || "-")}}</td><td>${{e(status)}}</td></tr>`;
      }}).join("")}}</tbody></table></section>`;
    }}).join("");
    return `<details class="exec-epic"${{epicIndex === 0 ? " open" : ""}}><summary><div class="exec-epic-left"><span class="metric-pill">EPIC</span><div class="issue-id">${{e(epic.epic_key || "-")}}</div><div class="issue-title">${{e(epic.epic_summary || "-")}}</div></div><div class="exec-epic-metrics"><span class="metric-pill">Stories: ${{storyNodes.length}}</span><span class="metric-pill">Subtasks: ${{storyNodes.reduce((acc, s0) => acc + s0.subtasks.size, 0)}}</span><span class="metric-pill">Start: ${{e(epicStart || "-")}}</span><span class="metric-pill">Due: ${{e(epicDue || "-")}}</span><span class="metric-pill">Planned: ${{n(epicPlannedHours).toFixed(2)}}h</span><span class="metric-pill">Actual: ${{n(epic.actual_hours).toFixed(2)}}h</span><span class="metric-pill">Done: ${{e(epicDone || "-")}}</span>${{epicNegativeCount > 0 ? `<span class="metric-pill exec-neg-pill">Penalty Subtasks: ${{epicNegativeCount}}</span>` : ""}}</div></summary><div class="exec-epic-body">${{storyHtml}}</div></details>`;
  }}).join("")}}</div>`;
}}
function toTitleCaseKey(key) {{
  const txt = String(key || "").replace(/_/g, " ").trim();
  if (!txt) return "Metric";
  return txt.replace(/\\w\\S*/g, (w) => w.charAt(0).toUpperCase() + w.slice(1).toLowerCase());
}}
function leaderboardCapacityValue(it) {{
  const managed = getManagedValue(it, ["capacity_available_for_more_work", "capacityavailableformorework", "capacity_available_more_work"]);
  if (Number.isFinite(managed)) return managed;
  return n(it && it.capacity_gap_hours);
}}
async function copyTextToClipboard(value) {{
  const text = String(value || "");
  if (!text) return false;
  if (navigator.clipboard && typeof navigator.clipboard.writeText === "function") {{
    try {{
      await navigator.clipboard.writeText(text);
      return true;
    }} catch (error) {{}}
  }}
  const textarea = document.createElement("textarea");
  textarea.value = text;
  textarea.setAttribute("readonly", "readonly");
  textarea.style.position = "fixed";
  textarea.style.opacity = "0";
  textarea.style.pointerEvents = "none";
  document.body.appendChild(textarea);
  textarea.focus();
  textarea.select();
  let copied = false;
  try {{
    copied = document.execCommand("copy");
  }} catch (error) {{
    copied = false;
  }}
  document.body.removeChild(textarea);
  return copied;
}}
function setLeaderboardActionStatus(text, tone) {{
  if (!leaderboardActionStatusEl) return;
  if (leaderboardActionStatusTimer) {{
    clearTimeout(leaderboardActionStatusTimer);
    leaderboardActionStatusTimer = null;
  }}
  leaderboardActionStatusEl.textContent = String(text || "");
  if (tone === "ok") leaderboardActionStatusEl.style.color = "#86efac";
  else if (tone === "warn") leaderboardActionStatusEl.style.color = "#facc15";
  else if (tone === "err") leaderboardActionStatusEl.style.color = "#fca5a5";
  else leaderboardActionStatusEl.style.color = "#9db1d8";
  if (text) {{
    leaderboardActionStatusTimer = setTimeout(() => {{
      if (!leaderboardActionStatusEl) return;
      leaderboardActionStatusEl.textContent = "";
      leaderboardActionStatusEl.style.color = "#9db1d8";
    }}, 3500);
  }}
}}
function setLeaderActionsMenuOpen(isOpen) {{
  if (!leaderActionsMenu || !leaderActionsToggle) return;
  const open = Boolean(isOpen);
  leaderActionsMenu.hidden = !open;
  leaderActionsToggle.setAttribute("aria-expanded", open ? "true" : "false");
}}
async function copyGapPeopleFromLeaderboard() {{
  const source = Array.isArray(lastLeaderboardViewItems) ? lastLeaderboardViewItems : [];
  const names = source
    .filter((it) => leaderboardCapacityValue(it) < 0)
    .map((it) => String(it && it.assignee || "").trim())
    .filter(Boolean);
  const uniqueNames = Array.from(new Set(names));
  if (!uniqueNames.length) {{
    setLeaderboardActionStatus("No gap people found in the current leaderboard filter.", "warn");
    return;
  }}
  const copied = await copyTextToClipboard(uniqueNames.join("\\n"));
  if (copied) setLeaderboardActionStatus(`Copied ${{uniqueNames.length}} gap people.`, "ok");
  else setLeaderboardActionStatus("Could not copy. Please allow clipboard access.", "err");
}}
function getManagedValue(item, candidates) {{
  const values = item && item.managed_values ? item.managed_values : {{}};
  const byKey = new Map(Object.entries(values).map(([k, v]) => [String(k || "").toLowerCase(), n(v)]));
  for (const raw of (Array.isArray(candidates) ? candidates : [])) {{
    const key = String(raw || "").toLowerCase();
    if (byKey.has(key)) return n(byKey.get(key));
  }}
  return NaN;
}}
function isScoreEligible(item) {{
  return n(item?.score_eligible) > 0
    || n(item?.planned_hours_assigned) > 0
    || n(item?.eligible_members) > 0
    || Number.isFinite(n(item?.avg_score));
}}
function scoreNumber(item, mode) {{
  if (!isScoreEligible(item)) return NaN;
  if (mode === "simple") return n(item?.simple_score);
  if (Number.isFinite(n(item?.final_score))) return n(item?.final_score);
  return n(item?.avg_score);
}}
function scoreText(item, mode) {{
  const value = scoreNumber(item, mode);
  return Number.isFinite(value) ? value.toFixed(1) : "N/A";
}}
function scoreSortValue(item, mode) {{
  const value = scoreNumber(item, mode);
  return Number.isFinite(value) ? value : Number.NEGATIVE_INFINITY;
}}
function renderSettingsLoadingState() {{
  const loadingHtml = '<div class="empty" style="padding:10px;">Loading performance settings before calculating scores.</div>';
  const leaderboardEl = document.getElementById("leaderboard");
  const detailEl = document.getElementById("detail");
  const scoreDrilldownEl = document.getElementById("score-drilldown");
  const teamChartHost = document.getElementById("team-performance-chart");
  const teamDetailHost = document.getElementById("selected-team-performance");
  if (document.getElementById("kpi-avg")) document.getElementById("kpi-avg").textContent = "Loading...";
  if (document.getElementById("kpi-top")) document.getElementById("kpi-top").textContent = "-";
  if (document.getElementById("kpi-risk")) document.getElementById("kpi-risk").textContent = "-";
  if (document.getElementById("kpi-pen")) document.getElementById("kpi-pen").textContent = "-";
  if (document.getElementById("kpi-rework")) document.getElementById("kpi-rework").textContent = "-";
  if (document.getElementById("header-average-performance-value")) document.getElementById("header-average-performance-value").textContent = "Loading...";
  if (document.getElementById("header-average-performance-mode")) document.getElementById("header-average-performance-mode").innerHTML = `<span class="material-symbols-outlined" aria-hidden="true">monitoring</span><span>Loading Settings</span>`;
  if (document.getElementById("header-average-performance-meta")) document.getElementById("header-average-performance-meta").textContent = "Waiting for performance settings.";
  if (document.getElementById("discover-insights")) document.getElementById("discover-insights").innerHTML = '<span class="pill">Loading settings...</span>';
  if (document.getElementById("top3-high")) document.getElementById("top3-high").innerHTML = loadingHtml;
  if (document.getElementById("top3-low")) document.getElementById("top3-low").innerHTML = loadingHtml;
  if (leaderboardEl) leaderboardEl.innerHTML = loadingHtml;
  if (detailEl) detailEl.innerHTML = '<div class="empty">Loading performance settings.</div>';
  if (scoreDrilldownEl) scoreDrilldownEl.innerHTML = '<div class="empty">Loading performance settings.</div>';
  if (teamChartHost) teamChartHost.innerHTML = loadingHtml;
  if (teamDetailHost) teamDetailHost.innerHTML = '<div class="empty">Loading performance settings.</div>';
}}
function render(items) {{
  if (!performanceSettingsReady) {{
    renderSettingsLoadingState();
    return;
  }}
  const activeMode = String(document.getElementById("leader-scoring-mode")?.value || "simple");
  const eligibleScoredItems = items.filter((i) => Number.isFinite(scoreNumber(i, activeMode)));
  const avgActiveScore = eligibleScoredItems.length
    ? (eligibleScoredItems.reduce((a, b) => a + scoreNumber(b, activeMode), 0) / eligibleScoredItems.length)
    : NaN;
  const topAdvancedItem = eligibleScoredItems
    .slice()
    .sort((a, b) => scoreSortValue(b, activeMode) - scoreSortValue(a, activeMode) || a.assignee.localeCompare(b.assignee))[0];
  document.getElementById("kpi-avg").textContent = Number.isFinite(avgActiveScore) ? avgActiveScore.toFixed(1) : "N/A";
  document.getElementById("kpi-top").textContent = topAdvancedItem?.assignee || "-";
  document.getElementById("kpi-risk").textContent = String(eligibleScoredItems.filter(i => n(i.final_score) < 60).length);
  document.getElementById("kpi-pen").textContent = items.reduce((a,b)=>a+n(b.total_penalty),0).toFixed(1);
  document.getElementById("kpi-rework").textContent = items.reduce((a,b)=>a+n(b.rework_hours),0).toFixed(1) + "h";
  const headerAverageValueEl = document.getElementById("header-average-performance-value");
  const headerAverageModeEl = document.getElementById("header-average-performance-mode");
  const headerAverageMetaEl = document.getElementById("header-average-performance-meta");
  if (headerAverageValueEl) headerAverageValueEl.textContent = Number.isFinite(avgActiveScore) ? `${{avgActiveScore.toFixed(1)}}%` : "N/A";
  if (headerAverageModeEl) headerAverageModeEl.innerHTML = `<span class="material-symbols-outlined" aria-hidden="true">monitoring</span><span>${{e(activeMode === "simple" ? "Simple Scoring" : "Advanced Scoring")}}</span>`;
  if (headerAverageMetaEl) headerAverageMetaEl.textContent = `Eligible assignees: ${{eligibleScoredItems.length}} of ${{items.length}}`;
  const totalAssignees = items.length;
  const atRiskCount = eligibleScoredItems.filter((i) => n(i.final_score) < 60).length;
  const ineligibleAssignees = items.filter((i) => !isScoreEligible(i)).length;
  const highMissed = items.filter((i) => n(i.missed_start_ratio) >= 30).length;
  const overloaded = items.filter((i) => (n(i.planned_hours_assigned) - n(i.employee_capacity_hours)) > 0).length;
  document.getElementById("discover-insights").innerHTML = [
    `<span class="pill">Assignees: ${{totalAssignees}}</span>`,
    `<span class="pill">At-Risk: ${{atRiskCount}}</span>`,
    `<span class="pill">Scoring N/A: ${{ineligibleAssignees}}</span>`,
    `<span class="pill">High Missed-Start (>=30%): ${{highMissed}}</span>`,
    `<span class="pill">Over Capacity: ${{overloaded}}</span>`
  ].join("");
  const rankedAdvancedItems = eligibleScoredItems
    .slice()
    .sort((a, b) => scoreSortValue(b, activeMode) - scoreSortValue(a, activeMode) || a.assignee.localeCompare(b.assignee));
  const topHigh = rankedAdvancedItems.slice(0, 3);
  const topLow = rankedAdvancedItems.slice(-3).reverse();
  document.getElementById("top3-high").innerHTML = topHigh.length
    ? topHigh.map((it, idx) => `<div class="top3-item high"><span class="nm">#${{idx + 1}} ${{e(it.assignee)}}</span><span class="sc">${{scoreText(it, "advanced")}}</span></div>`).join("")
    : '<div class="empty">No data.</div>';
  document.getElementById("top3-low").innerHTML = topLow.length
    ? topLow.map((it, idx) => `<div class="top3-item low"><span class="nm">#${{idx + 1}} ${{e(it.assignee)}}</span><span class="sc">${{scoreText(it, "advanced")}}</span></div>`).join("")
    : '<div class="empty">No data.</div>';
  const teamChartHost = document.getElementById("team-performance-chart");
  const teamDetailHost = document.getElementById("selected-team-performance");
  let teamRows = [];
  if (!teams.length) {{
    teamChartHost.innerHTML = '<div class="empty">No teams configured in settings.</div>';
    teamDetailHost.innerHTML = '<div class="empty">No teams configured in settings.</div>';
    selectedTeam = "";
  }} else {{
    const byName = new Map(items.map((it) => [String(it.assignee || "").toLowerCase(), it]));
    teamRows = teams.map((t) => {{
      const members = Array.isArray(t.assignees) ? t.assignees : [];
      const matched = members.map((m) => byName.get(String(m || "").toLowerCase())).filter(Boolean);
      const eligibleMatched = matched.filter((m) => isScoreEligible(m));
      const avgScore = eligibleMatched.length ? (eligibleMatched.reduce((a, b) => a + n(b.final_score), 0) / eligibleMatched.length) : NaN;
      const totalPenalty = matched.reduce((a, b) => a + n(b.total_penalty), 0);
      const atRisk = eligibleMatched.filter((m) => n(m.final_score) < 60).length;
      return {{
        team_name: String(t.team_name || ""),
        team_leader: String(t.team_leader || "-"),
        members: members,
        total_members: members.length,
        active_members: matched.length,
        eligible_members: eligibleMatched.length,
        avg_score: avgScore,
        total_penalty: totalPenalty,
        total_rework_hours: matched.reduce((a, b) => a + n(b.rework_hours), 0),
        at_risk: atRisk,
        planned_leave_count: matched.reduce((a, b) => a + n(b.planned_leave_count), 0),
        planned_leave_hours: matched.reduce((a, b) => a + n(b.planned_leave_hours), 0),
        planned_leave_days: matched.reduce((a, b) => a + n(b.planned_leave_days), 0),
        unplanned_leave_count: matched.reduce((a, b) => a + n(b.unplanned_leave_count), 0),
        unplanned_leave_hours: matched.reduce((a, b) => a + n(b.unplanned_leave_hours), 0),
        unplanned_leave_days: matched.reduce((a, b) => a + n(b.unplanned_leave_days), 0)
      }};
    }}).sort((a, b) => scoreSortValue(b, "advanced") - scoreSortValue(a, "advanced") || a.team_name.localeCompare(b.team_name));
    if (selectedTeam && !teamRows.some((tr) => tr.team_name === selectedTeam)) selectedTeam = "";
    teamChartHost.innerHTML = renderTeamChartSvg(teamRows, selectedTeam);
    Array.from(teamChartHost.querySelectorAll(".team-bar-hit")).forEach((node) => {{
      node.addEventListener("click", () => {{
        const clicked = String(node.getAttribute("data-team-name") || "");
        selectedTeam = selectedTeam === clicked ? "" : clicked;
        render(compute());
      }});
    }});
  }}
  const selectedTeamRow = selectedTeam ? teamRows.find((tr) => tr.team_name === selectedTeam) : null;
  if (selectedTeamRow) {{
    teamDetailHost.innerHTML = `
      <div class="team-card sel" style="cursor:default;transform:none;">
        <div class="team-head">
          <div>
            <div class="team-name">${{e(selectedTeamRow.team_name || "-")}}</div>
            <div class="team-sub">Lead: ${{e(selectedTeamRow.team_leader)}} | Members: ${{selectedTeamRow.active_members}}/${{selectedTeamRow.total_members}} | Eligible: ${{selectedTeamRow.eligible_members}}</div>
          </div>
          <div class="team-score">${{scoreText(selectedTeamRow, "advanced")}}</div>
        </div>
        <div class="team-metrics">
          <div>Total Penalty: ${{n(selectedTeamRow.total_penalty).toFixed(1)}}</div>
          <div>Total Rework: ${{n(selectedTeamRow.total_rework_hours).toFixed(1)}}h</div>
          <div>At Risk (&lt;60): ${{selectedTeamRow.at_risk}}</div>
          <div>Planned Leaves: ${{n(selectedTeamRow.planned_leave_count).toFixed(0)}} | ${{n(selectedTeamRow.planned_leave_hours).toFixed(2)}}h / ${{n(selectedTeamRow.planned_leave_days).toFixed(2)}}d</div>
          <div>Unplanned Leaves: ${{n(selectedTeamRow.unplanned_leave_count).toFixed(0)}} | ${{n(selectedTeamRow.unplanned_leave_hours).toFixed(2)}}h / ${{n(selectedTeamRow.unplanned_leave_days).toFixed(2)}}d</div>
        </div>
      </div>
    `;
  }} else if (teams.length) {{
    teamDetailHost.innerHTML = '<div class="empty">Select a team from chart.</div>';
  }}
  let viewItems = items.map((it) => ({{...it, capacity_gap_hours: n(it.employee_capacity_hours) - n(it.planned_hours_assigned)}}));
  const riskMode = String(document.getElementById("filter-risk")?.value || "all");
  const missedMode = String(document.getElementById("filter-missed")?.value || "all");
  const sortMode = String(document.getElementById("leader-sort")?.value || "score");
  const sortDirection = String(document.getElementById("leader-sort-direction")?.value || "desc").toLowerCase() === "asc" ? "asc" : "desc";
  const leaderSearchText = String(document.getElementById("leader-search")?.value || "").trim().toLowerCase();
  const leaderScoringMode = String(document.getElementById("leader-scoring-mode")?.value || "simple");
  activeScoringTab = leaderScoringMode;
  function lbScore(it) {{ return scoreNumber(it, leaderScoringMode); }}
  function compareLeaderboardRows(a, b, primaryValue, secondaryValue) {{
    const primaryDiff = n(primaryValue(a)) - n(primaryValue(b));
    if (primaryDiff !== 0) return sortDirection === "asc" ? primaryDiff : -primaryDiff;
    const secondaryDiff = n(secondaryValue(a)) - n(secondaryValue(b));
    if (secondaryDiff !== 0) return sortDirection === "asc" ? secondaryDiff : -secondaryDiff;
    return a.assignee.localeCompare(b.assignee);
  }}
  if (selectedTeamRow) {{
    const allowed = new Set((selectedTeamRow.members || []).map((m) => String(m || "").toLowerCase()));
    viewItems = viewItems.filter((it) => allowed.has(String(it.assignee || "").toLowerCase()));
  }}
  if (leaderSearchText) {{
    viewItems = viewItems.filter((it) => String(it.assignee || "").toLowerCase().includes(leaderSearchText));
  }}
  if (riskMode === "risk") viewItems = viewItems.filter((it) => Number.isFinite(lbScore(it)) && lbScore(it) < 60);
  if (missedMode === "missed") viewItems = viewItems.filter((it) => n(it.missed_start_count) > 0);
  if (sortMode === "rmis") viewItems.sort((a,b)=>compareLeaderboardRows(a, b, (item) => n(item.active_rmi_count), (item) => scoreSortValue(item, leaderScoringMode)));
  else if (sortMode === "missed") viewItems.sort((a,b)=>compareLeaderboardRows(a, b, (item) => n(item.missed_start_ratio), (item) => scoreSortValue(item, leaderScoringMode)));
  else if (sortMode === "capacity_gap") viewItems.sort((a,b)=>compareLeaderboardRows(a, b, (item) => n(item.capacity_gap_hours), (item) => scoreSortValue(item, leaderScoringMode)));
  else if (sortMode === "available_more_work") viewItems.sort((a,b)=>compareLeaderboardRows(a, b, (item) => leaderboardCapacityValue(item), (item) => scoreSortValue(item, leaderScoringMode)));
  else viewItems.sort((a,b)=>compareLeaderboardRows(a, b, (item) => scoreSortValue(item, leaderScoringMode), (item) => n(item.active_rmi_count)));
  lastLeaderboardViewItems = viewItems.slice();
  document.getElementById("leaderboard-title").textContent = selectedTeam ? `Leaderboard - ${{selectedTeam}}` : "Leaderboard";
  document.getElementById("leaderboard-filter").textContent = selectedTeam
    ? `Filtered by team "${{selectedTeam}}" (${{viewItems.length}} assignee${{viewItems.length === 1 ? "" : "s"}})`
    : "";
  const lb = document.getElementById("leaderboard");
  if (!viewItems.length) {{ lb.innerHTML = '<div class="empty" style="padding:10px;">No assignee activity for current filter.</div>'; document.getElementById("detail").innerHTML = '<div class="empty">No assignee activity for current filter.</div>'; document.getElementById("score-drilldown").innerHTML = '<div class="empty">No assignee activity for current filter.</div>'; return; }}
  lb.innerHTML = viewItems.map((it, i) => {{
    const capMore = leaderboardCapacityValue(it);
    const rowScore = lbScore(it);
    const rowScoreText = scoreText(it, leaderScoringMode);
    const ssPlanned = n(it.ss_total_estimate);
    const ssActual = (Array.isArray(it.ss_subtask_details) ? it.ss_subtask_details : []).reduce((acc, row) => acc + n(row?.actual), 0);
    const ssOverrun = n(it.simple_score_overrun_active);
    const simpleFormulaSub = isScoreEligible(it)
      ? `Simple: 100 x (1 - ${{simpleOverrunShortLabel()}}/Planned) | Planned: ${{ssPlanned.toFixed(1)}}h | Actual: ${{ssActual.toFixed(1)}}h | ${{simpleOverrunLabel()}}: ${{ssOverrun.toFixed(1)}}h`
      : `Simple: N/A | Planned Hours Assigned: ${{n(it.planned_hours_assigned).toFixed(1)}}h | Not eligible for scoring`;
    const advancedSub = `${{n(it.total_hours).toFixed(1)}}h logged | Missed: ${{n(it.missed_start_ratio).toFixed(1)}}% | Cap Gap: ${{n(it.capacity_gap_hours).toFixed(1)}}h`;
    const rowSub = leaderScoringMode === "simple" ? simpleFormulaSub : advancedSub;
    const refreshState = getEmployeeRefreshInlineState(it.assignee);
    const refreshHtml = refreshState ? `<div class="row-refresh${{refreshState.tone ? ` row-refresh-${{refreshState.tone}}` : ""}}"><div class="row-refresh-head"><span class="row-refresh-status">${{e(refreshState.statusText)}}</span><span class="row-refresh-pct">${{refreshState.progress}}%</span></div><div class="row-refresh-track" role="progressbar" aria-valuemin="0" aria-valuemax="100" aria-valuenow="${{refreshState.progress}}" aria-label="Refresh progress for ${{e(it.assignee)}}"><div class="row-refresh-fill" style="width:${{refreshState.progress}}%;"></div></div><div class="row-refresh-sub">${{e(refreshState.detailText)}}</div></div>` : "";
    const refreshDisabled = refreshState && refreshState.isBusy ? " disabled" : "";
    return `<div class="row${{it.assignee===selectedName?' sel':''}}" data-name="${{e(it.assignee)}}"><div class="rank">#${{i+1}}</div><div><div>${{e(it.assignee)}}</div><div class="leader-metrics"><button type="button" class="assignee-refresh-btn" data-assignee="${{e(it.assignee)}}" title="Refresh this assignee" aria-label="Refresh ${{e(it.assignee)}}"${{refreshDisabled}}><span class="material-symbols-outlined">refresh</span></button><span class="metric-chip"><span class="material-symbols-outlined">deployed_code</span><span class="metric-value">${{n(it.active_rmi_count).toFixed(0)}}</span></span><span class="metric-chip"><span class="material-symbols-outlined">sliders</span><span class="metric-value${{capMore < 0 ? " warn" : ""}}">${{capMore.toFixed(1)}}h</span></span><span class="metric-chip"><span class="material-symbols-outlined">award_star</span><span class="metric-value">${{rowScoreText}}</span></span></div><div class="sub">${{rowSub}}</div></div><div class="score"><span class="material-symbols-outlined">award_star</span>${{rowScoreText}}</div>${{refreshHtml}}</div>`;
  }}).join("");
  Array.from(lb.querySelectorAll(".row")).forEach((el)=>el.addEventListener("click", ()=>{{ selectedName = String(el.getAttribute("data-name") || ""); availabilityBreakdownForAssignee = ""; plannedHoursBreakdownForAssignee = ""; actualHoursBreakdownForAssignee = ""; rmiListForAssignee = ""; render(compute()); }}));
  Array.from(lb.querySelectorAll(".assignee-refresh-btn")).forEach((btn)=>btn.addEventListener("click", async (event)=>{{
    event.stopPropagation();
    const assignee = String(btn.getAttribute("data-assignee") || "");
    if (!assignee) return;
    if (typeof window !== "undefined" && typeof window.__startEmployeeRefreshRun === "function") {{
      await window.__startEmployeeRefreshRun(assignee, btn);
      return;
    }}
    setEmployeeRefreshStatus("Refresh monitor is unavailable on this page build.", "err");
  }}));
  let item = viewItems.find(x => x.assignee === selectedName); if (!item) {{ item = viewItems[0]; selectedName = item.assignee; }}
  const feed = (item.feed || []).map((v) => `<div class="i"><strong>${{e(v.label)}}</strong><br>${{n(v.hours).toFixed(2)}}h | <span class="neg">-${{n(v.points).toFixed(2)}}</span></div>`).join("") || '<div class="i empty">No violations.</div>';
  const hierarchyTable = renderHierarchyTable(item.assigned_hierarchy, item.due_compliance_items, item.missed_start_items);
  const executionHierarchyTable = renderExecutionHierarchyTable(item);
  const dueTable = renderDueTable(item.due_compliance_items);
  const missedTable = renderMissedTable(item.missed_start_items);
  const assignedHierarchyRows = Array.isArray(item.assigned_hierarchy) ? item.assigned_hierarchy : [];
  const assignmentCounts = assignedHierarchyRows.reduce((acc, row) => {{
    const baseType = String(row?.issue_type || "").toLowerCase();
    const hierarchyType = String(row?.hierarchy_type || "").toLowerCase();
    if (baseType === "epic") acc.epics += 1;
    else if (baseType === "story") acc.stories += 1;
    else if (hierarchyType === "bug_subtask") acc.bugs += 1;
    else if (baseType === "subtask") acc.subtasks += 1;
    return acc;
  }}, {{ epics: 0, stories: 0, subtasks: 0, bugs: 0 }});
  const assignmentScorecardsHtml = `<div class="assignment-scorecards"><div class="assignment-card epic"><div class="k">Epics</div><div class="v">${{assignmentCounts.epics}}</div></div><div class="assignment-card story"><div class="k">Stories</div><div class="v">${{assignmentCounts.stories}}</div></div><div class="assignment-card subtask"><div class="k">Subtasks</div><div class="v">${{assignmentCounts.subtasks}}</div></div><div class="assignment-card bug"><div class="k">Bugs</div><div class="v">${{assignmentCounts.bugs}}</div></div></div>`;
  const assignedMixChart = renderAssignedMixChart(item);
  const dueMixChart = renderDueComplianceChart(item.due_compliance_items);
  const managedCatalog = new Map((managedFields || []).map((mf) => [String(mf?.field_key || "").toLowerCase(), mf]));
  const managedEntries = Object.entries(item.managed_values || {{}})
    .map(([k, v]) => {{
      const meta = managedCatalog.get(String(k || "").toLowerCase()) || {{}};
      const label = String(meta?.label || "").trim() || toTitleCaseKey(k);
      const meaning = String(meta?.description || "").trim() || "Derived leadership metric from configured managed fields.";
      return {{ key: k, label, meaning, value: n(v) }};
    }})
    .sort((a, b) => a.label.localeCompare(b.label));
  const plannedAssignedEntry = {{
    key: "planned_hours_assigned_static",
    label: "Planned Hours Assigned",
    meaning: "Total original estimate hours for assigned non-bug subtasks within current filters (leave excluded).",
    value: n(item.planned_hours_assigned),
  }};
  const actualHoursSpentEntry = {{
    key: "actual_hours_spent_static",
    label: "Actual Hours Spent",
    meaning: extendedActualsEnabled
      ? "Total actual logged hours for selected subtasks (based on start OR due date in selected range), using full logged history."
      : "Total actual logged hours for assigned work items within current filters.",
    value: n(item.actual_hours_stats_total),
  }};
  const availabilityIndex = managedEntries.findIndex((entry) => {{
    const key = String(entry?.key || "").trim().toLowerCase();
    const label = String(entry?.label || "").trim().toLowerCase();
    return key === "availability" || label === "availability";
  }});
  if (availabilityIndex >= 0) managedEntries.splice(availabilityIndex + 1, 0, plannedAssignedEntry);
  else managedEntries.push(plannedAssignedEntry);
  const plannedAssignedIndex = managedEntries.findIndex((entry) => {{
    const key = String(entry?.key || "").trim().toLowerCase();
    const label = String(entry?.label || "").trim().toLowerCase();
    return key === "planned_hours_assigned_static" || label === "planned hours assigned";
  }});
  if (plannedAssignedIndex >= 0) managedEntries.splice(plannedAssignedIndex + 1, 0, actualHoursSpentEntry);
  else managedEntries.push(actualHoursSpentEntry);
  const managedMetricOrder = new Map([
    ["availability", 0],
    ["plannedhoursassignedstatic", 1],
    ["plannedhoursassigned", 1],
    ["actualhoursspentstatic", 2],
    ["actualhoursspent", 2],
    ["hoursrequiredtocompleteprojects", 3],
    ["capacityavailableformorework", 4],
  ]);
  managedEntries.sort((a, b) => {{
    const aRank = managedMetricOrder.get(normMetricToken(a?.key)) ?? managedMetricOrder.get(normMetricToken(a?.label)) ?? Number.MAX_SAFE_INTEGER;
    const bRank = managedMetricOrder.get(normMetricToken(b?.key)) ?? managedMetricOrder.get(normMetricToken(b?.label)) ?? Number.MAX_SAFE_INTEGER;
    if (aRank !== bRank) return aRank - bRank;
    return String(a?.label || "").localeCompare(String(b?.label || ""));
  }});
  function normMetricToken(text) {{
    return String(text || "").toLowerCase().replace(/[^a-z0-9]+/g, "");
  }}
  function extractFormulaTokens(expr) {{
    const src = String(expr || "");
    const found = src.match(/[A-Za-z_][A-Za-z0-9_]*/g) || [];
    const skip = new Set(["SUM", "COUNT", "MIN", "MAX", "AVG"]);
    const out = [];
    const seen = new Set();
    for (const token of found) {{
      const upper = String(token || "").toUpperCase();
      const norm = String(token || "").toLowerCase();
      if (!norm || skip.has(upper) || seen.has(norm)) continue;
      seen.add(norm);
      out.push(norm);
    }}
    return out;
  }}
  const availabilityEntry = managedEntries.find((entry) => {{
    const k = normMetricToken(entry?.key);
    const l = normMetricToken(entry?.label);
    return k === "availability" || l === "availability";
  }});
  const availabilityFormulaField = availabilityEntry
    ? (managedCatalog.get(String(availabilityEntry.key || "").toLowerCase())
      || Array.from(managedCatalog.values()).find((meta) => normMetricToken(meta?.label) === "availability")
      || null)
    : null;
  const availabilityFormula = String(availabilityFormulaField?.formula_expression || "").trim();
  const availabilityFormulaTokens = extractFormulaTokens(availabilityFormula);
  const availabilityScope = item?.managed_scope && typeof item.managed_scope === "object" ? item.managed_scope : {{}};
  const availabilityIngredients = availabilityFormulaTokens.map((token) => {{
    const key = String(token || "").toLowerCase();
    let source = "scope";
    let value = availabilityScope[key];
    if (!Number.isFinite(Number(value))) {{
      value = item?.entity_values?.[key];
      source = "entity";
    }}
    if (!Number.isFinite(Number(value))) {{
      value = item?.managed_values?.[key];
      source = "managed";
    }}
    const missing = !Number.isFinite(Number(value));
    return {{
      key,
      value: missing ? 0 : n(value),
      missing,
      source: missing ? "default" : source,
    }};
  }});
  const availabilityRawValue = n(availabilityEntry?.value);
  const otherMetricMax = Math.max(1, ...managedEntries
    .filter((entry) => {{
      const k = normMetricToken(entry?.key);
      const l = normMetricToken(entry?.label);
      return !(k === "availability" || l === "availability");
    }})
    .map((entry) => n(entry?.value)));
  function isCapacityAvailableForMoreWork(entry) {{
    const k = normMetricToken(entry?.key);
    const l = normMetricToken(entry?.label);
    return k === "capacityavailableformorework" || l === "capacityavailableformorework";
  }}
  function isPlannedHoursAssignedMetric(entry) {{
    const k = normMetricToken(entry?.key);
    const l = normMetricToken(entry?.label);
    return k === "plannedhoursassignedstatic" || l === "plannedhoursassigned";
  }}
  function isActualHoursSpentMetric(entry) {{
    const k = normMetricToken(entry?.key);
    const l = normMetricToken(entry?.label);
    return k === "actualhoursspentstatic" || l === "actualhoursspent";
  }}
  function barMeta(entry) {{
    const key = normMetricToken(entry?.key);
    const label = normMetricToken(entry?.label);
    const isAvailability = key === "availability" || label === "availability";
    const isPlannedAssigned = isPlannedHoursAssignedMetric(entry);
    const isActualSpent = isActualHoursSpentMetric(entry);
    const isCapacityAvail = isCapacityAvailableForMoreWork(entry);
    const maxValue = isAvailability
      ? 178
      : ((isPlannedAssigned || isActualSpent)
        ? 178
        : (isCapacityAvail
          ? Math.max(1, availabilityRawValue)
          : otherMetricMax));
    const rawValue = n(entry?.value);
    const clampedValue = Math.max(0, Math.min(rawValue, maxValue));
    const pct = maxValue > 0 ? (clampedValue / maxValue) * 100 : 0;
    const overflow = rawValue > maxValue;
    let fill = "linear-gradient(90deg,#38bdf8,#60a5fa)";
    if (overflow && (isPlannedAssigned || isActualSpent)) {{
      fill = "linear-gradient(90deg,#38bdf8 0%, #60a5fa 42%, #f97316 72%, #ef4444 100%)";
    }}
    return {{
      isAvailability,
      isPlannedAssigned,
      isActualSpent,
      maxValue,
      rawValue,
      pct: Math.max(0, Math.min(100, pct)),
      fill,
      note: isAvailability
        ? "Scale: 0 to 178h"
        : ((isPlannedAssigned || isActualSpent)
          ? (overflow ? `Scale: 0 to 178h (Exceeded by ${{(rawValue - 178).toFixed(2)}}h)` : "Scale: 0 to 178h")
          : `Scale: 0 to ${{maxValue.toFixed(2)}}`)
    }};
  }}
  const activeFrom = document.getElementById("from").value || defaultFrom;
  const activeTo = document.getElementById("to").value || defaultTo;
  const activeProfileForBreakdown = resolveActiveCapacityProfile(activeFrom, activeTo);
  const businessDaysInRange = computeBusinessDays(activeFrom, activeTo, activeProfileForBreakdown);
  function normIsoDay(value) {{
    const text = String(value || "").trim();
    if (!text) return "";
    const day = text.slice(0, 10);
    return /^\\d{{4}}-\\d{{2}}-\\d{{2}}$/.test(day) ? day : "";
  }}
  function formatScorecardDate(value) {{
    const iso = normIsoDay(value);
    if (!iso) return "-";
    const parts = iso.split("-");
    if (parts.length !== 3) return "-";
    const monthNames = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"];
    const monthIndex = Number(parts[1]) - 1;
    const monthName = monthNames[monthIndex];
    if (!monthName) return "-";
    return `${{parts[2]}} ${{monthName}} ${{parts[0]}}`;
  }}
  const activeRamadanStart = normIsoDay(activeProfileForBreakdown?.ramadan_start_date);
  const activeRamadanEnd = normIsoDay(activeProfileForBreakdown?.ramadan_end_date);
  function overlapsActiveRamadan(startIso, dueIso) {{
    if (!activeRamadanStart || !activeRamadanEnd) return false;
    const start = normIsoDay(startIso);
    const due = normIsoDay(dueIso);
    if (!start && !due) return false;
    if (start && due) {{
      const itemStart = start <= due ? start : due;
      const itemEnd = start <= due ? due : start;
      return itemStart <= activeRamadanEnd && itemEnd >= activeRamadanStart;
    }}
    const singleDate = start || due;
    return singleDate >= activeRamadanStart && singleDate <= activeRamadanEnd;
  }}
  const managedHtml = managedEntries.length
    ? `<div class="exec-metrics">${{managedEntries.map((m) => {{
      const b = barMeta(m);
      const action = b.isAvailability
        ? "toggle-availability-breakdown"
        : (b.isPlannedAssigned
          ? "toggle-planned-hours-breakdown"
          : (b.isActualSpent ? "toggle-actual-hours-breakdown" : ""));
      const toggleAttr = action ? ` data-action="${{action}}"` : "";
      const cardClass = action ? "exec-metric actionable" : "exec-metric";
      const isAvailabilityOpen = b.isAvailability && availabilityBreakdownForAssignee === String(item.assignee || "");
      const isPlannedOpen = b.isPlannedAssigned && plannedHoursBreakdownForAssignee === String(item.assignee || "");
      const isActualOpen = b.isActualSpent && actualHoursBreakdownForAssignee === String(item.assignee || "");
      const ingredientRows = availabilityIngredients.length
        ? availabilityIngredients.map((part) => `<div class="availability-line"><span class="availability-name">${{e(part.key)}}${{part.missing ? " (default)" : ""}}</span><span class="availability-num">${{n(part.value).toFixed(2)}}${{b.isAvailability ? "h" : ""}}</span></div>`).join("")
        : '<div class="availability-note">No ingredients detected from formula.</div>';
      const formulaBlock = b.isAvailability && isAvailabilityOpen
        ? `<div class="availability-breakdown"><div class="availability-line"><span class="availability-name">Formula</span><span class="availability-num">${{availabilityFormula ? e(availabilityFormula) : "Formula not configured"}}</span></div><div class="availability-line"><span class="availability-name">Business Days</span><span class="availability-num">${{n(businessDaysInRange).toFixed(0)}}d</span></div><div class="availability-note">Ingredients</div>${{ingredientRows}}<div class="availability-line"><span class="availability-name"><strong>Result</strong></span><span class="availability-num"><strong>${{b.rawValue.toFixed(2)}}h</strong></span></div></div>`
        : "";
      const plannedRows = (Array.isArray(item.assigned_hierarchy) ? item.assigned_hierarchy : [])
        .filter((row) => {{
          const t = String(row?.hierarchy_type || row?.issue_type || "").toLowerCase();
          return t === "subtask" || t === "bug_subtask";
        }});
      const plannedTotal = plannedRows.reduce((acc, row) => acc + n(row?.original_estimate_hours), 0);
      const plannedGroups = [
        {{
          key: "end_only",
          label: "Subtasks having only end date in the user-specified date range filter",
          rows: [],
        }},
        {{
          key: "both",
          label: "Subtasks having both start and end date in the user-specified date range filter",
          rows: [],
        }},
        {{
          key: "start_only",
          label: "Subtasks having only the start date in the user-specified date range filter",
          rows: [],
        }},
      ];
      const plannedFallbackRows = [];
      for (const row of plannedRows) {{
        const rawType = String(row?.hierarchy_type || row?.issue_type || "").toLowerCase();
        const issueKey = String(row?.issue_key || "").toUpperCase();
        const itemRef = workItemsByKey.get(issueKey) || {{}};
        const plannedStart = String(row?.start_date || itemRef?.start_date || "");
        const plannedDue = String(row?.due_date || itemRef?.due_date || "");
        const startInRange = !!(plannedStart && inRange(plannedStart, activeFrom, activeTo));
        const dueInRange = !!(plannedDue && inRange(plannedDue, activeFrom, activeTo));
        const enrichedRow = {{
          rawType,
          issueKey,
          plannedStart,
          plannedDue,
          originalEstimateHours: n(row?.original_estimate_hours),
          loggedHours: n(item.issue_logged_hours_by_issue?.[issueKey]),
        }};
        if (!startInRange && dueInRange) plannedGroups[0].rows.push(enrichedRow);
        else if (startInRange && dueInRange) plannedGroups[1].rows.push(enrichedRow);
        else if (startInRange && !dueInRange) plannedGroups[2].rows.push(enrichedRow);
        else plannedFallbackRows.push(enrichedRow);
      }}
      function plannedRowSort(a, b) {{
        const aAnchor = normIsoDay(a.plannedDue) || normIsoDay(a.plannedStart) || "9999-12-31";
        const bAnchor = normIsoDay(b.plannedDue) || normIsoDay(b.plannedStart) || "9999-12-31";
        return aAnchor.localeCompare(bAnchor) || a.issueKey.localeCompare(b.issueKey);
      }}
      function renderPlannedGroupRow(row) {{
        const ramadanChip = overlapsActiveRamadan(row.plannedStart, row.plannedDue)
          ? '<span class="ramadan-chip">Ramadan</span>'
          : "";
        const issueUrl = jiraIssueUrl(row.issueKey);
        const linkCell = issueUrl
          ? `<a class="jira-link-icon" href="${{e(issueUrl)}}" target="_blank" rel="noopener noreferrer" title="Open in Jira" onclick="event.stopPropagation();"><span class="material-symbols-outlined">open_in_new</span></a>`
          : `<span class="jira-link-disabled">-</span>`;
        return `<tr><td class="issue-id">${{e(row.issueKey || "-")}}</td><td>${{subtaskTypeIcon(row.rawType)}}</td><td>${{e(formatScorecardDate(row.plannedStart))}}</td><td>${{e(formatScorecardDate(row.plannedDue))}}</td><td>${{row.originalEstimateHours.toFixed(2)}}h${{ramadanChip}}</td><td>${{n(row.loggedHours).toFixed(2)}}h</td><td>${{linkCell}}</td></tr>`;
      }}
      const plannedTableBody = plannedRows.length
        ? plannedGroups.map((group) => {{
          const sortedRows = group.rows.slice().sort(plannedRowSort);
          const countLabel = sortedRows.length === 1 ? "1 subtask" : `${{sortedRows.length}} subtasks`;
          const groupRowsHtml = sortedRows.length
            ? sortedRows.map(renderPlannedGroupRow).join("")
            : '<tr><td colspan="7" class="empty">No subtasks in this group.</td></tr>';
          return `<tbody><tr class="ss-group-row"><td colspan="7"><strong>${{e(group.label)}}</strong><span style="margin-left:8px;opacity:.8;">${{e(countLabel)}}</span></td></tr>${{groupRowsHtml}}</tbody>`;
        }}).join("")
        + (plannedFallbackRows.length
          ? `<tbody><tr class="ss-group-row"><td colspan="7"><strong>Other subtasks in current scope</strong><span style="margin-left:8px;opacity:.8;">${{e(plannedFallbackRows.length === 1 ? "1 subtask" : `${{plannedFallbackRows.length}} subtasks`)}}</span></td></tr>${{plannedFallbackRows.slice().sort(plannedRowSort).map(renderPlannedGroupRow).join("")}}</tbody>`
          : "")
        : '<tbody><tr><td colspan="7" class="empty">No subtasks in current scope.</td></tr></tbody>';
      const plannedBreakdownBlock = b.isPlannedAssigned && isPlannedOpen
        ? `<div class="availability-breakdown"><div class="availability-note">Assigned subtasks in current filters (including bug subtasks), grouped by how their planned dates match the active date range.</div><div class="tbl-wrap" style="max-height:240px;overflow:auto;"><table class="ss-tbl"><thead><tr><th>Jira Subtask ID</th><th>Type</th><th>Start Date</th><th>End Date</th><th>Original Estimate</th><th>Logged Hours</th><th>Jira</th></tr></thead>${{plannedTableBody}}</table></div><div class="availability-line"><span class="availability-name"><strong>Total Original Estimates</strong></span><span class="availability-num"><strong>${{plannedTotal.toFixed(2)}}h</strong></span></div></div>`
        : "";
      const actualRows = Object.entries(item.issue_logged_hours_stats_by_issue || {{}})
        .map(([issueKeyRaw, loggedHoursRaw]) => {{
          const issueKey = String(issueKeyRaw || "").toUpperCase();
          const loggedHours = n(loggedHoursRaw);
          const wi = workItemsByKey.get(issueKey) || {{}};
          const rawType = String(wi.issue_type || wi.work_item_type || wi.jira_issue_type || "");
          const normalizedType = normalizeHierarchyType(rawType);
          return {{
            issue_key: issueKey,
            issue_type: normalizedType || "subtask",
            actual_hours: loggedHours,
          }};
        }})
        .filter((row) => n(row?.actual_hours) > 0)
        .sort((a, b) => String(a.issue_key || "").localeCompare(String(b.issue_key || "")));
      const actualTotal = actualRows.reduce((acc, row) => acc + n(row?.actual_hours), 0);
      const actualTableBody = actualRows.length
        ? actualRows.map((row) => {{
          const issueKey = String(row?.issue_key || "").toUpperCase();
          const issueUrl = jiraIssueUrl(issueKey);
          const linkCell = issueUrl
            ? `<a class="jira-link-icon" href="${{e(issueUrl)}}" target="_blank" rel="noopener noreferrer" title="Open in Jira" onclick="event.stopPropagation();"><span class="material-symbols-outlined">open_in_new</span></a>`
            : `<span class="jira-link-disabled">-</span>`;
          return `<tr><td class="issue-id">${{e(issueKey || "-")}}</td><td>${{subtaskTypeIcon(row?.issue_type || "subtask")}}</td><td>${{n(row?.actual_hours).toFixed(2)}}h</td><td>${{linkCell}}</td></tr>`;
        }}).join("")
        : '<tr><td colspan="4" class="empty">No logged subtasks in current scope.</td></tr>';
      const actualNote = extendedActualsEnabled
        ? "Selected subtasks (start OR due date in current range) with full logged history (including bug subtasks)."
        : "Subtasks with logged hours in current date/project filters (including bug subtasks).";
      const actualBreakdownBlock = b.isActualSpent && isActualOpen
        ? `<div class="availability-breakdown"><div class="availability-note">${{actualNote}}</div><div class="tbl-wrap" style="max-height:240px;overflow:auto;"><table class="ss-tbl"><thead><tr><th>Jira Subtask ID</th><th>Type</th><th>Actual Logged Hours</th><th>Jira</th></tr></thead><tbody>${{actualTableBody}}</tbody></table></div><div class="availability-line"><span class="availability-name"><strong>Total Actual Logged Hours</strong></span><span class="availability-num"><strong>${{actualTotal.toFixed(2)}}h</strong></span></div></div>`
        : "";
      const clickHint = b.isAvailability
        ? " | Click to view formula"
        : (b.isPlannedAssigned
          ? " | Click to view subtasks table"
          : (b.isActualSpent ? " | Click to view logged subtasks table" : ""));
      return `<div class="${{cardClass}}"${{toggleAttr}}><div class="exec-m-head"><div class="exec-m-name">${{e(m.label)}}</div><div class="exec-m-value">${{b.rawValue.toFixed(2)}}${{(b.isAvailability || b.isPlannedAssigned || b.isActualSpent) ? "h" : ""}}</div></div><div class="exec-m-meaning">${{e(m.meaning)}}</div><div class="exec-bar-track"><div class="exec-bar-fill" style="width:${{b.pct.toFixed(2)}}%;background:${{b.fill}};"></div></div><div class="exec-scale-note">${{e(b.note)}}${{clickHint}}</div>${{formulaBlock}}${{plannedBreakdownBlock}}${{actualBreakdownBlock}}</div>`;
    }}).join("")}}</div>`
    : '<div class="empty">No managed metrics configured.</div>';
  const activeProjectList = Array.from(selectedProjects()).sort();
  const activeProjectsText = activeProjectList.length ? activeProjectList.join(", ") : "All";
  const eligibleForScore = isScoreEligible(item);
  const healthTag = !eligibleForScore ? "Scoring N/A" : (n(item.final_score) < 60 ? "High Risk" : (n(item.missed_start_ratio) >= 30 ? "Start Discipline Risk" : "Stable"));
  const healthColor = !eligibleForScore ? "#94a3b8" : (n(item.final_score) < 60 ? "#f43f5e" : (n(item.missed_start_ratio) >= 30 ? "#f59e0b" : "#22c55e"));
  const summaryCapacityManaged = getManagedValue(item, ["capacity_available_for_more_work", "capacityavailableformorework", "capacity_available_more_work"]);
  const summaryCapacity = Number.isFinite(summaryCapacityManaged) ? summaryCapacityManaged : (n(item.employee_capacity_hours) - n(item.planned_hours_assigned));
  const summaryRmis = n(item.active_rmi_count);
  const summaryRmiKeys = Array.isArray(item.active_rmi_keys) ? item.active_rmi_keys : [];
  const isRmiListOpen = rmiListForAssignee === String(item.assignee || "");
  const rmiListRows = summaryRmiKeys.length
    ? summaryRmiKeys.map((epicKey) => {{
      const issueKey = String(epicKey || "").toUpperCase();
      const epicRow = workItemsByKey.get(issueKey) || {{}};
      const epicName = String(epicRow.summary || epicRow.item_summary || epicRow.epic_name || "").trim() || "-";
      const issueUrl = jiraIssueUrl(issueKey);
      const linkCell = issueUrl
        ? `<a class="jira-link-icon" href="${{e(issueUrl)}}" target="_blank" rel="noopener noreferrer" title="Open in Jira"><span class="material-symbols-outlined">open_in_new</span></a>`
        : `<span class="jira-link-disabled">-</span>`;
      return `<tr><td class="issue-id">${{e(issueKey || "-")}}</td><td class="issue-title">${{e(epicName)}}</td><td>${{linkCell}}</td></tr>`;
    }}).join("")
    : '<tr><td colspan="3" class="empty">No RMIs in selected date/project filters.</td></tr>';
  const rmiListHtml = isRmiListOpen
    ? `<div class="tbl-wrap rmi-list-wrap"><h3 class="tbl-title">RMIs In Current Scope</h3><table class="ss-tbl"><thead><tr><th>RMI (Epic Key)</th><th>RMI Name</th><th>Jira</th></tr></thead><tbody>${{rmiListRows}}</tbody></table></div>`
    : "";
  const summaryScore = n(item.final_score);
  const summaryScoreText = scoreText(item, "advanced");
  const baseSimpleScore = dueCompletionEnabled ? n(item.simple_score_due) : n(item.simple_score_raw);
  const overloadedApplied = n(item.simple_score_overloaded_applied) > 0;
  const activeSimpleScore = eligibleForScore
    ? n(item.simple_score)
    : NaN;
  const activeBigScore = activeScoringTab === "simple" ? activeSimpleScore : scoreNumber(item, "advanced");
  const activeBigScoreText = Number.isFinite(activeBigScore) ? activeBigScore.toFixed(1) : "N/A";
  const activeBigLabel = activeScoringTab === "simple" ? "Simple Score" : "Advanced Score";
  const activeBigSub = activeScoringTab === "simple"
    ? (eligibleForScore
      ? `Simple efficiency${{dueCompletionEnabled ? " (due-adjusted)" : ""}}${{overloadedApplied ? " + overloaded penalty" : ""}}${{planningRealismEnabled && overloadedApplied ? " + overload capping/planning realism cap" : ""}} | ${{simpleOverrunLabel()}}: ${{n(item.simple_score_overrun_active).toFixed(1)}}h | Planned: ${{n(item.ss_total_estimate).toFixed(1)}}h${{dueCompletionEnabled && n(item.ss_due_penalty_estimate) > 0 ? ` | Late estimate penalty: ${{n(item.ss_due_penalty_estimate).toFixed(1)}}h` : ""}}${{overloadedApplied ? ` | Overload score: ${{n(item.simple_score_overloaded).toFixed(1)}}% | Overload penalty: ${{n(item.simple_score_overloaded_penalty_pct).toFixed(1)}}%` : ""}}`
      : `Simple scoring is N/A because Planned Hours Assigned is ${{n(item.planned_hours_assigned).toFixed(1)}}h.`)
    : (eligibleForScore
      ? `Penalty-based | Raw ${{n(item.raw_score).toFixed(2)}} | Penalty -${{n(item.total_penalty).toFixed(2)}} | Base ${{n(settings.base_score).toFixed(0)}}`
      : `Advanced scoring is N/A because Planned Hours Assigned is ${{n(item.planned_hours_assigned).toFixed(1)}}h.`);
  const summaryMetricsHtml = `<div class="kpis" style="margin-top:8px;"><div class="kpi actionable" data-action="toggle-rmis-list"><div class="k"><span class="material-symbols-outlined" style="font-size:15px;vertical-align:middle;margin-right:4px;">deployed_code</span>RMIs</div><div class="v">${{summaryRmis.toFixed(0)}}</div><div class="kpi-note">${{isRmiListOpen ? "Click to hide RMIs list" : "Click to view RMIs list"}}</div></div><div class="kpi"><div class="k"><span class="material-symbols-outlined" style="font-size:15px;vertical-align:middle;margin-right:4px;">sliders</span>Capacity</div><div class="v">${{summaryCapacity.toFixed(1)}}h</div></div><div class="kpi"><div class="k"><span class="material-symbols-outlined" style="font-size:15px;vertical-align:middle;margin-right:4px;">award_star</span>Score</div><div class="v" id="summary-score-kpi">${{activeBigScoreText}}</div></div></div>`;
  const summaryScoreLabelHtml = activeScoringTab === "simple"
    ? `<button type="button" class="summary-score-trigger" id="summary-simple-score-trigger" aria-label="Open simple score details"><span class="score-label-text">${{e(activeBigLabel)}}</span><span class="material-symbols-outlined" aria-hidden="true">open_in_new</span></button>`
    : e(activeBigLabel);
  const summaryHtml = `<div class="card"><div style="display:flex;justify-content:space-between;align-items:end;"><div><div class="sub">Assignee</div><div style="font-size:1.1rem;font-weight:800;">${{e(item.assignee)}}</div></div><div><div class="big" id="summary-big-score">${{activeBigScoreText}}</div><div class="score-label" id="summary-score-label">${{summaryScoreLabelHtml}}</div></div></div>${{summaryMetricsHtml}}${{rmiListHtml}}<div class="sub" id="summary-score-sub">${{activeBigSub}}</div><div class="discover"><span class="pill" style="border-color:${{healthColor}};">Health: ${{healthTag}}</span><span class="pill">Capacity Gap: ${{(n(item.employee_capacity_hours)-n(item.planned_hours_assigned)).toFixed(1)}}h</span><span class="pill">Missed Starts: ${{n(item.missed_start_ratio).toFixed(1)}}%</span><span class="pill">Missed Due Dates: ${{n(item.missed_due_date_ratio).toFixed(1)}}%</span></div></div>`;
  const managedSectionHtml = `<div class="mini" style="margin:8px 0;"><h3>Managed Field Metrics - ${{e(item.assignee)}}</h3><div class="sub">Employee-only metrics within filters | Date: ${{e(activeFrom)}} to ${{e(activeTo)}} | Projects: ${{e(activeProjectsText)}}</div>${{managedHtml}}</div>`;
  const ssDetails = item.ss_subtask_details || [];
  function ssRowClass(row) {{
    const est = String(row.estimate_status || "");
    const due = String(row.due_completion_status || "");
    if (est === "no_estimate") return "";
    if (n(row.is_commitment)) return "ss-row-commitment";
    if (est === "within_estimate") return "ss-row-within";
    if (est === "over_estimate" && due === "late") return "ss-row-over-late";
    if (est === "over_estimate") return "ss-row-over";
    return "";
  }}
  function ssPillHtml(row) {{
    const est = String(row.estimate_status || "");
    if (est === "no_estimate") return `<span class="ss-status-pill ss-pill-noest">No Estimate</span>`;
    if (n(row.is_commitment)) return `<span class="ss-status-pill ss-pill-commitment">Commitment</span>`;
    if (est === "within_estimate") return `<span class="ss-status-pill ss-pill-within">Within Estimate</span>`;
    if (est === "over_estimate" && String(row.due_completion_status || "") === "late") return `<span class="ss-status-pill ss-pill-late">Over + Late</span>`;
    if (est === "over_estimate") return `<span class="ss-status-pill ss-pill-over">Over Estimate</span>`;
    return "";
  }}
  const ssTableRows = ssDetails.map((row) => {{
    const cls = `${{ssRowClass(row)}}${{row.is_penalized_for_due ? " penalized-row" : ""}}`.trim();
    return `<tr class="${{cls}}"><td class="issue-id">${{e(row.issue_key)}}</td><td class="issue-title">${{e(row.summary)}}</td><td>${{n(row.estimate).toFixed(1)}}h</td><td>${{n(row.actual).toFixed(1)}}h</td><td>${{n(row.overrun) > 0 ? n(row.overrun).toFixed(1) + "h" : "-"}}</td><td>${{row.planned_due_date ? e(row.planned_due_date) : (row.due_date ? e(row.due_date) : "-")}}</td><td>${{row.last_logged_date ? e(row.last_logged_date) : "-"}}</td><td>${{row.actual_complete_date ? e(row.actual_complete_date) : (row.effective_completion_date ? e(row.effective_completion_date) : "-")}}<div class="sub">${{e(actualCompletionSourceText(row.actual_complete_source))}}</div></td><td>${{ssPillHtml(row)}}<div class="sub">${{e(actualCompletionReason(row))}}</div></td></tr>`;
  }}).join("");
  const ssTableHtml = ssDetails.length ? `<div class="table-title" style="margin:10px 0 6px;font-weight:700;">Assignee Task Estimate vs Actual (Overrun & Status)</div><div class="tbl-wrap" style="max-height:320px;overflow:auto;"><table class="ss-tbl"><thead><tr><th>Key</th><th>Summary</th><th>Estimate</th><th>Actual</th><th>Overrun</th><th>Planned Due Date</th><th>Last Logged Date</th><th>Actual Complete Date</th><th>Status</th></tr></thead><tbody>${{ssTableRows}}</tbody></table></div>` : '<div class="empty">No subtask data for simple scoring.</div>';
  const ssWithin = n(item.ss_within_count);
  const ssOver = n(item.ss_over_count);
  const ssCommit = n(item.ss_commitment_count);
  const ssLatePenaltyCount = n(item.ss_due_penalty_count);
  const ssLatePenaltyEstimate = n(item.ss_due_penalty_estimate);
  const ssTotal = ssWithin + ssOver;
  const ssPlannedHours = n(item.ss_total_estimate);
  const ssActualHours = n(item.ss_total_actual);
  const ssOverrunHours = n(item.simple_score_overrun_active);
  const ssAdjustedOverrunHours = Math.max(0, ssOverrunHours - n(item.ss_commitment_overrun));
  const ssAppliedOverrunHours = dueCompletionEnabled ? (ssAdjustedOverrunHours + ssLatePenaltyEstimate) : ssOverrunHours;
  const ssFormulaText = dueCompletionEnabled
    ? `Score % = max(0, min(100, 100 x (1 - (${{simpleOverrunLabel()}} - Commitment Forgiven + Late Completion Estimate Penalty) / Planned Hours)))`
    : `Score % = max(0, min(100, 100 x (1 - ${{simpleOverrunLabel()}} / Planned Hours)))`;
  const ssOverloadedFormulaText = "Overloaded Penalty: if Planned > Capacity x (1 + N/100), overload score = Capacity/Planned x 100 and overload penalty = 100 - overload score";
  const ssPlanningRealismFormulaText = "Overload Capping/ Planning Realism: if ON, Final Simple Score = overload score; if OFF, Final Simple Score = Base Simple Score - overload penalty";
  const ssOverrunFormulaText = simpleOverrunMode === "total"
    ? "Total Overrun Hours = max(0, Total Actual Hours - Planned Hours)"
    : "Overrun Subtask Hours = sum(max(0, Subtask Actual - Subtask Planned))";
  const ssFormulaAppliedText = dueCompletionEnabled
    ? `100 x (1 - (${{ssOverrunHours.toFixed(1)}} - ${{(simpleOverrunMode === "total" ? 0 : n(item.ss_commitment_overrun)).toFixed(1)}} + ${{ssLatePenaltyEstimate.toFixed(1)}}) / ${{ssPlannedHours.toFixed(1)}})`
    : `100 x (1 - ${{ssOverrunHours.toFixed(1)}} / ${{ssPlannedHours.toFixed(1)}})`;
  const ssFormulaIngredientsHtml = `<section class="formula-guide"><div class="formula-head"><h3 class="formula-title">Simple Scoring Formula Guide</h3><div class="formula-head-actions"><button type="button" class="formula-toggle-btn" id="toggle-formula-guide" aria-controls="simple-formula-guide-body" aria-expanded="${{simpleFormulaGuideExpanded ? "true" : "false"}}"><span class="material-symbols-outlined">${{simpleFormulaGuideExpanded ? "expand_less" : "expand_more"}}</span><span>${{simpleFormulaGuideExpanded ? "Collapse" : "Expand"}}</span></button><div class="formula-score-pill">${{Number.isFinite(activeSimpleScore) ? activeSimpleScore.toFixed(1) + "%" : "N/A"}}</div></div></div><div class="formula-layout" id="simple-formula-guide-body" ${{simpleFormulaGuideExpanded ? "" : "hidden"}}><div class="formula-steps"><div class="formula-step step-gap"><div class="formula-kicker">Step 1: ${{simpleOverrunShortLabel()}}</div><p class="formula-eq">${{e(ssOverrunFormulaText)}}</p><div class="formula-mini-help">max(a, b) means: choose the bigger number.</div></div><div class="formula-step"><div class="formula-kicker">Step 2: Base Simple Score</div><p class="formula-eq">${{e(ssFormulaText)}}</p><p class="formula-applied">${{ssPlannedHours > 0 ? `Applied: ${{e(ssFormulaAppliedText)}}` : "Applied: Planned Hours is 0, so the employee is not eligible for scoring."}}</p></div><div class="formula-step"><div class="formula-kicker">Step 3: Overloaded Penalty (Optional)</div><p class="formula-eq">${{e(ssOverloadedFormulaText)}}</p><p class="formula-applied">${{overloadedPenaltyEnabled ? `Threshold N: ${{overloadedPenaltyThresholdPct.toFixed(1)}}% | Capacity: ${{n(item.employee_capacity_hours).toFixed(1)}}h | Max Planned Before Overload: ${{(n(item.employee_capacity_hours) * (1 + overloadedPenaltyThresholdPct / 100)).toFixed(1)}}h | Overload score: ${{n(item.simple_score_overloaded).toFixed(1)}}% | Overload penalty: ${{n(item.simple_score_overloaded_penalty_pct).toFixed(1)}}%${{overloadedApplied ? " (Applied)" : " (Not applied)"}}` : "Overloaded penalty is turned off."}}</p></div><div class="formula-step"><div class="formula-kicker">Step 4: Overload Capping/ Planning Realism (Optional)</div><p class="formula-eq">${{e(ssPlanningRealismFormulaText)}}</p><p class="formula-applied">${{planningRealismEnabled ? (overloadedApplied ? `Overload Capping/ Planning Realism is ON, so final simple score is capped to overload score ${{n(item.simple_score_overloaded).toFixed(1)}}%.` : "Overload Capping/ Planning Realism is ON, but overload threshold was not hit.") : (overloadedApplied ? `Overload Capping/ Planning Realism is OFF, so overload penalty of ${{n(item.simple_score_overloaded_penalty_pct).toFixed(1)}}% is deducted from base simple score.` : "Overload Capping/ Planning Realism is OFF.")}}</p></div><div class="formula-note">${{simpleOverrunMode === "total" ? "Total overrun mode uses only max(0, Total Actual Hours - Planned Hours). Subtask overruns are ignored if total actual hours stay within plan." : "Overrun subtask mode sums only positive subtask overruns. A good total actual can still carry penalty if individual subtasks overran."}}</div>${{dueCompletionEnabled ? `<div class="formula-note"><strong>Commitment Forgiven</strong> = ${{simpleOverrunMode === "total" ? "not used in total overrun mode." : "overrun hours from items finished on time are not counted in penalty."}}</div><div class="formula-note"><strong>Late Completion Estimate Penalty</strong> = each late-completed subtask adds its original estimate to the due-completion penalty pool.</div>` : ""}}<div class="formula-safeguards"><div class="formula-safeguards-title">Why we use max (simple examples)</div><div class="formula-safeguard-item">1. <strong>Stop negative overrun:</strong> <code>max(0, Actual - Planned)</code><br>Example: Planned 5h, Actual 3h -> Actual - Planned = -2h -> Overrun = max(0, -2) = 0h.</div><div class="formula-safeguard-item">2. <strong>Keep score between 0% and 100%:</strong> <code>max(0, min(100, ...))</code><br>Example: If math gives 108%, final shown score is 100%. If math gives -12%, final shown score is 0%.</div></div></div><div class="formula-metrics"><div class="formula-row"><span>Planned Hours</span><span>${{ssPlannedHours.toFixed(1)}}h</span></div><div class="formula-row"><span>Actual Hours Spent</span><span>${{ssActualHours.toFixed(1)}}h</span></div><div class="formula-row"><span>Employee Capacity</span><span>${{n(item.employee_capacity_hours).toFixed(1)}}h</span></div><div class="formula-row"><span>${{simpleOverrunLabel()}}</span><span>${{ssOverrunHours.toFixed(1)}}h</span></div>${{dueCompletionEnabled ? `<div class="formula-row"><span>Commitment Forgiven</span><span>${{(simpleOverrunMode === "total" ? 0 : n(item.ss_commitment_overrun)).toFixed(1)}}h</span></div><div class="formula-row"><span>Late Completion Estimate Penalty</span><span>${{ssLatePenaltyEstimate.toFixed(1)}}h across ${{ssLatePenaltyCount.toFixed(0)}} late item(s)</span></div><div class="formula-row"><span>Applied Penalty Hours</span><span>${{ssAppliedOverrunHours.toFixed(1)}}h</span></div>` : ""}}<div class="formula-row"><span>Overloaded Penalty</span><span>${{overloadedPenaltyEnabled ? (overloadedApplied ? "Applied" : "On") : "Off"}}</span></div><div class="formula-row"><span>Overload Capping/ Planning Realism</span><span>${{planningRealismEnabled ? (overloadedApplied ? "Cap Applied" : "On") : "Off"}}</span></div><div class="formula-row final"><span>Final Simple Score</span><span>${{Number.isFinite(activeSimpleScore) ? activeSimpleScore.toFixed(1) + "%" : "N/A"}}</span></div></div></div></section>`;
  function renderSsDonut(within, over, commit, total) {{
    if (total <= 0) return '<div class="empty">No data.</div>';
    const r = 50, cx = 60, cy = 60, sw = 18;
    const overNonCommit = Math.max(0, over - commit);
    const slices = [
      {{val:within, color:"#22c55e", label:"Within Estimate"}},
      {{val:commit, color:"#818cf8", label:"Commitment"}},
      {{val:overNonCommit, color:"#f97316", label:"Over Estimate"}},
    ].filter((s) => s.val > 0);
    let offset = 0;
    const circ = 2 * Math.PI * r;
    const paths = slices.map((s) => {{
      const pct = s.val / total;
      const dash = circ * pct;
      const gap = circ - dash;
      const o = offset;
      offset += pct;
      return `<circle cx="${{cx}}" cy="${{cy}}" r="${{r}}" fill="none" stroke="${{s.color}}" stroke-width="${{sw}}" stroke-dasharray="${{dash.toFixed(2)}} ${{gap.toFixed(2)}}" stroke-dashoffset="${{(-circ * o).toFixed(2)}}" transform="rotate(-90 ${{cx}} ${{cy}})"></circle>`;
    }}).join("");
    const legendItems = slices.map((s) => `<div><span class="ss-legend-dot" style="background:${{s.color}};"></span>${{s.label}}: ${{s.val}}</div>`).join("");
    return `<div class="ss-donut-wrap"><svg width="120" height="120" viewBox="0 0 120 120">${{paths}}<text x="${{cx}}" y="${{cy + 4}}" text-anchor="middle" fill="#e6f0ff" font-size="16" font-weight="900">${{total}}</text></svg><div class="ss-legend">${{legendItems}}</div></div>`;
  }}
  const ssDonutHtml = renderSsDonut(ssWithin, ssOver, ssCommit, ssTotal);
  const simpleScoringContent = `<div class="scoring-section"><div class="scoring-section-head"><div class="scoring-section-title">Simple Scoring</div><div class="ss-toggle"><span class="ss-toggle-label">Due Completion</span><label class="ss-switch"><input type="checkbox" id="due-completion-toggle" ${{dueCompletionEnabled ? "checked" : ""}}><span class="ss-slider"></span></label></div></div><div class="ss-big-score">${{Number.isFinite(activeSimpleScore) ? activeSimpleScore.toFixed(1) : "N/A"}}</div><div class="sub">${{eligibleForScore ? `Simple efficiency score${{dueCompletionEnabled ? " (due-adjusted)" : ""}}${{overloadedApplied ? " + overloaded penalty" : ""}}${{planningRealismEnabled && overloadedApplied ? " + overload capping/planning realism cap" : ""}} | Planned: ${{ssPlannedHours.toFixed(1)}}h | Actual: ${{ssActualHours.toFixed(1)}}h | ${{simpleOverrunLabel()}}: ${{ssOverrunHours.toFixed(1)}}h${{dueCompletionEnabled && ssCommit > 0 && simpleOverrunMode !== "total" ? ` | Commitment forgiven: ${{n(item.ss_commitment_overrun).toFixed(1)}}h` : ""}}${{dueCompletionEnabled && ssLatePenaltyEstimate > 0 ? ` | Late estimate penalty: ${{ssLatePenaltyEstimate.toFixed(1)}}h` : ""}}${{overloadedApplied ? ` | Overload score: ${{n(item.simple_score_overloaded).toFixed(1)}}% | Overload penalty: ${{n(item.simple_score_overloaded_penalty_pct).toFixed(1)}}%` : ""}}` : `Simple scoring is not available because Planned Hours Assigned is ${{n(item.planned_hours_assigned).toFixed(1)}}h.`}}</div><div class="ss-summary-row"><span class="ss-chip">Within: ${{ssWithin}}</span><span class="ss-chip">Over: ${{ssOver}}</span>${{dueCompletionEnabled ? `<span class="ss-chip">Commitment: ${{simpleOverrunMode === "total" ? 0 : ssCommit}}</span><span class="ss-chip">On Time: ${{n(item.ss_on_time_count)}}</span><span class="ss-chip">Late: ${{n(item.ss_late_count)}}</span><span class="ss-chip">Late Estimate Penalty: ${{ssLatePenaltyEstimate.toFixed(1)}}h</span>` : ""}}<span class="ss-chip">No Estimate: ${{n(item.ss_no_estimate_count)}}</span><span class="ss-chip">Overrun Basis: ${{simpleOverrunMode === "total" ? "Total" : "Subtasks"}}</span><span class="ss-chip">Overloaded Penalty: ${{overloadedPenaltyEnabled ? (overloadedApplied ? "Applied" : "On") : "Off"}}</span><span class="ss-chip">Overload Capping/ Planning Realism: ${{planningRealismEnabled ? (overloadedApplied ? "Cap Applied" : "On") : "Off"}}</span></div>${{ssFormulaIngredientsHtml}}${{ssDonutHtml}}${{ssTableHtml}}</div>`;
  const advPenaltyFormulaText = "Total Penalty = (Bug Hours x Points/Bug Hour) + (Bug Late Hours x Points/Bug Late Hour) + (Unplanned Leave Hours x Points/Unplanned Leave Hour) + (Subtask Late Hours x Points/Subtask Late Hour) + (Estimate Overrun Hours x Points/Estimate Overrun Hour) + (Missed Due Dates x Points/Missed Due Date)";
  const advRawFormulaText = "Raw Score = Base Score - Total Penalty";
  const advFinalFormulaText = "Final Score = clamp(Raw Score, Min Score, Max Score)";
  const advPenaltyAppliedText = `(${{n(item.bug_hours).toFixed(2)}} x ${{n(settings.points_per_bug_hour).toFixed(2)}}) + (${{n(item.bug_late_hours).toFixed(2)}} x ${{n(settings.points_per_bug_late_hour).toFixed(2)}}) + (${{n(item.unplanned_leave_hours).toFixed(2)}} x ${{n(settings.points_per_unplanned_leave_hour).toFixed(2)}}) + (${{n(item.subtask_late_hours).toFixed(2)}} x ${{n(settings.points_per_subtask_late_hour).toFixed(2)}}) + (${{n(item.estimate_overrun_hours).toFixed(2)}} x ${{n(settings.points_per_estimate_overrun_hour).toFixed(2)}}) + (${{n(item.missed_due_date_count).toFixed(0)}} x ${{n(settings.points_per_missed_due_date).toFixed(2)}})`;
  const advRawAppliedText = `${{n(settings.base_score).toFixed(2)}} - ${{n(item.total_penalty).toFixed(2)}}`;
  const advFinalAppliedText = `clamp(${{n(item.raw_score).toFixed(2)}}, ${{n(settings.min_score).toFixed(2)}}, ${{n(settings.max_score).toFixed(2)}})`;
  const advancedFormulaIngredientsHtml = `<section class="formula-guide"><div class="formula-head"><h3 class="formula-title">Advanced Scoring Formula Guide</h3><div class="formula-head-actions"><button type="button" class="formula-toggle-btn" id="toggle-advanced-formula-guide" aria-controls="advanced-formula-guide-body" aria-expanded="${{advancedFormulaGuideExpanded ? "true" : "false"}}"><span class="material-symbols-outlined">${{advancedFormulaGuideExpanded ? "expand_less" : "expand_more"}}</span><span>${{advancedFormulaGuideExpanded ? "Collapse" : "Expand"}}</span></button><div class="formula-score-pill">${{eligibleForScore ? `${{summaryScore.toFixed(1)}}%` : "N/A"}}</div></div></div><div class="formula-layout" id="advanced-formula-guide-body" ${{advancedFormulaGuideExpanded ? "" : "hidden"}}><div class="formula-steps"><div class="formula-step step-gap"><div class="formula-kicker">Step 1: Total Penalty</div><p class="formula-eq">${{e(advPenaltyFormulaText)}}</p><p class="formula-applied">Applied: ${{e(advPenaltyAppliedText)}}</p></div><div class="formula-step"><div class="formula-kicker">Step 2: Raw Score</div><p class="formula-eq">${{e(advRawFormulaText)}}</p><p class="formula-applied">Applied: ${{e(advRawAppliedText)}} = ${{n(item.raw_score).toFixed(2)}}</p></div><div class="formula-step"><div class="formula-kicker">Step 3: Final Score</div><p class="formula-eq">${{e(advFinalFormulaText)}}</p><p class="formula-applied">${{eligibleForScore ? `Applied: ${{e(advFinalAppliedText)}} = ${{summaryScore.toFixed(2)}}` : `Applied: Planned Hours Assigned is ${{n(item.planned_hours_assigned).toFixed(1)}}h, so advanced scoring is N/A.`}}</p></div><div class="formula-note">Penalty multipliers come from Performance Settings. Higher multipliers increase score deductions for the same workload pattern.</div><div class="formula-safeguards"><div class="formula-safeguards-title">Safeguards</div><div class="formula-safeguard-item">1. <strong>Penalty-first model:</strong> only <code>base_score</code> adds points; all configured factors reduce score.</div><div class="formula-safeguard-item">2. <strong>Bounds:</strong> <code>clamp(raw, min, max)</code> keeps final score within allowed range.</div></div></div><div class="formula-metrics"><div class="formula-row"><span>Base Score</span><span>${{n(settings.base_score).toFixed(2)}}</span></div><div class="formula-row"><span>Total Penalty</span><span>${{n(item.total_penalty).toFixed(2)}}</span></div><div class="formula-row"><span>Raw Score</span><span>${{n(item.raw_score).toFixed(2)}}</span></div><div class="formula-row"><span>Score Bounds</span><span>${{n(settings.min_score).toFixed(0)}} to ${{n(settings.max_score).toFixed(0)}}</span></div><div class="formula-row final"><span>Final Advanced Score</span><span>${{eligibleForScore ? `${{summaryScore.toFixed(1)}}%` : "N/A"}}</span></div></div></div></section>`;
  const advancedScoringContent = `<div class="scoring-section"><div class="scoring-section-head"><div class="scoring-section-title">Advanced Scoring <span class="beta-tag">beta</span></div></div><div class="ss-big-score">${{summaryScoreText}}</div><div class="sub">${{eligibleForScore ? `Penalty-based | Raw ${{n(item.raw_score).toFixed(2)}} | Total Penalty -${{n(item.total_penalty).toFixed(2)}} | Base ${{n(settings.base_score).toFixed(0)}}` : `Advanced scoring is not available because Planned Hours Assigned is ${{n(item.planned_hours_assigned).toFixed(1)}}h.`}}</div>${{advancedFormulaIngredientsHtml}}<div class="grid2" style="margin-top:8px;"><section class="mini"><h3>Score Breakdown</h3><div class="l"><span>Bug Hours</span><span class="neg">-${{n(item.penalties.bug).toFixed(2)}}</span></div><div class="l"><span>Bug Late Hours</span><span class="neg">-${{n(item.penalties.bug_late).toFixed(2)}}</span></div><div class="l"><span>Unplanned Leaves</span><span class="neg">-${{n(item.penalties.leave).toFixed(2)}}</span></div><div class="l"><span>Subtask Late Hours</span><span class="neg">-${{n(item.penalties.subtask_late).toFixed(2)}}</span></div><div class="l"><span>Missed Due Dates</span><span class="neg">-${{n(item.penalties.missed_due_date).toFixed(2)}}</span></div><div class="l"><span>Estimate Overrun</span><span class="neg">-${{n(item.penalties.estimate).toFixed(2)}}</span></div></section><section class="mini"><h3>Planning Scorecards</h3><div class="l"><span>Employee Capacity</span><span>${{n(item.employee_capacity_hours).toFixed(2)}}h</span></div><div class="l"><span>Planned Assigned</span><span>${{n(item.planned_hours_assigned).toFixed(2)}}h</span></div><div class="l"><span>Assigned (E/S/ST)</span><span>${{n(item.assigned_counts.epic).toFixed(0)}}/${{n(item.assigned_counts.story).toFixed(0)}}/${{n(item.assigned_counts.subtask).toFixed(0)}}</span></div><div class="l actionable" data-action="open-missed-starts"><span>Missed Starts</span><span><button type="button" class="metric-link-btn">View subtasks</button> ${{n(item.missed_start_count).toFixed(0)}} / ${{n(item.total_assigned_count).toFixed(0)}} (${{n(item.missed_start_ratio).toFixed(1)}}%)</span></div><div class="l actionable" data-action="open-missed-due"><span>Missed Due Dates</span><span><button type="button" class="metric-link-btn">View subtasks</button> ${{n(item.missed_due_date_count).toFixed(0)}} / ${{n(item.due_dated_assigned_count).toFixed(0)}} (${{n(item.missed_due_date_ratio).toFixed(1)}}%)</span></div><div class="l"><span>Planned Leaves</span><span>${{n(item.planned_leave_count).toFixed(0)}} | ${{n(item.planned_leave_hours).toFixed(2)}}h / ${{n(item.planned_leave_days).toFixed(2)}}d</span></div><div class="l"><span>Unplanned Leaves</span><span>${{n(item.unplanned_leave_count).toFixed(0)}} | ${{n(item.unplanned_leave_hours).toFixed(2)}}h / ${{n(item.unplanned_leave_days).toFixed(2)}}d</span></div></section></div><div class="kpi-charts"><div class="mini-chart"><h3 style="margin:0 0 6px;font-size:.82rem;">Assigned Mix Chart</h3>${{assignedMixChart}}</div><div class="mini-chart"><h3 style="margin:0 0 6px;font-size:.82rem;">Due Compliance Chart</h3>${{dueMixChart}}</div></div><div class="ts-card"><h3 style="margin:0 0 6px;font-size:.82rem;">Performance Over Days</h3>${{renderSeriesSvg(item.daily_series)}}</div></div>`;
  const scoringTabsHtml = `<div class="tabs" style="margin-top:10px;" data-tab-group="scoring"><button class="tab-btn tab-btn-score${{activeScoringTab === "simple" ? " active" : ""}}" data-tab="simple"><span class="tab-kicker">Mode</span><span class="tab-main-row"><span class="tab-title">Simple Scoring</span><span class="tab-score">${{Number.isFinite(activeSimpleScore) ? activeSimpleScore.toFixed(1) : "N/A"}}</span></span></button><button class="tab-btn tab-btn-score tab-btn-beta${{activeScoringTab === "advanced" ? " active" : ""}}" data-tab="advanced"><span class="tab-kicker">Mode</span><span class="tab-main-row"><span class="tab-title">Advanced Scoring <span class="tab-beta">beta</span></span><span class="tab-score">${{scoreText(item, "advanced")}}</span></span></button></div><div class="tab-pane${{activeScoringTab === "simple" ? " active" : ""}}" data-pane="simple" data-tab-group="scoring">${{simpleScoringContent}}</div><div class="tab-pane${{activeScoringTab === "advanced" ? " active" : ""}}" data-pane="advanced" data-tab-group="scoring">${{advancedScoringContent}}</div>`;
  const execPlanTabsHtml = `${{assignmentScorecardsHtml}}<div class="tabs" style="margin-top:10px;" data-tab-group="detail"><button class="tab-btn active" data-tab="planning">Planning</button><button class="tab-btn" data-tab="execution">Execution</button></div><div class="tab-pane active" data-pane="planning" data-tab-group="detail"><div class="tbl-wrap"><h3 class="tbl-title">Interactive Hierarchy Breakdown (Epic -> Story -> Subtask)</h3>${{hierarchyTable}}</div></div><div class="tab-pane" data-pane="execution" data-tab-group="detail"><div class="tbl-wrap"><h3 class="tbl-title">Execution Nested View (Epic -> Story -> Subtask) | Planned vs Actual</h3>${{executionHierarchyTable}}</div><div class="tbl-wrap" id="due-compliance-context"><h3 class="tbl-title">Due Compliance Table (Logged Items)</h3>${{dueTable}}</div><div class="tbl-wrap" id="missed-start-context"><h3 class="tbl-title">Missed Start Context Table</h3>${{missedTable}}</div><div class="feed">${{feed}}</div></div>`;
  document.getElementById("detail").innerHTML = `${{summaryHtml}}${{managedSectionHtml}}`;
  document.getElementById("score-drilldown").innerHTML = `${{scoringTabsHtml}}${{execPlanTabsHtml}}`;
  const detailHost = document.getElementById("detail");
  const scoreHost = document.getElementById("score-drilldown");
  function activateGroupTab(host, group, tab) {{
    if (!host) return;
    host.querySelectorAll(`.tabs[data-tab-group="${{group}}"] .tab-btn`).forEach((b) => b.classList.toggle("active", String(b.getAttribute("data-tab") || "") === tab));
    host.querySelectorAll(`.tab-pane[data-tab-group="${{group}}"]`).forEach((p) => p.classList.toggle("active", String(p.getAttribute("data-pane") || "") === tab));
  }}
  scoreHost?.querySelectorAll('.tabs[data-tab-group="scoring"] .tab-btn').forEach((btn) => {{
    btn.addEventListener("click", () => {{
      const tab = String(btn.getAttribute("data-tab") || "simple");
      activeScoringTab = tab;
      activateGroupTab(scoreHost, "scoring", tab);
      const bigEl = detailHost.querySelector("#summary-big-score");
      const lblEl = detailHost.querySelector("#summary-score-label");
      const subEl = detailHost.querySelector("#summary-score-sub");
      const kpiEl = detailHost.querySelector("#summary-score-kpi");
      if (tab === "simple") {{
        if (bigEl) bigEl.textContent = Number.isFinite(activeSimpleScore) ? activeSimpleScore.toFixed(1) : "N/A";
        if (lblEl) lblEl.innerHTML = '<button type="button" class="summary-score-trigger" id="summary-simple-score-trigger" aria-label="Open simple score details"><span class="score-label-text">Simple Score</span><span class="material-symbols-outlined" aria-hidden="true">open_in_new</span></button>';
        if (subEl) subEl.textContent = isScoreEligible(item)
          ? `Simple efficiency${{dueCompletionEnabled ? " (due-adjusted)" : ""}}${{overloadedApplied ? " + overloaded penalty" : ""}}${{planningRealismEnabled && overloadedApplied ? " + overload capping/planning realism cap" : ""}} | ${{simpleOverrunLabel()}}: ${{n(item.simple_score_overrun_active).toFixed(1)}}h | Planned: ${{n(item.ss_total_estimate).toFixed(1)}}h${{dueCompletionEnabled && n(item.ss_due_penalty_estimate) > 0 ? ` | Late estimate penalty: ${{n(item.ss_due_penalty_estimate).toFixed(1)}}h` : ""}}${{overloadedApplied ? ` | Overload score: ${{n(item.simple_score_overloaded).toFixed(1)}}% | Overload penalty: ${{n(item.simple_score_overloaded_penalty_pct).toFixed(1)}}%` : ""}}`
          : `Simple scoring is N/A because Planned Hours Assigned is ${{n(item.planned_hours_assigned).toFixed(1)}}h.`;
        if (kpiEl) kpiEl.textContent = Number.isFinite(activeSimpleScore) ? activeSimpleScore.toFixed(1) : "N/A";
      }} else {{
        if (bigEl) bigEl.textContent = scoreText(item, "advanced");
        if (lblEl) lblEl.textContent = "Advanced Score";
        if (subEl) subEl.textContent = `Penalty-based | Raw ${{n(item.raw_score).toFixed(2)}} | Penalty -${{n(item.total_penalty).toFixed(2)}} | Base ${{n(settings.base_score).toFixed(0)}}`;
        if (kpiEl) kpiEl.textContent = scoreText(item, "advanced");
      }}
      const summarySimpleScoreTrigger = detailHost.querySelector("#summary-simple-score-trigger");
      if (summarySimpleScoreTrigger) {{
        summarySimpleScoreTrigger.addEventListener("click", () => openScoreDrawerForAssignee(item));
      }}
      if (scoreDrawerAssignee && scoreDrawerAssignee === String(item.assignee || "") && tab === "simple") {{
        openScoreDrawerForAssignee(item);
      }}
    }});
  }});
  const summarySimpleScoreTrigger = detailHost.querySelector("#summary-simple-score-trigger");
  if (summarySimpleScoreTrigger) {{
    summarySimpleScoreTrigger.addEventListener("click", () => openScoreDrawerForAssignee(item));
  }}
  if (scoreDrawerAssignee && scoreDrawerAssignee === String(item.assignee || "") && activeScoringTab === "simple" && document.body.classList.contains("score-drawer-open")) {{
    openScoreDrawerForAssignee(item);
  }}
  scoreHost?.querySelectorAll('.tabs[data-tab-group="detail"] .tab-btn').forEach((btn) => {{
    btn.addEventListener("click", () => {{
      activateGroupTab(scoreHost, "detail", String(btn.getAttribute("data-tab") || "execution"));
    }});
  }});
  function focusContext(targetId) {{
    activateGroupTab(scoreHost, "detail", "execution");
    const target = scoreHost ? scoreHost.querySelector(targetId) : null;
    if (!target) return;
    target.classList.add("focus-pulse");
    target.scrollIntoView({{ behavior: "smooth", block: "start" }});
    setTimeout(() => target.classList.remove("focus-pulse"), 1300);
  }}
  const missedStartsTrigger = scoreHost?.querySelector('[data-action="open-missed-starts"]');
  if (missedStartsTrigger) {{
    missedStartsTrigger.addEventListener("click", () => {{
      focusContext("#missed-start-context");
    }});
  }}
  const missedDueTrigger = scoreHost?.querySelector('[data-action="open-missed-due"]');
  if (missedDueTrigger) {{
    missedDueTrigger.addEventListener("click", () => {{
      focusContext("#due-compliance-context");
    }});
  }}
  const rmiListTrigger = detailHost.querySelector('[data-action="toggle-rmis-list"]');
  if (rmiListTrigger) {{
    rmiListTrigger.addEventListener("click", () => {{
      const assigneeName = String(item.assignee || "");
      rmiListForAssignee = rmiListForAssignee === assigneeName ? "" : assigneeName;
      render(items);
    }});
  }}
  const availabilityTrigger = detailHost.querySelector('[data-action="toggle-availability-breakdown"]');
  if (availabilityTrigger) {{
    availabilityTrigger.addEventListener("click", () => {{
      const assigneeName = String(item.assignee || "");
      availabilityBreakdownForAssignee = availabilityBreakdownForAssignee === assigneeName ? "" : assigneeName;
      plannedHoursBreakdownForAssignee = "";
      actualHoursBreakdownForAssignee = "";
      render(items);
    }});
  }}
  const plannedHoursTrigger = detailHost.querySelector('[data-action="toggle-planned-hours-breakdown"]');
  if (plannedHoursTrigger) {{
    plannedHoursTrigger.addEventListener("click", () => {{
      const assigneeName = String(item.assignee || "");
      plannedHoursBreakdownForAssignee = plannedHoursBreakdownForAssignee === assigneeName ? "" : assigneeName;
      availabilityBreakdownForAssignee = "";
      actualHoursBreakdownForAssignee = "";
      render(items);
    }});
  }}
  const actualHoursTrigger = detailHost.querySelector('[data-action="toggle-actual-hours-breakdown"]');
  if (actualHoursTrigger) {{
    actualHoursTrigger.addEventListener("click", () => {{
      const assigneeName = String(item.assignee || "");
      actualHoursBreakdownForAssignee = actualHoursBreakdownForAssignee === assigneeName ? "" : assigneeName;
      availabilityBreakdownForAssignee = "";
      plannedHoursBreakdownForAssignee = "";
      render(items);
    }});
  }}
  const dueToggle = scoreHost?.querySelector("#due-completion-toggle");
  if (dueToggle) {{
    dueToggle.addEventListener("change", () => {{
      dueCompletionEnabled = dueToggle.checked;
      render(compute());
    }});
  }}
  const formulaToggle = scoreHost?.querySelector("#toggle-formula-guide");
  if (formulaToggle) {{
    formulaToggle.addEventListener("click", () => {{
      simpleFormulaGuideExpanded = !simpleFormulaGuideExpanded;
      const body = scoreHost.querySelector("#simple-formula-guide-body");
      if (body) body.hidden = !simpleFormulaGuideExpanded;
      const icon = formulaToggle.querySelector(".material-symbols-outlined");
      const label = formulaToggle.querySelector("span:last-child");
      formulaToggle.setAttribute("aria-expanded", simpleFormulaGuideExpanded ? "true" : "false");
      if (icon) icon.textContent = simpleFormulaGuideExpanded ? "expand_less" : "expand_more";
      if (label) label.textContent = simpleFormulaGuideExpanded ? "Collapse" : "Expand";
    }});
  }}
  const advancedFormulaToggle = scoreHost?.querySelector("#toggle-advanced-formula-guide");
  if (advancedFormulaToggle) {{
    advancedFormulaToggle.addEventListener("click", () => {{
      advancedFormulaGuideExpanded = !advancedFormulaGuideExpanded;
      const body = scoreHost.querySelector("#advanced-formula-guide-body");
      if (body) body.hidden = !advancedFormulaGuideExpanded;
      const icon = advancedFormulaToggle.querySelector(".material-symbols-outlined");
      const label = advancedFormulaToggle.querySelector("span:last-child");
      advancedFormulaToggle.setAttribute("aria-expanded", advancedFormulaGuideExpanded ? "true" : "false");
      if (icon) icon.textContent = advancedFormulaGuideExpanded ? "expand_less" : "expand_more";
      if (label) label.textContent = advancedFormulaGuideExpanded ? "Collapse" : "Expand";
    }});
  }}
}}
function renderAll() {{ availabilityBreakdownForAssignee = ""; plannedHoursBreakdownForAssignee = ""; actualHoursBreakdownForAssignee = ""; rmiListForAssignee = ""; render(compute()); }}
function setHeaderCollapsed(isCollapsed) {{
  const collapsed = Boolean(isCollapsed);
  if (headerSectionEl) {{
    headerSectionEl.classList.toggle("is-collapsed", collapsed);
  }}
  document.body.classList.toggle("header-collapsed", collapsed);
  if (headerToggleButton) {{
    headerToggleButton.setAttribute("aria-expanded", collapsed ? "false" : "true");
    headerToggleButton.textContent = collapsed ? "Expand Header" : "Collapse Header";
  }}
  if (headerExpandFabButton) {{
    headerExpandFabButton.setAttribute("aria-hidden", collapsed ? "false" : "true");
  }}
  localStorage.setItem(HEADER_COLLAPSED_STORAGE_KEY, collapsed ? "1" : "0");
}}
document.getElementById("projects").innerHTML = projects.map((p) => `<option value="${{e(p)}}" selected>${{e(p)}}</option>`).join("");
refreshCapacityProfileOptions();
document.getElementById("from").value = defaultFrom; document.getElementById("to").value = defaultTo;
document.getElementById("meta").textContent = `Generated: ${{payload.generated_at || "-"}} | Data window: ${{defaultFrom}} to ${{defaultTo}}`;
setHeaderCollapsed(localStorage.getItem(HEADER_COLLAPSED_STORAGE_KEY) === "1");
if (headerToggleButton) {{
  headerToggleButton.addEventListener("click", () => {{
    const currentlyCollapsed = headerSectionEl ? headerSectionEl.classList.contains("is-collapsed") : false;
    setHeaderCollapsed(!currentlyCollapsed);
  }});
}}
if (headerExpandFabButton) {{
  headerExpandFabButton.addEventListener("click", () => {{
    setHeaderCollapsed(false);
  }});
}}
if (scoreDrawerCloseEl) {{
  scoreDrawerCloseEl.addEventListener("click", closeScoreDrawer);
}}
if (scoreDrawerOverlayEl) {{
  scoreDrawerOverlayEl.addEventListener("click", closeScoreDrawer);
}}
document.addEventListener("keydown", (event) => {{
  if (event.key === "Escape" && document.body.classList.contains("score-drawer-open")) {{
    closeScoreDrawer();
  }}
}});
if (advFilterToggleButton && advFilterMenu) {{
  advFilterToggleButton.addEventListener("click", () => {{
    const expanded = advFilterToggleButton.getAttribute("aria-expanded") === "true";
    setAdvancedFilterMenuOpen(!expanded);
  }});
  advFilterMenu.querySelectorAll(".adv-filter-item").forEach((btn) => {{
    btn.addEventListener("click", () => {{
      const preset = String(btn.getAttribute("data-preset") || "");
      if (preset === "last30") applyDateShortcut("last_30_days");
      else if (preset === "lastMonth") applyDateShortcut("previous_month");
      else if (preset === "currentMonth") applyDateShortcut("current_month");
      else if (preset === "last90") applyDateShortcut("last_90_days");
      else if (preset === "lastQuarter") applyDateShortcut("last_quarter");
      else if (preset === "currentQuarter") applyDateShortcut("current_quarter");
      setDateFilterStatus(`Preset applied: ${{btn.textContent || preset}}`);
      setAdvancedFilterMenuOpen(false);
      renderAll();
    }});
  }});
  document.addEventListener("click", (event) => {{
    const target = event.target;
    if (!(target instanceof Element)) return;
    if (target.closest(".adv-filter-wrap")) return;
    setAdvancedFilterMenuOpen(false);
  }});
  document.addEventListener("keydown", (event) => {{
    if (event.key === "Escape") setAdvancedFilterMenuOpen(false);
  }});
}}
document.getElementById("toggle-performance-kpis").addEventListener("click", () => {{
  const wrap = document.getElementById("performance-kpis-wrap");
  const collapsed = wrap.classList.toggle("is-collapsed");
  document.querySelector("#toggle-performance-kpis .hint").textContent = collapsed ? "click to expand" : "click to collapse";
}});
document.getElementById("toggle-team-performance").addEventListener("click", () => {{
  const sec = document.getElementById("team-performance-section");
  const collapsed = sec.classList.toggle("is-collapsed");
  document.querySelector("#toggle-team-performance .hint").textContent = collapsed ? "click to expand" : "click to collapse";
}});
document.getElementById("apply").addEventListener("click", ()=>{{ setDateFilterStatus(""); renderAll(); }});
if (simpleOverrunModeEl) {{
  simpleOverrunModeEl.addEventListener("change", () => {{
    syncSimpleOverrunMode(simpleOverrunModeEl.value);
    renderAll();
  }});
}}
if (capacityProfileSelectEl) {{
  capacityProfileSelectEl.addEventListener("change", () => {{
    syncCapacityProfileSelection(capacityProfileSelectEl.value, "header");
    renderAll();
  }});
}}
if (capacityProfileTopSelectEl) {{
  capacityProfileTopSelectEl.addEventListener("change", () => {{
    syncCapacityProfileSelection(capacityProfileTopSelectEl.value, "top");
    renderAll();
  }});
}}
if (assigneeExtendedActualsToggleEl) {{
  assigneeExtendedActualsToggleEl.addEventListener("change", () => {{
    extendedActualsEnabled = Boolean(assigneeExtendedActualsToggleEl.checked);
    actualHoursBreakdownForAssignee = "";
    renderAll();
  }});
}}
if (assigneeOverloadedPenaltyToggleEl) {{
  assigneeOverloadedPenaltyToggleEl.checked = overloadedPenaltyEnabled;
  assigneeOverloadedPenaltyToggleEl.addEventListener("change", () => {{
    overloadedPenaltyEnabled = Boolean(assigneeOverloadedPenaltyToggleEl.checked);
    renderAll();
  }});
}}
if (assigneePlanningRealismToggleEl) {{
  assigneePlanningRealismToggleEl.checked = planningRealismEnabled;
  assigneePlanningRealismToggleEl.addEventListener("change", () => {{
    planningRealismEnabled = Boolean(assigneePlanningRealismToggleEl.checked);
    renderAll();
  }});
}}
if (leaderActionsToggle && leaderActionsMenu) {{
  leaderActionsToggle.addEventListener("click", () => {{
    const expanded = leaderActionsToggle.getAttribute("aria-expanded") === "true";
    setLeaderActionsMenuOpen(!expanded);
  }});
  leaderActionsMenu.addEventListener("click", async (event) => {{
    const target = event.target;
    if (!(target instanceof Element)) return;
    const actionButton = target.closest("[data-action]");
    if (!actionButton) return;
    const action = String(actionButton.getAttribute("data-action") || "");
    if (action === "copy-gap-people") {{
      await copyGapPeopleFromLeaderboard();
      setLeaderActionsMenuOpen(false);
    }}
  }});
  document.addEventListener("click", (event) => {{
    const target = event.target;
    if (!(target instanceof Element)) return;
    if (target.closest(".leader-actions-menu-wrap")) return;
    setLeaderActionsMenuOpen(false);
  }});
  document.addEventListener("keydown", (event) => {{
    if (event.key === "Escape") setLeaderActionsMenuOpen(false);
  }});
}}
document.getElementById("leader-scoring-mode").addEventListener("change", renderAll);
document.getElementById("leader-sort").addEventListener("change", renderAll);
document.getElementById("leader-sort-direction").addEventListener("change", renderAll);
document.getElementById("filter-risk").addEventListener("change", renderAll);
document.getElementById("filter-missed").addEventListener("change", renderAll);
document.getElementById("leader-search").addEventListener("input", renderAll);
document.getElementById("reset").addEventListener("click", ()=>{{ document.getElementById("from").value=defaultFrom; document.getElementById("to").value=defaultTo; document.getElementById("search").value=\"\"; document.getElementById("leader-search").value=\"\"; document.getElementById("leader-sort").value=\"score\"; document.getElementById("leader-sort-direction").value=\"desc\"; document.getElementById("leader-scoring-mode").value=\"simple\"; document.getElementById("filter-risk").value=\"all\"; document.getElementById("filter-missed").value=\"all\"; syncSimpleOverrunMode(\"subtasks\"); if (assigneeExtendedActualsToggleEl) assigneeExtendedActualsToggleEl.checked = false; extendedActualsEnabled = false; applyPerformanceSettings(settings); syncCapacityProfileSelection(\"auto\", \"\"); selectedTeam = \"\"; setDateFilterStatus(""); Array.from(document.getElementById("projects").options).forEach(o => o.selected=true); renderAll(); }});
document.getElementById("shortcut-current-month").addEventListener("click", ()=>{{ applyDateShortcut("current_month"); renderAll(); }});
document.getElementById("shortcut-previous-month").addEventListener("click", ()=>{{ applyDateShortcut("previous_month"); renderAll(); }});
document.getElementById("shortcut-last-30-days").addEventListener("click", ()=>{{ applyDateShortcut("last_30_days"); renderAll(); }});
document.getElementById("shortcut-quarter-to-date").addEventListener("click", ()=>{{ applyDateShortcut("quarter_to_date"); renderAll(); }});
document.getElementById("shortcut-reset").addEventListener("click", ()=>{{ applyDateShortcut("reset"); renderAll(); }});
if (employeeRefreshBtn) {{
  function clearEmployeeRefreshPoll() {{
    if (employeeRefreshPollHandle) {{
      clearTimeout(employeeRefreshPollHandle);
      employeeRefreshPollHandle = null;
    }}
  }}

  function setEmployeeRefreshUiState(isRunning) {{
    employeeRefreshBtn.disabled = Boolean(isRunning);
    if (employeeRefreshCancelBtn) {{
      employeeRefreshCancelBtn.style.display = "";
      employeeRefreshCancelBtn.disabled = !isRunning;
    }}
  }}

  function refreshDetailsText(run) {{
    const row = run && typeof run === "object" ? run : {{}};
    const stats = row.stats && typeof row.stats === "object" ? row.stats : {{}};
    const sources = stats.sources && typeof stats.sources === "object" ? stats.sources : {{}};
    const wi = sources.work_items || {{}};
    const wl = sources.worklogs || {{}};
    const lv = sources.leaves || {{}};
    const ep = stats.employee_progress && typeof stats.employee_progress === "object" ? stats.employee_progress : {{}};
    const out = [];
    out.push(`Run ID: ${{String(row.run_id || "-")}}`);
    if (Number.isFinite(Number(row.progress))) out.push(`Progress: ${{Math.max(0, Math.min(100, Math.round(Number(row.progress))))}}%`);
    if (Number.isFinite(Number(ep.total_employees)) && Number(ep.total_employees) > 0) {{
      const fetched = Number.isFinite(Number(ep.fetched_employees)) ? n(ep.fetched_employees).toFixed(0) : "0";
      const remaining = Number.isFinite(Number(ep.remaining_employees)) ? n(ep.remaining_employees).toFixed(0) : "?";
      out.push(`Employees -> Total: ${{n(ep.total_employees).toFixed(0)}}, Fetched: ${{fetched}}, Remaining: ${{remaining}}`);
      if (String(ep.current_assignee || "").trim()) out.push(`Current: ${{String(ep.current_assignee)}}`);
    }}
    if (Number.isFinite(Number(stats.duration_sec))) out.push(`Duration: ${{Number(stats.duration_sec).toFixed(2)}}s`);
    if (row.started_at_utc) out.push(`Started: ${{String(row.started_at_utc)}}`);
    if (row.ended_at_utc) out.push(`Ended: ${{String(row.ended_at_utc)}}`);
    out.push(`Rows -> WI: ${{n(wi.rows).toFixed(0)}}, WL: ${{n(wl.rows).toFixed(0)}}, Leaves: ${{n(lv.rows).toFixed(0)}}`);
    if (String(row.error || "").trim()) out.push(`Error: ${{String(row.error).trim()}}`);
    return out.join(" | ");
  }}

  function refreshToneByStatus(status) {{
    const key = String(status || "").toLowerCase();
    if (key === "success") return "ok";
    if (key === "failed" || key === "canceled") return "err";
    return "";
  }}

  function refreshStatusText(run) {{
    const status = String((run && run.status) || "").toLowerCase();
    const step = String((run && run.step) || "").replace(/_/g, " ").trim();
    const progress = Number(run && run.progress);
    const stats = (run && run.stats && typeof run.stats === "object") ? run.stats : {{}};
    const ep = (stats.employee_progress && typeof stats.employee_progress === "object") ? stats.employee_progress : {{}};
    if (status === "running") {{
      const pct = Number.isFinite(progress) ? `${{Math.max(0, Math.min(100, Math.round(progress)))}}%` : "";
      let empText = "";
      if (Number.isFinite(Number(ep.total_employees)) && Number(ep.total_employees) > 0) {{
        const fetched = Number.isFinite(Number(ep.fetched_employees)) ? Number(ep.fetched_employees) : 0;
        empText = ` | Employee ${{fetched}}/${{Number(ep.total_employees)}}`;
        if (String(ep.current_assignee || "").trim()) empText += ` (${{String(ep.current_assignee)}})`;
      }}
      const stepText = step ? ` (${{step}})` : "";
      return `Refresh running${{stepText}}${{empText}}${{pct ? ` - ${{pct}}` : ""}}`;
    }}
    if (status === "canceled") return String((run && run.error) || "Refresh canceled. Previous snapshot retained.");
    if (status === "failed") return String((run && run.error) || "Refresh failed.");
    if (status === "success") {{
      let msg = "Refresh complete. Reloading...";
      if (Number.isFinite(Number(ep.total_employees)) && Number(ep.total_employees) > 0) {{
        const fetched = Number.isFinite(Number(ep.fetched_employees)) ? Number(ep.fetched_employees) : 0;
        msg = `Refresh complete (${{fetched}}/${{Number(ep.total_employees)}} employees). Reloading...`;
      }}
      return msg;
    }}
    if (status === "cancel_requested") return "Cancel requested. Waiting for safe stop...";
    return String((run && run.error) || "Refresh status unavailable.");
  }}

  async function pollEmployeeRefresh(runId) {{
    if (!runId) return;
    try {{
      const response = await fetch(`/api/employee-performance/refresh/${{encodeURIComponent(runId)}}`);
      const body = await response.json().catch(() => ({{ ok: false, error: "Invalid refresh status response." }}));
      if (!response.ok || !body.ok || !body.run) {{
        const msg = body && body.error ? String(body.error) : `Failed to check refresh status (${{response.status}})`;
        setEmployeeRefreshStatus(msg, "err");
        setEmployeeRefreshUiState(false);
        setEmployeeRefreshInlineState(employeeRefreshInlineRun ? {{ ...employeeRefreshInlineRun, status: "failed", error: msg }} : null, employeeRefreshInlineAssignee || "", true);
        clearEmployeeRefreshPoll();
        return;
      }}
      const run = body.run || {{}};
      const status = String(run.status || "").toLowerCase();
      setEmployeeRefreshStatus(refreshStatusText(run), refreshToneByStatus(status));
      setEmployeeRefreshDetails(refreshDetailsText(run), refreshToneByStatus(status));
      setEmployeeRefreshInlineState(run, employeeRefreshInlineAssignee || "", true);
      if (status === "running" || status === "cancel_requested") {{
        setEmployeeRefreshUiState(true);
        clearEmployeeRefreshPoll();
        employeeRefreshPollHandle = setTimeout(() => pollEmployeeRefresh(runId), 1500);
        return;
      }}
      setEmployeeRefreshUiState(false);
      clearEmployeeRefreshPoll();
      if (status === "success") {{
        setTimeout(() => window.location.reload(), 800);
      }}
    }} catch (error) {{
      setEmployeeRefreshStatus(error && error.message ? error.message : String(error), "err");
      setEmployeeRefreshUiState(false);
      setEmployeeRefreshInlineState(employeeRefreshInlineRun ? {{ ...employeeRefreshInlineRun, status: "failed", error: (error && error.message ? error.message : String(error)) }} : null, employeeRefreshInlineAssignee || "", true);
      clearEmployeeRefreshPoll();
    }}
  }}

  async function resumeEmployeeRefreshIfRunning() {{
    if (typeof window === "undefined" || !window.location || window.location.protocol === "file:") return;
    try {{
      const response = await fetch("/api/employee-performance/refresh/current");
      const body = await response.json().catch(() => ({{ ok: false }}));
      const run = body && body.ok ? body.run : null;
      if (!run) {{
        setEmployeeRefreshUiState(false);
        setEmployeeRefreshDetails("", "");
        setEmployeeRefreshInlineState(null, "", true);
        return;
      }}
      const status = String(run.status || "").toLowerCase();
      setEmployeeRefreshStatus(refreshStatusText(run), refreshToneByStatus(status));
      setEmployeeRefreshDetails(refreshDetailsText(run), refreshToneByStatus(status));
      setEmployeeRefreshInlineState(run, String(run.assignee || ""), true);
      if (status === "running" || status === "cancel_requested") {{
        setEmployeeRefreshUiState(true);
        clearEmployeeRefreshPoll();
        employeeRefreshPollHandle = setTimeout(() => pollEmployeeRefresh(String(run.run_id || "")), 1200);
      }} else {{
        setEmployeeRefreshUiState(false);
      }}
    }} catch (_err) {{
      setEmployeeRefreshUiState(false);
    }}
  }}

  window.__startEmployeeRefreshRun = async (assigneeName, triggerButton) => {{
    if (typeof window !== "undefined" && window.location && window.location.protocol === "file:") {{
      setEmployeeRefreshStatus("Offline mode: refresh API unavailable.", "err");
      return;
    }}
    const targetAssignee = String(assigneeName || "").trim();
    if (triggerButton) triggerButton.disabled = true;
    setEmployeeRefreshUiState(true);
    setEmployeeRefreshStatus(targetAssignee ? `Starting refresh for ${{targetAssignee}}...` : "Starting refresh...", "");
    setEmployeeRefreshDetails("", "");
    setEmployeeRefreshInlineState(targetAssignee ? {{ run_id: "", status: "running", step: "initializing", progress: 0, assignee: targetAssignee, stats: {{}} }} : null, targetAssignee, true);
    try {{
      const response = await fetch("/api/employee-performance/refresh", {{
        method: "POST",
        headers: {{"Content-Type":"application/json"}},
        body: JSON.stringify(targetAssignee ? {{ assignee: targetAssignee, replace_running: true }} : {{ replace_running: true }}),
      }});
      const payload = await response.json().catch(() => null);
      if (!response.ok || !payload || !payload.ok) {{
        const legacyResp = await fetch("/api/report/refresh", {{
          method: "POST",
          headers: {{"Content-Type":"application/json"}},
          body: JSON.stringify(
            targetAssignee
              ? {{ report: "employee_performance", assignee: targetAssignee, isolated: true }}
              : {{ report: "employee_performance", isolated: true }}
          ),
        }});
        const legacyPayload = await legacyResp.json().catch(() => null);
        if (!legacyResp.ok || !legacyPayload || !legacyPayload.ok) {{
          const errMsg = (legacyPayload && legacyPayload.error)
            ? String(legacyPayload.error)
            : (payload && payload.error)
              ? String(payload.error)
              : `Refresh failed (${{response.status}})`;
          throw new Error(errMsg);
        }}
        setEmployeeRefreshStatus("Refresh complete. Reloading...", "ok");
        setEmployeeRefreshDetails("", "");
        setEmployeeRefreshInlineState(targetAssignee ? {{ run_id: "", status: "success", step: "done", progress: 100, assignee: targetAssignee, stats: {{}} }} : null, targetAssignee, true);
        setTimeout(() => window.location.reload(), 800);
        return;
      }}
      const runId = String(payload.run_id || (payload.run && payload.run.run_id) || "");
      if (!runId) throw new Error("Refresh started without run_id.");
      const run = payload.run && typeof payload.run === "object" ? payload.run : {{}};
      setEmployeeRefreshStatus(targetAssignee ? `Refresh started for ${{targetAssignee}}. Monitoring run...` : "Refresh started. Monitoring run...", "");
      setEmployeeRefreshDetails(refreshDetailsText(run), "");
      setEmployeeRefreshInlineState(run, targetAssignee, true);
      clearEmployeeRefreshPoll();
      employeeRefreshPollHandle = setTimeout(() => pollEmployeeRefresh(runId), 700);
    }} catch (error) {{
      setEmployeeRefreshStatus(error && error.message ? error.message : String(error), "err");
      setEmployeeRefreshUiState(false);
      setEmployeeRefreshInlineState(targetAssignee ? {{ run_id: "", status: "failed", step: "failed", progress: 100, assignee: targetAssignee, error: (error && error.message ? error.message : String(error)), stats: {{}} }} : null, targetAssignee, true);
      clearEmployeeRefreshPoll();
    }} finally {{
      if (triggerButton) triggerButton.disabled = false;
    }}
  }};

  employeeRefreshBtn.addEventListener("click", async () => {{
    await window.__startEmployeeRefreshRun("", employeeRefreshBtn);
  }});

  if (employeeRefreshCancelBtn) {{
    employeeRefreshCancelBtn.addEventListener("click", async () => {{
      employeeRefreshCancelBtn.disabled = true;
      try {{
        const response = await fetch("/api/employee-performance/cancel", {{
          method: "POST",
          headers: {{"Content-Type":"application/json"}},
          body: JSON.stringify({{}}),
        }});
        const payload = await response.json().catch(() => ({{ ok: false, error: "Invalid cancel response." }}));
        if (!response.ok || !payload.ok) {{
          const msg = payload && payload.error ? String(payload.error) : `Cancel failed (${{response.status}})`;
          throw new Error(msg);
        }}
        setEmployeeRefreshStatus(String(payload.message || "Cancel requested. Waiting for safe stop..."), "");
        setEmployeeRefreshDetails(String(payload.message || ""), "");
      }} catch (error) {{
        setEmployeeRefreshStatus(error && error.message ? error.message : String(error), "err");
      }} finally {{
        employeeRefreshCancelBtn.disabled = false;
      }}
    }});
  }}

  setEmployeeRefreshUiState(false);
  resumeEmployeeRefreshIfRunning();
}}
hydratePerformanceSettings().finally(() => {{
  performanceSettingsReady = true;
  renderAll();
}});
</script>
<script>
if (typeof window !== "undefined" && window.location && window.location.protocol !== "file:") {{
  const navScript = document.createElement("script");
  navScript.src = "shared-nav.js";
  document.body.appendChild(navScript);
}}
</script>
</body>
</html>"""


def _resolve_runtime_paths(base_dir: Path) -> dict[str, Path]:
    worklog_name = os.getenv("JIRA_WORKLOG_XLSX_PATH", DEFAULT_WORKLOG_INPUT_XLSX).strip() or DEFAULT_WORKLOG_INPUT_XLSX
    work_items_name = os.getenv("JIRA_EXPORT_XLSX_PATH", DEFAULT_WORK_ITEMS_INPUT_XLSX).strip() or DEFAULT_WORK_ITEMS_INPUT_XLSX
    leave_name = os.getenv("JIRA_LEAVE_REPORT_XLSX_PATH", DEFAULT_LEAVE_REPORT_INPUT_XLSX).strip() or DEFAULT_LEAVE_REPORT_INPUT_XLSX
    html_name = os.getenv("JIRA_EMPLOYEE_PERFORMANCE_HTML_PATH", DEFAULT_HTML_OUTPUT).strip() or DEFAULT_HTML_OUTPUT
    db_name = os.getenv("JIRA_ASSIGNEE_HOURS_CAPACITY_DB_PATH", DEFAULT_CAPACITY_DB).strip() or DEFAULT_CAPACITY_DB
    source_mode = _to_text(os.getenv("JIRA_EMP_PERF_INPUT_SOURCE", "xlsx")).lower() or "xlsx"
    run_id = _to_text(os.getenv("JIRA_EMP_PERF_RUN_ID"))
    return {
        "worklog_path": _resolve_path(worklog_name, base_dir),
        "work_items_path": _resolve_path(work_items_name, base_dir),
        "leave_report_path": _resolve_path(leave_name, base_dir),
        "html_path": _resolve_path(html_name, base_dir),
        "db_path": _resolve_path(db_name, base_dir),
        "source_mode": source_mode,
        "run_id": run_id,
    }


def main() -> None:
    base_dir = Path(__file__).resolve().parent
    paths = _resolve_runtime_paths(base_dir)
    _init_performance_settings_db(paths["db_path"])
    settings = _load_performance_settings(paths["db_path"])
    teams = _list_performance_teams(paths["db_path"])
    entities_catalog = load_report_entities(paths["db_path"])
    managed_fields = load_manage_fields(paths["db_path"], include_inactive=False)
    capacity_profiles = _list_capacity_profiles(paths["db_path"])
    source_mode = _to_text(paths.get("source_mode")).lower() or "xlsx"
    if source_mode == "db":
        requested_run = _to_text(paths.get("run_id"))
        run_id = _resolve_epf_run_id(paths["db_path"], requested_run)
        if not run_id:
            raise ValueError("DB source mode selected but no active epf run_id found.")
        work_items = _load_work_items_from_epf_db(paths["db_path"], run_id)
        worklogs = _load_worklogs_from_epf_db(paths["db_path"], run_id, work_items)
        leave_rows = _load_unplanned_leave_rows_from_epf_db(paths["db_path"], run_id)
        leave_issue_keys = _load_leave_issue_keys_from_epf_db(paths["db_path"], run_id)
    else:
        work_items = _load_work_items(paths["work_items_path"])
        worklogs = _load_worklogs(paths["worklog_path"], work_items)
        leave_rows = _load_unplanned_leave_rows(paths["leave_report_path"])
        leave_issue_keys = _load_leave_issue_keys(paths["leave_report_path"])
    simple_scoring = _precompute_simple_scoring(paths["db_path"], work_items, worklogs)
    payload = _build_payload(
        worklogs,
        list(work_items.values()),
        leave_rows,
        settings,
        teams,
        entities_catalog=entities_catalog,
        managed_fields=managed_fields,
        capacity_profiles=capacity_profiles,
        leave_issue_keys=leave_issue_keys,
        simple_scoring=simple_scoring,
    )
    paths["html_path"].write_text(_build_html(payload), encoding="utf-8")
    print(f"Employee performance logs: {len(worklogs)}")
    print(f"Wrote HTML report: {paths['html_path']}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Generate employee performance report.")
    parser.parse_args()
    main()
