from __future__ import annotations

from datetime import date, datetime, timedelta
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


HEADER_FILL = PatternFill("solid", fgColor="173F70")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TOTAL_FILL = PatternFill("solid", fgColor="DCEBFA")


def _text(value: Any) -> str:
    return str(value or "").strip()


def _key(value: Any) -> str:
    return _text(value).casefold()


def _number(value: Any) -> float:
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def _excel_safe(value: Any) -> Any:
    if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
        return "'" + value
    return value


def _month_range(month: str) -> tuple[str, str]:
    try:
        start = datetime.strptime(month, "%Y-%m").date().replace(day=1)
    except ValueError as exc:
        raise ValueError("month must use YYYY-MM format") from exc
    next_month = (start.replace(day=28) + timedelta(days=4)).replace(day=1)
    return start.isoformat(), (next_month - timedelta(days=1)).isoformat()


def _selected_profile(profiles: list[dict], profile_index: str, month_start: str) -> dict:
    if profile_index != "":
        try:
            return profiles[int(profile_index)]
        except (ValueError, IndexError):
            pass
    return next(
        (
            profile
            for profile in profiles
            if _text(profile.get("from_date")) <= month_start <= _text(profile.get("to_date"))
        ),
        profiles[0] if profiles else {},
    )


def _calendar_rows(profile: dict, start: str, end: str) -> list[dict]:
    current = date.fromisoformat(start)
    last = date.fromisoformat(end)
    holidays = {_text(value) for value in profile.get("holiday_dates") or []}
    rows: list[dict] = []
    while current <= last:
        iso = current.isoformat()
        weekend = current.weekday() >= 5
        holiday = iso in holidays
        ramadan = bool(_text(profile.get("ramadan_start_date"))) and (
            _text(profile.get("ramadan_start_date")) <= iso <= _text(profile.get("ramadan_end_date"))
        )
        hours = 0.0
        classification = "Weekend" if weekend else "Official leave" if holiday else "Workday"
        if not weekend and not holiday:
            if ramadan:
                classification = "Ramadan workday"
                hours = _number(profile.get("ramadan_hours_per_day"))
            else:
                hours = _number(profile.get("standard_hours_per_day") or 8)
        rows.append({"date": iso, "classification": classification, "hours": round(hours, 2)})
        current += timedelta(days=1)
    return rows


def _epic_for_issue(issue_key: str, issue_map: dict[str, dict]) -> tuple[str, str]:
    current = _text(issue_key).upper()
    visited: set[str] = set()
    for _ in range(8):
        if not current or current in visited:
            break
        visited.add(current)
        item = issue_map.get(current) or {}
        if "epic" in _text(item.get("issue_type")).lower():
            return current, _text(item.get("summary"))
        current = _text(item.get("parent_issue_key")).upper()
    return "", ""


def _visible_names(payload: dict, filters: dict) -> list[str]:
    issues = payload.get("issues") or []
    worklogs = payload.get("worklogs") or []
    leaves = payload.get("leaves") or []
    resources = payload.get("resources") or {}
    display_by_key: dict[str, str] = {}
    for name in list(resources) + [row.get("assignee") for row in issues] + [
        row.get("worklog_author") or row.get("issue_assignee") for row in worklogs
    ] + [row.get("assignee") for row in leaves]:
        if _key(name):
            display_by_key.setdefault(_key(name), _text(name))

    selected_teams = {_key(value) for value in filters.get("selected_teams") or []}
    teams = payload.get("teams") or []
    resource_by_key = {_key(name): row for name, row in resources.items()}
    show_resigned = bool(filters.get("display_resigned"))
    visible: list[str] = []
    for name_key, display_name in display_by_key.items():
        memberships = [
            team for team in teams if any(_key(member) == name_key for member in team.get("assignees") or [])
        ]
        if memberships and not any(_key(team.get("team_name")) in selected_teams for team in memberships):
            continue
        if not show_resigned and bool((resource_by_key.get(name_key) or {}).get("resigned")):
            continue
        visible.append(display_name)
    return sorted(visible, key=str.casefold)


def _style_sheet(ws, table_name: str) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 24
    for column in ws.columns:
        letter = column[0].column_letter
        width = max((len(_text(cell.value)) for cell in column), default=8) + 2
        ws.column_dimensions[letter].width = min(max(width, 11), 48)
    if ws.max_row >= 2 and ws.max_column >= 1:
        table = Table(displayName=table_name, ref=ws.dimensions)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
            showRowStripes=True, showColumnStripes=False,
        )
        ws.add_table(table)


def _write_sheet(wb: Workbook, title: str, headers: list[str], rows: list[list[Any]], table_name: str):
    ws = wb.create_sheet(title)
    ws.append([_excel_safe(value) for value in headers])
    for row in rows:
        ws.append([_excel_safe(value) for value in row])
    _style_sheet(ws, table_name)
    return ws


def build_employee_capacity_utilization_workbook(payload: dict, filters: dict) -> tuple[BytesIO, str]:
    month = _text(filters.get("month"))
    start, end = _month_range(month)
    scope = _text(filters.get("scope")) or "any"
    include_leaves = bool(filters.get("include_leaves"))
    profiles = payload.get("profiles") or []
    profile = _selected_profile(profiles, _text(filters.get("profile_index")), start)
    calendar = _calendar_rows(profile, start, end)
    capacity_per_employee = round(sum(row["hours"] for row in calendar), 2)
    official_days = sum(1 for row in calendar if row["classification"] == "Official leave")
    names = _visible_names(payload, filters)
    name_keys = {_key(name) for name in names}
    issues = payload.get("issues") or []
    issue_map = {_text(row.get("issue_key")).upper(): row for row in issues}
    support = {_key(name) for name in payload.get("support") or []}
    resources = {_key(name): row for name, row in (payload.get("resources") or {}).items()}
    teams = payload.get("teams") or []
    browse_base = _text(payload.get("jira_browse_base")).rstrip("/")

    leaves = [
        row for row in payload.get("leaves") or []
        if _key(row.get("assignee")) in name_keys and start <= _text(row.get("period_day")) <= end
    ]
    booked = [
        row for row in issues
        if _key(row.get("assignee")) in name_keys
        and "sub" in _text(row.get("issue_type")).lower()
        and not _text(row.get("issue_key")).upper().startswith("RLT-")
        and _text(row.get("start_date")) and _text(row.get("due_date"))
        and _text(row.get("start_date")) <= end and _text(row.get("due_date")) >= start
    ]
    worklogs = []
    for row in payload.get("worklogs") or []:
        author_key = _key(row.get("worklog_author") or row.get("issue_assignee"))
        if author_key not in name_keys or not (start <= _text(row.get("worklog_date")) <= end):
            continue
        is_leave_worklog = (
            _key(row.get("project_key")) == "rlt"
            or _text(row.get("issue_id") or row.get("issue_key")).upper().startswith("RLT-")
        )
        if is_leave_worklog and not include_leaves:
            continue
        if scope == "assigned" and not (
            _key(row.get("item_assignee")) == author_key
            and "sub" in _text(row.get("item_issue_type")).lower()
        ):
            continue
        worklogs.append(row)

    summary_rows: list[list[Any]] = []
    for name in names:
        name_key = _key(name)
        taken = round(sum(
            _number(row.get("planned_taken_hours")) + _number(row.get("unplanned_taken_hours"))
            for row in leaves if _key(row.get("assignee")) == name_key
        ), 2)
        booked_hours = round(sum(
            _number(row.get("original_estimate_hours")) for row in booked if _key(row.get("assignee")) == name_key
        ), 2)
        logged_hours = round(sum(
            _number(row.get("hours_logged")) for row in worklogs
            if _key(row.get("worklog_author") or row.get("issue_assignee")) == name_key
        ), 2)
        availability = max(0.0, capacity_per_employee - taken)
        summary_rows.append([
            name, capacity_per_employee, official_days, taken, availability,
            booked_hours, logged_hours, round(logged_hours / availability * 100, 1) if availability else 0.0,
        ])
    total_columns = [sum(_number(row[index]) for row in summary_rows) for index in range(1, 7)]
    total_availability = total_columns[3]
    total_logged = total_columns[5]
    summary_rows.append([
        "Grand Total", total_columns[0], int(total_columns[1]), total_columns[2], total_availability,
        total_columns[4], total_logged, round(total_logged / total_availability * 100, 1) if total_availability else 0.0,
    ])

    wb = Workbook()
    wb.remove(wb.active)
    summary_ws = _write_sheet(
        wb, "Summary",
        ["Employee Name", "Capacity (Hours)", "Official Leaves (Days)", "Leaves Taken (Hours)",
         "Availability (Hours)", "Booked Manhours", "Logged Hours", "Utilization (%)"],
        summary_rows, "UtilizationSummary",
    )
    for cell in summary_ws[summary_ws.max_row]:
        cell.fill = TOTAL_FILL
        cell.font = Font(bold=True)
    for row in summary_ws.iter_rows(min_row=2, min_col=2, max_col=8):
        for cell in row:
            cell.number_format = "0.00"

    worklog_rows = []
    for row in worklogs:
        epic_key = _text(row.get("epic_key"))
        epic_name = _text(row.get("epic_summary"))
        if not epic_key:
            epic_key, epic_name = _epic_for_issue(_text(row.get("issue_id")), issue_map)
        issue_key = _text(row.get("issue_id")).upper()
        worklog_rows.append([
            row.get("worklog_author") or row.get("issue_assignee"), issue_key,
            f"{browse_base}/{issue_key}" if browse_base and issue_key else "", row.get("item_summary"),
            epic_key, epic_name, f"{browse_base}/{epic_key}" if browse_base and epic_key else "",
            row.get("item_assignee") or row.get("issue_assignee"), row.get("worklog_date"),
            _number(row.get("hours_logged")),
        ])
    worklog_ws = _write_sheet(
        wb, "Worklogs",
        ["Employee", "Work Item", "Work Item Link", "Work Item Title", "Epic", "Epic Name",
         "Epic Jira Link", "Assigned Resource", "Worklog Date", "Hours Logged"],
        worklog_rows, "UtilizationWorklogs",
    )

    booked_rows = []
    for row in booked:
        issue_key = _text(row.get("issue_key")).upper()
        epic_key, epic_name = _epic_for_issue(issue_key, issue_map)
        booked_rows.append([
            row.get("assignee"), issue_key, f"{browse_base}/{issue_key}" if browse_base else "", row.get("summary"),
            epic_key, epic_name, f"{browse_base}/{epic_key}" if browse_base and epic_key else "",
            row.get("start_date"), row.get("due_date"), _number(row.get("original_estimate_hours")),
        ])
    booked_ws = _write_sheet(
        wb, "Booked Subtasks",
        ["Employee", "Work Item", "Work Item Link", "Title", "Epic", "Epic Name", "Epic Jira Link",
         "Start Date", "Due Date", "Original Estimate (Hours)"],
        booked_rows, "UtilizationBooked",
    )

    leave_rows = [[
        row.get("assignee"), row.get("period_day"), row.get("jira_task_ids"), row.get("jira_task_links"),
        _number(row.get("planned_taken_hours")), _number(row.get("unplanned_taken_hours")),
        _number(row.get("planned_taken_hours")) + _number(row.get("unplanned_taken_hours")),
    ] for row in leaves]
    leave_ws = _write_sheet(
        wb, "Leave Records",
        ["Employee", "Date", "Jira Leave Item", "Jira Link", "Planned Hours", "Unplanned Hours", "Taken Hours"],
        leave_rows, "UtilizationLeaves",
    )
    _write_sheet(
        wb, "Capacity Calendar",
        ["Date", "Classification", "Hours per Employee", "Employees", "Total Capacity Hours"],
        [[row["date"], row["classification"], row["hours"], len(names), row["hours"] * len(names)] for row in calendar],
        "UtilizationCalendar",
    )
    employee_rows = []
    for name in names:
        memberships = [
            _text(team.get("team_name")) for team in teams
            if any(_key(member) == _key(name) for member in team.get("assignees") or [])
        ]
        employee_rows.append([
            name, ", ".join(memberships) or "Unassigned",
            "Resigned" if bool((resources.get(_key(name)) or {}).get("resigned")) else "Active",
            "Yes" if _key(name) in support else "No",
        ])
    _write_sheet(
        wb, "Employees", ["Employee", "Teams", "Resource Status", "Support Resource"],
        employee_rows, "UtilizationEmployees",
    )

    for ws in (worklog_ws, booked_ws, leave_ws):
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith(("http://", "https://")):
                    cell.hyperlink = cell.value
                    cell.style = "Hyperlink"
    metadata = wb.create_sheet("Export Info", 0)
    metadata.append(["Employee Capacity & Utilization Export"])
    metadata.append(["Month", month])
    metadata.append([
        "Logged Hours Scope",
        "Work logged on their assigned subtasks" if scope == "assigned" else "All work logged by employee",
    ])
    metadata.append(["Include Leaves in Logged Hours", "Yes" if include_leaves else "No"])
    metadata.append(["Employees", len(names)])
    metadata.append(["Canonical Run", payload.get("canonical_run_id")])
    metadata.append(["Generated At", payload.get("generated_at")])
    metadata.column_dimensions["A"].width = 26
    metadata.column_dimensions["B"].width = 42
    metadata["A1"].font = Font(size=16, bold=True, color="173F70")
    metadata.merge_cells("A1:B1")

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output, f"employee_capacity_utilization_{month}.xlsx"
