"""
Generate an interactive gantt chart HTML report from nested view.xlsx.
"""
from __future__ import annotations

import json
import os
from datetime import date, datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

from canonical_report_data import load_canonical_actuals_by_issue, load_nested_rows_from_canonical, resolve_canonical_run_id

REQUIRED_HEADERS = [
    "Aspect",
    "Man-days",
    "Man-hours",
    "Actual Hours",
    "Planned Start Date",
    "Planned End Date",
]

DEFAULT_INPUT_XLSX = "nested view.xlsx"
DEFAULT_OUTPUT_HTML = "gantt_chart_report.html"
DEFAULT_WORK_ITEMS_XLSX = "1_jira_work_items_export.xlsx"


def _resolve_path(value: str, base_dir: Path) -> Path:
    path = Path(value)
    if path.is_absolute():
        return path
    return base_dir / path


def _to_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _to_number_or_blank(value):
    if value in (None, ""):
        return ""
    try:
        return round(float(value), 2)
    except (TypeError, ValueError):
        return ""


def _parse_to_iso_date(value) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()

    text = _to_text(value)
    if not text:
        return ""

    formats = (
        "%Y-%m-%d",
        "%Y/%m/%d",
        "%d-%b-%Y",
        "%d-%B-%Y",
        "%m/%d/%Y",
        "%d/%m/%Y",
    )
    for fmt in formats:
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue

    try:
        parsed = datetime.fromisoformat(text.replace("Z", "+00:00"))
        return parsed.date().isoformat()
    except ValueError:
        return ""


def _row_type_from_level(level: int) -> str:
    mapping = {
        1: "project",
        2: "product",
        3: "rmi",
        4: "story",
        5: "subtask",
        6: "assignee",
    }
    return mapping.get(level, "unknown")


def _project_key_from_aspect(aspect: str) -> str:
    text = _to_text(aspect)
    if " - " in text:
        return text.split(" - ", 1)[0].strip()
    return text.strip()


def _detect_bug_label(row_type: str, aspect: str) -> bool:
    if row_type not in ("story", "subtask"):
        return False
    return "bug" in _to_text(aspect).lower()


def _type_label(row_type: str, aspect: str) -> str:
    if row_type == "project":
        return "Project"
    if row_type == "product":
        return "Product Categorization"
    if row_type == "rmi":
        return "Epic"
    if row_type == "story":
        return "Bug" if _detect_bug_label("story", aspect) else "Story"
    if row_type == "subtask":
        return "Bug" if _detect_bug_label("subtask", aspect) else "Subtask"
    if row_type == "assignee":
        return "Assignee"
    return row_type.capitalize()


def _load_actual_date_lookup(work_items_path: Path) -> dict:
    if not work_items_path.exists():
        return {}

    wb = load_workbook(work_items_path, read_only=True, data_only=True)
    ws = wb.active
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header:
        wb.close()
        return {}

    headers = [_to_text(c) for c in header]
    idx = {h: i for i, h in enumerate(headers)}
    required = [
        "project_key",
        "jira_issue_type",
        "summary",
        "parent_issue_key",
        "issue_key",
        "actual_start_date",
        "actual_end_date",
    ]
    if any(name not in idx for name in required):
        wb.close()
        return {}

    rows_by_key: dict[str, dict] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        issue_key = _to_text(row[idx["issue_key"]])
        if not issue_key:
            continue
        rows_by_key[issue_key] = {
            "project_key": _to_text(row[idx["project_key"]]),
            "jira_issue_type": _to_text(row[idx["jira_issue_type"]]).lower(),
            "summary": _to_text(row[idx["summary"]]),
            "parent_issue_key": _to_text(row[idx["parent_issue_key"]]),
            "actual_start": _parse_to_iso_date(row[idx["actual_start_date"]]),
            "actual_end": _parse_to_iso_date(row[idx["actual_end_date"]]),
        }

    epic_lookup: dict[tuple, tuple[str, str]] = {}
    story_lookup: dict[tuple, tuple[str, str]] = {}
    subtask_lookup: dict[tuple, tuple[str, str]] = {}

    for issue_key, rec in rows_by_key.items():
        typ = rec["jira_issue_type"]
        project_key = rec["project_key"]
        summary = rec["summary"]
        parent_key = rec["parent_issue_key"]
        actual_pair = (rec["actual_start"], rec["actual_end"])

        if "epic" in typ:
            epic_lookup[(project_key, summary)] = actual_pair

    for issue_key, rec in rows_by_key.items():
        typ = rec["jira_issue_type"]
        if typ != "story":
            continue
        project_key = rec["project_key"]
        summary = rec["summary"]
        parent_key = rec["parent_issue_key"]
        epic_summary = rows_by_key.get(parent_key, {}).get("summary", "")
        story_lookup[(project_key, epic_summary, summary)] = (rec["actual_start"], rec["actual_end"])

    for issue_key, rec in rows_by_key.items():
        typ = rec["jira_issue_type"]
        if typ not in ("sub-task", "subtask"):
            continue
        project_key = rec["project_key"]
        summary = rec["summary"]
        parent_key = rec["parent_issue_key"]
        story = rows_by_key.get(parent_key, {})
        story_summary = story.get("summary", "")
        epic_key = story.get("parent_issue_key", "")
        epic_summary = rows_by_key.get(epic_key, {}).get("summary", "")
        subtask_lookup[(project_key, epic_summary, story_summary, summary)] = (
            rec["actual_start"],
            rec["actual_end"],
        )

    wb.close()
    return {
        "epic": epic_lookup,
        "story": story_lookup,
        "subtask": subtask_lookup,
    }


def _load_rows(input_path: Path, work_items_path: Path) -> list[dict]:
    if not input_path.exists():
        raise FileNotFoundError(f"Nested view workbook not found: {input_path}")

    wb = load_workbook(input_path, read_only=False, data_only=True)
    ws = wb["NestedView"] if "NestedView" in wb.sheetnames else wb.active

    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header:
        wb.close()
        raise ValueError("Nested view workbook has no header row.")

    found_headers = [_to_text(cell) for cell in header]
    header_index = {h.lower(): i for i, h in enumerate(found_headers)}
    missing = [h for h in REQUIRED_HEADERS if h.lower() not in header_index]
    if missing:
        wb.close()
        raise ValueError(
            "Nested view workbook headers are missing required columns. "
            f"Required: {REQUIRED_HEADERS}, missing: {missing}, found: {found_headers}"
        )

    idx_aspect = header_index["aspect"]
    idx_man_days = header_index["man-days"]
    idx_man_hours = header_index["man-hours"]
    idx_actual_hours = header_index["actual hours"]
    idx_planned_start = header_index["planned start date"]
    idx_planned_end = header_index["planned end date"]

    actual_lookup = _load_actual_date_lookup(work_items_path)
    rows: list[dict] = []
    stack: dict[int, int] = {}
    node_by_id: dict[int, dict] = {}
    next_id = 1

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        level = int(getattr(ws.row_dimensions[row_idx], "outlineLevel", 0) or 0)
        if level <= 0:
            level = 1

        for key in list(stack):
            if key >= level:
                del stack[key]

        parent_id = stack.get(level - 1)
        row_id = next_id
        next_id += 1
        stack[level] = row_id

        aspect = _to_text(row[idx_aspect] if len(row) > idx_aspect else "")
        row_type = _row_type_from_level(level)
        type_label = _type_label(row_type, aspect)
        planned_start = _parse_to_iso_date(row[idx_planned_start] if len(row) > idx_planned_start else "")
        planned_end = _parse_to_iso_date(row[idx_planned_end] if len(row) > idx_planned_end else "")

        start_ordinal = None
        end_ordinal = None
        duration_days = ""
        if planned_start and planned_end:
            try:
                start_date = datetime.strptime(planned_start, "%Y-%m-%d").date()
                end_date = datetime.strptime(planned_end, "%Y-%m-%d").date()
                if start_date <= end_date:
                    start_ordinal = start_date.toordinal()
                    end_ordinal = end_date.toordinal()
                    duration_days = (end_date - start_date).days + 1
            except ValueError:
                pass

        project_key = ""
        epic_name = ""
        story_name = ""
        if level == 1:
            project_key = _project_key_from_aspect(aspect)
        elif parent_id:
            parent = node_by_id.get(parent_id)
            if parent:
                project_key = parent.get("project_key", "")
                if row_type == "story":
                    epic_name = parent.get("aspect", "") if parent.get("row_type") == "rmi" else parent.get("epic_name", "")
                elif row_type == "subtask":
                    story_name = parent.get("aspect", "") if parent.get("row_type") == "story" else parent.get("story_name", "")
                    epic_name = parent.get("epic_name", "")
                else:
                    epic_name = parent.get("epic_name", "")
                    story_name = parent.get("story_name", "")

        if row_type == "rmi":
            epic_name = aspect
        if row_type == "story":
            story_name = aspect

        actual_start = ""
        actual_end = ""
        if row_type == "rmi":
            actual_start, actual_end = actual_lookup.get("epic", {}).get((project_key, aspect), ("", ""))
        elif row_type == "story":
            key = (project_key, epic_name, aspect)
            actual_start, actual_end = actual_lookup.get("story", {}).get(key, ("", ""))
        elif row_type == "subtask":
            key = (project_key, epic_name, story_name, aspect)
            actual_start, actual_end = actual_lookup.get("subtask", {}).get(key, ("", ""))

        actual_start_ordinal = None
        actual_end_ordinal = None
        actual_duration_days = ""
        if actual_start and actual_end:
            try:
                actual_start_date = datetime.strptime(actual_start, "%Y-%m-%d").date()
                actual_end_date = datetime.strptime(actual_end, "%Y-%m-%d").date()
                if actual_start_date <= actual_end_date:
                    actual_start_ordinal = actual_start_date.toordinal()
                    actual_end_ordinal = actual_end_date.toordinal()
                    actual_duration_days = (actual_end_date - actual_start_date).days + 1
            except ValueError:
                pass

        row_data = {
                "id": row_id,
                "parent_id": parent_id,
                "level": level,
                "row_type": row_type,
                "type_label": type_label,
                "aspect": aspect,
                "man_days": _to_number_or_blank(row[idx_man_days] if len(row) > idx_man_days else ""),
                "man_hours": _to_number_or_blank(row[idx_man_hours] if len(row) > idx_man_hours else ""),
                "actual_hours": _to_number_or_blank(row[idx_actual_hours] if len(row) > idx_actual_hours else ""),
                "planned_start": planned_start,
                "planned_end": planned_end,
                "start_ordinal": start_ordinal,
                "end_ordinal": end_ordinal,
                "duration_days": duration_days,
                "has_range": bool(start_ordinal and end_ordinal),
                "actual_start": actual_start,
                "actual_end": actual_end,
                "actual_start_ordinal": actual_start_ordinal,
                "actual_end_ordinal": actual_end_ordinal,
                "actual_duration_days": actual_duration_days,
                "has_actual_range": bool(actual_start_ordinal and actual_end_ordinal),
                "project_key": project_key,
                "epic_name": epic_name,
                "story_name": story_name,
            }
        rows.append(row_data)
        node_by_id[row_id] = row_data

    wb.close()
    return _merge_subtask_assignee_rows(rows)


def _load_rows_from_canonical(db_path: Path, run_id: str = "") -> list[dict]:
    effective_run_id = resolve_canonical_run_id(db_path, run_id)
    if not effective_run_id:
        return []
    nested_rows = load_nested_rows_from_canonical(db_path, effective_run_id)
    actuals_by_issue = load_canonical_actuals_by_issue(db_path, effective_run_id)
    rows: list[dict] = []
    for row in nested_rows:
        row_type = _to_text(row.get("row_type"))
        if row_type not in {"project", "product", "rmi", "story", "subtask", "assignee"}:
            continue
        jira_key = _to_text(row.get("jira_key")).upper()
        actual = actuals_by_issue.get(jira_key, {})
        actual_start = _to_text(actual.get("first_worklog_date"))
        actual_end = _to_text(actual.get("actual_complete_date") or actual.get("last_worklog_date"))
        actual_start_ordinal = None
        actual_end_ordinal = None
        actual_duration_days = ""
        if actual_start and actual_end:
            try:
                actual_start_date = date.fromisoformat(actual_start)
                actual_end_date = date.fromisoformat(actual_end)
                if actual_start_date <= actual_end_date:
                    actual_start_ordinal = actual_start_date.toordinal()
                    actual_end_ordinal = actual_end_date.toordinal()
                    actual_duration_days = (actual_end_date - actual_start_date).days + 1
            except ValueError:
                pass
        rows.append(
            {
                "id": row.get("id"),
                "parent_id": row.get("parent_id"),
                "level": row.get("level"),
                "row_type": row_type,
                "type_label": _type_label(row_type, _to_text(row.get("aspect"))),
                "aspect": _to_text(row.get("aspect")),
                "man_days": _to_number_or_blank(row.get("man_days") or row.get("approved_days")),
                "man_hours": _to_number_or_blank(row.get("man_hours") or row.get("approved_hours")),
                "actual_hours": _to_number_or_blank(row.get("actual_hours")),
                "planned_start": _to_text(row.get("planned_start")),
                "planned_end": _to_text(row.get("planned_end")),
                "start_ordinal": row.get("start_ordinal"),
                "end_ordinal": row.get("end_ordinal"),
                "duration_days": row.get("duration_days") or "",
                "has_range": bool(row.get("start_ordinal") and row.get("end_ordinal")),
                "actual_start": actual_start,
                "actual_end": actual_end,
                "actual_start_ordinal": actual_start_ordinal,
                "actual_end_ordinal": actual_end_ordinal,
                "actual_duration_days": actual_duration_days,
                "has_actual_range": bool(actual_start_ordinal and actual_end_ordinal),
                "project_key": _to_text(row.get("project_key")),
                "epic_name": "",
                "story_name": "",
                "jira_key": jira_key,
                "assignee": _to_text(row.get("assignee")),
            }
        )
    return rows


def _merge_subtask_assignee_rows(rows: list[dict]) -> list[dict]:
    children_by_parent: dict[int | None, list[dict]] = {}
    by_id = {row["id"]: row for row in rows}
    for row in rows:
        children_by_parent.setdefault(row.get("parent_id"), []).append(row)

    assignee_by_subtask: dict[int, str] = {}
    for row in rows:
        if row.get("row_type") != "subtask":
            continue
        assignees = [
            child for child in children_by_parent.get(row["id"], [])
            if child.get("row_type") == "assignee"
        ]
        if assignees:
            # Expected 1:1 subtask-assignee cardinality.
            assignee_by_subtask[row["id"]] = _to_text(assignees[0].get("aspect"))

    merged: list[dict] = []
    for row in rows:
        if row.get("row_type") == "assignee":
            parent = by_id.get(row.get("parent_id"))
            if parent and parent.get("row_type") == "subtask":
                continue
        item = dict(row)
        if item.get("row_type") == "subtask":
            item["assignee"] = assignee_by_subtask.get(item["id"], "")
        else:
            item["assignee"] = ""
        merged.append(item)
    return merged


def _build_html(data: dict) -> str:
    payload = json.dumps(data, ensure_ascii=True)
    return f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>gantt chart</title>
  <link href="https://fonts.googleapis.com/icon?family=Material+Icons+Outlined" rel="stylesheet">
  <script>
    (function () {{
      const storageKey = "rmi-gantt-report-theme";
      const stored = localStorage.getItem(storageKey);
      const theme = (stored === "dark" || stored === "light") ? stored : "light";
      document.documentElement.setAttribute("data-theme", theme);
    }})();
  </script>
  <style>
    :root {{
      --bg: #f3f6f9;
      --panel: #ffffff;
      --text: #1f2937;
      --muted: #6b7280;
      --line: #dbe3ea;
      --head: #0f4c5c;
      --head-text: #ffffff;
      --left-1: 430px;
      --left-2: 190px;
    }}
    :root {{ color-scheme: light; }}
    html[data-theme="dark"] {{ color-scheme: dark; }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      font-family: "Segoe UI", Tahoma, Verdana, sans-serif;
      color: var(--text);
      background:
        radial-gradient(1000px 280px at 10% -5%, #d7eef6 0%, transparent 60%),
        linear-gradient(180deg, #eef4f7, var(--bg));
    }}
    .page {{
      max-width: 1800px;
      margin: 0 auto;
      padding: 16px;
    }}
    .panel {{
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 12px;
      padding: 14px 16px;
      margin-bottom: 12px;
    }}
    .title {{
      margin: 0 0 6px;
      font-size: 1.45rem;
      font-weight: 700;
      color: #0b3142;
    }}
    .meta {{
      margin: 0;
      color: var(--muted);
      font-size: 1rem;
    }}
    .toolbar {{
      margin-top: 10px;
      display: flex;
      flex-wrap: wrap;
      gap: 10px;
      align-items: center;
    }}
    .search {{
      min-width: 320px;
      flex: 1 1 420px;
      border: 1px solid #b6c7d2;
      border-radius: 8px;
      padding: 9px 12px;
      font-size: 0.98rem;
      color: #12313f;
      background: #fff;
    }}
    .search:focus {{
      outline: none;
      border-color: #2a6274;
      box-shadow: 0 0 0 2px rgba(42, 98, 116, 0.15);
    }}
    .btn {{
      border: 1px solid #255f73;
      background: #0f4c5c;
      color: #fff;
      border-radius: 8px;
      padding: 9px 14px;
      cursor: pointer;
      font-size: 0.94rem;
    }}
    .btn .material-icons-outlined {{
      font-size: 0.92rem;
      vertical-align: -2px;
      margin-right: 0.2rem;
    }}
    #theme-toggle {{
      background: rgba(191, 219, 254, 0.45);
      border-color: #93c5fd;
      color: #1e3a8a;
    }}
    #theme-toggle .material-icons-outlined {{
      color: #1d4ed8;
    }}
    #theme-toggle:hover {{
      background: rgba(191, 219, 254, 0.62);
    }}
    html[data-theme="dark"] #theme-toggle {{
      background: rgba(254, 243, 199, 0.45);
      border-color: #fcd34d;
      color: #fff;
    }}
    html[data-theme="dark"] #theme-toggle .material-icons-outlined {{
      color: #facc15;
    }}
    html[data-theme="dark"] #theme-toggle:hover {{
      background: rgba(254, 243, 199, 0.62);
    }}
    .btn.alt {{
      background: #fff;
      color: #255f73;
    }}
    .btn:hover {{ filter: brightness(1.05); }}

    .gantt-wrap {{
      background: #fff;
      border: 1px solid var(--line);
      border-radius: 12px;
      overflow: auto;
      max-height: 82vh;
    }}
    .gantt {{
      min-width: 980px;
    }}
    .grid-row {{
      display: grid;
      grid-template-columns: var(--left-1) var(--left-2) 1fr;
      min-width: max-content;
    }}
    .cell {{
      border-top: 1px solid var(--line);
      padding: 10px 12px;
      font-size: 0.98rem;
      background: #fff;
    }}
    .head .cell {{
      border-top: 0;
      background: var(--head);
      color: var(--head-text);
      font-weight: 700;
      white-space: nowrap;
      position: sticky;
      top: 0;
      z-index: 30;
      padding-top: 10px;
      padding-bottom: 10px;
    }}
    .timeline-head {{
      position: sticky;
      top: 0;
      z-index: 25;
      overflow: hidden;
      padding: 0 !important;
    }}
    .timeline-months {{
      position: relative;
      min-height: 52px;
      background: #f7fbff;
      color: #243745;
      border-bottom: 1px solid #d2dce5;
    }}
    .month-block {{
      position: absolute;
      top: 0;
      bottom: 0;
      border-right: 1px solid #d2dce5;
      background: rgba(214, 229, 241, 0.35);
      padding: 6px 8px;
      font-size: 0.82rem;
      color: #2b4556;
      white-space: nowrap;
      overflow: hidden;
      text-overflow: ellipsis;
    }}

    .sticky-1 {{
      position: sticky;
      left: 0;
      z-index: 20;
      background: #fff;
      border-right: 1px solid var(--line);
    }}
    .sticky-2 {{
      position: sticky;
      left: var(--left-1);
      z-index: 19;
      background: #fff;
      border-right: 1px solid var(--line);
    }}
    .head .sticky-1, .head .sticky-2 {{
      z-index: 40;
      background: var(--head);
    }}
    .head .sticky-2 {{
      z-index: 39;
    }}
    .head .timeline-head {{
      z-index: 38;
    }}

    .aspect-cell {{
      display: flex;
      align-items: center;
      gap: 6px;
      min-width: 0;
    }}
    .toggle {{
      width: 22px;
      height: 22px;
      border: 1px solid #a5b7c3;
      border-radius: 5px;
      background: #fff;
      color: #294b5a;
      cursor: pointer;
      line-height: 1;
      font-size: 14px;
      padding: 0;
      flex: 0 0 auto;
    }}
    .toggle.placeholder {{
      visibility: hidden;
      cursor: default;
    }}
    .node-text {{
      display: block;
      flex: 1 1 auto;
      min-width: 0;
      overflow: hidden;
      text-overflow: ellipsis;
      white-space: nowrap;
    }}
    .lvl-1 .node-text {{ font-weight: 700; color: #0b3142; }}
    .lvl-2 .node-text {{ font-weight: 700; color: #264653; }}
    .lvl-3 .node-text {{ font-weight: 600; color: #2a6274; }}
    .lvl-4 .node-text {{ font-weight: 600; }}
    .lvl-5 .node-text {{ color: #374151; }}
    .lvl-6 .node-text {{ color: #4b5563; font-style: italic; }}

    .row-type-project .sticky-1, .row-type-project .sticky-2 {{ background: #eaf4ff; }}
    .row-type-product .sticky-1, .row-type-product .sticky-2 {{ background: #f5f9c9; }}
    .row-type-rmi .sticky-1, .row-type-rmi .sticky-2 {{ background: #f1e9ff; }}
    .row-type-story .sticky-1, .row-type-story .sticky-2 {{ background: #e9f9e9; }}
    .row-type-subtask .sticky-1, .row-type-subtask .sticky-2 {{ background: #f3fff3; }}
    .row-type-bug .sticky-1, .row-type-bug .sticky-2 {{ background: #ffeedd; }}
    .row-type-assignee .sticky-1, .row-type-assignee .sticky-2 {{ background: #f8fafc; }}

    .type-pill {{
      display: inline-flex;
      align-items: center;
      font-size: 0.8rem;
      line-height: 1.1;
      padding: 4px 8px;
      border-radius: 999px;
      border: 1px solid #d6e2eb;
      background: #f9fcff;
      color: #334b5c;
      font-weight: 700;
      min-width: 86px;
      justify-content: center;
    }}
    .project-pill {{
      display: inline-block;
      font-size: 0.82rem;
      line-height: 1;
      padding: 3px 6px;
      border-radius: 999px;
      border: 1px solid #ecdba3;
      background: #fff8df;
      color: #7a6100;
      margin-left: 8px;
    }}
    .assignee-pill {{
      display: inline-flex;
      align-items: center;
      font-size: 0.8rem;
      line-height: 1;
      padding: 3px 7px;
      border-radius: 999px;
      border: 1px solid #bfdbfe;
      background: #eff6ff;
      color: #1d4ed8;
      font-weight: 700;
      margin-left: 8px;
      max-width: 180px;
      overflow: hidden;
      text-overflow: ellipsis;
      white-space: nowrap;
    }}

    .timeline-cell {{
      position: relative;
      min-height: 52px;
      padding: 0;
      background: #fff;
      overflow: hidden;
    }}
    .grid-line {{
      position: absolute;
      top: 0;
      bottom: 0;
      width: 1px;
      background: #eef3f7;
      pointer-events: none;
    }}
    .bar {{
      position: absolute;
      height: 15px;
      border-radius: 999px;
      border: 1px solid transparent;
      min-width: 2px;
      cursor: default;
    }}
    .bar-planned {{
      top: 9px;
    }}
    .bar-actual {{
      top: 28px;
    }}
    .bar.project {{
      background: rgba(37, 99, 235, 0.16);
      border-color: rgba(37, 99, 235, 0.6);
    }}
    .bar.product {{
      background: rgba(120, 113, 108, 0.16);
      border-color: rgba(120, 113, 108, 0.45);
    }}
    .bar.rmi {{
      background: rgba(147, 51, 234, 0.16);
      border-color: rgba(147, 51, 234, 0.58);
    }}
    .bar.story {{
      background: rgba(22, 163, 74, 0.16);
      border-color: rgba(22, 163, 74, 0.58);
    }}
    .bar.subtask {{
      background: rgba(14, 116, 144, 0.16);
      border-color: rgba(14, 116, 144, 0.58);
    }}
    .bar.assignee {{
      background: rgba(107, 114, 128, 0.16);
      border-color: rgba(107, 114, 128, 0.55);
    }}
    .bar.actual-rmi {{
      background: rgba(147, 51, 234, 0.34);
      border-color: rgba(107, 33, 168, 0.8);
    }}
    .bar.actual-story {{
      background: rgba(22, 163, 74, 0.35);
      border-color: rgba(21, 128, 61, 0.8);
    }}
    .bar.actual-subtask {{
      background: rgba(14, 116, 144, 0.35);
      border-color: rgba(14, 116, 144, 0.85);
    }}
    .empty {{
      padding: 22px;
      color: #607282;
      font-size: 1rem;
    }}
    .legend {{
      margin-top: 8px;
      display: flex;
      flex-wrap: wrap;
      gap: 8px;
      color: #506575;
      font-size: 0.9rem;
    }}
    .legend-pill {{
      display: inline-flex;
      align-items: center;
      gap: 6px;
      border: 1px solid #d8e4ee;
      border-radius: 999px;
      padding: 3px 8px;
      background: #f8fbff;
    }}
    .legend-chip {{
      width: 18px;
      height: 8px;
      border-radius: 999px;
      border: 1px solid #8aa2b4;
      background: rgba(37, 99, 235, 0.2);
    }}
    .legend-chip.actual {{
      background: rgba(37, 99, 235, 0.38);
      border-color: #436b8a;
    }}
    html[data-theme="dark"] body {{
      color: #e5e7eb;
      background:
        radial-gradient(1000px 280px at 10% -5%, #102a43 0%, transparent 65%),
        linear-gradient(180deg, #0f172a, #0b1220);
    }}
    html[data-theme="dark"] .panel,
    html[data-theme="dark"] .gantt-wrap {{
      background: #111827;
      border-color: #1f2937;
    }}
    html[data-theme="dark"] .title {{ color: #f3f4f6; }}
    html[data-theme="dark"] .meta,
    html[data-theme="dark"] .legend {{ color: #94a3b8; }}
    html[data-theme="dark"] .search {{
      background: #0f172a;
      color: #e5e7eb;
      border-color: #374151;
    }}
    html[data-theme="dark"] .search:focus {{
      border-color: #60a5fa;
      box-shadow: 0 0 0 2px rgba(96, 165, 250, 0.25);
    }}
    html[data-theme="dark"] .btn.alt {{
      background: #0f172a;
      color: #cbd5e1;
      border-color: #374151;
    }}
    html[data-theme="dark"] .head .cell,
    html[data-theme="dark"] .head .sticky-1,
    html[data-theme="dark"] .head .sticky-2 {{
      background: #1e293b;
      color: #f3f4f6;
    }}
    html[data-theme="dark"] .timeline-months {{
      background: #0f172a;
      border-bottom-color: #334155;
      color: #cbd5e1;
    }}
    html[data-theme="dark"] .month-block {{
      background: rgba(30, 41, 59, 0.7);
      border-right-color: #334155;
      color: #cbd5e1;
    }}
    html[data-theme="dark"] .cell {{
      background: #111827;
      border-top-color: #334155;
      color: #d1d5db;
    }}
    html[data-theme="dark"] .sticky-1,
    html[data-theme="dark"] .sticky-2 {{
      background: #111827;
      border-right-color: #334155;
    }}
    html[data-theme="dark"] .toggle {{
      background: #111827;
      border-color: #475569;
      color: #d1d5db;
    }}
    html[data-theme="dark"] .lvl-1 .node-text,
    html[data-theme="dark"] .lvl-2 .node-text,
    html[data-theme="dark"] .lvl-3 .node-text,
    html[data-theme="dark"] .lvl-4 .node-text,
    html[data-theme="dark"] .lvl-5 .node-text,
    html[data-theme="dark"] .lvl-6 .node-text {{
      color: #e5e7eb;
    }}
    html[data-theme="dark"] .row-type-project .sticky-1, html[data-theme="dark"] .row-type-project .sticky-2 {{ background: #102a43; }}
    html[data-theme="dark"] .row-type-product .sticky-1, html[data-theme="dark"] .row-type-product .sticky-2 {{ background: #3a330a; }}
    html[data-theme="dark"] .row-type-rmi .sticky-1, html[data-theme="dark"] .row-type-rmi .sticky-2 {{ background: #2e1f4f; }}
    html[data-theme="dark"] .row-type-story .sticky-1, html[data-theme="dark"] .row-type-story .sticky-2 {{ background: #0f3a2c; }}
    html[data-theme="dark"] .row-type-subtask .sticky-1, html[data-theme="dark"] .row-type-subtask .sticky-2 {{ background: #1f2d4d; }}
    html[data-theme="dark"] .row-type-bug .sticky-1, html[data-theme="dark"] .row-type-bug .sticky-2 {{ background: #3f2a1d; }}
    html[data-theme="dark"] .row-type-assignee .sticky-1, html[data-theme="dark"] .row-type-assignee .sticky-2 {{ background: #1e293b; }}
    html[data-theme="dark"] .type-pill {{
      background: #1e293b;
      border-color: #334155;
      color: #cbd5e1;
    }}
    html[data-theme="dark"] .project-pill {{
      background: #3a330a;
      border-color: #a16207;
      color: #fde68a;
    }}
    html[data-theme="dark"] .assignee-pill {{
      background: #172554;
      border-color: #1d4ed8;
      color: #bfdbfe;
    }}
    html[data-theme="dark"] .timeline-cell {{
      background: #111827;
    }}
    html[data-theme="dark"] .grid-line {{ background: #1f2937; }}
    html[data-theme="dark"] .empty {{ color: #94a3b8; }}
    html[data-theme="dark"] .legend-pill {{
      border-color: #334155;
      background: #0f172a;
      color: #cbd5e1;
    }}
  </style>
  <link rel="stylesheet" href="shared-nav.css">
</head>
<body>
  <div class="page">
    <section class="panel">
      <h1 class="title">gantt chart</h1>
      <p class="meta">Generated: <span id="generated-at"></span> | Source: <span id="source-file"></span> | Visible Rows: <span id="row-count"></span></p>
      <div class="toolbar">
        <input class="search" id="search-input" type="text" placeholder="Search Aspect or Type">
        <button class="btn alt" type="button" id="theme-toggle"><span class="material-icons-outlined" aria-hidden="true">dark_mode</span>Dark mode</button>
        <button class="btn" type="button" id="expand-all">Expand All</button>
        <button class="btn alt" type="button" id="collapse-projects">Collapse To Projects</button>
        <button class="btn alt" type="button" id="fit-range">Fit To Range</button>
        <button class="btn alt" type="button" id="reset-zoom">Reset Zoom</button>
      </div>
      <div class="legend">
        <span class="legend-pill"><span class="legend-chip"></span>Planned date range</span>
        <span class="legend-pill"><span class="legend-chip actual"></span>Actual date range</span>
      </div>
    </section>
    <section class="gantt-wrap" id="gantt-wrap">
      <div id="gantt-root" class="gantt"></div>
    </section>
  </div>
  <script>
    const reportData = {payload};
    const allRows = Array.isArray(reportData.rows) ? reportData.rows : [];
    const generatedNode = document.getElementById("generated-at");
    const sourceNode = document.getElementById("source-file");
    const rowCountNode = document.getElementById("row-count");
    const searchInput = document.getElementById("search-input");
    const themeToggleButton = document.getElementById("theme-toggle");
    const ganttRoot = document.getElementById("gantt-root");
    const ganttWrap = document.getElementById("gantt-wrap");
    const THEME_STORAGE_KEY = "rmi-gantt-report-theme";

    generatedNode.textContent = reportData.generated_at || "-";
    sourceNode.textContent = reportData.source_file || "-";

    const childrenByParent = new Map();
    const rowsById = new Map();
    const collapsed = new Set();
    const DEFAULT_DAY_PX = 12;
    let dayPx = DEFAULT_DAY_PX;
    let query = "";

    function setTheme(theme) {{
      const nextTheme = theme === "dark" ? "dark" : "light";
      document.documentElement.setAttribute("data-theme", nextTheme);
      localStorage.setItem(THEME_STORAGE_KEY, nextTheme);
      updateThemeToggleLabel(nextTheme);
    }}

    function getInitialTheme() {{
      const stored = localStorage.getItem(THEME_STORAGE_KEY);
      if (stored === "dark" || stored === "light") {{
        return stored;
      }}
      return "light";
    }}

    function updateThemeToggleLabel(theme) {{
      if (!themeToggleButton) {{
        return;
      }}
      themeToggleButton.innerHTML = theme === "dark"
        ? '<span class="material-icons-outlined" aria-hidden="true">light_mode</span>Light mode'
        : '<span class="material-icons-outlined" aria-hidden="true">dark_mode</span>Dark mode';
    }}

    function initializeThemeToggle() {{
      const initialTheme = getInitialTheme();
      document.documentElement.setAttribute("data-theme", initialTheme);
      updateThemeToggleLabel(initialTheme);
      if (!themeToggleButton) {{
        return;
      }}
      themeToggleButton.addEventListener("click", () => {{
        const currentTheme = document.documentElement.getAttribute("data-theme") === "dark" ? "dark" : "light";
        setTheme(currentTheme === "dark" ? "light" : "dark");
      }});
    }}

    initializeThemeToggle();

    for (const row of allRows) {{
      rowsById.set(row.id, row);
      const parent = row.parent_id || null;
      if (!childrenByParent.has(parent)) {{
        childrenByParent.set(parent, []);
      }}
      childrenByParent.get(parent).push(row.id);
    }}

    function childrenOf(id) {{
      return childrenByParent.get(id) || [];
    }}

    function hasChildren(id) {{
      return childrenOf(id).length > 0;
    }}

    function ancestorsVisible(row) {{
      let current = row;
      while (current && current.parent_id) {{
        if (collapsed.has(current.parent_id)) {{
          return false;
        }}
        current = rowsById.get(current.parent_id) || null;
      }}
      return true;
    }}

    function searchMatch(row, q) {{
      if (!q) {{
        return true;
      }}
      const blob = [row.aspect || "", row.type_label || "", row.row_type || "", row.assignee || ""].join(" ").toLowerCase();
      return blob.includes(q);
    }}

    function collectAncestors(id, keepSet) {{
      let current = rowsById.get(id);
      while (current && current.parent_id) {{
        keepSet.add(current.parent_id);
        current = rowsById.get(current.parent_id) || null;
      }}
    }}

    function filteredVisibleRows() {{
      const q = query.trim().toLowerCase();
      if (!q) {{
        return allRows.filter((row) => ancestorsVisible(row));
      }}
      const keep = new Set();
      for (const row of allRows) {{
        if (searchMatch(row, q)) {{
          keep.add(row.id);
          collectAncestors(row.id, keep);
        }}
      }}
      return allRows.filter((row) => keep.has(row.id) && ancestorsVisible(row));
    }}

    function monthLabelFromOrdinal(ord) {{
      const millis = (Number(ord) - 719163) * 86400000;
      const d = new Date(millis);
      return d.toLocaleString(undefined, {{ month: "short", year: "numeric" }});
    }}

    function buildMonthSegments(minOrd, maxOrd, totalDays) {{
      const segments = [];
      const minMillis = (Number(minOrd) - 719163) * 86400000;
      const maxMillis = (Number(maxOrd) - 719163) * 86400000;
      const minDate = new Date(minMillis);
      const maxDate = new Date(maxMillis);

      let cursor = new Date(minDate.getFullYear(), minDate.getMonth(), 1);
      if (cursor > minDate) {{
        cursor = new Date(minDate.getFullYear(), minDate.getMonth() - 1, 1);
      }}

      while (cursor <= maxDate) {{
        const start = new Date(cursor.getFullYear(), cursor.getMonth(), 1);
        const end = new Date(cursor.getFullYear(), cursor.getMonth() + 1, 0);
        const startOrd = Math.floor(start.getTime() / 86400000) + 719163;
        const endOrd = Math.floor(end.getTime() / 86400000) + 719163;
        const segStart = Math.max(startOrd, minOrd);
        const segEnd = Math.min(endOrd, maxOrd);
        if (segStart <= segEnd) {{
          const leftDays = segStart - minOrd;
          const widthDays = (segEnd - segStart) + 1;
          segments.push({{
            label: monthLabelFromOrdinal(startOrd),
            leftPct: (leftDays / totalDays) * 100,
            widthPct: (widthDays / totalDays) * 100,
          }});
        }}
        cursor = new Date(cursor.getFullYear(), cursor.getMonth() + 1, 1);
      }}
      return segments;
    }}

    function formatNumber(value) {{
      if (value === "" || value === null || value === undefined) {{
        return "-";
      }}
      const n = Number(value);
      if (!Number.isFinite(n)) {{
        return "-";
      }}
      return String(n.toFixed(2)).replace(/\\.00$/, "");
    }}

    function formatDataLabelDate(value) {{
      const raw = String(value || "").trim();
      if (!raw) {{
        return "-";
      }}
      const d = new Date(raw);
      if (Number.isNaN(d.getTime())) {{
        return raw;
      }}
      const day = String(d.getDate()).padStart(2, "0");
      const mon = d.toLocaleString(undefined, {{ month: "short" }});
      const year = String(d.getFullYear()).slice(-2);
      return `${{day}} ${{mon}} ${{year}}`;
    }}

    function render() {{
      const visibleRows = filteredVisibleRows();
      rowCountNode.textContent = String(visibleRows.length);

      const rangeRows = visibleRows.filter((r) => r.has_range);
      if (!visibleRows.length || !rangeRows.length) {{
        ganttRoot.innerHTML = '<div class="empty">No rows available for current filters.</div>';
        return;
      }}

      const minOrd = Math.min(...rangeRows.map((r) => Number(r.start_ordinal)));
      const maxOrd = Math.max(...rangeRows.map((r) => Number(r.end_ordinal)));
      const totalDays = Math.max(1, (maxOrd - minOrd) + 1);
      const timelineWidth = Math.max(760, Math.floor(totalDays * dayPx));
      const months = buildMonthSegments(minOrd, maxOrd, totalDays);

      const monthHtml = months.map((m) =>
        `<div class="month-block" style="left:${{m.leftPct}}%;width:${{m.widthPct}}%;">${{m.label}}</div>`
      ).join("");

      const lineEvery = Math.max(1, Math.floor(totalDays / 42));
      const gridLines = [];
      for (let day = 0; day <= totalDays; day += lineEvery) {{
        const pct = (day / totalDays) * 100;
        gridLines.push(`<span class="grid-line" style="left:${{pct}}%"></span>`);
      }}
      const gridLineHtml = gridLines.join("");

      const rowHtml = visibleRows.map((row) => {{
        const level = Number(row.level || 1);
        const pad = Math.max(0, (level - 1) * 18);
        const ids = childrenOf(row.id);
        const isBug = row.type_label === "Bug";
        const typeClass = isBug ? "bug" : (row.row_type || "unknown");
        const hasToggle = ids.length > 0;
        const toggleText = collapsed.has(row.id) ? "+" : "-";

        let barHtml = "";
        if (row.has_range) {{
          const leftDays = Number(row.start_ordinal) - minOrd;
          const spanDays = (Number(row.end_ordinal) - Number(row.start_ordinal)) + 1;
          const leftPct = (leftDays / totalDays) * 100;
          const widthPct = (Math.max(1, spanDays) / totalDays) * 100;
          const tip = [
            `Aspect: ${{row.aspect || "-"}}`,
            `Type: ${{row.type_label || "-"}}`,
            `Assignee: ${{row.assignee || "-"}}`,
            `Planned Start: ${{formatDataLabelDate(row.planned_start)}}`,
            `Planned End: ${{formatDataLabelDate(row.planned_end)}}`,
            `Duration: ${{row.duration_days || "-"}} day(s)`,
            `Actual Start: ${{formatDataLabelDate(row.actual_start)}}`,
            `Actual End: ${{formatDataLabelDate(row.actual_end)}}`,
            `Actual Duration: ${{row.actual_duration_days || "-"}} day(s)`,
            `Man-days: ${{formatNumber(row.man_days)}}`,
            `Man-hours: ${{formatNumber(row.man_hours)}}`,
            `Actual Hours: ${{formatNumber(row.actual_hours)}}`,
          ].join("\\n");
          barHtml = `<div class="bar bar-planned ${{row.row_type || "unknown"}}" style="left:${{leftPct}}%;width:${{widthPct}}%;" title="${{tip}}"></div>`;
          if (row.has_actual_range && (row.row_type === "rmi" || row.row_type === "story" || row.row_type === "subtask")) {{
            const actualLeftDays = Number(row.actual_start_ordinal) - minOrd;
            const actualSpanDays = (Number(row.actual_end_ordinal) - Number(row.actual_start_ordinal)) + 1;
            const actualLeftPct = (actualLeftDays / totalDays) * 100;
            const actualWidthPct = (Math.max(1, actualSpanDays) / totalDays) * 100;
            const actualCls = row.row_type === "rmi" ? "actual-rmi" : (row.row_type === "story" ? "actual-story" : "actual-subtask");
            barHtml += `<div class="bar bar-actual ${{actualCls}}" style="left:${{actualLeftPct}}%;width:${{actualWidthPct}}%;" title="${{tip}}"></div>`;
          }}
        }}

        const projectPill = row.row_type === "project" ? '<span class="project-pill">Project</span>' : "";
        const assigneePill = row.assignee ? `<span class="assignee-pill" title="${{row.assignee}}">${{row.assignee}}</span>` : "";
        const toggleButton = hasToggle
          ? `<button class="toggle" type="button" data-toggle-id="${{row.id}}">${{toggleText}}</button>`
          : '<button class="toggle placeholder" type="button" disabled>.</button>';

        return `
          <div class="grid-row lvl-${{level}} row-type-${{typeClass}}">
            <div class="cell sticky-1">
              <div class="aspect-cell" style="padding-left:${{pad}}px;">
                ${{toggleButton}}
                <span class="node-text" title="${{row.aspect || ""}}">${{row.aspect || ""}}</span>
                ${{projectPill}}
                ${{assigneePill}}
              </div>
            </div>
            <div class="cell sticky-2"><span class="type-pill">${{formatNumber(row.man_days)}} md</span></div>
            <div class="cell timeline-cell" style="width:${{timelineWidth}}px; min-height: 54px;">
              ${{gridLineHtml}}
              ${{barHtml}}
            </div>
          </div>
        `;
      }}).join("");

      ganttRoot.innerHTML = `
        <div class="gantt" style="width: calc(var(--left-1) + var(--left-2) + ${{timelineWidth}}px);">
          <div class="grid-row head">
            <div class="cell sticky-1">Aspect</div>
            <div class="cell sticky-2">Man-days</div>
            <div class="cell timeline-head">
              <div class="timeline-months">${{monthHtml}}</div>
            </div>
          </div>
          ${{rowHtml}}
        </div>
      `;

      for (const btn of ganttRoot.querySelectorAll("[data-toggle-id]")) {{
        btn.addEventListener("click", () => {{
          const id = Number(btn.getAttribute("data-toggle-id"));
          if (!Number.isFinite(id)) {{
            return;
          }}
          if (collapsed.has(id)) {{
            collapsed.delete(id);
          }} else {{
            collapsed.add(id);
          }}
          render();
        }});
      }}
    }}

    function collapseToProjects() {{
      collapsed.clear();
      for (const row of allRows) {{
        if (row.row_type === "project" && hasChildren(row.id)) {{
          collapsed.add(row.id);
        }}
      }}
      render();
    }}

    function expandAll() {{
      collapsed.clear();
      render();
    }}

    function fitToRange() {{
      const rows = filteredVisibleRows().filter((r) => r.has_range);
      if (!rows.length) {{
        return;
      }}
      const minOrd = Math.min(...rows.map((r) => Number(r.start_ordinal)));
      const maxOrd = Math.max(...rows.map((r) => Number(r.end_ordinal)));
      const totalDays = Math.max(1, (maxOrd - minOrd) + 1);
      const available = Math.max(380, ganttWrap.clientWidth - 560);
      dayPx = Math.max(2, Math.min(18, available / totalDays));
      render();
    }}

    function resetZoom() {{
      dayPx = DEFAULT_DAY_PX;
      render();
    }}

    searchInput.addEventListener("input", () => {{
      query = searchInput.value || "";
      render();
    }});
    document.getElementById("expand-all").addEventListener("click", expandAll);
    document.getElementById("collapse-projects").addEventListener("click", collapseToProjects);
    document.getElementById("fit-range").addEventListener("click", fitToRange);
    document.getElementById("reset-zoom").addEventListener("click", resetZoom);

    render();
  </script>
<script src="shared-nav.js"></script>
</body>
</html>
"""


def main() -> None:
    base_dir = Path(__file__).resolve().parent
    input_name = os.getenv("JIRA_NESTED_VIEW_XLSX_PATH", DEFAULT_INPUT_XLSX).strip() or DEFAULT_INPUT_XLSX
    output_name = os.getenv("JIRA_GANTT_HTML_PATH", DEFAULT_OUTPUT_HTML).strip() or DEFAULT_OUTPUT_HTML
    work_items_name = os.getenv("JIRA_EXPORT_XLSX_PATH", DEFAULT_WORK_ITEMS_XLSX).strip() or DEFAULT_WORK_ITEMS_XLSX
    db_name = os.getenv("JIRA_ASSIGNEE_HOURS_CAPACITY_DB_PATH", "assignee_hours_capacity.db").strip() or "assignee_hours_capacity.db"
    canonical_run_id = os.getenv("JIRA_CANONICAL_RUN_ID", "").strip()

    input_path = _resolve_path(input_name, base_dir)
    output_path = _resolve_path(output_name, base_dir)
    work_items_path = _resolve_path(work_items_name, base_dir)
    db_path = _resolve_path(db_name, base_dir)

    rows = _load_rows_from_canonical(db_path, canonical_run_id)
    source_file = "canonical_db"
    if not rows:
        rows = _load_rows(input_path, work_items_path)
        source_file = str(input_path)
    data = {
        "generated_at": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC"),
        "source_file": source_file,
        "rows": rows,
    }
    output_path.write_text(_build_html(data), encoding="utf-8")

    print(f"Source data: {source_file}")
    print(f"Rows loaded: {len(rows)}")
    print(f"Gantt chart report written: {output_path}")


if __name__ == "__main__":
    main()

