from __future__ import annotations

import sqlite3
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook

from generate_employee_performance_report import (
    DEFAULT_PERFORMANCE_SETTINGS,
    _apply_managed_project_display_names,
    _derive_actual_completion,
    _build_html,
    _build_payload,
    _init_performance_settings_db,
    _load_leave_issue_keys,
    _load_managed_project_display_names,
    _load_unplanned_leave_rows,
    _load_simple_scoring,
    _load_work_items,
    _load_worklogs,
    _load_performance_settings,
    _normalize_performance_settings,
    _precompute_simple_scoring,
    _resolve_employee_performance_source_mode,
    _resolve_runtime_paths,
    _save_performance_settings,
)
import report_server
from report_server import create_report_server_app


def _seed_canonical_run(db_path: Path, run_id: str = "canonical-test-run") -> str:
    with sqlite3.connect(db_path) as conn:
        now = "2026-03-10T00:00:00+00:00"
        conn.execute(
            """
            INSERT OR REPLACE INTO canonical_refresh_runs(
                run_id, scope_year, managed_project_keys_json, started_at_utc, ended_at_utc,
                status, trigger_source, error_message, stats_json,
                progress_step, progress_pct, cancel_requested, updated_at_utc
            ) VALUES (?, 2026, '["O2","RLT"]', ?, ?, 'success', 'test', '', '{}', 'done', 100, 0, ?)
            """,
            (run_id, now, now, now),
        )
        conn.execute(
            "UPDATE canonical_refresh_state SET active_run_id=?, last_success_run_id=?, updated_at_utc=? WHERE id=1",
            (run_id, run_id, now),
        )
        issues = [
            {
                "issue_id": "1",
                "issue_key": "O2-100",
                "project_key": "O2",
                "issue_type": "Story",
                "fix_type": "",
                "summary": "Story",
                "status": "Open",
                "assignee": "Alice",
                "start_date": "2026-02-01",
                "due_date": "2026-02-12",
                "resolved_stable_since_date": "",
                "original_estimate_hours": 8.0,
                "total_hours_logged": 0.0,
                "parent_issue_key": "",
                "story_key": "O2-100",
                "epic_key": "O2-100",
                "raw_payload_json": "{}",
            },
            {
                "issue_id": "2",
                "issue_key": "O2-101",
                "project_key": "O2",
                "issue_type": "Sub-task",
                "fix_type": "",
                "summary": "Subtask",
                "status": "Open",
                "assignee": "Alice",
                "start_date": "2026-02-02",
                "due_date": "2026-02-10",
                "resolved_stable_since_date": "",
                "original_estimate_hours": 8.0,
                "total_hours_logged": 2.0,
                "parent_issue_key": "O2-100",
                "story_key": "O2-100",
                "epic_key": "O2-100",
                "raw_payload_json": "{}",
            },
            {
                "issue_id": "3",
                "issue_key": "RLT-1",
                "project_key": "RLT",
                "issue_type": "Task",
                "fix_type": "",
                "summary": "Leave parent",
                "status": "Open",
                "assignee": "Alice",
                "start_date": "2026-02-11",
                "due_date": "2026-02-11",
                "resolved_stable_since_date": "",
                "original_estimate_hours": 8.0,
                "total_hours_logged": 0.0,
                "parent_issue_key": "",
                "story_key": "",
                "epic_key": "",
                "raw_payload_json": "{}",
            },
            {
                "issue_id": "4",
                "issue_key": "RLT-2",
                "project_key": "RLT",
                "issue_type": "Sub-task",
                "fix_type": "",
                "summary": "Planned Leave",
                "status": "Planned Leave",
                "assignee": "Alice",
                "start_date": "2026-02-11",
                "due_date": "2026-02-11",
                "resolved_stable_since_date": "",
                "original_estimate_hours": 8.0,
                "total_hours_logged": 0.0,
                "parent_issue_key": "RLT-1",
                "story_key": "",
                "epic_key": "",
                "raw_payload_json": '{"fields":{"customfield_10584":"Planned Leave"}}',
            },
        ]
        for row in issues:
            conn.execute(
                """
                INSERT OR REPLACE INTO canonical_issues(
                    run_id, issue_id, issue_key, project_key, issue_type, fix_type, summary, status, assignee,
                    start_date, due_date, resolved_stable_since_date, original_estimate_hours, total_hours_logged,
                    created_utc, updated_utc, parent_issue_key, story_key, epic_key, raw_payload_json
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    run_id,
                    row["issue_id"],
                    row["issue_key"],
                    row["project_key"],
                    row["issue_type"],
                    row["fix_type"],
                    row["summary"],
                    row["status"],
                    row["assignee"],
                    row["start_date"],
                    row["due_date"],
                    row["resolved_stable_since_date"],
                    row["original_estimate_hours"],
                    row["total_hours_logged"],
                    now,
                    now,
                    row["parent_issue_key"],
                    row["story_key"],
                    row["epic_key"],
                    row["raw_payload_json"],
                ),
            )
        conn.execute(
            """
            INSERT OR REPLACE INTO canonical_worklogs(
                run_id, worklog_id, issue_key, project_key, worklog_author, issue_assignee,
                started_utc, started_date, updated_utc, hours_logged
            ) VALUES (?, 'wl-1', 'O2-101', 'O2', 'Alice', 'Alice', '2026-02-11T10:00:00+00:00', '2026-02-11', '2026-02-11T10:00:00+00:00', 2.0)
            """,
            (run_id,),
        )
        conn.execute(
            """
            INSERT OR REPLACE INTO canonical_worklogs(
                run_id, worklog_id, issue_key, project_key, worklog_author, issue_assignee,
                started_utc, started_date, updated_utc, hours_logged
            ) VALUES (?, 'wl-2', 'RLT-2', 'RLT', 'Alice', 'Alice', '2026-02-11T10:00:00+00:00', '2026-02-11', '2026-02-11T10:00:00+00:00', 8.0)
            """,
            (run_id,),
        )
        conn.commit()
    return run_id


class EmployeePerformanceReportTests(unittest.TestCase):
    def test_settings_validation(self):
        valid = _normalize_performance_settings(
            {
                "base_score": 100,
                "min_score": 0,
                "max_score": 100,
                "points_per_bug_hour": 1,
                "points_per_bug_late_hour": 2,
                "points_per_unplanned_leave_hour": 3,
                "points_per_subtask_late_hour": 4,
                "points_per_estimate_overrun_hour": 5,
                "points_per_missed_due_date": 2,
                "overloaded_penalty_enabled": 1,
                "planning_realism_enabled": 0,
                "overloaded_penalty_threshold_pct": 10,
            }
        )
        self.assertEqual(valid["base_score"], 100)

        with self.assertRaises(ValueError):
            _normalize_performance_settings(
                {
                    "base_score": 100,
                    "min_score": 0,
                    "max_score": 100,
                    "points_per_bug_hour": -1,
                    "points_per_bug_late_hour": 2,
                    "points_per_unplanned_leave_hour": 3,
                    "points_per_subtask_late_hour": 4,
                    "points_per_estimate_overrun_hour": 5,
                    "points_per_missed_due_date": 2,
                    "overloaded_penalty_enabled": 1,
                    "planning_realism_enabled": 0,
                    "overloaded_penalty_threshold_pct": 10,
                }
            )

    def test_settings_db_roundtrip(self):
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            db = Path(td) / "assignee_hours_capacity.db"
            _init_performance_settings_db(db)
            initial = _load_performance_settings(db)
            self.assertIn("base_score", initial)
            self.assertEqual(initial["overloaded_penalty_enabled"], 0)
            self.assertEqual(initial["planning_realism_enabled"], 0)
            saved = _save_performance_settings(
                db,
                {
                    "base_score": 90,
                    "min_score": 0,
                    "max_score": 100,
                    "points_per_bug_hour": 0.8,
                    "points_per_bug_late_hour": 1.9,
                    "points_per_unplanned_leave_hour": 0.6,
                    "points_per_subtask_late_hour": 1.2,
                    "points_per_estimate_overrun_hour": 1.4,
                    "points_per_missed_due_date": 2.0,
                    "overloaded_penalty_enabled": 1,
                    "planning_realism_enabled": 1,
                    "overloaded_penalty_threshold_pct": 12.5,
                },
            )
            loaded = _load_performance_settings(db)
            self.assertEqual(saved["base_score"], loaded["base_score"])
            self.assertEqual(saved["points_per_bug_late_hour"], loaded["points_per_bug_late_hour"])
            self.assertEqual(saved["overloaded_penalty_enabled"], loaded["overloaded_penalty_enabled"])
            self.assertEqual(saved["planning_realism_enabled"], loaded["planning_realism_enabled"])
            self.assertEqual(saved["overloaded_penalty_threshold_pct"], loaded["overloaded_penalty_threshold_pct"])

    def test_html_contains_core_controls(self):
        payload = _build_payload([], [], [], dict(DEFAULT_PERFORMANCE_SETTINGS), [], [], [], [])
        html = _build_html(payload)
        self.assertIn('id="from"', html)
        self.assertIn('id="to"', html)
        self.assertIn('id="projects"', html)
        self.assertIn('id="leaderboard"', html)
        self.assertIn('id="leader-sort"', html)
        self.assertIn('id="leader-sort-direction"', html)
        self.assertIn("Available for more work", html)
        self.assertIn('<option value="desc" selected>Descending</option>', html)
        self.assertIn('href="/settings/performance"', html)
        self.assertIn("Planned Leaves", html)
        self.assertIn("Unplanned Leaves", html)
        self.assertIn('id="shortcut-current-month"', html)
        self.assertIn('id="shortcut-previous-month"', html)
        self.assertIn('id="shortcut-last-30-days"', html)
        self.assertIn('id="shortcut-quarter-to-date"', html)
        self.assertIn('id="assignee-overloaded-penalty-toggle"', html)
        self.assertIn('id="assignee-planning-realism-toggle"', html)
        self.assertIn('id="header-performance-controls-trigger"', html)
        self.assertIn('id="header-performance-controls-popover"', html)
        self.assertIn('id="top-overburn-mode"', html)
        self.assertIn('id="top-efficiency-mode"', html)
        self.assertIn("Overburn Per Task", html)
        self.assertIn("Overburn Total", html)
        self.assertIn("Penalty Inclusive Efficiency", html)
        self.assertIn("Simple Efficiency", html)
        self.assertIn('id="header-average-performance-value"', html)
        self.assertIn('id="header-average-performance-mode"', html)
        self.assertIn('id="header-average-performance-meta"', html)
        self.assertIn('id="header-efficiency-value"', html)
        self.assertIn('id="header-efficiency-mode"', html)
        self.assertIn('id="header-efficiency-meta"', html)
        self.assertIn('id="header-total-planned-hours-value"', html)
        self.assertIn('id="header-total-actual-hours-value"', html)
        self.assertIn('fetch("/api/performance/settings")', html)
        self.assertIn("function applyPerformanceSettings(nextSettings)", html)
        self.assertIn("function syncEfficiencyScorecardMode(nextMode)", html)
        self.assertIn("function syncHeaderPerformanceControls()", html)
        self.assertIn("hydratePerformanceSettings().finally(() => {", html)
        self.assertIn("let performanceSettingsReady = false;", html)
        self.assertIn("Loading performance settings before calculating scores.", html)
        self.assertIn("if (!performanceSettingsReady) {", html)
        self.assertIn("performanceSettingsReady = true;", html)
        self.assertIn('document.getElementById("leader-sort-direction").addEventListener("change", renderAll);', html)
        self.assertIn('document.getElementById("leader-sort-direction").value="desc";', html)

    def test_html_applies_subtask_only_scope_for_performance_kpis(self):
        payload = _build_payload([], [], [], dict(DEFAULT_PERFORMANCE_SETTINGS), [], [], [], [])
        html = _build_html(payload)
        self.assertIn("function issueTypeLabel(t)", html)
        self.assertIn("function isSubtaskPerformanceType(t)", html)
        self.assertIn('const issueType = String(r.item_issue_type || r.issue_type || "");', html)
        self.assertIn("return isSubtaskPerformanceType(issueType);", html)
        self.assertIn(
            'const issueType = String(r.jira_issue_type || r.issue_type || r.work_item_type || "");',
            html,
        )
        self.assertIn("if (!isSubtaskPerformanceType(issueType)) return false;", html)
        self.assertIn("const leaveIssueKeySet = new Set(", html)
        self.assertIn("function isLeaveIssueKey(issueKey)", html)
        self.assertNotIn("if (!matchesPlannedRange(r, from, to)) return false;", html)
        self.assertIn("const assignedItemsWork = assignedItems.filter((r) => !isLeaveIssueKey(String(r.issue_key || \"\")));", html)
        self.assertIn("value: n(item.planned_hours_assigned),", html)
        self.assertIn("value: n(item.actual_hours_stats_total),", html)
        self.assertIn("toggle-actual-hours-breakdown", html)
        self.assertIn("Object.entries(item.issue_logged_hours_stats_by_issue || {})", html)
        self.assertIn('id="assignee-extended-actuals-toggle"', html)

    def test_html_subtask_type_helper_includes_subtask_and_bug_subtask_patterns(self):
        payload = _build_payload([], [], [], dict(DEFAULT_PERFORMANCE_SETTINGS), [], [], [], [])
        html = _build_html(payload)
        self.assertIn('if (label.includes("sub-task") || label.includes("subtask")) return true;', html)
        self.assertIn('return label.includes("bug") && (label.includes("sub-task") || label.includes("subtask"));', html)

    def test_html_leaderboard_search_prefers_name_token_matches(self):
        payload = _build_payload([], [], [], dict(DEFAULT_PERFORMANCE_SETTINGS), [], [], [], [])
        html = _build_html(payload)
        self.assertIn("function assigneeSearchRank(name, query)", html)
        self.assertIn('else if (parts.some((part) => part.startsWith(term))) score += 30;', html)
        self.assertIn('const strongSearchMatchExists = viewItems.some((it) => assigneeSearchRank(String(it.assignee || ""), leaderSearchText) >= 20);', html)
        self.assertIn('const searchRank = assigneeSearchRank(String(it.assignee || ""), leaderSearchText);', html)
        self.assertIn('return !strongSearchMatchExists || searchRank >= 20;', html)
        self.assertIn('const searchDiff = assigneeSearchRank(String(b.assignee || ""), leaderSearchText) - assigneeSearchRank(String(a.assignee || ""), leaderSearchText);', html)

    def test_html_marks_zero_planned_hours_as_score_na(self):
        payload = _build_payload([], [], [], dict(DEFAULT_PERFORMANCE_SETTINGS), [], [], [], [])
        html = _build_html(payload)
        self.assertIn("function isScoreEligible(item)", html)
        self.assertIn('return Number.isFinite(value) ? value.toFixed(1) : "N/A";', html)
        self.assertIn("Scoring N/A", html)
        self.assertIn("not eligible for scoring", html)
        self.assertIn("Planned Hours Assigned is", html)

    def test_html_planned_hours_breakdown_excludes_bug_subtasks(self):
        payload = _build_payload([], [], [], dict(DEFAULT_PERFORMANCE_SETTINGS), [], [], [], [])
        html = _build_html(payload)
        self.assertIn('return t === "subtask";', html)
        self.assertIn("Assigned non-bug subtasks in current filters", html)
        self.assertNotIn("Assigned subtasks in current filters (including bug subtasks)", html)

    def test_html_contains_simple_score_drawer_controls(self):
        payload = _build_payload([], [], [], dict(DEFAULT_PERFORMANCE_SETTINGS), [], [], [], [])
        html = _build_html(payload)
        self.assertIn('id="score-detail-drawer"', html)
        self.assertIn('id="score-detail-drawer-overlay"', html)
        self.assertIn('id="summary-simple-score-trigger"', html)
        self.assertIn("openScoreDrawerForAssignee(item)", html)
        self.assertIn("Simple Score Details", html)
        self.assertIn("Actual Complete Date", html)
        self.assertIn("Last Logged Date", html)
        self.assertIn("Planned Due Date", html)

    def test_html_overloaded_penalty_uses_capacity_threshold_formula(self):
        payload = _build_payload([], [], [], dict(DEFAULT_PERFORMANCE_SETTINGS), [], [], [], [])
        html = _build_html(payload)
        self.assertIn("const effectiveCapacity = Math.max(0, n(it.employee_capacity_hours));", html)
        self.assertIn("plannedHrs > effectiveCapacity", html)
        self.assertIn("safeThreshold", html)
        self.assertIn("actualHrs < safeThreshold", html)
        self.assertIn("shortfall", html)
        self.assertIn("planningRealismEnabled", html)
        self.assertIn("simple_score_overloaded_penalty_pct", html)
        self.assertIn("simple_score_overrun_active", html)
        self.assertIn(
            'it.employee_capacity_hours = Math.max(0, n(it.base_capacity_hours) - n(it.planned_leave_hours));',
            html,
        )
        self.assertNotIn(
            'it.employee_capacity_hours = Math.max(0, n(it.base_capacity_hours) - n(it.planned_leave_hours) - n(it.unplanned_leave_hours));',
            html,
        )

    def test_html_availability_breakdown_shows_unplanned_leave_as_not_deducted(self):
        payload = _build_payload([], [], [], dict(DEFAULT_PERFORMANCE_SETTINGS), [], [], [], [])
        html = _build_html(payload)
        self.assertIn("Unplanned Leaves (Not Deducted)", html)
        self.assertIn("const availabilityDisplayIngredients = [...availabilityIngredients];", html)

    def test_html_recomputes_simple_score_when_extended_actuals_toggle_changes_actuals(self):
        payload = _build_payload([], [], [], dict(DEFAULT_PERFORMANCE_SETTINGS), [], [], [], [])
        html = _build_html(payload)
        self.assertIn("const actual = n(statsByIssue[issueKey]);", html)
        self.assertIn("const overrun = estimate > 0 ? Math.max(0, actual - estimate) : 0;", html)
        self.assertIn('const isCommitment = estimateStatus === "over_estimate" && dueStatus === "on_time" ? 1 : 0;', html)
        self.assertIn("it.ss_total_actual = liveTotalActual;", html)
        self.assertIn("it.ss_total_overrun = liveTotalOverrun;", html)
        self.assertIn('simpleOverrunMode === "total" ? totalOverTotal : totalOverSubtasks', html)
        self.assertIn('syncSimpleOverrunMode("subtasks")', html)
        self.assertIn('syncEfficiencyScorecardMode("penalty_inclusive")', html)

    def test_html_uses_shared_scoped_subtasks_endpoint_for_filter_refresh(self):
        payload = _build_payload([], [], [], dict(DEFAULT_PERFORMANCE_SETTINGS), [], [], [], [])
        html = _build_html(payload)
        self.assertIn('const SCOPED_SUBTASKS_ENDPOINT = "/api/scoped-subtasks";', html)
        self.assertIn("async function loadScopedSubtasksForCurrentFilters()", html)
        self.assertIn('params.set("mode", extendedActualsEnabled ? "extended" : "log_date");', html)
        self.assertIn('async function renderAll()', html)
        self.assertIn("await loadScopedSubtasksForCurrentFilters();", html)
        self.assertIn("const scopedIssueSet = scopedSubtasksIssueKeySet instanceof Set ? scopedSubtasksIssueKeySet : null;", html)

    def test_html_due_completion_mode_mentions_late_completion_penalties(self):
        payload = _build_payload([], [], [], dict(DEFAULT_PERFORMANCE_SETTINGS), [], [], [], [])
        html = _build_html(payload)
        self.assertIn("Late Completion Rule", html)
        self.assertIn("Late-completed subtasks add their original estimate as penalty input", html)
        self.assertIn("Late Completion Estimate Penalty", html)
        self.assertIn("Late estimate penalty:", html)

    def test_html_due_compliance_table_uses_updated_completion_headers(self):
        payload = _build_payload([], [], [], dict(DEFAULT_PERFORMANCE_SETTINGS), [], [], [], [])
        html = _build_html(payload)
        self.assertIn("<th>Due</th><th>Last Logged Date</th><th>Actual Completed Date</th><th>Bucket</th>", html)
        self.assertIn("Actual complete date came from last logged date.", html)
        self.assertNotIn("Stable Resolved", html)
        self.assertNotIn("Status Resolved Date", html)
        self.assertNotIn("Actual complete date came from resolved date.", html)

    def test_html_simple_scoring_table_declares_late_completed_status(self):
        payload = _build_payload([], [], [], dict(DEFAULT_PERFORMANCE_SETTINGS), [], [], [], [])
        html = _build_html(payload)
        self.assertIn('if (est === "within_estimate" && due === "late") return `<span class="ss-status-pill ss-pill-late">Late Completed</span>`;', html)
        self.assertIn('if (est === "over_estimate" && due === "late") return `<span class="ss-status-pill ss-pill-late">Over + Late Completed</span>`;', html)

    def test_runtime_paths_default_employee_performance_source_is_auto(self):
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            tdp = Path(td)
            with patch.dict("os.environ", {}, clear=True):
                paths = _resolve_runtime_paths(tdp)
        self.assertEqual(paths["source_mode"], "auto")

    def test_source_mode_auto_prefers_latest_canonical_snapshot(self):
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            db_path = Path(td) / "assignee_hours_capacity.db"
            _init_performance_settings_db(db_path)
            report_server._init_epf_refresh_db(db_path)
            report_server._init_canonical_refresh_db(db_path)
            canonical_run_id = _seed_canonical_run(db_path)
            with sqlite3.connect(db_path) as conn:
                conn.execute(
                    "UPDATE epf_refresh_state SET active_run_id = ?, last_success_run_id = ?, updated_at_utc = ? WHERE id = 1",
                    ("epf-active-run", "epf-active-run", "2026-03-11T00:00:00+00:00"),
                )
                conn.commit()
            mode, run_id = _resolve_employee_performance_source_mode(
                {"source_mode": "auto", "db_path": db_path, "run_id": "", "canonical_run_id": ""}
            )
        self.assertEqual(mode, "canonical_db")
        self.assertEqual(run_id, canonical_run_id)

    def test_source_mode_auto_falls_back_to_active_epf_snapshot_when_canonical_missing(self):
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            db_path = Path(td) / "assignee_hours_capacity.db"
            _init_performance_settings_db(db_path)
            report_server._init_epf_refresh_db(db_path)
            with sqlite3.connect(db_path) as conn:
                conn.execute(
                    "UPDATE epf_refresh_state SET active_run_id = ?, last_success_run_id = ?, updated_at_utc = ? WHERE id = 1",
                    ("epf-active-run", "epf-active-run", "2026-03-11T00:00:00+00:00"),
                )
                conn.commit()
            mode, run_id = _resolve_employee_performance_source_mode(
                {"source_mode": "auto", "db_path": db_path, "run_id": "", "canonical_run_id": ""}
            )
        self.assertEqual(mode, "db")
        self.assertEqual(run_id, "epf-active-run")

    def test_source_mode_preserves_explicit_db_requirement(self):
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            db_path = Path(td) / "assignee_hours_capacity.db"
            _init_performance_settings_db(db_path)
            report_server._init_epf_refresh_db(db_path)
            report_server._init_canonical_refresh_db(db_path)
            _seed_canonical_run(db_path)
            with sqlite3.connect(db_path) as conn:
                conn.execute(
                    "UPDATE epf_refresh_state SET active_run_id = ?, last_success_run_id = ?, updated_at_utc = ? WHERE id = 1",
                    ("epf-active-run", "epf-active-run", "2026-03-11T00:00:00+00:00"),
                )
                conn.commit()
            mode, run_id = _resolve_employee_performance_source_mode(
                {"source_mode": "db", "db_path": db_path, "run_id": "", "canonical_run_id": ""}
            )
        self.assertEqual(mode, "db")
        self.assertEqual(run_id, "epf-active-run")

    def test_source_mode_falls_back_to_xlsx_when_no_epf_snapshot_exists(self):
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            db_path = Path(td) / "assignee_hours_capacity.db"
            _init_performance_settings_db(db_path)
            mode, run_id = _resolve_employee_performance_source_mode(
                {"source_mode": "auto", "db_path": db_path, "run_id": "", "canonical_run_id": ""}
            )
        self.assertEqual(mode, "xlsx")
        self.assertEqual(run_id, "")

    def test_source_mode_preserves_explicit_canonical_db_requirement(self):
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            db_path = Path(td) / "assignee_hours_capacity.db"
            report_server._init_canonical_refresh_db(db_path)
            canonical_run_id = _seed_canonical_run(db_path)
            mode, run_id = _resolve_employee_performance_source_mode(
                {"source_mode": "canonical_db", "db_path": db_path, "run_id": "", "canonical_run_id": canonical_run_id}
            )
        self.assertEqual(mode, "canonical_db")
        self.assertEqual(run_id, canonical_run_id)

    def test_derive_actual_completion_uses_last_log(self):
        meta = _derive_actual_completion("2026-03-05", "2026-03-07")
        self.assertEqual(meta["actual_complete_date"], "2026-03-07")
        self.assertEqual(meta["actual_complete_source"], "last_logged_date")
        self.assertEqual(meta["completion_bucket"], "after_due")

    def test_derive_actual_completion_handles_missing_dates(self):
        only_last = _derive_actual_completion("2026-03-05", "2026-03-04")
        none = _derive_actual_completion("2026-03-05", "")
        self.assertEqual(only_last["actual_complete_date"], "2026-03-04")
        self.assertEqual(none["actual_complete_date"], "")
        self.assertEqual(none["completion_bucket"], "not_completed")

    def test_precompute_simple_scoring_persists_actual_completion_fields(self):
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            db = Path(td) / "assignee_hours_capacity.db"
            work_items = {
                "O2-101": {
                    "issue_type": "Sub-task",
                    "assignee": "Alice",
                    "original_estimate_hours": 8,
                    "due_date": "2026-03-05",
                    "status": "Done",
                },
                "O2-102": {
                    "issue_type": "Sub-task",
                    "assignee": "Alice",
                    "original_estimate_hours": 8,
                    "due_date": "2026-03-08",
                    "status": "Done",
                },
            }
            worklogs = [
                {"issue_id": "O2-101", "worklog_date": "2026-03-07", "hours_logged": 2},
                {"issue_id": "O2-102", "worklog_date": "2026-03-06", "hours_logged": 2},
            ]

            _precompute_simple_scoring(db, work_items, worklogs)
            rows = {row["issue_key"]: row for row in _load_simple_scoring(db)}

            self.assertEqual(rows["O2-101"]["planned_due_date"], "2026-03-05")
            self.assertEqual(rows["O2-101"]["last_logged_date"], "2026-03-07")
            self.assertEqual(rows["O2-101"]["actual_complete_date"], "2026-03-07")
            self.assertEqual(rows["O2-101"]["actual_complete_source"], "last_logged_date")
            self.assertEqual(rows["O2-101"]["due_completion_status"], "late")

            self.assertEqual(rows["O2-102"]["actual_complete_date"], "2026-03-06")
            self.assertEqual(rows["O2-102"]["actual_complete_source"], "last_logged_date")
            self.assertEqual(rows["O2-102"]["due_completion_status"], "on_time")

    def test_load_leave_issue_keys_prefers_raw_subtasks_and_normalizes(self):
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            path = Path(td) / "rlt_leave_report.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "Raw_Subtasks"
            ws.append(["issue_key", "assignee"])
            ws.append(["rlt-172", "Maria"])
            ws.append(["RLT-172", "Maria"])
            ws.append([" RLT-173 ", "Maria"])
            ws.append(["", "Alice"])
            wb.save(path)

            keys = _load_leave_issue_keys(path)
            self.assertEqual(keys, ["RLT-172", "RLT-173"])

    def test_build_payload_includes_leave_issue_keys(self):
        payload = _build_payload(
            [],
            [],
            [],
            dict(DEFAULT_PERFORMANCE_SETTINGS),
            [],
            [],
            [],
            [],
            leave_issue_keys=["RLT-172"],
        )
        self.assertEqual(payload["leave_issue_keys"], ["RLT-172"])

    def test_build_payload_normalizes_leave_issue_keys_from_set(self):
        payload = _build_payload(
            [],
            [],
            [],
            dict(DEFAULT_PERFORMANCE_SETTINGS),
            [],
            [],
            [],
            [],
            leave_issue_keys={"rlt-173", "RLT-172", "rlt-172"},
        )
        self.assertEqual(payload["leave_issue_keys"], ["RLT-172", "RLT-173"])

    def test_apply_managed_project_display_names_updates_simple_score_project_labels(self):
        work_items = {
            "ABC-1": {"issue_key": "ABC-1", "project_key": "ABC", "project_name": "ABC"},
            "XYZ-2": {"issue_key": "XYZ-2", "project_key": "XYZ", "project_name": "XYZ"},
        }
        simple_scoring_rows = [
            {"issue_key": "ABC-11", "project_key": "ABC", "project_name": "ABC"},
            {"issue_key": "XYZ-22", "project_key": "XYZ", "project_name": "XYZ"},
        ]

        _apply_managed_project_display_names(
            work_items,
            simple_scoring_rows,
            {"ABC": "Alpha Display"},
        )

        self.assertEqual(work_items["ABC-1"]["project_name"], "Alpha Display")
        self.assertEqual(simple_scoring_rows[0]["project_name"], "Alpha Display")
        self.assertEqual(work_items["XYZ-2"]["project_name"], "XYZ")
        self.assertEqual(simple_scoring_rows[1]["project_name"], "XYZ")

    def test_load_managed_project_display_names_reads_active_display_names(self):
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            db = Path(td) / "assignee_hours_capacity.db"
            with sqlite3.connect(db) as conn:
                conn.execute(
                    """
                    CREATE TABLE managed_projects (
                        id INTEGER PRIMARY KEY,
                        project_key TEXT NOT NULL UNIQUE,
                        project_name TEXT NOT NULL,
                        display_name TEXT NOT NULL,
                        color_hex TEXT NOT NULL,
                        is_active INTEGER NOT NULL,
                        created_at_utc TEXT NOT NULL,
                        updated_at_utc TEXT NOT NULL
                    )
                    """
                )
                conn.execute(
                    """
                    INSERT INTO managed_projects (
                        project_key, project_name, display_name, color_hex, is_active, created_at_utc, updated_at_utc
                    ) VALUES (?, ?, ?, ?, ?, ?, ?)
                    """,
                    ("ABC", "Alpha Raw", "Alpha Display", "#123456", 1, "2026-03-10T00:00:00Z", "2026-03-10T00:00:00Z"),
                )
                conn.execute(
                    """
                    INSERT INTO managed_projects (
                        project_key, project_name, display_name, color_hex, is_active, created_at_utc, updated_at_utc
                    ) VALUES (?, ?, ?, ?, ?, ?, ?)
                    """,
                    ("XYZ", "Xray Raw", "Xray Display", "#654321", 0, "2026-03-10T00:00:00Z", "2026-03-10T00:00:00Z"),
                )
                conn.commit()

            self.assertEqual(_load_managed_project_display_names(db), {"ABC": "Alpha Display"})

    def test_load_unplanned_leave_rows_from_daily_assignee(self):
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            path = Path(td) / "rlt_leave_report.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "Daily_Assignee"
            ws.append(["assignee", "period_day", "planned_taken_hours", "unplanned_taken_hours"])
            ws.append(["Alice", "2026-02-01", 8, 0])
            ws.append(["Bob", "2026-02-01", 0, 4])
            wb.save(path)

            rows = _load_unplanned_leave_rows(path)
            self.assertEqual(len(rows), 2)
            self.assertEqual(rows[0]["assignee"], "Alice")
            self.assertEqual(rows[0]["planned_taken_hours"], 8)
            self.assertEqual(rows[1]["unplanned_taken_hours"], 4)

    def test_load_unplanned_leave_rows_from_dedicated_leave_sheet(self):
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            path = Path(td) / "employee_performance_leaves.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "Leaves"
            ws.append(["Employee", "Date", "Leave Type", "Leave Hours"])
            ws.append(["Alice", "2026-02-03", "Planned", 8])
            ws.append(["Alice", "2026-02-04", "Unplanned", 4])
            wb.save(path)

            rows = _load_unplanned_leave_rows(path)
            self.assertEqual(len(rows), 2)
            self.assertEqual(rows[0]["planned_taken_hours"], 8)
            self.assertEqual(rows[0]["unplanned_taken_hours"], 0)
            self.assertEqual(rows[1]["planned_taken_hours"], 0)
            self.assertEqual(rows[1]["unplanned_taken_hours"], 4)

    def test_performance_settings_api(self):
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            tdp = Path(td)
            (tdp / "report_html").mkdir(parents=True, exist_ok=True)
            (tdp / "report_html" / "dashboard.html").write_text("<html><body>ok</body></html>", encoding="utf-8")
            wb = Workbook()
            ws = wb.active
            ws.append(["project_key", "worklog_date", "period_day", "period_week", "period_month", "issue_assignee", "hours_logged"])
            ws.append(["O2", "2026-02-01", "2026-02-01", "2026-W05", "2026-02", "Alice", 2.0])
            wb.save(tdp / "assignee_hours_report.xlsx")
            app = create_report_server_app(base_dir=tdp, folder_raw="report_html")
            client = app.test_client()

            get_resp = client.get("/api/performance/settings")
            self.assertEqual(get_resp.status_code, 200)
            body = get_resp.get_json()
            self.assertIn("settings", body)

            post_resp = client.post(
                "/api/performance/settings",
                json={
                    "base_score": 88,
                    "min_score": 0,
                    "max_score": 100,
                    "points_per_bug_hour": 1.1,
                    "points_per_bug_late_hour": 2.1,
                    "points_per_unplanned_leave_hour": 0.9,
                    "points_per_subtask_late_hour": 1.2,
                    "points_per_estimate_overrun_hour": 1.3,
                    "points_per_missed_due_date": 2.0,
                    "overloaded_penalty_enabled": 1,
                    "planning_realism_enabled": 0,
                    "overloaded_penalty_threshold_pct": 10.0,
                },
            )
            self.assertEqual(post_resp.status_code, 200)
            saved = post_resp.get_json()
            self.assertEqual(saved["settings"]["base_score"], 88)

    def test_performance_settings_page_uses_planning_realism_labels(self):
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            tdp = Path(td)
            (tdp / "report_html").mkdir(parents=True, exist_ok=True)
            (tdp / "report_html" / "dashboard.html").write_text("<html><body>ok</body></html>", encoding="utf-8")
            wb = Workbook()
            ws = wb.active
            ws.append(["project_key", "worklog_date", "period_day", "period_week", "period_month", "issue_assignee", "hours_logged"])
            ws.append(["O2", "2026-02-01", "2026-02-01", "2026-W05", "2026-02", "Alice", 2.0])
            wb.save(tdp / "assignee_hours_report.xlsx")
            app = create_report_server_app(base_dir=tdp, folder_raw="report_html")
            client = app.test_client()

            resp = client.get("/settings/performance")
            self.assertEqual(resp.status_code, 200)
            html = resp.get_data(as_text=True)
            self.assertIn("Overloaded Penalty", html)
            self.assertIn("Overload Capping/ Planning Realism", html)
            self.assertIn("Overloaded Threshold (%)", html)
            self.assertIn("shortfall penalty is deducted", html)
            self.assertIn("Overload Capping/ Planning Realism = ", html)
            self.assertIn("Default is OFF", html)

    def test_performance_team_management_api(self):
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            tdp = Path(td)
            (tdp / "report_html").mkdir(parents=True, exist_ok=True)
            (tdp / "report_html" / "dashboard.html").write_text("<html><body>ok</body></html>", encoding="utf-8")

            wb = Workbook()
            ws = wb.active
            ws.append(["project_key", "worklog_date", "period_day", "period_week", "period_month", "issue_assignee", "hours_logged"])
            ws.append(["O2", "2026-02-01", "2026-02-01", "2026-W05", "2026-02", "Alice", 2.0])
            ws.append(["O2", "2026-02-02", "2026-02-02", "2026-W05", "2026-02", "Bob", 3.0])
            wb.save(tdp / "assignee_hours_report.xlsx")

            app = create_report_server_app(base_dir=tdp, folder_raw="report_html")
            client = app.test_client()

            assignees_resp = client.get("/api/performance/assignees")
            self.assertEqual(assignees_resp.status_code, 200)
            assignees_json = assignees_resp.get_json()
            self.assertIn("Alice", assignees_json["assignees"])

            create_resp = client.post(
                "/api/performance/teams",
                json={"team_name": "Core Team", "team_leader": "Alice", "assignees": ["Alice", "Bob"]},
            )
            self.assertEqual(create_resp.status_code, 200)
            created = create_resp.get_json()
            self.assertEqual(created["team"]["team_name"], "Core Team")
            self.assertEqual(created["team"]["team_leader"], "Alice")

            update_resp = client.put(
                "/api/performance/teams/Core%20Team",
                json={"team_name": "Core Platform", "team_leader": "Bob", "assignees": ["Bob"]},
            )
            self.assertEqual(update_resp.status_code, 200)
            updated = update_resp.get_json()
            self.assertEqual(updated["team"]["team_name"], "Core Platform")
            self.assertEqual(updated["team"]["team_leader"], "Bob")
            self.assertEqual(updated["team"]["assignees"], ["Bob"])

            list_resp = client.get("/api/performance/teams")
            self.assertEqual(list_resp.status_code, 200)
            teams = list_resp.get_json()["teams"]
            self.assertFalse(any(t["team_name"] == "Core Team" for t in teams))
            self.assertTrue(any(t["team_name"] == "Core Platform" for t in teams))

            del_resp = client.delete("/api/performance/teams/Core%20Platform")
            self.assertEqual(del_resp.status_code, 200)
            self.assertTrue(del_resp.get_json()["deleted"])

    def test_employee_refresh_cancel_marks_running_run(self):
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            tdp = Path(td)
            (tdp / "report_html").mkdir(parents=True, exist_ok=True)
            (tdp / "report_html" / "dashboard.html").write_text("<html><body>ok</body></html>", encoding="utf-8")
            wb = Workbook()
            ws = wb.active
            ws.append(["project_key", "worklog_date", "period_day", "period_week", "period_month", "issue_assignee", "hours_logged"])
            ws.append(["O2", "2026-02-01", "2026-02-01", "2026-W05", "2026-02", "Alice", 2.0])
            wb.save(tdp / "assignee_hours_report.xlsx")
            app = create_report_server_app(base_dir=tdp, folder_raw="report_html")
            client = app.test_client()
            db_path = tdp / "assignee_hours_capacity.db"

            with sqlite3.connect(db_path) as conn:
                conn.execute(
                    """
                    INSERT INTO epf_refresh_runs(
                        run_id, started_at_utc, status, trigger_source, error_message, stats_json,
                        progress_step, progress_pct, cancel_requested, updated_at_utc
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        "epf-test-run-1",
                        "2026-03-05T00:00:00+00:00",
                        "running",
                        "api_refresh_async",
                        "",
                        "{}",
                        "fetching_worklogs",
                        20,
                        0,
                        "2026-03-05T00:00:00+00:00",
                    ),
                )
                conn.commit()

            cancel_resp = client.post("/api/employee-performance/cancel", json={"run_id": "epf-test-run-1"})
            self.assertEqual(cancel_resp.status_code, 200)
            body = cancel_resp.get_json() or {}
            self.assertTrue(body.get("ok"))
            self.assertIn(body.get("status"), ("cancel_requested", "canceled"))

            with sqlite3.connect(db_path) as conn:
                row = conn.execute("SELECT status FROM epf_refresh_runs WHERE run_id = ?", ("epf-test-run-1",)).fetchone()
            self.assertIsNotNone(row)
            self.assertIn(row[0], ("running", "canceled"))

    def test_employee_refresh_uses_canonical_source_scripts(self):
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            tdp = Path(td)
            (tdp / "report_html").mkdir(parents=True, exist_ok=True)
            (tdp / "report_html" / "dashboard.html").write_text("<html><body>ok</body></html>", encoding="utf-8")

            wb = Workbook()
            ws = wb.active
            ws.append(["project_key", "worklog_date", "period_day", "period_week", "period_month", "issue_assignee", "hours_logged"])
            ws.append(["O2", "2026-02-01", "2026-02-01", "2026-W05", "2026-02", "Alice", 2.0])
            wb.save(tdp / "assignee_hours_report.xlsx")

            app = create_report_server_app(base_dir=tdp, folder_raw="report_html")
            refresh_fn = app.view_functions["employee_performance_refresh"]
            refresh_globals = refresh_fn.__globals__
            db_path = tdp / "assignee_hours_capacity.db"
            canonical_run_id = _seed_canonical_run(db_path)

            run_envs: list[tuple[str, dict[str, str]]] = []

            def fake_run_script_interruptible(script_name, _cwd, env_overrides=None, extra_args=None, cancel_check=None):
                run_envs.append((script_name, dict(env_overrides or {})))
                if script_name == "generate_rlt_leave_report.py":
                    book = Workbook()
                    daily = book.active
                    daily.title = "Daily_Assignee"
                    daily.append(["assignee", "period_day", "unplanned_taken_hours", "planned_taken_hours"])
                    daily.append(["Alice", "2026-02-11", 0, 0])
                    raw_sub = book.create_sheet("Raw_Subtasks")
                    raw_sub.append(["issue_key"])
                    book.save(tdp / "rlt_leave_report.xlsx")
                    return 0, "ok", ""

                if script_name == "generate_employee_performance_report.py":
                    (tdp / "employee_performance_report.html").write_text("<html><body>ok</body></html>", encoding="utf-8")
                    return 0, "ok", ""

                return 0, "ok", ""

            with patch.dict(refresh_globals, {"_run_script_interruptible": fake_run_script_interruptible, "sync_report_html": lambda *_args, **_kwargs: None}):
                client = app.test_client()
                response = client.post("/api/employee-performance/refresh", json={"assignee": "Alice", "replace_running": False})
                self.assertEqual(response.status_code, 202)

                import time

                deadline = time.time() + 5
                while time.time() < deadline:
                    with sqlite3.connect(db_path) as conn:
                        row = conn.execute("SELECT status FROM epf_refresh_runs ORDER BY started_at_utc DESC LIMIT 1").fetchone()
                    if row and row[0] != "running":
                        break
                    time.sleep(0.05)

            env_by_script = {name: env for name, env in run_envs}
            self.assertEqual(env_by_script["generate_rlt_leave_report.py"].get("JIRA_LEAVE_REPORT_SOURCE"), "canonical_db")
            self.assertEqual(env_by_script["generate_rlt_leave_report.py"].get("JIRA_CANONICAL_RUN_ID"), canonical_run_id)
            self.assertEqual(env_by_script["generate_employee_performance_report.py"].get("JIRA_EMP_PERF_INPUT_SOURCE"), "canonical_db")
            self.assertEqual(env_by_script["generate_employee_performance_report.py"].get("JIRA_EMP_PERF_CANONICAL_RUN_ID"), canonical_run_id)
            self.assertNotIn("export_jira_subtask_worklogs.py", env_by_script)
            self.assertNotIn("export_jira_work_items.py", env_by_script)

            with sqlite3.connect(db_path) as conn:
                run_row = conn.execute(
                    "SELECT run_id, status FROM epf_refresh_runs ORDER BY started_at_utc DESC LIMIT 1"
                ).fetchone()
            self.assertIsNotNone(run_row)
            run_id = str(run_row[0] or "")
            self.assertTrue(run_id)

            with sqlite3.connect(db_path) as conn:
                stats_row = conn.execute(
                    "SELECT stats_json FROM epf_refresh_runs WHERE run_id = ?",
                    (run_id,),
                ).fetchone()
            self.assertIsNotNone(stats_row)
            self.assertIn("canonical_db", str(stats_row[0] or ""))

    def test_employee_full_refresh_via_generic_refresh_uses_canonical_scripts(self):
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            tdp = Path(td)
            (tdp / "report_html").mkdir(parents=True, exist_ok=True)
            (tdp / "report_html" / "dashboard.html").write_text("<html><body>ok</body></html>", encoding="utf-8")

            app = create_report_server_app(base_dir=tdp, folder_raw="report_html")
            client = app.test_client()
            db_path = tdp / "assignee_hours_capacity.db"
            canonical_run_id = _seed_canonical_run(db_path)

            original_run_script = report_server._run_script
            original_sync_report_html = report_server.sync_report_html
            calls: list[tuple[str, dict[str, str]]] = []

            def _fake_run_script(script_name, _cwd, extra_args=None, env_overrides=None):
                calls.append((script_name, dict(env_overrides or {})))
                env_overrides = env_overrides or {}
                if script_name == "generate_rlt_leave_report.py":
                    book = Workbook()
                    daily = book.active
                    daily.title = "Daily_Assignee"
                    daily.append(["assignee", "period_day", "unplanned_taken_hours", "planned_taken_hours"])
                    daily.append(["Alice", "2026-02-11", 0.0, 0.0])
                    raw_sub = book.create_sheet("Raw_Subtasks")
                    raw_sub.append(["issue_key"])
                    raw_sub.append(["O2-101"])
                    book.save(tdp / "rlt_leave_report.xlsx")
                    return 0, "ok", ""

                if script_name == "generate_employee_performance_report.py":
                    (tdp / "employee_performance_report.html").write_text("<html><body>ok</body></html>", encoding="utf-8")
                    return 0, "ok", ""

                return 0, "ok", ""

            report_server._run_script = _fake_run_script
            report_server.sync_report_html = lambda *_args, **_kwargs: None
            try:
                resp = client.post(
                    "/api/report/refresh",
                    json={"report": "employee_performance"},
                )
                self.assertEqual(resp.status_code, 200)
                body = resp.get_json() or {}
                self.assertTrue(body.get("ok"))
                self.assertEqual(body.get("canonical_run_id"), canonical_run_id)
                env_by_script = {name: env for name, env in calls}
                self.assertEqual(env_by_script["generate_rlt_leave_report.py"].get("JIRA_LEAVE_REPORT_SOURCE"), "canonical_db")
                self.assertEqual(env_by_script["generate_rlt_leave_report.py"].get("JIRA_CANONICAL_RUN_ID"), canonical_run_id)
                self.assertEqual(env_by_script["generate_employee_performance_report.py"].get("JIRA_EMP_PERF_INPUT_SOURCE"), "canonical_db")
                self.assertEqual(env_by_script["generate_employee_performance_report.py"].get("JIRA_EMP_PERF_CANONICAL_RUN_ID"), canonical_run_id)
            finally:
                report_server._run_script = original_run_script
                report_server.sync_report_html = original_sync_report_html

    def test_assignee_scoped_refresh_preserves_reassigned_subtasks_and_updates_exports(self):
        def make_issue(
            key: str,
            issue_type: str,
            assignee: str,
            *,
            parent: str = "",
            epic: str = "",
            summary: str = "",
            status: str = "In Progress",
            estimate_seconds: int = 28800,
        ) -> dict:
            fields = {
                "summary": summary or key,
                "status": {"name": status},
                "assignee": {"displayName": assignee},
                "issuetype": {"name": issue_type},
                "project": {"key": "O2"},
                "parent": {"key": parent} if parent else {},
                "timeoriginalestimate": estimate_seconds,
                "aggregatetimespent": estimate_seconds,
                "timespent": estimate_seconds,
                "created": "2026-03-01T00:00:00+00:00",
                "updated": "2026-03-02T00:00:00+00:00",
            }
            if epic:
                fields["customfield_10014"] = epic
            return {"id": key.replace("-", ""), "key": key, "fields": fields}

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            tdp = Path(td)
            (tdp / "report_html").mkdir(parents=True, exist_ok=True)
            (tdp / "report_html" / "dashboard.html").write_text("<html><body>ok</body></html>", encoding="utf-8")
            app = create_report_server_app(base_dir=tdp, folder_raw="report_html")
            client = app.test_client()
            db_path = tdp / "assignee_hours_capacity.db"
            previous_run_id = "canonical-prev-run"
            now = "2026-03-10T00:00:00+00:00"

            with sqlite3.connect(db_path) as conn:
                conn.execute(
                    """
                    INSERT OR REPLACE INTO canonical_refresh_runs(
                        run_id, scope_year, managed_project_keys_json, started_at_utc, ended_at_utc,
                        status, trigger_source, error_message, stats_json,
                        progress_step, progress_pct, cancel_requested, updated_at_utc
                    ) VALUES (?, 2026, '["O2"]', ?, ?, 'success', 'test', '', '{}', 'done', 100, 0, ?)
                    """,
                    (previous_run_id, now, now, now),
                )
                conn.execute(
                    "UPDATE canonical_refresh_state SET active_run_id=?, last_success_run_id=?, updated_at_utc=? WHERE id=1",
                    (previous_run_id, previous_run_id, now),
                )
                issue_rows = [
                    (
                        previous_run_id, "1", "O2-EP1", "O2", "Epic", "Epic", "Open", "Lead", "2026-03-01", "2026-03-31",
                        now, now, "", 0.0, 0.0, "", "", "", "O2-EP1", "{}",
                    ),
                    (
                        previous_run_id, "2", "O2-ST1", "O2", "Story", "Story", "Open", "Lead", "2026-03-01", "2026-03-20",
                        now, now, "", 16.0, 0.0, "", "O2-EP1", "O2-ST1", "O2-EP1", "{}",
                    ),
                    (
                        previous_run_id, "3", "O2-101", "O2", "Sub-task", "Old Alice Task", "Open", "Alice", "2026-03-02", "2026-03-10",
                        now, now, "", 8.0, 2.0, "", "O2-ST1", "O2-ST1", "O2-EP1", "{}",
                    ),
                ]
                conn.executemany(
                    """
                    INSERT INTO canonical_issues(
                        run_id, issue_id, issue_key, project_key, issue_type, summary, status, assignee,
                        start_date, due_date, created_utc, updated_utc, resolved_stable_since_date,
                        original_estimate_hours, total_hours_logged, fix_type, parent_issue_key,
                        story_key, epic_key, raw_payload_json
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    issue_rows,
                )
                conn.executemany(
                    """
                    INSERT INTO canonical_issue_links(
                        run_id, issue_key, parent_issue_key, story_key, epic_key, hierarchy_level
                    ) VALUES (?, ?, ?, ?, ?, ?)
                    """,
                    [
                        (previous_run_id, "O2-EP1", "", "", "O2-EP1", "epic"),
                        (previous_run_id, "O2-ST1", "O2-EP1", "O2-ST1", "O2-EP1", "story"),
                        (previous_run_id, "O2-101", "O2-ST1", "O2-ST1", "O2-EP1", "subtask"),
                    ],
                )
                conn.execute(
                    """
                    INSERT INTO canonical_worklogs(
                        run_id, worklog_id, issue_key, project_key, worklog_author, issue_assignee,
                        started_utc, started_date, updated_utc, hours_logged
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (previous_run_id, "wl-prev-1", "O2-101", "O2", "Alice", "Alice", "2026-03-03T00:00:00+00:00", "2026-03-03", now, 2.0),
                )
                conn.commit()

            current_alice_subtask = make_issue("O2-102", "Sub-task", "Alice", parent="O2-ST1", summary="New Alice Task")
            reassigned_subtask = make_issue("O2-101", "Sub-task", "Bob", parent="O2-ST1", summary="Moved To Bob")
            story_issue = make_issue("O2-ST1", "Story", "Lead", parent="O2-EP1", epic="O2-EP1", summary="Story")
            epic_issue = make_issue("O2-EP1", "Epic", "Lead", summary="Epic", estimate_seconds=0)
            issue_map = {
                "O2-101": reassigned_subtask,
                "O2-102": current_alice_subtask,
                "O2-ST1": story_issue,
                "O2-EP1": epic_issue,
            }

            def fake_fetch_by_keys(_session, issue_keys, _fields):
                return [issue_map[key] for key in issue_keys if key in issue_map]

            def fake_fetch_worklogs(_session, issue_key):
                key = str(issue_key or "").upper()
                if key == "O2-101":
                    return [{"id": "wl-101", "started": "2026-03-04T00:00:00+00:00", "updated": now, "author": {"displayName": "Bob"}}]
                if key == "O2-102":
                    return [{"id": "wl-102", "started": "2026-03-05T00:00:00+00:00", "updated": now, "author": {"displayName": "Alice"}}]
                return []

            def fake_run_script_interruptible(script_name, _cwd, env_overrides=None, extra_args=None, cancel_check=None):
                if script_name == "generate_employee_performance_report.py":
                    (tdp / "employee_performance_report.html").write_text("<html><body>ok</body></html>", encoding="utf-8")
                if script_name == "generate_rlt_leave_report.py":
                    book = Workbook()
                    daily = book.active
                    daily.title = "Daily_Assignee"
                    daily.append(["assignee", "period_day", "unplanned_taken_hours", "planned_taken_hours"])
                    book.save(tdp / "rlt_leave_report.xlsx")
                return 0, "ok", ""

            with (
                patch.dict("os.environ", {"JIRA_EXPORTS_DB_PATH": str(tdp / "jira_exports.db")}, clear=False),
                patch.object(report_server, "get_session", return_value=object()),
                patch.object(report_server, "resolve_jira_start_date_field_id", return_value="customfield_start"),
                patch.object(report_server, "resolve_jira_end_date_field_ids", return_value=["duedate"]),
                patch.object(report_server, "export_resolve_fix_type_field_id", return_value="customfield_fix_type"),
                patch.object(report_server, "_fetch_subtask_issues_for_assignee", return_value=[current_alice_subtask]),
                patch.object(report_server, "_fetch_jira_issues_by_keys", side_effect=fake_fetch_by_keys),
                patch.object(report_server, "export_fetch_worklogs_for_issue", side_effect=fake_fetch_worklogs),
                patch.object(report_server, "_run_script_interruptible", side_effect=fake_run_script_interruptible),
                patch.object(report_server, "sync_report_html", side_effect=lambda *_args, **_kwargs: None),
            ):
                response = client.post("/api/employee-performance/assignee-refresh", json={"assignee": "Alice"})
                self.assertEqual(response.status_code, 202)
                run_id = str((response.get_json() or {}).get("run_id") or "")
                self.assertTrue(run_id)

                import time

                deadline = time.time() + 5
                status = "running"
                while time.time() < deadline and status == "running":
                    poll = client.get(f"/api/employee-performance/assignee-refresh/{run_id}")
                    self.assertEqual(poll.status_code, 200)
                    run_payload = poll.get_json() or {}
                    status = str(((run_payload.get("run") or {}).get("status")) or "")
                    if status == "success":
                        break
                    time.sleep(0.05)
                self.assertEqual(status, "success")

            with sqlite3.connect(db_path) as conn:
                state_row = conn.execute("SELECT last_success_run_id FROM canonical_refresh_state WHERE id = 1").fetchone()
                latest_run_id = str(state_row[0] or "")
                self.assertEqual(latest_run_id, run_id)
                moved_row = conn.execute(
                    "SELECT assignee FROM canonical_issues WHERE run_id = ? AND issue_key = 'O2-101'",
                    (run_id,),
                ).fetchone()
                new_row = conn.execute(
                    "SELECT assignee FROM canonical_issues WHERE run_id = ? AND issue_key = 'O2-102'",
                    (run_id,),
                ).fetchone()
                derived_row = conn.execute(
                    "SELECT COUNT(*) FROM canonical_issue_actuals WHERE run_id = ?",
                    (run_id,),
                ).fetchone()
            self.assertIsNotNone(moved_row)
            self.assertEqual(str(moved_row[0] or ""), "Bob")
            self.assertIsNotNone(new_row)
            self.assertEqual(str(new_row[0] or ""), "Alice")
            self.assertGreater(int(derived_row[0] or 0), 0)

            exports_db = tdp / "jira_exports.db"
            with sqlite3.connect(exports_db) as conn:
                exported_moved = conn.execute("SELECT assignee FROM work_items WHERE issue_key = 'O2-101'").fetchone()
                exported_new = conn.execute("SELECT assignee FROM work_items WHERE issue_key = 'O2-102'").fetchone()
            self.assertIsNotNone(exported_moved)
            self.assertEqual(str(exported_moved[0] or ""), "Bob")
            self.assertIsNotNone(exported_new)
            self.assertEqual(str(exported_new[0] or ""), "Alice")

    def test_fix_type_rework_flows_from_work_items_to_worklogs(self):
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            tdp = Path(td)
            work_items_xlsx = tdp / "1_jira_work_items_export.xlsx"
            worklogs_xlsx = tdp / "2_jira_subtask_worklogs.xlsx"

            wi_wb = Workbook()
            wi_ws = wi_wb.active
            wi_ws.append(
                [
                    "project_key",
                    "issue_key",
                    "work_item_id",
                    "work_item_type",
                    "jira_issue_type",
                    "fix_type",
                    "summary",
                    "status",
                    "end_date",
                    "original_estimate_hours",
                    "parent_issue_key",
                ]
            )
            wi_ws.append(
                ["O2", "O2-101", "O2-101", "Subtask", "Sub-task", "rework", "A", "Open", "2026-02-10", 8, "O2-100"]
            )
            wi_wb.save(work_items_xlsx)

            wl_wb = Workbook()
            wl_ws = wl_wb.active
            wl_ws.append(["issue_id", "issue_assignee", "worklog_started", "hours_logged", "issue_type", "parent_story_id"])
            wl_ws.append(["O2-101", "Alice", "2026-02-11T10:00:00+00:00", 2, "Sub-task", "O2-100"])
            wl_wb.save(worklogs_xlsx)

            work_items = _load_work_items(work_items_xlsx)
            self.assertEqual(work_items["O2-101"]["fix_type"], "rework")

            worklogs = _load_worklogs(worklogs_xlsx, work_items)
            self.assertEqual(len(worklogs), 1)
            self.assertEqual(worklogs[0]["fix_type"], "rework")


if __name__ == "__main__":
    unittest.main()
