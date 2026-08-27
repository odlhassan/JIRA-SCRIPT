import tempfile
import unittest
from io import BytesIO
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

from employee_capacity_utilization_export import build_employee_capacity_utilization_workbook
from generate_employee_capacity_utilization_report import _epic_for_issue, _html
from report_server import REPORT_IDS_WITHOUT_REFRESH_WIDGET, _base_page_catalog, create_report_server_app


class EmployeeCapacityUtilizationReportTests(unittest.TestCase):
    def test_report_has_standalone_filters_and_canonical_payload(self):
        html = _html({"issues": [], "worklogs": [], "leaves": [], "teams": [], "profiles": [], "support": [], "resources": {}})
        self.assertIn("Employee Capacity &amp; Utilization", html)
        self.assertIn('id="month"', html)
        self.assertIn('id="profile"', html)
        self.assertIn('id="teams"', html)
        self.assertIn('id="resigned"', html)
        self.assertIn('id="support"', html)
        self.assertIn("Count logged hours from", html)
        self.assertIn("All work logged by employee", html)
        self.assertIn("Work logged on their assigned subtasks", html)
        self.assertIn("Assigned-subtask work only", html)
        self.assertIn("Logged hours (assigned subtasks)", html)
        self.assertIn('id="include-leaves-logged"', html)
        self.assertIn("isLeaveWorklog", html)
        self.assertIn("include_leaves:includeLeaves.checked", html)
        self.assertIn("T12:00:00", html)
        self.assertIn("previous UTC date", html)
        self.assertIn("process team", html)
        self.assertIn("Grand Total", html)
        self.assertIn("Canonical Jira worklogs", html)
        self.assertIn("/api/employee-capacity-utilization/data", html)
        self.assertIn("filters apply automatically", html)
        self.assertNotIn(">Refresh<", html)
        self.assertIn("All except Process Team", html)
        self.assertIn("#teams .team-option input[type=checkbox]", html)
        self.assertIn("width:17px!important", html)
        self.assertIn("position:sticky", html)
        self.assertIn("event.key==='Escape'", html)
        self.assertIn('id="util-a"', html)
        self.assertIn('id="util-b"', html)
        self.assertIn('id="util-c"', html)
        self.assertIn('name="color-scope"', html)
        self.assertIn("employeeCapacityUtilization.colorRules.v1", html)
        self.assertIn("utilizationBand", html)
        self.assertIn("Enter increasing values", html)
        self.assertIn("ecu-drawer", html)
        self.assertIn("ecu-resize", html)
        self.assertIn("pointermove", html)
        self.assertIn("wireDetailCells", html)
        self.assertIn("Work item title", html)
        self.assertIn("Hours logged", html)
        self.assertIn("jira_browse_base", html)
        self.assertIn("Download Excel", html)
        self.assertIn("/api/employee-capacity-utilization/export", html)
        self.assertIn("metric==='employee'?'logged':metric", html)
        self.assertIn("'Epic name'", html)

    def test_epic_is_resolved_through_parent_hierarchy(self):
        items = {
            "SUB-1": {"issue_type": "Sub-task", "parent_issue_key": "STORY-1"},
            "STORY-1": {"issue_type": "Story", "parent_issue_key": "EPIC-1"},
            "EPIC-1": {"issue_type": "Epic", "summary": "Customer onboarding"},
        }
        self.assertEqual(_epic_for_issue("SUB-1", items), ("EPIC-1", "Customer onboarding"))

    def test_report_is_available_to_page_categorization(self):
        pages = _base_page_catalog()
        self.assertTrue(any(page["page_key"] == "employee_capacity_utilization" for page in pages))

    def test_report_has_no_injected_refresh_widget(self):
        self.assertIn("employee_capacity_utilization", REPORT_IDS_WITHOUT_REFRESH_WIDGET)

    def test_live_api_returns_runtime_canonical_payload(self):
        payload = {
            "issues": [], "worklogs": [{"worklog_author": "Alice", "hours_logged": 7.5}],
            "leaves": [], "teams": [], "profiles": [], "support": [], "resources": {},
            "canonical_run_id": "production-run", "source": "canonical_database",
        }
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            (root / "report_html").mkdir()
            with patch("report_server.build_employee_capacity_utilization_payload", return_value=payload) as builder:
                client = create_report_server_app(root, "report_html").test_client()
                response = client.get("/api/employee-capacity-utilization/data")
            self.assertEqual(response.status_code, 200)
            self.assertEqual(response.get_json()["worklogs"][0]["hours_logged"], 7.5)
            self.assertEqual(response.get_json()["canonical_run_id"], "production-run")
            builder.assert_called_once_with(root / "assignee_hours_capacity.db")

    def test_excel_export_contains_structured_sheets_and_epic_columns(self):
        payload = {
            "issues": [
                {"issue_key": "EPIC-1", "issue_type": "Epic", "summary": "Customer onboarding", "assignee": "Alice", "parent_issue_key": ""},
                {"issue_key": "STORY-1", "issue_type": "Story", "summary": "Portal", "assignee": "Alice", "parent_issue_key": "EPIC-1"},
                {"issue_key": "SUB-1", "issue_type": "Sub-task", "summary": "Build form", "assignee": "Alice", "parent_issue_key": "STORY-1", "start_date": "2026-08-01", "due_date": "2026-08-31", "original_estimate_hours": 16},
            ],
            "worklogs": [
                {"issue_id": "SUB-1", "project_key": "APP", "worklog_author": "Alice", "issue_assignee": "Alice", "item_assignee": "Alice", "item_issue_type": "Sub-task", "item_summary": "Build form", "worklog_date": "2026-08-10", "hours_logged": 6, "epic_key": "EPIC-1", "epic_summary": "Customer onboarding"},
                {"issue_id": "TASK-1", "project_key": "APP", "worklog_author": "Alice", "issue_assignee": "Alice", "item_assignee": "Alice", "item_issue_type": "Task", "item_summary": "Team coordination", "worklog_date": "2026-08-12", "hours_logged": 4},
                {"issue_id": "RLT-1", "project_key": "RLT", "worklog_author": "Alice", "issue_assignee": "Alice", "item_assignee": "Alice", "item_issue_type": "Task", "item_summary": "Annual leave", "worklog_date": "2026-08-11", "hours_logged": 8},
            ],
            "leaves": [], "teams": [], "profiles": [], "support": [], "resources": {"Alice": {"resigned": False}},
            "canonical_run_id": "production-run", "generated_at": "2026-08-25T00:00:00Z", "jira_browse_base": "https://jira.example/browse",
        }
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            (root / "report_html").mkdir()
            with patch("report_server.build_employee_capacity_utilization_payload", return_value=payload):
                client = create_report_server_app(root, "report_html").test_client()
                response = client.post("/api/employee-capacity-utilization/export", json={"month": "2026-08", "scope": "any", "selected_teams": [], "display_resigned": False})
                assigned_response = client.post("/api/employee-capacity-utilization/export", json={"month": "2026-08", "scope": "assigned", "selected_teams": [], "display_resigned": False})
                response_with_leaves = client.post("/api/employee-capacity-utilization/export", json={"month": "2026-08", "scope": "any", "selected_teams": [], "display_resigned": False, "include_leaves": True})
        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.data), read_only=False, data_only=False)
        self.assertEqual(workbook.sheetnames, ["Export Info", "Summary", "Worklogs", "Booked Subtasks", "Leave Records", "Capacity Calendar", "Employees"])
        worklog_headers = [cell.value for cell in workbook["Worklogs"][1]]
        self.assertIn("Epic Name", worklog_headers)
        self.assertIn("Epic Jira Link", worklog_headers)
        self.assertEqual(workbook["Worklogs"]["F2"].value, "Customer onboarding")
        self.assertEqual(workbook["Worklogs"]["G2"].hyperlink.target, "https://jira.example/browse/EPIC-1")
        summary_headers = {cell.value: cell.column for cell in workbook["Summary"][1]}
        self.assertEqual(workbook["Summary"].cell(2, summary_headers["Logged Hours"]).value, 10)
        self.assertEqual(workbook["Worklogs"].max_row, 3)
        self.assertEqual(workbook["Export Info"]["B4"].value, "No")

        assigned_workbook = load_workbook(BytesIO(assigned_response.data), read_only=False, data_only=False)
        assigned_headers = {cell.value: cell.column for cell in assigned_workbook["Summary"][1]}
        self.assertEqual(assigned_workbook["Summary"].cell(2, assigned_headers["Logged Hours"]).value, 6)
        self.assertEqual(assigned_workbook["Worklogs"].max_row, 2)
        self.assertEqual(assigned_workbook["Export Info"]["B3"].value, "Work logged on their assigned subtasks")

        workbook_with_leaves = load_workbook(BytesIO(response_with_leaves.data), read_only=False, data_only=False)
        included_summary_headers = {cell.value: cell.column for cell in workbook_with_leaves["Summary"][1]}
        self.assertEqual(workbook_with_leaves["Summary"].cell(2, included_summary_headers["Logged Hours"]).value, 18)
        self.assertEqual(workbook_with_leaves["Worklogs"].max_row, 4)
        self.assertEqual(workbook_with_leaves["Export Info"]["B4"].value, "Yes")

    def test_official_weekday_leave_reduces_availability(self):
        payload = {
            "issues": [], "worklogs": [], "leaves": [], "teams": [], "support": [],
            "resources": {"Alice": {"resigned": False}},
            "profiles": [{
                "from_date": "2026-08-01", "to_date": "2026-08-31",
                "standard_hours_per_day": 8, "ramadan_start_date": "",
                "ramadan_end_date": "", "ramadan_hours_per_day": 6.5,
                "holiday_dates": ["2026-08-14"],
            }],
        }
        workbook_data, _ = build_employee_capacity_utilization_workbook(
            payload,
            {"month": "2026-08", "scope": "any", "selected_teams": [], "display_resigned": False},
        )
        workbook = load_workbook(workbook_data, read_only=False, data_only=False)
        headers = {cell.value: cell.column for cell in workbook["Summary"][1]}
        row = 2
        self.assertEqual(workbook["Summary"].cell(row, headers["Capacity (Hours)"]).value, 160)
        self.assertEqual(workbook["Summary"].cell(row, headers["Official Leaves (Days)"]).value, 1)
        self.assertEqual(workbook["Summary"].cell(row, headers["Availability (Hours)"]).value, 160)
