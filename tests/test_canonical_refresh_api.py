from __future__ import annotations

import sqlite3
import tempfile
import time
import unittest
from unittest.mock import patch
from pathlib import Path

from openpyxl import load_workbook

import report_server
from report_server import create_report_server_app


def _seed_managed_projects(db_path: Path) -> None:
    with sqlite3.connect(db_path) as conn:
        conn.execute(
            """
            INSERT INTO managed_projects(
                project_key, project_name, display_name, color_hex, is_active, created_at_utc, updated_at_utc
            ) VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            ("O2", "Octopus 2", "Octopus 2", "#336699", 1, "2026-01-01 00:00:00", "2026-01-01 00:00:00"),
        )
        conn.execute(
            """
            INSERT INTO managed_projects(
                project_key, project_name, display_name, color_hex, is_active, created_at_utc, updated_at_utc
            ) VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            ("FF", "Fast Forward", "Fast Forward", "#663399", 1, "2026-01-01 00:00:00", "2026-01-01 00:00:00"),
        )
        conn.execute(
            """
            INSERT INTO managed_projects(
                project_key, project_name, display_name, color_hex, is_active, created_at_utc, updated_at_utc
            ) VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            ("MN", "Moon", "Moon", "#993333", 0, "2026-01-01 00:00:00", "2026-01-01 00:00:00"),
        )
        conn.commit()


class CanonicalRefreshApiTests(unittest.TestCase):
    def test_schema_and_refresh_run_use_only_active_managed_projects(self):
        with patch.dict("os.environ", {"JIRA_PROJECT_KEYS": ""}, clear=False), tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            (root / "report_html").mkdir(parents=True, exist_ok=True)
            (root / "report_html" / "dashboard.html").write_text("<html><body>ok</body></html>", encoding="utf-8")
            app = create_report_server_app(base_dir=root, folder_raw="report_html")
            db_path = root / "assignee_hours_capacity.db"
            _seed_managed_projects(db_path)
            client = app.test_client()

            issue_catalog = {
                "O2-EP1": {
                    "id": "1001",
                    "key": "O2-EP1",
                    "fields": {
                        "project": {"key": "O2"},
                        "issuetype": {"name": "Epic"},
                        "summary": "Epic One",
                        "status": {"name": "Open"},
                        "assignee": {"displayName": "Alice"},
                        "parent": {},
                        "customfield_10014": "",
                        "customfield_1": "2026-01-05",
                        "duedate": "2026-03-01",
                        "created": "2025-12-20T00:00:00+00:00",
                        "updated": "2026-01-10T00:00:00+00:00",
                        "timeoriginalestimate": 28800,
                        "aggregatetimespent": 7200,
                    },
                    "changelog": {"histories": []},
                },
                "O2-ST1": {
                    "id": "1002",
                    "key": "O2-ST1",
                    "fields": {
                        "project": {"key": "O2"},
                        "issuetype": {"name": "Story"},
                        "summary": "Story One",
                        "status": {"name": "In Progress"},
                        "assignee": {"displayName": "Alice"},
                        "parent": {"key": "O2-EP1"},
                        "customfield_10014": "",
                        "customfield_1": "2026-02-01",
                        "duedate": "2026-02-15",
                        "created": "2026-01-02T00:00:00+00:00",
                        "updated": "2026-02-03T00:00:00+00:00",
                        "timeoriginalestimate": 14400,
                        "aggregatetimespent": 3600,
                    },
                    "changelog": {"histories": []},
                },
                "O2-SUB1": {
                    "id": "1003",
                    "key": "O2-SUB1",
                    "fields": {
                        "project": {"key": "O2"},
                        "issuetype": {"name": "Sub-task"},
                        "summary": "Subtask One",
                        "status": {"name": "Done"},
                        "assignee": {"displayName": "Bob"},
                        "parent": {"key": "O2-ST1"},
                        "customfield_10014": "",
                        "customfield_1": "",
                        "duedate": "2026-02-10",
                        "created": "2026-02-01T00:00:00+00:00",
                        "updated": "2026-02-11T00:00:00+00:00",
                        "timeoriginalestimate": 7200,
                        "aggregatetimespent": 7200,
                    },
                    "changelog": {"histories": []},
                },
                "FF-T1": {
                    "id": "1004",
                    "key": "FF-T1",
                    "fields": {
                        "project": {"key": "FF"},
                        "issuetype": {"name": "Task"},
                        "summary": "Task One",
                        "status": {"name": "Open"},
                        "assignee": {"displayName": "Carol"},
                        "parent": {},
                        "customfield_10014": "",
                        "customfield_1": "",
                        "duedate": "2026-08-05",
                        "created": "2026-07-01T00:00:00+00:00",
                        "updated": "2026-08-03T00:00:00+00:00",
                        "timeoriginalestimate": 21600,
                        "aggregatetimespent": 0,
                    },
                    "changelog": {"histories": []},
                },
                "MN-EP1": {
                    "id": "2001",
                    "key": "MN-EP1",
                    "fields": {
                        "project": {"key": "MN"},
                        "issuetype": {"name": "Epic"},
                        "summary": "Moon Epic",
                        "status": {"name": "Open"},
                        "assignee": {"displayName": "Dana"},
                        "parent": {},
                        "customfield_10014": "",
                        "customfield_1": "2026-01-01",
                        "duedate": "2026-01-30",
                        "created": "2026-01-01T00:00:00+00:00",
                        "updated": "2026-01-02T00:00:00+00:00",
                        "timeoriginalestimate": 3600,
                        "aggregatetimespent": 0,
                    },
                    "changelog": {"histories": []},
                },
            }

            def _fake_fetch_issues(_session, jql, fields):
                if "worklogDate" in jql:
                    return [issue_catalog["O2-SUB1"]]
                if "updated >=" in jql:
                    return [issue_catalog["FF-T1"]]
                return [issue_catalog["O2-ST1"], issue_catalog["MN-EP1"]]

            def _fake_fetch_issues_by_keys(_session, issue_keys, fields):
                return [issue_catalog[key] for key in issue_keys if key in issue_catalog]

            def _fake_fetch_stories(_session, epic_keys, fields, project_keys=None):
                if "O2-EP1" in epic_keys:
                    return [issue_catalog["O2-ST1"]]
                return []

            def _fake_fetch_subtasks(_session, story_keys, fields, project_keys=None):
                if "O2-ST1" in story_keys:
                    return [issue_catalog["O2-SUB1"]]
                return []

            def _fake_worklogs(_session, issue_key, delay_seconds=0.2, max_retries=5, request_timeout_seconds=30.0):
                if issue_key == "O2-SUB1":
                    return [
                        {
                            "id": "wl-1",
                            "started": "2026-02-09T10:00:00.000+0000",
                            "updated": "2026-02-09T11:00:00.000+0000",
                            "timeSpentSeconds": 3600,
                            "author": {"displayName": "Bob"},
                        }
                    ]
                if issue_key == "FF-T1":
                    return [
                        {
                            "id": "wl-2",
                            "started": "2026-08-02T10:00:00.000+0000",
                            "updated": "2026-08-02T11:00:00.000+0000",
                            "timeSpentSeconds": 1800,
                            "author": {"displayName": "Carol"},
                        }
                    ]
                return []

            with (
                patch.object(report_server, "get_session", return_value=object()),
                patch.object(report_server, "resolve_jira_start_date_field_id", return_value="customfield_1"),
                patch.object(report_server, "resolve_jira_end_date_field_ids", return_value=["duedate"]),
                patch.object(report_server, "export_resolve_fix_type_field_id", return_value=""),
                patch.object(report_server, "export_fetch_issues", side_effect=_fake_fetch_issues),
                patch.object(report_server, "export_fetch_issues_by_keys", side_effect=_fake_fetch_issues_by_keys),
                patch.object(report_server, "_fetch_story_issues_for_epics", side_effect=_fake_fetch_stories),
                patch.object(report_server, "_fetch_subtask_issues_for_stories", side_effect=_fake_fetch_subtasks),
                patch.object(report_server, "export_fetch_worklogs_for_issue", side_effect=_fake_worklogs),
            ):
                resp = client.post("/api/canonical-refresh", json={"year": 2026})
                self.assertEqual(resp.status_code, 202)
                run_id = str((resp.get_json() or {}).get("run_id") or "")
                self.assertTrue(run_id)

                final_status = ""
                payload = {}
                for _ in range(60):
                    status_resp = client.get(f"/api/canonical-refresh/{run_id}")
                    self.assertEqual(status_resp.status_code, 200)
                    payload = status_resp.get_json() or {}
                    run = payload.get("run") or {}
                    final_status = str(run.get("status") or "").lower()
                    if final_status in {"success", "failed", "canceled"}:
                        break
                    time.sleep(0.05)
                self.assertEqual(final_status, "success")
                run = payload.get("run") or {}
                self.assertEqual(int(run.get("scope_year") or 0), 2026)
                self.assertEqual(sorted(run.get("managed_project_keys") or []), ["FF", "O2"])

                current_resp = client.get("/api/canonical-refresh/current")
                self.assertEqual(current_resp.status_code, 200)
                current_run = ((current_resp.get_json() or {}).get("run") or {})
                self.assertEqual(str(current_run.get("run_id") or ""), run_id)

            with sqlite3.connect(db_path) as conn:
                tables = {
                    str(row[0])
                    for row in conn.execute("SELECT name FROM sqlite_master WHERE type = 'table'").fetchall()
                }
                state = conn.execute(
                    "SELECT active_run_id, last_success_run_id FROM canonical_refresh_state WHERE id = 1"
                ).fetchone()
                issue_keys = [
                    str(row[0])
                    for row in conn.execute(
                        "SELECT issue_key FROM canonical_issues WHERE run_id = ? ORDER BY issue_key",
                        (run_id,),
                    ).fetchall()
                ]
                worklog_ids = [
                    str(row[0])
                    for row in conn.execute(
                        "SELECT worklog_id FROM canonical_worklogs WHERE run_id = ? ORDER BY worklog_id",
                        (run_id,),
                    ).fetchall()
                ]
                reasons = {
                    (str(row[0]), str(row[1]))
                    for row in conn.execute(
                        "SELECT issue_key, reason FROM canonical_issue_scope_reasons WHERE run_id = ?",
                        (run_id,),
                    ).fetchall()
                }
                issue_actuals = {
                    str(row[0]): {
                        "actual_complete_date": str(row[1] or ""),
                        "due_completion_bucket": str(row[2] or ""),
                        "total_worklog_hours": float(row[3] or 0),
                    }
                    for row in conn.execute(
                        """
                        SELECT issue_key, actual_complete_date, due_completion_bucket, total_worklog_hours
                        FROM canonical_issue_actuals
                        WHERE run_id = ?
                        """,
                        (run_id,),
                    ).fetchall()
                }
                assignee_periods = {
                    (str(row[0]), str(row[1]), str(row[2]), str(row[3])): float(row[4] or 0)
                    for row in conn.execute(
                        """
                        SELECT project_key, assignee, period_type, period_value, total_hours
                        FROM canonical_assignee_period_hours
                        WHERE run_id = ?
                        """,
                        (run_id,),
                    ).fetchall()
                }
                planning_flags = {
                    str(row[0]): {
                        "missing_start_date": int(row[1] or 0),
                        "missing_due_date": int(row[2] or 0),
                        "missing_estimate": int(row[3] or 0),
                        "has_any_worklog": int(row[4] or 0),
                    }
                    for row in conn.execute(
                        """
                        SELECT issue_key, missing_start_date, missing_due_date, missing_estimate, has_any_worklog
                        FROM canonical_issue_planning_flags
                        WHERE run_id = ?
                        """,
                        (run_id,),
                    ).fetchall()
                }
                hierarchy_rows = {
                    str(row[0]): {
                        "direct_child_count": int(row[1] or 0),
                        "descendant_issue_count": int(row[2] or 0),
                    }
                    for row in conn.execute(
                        """
                        SELECT issue_key, direct_child_count, descendant_issue_count
                        FROM canonical_hierarchy_summary
                        WHERE run_id = ?
                        """,
                        (run_id,),
                    ).fetchall()
                }
                project_assignee_rows = {
                    (str(row[0]), str(row[1])): {
                        "issue_count": int(row[2] or 0),
                        "worklog_hours": float(row[3] or 0),
                        "completed_issue_count": int(row[4] or 0),
                    }
                    for row in conn.execute(
                        """
                        SELECT project_key, assignee, issue_count, worklog_hours, completed_issue_count
                        FROM canonical_project_assignee_summary
                        WHERE run_id = ?
                        """,
                        (run_id,),
                    ).fetchall()
                }
            self.assertIn("canonical_refresh_runs", tables)
            self.assertIn("canonical_issues", tables)
            self.assertIn("canonical_issue_links", tables)
            self.assertIn("canonical_worklogs", tables)
            self.assertIn("canonical_issue_scope_reasons", tables)
            self.assertIn("canonical_sync_state", tables)
            self.assertIn("canonical_issue_actuals", tables)
            self.assertIn("canonical_assignee_period_hours", tables)
            self.assertIn("canonical_issue_planning_flags", tables)
            self.assertIn("canonical_hierarchy_summary", tables)
            self.assertIn("canonical_project_assignee_summary", tables)
            self.assertIsNotNone(state)
            self.assertEqual(str(state[0] or ""), run_id)
            self.assertEqual(str(state[1] or ""), run_id)
            self.assertEqual(issue_keys, ["FF-T1", "O2-EP1", "O2-ST1", "O2-SUB1"])
            self.assertEqual(worklog_ids, ["wl-1", "wl-2"])
            self.assertIn(("O2-ST1", "planned_date_in_scope"), reasons)
            self.assertIn(("FF-T1", "updated_in_scope"), reasons)
            self.assertIn(("O2-SUB1", "worklog_in_scope"), reasons)
            self.assertIn(("O2-EP1", "parent_of_included_child"), reasons)
            self.assertEqual(issue_actuals["O2-SUB1"]["actual_complete_date"], "2026-02-09")
            self.assertEqual(issue_actuals["O2-SUB1"]["due_completion_bucket"], "before_due")
            self.assertEqual(issue_actuals["FF-T1"]["total_worklog_hours"], 0.5)
            self.assertEqual(assignee_periods[("O2", "Bob", "day", "2026-02-09")], 1.0)
            self.assertEqual(assignee_periods[("FF", "Carol", "month", "2026-08")], 0.5)
            self.assertEqual(planning_flags["FF-T1"]["missing_start_date"], 1)
            self.assertEqual(planning_flags["FF-T1"]["has_any_worklog"], 1)
            self.assertEqual(hierarchy_rows["O2-EP1"]["direct_child_count"], 1)
            self.assertEqual(hierarchy_rows["O2-EP1"]["descendant_issue_count"], 2)
            self.assertEqual(project_assignee_rows[("O2", "Bob")]["issue_count"], 1)
            self.assertEqual(project_assignee_rows[("O2", "Bob")]["worklog_hours"], 1.0)
            self.assertEqual(project_assignee_rows[("O2", "Bob")]["completed_issue_count"], 1)
            work_items_xlsx = root / "1_jira_work_items_export.xlsx"
            worklogs_xlsx = root / "2_jira_subtask_worklogs.xlsx"
            rollup_xlsx = root / "3_jira_subtask_worklog_rollup.xlsx"
            nested_view_xlsx = root / "nested view.xlsx"
            self.assertTrue(work_items_xlsx.exists())
            self.assertTrue(worklogs_xlsx.exists())
            self.assertTrue(rollup_xlsx.exists())
            self.assertTrue(nested_view_xlsx.exists())

            wb = load_workbook(work_items_xlsx, read_only=True, data_only=True)
            ws = wb.active
            header = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
            first_row = list(next(ws.iter_rows(min_row=2, max_row=2, values_only=True)))
            wb.close()
            self.assertIn("issue_key", header)
            self.assertIn("actual_end_date", header)
            self.assertIn(first_row[1], {"FF-T1", "O2-EP1", "O2-ST1", "O2-SUB1"})

            wb = load_workbook(worklogs_xlsx, read_only=True, data_only=True)
            ws = wb.active
            header = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            wb.close()
            self.assertIn("worklog_author", header)
            self.assertEqual(len(data_rows), 2)

            wb = load_workbook(rollup_xlsx, read_only=True, data_only=True)
            ws = wb.active
            header = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            wb.close()
            self.assertIn("total hours_logged", header)
            self.assertEqual(len(data_rows), 1)

            wb = load_workbook(nested_view_xlsx, read_only=True, data_only=True)
            ws = wb.active
            header = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            wb.close()
            self.assertEqual(header[0], "Aspect")
            self.assertGreaterEqual(len(data_rows), 4)

    def test_refresh_requires_active_managed_projects(self):
        with patch.dict("os.environ", {"JIRA_PROJECT_KEYS": ""}, clear=False), tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            (root / "report_html").mkdir(parents=True, exist_ok=True)
            (root / "report_html" / "dashboard.html").write_text("<html><body>ok</body></html>", encoding="utf-8")
            app = create_report_server_app(base_dir=root, folder_raw="report_html")
            client = app.test_client()

            resp = client.post("/api/canonical-refresh", json={"year": 2026})
            self.assertEqual(resp.status_code, 400)
            body = resp.get_json() or {}
            self.assertFalse(body.get("ok"))
            self.assertIn("No active managed projects", str(body.get("error") or ""))

    def test_cancel_marks_running_canonical_refresh(self):
        with patch.dict("os.environ", {"JIRA_PROJECT_KEYS": ""}, clear=False), tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            (root / "report_html").mkdir(parents=True, exist_ok=True)
            (root / "report_html" / "dashboard.html").write_text("<html><body>ok</body></html>", encoding="utf-8")
            app = create_report_server_app(base_dir=root, folder_raw="report_html")
            db_path = root / "assignee_hours_capacity.db"
            _seed_managed_projects(db_path)
            client = app.test_client()

            original_runner = report_server._run_canonical_phase1_refresh

            def _slow_runner(db_path, run_id, scope_year, managed_project_keys, trigger_source="api_refresh_async"):
                started = report_server._canonical_now_utc()
                with sqlite3.connect(db_path) as conn:
                    conn.execute(
                        """
                        INSERT OR REPLACE INTO canonical_refresh_runs(
                            run_id, scope_year, managed_project_keys_json, started_at_utc, ended_at_utc,
                            status, trigger_source, error_message, stats_json,
                            progress_step, progress_pct, cancel_requested, updated_at_utc
                        ) VALUES (?, ?, ?, ?, NULL, 'running', ?, '', '{}', 'loading_managed_project_scope', 5, 0, ?)
                        """,
                        (
                            run_id,
                            int(scope_year or 0),
                            '["FF","O2"]',
                            started,
                            trigger_source,
                            started,
                        ),
                    )
                    conn.execute(
                        "UPDATE canonical_refresh_state SET active_run_id = ?, updated_at_utc = ? WHERE id = 1",
                        (run_id, started),
                    )
                    conn.commit()
                for _ in range(40):
                    if report_server._canonical_is_cancel_requested(db_path, run_id):
                        return {"ok": True}, 200
                    time.sleep(0.05)
                return {"ok": True}, 200

            report_server._run_canonical_phase1_refresh = _slow_runner
            try:
                start_resp = client.post("/api/canonical-refresh", json={"year": 2026})
                self.assertEqual(start_resp.status_code, 202)
                run_id = str((start_resp.get_json() or {}).get("run_id") or "")
                self.assertTrue(run_id)

                cancel_resp = client.post("/api/canonical-refresh/cancel", json={"run_id": run_id})
                self.assertEqual(cancel_resp.status_code, 200)
                cancel_body = cancel_resp.get_json() or {}
                self.assertTrue(cancel_body.get("ok"))
                self.assertEqual(str(cancel_body.get("status") or ""), "cancel_requested")

                cancel_requested = 0
                for _ in range(30):
                    with sqlite3.connect(db_path) as conn:
                        row = conn.execute(
                            "SELECT cancel_requested FROM canonical_refresh_runs WHERE run_id = ?",
                            (run_id,),
                        ).fetchone()
                    cancel_requested = int((row[0] if row else 0) or 0)
                    if cancel_requested == 1:
                        break
                    time.sleep(0.05)
                self.assertEqual(cancel_requested, 1)
            finally:
                report_server._run_canonical_phase1_refresh = original_runner


if __name__ == "__main__":
    unittest.main()
