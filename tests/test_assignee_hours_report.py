from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from generate_assignee_hours_report import (
    DAY_CROSSTAB_SHEET,
    MONTH_CROSSTAB_SHEET,
    SUMMARY_SHEET,
    WEEK_CROSSTAB_SHEET,
    _load_leave_subtask_rows,
    _load_worklog_rows,
    _read_summary_xlsx,
    _write_summary_xlsx,
    _save_capacity_settings,
    calculate_capacity_metrics,
    create_server_app,
    aggregate_rows,
    build_crosstab,
    day_boundary_flags,
    extract_project_key,
    iso_week_code,
    month_code,
    parse_worklog_date,
    period_sort_key,
)


class AssigneeHoursReportTests(unittest.TestCase):
    def test_parse_worklog_date(self):
        self.assertEqual(parse_worklog_date("2026-02-20T11:15:00.000+0500"), "2026-02-20")
        self.assertEqual(parse_worklog_date("2026-02-20T11:15:00+05:00"), "2026-02-20")
        self.assertEqual(parse_worklog_date("invalid"), "")

    def test_iso_week_boundary(self):
        self.assertEqual(iso_week_code("2021-01-01"), "2020-W53")
        self.assertEqual(iso_week_code("2026-02-20"), "2026-W08")

    def test_extract_project_key(self):
        self.assertEqual(extract_project_key("O2-123"), "O2")
        self.assertEqual(extract_project_key("digitallog-22"), "DIGITALLOG")
        self.assertEqual(extract_project_key("bad key"), "UNKNOWN")

    def test_load_worklog_rows_normalization(self):
        with tempfile.TemporaryDirectory() as td:
            path = Path(td) / "worklogs.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["issue_id", "issue_assignee", "worklog_author", "worklog_started", "hours_logged"])
            ws.append(["O2-10", "Alice", "Hassan Malik", "2026-02-19T10:00:00+0500", 2.5])
            ws.append(["FF-9", "", "", "2026-02-20T09:00:00+0500", 1.25])
            ws.append(["O2-11", "Bob", "", 3])
            wb.save(path)

            rows = _load_worklog_rows(path)
            self.assertEqual(len(rows), 2)
            self.assertEqual(rows[0]["project_key"], "O2")
            self.assertEqual(rows[0]["period_day"], "2026-02-19")
            self.assertEqual(rows[0]["period_month"], month_code("2026-02-19"))
            self.assertEqual(rows[1]["issue_assignee"], "Unassigned")
            self.assertEqual(rows[0]["worklog_author"], "Hassan Malik")
            self.assertEqual(rows[1]["worklog_author"], "Unassigned")

    def test_load_leave_subtask_rows(self):
        with tempfile.TemporaryDirectory() as td:
            path = Path(td) / "rlt_leave_report.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "Raw_Subtasks"
            ws.append(
                [
                    "issue_key",
                    "assignee",
                    "start_date",
                    "due_date",
                    "original_estimate_hours",
                    "total_worklog_hours",
                    "leave_classification",
                ]
            )
            ws.append(["RLT-172", "Maria Sharafat", "2026-01-01", "2026-03-31", 528, 528, "Planned"])
            ws.append(["RLT-173", "Maria Sharafat", "bad", None, 176, 176, "Planned"])
            wb.save(path)

            rows = _load_leave_subtask_rows(path)
            self.assertEqual(len(rows), 2)
            self.assertEqual(rows[0]["issue_key"], "RLT-172")
            self.assertEqual(rows[0]["start_date"], "2026-01-01")
            self.assertEqual(rows[0]["due_date"], "2026-03-31")
            self.assertEqual(rows[0]["original_estimate_hours"], 528.0)
            self.assertEqual(rows[0]["total_worklog_hours"], 528.0)
            self.assertEqual(rows[1]["start_date"], "")

    def test_aggregate_rows_period_assignee(self):
        rows = [
            {
                "project_key": "O2",
                "worklog_date": "2026-02-18",
                "period_day": "2026-02-18",
                "period_week": "2026-W08",
                "period_month": "2026-02",
                "issue_assignee": "Alice",
                "worklog_author": "Hassan Malik",
                "hours_logged": 2.0,
            },
            {
                "project_key": "O2",
                "worklog_date": "2026-02-19",
                "period_day": "2026-02-19",
                "period_week": "2026-W08",
                "period_month": "2026-02",
                "issue_assignee": "Alice",
                "worklog_author": "Hassan Malik",
                "hours_logged": 3.0,
            },
            {
                "project_key": "FF",
                "worklog_date": "2026-02-20",
                "period_day": "2026-02-20",
                "period_week": "2026-W08",
                "period_month": "2026-02",
                "issue_assignee": "Bob",
                "hours_logged": 4.0,
            },
        ]
        grouped = aggregate_rows(
            rows=rows,
            granularity="week",
            from_date="2026-02-17",
            to_date="2026-02-20",
            selected_projects={"O2"},
        )
        self.assertEqual(len(grouped), 1)
        self.assertEqual(grouped[0]["period"], "2026-W08")
        self.assertEqual(grouped[0]["assignee"], "Hassan Malik")
        self.assertEqual(grouped[0]["total_hours"], 5.0)

    def test_day_boundary_flags(self):
        self.assertEqual(
            day_boundary_flags("2026-03-01"),
            {"is_week_end": True, "is_month_end": False, "is_both": False},
        )
        self.assertEqual(
            day_boundary_flags("2026-02-28"),
            {"is_week_end": False, "is_month_end": True, "is_both": False},
        )
        self.assertEqual(
            day_boundary_flags("2026-05-31"),
            {"is_week_end": True, "is_month_end": True, "is_both": True},
        )

    def test_period_sort_key_week(self):
        periods = ["2026-W02", "2025-W52", "2026-W01"]
        sorted_periods = sorted(periods, key=lambda p: period_sort_key(p, "week"))
        self.assertEqual(sorted_periods, ["2025-W52", "2026-W01", "2026-W02"])

    def test_build_crosstab_day_totals(self):
        rows = [
            {
                "project_key": "O2",
                "worklog_date": "2026-02-20",
                "period_day": "2026-02-20",
                "period_week": "2026-W08",
                "period_month": "2026-02",
                "issue_assignee": "Alice",
                "worklog_author": "Hassan Malik",
                "hours_logged": 2.5,
            },
            {
                "project_key": "O2",
                "worklog_date": "2026-02-21",
                "period_day": "2026-02-21",
                "period_week": "2026-W08",
                "period_month": "2026-02",
                "issue_assignee": "Alice",
                "worklog_author": "Hassan Malik",
                "hours_logged": 1.5,
            },
            {
                "project_key": "FF",
                "worklog_date": "2026-02-20",
                "period_day": "2026-02-20",
                "period_week": "2026-W08",
                "period_month": "2026-02",
                "issue_assignee": "Bob",
                "hours_logged": 4.0,
            },
        ]
        crosstab = build_crosstab(
            rows=rows,
            granularity="day",
            from_date="2026-02-20",
            to_date="2026-02-21",
            selected_projects={"O2"},
        )
        self.assertEqual(crosstab["columns"], ["2026-02-20", "2026-02-21"])
        self.assertEqual(len(crosstab["row_items"]), 1)
        self.assertEqual(crosstab["row_items"][0]["assignee"], "Hassan Malik")
        self.assertEqual(crosstab["row_items"][0]["values"]["2026-02-20"], 2.5)
        self.assertEqual(crosstab["row_items"][0]["values"]["2026-02-21"], 1.5)
        self.assertEqual(crosstab["row_items"][0]["total_hours"], 4.0)
        self.assertEqual(crosstab["grand_totals"]["columns"]["2026-02-20"], 2.5)
        self.assertEqual(crosstab["grand_totals"]["columns"]["2026-02-21"], 1.5)
        self.assertEqual(crosstab["grand_totals"]["overall_total"], 4.0)

    def test_build_crosstab_month_order(self):
        rows = [
            {
                "project_key": "O2",
                "worklog_date": "2025-12-15",
                "period_day": "2025-12-15",
                "period_week": "2025-W51",
                "period_month": "2025-12",
                "issue_assignee": "Alice",
                "worklog_author": "Hassan Malik",
                "hours_logged": 1.0,
            },
            {
                "project_key": "O2",
                "worklog_date": "2026-01-12",
                "period_day": "2026-01-12",
                "period_week": "2026-W03",
                "period_month": "2026-01",
                "issue_assignee": "Alice",
                "worklog_author": "Hassan Malik",
                "hours_logged": 2.0,
            },
        ]
        crosstab = build_crosstab(
            rows=rows,
            granularity="month",
            from_date="2025-12-01",
            to_date="2026-01-31",
        )
        self.assertEqual(crosstab["columns"], ["2025-12", "2026-01"])

    def test_calculate_capacity_metrics(self):
        result = calculate_capacity_metrics(
            {
                "from_date": "2026-03-01",
                "to_date": "2026-03-10",
                "employee_count": 10,
                "standard_hours_per_day": 8,
                "ramadan_start_date": "2026-03-04",
                "ramadan_end_date": "2026-03-06",
                "ramadan_hours_per_day": 6.5,
                "holiday_dates": ["2026-03-03", "2026-03-08"],
            }
        )
        self.assertEqual(result["metrics"]["total_weekdays"], 7)
        self.assertEqual(result["metrics"]["holiday_weekdays"], 1)
        self.assertEqual(result["metrics"]["ramadan_weekdays"], 3)
        self.assertEqual(result["metrics"]["non_ramadan_weekdays"], 3)
        self.assertEqual(result["metrics"]["available_capacity_hours"], 435.0)

    def test_capacity_api_get_post(self):
        with tempfile.TemporaryDirectory() as td:
            td_path = Path(td)
            worklog_path = td_path / "worklogs.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["issue_id", "issue_assignee", "worklog_started", "hours_logged"])
            ws.append(["O2-1", "Alice", "2026-03-02T10:00:00+0500", 2.0])
            wb.save(worklog_path)

            paths = {
                "base_dir": td_path,
                "input_path": worklog_path,
                "work_items_path": td_path / "1_jira_work_items_export.xlsx",
                "summary_path": td_path / "assignee_hours_report.xlsx",
                "html_path": td_path / "assignee_hours_report.html",
                "db_path": td_path / "assignee_hours_capacity.db",
                "leave_report_path": td_path / "rlt_leave_report.xlsx",
            }
            app = create_server_app(paths)
            client = app.test_client()

            get_resp = client.get("/api/capacity?from=2026-03-01&to=2026-03-31")
            self.assertEqual(get_resp.status_code, 200)
            get_json = get_resp.get_json()
            self.assertEqual(get_json["settings"]["employee_count"], 0)
            self.assertIn("leave_metrics", get_json)

            post_resp = client.post(
                "/api/capacity",
                json={
                    "from_date": "2026-03-01",
                    "to_date": "2026-03-31",
                    "employee_count": 5,
                    "standard_hours_per_day": 8,
                    "ramadan_start_date": "2026-03-10",
                    "ramadan_end_date": "2026-03-20",
                    "ramadan_hours_per_day": 6.5,
                    "holiday_dates": ["2026-03-17"],
                },
            )
            self.assertEqual(post_resp.status_code, 200)
            post_json = post_resp.get_json()
            self.assertEqual(post_json["settings"]["employee_count"], 5)
            self.assertIn("leave_metrics", post_json)

            get_resp_2 = client.get("/api/capacity?from=2026-03-01&to=2026-03-31")
            self.assertEqual(get_resp_2.status_code, 200)
            get_json_2 = get_resp_2.get_json()
            self.assertEqual(get_json_2["settings"]["employee_count"], 5)

            profiles_resp = client.get("/api/capacity/profiles")
            self.assertEqual(profiles_resp.status_code, 200)
            profiles_json = profiles_resp.get_json()
            self.assertTrue(any(p["from_date"] == "2026-03-01" and p["to_date"] == "2026-03-31" for p in profiles_json["profiles"]))

    def test_capacity_api_validation(self):
        with tempfile.TemporaryDirectory() as td:
            td_path = Path(td)
            worklog_path = td_path / "worklogs.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["issue_id", "issue_assignee", "worklog_started", "hours_logged"])
            ws.append(["O2-1", "Alice", "2026-03-02T10:00:00+0500", 2.0])
            wb.save(worklog_path)

            paths = {
                "base_dir": td_path,
                "input_path": worklog_path,
                "work_items_path": td_path / "1_jira_work_items_export.xlsx",
                "summary_path": td_path / "assignee_hours_report.xlsx",
                "html_path": td_path / "assignee_hours_report.html",
                "db_path": td_path / "assignee_hours_capacity.db",
                "leave_report_path": td_path / "rlt_leave_report.xlsx",
            }
            app = create_server_app(paths)
            client = app.test_client()

            post_resp = client.post(
                "/api/capacity",
                json={
                    "from_date": "2026-03-01",
                    "to_date": "2026-03-31",
                    "employee_count": 3,
                    "standard_hours_per_day": 8,
                    "ramadan_start_date": "2026-03-20",
                    "ramadan_end_date": "",
                    "ramadan_hours_per_day": 6.5,
                    "holiday_dates": [],
                },
            )
            self.assertEqual(post_resp.status_code, 400)

    def test_summary_xlsx_roundtrip(self):
        rows = [
            {
                "project_key": "O2",
                "worklog_date": "2026-02-20",
                "period_day": "2026-02-20",
                "period_week": "2026-W08",
                "period_month": "2026-02",
                "issue_assignee": "Alice",
                "worklog_author": "Hassan Malik",
                "hours_logged": 2.5,
            }
        ]
        with tempfile.TemporaryDirectory() as td:
            path = Path(td) / "assignee_hours_report.xlsx"
            _write_summary_xlsx(rows, path)
            loaded = _read_summary_xlsx(path)
            self.assertEqual(len(loaded), 1)
            self.assertEqual(loaded[0]["project_key"], "O2")
            self.assertEqual(loaded[0]["issue_assignee"], "Alice")
            self.assertEqual(loaded[0]["worklog_author"], "Hassan Malik")
            self.assertEqual(loaded[0]["hours_logged"], 2.5)

            wb = load_workbook(path, data_only=True)
            self.assertIn(SUMMARY_SHEET, wb.sheetnames)
            self.assertIn(DAY_CROSSTAB_SHEET, wb.sheetnames)
            self.assertIn(WEEK_CROSSTAB_SHEET, wb.sheetnames)
            self.assertIn(MONTH_CROSSTAB_SHEET, wb.sheetnames)

            day_ws = wb[DAY_CROSSTAB_SHEET]
            self.assertEqual(day_ws.cell(row=1, column=1).value, "User")
            self.assertEqual(day_ws.cell(row=2, column=1).value, "Hassan Malik")
            self.assertEqual(day_ws.cell(row=2, column=3).value, 2.5)
            self.assertEqual(day_ws.cell(row=3, column=1).value, "Grand Total")
            self.assertEqual(day_ws.cell(row=3, column=3).value, 2.5)
            wb.close()


if __name__ == "__main__":
    unittest.main()
