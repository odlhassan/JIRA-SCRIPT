from __future__ import annotations

import sqlite3
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from generate_assignee_hours_report import _build_html as build_assignee_html
from generate_employee_performance_report import _build_html as build_employee_perf_html
from generate_nested_view_html import _build_html as build_nested_html
from generate_phase_rmi_gantt_html import _build_html as build_team_rmi_gantt_html
from generate_planned_rmis_html import _build_html as build_planned_rmis_html
from generate_rnd_data_story import _build_html as build_rnd_story_html
from openpyxl import Workbook
from report_server import _seating_planner_html, create_report_server_app


def _write_minimal_assignee_workbook(root: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(["project_key", "worklog_date", "period_day", "period_week", "period_month", "issue_assignee", "hours_logged"])
    ws.append(["O2", "2026-02-01", "2026-02-01", "2026-W05", "2026-02", "Alice", 1.0])
    wb.save(root / "assignee_hours_report.xlsx")


def _write_epics_import_source(path: Path, *, total_override: float | None = None) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "OmniConnect RMI"
    epic_url = "https://octopusdtlsupport.atlassian.net/browse/O2-321"
    headers = [
        "Sr #", "Category", "Components", "Road Map Items", "Jira ID", "Originator",
        "Value", "Priority", "Priority", "Priority", "Priority", "Plan Status", "Work Status",
        "Prc Design", "R/URS", "R/DDS", "Dev", "SQA", "Prc Test", "Doc", "Reg SQA", "Release",
        "Man Days", "Optimistic (50%)", "Pessimistic (10%)", "Est Formula", "TK's TARGET",
        "Start Date", "Dev End", "SQA HO", "Prod Date",
        "Prc Design", "R/URS (5%)", "R/DDS (10%)", "Dev (40%)", "Handover", "SQA (15%)",
        "Bug Fixing (15%)", "Prc Test", "Doc (5%)", "Reg SQA (10%)", "Release",
    ]
    ws.merge_cells("A1:D1")
    ws["A1"] = "OmniConnect Roadmap Items"
    ws.merge_cells("N1:W1")
    ws["N1"] = "RnD Most likely"
    ws.merge_cells("X1:AP1")
    ws["X1"] = "TK Budgeted"
    ws.append(headers)
    phase_values = [1, 2, 3, 4, 5, 6, 7, 8, 9]
    ws.append([
        1, "Input", "Streaming", "Streaming Epic", "[O2-321] Streaming Pub/Sub Architecture - Jira",
        "RnD", None, None, None, None, None, "Plan", "Ready",
        *phase_values, total_override if total_override is not None else sum(phase_values),
    ])
    ws["E3"].hyperlink = epic_url
    ws.append([
        2, None, None, "Streaming child without Jira", None,
        "Imp Team", None, None, None, None, None, "UnPlan", "Blocked",
    ])
    ws.merge_cells("B3:B4")
    ws.merge_cells("C3:C4")
    wb.save(path)


class ReportUiSmokeTests(unittest.TestCase):
    def test_seating_planner_pdf_export_uses_single_page_vector_print_view(self):
        html = _seating_planner_html()

        self.assertIn("@page{size:A4 landscape;margin:8mm;}", html)
        self.assertIn('id="spPrintPage"', html)
        self.assertIn('id="spPrintStage"', html)
        self.assertIn("function buildVectorPrintView()", html)
        self.assertIn("function exportSinglePagePdf()", html)
        self.assertIn("addEventListener('click',exportSinglePagePdf)", html)
        self.assertIn("cloneNode(true)", html)
        self.assertIn("body > :not(#spPrintPage)", html)

    def test_seating_planner_zoom_prefers_layout_zoom_for_crisp_canvas(self):
        html = _seating_planner_html()

        self.assertIn("if ('zoom' in canvas.style)", html)
        self.assertIn("canvas.style.zoom = String(G.zoom)", html)
        self.assertIn("zoom-transform-fallback", html)
        self.assertIn("canvas.style.transform = `scale(${G.zoom})`", html)

    def test_assignee_header_and_drawer_controls_exist(self):
        payload = {
            "rows": [],
            "projects": [],
            "default_from": "2026-01-01",
            "default_to": "2026-01-31",
            "capacity_profiles": [],
            "leave_daily_rows": [],
            "generated_at": "2026-02-21 00:00 UTC",
        }
        html = build_assignee_html(payload)
        self.assertIn('class="enterprise-header"', html)
        self.assertIn('id="open-capacity-settings"', html)
        self.assertIn('id="settings-drawer"', html)
        self.assertIn('id="settings-drawer-overlay"', html)
        self.assertIn('id="capacity-employees"', html)
        self.assertIn('id="capacity-profile-select"', html)
        self.assertIn('id="capacity-profile-apply"', html)
        self.assertIn('id="actual-hours-mode"', html)
        self.assertIn("/api/capacity/profiles", html)
        self.assertIn("/api/capacity/calculate", html)
        self.assertIn("/api/actual-hours/aggregate", html)

    def test_employee_performance_simple_score_drawer_controls_exist(self):
        payload = {
            "worklogs": [],
            "work_items": [],
            "leave_rows": [],
            "settings": {},
            "teams": [],
            "projects": [],
            "default_from": "2026-01-01",
            "default_to": "2026-01-31",
            "leave_hours_per_day": 8,
            "entities_catalog": [],
            "managed_fields": [],
            "capacity_profiles": [],
            "simple_scoring": [],
            "jira_browse_base": "https://example.atlassian.net/browse",
            "generated_at": "2026-02-21 00:00 UTC",
        }
        html = build_employee_perf_html(payload)
        self.assertIn('id="score-detail-drawer"', html)
        self.assertIn('id="score-detail-drawer-close"', html)
        self.assertIn('id="score-detail-drawer-body"', html)
        self.assertIn('id="header-average-performance-value"', html)
        self.assertIn("Simple Score Details", html)
        self.assertIn("Planned Due Date", html)
        self.assertIn("Last Logged Date", html)
        self.assertIn("Actual Complete Date", html)

    def test_nested_view_options_and_profile_controls_exist(self):
        payload = {
            "generated_at": "2026-02-21 00:00 UTC",
            "source_file": "nested view.xlsx",
            "rows": [],
            "capacity_profiles": [],
            "leave_daily_rows": [],
            "leave_subtask_rows": [],
        }
        html = build_nested_html(payload)
        self.assertIn('class="scorecards"', html)
        self.assertIn('id="view-options"', html)
        self.assertIn('id="view-options-toggle"', html)
        self.assertIn('id="view-options-menu"', html)
        self.assertIn('id="theme-toggle"', html)
        self.assertIn('id="toggle-density"', html)
        self.assertIn('id="toggle-no-entry"', html)
        self.assertIn('id="toggle-product"', html)
        self.assertIn('id="date-filter-from"', html)
        self.assertIn('id="date-filter-to"', html)
        self.assertIn('id="planned-hours-source"', html)
        self.assertIn('id="extended-actual-hours-toggle"', html)
        self.assertIn('id="project-filter-progress"', html)
        self.assertIn('id="team-filter-progress"', html)
        self.assertIn('id="actual-hours-mode"', html)
        self.assertIn('href="/settings/capacity"', html)
        self.assertIn('id="capacity-profile-expanded"', html)
        self.assertIn('id="capacity-profile-resize-handle"', html)
        self.assertIn('id="score-total-capacity-formula"', html)
        self.assertIn('id="score-total-capacity-formula-hours"', html)
        self.assertIn('id="score-total-leaves-planned-formula"', html)
        self.assertIn("Availability", html)
        self.assertIn("Total Capacity (Hours) - Total Leaves Planned", html)
        self.assertIn("Capacity Profile Calendar", html)
        self.assertIn("function renderCapacityProfileExpanded(profile)", html)
        self.assertIn("function setCapacityDrawerWidth(widthVw)", html)
        self.assertIn("function startCapacityDrawerResize(event)", html)
        self.assertIn("renderCapacityProfileDetails();", html)
        self.assertIn("function subtaskMatchesActualHoursMode(row, bounds)", html)
        self.assertIn(
            "Sum(All Logged Hours for subtasks whose planned Start OR Due date is within selected range)",
            html,
        )
        self.assertIn(
            "Sum(Logged Hours in selected date range for subtasks with worklog dates in selected range)",
            html,
        )
        self.assertIn("Approved Days", html)
        self.assertIn("Approved Hours", html)
        self.assertIn("Planned Days", html)
        self.assertIn("Planned Hours", html)

    def test_nested_capacity_endpoints_present(self):
        payload = {
            "generated_at": "2026-02-21 00:00 UTC",
            "source_file": "nested view.xlsx",
            "rows": [],
            "capacity_profiles": [],
            "leave_daily_rows": [],
            "leave_subtask_rows": [],
        }
        html = build_nested_html(payload)
        self.assertIn("/api/capacity/profiles", html)
        self.assertIn("/api/nested-view/actual-hours", html)
        self.assertIn("/api/actual-hours/aggregate", html)
        self.assertIn("/api/manage-fields?include_inactive=0", html)
        self.assertIn("hasCapacityApi", html)
        self.assertIn("hasManagedFieldsApi", html)
        self.assertIn("refreshManagedFieldsFromApi", html)
        self.assertIn("evaluateManagedField", html)
        self.assertIn("leave_subtask_rows", html)

    def test_nested_project_filter_refreshes_after_live_tree_reload(self):
        payload = {
            "generated_at": "2026-02-21 00:00 UTC",
            "source_file": "nested view.xlsx",
            "rows": [],
            "capacity_profiles": [],
            "leave_daily_rows": [],
            "leave_subtask_rows": [],
        }
        html = build_nested_html(payload)
        self.assertIn("function refreshProjectFilterOptions(options)", html)
        self.assertIn("refreshProjectFilterOptions({ preserveSelection: false });", html)
        self.assertIn("refreshProjectFilterOptions({ preserveSelection: true });", html)
        self.assertIn('fetch("/api/nested-view/tree"', html)

    def test_nested_subtask_logs_scope_uses_worklog_date_basis(self):
        payload = {
            "generated_at": "2026-02-21 00:00 UTC",
            "source_file": "nested view.xlsx",
            "rows": [],
            "capacity_profiles": [],
            "leave_daily_rows": [],
            "leave_subtask_rows": [],
        }
        html = build_nested_html(payload)
        self.assertIn("function resolveScopedSubtaskBasis(plannedHoursSource)", html)
        self.assertIn('? \"log_date\"', html)
        self.assertIn(': \"planned_dates\";', html)
        self.assertIn('"&scope_basis=" + scopeBasisParam', html)
        self.assertIn('if (scorecardPlannedHoursSource === "subtask_logs") {', html)
        self.assertIn("return subtaskLoggedInRange(row);", html)

    def test_nested_leave_scorecards_prefer_distributed_buckets_for_date_accuracy(self):
        payload = {
            "generated_at": "2026-02-21 00:00 UTC",
            "source_file": "nested view.xlsx",
            "rows": [],
            "capacity_profiles": [],
            "leave_daily_rows": [],
            "leave_distributed_rows": [
                {
                    "issue_key": "RLT-172",
                    "assignee": "Alice",
                    "start_date": "2026-03-01",
                    "due_date": "2026-03-01",
                    "planned_date_for_bucket": "2026-03-01",
                    "original_estimate_hours": 8,
                    "total_worklog_hours": 8,
                    "leave_classification": "Planned",
                }
            ],
            "leave_subtask_rows": [
                {
                    "issue_key": "RLT-172",
                    "assignee": "Alice",
                    "start_date": "2026-01-01",
                    "due_date": "2026-03-31",
                    "original_estimate_hours": 528,
                    "total_worklog_hours": 528,
                    "leave_classification": "Planned",
                }
            ],
        }
        html = build_nested_html(payload)
        self.assertIn("const leaveDistributedRows = Array.isArray(reportData.leave_distributed_rows)", html)
        self.assertIn("function computeDistributedLeaveMetricsForRange(bounds)", html)
        self.assertIn("? distributedLeaveMetrics", html)
        self.assertIn("Subtasks_Distributed buckets", html)
        self.assertIn("Sum(bucketed RLT leave original estimates)", html)

    def test_nested_view_actual_rollup_uses_subtask_leaves_only(self):
        payload = {
            "generated_at": "2026-02-21 00:00 UTC",
            "source_file": "nested view.xlsx",
            "rows": [],
            "capacity_profiles": [],
            "leave_daily_rows": [],
            "leave_subtask_rows": [],
        }
        html = build_nested_html(payload)
        self.assertIn('const LEAF_WORK_ROW_TYPES = new Set(["subtask"]);', html)
        self.assertIn("function hasLeafWorkChildren(row)", html)
        self.assertIn("function sumPlannedLeafHours(parentId)", html)
        self.assertIn("row.planned_hours = plannedHours;", html)
        self.assertNotIn('const WORK_ROW_TYPES = new Set(["subtask", "story"]);', html)

    def test_nested_date_filter_uses_active_selection_bounds(self):
        payload = {
            "generated_at": "2026-02-21 00:00 UTC",
            "source_file": "nested view.xlsx",
            "rows": [],
            "capacity_profiles": [],
            "leave_daily_rows": [],
            "leave_subtask_rows": [],
        }
        html = build_nested_html(payload)
        self.assertIn(
            "const bounds = getDateFilterBoundsFor(activeSelection.dateFrom, activeSelection.dateTo);",
            html,
        )
        self.assertIn("function matchesDateFilter(row, selection)", html)
        self.assertIn("const activeSelection = buildScorecardSelectionSnapshot(selection);", html)

    def test_rnd_story_controls_exist(self):
        payload = {
            "department_name": "Research and Development (RnD)",
            "generated_at": "2026-02-21 00:00 UTC",
            "source_files": {},
            "defaults": {"from_date": "2026-02-01", "to_date": "2026-02-28"},
            "epics": [],
            "epic_logged_hours_by_key": {},
            "worklog_rows": [],
            "capacity_profiles": [],
            "leave_daily_rows": [],
        }
        html = build_rnd_story_html(payload)
        self.assertIn('id="from-date"', html)
        self.assertIn('id="to-date"', html)
        self.assertIn('id="capacity-profile-select"', html)
        self.assertIn('id="apply-profile-btn"', html)
        self.assertIn('id="actual-hours-mode"', html)
        self.assertIn('id="kpi-capacity-after-leaves"', html)
        self.assertIn('id="kpi-hours-required-projects"', html)
        self.assertIn("funnel-hours-required-track", html)
        self.assertIn("funnel-hours-required-val", html)
        self.assertIn("/api/capacity?from=", html)
        self.assertIn("/api/actual-hours/aggregate", html)
        self.assertIn("/api/scoped-subtasks", html)
        self.assertIn("/api/manage-fields?include_inactive=0", html)
        self.assertIn("evaluateManagedField", html)
        self.assertIn("managedFieldFormulaText", html)

    def test_ipp_meeting_drawer_scopes_mini_gantt_to_linked_phases(self):
        html = (Path(__file__).resolve().parents[1] / "ipp_meeting_dashboard_template.html").read_text(encoding="utf-8")
        self.assertIn("function renderMiniGantt(row, visiblePhaseNames)", html)
        self.assertIn("const linkedPhaseNames = mappedPhaseRows.map((item) => item.phase);", html)
        self.assertIn("${renderMiniGantt(row, linkedPhaseNames)}", html)
        self.assertIn("No linked phases available for this RMI.", html)

    def test_planned_rmis_actual_mode_controls_exist(self):
        payload = {
            "rows": [],
            "generated_at": "2026-02-24 00:00 UTC",
            "source_file": "nested view.xlsx",
            "default_from": "2026-02-01",
            "default_to": "2026-02-28",
        }
        html = build_planned_rmis_html(payload)
        self.assertIn('id=\'actual-hours-mode\'', html)
        self.assertIn('id=\'actual-hours-status\'', html)
        self.assertIn("/api/actual-hours/aggregate", html)

    def test_team_rmi_gantt_contains_team_lanes_and_clickable_epic_links(self):
        payload = {
            "generated_at": "2026-03-02 00:00 UTC",
            "source_file": "1_jira_work_items_export.xlsx",
            "team_names": ["Technical Writing", "Unmapped Team"],
            "items": [
                {
                    "team_name": "Technical Writing",
                    "epic_key": "P1-100",
                    "epic_name": "Epic Alpha",
                    "epic_url": "https://jira.example/browse/P1-100",
                    "epic_status": "In Progress",
                    "project_key": "P1",
                    "planned_start": "2026-02-01",
                    "planned_end": "2026-02-20",
                    "planned_hours": 24.0,
                    "planned_man_days": 3.0,
                    "story_count": 2,
                    "is_unmapped_team": 0,
                    "snapshot_utc": "2026-03-02 00:00:00",
                }
            ],
            "snapshot_meta": {
                "snapshot_utc": "2026-03-02 00:00:00",
                "source_work_items_path": "1_jira_work_items_export.xlsx",
                "total_story_rows": 6,
                "included_story_rows": 3,
                "excluded_missing_epic": 1,
                "excluded_missing_dates": 1,
                "excluded_missing_estimate": 1,
            },
        }
        html = build_team_rmi_gantt_html(payload)
        self.assertIn("Team Owner RMI Gantt", html)
        self.assertIn("Technical Writing", html)
        self.assertIn("Unmapped Team", html)
        self.assertIn("Cards open Jira epic links", html)
        self.assertIn("target=\"_blank\"", html)
        self.assertIn("team_names", html)
        self.assertIn("epic_status", html)
        self.assertIn("status-pill", html)
        self.assertIn("function statusStyle(status)", html)
        self.assertIn("In Progress", html)
        self.assertIn('id="shortcut-this-year"', html)
        self.assertIn('id="shortcut-this-month"', html)
        self.assertIn('id="shortcut-previous-month"', html)
        self.assertIn('id="shortcut-this-quarter"', html)
        self.assertIn('id="shortcut-this-week"', html)
        self.assertIn('id="shortcut-last-week"', html)
        self.assertIn('id="shift-range-back"', html)
        self.assertIn('id="shift-range-forward"', html)
        self.assertIn('data-range-preset="this-year"', html)
        self.assertIn('data-range-preset="last-week"', html)
        self.assertIn("function resolvePresetRange(presetKey)", html)
        self.assertIn("function setActiveQuickFilter(presetKey)", html)
        self.assertIn("function shiftCurrentRangeByMonths(monthOffset)", html)
        self.assertIn("function addMonths(d, months)", html)
        self.assertIn('shiftRangeBackButton.addEventListener("click"', html)
        self.assertIn('shiftRangeForwardButton.addEventListener("click"', html)
        self.assertIn("quickFilterButtons.forEach((button) => {", html)

    def test_employee_performance_controls_exist(self):
        payload = {
            "worklogs": [],
            "leave_rows": [],
            "projects": [],
            "default_from": "2026-02-01",
            "default_to": "2026-02-28",
            "settings": {
                "base_score": 100,
                "min_score": 0,
                "max_score": 100,
                "points_per_bug_hour": 0.5,
                "points_per_bug_late_hour": 1.5,
                "points_per_unplanned_leave_hour": 0.75,
                "points_per_subtask_late_hour": 1.0,
                "points_per_estimate_overrun_hour": 1.25,
            },
            "generated_at": "2026-02-22 00:00 UTC",
        }
        html = build_employee_perf_html(payload)
        self.assertIn('id="from"', html)
        self.assertIn('id="to"', html)
        self.assertIn('id="projects"', html)
        self.assertIn('id="leaderboard"', html)
        self.assertIn("/settings/performance", html)
        self.assertIn('id="shortcut-current-month"', html)
        self.assertIn('id="shortcut-previous-month"', html)
        self.assertIn('id="shortcut-last-30-days"', html)
        self.assertIn('id="shortcut-quarter-to-date"', html)
        self.assertIn('id="assignee-extended-actuals-toggle"', html)
        self.assertIn('data-score-drawer-accordion="rules"', html)
        self.assertIn('id="score-drawer-rules-content" class="score-drawer-section-content" hidden', html)
        self.assertIn('aria-expanded="false"', html)
        self.assertIn('id="score-subtask-epic-filter"', html)
        self.assertIn('id="score-subtask-project-filter"', html)
        self.assertIn('id="score-subtask-table-body"', html)
        self.assertIn("Actual Completed Date", html)
        self.assertIn("due-status-pill", html)
        self.assertIn('id="assignee-detail-refresh-btn"', html)
        self.assertIn("/api/employee-performance/assignee-refresh", html)

    def test_report_entities_formula_editor_controls_exist(self):
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            (root / "report_html").mkdir(parents=True, exist_ok=True)
            (root / "report_html" / "dashboard.html").write_text("<html><body>ok</body></html>", encoding="utf-8")
            wb = Workbook()
            ws = wb.active
            ws.append(["project_key", "worklog_date", "period_day", "period_week", "period_month", "issue_assignee", "hours_logged"])
            ws.append(["O2", "2026-02-01", "2026-02-01", "2026-W05", "2026-02", "Alice", 1.0])
            wb.save(root / "assignee_hours_report.xlsx")
            app = create_report_server_app(base_dir=root, folder_raw="report_html")
            client = app.test_client()
            resp = client.get("/settings/report-entities")
            self.assertEqual(resp.status_code, 200)
            html = resp.get_data(as_text=True)
            self.assertIn('id="e-formula-expression"', html)
            self.assertIn('id="formula-suggestions"', html)
            self.assertIn('id="formula-validation"', html)
            self.assertIn('id="formula-quick-insert"', html)

    def test_canonical_refresh_settings_formats_timestamps_for_display(self):
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            (root / "report_html").mkdir(parents=True, exist_ok=True)
            (root / "report_html" / "dashboard.html").write_text("<html><body>ok</body></html>", encoding="utf-8")
            app = create_report_server_app(base_dir=root, folder_raw="report_html")
            client = app.test_client()

            resp = client.get("/settings/canonical-refresh")
            self.assertEqual(resp.status_code, 200)
            html = resp.get_data(as_text=True)

            self.assertIn("formatFriendlyTimestamp", html)
            self.assertIn("formatTimestampHtml", html)
            self.assertIn(".timestamp-display", html)
            self.assertIn('metaStarted.innerHTML = formatTimestampHtml(item.started_at_utc);', html)
            self.assertIn('title="UTC: ${esc(raw)}"', html)
            self.assertIn('id="create-db-backup" type="checkbox"', html)
            self.assertNotIn('id="create-db-backup" type="checkbox" checked', html)
            self.assertIn("create_db_backup: createDbBackup", html)
            self.assertIn('id="meta-db-backup"', html)

    def test_manage_fields_page_and_settings_links_exist(self):
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            (root / "report_html").mkdir(parents=True, exist_ok=True)
            (root / "report_html" / "dashboard.html").write_text("<html><body>ok</body></html>", encoding="utf-8")
            wb = Workbook()
            ws = wb.active
            ws.append(["project_key", "worklog_date", "period_day", "period_week", "period_month", "issue_assignee", "hours_logged"])
            ws.append(["O2", "2026-02-01", "2026-02-01", "2026-W05", "2026-02", "Alice", 1.0])
            wb.save(root / "assignee_hours_report.xlsx")
            app = create_report_server_app(base_dir=root, folder_raw="report_html")
            client = app.test_client()

            manage = client.get("/settings/manage-fields")
            self.assertEqual(manage.status_code, 200)
            manage_html = manage.get_data(as_text=True)
            self.assertIn("Manage Fields", manage_html)
            self.assertIn('id="mf-formula-expression"', manage_html)
            self.assertIn('id="mf-formula-suggestions"', manage_html)
            self.assertIn('id="mf-formula-quick-insert"', manage_html)
            self.assertIn('id="mf-formula-validation"', manage_html)
            self.assertIn('id="mf-field-key" class="mono" readonly', manage_html)
            self.assertIn("Auto-generated from Label", manage_html)
            self.assertIn("updateFormulaMetaFromReferences", manage_html)
            self.assertIn("meta.references", manage_html)
            self.assertIn("/api/manage-fields", manage_html)

            capacity_html = client.get("/settings/capacity").get_data(as_text=True)
            self.assertNotIn('href="/settings/manage-fields"', capacity_html)
            perf_html = client.get("/settings/performance").get_data(as_text=True)
            self.assertNotIn('href="/settings/manage-fields"', perf_html)
            self.assertIn('id="team-unassigned-list"', perf_html)
            self.assertIn("Assignees Not in Any Team", perf_html)
            entities_html = client.get("/settings/report-entities").get_data(as_text=True)
            self.assertNotIn('href="/settings/manage-fields"', entities_html)
            self.assertNotIn('href="/settings/projects"', entities_html)
            self.assertNotIn('href="/settings/epics-dropdown-options"', entities_html)
            self.assertNotIn('href="/settings/epic-phases"', entities_html)
            self.assertNotIn('href="/settings/epics-management"', entities_html)

            dropdowns = client.get("/settings/epics-dropdown-options")
            self.assertEqual(dropdowns.status_code, 200)
            dropdowns_html = dropdowns.get_data(as_text=True)
            self.assertIn("Epic Dropdown Options", dropdowns_html)
            self.assertIn('id="product-category-options"', dropdowns_html)
            self.assertIn('id="component-options"', dropdowns_html)
            self.assertIn("/api/epics-management/dropdown-options", dropdowns_html)

            projects = client.get("/settings/projects")
            self.assertEqual(projects.status_code, 200)
            projects_html = projects.get_data(as_text=True)
            self.assertIn("Managed Projects", projects_html)
            self.assertIn('id="jira-search"', projects_html)
            self.assertIn('id="jira-search-results"', projects_html)
            self.assertIn('id="project-key"', projects_html)
            self.assertIn('id="project-name"', projects_html)
            self.assertIn('id="display-name"', projects_html)
            self.assertIn('id="color-hex"', projects_html)
            self.assertIn("/api/projects", projects_html)
            self.assertIn("/api/jira/projects/search", projects_html)

            epic_phases = client.get("/settings/epic-phases")
            self.assertEqual(epic_phases.status_code, 200)
            epic_phases_html = epic_phases.get_data(as_text=True)
            self.assertIn("Manage Epic Phases", epic_phases_html)
            self.assertIn("Epic Plan Columns are managed here as Epic Phases", epic_phases_html)
            self.assertIn('id="phase-name"', epic_phases_html)
            self.assertIn('id="phase-position"', epic_phases_html)
            self.assertIn('id="phase-jira-enabled"', epic_phases_html)
            self.assertIn('id="add-phase-btn"', epic_phases_html)
            self.assertIn("Phase Role", epic_phases_html)
            self.assertIn("Formula", epic_phases_html)
            self.assertIn("Lock", epic_phases_html)
            self.assertIn("data-phase-lock-input", epic_phases_html)
            self.assertIn("Formula-managed", epic_phases_html)
            self.assertIn('id="tab-active"', epic_phases_html)
            self.assertIn('id="tab-deleted"', epic_phases_html)
            self.assertIn('id="phases-tbody"', epic_phases_html)
            self.assertIn("/api/epics-management/plan-columns", epic_phases_html)
            self.assertIn('data-phase-label-input', epic_phases_html)
            self.assertIn('data-phase-original-label', epic_phases_html)
            self.assertIn('data-phase-jira-input', epic_phases_html)
            self.assertIn("savePhaseNameCell", epic_phases_html)
            self.assertIn("savePhaseToggleCell", epic_phases_html)
            self.assertIn("Press Enter or leave the field to save", epic_phases_html)
            self.assertIn("phaseJiraEnabledEl.checked = true;", epic_phases_html)
            self.assertIn("/api/epics-management/plan-columns/order", epic_phases_html)
            self.assertIn("/restore", epic_phases_html)

            epics = client.get("/settings/epics-management")
            self.assertEqual(epics.status_code, 200)
            epics_html = epics.get_data(as_text=True)
            self.assertIn("Epics Planner", epics_html)
            self.assertIn("Quick add epic", epics_html)
            self.assertIn("<kbd>Shift</kbd>", epics_html)
            self.assertIn('id="epics-tbody"', epics_html)
            self.assertIn('id="executive-summary-tbody"', epics_html)
            self.assertIn("Total Planned", epics_html)
            self.assertIn("Unplanned", epics_html)
            self.assertIn("Onhold", epics_html)
            self.assertIn('class="executive-summary-table"', epics_html)
            self.assertIn("renderExecutiveSummary", epics_html)
            self.assertIn("click any epic row to open its phase matrix", epics_html)
            self.assertIn('id="seal-epics-btn"', epics_html)
            self.assertIn('id="populate-jira-btn"', epics_html)
            self.assertIn('id="populate-jira-modal"', epics_html)
            self.assertIn('id="populate-jira-confirm"', epics_html)
            self.assertIn('id="populate-jira-history-btn"', epics_html)
            self.assertIn('id="populate-jira-report"', epics_html)
            self.assertIn('id="populate-jira-history"', epics_html)
            self.assertIn("Retry Failed Work Items", epics_html)
            self.assertIn('/api/epics-management/populate-jira/preflight', epics_html)
            self.assertIn('/api/epics-management/populate-jira/execute', epics_html)
            self.assertIn('/api/epics-management/populate-jira/reports', epics_html)
            self.assertIn("Manage sealed budgets", epics_html)
            self.assertIn("User Manual Plan", epics_html)
            self.assertIn("Most Likely", epics_html)
            self.assertIn("TK Budgeted", epics_html)
            self.assertIn("TK Approved", epics_html)
            self.assertIn("Plan Overview", epics_html)
            self.assertIn('data-accordion-row="1"', epics_html)
            self.assertIn("epic-accordion-toggle", epics_html)
            self.assertIn("epic-overview-split", epics_html)
            self.assertIn('title="Most Likely"', epics_html)
            self.assertIn('title="TK Approved"', epics_html)
            self.assertIn('aria-label="Epic phase matrix"', epics_html)
            self.assertIn("<th>Phase</th><th>Most likely</th><th>TK Budgeted</th>", epics_html)
            self.assertNotIn("Click the epic row to view the phase matrix.", epics_html)
            self.assertNotIn("Most Likely input", epics_html)
            self.assertIn("plan-layer-most-likely", epics_html)
            self.assertIn("plan-layer-tk-budgeted", epics_html)
            self.assertIn("plan-pair-start", epics_html)
            self.assertIn("plan-pair-end", epics_html)
            self.assertIn("--most-likely-bg:#ffedd5", epics_html)
            self.assertIn("--tk-budgeted-bg:#dcfce7", epics_html)
            self.assertIn('id="plan-dialog"', epics_html)
            self.assertIn('id="plan-mandays-wrap"', epics_html)
            self.assertIn('id="add-epic-btn"', epics_html)
            self.assertIn('id="add-plan-column-btn"', epics_html)
            self.assertIn('id="manage-plan-columns-btn"', epics_html)
            self.assertIn('id="epic-dialog"', epics_html)
            self.assertIn('id="epic-project-select"', epics_html)
            self.assertIn("tree-level-category", epics_html)
            self.assertIn("tree-level-component", epics_html)
            self.assertIn("tree-level-epic", epics_html)
            self.assertNotIn('id="epic-product-category"', epics_html)
            self.assertNotIn('id="epic-component"', epics_html)
            self.assertIn("Jira URL (optional)", epics_html)
            self.assertIn('epic_key: resolvedEpicKey || ""', epics_html)
            self.assertNotIn("Jira URL must include an epic key like /browse/O2-1234.", epics_html)
            self.assertIn("/api/epics-management/dropdown-options", epics_html)
            self.assertIn("/api/epics-management/plan-columns", epics_html)
            self.assertIn("/api/epics-management/plan-columns/order", epics_html)
            self.assertIn("/restore", epics_html)
            self.assertIn('id="manage-plan-columns-btn"', epics_html)
            self.assertIn('id="plan-column-dialog"', epics_html)
            self.assertIn('id="manage-columns-dialog"', epics_html)
            self.assertIn('id="plan-column-name"', epics_html)
            self.assertIn('id="plan-column-position"', epics_html)
            self.assertIn('id="plan-column-jira-enabled"', epics_html)
            self.assertIn('id="plan-column-restore-hint"', epics_html)
            self.assertIn('id="manage-columns-dialog"', epics_html)
            self.assertIn("data-sync-epic-row", epics_html)
            self.assertIn('id="sync-jira-modal"', epics_html)
            self.assertIn('id="sync-epic-mandays"', epics_html)
            self.assertIn('id="sync-phase-mandays"', epics_html)
            self.assertIn('id="sync-epic-dates"', epics_html)
            self.assertIn('id="sync-phase-dates"', epics_html)
            self.assertIn('id="epic-research-urs-plan-jira-url"', epics_html)
            self.assertIn('id="epic-dds-plan-jira-url"', epics_html)
            self.assertIn('id="epic-development-plan-jira-url"', epics_html)
            self.assertIn('id="epic-sqa-plan-jira-url"', epics_html)
            self.assertIn('id="epic-user-manual-plan-jira-url"', epics_html)
            self.assertIn('id="epic-production-plan-jira-url"', epics_html)
            self.assertIn('id="dynamic-plan-fields"', epics_html)
            self.assertIn('DATE_ONLY_FORMULA_PLAN_KEYS', epics_html)
            self.assertIn('Edit Planned Dates for', epics_html)
            self.assertIn("hintEl.textContent = 'No Jira issues are marked as created by Populate Jira", epics_html)
            self.assertIn('only "Planner only" applies until then.\';', epics_html)

    def test_page_categories_page_contains_report_display_name_controls(self):
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            (root / "report_html").mkdir(parents=True, exist_ok=True)
            (root / "report_html" / "dashboard.html").write_text("<html><body>ok</body></html>", encoding="utf-8")
            wb = Workbook()
            ws = wb.active
            ws.append(["project_key", "worklog_date", "period_day", "period_week", "period_month", "issue_assignee", "hours_logged"])
            ws.append(["O2", "2026-02-01", "2026-02-01", "2026-W05", "2026-02", "Alice", 1.0])
            wb.save(root / "assignee_hours_report.xlsx")
            app = create_report_server_app(base_dir=root, folder_raw="report_html")
            client = app.test_client()

            resp = client.get("/settings/page-categories")
            self.assertEqual(resp.status_code, 200)
            html = resp.get_data(as_text=True)
            self.assertIn("Report display names are editable here; slugs stay fixed.", html)
            self.assertIn("page_overrides", html)
            self.assertIn("data-page-display-name", html)

    def test_epics_management_create_and_update_api(self):
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            (root / "report_html").mkdir(parents=True, exist_ok=True)
            (root / "report_html" / "dashboard.html").write_text("<html><body>ok</body></html>", encoding="utf-8")
            wb = Workbook()
            ws = wb.active
            ws.append(["project_key", "worklog_date", "period_day", "period_week", "period_month", "issue_assignee", "hours_logged"])
            ws.append(["O2", "2026-02-01", "2026-02-01", "2026-W05", "2026-02", "Alice", 1.0])
            wb.save(root / "assignee_hours_report.xlsx")
            app = create_report_server_app(base_dir=root, folder_raw="report_html")
            client = app.test_client()

            columns_resp = client.get("/api/epics-management/plan-columns")
            self.assertEqual(columns_resp.status_code, 200)
            columns_body = columns_resp.get_json()
            seeded = {str(item.get("key")) for item in columns_body.get("columns", [])}
            self.assertIn("epic_plan", seeded)
            self.assertIn("production_plan", seeded)
            self.assertIn("process_design", seeded)
            self.assertIn("regression_sqa_testing", seeded)
            research_col = next(item for item in columns_body.get("columns", []) if str(item.get("key")) == "research_urs_plan")
            self.assertEqual(research_col.get("label"), "R/URS")
            self.assertTrue(research_col.get("most_likely_enabled"))
            self.assertTrue(research_col.get("tk_budgeted_enabled"))
            self.assertTrue(research_col.get("is_locked"))
            self.assertEqual(research_col.get("formula_role"), "percentage_if_input")
            handover_col = next(item for item in columns_body.get("columns", []) if str(item.get("key")) == "qa_handover")
            self.assertEqual(handover_col.get("phase_role"), "formula_managed")
            self.assertFalse(handover_col.get("most_likely_enabled"))
            self.assertTrue(handover_col.get("jira_link_enabled"))
            bug_fixing_col = next(item for item in columns_body.get("columns", []) if str(item.get("key")) == "bug_fixing")
            self.assertTrue(bug_fixing_col.get("jira_link_enabled"))
            production_col = next(item for item in columns_body.get("columns", []) if str(item.get("key")) == "production_plan")
            self.assertTrue(production_col.get("jira_link_enabled"))

            rename_default_resp = client.put(
                "/api/epics-management/plan-columns/research_urs_plan",
                json={"label": "Discovery URS", "jira_link_enabled": True},
            )
            self.assertEqual(rename_default_resp.status_code, 200)
            columns_after_default_rename = client.get("/api/epics-management/plan-columns")
            self.assertEqual(columns_after_default_rename.status_code, 200)
            renamed_default = next(
                item
                for item in columns_after_default_rename.get_json().get("columns", [])
                if str(item.get("key")) == "research_urs_plan"
            )
            self.assertEqual(renamed_default.get("label"), "Discovery URS")
            restore_default_name_resp = client.put(
                "/api/epics-management/plan-columns/research_urs_plan",
                json={"label": "R/URS", "jira_link_enabled": True, "is_locked": True},
            )
            self.assertEqual(restore_default_name_resp.status_code, 200)

            add_column_resp = client.post(
                "/api/epics-management/plan-columns",
                json={"label": "Security Plan", "jira_link_enabled": True, "insert_position": 2},
            )
            self.assertEqual(add_column_resp.status_code, 201)
            add_column_body = add_column_resp.get_json()
            security_key = add_column_body["column"]["key"]
            self.assertTrue(security_key.startswith("security_plan"))

            rename_column_resp = client.put(
                f"/api/epics-management/plan-columns/{security_key}",
                json={"label": "Security Review Plan"},
            )
            self.assertEqual(rename_column_resp.status_code, 200)
            renamed_column = (rename_column_resp.get_json() or {}).get("column") or {}
            self.assertEqual(str(renamed_column.get("key")), security_key)
            self.assertEqual(str(renamed_column.get("label")), "Security Review Plan")

            reordered_resp = client.put(
                "/api/epics-management/plan-columns/order",
                json={
                    "ordered_keys": [
                        "research_urs_plan",
                        "epic_plan",
                        security_key,
                        "process_design",
                        "dds_plan",
                        "development_plan",
                        "qa_handover",
                        "sqa_plan",
                        "bug_fixing",
                        "process_qa_testing",
                        "user_manual_plan",
                        "regression_sqa_testing",
                        "production_plan",
                    ]
                },
            )
            self.assertEqual(reordered_resp.status_code, 200)
            reordered_body = reordered_resp.get_json()
            reordered_keys = [str(item.get("key")) for item in reordered_body.get("columns", [])]
            self.assertEqual(
                reordered_keys,
                [
                    "research_urs_plan",
                    "epic_plan",
                    security_key,
                    "process_design",
                    "dds_plan",
                    "development_plan",
                    "qa_handover",
                    "sqa_plan",
                    "bug_fixing",
                    "process_qa_testing",
                    "user_manual_plan",
                    "regression_sqa_testing",
                    "production_plan",
                ],
            )

            delete_default_resp = client.delete("/api/epics-management/plan-columns/epic_plan")
            self.assertEqual(delete_default_resp.status_code, 400)

            create_resp = client.post(
                "/api/epics-management/rows",
                json={
                    "epic_key": "O2-999",
                    "project_key": "O2",
                    "project_name": "O2 Project",
                    "product_category": "Core",
                    "epic_name": "Payments Revamp",
                    "description": "Initial epic",
                    "originator": "Lead A",
                    "priority": "High",
                    "ipp_meeting_planned": "Yes",
                    "jira_url": "https://jira.example.com/browse/O2-999",
                    "plans": {
                        "epic_plan": {"man_days": 8, "start_date": "2026-02-01", "due_date": "2026-02-10"},
                        "research_urs_plan": {
                            "man_days": 2,
                            "start_date": "2026-02-01",
                            "due_date": "2026-02-03",
                            "jira_url": "https://jira.example.com/browse/O2-1101",
                        },
                        "user_manual_plan": {
                            "jira_url": "https://jira.example.com/browse/O2-1199",
                        },
                        security_key: {
                            "jira_url": "https://jira.example.com/browse/O2-1205",
                            "man_days": 1.5,
                            "start_date": "2026-02-05",
                            "due_date": "2026-02-06",
                        },
                    },
                },
            )
            self.assertEqual(create_resp.status_code, 201)
            create_body = create_resp.get_json()
            created_row_id = create_body["row"]["id"]
            self.assertTrue(created_row_id, "Row must have an id")
            self.assertEqual(create_body["row"]["epic_key"], "O2-999")
            self.assertEqual(create_body["row"]["priority"], "High")
            self.assertEqual(create_body["row"]["ipp_meeting_planned"], "Yes")
            self.assertEqual(
                create_body["row"]["plans"]["research_urs_plan"]["jira_url"],
                "https://jira.example.com/browse/O2-1101",
            )
            self.assertEqual(
                create_body["row"]["plans"]["user_manual_plan"]["jira_url"],
                "https://jira.example.com/browse/O2-1199",
            )
            self.assertEqual(
                create_body["row"]["plans"][security_key]["jira_url"],
                "https://jira.example.com/browse/O2-1205",
            )
            self.assertEqual(create_body["row"]["plans"]["research_urs_plan"]["most_likely_man_days"], 2.0)
            self.assertEqual(create_body["row"]["plans"]["research_urs_plan"]["tk_budgeted_man_days"], 0.08)
            self.assertEqual(create_body["row"]["plans"]["research_urs_plan"]["man_days"], 0.08)
            self.assertEqual(create_body["row"]["plans"]["research_urs_plan"]["tk_budgeted_start_date"], "2026-02-01")
            self.assertEqual(create_body["row"]["plans"]["research_urs_plan"]["tk_budgeted_due_date"], "2026-02-03")
            self.assertEqual(create_body["row"]["plans"][security_key]["most_likely_man_days"], 1.5)
            self.assertEqual(create_body["row"]["plans"][security_key]["man_days"], 1.5)
            self.assertEqual(create_body["row"]["plans"]["epic_plan"]["most_likely_man_days"], 3.5)
            self.assertEqual(create_body["row"]["plans"]["epic_plan"]["optimistic_man_days"], 1.75)
            self.assertEqual(create_body["row"]["plans"]["epic_plan"]["pessimistic_man_days"], 3.85)
            self.assertEqual(create_body["row"]["plans"]["epic_plan"]["calculated_man_days"], 3.27)
            self.assertEqual(create_body["row"]["plans"]["epic_plan"]["tk_approved_man_days"], 1.64)

            update_resp = client.put(
                f"/api/epics-management/rows/{created_row_id}",
                json={
                    "description": "Updated epic",
                    "priority": "Highest",
                    "ipp_meeting_planned": "No",
                    "plans": {
                        "epic_plan": {"man_days": 10, "start_date": "2026-02-01", "due_date": "2026-02-12"},
                        "dds_plan": {"jira_url": "https://jira.example.com/browse/O2-1201"},
                        security_key: {"man_days": 2, "start_date": "2026-02-07", "due_date": "2026-02-09"},
                    },
                },
            )
            self.assertEqual(update_resp.status_code, 200)
            update_body = update_resp.get_json()
            self.assertEqual(update_body["row"]["description"], "Updated epic")
            self.assertEqual(update_body["row"]["priority"], "Highest")
            self.assertEqual(update_body["row"]["ipp_meeting_planned"], "No")
            self.assertEqual(update_body["row"]["plans"]["epic_plan"]["most_likely_man_days"], 4.0)
            self.assertEqual(update_body["row"]["plans"]["epic_plan"]["tk_approved_man_days"], 1.87)
            self.assertEqual(
                update_body["row"]["plans"]["dds_plan"]["jira_url"],
                "https://jira.example.com/browse/O2-1201",
            )
            self.assertEqual(update_body["row"]["plans"][security_key]["most_likely_man_days"], 2.0)
            self.assertEqual(update_body["row"]["plans"][security_key]["man_days"], 2.0)

            formula_dates_resp = client.post(
                "/api/epics-management/rows",
                json={
                    "epic_key": "O2-1001",
                    "project_key": "O2",
                    "project_name": "O2 Project",
                    "product_category": "Core",
                    "epic_name": "Formula Phase Dates",
                    "plans": {
                        "development_plan": {
                            "man_days": 4,
                            "start_date": "2026-03-01",
                            "due_date": "2026-03-05",
                        },
                        "qa_handover": {
                            "start_date": "2026-03-06",
                            "due_date": "2026-03-06",
                        },
                        "bug_fixing": {
                            "start_date": "2026-03-07",
                            "due_date": "2026-03-08",
                        },
                        "production_plan": {
                            "start_date": "2026-03-09",
                            "due_date": "2026-03-09",
                        },
                    },
                },
            )
            self.assertEqual(formula_dates_resp.status_code, 201)
            formula_dates_body = formula_dates_resp.get_json()
            formula_row = formula_dates_body["row"]
            self.assertEqual(formula_row["plans"]["qa_handover"]["start_date"], "2026-03-06")
            self.assertEqual(formula_row["plans"]["qa_handover"]["due_date"], "2026-03-06")
            self.assertEqual(formula_row["plans"]["qa_handover"]["tk_budgeted_start_date"], "2026-03-06")
            self.assertEqual(formula_row["plans"]["qa_handover"]["tk_budgeted_due_date"], "2026-03-06")
            self.assertEqual(formula_row["plans"]["qa_handover"]["tk_budgeted_man_days"], 0.5)
            self.assertEqual(formula_row["plans"]["qa_handover"]["man_days"], 0.5)
            self.assertEqual(formula_row["plans"]["qa_handover"]["most_likely_man_days"], "")
            self.assertEqual(formula_row["plans"]["bug_fixing"]["start_date"], "2026-03-07")
            self.assertEqual(formula_row["plans"]["bug_fixing"]["due_date"], "2026-03-08")
            self.assertEqual(formula_row["plans"]["bug_fixing"]["tk_budgeted_start_date"], "2026-03-07")
            self.assertEqual(formula_row["plans"]["bug_fixing"]["tk_budgeted_due_date"], "2026-03-08")
            self.assertGreater(formula_row["plans"]["bug_fixing"]["tk_budgeted_man_days"], 0)
            self.assertEqual(formula_row["plans"]["bug_fixing"]["man_days"], formula_row["plans"]["bug_fixing"]["tk_budgeted_man_days"])
            self.assertEqual(formula_row["plans"]["bug_fixing"]["most_likely_man_days"], "")
            self.assertEqual(formula_row["plans"]["production_plan"]["start_date"], "2026-03-09")
            self.assertEqual(formula_row["plans"]["production_plan"]["due_date"], "2026-03-09")
            self.assertEqual(formula_row["plans"]["production_plan"]["tk_budgeted_start_date"], "2026-03-09")
            self.assertEqual(formula_row["plans"]["production_plan"]["tk_budgeted_due_date"], "2026-03-09")
            self.assertEqual(formula_row["plans"]["production_plan"]["tk_budgeted_man_days"], 2.0)
            self.assertEqual(formula_row["plans"]["production_plan"]["man_days"], 2.0)
            self.assertEqual(formula_row["plans"]["production_plan"]["most_likely_man_days"], "")

            planner_columns_after_rename = client.get("/api/epics-management/plan-columns")
            self.assertEqual(planner_columns_after_rename.status_code, 200)
            planner_columns = planner_columns_after_rename.get_json().get("columns", [])
            security_columns = [item for item in planner_columns if str(item.get("key")) == security_key]
            self.assertEqual(len(security_columns), 1)
            self.assertEqual(str(security_columns[0].get("label")), "Security Review Plan")

            delete_dynamic_resp = client.delete(f"/api/epics-management/plan-columns/{security_key}")
            self.assertEqual(delete_dynamic_resp.status_code, 200)
            delete_dynamic_body = delete_dynamic_resp.get_json()
            keys_after_delete = {str(item.get("key")) for item in delete_dynamic_body.get("columns", [])}
            self.assertNotIn(security_key, keys_after_delete)

            restore_resp = client.post(f"/api/epics-management/plan-columns/{security_key}/restore")
            self.assertEqual(restore_resp.status_code, 200)
            restored_column = (restore_resp.get_json() or {}).get("column") or {}
            self.assertEqual(str(restored_column.get("key")), security_key)

            create_default_resp = client.post(
                "/api/epics-management/rows",
                json={
                    "epic_key": "O2-1000",
                    "project_key": "O2",
                    "project_name": "O2 Project",
                    "product_category": "Core",
                    "epic_name": "Default Planner Flag",
                },
            )
            self.assertEqual(create_default_resp.status_code, 201)
            create_default_body = create_default_resp.get_json()
            self.assertEqual(create_default_body["row"]["ipp_meeting_planned"], "No")

            add_delete_candidate_resp = client.post(
                "/api/epics-management/plan-columns",
                json={"label": "Deprecation Plan", "jira_link_enabled": False},
            )
            self.assertEqual(add_delete_candidate_resp.status_code, 201)
            delete_candidate_key = str(add_delete_candidate_resp.get_json()["column"]["key"])

            delete_column_resp = client.delete(f"/api/epics-management/plan-columns/{delete_candidate_key}")
            self.assertEqual(delete_column_resp.status_code, 200)
            delete_column_body = delete_column_resp.get_json()
            remaining_keys = [str(item.get("key")) for item in delete_column_body.get("columns", [])]
            self.assertNotIn(delete_candidate_key, remaining_keys)

            delete_default_resp = client.delete("/api/epics-management/plan-columns/epic_plan")
            self.assertEqual(delete_default_resp.status_code, 400)

            default_jira_resp = client.post(
                "/api/epics-management/plan-columns",
                json={"label": "Auto Jira Default Plan"},
            )
            self.assertEqual(default_jira_resp.status_code, 201)
            default_jira_column = (default_jira_resp.get_json() or {}).get("column") or {}
            self.assertTrue(bool(default_jira_column.get("jira_link_enabled")))

            locked_dynamic_resp = client.put(
                f"/api/epics-management/plan-columns/{delete_candidate_key}",
                json={"label": "Deprecation Plan", "is_locked": True},
            )
            self.assertEqual(locked_dynamic_resp.status_code, 200)
            self.assertTrue(bool((locked_dynamic_resp.get_json() or {}).get("column", {}).get("is_locked")))
            delete_locked_resp = client.delete(f"/api/epics-management/plan-columns/{delete_candidate_key}")
            self.assertEqual(delete_locked_resp.status_code, 400)

            unlock_dynamic_resp = client.put(
                f"/api/epics-management/plan-columns/{delete_candidate_key}",
                json={"label": "Deprecation Plan", "is_locked": False},
            )
            self.assertEqual(unlock_dynamic_resp.status_code, 200)
            self.assertFalse(bool((unlock_dynamic_resp.get_json() or {}).get("column", {}).get("is_locked")))
            delete_unlocked_resp = client.delete(f"/api/epics-management/plan-columns/{delete_candidate_key}")
            self.assertEqual(delete_unlocked_resp.status_code, 200)

            unlock_default_resp = client.put(
                "/api/epics-management/plan-columns/epic_plan",
                json={"label": "Epic Plan", "is_locked": False},
            )
            self.assertEqual(unlock_default_resp.status_code, 200)
            self.assertFalse(bool((unlock_default_resp.get_json() or {}).get("column", {}).get("is_locked")))
            delete_unlocked_default_resp = client.delete("/api/epics-management/plan-columns/epic_plan")
            self.assertEqual(delete_unlocked_default_resp.status_code, 200)
            restore_default_resp = client.post("/api/epics-management/plan-columns/epic_plan/restore")
            self.assertEqual(restore_default_resp.status_code, 200)
            relock_default_resp = client.put(
                "/api/epics-management/plan-columns/epic_plan",
                json={"label": "Epic Plan", "is_locked": True},
            )
            self.assertEqual(relock_default_resp.status_code, 200)
            self.assertTrue(bool((relock_default_resp.get_json() or {}).get("column", {}).get("is_locked")))

    def test_epics_import_preview_parses_work_status_and_rnd_phase_group(self):
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            (root / "report_html").mkdir(parents=True, exist_ok=True)
            (root / "report_html" / "dashboard.html").write_text("<html><body>ok</body></html>", encoding="utf-8")
            _write_minimal_assignee_workbook(root)
            source_path = root / "Epic Estimates Approved Plan.xlsx"
            _write_epics_import_source(source_path, total_override=99)
            app = create_report_server_app(base_dir=root, folder_raw="report_html")
            client = app.test_client()

            with source_path.open("rb") as handle:
                upload_resp = client.post(
                    "/api/epics-management/import/upload",
                    data={"workbook": (handle, source_path.name)},
                    content_type="multipart/form-data",
                )
            self.assertEqual(upload_resp.status_code, 200)
            upload_token = upload_resp.get_json()["upload_token"]
            preview_resp = client.get(f"/api/epics-management/import/preview?fetch_jira=0&upload_token={upload_token}")
            self.assertEqual(preview_resp.status_code, 200)
            body = preview_resp.get_json()
            rows = body["rows"]
            self.assertEqual(len(rows), 2)
            row = rows[0]
            self.assertEqual(row["project_name"], "OmniConnect")
            self.assertEqual(row["jira_url"], "https://octopusdtlsupport.atlassian.net/browse/O2-321")
            self.assertEqual(row["epic_key"], "O2-321")
            self.assertEqual(row["work_status"], "Ready")
            self.assertEqual(row["category"], "Input")
            self.assertEqual(row["component"], "Streaming")
            self.assertEqual(row["phases"]["process_design"], 1)
            self.assertEqual(row["phases"]["research_urs_plan"], 2)
            self.assertEqual(row["phases"]["production_plan"], 9)
            self.assertEqual(row["phase_sum"], 45.0)
            self.assertEqual(row["man_days_total"], 99)
            self.assertFalse(row["total_matches"])
            self.assertIn("Phase total 45.0 differs from Man Days 99", row["warnings"][0])
            self.assertFalse(rows[1]["can_import"])

            page_resp = client.get("/settings/epics-management/import")
            self.assertEqual(page_resp.status_code, 200)
            page_html = page_resp.get_data(as_text=True)
            self.assertIn("Epics Planner Import", page_html)
            self.assertIn("/api/epics-management/import/upload", page_html)
            self.assertIn("/api/epics-management/import/preview", page_html)
            self.assertIn("Work:", page_html)

    @patch("report_server._fetch_jira_issues_for_jql")
    @patch("report_server.resolve_jira_end_date_field_ids")
    @patch("report_server.resolve_jira_start_date_field_id")
    @patch("report_server.get_session")
    def test_epics_import_preview_fetches_jira_and_suggests_phase_links(
        self,
        mock_get_session,
        mock_start_field,
        mock_end_fields,
        mock_fetch_jql,
    ):
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            (root / "report_html").mkdir(parents=True, exist_ok=True)
            (root / "report_html" / "dashboard.html").write_text("<html><body>ok</body></html>", encoding="utf-8")
            _write_minimal_assignee_workbook(root)
            source_path = root / "Epic Estimates Approved Plan.xlsx"
            _write_epics_import_source(source_path)
            app = create_report_server_app(base_dir=root, folder_raw="report_html")
            client = app.test_client()

            mock_get_session.return_value = object()
            mock_start_field.return_value = "customfield_start"
            mock_end_fields.return_value = ["customfield_end"]
            mock_fetch_jql.side_effect = [
                [
                    {
                        "key": "O2-321",
                        "fields": {
                            "summary": "Jira Epic Summary",
                            "description": {"content": [{"content": [{"text": "Jira description"}]}]},
                            "issuetype": {"name": "Epic"},
                            "status": {"name": "In Progress"},
                            "customfield_start": "2026-05-01",
                            "customfield_end": "2026-05-31",
                            "timeoriginalestimate": 28800,
                        },
                    }
                ],
                [
                    {
                        "key": "O2-401",
                        "fields": {
                            "summary": "Development build story",
                            "issuetype": {"name": "Story"},
                            "status": {"name": "To Do"},
                            "customfield_10014": "O2-321",
                            "customfield_start": "2026-05-03",
                            "customfield_end": "2026-05-10",
                            "timeoriginalestimate": 57600,
                        },
                    },
                    {
                        "key": "O2-402",
                        "fields": {
                            "summary": "SQA validation story",
                            "issuetype": {"name": "Story"},
                            "status": {"name": "To Do"},
                            "customfield_10014": "O2-321",
                            "customfield_start": "2026-05-11",
                            "customfield_end": "2026-05-12",
                            "timeoriginalestimate": 14400,
                        },
                    },
                ],
                [],
            ]

            with source_path.open("rb") as handle:
                upload_resp = client.post(
                    "/api/epics-management/import/upload",
                    data={"workbook": (handle, source_path.name)},
                    content_type="multipart/form-data",
                )
            self.assertEqual(upload_resp.status_code, 200)
            upload_token = upload_resp.get_json()["upload_token"]
            preview_resp = client.get(f"/api/epics-management/import/preview?upload_token={upload_token}")
            self.assertEqual(preview_resp.status_code, 200)
            row = preview_resp.get_json()["rows"][0]
            self.assertEqual(row["epic_name"], "Jira Epic Summary")
            self.assertIn("Jira description", row["description"])
            dev_suggestion = row["phase_suggestions"]["development_plan"]
            self.assertTrue(dev_suggestion["accepted"])
            self.assertEqual(dev_suggestion["issue_key"], "O2-401")
            self.assertEqual(dev_suggestion["jira_url"], "https://octopusdtlsupport.atlassian.net/browse/O2-401")
            self.assertEqual(dev_suggestion["start_date"], "2026-05-03")

    def test_epics_import_submit_backs_up_rebudgets_and_writes_rows(self):
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            (root / "report_html").mkdir(parents=True, exist_ok=True)
            (root / "report_html" / "dashboard.html").write_text("<html><body>ok</body></html>", encoding="utf-8")
            _write_minimal_assignee_workbook(root)
            app = create_report_server_app(base_dir=root, folder_raw="report_html")
            client = app.test_client()

            create_existing = client.post(
                "/api/epics-management/rows",
                json={
                    "epic_key": "O2-321",
                    "project_key": "O2",
                    "project_name": "Old Project",
                    "product_category": "Old",
                    "component": "Old Component",
                    "epic_name": "Old Epic",
                    "priority": "Low",
                    "plan_status": "Not Planned Yet",
                    "jira_url": "https://octopusdtlsupport.atlassian.net/browse/O2-321",
                    "plans": {
                        "sqa_plan": {
                            "most_likely_man_days": 1,
                            "man_days": 1,
                            "jira_url": "https://octopusdtlsupport.atlassian.net/browse/O2-990",
                        }
                    },
                },
            )
            self.assertEqual(create_existing.status_code, 201)
            seal_resp = client.post("/api/epics-management/seal", json={"epic_keys": ["O2-321"]})
            self.assertEqual(seal_resp.status_code, 200)
            source_path = root / "Epic Estimates Approved Plan.xlsx"
            _write_epics_import_source(source_path)
            with source_path.open("rb") as handle:
                upload_resp = client.post(
                    "/api/epics-management/import/upload",
                    data={"workbook": (handle, source_path.name)},
                    content_type="multipart/form-data",
                )
            self.assertEqual(upload_resp.status_code, 200)
            upload_token = upload_resp.get_json()["upload_token"]

            submit_resp = client.post(
                "/api/epics-management/import/submit",
                json={
                    "upload_token": upload_token,
                    "rows": [
                        {
                            "include": True,
                            "source_id": "OmniConnect RMI:3",
                            "epic_key": "O2-321",
                            "project_key": "O2",
                            "project_name": "OmniConnect",
                            "category": "Input",
                            "component": "Streaming",
                            "jira_url": "https://octopusdtlsupport.atlassian.net/browse/O2-321",
                            "epic_name": "Jira Epic Summary",
                            "description": "Jira description",
                            "originator": "RnD",
                            "phases": {
                                "process_design": 1,
                                "research_urs_plan": 2,
                                "dds_plan": 3,
                                "development_plan": 4,
                                "sqa_plan": 5,
                                "process_qa_testing": 6,
                                "user_manual_plan": 7,
                                "regression_sqa_testing": 8,
                                "production_plan": 9,
                            },
                            "phase_reviews": {
                                "development_plan": {
                                    "accepted": True,
                                    "jira_url": "https://octopusdtlsupport.atlassian.net/browse/O2-401",
                                    "start_date": "2026-05-03",
                                    "due_date": "2026-05-10",
                                },
                                "sqa_plan": {
                                    "accepted": False,
                                    "jira_url": "https://octopusdtlsupport.atlassian.net/browse/O2-402",
                                    "start_date": "2026-05-11",
                                    "due_date": "2026-05-12",
                                },
                            },
                        },
                        {
                            "include": True,
                            "source_id": "Fintech Fuel RMI:22",
                            "epic_key": "FF-541",
                            "project_key": "FF",
                            "project_name": "Fintech Fuel",
                            "category": "Administration",
                            "component": "URM",
                            "jira_url": "https://octopusdtlsupport.atlassian.net/browse/FF-541",
                            "epic_name": "New FF Epic",
                            "description": "New description",
                            "originator": "PGL",
                            "phases": {"development_plan": 6, "sqa_plan": 3},
                            "phase_reviews": {},
                        },
                    ]
                },
            )
            self.assertEqual(submit_resp.status_code, 200)
            body = submit_resp.get_json()
            self.assertEqual(body["updated"], 1)
            self.assertEqual(body["inserted"], 1)
            self.assertEqual(body["rebudgeted"], 1)
            self.assertTrue(Path(body["backup_path"]).exists())

            rows_resp = client.get("/api/epics-management/rows")
            self.assertEqual(rows_resp.status_code, 200)
            rows = {row["epic_key"]: row for row in rows_resp.get_json()["rows"]}
            updated = rows["O2-321"]
            self.assertEqual(updated["project_name"], "OmniConnect")
            self.assertEqual(updated["product_category"], "Input")
            self.assertEqual(updated["component"], "Streaming")
            self.assertEqual(updated["originator"], "RnD")
            self.assertEqual(updated["priority"], "High")
            self.assertEqual(updated["plan_status"], "Planned")
            self.assertEqual(updated["is_sealed"], 0)
            self.assertEqual(updated["is_tk_epic"], 1)
            self.assertEqual(updated["plans"]["development_plan"]["most_likely_man_days"], 4.0)
            self.assertEqual(updated["plans"]["development_plan"]["jira_url"], "https://octopusdtlsupport.atlassian.net/browse/O2-401")
            self.assertEqual(updated["plans"]["development_plan"]["start_date"], "2026-05-03")
            self.assertEqual(updated["plans"]["sqa_plan"]["most_likely_man_days"], 5.0)
            self.assertEqual(updated["plans"]["sqa_plan"]["jira_url"], "https://octopusdtlsupport.atlassian.net/browse/O2-990")
            self.assertIn("FF-541", rows)
            self.assertEqual(rows["FF-541"]["is_tk_epic"], 1)

    def test_epics_import_preview_and_submit_require_upload_token(self):
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            (root / "report_html").mkdir(parents=True, exist_ok=True)
            (root / "report_html" / "dashboard.html").write_text("<html><body>ok</body></html>", encoding="utf-8")
            _write_minimal_assignee_workbook(root)
            app = create_report_server_app(base_dir=root, folder_raw="report_html")
            client = app.test_client()

            preview_resp = client.get("/api/epics-management/import/preview?fetch_jira=0")
            self.assertEqual(preview_resp.status_code, 400)
            self.assertIn("upload_token", (preview_resp.get_json() or {}).get("error", ""))

            submit_resp = client.post("/api/epics-management/import/submit", json={"rows": []})
            self.assertEqual(submit_resp.status_code, 400)
            self.assertIn("upload_token", (submit_resp.get_json() or {}).get("error", ""))

    def test_epics_management_tk_flag_toggle_api(self):
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            (root / "report_html").mkdir(parents=True, exist_ok=True)
            (root / "report_html" / "dashboard.html").write_text("<html><body>ok</body></html>", encoding="utf-8")
            _write_minimal_assignee_workbook(root)
            app = create_report_server_app(base_dir=root, folder_raw="report_html")
            client = app.test_client()
            create_resp = client.post(
                "/api/epics-management/rows",
                json={
                    "epic_key": "O2-777",
                    "project_key": "O2",
                    "project_name": "OmniConnect",
                    "product_category": "Input",
                    "component": "Streaming",
                    "epic_name": "TK Toggle Epic",
                    "priority": "Low",
                    "plan_status": "Planned",
                    "jira_url": "https://octopusdtlsupport.atlassian.net/browse/O2-777",
                    "plans": {},
                },
            )
            self.assertEqual(create_resp.status_code, 201)

            mark_resp = client.post("/api/epics-management/rows/O2-777/tk-flag", json={"is_tk_epic": True})
            self.assertEqual(mark_resp.status_code, 200)
            self.assertEqual((mark_resp.get_json() or {}).get("row", {}).get("is_tk_epic"), 1)

            unmark_resp = client.post("/api/epics-management/rows/O2-777/tk-flag", json={"is_tk_epic": False})
            self.assertEqual(unmark_resp.status_code, 200)
            self.assertEqual((unmark_resp.get_json() or {}).get("row", {}).get("is_tk_epic"), 0)

    def test_epics_management_seal_blocks_updates_until_rebudget(self):
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            (root / "report_html").mkdir(parents=True, exist_ok=True)
            (root / "report_html" / "dashboard.html").write_text("<html><body>ok</body></html>", encoding="utf-8")
            wb = Workbook()
            ws = wb.active
            ws.append(["project_key", "worklog_date", "period_day", "period_week", "period_month", "issue_assignee", "hours_logged"])
            ws.append(["O2", "2026-02-01", "2026-02-01", "2026-W05", "2026-02", "Alice", 1.0])
            wb.save(root / "assignee_hours_report.xlsx")
            app = create_report_server_app(base_dir=root, folder_raw="report_html")
            client = app.test_client()

            create_resp = client.post(
                "/api/epics-management/rows",
                json={
                    "epic_key": "O2-555",
                    "project_key": "O2",
                    "project_name": "O2 Project",
                    "product_category": "Core",
                    "epic_name": "Seal Protected Epic",
                    "jira_url": "https://jira.example.com/browse/O2-555",
                    "plans": {
                        "development_plan": {
                            "man_days": 8,
                            "start_date": "2026-02-01",
                            "due_date": "2026-02-10",
                        }
                    },
                },
            )
            self.assertEqual(create_resp.status_code, 201)

            seal_resp = client.post("/api/epics-management/seal", json={"epic_keys": ["O2-555"]})
            self.assertEqual(seal_resp.status_code, 200)
            seal_body = seal_resp.get_json() or {}
            self.assertEqual(seal_body.get("sealed_count"), 1)

            rows_after_seal = client.get("/api/epics-management/rows").get_json()["rows"]
            sealed_row = next(item for item in rows_after_seal if item["epic_key"] == "O2-555")
            self.assertEqual(sealed_row["is_sealed"], 1)
            self.assertEqual(sealed_row["plans"]["development_plan"]["most_likely_man_days"], 8.0)
            self.assertEqual(sealed_row["plans"]["development_plan"]["tk_budgeted_start_date"], "2026-02-01")
            self.assertEqual(sealed_row["plans"]["epic_plan"]["tk_approved_man_days"], 3.74)

            update_resp = client.put("/api/epics-management/rows/O2-555", json={"description": "Blocked"})
            self.assertEqual(update_resp.status_code, 423)

            delete_resp = client.delete("/api/epics-management/rows/O2-555")
            self.assertEqual(delete_resp.status_code, 423)

            sync_resp = client.post(
                "/api/epics-management/rows/O2-555/sync-jira-plan",
                json={"jira_url": "https://jira.example.com/browse/O2-555"},
            )
            self.assertEqual(sync_resp.status_code, 423)

            rebudget_resp = client.post("/api/epics-management/rows/O2-555/re-budget")
            self.assertEqual(rebudget_resp.status_code, 200)

            update_after_rebudget = client.put("/api/epics-management/rows/O2-555", json={"description": "Unlocked"})
            self.assertEqual(update_after_rebudget.status_code, 200)
            self.assertEqual(update_after_rebudget.get_json()["row"]["description"], "Unlocked")

            delete_after_rebudget = client.delete("/api/epics-management/rows/O2-555")
            self.assertEqual(delete_after_rebudget.status_code, 200)

    @patch("report_server._jira_write_permissions_summary")
    def test_epics_management_populate_jira_preflight_returns_duplicate_warning_and_month_split_subtasks(
        self,
        mock_permissions,
    ):
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            (root / "report_html").mkdir(parents=True, exist_ok=True)
            (root / "report_html" / "dashboard.html").write_text("<html><body>ok</body></html>", encoding="utf-8")
            _write_minimal_assignee_workbook(root)
            app = create_report_server_app(base_dir=root, folder_raw="report_html")
            client = app.test_client()

            mock_permissions.return_value = {
                "ok": True,
                "project_key": "O2",
                "permissions": {},
                "missing_permissions": [],
            }

            create_resp = client.post(
                "/api/epics-management/rows",
                json={
                    "epic_key": "O2-880",
                    "project_key": "O2",
                    "project_name": "OmniConnect",
                    "product_category": "Input",
                    "component": "Streaming",
                    "epic_name": "Publish Preview Epic",
                    "description": "Preview me",
                    "plans": {
                        "epic_plan": {"man_days": 12, "start_date": "2026-01-10", "due_date": "2026-03-20"},
                        "development_plan": {"man_days": 9, "start_date": "2026-01-15", "due_date": "2026-03-10"},
                    },
                },
            )
            self.assertEqual(create_resp.status_code, 201)

            db_path = root / "assignee_hours_capacity.db"
            conn = sqlite3.connect(db_path)
            try:
                conn.execute(
                    """
                    INSERT INTO canonical_refresh_state(id, active_run_id, last_success_run_id, updated_at_utc)
                    VALUES (1, ?, ?, ?)
                    ON CONFLICT(id) DO UPDATE SET
                        active_run_id=excluded.active_run_id,
                        last_success_run_id=excluded.last_success_run_id,
                        updated_at_utc=excluded.updated_at_utc
                    """,
                    ("run-preview", "run-preview", "2026-04-30T00:00:00+00:00"),
                )
                conn.execute(
                    """
                    INSERT INTO canonical_issues(
                        run_id, issue_id, issue_key, project_key, issue_type, summary, status,
                        assignee, start_date, due_date, created_utc, updated_utc,
                        resolved_stable_since_date, original_estimate_hours, total_hours_logged,
                        fix_type, parent_issue_key, story_key, epic_key, raw_payload_json
                    )
                    VALUES (?, '', ?, ?, ?, ?, ?, '', '', '', '', '', '', 0, 0, '', '', '', '', '{}')
                    """,
                    ("run-preview", "O2-998", "O2", "Epic", "Publish Preview Epic", "In Progress"),
                )
                conn.commit()
            finally:
                conn.close()

            preview_resp = client.post(
                "/api/epics-management/populate-jira/preflight",
                json={"epic_keys": ["O2-880"]},
            )
            self.assertEqual(preview_resp.status_code, 200)
            body = preview_resp.get_json() or {}
            self.assertTrue(body.get("can_execute"))
            self.assertEqual(body.get("source"), "epics_management_populate_jira")

            epics = body.get("epics") or []
            self.assertEqual(len(epics), 1)
            item = epics[0]
            self.assertTrue((item.get("duplicate_warning") or {}).get("found"))
            self.assertEqual((item.get("permission") or {}).get("project_key"), "O2")

            stories = item.get("stories") or []
            self.assertGreaterEqual(len(stories), 1)
            story = next(entry for entry in stories if entry.get("phase_key") == "development_plan")
            self.assertEqual(story.get("phase_key"), "development_plan")
            subtasks = story.get("subtasks") or []
            self.assertEqual(len(subtasks), 3)
            self.assertEqual(subtasks[0].get("month_label"), "Jan effort")
            self.assertEqual(subtasks[0].get("start_date"), "2026-01-15")
            self.assertEqual(subtasks[0].get("due_date"), "2026-01-31")
            self.assertEqual(subtasks[1].get("month_label"), "Feb effort")
            self.assertEqual(subtasks[2].get("month_label"), "Mar effort")

    @patch("report_server._jira_create_issue")
    @patch("report_server._resolve_jira_field_id_by_name")
    @patch("report_server._resolve_jira_publish_session")
    @patch("report_server.resolve_jira_end_date_field_ids")
    @patch("report_server.resolve_jira_start_date_field_id")
    def test_epics_management_populate_jira_execute_persists_links_for_sealed_epic(
        self,
        mock_start_field,
        mock_end_fields,
        mock_resolve_publish_session,
        mock_resolve_field_id,
        mock_create_issue,
    ):
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            (root / "report_html").mkdir(parents=True, exist_ok=True)
            (root / "report_html" / "dashboard.html").write_text("<html><body>ok</body></html>", encoding="utf-8")
            _write_minimal_assignee_workbook(root)
            app = create_report_server_app(base_dir=root, folder_raw="report_html")
            client = app.test_client()

            mock_session = object()
            mock_start_field.return_value = "customfield_start"
            mock_end_fields.return_value = ["customfield_end"]
            mock_resolve_publish_session.return_value = (
                mock_session,
                {
                    "ok": True,
                    "project_key": "O2",
                    "permissions": {},
                    "missing_permissions": [],
                    "token_source": "default",
                },
            )
            mock_resolve_field_id.side_effect = lambda _session, field_name: {
                "Epic Name": "customfield_epic_name",
                "Epic Link": "customfield_10014",
            }.get(field_name, "")
            issued_keys: list[str] = []

            def _fake_create_issue(_session, _fields):
                issue_key = f"O2-90{len(issued_keys) + 1}"
                issued_keys.append(issue_key)
                return {"issue_key": issue_key, "jira_url": f"https://jira.example.com/browse/{issue_key}"}

            mock_create_issue.side_effect = _fake_create_issue

            create_resp = client.post(
                "/api/epics-management/rows",
                json={
                    "epic_key": "O2-556",
                    "project_key": "O2",
                    "project_name": "OmniConnect",
                    "product_category": "Input",
                    "component": "Streaming",
                    "epic_name": "Sealed Publish Epic",
                    "plans": {
                        "epic_plan": {"man_days": 6, "start_date": "2026-02-01", "due_date": "2026-02-12"},
                        "research_urs_plan": {"man_days": 4, "start_date": "2026-02-02", "due_date": "2026-02-10"},
                    },
                },
            )
            self.assertEqual(create_resp.status_code, 201)

            seal_resp = client.post("/api/epics-management/seal", json={"epic_keys": ["O2-556"]})
            self.assertEqual(seal_resp.status_code, 200)

            execute_resp = client.post(
                "/api/epics-management/populate-jira/execute",
                json={"epics": [{"epic_key": "O2-556", "mode": "create", "allow_duplicate": True}]},
            )
            self.assertEqual(execute_resp.status_code, 200)
            body = execute_resp.get_json() or {}
            self.assertEqual(body.get("source"), "epics_management_populate_jira")
            results = body.get("results") or []
            self.assertEqual(len(results), 1)
            result = results[0]
            self.assertEqual(result.get("epic_jira_key"), "O2-901")
            self.assertEqual(result.get("story_count"), 3)

            rows_resp = client.get("/api/epics-management/rows")
            self.assertEqual(rows_resp.status_code, 200)
            row = next(item for item in (rows_resp.get_json() or {}).get("rows", []) if item.get("epic_key") == "O2-556")
            self.assertEqual(row.get("is_sealed"), 1)
            self.assertEqual(row.get("epr_jira_epic_created"), 1)
            self.assertEqual(row.get("epr_created_jira_issue_count"), 4)
            self.assertEqual(row.get("jira_url"), "https://jira.example.com/browse/O2-901")
            self.assertEqual(
                ((row.get("plans") or {}).get("research_urs_plan") or {}).get("jira_url"),
                "https://jira.example.com/browse/O2-902",
            )
            self.assertEqual(
                ((row.get("plans") or {}).get("production_plan") or {}).get("jira_url"),
                "https://jira.example.com/browse/O2-904",
            )

            db_path = root / "assignee_hours_capacity.db"
            conn = sqlite3.connect(db_path)
            try:
                publish_rows = conn.execute(
                    "SELECT issue_level, jira_issue_key, jira_url, created_via_epr FROM epics_management_jira_publish WHERE epic_key=? ORDER BY issue_level, phase_key",
                    ("O2-556",),
                ).fetchall()
                self.assertEqual(len(publish_rows), 4)
                self.assertTrue(all(int(pr[3] or 0) == 1 for pr in publish_rows))
                epic_records = [row for row in publish_rows if row[0] == "epic"]
                story_records = [row for row in publish_rows if row[0] == "story"]
                self.assertEqual(len(epic_records), 1)
                self.assertEqual(epic_records[0][1], "O2-901")
                self.assertEqual({row[1] for row in story_records}, {"O2-902", "O2-903", "O2-904"})
            finally:
                conn.close()

    @patch("report_server._jira_update_issue")
    @patch("report_server._jira_create_issue")
    @patch("report_server._resolve_jira_field_id_by_name")
    @patch("report_server._resolve_jira_publish_session")
    @patch("report_server.resolve_jira_end_date_field_ids")
    @patch("report_server.resolve_jira_start_date_field_id")
    def test_epics_management_populate_jira_report_records_failure_history_and_retry(
        self,
        mock_start_field,
        mock_end_fields,
        mock_resolve_publish_session,
        mock_resolve_field_id,
        mock_create_issue,
        mock_update_issue,
    ):
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            (root / "report_html").mkdir(parents=True, exist_ok=True)
            (root / "report_html" / "dashboard.html").write_text("<html><body>ok</body></html>", encoding="utf-8")
            _write_minimal_assignee_workbook(root)
            app = create_report_server_app(base_dir=root, folder_raw="report_html")
            client = app.test_client()

            mock_session = object()
            mock_start_field.return_value = "customfield_start"
            mock_end_fields.return_value = ["customfield_end"]
            mock_resolve_publish_session.return_value = (
                mock_session,
                {
                    "ok": True,
                    "project_key": "O2",
                    "permissions": {},
                    "missing_permissions": [],
                    "token_source": "default",
                },
            )
            mock_resolve_field_id.side_effect = lambda _session, field_name: {
                "Epic Name": "customfield_epic_name",
                "Epic Link": "customfield_10014",
            }.get(field_name, "")
            mock_update_issue.side_effect = lambda _session, issue_key, _fields: {
                "issue_key": issue_key,
                "jira_url": f"https://jira.example.com/browse/{issue_key}",
            }
            issued_keys: list[str] = []
            failed_dev_once = {"value": False}

            def _fake_create_issue(_session, fields):
                issue_type = (fields.get("issuetype") or {}).get("name")
                summary = fields.get("summary")
                if issue_type == "Story" and summary == "Dev" and not failed_dev_once["value"]:
                    failed_dev_once["value"] = True
                    raise ValueError("Jira create issue failed (400): summary: Dev blocked")
                issue_key = f"O2-92{len(issued_keys) + 1}"
                issued_keys.append(issue_key)
                return {"issue_key": issue_key, "jira_url": f"https://jira.example.com/browse/{issue_key}"}

            mock_create_issue.side_effect = _fake_create_issue

            create_resp = client.post(
                "/api/epics-management/rows",
                json={
                    "epic_key": "O2-557",
                    "project_key": "O2",
                    "project_name": "OmniConnect",
                    "product_category": "Input",
                    "component": "Streaming",
                    "epic_name": "Failure Report Epic",
                    "plans": {
                        "epic_plan": {"man_days": 6, "start_date": "2026-02-01", "due_date": "2026-02-12"},
                        "development_plan": {"man_days": 4, "start_date": "2026-02-02", "due_date": "2026-02-10"},
                    },
                },
            )
            self.assertEqual(create_resp.status_code, 201)

            execute_resp = client.post(
                "/api/epics-management/populate-jira/execute",
                json={"epics": [{"epic_key": "O2-557", "mode": "create", "allow_duplicate": True}]},
            )
            self.assertEqual(execute_resp.status_code, 200)
            body = execute_resp.get_json() or {}
            report = body.get("report") or {}
            self.assertEqual(report.get("status"), "partial_failed")
            self.assertGreater((report.get("summary") or {}).get("failed", 0), 0)
            failed_story = next(
                item
                for item in (report.get("items") or [])
                if item.get("issue_level") == "story" and item.get("phase_key") == "development_plan"
            )
            self.assertEqual(failed_story.get("status"), "failed")
            self.assertTrue(failed_story.get("can_retry"))
            self.assertIn("Dev blocked", failed_story.get("error", ""))

            history_resp = client.get("/api/epics-management/populate-jira/reports")
            self.assertEqual(history_resp.status_code, 200)
            history = (history_resp.get_json() or {}).get("reports") or []
            self.assertEqual(history[0].get("report_id"), report.get("report_id"))

            detail_resp = client.get(f"/api/epics-management/populate-jira/reports/{report.get('report_id')}")
            self.assertEqual(detail_resp.status_code, 200)
            detail_report = (detail_resp.get_json() or {}).get("report") or {}
            self.assertEqual(detail_report.get("report_id"), report.get("report_id"))

            retry_resp = client.post(f"/api/epics-management/populate-jira/reports/{report.get('report_id')}/retry")
            self.assertEqual(retry_resp.status_code, 200)
            retry_report = (retry_resp.get_json() or {}).get("report") or {}
            self.assertEqual(retry_report.get("status"), "completed")
            retry_request = ((retry_report.get("request") or {}).get("epics") or [])[0]
            self.assertEqual(retry_request.get("mode"), "update")
            self.assertEqual(retry_request.get("phase_keys"), ["development_plan"])
            self.assertEqual(retry_report.get("request", {}).get("retry_of_report_id"), report.get("report_id"))

    def test_epics_management_delete_row_clears_jira_publish(self):
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            (root / "report_html").mkdir(parents=True, exist_ok=True)
            (root / "report_html" / "dashboard.html").write_text("<html><body>ok</body></html>", encoding="utf-8")
            _write_minimal_assignee_workbook(root)
            app = create_report_server_app(base_dir=root, folder_raw="report_html")
            client = app.test_client()

            create_resp = client.post(
                "/api/epics-management/rows",
                json={
                    "epic_key": "O2-888",
                    "project_key": "O2",
                    "project_name": "OmniConnect",
                    "product_category": "Input",
                    "component": "Del",
                    "epic_name": "Delete publish cleanup",
                },
            )
            self.assertEqual(create_resp.status_code, 201)
            row = (create_resp.get_json() or {}).get("row") or {}
            row_id = str(row.get("id") or "").strip()
            self.assertTrue(row_id)

            db_path = root / "assignee_hours_capacity.db"
            conn = sqlite3.connect(db_path)
            try:
                conn.execute(
                    """
                    INSERT INTO epics_management_jira_publish (
                      epic_row_id, epic_key, phase_key, issue_level, jira_issue_key, jira_url,
                      parent_jira_key, month_label, man_days, start_date, due_date,
                      published_at_utc, updated_at_utc, created_via_epr
                    ) VALUES (?, ?, '', 'epic', 'O2-800', '', '', '', 0, '', '', '', '', 1)
                    """,
                    (row_id, "O2-888"),
                )
                conn.commit()
            finally:
                conn.close()

            del_resp = client.delete("/api/epics-management/rows/O2-888", json={"delete_jira": False})
            self.assertEqual(del_resp.status_code, 200)
            body = del_resp.get_json() or {}
            self.assertEqual(body.get("deleted"), True)
            self.assertEqual(body.get("epic_key"), "O2-888")
            self.assertEqual(body.get("jira_issues_deleted"), [])

            conn2 = sqlite3.connect(db_path)
            try:
                n = conn2.execute(
                    "SELECT COUNT(*) FROM epics_management_jira_publish WHERE epic_row_id=?",
                    (row_id,),
                ).fetchone()[0]
                self.assertEqual(int(n), 0)
                left = conn2.execute("SELECT COUNT(*) FROM epics_management WHERE id=?", (row_id,)).fetchone()[0]
                self.assertEqual(int(left), 0)
            finally:
                conn2.close()

    @patch("report_server._jira_delete_issue")
    @patch("report_server._resolve_jira_issue_delete_session")
    def test_epics_management_delete_with_jira_only_created_keys_in_order(
        self, mock_resolve_delete_session, mock_jira_delete
    ):
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            (root / "report_html").mkdir(parents=True, exist_ok=True)
            (root / "report_html" / "dashboard.html").write_text("<html><body>ok</body></html>", encoding="utf-8")
            _write_minimal_assignee_workbook(root)
            app = create_report_server_app(base_dir=root, folder_raw="report_html")
            client = app.test_client()

            mock_resolve_delete_session.return_value = (
                object(),
                {"ok": True, "missing_permissions": []},
            )
            mock_jira_delete.return_value = None

            create_resp = client.post(
                "/api/epics-management/rows",
                json={
                    "epic_key": "O2-887",
                    "project_key": "O2",
                    "project_name": "OmniConnect",
                    "product_category": "Input",
                    "component": "JD",
                    "epic_name": "Jira delete order",
                },
            )
            self.assertEqual(create_resp.status_code, 201)
            row = (create_resp.get_json() or {}).get("row") or {}
            row_id = str(row.get("id") or "").strip()

            db_path = root / "assignee_hours_capacity.db"
            now = "2026-04-01T12:00:00Z"
            conn = sqlite3.connect(db_path)
            try:
                # Insert epic first (lower id): ordering must still prefer sub-task -> story -> epic.
                conn.execute(
                    """
                    INSERT INTO epics_management_jira_publish (
                      epic_row_id, epic_key, phase_key, issue_level, jira_issue_key, jira_url,
                      parent_jira_key, month_label, man_days, start_date, due_date,
                      published_at_utc, updated_at_utc, created_via_epr
                    ) VALUES (?, ?, '', 'epic', 'O2-E1', '', '', '', 0, '', '', ?, ?, 1)
                    """,
                    (row_id, "O2-887", now, now),
                )
                conn.execute(
                    """
                    INSERT INTO epics_management_jira_publish (
                      epic_row_id, epic_key, phase_key, issue_level, jira_issue_key, jira_url,
                      parent_jira_key, month_label, man_days, start_date, due_date,
                      published_at_utc, updated_at_utc, created_via_epr
                    ) VALUES (?, ?, 'research_urs_plan', 'story', 'O2-S1', '', '', '', 0, '', '', ?, ?, 1)
                    """,
                    (row_id, "O2-887", now, now),
                )
                conn.execute(
                    """
                    INSERT INTO epics_management_jira_publish (
                      epic_row_id, epic_key, phase_key, issue_level, jira_issue_key, jira_url,
                      parent_jira_key, month_label, man_days, start_date, due_date,
                      published_at_utc, updated_at_utc, created_via_epr
                    ) VALUES (?, ?, 'm1', 'subtask', 'O2-ST1', 'https://jira.example/browse/O2-ST1', 'O2-S1', 'Jan', 0, '', '', ?, ?, 1)
                    """,
                    (row_id, "O2-887", now, now),
                )
                conn.execute(
                    """
                    INSERT INTO epics_management_jira_publish (
                      epic_row_id, epic_key, phase_key, issue_level, jira_issue_key, jira_url,
                      parent_jira_key, month_label, man_days, start_date, due_date,
                      published_at_utc, updated_at_utc, created_via_epr
                    ) VALUES (?, ?, '', 'story', 'O2-LINKED', '', '', '', 0, '', '', ?, ?, 0)
                    """,
                    (row_id, "O2-887", now, now),
                )
                conn.commit()
            finally:
                conn.close()

            del_resp = client.delete("/api/epics-management/rows/O2-887", json={"delete_jira": True})
            self.assertEqual(del_resp.status_code, 200)
            self.assertEqual((del_resp.get_json() or {}).get("jira_issues_deleted"), ["O2-ST1", "O2-S1", "O2-E1"])

            issued = [call.args[1] for call in mock_jira_delete.call_args_list]
            self.assertEqual(issued, ["O2-ST1", "O2-S1", "O2-E1"])
            conn3 = sqlite3.connect(db_path)
            try:
                self.assertEqual(
                    conn3.execute("SELECT COUNT(*) FROM epics_management_jira_publish WHERE epic_row_id=?", (row_id,)).fetchone()[
                        0
                    ],
                    0,
                )
            finally:
                conn3.close()

    def test_jira_pick_select_option_for_yes_no_prefers_id(self):
        from report_server import _jira_pick_select_option_for_yes_no

        allowed = [{"id": "10650-y", "value": "Yes"}, {"id": "10650-n", "value": "No"}]
        self.assertEqual(_jira_pick_select_option_for_yes_no(allowed, "yes"), {"id": "10650-y"})
        self.assertEqual(_jira_pick_select_option_for_yes_no(allowed, "No"), {"id": "10650-n"})

    def test_jira_pick_select_option_matches_planned_labels(self):
        from report_server import _jira_pick_select_option_for_yes_no

        allowed = [{"id": "a", "value": "Planned"}, {"id": "b", "value": "Not planned"}]
        self.assertEqual(_jira_pick_select_option_for_yes_no(allowed, "Yes"), {"id": "a"})
        self.assertEqual(_jira_pick_select_option_for_yes_no(allowed, "No"), {"id": "b"})

    def test_jira_rmi_planned_fields_fallback_value_when_no_createmeta_allowed_values(self):
        from report_server import _jira_rmi_planned_fields_for_issue

        with patch("report_server._resolve_jira_field_id_by_name", return_value="customfield_10650"):
            with patch("report_server._jira_createmeta_issue_fields", return_value={}):
                out = _jira_rmi_planned_fields_for_issue(object(), "WOM", "Epic", "y")
        self.assertEqual(out, {"customfield_10650": {"value": "Yes"}})

    def test_reference_tk_estimates_folder_contains_expected_files(self):
        root = Path(__file__).resolve().parents[1] / "Reference TK Estimates Folder"
        expected = [
            root / "extract_rmi_jira_to_sqlite.py",
            root / "generate_rmi_gantt_html.py",
            root / "run_rmi_pipeline.py",
            root / "populate_ipp_from_jira.py",
            root / "tk_approved_for_ipp.py",
            root / "tests" / "test_extract_rmi_jira_to_sqlite.py",
            root / "tests" / "test_generate_rmi_gantt_html.py",
            root / "tests" / "test_run_rmi_pipeline.py",
            root / "IPP Meeting Reports" / "Epic Estimates Approved Plan.xlsx",
            root / "IPP Meeting Reports" / "Epic Estimates Approved Plan - backup before bug fixing update.xlsx",
            root / "IPP Meeting Reports" / "Epic Estimates Approved Plan - backup before handover update.xlsx",
            root / "IPP Meeting Reports" / "Epic Estimates Approved Plan.codex-copy.xlsx",
            root / "IPP Meeting Reports" / "rmi_jira_extract.db",
            root / "IPP Meeting Reports" / "rmi_jira_gantt.html",
            root / "IPP Meeting Reports" / "rmi_jira_gantt.backup.html",
            root / "IPP Meeting Reports" / "RMI Jira Gantt Report Documentation.md",
        ]
        missing = [str(path.relative_to(root)) for path in expected if not path.is_file()]
        self.assertEqual(missing, [])

    def test_epics_dropdown_options_api(self):
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            (root / "report_html").mkdir(parents=True, exist_ok=True)
            (root / "report_html" / "dashboard.html").write_text("<html><body>ok</body></html>", encoding="utf-8")
            wb = Workbook()
            ws = wb.active
            ws.append(["project_key", "worklog_date", "period_day", "period_week", "period_month", "issue_assignee", "hours_logged"])
            ws.append(["O2", "2026-02-01", "2026-02-01", "2026-W05", "2026-02", "Alice", 1.0])
            wb.save(root / "assignee_hours_report.xlsx")
            app = create_report_server_app(base_dir=root, folder_raw="report_html")
            client = app.test_client()

            get_initial = client.get("/api/epics-management/dropdown-options")
            self.assertEqual(get_initial.status_code, 200)
            self.assertEqual(get_initial.get_json()["product_category_options"], [])
            self.assertEqual(get_initial.get_json()["component_options"], [])
            self.assertEqual(get_initial.get_json()["plan_status_options"], [])

            save_resp = client.put(
                "/api/epics-management/dropdown-options",
                json={
                    "product_category": ["Core", "Payments", "core"],
                    "components": ["Checkout API", "Portal"],
                    "plan_statuses": ["Planned", "Not Planned Yet", "planned"],
                },
            )
            self.assertEqual(save_resp.status_code, 200)
            body = save_resp.get_json()
            self.assertEqual(body["product_category_options"], ["Core", "Payments"])
            self.assertEqual(body["component_options"], ["Checkout API", "Portal"])
            self.assertEqual(body["plan_status_options"], ["Planned", "Not Planned Yet"])

            get_saved = client.get("/api/epics-management/dropdown-options")
            self.assertEqual(get_saved.status_code, 200)
            saved_body = get_saved.get_json()
            self.assertEqual(saved_body["product_category_options"], ["Core", "Payments"])
            self.assertEqual(saved_body["component_options"], ["Checkout API", "Portal"])
            self.assertEqual(saved_body["plan_status_options"], ["Planned", "Not Planned Yet"])

    def test_epics_management_tmp_key_orphan_create_and_jira_key_promotion(self):
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            (root / "report_html").mkdir(parents=True, exist_ok=True)
            (root / "report_html" / "dashboard.html").write_text("<html><body>ok</body></html>", encoding="utf-8")
            wb = Workbook()
            ws = wb.active
            ws.append(["project_key", "worklog_date", "period_day", "period_week", "period_month", "issue_assignee", "hours_logged"])
            ws.append(["O2", "2026-02-01", "2026-02-01", "2026-W05", "2026-02", "Alice", 1.0])
            wb.save(root / "assignee_hours_report.xlsx")
            app = create_report_server_app(base_dir=root, folder_raw="report_html")
            client = app.test_client()

            create_resp = client.post(
                "/api/epics-management/rows",
                json={"epic_name": "Planning Draft Epic"},
            )
            self.assertEqual(create_resp.status_code, 201)
            body = create_resp.get_json()
            tmp_key = str(body["row"]["epic_key"])
            self.assertRegex(tmp_key, r"^TMP-\d{8}T\d{6}Z-[A-Z0-9]{6}$")
            self.assertEqual(body["row"]["project_key"], "ORPHAN")
            self.assertEqual(body["row"]["project_name"], "Orphan")
            self.assertEqual(body["row"]["product_category"], "Orphan")

            update_resp = client.put(
                f"/api/epics-management/rows/{tmp_key}",
                json={"jira_url": "https://jira.example.com/browse/O2-4242"},
            )
            self.assertEqual(update_resp.status_code, 200)
            update_body = update_resp.get_json()
            self.assertEqual(update_body["row"]["epic_key"], "O2-4242")
            self.assertEqual(update_body["row"]["jira_url"], "https://jira.example.com/browse/O2-4242")

            rows_resp = client.get("/api/epics-management/rows")
            self.assertEqual(rows_resp.status_code, 200)
            keys = {str(item.get("epic_key")) for item in rows_resp.get_json().get("rows", [])}
            self.assertIn("O2-4242", keys)
            self.assertNotIn(tmp_key, keys)

    def test_epics_management_tmp_key_conflict_offers_vacant_key_reuse(self):
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            (root / "report_html").mkdir(parents=True, exist_ok=True)
            (root / "report_html" / "dashboard.html").write_text("<html><body>ok</body></html>", encoding="utf-8")
            wb = Workbook()
            ws = wb.active
            ws.append(["project_key", "worklog_date", "period_day", "period_week", "period_month", "issue_assignee", "hours_logged"])
            ws.append(["O2", "2026-02-01", "2026-02-01", "2026-W05", "2026-02", "Alice", 1.0])
            wb.save(root / "assignee_hours_report.xlsx")
            app = create_report_server_app(base_dir=root, folder_raw="report_html")
            client = app.test_client()

            vacant_tmp_key = "TMP-20260101T000000Z-ABC123"
            seed_resp = client.post(
                "/api/epics-management/rows",
                json={
                    "epic_key": vacant_tmp_key,
                    "epic_name": vacant_tmp_key,
                    "project_key": "ORPHAN",
                    "project_name": "Orphan",
                    "product_category": "Orphan",
                    "component": "",
                    "description": "",
                    "originator": "",
                    "jira_url": "",
                    "plans": {},
                },
            )
            self.assertEqual(seed_resp.status_code, 201)

            conflict_resp = client.post(
                "/api/epics-management/rows",
                json={
                    "epic_key": vacant_tmp_key,
                    "epic_name": "New Planned Epic",
                },
            )
            self.assertEqual(conflict_resp.status_code, 409)
            conflict_body = conflict_resp.get_json() or {}
            self.assertEqual(conflict_body.get("code"), "epic_key_exists")
            self.assertEqual(conflict_body.get("vacant_tmp_key"), vacant_tmp_key)
            self.assertTrue(conflict_body.get("can_reuse_vacant_tmp_key"))
            # Error message must not expose backend epic keys to the user
            self.assertNotIn("TMP-", conflict_body.get("error", ""))

            reuse_resp = client.put(
                f"/api/epics-management/rows/{vacant_tmp_key}",
                json={
                    "epic_name": "New Planned Epic",
                    "description": "Saved by reusing vacant TMP key",
                },
            )
            self.assertEqual(reuse_resp.status_code, 200)
            reuse_body = reuse_resp.get_json() or {}
            self.assertEqual(reuse_body.get("row", {}).get("epic_key"), vacant_tmp_key)
            self.assertEqual(reuse_body.get("row", {}).get("epic_name"), "New Planned Epic")

    @patch("report_server._fetch_jira_issues_for_jql")
    @patch("report_server.resolve_jira_end_date_field_ids")
    @patch("report_server.resolve_jira_start_date_field_id")
    @patch("report_server.get_session")
    @patch("report_server.extract_jira_key_from_url")
    def test_epics_management_sync_persists_epic_and_story_rows(
        self,
        mock_extract_key,
        mock_get_session,
        mock_start_field,
        mock_end_fields,
        mock_fetch_jql,
    ):
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            (root / "report_html").mkdir(parents=True, exist_ok=True)
            (root / "report_html" / "dashboard.html").write_text("<html><body>ok</body></html>", encoding="utf-8")
            wb = Workbook()
            ws = wb.active
            ws.append(["project_key", "worklog_date", "period_day", "period_week", "period_month", "issue_assignee", "hours_logged"])
            ws.append(["O2", "2026-02-01", "2026-02-01", "2026-W05", "2026-02", "Alice", 1.0])
            wb.save(root / "assignee_hours_report.xlsx")
            app = create_report_server_app(base_dir=root, folder_raw="report_html")
            client = app.test_client()

            create_resp = client.post(
                "/api/epics-management/rows",
                json={
                    "epic_key": "O2-321",
                    "project_key": "O2",
                    "project_name": "O2 Project",
                    "product_category": "Core",
                    "epic_name": "Existing Epic",
                    "jira_url": "https://jira.example.com/browse/O2-321",
                    "plans": {
                        "research_urs_plan": {"jira_url": "https://jira.example.com/browse/O2-101"},
                        "dds_plan": {"jira_url": "https://jira.example.com/browse/O2-102"},
                    },
                },
            )
            self.assertEqual(create_resp.status_code, 201)

            mock_extract_key.side_effect = lambda url: str(url or "").rstrip("/").split("/")[-1]
            mock_get_session.return_value = object()
            mock_start_field.return_value = "customfield_start"
            mock_end_fields.return_value = ["customfield_end"]
            mock_fetch_jql.side_effect = [
                [
                    {
                        "key": "O2-321",
                        "fields": {
                            "issuetype": {"name": "Epic"},
                            "summary": "Jira Epic Summary",
                            "description": {
                                "type": "doc",
                                "content": [
                                    {
                                        "type": "paragraph",
                                        "content": [{"type": "text", "text": "Epic desc from Jira"}],
                                    }
                                ],
                            },
                            "timeoriginalestimate": 28800,
                            "customfield_start": "2026-02-03",
                            "customfield_end": "2026-02-20",
                        },
                    }
                ],
                [
                    {
                        "key": "O2-101",
                        "fields": {
                            "issuetype": {"name": "Story"},
                            "customfield_10014": "O2-321",
                            "summary": "Story 101",
                            "status": {"name": "In Progress"},
                            "timeoriginalestimate": 14400,
                            "customfield_start": "2026-02-01",
                            "customfield_end": "2026-02-12",
                        },
                    },
                    {
                        "key": "O2-102",
                        "fields": {
                            "issuetype": {"name": "Story"},
                            "customfield_10014": "O2-321",
                            "summary": "Story 102",
                            "status": {"name": "To Do"},
                            "timeoriginalestimate": 7200,
                            "customfield_start": "2026-02-05",
                            "customfield_end": "2026-02-25",
                        },
                    },
                    {
                        "key": "O2-103",
                        "fields": {
                            "issuetype": {"name": "Sub-task"},
                            "customfield_10014": "O2-321",
                            "summary": "Subtask 103",
                            "status": {"name": "Done"},
                            "timeoriginalestimate": 3600,
                            "customfield_start": "2026-02-02",
                            "customfield_end": "2026-02-10",
                        },
                    },
                ],
            ]

            sync_resp = client.post(
                "/api/epics-management/rows/O2-321/sync-jira-plan",
                json={"jira_url": "https://jira.example.com/browse/O2-321"},
            )
            self.assertEqual(sync_resp.status_code, 200)
            body = sync_resp.get_json()
            self.assertEqual(body["synced_story_count"], 2)
            self.assertEqual(body["row"]["epic_name"], "Jira Epic Summary")
            self.assertIn("Epic desc from Jira", body["row"]["description"])
            self.assertEqual(body["row"]["plans"]["research_urs_plan"]["most_likely_man_days"], 0.5)
            self.assertEqual(body["row"]["plans"]["research_urs_plan"]["man_days"], 0.02)
            self.assertEqual(body["row"]["plans"]["research_urs_plan"]["tk_budgeted_man_days"], 0.02)
            self.assertEqual(body["row"]["plans"]["research_urs_plan"]["start_date"], "2026-02-01")
            self.assertEqual(body["row"]["plans"]["research_urs_plan"]["due_date"], "2026-02-12")
            self.assertEqual(body["row"]["plans"]["research_urs_plan"]["tk_budgeted_start_date"], "2026-02-01")
            self.assertEqual(body["row"]["plans"]["research_urs_plan"]["tk_budgeted_due_date"], "2026-02-12")
            self.assertEqual(body["row"]["plans"]["dds_plan"]["most_likely_man_days"], 0.25)
            self.assertEqual(body["row"]["plans"]["dds_plan"]["man_days"], 0.04)
            self.assertEqual(body["row"]["plans"]["dds_plan"]["tk_budgeted_man_days"], 0.04)
            self.assertEqual(body["row"]["plans"]["dds_plan"]["start_date"], "2026-02-05")
            self.assertEqual(body["row"]["plans"]["dds_plan"]["due_date"], "2026-02-25")

            db_path = root / "assignee_hours_capacity.db"
            conn = sqlite3.connect(db_path)
            try:
                row = conn.execute(
                    "SELECT epic_name, description FROM epics_management WHERE epic_key=?",
                    ("O2-321",),
                ).fetchone()
                self.assertIsNotNone(row)
                self.assertEqual(row[0], "Jira Epic Summary")
                self.assertIn("Epic desc from Jira", row[1])

                story_rows = conn.execute(
                    "SELECT story_key, epic_key, story_name, story_status FROM epics_management_story_sync WHERE epic_key=? ORDER BY story_key",
                    ("O2-321",),
                ).fetchall()
                self.assertEqual(len(story_rows), 2)
                self.assertEqual(story_rows[0][0], "O2-101")
                self.assertEqual(story_rows[0][1], "O2-321")
                self.assertEqual(story_rows[0][2], "Story 101")
                self.assertEqual(story_rows[1][0], "O2-102")
            finally:
                conn.close()

    @patch("report_server._fetch_jira_issues_for_jql")
    @patch("report_server.resolve_jira_end_date_field_ids")
    @patch("report_server.resolve_jira_start_date_field_id")
    @patch("report_server.get_session")
    @patch("report_server.extract_jira_key_from_url")
    def test_epics_management_sync_epic_only_scope_preserves_phase_values(
        self,
        mock_extract_key,
        mock_get_session,
        mock_start_field,
        mock_end_fields,
        mock_fetch_jql,
    ):
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            (root / "report_html").mkdir(parents=True, exist_ok=True)
            (root / "report_html" / "dashboard.html").write_text("<html><body>ok</body></html>", encoding="utf-8")
            _write_minimal_assignee_workbook(root)
            app = create_report_server_app(base_dir=root, folder_raw="report_html")
            client = app.test_client()

            create_resp = client.post(
                "/api/epics-management/rows",
                json={
                    "epic_key": "O2-654",
                    "project_key": "O2",
                    "project_name": "O2 Project",
                    "product_category": "Core",
                    "epic_name": "Scoped Sync Epic",
                    "jira_url": "https://jira.example.com/browse/O2-654",
                    "plans": {
                        "research_urs_plan": {
                            "jira_url": "https://jira.example.com/browse/O2-201",
                            "man_days": 2,
                            "start_date": "2026-03-01",
                            "due_date": "2026-03-05",
                        }
                    },
                },
            )
            self.assertEqual(create_resp.status_code, 201)

            no_scope_resp = client.post(
                "/api/epics-management/rows/O2-654/sync-jira-plan",
                json={
                    "jira_url": "https://jira.example.com/browse/O2-654",
                    "sync_epic_mandays": False,
                    "sync_phase_mandays": False,
                    "sync_epic_dates": False,
                    "sync_phase_dates": False,
                },
            )
            self.assertEqual(no_scope_resp.status_code, 400)

            mock_extract_key.side_effect = lambda url: str(url or "").rstrip("/").split("/")[-1]
            mock_get_session.return_value = object()
            mock_start_field.return_value = "customfield_start"
            mock_end_fields.return_value = ["customfield_end"]
            mock_fetch_jql.side_effect = [
                [
                    {
                        "key": "O2-654",
                        "fields": {
                            "issuetype": {"name": "Epic"},
                            "summary": "Scoped Sync Epic from Jira",
                            "description": {
                                "type": "doc",
                                "content": [
                                    {
                                        "type": "paragraph",
                                        "content": [{"type": "text", "text": "Scoped sync desc"}],
                                    }
                                ],
                            },
                            "timeoriginalestimate": 28800,
                            "customfield_start": "2026-04-10",
                            "customfield_end": "2026-04-20",
                        },
                    }
                ],
                [
                    {
                        "key": "O2-201",
                        "fields": {
                            "issuetype": {"name": "Story"},
                            "customfield_10014": "O2-654",
                            "summary": "Story 201",
                            "status": {"name": "In Progress"},
                            "timeoriginalestimate": 14400,
                            "customfield_start": "2026-05-01",
                            "customfield_end": "2026-05-08",
                        },
                    }
                ],
            ]

            sync_resp = client.post(
                "/api/epics-management/rows/O2-654/sync-jira-plan",
                json={
                    "jira_url": "https://jira.example.com/browse/O2-654",
                    "sync_epic_mandays": True,
                    "sync_phase_mandays": False,
                    "sync_epic_dates": True,
                    "sync_phase_dates": False,
                },
            )
            self.assertEqual(sync_resp.status_code, 200)
            body = sync_resp.get_json() or {}
            row = body["row"]
            self.assertEqual(row["epic_name"], "Scoped Sync Epic from Jira")
            self.assertEqual(row["plans"]["research_urs_plan"]["most_likely_man_days"], 2.0)
            self.assertEqual(row["plans"]["research_urs_plan"]["start_date"], "2026-03-01")
            self.assertEqual(row["plans"]["research_urs_plan"]["due_date"], "2026-03-05")
            self.assertEqual(row["plans"]["epic_plan"]["most_likely_man_days"], 1.0)
            self.assertEqual(row["plans"]["epic_plan"]["start_date"], "2026-04-10")
            self.assertEqual(row["plans"]["epic_plan"]["due_date"], "2026-04-20")

            rows_resp = client.get("/api/epics-management/rows")
            self.assertEqual(rows_resp.status_code, 200)
            rows = rows_resp.get_json()["rows"]
            saved_row = next(item for item in rows if item["epic_key"] == "O2-654")
            self.assertEqual(saved_row["plans"]["research_urs_plan"]["most_likely_man_days"], 2.0)
            self.assertEqual(saved_row["plans"]["research_urs_plan"]["start_date"], "2026-03-01")
            self.assertEqual(saved_row["plans"]["research_urs_plan"]["due_date"], "2026-03-05")
            self.assertEqual(saved_row["plans"]["epic_plan"]["most_likely_man_days"], 1.0)
            self.assertEqual(saved_row["plans"]["epic_plan"]["start_date"], "2026-04-10")
            self.assertEqual(saved_row["plans"]["epic_plan"]["due_date"], "2026-04-20")

    def test_dashboard_template_uses_planner_validation_alerts(self):
        template_path = Path(__file__).resolve().parents[1] / "dashboard_template.html"
        html = template_path.read_text(encoding="utf-8")
        self.assertIn("Planner Validation:", html)
        self.assertIn("Planner Dates:", html)
        self.assertIn("Planner Hours:", html)
        self.assertIn("kind === 'story'", html)
        self.assertIn("storyPlannerStartCell", html)
        self.assertIn("storyPlannerEndCell", html)
        self.assertIn("Jira planned dates/hours differ from Epics Planner epic plan.", html)
        self.assertIn("mismatch-planner-btn", html)
        self.assertIn("/settings/epics-management?epic_key=", html)
        self.assertIn("reason=planner_mismatch", html)
        self.assertNotIn("Alert: Jira planned dates differ from IPP meeting dates.", html)

    def test_dashboard_template_releases_drawer_controls_exist(self):
        template_path = Path(__file__).resolve().parents[1] / "dashboard_template.html"
        html = template_path.read_text(encoding="utf-8")
        self.assertIn('id="releases-drawer-open"', html)
        self.assertIn('>Planned Releases</button>', html)
        self.assertIn('id="releases-drawer"', html)
        self.assertIn('id="releases-drawer-overlay"', html)
        self.assertIn('id="releases-drawer-resize-handle"', html)
        self.assertIn('>Planned Releases</h2>', html)
        # New PR calendar controls
        self.assertIn('id="pr-proj-filter"', html)
        self.assertIn('id="pr-prev-yr"', html)
        self.assertIn('id="pr-next-yr"', html)
        self.assertIn('id="pr-btn-cal"', html)
        self.assertIn('id="pr-btn-rm"', html)
        self.assertIn('id="pr-cal-grid"', html)
        self.assertIn('id="pr-list-pane"', html)
        self.assertIn('id="pr-rm-content"', html)
        # New PR JS functions
        self.assertIn("function prRender()", html)
        self.assertIn("async function prLoadData()", html)
        self.assertIn("function renderReleasesDrawerContent()", html)
        self.assertIn("function startReleasesDrawerResize(event)", html)
        # New PR API
        self.assertIn("/api/product-releases", html)
        # New PR CSS
        self.assertIn(".pr-cal-grid {", html)
        self.assertIn(".pr-list-item {", html)
        self.assertIn(".pr-chip {", html)
        # Structural constants still present
        self.assertIn("const EPIC_REFRESH_ENDPOINT_BASE = '/api/dashboard/refresh-epic';", html)
        self.assertIn("epicRefreshRuns", html)
        self.assertIn("await fetch(`${EPIC_REFRESH_ENDPOINT_BASE}/${encodeURIComponent(runId)}`", html)
        self.assertIn("Run scoped colossal refresh for this epic", html)

    def test_dashboard_template_does_not_include_colossal_refresh_launcher(self):
        template_path = Path(__file__).resolve().parents[1] / "dashboard_template.html"
        html_path = Path(__file__).resolve().parents[1] / "report_html" / "dashboard.html"
        template_html = template_path.read_text(encoding="utf-8")
        report_html = html_path.read_text(encoding="utf-8")
        for html in (template_html, report_html):
            self.assertNotIn('id="dashboard-refresh-trigger"', html)
            self.assertNotIn('id="dashboard-refresh-dropdown"', html)
            self.assertNotIn('id="dashboard-refresh-cancel"', html)
            self.assertNotIn('id="dashboard-refresh-progress-wrap"', html)
            self.assertNotIn('id="dashboard-refresh-resume-banner"', html)
            self.assertNotIn('initializeDashboardRefreshPanel()', html)
            self.assertNotIn("startRefresh(mode, resume)", html)

    def test_planned_vs_dispensed_page_controls_exist(self):
        html_path = Path(__file__).resolve().parents[1] / "report_html" / "planned_vs_dispensed_report.html"
        self.assertTrue(html_path.exists())
        html = html_path.read_text(encoding="utf-8")
        self.assertIn('id="date-filter-from"', html)
        self.assertIn('id="date-filter-to"', html)
        self.assertIn('id="date-filter-apply"', html)
        self.assertIn('id="date-filter-reset"', html)
        self.assertIn('id="adv-filter-menu"', html)
        self.assertIn('id="planned-hours-source"', html)
        self.assertIn('id="plan-source"', html)
        self.assertIn('id="projects-trigger"', html)
        self.assertIn('id="projects-menu"', html)
        self.assertIn('id="projects-select-all"', html)
        self.assertIn('id="projects-clear-all"', html)
        self.assertIn('id="projects-options"', html)
        self.assertIn("By Log Date", html)
        self.assertIn("By Planned Date", html)
        self.assertIn("col-resize-handle", html)
        self.assertIn("/api/approved-vs-planned-hours/ui-settings", html)
        self.assertIn("Approved vs Planned Hours Report", html)
        self.assertIn("Total Approved Hours", html)
        self.assertIn("Total Planned Hours", html)
        self.assertIn("ACTUAL HOURS", html)
        self.assertIn('id="pvd-total-actual-hours"', html)
        self.assertIn("Planned Hours (Subtask Original Estimates)", html)
        self.assertIn("Actual Hours (Subtask and Bug Subtask Worklogs)", html)
        self.assertIn("Epic Drill-down", html)
        self.assertNotIn("details.epic", html)
        self.assertNotIn("details.story", html)
        self.assertNotIn('<details class="epic"', html)
        self.assertNotIn('<details class="story"', html)
        self.assertIn('id="pvd-comparison-chart"', html)
        self.assertIn('id="pvd-detail-root"', html)
        self.assertIn("/api/approved-vs-planned-hours/summary", html)
        self.assertIn("/api/approved-vs-planned-hours/details", html)

    def test_planned_actual_table_view_page_controls_exist(self):
        html_path = Path(__file__).resolve().parents[1] / "report_html" / "planned_actual_table_view.html"
        self.assertTrue(html_path.exists())
        html = html_path.read_text(encoding="utf-8")
        self.assertIn('id="from-date"', html)
        self.assertIn('id="to-date"', html)
        self.assertIn('id="mode"', html)
        self.assertIn('id="projects"', html)
        self.assertIn('id="statuses"', html)
        self.assertIn('id="assignees"', html)
        self.assertIn('id="load-btn"', html)
        self.assertIn('id="fetch-btn"', html)
        self.assertIn("/api/planned-actual-table-view/summary", html)
        self.assertIn("/api/planned-actual-table-view/refresh", html)
        self.assertIn("/api/planned-actual-table-view/filter-options", html)
        self.assertIn("/api/planned-actual-table-view/queue", html)
        self.assertIn("/api/planned-actual-table-view/cancel", html)
        self.assertIn("/api/planned-actual-table-view/history", html)
        self.assertIn("/api/planned-actual-table-view/diff", html)
        self.assertIn("/api/planned-actual-table-view/export", html)
        self.assertIn("Fetch Queue", html)
        self.assertIn("Cancel and Rollback", html)

    def test_delayed_epic_chain_gantt_page_controls_exist(self):
        html_path = Path(__file__).resolve().parents[1] / "delayed_epic_chain_gantt_report.html"
        self.assertTrue(html_path.exists())
        html = html_path.read_text(encoding="utf-8")
        self.assertIn('id="from-date"', html)
        self.assertIn('id="to-date"', html)
        self.assertIn('id="assignee-filter"', html)
        self.assertIn('id="assignee-mode"', html)
        self.assertIn('id="show-full-year"', html)
        self.assertIn('id="week-width"', html)
        self.assertIn('id="gantt-root"', html)
        self.assertIn("/api/delayed-epic-chain-gantt/filter-options", html)
        self.assertIn("/api/delayed-epic-chain-gantt/data", html)
        self.assertIn("/api/delayed-epic-chain-gantt/ui-settings", html)
        self.assertIn("/api/report-date-filter", html)
        self.assertIn("range-box", html)
        self.assertIn("week-cell", html)
        self.assertIn("bar planned", html)
        self.assertIn("bar actual", html)

    def test_executive_dashboard_page_controls_exist(self):
        html_path = Path(__file__).resolve().parents[1] / "executive_dashboard.html"
        self.assertTrue(html_path.exists())
        html = html_path.read_text(encoding="utf-8")
        self.assertIn('id="date-filter-from"', html)
        self.assertIn('id="date-filter-to"', html)
        self.assertIn('id="project-filter"', html)
        self.assertIn('id="date-filter-apply"', html)
        self.assertIn('id="date-filter-reset"', html)
        self.assertIn('id="metric-total-committed"', html)
        self.assertIn('id="metric-completed-items"', html)
        self.assertIn('id="metric-estimation-accuracy"', html)
        self.assertIn('id="metric-cycle-time"', html)
        self.assertIn('id="blocked-cycle-table"', html)
        self.assertIn("/api/executive-dashboard/summary", html)
        self.assertIn("/settings/executive-dashboard", html)

    def test_executive_dashboard_settings_page_exists(self):
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            (root / "report_html").mkdir(parents=True, exist_ok=True)
            (root / "report_html" / "dashboard.html").write_text("<html><body>ok</body></html>", encoding="utf-8")
            wb = Workbook()
            ws = wb.active
            ws.append(["project_key", "worklog_date", "period_day", "period_week", "period_month", "issue_assignee", "hours_logged"])
            ws.append(["O2", "2026-02-01", "2026-02-01", "2026-W05", "2026-02", "Alice", 1.0])
            wb.save(root / "assignee_hours_report.xlsx")
            app = create_report_server_app(base_dir=root, folder_raw="report_html")
            client = app.test_client()
            resp = client.get("/settings/executive-dashboard")
            self.assertEqual(resp.status_code, 200)
            html = resp.get_data(as_text=True)
            self.assertIn("Executive Dashboard Settings", html)
            self.assertIn('id="estimation-basis"', html)
            self.assertIn("/api/executive-dashboard/settings", html)

    def test_original_estimates_hierarchy_page_controls_exist(self):
        html_path = Path(__file__).resolve().parents[1] / "report_html" / "original_estimates_hierarchy_report.html"
        self.assertTrue(html_path.exists())
        html = html_path.read_text(encoding="utf-8")
        self.assertIn('id="from-date"', html)
        self.assertIn('id="to-date"', html)
        self.assertIn('id="projects"', html)
        self.assertIn('id="statuses"', html)
        self.assertIn('id="assignees"', html)
        self.assertIn('id="search-anything"', html)
        self.assertIn('id="apply-btn"', html)
        self.assertIn('id="reset-btn"', html)
        self.assertNotIn('id="fetch-btn"', html)
        self.assertIn('id="table-body"', html)
        self.assertIn("/api/original-estimates/filter-options", html)
        self.assertIn("/api/original-estimates/summary", html)
        self.assertNotIn("/api/original-estimates/refresh", html)
        self.assertNotIn("refresh-epic", html)
        self.assertNotIn("codex-refresh-widget", html)


if __name__ == "__main__":
    unittest.main()
