from __future__ import annotations

import json
import os
import sqlite3
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook

import fetch_jira_dashboard as dashboard_gen


def _create_work_items_xlsx(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "WorkItems"
    ws.append(
        [
            "project_key",
            "issue_key",
            "parent_issue_key",
            "jira_issue_type",
            "summary",
            "status",
            "start_date",
            "end_date",
            "actual_start_date",
            "actual_end_date",
            "original_estimate",
            "total_hours_logged",
            "assignee",
            "jira_url",
        ]
    )
    ws.append(["O2", "O2-1", "", "Epic", "Matched Epic", "In Progress", "2026-02-01", "2026-02-10", "", "", "80", 80.0, "Alice", "https://jira.example.com/browse/O2-1"])
    ws.append(["O2", "O2-2", "", "Epic", "Date Mismatch Epic", "In Progress", "2026-02-01", "2026-02-10", "", "", "80", 80.0, "Bob", "https://jira.example.com/browse/O2-2"])
    ws.append(["O2", "O2-3", "", "Epic", "Hours Mismatch Epic", "In Progress", "2026-02-01", "2026-02-10", "", "", "80", 80.0, "Charlie", "https://jira.example.com/browse/O2-3"])
    ws.append(["O2", "O2-4", "", "Epic", "Incomplete Planner Epic", "In Progress", "2026-02-01", "2026-02-10", "", "", "80", 80.0, "Diana", "https://jira.example.com/browse/O2-4"])
    ws.append(["O2", "O2-5", "", "Epic", "No Planner Epic", "In Progress", "2026-02-01", "2026-02-10", "", "", "80", 80.0, "Evan", "https://jira.example.com/browse/O2-5"])
    ws.append(["O2", "O2-101", "O2-1", "Story", "Planner Story", "In Progress", "2026-03-01", "2026-03-05", "", "", "16", 8.0, "Frank", "https://jira.example.com/browse/O2-101"])
    wb.save(path)


def _create_empty_rows_xlsx(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(["placeholder"])
    wb.save(path)


def _create_minimal_ipp_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Latest"
    ws.append(["Jira Task ID", "Planned Start Date", "Planned End Date"])
    wb.save(path)


def _create_epics_db(path: Path) -> None:
    conn = sqlite3.connect(path)
    try:
        conn.execute(
            """
            CREATE TABLE epics_management (
                epic_key TEXT PRIMARY KEY,
                epic_plan_json TEXT NOT NULL DEFAULT '{}'
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE epics_management_story_sync (
                story_key TEXT PRIMARY KEY,
                epic_key TEXT NOT NULL,
                start_date TEXT NOT NULL DEFAULT '',
                due_date TEXT NOT NULL DEFAULT ''
            )
            """
        )
        rows = [
            ("O2-1", {"start_date": "2026-02-01", "due_date": "2026-02-10", "man_days": 10}),
            ("O2-2", {"start_date": "2026-02-01", "due_date": "2026-02-11", "man_days": 10}),
            ("O2-3", {"start_date": "2026-02-01", "due_date": "2026-02-10", "man_days": 9.5}),
            ("O2-4", {"start_date": "2026-02-01", "due_date": "", "man_days": ""}),
        ]
        conn.executemany(
            "INSERT INTO epics_management (epic_key, epic_plan_json) VALUES (?, ?)",
            [(epic_key, json.dumps(plan)) for epic_key, plan in rows],
        )
        conn.execute(
            "INSERT INTO epics_management_story_sync (story_key, epic_key, start_date, due_date) VALUES (?, ?, ?, ?)",
            ("O2-101", "O2-1", "2026-02-02", "2026-02-08"),
        )
        conn.commit()
    finally:
        conn.close()


class FetchDashboardPlannerValidationTests(unittest.TestCase):
    def test_planner_validation_statuses_and_fields(self):
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            export_path = root / "1.xlsx"
            worklog_path = root / "2.xlsx"
            rollup_path = root / "3.xlsx"
            ipp_path = root / "ipp.xlsx"
            db_path = root / "planner.db"
            _create_work_items_xlsx(export_path)
            _create_empty_rows_xlsx(worklog_path)
            _create_empty_rows_xlsx(rollup_path)
            _create_minimal_ipp_workbook(ipp_path)
            _create_epics_db(db_path)

            env = {
                "JIRA_EXPORT_XLSX_PATH": str(export_path),
                "JIRA_WORKLOG_XLSX_PATH": str(worklog_path),
                "JIRA_SUBTASK_ROLLUP_XLSX_PATH": str(rollup_path),
                "IPP_MEETING_XLSX_PATH": str(ipp_path),
                "JIRA_ASSIGNEE_HOURS_CAPACITY_DB_PATH": str(db_path),
            }
            with patch.dict(os.environ, env, clear=False):
                payload = dashboard_gen.fetch_dashboard_data()

            by_key = {str(item.get("issue_key")): item for item in payload.get("epics", [])}
            self.assertEqual(by_key["O2-1"]["planner_validation_status"], "Matched")
            self.assertEqual(by_key["O2-1"]["planner_dates_match"], "Yes")
            self.assertEqual(by_key["O2-1"]["planner_hours_match"], "Yes")

            self.assertEqual(by_key["O2-2"]["planner_validation_status"], "Mismatch")
            self.assertEqual(by_key["O2-2"]["planner_dates_match"], "No")
            self.assertEqual(by_key["O2-2"]["planner_hours_match"], "Yes")

            self.assertEqual(by_key["O2-3"]["planner_validation_status"], "Mismatch")
            self.assertEqual(by_key["O2-3"]["planner_dates_match"], "Yes")
            self.assertEqual(by_key["O2-3"]["planner_hours_match"], "No")

            self.assertEqual(by_key["O2-4"]["planner_validation_status"], "Incomplete")
            self.assertEqual(by_key["O2-4"]["planner_dates_match"], "N/A")
            self.assertEqual(by_key["O2-4"]["planner_hours_match"], "N/A")

            self.assertEqual(by_key["O2-5"]["planner_validation_status"], "No Planner Entry")
            self.assertEqual(by_key["O2-5"]["planner_dates_match"], "N/A")
            self.assertEqual(by_key["O2-5"]["planner_hours_match"], "N/A")

            stories_by_key = {str(item.get("issue_key")): item for item in payload.get("stories", [])}
            self.assertEqual(stories_by_key["O2-101"]["jira_start_date"], "2026-03-01")
            self.assertEqual(stories_by_key["O2-101"]["jira_end_date"], "2026-03-05")
            self.assertEqual(stories_by_key["O2-101"]["planner_story_start_date"], "2026-02-02")
            self.assertEqual(stories_by_key["O2-101"]["planner_story_end_date"], "2026-02-08")

    def test_missing_planner_db_or_table_defaults_to_no_planner_entry(self):
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            export_path = root / "1.xlsx"
            worklog_path = root / "2.xlsx"
            rollup_path = root / "3.xlsx"
            ipp_path = root / "ipp.xlsx"
            db_path = root / "missing.db"
            _create_work_items_xlsx(export_path)
            _create_empty_rows_xlsx(worklog_path)
            _create_empty_rows_xlsx(rollup_path)
            _create_minimal_ipp_workbook(ipp_path)

            env = {
                "JIRA_EXPORT_XLSX_PATH": str(export_path),
                "JIRA_WORKLOG_XLSX_PATH": str(worklog_path),
                "JIRA_SUBTASK_ROLLUP_XLSX_PATH": str(rollup_path),
                "IPP_MEETING_XLSX_PATH": str(ipp_path),
                "JIRA_ASSIGNEE_HOURS_CAPACITY_DB_PATH": str(db_path),
            }
            with patch.dict(os.environ, env, clear=False):
                payload = dashboard_gen.fetch_dashboard_data()

            for epic in payload.get("epics", []):
                self.assertEqual(epic.get("planner_validation_status"), "No Planner Entry")
                self.assertEqual(epic.get("planner_dates_match"), "N/A")
                self.assertEqual(epic.get("planner_hours_match"), "N/A")


if __name__ == "__main__":
    unittest.main()
