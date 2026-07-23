from __future__ import annotations

from io import BytesIO
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

from report_server import (
    RND_MUSCLE_UTILIZATION_SETTINGS_ROUTE,
    RND_MUSCLE_UTILIZATION_VIEW_ROUTE,
    _resolve_capacity_runtime_paths,
    create_report_server_app,
)


def _build_app(root: Path):
    (root / "report_html").mkdir(parents=True, exist_ok=True)
    (root / "report_html" / "dashboard.html").write_text("<html><body>ok</body></html>", encoding="utf-8")
    wb = Workbook()
    ws = wb.active
    ws.append(["project_key", "worklog_date", "period_day", "period_week", "period_month", "issue_assignee", "hours_logged"])
    ws.append(["O2", "2026-02-01", "2026-02-01", "2026-W05", "2026-02", "Alice", 1.0])
    wb.save(root / "assignee_hours_report.xlsx")
    return create_report_server_app(base_dir=root, folder_raw="report_html")


def _seed_canonical_resource(root: Path) -> None:
    import sqlite3

    conn = sqlite3.connect(root / "assignee_hours_capacity.db")
    try:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS canonical_issues (
                run_id TEXT NOT NULL,
                issue_key TEXT NOT NULL,
                project_key TEXT NOT NULL,
                issue_type TEXT NOT NULL,
                summary TEXT NOT NULL,
                status TEXT NOT NULL,
                assignee TEXT NOT NULL
            )
            """
        )
        conn.execute(
            """
            INSERT INTO canonical_issues(run_id, issue_key, project_key, issue_type, summary, status, assignee)
            VALUES('r1','O2-1','O2','Task','One','Done','Ayesha Khan')
            """
        )
        conn.commit()
    finally:
        conn.close()


def _seed_resource_resignation(root: Path, assignee_name: str, resignation_date: str) -> None:
    import sqlite3

    conn = sqlite3.connect(root / "assignee_hours_capacity.db")
    try:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS performance_resource_resignations (
                assignee_name TEXT PRIMARY KEY,
                resignation_date TEXT,
                updated_at TEXT NOT NULL DEFAULT ''
            )
            """
        )
        conn.execute(
            """
            INSERT OR REPLACE INTO performance_resource_resignations(assignee_name, resignation_date, updated_at)
            VALUES(?,?,?)
            """,
            (assignee_name, resignation_date, "2026-07-08T00:00:00Z"),
        )
        conn.commit()
    finally:
        conn.close()


def _seed_epics_and_resources(root: Path) -> None:
    import sqlite3

    capacity_conn = sqlite3.connect(root / "assignee_hours_capacity.db")
    try:
        capacity_conn.execute(
            """
            CREATE TABLE epics_management (
                id TEXT PRIMARY KEY,
                epic_key TEXT NOT NULL,
                project_key TEXT NOT NULL,
                project_name TEXT NOT NULL,
                epic_name TEXT NOT NULL,
                priority TEXT NOT NULL DEFAULT 'Low',
                start_date TEXT NOT NULL DEFAULT '',
                due_date TEXT NOT NULL DEFAULT '',
                jira_url TEXT NOT NULL DEFAULT ''
            )
            """
        )
        capacity_conn.executemany(
            """
            INSERT INTO epics_management(id, epic_key, project_key, project_name, epic_name, priority)
            VALUES(?,?,?,?,?,?)
            """,
            [
                ("row-1", "O2-100", "O2", "OmniConnect", "First Epic", "High"),
                ("row-2", "FF-200", "FF", "Fintech Fuel", "Second Epic", "Low"),
            ],
        )
        capacity_conn.executemany(
            """
            INSERT INTO rnd_muscle_resources(resource_id, display_name, initials, email, team_id)
            VALUES(?,?,?,?,?)
            """,
            [
                ("res-1", "Hassan Malik", "HM", "hassan@example.com", ""),
                ("res-2", "Ayesha Khan", "AK", "ayesha@example.com", ""),
            ],
        )
        capacity_conn.commit()
    finally:
        capacity_conn.close()


class RndMuscleUtilizationApiTests(unittest.TestCase):
    def test_rnd_db_path_strips_accidental_quotes(self):
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            expected = root / "runtime" / "rnd.db"

            with patch.dict("os.environ", {"JIRA_RND_MUSCLE_UTILIZATION_DB_PATH": f'"{expected}"'}, clear=False):
                resolved = _resolve_capacity_runtime_paths(root)["rnd_muscle_db_path"]

            self.assertEqual(resolved, expected)
            self.assertTrue(expected.parent.exists())

    def test_rnd_db_defaults_to_capacity_db_when_not_explicitly_configured(self):
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)

            with patch.dict("os.environ", {"JIRA_RND_MUSCLE_UTILIZATION_DB_PATH": ""}, clear=False):
                resolved = _resolve_capacity_runtime_paths(root)

            self.assertEqual(resolved["rnd_muscle_db_path"], resolved["db_path"])

    def test_explicit_rnd_db_path_falls_back_to_home_data_when_not_writable(self):
        with (
            tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as app_td,
            tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as home_td,
        ):
            app_root = Path(app_td)
            home_root = Path(home_td)
            (app_root / "rnd_muscle_utilization.db").mkdir()

            with patch.dict(
                "os.environ",
                {
                    "HOME": str(home_root),
                    "JIRA_RND_MUSCLE_UTILIZATION_DB_PATH": "rnd_muscle_utilization.db",
                },
                clear=False,
            ):
                resolved = _resolve_capacity_runtime_paths(app_root)["rnd_muscle_db_path"]

            self.assertEqual(resolved, home_root / "data" / "rnd_muscle_utilization.db")
            self.assertTrue(resolved.parent.exists())

    def test_default_rnd_tables_are_created_in_existing_capacity_database(self):
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            app = _build_app(root)
            client = app.test_client()

            resp = client.get("/api/rnd-muscle-utilization")

            import sqlite3

            with sqlite3.connect(root / "assignee_hours_capacity.db") as capacity_conn:
                capacity_tables = {
                    row[0]
                    for row in capacity_conn.execute(
                        "SELECT name FROM sqlite_master WHERE type='table' AND name LIKE 'rnd_muscle_%'"
                    ).fetchall()
                }

        self.assertEqual(resp.status_code, 200)
        self.assertIn("rnd_muscle_resources", capacity_tables)

    def test_page_controls_are_wired_to_client_handlers(self):
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            app = _build_app(Path(td))
            client = app.test_client()

            resp = client.get(RND_MUSCLE_UTILIZATION_SETTINGS_ROUTE)

        self.assertEqual(resp.status_code, 200)
        html = resp.get_data(as_text=True)
        self.assertIn('id="rnd-create-team-btn"', html)
        self.assertIn('id="rnd-add-skill-btn"', html)
        self.assertIn('id="rnd-configure-projects-btn"', html)
        self.assertIn('id="rnd-theme-mode"', html)
        self.assertIn('role="switch"', html)
        self.assertIn('class="theme-switch"', html)
        self.assertIn('id="rnd-theme-color"', html)
        self.assertIn('id="rnd-export-mappings-btn"', html)
        self.assertIn('id="rnd-import-mappings-btn"', html)
        self.assertIn('id="rnd-import-mappings-file"', html)
        self.assertIn('accept=".xlsx,application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"', html)
        self.assertIn('function exportMappingWorkbook()', html)
        self.assertIn('function importMappingWorkbook(file)', html)
        self.assertIn('API + "/mappings/export"', html)
        self.assertIn('API + "/mappings/import"', html)
        self.assertIn('formData.append("workbook", file, file.name)', html)
        self.assertIn('settlePlannerSavesBeforeWorkbookAction', html)
        self.assertIn('id="rnd-team-color-grid"', html)
        self.assertIn('TEAM_COLOR_PALETTE', html)
        self.assertIn('rnd-muscle-theme-mode-v1', html)
        self.assertIn('rnd-muscle-theme-color-v1', html)
        self.assertIn('Theme saved for configuration and view mode.', html)
        self.assertIn('--scrollbar-track:var(--panel-2)', html)
        self.assertIn('body[data-rnd-theme="light"]', html)
        self.assertIn('--scrollbar-track:#f3f3f3', html)
        self.assertIn('body[data-rnd-theme="dark"]', html)
        self.assertIn('--scrollbar-track:#2d2d30', html)
        self.assertIn('body[data-rnd-theme] *::-webkit-scrollbar-thumb', html)
        self.assertIn('scrollbar-color:var(--scrollbar-thumb) var(--scrollbar-track)', html)
        self.assertIn('function epicPriorityClass(priority)', html)
        self.assertIn('epic-priority-1', html)
        self.assertIn('--priority-bg:#4c1d95', html)
        self.assertIn('epic-priority-2', html)
        self.assertIn('--priority-bg:#7c3aed', html)
        self.assertIn('epic-priority-3', html)
        self.assertIn('--priority-bg:#ede9fe', html)
        self.assertIn('body[data-rnd-theme="dark"] .epic-priority-1', html)
        self.assertIn('--priority-bg:#452064', html)
        self.assertIn('body[data-rnd-theme="dark"] .epic-priority-2', html)
        self.assertIn('--priority-bg:#352640', html)
        self.assertIn('body[data-rnd-theme="dark"] .epic-priority-3', html)
        self.assertIn('--priority-bg:#2b2931', html)
        self.assertIn('row.className = "row " + epicPriorityClass(epic.priority)', html)
        self.assertIn('card.className = "staged-card " + epicPriorityClass(epic && epic.priority)', html)
        self.assertIn('tr.className = epicPriorityClass(item.priority)', html)
        self.assertIn('btn.className = "cluster-epic " + epicPriorityClass(epic && epic.priority)', html)
        self.assertNotIn('--priority-accent-width', html)
        self.assertNotIn('--priority-shadow', html)
        self.assertNotIn('.backlog tr.epic-priority td:first-child', html)
        self.assertIn('resource-row.team-colored', html)
        self.assertIn('id="rnd-resource-search"', html)
        self.assertIn('resourceSearchText', html)
        self.assertIn('function teamByResourceId()', html)
        self.assertIn('function teamForResource(resource, teamsByIdMap, teamByResourceIdMap)', html)
        self.assertIn('teamForResource(resource, teamsById, teamsByResourceId)', html)
        self.assertIn('resource-team-group', html)
        self.assertIn('resource-team-head', html)
        self.assertIn('a.key === "__no_team__"', html)
        self.assertIn('if (resource.resigned) return false;', html)
        self.assertIn('resource-resigned-chip', html)
        self.assertIn('resource-booked-chip', html)
        self.assertIn('resource-booked-menu', html)
        self.assertIn('position:fixed; z-index:100000', html)
        self.assertIn('document.body.appendChild(menu)', html)
        self.assertIn('positionBookedMenu(bookedChip, menu)', html)
        self.assertIn('window.addEventListener("scroll", () => hideOpenBookedMenus(), true)', html)
        self.assertIn('epicBookedDetailsForResource', html)
        self.assertIn('hideOpenBookedMenus', html)
        self.assertIn('clearBookedMenus', html)
        self.assertIn('booked on " + bookedEpics.length', html)
        self.assertIn('Booked epics', html)
        self.assertIn('No active resources match this search.', html)
        self.assertIn('function treemapLayouts(groups)', html)
        self.assertIn('tile.style.width = "calc(" + layout.w.toFixed(3) + "% - 8px)"', html)
        self.assertIn('tile.style.height = "calc(" + layout.h.toFixed(3) + "% - 8px)"', html)
        self.assertIn('compareInsightGroups', html)
        self.assertIn('--resource-team-soft', html)
        self.assertIn('--resource-team-accent', html)
        self.assertIn('--resource-team-text', html)
        self.assertIn('const strong = lightMode ? base : mixColor(base, "#181818", 0.38)', html)
        self.assertIn('strongText: readableTextColor(strong)', html)
        self.assertIn('card.style.background = tone.strong', html)
        self.assertIn('card.style.setProperty("--resource-card-accent", tone.base)', html)
        self.assertIn('bubble.style.background = tone.strong', html)
        self.assertIn('row.className = "row resource-row" + (team.team_id ? " team-colored" : "")', html)
        self.assertIn('renderTeamColorPalette', html)
        self.assertIn('setTeamColor', html)
        self.assertIn('byId("rnd-theme-mode").checked ? "light" : "dark"', html)
        self.assertIn('@keyframes rndDropPulse', html)
        self.assertIn('card.classList.add("dragging")', html)
        self.assertIn('row.classList.add("dragging")', html)
        self.assertIn('"/api/rnd-muscle-utilization"', html)
        self.assertIn('addEventListener("click"', html)
        self.assertIn('showModal()', html)
        self.assertIn('role="status"', html)
        self.assertIn('id="rnd-existing-skills"', html)
        self.assertIn('id="rnd-project-form"', html)
        self.assertIn(RND_MUSCLE_UTILIZATION_VIEW_ROUTE, html)
        self.assertIn('id="rnd-staged-resources"', html)
        self.assertIn('id="rnd-save-progress"', html)
        self.assertIn('role="progressbar"', html)
        self.assertIn('const SAVE_IDLE_DELAY_MS = 2500;', html)
        self.assertIn('function scheduleDeferredSave', html)
        self.assertIn('function flushDeferredSaves', html)
        self.assertIn('setEpicResourceMappings(epicKey, resourceIds, allocation);', html)
        self.assertIn('setPlannerEpicOrder(epicKeys);', html)
        self.assertIn('id="rnd-cluster-stage"', html)
        self.assertIn('id="rnd-cluster-legend"', html)
        self.assertIn('<strong>Team colors</strong>', html)
        self.assertIn('visibleName.textContent = resourceName', html)
        self.assertNotIn('bubble.textContent = (resource && resource.initials)', html)
        self.assertIn('id="rnd-view-product"', html)
        self.assertIn('>Product wise</button>', html)
        self.assertIn('id="rnd-product-planner"', html)
        self.assertIn('id="rnd-product-projects"', html)
        self.assertIn('id="rnd-product-people"', html)
        self.assertIn('id="rnd-product-team-legend"', html)
        self.assertIn('class="product-layout"', html)
        self.assertIn('let currentView = "hierarchical";', html)
        self.assertIn('id="rnd-view-hierarchical" class="tab active"', html)
        self.assertIn('id="rnd-view-product" class="tab "', html)
        self.assertIn('grid-template-columns:repeat(auto-fill, minmax(240px, 1fr))', html)
        self.assertIn('overflow-y:auto', html)
        self.assertIn('PRODUCT_PANEL_ORDER_KEY', html)
        self.assertIn('application/x-rnd-product-project', html)
        self.assertIn('id="rnd-product-people-btn"', html)
        self.assertIn('id="rnd-product-people-popover"', html)
        self.assertIn('class="product-wise-head"', html)
        self.assertIn('function setProductPeopleOpen(open)', html)
        self.assertIn('id="rnd-opt-show-logos"', html)
        self.assertIn('id="rnd-opt-apply-colors"', html)
        self.assertIn('SHOW_PROJECT_LOGOS_KEY', html)
        self.assertIn('APPLY_PROJECT_COLORS_KEY', html)
        self.assertIn('function contrastTextColor(hex)', html)
        self.assertIn('function renderProductWiseView()', html)
        self.assertIn('function renderProductPeople()', html)
        self.assertIn('function productEpicMatchesPeople(epicKey, mappings)', html)
        self.assertIn('id="rnd-product-people-search"', html)
        self.assertIn('id="rnd-product-clear-people"', html)
        self.assertIn('const selectedProductResourceIds = new Set()', html)
        self.assertIn('selectedProductResourceIds.has(mapping.resource_id)', html)
        self.assertIn('selectedProductResourceIds.add(resource.resource_id)', html)
        self.assertIn('selectedProductResourceIds.delete(resource.resource_id)', html)
        self.assertIn('selectedProductResourceIds.clear()', html)
        self.assertIn('card.setAttribute("aria-pressed"', html)
        self.assertIn('const epicResourceIds = selectedProductEpicKey', html)
        self.assertIn('mapping.epic_key === selectedProductEpicKey', html)
        self.assertIn('const peopleForEpic = epicResourceIds', html)
        self.assertIn('People involved in " + selectedEpicLabel', html)
        self.assertIn('Epics match any selected person.', html)
        self.assertIn('No active people are mapped to this epic.', html)
        self.assertIn('Showing epics involving any selected person.', html)
        self.assertIn('No epics involve the selected people.', html)
        self.assertIn('projectEpics.length + " of " + allProjectEpics.length', html)
        self.assertIn('selectedProductEpicKey = wasSelected ? "" : epic.epic_key', html)
        self.assertIn('Select people to show epics involving any of them.', html)
        self.assertIn('No people match your search in this view.', html)
        self.assertIn('findPlannerEpic', html)
        self.assertIn('id="rnd-resource-skill-dialog"', html)
        self.assertIn('id="rnd-resource-direct-skills"', html)
        self.assertIn('openResourceSkills', html)
        self.assertIn('Resource skills', html)
        self.assertNotIn('Inherited from team', html)
        self.assertIn('"/resources/" + encodeURIComponent(resourceId) + "/skills"', html)
        self.assertIn('Edit skills', html)
        self.assertIn('[hidden] { display:none !important; }', html)
        self.assertIn('<h3>Mapped epics</h3>', html)
        self.assertIn('<h3>Mapped resources</h3>', html)
        self.assertIn('id="rnd-backlog-drop-zone"', html)
        self.assertIn('id="rnd-backlog-body"', html)
        self.assertIn('epicLabel.className = "backlog-epic"', html)
        self.assertIn('epicName.textContent = item.epic_name', html)
        self.assertIn('item.epic_key + (item.epic_name ? " - " + item.epic_name : "")', html)
        self.assertIn('Add to planner', html)
        self.assertIn('id="rnd-mapped-epics-caption"', html)
        self.assertIn('API + "/planner"', html)
        self.assertIn('API + "/planner/reorder"', html)
        self.assertIn('selectedCanvasEpicKey = canvasEpicKeys.find((key) => mappedEpicKeys.has(key)) || canvasEpicKeys[0];', html)
        self.assertIn('draggable = true', html)
        self.assertIn('application/x-rnd-epic', html)
        self.assertIn('application/x-rnd-planner-epic', html)
        self.assertIn('application/x-rnd-resource', html)
        self.assertIn('reorderPlannerEpic', html)
        self.assertIn('staged-drop-placeholder', html)
        self.assertIn('showPlannerEpicDropPlaceholder', html)
        self.assertIn('targetPlaceholder', html)
        self.assertIn('!targetCard && !targetPlaceholder', html)
        self.assertIn('movePlannerEpicToPosition', html)
        self.assertIn('Drop epic here', html)
        self.assertIn('reorderPlannerEpic(draggedEpicKey, dropTarget.targetEpicKey, dropTarget.position)', html)
        self.assertIn('cluster-bubble-detail', html)
        self.assertIn('resourceName + " - " + teamName', html)
        self.assertIn('detailTeam.textContent = teamName', html)
        self.assertIn('epicKeyFromDrag(event)', html)
        self.assertIn('dragHasEpic(event)', html)
        self.assertIn('byId("rnd-backlog-drop-zone").addEventListener("drop"', html)

    def test_view_mode_route_renders_planner_without_side_catalogs(self):
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            app = _build_app(Path(td))
            client = app.test_client()

            resp = client.get(RND_MUSCLE_UTILIZATION_VIEW_ROUTE)

        self.assertEqual(resp.status_code, 200)
        html = resp.get_data(as_text=True)
        self.assertIn('class="page planner-only"', html)
        self.assertIn('id="rnd-hierarchical-planner"', html)
        self.assertIn('id="rnd-product-planner"', html)
        self.assertIn('let currentView = "product";', html)
        self.assertIn('id="rnd-view-product" class="tab active"', html)
        self.assertIn('id="rnd-view-hierarchical" class="tab "', html)
        self.assertIn('id="rnd-hierarchical-planner" class="hierarchical-planner" hidden', html)
        self.assertIn('id="rnd-product-planner" class="product-planner" >', html)
        self.assertIn('id="rnd-theme-mode"', html)
        self.assertIn('id="rnd-theme-color"', html)
        self.assertIn('Configuration', html)
        self.assertIn('<body class="rnd-planner-only">', html)
        self.assertIn('body.rnd-planner-only { overflow:auto; }', html)
        self.assertIn('.page.planner-only .planner { height:auto; grid-template-rows:auto 100vh auto; }', html)
        self.assertIn('.page.planner-only .canvas { height:100vh; min-height:100vh; }', html)

    def test_planner_order_apis_return_refreshable_state(self):
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            app = _build_app(root)
            client = app.test_client()
            client.get("/api/rnd-muscle-utilization")
            _seed_epics_and_resources(root)

            self.assertEqual(client.post("/api/rnd-muscle-utilization/backlog", json={"epic_key": "O2-100"}).status_code, 201)
            self.assertEqual(client.post("/api/rnd-muscle-utilization/backlog", json={"epic_key": "FF-200"}).status_code, 201)
            self.assertEqual(client.post("/api/rnd-muscle-utilization/planner", json={"epic_key": "O2-100"}).status_code, 201)
            self.assertEqual(client.post("/api/rnd-muscle-utilization/planner", json={"epic_key": "FF-200"}).status_code, 201)
            backlog_resp = client.post("/api/rnd-muscle-utilization/backlog/reorder", json={"epic_keys": ["FF-200", "O2-100"]})
            epic_resp = client.post("/api/rnd-muscle-utilization/planner/reorder", json={"epic_keys": ["FF-200", "O2-100"]})
            map_resp = client.post(
                "/api/rnd-muscle-utilization/mappings",
                json={"epic_key": "O2-100", "resource_ids": ["res-1", "res-2"], "allocation_hours_by_resource_id": {}},
            )
            resource_resp = client.post(
                "/api/rnd-muscle-utilization/mappings/reorder",
                json={"epic_key": "O2-100", "resource_ids": ["res-2", "res-1"]},
            )

        self.assertEqual(backlog_resp.status_code, 200)
        self.assertEqual(epic_resp.status_code, 200)
        self.assertEqual([item["epic_key"] for item in backlog_resp.get_json()["state"]["planner"]["backlog"]], ["FF-200", "O2-100"])
        self.assertEqual([item["epic_key"] for item in epic_resp.get_json()["state"]["planner"]["planner_epics"]], ["FF-200", "O2-100"])
        planner_by_key = {item["epic_key"]: item for item in epic_resp.get_json()["state"]["planner"]["planner_epics"]}
        self.assertEqual(planner_by_key["O2-100"]["epic_name"], "First Epic")
        self.assertEqual(planner_by_key["O2-100"]["project_key"], "O2")
        self.assertEqual(planner_by_key["O2-100"]["project_name"], "OmniConnect")
        self.assertEqual(map_resp.status_code, 200)
        self.assertEqual(resource_resp.status_code, 200)
        self.assertEqual(
            [m["resource_id"] for m in resource_resp.get_json()["state"]["planner"]["mappings"] if m["epic_key"] == "O2-100"],
            ["res-2", "res-1"],
        )

    def test_project_tabs_carry_managed_project_color_and_images(self):
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            app = _build_app(root)
            client = app.test_client()
            client.get("/api/rnd-muscle-utilization")
            _seed_epics_and_resources(root)

            existing = client.get("/api/projects").get_json()["projects"]
            if any(item["project_key"] == "O2" for item in existing):
                update_resp = client.put(
                    "/api/projects/O2",
                    json={"display_name": "OmniConnect", "color_hex": "#AB12CD"},
                )
                self.assertEqual(update_resp.status_code, 200)
            else:
                create_resp = client.post(
                    "/api/projects",
                    json={"project_key": "O2", "project_name": "OmniConnect", "display_name": "OmniConnect", "color_hex": "#AB12CD"},
                )
                self.assertEqual(create_resp.status_code, 200)

            state_resp = client.get("/api/rnd-muscle-utilization")
            tabs = state_resp.get_json()["state"]["project_tabs"]

        tabs_by_key = {tab["project_key"]: tab for tab in tabs if not tab.get("is_all_tab")}
        self.assertEqual(tabs_by_key["O2"]["color_hex"], "#AB12CD")
        self.assertIn("thumbnail_url", tabs_by_key["O2"])
        self.assertIn("logo_url", tabs_by_key["O2"])
        self.assertIsNone(tabs_by_key["O2"]["thumbnail_url"])
        all_tab = next(tab for tab in tabs if tab.get("is_all_tab"))
        self.assertNotIn("color_hex", all_tab)

    def test_mapping_workbook_export_can_be_imported_back_to_restore_mappings(self):
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            app = _build_app(root)
            client = app.test_client()
            client.get("/api/rnd-muscle-utilization")
            _seed_epics_and_resources(root)
            client.post("/api/rnd-muscle-utilization/planner", json={"epic_key": "O2-100"})
            client.post("/api/rnd-muscle-utilization/planner", json={"epic_key": "FF-200"})
            client.post(
                "/api/rnd-muscle-utilization/mappings",
                json={
                    "epic_key": "O2-100",
                    "resource_ids": ["res-1", "res-2"],
                    "allocation_hours_by_resource_id": {"res-1": 6.5, "res-2": 2},
                },
            )

            export_resp = client.get("/api/rnd-muscle-utilization/mappings/export")
            exported_bytes = export_resp.data
            workbook = load_workbook(BytesIO(exported_bytes), data_only=True)
            mapping_rows = list(workbook["Mappings"].iter_rows(values_only=True))
            sheet_names = workbook.sheetnames
            workbook.close()

            client.post(
                "/api/rnd-muscle-utilization/mappings",
                json={
                    "epic_key": "O2-100",
                    "resource_ids": ["res-2"],
                    "allocation_hours_by_resource_id": {"res-2": 9},
                },
            )
            import_resp = client.post(
                "/api/rnd-muscle-utilization/mappings/import",
                data={"workbook": (BytesIO(exported_bytes), "roundtrip.xlsx")},
                content_type="multipart/form-data",
            )

        self.assertEqual(export_resp.status_code, 200)
        self.assertIn("rnd-muscle-mappings-", export_resp.headers["Content-Disposition"])
        self.assertEqual(sheet_names, ["Mappings", "Resources", "Skills", "Teams", "Instructions"])
        self.assertEqual(
            mapping_rows[0],
            (
                "Epic Key",
                "Epic Name",
                "Project Key",
                "Project Name",
                "Resource ID",
                "Resource Name",
                "Resource Email",
                "Team",
                "Allocation Hours",
                "Sort Order",
            ),
        )
        self.assertTrue(any(row[0] == "FF-200" and not row[4] for row in mapping_rows[1:]))
        self.assertEqual(import_resp.status_code, 200)
        import_body = import_resp.get_json()
        self.assertEqual(import_body["imported"], {"epic_count": 2, "mapping_count": 2})
        restored = [
            item
            for item in import_body["state"]["planner"]["mappings"]
            if item["epic_key"] == "O2-100"
        ]
        self.assertEqual([item["resource_id"] for item in restored], ["res-1", "res-2"])
        self.assertEqual([item["allocation_hours"] for item in restored], [6.5, 2.0])
        self.assertFalse(any(item["epic_key"] == "FF-200" for item in import_body["state"]["planner"]["mappings"]))

    def test_mapping_workbook_export_and_import_round_trips_teams_and_skills(self):
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            app = _build_app(root)
            client = app.test_client()
            client.get("/api/rnd-muscle-utilization")
            _seed_epics_and_resources(root)

            skill_resp = client.post("/api/rnd-muscle-utilization/skills", json={"name": "Data Engineering"})
            self.assertEqual(skill_resp.status_code, 201)
            skill_id = next(
                s["skill_id"] for s in skill_resp.get_json()["state"]["skills"] if s["name"] == "Data Engineering"
            )
            team_resp = client.post(
                "/api/rnd-muscle-utilization/teams",
                json={"name": "Platform Squad", "color_hex": "#2563eb", "skill_ids": [skill_id], "resource_ids": ["res-1"]},
            )
            self.assertEqual(team_resp.status_code, 200)

            export_resp = client.get("/api/rnd-muscle-utilization/mappings/export")
            exported_bytes = export_resp.data
            workbook = load_workbook(BytesIO(exported_bytes), data_only=True)
            skills_rows = list(workbook["Skills"].iter_rows(values_only=True))
            teams_rows = list(workbook["Teams"].iter_rows(values_only=True))
            workbook.close()

            # Fresh app/DB: nothing exists yet.
            root2 = Path(td) / "fresh"
            app2 = _build_app(root2)
            client2 = app2.test_client()
            client2.get("/api/rnd-muscle-utilization")
            _seed_epics_and_resources(root2)

            import_resp = client2.post(
                "/api/rnd-muscle-utilization/mappings/import",
                data={"workbook": (BytesIO(exported_bytes), "teams-skills.xlsx")},
                content_type="multipart/form-data",
            )

        self.assertIn(("Data Engineering",), skills_rows[1:])
        self.assertEqual(teams_rows[0], ("Team Name", "Color Hex", "Skill Names", "Resource Names"))
        self.assertIn(("Platform Squad", "#2563eb", "Data Engineering", "Hassan Malik"), teams_rows[1:])

        self.assertEqual(import_resp.status_code, 200)
        imported = import_resp.get_json()["imported"]
        self.assertEqual(imported["skills_added"], 1)
        self.assertEqual(imported["teams_created"], 1)
        state = import_resp.get_json()["state"]
        new_team = next(t for t in state["teams"] if t["name"] == "Platform Squad")
        self.assertEqual(new_team["color_hex"], "#2563eb")
        new_skill_names = {s["name"] for s in state["skills"]}
        self.assertIn("Data Engineering", new_skill_names)
        resources_by_name = {r["display_name"]: r["resource_id"] for r in state["resources"]}
        self.assertEqual(new_team["resource_ids"], [resources_by_name["Hassan Malik"]])

    def test_mapping_workbook_import_rejects_unknown_resource_without_partial_changes(self):
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            app = _build_app(root)
            client = app.test_client()
            client.get("/api/rnd-muscle-utilization")
            _seed_epics_and_resources(root)
            client.post("/api/rnd-muscle-utilization/planner", json={"epic_key": "O2-100"})
            client.post(
                "/api/rnd-muscle-utilization/mappings",
                json={
                    "epic_key": "O2-100",
                    "resource_ids": ["res-1"],
                    "allocation_hours_by_resource_id": {"res-1": 4},
                },
            )
            export_resp = client.get("/api/rnd-muscle-utilization/mappings/export")
            workbook = load_workbook(BytesIO(export_resp.data))
            workbook["Mappings"].cell(row=2, column=5, value="missing-resource")
            invalid_payload = BytesIO()
            workbook.save(invalid_payload)
            workbook.close()
            invalid_payload.seek(0)

            import_resp = client.post(
                "/api/rnd-muscle-utilization/mappings/import",
                data={"workbook": (invalid_payload, "invalid-mappings.xlsx")},
                content_type="multipart/form-data",
            )
            state_resp = client.get("/api/rnd-muscle-utilization")

        self.assertEqual(import_resp.status_code, 400)
        self.assertIn("Unknown resource_id", import_resp.get_json()["error"])
        current = [
            item
            for item in state_resp.get_json()["state"]["planner"]["mappings"]
            if item["epic_key"] == "O2-100"
        ]
        self.assertEqual([(item["resource_id"], item["allocation_hours"]) for item in current], [("res-1", 4.0)])

    def test_mapping_workbook_blank_resource_row_clears_that_epic_mapping(self):
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            app = _build_app(root)
            client = app.test_client()
            client.get("/api/rnd-muscle-utilization")
            _seed_epics_and_resources(root)
            client.post("/api/rnd-muscle-utilization/planner", json={"epic_key": "O2-100"})
            client.post(
                "/api/rnd-muscle-utilization/mappings",
                json={
                    "epic_key": "O2-100",
                    "resource_ids": ["res-1"],
                    "allocation_hours_by_resource_id": {"res-1": 4},
                },
            )
            export_resp = client.get("/api/rnd-muscle-utilization/mappings/export")
            workbook = load_workbook(BytesIO(export_resp.data))
            mapping_sheet = workbook["Mappings"]
            for column in range(5, 11):
                mapping_sheet.cell(row=2, column=column).value = None
            clear_payload = BytesIO()
            workbook.save(clear_payload)
            workbook.close()
            clear_payload.seek(0)

            import_resp = client.post(
                "/api/rnd-muscle-utilization/mappings/import",
                data={"workbook": (clear_payload, "clear-o2.xlsx")},
                content_type="multipart/form-data",
            )

        self.assertEqual(import_resp.status_code, 200)
        body = import_resp.get_json()
        self.assertEqual(body["imported"], {"epic_count": 1, "mapping_count": 0})
        self.assertFalse(any(item["epic_key"] == "O2-100" for item in body["state"]["planner"]["mappings"]))

    def test_state_skill_and_team_apis_return_refreshable_state(self):
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            app = _build_app(root)
            _seed_canonical_resource(root)
            _seed_resource_resignation(root, "Ayesha Khan", "2026-06-30")
            client = app.test_client()

            state_resp = client.get("/api/rnd-muscle-utilization")
            self.assertEqual(state_resp.status_code, 200)
            initial_body = state_resp.get_json()
            self.assertTrue(initial_body["ok"])
            self.assertGreaterEqual(len(initial_body["state"]["skills"]), 11)
            resource = next(item for item in initial_body["state"]["resources"] if item["display_name"] == "Ayesha Khan")
            self.assertTrue(resource["resigned"])
            self.assertEqual(resource["resignation_date"], "2026-06-30")

            skill_resp = client.post("/api/rnd-muscle-utilization/skills", json={"name": "Data Engineering"})
            self.assertEqual(skill_resp.status_code, 201)
            skill_body = skill_resp.get_json()
            skill = next(item for item in skill_body["state"]["skills"] if item["name"] == "Data Engineering")

            team_resp = client.post(
                "/api/rnd-muscle-utilization/teams",
                json={"name": "Platform", "color_hex": "#2563eb", "skill_ids": [skill["skill_id"]], "resource_ids": [resource["resource_id"]]},
            )
            self.assertEqual(team_resp.status_code, 200)
            team_body = team_resp.get_json()
            self.assertTrue(team_body["ok"])
            team = next(item for item in team_body["state"]["teams"] if item["name"] == "Platform")
            self.assertEqual(team["skill_ids"], [skill["skill_id"]])
            self.assertEqual(team["resource_ids"], [resource["resource_id"]])

            direct_skill_resp = client.post("/api/rnd-muscle-utilization/skills", json={"name": "Architecture"})
            self.assertEqual(direct_skill_resp.status_code, 201)
            direct_skill = next(item for item in direct_skill_resp.get_json()["state"]["skills"] if item["name"] == "Architecture")

            resource_skill_resp = client.put(
                f"/api/rnd-muscle-utilization/resources/{resource['resource_id']}/skills",
                json={"skill_ids": [direct_skill["skill_id"]]},
            )
            self.assertEqual(resource_skill_resp.status_code, 200)
            resource_after = next(
                item
                for item in resource_skill_resp.get_json()["state"]["resources"]
                if item["resource_id"] == resource["resource_id"]
            )
            self.assertEqual(resource_after["skill_ids"], [direct_skill["skill_id"]])

    def test_invalid_team_payload_returns_400_instead_of_silent_halt(self):
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            app = _build_app(Path(td))
            client = app.test_client()

            resp = client.post(
                "/api/rnd-muscle-utilization/teams",
                json={"name": "Bad Color", "color_hex": "blue"},
            )

        self.assertEqual(resp.status_code, 400)
        self.assertFalse(resp.get_json()["ok"])
        self.assertIn("Invalid color_hex", resp.get_json()["error"])

    def test_non_palette_team_color_returns_400(self):
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            app = _build_app(Path(td))
            client = app.test_client()

            resp = client.post(
                "/api/rnd-muscle-utilization/teams",
                json={"name": "Off Palette", "color_hex": "#123456"},
            )

        self.assertEqual(resp.status_code, 400)
        self.assertFalse(resp.get_json()["ok"])
        self.assertIn("10 supported RnD team colors", resp.get_json()["error"])


if __name__ == "__main__":
    unittest.main()
