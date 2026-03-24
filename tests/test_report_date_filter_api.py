from __future__ import annotations

import tempfile
import unittest
from time import sleep
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook

from report_server import create_report_server_app


def _build_app(root: Path):
    (root / "report_html").mkdir(parents=True, exist_ok=True)
    (root / "report_html" / "dashboard.html").write_text("<html><body>ok</body></html>", encoding="utf-8")
    (root / "executive_dashboard.html").write_text("<html><body>Executive</body><script src=\"/shared-nav.js\"></script></html>", encoding="utf-8")
    (root / "report_html" / "shared-nav.js").write_text("console.log('nav');", encoding="utf-8")
    (root / "report_html" / "shared-nav.css").write_text("body{}", encoding="utf-8")
    (root / "report_html" / "shared-date-filter.js").write_text("console.log('date-filter');", encoding="utf-8")

    wb = Workbook()
    ws = wb.active
    ws.append(["project_key", "worklog_date", "period_day", "period_week", "period_month", "issue_assignee", "hours_logged"])
    ws.append(["O2", "2026-02-01", "2026-02-01", "2026-W05", "2026-02", "Alice", 1.0])
    wb.save(root / "assignee_hours_report.xlsx")

    work_items = Workbook()
    wi = work_items.active
    wi.append(["project_key", "issue_key", "jira_issue_type", "parent_issue_key", "start_date", "end_date"])
    wi.append(["O2", "O2-EP1", "Epic", "", "2026-02-01", "2026-02-28"])
    wi.append(["O2", "O2-ST1", "Story", "O2-EP1", "", ""])
    wi.append(["O2", "O2-SUB1", "Sub-task", "O2-ST1", "", ""])
    work_items.save(root / "1_jira_work_items_export.xlsx")

    worklogs = Workbook()
    wl = worklogs.active
    wl.append(["issue_id", "parent_epic_id", "issue_assignee", "worklog_author", "worklog_started", "hours_logged"])
    wl.append(["O2-SUB1", "O2-EP1", "Alice", "Alice", "2026-02-10T10:00:00+0500", 4])
    worklogs.save(root / "2_jira_subtask_worklogs.xlsx")
    return create_report_server_app(base_dir=root, folder_raw="report_html")


class ReportDateFilterApiTests(unittest.TestCase):
    def test_planned_vs_dispensed_ui_settings_get_post(self):
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            app = _build_app(root)
            client = app.test_client()

            get_default = client.get("/api/planned-vs-dispensed/ui-settings")
            self.assertEqual(get_default.status_code, 200)
            default_payload = get_default.get_json()
            self.assertTrue(default_payload.get("ok"))
            self.assertEqual(default_payload["settings"]["first_column_width_px"], 420)

            bad = client.post(
                "/api/planned-vs-dispensed/ui-settings",
                json={"first_column_width_px": 120},
            )
            self.assertEqual(bad.status_code, 400)
            self.assertFalse(bad.get_json().get("ok"))

            save = client.post(
                "/api/planned-vs-dispensed/ui-settings",
                json={"first_column_width_px": 560},
            )
            self.assertEqual(save.status_code, 200)
            saved_payload = save.get_json()
            self.assertTrue(saved_payload.get("ok"))
            self.assertEqual(saved_payload["settings"]["first_column_width_px"], 560)

            get_saved = client.get("/api/planned-vs-dispensed/ui-settings")
            self.assertEqual(get_saved.status_code, 200)
            latest_payload = get_saved.get_json()
            self.assertTrue(latest_payload.get("ok"))
            self.assertEqual(latest_payload["settings"]["first_column_width_px"], 560)

    def test_get_post_report_date_filter(self):
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            app = _build_app(root)
            client = app.test_client()

            get_empty = client.get("/api/report-date-filter")
            self.assertEqual(get_empty.status_code, 200)
            self.assertEqual(get_empty.get_json(), {"ok": True, "filter": None})

            bad = client.post("/api/report-date-filter", json={"from_date": "2026-02-20", "to_date": "2026-02-10"})
            self.assertEqual(bad.status_code, 400)
            self.assertFalse(bad.get_json().get("ok"))

            save = client.post(
                "/api/report-date-filter",
                json={"from_date": "2026-02-01", "to_date": "2026-02-28", "source_page": "approved_vs_planned_hours_report"},
            )
            self.assertEqual(save.status_code, 200)
            saved_payload = save.get_json()
            self.assertTrue(saved_payload.get("ok"))
            self.assertEqual(saved_payload["filter"]["from_date"], "2026-02-01")
            self.assertEqual(saved_payload["filter"]["to_date"], "2026-02-28")
            self.assertEqual(saved_payload["filter"]["source_page"], "approved_vs_planned_hours_report")

            get_saved = client.get("/api/report-date-filter")
            self.assertEqual(get_saved.status_code, 200)
            latest = get_saved.get_json()
            self.assertTrue(latest.get("ok"))
            self.assertEqual(latest["filter"]["from_date"], "2026-02-01")
            self.assertEqual(latest["filter"]["to_date"], "2026-02-28")

    def test_fallback_range_used_when_from_to_missing(self):
        hierarchy = {
            "epics": [
                {
                    "issue_key": "O2-EP1",
                    "project_key": "O2",
                    "summary": "Epic One",
                    "status": "In Progress",
                    "assignee": "Alice",
                    "estimate_hours": 8.0,
                    "planned_start": "2026-02-01",
                    "planned_due": "2026-02-15",
                }
            ],
            "stories": [],
            "subtasks": [],
        }
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            app = _build_app(root)
            client = app.test_client()

            save = client.post(
                "/api/report-date-filter",
                json={"from_date": "2026-02-01", "to_date": "2026-02-28", "source_page": "dashboard"},
            )
            self.assertEqual(save.status_code, 200)

            actual_resp = client.get("/api/actual-hours/aggregate?mode=log_date&report=test")
            self.assertEqual(actual_resp.status_code, 409)
            actual_json = actual_resp.get_json()
            self.assertFalse(actual_json.get("ok"))
            self.assertIn("No successful canonical refresh found", actual_json.get("error", ""))

            with (
                patch("report_server.get_session", return_value=object()),
                patch("report_server._load_planned_vs_dispensed_hierarchy", return_value=hierarchy),
            ):
                summary_resp = client.get("/api/planned-vs-dispensed/summary?mode=log_date&projects=O2")
                self.assertEqual(summary_resp.status_code, 409)
                summary_json = summary_resp.get_json()
                self.assertFalse(summary_json.get("ok"))
                self.assertIn("No successful canonical refresh found", summary_json.get("error", ""))

            capacity_resp = client.get("/api/capacity")
            self.assertEqual(capacity_resp.status_code, 200)
            cap_json = capacity_resp.get_json()
            self.assertEqual(cap_json["settings"]["from_date"], "2026-02-01")
            self.assertEqual(cap_json["settings"]["to_date"], "2026-02-28")

    def test_partial_range_returns_400(self):
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            app = _build_app(root)
            client = app.test_client()
            partial = client.get("/api/actual-hours/aggregate?from=2026-02-01&mode=log_date")
            self.assertEqual(partial.status_code, 400)
            self.assertIn("provided together", partial.get_json().get("error", ""))

    def test_report_html_serves_shared_date_filter_script_tag(self):
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            app = _build_app(root)
            client = app.test_client()
            resp = client.get("/dashboard.html")
            self.assertEqual(resp.status_code, 200)
            html = resp.get_data(as_text=True)
            self.assertIn("shared-date-filter.js", html)
            exec_resp = client.get("/executive_dashboard.html")
            self.assertEqual(exec_resp.status_code, 200)
            self.assertIn("shared-date-filter.js", exec_resp.get_data(as_text=True))

    def test_dashboard_html_promotes_newer_root_source_before_serving(self):
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            app = _build_app(root)
            client = app.test_client()
            report_dashboard = root / "report_html" / "dashboard.html"
            report_dashboard.write_text("<html><body>old dashboard</body></html>", encoding="utf-8")
            sleep(1.1)
            (root / "dashboard.html").write_text("<html><body>new dashboard</body></html>", encoding="utf-8")

            resp = client.get("/dashboard.html")
            self.assertEqual(resp.status_code, 200)
            html = resp.get_data(as_text=True)
            self.assertIn("new dashboard", html)
            self.assertIn("no-store", resp.headers.get("Cache-Control", ""))
            self.assertIn("new dashboard", report_dashboard.read_text(encoding="utf-8"))

    def test_employee_performance_html_promotes_newer_root_source_before_serving(self):
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            app = _build_app(root)
            client = app.test_client()
            report_path = root / "report_html" / "employee_performance_report.html"
            report_path.write_text("<html><body>old employee report</body></html>", encoding="utf-8")
            sleep(1.1)
            (root / "employee_performance_report.html").write_text(
                "<html><body>new employee report</body></html>", encoding="utf-8"
            )

            resp = client.get("/employee_performance_report.html")
            self.assertEqual(resp.status_code, 200)
            self.assertIn("new employee report", resp.get_data(as_text=True))
            self.assertIn("new employee report", report_path.read_text(encoding="utf-8"))

    def test_report_html_keeps_current_copy_when_root_source_is_older(self):
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            (root / "dashboard.html").write_text("<html><body>older root</body></html>", encoding="utf-8")
            sleep(1.1)
            app = _build_app(root)
            client = app.test_client()
            report_dashboard = root / "report_html" / "dashboard.html"
            report_dashboard.write_text("<html><body>current published</body></html>", encoding="utf-8")

            resp = client.get("/dashboard.html")
            self.assertEqual(resp.status_code, 200)
            html = resp.get_data(as_text=True)
            self.assertIn("current published", html)
            self.assertNotIn("older root", html)

    def test_report_html_falls_back_when_only_published_copy_exists(self):
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as td:
            root = Path(td)
            app = _build_app(root)
            client = app.test_client()

            resp = client.get("/dashboard.html")
            self.assertEqual(resp.status_code, 200)
            self.assertIn("ok", resp.get_data(as_text=True))
            self.assertIn("no-store", resp.headers.get("Cache-Control", ""))


if __name__ == "__main__":
    unittest.main()
