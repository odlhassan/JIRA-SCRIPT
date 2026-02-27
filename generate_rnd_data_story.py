from __future__ import annotations

import argparse
import json
import os
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from generate_assignee_hours_report import (
    DEFAULT_CAPACITY_DB,
    DEFAULT_LEAVE_REPORT_INPUT_XLSX,
    _list_capacity_profiles,
    _load_leave_daily_rows,
)
from generate_nested_view_html import DEFAULT_INPUT_XLSX as DEFAULT_NESTED_VIEW_INPUT_XLSX, _load_nested_rows

DEFAULT_WORK_ITEMS_INPUT_XLSX = "1_jira_work_items_export.xlsx"
DEFAULT_WORKLOG_INPUT_XLSX = "2_jira_subtask_worklogs.xlsx"
DEFAULT_ASSIGNEE_HOURS_INPUT_XLSX = "assignee_hours_report.xlsx"
DEFAULT_HTML_OUTPUT = "rnd_data_story.html"
DEFAULT_DATASET_OUTPUT = "rnd_data_story_dataset.xlsx"


def _to_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _to_float(value: Any) -> float:
    text = _to_text(value)
    if not text:
        return 0.0
    try:
        out = float(text)
    except ValueError:
        return 0.0
    return round(out, 2)


def _resolve_path(raw: str, base_dir: Path) -> Path:
    path = Path(raw)
    if path.is_absolute():
        return path
    return base_dir / path


def _parse_iso_date(text: str) -> date | None:
    value = _to_text(text)
    if not value:
        return None
    try:
        return date.fromisoformat(value)
    except ValueError:
        return None


def epic_in_range(start_date: str, end_date: str, from_date: str, to_date: str) -> bool:
    from_day = _parse_iso_date(from_date)
    to_day = _parse_iso_date(to_date)
    if not from_day or not to_day:
        return False
    start_day = _parse_iso_date(start_date)
    end_day = _parse_iso_date(end_date)
    start_in = bool(start_day and from_day <= start_day <= to_day)
    end_in = bool(end_day and from_day <= end_day <= to_day)
    return start_in or end_in


def classify_status(status: str) -> str:
    text = _to_text(status).lower().replace("-", " ").replace("_", " ")
    if "resolved" in text:
        return "closed"
    if "in progress" in text:
        return "open"
    return "other"


def pending_hours(estimate_hours: float, epic_logged_hours: float) -> float:
    return round(max(float(estimate_hours or 0.0) - float(epic_logged_hours or 0.0), 0.0), 2)


def planned_committed_hours(epics: list[dict[str, Any]], from_date: str, to_date: str, excluded_project_key: str = "RLT") -> float:
    total = 0.0
    excluded = _to_text(excluded_project_key).upper()
    for epic in epics:
        if _to_text(epic.get("project_key")).upper() == excluded:
            continue
        if not epic_in_range(_to_text(epic.get("start_date")), _to_text(epic.get("end_date")), from_date, to_date):
            continue
        total += _to_float(epic.get("original_estimate_hours"))
    return round(total, 2)


def _load_epics(work_items_path: Path) -> list[dict[str, Any]]:
    if not work_items_path.exists():
        raise FileNotFoundError(f"Work items file not found: {work_items_path}")

    wb = load_workbook(work_items_path, read_only=True, data_only=True)
    try:
        ws = wb.active
        header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header:
            raise ValueError("Work items workbook has no header row.")
        headers = [_to_text(h) for h in header]
        idx = {name: i for i, name in enumerate(headers)}
        required = ["issue_key", "summary", "status", "start_date", "end_date", "original_estimate_hours", "jira_issue_type"]
        missing = [name for name in required if name not in idx]
        if missing:
            raise ValueError("Work items workbook missing required columns: " + ", ".join(missing))

        out: list[dict[str, Any]] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            issue_type = _to_text(row[idx["jira_issue_type"]]).lower()
            if "epic" not in issue_type:
                continue
            issue_key = _to_text(row[idx["issue_key"]]).upper()
            if not issue_key:
                continue
            out.append(
                {
                    "issue_key": issue_key,
                    "project_key": (_to_text(row[idx["project_key"]]).upper() if "project_key" in idx else issue_key.split("-", 1)[0]),
                    "summary": _to_text(row[idx["summary"]]),
                    "status": _to_text(row[idx["status"]]),
                    "start_date": _to_text(row[idx["start_date"]]),
                    "end_date": _to_text(row[idx["end_date"]]),
                    "original_estimate_hours": _to_float(row[idx["original_estimate_hours"]]),
                }
            )
        return out
    finally:
        wb.close()


def _load_subtask_worklogs_for_epics(worklog_path: Path) -> list[dict[str, Any]]:
    if not worklog_path.exists():
        raise FileNotFoundError(f"Subtask worklog file not found: {worklog_path}")

    wb = load_workbook(worklog_path, read_only=True, data_only=True)
    try:
        ws = wb.active
        header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header:
            raise ValueError("Subtask worklog workbook has no header row.")
        headers = [_to_text(h) for h in header]
        idx = {name: i for i, name in enumerate(headers)}
        required = ["parent_epic_id", "hours_logged"]
        missing = [name for name in required if name not in idx]
        if missing:
            raise ValueError("Subtask worklog workbook missing required columns: " + ", ".join(missing))

        out: list[dict[str, Any]] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            parent_epic_id = _to_text(row[idx["parent_epic_id"]]).upper()
            if not parent_epic_id:
                continue
            out.append({"parent_epic_id": parent_epic_id, "hours_logged": _to_float(row[idx["hours_logged"]])})
        return out
    finally:
        wb.close()


def aggregate_epic_logged_hours(worklog_rows: list[dict[str, Any]]) -> dict[str, float]:
    totals: dict[str, float] = {}
    for row in worklog_rows:
        key = _to_text(row.get("parent_epic_id")).upper()
        if not key:
            continue
        totals[key] = round(totals.get(key, 0.0) + _to_float(row.get("hours_logged")), 2)
    return totals


def _load_assignee_worklog_rows(summary_path: Path) -> list[dict[str, Any]]:
    if not summary_path.exists():
        raise FileNotFoundError(f"Assignee hours file not found: {summary_path}")

    wb = load_workbook(summary_path, read_only=True, data_only=True)
    try:
        if "AssigneeHours" not in wb.sheetnames:
            raise ValueError("Assignee hours workbook missing 'AssigneeHours' sheet.")
        ws = wb["AssigneeHours"]
        header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header:
            raise ValueError("AssigneeHours sheet has no header row.")
        headers = [_to_text(h) for h in header]
        idx = {name: i for i, name in enumerate(headers)}
        required = ["project_key", "worklog_date", "hours_logged"]
        missing = [name for name in required if name not in idx]
        if missing:
            raise ValueError("AssigneeHours sheet missing required columns: " + ", ".join(missing))

        out: list[dict[str, Any]] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            worklog_date = _to_text(row[idx["worklog_date"]])
            if not _parse_iso_date(worklog_date):
                continue
            hours_logged = _to_float(row[idx["hours_logged"]])
            if hours_logged <= 0:
                continue
            assignee = "Unassigned"
            if "issue_assignee" in idx:
                assignee = _to_text(row[idx["issue_assignee"]]) or "Unassigned"
            out.append({"project_key": _to_text(row[idx["project_key"]]).upper() or "UNKNOWN", "worklog_date": worklog_date, "hours_logged": hours_logged, "issue_assignee": assignee})
        return out
    finally:
        wb.close()


def _load_planned_epic_rows(nested_view_path: Path) -> list[dict[str, Any]]:
    if not nested_view_path.exists():
        return []
    out: list[dict[str, Any]] = []
    for row in _load_nested_rows(nested_view_path):
        if _to_text(row.get("row_type")) != "rmi":
            continue
        man_hours = _to_float(row.get("man_hours"))
        if man_hours <= 0:
            man_hours = round(_to_float(row.get("man_days")) * 8.0, 2)
        out.append(
            {
                "project_key": _to_text(row.get("project_key")).upper(),
                "planned_start": _to_text(row.get("planned_start")),
                "planned_end": _to_text(row.get("planned_end")),
                "planned_hours": man_hours,
            }
        )
    return out


def _load_project_actual_rows(nested_view_path: Path) -> list[dict[str, Any]]:
    if not nested_view_path.exists():
        return []
    out: list[dict[str, Any]] = []
    for row in _load_nested_rows(nested_view_path):
        if _to_text(row.get("row_type")) != "project":
            continue
        out.append(
            {
                "project_key": _to_text(row.get("project_key")).upper(),
                "actual_hours": _to_float(row.get("actual_hours")),
            }
        )
    return out


def _default_range(worklog_rows: list[dict[str, Any]]) -> tuple[str, str]:
    now = datetime.now(timezone.utc).date()
    current_start = date(now.year, now.month, 1)
    if now.month == 12:
        next_month_start = date(now.year + 1, 1, 1)
    else:
        next_month_start = date(now.year, now.month + 1, 1)
    current_end = next_month_start - timedelta(days=1)

    if now.month == 1:
        prev_start = date(now.year - 1, 12, 1)
    else:
        prev_start = date(now.year, now.month - 1, 1)

    valid = [d for d in (_parse_iso_date(_to_text(r.get("worklog_date"))) for r in worklog_rows) if d is not None]
    if valid:
        min_row_date = min(valid)
        max_row_date = max(valid)
        start = max(prev_start, min_row_date)
        end = min(current_end, max_row_date)
        if end < start:
            start, end = min_row_date, max_row_date
        return start.isoformat(), end.isoformat()

    return prev_start.isoformat(), current_end.isoformat()


def _benchmark_hours_by_due_date(planned_hours: float, from_date: str, due_date: str, as_of: date | None = None) -> float:
    due_day = _parse_iso_date(due_date)
    from_day = _parse_iso_date(from_date)
    if planned_hours <= 0:
        return 0.0
    if not due_day or not from_day:
        return round(planned_hours, 2)
    if due_day <= from_day:
        return round(planned_hours, 2)
    current_day = as_of or datetime.now(timezone.utc).date()
    if current_day <= from_day:
        return 0.0
    elapsed_days = (min(current_day, due_day) - from_day).days + 1
    total_days = (due_day - from_day).days + 1
    ratio = max(0.0, min(float(elapsed_days) / float(total_days), 1.0))
    return round(planned_hours * ratio, 2)


def _project_compare_rows(
    epics: list[dict[str, Any]],
    planned_epic_rows: list[dict[str, Any]],
    project_actual_rows: list[dict[str, Any]],
    from_date: str,
    to_date: str,
) -> list[dict[str, Any]]:
    planned_by_project: dict[str, float] = {}
    actual_by_project: dict[str, float] = {}
    due_date_by_project: dict[str, str] = {}

    for row in planned_epic_rows:
        project_key = _to_text(row.get("project_key")).upper()
        if not project_key or project_key == "RLT":
            continue
        planned_start = _to_text(row.get("planned_start"))
        planned_end = _to_text(row.get("planned_end"))
        if not epic_in_range(planned_start, planned_end, from_date, to_date):
            continue
        planned_by_project[project_key] = round(planned_by_project.get(project_key, 0.0) + _to_float(row.get("planned_hours")), 2)

    for row in project_actual_rows:
        project_key = _to_text(row.get("project_key")).upper()
        if not project_key or project_key == "RLT":
            continue
        actual_by_project[project_key] = round(actual_by_project.get(project_key, 0.0) + _to_float(row.get("actual_hours")), 2)

    for epic in epics:
        project_key = _to_text(epic.get("project_key")).upper()
        if not project_key or project_key == "RLT":
            continue
        start_date = _to_text(epic.get("start_date"))
        end_date = _to_text(epic.get("end_date"))
        if not epic_in_range(start_date, end_date, from_date, to_date):
            continue
        due_day = _parse_iso_date(end_date)
        if not due_day:
            continue
        previous = _parse_iso_date(due_date_by_project.get(project_key, ""))
        if previous is None or due_day < previous:
            due_date_by_project[project_key] = due_day.isoformat()

    output: list[dict[str, Any]] = []
    for project_key in sorted(set(planned_by_project) | set(actual_by_project)):
        planned_hours = round(planned_by_project.get(project_key, 0.0), 2)
        actual_hours = round(actual_by_project.get(project_key, 0.0), 2)
        hours_required = round(planned_hours - actual_hours, 2)
        due_date = _to_text(due_date_by_project.get(project_key))
        benchmark_hours = _benchmark_hours_by_due_date(planned_hours, from_date, due_date)
        completion_pct = round((actual_hours / planned_hours) * 100, 2) if planned_hours > 0 else 0.0
        output.append(
            {
                "project_key": project_key,
                "planned_hours": planned_hours,
                "actual_hours": actual_hours,
                "hours_required_to_complete": hours_required,
                "benchmark_hours_due_date": benchmark_hours,
                "project_due_date": due_date,
                "completion_pct": completion_pct,
            }
        )

    output.sort(key=lambda item: (float(item.get("planned_hours") or 0.0), str(item.get("project_key") or "")), reverse=True)
    return output


def _build_page1_dataset(
    epics: list[dict[str, Any]],
    defaults: dict[str, str],
    planned_epic_rows: list[dict[str, Any]],
    project_actual_rows: list[dict[str, Any]],
) -> dict[str, Any]:
    from_date = _to_text(defaults.get("from_date"))
    to_date = _to_text(defaults.get("to_date"))
    compare_rows = _project_compare_rows(epics, planned_epic_rows, project_actual_rows, from_date, to_date)
    total_planned = round(sum(_to_float(row.get("planned_hours")) for row in compare_rows), 2)
    total_actual = round(sum(_to_float(row.get("actual_hours")) for row in compare_rows), 2)
    total_required = round(total_planned - total_actual, 2)
    return {
        "page_name": "Total work by projects",
        "date_range": {"from_date": from_date, "to_date": to_date},
        "scorecard": {
            "capacity_hours": 0.0,
            "planned_hours": total_planned,
            "actual_hours": total_actual,
            "hours_required_to_complete_projects": total_required,
        },
        "project_compare_rows": compare_rows,
    }


def _write_page1_dataset_xlsx(output_path: Path, dataset: dict[str, Any]) -> None:
    wb = Workbook()
    ws_score = wb.active
    ws_score.title = "Page1_Scorecard"
    ws_score.append(["page_name", "from_date", "to_date", "capacity_hours", "planned_hours", "actual_hours", "hours_required_to_complete_projects"])
    date_range = dataset.get("date_range") or {}
    scorecard = dataset.get("scorecard") or {}
    ws_score.append(
        [
            _to_text(dataset.get("page_name")),
            _to_text(date_range.get("from_date")),
            _to_text(date_range.get("to_date")),
            _to_float(scorecard.get("capacity_hours")),
            _to_float(scorecard.get("planned_hours")),
            _to_float(scorecard.get("actual_hours")),
            _to_float(scorecard.get("hours_required_to_complete_projects")),
        ]
    )

    ws_compare = wb.create_sheet("Page1_ProjectCompare")
    ws_compare.append(["project_key", "planned_hours", "actual_hours", "hours_required_to_complete", "benchmark_hours_due_date", "project_due_date", "completion_pct"])
    for row in dataset.get("project_compare_rows") or []:
        ws_compare.append(
            [
                _to_text(row.get("project_key")),
                _to_float(row.get("planned_hours")),
                _to_float(row.get("actual_hours")),
                _to_float(row.get("hours_required_to_complete")),
                _to_float(row.get("benchmark_hours_due_date")),
                _to_text(row.get("project_due_date")),
                _to_float(row.get("completion_pct")),
            ]
        )

    wb.save(output_path)


def _build_payload(paths: dict[str, Path]) -> dict[str, Any]:
    epics = _load_epics(paths["work_items_path"])
    subtask_logs = _load_subtask_worklogs_for_epics(paths["worklog_path"])
    epic_logged_hours_by_key = aggregate_epic_logged_hours(subtask_logs)
    worklog_rows = _load_assignee_worklog_rows(paths["assignee_hours_path"])
    planned_epic_rows = _load_planned_epic_rows(paths["nested_view_path"])
    project_actual_rows = _load_project_actual_rows(paths["nested_view_path"])
    default_from, default_to = _default_range(worklog_rows)
    defaults = {"from_date": default_from, "to_date": default_to}
    page1_dataset = _build_page1_dataset(epics, defaults, planned_epic_rows, project_actual_rows)
    assignee_names = {(_to_text(r.get("issue_assignee")) or "Unassigned") for r in worklog_rows}
    assignee_names = {name for name in assignee_names if name and name.lower() != "unassigned"}
    default_employee_count = len(assignee_names)
    return {
        "department_name": "Research and Development (RnD)",
        "generated_at": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC"),
        "source_files": {
            "work_items": str(paths["work_items_path"]),
            "subtask_worklogs": str(paths["worklog_path"]),
            "assignee_hours": str(paths["assignee_hours_path"]),
            "leave_report": str(paths["leave_report_path"]),
            "capacity_db": str(paths["capacity_db_path"]),
            "page1_dataset": str(paths["output_dataset_path"]),
        },
        "defaults": defaults,
        "default_employee_count": default_employee_count,
        "epics": epics,
        "epic_logged_hours_by_key": epic_logged_hours_by_key,
        "worklog_rows": worklog_rows,
        "planned_epic_rows": planned_epic_rows,
        "project_actual_rows": project_actual_rows,
        "page1_dataset": page1_dataset,
        "capacity_profiles": _list_capacity_profiles(paths["capacity_db_path"]),
        "leave_daily_rows": _load_leave_daily_rows(paths["leave_report_path"]),
    }


def _build_html(payload: dict[str, Any]) -> str:
    data = json.dumps(payload, ensure_ascii=True)
    template = """<!doctype html>
<html lang="en"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1"><title>RnD Capacity and Workload Data Story</title><link rel="stylesheet" href="https://fonts.googleapis.com/css2?family=Material+Symbols+Outlined:opsz,wght,FILL,GRAD@20..48,500,0,0">
<style>
:root{--bg:#f2f7f7;--panel:#fff;--line:#d0dde0;--text:#0f172a;--kpi-capacity-bg:#fee2e2;--kpi-capacity-line:#fca5a5;--kpi-capacity-ink:#7f1d1d;--kpi-planned-bg:#dbeafe;--kpi-planned-line:#93c5fd;--kpi-planned-ink:#1e3a8a;--kpi-actual-bg:#f3e8ff;--kpi-actual-line:#d8b4fe;--kpi-actual-ink:#a855f7;--kpi-gap-bg:#e0f2fe;--kpi-gap-line:#7dd3fc;--kpi-gap-ink:#2563eb;--kpi-leaves-bg:#fecaca;--kpi-leaves-line:#f87171;--kpi-leaves-ink:#b91c1c;--kpi-capacity-gap-bg:#fef2f2;--kpi-capacity-gap-line:#fecaca;--kpi-capacity-gap-ink:#ef4444;font-family:"Segoe UI",Tahoma,sans-serif;color:var(--text);background:linear-gradient(180deg,#f7fbfb,var(--bg))}.page{max-width:1400px;margin:0 auto;padding:16px}.panel{background:var(--panel);border:1px solid var(--line);border-radius:12px;padding:12px;margin-bottom:12px}.hero{background:linear-gradient(130deg,#083344,#0f766e);color:#eaffff}.hero h1{margin:0;font-size:1.28rem}.muted{color:#d7f5f5;font-size:.84rem;margin-top:4px}.filters-toolbar{border:1px solid #d7e3e8;border-radius:10px;background:#fbfeff;padding:8px}.filters{display:grid;gap:8px;grid-template-columns:minmax(320px,1.9fr) minmax(240px,1.2fr) auto;align-items:end}.filter-field{padding:6px 8px;border:1px solid #dbe7eb;border-radius:8px;background:#ffffff}.filters label{display:block;font-size:.62rem;font-weight:700;color:#4b6470;margin-bottom:4px;letter-spacing:.05em;text-transform:uppercase}.date-range-control{display:grid;grid-template-columns:minmax(0,1fr) auto minmax(0,1fr);gap:6px;align-items:center}.range-sep{font-size:.65rem;font-weight:700;color:#64748b;text-transform:uppercase;letter-spacing:.03em}.filter-control{position:relative}.filter-control input[type=date],.filter-control select{width:100%;border:1px solid #c8d8df;border-radius:7px;padding:7px 30px 7px 9px;font-size:.82rem;font-weight:600;color:#0f172a;background:#ffffff;transition:border-color .12s ease,box-shadow .12s ease;outline:none}.filter-control input[type=date]:hover,.filter-control select:hover{border-color:#9fb8c2}.filter-control input[type=date]:focus,.filter-control select:focus{border-color:#0e7490;box-shadow:0 0 0 2px rgba(14,116,144,.12)}.filter-control.select::after{content:"";position:absolute;right:10px;top:50%;width:7px;height:7px;border-right:2px solid #4b6470;border-bottom:2px solid #4b6470;transform:translateY(-60%) rotate(45deg);pointer-events:none}.btns{display:flex;gap:6px;align-items:flex-end;flex-wrap:wrap}.btn{border:1px solid #0f766e;border-radius:7px;background:#0f766e;color:#f0fdff;padding:7px 11px;font-size:.78rem;font-weight:700;cursor:pointer;box-shadow:none}.btn:hover{filter:brightness(.98)}.btn.alt{background:#ffffff;color:#0f2d3a;border-color:#c3d4db}@media (max-width:980px){.filters{grid-template-columns:1fr}}.story-nav{display:flex;align-items:center;justify-content:space-between;gap:10px}.nav-btn{width:42px;height:42px;border-radius:50%;border:1px solid #a9bdc5;background:#fff;color:#0f172a;font-size:1.2rem;font-weight:700;cursor:pointer}.nav-btn:disabled{opacity:.4;cursor:not-allowed}.nav-label{font-weight:700;color:#334155;font-size:.9rem}.story-page{display:none}.story-page.active{display:block}.page-title{margin:0 0 10px;color:#334155}.kpis{display:grid;gap:8px;grid-template-columns:repeat(3,minmax(0,1fr))}.kpi{border:1px solid #d7e4e8;border-radius:10px;padding:10px;background:linear-gradient(180deg,#fff,#f8fbfc);position:relative}.kpi .label{font-size:.72rem;text-transform:uppercase;letter-spacing:.03em;color:#475569;font-weight:700}.kpi .value{font-size:1.3rem;font-weight:800;margin-top:6px}.leadership-cards{display:grid;gap:8px;grid-template-columns:repeat(auto-fit,minmax(240px,1fr));margin-bottom:10px}.kpi.metric-capacity{background:var(--kpi-capacity-bg);border-color:var(--kpi-capacity-line)}.kpi.metric-capacity .value{color:var(--kpi-capacity-ink)}.kpi.metric-planned{background:var(--kpi-planned-bg);border-color:var(--kpi-planned-line)}.kpi.metric-planned .value{color:var(--kpi-planned-ink)}.kpi.metric-actual{background:var(--kpi-actual-bg);border-color:var(--kpi-actual-line)}.kpi.metric-actual .value{color:var(--kpi-actual-ink)}.kpi.metric-required{background:var(--kpi-gap-bg);border-color:var(--kpi-gap-line)}.kpi.metric-required .value{color:var(--kpi-gap-ink)}.kpi.metric-available{background:var(--kpi-capacity-gap-bg);border-color:var(--kpi-capacity-gap-line)}.kpi.metric-available .value{color:var(--kpi-capacity-gap-ink)}.metric-bar-wrap{margin-top:8px;height:8px;border-radius:999px;background:rgba(148,163,184,.28);overflow:hidden}.metric-bar{height:100%;width:0%;transition:width .35s ease}.metric-bar.capacity{background:#ef4444}.metric-bar.planned{background:#3b82f6}.metric-bar.actual{background:#a855f7}.metric-bar.required{background:#0ea5e9}.metric-bar.available{background:#ef4444}.leadership-mermaid-card{margin-bottom:10px}.leadership-mermaid-wrap{border:1px solid #d7e4e8;border-radius:10px;background:linear-gradient(180deg,#ffffff,#f8fbfc);padding:8px;overflow:auto}.leadership-mermaid-wrap .mermaid{min-width:760px}.leadership-mermaid-note{font-size:.75rem;color:#64748b;margin-top:6px}.leadership-flow{display:grid;grid-template-columns:minmax(0,1fr) 42px minmax(0,1fr) 42px minmax(0,1fr);grid-template-rows:auto auto;gap:12px 10px;align-items:stretch}.leadership-flow>*{min-width:0}.leadership-flow>:nth-child(1){grid-column:1;grid-row:1}.leadership-flow>:nth-child(2){grid-column:2;grid-row:1}.leadership-flow>:nth-child(3){grid-column:3;grid-row:1}.leadership-flow>:nth-child(4){grid-column:4;grid-row:1}.leadership-flow>:nth-child(5){grid-column:5;grid-row:1}.leadership-flow>:nth-child(6){grid-column:5;grid-row:2}.leadership-flow>:nth-child(7){grid-column:4;grid-row:2}.leadership-flow>:nth-child(8){grid-column:3;grid-row:2}.leadership-flow>:nth-child(9){grid-column:2;grid-row:2}.flow-card{display:flex;flex-direction:column;justify-content:center;min-height:102px}.flow-arrow{display:flex;align-items:center;justify-content:center;position:relative}.flow-arrow::before{content:"";position:absolute;left:6px;right:6px;top:50%;height:6px;transform:translateY(-50%);background:#0f766e;border-radius:999px;box-shadow:0 0 0 1px rgba(255,255,255,.45) inset}.flow-arrow span{position:relative;z-index:1;font-size:1.9rem;font-weight:900;line-height:1;color:#0f766e;text-shadow:0 0 0 #0f766e}.leadership-flow>.flow-arrow:nth-child(6)::before{left:50%;right:auto;top:6px;bottom:6px;width:6px;height:auto;transform:translateX(-50%)}.leadership-flow>.flow-arrow:nth-child(6) span{transform:rotate(90deg)}.leadership-flow>.flow-arrow:nth-child(8) span{transform:rotate(180deg)}.flow-card.capacity{background:var(--kpi-capacity-bg);border-color:var(--kpi-capacity-line)}.flow-card.capacity .value{color:var(--kpi-capacity-ink)}.flow-card.planned{background:var(--kpi-planned-bg);border-color:var(--kpi-planned-line)}.flow-card.planned .value{color:var(--kpi-planned-ink)}.flow-card.actual{background:var(--kpi-actual-bg);border-color:var(--kpi-actual-line)}.flow-card.actual .value{color:var(--kpi-actual-ink)}.flow-card.required{background:var(--kpi-gap-bg);border-color:var(--kpi-gap-line)}.flow-card.required .value{color:var(--kpi-gap-ink)}.flow-card.available{background:var(--kpi-capacity-gap-bg);border-color:var(--kpi-capacity-gap-line)}.flow-card.available .value{color:var(--kpi-capacity-gap-ink)}.kpi-epic-status{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:8px;margin-top:10px}.kpi-epic-status .kpi.closed{background:#dcfce7;border-color:#86efac}.kpi-epic-status .kpi.closed .value{color:#166534}.kpi-epic-status .kpi.open{background:#ffedd5;border-color:#fdba74}.kpi-epic-status .kpi.open .value{color:#9a3412}.barrel-enterprise{display:grid;grid-template-columns:1fr;gap:12px;align-items:stretch}.barrel-canvas-wrap{position:relative;width:min(70vw,100%);max-width:70vw;margin:0 auto;height:min(72vh,760px);min-height:520px;border:1px solid #d9e6ea;border-radius:12px;background:radial-gradient(circle at 18% 10%,#ffffff 0,#f4f9fb 52%,#ecf3f7 100%);padding:14px;box-shadow:inset 0 0 0 1px rgba(255,255,255,.65)}.barrel-chart{display:block;width:100%;height:100%}.barrel-stat-grid{display:grid;gap:8px;grid-template-columns:repeat(auto-fit,minmax(220px,1fr))}.barrel-stat{border:1px solid #d7e4e8;border-radius:10px;background:#f8fbfc;padding:10px}.barrel-stat .k{display:block;font-size:.72rem;color:#64748b;font-weight:700;letter-spacing:.01em}.barrel-stat .v{display:block;font-size:1rem;font-weight:800;color:#0f172a;margin-top:4px}.barrel-stat .formula{margin-top:6px}.story-grid{display:grid;gap:10px;grid-template-columns:repeat(auto-fit,minmax(280px,1fr))}.story-card{border:1px solid #d7e4e8;border-radius:10px;padding:10px;position:relative}.story-title{margin:0 0 8px;font-size:.88rem;color:#334155}.bar-wrap{background:#edf3f5;border:1px solid #d3e0e4;border-radius:999px;height:14px;overflow:hidden}.bar{height:100%}.a{background:#0e7490}.b{background:#f59e0b}.c{background:#16a34a}.d{background:#7c3aed}.legend{margin-top:8px;font-size:.78rem;color:#334155;display:flex;justify-content:space-between;gap:8px}.funnel{display:grid;gap:10px;margin-top:6px}.frow{display:grid;grid-template-columns:220px 1fr;gap:10px;align-items:center}.flabel{font-size:.78rem;color:#334155;font-weight:700}.ftrack{position:relative;height:22px;background:#edf3f5;border:1px solid #d3e0e4;border-radius:999px;overflow:hidden}.fbar{position:absolute;left:50%;transform:translateX(-50%);height:100%;border-radius:999px}.f-available{background:#0e7490}.f-after{background:#3b82f6}.f-booked{background:#f59e0b}.f-remaining{background:#16a34a}.f-required{background:#dc2626}.fval{position:absolute;inset:0;display:flex;align-items:center;justify-content:center;font-size:.74rem;font-weight:800;color:#f8fafc;text-shadow:0 1px 1px rgba(15,23,42,.55)}#funnel-hours-required-track{overflow:visible}#funnel-hours-required-val{inset:auto;left:calc(100% + 8px);top:50%;transform:translateY(-50%);display:inline-flex;align-items:center;justify-content:center;height:20px;min-width:52px;padding:0 8px;border-radius:999px;border:1px solid #fecaca;background:#fff;color:#991b1b;text-shadow:none}.fnote{margin-top:8px;font-size:.8rem;color:#334155}.compare-grid{display:grid;gap:10px}.compare-item{border:1px solid #d7e4e8;border-radius:10px;padding:8px;position:relative}.compare-head{display:flex;justify-content:space-between;font-size:.78rem;font-weight:700;color:#334155;margin-bottom:6px}.formula{display:block;margin-top:4px;font-size:.68rem;color:#64748b}.ratio-wrap{display:flex;gap:12px;align-items:center;margin-top:10px}.ratio-ring{width:88px;height:88px;border-radius:50%;display:grid;place-items:center;background:conic-gradient(#16a34a 0%, #e2e8f0 0)}.ratio-ring span{width:64px;height:64px;border-radius:50%;background:#fff;display:grid;place-items:center;font-size:.84rem;font-weight:800;color:#0f172a}.ratio-text{font-size:.82rem;color:#334155}.breakdown{display:grid;grid-template-columns:repeat(auto-fit,minmax(150px,1fr));gap:8px;margin-top:10px}.break-cell{border:1px solid #d7e4e8;border-radius:8px;padding:8px;background:#f8fbfc;position:relative}.break-cell .k{display:block;font-size:.72rem;color:#64748b}.break-cell .v{display:block;font-size:.92rem;font-weight:800;color:#0f172a}.insights{display:grid;gap:8px;grid-template-columns:repeat(auto-fit,minmax(260px,1fr))}.insight{border-radius:10px;padding:10px;border:1px solid #d8e4e8;background:#f8fbfc;position:relative}.insight.ok{border-color:#bbf7d0;background:#f0fdf4}.insight.warn{border-color:#fde68a;background:#fffbeb}.insight.risk{border-color:#fecaca;background:#fef2f2}.insight h3{margin:0;font-size:.88rem}.insight p{margin:6px 0 0;font-size:.86rem;color:#334155}.profile-status{font-size:.8rem;color:#334155;margin-top:6px}.vchart{display:grid;gap:8px}.vrow{display:grid;grid-template-columns:minmax(120px,2fr) 6fr minmax(60px,1fr);gap:8px;align-items:center}.vlabel{font-size:.78rem;color:#334155;font-weight:700;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}.vbar-shell{height:12px;border-radius:999px;background:#edf3f5;border:1px solid #d3e0e4;overflow:hidden}.vbar{height:100%;background:linear-gradient(90deg,#2563eb,#3b82f6)}.vvalue{font-size:.76rem;color:#1f2937;text-align:right}.summary-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(260px,1fr));gap:10px}.summary-card{border:1px solid #d7e4e8;border-radius:10px;padding:12px;background:#f8fbfc;position:relative}.summary-card h4{margin:0 0 8px}.summary-card p{margin:0;color:#334155}.card-i-wrap{position:absolute;top:8px;right:8px;display:inline-flex;align-items:center;justify-content:center;width:18px;height:18px;border-radius:999px;border:1px solid #94a3b8;background:#f8fafc;color:#334155;font-size:11px;font-weight:700;cursor:default;z-index:5}.card-i-tip{position:absolute;top:calc(100% + 8px);right:0;min-width:210px;max-width:280px;padding:8px 10px;border-radius:8px;border:1px solid #cbd5e1;background:#fff;color:#1f2937;font-size:12px;line-height:1.35;white-space:pre-line;box-shadow:0 10px 24px rgba(15,23,42,.18);opacity:0;visibility:hidden;pointer-events:none;transition:opacity .14s ease}.card-i-wrap:hover .card-i-tip{opacity:1;visibility:visible}@media (max-width:1200px){.leadership-flow{grid-template-columns:1fr;grid-template-rows:none;gap:8px}.leadership-flow>:nth-child(n){grid-column:auto;grid-row:auto}.flow-arrow{min-height:30px}.flow-arrow::before{left:50%;right:auto;top:6px;bottom:6px;width:6px;height:auto;transform:translateX(-50%)}.flow-arrow span{transform:rotate(90deg)}.leadership-flow>.flow-arrow:nth-child(8) span{transform:rotate(90deg)}.barrel-canvas-wrap{width:min(70vw,100%);max-width:70vw;height:min(64vh,620px);min-height:420px}}@media (max-width:760px){.kpi-epic-status{grid-template-columns:1fr}}
</style><style>
.range-field{display:grid;gap:8px}
.date-presets{display:flex;flex-wrap:wrap;gap:6px}
.chip{border:1px solid #bfced5;border-radius:999px;background:#fff;color:#164e63;padding:4px 10px;font-size:.7rem;font-weight:700;letter-spacing:.01em;cursor:pointer;transition:border-color .12s ease,background .12s ease,color .12s ease,box-shadow .12s ease}
.chip:hover{border-color:#0e7490;background:#f0f9ff}
.chip.is-active{border-color:#0e7490;background:#0e7490;color:#ecfeff;box-shadow:0 0 0 2px rgba(14,116,144,.12)}
.material-symbols-outlined{font-variation-settings:"FILL" 0,"wght" 500,"GRAD" 0,"opsz" 24;font-size:1rem;line-height:1;vertical-align:-0.14em}
.btn .material-symbols-outlined{font-size:.95rem}
.nav-btn{display:inline-flex;align-items:center;justify-content:center}
.date-hint{font-size:.73rem;color:#475569}
.date-hint.ok{color:#0f766e}
.date-hint.warn{color:#b45309}
.btn[disabled]{opacity:.72;cursor:wait}
.filters-toolbar.is-dirty{border-color:#99b8c6;box-shadow:0 0 0 2px rgba(14,116,144,.08) inset}
.project-compare-chart{display:grid;gap:12px}
.bc-row{display:grid;grid-template-columns:minmax(120px,1.5fr) minmax(120px,1.3fr) minmax(0,6fr) minmax(180px,1.7fr);gap:12px;align-items:center}
.bc-head{font-size:.7rem;font-weight:800;color:#475569;text-transform:uppercase;letter-spacing:.04em;padding-bottom:2px;border-bottom:1px dashed #d3e0e4}
.bc-label{font-size:.76rem;color:#334155;font-weight:700;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.bc-due{font-size:.74rem;color:#334155;font-weight:700;white-space:nowrap}
.bc-track{position:relative;height:20px;border-radius:999px;background:#edf3f5;border:1px solid #d3e0e4;overflow:visible;margin-top:8px;margin-bottom:8px}
.bc-actual{position:absolute;left:0;top:0;bottom:0;border-radius:999px;background:linear-gradient(90deg,#7e22ce,#a855f7)}
.bc-benchmark{position:absolute;top:-4px;bottom:-4px;width:2px;background:#dc2626;border-radius:2px}
.bc-bar-label{position:absolute;top:-28px;right:0;font-size:.68rem;font-weight:800;color:#6b21a8;background:#fff;border:1px solid #e9d5ff;border-radius:999px;padding:1px 6px;line-height:1.2}
.bc-benchmark-label{position:absolute;top:24px;transform:translateX(-50%);font-size:.68rem;font-weight:800;color:#991b1b;background:#fff;border:1px solid #fecaca;border-radius:999px;padding:1px 6px;line-height:1.2;white-space:nowrap}
.bc-values{font-size:.72rem;color:#1f2937;text-align:right;font-weight:700}
.kpi-sub{font-size:.72rem;color:#64748b;font-weight:700;margin-top:4px}
.monthly-stack-chart{display:grid;gap:12px}
.ms-head,.ms-row{display:grid;grid-template-columns:minmax(120px,1.4fr) minmax(0,6fr) minmax(240px,2.1fr);gap:12px;align-items:center}
.ms-head{font-size:.7rem;font-weight:800;color:#475569;text-transform:uppercase;letter-spacing:.04em;padding-bottom:4px;border-bottom:1px dashed #d3e0e4}
.ms-project{font-size:.76rem;color:#334155;font-weight:700;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.ms-track{position:relative;height:22px;border-radius:999px;background:#edf3f5;border:1px solid #d3e0e4;overflow:hidden}
.ms-bar{height:100%;display:flex;border-radius:999px;overflow:hidden}
.ms-seg{height:100%;display:flex;align-items:center;justify-content:center;font-size:.66rem;font-weight:800;color:#f8fafc;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.ms-prev{background:#0f766e}
.ms-current{background:#2563eb}
.ms-upcoming{background:#9333ea}
.ms-values{font-size:.72rem;color:#1f2937;text-align:right;font-weight:700}
.ms-legend{display:flex;gap:8px;flex-wrap:wrap;font-size:.72rem;color:#334155;margin-top:6px}
.ms-pill{display:inline-flex;align-items:center;gap:6px;padding:2px 8px;border-radius:999px;background:#f8fafc;border:1px solid #dbe7eb}
.ms-dot{width:8px;height:8px;border-radius:999px;display:inline-block}
.ms-dot.prev{background:#0f766e}.ms-dot.current{background:#2563eb}.ms-dot.upcoming{background:#9333ea}
.project-scorecards{display:grid;grid-template-columns:repeat(auto-fit,minmax(170px,1fr));gap:8px}
.project-scorecard{border:1px solid #d7e4e8;border-radius:10px;background:#f8fbfc;padding:10px;cursor:pointer;transition:border-color .12s ease,transform .12s ease,box-shadow .12s ease;position:relative}
.project-scorecard:hover{border-color:#0e7490;transform:translateY(-1px)}
.project-scorecard.active{border-color:#0e7490;box-shadow:0 0 0 2px rgba(14,116,144,.12);background:#ecfeff}
.project-scorecard .k{display:block;font-size:.72rem;font-weight:700;color:#475569;text-transform:uppercase;letter-spacing:.03em}
.project-scorecard .v{display:block;margin-top:6px;font-size:1.1rem;font-weight:800;color:#0f172a}
.project-scorecard .n{display:block;margin-top:4px;font-size:.72rem;color:#64748b}
.project-scorecards-empty{font-size:.78rem;color:#64748b}
.epic-project-filter{margin-bottom:8px;font-size:.78rem;color:#334155}
.epic-bars{display:grid;gap:8px}
.epic-row{display:grid;grid-template-columns:minmax(180px,2.3fr) minmax(0,8fr);gap:8px;align-items:center}
.epic-label{font-size:.75rem;color:#334155;font-weight:700;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.epic-track{height:16px;background:#edf3f5;border:1px solid #d3e0e4;border-radius:999px;overflow:visible;position:relative}
.epic-bar{height:100%;border-radius:999px;position:relative}
.epic-value-pin{position:absolute;top:50%;transform:translateY(-50%);font-size:.74rem;font-weight:700;color:#1f2937;white-space:nowrap;pointer-events:none}
.epic-project-tag{display:inline-block;margin-right:6px;padding:1px 6px;border-radius:999px;border:1px solid #dbe7eb;background:#fff;font-size:.67rem;color:#475569;font-weight:700}
.epic-empty{font-size:.78rem;color:#64748b}
.pa-toolbar{display:grid;gap:8px;grid-template-columns:minmax(0,1fr) minmax(0,1fr) minmax(180px,220px) auto;align-items:end;margin-bottom:10px}
.pa-status{font-size:.76rem;color:#64748b}
.pa-link{display:inline-flex;align-items:center;justify-content:center;border:1px solid #0e7490;border-radius:8px;padding:7px 10px;background:#ecfeff;color:#0f766e;font-size:.78rem;font-weight:700;text-decoration:none}
.pa-chart{display:grid;gap:8px}
.pa-head,.pa-row{display:grid;grid-template-columns:minmax(140px,1.5fr) minmax(0,6fr) minmax(220px,2fr);gap:10px;align-items:center}
.pa-head{font-size:.7rem;font-weight:800;color:#475569;text-transform:uppercase;letter-spacing:.04em;padding-bottom:3px;border-bottom:1px dashed #d3e0e4}
.pa-project{font-size:.76rem;color:#334155;font-weight:700;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.pa-track{height:18px;border-radius:999px;background:#edf3f5;border:1px solid #d3e0e4;overflow:hidden;position:relative}
.pa-stack{height:100%;display:flex;border-radius:999px;overflow:hidden}
.pa-seg{height:100%;display:flex;align-items:center;justify-content:center;font-size:.66rem;font-weight:800;color:#f8fafc;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.pa-seg.eq{background:#0f766e}
.pa-seg.over{background:#dc2626}
.pa-seg.under{background:#2563eb}
.pa-values{font-size:.72rem;color:#1f2937;text-align:right;font-weight:700}
.pa-legend{display:flex;gap:8px;flex-wrap:wrap;margin-top:6px}
.pa-pill{display:inline-flex;align-items:center;gap:6px;padding:2px 8px;border-radius:999px;background:#f8fafc;border:1px solid #dbe7eb;font-size:.72rem;color:#334155}
.pa-dot{width:8px;height:8px;border-radius:999px;display:inline-block}
.pa-dot.eq{background:#0f766e}.pa-dot.over{background:#dc2626}.pa-dot.under{background:#2563eb}
.pa-empty{font-size:.78rem;color:#64748b}
@media (max-width:760px){.date-range-control{grid-template-columns:1fr}.range-sep{display:none}.bc-row{grid-template-columns:minmax(90px,1.2fr) minmax(90px,1fr) minmax(0,4fr) minmax(130px,1.2fr);gap:8px}.bc-track{height:16px}.bc-bar-label{top:-24px}.bc-benchmark-label{top:20px}.ms-head,.ms-row{grid-template-columns:minmax(90px,1.2fr) minmax(0,4fr) minmax(130px,1.2fr);gap:8px}.ms-track{height:18px}.epic-row{grid-template-columns:minmax(140px,2fr) minmax(0,4fr)}.pa-toolbar{grid-template-columns:1fr}.pa-head,.pa-row{grid-template-columns:minmax(90px,1.2fr) minmax(0,3fr) minmax(130px,1.5fr)}}
</style>  <link rel="stylesheet" href="shared-nav.css">
</head><body><div class="page"> 
<section class="panel hero"><h1>RnD Capacity and Workload Data Story</h1><div class="muted" id="hero-meta"></div></section>
<section class="panel"><div class="filters-toolbar" id="filters-toolbar"><div class="filters"><div class="filter-field range-field"><label for="from-date">Date Range</label><div class="date-range-control"><div class="filter-control date"><input id="from-date" type="date" aria-label="From date"></div><span class="range-sep">to</span><div class="filter-control date"><input id="to-date" type="date" aria-label="To date"></div></div><div class="date-presets" role="group" aria-label="Quick date ranges"><button class="chip" type="button" data-preset="7d">Last 7 days</button><button class="chip" type="button" data-preset="14d">Last 14 days</button><button class="chip" type="button" data-preset="30d">Last 30 days</button><button class="chip" type="button" data-preset="month">This month</button><button class="chip" type="button" data-preset="prev-month">Last month</button></div><div class="date-hint" id="date-range-hint">Select a valid date range.</div></div><div class="filter-field"><label for="capacity-profile-select">Capacity Profile</label><div class="filter-control select"><select id="capacity-profile-select"></select></div></div><div class="filter-field"><label for="actual-hours-mode">Actual Hours Mode</label><div class="filter-control select"><select id="actual-hours-mode"><option value="log_date">By Log Date</option><option value="planned_dates">By Planned Dates</option></select></div></div><div class="btns"><button class="btn alt" id="apply-profile-btn" type="button"><span class="material-symbols-outlined" aria-hidden="true">publish</span> Apply Profile</button><button class="btn" id="apply-btn" type="button"><span class="material-symbols-outlined" aria-hidden="true">tune</span> Apply</button><button class="btn alt" id="reset-btn" type="button"><span class="material-symbols-outlined" aria-hidden="true">restart_alt</span> Reset</button></div></div></div><div class="profile-status" id="profile-status">Capacity source: default profile logic.</div></section>
<section class="panel story-nav"><button id="story-prev" class="nav-btn" type="button" aria-label="Previous story page"><span class="material-symbols-outlined" aria-hidden="true">arrow_back</span></button><div id="story-page-label" class="nav-label">Page 1 of 1</div><button id="story-next" class="nav-btn" type="button" aria-label="Next story page"><span class="material-symbols-outlined" aria-hidden="true">arrow_forward</span></button></section>
<section class="panel story-page" data-story-page="1"><h2 class="page-title">1. Total work by projects</h2><div class="leadership-cards"><article class="kpi metric-capacity"><div class="label">Capacity</div><div class="value" id="kpi-capacity-after-leaves">0h</div><div class="metric-bar-wrap"><div class="metric-bar capacity" id="card-bar-capacity"></div></div></article><article class="kpi metric-planned"><div class="label">Planned Hours</div><div class="value" id="kpi-work-on-plate">0h</div><div class="metric-bar-wrap"><div class="metric-bar planned" id="card-bar-planned"></div></div></article><article class="kpi metric-actual"><div class="label">Actual Hours</div><div class="value" id="kpi-total-actual-project-hours">0h</div><div class="metric-bar-wrap"><div class="metric-bar actual" id="card-bar-actual"></div></div></article><article class="kpi metric-required"><div class="label">Hours required to complete projects</div><div class="value" id="kpi-hours-required-projects">0h</div><div class="metric-bar-wrap"><div class="metric-bar required" id="card-bar-required"></div></div></article></div><article class="story-card"><h3 class="story-title">Project Bullet Chart (Actual bar vs Planned red benchmark)</h3><div class="project-compare-chart" id="project-plan-vs-actual-chart"><div class="vlabel">No project data available.</div></div></article></section>
<section class="panel story-page" data-story-page="2"><h2 class="page-title">2. Planned hours by month (stacked by project)</h2><div class="leadership-cards"><article class="kpi metric-planned"><div class="label">Total Planned Hours</div><div class="value" id="kpi-monthly-total-planned">0h</div></article><article class="kpi metric-planned"><div class="label">Previous Month Planned Hours</div><div class="kpi-sub" id="kpi-monthly-prev-name">-</div><div class="value" id="kpi-monthly-prev-planned">0h</div></article><article class="kpi metric-planned"><div class="label">Current Month Planned Hours</div><div class="kpi-sub" id="kpi-monthly-current-name">-</div><div class="value" id="kpi-monthly-current-planned">0h</div></article><article class="kpi metric-planned"><div class="label">Upcoming Planned Hours</div><div class="kpi-sub" id="kpi-monthly-upcoming-name">-</div><div class="value" id="kpi-monthly-upcoming-planned">0h</div></article></div><article class="story-card"><h3 class="story-title">Project Planned Hours Stack Chart (Prev vs Current vs Upcoming)</h3><div class="monthly-stack-chart" id="project-monthly-planned-chart"><div class="vlabel">No planned project data available.</div></div></article></section>
<section class="panel story-page" data-story-page="3"><h2 class="page-title">3. Project-wise epic planned hours</h2><article class="story-card"><h3 class="story-title">Project Planned Hours Scorecards (click to filter epics)</h3><div class="project-scorecards" id="project-planned-scorecards"><div class="project-scorecards-empty">No project scorecards available for the selected range.</div></div></article><article class="story-card"><h3 class="story-title">Epic Planned Hours Bar Chart (data labels shown)</h3><div class="epic-project-filter" id="epic-project-filter-label">Showing all projects</div><div class="epic-bars" id="epic-planned-hours-chart"><div class="epic-empty">No epics matched the selected range.</div></div></article></section>
<section class="panel story-page" data-story-page="4"><h2 class="page-title">4. Planned vs Actual per Project</h2><article class="story-card"><div class="pa-toolbar"><div class="filter-field"><label for="from-date-page4">Date Filter From</label><div class="filter-control date"><input id="from-date-page4" type="date" aria-label="From date page 4"></div></div><div class="filter-field"><label for="to-date-page4">Date Filter To</label><div class="filter-control date"><input id="to-date-page4" type="date" aria-label="To date page 4"></div></div><div class="filter-field"><label for="actual-hours-mode-page4">Actual Hours Mode</label><div class="filter-control select"><select id="actual-hours-mode-page4"><option value="planned_dates">By Planned Dates</option><option value="log_date">By Log Date</option></select></div></div><div class="filter-field"><label for="status-page4">Status</label><div class="filter-control select"><select id="status-page4"><option value="both">Both</option><option value="in_progress">In Progress</option><option value="resolved">Resolved</option></select></div></div></div><div class="btns"><button class="btn" id="apply-page4-btn" type="button"><span class="material-symbols-outlined" aria-hidden="true">tune</span> Apply</button><button class="btn alt" id="reset-page4-btn" type="button"><span class="material-symbols-outlined" aria-hidden="true">restart_alt</span> Reset</button><a class="pa-link" href="/settings/report-entities?focus=planned_actual_equality_tolerance_hours">Configure Equality Tolerance</a></div><div class="pa-status" id="page4-status">Tolerance: 0h (exact equality).</div></article><article class="story-card"><h3 class="story-title">Stacked Distribution by Project (Epic Counts)</h3><div class="pa-chart" id="planned-actual-project-stack-chart"><div class="pa-empty">No project data available.</div></div></article></section>



</div>\n<script src="https://cdn.jsdelivr.net/npm/echarts@5/dist/echarts.min.js"></script>\n<script src="https://cdn.jsdelivr.net/npm/mermaid@11/dist/mermaid.min.js"></script>\n<script>\nconst payload=__RND_DATA_STORY_PAYLOAD__;
const $=(id)=>document.getElementById(id);
const txt=(v)=>v===undefined||v===null?"":String(v).trim();
const num=(v)=>{const n=Number(v||0);return Number.isFinite(n)?n:0};
const parseIso=(v)=>{const t=txt(v);if(!t)return null;const d=new Date(t+"T00:00:00");return Number.isNaN(d.getTime())?null:d};
const inRange=(v,f,t)=>{const d=parseIso(v);return !!(d&&f&&t&&d>=f&&d<=t)};
const statusClass=(s)=>{const t=txt(s).toLowerCase().replace(/[-_]+/g," ");if(t.includes("resolved"))return "closed";if(t.includes("in progress"))return "open";return "other"};
const hTxt=(v)=>{const n=num(v);if(Math.abs(n-Math.round(n))<0.001)return String(Math.round(n));return n.toFixed(1).replace(/\\.0$/,"")};
let activeProfileKey="";let storyIndex=0;let isApplyingStory=false;let managedFieldsByKey=new Map();let entityCatalogByKey=new Map();let selectedEpicProjectKey="";let latestEpicRows=[];let page4ToleranceHours=0;const page4Filters={from_date:"",to_date:"",actual_mode:"planned_dates",status:"both"};const MANAGED_FIELDS_ENDPOINT="/api/manage-fields?include_inactive=0";const ACTUAL_MODE_STORAGE_KEY="actual-hours-mode:rnd-story";const ACTUAL_MODE_DEFAULT="log_date";
const defaultEmployeeCount=Math.max(num(payload.default_employee_count||0),0);
function countAssigneesInRange(fromDate,toDate){const f=parseIso(fromDate),t=parseIso(toDate);if(!f||!t)return defaultEmployeeCount;const names=new Set();for(const row of (Array.isArray(payload.worklog_rows)?payload.worklog_rows:[])){const d=parseIso(row.worklog_date);if(!d||d<f||d>t)continue;const n=txt(row.issue_assignee||"");if(!n||n.toLowerCase()==="unassigned")continue;names.add(n)}return names.size||defaultEmployeeCount;}
function leaveFromEmbedded(settings){const rows=Array.isArray(payload.leave_daily_rows)?payload.leave_daily_rows:[];const f=parseIso(settings.from_date),t=parseIso(settings.to_date);let pt=0,pn=0,ut=0;if(!f||!t)return{planned_taken_hours:0,planned_not_taken_hours:0,unplanned_taken_hours:0,taken_hours:0,not_yet_taken_hours:0};for(const r of rows){const d=parseIso(r.period_day);if(!d||d<f||d>t)continue;pt+=num(r.planned_taken_hours);pn+=num(r.planned_not_taken_hours);ut+=num(r.unplanned_taken_hours)}return{planned_taken_hours:+pt.toFixed(2),planned_not_taken_hours:+pn.toFixed(2),unplanned_taken_hours:+ut.toFixed(2),taken_hours:+(pt+ut).toFixed(2),not_yet_taken_hours:+pn.toFixed(2)}}
function calcCapacityClient(settings){const f=parseIso(settings.from_date),t=parseIso(settings.to_date);if(!f||!t||t<f)return{settings,metrics:{available_capacity_hours:0},leave_metrics:leaveFromEmbedded(settings)};const holidays=new Set(Array.isArray(settings.holiday_dates)?settings.holiday_dates.map((x)=>txt(x)):[]);const rs=parseIso(settings.ramadan_start_date),re=parseIso(settings.ramadan_end_date);let n=0,r=0;const c=new Date(f);while(c<=t){const weekday=c.getDay()>=1&&c.getDay()<=5;const iso=c.toISOString().slice(0,10);if(weekday&&!holidays.has(iso)){const inRamadan=!!(rs&&re&&c>=rs&&c<=re);if(inRamadan)r+=1;else n+=1}c.setDate(c.getDate()+1)}const av=num(settings.employee_count)*(n*num(settings.standard_hours_per_day||8)+r*num(settings.ramadan_hours_per_day||6.5));const leave=leaveFromEmbedded(settings);const rem=av-num(leave.taken_hours)-num(leave.not_yet_taken_hours);return{settings,metrics:{available_capacity_hours:+av.toFixed(2)},leave_metrics:{...leave,remaining_balance_hours:+rem.toFixed(2)}}}
function getProfiles(){return Array.isArray(payload.capacity_profiles)?payload.capacity_profiles:[]}
function profileKey(p){return `${txt(p.from_date)}|${txt(p.to_date)}`}
function findProfileByKey(key){return getProfiles().find((p)=>profileKey(p)===key)||null}
function populateProfileSelect(){const select=$("capacity-profile-select");const profiles=getProfiles();if(!profiles.length){select.innerHTML='<option value="">No saved profiles</option>';select.disabled=true;return}select.disabled=false;select.innerHTML=profiles.map((p)=>`<option value="${profileKey(p)}">${txt(p.from_date)} to ${txt(p.to_date)} (Emp:${num(p.employee_count)}, Std:${hTxt(p.standard_hours_per_day)}h, Ram:${hTxt(p.ramadan_hours_per_day)}h)</option>`).join("")}
function normalizeSettings(base,fromDate,toDate){const s={...(base||{})};s.from_date=fromDate;s.to_date=toDate;s.employee_count=num(s.employee_count);if(s.employee_count<=0){s.employee_count=Math.max(countAssigneesInRange(fromDate,toDate),1)}s.standard_hours_per_day=num(s.standard_hours_per_day||8)||8;s.ramadan_hours_per_day=num(s.ramadan_hours_per_day||6.5)||6.5;s.ramadan_start_date=txt(s.ramadan_start_date);s.ramadan_end_date=txt(s.ramadan_end_date);s.holiday_dates=Array.isArray(s.holiday_dates)?s.holiday_dates:[];return s}
async function loadCapacity(fromDate,toDate){if(activeProfileKey){const p=findProfileByKey(activeProfileKey);if(p){return calcCapacityClient(normalizeSettings(p,fromDate,toDate))}}const profiles=getProfiles();if((window.location.protocol||"").startsWith("http")){try{const res=await fetch(`/api/capacity?from=${encodeURIComponent(fromDate)}&to=${encodeURIComponent(toDate)}`);if(res.ok){const data=await res.json();const emp=num(data?.settings?.employee_count);if(emp>0)return data;if(profiles.length){return calcCapacityClient(normalizeSettings(profiles[0],fromDate,toDate))}return calcCapacityClient(normalizeSettings(data?.settings||{},fromDate,toDate))}}catch(e){}}const fallback=profiles[0]||{from_date:fromDate,to_date:toDate,employee_count:0,standard_hours_per_day:8,ramadan_start_date:"",ramadan_end_date:"",ramadan_hours_per_day:6.5,holiday_dates:[]};return calcCapacityClient(normalizeSettings(fallback,fromDate,toDate))}
function plannedWorkOnPlate(from,to){const rows=Array.isArray(payload.planned_epic_rows)?payload.planned_epic_rows:[];if(rows.length){let total=0;for(const row of rows){if(txt(row.project_key).toUpperCase()==="RLT")continue;if(!(inRange(row.planned_start,from,to)||inRange(row.planned_end,from,to)))continue;total+=num(row.planned_hours)}return +total.toFixed(2)}let fallback=0;for(const epic of (Array.isArray(payload.epics)?payload.epics:[])){if(txt(epic.project_key).toUpperCase()==="RLT")continue;if(!(inRange(epic.start_date,from,to)||inRange(epic.end_date,from,to)))continue;fallback+=num(epic.original_estimate_hours)}return +fallback.toFixed(2)}function totalActualProjectHours(from,to,aggregatePayload){const projectHours=aggregatePayload&&aggregatePayload.project_hours_by_key&&typeof aggregatePayload.project_hours_by_key==="object"?aggregatePayload.project_hours_by_key:null;if(projectHours){let total=0;for(const [key,hours] of Object.entries(projectHours)){if(txt(key).toUpperCase()==="RLT")continue;total+=num(hours)}return +total.toFixed(2)}let total=0;for(const row of (Array.isArray(payload.worklog_rows)?payload.worklog_rows:[])){if(txt(row.project_key).toUpperCase()==="RLT")continue;const d=parseIso(row.worklog_date);if(!d||!from||!to||d<from||d>to)continue;total+=num(row.hours_logged)}if(total>0||Array.isArray(payload.worklog_rows))return +total.toFixed(2);const rows=Array.isArray(payload.project_actual_rows)?payload.project_actual_rows:[];for(const row of rows){if(txt(row.project_key).toUpperCase()==="RLT")continue;total+=num(row.actual_hours)}return +total.toFixed(2)}
function benchmarkHoursByDueDate(plannedHours,fromDate,dueDate){const planned=Math.max(num(plannedHours),0);if(planned<=0)return 0;const from=parseIso(fromDate);const due=parseIso(dueDate);if(!from||!due||due<=from)return planned;const now=new Date();const today=new Date(now.getFullYear(),now.getMonth(),now.getDate());if(today<=from)return 0;const capped=today<due?today:due;const elapsed=Math.floor((capped.getTime()-from.getTime())/(24*60*60*1000))+1;const total=Math.floor((due.getTime()-from.getTime())/(24*60*60*1000))+1;const ratio=Math.max(0,Math.min(elapsed/Math.max(total,1),1));return +(planned*ratio).toFixed(2)}
function projectCompareRows(from,to,aggregatePayload){const plannedByProject={};for(const row of (Array.isArray(payload.planned_epic_rows)?payload.planned_epic_rows:[])){const key=txt(row.project_key).toUpperCase();if(!key||key==="RLT")continue;if(!(inRange(row.planned_start,from,to)||inRange(row.planned_end,from,to)))continue;plannedByProject[key]=(plannedByProject[key]||0)+num(row.planned_hours)}const actualByProject={};const projectHours=aggregatePayload&&aggregatePayload.project_hours_by_key&&typeof aggregatePayload.project_hours_by_key==="object"?aggregatePayload.project_hours_by_key:null;if(projectHours){for(const [keyRaw,hours] of Object.entries(projectHours)){const key=txt(keyRaw).toUpperCase();if(!key||key==="RLT")continue;actualByProject[key]=(actualByProject[key]||0)+num(hours)}}else{const worklogs=Array.isArray(payload.worklog_rows)?payload.worklog_rows:[];if(worklogs.length){for(const row of worklogs){const key=txt(row.project_key).toUpperCase();if(!key||key==="RLT")continue;const d=parseIso(row.worklog_date);if(!d||!from||!to||d<from||d>to)continue;actualByProject[key]=(actualByProject[key]||0)+num(row.hours_logged)}}else{for(const row of (Array.isArray(payload.project_actual_rows)?payload.project_actual_rows:[])){const key=txt(row.project_key).toUpperCase();if(!key||key==="RLT")continue;actualByProject[key]=(actualByProject[key]||0)+num(row.actual_hours)}}}const dueByProject={};for(const epic of (Array.isArray(payload.epics)?payload.epics:[])){const key=txt(epic.project_key).toUpperCase();if(!key||key==="RLT")continue;if(!(inRange(epic.start_date,from,to)||inRange(epic.end_date,from,to)))continue;const dueTxt=txt(epic.end_date);const due=parseIso(dueTxt);if(!due)continue;const prev=parseIso(dueByProject[key]||"");if(!prev||due<prev)dueByProject[key]=dueTxt}const rows=[];for(const key of new Set([...Object.keys(plannedByProject),...Object.keys(actualByProject)])){const planned=+(num(plannedByProject[key]).toFixed(2));const actual=+(num(actualByProject[key]).toFixed(2));const requiredDefault=+(planned-actual).toFixed(2);const requiredEval=evaluateManagedField("hours_required_to_complete_projects",requiredDefault,{capacity:0,planned_hours:planned,actual_hours:actual,planned_leaves:0});const required=+num(requiredEval.value).toFixed(2);const dueDate=txt(dueByProject[key]);const benchmark=benchmarkHoursByDueDate(planned,isoDateLocal(from),dueDate);rows.push({project_key:key,planned_hours:planned,actual_hours:actual,hours_required_to_complete:required,benchmark_hours_due_date:benchmark,project_due_date:dueDate,completion_pct:planned>0?+((actual/planned)*100).toFixed(2):0})}rows.sort((a,b)=>num(b.planned_hours)-num(a.planned_hours));return rows}
function monthStartLocal(d){return new Date(d.getFullYear(),d.getMonth(),1)}
function addMonthsLocal(d,delta){return new Date(d.getFullYear(),d.getMonth()+delta,1)}
function monthLabel(d){return d.toLocaleDateString(undefined,{month:"long",year:"numeric"})}
function monthlyPlannedByProject(){const now=new Date();const currentStart=monthStartLocal(now);const previousStart=addMonthsLocal(currentStart,-1);const nextStart=addMonthsLocal(currentStart,1);const byProject={};for(const row of (Array.isArray(payload.planned_epic_rows)?payload.planned_epic_rows:[])){const key=txt(row.project_key).toUpperCase();if(!key||key==="RLT")continue;const anchor=parseIso(row.planned_end)||parseIso(row.planned_start);if(!anchor)continue;const planned=Math.max(num(row.planned_hours),0);if(planned<=0)continue;if(!Object.prototype.hasOwnProperty.call(byProject,key))byProject[key]={project_key:key,previous:0,current:0,upcoming:0,total:0};if(anchor>=previousStart&&anchor<currentStart)byProject[key].previous+=planned;else if(anchor>=currentStart&&anchor<nextStart)byProject[key].current+=planned;else if(anchor>=nextStart)byProject[key].upcoming+=planned}const rows=[];let previousTotal=0,currentTotal=0,upcomingTotal=0;for(const key of Object.keys(byProject)){const rec=byProject[key];rec.previous=+num(rec.previous).toFixed(2);rec.current=+num(rec.current).toFixed(2);rec.upcoming=+num(rec.upcoming).toFixed(2);rec.total=+(rec.previous+rec.current+rec.upcoming).toFixed(2);if(rec.total<=0)continue;rows.push(rec);previousTotal+=rec.previous;currentTotal+=rec.current;upcomingTotal+=rec.upcoming}rows.sort((a,b)=>num(b.total)-num(a.total));const totalDefault=+(previousTotal+currentTotal+upcomingTotal).toFixed(2);const previousDefault=+num(previousTotal).toFixed(2);const currentDefault=+num(currentTotal).toFixed(2);const upcomingDefault=+num(upcomingTotal).toFixed(2);const baseContext={capacity:0,actual_hours:0,planned_leaves:0};const totalEval=evaluateManagedField("total_planned_hours",totalDefault,{...baseContext,planned_hours:totalDefault});const prevEval=evaluateManagedField("previous_month_planned_hours",previousDefault,{...baseContext,planned_hours:previousDefault});const currentEval=evaluateManagedField("current_month_planned_hours",currentDefault,{...baseContext,planned_hours:currentDefault});const upcomingEval=evaluateManagedField("upcoming_planned_hours",upcomingDefault,{...baseContext,planned_hours:upcomingDefault});return{rows,meta:{previous_label:monthLabel(previousStart),current_label:monthLabel(currentStart),upcoming_label:`${monthLabel(nextStart)} onward`,total:+num(totalEval.value).toFixed(2),previous:+num(prevEval.value).toFixed(2),current:+num(currentEval.value).toFixed(2),upcoming:+num(upcomingEval.value).toFixed(2)}}}
function epicMetrics(from,to){const by=payload.epic_logged_hours_by_key||{};const out=[];let pending=0,closed=0,open=0,totalEstimate=0,totalLogged=0;for(const epic of (Array.isArray(payload.epics)?payload.epics:[])){if(txt(epic.project_key).toUpperCase()==="RLT")continue;if(!(inRange(epic.start_date,from,to)||inRange(epic.end_date,from,to)))continue;const key=txt(epic.issue_key).toUpperCase();const logged=num(by[key]);const est=num(epic.original_estimate_hours);const p=Math.max(est-logged,0);const s=statusClass(epic.status);if(s==="closed")closed+=1;if(s==="open")open+=1;pending+=p;totalEstimate+=est;totalLogged+=logged;out.push({...epic,epic_logged_hours:+logged.toFixed(2),pending_hours:+p.toFixed(2),status_bucket:s})}const required=totalEstimate-totalLogged;return{filtered:out,pending_required_hours:+pending.toFixed(2),hours_required_to_complete_epics:+required.toFixed(2),total_estimate_hours:+totalEstimate.toFixed(2),total_logged_hours:+totalLogged.toFixed(2),closed_resolved_count:closed,open_in_progress_count:open}}
function loadManagedFieldsFromPayloadOrApi(){const fromPayload=Array.isArray(payload.managed_fields)?payload.managed_fields:[];if(fromPayload.length){applyManagedFieldsPayload({fields:fromPayload,entity_catalog:Array.isArray(payload.entity_catalog)?payload.entity_catalog:[]});return Promise.resolve(true)}if(!String(window.location.protocol||"").startsWith("http"))return Promise.resolve(false);return fetch(MANAGED_FIELDS_ENDPOINT,{method:"GET"}).then((res)=>res.ok?res.json():null).then((data)=>{if(!data)return false;applyManagedFieldsPayload(data);return true}).catch(()=>false)}
function applyManagedFieldsPayload(data){const fields=Array.isArray(data&&data.fields)?data.fields:[];const entities=Array.isArray(data&&data.entity_catalog)?data.entity_catalog:[];const nextFields=new Map();for(const f of fields){const key=txt(f&&f.field_key).toLowerCase();if(!key||!f||!f.is_active)continue;nextFields.set(key,f)}managedFieldsByKey=nextFields;const nextEntities=new Map();for(const e of entities){const key=txt(e&&e.entity_key).toLowerCase();if(!key)continue;nextEntities.set(key,e)}entityCatalogByKey=nextEntities}
function evaluateExpressionWithContext(expression,context){const text=txt(expression);if(!text)return{ok:false,value:NaN,error:"Empty expression."};const src=text;const tokens=[];let i=0;while(i<src.length){const ch=src[i];if(/\\s/.test(ch)){i+=1;continue}if("+-*/".includes(ch)){tokens.push({t:"op",v:ch,p:i});i+=1;continue}if(ch==="("){tokens.push({t:"lparen",v:ch,p:i});i+=1;continue}if(ch===")"){tokens.push({t:"rparen",v:ch,p:i});i+=1;continue}if(ch===","){tokens.push({t:"comma",v:ch,p:i});i+=1;continue}if(/[A-Za-z_]/.test(ch)){const s=i;i+=1;while(i<src.length&&/[A-Za-z0-9_]/.test(src[i]))i+=1;tokens.push({t:"ident",v:src.slice(s,i),p:s});continue}return{ok:false,value:NaN,error:`Invalid character at position ${i+1}`}}tokens.push({t:"eof",v:"",p:src.length});let p=0;const peek=()=>tokens[p];const eat=(expected)=>{const tk=tokens[p];if(expected&&tk.t!==expected)throw new Error(`Expected ${expected} at position ${tk.p+1}.`);p+=1;return tk};const fn=(name,arg)=>{const n=txt(name).toLowerCase();const v=num(arg);if(n==="sum"||n==="min"||n==="max"||n==="average")return v;if(n==="count")return v!==0?1:0;throw new Error(`Unknown function '${name}'.`)};const expr=()=>{let v=term();while(peek().t==="op"&&(peek().v==="+"||peek().v==="-")){const op=eat("op").v;const rhs=term();v=op==="+"?v+rhs:v-rhs}return v};const term=()=>{let v=factor();while(peek().t==="op"&&(peek().v==="*"||peek().v==="/")){const op=eat("op").v;const rhs=factor();if(op==="*")v*=rhs;else v=rhs===0?0:v/rhs}return v};const factor=()=>{const tk=peek();if(tk.t==="ident"){const id=eat("ident");const key=txt(id.v).toLowerCase();if(peek().t==="lparen"){eat("lparen");const arg=expr();if(peek().t==="comma")throw new Error(`Function '${id.v}' accepts one argument at position ${peek().p+1}.`);eat("rparen");return fn(id.v,arg)}if(!Object.prototype.hasOwnProperty.call(context,key))throw new Error(`Unknown identifier '${id.v}'.`);return num(context[key])}if(tk.t==="lparen"){eat("lparen");const v=expr();eat("rparen");return v}throw new Error(`Unexpected token at position ${tk.p+1}.`)};try{const value=expr();if(peek().t!=="eof")throw new Error(`Unexpected token at position ${peek().p+1}.`);return{ok:true,value:+num(value).toFixed(2),error:""}}catch(err){return{ok:false,value:NaN,error:txt(err&&err.message||err)}}}
function managedFieldFormulaText(fieldKey,fallback){const key=txt(fieldKey).toLowerCase();const item=managedFieldsByKey.get(key);const formula=txt(item&&item.formula_expression);return formula||fallback}
function evaluateManagedField(fieldKey,fallbackValue,context){const key=txt(fieldKey).toLowerCase();const item=managedFieldsByKey.get(key);const formula=txt(item&&item.formula_expression);if(!formula)return{value:fallbackValue,usedManagedField:false,formulaText:"",error:""};const out=evaluateExpressionWithContext(formula,context||{});if(!out.ok)return{value:fallbackValue,usedManagedField:false,formulaText:formula,error:out.error};return{value:out.value,usedManagedField:true,formulaText:formula,error:""}}
function width(id,n,d){const e=$(id);if(!e)return;const pct=d>0?Math.max(0,Math.min((n/d)*100,100)):0;e.style.width=`${pct.toFixed(2)}%`}
function setTxt(id,val){const e=$(id);if(e)e.textContent=val}
let mermaidReady=false;
function escMermaid(v){return String(v??"").replace(/"/g,"'").replace(/[\\[\\]{}]/g,"");}
function renderLeadershipMermaid(values){const host=$("leadership-mermaid");if(!host||!mermaidReady||typeof mermaid==="undefined")return;const def=`flowchart TB
subgraph ROW1[Capacity Flow Top]
 direction LR
 A[Total Capacity\n${escMermaid(values.after)}h] --> B[Planned Projects\n${escMermaid(values.work)}h] --> C[Actual Project Hours\n${escMermaid(values.actual)}h]
end
subgraph ROW2[Capacity Flow Bottom]
 direction LR
 E[Capacity Available\n${escMermaid(values.invest)}h] --> D[Hours Required\n${escMermaid(values.required)}h]
end
C --> D
`;
host.removeAttribute("data-processed");host.textContent=def;mermaid.run({nodes:[host]}).catch(()=>{});}
let barrelChart=null;
function ensureBarrelChart(){const host=$("barrel-enterprise-chart");if(!host||typeof echarts==="undefined")return null;if(!barrelChart){barrelChart=echarts.init(host,null,{renderer:"canvas"});window.addEventListener("resize",()=>{if(barrelChart)barrelChart.resize()})}return barrelChart}
function renderLeadershipBarrel(capacity,planned,actual,hoursRequired,capacityAvailable){
const cap=Math.max(num(capacity),0);
const target=Math.max(num(planned),0);
const fluid=Math.max(num(actual),0);
const capSafe=Math.max(cap,1);
const targetClamped=Math.max(0,Math.min(target,capSafe));
const fluidClamped=Math.max(0,Math.min(fluid,capSafe));
const chart=ensureBarrelChart();
if(chart){
chart.setOption({animationDuration:520,animationDurationUpdate:420,grid:{left:66,right:220,top:20,bottom:30,containLabel:false},xAxis:{type:"category",data:["barrel"],axisLine:{show:false},axisTick:{show:false},axisLabel:{show:false},splitLine:{show:false}},yAxis:{type:"value",min:0,max:capSafe,splitNumber:5,axisLine:{show:false},axisTick:{show:false},splitLine:{lineStyle:{color:"#dce8ee",type:"dashed"}},axisLabel:{fontSize:11,fontWeight:700,color:"#64748b",formatter:(v)=>`${hTxt(v)}h`}},series:[{name:"Fluid",type:"bar",data:[fluidClamped],barWidth:"48%",z:2,itemStyle:{borderRadius:[999,999,999,999],color:new echarts.graphic.LinearGradient(0,0,0,1,[{offset:0,color:"#a855f7"},{offset:1,color:"#6d28d9"}]),shadowBlur:8,shadowColor:"rgba(109,40,217,.25)"},label:{show:true,position:["56%",-2],formatter:`Actual: ${hTxt(fluidClamped)}h`,backgroundColor:"rgba(255,255,255,.95)",borderColor:"#e9d5ff",borderWidth:1,borderRadius:6,padding:[3,6],color:"#6b21a8",fontWeight:800,fontSize:12}},{name:"Shell",type:"bar",data:[capSafe],barWidth:"48%",barGap:"-100%",z:3,itemStyle:{color:"rgba(248,250,252,.1)",borderColor:"#334155",borderWidth:4,borderRadius:[999,999,999,999]}},{name:"Capacity Line",type:"line",data:[capSafe],symbol:"none",z:5,lineStyle:{color:"#0f766e",width:2},label:{show:true,position:["56%",-24],formatter:`Capacity: ${hTxt(capSafe)}h`,backgroundColor:"rgba(255,255,255,.95)",borderColor:"#99f6e4",borderWidth:1,borderRadius:6,padding:[3,6],color:"#115e59",fontWeight:800,fontSize:12}},{name:"Planned Line",type:"line",data:[targetClamped],symbol:"none",z:5,lineStyle:{color:"#dc2626",width:3},label:{show:true,position:["56%",-2],formatter:`Planned: ${hTxt(targetClamped)}h`,backgroundColor:"rgba(255,255,255,.95)",borderColor:"#fecaca",borderWidth:1,borderRadius:6,padding:[3,6],color:"#991b1b",fontWeight:800,fontSize:12}}]},true)
}
setTxt("barrel-capacity-val",`${hTxt(cap)}h`);
setTxt("barrel-planned-val",`${hTxt(target)}h`);
setTxt("barrel-fluid-val",`${hTxt(fluid)}h`);
setTxt("barrel-hours-required-val",`${hTxt(hoursRequired)}h`);
setTxt("barrel-capacity-available-val",`${hTxt(capacityAvailable)}h`);
setTxt("barrel-hours-required-note",`Target - Fluid = ${hTxt(hoursRequired)}h`);
setTxt("barrel-capacity-available-note",`Capacity - Target = ${hTxt(capacityAvailable)}h`);
}

function renderLeadershipCards(capacity,planned,actual,required){const maxH=Math.max(num(capacity),num(planned),num(actual),Math.max(num(required),0),1);width("card-bar-capacity",Math.max(num(capacity),0),maxH);width("card-bar-planned",Math.max(num(planned),0),maxH);width("card-bar-actual",Math.max(num(actual),0),maxH);width("card-bar-required",Math.max(num(required),0),maxH);}
function renderProjectPlanVsActual(rows){const host=$("project-plan-vs-actual-chart");if(!host)return;const source=Array.isArray(rows)?rows:[];if(!source.length){host.innerHTML='<div class="vlabel">No project data available for the selected range.</div>';return}const top=source.slice(0,12);const max=Math.max(...top.map((r)=>Math.max(num(r.planned_hours),num(r.actual_hours))),1);const head=`<div class="bc-row"><div class="bc-head">RMIs</div><div class="bc-head">Due Date</div><div class="bc-head">Bullet Chart</div><div class="bc-head" style="text-align:right">Data Labels</div></div>`;const body=top.map((r)=>{const actual=num(r.actual_hours);const planned=num(r.planned_hours);const actualPct=((actual/max)*100).toFixed(2);const plannedPct=((planned/max)*100).toFixed(2);const due=txt(r.project_due_date)||"-";const plannedLabelPct=plannedPct==="0.00"?"0.80":plannedPct;return `<div class="bc-row"><div class="bc-label">${txt(r.project_key)}</div><div class="bc-due">${due}</div><div class="bc-track"><div class="bc-actual" style="width:${actualPct}%"><div class="bc-bar-label">${hTxt(actual)}h</div></div><div class="bc-benchmark" style="left:${plannedPct}%"></div><div class="bc-benchmark-label" style="left:${plannedLabelPct}%">${hTxt(planned)}h</div></div><div class="bc-values">Actual ${hTxt(actual)}h | Planned ${hTxt(planned)}h</div></div>`}).join("");host.innerHTML=head+body}
function renderMonthlyPlannedKpis(meta){const m=meta||{};setTxt("kpi-monthly-total-planned",`${hTxt(m.total)}h`);setTxt("kpi-monthly-prev-name",txt(m.previous_label)||"-");setTxt("kpi-monthly-current-name",txt(m.current_label)||"-");setTxt("kpi-monthly-upcoming-name",txt(m.upcoming_label)||"-");setTxt("kpi-monthly-prev-planned",`${hTxt(m.previous)}h`);setTxt("kpi-monthly-current-planned",`${hTxt(m.current)}h`);setTxt("kpi-monthly-upcoming-planned",`${hTxt(m.upcoming)}h`)}
function renderMonthlyPlannedStack(rows,meta){const host=$("project-monthly-planned-chart");if(!host)return;const source=Array.isArray(rows)?rows:[];const m=meta||{};if(!source.length){host.innerHTML='<div class="vlabel">No planned project data available for monthly buckets.</div>';return}const top=source.slice(0,12);const max=Math.max(...top.map((r)=>num(r.total)),1);const head=`<div class="ms-head"><div>Projects</div><div>Stacked Planned Hours</div><div style="text-align:right">Data Labels</div></div>`;const legend=`<div class="ms-legend"><span class="ms-pill"><span class="ms-dot prev"></span>${txt(m.previous_label)||"Previous Month"}</span><span class="ms-pill"><span class="ms-dot current"></span>${txt(m.current_label)||"Current Month"}</span><span class="ms-pill"><span class="ms-dot upcoming"></span>${txt(m.upcoming_label)||"Upcoming"}</span></div>`;const body=top.map((r)=>{const total=Math.max(num(r.total),0);const barPct=((total/max)*100).toFixed(2);const prevPct=total>0?((num(r.previous)/total)*100):0;const currentPct=total>0?((num(r.current)/total)*100):0;const upcomingPct=Math.max(0,100-prevPct-currentPct);const prevLabel=prevPct>=16?`${hTxt(r.previous)}h`:"";const currentLabel=currentPct>=16?`${hTxt(r.current)}h`:"";const upcomingLabel=upcomingPct>=16?`${hTxt(r.upcoming)}h`:"";return `<div class="ms-row"><div class="ms-project">${txt(r.project_key)}</div><div class="ms-track"><div class="ms-bar" style="width:${barPct}%"><span class="ms-seg ms-prev" style="width:${prevPct.toFixed(2)}%" title="Previous: ${hTxt(r.previous)}h">${prevLabel}</span><span class="ms-seg ms-current" style="width:${currentPct.toFixed(2)}%" title="Current: ${hTxt(r.current)}h">${currentLabel}</span><span class="ms-seg ms-upcoming" style="width:${upcomingPct.toFixed(2)}%" title="Upcoming: ${hTxt(r.upcoming)}h">${upcomingLabel}</span></div></div><div class="ms-values">Prev ${hTxt(r.previous)}h | Curr ${hTxt(r.current)}h | Upc ${hTxt(r.upcoming)}h | Total ${hTxt(total)}h</div></div>`}).join("");host.innerHTML=head+body+legend}
function epicPlannedRowsByProject(from,to){const epics=Array.isArray(payload.epics)?payload.epics:[];const rows=[];const totalsByProject={};for(const epic of epics){const projectKey=txt(epic.project_key).toUpperCase();if(!projectKey||projectKey==="RLT")continue;if(!(inRange(epic.start_date,from,to)||inRange(epic.end_date,from,to)))continue;const planned=Math.max(num(epic.original_estimate_hours),0);if(planned<=0)continue;const row={project_key:projectKey,issue_key:txt(epic.issue_key),summary:txt(epic.summary),planned_hours:+planned.toFixed(2)};rows.push(row);totalsByProject[projectKey]=(totalsByProject[projectKey]||0)+planned}const projectOrder=Object.entries(totalsByProject).sort((a,b)=>num(b[1])-num(a[1])).map(([key])=>key);const projectPos=new Map(projectOrder.map((key,idx)=>[key,idx]));rows.sort((a,b)=>{const projectDiff=(projectPos.get(a.project_key)||9999)-(projectPos.get(b.project_key)||9999);if(projectDiff!==0)return projectDiff;const plannedDiff=num(b.planned_hours)-num(a.planned_hours);if(plannedDiff!==0)return plannedDiff;return txt(a.issue_key).localeCompare(txt(b.issue_key))});return rows}
function projectColor(projectKey){const key=txt(projectKey).toUpperCase();if(key.includes("OMNICONNECT")||key==="OMNI"||key==="OC")return "#16a34a";if(key.includes("DIGITAL")||key.includes("DLOG")||key==="DL")return "#2563eb";const palette=["#0e7490","#7c3aed","#f59e0b","#dc2626","#0891b2","#65a30d","#4f46e5","#b45309"];let hash=0;for(const ch of key)hash=((hash<<5)-hash)+ch.charCodeAt(0);const idx=Math.abs(hash)%palette.length;return palette[idx]}
function projectScorecardRows(epicRows){const totals={};for(const row of (Array.isArray(epicRows)?epicRows:[])){const key=txt(row.project_key).toUpperCase();if(!key)continue;if(!Object.prototype.hasOwnProperty.call(totals,key))totals[key]={project_key:key,planned_hours:0,epic_count:0};totals[key].planned_hours+=num(row.planned_hours);totals[key].epic_count+=1}return Object.values(totals).map((r)=>({...r,planned_hours:+num(r.planned_hours).toFixed(2)})).sort((a,b)=>num(b.planned_hours)-num(a.planned_hours))}
function renderProjectPlannedScorecards(rows){const host=$("project-planned-scorecards");if(!host)return;const source=Array.isArray(rows)?rows:[];if(!source.length){selectedEpicProjectKey="";host.innerHTML='<div class="project-scorecards-empty">No project scorecards available for the selected range.</div>';return}if(selectedEpicProjectKey&&!source.some((r)=>txt(r.project_key).toUpperCase()===selectedEpicProjectKey))selectedEpicProjectKey="";const allActive=!selectedEpicProjectKey;const allPlanned=source.reduce((sum,r)=>sum+num(r.planned_hours),0);const allCount=source.reduce((sum,r)=>sum+num(r.epic_count),0);const allCard=`<button type="button" class="project-scorecard ${allActive?"active":""}" data-project-key=""><span class="k">All Projects</span><span class="v">${hTxt(allPlanned)}h</span><span class="n">${Math.round(allCount)} epics</span></button>`;const cards=source.map((r)=>{const key=txt(r.project_key).toUpperCase();const active=selectedEpicProjectKey===key?"active":"";return `<button type="button" class="project-scorecard ${active}" data-project-key="${key}" style="border-left:5px solid ${projectColor(key)}"><span class="k">${key}</span><span class="v">${hTxt(r.planned_hours)}h</span><span class="n">${Math.round(num(r.epic_count))} epics</span></button>`}).join("");host.innerHTML=allCard+cards;for(const btn of Array.from(host.querySelectorAll(".project-scorecard"))){btn.addEventListener("click",()=>{const key=txt(btn.getAttribute("data-project-key")).toUpperCase();selectedEpicProjectKey=key||"";renderProjectPlannedScorecards(projectScorecardRows(latestEpicRows));renderEpicPlannedHoursBars(latestEpicRows,selectedEpicProjectKey)})}}
function renderEpicPlannedHoursBars(epicRows,projectFilter){const host=$("epic-planned-hours-chart");const labelHost=$("epic-project-filter-label");if(!host)return;const filter=txt(projectFilter).toUpperCase();const source=Array.isArray(epicRows)?epicRows:[];const filtered=filter?source.filter((row)=>txt(row.project_key).toUpperCase()===filter):source;if(labelHost)labelHost.textContent=filter?`Showing project: ${filter}`:"Showing all projects";if(!filtered.length){host.innerHTML='<div class="epic-empty">No epics matched the selected range.</div>';return}const max=Math.max(...filtered.map((r)=>num(r.planned_hours)),1);host.innerHTML=filtered.map((r)=>{const planned=num(r.planned_hours);const pctValue=(planned/max)*100;const pct=pctValue.toFixed(2);const labelLeft=Math.min(Math.max(pctValue+0.8,1.2),96);const projectKey=txt(r.project_key).toUpperCase();const color=projectColor(projectKey);const label=`${txt(r.issue_key)} - ${txt(r.summary)||"-"}`;return `<div class="epic-row"><div class="epic-label"><span class="epic-project-tag">${projectKey}</span>${label}</div><div class="epic-track"><div class="epic-bar" style="width:${pct}%;background:${color}"></div><div class="epic-value-pin" style="left:${labelLeft}%">${hTxt(planned)}h</div></div></div>`}).join("")}
function statusBucketForFilter(epicStatus){const s=statusClass(epicStatus);if(s==="open")return"in_progress";if(s==="closed")return"resolved";return"other"}
function isStatusIncluded(bucket,statusFilter){const filter=txt(statusFilter).toLowerCase();if(filter==="in_progress")return bucket==="in_progress";if(filter==="resolved")return bucket==="resolved";if(filter==="both")return bucket==="in_progress"||bucket==="resolved";return false}
async function loadPlannedActualTolerance(){if(!String(window.location.protocol||"").startsWith("http")){page4ToleranceHours=0;return page4ToleranceHours}try{const resp=await fetch("/api/report-entities",{method:"GET"});const data=await resp.json().catch(()=>({}));if(resp.ok&&data&&data.global_settings){const raw=Number(data.global_settings.planned_actual_equality_tolerance_hours||0);if(Number.isFinite(raw)&&raw>=0){page4ToleranceHours=raw;return page4ToleranceHours}}}catch(_err){}page4ToleranceHours=0;return page4ToleranceHours}
function setPage4InputsFromState(){$("from-date-page4").value=txt(page4Filters.from_date);$("to-date-page4").value=txt(page4Filters.to_date);$("actual-hours-mode-page4").value=txt(page4Filters.actual_mode)||"planned_dates";$("status-page4").value=txt(page4Filters.status)||"both"}
function readPage4InputsToState(){page4Filters.from_date=txt($("from-date-page4").value);page4Filters.to_date=txt($("to-date-page4").value);page4Filters.actual_mode=txt($("actual-hours-mode-page4").value)||"planned_dates";page4Filters.status=txt($("status-page4").value)||"both"}
function resetPage4Filters(){const d=payload.defaults||{};page4Filters.from_date=txt(d.from_date);page4Filters.to_date=txt(d.to_date);page4Filters.actual_mode="planned_dates";page4Filters.status="both";setPage4InputsFromState()}
function epicActualHoursByModeAndRange(aggregatePayload){const by={};const source=aggregatePayload&&aggregatePayload.epic_hours_by_issue&&typeof aggregatePayload.epic_hours_by_issue==="object"?aggregatePayload.epic_hours_by_issue:payload.epic_logged_hours_by_key||{};for(const [key,val] of Object.entries(source)){const issueKey=txt(key).toUpperCase();if(!issueKey)continue;by[issueKey]=num(val)}return by}
function projectPlannedActualDistribution(filters,actualAgg,toleranceHours){const from=parseIso(filters.from_date),to=parseIso(filters.to_date);if(!from||!to||to<from)return[];const tol=Math.max(num(toleranceHours),0);const actualByEpic=epicActualHoursByModeAndRange(actualAgg);const byProject={};for(const epic of (Array.isArray(payload.epics)?payload.epics:[])){const projectKey=txt(epic.project_key).toUpperCase();if(!projectKey||projectKey==="RLT")continue;if(!(inRange(epic.start_date,from,to)||inRange(epic.end_date,from,to)))continue;const bucket=statusBucketForFilter(epic.status);if(!isStatusIncluded(bucket,filters.status))continue;const planned=Math.max(num(epic.original_estimate_hours),0);if(planned<=0)continue;const issueKey=txt(epic.issue_key).toUpperCase();const actual=Math.max(num(actualByEpic[issueKey]),0);if(!Object.prototype.hasOwnProperty.call(byProject,projectKey))byProject[projectKey]={project_key:projectKey,equal_count:0,overrun_count:0,under_count:0,total_count:0};const rec=byProject[projectKey];if(Math.abs(planned-actual)<=tol)rec.equal_count+=1;else if(planned<actual)rec.overrun_count+=1;else rec.under_count+=1;rec.total_count+=1}const rows=Object.values(byProject).filter((r)=>num(r.total_count)>0);rows.sort((a,b)=>num(b.total_count)-num(a.total_count));return rows}
function renderPlannedActualProjectStack(rows){const host=$("planned-actual-project-stack-chart");if(!host)return;const source=Array.isArray(rows)?rows:[];if(!source.length){host.innerHTML='<div class="pa-empty">No project data matched the selected filters.</div>';return}const maxTotal=Math.max(...source.map((r)=>num(r.total_count)),1);const head='<div class="pa-head"><div>Projects</div><div>Stacked Distribution</div><div style="text-align:right">Data Labels</div></div>';const body=source.map((r)=>{const total=Math.max(num(r.total_count),0);const trackPct=((total/maxTotal)*100).toFixed(2);const eqPct=total>0?((num(r.equal_count)/total)*100):0;const overPct=total>0?((num(r.overrun_count)/total)*100):0;const underPct=Math.max(0,100-eqPct-overPct);const eqLabel=eqPct>=15?String(Math.round(num(r.equal_count))):"";const overLabel=overPct>=15?String(Math.round(num(r.overrun_count))):"";const underLabel=underPct>=15?String(Math.round(num(r.under_count))):"";return `<div class="pa-row"><div class="pa-project">${txt(r.project_key)}</div><div class="pa-track"><div class="pa-stack" style="width:${trackPct}%"><span class="pa-seg eq" style="width:${eqPct.toFixed(2)}%" title="Planned == Actual: ${Math.round(num(r.equal_count))}">${eqLabel}</span><span class="pa-seg over" style="width:${overPct.toFixed(2)}%" title="Planned &lt; Actual: ${Math.round(num(r.overrun_count))}">${overLabel}</span><span class="pa-seg under" style="width:${underPct.toFixed(2)}%" title="Planned &gt; Actual: ${Math.round(num(r.under_count))}">${underLabel}</span></div></div><div class="pa-values">== ${Math.round(num(r.equal_count))} | &lt; ${Math.round(num(r.overrun_count))} | &gt; ${Math.round(num(r.under_count))} | Total ${Math.round(total)}</div></div>`}).join("");const legend='<div class="pa-legend"><span class="pa-pill"><span class="pa-dot eq"></span>Planned == Actual</span><span class="pa-pill"><span class="pa-dot over"></span>Planned &lt; Actual</span><span class="pa-pill"><span class="pa-dot under"></span>Planned &gt; Actual</span></div>';host.innerHTML=head+body+legend}
async function applyPage4(){readPage4InputsToState();const from=parseIso(page4Filters.from_date),to=parseIso(page4Filters.to_date);if(!from||!to||to<from){$("page4-status").textContent="Select a valid date range for Page 4.";renderPlannedActualProjectStack([]);return}const actualAgg=await loadActualAggregate(page4Filters.from_date,page4Filters.to_date,page4Filters.actual_mode).catch(()=>null);const tol=await loadPlannedActualTolerance();const rows=projectPlannedActualDistribution(page4Filters,actualAgg,tol);$("page4-status").textContent=`Tolerance: ${hTxt(tol)}h (${tol===0?"exact equality":"near-equality"}). Status: ${txt(page4Filters.status)}. Mode: ${txt(page4Filters.actual_mode)}.`;renderPlannedActualProjectStack(rows)}
function renderPendingBreakdown(invest,pending,totalEstimate,totalLogged,hoursRequired){const maxBase=Math.max(Math.abs(invest),pending,1);width("bar-investable-coverage",Math.max(invest,0),maxBase);width("bar-pending-required",Math.max(pending,0),maxBase);setTxt("legend-investable-2",`${hTxt(invest)}h`);setTxt("legend-pending",`${hTxt(pending)}h`);setTxt("cmp-total-estimate",`${hTxt(totalEstimate)}h`);setTxt("cmp-total-logged",`${hTxt(totalLogged)}h`);setTxt("cmp-hours-required",`${hTxt(hoursRequired)}h`);const gap=invest-pending;setTxt("cmp-gap",`${hTxt(gap)}h`);const pct=pending>0?(invest/pending)*100:100;const pctClamped=Math.max(0,Math.min(pct,100));const ring=$("coverage-ring");if(ring){const color=gap>=0?"#16a34a":"#dc2626";ring.style.background=`conic-gradient(${color} ${pctClamped.toFixed(2)}%, #e2e8f0 0)`}setTxt("coverage-pct",`${hTxt(pct)}%`);const coverageText=gap>=0?`Good: we have ${hTxt(gap)}h more than needed.`:`Need ${hTxt(Math.abs(gap))}h more to finish pending work.`;setTxt("coverage-text",coverageText)}
function insight(id,msg,kind){const c=$(id);c.classList.remove("ok","warn","risk");if(kind)c.classList.add(kind);$(id+"-text").textContent=msg}
function renderTopPending(rows){const host=$("top-pending-chart");if(!rows.length){host.innerHTML='<div class="vlabel">No epics matched the selected range.</div>';return}const top=rows.slice().sort((a,b)=>num(b.pending_hours)-num(a.pending_hours)).slice(0,8);const max=Math.max(...top.map((r)=>num(r.pending_hours)),1);host.innerHTML=top.map((r)=>`<div class="vrow"><div class="vlabel">${txt(r.issue_key)} - ${txt(r.summary)||"-"}</div><div class="vbar-shell"><div class="vbar" style="width:${((num(r.pending_hours)/max)*100).toFixed(2)}%"></div></div><div class="vvalue">${hTxt(r.pending_hours)}h</div></div>`).join("")}
function setProfileStatus(text){$("profile-status").textContent=text}
async function loadActualAggregate(fromIso,toIso,mode){if(!String(window.location.protocol||"").startsWith("http"))return null;const query=`from=${encodeURIComponent(fromIso)}&to=${encodeURIComponent(toIso)}&mode=${encodeURIComponent(mode)}&report=rnd_story`;const response=await fetch(`/api/actual-hours/aggregate?${query}`,{method:"GET"});const data=await response.json().catch(()=>({}));if(!response.ok||!data||data.ok===false)throw new Error(String(data&&data.error||"Failed to fetch actual-hour aggregate."));return data}
function setApplyBusy(busy){isApplyingStory=!!busy;const btn=$("apply-btn");if(!btn)return;btn.disabled=busy;btn.innerHTML=busy?'<span class="material-symbols-outlined" aria-hidden="true">hourglass_top</span> Applying...':'<span class="material-symbols-outlined" aria-hidden="true">tune</span> Apply'}
function addDays(dateObj,days){const d=new Date(dateObj);d.setDate(d.getDate()+days);return d}
function isoDateLocal(dateObj){const y=dateObj.getFullYear();const m=String(dateObj.getMonth()+1).padStart(2,"0");const d=String(dateObj.getDate()).padStart(2,"0");return `${y}-${m}-${d}`}
function dayDiffInclusive(from,to){const ms=Math.floor((to.getTime()-from.getTime())/(24*60*60*1000));return ms+1}
function syncPresetState(){const fromDate=txt($("from-date")?.value),toDate=txt($("to-date")?.value);const from=parseIso(fromDate),to=parseIso(toDate);let active="";if(from&&to&&to>=from){const today=new Date();const todayIso=isoDateLocal(today);const thisMonthStart=new Date(today.getFullYear(),today.getMonth(),1);const prevMonthStart=new Date(today.getFullYear(),today.getMonth()-1,1);const prevMonthEnd=new Date(today.getFullYear(),today.getMonth(),0);if(fromDate===isoDateLocal(addDays(today,-6))&&toDate===todayIso)active="7d";else if(fromDate===isoDateLocal(addDays(today,-13))&&toDate===todayIso)active="14d";else if(fromDate===isoDateLocal(addDays(today,-29))&&toDate===todayIso)active="30d";else if(fromDate===isoDateLocal(thisMonthStart)&&toDate===todayIso)active="month";else if(fromDate===isoDateLocal(prevMonthStart)&&toDate===isoDateLocal(prevMonthEnd))active="prev-month"}for(const chip of Array.from(document.querySelectorAll(".chip[data-preset]"))){chip.classList.toggle("is-active",txt(chip.getAttribute("data-preset"))===active)}}
function updateDateRangeHint(){const hint=$("date-range-hint");if(!hint)return false;const fromDate=txt($("from-date")?.value),toDate=txt($("to-date")?.value);const from=parseIso(fromDate),to=parseIso(toDate);hint.classList.remove("ok","warn");if(!from||!to){hint.textContent="Select both start and end dates.";hint.classList.add("warn");return false}if(to<from){hint.textContent="End date must be on or after start date.";hint.classList.add("warn");return false}const days=dayDiffInclusive(from,to);hint.textContent=`${fromDate} to ${toDate} (${days} day${days===1?"":"s"})`;hint.classList.add("ok");syncPresetState();return true}
function setToolbarDirty(isDirty){$("filters-toolbar")?.classList.toggle("is-dirty",!!isDirty)}
function applyPreset(preset){const today=new Date();let from=today;let to=today;if(preset==="7d"){from=addDays(today,-6)}else if(preset==="14d"){from=addDays(today,-13)}else if(preset==="30d"){from=addDays(today,-29)}else if(preset==="month"){from=new Date(today.getFullYear(),today.getMonth(),1)}else if(preset==="prev-month"){from=new Date(today.getFullYear(),today.getMonth()-1,1);to=new Date(today.getFullYear(),today.getMonth(),0)}$("from-date").value=isoDateLocal(from);$("to-date").value=isoDateLocal(to);updateDateRangeHint();setToolbarDirty(false);applyStory()}
function attachCardInfo(){const availabilityFormula=managedFieldFormulaText("availability","capacity-planned_leaves");const moreWorkFormula=managedFieldFormulaText("capacity_available_for_more_work","capacity-planned_hours-planned_leaves");const requiredFormula=managedFieldFormulaText("hours_required_to_complete_projects","planned_hours-actual_hours");const monthlyTotalFormula=managedFieldFormulaText("total_planned_hours","sum(previous_month_planned_hours+current_month_planned_hours+upcoming_planned_hours)");const previousMonthlyFormula=managedFieldFormulaText("previous_month_planned_hours","sum(previous_month_planned_hours)");const currentMonthlyFormula=managedFieldFormulaText("current_month_planned_hours","sum(current_month_planned_hours)");const upcomingMonthlyFormula=managedFieldFormulaText("upcoming_planned_hours","sum(upcoming_planned_hours)");const formulaByTitle={"total capacity (leaves adjusted)":"Formula: Total Capacity (Leaves Adjusted) = "+availabilityFormula,"total planned projects (hours)":"Formula: Total Planned Projects (Hours) = Sum(Epic Original Estimate Hours), where Epic Start OR End is in selected range, excluding RLT","capacity available for more work":"Formula: Capacity available for more work = "+moreWorkFormula,"total actual project hours":"Formula: Total Actual Project Hours = Sum(Project Actual Hours), excluding RLT","hours required to complete projects":"Formula: Hours Required To Complete Projects = "+requiredFormula,"hours logged (invested hours)":"Formula: Hours Logged (Invested hours) = Sum(Actual Hours Logged against all epics/projects), excluding RLT RnD Leave Tracker","capacity barrel view":"Formula: Barrel Capacity = Total Capacity (Leaves Adjusted); Target Mark = Total Planned Projects; Fluid = Total Actual Project Hours","fluid level (total actual project hours)":"Formula: Fluid Level = Total Actual Project Hours","pending hours required":"Formula: Pending Hours Required = Sum(max(Epic Original Estimate Hours - Epic Logged Hours, 0))","closed (resolved) epics":"Formula: Count of epics with status containing Resolved in selected range","open (in progress) epics":"Formula: Count of epics with status containing In Progress in selected range","capacity commitment funnel":"Formula: Visual flow of Available Capacity -> Available Capacity (Leaves Adjusted) -> Booked (Work) -> Hours Logged (Invested hours) -> Hours Required To Complete Projects -> Capacity available for more work","do we have enough hours?":"Formula: Coverage Gap = Investable More Hours - Pending Hours Required","top epics needing hours":"Formula: Top by Pending Hours per epic = max(Epic Original Estimate Hours - Epic Logged Hours, 0)","epic status split":"Formula: Resolved vs In Progress epic counts in selected range","capacity surplus/deficit":"Formula: Surplus/Deficit = Investable More Hours","pending demand coverage gap":"Formula: Coverage Gap = Investable More Hours - Pending Hours Required","epic status pressure":"Formula: Pressure compares Open (In Progress) count vs Closed (Resolved) count","total planned hours":"Formula: Sum(Epic Original Estimate Hours) for selected epics","hours already logged":"Formula: Sum(Epic Logged Hours) for selected epics","hours still needed":"Formula: Hours still needed = Total planned hours - Hours already logged","difference":"Formula: Difference = Investable More Hours - Pending Hours Required","capacity decision":"Formula: Decision from sign of Investable More Hours","demand decision":"Formula: Decision from sign of (Investable More Hours - Pending Hours Required)","flow decision":"Formula: Decision from comparison of Open vs Closed epic counts","previous month planned hours":"Formula: Previous Month Planned Hours = "+previousMonthlyFormula,"current month planned hours":"Formula: Current Month Planned Hours = "+currentMonthlyFormula,"upcoming planned hours":"Formula: Upcoming Planned Hours = "+upcomingMonthlyFormula,"project planned hours stack chart (prev vs current vs upcoming)":"Formula: Total Planned Hours = "+monthlyTotalFormula,"project planned hours scorecards (click to filter epics)":"Formula: Project Planned Hours = Sum(Epic Original Estimate Hours) grouped by project in selected date range.","epic planned hours bar chart (data labels shown)":"Formula: Epic Planned Hours = Epic Original Estimate Hours, grouped project-wise and sorted descending within each project; color is project-specific.","stacked distribution by project (epic counts)":"Formula: For each project, count selected epics by three buckets: Planned == Actual, Planned < Actual, Planned > Actual. Equality uses admin tolerance (hours).","planned vs actual per project":"Formula: Uses page-specific date, actual-hours mode, and status filter; Both = In Progress + Resolved only."};const cards=document.querySelectorAll(".kpi,.story-card,.compare-item,.break-cell,.insight,.summary-card");for(const card of cards){const old=card.querySelector(".card-i-wrap");if(old)old.remove();const titleNode=card.querySelector(".label,.story-title,h3,h4,.k");const title=titleNode?txt(titleNode.textContent):"This card";const key=title.toLowerCase().replace(/`/g,"").replace(/\\s+/g," ").trim();const formula=formulaByTitle[key]||"Formula: Value is computed from selected date range, RnD scope, and applied capacity profile.";const host=document.createElement("span");host.className="card-i-wrap";host.textContent="i";const tip=document.createElement("span");tip.className="card-i-tip";tip.textContent=title+"\\n"+formula;host.appendChild(tip);card.appendChild(host);}}
function storyPages(){return Array.from(document.querySelectorAll(".story-page"))}
function showStoryPage(index){const pages=storyPages();if(!pages.length)return;storyIndex=Math.max(0,Math.min(index,pages.length-1));pages.forEach((p,i)=>p.classList.toggle("active",i===storyIndex));$("story-page-label").textContent=`Page ${storyIndex+1} of ${pages.length}`;$("story-prev").disabled=storyIndex===0;$("story-next").disabled=storyIndex===pages.length-1}
function updateSummary(after,work,invest,pending,closed,open){$("summary-capacity").textContent=invest>=0?`Capacity can absorb current commitment with ${hTxt(invest)}h headroom.`:`Capacity is over-committed by ${hTxt(Math.abs(invest))}h.`;const gap=invest-pending;$("summary-demand").textContent=gap>=0?`Pending demand is fully coverable with ${hTxt(gap)}h remaining.`:`Pending demand exceeds investable hours by ${hTxt(Math.abs(gap))}h.`;$("summary-flow").textContent=open>closed?`Delivery pressure is high: ${Math.round(open)} in progress vs ${Math.round(closed)} resolved.`:`Flow is stable: ${Math.round(closed)} resolved vs ${Math.round(open)} in progress.`}
async function applyStory(){
const fromDate=txt($("from-date").value),toDate=txt($("to-date").value);
const from=parseIso(fromDate),to=parseIso(toDate);
if(!from||!to||to<from){updateDateRangeHint();setProfileStatus("Pick a valid date range before applying.");return}
const modeSel=$("actual-hours-mode");const actualMode=modeSel?txt(modeSel.value)||ACTUAL_MODE_DEFAULT:ACTUAL_MODE_DEFAULT;localStorage.setItem(ACTUAL_MODE_STORAGE_KEY,actualMode);
setApplyBusy(true);
setToolbarDirty(false);
try{
const cap=await loadCapacity(fromDate,toDate);
const actualAgg=await loadActualAggregate(fromDate,toDate,actualMode).catch((err)=>{setProfileStatus(`Actual-hours mode fallback: ${txt(err?.message||err)}`);return null;});
const available=num(cap?.metrics?.available_capacity_hours);
const leave=cap?.leave_metrics||{};
const leaveTotal=num(leave.planned_taken_hours)+num(leave.planned_not_taken_hours||leave.not_yet_taken_hours)+num(leave.unplanned_taken_hours);
const afterDefault=+(available-leaveTotal).toFixed(2);
const work=plannedWorkOnPlate(from,to);
const actualProjects=totalActualProjectHours(from,to,actualAgg);
const hoursRequiredDefault=+(work-actualProjects).toFixed(2);
const context={capacity:available,planned_hours:work,actual_hours:actualProjects,planned_leaves:leaveTotal};
const availabilityEval=evaluateManagedField("availability",afterDefault,context);
const after=+num(availabilityEval.value).toFixed(2);
const requiredEval=evaluateManagedField("hours_required_to_complete_projects",hoursRequiredDefault,context);
const hoursRequiredProjects=+num(requiredEval.value).toFixed(2);
const capacityMoreWorkDefault=+(after-work).toFixed(2);
const capacityMoreWorkEval=evaluateManagedField("capacity_available_for_more_work",capacityMoreWorkDefault,context);
const capacityMoreWork=+num(capacityMoreWorkEval.value).toFixed(2);
const compareRows=projectCompareRows(from,to,actualAgg);
const monthlyPlanned=monthlyPlannedByProject();
$("kpi-capacity-after-leaves").textContent=`${hTxt(after)}h`;
$("kpi-work-on-plate").textContent=`${hTxt(work)}h`;
$("kpi-total-actual-project-hours").textContent=`${hTxt(actualProjects)}h`;
$("kpi-hours-required-projects").textContent=`${hTxt(hoursRequiredProjects)}h`;
renderLeadershipCards(after,work,actualProjects,hoursRequiredProjects);
renderLeadershipBarrel(after,work,actualProjects,hoursRequiredProjects,capacityMoreWork);
renderProjectPlanVsActual(compareRows);
renderMonthlyPlannedKpis(monthlyPlanned.meta);
renderMonthlyPlannedStack(monthlyPlanned.rows,monthlyPlanned.meta);
latestEpicRows=epicPlannedRowsByProject(from,to);
renderProjectPlannedScorecards(projectScorecardRows(latestEpicRows));
renderEpicPlannedHoursBars(latestEpicRows,selectedEpicProjectKey);
attachCardInfo();
}catch(err){setProfileStatus(`Unable to apply filters: ${txt(err?.message||err)||"unknown error"}`)}
finally{setApplyBusy(false)}
}
function resetFilters(){const d=payload.defaults||{};$("from-date").value=txt(d.from_date);$("to-date").value=txt(d.to_date);activeProfileKey="";setProfileStatus("Capacity source: default profile logic.");updateDateRangeHint();setToolbarDirty(false);applyStory()}
function applySelectedProfile(){const select=$("capacity-profile-select");const key=txt(select.value);if(!key){setProfileStatus("No profile selected.");return}activeProfileKey=key;const p=findProfileByKey(key);if(!p){setProfileStatus("Selected profile is unavailable.");return}setProfileStatus(`Capacity source: applied profile ${txt(p.from_date)} to ${txt(p.to_date)}.`);setToolbarDirty(false);applyStory()}
(function init(){if(typeof mermaid!=="undefined"){mermaid.initialize({startOnLoad:false,securityLevel:"loose",theme:"base",flowchart:{htmlLabels:true,curve:"basis"},themeVariables:{primaryColor:"#dbeafe",primaryBorderColor:"#93c5fd",lineColor:"#0f766e",fontFamily:"Segoe UI"}});mermaidReady=true;}const s=payload.source_files||{};$("hero-meta").textContent=[`Department: ${txt(payload.department_name)||"RnD"}`,`Generated: ${txt(payload.generated_at)||"-"}`,`WorkItems: ${txt(s.work_items)||"-"}`,`Worklogs: ${txt(s.subtask_worklogs)||"-"}`].join(" | ");populateProfileSelect();const modeEl=$("actual-hours-mode");if(modeEl){const stored=localStorage.getItem(ACTUAL_MODE_STORAGE_KEY);modeEl.value=(stored==="planned_dates"||stored==="log_date")?stored:ACTUAL_MODE_DEFAULT;modeEl.addEventListener("change",()=>{setToolbarDirty(false);applyStory();});}$("apply-profile-btn").addEventListener("click",applySelectedProfile);$("apply-btn").addEventListener("click",applyStory);$("reset-btn").addEventListener("click",resetFilters);for(const inputId of ["from-date","to-date"]){const input=$(inputId);if(!input)continue;input.addEventListener("input",()=>{updateDateRangeHint();setToolbarDirty(true)});input.addEventListener("change",()=>{updateDateRangeHint();setToolbarDirty(true)});input.addEventListener("keydown",(e)=>{if(e.key==="Enter"){e.preventDefault();applyStory()}})}for(const chip of Array.from(document.querySelectorAll(".chip[data-preset]"))){chip.addEventListener("click",()=>applyPreset(txt(chip.getAttribute("data-preset"))))}if($("apply-page4-btn"))$("apply-page4-btn").addEventListener("click",()=>{applyPage4()});if($("reset-page4-btn"))$("reset-page4-btn").addEventListener("click",()=>{resetPage4Filters();applyPage4()});for(const page4InputId of ["from-date-page4","to-date-page4"]){const input=$(page4InputId);if(!input)continue;input.addEventListener("keydown",(e)=>{if(e.key==="Enter"){e.preventDefault();applyPage4()}})}if($("actual-hours-mode-page4"))$("actual-hours-mode-page4").addEventListener("change",()=>applyPage4());if($("status-page4"))$("status-page4").addEventListener("change",()=>applyPage4());$("story-prev").addEventListener("click",()=>showStoryPage(storyIndex-1));$("story-next").addEventListener("click",()=>showStoryPage(storyIndex+1));document.addEventListener("keydown",(e)=>{if(e.key==="ArrowLeft")showStoryPage(storyIndex-1);if(e.key==="ArrowRight")showStoryPage(storyIndex+1)});showStoryPage(0);resetPage4Filters();setPage4InputsFromState();loadManagedFieldsFromPayloadOrApi().then(()=>{attachCardInfo();resetFilters();applyPage4()}).catch(()=>{attachCardInfo();resetFilters();applyPage4()})})();
</script><script src="shared-nav.js"></script>
</body></html>
"""
    return template.replace("__RND_DATA_STORY_PAYLOAD__", data)


def _resolve_runtime_paths(base_dir: Path) -> dict[str, Path]:
    work_items_name = os.getenv("JIRA_EXPORT_XLSX_PATH", DEFAULT_WORK_ITEMS_INPUT_XLSX).strip() or DEFAULT_WORK_ITEMS_INPUT_XLSX
    worklog_name = os.getenv("JIRA_WORKLOG_XLSX_PATH", DEFAULT_WORKLOG_INPUT_XLSX).strip() or DEFAULT_WORKLOG_INPUT_XLSX
    assignee_hours_name = os.getenv("JIRA_ASSIGNEE_HOURS_XLSX_PATH", DEFAULT_ASSIGNEE_HOURS_INPUT_XLSX).strip() or DEFAULT_ASSIGNEE_HOURS_INPUT_XLSX
    leave_name = os.getenv("JIRA_LEAVE_REPORT_XLSX_PATH", DEFAULT_LEAVE_REPORT_INPUT_XLSX).strip() or DEFAULT_LEAVE_REPORT_INPUT_XLSX
    db_name = os.getenv("JIRA_ASSIGNEE_HOURS_CAPACITY_DB_PATH", DEFAULT_CAPACITY_DB).strip() or DEFAULT_CAPACITY_DB
    nested_name = os.getenv("JIRA_NESTED_INPUT_XLSX_PATH", DEFAULT_NESTED_VIEW_INPUT_XLSX).strip() or DEFAULT_NESTED_VIEW_INPUT_XLSX
    html_name = os.getenv("JIRA_RND_STORY_HTML_PATH", DEFAULT_HTML_OUTPUT).strip() or DEFAULT_HTML_OUTPUT
    dataset_name = os.getenv("JIRA_RND_STORY_DATASET_XLSX_PATH", DEFAULT_DATASET_OUTPUT).strip() or DEFAULT_DATASET_OUTPUT
    return {
        "work_items_path": _resolve_path(work_items_name, base_dir),
        "worklog_path": _resolve_path(worklog_name, base_dir),
        "assignee_hours_path": _resolve_path(assignee_hours_name, base_dir),
        "leave_report_path": _resolve_path(leave_name, base_dir),
        "capacity_db_path": _resolve_path(db_name, base_dir),
        "nested_view_path": _resolve_path(nested_name, base_dir),
        "output_html_path": _resolve_path(html_name, base_dir),
        "output_dataset_path": _resolve_path(dataset_name, base_dir),
    }


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate RnD capacity and workload data story HTML.")
    parser.add_argument("--output-html", default="", help="Optional output HTML path override.")
    parser.add_argument("--output-dataset", default="", help="Optional output dataset XLSX path override.")
    return parser.parse_args()


def main() -> None:
    args = _parse_args()
    base_dir = Path(__file__).resolve().parent
    paths = _resolve_runtime_paths(base_dir)
    output_path = paths["output_html_path"]
    output_dataset_path = paths["output_dataset_path"]
    if args.output_html.strip():
        output_path = _resolve_path(args.output_html.strip(), base_dir)
    if args.output_dataset.strip():
        output_dataset_path = _resolve_path(args.output_dataset.strip(), base_dir)
    paths["output_dataset_path"] = output_dataset_path

    payload = _build_payload(paths)
    _write_page1_dataset_xlsx(output_dataset_path, payload.get("page1_dataset") or {})
    output_path.write_text(_build_html(payload), encoding="utf-8")
    print("RnD data story generated")
    print(f"Dataset: {output_dataset_path}")
    print(f"Output: {output_path}")


if __name__ == "__main__":
    main()




























