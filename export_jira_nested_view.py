"""
Export Jira issues into a hierarchical nested-view Excel report.

Hierarchy:
1) Project Key - Project Name
2) Product Categorization
3) RMI name (Epic title)
4) Story name
5) Sub Task name
6) Assignee name of the sub task
"""
from __future__ import annotations

import argparse
import os
import sqlite3
import time
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook, load_workbook

from ipp_meeting_utils import normalize_issue_key
from ipp_meeting_utils import resolve_jira_start_date_field_id
from jira_incremental_cache import (
    apply_overlap,
    bootstrap_default_checkpoint,
    get_cached_issue_payloads,
    get_changed_or_new_issue_keys,
    get_db_path,
    get_or_init_checkpoint,
    init_db,
    mark_full_sync,
    mark_missing_issues_deleted,
    needs_full_sync,
    parse_iso_utc,
    set_checkpoint,
    upsert_issue_index,
    upsert_issue_payloads,
    utc_now_iso,
)
from jira_client import BASE_URL, get_session
from jira_export_db import connect as export_db_connect
from jira_export_db import ensure_schema as ensure_exports_schema
from jira_export_db import has_subtask_worklog_rollup
from jira_export_db import read_subtask_worklog_rollup as read_subtask_worklog_rollup_db
from jira_export_db import record_export_run
from jira_export_db import write_nested_view_nodes as write_nested_view_nodes_db

DEFAULT_PROJECT_KEYS = ["DIGITALLOG", "FF", "O2", "ODL", "MN"]
DEFAULT_OUTPUT = "nested view.xlsx"
DEFAULT_ROLLUP_INPUT = "3_jira_subtask_worklog_rollup.xlsx"
WORK_ITEM_TYPES = ["Epic", "Story", "Task", "Sub-task", "Subtask", "Bug Task", "Bug Subtask"]

PROJECT_NAME_BY_KEY = {
    "FF": "Fintech Fuel",
    "O2": "OmniConnect",
    "MN": "Omni Chat",
    "ODL": "ODL Miscellaneous",
    "DIGITALLOG": "Digital Log",
}

HEADERS = [
    "Aspect",
    "Man-days",
    "Man-hours",
    "Actual Hours",
    "Actual Days",
    "Planned Start Date",
    "Planned End Date",
]


@dataclass
class IssueNode:
    key: str
    kind: str
    project_key: str
    summary: str
    parent_key: str
    assignee: str
    product_category: str
    epic_key: str
    story_key: str
    man_hours: float
    man_days: float
    actual_hours: float
    planned_start: str
    planned_end: str


def _resolve_capacity_settings_db_path() -> Path:
    value = (os.getenv("JIRA_ASSIGNEE_HOURS_CAPACITY_DB_PATH", "assignee_hours_capacity.db") or "").strip()
    path = Path(value or "assignee_hours_capacity.db")
    if path.is_absolute():
        return path
    return Path(__file__).resolve().parent / path


def _load_project_keys_from_managed_db(db_path: Path) -> list[str]:
    if not db_path.exists():
        return []
    conn = sqlite3.connect(db_path)
    try:
        rows = conn.execute(
            """
            SELECT project_key
            FROM managed_projects
            WHERE is_active = 1
            ORDER BY project_key ASC
            """
        ).fetchall()
    except sqlite3.Error:
        return []
    finally:
        conn.close()
    return [str(row[0]).strip().upper() for row in rows if str(row[0]).strip()]


def _get_project_keys() -> tuple[list[str], str]:
    db_path = _resolve_capacity_settings_db_path()
    db_keys = _load_project_keys_from_managed_db(db_path)
    if db_keys:
        return db_keys, "managed_projects_db"
    raw = os.getenv("JIRA_PROJECT_KEYS", "")
    if not raw.strip():
        return DEFAULT_PROJECT_KEYS, "env_fallback"
    return [key.strip() for key in raw.split(",") if key.strip()], "env_fallback"


def _is_incremental_disabled(enable_incremental: bool) -> bool:
    if enable_incremental:
        return False
    return (os.getenv("JIRA_INCREMENTAL_DISABLE", "1").strip() or "1") == "1"


def _get_overlap_minutes() -> int:
    raw = os.getenv("JIRA_INCREMENTAL_OVERLAP_MINUTES", "5").strip() or "5"
    try:
        return max(int(raw), 0)
    except ValueError:
        return 5


def _get_force_full_days() -> int:
    raw = os.getenv("JIRA_FORCE_FULL_SYNC_DAYS", "7").strip() or "7"
    try:
        return max(int(raw), 1)
    except ValueError:
        return 7


def _get_bootstrap_days() -> int:
    raw = os.getenv("JIRA_INCREMENTAL_BOOTSTRAP_DAYS", "365").strip() or "365"
    try:
        return max(int(raw), 1)
    except ValueError:
        return 365


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Export Jira nested-view report to Excel.")
    parser.add_argument(
        "--incremental",
        action="store_true",
        help="Enable smart incremental fetch (default: full fetch).",
    )
    return parser.parse_args()


def _to_jql_updated_value(value: str) -> str:
    return parse_iso_utc(value).strftime("%Y-%m-%d %H:%M")


def _fetch_issues(session, jql: str, fields: list[str]) -> list[dict]:
    url = f"{BASE_URL}/rest/api/3/search/jql"
    issues: list[dict] = []
    next_page_token = None

    while True:
        payload = {"jql": jql, "maxResults": 100, "fields": fields}
        if next_page_token:
            payload["nextPageToken"] = next_page_token

        response = None
        for attempt in range(6):
            response = session.post(url, json=payload)
            if response.status_code == 429:
                retry_after = response.headers.get("Retry-After")
                try:
                    retry_after_seconds = float(retry_after) if retry_after else 0.0
                except ValueError:
                    retry_after_seconds = 0.0
                backoff_seconds = max((0.5 * (2**attempt)), retry_after_seconds, 0.5)
                time.sleep(backoff_seconds)
                continue
            response.raise_for_status()
            break
        if response is None:
            break
        if response.status_code == 429:
            response.raise_for_status()
        data = response.json()

        issues.extend(data.get("issues", []))
        next_page_token = data.get("nextPageToken")
        if not next_page_token:
            break

    return issues


def _get_all_fields(session) -> list[dict]:
    response = session.get(f"{BASE_URL}/rest/api/3/field")
    response.raise_for_status()
    return response.json()


def _get_product_categorization_field_id(all_fields: list[dict]) -> str:
    env_field = os.getenv("JIRA_PRODUCT_CATEGORIZATION_FIELD_ID", "").strip()
    if env_field:
        return env_field

    for field in all_fields:
        name = (field.get("name") or "").strip().lower()
        if name == "product categorization":
            return field.get("id", "")
    return ""


def _build_discovery_jql(project_keys: list[str], from_updated_utc: str | None) -> str:
    keys_str = ", ".join(project_keys)
    base = (
        f'project in ({keys_str}) AND issuetype in ("Epic", "Story", "Task", '
        f'"Sub-task", "Subtask", "Bug Task", "Bug Subtask")'
    )
    if from_updated_utc:
        return f'{base} AND updated >= "{_to_jql_updated_value(from_updated_utc)}" ORDER BY updated ASC'
    return f"{base} ORDER BY updated ASC"


def _candidate_rows_from_issues(issues: list[dict]) -> list[dict]:
    now_utc = utc_now_iso()
    rows: list[dict] = []
    for issue in issues:
        fields = issue.get("fields", {}) or {}
        rows.append(
            {
                "issue_id": str(issue.get("id", "")).strip(),
                "issue_key": str(issue.get("key", "")).strip(),
                "updated_utc": str(fields.get("updated", "")).strip(),
                "issue_type": str((fields.get("issuetype") or {}).get("name", "")).strip(),
                "project_key": str((fields.get("project") or {}).get("key", "")).strip(),
                "last_seen_utc": now_utc,
                "is_deleted": 0,
            }
        )
    return [row for row in rows if row["issue_id"] and row["issue_key"]]


def _fetch_issues_by_keys(session, issue_keys: list[str], fields: list[str]) -> list[dict]:
    if not issue_keys:
        return []
    results: list[dict] = []
    for offset in range(0, len(issue_keys), 500):
        chunk = issue_keys[offset : offset + 500]
        keys_clause = ", ".join(f'"{k}"' for k in chunk)
        jql = f"key in ({keys_clause})"
        results.extend(_fetch_issues(session, jql=jql, fields=fields))
    return results


def _get_active_issue_keys(conn: sqlite3.Connection, project_keys: list[str], issue_types: list[str]) -> list[str]:
    if not project_keys or not issue_types:
        return []
    key_ph = ",".join("?" for _ in project_keys)
    type_ph = ",".join("?" for _ in issue_types)
    rows = conn.execute(
        f"""
        SELECT issue_key
        FROM issue_index
        WHERE is_deleted = 0
          AND project_key IN ({key_ph})
          AND issue_type IN ({type_ph})
        ORDER BY issue_key
        """,
        tuple(project_keys + issue_types),
    ).fetchall()
    return [str(row[0]) for row in rows]


def _extract_parent_key(fields: dict) -> str:
    parent = fields.get("parent") or {}
    parent_key = parent.get("key")
    if parent_key:
        return str(parent_key).strip()

    epic_link = fields.get("customfield_10014")
    if isinstance(epic_link, str) and epic_link.strip():
        return epic_link.strip()
    if isinstance(epic_link, dict):
        epic_key = epic_link.get("key")
        if epic_key:
            return str(epic_key).strip()
    return ""


def _normalize_kind(issue_type_name: str) -> str:
    name = (issue_type_name or "").strip().lower()
    if "epic" in name:
        return "epic"
    if "sub-task" in name or "subtask" in name or ("bug" in name and "sub" in name):
        return "subtask"
    if "story" in name or "task" in name:
        return "story"
    return "other"


def _seconds_to_hours(seconds_value) -> float:
    if seconds_value in (None, ""):
        return 0.0
    try:
        return round(float(seconds_value) / 3600.0, 2)
    except (TypeError, ValueError):
        return 0.0


def _hours_to_days(hours: float) -> float:
    try:
        return round(float(hours) / 8.0, 2)
    except (TypeError, ValueError):
        return 0.0


def _text_or_default(value, default_text: str = "") -> str:
    if value is None:
        return default_text
    text = str(value).strip()
    return text if text else default_text


def _parse_product_category(value) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, dict):
        for key in ("value", "name", "displayName", "label"):
            candidate = value.get(key)
            if candidate:
                return str(candidate).strip()
        return ""
    if isinstance(value, list):
        parts = []
        for item in value:
            parsed = _parse_product_category(item)
            if parsed:
                parts.append(parsed)
        return ", ".join(parts)
    return str(value).strip()


def _load_subtask_rollup_values(rollup_path: Path) -> dict[str, dict[str, object]]:
    if not rollup_path.exists():
        print(f"Warning: Rollup workbook not found. Subtask actual hours default to 0: {rollup_path}")
        return {}

    wb = load_workbook(rollup_path, read_only=True, data_only=True)
    ws = wb.active
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header:
        wb.close()
        print("Warning: Rollup workbook has no header row. Subtask actual hours default to 0.")
        return {}

    headers = [str(h).strip() if h is not None else "" for h in header]
    required = ["issue_id", "total hours_logged", "planned start date", "planned end date"]
    missing = [name for name in required if name not in headers]
    if missing:
        wb.close()
        print(f"Warning: Rollup workbook missing required columns {missing}. Subtask actual hours default to 0.")
        return {}

    issue_idx = headers.index("issue_id")
    actual_idx = headers.index("total hours_logged")
    planned_start_idx = headers.index("planned start date")
    planned_end_idx = headers.index("planned end date")

    result: dict[str, dict[str, object]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if issue_idx >= len(row):
            continue
        issue_key = normalize_issue_key(row[issue_idx])
        if not issue_key:
            continue
        value = row[actual_idx] if actual_idx < len(row) else 0
        planned_start_value = row[planned_start_idx] if planned_start_idx < len(row) else ""
        planned_end_value = row[planned_end_idx] if planned_end_idx < len(row) else ""
        try:
            actual_hours = round(float(value or 0), 2)
        except (TypeError, ValueError):
            actual_hours = 0.0
        result[issue_key] = {
            "actual_hours": actual_hours,
            "planned_start": _text_or_default(planned_start_value),
            "planned_end": _text_or_default(planned_end_value),
        }

    wb.close()
    return result


def _resolve_output_path(value: str, script_dir: Path) -> Path:
    path = Path(value)
    if path.is_absolute():
        return path
    return script_dir / path


def _build_nodes(
    issues: list[dict],
    start_date_field_id: str,
    product_field_id: str,
    subtask_rollup_values: dict[str, dict[str, object]],
) -> tuple[dict[str, IssueNode], dict[str, IssueNode], dict[str, IssueNode]]:
    epics: dict[str, IssueNode] = {}
    stories: dict[str, IssueNode] = {}
    subtasks: dict[str, IssueNode] = {}

    raw_items: dict[str, dict] = {}
    for issue in issues:
        key = normalize_issue_key(issue.get("key", ""))
        if not key:
            continue
        fields = issue.get("fields", {}) or {}
        issue_type_name = ((fields.get("issuetype") or {}).get("name") or "").strip()
        kind = _normalize_kind(issue_type_name)
        if kind == "other":
            continue

        project_key = _text_or_default((fields.get("project") or {}).get("key", ""), "UNKNOWN")
        summary = _text_or_default(fields.get("summary"), key)
        parent_key = normalize_issue_key(_extract_parent_key(fields))
        assignee = _text_or_default((fields.get("assignee") or {}).get("displayName"), "Unassigned")
        # Planned dates in nested view come from Jira planned start and Jira due date.
        start_date = _text_or_default(fields.get(start_date_field_id) if start_date_field_id else "")
        due_date = _text_or_default(fields.get("duedate"))

        man_hours = _seconds_to_hours(fields.get("timeoriginalestimate"))
        man_days = _hours_to_days(man_hours)
        # "Actual" metrics in nested view represent logged time from Jira/worklog rollups.
        own_actual_hours = _seconds_to_hours(fields.get("aggregatetimespent") or fields.get("timespent"))

        product_value = ""
        if product_field_id:
            product_value = _parse_product_category(fields.get(product_field_id))

        raw_items[key] = {
            "key": key,
            "kind": kind,
            "project_key": project_key,
            "summary": summary,
            "parent_key": parent_key,
            "assignee": assignee,
            "product_category": product_value,
            "man_hours": man_hours,
            "man_days": man_days,
            "actual_hours": own_actual_hours,
            "planned_start": start_date,
            "planned_end": due_date,
        }

    for key, item in raw_items.items():
        if item["kind"] != "epic":
            continue
        product_category = _text_or_default(item["product_category"], "Uncategorized")
        epics[key] = IssueNode(
            key=key,
            kind="epic",
            project_key=item["project_key"],
            summary=item["summary"],
            parent_key="",
            assignee=item["assignee"],
            product_category=product_category,
            epic_key=key,
            story_key="",
            man_hours=item["man_hours"],
            man_days=item["man_days"],
            actual_hours=item["actual_hours"],
            planned_start=item["planned_start"],
            planned_end=item["planned_end"],
        )

    for key, item in raw_items.items():
        if item["kind"] != "story":
            continue
        parent_key = normalize_issue_key(item["parent_key"])
        epic_key = parent_key if parent_key in epics else ""
        epic = epics.get(epic_key)
        product_category = _text_or_default(item["product_category"]) or (epic.product_category if epic else "")
        product_category = _text_or_default(product_category, "Uncategorized")

        stories[key] = IssueNode(
            key=key,
            kind="story",
            project_key=item["project_key"],
            summary=item["summary"],
            parent_key=parent_key,
            assignee=item["assignee"],
            product_category=product_category,
            epic_key=epic_key,
            story_key=key,
            man_hours=item["man_hours"],
            man_days=item["man_days"],
            actual_hours=item["actual_hours"],
            planned_start=item["planned_start"],
            planned_end=item["planned_end"],
        )

    for key, item in raw_items.items():
        if item["kind"] != "subtask":
            continue
        parent_story_key = normalize_issue_key(item["parent_key"])
        story = stories.get(parent_story_key)
        epic = epics.get(story.epic_key) if story and story.epic_key else None
        product_category = _text_or_default(item["product_category"]) or (story.product_category if story else "")
        if not product_category and epic:
            product_category = epic.product_category
        product_category = _text_or_default(product_category, "Uncategorized")

        normalized_key = normalize_issue_key(key)
        rollup = subtask_rollup_values.get(normalized_key, {})
        # Subtask actual hours are sourced from worklog rollup for logged-time accuracy.
        actual_hours = round(float(rollup.get("actual_hours", 0.0) or 0.0), 2)
        planned_start = _text_or_default(rollup.get("planned_start")) or item["planned_start"]
        planned_end = _text_or_default(rollup.get("planned_end")) or item["planned_end"]

        subtasks[key] = IssueNode(
            key=key,
            kind="subtask",
            project_key=item["project_key"] if item["project_key"] != "UNKNOWN" else (story.project_key if story else "UNKNOWN"),
            summary=item["summary"],
            parent_key=parent_story_key,
            assignee=item["assignee"],
            product_category=product_category,
            epic_key=story.epic_key if story else "",
            story_key=parent_story_key,
            man_hours=item["man_hours"],
            man_days=item["man_days"],
            actual_hours=round(actual_hours, 2),
            planned_start=planned_start,
            planned_end=planned_end,
        )

    return epics, stories, subtasks


def _as_metric(value: float | str) -> float | str:
    if value == "":
        return ""
    try:
        return round(float(value), 2)
    except (TypeError, ValueError):
        return ""


def _append_row(
    ws,
    outline_level: int,
    aspect: str,
    man_days: float | str,
    man_hours: float | str,
    actual_hours: float | str,
    actual_days: float | str,
    planned_start: str,
    planned_end: str,
) -> int:
    ws.append(
        [
            aspect,
            _as_metric(man_days),
            _as_metric(man_hours),
            _as_metric(actual_hours),
            _as_metric(actual_days),
            planned_start,
            planned_end,
        ]
    )
    row_num = ws.max_row
    row_dim = ws.row_dimensions[row_num]
    row_dim.outlineLevel = outline_level
    row_dim.hidden = False
    row_dim.collapsed = False
    return row_num


def _sorted_keys_with_preferred_order(keys: set[str], preferred: list[str]) -> list[str]:
    preferred_rank = {value: index for index, value in enumerate(preferred)}
    return sorted(keys, key=lambda k: (preferred_rank.get(k, 9999), k))


def _parse_iso_date(value: str) -> datetime | None:
    text = _text_or_default(value)
    if not text:
        return None
    try:
        return datetime.fromisoformat(text)
    except ValueError:
        pass
    try:
        return datetime.fromisoformat(f"{text}T00:00:00")
    except ValueError:
        return None


def _merge_date_bounds(values: list[str]) -> tuple[str, str]:
    parsed = [dt for dt in (_parse_iso_date(value) for value in values) if dt is not None]
    if not parsed:
        return "", ""
    return min(parsed).date().isoformat(), max(parsed).date().isoformat()


def _aggregate_metrics(items: list[dict[str, object]]) -> dict[str, object]:
    man_days = round(sum(float(item.get("man_days", 0.0) or 0.0) for item in items), 2)
    man_hours = round(sum(float(item.get("man_hours", 0.0) or 0.0) for item in items), 2)
    actual_hours = round(sum(float(item.get("actual_hours", 0.0) or 0.0) for item in items), 2)

    starts = [_text_or_default(item.get("planned_start")) for item in items]
    ends = [_text_or_default(item.get("planned_end")) for item in items]
    planned_start, _ = _merge_date_bounds(starts)
    _, planned_end = _merge_date_bounds(ends)

    return {
        "man_days": man_days,
        "man_hours": man_hours,
        "actual_hours": actual_hours,
        "planned_start": planned_start,
        "planned_end": planned_end,
    }


def _write_nested_view(
    epics: dict[str, IssueNode],
    stories: dict[str, IssueNode],
    subtasks: dict[str, IssueNode],
    project_keys_preferred: list[str],
    output_path: Path,
) -> int:
    wb = Workbook()
    ws = wb.active
    ws.title = "NestedView"
    ws.append(HEADERS)

    ws.freeze_panes = "A2"
    ws.sheet_view.showOutlineSymbols = True
    ws.sheet_properties.outlinePr.summaryBelow = False
    ws.sheet_properties.outlinePr.summaryRight = True

    ws.column_dimensions["A"].width = 56
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 20
    ws.column_dimensions["G"].width = 20

    stories_by_epic: dict[str, list[IssueNode]] = {}
    for story in stories.values():
        key = story.epic_key or f"__NO_RMI__::{story.project_key}::{story.product_category}"
        stories_by_epic.setdefault(key, []).append(story)

    subtasks_by_story: dict[str, list[IssueNode]] = {}
    for subtask in subtasks.values():
        subtasks_by_story.setdefault(subtask.story_key, []).append(subtask)

    projects_found: set[str] = set()
    products_by_project: dict[str, set[str]] = {}

    for epic in epics.values():
        projects_found.add(epic.project_key)
        products_by_project.setdefault(epic.project_key, set()).add(epic.product_category)
    for story in stories.values():
        projects_found.add(story.project_key)
        products_by_project.setdefault(story.project_key, set()).add(story.product_category)
    for subtask in subtasks.values():
        projects_found.add(subtask.project_key)
        products_by_project.setdefault(subtask.project_key, set()).add(subtask.product_category)

    project_order = _sorted_keys_with_preferred_order(projects_found, project_keys_preferred)

    story_metrics: dict[str, dict[str, object]] = {}
    for story in stories.values():
        related_subtasks = subtasks_by_story.get(story.key, [])
        fallback_start, _ = _merge_date_bounds([subtask.planned_start for subtask in related_subtasks])
        _, fallback_end = _merge_date_bounds([subtask.planned_end for subtask in related_subtasks])
        planned_start = story.planned_start or fallback_start
        planned_end = story.planned_end or fallback_end
        story_metrics[story.key] = {
            "man_days": story.man_days,
            "man_hours": story.man_hours,
            "actual_hours": story.actual_hours,
            "planned_start": planned_start,
            "planned_end": planned_end,
        }

    row_count = 0

    for project_key in project_order:
        project_name = PROJECT_NAME_BY_KEY.get(project_key, project_key)
        product_values = sorted(products_by_project.get(project_key, {"Uncategorized"}))
        product_blocks: list[dict[str, object]] = []

        for product_category in product_values:
            product_label = _text_or_default(product_category, "Uncategorized")

            epics_in_group: list[IssueNode] = [
                epic
                for epic in epics.values()
                if epic.project_key == project_key and epic.product_category == product_label
            ]

            orphan_epic_placeholder_key = f"__NO_RMI__::{project_key}::{product_label}"
            orphan_stories = stories_by_epic.get(orphan_epic_placeholder_key, [])
            if orphan_stories:
                placeholder = IssueNode(
                    key=orphan_epic_placeholder_key,
                    kind="epic",
                    project_key=project_key,
                    summary="No RMI",
                    parent_key="",
                    assignee="",
                    product_category=product_label,
                    epic_key="",
                    story_key="",
                    man_hours=0.0,
                    man_days=0.0,
                    actual_hours=0.0,
                    planned_start="",
                    planned_end="",
                )
                epics_in_group.append(placeholder)

            epics_in_group = sorted(epics_in_group, key=lambda e: (e.summary.lower(), e.key))
            epic_display_metrics: dict[str, dict[str, object]] = {}
            for epic in epics_in_group:
                epic_story_key = epic.key if epic.key else f"__NO_RMI__::{project_key}::{product_label}"
                related_stories = stories_by_epic.get(epic_story_key, [])
                related_story_metrics = [story_metrics[story.key] for story in related_stories if story.key in story_metrics]

                fallback_start, _ = _merge_date_bounds([str(metric.get("planned_start", "") or "") for metric in related_story_metrics])
                _, fallback_end = _merge_date_bounds([str(metric.get("planned_end", "") or "") for metric in related_story_metrics])

                epic_display_metrics[epic.key] = {
                    "man_days": epic.man_days,
                    "man_hours": epic.man_hours,
                    "actual_hours": epic.actual_hours,
                    "planned_start": epic.planned_start or fallback_start,
                    "planned_end": epic.planned_end or fallback_end,
                }

            product_metric = _aggregate_metrics(list(epic_display_metrics.values()))
            product_blocks.append(
                {
                    "label": product_label,
                    "epics": epics_in_group,
                    "epic_metrics": epic_display_metrics,
                    "metric": product_metric,
                }
            )

        project_metric = _aggregate_metrics([block["metric"] for block in product_blocks])
        _append_row(
            ws,
            outline_level=1,
            aspect=f"{project_key} - {project_name}",
            man_days=project_metric["man_days"],
            man_hours=project_metric["man_hours"],
            actual_hours=project_metric["actual_hours"],
            actual_days=_hours_to_days(project_metric["actual_hours"]),
            planned_start=str(project_metric["planned_start"] or ""),
            planned_end=str(project_metric["planned_end"] or ""),
        )
        row_count += 1

        for block in product_blocks:
            _append_row(
                ws,
                outline_level=2,
                aspect=str(block["label"]),
                man_days=block["metric"]["man_days"],
                man_hours=block["metric"]["man_hours"],
                actual_hours=block["metric"]["actual_hours"],
                actual_days=_hours_to_days(block["metric"]["actual_hours"]),
                planned_start=str(block["metric"]["planned_start"] or ""),
                planned_end=str(block["metric"]["planned_end"] or ""),
            )
            row_count += 1

            for epic in block["epics"]:
                metrics = block["epic_metrics"].get(epic.key, {})
                _append_row(
                    ws,
                    outline_level=3,
                    aspect=epic.summary,
                    man_days=metrics.get("man_days", 0.0),
                    man_hours=metrics.get("man_hours", 0.0),
                    actual_hours=metrics.get("actual_hours", 0.0),
                    actual_days=_hours_to_days(metrics.get("actual_hours", 0.0)),
                    planned_start=str(metrics.get("planned_start", "") or ""),
                    planned_end=str(metrics.get("planned_end", "") or ""),
                )
                row_count += 1

                orphan_epic_placeholder_key = f"__NO_RMI__::{project_key}::{block['label']}"
                epic_story_key = epic.key if epic.key else orphan_epic_placeholder_key
                related_stories = stories_by_epic.get(epic_story_key, [])
                related_stories = sorted(related_stories, key=lambda s: (s.summary.lower(), s.key))

                for story in related_stories:
                    story_metric = story_metrics.get(
                        story.key,
                        {
                            "man_days": story.man_days,
                            "man_hours": story.man_hours,
                            "actual_hours": story.actual_hours,
                            "planned_start": story.planned_start,
                            "planned_end": story.planned_end,
                        },
                    )
                    _append_row(
                        ws,
                        outline_level=4,
                        aspect=story.summary,
                        man_days=story_metric["man_days"],
                        man_hours=story_metric["man_hours"],
                        actual_hours=story_metric["actual_hours"],
                        actual_days=_hours_to_days(story_metric["actual_hours"]),
                        planned_start=str(story_metric["planned_start"] or ""),
                        planned_end=str(story_metric["planned_end"] or ""),
                    )
                    row_count += 1

                    related_subtasks = subtasks_by_story.get(story.key, [])
                    related_subtasks = sorted(related_subtasks, key=lambda t: (t.summary.lower(), t.key))
                    for subtask in related_subtasks:
                        _append_row(
                            ws,
                            outline_level=5,
                            aspect=subtask.summary,
                            man_days=subtask.man_days,
                            man_hours=subtask.man_hours,
                            actual_hours=subtask.actual_hours,
                            actual_days=_hours_to_days(subtask.actual_hours),
                            planned_start=subtask.planned_start,
                            planned_end=subtask.planned_end,
                        )
                        row_count += 1

                        _append_row(
                            ws,
                            outline_level=6,
                            aspect=subtask.assignee,
                            man_days=subtask.man_days,
                            man_hours=subtask.man_hours,
                            actual_hours=subtask.actual_hours,
                            actual_days=_hours_to_days(subtask.actual_hours),
                            planned_start=subtask.planned_start,
                            planned_end=subtask.planned_end,
                        )
                        row_count += 1

    ws.auto_filter.ref = f"A1:G{ws.max_row}"
    wb.save(output_path)
    return row_count


def main() -> None:
    args = _parse_args()
    project_keys, project_source = _get_project_keys()
    keys_str = ", ".join(project_keys)

    session = get_session()
    incremental_disabled = _is_incremental_disabled(args.incremental)
    overlap_minutes = _get_overlap_minutes()
    force_full_days = _get_force_full_days()
    bootstrap_days = _get_bootstrap_days()
    db_path = get_db_path()
    db_path.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(db_path)
    init_db(conn)
    default_checkpoint = bootstrap_default_checkpoint(bootstrap_days)
    checkpoint = get_or_init_checkpoint(conn, "nested_view", default_checkpoint)
    now_utc = utc_now_iso()
    force_full_sync = incremental_disabled or needs_full_sync(conn, "nested_view", now_utc, force_full_days)
    from_updated = None if force_full_sync else apply_overlap(checkpoint, overlap_minutes)
    print(f"Incremental mode: {'OFF' if incremental_disabled else 'ON'}")
    print(f"Sync DB: {db_path}")
    print(f"Projects source: {project_source}")
    print(f"Force full sync: {'Yes' if force_full_sync else 'No'}")
    if from_updated:
        print(f"Discovery updated >= {from_updated} (checkpoint={checkpoint}, overlap={overlap_minutes}m)")

    all_fields = _get_all_fields(session)

    start_date_field_id = resolve_jira_start_date_field_id(session, BASE_URL, project_keys=project_keys)
    if start_date_field_id:
        print(f"Using Jira start date field: {start_date_field_id}")
    else:
        print("Start date field not found; Planned Start Date may be blank.")

    product_field_id = _get_product_categorization_field_id(all_fields)
    if product_field_id:
        print(f"Using Product Categorization field: {product_field_id}")
    else:
        print("Product Categorization field not found; defaulting to 'Uncategorized'.")

    script_dir = Path(__file__).resolve().parent
    rollup_input_name = os.getenv("JIRA_SUBTASK_ROLLUP_XLSX_PATH", DEFAULT_ROLLUP_INPUT).strip() or DEFAULT_ROLLUP_INPUT
    rollup_path = _resolve_output_path(rollup_input_name, script_dir)

    # Primary: load subtask rollup from exports DB; fallback to xlsx
    export_conn = export_db_connect()
    ensure_exports_schema(export_conn)
    if has_subtask_worklog_rollup(export_conn):
        subtask_rollup_values = read_subtask_worklog_rollup_db(export_conn)
        print("Subtask rollup loaded from DB")
    else:
        subtask_rollup_values = _load_subtask_rollup_values(rollup_path)
        print("Subtask rollup loaded from xlsx")
    print(f"Subtask rollup rows loaded: {len(subtask_rollup_values)}")

    detail_fields = [
        "project",
        "summary",
        "issuetype",
        "parent",
        "customfield_10014",
        "assignee",
        "duedate",
        "timeoriginalestimate",
        "timespent",
        "aggregatetimespent",
        "updated",
    ]
    if start_date_field_id:
        detail_fields.append(start_date_field_id)
    if product_field_id:
        detail_fields.append(product_field_id)

    discovery_fields = [
        "project",
        "issuetype",
        "updated",
    ]
    discovery_jql = _build_discovery_jql(project_keys, from_updated_utc=from_updated)
    print(f"Running discovery query for projects: {keys_str}")
    discovered = _fetch_issues(session, jql=discovery_jql, fields=discovery_fields)
    candidates = _candidate_rows_from_issues(discovered)
    changed_issue_keys = get_changed_or_new_issue_keys(conn, candidates)
    upsert_issue_index(conn, candidates)

    active_ids = {row["issue_id"] for row in candidates}
    deleted_count = 0
    if force_full_sync:
        deleted_count = mark_missing_issues_deleted(conn, project_keys, WORK_ITEM_TYPES, active_ids)
        if deleted_count:
            print(f"Marked deleted/inaccessible work items: {deleted_count}")

    discovered_keys = [str(item.get("issue_key", "")).strip() for item in candidates if str(item.get("issue_key", "")).strip()]
    detail_fetch_keys = discovered_keys if force_full_sync else changed_issue_keys
    detail_fetch_keys = sorted({key for key in detail_fetch_keys if key})
    if detail_fetch_keys:
        print(f"Fetching full work-item payloads for {len(detail_fetch_keys)} changed/new issues")
        detailed = _fetch_issues_by_keys(session, detail_fetch_keys, detail_fields)
        payload_rows: list[dict] = []
        index_rows: list[dict] = []
        now_seen = utc_now_iso()
        for issue in detailed:
            issue_id = str(issue.get("id", "")).strip()
            issue_key = str(issue.get("key", "")).strip()
            fields = issue.get("fields", {}) or {}
            updated_utc = str(fields.get("updated", "")).strip()
            issue_type = str((fields.get("issuetype") or {}).get("name", "")).strip()
            project_key = str((fields.get("project") or {}).get("key", "")).strip()
            if not issue_id or not issue_key:
                continue
            payload_rows.append(
                {
                    "issue_id": issue_id,
                    "issue_key": issue_key,
                    "updated_utc": updated_utc,
                    "payload": issue,
                }
            )
            index_rows.append(
                {
                    "issue_id": issue_id,
                    "issue_key": issue_key,
                    "updated_utc": updated_utc,
                    "issue_type": issue_type,
                    "project_key": project_key,
                    "last_seen_utc": now_seen,
                    "is_deleted": 0,
                }
            )
        upsert_issue_payloads(conn, payload_rows)
        upsert_issue_index(conn, index_rows)

    active_issue_keys = _get_active_issue_keys(conn, project_keys, WORK_ITEM_TYPES)
    cached_payloads = get_cached_issue_payloads(conn, project_keys=project_keys, issue_types=WORK_ITEM_TYPES)
    cached_payload_by_key = {str(item.get("key", "")).strip(): item for item in cached_payloads if str(item.get("key", "")).strip()}
    missing_payload_keys = [key for key in active_issue_keys if key not in cached_payload_by_key]
    if missing_payload_keys:
        print(f"Backfilling missing cached payloads for {len(missing_payload_keys)} work items")
        missing_detailed = _fetch_issues_by_keys(session, sorted(missing_payload_keys), detail_fields)
        payload_rows = []
        index_rows = []
        now_seen = utc_now_iso()
        for issue in missing_detailed:
            issue_id = str(issue.get("id", "")).strip()
            issue_key = str(issue.get("key", "")).strip()
            fields = issue.get("fields", {}) or {}
            updated_utc = str(fields.get("updated", "")).strip()
            issue_type = str((fields.get("issuetype") or {}).get("name", "")).strip()
            project_key = str((fields.get("project") or {}).get("key", "")).strip()
            if not issue_id or not issue_key:
                continue
            payload_rows.append(
                {
                    "issue_id": issue_id,
                    "issue_key": issue_key,
                    "updated_utc": updated_utc,
                    "payload": issue,
                }
            )
            index_rows.append(
                {
                    "issue_id": issue_id,
                    "issue_key": issue_key,
                    "updated_utc": updated_utc,
                    "issue_type": issue_type,
                    "project_key": project_key,
                    "last_seen_utc": now_seen,
                    "is_deleted": 0,
                }
            )
        upsert_issue_payloads(conn, payload_rows)
        upsert_issue_index(conn, index_rows)

    issues = get_cached_issue_payloads(conn, project_keys=project_keys, issue_types=WORK_ITEM_TYPES)
    print(f"Cached work items available: {len(issues)}")

    max_updated_seen = ""
    for row in candidates:
        updated_utc = str(row.get("updated_utc", "")).strip()
        if not updated_utc:
            continue
        if not max_updated_seen or parse_iso_utc(updated_utc) > parse_iso_utc(max_updated_seen):
            max_updated_seen = updated_utc
    if max_updated_seen:
        set_checkpoint(conn, "nested_view", max_updated_seen)
    if force_full_sync:
        mark_full_sync(conn, "nested_view", utc_now_iso())

    epics, stories, subtasks = _build_nodes(
        issues=issues,
        start_date_field_id=start_date_field_id,
        product_field_id=product_field_id,
        subtask_rollup_values=subtask_rollup_values,
    )
    print(f"Hierarchy nodes: epics={len(epics)}, stories={len(stories)}, subtasks={len(subtasks)}")

    output_name = os.getenv("JIRA_NESTED_VIEW_XLSX_PATH", DEFAULT_OUTPUT).strip() or DEFAULT_OUTPUT
    output_path = _resolve_output_path(output_name, script_dir)

    # Primary: write nodes to exports DB
    node_count = len(epics) + len(stories) + len(subtasks)
    write_nested_view_nodes_db(export_conn, epics, stories, subtasks)
    record_export_run(export_conn, "nested_view", node_count)
    export_conn.close()

    # Secondary: write xlsx
    data_rows = _write_nested_view(
        epics=epics,
        stories=stories,
        subtasks=subtasks,
        project_keys_preferred=project_keys,
        output_path=output_path,
    )

    generated_at = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    print(f"Export written: {output_path}")
    print(f"Rows written (excluding header): {data_rows}")
    print(f"Generated at: {generated_at}")
    conn.close()


if __name__ == "__main__":
    main()
