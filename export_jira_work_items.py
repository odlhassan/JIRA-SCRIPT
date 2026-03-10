"""
Export Jira work items for selected projects into a flat Excel file.

Each work item is a row. Parent/child nesting is represented with parent columns.
"""
from __future__ import annotations

import argparse
import os
import sqlite3
import time
import uuid
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook, load_workbook
from requests import exceptions as requests_exceptions

from ipp_meeting_utils import (
    fetch_jira_issue_planned_dates,
    load_ipp_actual_and_remarks_by_key,
    load_ipp_issue_keys,
    load_ipp_planned_dates_by_key,
    normalize_issue_key,
    resolve_jira_end_date_field_id,
    resolve_jira_end_date_field_ids,
    resolve_jira_start_date_field_id,
    yes_no_dates_altered,
    yes_no_ipp_actual_matches_jira_end,
    yes_no_in_ipp,
)
from jira_incremental_cache import (
    apply_overlap,
    bootstrap_default_checkpoint,
    get_cached_issue_payloads,
    get_db_path,
    get_or_init_checkpoint,
    init_db,
    mark_full_sync,
    mark_missing_issues_deleted,
    needs_full_sync,
    parse_iso_utc,
    record_pipeline_artifact,
    set_checkpoint,
    upsert_issue_index,
    upsert_issue_payloads,
    utc_now_iso,
)
from jira_client import BASE_URL, extract_jira_key_from_url, get_session
from jira_export_db import connect as export_db_connect
from jira_export_db import ensure_schema as ensure_exports_schema
from jira_export_db import has_subtask_worklogs
from jira_export_db import read_worklogs_actual_dates
from jira_export_db import record_export_run
from jira_export_db import write_work_items as write_work_items_db

DEFAULT_PROJECT_KEYS = ["DIGITALLOG", "FF", "O2", "ODL", "MN"]
DEFAULT_OUTPUT = "1_jira_work_items_export.xlsx"
WORK_ITEM_TYPES = ["Epic", "Story", "Task", "Sub-task", "Subtask", "Bug Task", "Bug Subtask"]


def _resolve_capacity_settings_db_path() -> Path:
    value = (os.getenv("JIRA_ASSIGNEE_HOURS_CAPACITY_DB_PATH", "assignee_hours_capacity.db") or "").strip()
    path = Path(value or "assignee_hours_capacity.db")
    if path.is_absolute():
        return path
    return Path(__file__).resolve().parent / path


def _load_project_keys_from_managed_db() -> list[str]:
    db_path = _resolve_capacity_settings_db_path()
    if not db_path.exists():
        return []
    conn = sqlite3.connect(db_path)
    try:
        rows = conn.execute(
            """
            SELECT project_key
            FROM managed_projects
            WHERE is_active = 1
            ORDER BY project_key ASC
            """
        ).fetchall()
    except sqlite3.Error:
        return []
    finally:
        conn.close()
    return [str(row[0]).strip().upper() for row in rows if str(row[0]).strip()]


def _get_project_keys() -> tuple[list[str], str]:
    db_keys = _load_project_keys_from_managed_db()
    if db_keys:
        return db_keys, "managed_projects_db"
    raw = os.getenv("JIRA_PROJECT_KEYS", "")
    if not raw.strip():
        return DEFAULT_PROJECT_KEYS, "env_fallback"
    return [key.strip() for key in raw.split(",") if key.strip()], "env_fallback"


def _is_incremental_disabled(enable_incremental: bool) -> bool:
    if enable_incremental:
        return False
    return (os.getenv("JIRA_INCREMENTAL_DISABLE", "1").strip() or "1") == "1"


def _get_overlap_minutes() -> int:
    raw = os.getenv("JIRA_INCREMENTAL_OVERLAP_MINUTES", "5").strip() or "5"
    try:
        return max(int(raw), 0)
    except ValueError:
        return 5


def _get_force_full_days() -> int:
    raw = os.getenv("JIRA_FORCE_FULL_SYNC_DAYS", "7").strip() or "7"
    try:
        return max(int(raw), 1)
    except ValueError:
        return 7


def _get_bootstrap_days() -> int:
    raw = os.getenv("JIRA_INCREMENTAL_BOOTSTRAP_DAYS", "365").strip() or "365"
    try:
        return max(int(raw), 1)
    except ValueError:
        return 365


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Export Jira work items to Excel.")
    parser.add_argument(
        "--incremental",
        action="store_true",
        help="Enable smart incremental fetch (default: full fetch).",
    )
    return parser.parse_args()


def _target_assignee() -> str:
    return (os.getenv("JIRA_TARGET_ASSIGNEE", "") or "").strip()


def _jql_escape(value: str) -> str:
    return (value or "").replace("\\", "\\\\").replace('"', '\\"')


def _to_jql_updated_value(value: str) -> str:
    return parse_iso_utc(value).strftime("%Y-%m-%d %H:%M")


def _normalize_type(issue_type_name: str) -> str:
    name = (issue_type_name or "").strip().lower()
    if "epic" in name:
        return "Epic"
    if "story" in name:
        return "Story"
    if "bug" in name and "sub" in name:
        return "Bug Subtask"
    if "bug" in name and "task" in name:
        return "Bug Task"
    if "sub-task" in name or "subtask" in name:
        return "Subtask"
    if "task" in name:
        return "Task"
    return issue_type_name or "Unknown"


def _extract_parent_key(fields: dict) -> str:
    parent = fields.get("parent") or {}
    parent_key = parent.get("key")
    if parent_key:
        return parent_key

    # Epic Link for story/task-like items.
    epic_link = fields.get("customfield_10014")
    if isinstance(epic_link, str) and epic_link.strip():
        return epic_link.strip()
    if isinstance(epic_link, dict):
        epic_key = epic_link.get("key")
        if epic_key:
            return epic_key
    return ""


def _resolve_fix_type_field_id(session) -> str:
    configured = os.getenv("JIRA_FIX_TYPE_FIELD_ID", "").strip()
    if configured:
        return configured
    try:
        response = session.get(f"{BASE_URL}/rest/api/3/field")
        response.raise_for_status()
        all_fields = response.json()
    except Exception:
        return ""

    for field in all_fields:
        name = str(field.get("name", "")).strip().lower()
        field_id = str(field.get("id", "")).strip()
        if not field_id:
            continue
        if name == "fix type":
            return field_id
    for field in all_fields:
        name = str(field.get("name", "")).strip().lower()
        field_id = str(field.get("id", "")).strip()
        if not field_id:
            continue
        if "fix type" in name:
            return field_id
    return ""


def _extract_fix_type_value(value) -> str:
    if value is None:
        return ""
    if isinstance(value, dict):
        for key in ("value", "name"):
            candidate = str(value.get(key, "")).strip()
            if candidate:
                return candidate
        return ""
    if isinstance(value, list):
        parts = [_extract_fix_type_value(item) for item in value]
        return ", ".join([part for part in parts if part])
    return str(value).strip()


def _fetch_issues(session, jql: str, fields: list[str]) -> list[dict]:
    url = f"{BASE_URL}/rest/api/3/search/jql"
    issues: list[dict] = []
    next_page_token = None

    while True:
        payload = {"jql": jql, "maxResults": 100, "fields": fields, "expand": "changelog"}
        if next_page_token:
            payload["nextPageToken"] = next_page_token

        response = None
        last_exception: Exception | None = None
        for attempt in range(6):
            try:
                response = session.post(url, json=payload, timeout=(10, 120))
            except (
                requests_exceptions.ConnectionError,
                requests_exceptions.Timeout,
                requests_exceptions.ChunkedEncodingError,
            ) as exc:
                last_exception = exc
                # Retry transient network failures that can occur under load.
                backoff_seconds = max(0.5 * (2**attempt), 0.5)
                jitter = 0.1 * ((attempt % 3) + 1)
                time.sleep(backoff_seconds + jitter)
                continue
            if response.status_code == 429:
                retry_after = response.headers.get("Retry-After")
                try:
                    retry_after_seconds = float(retry_after) if retry_after else 0.0
                except ValueError:
                    retry_after_seconds = 0.0
                backoff_seconds = max((0.5 * (2**attempt)), retry_after_seconds, 0.5)
                time.sleep(backoff_seconds)
                continue
            response.raise_for_status()
            break
        if response is None:
            if last_exception:
                raise last_exception
            break
        if response.status_code == 429:
            response.raise_for_status()
        data = response.json()

        page_issues = data.get("issues", [])
        issues.extend(page_issues)

        next_page_token = data.get("nextPageToken")
        if not next_page_token:
            break

    return issues


def _stable_resolved_since(issue: dict) -> str:
    changelog = issue.get("changelog") or {}
    histories = changelog.get("histories") if isinstance(changelog, dict) else None
    if not isinstance(histories, list):
        return ""
    stable_resolved = ""
    in_resolved = False
    for history in histories:
        if not isinstance(history, dict):
            continue
        created_text = str(history.get("created", "")).strip()
        created_iso = ""
        if created_text:
            dt = _parse_worklog_timestamp(created_text)
            if dt:
                created_iso = dt.date().isoformat()
        items = history.get("items")
        if not isinstance(items, list):
            continue
        for item in items:
            if not isinstance(item, dict):
                continue
            field_name = str(item.get("field", "")).strip().lower()
            if field_name != "status":
                continue
            to_status = str(item.get("toString", "")).strip().lower()
            from_status = str(item.get("fromString", "")).strip().lower()
            moved_to_resolved = "resolved" in to_status
            moved_out_of_resolved = ("resolved" in from_status) and ("resolved" not in to_status)
            if moved_to_resolved:
                in_resolved = True
                stable_resolved = created_iso or stable_resolved
            if moved_out_of_resolved:
                in_resolved = False
                stable_resolved = ""
    return stable_resolved if in_resolved else ""


def _fetch_issues_by_keys(session, issue_keys: list[str], fields: list[str]) -> list[dict]:
    if not issue_keys:
        return []
    results: list[dict] = []
    for offset in range(0, len(issue_keys), 500):
        chunk = issue_keys[offset : offset + 500]
        keys_clause = ", ".join(f'"{k}"' for k in chunk)
        jql = f"key in ({keys_clause})"
        results.extend(_fetch_issues(session, jql=jql, fields=fields))
    return results


def _payload_has_required_detail_fields(
    issue: dict,
    detail_fields: list[str],
    start_date_field_id: str,
    end_date_field_ids: list[str],
    fix_type_field_id: str,
) -> bool:
    fields = issue.get("fields", {}) or {}
    if not isinstance(fields, dict):
        return False
    required_fields = {
        "project",
        "summary",
        "status",
        "assignee",
        "priority",
        "timetracking",
        "timeoriginalestimate",
        "timespent",
        "aggregatetimespent",
        "issuetype",
        "parent",
        "customfield_10014",
        "created",
        "updated",
    }
    required_fields.update(str(field_id).strip() for field_id in detail_fields if str(field_id).strip())
    if start_date_field_id:
        required_fields.add(start_date_field_id)
    for field_id in end_date_field_ids:
        field_text = str(field_id).strip()
        if field_text:
            required_fields.add(field_text)
    if fix_type_field_id:
        required_fields.add(fix_type_field_id)
    return all(field_name in fields for field_name in required_fields)


def _build_discovery_jql(project_keys: list[str], from_updated_utc: str | None, assignee_name: str = "") -> str:
    keys_str = ", ".join(project_keys)
    base = (
        f'project in ({keys_str}) AND issuetype in ("Epic", "Story", "Task", '
        f'"Sub-task", "Subtask", "Bug Task", "Bug Subtask")'
    )
    if assignee_name:
        base += f' AND assignee = "{_jql_escape(assignee_name)}"'
    if from_updated_utc:
        return f'{base} AND updated >= "{_to_jql_updated_value(from_updated_utc)}" ORDER BY updated ASC'
    return f"{base} ORDER BY updated ASC"


def _candidate_rows_from_issues(issues: list[dict]) -> list[dict]:
    now_utc = utc_now_iso()
    rows: list[dict] = []
    for issue in issues:
        fields = issue.get("fields", {}) or {}
        rows.append(
            {
                "issue_id": str(issue.get("id", "")).strip(),
                "issue_key": str(issue.get("key", "")).strip(),
                "updated_utc": str(fields.get("updated", "")).strip(),
                "issue_type": str((fields.get("issuetype") or {}).get("name", "")).strip(),
                "project_key": str((fields.get("project") or {}).get("key", "")).strip(),
                "last_seen_utc": now_utc,
                "is_deleted": 0,
            }
        )
    return [row for row in rows if row["issue_id"] and row["issue_key"]]


def _classify_candidates(conn: sqlite3.Connection, candidates: list[dict]) -> tuple[list[str], int, int]:
    issue_ids = [str(item.get("issue_id", "")).strip() for item in candidates if str(item.get("issue_id", "")).strip()]
    existing_updated: dict[str, str] = {}
    for offset in range(0, len(issue_ids), 900):
        chunk = issue_ids[offset : offset + 900]
        if not chunk:
            continue
        placeholders = ",".join("?" for _ in chunk)
        rows = conn.execute(
            f"SELECT issue_id, updated_utc FROM issue_index WHERE issue_id IN ({placeholders})",
            tuple(chunk),
        ).fetchall()
        for row in rows:
            existing_updated[str(row[0])] = str(row[1] or "")

    changed_keys: list[str] = []
    seen: set[str] = set()
    new_count = 0
    changed_existing = 0
    for item in candidates:
        issue_id = str(item.get("issue_id", "")).strip()
        issue_key = str(item.get("issue_key", "")).strip()
        updated_utc = str(item.get("updated_utc", "")).strip()
        if not issue_id or not issue_key:
            continue
        if issue_id not in existing_updated:
            new_count += 1
            if issue_key not in seen:
                changed_keys.append(issue_key)
                seen.add(issue_key)
            continue
        if existing_updated.get(issue_id, "") != updated_utc:
            changed_existing += 1
            if issue_key not in seen:
                changed_keys.append(issue_key)
                seen.add(issue_key)
    return changed_keys, new_count, changed_existing


def _get_active_issue_keys(conn: sqlite3.Connection, project_keys: list[str], issue_types: list[str]) -> list[str]:
    if not project_keys or not issue_types:
        return []
    key_ph = ",".join("?" for _ in project_keys)
    type_ph = ",".join("?" for _ in issue_types)
    rows = conn.execute(
        f"""
        SELECT issue_key
        FROM issue_index
        WHERE is_deleted = 0
          AND project_key IN ({key_ph})
          AND issue_type IN ({type_ph})
        ORDER BY issue_key
        """,
        tuple(project_keys + issue_types),
    ).fetchall()
    return [str(row[0]) for row in rows]


def _seconds_to_hours(seconds_value) -> float:
    if seconds_value in (None, ""):
        return 0.0
    try:
        return round(float(seconds_value) / 3600.0, 2)
    except (TypeError, ValueError):
        return 0.0


def _parse_worklog_timestamp(value):
    text = str(value or "").strip()
    if not text:
        return None
    formats = [
        "%Y-%m-%dT%H:%M:%S.%f%z",
        "%Y-%m-%dT%H:%M:%S%z",
        "%Y-%m-%d",
    ]
    for fmt in formats:
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def _update_min_max(bucket: dict, key: str, started_text: str) -> None:
    if not key:
        return
    started_dt = _parse_worklog_timestamp(started_text)
    if not started_dt:
        return
    current = bucket.get(key)
    if not current:
        bucket[key] = {"min_dt": started_dt, "max_dt": started_dt}
        return
    if started_dt < current["min_dt"]:
        current["min_dt"] = started_dt
    if started_dt > current["max_dt"]:
        current["max_dt"] = started_dt


def _format_dt(value) -> str:
    if not value:
        return ""
    return value.isoformat()


def _load_actual_dates_from_worklogs(worklog_path: Path):
    subtask_dates: dict[str, dict] = {}
    story_dates: dict[str, dict] = {}
    epic_dates: dict[str, dict] = {}

    if not worklog_path.exists():
        print(f"Warning: Worklog export not found for actual date rollup: {worklog_path}")
        return subtask_dates, story_dates, epic_dates

    wb = load_workbook(worklog_path, read_only=True, data_only=True)
    ws = wb.active
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        wb.close()
        return subtask_dates, story_dates, epic_dates

    headers = [str(h).strip() if h is not None else "" for h in header_row]
    required = ["issue_id", "parent_story_id", "parent_epic_id", "worklog_started"]
    missing = [name for name in required if name not in headers]
    if missing:
        wb.close()
        print(f"Warning: Worklog export missing required columns for actual date rollup: {missing}")
        return subtask_dates, story_dates, epic_dates

    idx = {name: headers.index(name) for name in required}
    for row in ws.iter_rows(min_row=2, values_only=True):
        subtask_key = normalize_issue_key(row[idx["issue_id"]])
        story_key = normalize_issue_key(row[idx["parent_story_id"]])
        epic_key = normalize_issue_key(row[idx["parent_epic_id"]])
        started_text = str(row[idx["worklog_started"]] or "").strip()
        if not started_text:
            continue
        _update_min_max(subtask_dates, subtask_key, started_text)
        _update_min_max(story_dates, story_key, started_text)
        _update_min_max(epic_dates, epic_key, started_text)

    wb.close()
    return subtask_dates, story_dates, epic_dates


def _fetch_story_to_epic_map(session, story_keys: set[str]) -> dict[str, str]:
    if not story_keys:
        return {}
    story_to_epic: dict[str, str] = {}
    sorted_keys = sorted({normalize_issue_key(k) for k in story_keys if normalize_issue_key(k)})
    for offset in range(0, len(sorted_keys), 500):
        chunk = sorted_keys[offset : offset + 500]
        keys_clause = ", ".join(f'"{key}"' for key in chunk)
        jql = f"key in ({keys_clause})"
        issues = _fetch_issues(session, jql=jql, fields=["parent", "customfield_10014"])
        for issue in issues:
            issue_key = normalize_issue_key(issue.get("key", ""))
            fields = issue.get("fields", {}) or {}
            story_to_epic[issue_key] = normalize_issue_key(_extract_parent_key(fields))
    return story_to_epic


def _build_story_to_epic_from_cached_issues(issues: list[dict], story_keys: set[str]) -> tuple[dict[str, str], set[str]]:
    if not story_keys:
        return {}, set()
    issue_by_key = {normalize_issue_key(issue.get("key", "")): issue for issue in issues}
    story_to_epic: dict[str, str] = {}
    missing: set[str] = set()
    for story_key in story_keys:
        normalized_story_key = normalize_issue_key(story_key)
        if not normalized_story_key:
            continue
        issue = issue_by_key.get(normalized_story_key)
        if not issue:
            missing.add(normalized_story_key)
            continue
        fields = issue.get("fields", {}) or {}
        story_to_epic[normalized_story_key] = normalize_issue_key(_extract_parent_key(fields))
    return story_to_epic, missing


def _build_epic_dates_from_cached_issues(
    issues: list[dict],
    start_date_field_id: str,
    end_date_field_ids: list[str],
    epic_keys: set[str],
) -> tuple[dict[str, dict[str, str]], set[str]]:
    wanted = {normalize_issue_key(key) for key in epic_keys if normalize_issue_key(key)}
    if not wanted:
        return {}, set()
    result: dict[str, dict[str, str]] = {}
    for issue in issues:
        key = normalize_issue_key(issue.get("key", ""))
        if key not in wanted:
            continue
        fields = issue.get("fields", {}) or {}
        issue_type = str((fields.get("issuetype") or {}).get("name", "")).strip().lower()
        if "epic" not in issue_type:
            continue
        planned_end = ""
        for field_id in end_date_field_ids:
            value = fields.get(field_id, "")
            text = str(value or "").strip()
            if text:
                planned_end = text
                break
        result[key] = {
            "planned_start": str(fields.get(start_date_field_id, "") if start_date_field_id else ""),
            "planned_end": planned_end,
        }
    missing = wanted - set(result.keys())
    return result, missing


def _resolve_epic_key_for_issue(issue_key: str, issue_type_name: str, parent_key: str, story_to_epic: dict[str, str]) -> str:
    issue_type = (issue_type_name or "").strip().lower()
    normalized_issue_key = normalize_issue_key(issue_key)
    normalized_parent_key = normalize_issue_key(parent_key)

    if "epic" in issue_type:
        return normalized_issue_key
    if "sub-task" in issue_type or "subtask" in issue_type:
        return story_to_epic.get(normalized_parent_key, "")
    return normalized_parent_key


def _build_rows(
    issues: list[dict],
    start_date_field_id: str,
    end_date_field_ids: list[str],
    fix_type_field_id: str,
    ipp_issue_keys: set[str],
    ipp_planned_dates: dict[str, dict[str, str]],
    ipp_actual_by_key: dict[str, dict[str, str]],
    jira_epic_dates: dict[str, dict[str, str]],
    story_to_epic: dict[str, str],
    subtask_actual_dates: dict[str, dict],
    story_actual_dates: dict[str, dict],
    epic_actual_dates: dict[str, dict],
) -> list[list[str]]:
    rows: list[list[str]] = []
    for issue in issues:
        fields = issue.get("fields", {})
        issue_key = issue.get("key", "")
        issue_url = f"{BASE_URL}/browse/{issue_key}" if issue_key else ""
        work_item_id = extract_jira_key_from_url(issue_url) or issue_key

        issue_type_name = (fields.get("issuetype") or {}).get("name", "")
        issue_type = _normalize_type(issue_type_name)

        parent_key = _extract_parent_key(fields)
        parent_url = f"{BASE_URL}/browse/{parent_key}" if parent_key else ""
        parent_work_item_id = extract_jira_key_from_url(parent_url) or parent_key
        epic_key_for_compare = _resolve_epic_key_for_issue(issue_key, issue_type_name, parent_key, story_to_epic)
        ipp_actual_data = ipp_actual_by_key.get(normalize_issue_key(epic_key_for_compare), {})
        normalized_issue_key = normalize_issue_key(issue_key)
        issue_type_lower = str(issue_type_name or "").strip().lower()
        actual_start_date = ""
        actual_end_date = ""
        if "epic" in issue_type_lower:
            epic_actual = epic_actual_dates.get(normalized_issue_key)
            if epic_actual:
                actual_start_date = _format_dt(epic_actual.get("min_dt"))
                actual_end_date = _format_dt(epic_actual.get("max_dt"))
        elif "story" in issue_type_lower:
            story_actual = story_actual_dates.get(normalized_issue_key)
            if story_actual:
                actual_start_date = _format_dt(story_actual.get("min_dt"))
                actual_end_date = _format_dt(story_actual.get("max_dt"))
        elif "sub-task" in issue_type_lower or "subtask" in issue_type_lower:
            subtask_actual = subtask_actual_dates.get(normalized_issue_key)
            if subtask_actual:
                actual_start_date = _format_dt(subtask_actual.get("min_dt"))
                actual_end_date = _format_dt(subtask_actual.get("max_dt"))

        status = (fields.get("status") or {}).get("name", "")
        assignee = (fields.get("assignee") or {}).get("displayName", "Unassigned")
        project_key = (fields.get("project") or {}).get("key", "")
        summary = fields.get("summary", "")
        start_date = fields.get(start_date_field_id, "") if start_date_field_id else ""
        end_date = ""
        for field_id in end_date_field_ids:
            value = fields.get(field_id, "")
            text = str(value or "").strip()
            if text:
                end_date = text
                break
        timetracking = fields.get("timetracking") or {}
        original_estimate = timetracking.get("originalEstimate", "")
        original_estimate_hours = _seconds_to_hours(fields.get("timeoriginalestimate"))
        total_hours_logged = _seconds_to_hours(fields.get("aggregatetimespent") or fields.get("timespent"))
        created = fields.get("created", "")
        updated = fields.get("updated", "")
        priority = (fields.get("priority") or {}).get("name", "")
        fix_type = _extract_fix_type_value(fields.get(fix_type_field_id)) if fix_type_field_id else ""
        resolved_stable_since_date = _stable_resolved_since(issue)

        rows.append(
            [
                project_key,
                issue_key,
                work_item_id,
                issue_type,
                issue_type_name,
                fix_type,
                summary,
                status,
                resolved_stable_since_date,
                start_date,
                end_date,
                actual_start_date,
                actual_end_date,
                original_estimate,
                original_estimate_hours,
                assignee,
                total_hours_logged,
                priority,
                parent_key,
                parent_work_item_id,
                parent_url,
                issue_url,
                yes_no_in_ipp(issue_key, ipp_issue_keys),
                yes_no_dates_altered(epic_key_for_compare, ipp_planned_dates, jira_epic_dates),
                ipp_actual_data.get("ipp_actual_date", ""),
                ipp_actual_data.get("ipp_remarks", ""),
                yes_no_ipp_actual_matches_jira_end(epic_key_for_compare, ipp_actual_by_key, jira_epic_dates),
                created,
                updated,
            ]
        )
    return rows


def _write_excel(rows: list[list[str]], output_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "WorkItems"

    headers = [
        "project_key",
        "issue_key",
        "work_item_id",
        "work_item_type",
        "jira_issue_type",
        "fix_type",
        "summary",
        "status",
        "resolved_stable_since_date",
        "start_date",
        "end_date",
        "actual_start_date",
        "actual_end_date",
        "original_estimate",
        "original_estimate_hours",
        "assignee",
        "total_hours_logged",
        "priority",
        "parent_issue_key",
        "parent_work_item_id",
        "parent_jira_url",
        "jira_url",
        "Latest IPP Meeting",
        "Jira IPP RMI Dates Altered",
        "IPP Actual Date (Production Date)",
        "IPP Remarks",
        "IPP Actual Date Matches Jira End Date",
        "created",
        "updated",
    ]
    sheet.append(headers)

    for row in rows:
        sheet.append(row)

    workbook.save(output_path)


def main() -> None:
    args = _parse_args()
    run_started = datetime.now(timezone.utc)
    run_id = uuid.uuid4().hex
    project_keys, project_source = _get_project_keys()
    target_assignee = _target_assignee()
    session = get_session()
    incremental_disabled = _is_incremental_disabled(args.incremental)
    overlap_minutes = _get_overlap_minutes()
    force_full_days = _get_force_full_days()
    bootstrap_days = _get_bootstrap_days()
    db_path = get_db_path()
    db_path.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(db_path)
    init_db(conn)

    default_checkpoint = bootstrap_default_checkpoint(bootstrap_days)
    checkpoint = get_or_init_checkpoint(conn, "work_items", default_checkpoint)
    now_utc = utc_now_iso()
    force_full_sync = incremental_disabled or needs_full_sync(conn, "work_items", now_utc, force_full_days)
    from_updated = None if force_full_sync else apply_overlap(checkpoint, overlap_minutes)

    print(f"Incremental mode: {'OFF' if incremental_disabled else 'ON'}")
    print(f"Sync DB: {db_path}")
    print(f"Projects source: {project_source}")
    print(f"Force full sync: {'Yes' if force_full_sync else 'No'}")
    if from_updated:
        print(f"Discovery updated >= {from_updated} (checkpoint={checkpoint}, overlap={overlap_minutes}m)")

    start_date_field_id = resolve_jira_start_date_field_id(session, BASE_URL, project_keys=project_keys)
    if start_date_field_id:
        print(f"Using Jira start date field: {start_date_field_id}")
    else:
        print("Start date field not found; start_date will be blank where unavailable.")

    end_date_field_id = resolve_jira_end_date_field_id(session, BASE_URL, project_keys=project_keys)
    end_date_field_ids = resolve_jira_end_date_field_ids(session, BASE_URL, project_keys=project_keys)
    if end_date_field_ids:
        print(f"Using Jira end date fields (priority order): {', '.join(end_date_field_ids)}")
    elif end_date_field_id:
        end_date_field_ids = [end_date_field_id]
        print(f"Using Jira end date field: {end_date_field_id}")
    else:
        end_date_field_ids = ["duedate"]
        print("End date field not found; end_date will be blank where unavailable.")

    discovery_fields = [
        "project",
        "issuetype",
        "updated",
    ]
    discovery_jql = _build_discovery_jql(project_keys, from_updated_utc=from_updated, assignee_name=target_assignee)
    print(f"Running discovery query for projects: {', '.join(project_keys)}")
    discovered = _fetch_issues(session, jql=discovery_jql, fields=discovery_fields)
    candidates = _candidate_rows_from_issues(discovered)
    changed_issue_keys, new_issue_count, changed_existing_count = _classify_candidates(conn, candidates)
    upsert_issue_index(conn, candidates)

    active_ids = {row["issue_id"] for row in candidates}
    deleted_count = 0
    if force_full_sync and not target_assignee:
        deleted_count = mark_missing_issues_deleted(conn, project_keys, WORK_ITEM_TYPES, active_ids)
        if deleted_count:
            print(f"Marked deleted/inaccessible work items: {deleted_count}")

    discovered_keys = [str(item.get("issue_key", "")).strip() for item in candidates if str(item.get("issue_key", "")).strip()]
    detail_fetch_keys = discovered_keys if force_full_sync else changed_issue_keys
    detail_fetch_keys = sorted({key for key in detail_fetch_keys if key})

    detail_fields = [
        "project",
        "summary",
        "status",
        "duedate",
        "assignee",
        "priority",
        "timetracking",
        "timeoriginalestimate",
        "timespent",
        "aggregatetimespent",
        "issuetype",
        "parent",
        "customfield_10014",
        "created",
        "updated",
    ]
    if start_date_field_id:
        detail_fields.append(start_date_field_id)
    for field_id in end_date_field_ids:
        if field_id not in detail_fields:
            detail_fields.append(field_id)
    fix_type_field_id = _resolve_fix_type_field_id(session)
    if fix_type_field_id:
        detail_fields.append(fix_type_field_id)
        print(f"Using Jira Fix Type field: {fix_type_field_id}")
    else:
        print("Fix Type field not found; fix_type will be blank.")
    if detail_fetch_keys:
        print(f"Fetching full work-item payloads for {len(detail_fetch_keys)} changed/new issues")
        detailed = _fetch_issues_by_keys(session, detail_fetch_keys, detail_fields)
        payload_rows: list[dict] = []
        index_rows: list[dict] = []
        now_seen = utc_now_iso()
        for issue in detailed:
            issue_id = str(issue.get("id", "")).strip()
            issue_key = str(issue.get("key", "")).strip()
            fields = issue.get("fields", {}) or {}
            updated_utc = str(fields.get("updated", "")).strip()
            issue_type = str((fields.get("issuetype") or {}).get("name", "")).strip()
            project_key = str((fields.get("project") or {}).get("key", "")).strip()
            if not issue_id or not issue_key:
                continue
            payload_rows.append(
                {
                    "issue_id": issue_id,
                    "issue_key": issue_key,
                    "updated_utc": updated_utc,
                    "payload": issue,
                }
            )
            index_rows.append(
                {
                    "issue_id": issue_id,
                    "issue_key": issue_key,
                    "updated_utc": updated_utc,
                    "issue_type": issue_type,
                    "project_key": project_key,
                    "last_seen_utc": now_seen,
                    "is_deleted": 0,
                }
            )
        upsert_issue_payloads(conn, payload_rows)
        upsert_issue_index(conn, index_rows)

    active_issue_keys = discovered_keys if target_assignee else _get_active_issue_keys(conn, project_keys, WORK_ITEM_TYPES)
    cached_payloads = get_cached_issue_payloads(conn, project_keys=project_keys, issue_types=WORK_ITEM_TYPES)
    if target_assignee:
        target_assignee_lower = target_assignee.casefold()
        cached_payloads = [
            item
            for item in cached_payloads
            if str((((item.get("fields") or {}).get("assignee") or {}).get("displayName") or "")).strip().casefold() == target_assignee_lower
        ]
    cached_payload_by_key = {str(item.get("key", "")).strip(): item for item in cached_payloads if str(item.get("key", "")).strip()}
    missing_payload_keys = [key for key in active_issue_keys if key not in cached_payload_by_key]
    incomplete_payload_keys = [
        key
        for key, payload in cached_payload_by_key.items()
        if not _payload_has_required_detail_fields(
            payload,
            detail_fields=detail_fields,
            start_date_field_id=start_date_field_id,
            end_date_field_ids=end_date_field_ids,
            fix_type_field_id=fix_type_field_id,
        )
    ]
    refill_payload_keys = sorted({*missing_payload_keys, *incomplete_payload_keys})
    if refill_payload_keys:
        if incomplete_payload_keys:
            print(f"Refetching {len(incomplete_payload_keys)} incomplete cached payloads with full detail fields")
        if missing_payload_keys:
            print(f"Backfilling missing cached payloads for {len(missing_payload_keys)} work items")
        missing_detailed = _fetch_issues_by_keys(session, refill_payload_keys, detail_fields)
        payload_rows: list[dict] = []
        index_rows: list[dict] = []
        now_seen = utc_now_iso()
        for issue in missing_detailed:
            issue_id = str(issue.get("id", "")).strip()
            issue_key = str(issue.get("key", "")).strip()
            fields = issue.get("fields", {}) or {}
            updated_utc = str(fields.get("updated", "")).strip()
            issue_type = str((fields.get("issuetype") or {}).get("name", "")).strip()
            project_key = str((fields.get("project") or {}).get("key", "")).strip()
            if not issue_id or not issue_key:
                continue
            payload_rows.append(
                {
                    "issue_id": issue_id,
                    "issue_key": issue_key,
                    "updated_utc": updated_utc,
                    "payload": issue,
                }
            )
            index_rows.append(
                {
                    "issue_id": issue_id,
                    "issue_key": issue_key,
                    "updated_utc": updated_utc,
                    "issue_type": issue_type,
                    "project_key": project_key,
                    "last_seen_utc": now_seen,
                    "is_deleted": 0,
                }
            )
        upsert_issue_payloads(conn, payload_rows)
        upsert_issue_index(conn, index_rows)

    issues = get_cached_issue_payloads(conn, project_keys=project_keys, issue_types=WORK_ITEM_TYPES)
    if target_assignee:
        target_assignee_lower = target_assignee.casefold()
        issues = [
            item
            for item in issues
            if str((((item.get("fields") or {}).get("assignee") or {}).get("displayName") or "")).strip().casefold() == target_assignee_lower
        ]
    print(f"Cached work items available: {len(issues)}")

    ipp_issue_keys = load_ipp_issue_keys()
    print(f"IPP issue keys loaded: {len(ipp_issue_keys)}")
    ipp_planned_dates = load_ipp_planned_dates_by_key()
    print(f"IPP planned-date keys loaded: {len(ipp_planned_dates)}")
    ipp_actual_by_key = load_ipp_actual_and_remarks_by_key()
    print(f"IPP actual-date/remarks keys loaded: {len(ipp_actual_by_key)}")

    subtask_parent_story_keys: set[str] = set()
    epic_keys_for_compare: set[str] = set()
    for issue in issues:
        issue_key = normalize_issue_key(issue.get("key", ""))
        fields = issue.get("fields", {}) or {}
        issue_type_name = ((fields.get("issuetype") or {}).get("name") or "").strip().lower()
        parent_key = normalize_issue_key(_extract_parent_key(fields))

        if "epic" in issue_type_name:
            if issue_key:
                epic_keys_for_compare.add(issue_key)
            continue
        if "sub-task" in issue_type_name or "subtask" in issue_type_name:
            if parent_key:
                subtask_parent_story_keys.add(parent_key)
            continue
        if parent_key:
            epic_keys_for_compare.add(parent_key)

    story_to_epic, missing_story_keys = _build_story_to_epic_from_cached_issues(issues, subtask_parent_story_keys)
    if missing_story_keys:
        story_to_epic.update(_fetch_story_to_epic_map(session, missing_story_keys))
    for epic_key in story_to_epic.values():
        if epic_key:
            epic_keys_for_compare.add(epic_key)

    jira_epic_dates, missing_epic_dates = _build_epic_dates_from_cached_issues(
        issues,
        start_date_field_id=start_date_field_id,
        end_date_field_ids=end_date_field_ids,
        epic_keys=epic_keys_for_compare,
    )
    if missing_epic_dates:
        jira_epic_dates.update(
            fetch_jira_issue_planned_dates(
                session,
                BASE_URL,
                missing_epic_dates,
                start_date_field_id=start_date_field_id,
                end_date_field_id=end_date_field_id,
                end_date_field_ids=end_date_field_ids,
                project_keys=project_keys,
            )
        )
    print(f"Jira epic planned-date keys loaded: {len(jira_epic_dates)}")

    # Primary: load actual dates from exports DB; fallback to worklogs xlsx
    export_conn = export_db_connect()
    ensure_exports_schema(export_conn)
    worklog_path_value = os.getenv("JIRA_WORKLOG_XLSX_PATH", "2_jira_subtask_worklogs.xlsx").strip() or "2_jira_subtask_worklogs.xlsx"
    worklog_path = Path(worklog_path_value)
    if not worklog_path.is_absolute():
        worklog_path = Path(__file__).resolve().parent / worklog_path
    if has_subtask_worklogs(export_conn):
        subtask_actual_dates, story_actual_dates, epic_actual_dates = read_worklogs_actual_dates(export_conn)
        print("Actual date rollups loaded from DB")
    else:
        subtask_actual_dates, story_actual_dates, epic_actual_dates = _load_actual_dates_from_worklogs(worklog_path)
        print("Actual date rollups loaded from worklogs xlsx")
    print(
        "Actual date rollups: "
        f"subtasks={len(subtask_actual_dates)}, stories={len(story_actual_dates)}, epics={len(epic_actual_dates)}"
    )

    rows = _build_rows(
        issues,
        start_date_field_id=start_date_field_id,
        end_date_field_ids=end_date_field_ids,
        fix_type_field_id=fix_type_field_id,
        ipp_issue_keys=ipp_issue_keys,
        ipp_planned_dates=ipp_planned_dates,
        ipp_actual_by_key=ipp_actual_by_key,
        jira_epic_dates=jira_epic_dates,
        story_to_epic=story_to_epic,
        subtask_actual_dates=subtask_actual_dates,
        story_actual_dates=story_actual_dates,
        epic_actual_dates=epic_actual_dates,
    )

    output_name = os.getenv("JIRA_EXPORT_XLSX_PATH", DEFAULT_OUTPUT).strip() or DEFAULT_OUTPUT
    output_path = Path(output_name)
    if not output_path.is_absolute():
        output_path = Path(__file__).resolve().parent / output_path

    # Primary: write to exports DB
    write_work_items_db(export_conn, rows)
    run_ended = datetime.now(timezone.utc)
    run_ended_iso = run_ended.replace(microsecond=0).isoformat().replace("+00:00", "Z")
    run_started_iso = run_started.replace(microsecond=0).isoformat().replace("+00:00", "Z")
    record_export_run(
        export_conn,
        "work_items",
        len(rows),
        status="success",
        run_id=run_id,
        started_at_utc=run_started_iso,
        finished_at_utc=run_ended_iso,
    )
    export_conn.close()

    # Secondary: write xlsx
    _write_excel(rows, output_path)
    generated_at = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    print(f"Export written: {output_path}")
    print(f"Generated at: {generated_at}")

    max_updated_seen = ""
    for row in candidates:
        updated_utc = str(row.get("updated_utc", "")).strip()
        if not updated_utc:
            continue
        if not max_updated_seen or parse_iso_utc(updated_utc) > parse_iso_utc(max_updated_seen):
            max_updated_seen = updated_utc
    if max_updated_seen:
        set_checkpoint(conn, "work_items", max_updated_seen)
    if force_full_sync:
        mark_full_sync(conn, "work_items", utc_now_iso())

    run_ended = datetime.now(timezone.utc)
    record_pipeline_artifact(
        conn,
        {
            "run_id": run_id,
            "pipeline": "work_items",
            "started_at_utc": run_started.replace(microsecond=0).isoformat().replace("+00:00", "Z"),
            "ended_at_utc": run_ended.replace(microsecond=0).isoformat().replace("+00:00", "Z"),
            "issues_scanned": len(candidates),
            "issues_changed": changed_existing_count,
            "new_issues": new_issue_count,
            "detail_fetches": len(detail_fetch_keys),
            "worklog_fetches": 0,
            "duration_ms": int((run_ended - run_started).total_seconds() * 1000),
        },
    )
    conn.close()


if __name__ == "__main__":
    main()
