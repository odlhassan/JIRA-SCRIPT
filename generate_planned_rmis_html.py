
from __future__ import annotations

import json
import os
from collections import defaultdict, deque
from datetime import date, datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

DEFAULT_INPUT_XLSX = "nested view.xlsx"
DEFAULT_OUTPUT_HTML = "planned_rmis_report.html"
DEFAULT_WORK_ITEMS_XLSX = "1_jira_work_items_export.xlsx"
REQUIRED_HEADERS = ["Aspect", "Man-hours", "Planned Start Date", "Planned End Date"]
OPTIONAL_ACTUAL_HEADER = "Actual Hours"


def _resolve_path(value: str, base_dir: Path) -> Path:
    p = Path(value)
    return p if p.is_absolute() else base_dir / p


def _to_text(value: object) -> str:
    return "" if value is None else str(value).strip()


def _to_number_or_blank(value: object):
    if value in (None, ""):
        return ""
    try:
        return round(float(value), 2)
    except (TypeError, ValueError):
        return ""


def _parse_to_iso_date(value: object) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = _to_text(value)
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%d-%b-%Y", "%d-%B-%Y", "%m/%d/%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00")).date().isoformat()
    except ValueError:
        return ""


def _project_key_from_aspect(aspect: str) -> str:
    txt = _to_text(aspect)
    return txt.split(" - ", 1)[0].strip() if " - " in txt else txt


def _to_key_text(value: object) -> str:
    return _to_text(value).strip().lower()


def _load_epic_key_index(work_items_path: Path) -> dict[tuple[str, str], deque[str]]:
    out: dict[tuple[str, str], deque[str]] = defaultdict(deque)
    if not work_items_path.exists():
        return out
    wb = load_workbook(work_items_path, read_only=True, data_only=True)
    try:
        ws = wb.active
        header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header:
            return out
        headers = [_to_text(v) for v in header]
        idx = {name: i for i, name in enumerate(headers)}
        required = ["project_key", "issue_key", "jira_issue_type", "summary"]
        if any(name not in idx for name in required):
            return out
        for row in ws.iter_rows(min_row=2, values_only=True):
            project_key = _to_text(row[idx["project_key"]]).upper()
            issue_key = _to_text(row[idx["issue_key"]]).upper()
            issue_type = _to_text(row[idx["jira_issue_type"]]).lower()
            summary = _to_key_text(row[idx["summary"]])
            if "epic" not in issue_type or not project_key or not issue_key or not summary:
                continue
            out[(project_key, summary)].append(issue_key)
        return out
    finally:
        wb.close()


def _load_payload(input_path: Path, work_items_path: Path) -> dict:
    wb = load_workbook(input_path, read_only=False, data_only=True)
    ws = wb["NestedView"] if "NestedView" in wb.sheetnames else wb.active

    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None) or ()
    idx = {_to_text(v).lower(): i for i, v in enumerate(header)}
    miss = [h for h in REQUIRED_HEADERS if h.lower() not in idx]
    if miss:
        wb.close()
        raise ValueError(f"Missing headers: {miss}")

    ia, ih, isd, ied = idx["aspect"], idx["man-hours"], idx["planned start date"], idx["planned end date"]
    iact = idx.get(OPTIONAL_ACTUAL_HEADER.lower(), -1)
    projects: list[dict] = []
    epic_key_index = _load_epic_key_index(work_items_path)
    stack: dict[int, int] = {}

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        level = int(getattr(ws.row_dimensions[row_idx], "outlineLevel", 0) or 1)
        for k in list(stack):
            if k >= level:
                del stack[k]
        aspect = _to_text(row[ia] if len(row) > ia else "")
        if not aspect:
            continue
        if level == 1:
            projects.append({"name": aspect, "project_key": _project_key_from_aspect(aspect), "epics": []})
            stack[1] = len(projects) - 1
            continue
        if level != 3 or 1 not in stack:
            continue

        hours = _to_number_or_blank(row[ih] if len(row) > ih else "")
        actual_hours = _to_number_or_blank(row[iact] if iact >= 0 and len(row) > iact else "")
        pstart = _parse_to_iso_date(row[isd] if len(row) > isd else "")
        pend = _parse_to_iso_date(row[ied] if len(row) > ied else "")
        has_range = False
        so, eo = "", ""
        if pstart and pend:
            try:
                sd, ed = datetime.strptime(pstart, "%Y-%m-%d").date(), datetime.strptime(pend, "%Y-%m-%d").date()
                if ed >= sd:
                    has_range = True
                    so, eo = sd.toordinal(), ed.toordinal()
            except ValueError:
                pass
        project_key = projects[stack[1]]["project_key"].upper()
        jira_key = ""
        summary_key = _to_key_text(aspect)
        queue = epic_key_index.get((project_key, summary_key))
        if queue:
            jira_key = queue.popleft()
        projects[stack[1]]["epics"].append({
            "name": aspect,
            "jira_key": jira_key,
            "planned_hours": hours,
            "actual_hours": actual_hours,
            "original_actual_hours": actual_hours,
            "planned_start": pstart,
            "planned_end": pend,
            "has_range": has_range,
            "start_ordinal": so,
            "end_ordinal": eo,
        })

    wb.close()

    rows: list[dict] = []
    rid = 1
    for p in projects:
        planned = [e for e in p["epics"] if e["has_range"]]
        unplanned = [e for e in p["epics"] if not e["has_range"]]
        pmin = min((int(e["start_ordinal"]) for e in planned), default=None)
        pmax = max((int(e["end_ordinal"]) for e in planned), default=None)
        total_hours = round(sum(float(e["planned_hours"]) for e in p["epics"] if isinstance(e["planned_hours"], (int, float))), 2)
        total_actual_hours = round(sum(float(e["actual_hours"]) for e in p["epics"] if isinstance(e["actual_hours"], (int, float))), 2)

        pid = rid
        rid += 1
        rows.append({
            "id": pid,
            "parent_id": None,
            "level": 1,
            "row_kind": "project",
            "group_type": "",
            "rmi_label": p["name"],
            "planned_hours": total_hours,
            "actual_hours": total_actual_hours,
            "planned_start": date.fromordinal(pmin).isoformat() if pmin else "",
            "planned_end": date.fromordinal(pmax).isoformat() if pmax else "",
            "has_range": bool(pmin and pmax),
            "start_ordinal": pmin or "",
            "end_ordinal": pmax or "",
            "project_key": p["project_key"],
        })

        gid = rid
        rid += 1
        rows.append({"id": gid, "parent_id": pid, "level": 2, "row_kind": "group", "group_type": "planned", "rmi_label": "Planned Epics", "planned_hours": round(sum(float(e["planned_hours"]) for e in planned if isinstance(e["planned_hours"], (int, float))), 2), "actual_hours": round(sum(float(e["actual_hours"]) for e in planned if isinstance(e["actual_hours"], (int, float))), 2), "planned_start": "", "planned_end": "", "has_range": False, "start_ordinal": "", "end_ordinal": "", "project_key": p["project_key"]})
        for e in planned:
            rows.append({"id": rid, "parent_id": gid, "level": 3, "row_kind": "epic", "group_type": "planned", "rmi_label": e["name"], "jira_key": e["jira_key"], "planned_hours": e["planned_hours"], "actual_hours": e["actual_hours"], "original_actual_hours": e["original_actual_hours"], "planned_start": e["planned_start"], "planned_end": e["planned_end"], "has_range": True, "start_ordinal": e["start_ordinal"], "end_ordinal": e["end_ordinal"], "project_key": p["project_key"]})
            rid += 1

        ugid = rid
        rid += 1
        rows.append({"id": ugid, "parent_id": pid, "level": 2, "row_kind": "group", "group_type": "not_planned", "rmi_label": "Not Planned Yet Epics", "planned_hours": round(sum(float(e["planned_hours"]) for e in unplanned if isinstance(e["planned_hours"], (int, float))), 2), "actual_hours": round(sum(float(e["actual_hours"]) for e in unplanned if isinstance(e["actual_hours"], (int, float))), 2), "planned_start": "", "planned_end": "", "has_range": False, "start_ordinal": "", "end_ordinal": "", "project_key": p["project_key"]})
        for e in unplanned:
            rows.append({"id": rid, "parent_id": ugid, "level": 3, "row_kind": "epic", "group_type": "not_planned", "rmi_label": e["name"], "jira_key": e["jira_key"], "planned_hours": e["planned_hours"], "actual_hours": e["actual_hours"], "original_actual_hours": e["original_actual_hours"], "planned_start": "", "planned_end": "", "has_range": False, "start_ordinal": "", "end_ordinal": "", "project_key": p["project_key"]})
            rid += 1

    today = datetime.now(timezone.utc).date()
    return {
        "rows": rows,
        "source_file": str(input_path),
        "default_from": date(today.year, 1, 1).isoformat(),
        "default_to": today.isoformat(),
    }


def _build_html(payload: dict) -> str:
    data = json.dumps(payload, ensure_ascii=True)
    return """<!doctype html>
<html lang='en'>
<head>
<meta charset='utf-8'>
<meta name='viewport' content='width=device-width,initial-scale=1'>
<title>Planned RMIs</title>
<link rel='stylesheet' href='shared-nav.css'>
<link rel='stylesheet' href='https://fonts.googleapis.com/css2?family=Material+Symbols+Outlined:opsz,wght,FILL,GRAD@20..48,500,0,0'>
<style>
:root{--bg:#f3f6f9;--panel:#ffffff;--text:#1f2937;--muted:#6b7280;--line:#dbe3ea;--head:#0f4c5c;--head-text:#ffffff;--left1:420px;--left2:150px}
*{box-sizing:border-box}body{margin:0;padding:18px;font-family:"Segoe UI",Tahoma,Verdana,sans-serif;color:var(--text);background:radial-gradient(1000px 280px at 10% -5%, #d7eef6 0%, transparent 60%),linear-gradient(180deg, #eef4f7, var(--bg))}
.page{max-width:1500px;margin:0 auto;display:grid;gap:12px}.panel{background:var(--panel);border:1px solid var(--line);border-radius:12px;padding:14px 16px}
.title{margin:0;font-size:1.25rem;font-weight:700;color:#0b3142}.meta{margin:0;color:var(--muted);font-size:.9rem}
.toolbar{margin-top:12px;display:flex;gap:8px;flex-wrap:wrap;align-items:end}.control{display:grid;gap:4px}.control label{font-size:.74rem;color:#355564;text-transform:uppercase;font-weight:700}.control input,.control select{border:1px solid #b6c7d2;border-radius:8px;padding:7px 9px;font-size:.86rem;min-width:150px;background:#fff;color:#12313f}
.btn{display:inline-flex;align-items:center;gap:6px;border:1px solid #255f73;background:#0f4c5c;color:#fff;border-radius:8px;padding:7px 12px;font-size:.9rem;cursor:pointer}.btn.alt{background:#fff;color:#255f73}
.scorecards{margin-top:10px;display:grid;grid-template-columns:repeat(auto-fit,minmax(200px,1fr));gap:8px}.score{border:1px solid #d7e3ea;border-radius:10px;background:#f8fbfd;padding:8px 10px;min-height:66px}.sl{color:#355564;font-size:.75rem;font-weight:700;text-transform:uppercase}.sv{color:#0b3142;margin-top:3px;font-size:1.15rem;font-weight:800}
#score-projects-card{background:#ecfeff;border-color:#99f6e4}.scorecards #score-projects-card .sv{color:#0f766e}
#score-planned-epics-card{background:#ecfdf3;border-color:#86efac}.scorecards #score-planned-epics-card .sv{color:#166534}
#score-not-planned-epics-card{background:#fff1f2;border-color:#fda4af}.scorecards #score-not-planned-epics-card .sv{color:#be123c}
#score-planned-hours-card{background:#eff6ff;border-color:#93c5fd}.scorecards #score-planned-hours-card .sv{color:#1d4ed8}
#score-not-planned-hours-card{background:#fef2f2;border-color:#fecaca}.scorecards #score-not-planned-hours-card .sv{color:#b91c1c}
#score-capacity-card{background:#fffbeb;border-color:#fcd34d}.scorecards #score-capacity-card .sv{color:#a16207}
#score-actual-hours-card{background:#f3e8ff;border-color:#d8b4fe}.scorecards #score-actual-hours-card .sv{color:#7c3aed}
.gw{background:var(--panel);border:1px solid var(--line);border-radius:12px;max-height:calc(100vh - 220px);overflow:auto;width:100%}.g{width:max-content;min-width:calc(var(--left1)+var(--left2)+900px)}
.r{display:grid;grid-template-columns:var(--left1) var(--left2) 1fr;min-height:40px;border-bottom:1px solid var(--line)}.r.h{position:sticky;top:0;z-index:30}.c{padding:8px 10px;font-size:.82rem}
.s1{position:sticky;left:0;z-index:20;background:#fff;border-right:1px solid var(--line)}.s2{position:sticky;left:var(--left1);z-index:20;background:#fff;border-right:1px solid var(--line);text-align:right;font-variant-numeric:tabular-nums}
.r.h .c{font-size:.9rem;background:var(--head);color:var(--head-text);text-align:left;padding:10px 10px;border-bottom:1px solid #0a3946}.r.h .s1,.r.h .s2{z-index:40;background:var(--head)}
.row-project .s1,.row-project .s2{background:#dbeeff}.row-group .s1,.row-group .s2{background:#f8fbfd}.group-planned .s1{border-left:3px solid #2da365}.group-not-planned .s1{border-left:3px solid #d4677f}
.rmi{display:flex;align-items:center;gap:6px;min-width:0}.lbl{min-width:0;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
.focus-link{border:none;background:transparent;padding:0;min-width:0;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;text-align:left;color:#0b4a6f;font-weight:700;cursor:pointer}
.focus-link:hover{text-decoration:underline}.focus-link:focus{outline:2px solid rgba(15,76,92,.25);outline-offset:2px;border-radius:4px}
.toggle{width:22px;height:22px;border:1px solid #adc0d8;border-radius:6px;background:#fff;color:#334155;cursor:pointer;line-height:1}.toggle.placeholder{visibility:hidden}
.ch{border-radius:999px;font-size:.66rem;font-weight:700;padding:2px 7px;border:1px solid transparent}.ch.p{background:#e8f7ed;border-color:#8fd6a1;color:#1a6e3e}.ch.u{background:#ffeef0;border-color:#f5b3bc;color:#8d3146}
.th{position:relative;min-height:54px;background:#f8fbfd;border-left:1px solid var(--line)}.months,.weeks{position:absolute;left:0;right:0;overflow:hidden;color:#334a67}.months{top:0;height:27px;border-bottom:1px solid var(--line);font-size:.68rem}.weeks{top:27px;height:27px;font-size:.62rem}
.mb,.wb{position:absolute;top:0;bottom:0;display:flex;align-items:center;justify-content:center;border-right:1px solid #cfdaea;white-space:nowrap;padding:0 4px}.wb{background:rgba(255,255,255,.5);color:#607793}
.tc{border-left:1px solid var(--line);position:relative}.slot{position:relative;min-height:40px}.wl{position:absolute;top:0;bottom:0;width:1px;background:#e4ebf5}
.bar{position:absolute;top:10px;height:20px;border-radius:999px;border:1px solid rgba(19,71,180,.35);background:linear-gradient(180deg,#3f7bff,#2157d6);box-shadow:0 2px 6px rgba(33,87,214,.24)}.bar.project{background:linear-gradient(180deg,#3bc8de,#0b8ca8);border-color:rgba(11,140,168,.45);top:13px;height:14px;opacity:.65}.bar.no{background:#c6d3e3;border-color:#b4c3d8;box-shadow:none}
.empty{padding:14px;color:var(--muted);font-size:.88rem}
</style>
</head>
<body>
<div class='page'><section class='panel'>
<h1 class='title'>Planned RMIs</h1>
<p class='meta'>Generated: <span id='gen'></span> | Source: <span id='src'></span> | Visible Epics: <span id='ve'></span></p>
<div class='toolbar'>
<div class='control'><label for='actual-hours-mode'>Actual Hours Mode</label><select id='actual-hours-mode'><option value='log_date'>By Log Date</option><option value='planned_dates'>By Planned Dates</option></select></div>
<div class='control'><label for='from'>From</label><input id='from' type='date'></div>
<div class='control'><label for='to'>To</label><input id='to' type='date'></div>
<button id='apply' class='btn alt' type='button'><span class='material-symbols-outlined'>filter_alt</span>Apply</button>
<button id='reset' class='btn' type='button'><span class='material-symbols-outlined'>restart_alt</span>Reset</button>
<button id='collapse' class='btn alt' type='button'><span class='material-symbols-outlined'>unfold_less</span>Collapse Projects</button>
<button id='expand' class='btn alt' type='button'><span class='material-symbols-outlined'>unfold_more</span>Expand All</button>
<span id='actual-hours-status' class='meta'></span>
</div>
<div class='scorecards'>
<div class='score' id='score-projects-card'><div class='sl'>Projects In View</div><div class='sv' id='sp'>0</div></div>
<div class='score' id='score-planned-epics-card'><div class='sl'>Planned Epics</div><div class='sv' id='se'>0</div></div>
<div class='score' id='score-not-planned-epics-card'><div class='sl'>Not Planned Yet Epics</div><div class='sv' id='su'>0</div></div>
<div class='score' id='score-planned-hours-card'><div class='sl'>Planned Hours (Planned)</div><div class='sv' id='sh'>0</div></div>
<div class='score' id='score-not-planned-hours-card'><div class='sl'>Planned Hours (Not Planned)</div><div class='sv' id='shu'>0</div></div>
<div class='score' id='score-capacity-card'><div class='sl'>Capacity (Hours Required)</div><div class='sv' id='scap'>0</div></div>
<div class='score' id='score-actual-hours-card'><div class='sl'>Actual Hours</div><div class='sv' id='sact'>0</div></div>
</div></section><section class='gw'><div class='g' id='root'></div></section></div>
<script>
const d=__DATA__,rows=Array.isArray(d.rows)?d.rows:[];
const byId=new Map(),kids=new Map(),collapsed=new Set();
for(const r of rows){byId.set(r.id,r);const p=r.parent_id||null;if(!kids.has(p))kids.set(p,[]);kids.get(p).push(r.id)}
const DAY=86400000,MINW=920,DAYPX=12;
const $=id=>document.getElementById(id);
const gen=$("gen"),src=$("src"),ve=$("ve"),from=$("from"),to=$("to"),root=$("root"),modeSel=$("actual-hours-mode"),statusEl=$("actual-hours-status");
const sp=$("sp"),se=$("se"),su=$("su"),sh=$("sh"),shu=$("shu"),scap=$("scap"),sact=$("sact");
const ACTUAL_MODE_STORAGE_KEY="actual-hours-mode:planned-rmis";
const ACTUAL_MODE_DEFAULT="log_date";
const ACTUAL_AGG_ENDPOINT="/api/actual-hours/aggregate";
gen.textContent=d.generated_at||"-";src.textContent=d.source_file||"-";from.value=d.default_from||"";to.value=d.default_to||"";
const storedMode=localStorage.getItem(ACTUAL_MODE_STORAGE_KEY);
if(modeSel){modeSel.value=(storedMode==="planned_dates"||storedMode==="log_date")?storedMode:ACTUAL_MODE_DEFAULT;}
const parseIso=s=>{if(!s)return null;const x=new Date(`${s}T00:00:00`);return Number.isNaN(x.getTime())?null:x};
const iso=x=>`${x.getFullYear()}-${String(x.getMonth()+1).padStart(2,"0")}-${String(x.getDate()).padStart(2,"0")}`;
const num=v=>{const n=Number(v);return Number.isFinite(n)?n:0};
const fh=v=>{const n=Number(v);return Number.isFinite(n)?String(n.toFixed(2)).replace(/\\.00$/,""):"-"};
const setStatus=t=>{if(statusEl)statusEl.textContent=String(t||"");};
const hasChildren=id=>(kids.get(id)||[]).length>0;
const overlap=(a,b,c,d)=>a<=d&&b>=c;
const addDays=(x,d)=>new Date(x.getTime()+d*DAY);
const mStart=d=>new Date(d.getFullYear(),d.getMonth(),1);
const mEnd=d=>new Date(d.getFullYear(),d.getMonth()+1,0);
const wStart=d=>{const x=new Date(d.getTime());const day=x.getDay();x.setDate(x.getDate()-(day===0?6:day-1));x.setHours(0,0,0,0);return x};
function expanded(row){let cur=row;while(cur&&cur.parent_id){if(collapsed.has(cur.parent_id))return false;cur=byId.get(cur.parent_id)||null}return true}
function collectAnc(id,keep){let cur=byId.get(id);while(cur&&cur.parent_id){keep.add(cur.parent_id);cur=byId.get(cur.parent_id)||null}}
function descendantsOf(id){const out=[];const stack=[id];while(stack.length){const cur=stack.pop();for(const child of (kids.get(cur)||[])){out.push(child);stack.push(child)}}return out}
function rangeForFocus(id){const target=byId.get(id);if(!target)return null;let min=null,max=null;const consider=row=>{if(row&&row.has_range&&Number.isFinite(Number(row.start_ordinal))&&Number.isFinite(Number(row.end_ordinal))){const s=Number(row.start_ordinal),e=Number(row.end_ordinal);min=min===null?s:Math.min(min,s);max=max===null?e:Math.max(max,e)}};consider(target);for(const did of descendantsOf(id)){consider(byId.get(did))}if(min===null||max===null)return null;return{min,max}}
function focusRow(id){const rg=rangeForFocus(id);if(!rg)return;from.value=iso(new Date((rg.min-719163)*DAY));to.value=iso(new Date((rg.max-719163)*DAY));render();setTimeout(()=>{const b=root.querySelector(`[data-bar-id='${id}']`)||root.querySelector(`[data-row-id='${id}'] .bar`);if(b&&b.scrollIntoView)b.scrollIntoView({behavior:"smooth",block:"nearest",inline:"center"})},0)}
function parseRange(){let a=parseIso(from.value),b=parseIso(to.value);if(!a)a=parseIso(d.default_from);if(!b)b=parseIso(d.default_to);if(!a||!b||b<a){const n=new Date();a=new Date(n.getFullYear(),0,1);b=new Date(n.getFullYear(),n.getMonth(),n.getDate())}from.value=iso(a);to.value=iso(b);return{a,b}}
function monthBlocks(a,b){const out=[],td=Math.max(1,Math.floor((b-a)/DAY)+1);let c=mStart(a);while(c<=b){const s=new Date(c.getTime()),e=mEnd(c),cs=s<a?a:s,ce=e>b?b:e,ld=Math.floor((cs-a)/DAY),wd=Math.floor((ce-cs)/DAY)+1;out.push({l:s.toLocaleDateString(undefined,{month:"short",year:"numeric"}),x:(ld/td)*100,w:(wd/td)*100});c=new Date(c.getFullYear(),c.getMonth()+1,1)}return out}
function weekBlocks(a,b){const out=[],td=Math.max(1,Math.floor((b-a)/DAY)+1);let c=wStart(a);while(c<=b){const s=new Date(c.getTime()),e=addDays(s,6),cs=s<a?a:s,ce=e>b?b:e,ld=Math.floor((cs-a)/DAY),wd=Math.floor((ce-cs)/DAY)+1;out.push({l:`Wk ${cs.toLocaleDateString(undefined,{day:"2-digit",month:"short"})}`,x:(ld/td)*100,w:(wd/td)*100});c=addDays(c,7)}return out}
function visibleRows(r){const keep=new Set();for(const x of rows){if(x.row_kind!=="epic")continue;if(x.group_type==="not_planned"){keep.add(x.id);collectAnc(x.id,keep);continue}const s=parseIso(x.planned_start),e=parseIso(x.planned_end);if(s&&e&&overlap(s,e,r.a,r.b)){keep.add(x.id);collectAnc(x.id,keep)}}return rows.filter(x=>keep.has(x.id)&&expanded(x))}
async function fetchActualHoursByMode(fromIso,toIso,mode){const query=`from=${encodeURIComponent(fromIso)}&to=${encodeURIComponent(toIso)}&mode=${encodeURIComponent(mode)}&report=planned_rmis`;const response=await fetch(`${ACTUAL_AGG_ENDPOINT}?${query}`,{method:"GET"});const payload=await response.json().catch(()=>({}));if(!response.ok||!payload||payload.ok===false){throw new Error(String(payload&&payload.error||"Failed to fetch actual hours."));}return payload;}
function applyFetchedActualHours(payload){const epicHours=payload&&payload.epic_hours_by_issue&&typeof payload.epic_hours_by_issue==="object"?payload.epic_hours_by_issue:{};for(const row of rows){if(row.row_kind!=="epic")continue;const key=String(row&&row.jira_key||"").trim().toUpperCase();if(key){row.actual_hours=num(epicHours[key]);}else{row.actual_hours=num(row.original_actual_hours);}}}
function render(){const r=parseRange();const vr=visibleRows(r);if(!vr.length){root.innerHTML="<div class='empty'>No rows found for selected filters.</div>";ve.textContent=sp.textContent=se.textContent=su.textContent="0";sh.textContent=shu.textContent=scap.textContent=sact.textContent="0";return}
const ep=vr.filter(x=>x.row_kind==="epic"),pep=ep.filter(x=>x.group_type==="planned"),uep=ep.filter(x=>x.group_type==="not_planned");
const pids=new Set(ep.map(x=>{const g=byId.get(x.parent_id),p=g&&g.row_kind==="group"?byId.get(g.parent_id):null;return p?p.id:null}).filter(x=>x!==null));
const plannedHours=pep.reduce((s,x)=>s+num(x.planned_hours),0),unplannedHours=uep.reduce((s,x)=>s+num(x.planned_hours),0),actualHours=ep.reduce((s,x)=>s+num(x.actual_hours),0);
ve.textContent=String(ep.length);sp.textContent=String(pids.size);se.textContent=String(pep.length);su.textContent=String(uep.length);sh.textContent=fh(plannedHours);shu.textContent=fh(unplannedHours);scap.textContent=fh(plannedHours+unplannedHours);sact.textContent=fh(actualHours);
const rr=vr.filter(x=>x.has_range&&Number.isFinite(Number(x.start_ordinal))&&Number.isFinite(Number(x.end_ordinal)));
const min=rr.length?Math.min(...rr.map(x=>Number(x.start_ordinal))):(Math.floor(r.a.getTime()/DAY)+719163);
const max=rr.length?Math.max(...rr.map(x=>Number(x.end_ordinal))):(Math.floor(r.b.getTime()/DAY)+719163);
const days=Math.max(1,(max-min)+1),tw=Math.max(MINW,Math.floor(days*DAYPX)),a=new Date((min-719163)*DAY),b=new Date((max-719163)*DAY);
const mh=monthBlocks(a,b).map(m=>`<div class='mb' style='left:${m.x}%;width:${m.w}%'>${m.l}</div>`).join(""),wb=weekBlocks(a,b),wh=wb.map(w=>`<div class='wb' style='left:${w.x}%;width:${w.w}%'>${w.l}</div>`).join(""),wln=wb.map(w=>`<span class='wl' style='left:${w.x}%'></span>`).join("");
const rh=vr.map(x=>{const lv=Number(x.level||1),pad=Math.max(0,(lv-1)*20),tg=hasChildren(x.id)?`<button class='toggle' data-i='${x.id}' type='button'>${collapsed.has(x.id)?"+":"-"}</button>`:"<button class='toggle placeholder' type='button' disabled>+</button>",rc=x.row_kind==="project"?"row-project":(x.row_kind==="group"?"row-group":"row-epic"),gc=x.group_type==="planned"?"group-planned":(x.group_type==="not_planned"?"group-not-planned":""),chip=x.row_kind==="epic"?(x.group_type==="planned"?"<span class='ch p'>Planned</span>":"<span class='ch u'>Not Planned</span>"):"";
let bar="";if(x.has_range&&Number.isFinite(Number(x.start_ordinal))&&Number.isFinite(Number(x.end_ordinal))){const ld=Number(x.start_ordinal)-min,sd=(Number(x.end_ordinal)-Number(x.start_ordinal))+1,l=(ld/days)*100,w=(Math.max(1,sd)/days)*100,bc=x.row_kind==="project"?"bar project":"bar";bar=`<div class='${bc}' data-bar-id='${x.id}' style='left:${l}%;width:${w}%' title='${x.rmi_label||""}'></div>`}else if(x.row_kind==="epic"&&x.group_type==="not_planned"){bar="<div class='bar no' style='left:0;width:42px' title='Not planned yet'></div>"}
const label=(x.row_kind==="epic"||x.row_kind==="project")?`<button class='focus-link' type='button' data-focus-id='${x.id}' title='Focus timeline'>${x.rmi_label||""}</button>`:`<span class='lbl' title='${x.rmi_label||""}'>${x.rmi_label||""}</span>`;
return `<div class='r ${rc} ${gc}' data-row-id='${x.id}'><div class='c s1'><div class='rmi' style='padding-left:${pad}px'>${tg}${label}${chip}</div></div><div class='c s2'>${fh(x.planned_hours)}</div><div class='c tc'><div class='slot' style='width:${tw}px'>${wln}${bar}</div></div></div>`}).join("");
root.innerHTML=`<div style='width:calc(var(--left1)+var(--left2)+${tw}px)'><div class='r h'><div class='c s1'>RMIs</div><div class='c s2'>Planned Hours</div><div class='c th'><div class='months'>${mh}</div><div class='weeks'>${wh}</div></div></div>${rh}</div>`;
for(const b of root.querySelectorAll("[data-i]")){b.onclick=()=>{const id=Number(b.getAttribute("data-i"));if(collapsed.has(id))collapsed.delete(id);else collapsed.add(id);render()}}
for(const f of root.querySelectorAll("[data-focus-id]")){f.onclick=()=>{const id=Number(f.getAttribute("data-focus-id"));if(Number.isFinite(id))focusRow(id)}}
}
async function applyActualMode(){const r=parseRange();const mode=modeSel?(String(modeSel.value||ACTUAL_MODE_DEFAULT)):"log_date";if(modeSel){localStorage.setItem(ACTUAL_MODE_STORAGE_KEY,mode);}if(!(window.location.protocol||"").startsWith("http")){setStatus("Static mode: API unavailable, using embedded actual hours.");render();return;}try{setStatus("Recomputing actual hours...");const payload=await fetchActualHoursByMode(iso(r.a),iso(r.b),mode);applyFetchedActualHours(payload);setStatus("");}catch(err){setStatus(String(err&&err.message||err||"Failed to fetch actual hours."));}render();}
$("apply").onclick=()=>{applyActualMode();};$("reset").onclick=()=>{from.value=d.default_from||"";to.value=d.default_to||"";applyActualMode();};$("collapse").onclick=()=>{collapsed.clear();for(const r of rows){if(r.row_kind==="project"&&hasChildren(r.id))collapsed.add(r.id)}render()};$("expand").onclick=()=>{collapsed.clear();render()};if(modeSel){modeSel.addEventListener("change",()=>{applyActualMode();});}applyActualMode();
</script>
<script src='shared-nav.js'></script>
</body>
</html>""".replace("__DATA__", data)


def main() -> None:
    base_dir = Path(__file__).resolve().parent
    input_name = os.getenv("JIRA_PLANNED_RMIS_INPUT_XLSX_PATH", DEFAULT_INPUT_XLSX).strip() or DEFAULT_INPUT_XLSX
    work_items_name = os.getenv("JIRA_EXPORT_XLSX_PATH", DEFAULT_WORK_ITEMS_XLSX).strip() or DEFAULT_WORK_ITEMS_XLSX
    output_name = os.getenv("JIRA_PLANNED_RMIS_HTML_PATH", DEFAULT_OUTPUT_HTML).strip() or DEFAULT_OUTPUT_HTML
    input_path = _resolve_path(input_name, base_dir)
    work_items_path = _resolve_path(work_items_name, base_dir)
    if not input_path.exists():
        raise FileNotFoundError(f"Nested view workbook not found: {input_path}")
    payload = _load_payload(input_path, work_items_path)
    payload["generated_at"] = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    output_path = _resolve_path(output_name, base_dir)
    output_path.write_text(_build_html(payload), encoding="utf-8")
    print(f"Report written: {output_path}")


if __name__ == "__main__":
    main()

