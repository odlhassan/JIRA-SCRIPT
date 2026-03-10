"""
Roll up sub-task worklog rows into one row per issue.

Input: Excel produced by export_jira_subtask_worklogs.py
Output: Excel with actual min/max worklog timestamps and total logged hours per sub-task.
"""
from __future__ import annotations

import os
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook, load_workbook

from ipp_meeting_utils import (
    fetch_jira_issue_planned_dates,
    load_ipp_actual_and_remarks_by_key,
    load_ipp_issue_keys,
    load_ipp_planned_dates_by_key,
    resolve_jira_start_date_field_id,
    yes_no_dates_altered,
    yes_no_ipp_actual_matches_jira_end,
    yes_no_in_ipp,
)
from jira_export_db import (
    connect as export_db_connect,
    ensure_schema as ensure_exports_schema,
    has_subtask_worklogs,
    read_subtask_worklogs as read_subtask_worklogs_db,
    record_export_run,
    write_subtask_worklog_rollup as write_subtask_worklog_rollup_db,
)
from jira_client import BASE_URL, get_session

DEFAULT_INPUT = "2_jira_subtask_worklogs.xlsx"
DEFAULT_OUTPUT = "3_jira_subtask_worklog_rollup.xlsx"

INPUT_HEADERS = [
    "issue_link",
    "issue_id",
    "issue_title",
    "issue_type",
    "parent_story_link",
    "parent_story_id",
    "parent_epic_id",
    "issue_assignee",
    "worklog_started",
    "hours_logged",
]

OUTPUT_HEADERS = [
    "issue_link",
    "issue_id",
    "issue_title",
    "issue_type",
    "parent_story_link",
    "parent_story_id",
    "parent_epic_id",
    "issue_assignee",
    "Latest IPP Meeting",
    "Jira IPP RMI Dates Altered",
    "IPP Actual Date (Production Date)",
    "IPP Remarks",
    "IPP Actual Date Matches Jira End Date",
    "planned start date",
    "planned end date",
    "actual start date",
    "actual end date",
    "total hours_logged",
]


def _parse_worklog_ts(value) -> datetime | None:
    if not value:
        return None
    text = str(value).strip()
    if not text:
        return None
    formats = [
        "%Y-%m-%dT%H:%M:%S.%f%z",
        "%Y-%m-%dT%H:%M:%S%z",
    ]
    for fmt in formats:
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def _first_non_empty(*values):
    for value in values:
        if value is None:
            continue
        text = str(value).strip()
        if text:
            return value
    return ""


def _read_input_rows(input_path: Path) -> list[dict]:
    wb = load_workbook(input_path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(min_row=1, max_row=1, values_only=True)
    header_row = next(rows_iter, None)
    if not header_row:
        wb.close()
        raise ValueError("Input workbook has no header row")

    headers = [str(h).strip() if h is not None else "" for h in header_row]
    missing = [h for h in INPUT_HEADERS if h not in headers]
    if missing:
        wb.close()
        raise ValueError(f"Input workbook missing required columns: {missing}")

    index_map = {name: headers.index(name) for name in INPUT_HEADERS}
    data_rows: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        issue_id = row[index_map["issue_id"]]
        if issue_id is None or not str(issue_id).strip():
            continue
        item = {name: row[index_map[name]] for name in INPUT_HEADERS}
        data_rows.append(item)

    wb.close()
    return data_rows


def _fetch_issues(session, jql: str, fields: list[str]) -> list[dict]:
    url = f"{BASE_URL}/rest/api/3/search/jql"
    issues: list[dict] = []
    next_page_token = None
    while True:
        payload = {"jql": jql, "maxResults": 100, "fields": fields}
        if next_page_token:
            payload["nextPageToken"] = next_page_token
        response = session.post(url, json=payload)
        response.raise_for_status()
        data = response.json()
        issues.extend(data.get("issues", []))
        next_page_token = data.get("nextPageToken")
        if not next_page_token:
            break
    return issues


def _fetch_story_planned_dates(session, story_ids: set[str]) -> dict[str, dict]:
    if not story_ids:
        return {}
    start_field_id = resolve_jira_start_date_field_id(session, BASE_URL)
    fields = ["duedate"]
    if start_field_id:
        fields.append(start_field_id)

    result: dict[str, dict] = {}
    story_list = sorted(story_ids)
    for offset in range(0, len(story_list), 500):
        chunk = story_list[offset : offset + 500]
        keys = ", ".join(f'"{story_id}"' for story_id in chunk)
        jql = f"key in ({keys})"
        issues = _fetch_issues(session, jql=jql, fields=fields)
        for issue in issues:
            story_key = issue.get("key", "")
            issue_fields = issue.get("fields", {}) or {}
            planned_start = issue_fields.get(start_field_id, "") if start_field_id else ""
            planned_end = issue_fields.get("duedate", "")
            result[story_key] = {
                "planned start date": planned_start or "",
                "planned end date": planned_end or "",
            }
    return result


def _rollup_rows(
    rows: list[dict],
    planned_dates: dict[str, dict],
    ipp_issue_keys: set[str],
    ipp_planned_dates: dict[str, dict[str, str]],
    ipp_actual_by_key: dict[str, dict[str, str]],
    jira_epic_dates: dict[str, dict[str, str]],
) -> list[list[object]]:
    grouped: dict[str, dict] = {}

    for row in rows:
        issue_id = str(row.get("issue_id", "")).strip()
        if not issue_id:
            continue

        group = grouped.get(issue_id)
        if group is None:
            group = {
                "issue_link": row.get("issue_link", ""),
                "issue_id": issue_id,
                "issue_title": row.get("issue_title", ""),
                "issue_type": row.get("issue_type", ""),
                "parent_story_link": row.get("parent_story_link", ""),
                "parent_story_id": row.get("parent_story_id", ""),
                "parent_epic_id": row.get("parent_epic_id", ""),
                "issue_assignee": row.get("issue_assignee", ""),
                "actual_min_dt": None,
                "actual_min_raw": "",
                "actual_max_dt": None,
                "actual_max_raw": "",
                "total_hours": 0.0,
            }
            grouped[issue_id] = group
        else:
            group["issue_link"] = _first_non_empty(group["issue_link"], row.get("issue_link", ""))
            group["issue_title"] = _first_non_empty(group["issue_title"], row.get("issue_title", ""))
            group["issue_type"] = _first_non_empty(group["issue_type"], row.get("issue_type", ""))
            group["parent_story_link"] = _first_non_empty(group["parent_story_link"], row.get("parent_story_link", ""))
            group["parent_story_id"] = _first_non_empty(group["parent_story_id"], row.get("parent_story_id", ""))
            group["parent_epic_id"] = _first_non_empty(group["parent_epic_id"], row.get("parent_epic_id", ""))
            group["issue_assignee"] = _first_non_empty(group["issue_assignee"], row.get("issue_assignee", ""))

        ts_raw = row.get("worklog_started", "")
        ts_dt = _parse_worklog_ts(ts_raw)
        if ts_dt is not None:
            if group["actual_min_dt"] is None or ts_dt < group["actual_min_dt"]:
                group["actual_min_dt"] = ts_dt
                group["actual_min_raw"] = ts_raw
            if group["actual_max_dt"] is None or ts_dt > group["actual_max_dt"]:
                group["actual_max_dt"] = ts_dt
                group["actual_max_raw"] = ts_raw

        hours_value = row.get("hours_logged", 0)
        try:
            group["total_hours"] += float(hours_value or 0)
        except (TypeError, ValueError):
            pass

    output_rows: list[list[object]] = []
    for issue_id in sorted(grouped):
        item = grouped[issue_id]
        parent_story_id = str(item.get("parent_story_id", "")).strip()
        parent_epic_id = str(item.get("parent_epic_id", "")).strip()
        parent_epic_key = parent_epic_id.upper()
        planned = planned_dates.get(parent_story_id, {})
        ipp_actual_data = ipp_actual_by_key.get(parent_epic_key, {})
        output_rows.append(
            [
                item["issue_link"],
                item["issue_id"],
                item["issue_title"],
                item["issue_type"],
                item["parent_story_link"],
                item["parent_story_id"],
                item["parent_epic_id"],
                item["issue_assignee"],
                yes_no_in_ipp(item["issue_id"], ipp_issue_keys),
                yes_no_dates_altered(parent_epic_id, ipp_planned_dates, jira_epic_dates),
                ipp_actual_data.get("ipp_actual_date", ""),
                ipp_actual_data.get("ipp_remarks", ""),
                yes_no_ipp_actual_matches_jira_end(parent_epic_key, ipp_actual_by_key, jira_epic_dates),
                planned.get("planned start date", ""),
                planned.get("planned end date", ""),
                item["actual_min_raw"],
                item["actual_max_raw"],
                round(item["total_hours"], 2),
            ]
        )
    return output_rows


def _write_output(rows: list[list[object]], output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "SubtaskRollup"
    ws.append(OUTPUT_HEADERS)
    for row in rows:
        ws.append(row)
    wb.save(output_path)


def main() -> None:
    input_name = os.getenv("JIRA_SUBTASK_WORKLOG_INPUT_XLSX_PATH", DEFAULT_INPUT).strip() or DEFAULT_INPUT
    output_name = os.getenv("JIRA_SUBTASK_ROLLUP_XLSX_PATH", DEFAULT_OUTPUT).strip() or DEFAULT_OUTPUT

    input_path = Path(input_name)
    if not input_path.is_absolute():
        input_path = Path(__file__).resolve().parent / input_path

    output_path = Path(output_name)
    if not output_path.is_absolute():
        output_path = Path(__file__).resolve().parent / output_path

    # Primary: read worklogs from exports DB; fallback to xlsx
    export_conn = export_db_connect()
    ensure_exports_schema(export_conn)
    if has_subtask_worklogs(export_conn):
        input_rows = read_subtask_worklogs_db(export_conn)
        print("Reading input worklogs: DB")
    else:
        if not input_path.exists():
            raise FileNotFoundError(f"Input file not found (and DB has no worklogs): {input_path}")
        print(f"Reading input worklogs: {input_path}")
        input_rows = _read_input_rows(input_path)
    print(f"Input rows: {len(input_rows)}")

    story_ids = {
        str(row.get("parent_story_id", "")).strip()
        for row in input_rows
        if row.get("parent_story_id") and str(row.get("parent_story_id", "")).strip()
    }
    print(f"Parent stories to fetch: {len(story_ids)}")

    session = get_session()
    planned_dates = _fetch_story_planned_dates(session, story_ids)
    print(f"Parent story plans fetched: {len(planned_dates)}")
    ipp_issue_keys = load_ipp_issue_keys()
    print(f"IPP issue keys loaded: {len(ipp_issue_keys)}")
    ipp_planned_dates = load_ipp_planned_dates_by_key()
    print(f"IPP planned-date keys loaded: {len(ipp_planned_dates)}")
    ipp_actual_by_key = load_ipp_actual_and_remarks_by_key()
    print(f"IPP actual-date/remarks keys loaded: {len(ipp_actual_by_key)}")
    epic_ids = {
        str(row.get("parent_epic_id", "")).strip()
        for row in input_rows
        if row.get("parent_epic_id") and str(row.get("parent_epic_id", "")).strip()
    }
    jira_epic_dates = fetch_jira_issue_planned_dates(session, BASE_URL, epic_ids)
    print(f"Jira epic planned-date keys loaded: {len(jira_epic_dates)}")

    output_rows = _rollup_rows(
        input_rows,
        planned_dates,
        ipp_issue_keys=ipp_issue_keys,
        ipp_planned_dates=ipp_planned_dates,
        ipp_actual_by_key=ipp_actual_by_key,
        jira_epic_dates=jira_epic_dates,
    )
    print(f"Output rollup rows: {len(output_rows)}")

    # Primary: write rollup to exports DB
    write_subtask_worklog_rollup_db(export_conn, output_rows)
    run_ended = datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")
    record_export_run(export_conn, "subtask_rollup", len(output_rows), finished_at_utc=run_ended)
    export_conn.close()

    # Secondary: write xlsx
    _write_output(output_rows, output_path)
    print(f"Export written: {output_path}")


if __name__ == "__main__":
    main()
