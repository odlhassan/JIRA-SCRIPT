import openpyxl
wb = openpyxl.load_workbook(r'c:\Users\hmalik\Downloads\Epic Estimates Approved Plan (3).xlsx', data_only=True)
rmi_sheets = [s for s in wb.sheetnames if 'RMI' in s.upper()]
print("RMI sheets:", rmi_sheets)
if rmi_sheets:
    ws = wb[rmi_sheets[0]]
    print("\nRow 1 non-empty:")
    for c in range(1, ws.max_column+1):
        v = ws.cell(1,c).value
        if v is not None: print(f"  col {c}: {repr(v)}")
    print("\nRow 2 non-empty:")
    for c in range(1, ws.max_column+1):
        v = ws.cell(2,c).value
        if v is not None: print(f"  col {c}: {repr(v)}")
    print("\nSearching for O2-1461...")
    for ri in range(3, ws.max_row+1):
        row_vals = [ws.cell(ri,c).value for c in range(1, ws.max_column+1)]
        row_str = ' '.join(str(v) for v in row_vals if v is not None)
        if 'O2-1461' in row_str:
            print(f"Found at row {ri}:")
            for c,v in enumerate(row_vals, 1): 
                if v is not None: print(f"  col {c}: {repr(v)}")
            break
