import openpyxl

# Load the workbook
wb = openpyxl.load_workbook(r'c:\Users\hmalik\Downloads\Epic Estimates Approved Plan (3).xlsx', data_only=True)

# Find all sheets with "RMI" in the title
rmi_sheets = [sheet for sheet in wb.sheetnames if 'RMI' in sheet]

print(f"Sheets with 'RMI' in title: {rmi_sheets}")
print()

if not rmi_sheets:
    print("No sheets with 'RMI' found!")
else:
    # Work with the first RMI sheet
    sheet_name = rmi_sheets[0]
    ws = wb[sheet_name]
    print(f"Working with sheet: {sheet_name}")
    print("=" * 80)
    
    # Task 1: Print ALL non-empty cell values in rows 1 and 2
    print("\n1. NON-EMPTY CELLS IN ROWS 1 AND 2:")
    print("-" * 80)
    for row_num in [1, 2]:
        print(f"\nRow {row_num}:")
        for col_idx, cell in enumerate(ws[row_num], start=1):
            if cell.value is not None:
                print(f"  Col {col_idx}: {cell.value}")
    
    # Task 2 & 3: Find epic "O2-1461" and print all columns + special columns
    print("\n\n2. SEARCHING FOR EPIC 'O2-1461':")
    print("-" * 80)
    
    # First, get the header row (row 1)
    headers = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        if cell.value is not None:
            headers[col_idx] = str(cell.value).lower()
    
    # Find columns with "date", "start", or "prod" in header
    special_cols = {}
    for col_idx, header in headers.items():
        if any(keyword in header for keyword in ["date", "start", "prod"]):
            special_cols[col_idx] = headers[col_idx]
    
    print(f"Special columns (containing 'date', 'start', or 'prod'): {special_cols}")
    
    # Search for O2-1461
    epic_found = False
    for row in ws.iter_rows(min_row=2, values_only=False):
        for cell in row:
            if cell.value == "O2-1461":
                epic_found = True
                row_num = cell.row
                print(f"\nFound 'O2-1461' at row {row_num}")
                
                # Task 2: Print ALL column values for this row
                print("\nALL COLUMN VALUES FOR THIS ROW:")
                for col_idx in range(1, ws.max_column + 1):
                    cell_val = ws.cell(row=row_num, column=col_idx).value
                    header = headers.get(col_idx, f"Col{col_idx}")
                    print(f"  Col {col_idx} ({header}): {cell_val}")
                
                # Task 3: Print raw cell values for special columns
                print("\nRAW VALUES FOR SPECIAL COLUMNS (date/start/prod):")
                for col_idx in sorted(special_cols.keys()):
                    cell = ws.cell(row=row_num, column=col_idx)
                    print(f"  Col {col_idx} ({special_cols[col_idx]}): {repr(cell.value)} (type: {type(cell.value).__name__})")
                
                break
        if epic_found:
            break
    
    if not epic_found:
        print("\nEpic 'O2-1461' NOT FOUND in this sheet!")
