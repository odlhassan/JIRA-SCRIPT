"""
Populate IPP Meeting Work Items Estimates.xlsx from Epic Estimates Approved Plan.xlsx
by extracting RMI rows with dates and matching them against Jira.
"""

from __future__ import annotations

import base64
import re
from difflib import SequenceMatcher
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill


BASE_DIR = Path(__file__).parent
REPORTS_DIR = BASE_DIR / "IPP Meeting Reports"
EPIC_PLAN_FILE = REPORTS_DIR / "Epic Estimates Approved Plan.xlsx"
IPP_OUTPUT_FILE = REPORTS_DIR / "IPP Meeting Work Items Estimates.xlsx"
UNMATCHED_REPORT_FILE = REPORTS_DIR / "Unmatched_RMIs_Report.xlsx"
STORY_SHEET_NAME = "Matched Epic Stories"
ENV_CANDIDATES = [
    BASE_DIR / ".env",
    BASE_DIR / "Documentation" / ".env",
]

RMI_NAME_COL = "D"
DATE_COLS = {
    "Z": "Start Date",
    "AA": "Dev End",
    "AB": "SQA HO",
    "AC": "Prod Date",
}
ESTIMATE_COLS = {
    "U": "Man Days",
    "V": "Optimistic (50%)",
    "W": "Pessimistic (10%)",
    "X": "Est Formula",
    "Y": "TK's TARGET",
}

MISSING_FIELD_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")


def load_env_config() -> Dict[str, str]:
    env_path = next((path for path in ENV_CANDIDATES if path.exists()), None)
    if env_path is None:
        raise FileNotFoundError("Could not find a .env file with Jira credentials.")

    config: Dict[str, str] = {}
    for raw_line in env_path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        config[key.strip()] = value.strip()

    required = ["JIRA_SITE", "JIRA_EMAIL", "JIRA_API_TOKEN", "JIRA_PROJECT_KEYS"]
    missing = [key for key in required if not config.get(key)]
    if missing:
        raise ValueError(f"Missing Jira settings in {env_path}: {', '.join(missing)}")

    config["ENV_PATH"] = str(env_path)
    return config


ENV_CONFIG = load_env_config()
JIRA_SITE = ENV_CONFIG["JIRA_SITE"]
JIRA_EMAIL = ENV_CONFIG["JIRA_EMAIL"]
JIRA_API_TOKEN = ENV_CONFIG["JIRA_API_TOKEN"]
JIRA_PROJECT_KEYS = [value.strip() for value in ENV_CONFIG["JIRA_PROJECT_KEYS"].split(",") if value.strip()]


class JiraAPIClient:
    def __init__(self, site: str, email: str, token: str):
        self.site = site
        self.email = email
        self.token = token
        self.base_url_v3 = f"https://{site}.atlassian.net/rest/api/3"
        self.search_url = f"{self.base_url_v3}/search/jql"
        self.session = requests.Session()

        auth_str = f"{email}:{token}"
        auth_b64 = base64.b64encode(auth_str.encode("utf-8")).decode("utf-8")
        self.session.headers.update(
            {
                "Authorization": f"Basic {auth_b64}",
                "Accept": "application/json",
                "Content-Type": "application/json",
            }
        )

    @staticmethod
    def _escape_jql_phrase(text: str) -> str:
        return text.replace("\\", "\\\\").replace('"', '\\"')

    @staticmethod
    def map_issue(issue: Dict) -> Dict:
        timetracking = issue["fields"].get("timetracking") or {}
        return {
            "key": issue["key"],
            "summary": issue["fields"]["summary"],
            "issue_type": issue["fields"]["issuetype"]["name"],
            "status": issue["fields"]["status"]["name"],
            "priority": issue["fields"]["priority"]["name"] if issue["fields"].get("priority") else "N/A",
            "story_points": issue["fields"].get("customfield_10016", 0) or 0,
            "jira_start_date": issue["fields"].get("customfield_10015"),
            "jira_due_date": issue["fields"].get("duedate"),
            "jira_original_estimate": timetracking.get("originalEstimate"),
            "jira_original_estimate_seconds": timetracking.get("originalEstimateSeconds")
            or issue["fields"].get("timeoriginalestimate"),
            "jira_aggregate_original_estimate_seconds": issue["fields"].get("aggregatetimeoriginalestimate"),
        }

    def test_connection(self) -> bool:
        try:
            response = self.session.get(f"{self.base_url_v3}/myself", timeout=15)
            response.raise_for_status()
            display_name = response.json().get("displayName", "User")
            print(f"Connected to Jira as: {display_name}")
            return True
        except requests.exceptions.RequestException as exc:
            print(f"Failed to connect to Jira: {exc}")
            return False

    def get_issue(self, issue_key: str) -> Dict:
        response = self.session.get(
            f"{self.base_url_v3}/issue/{issue_key}",
            params={
                "fields": ",".join(
                    [
                        "summary",
                        "issuetype",
                        "status",
                        "priority",
                        "customfield_10015",
                        "customfield_10016",
                        "duedate",
                        "timetracking",
                        "timeoriginalestimate",
                        "aggregatetimeoriginalestimate",
                    ]
                )
            },
            timeout=20,
        )
        response.raise_for_status()
        return self.map_issue(response.json())

    def search_by_text(self, text: str, project_keys: List[str], max_results: int = 10) -> List[Dict]:
        project_filter = " OR ".join(f"project = {key}" for key in project_keys)
        escaped_text = self._escape_jql_phrase(text)
        fields = [
            "summary",
            "issuetype",
            "status",
            "priority",
            "customfield_10015",
            "customfield_10016",
            "labels",
            "duedate",
            "timetracking",
            "timeoriginalestimate",
            "aggregatetimeoriginalestimate",
        ]
        queries = [
            f'({project_filter}) AND summary ~ "\\"{escaped_text}\\"" ORDER BY updated DESC',
            f'({project_filter}) AND text ~ "\\"{escaped_text}\\"" ORDER BY updated DESC',
        ]

        results: List[Dict] = []
        try:
            for jql in queries:
                response = self.session.post(
                    self.search_url,
                    json={"jql": jql, "maxResults": max_results, "fields": fields},
                    timeout=20,
                )
                response.raise_for_status()

                for issue in response.json().get("issues", []):
                    entry = self.map_issue(issue)
                    if entry["key"] not in {item["key"] for item in results}:
                        results.append(entry)

                if results:
                    break
        except requests.exceptions.RequestException as exc:
            print(f"Jira search error for {text}: {exc}")

        return results

    def get_child_stories(self, epic_key: str, max_results: int = 100) -> List[Dict]:
        fields = [
            "summary",
            "issuetype",
            "status",
            "priority",
            "customfield_10015",
            "duedate",
            "timetracking",
            "timeoriginalestimate",
            "aggregatetimeoriginalestimate",
            "parent",
        ]
        queries = [
            f'"Epic Link" = {epic_key} ORDER BY Rank ASC',
            f'parent = {epic_key} ORDER BY Rank ASC',
            f'parentEpic = {epic_key} AND issuekey != {epic_key} ORDER BY Rank ASC',
        ]

        for jql in queries:
            try:
                response = self.session.post(
                    self.search_url,
                    json={"jql": jql, "maxResults": max_results, "fields": fields},
                    timeout=20,
                )
                response.raise_for_status()
                issues = response.json().get("issues", [])
                if issues:
                    results = []
                    for issue in issues:
                        entry = self.map_issue(issue)
                        parent = issue["fields"].get("parent") or {}
                        parent_fields = parent.get("fields") or {}
                        entry["parent_key"] = parent.get("key")
                        entry["parent_summary"] = parent_fields.get("summary")
                        results.append(entry)
                    return results
            except requests.exceptions.RequestException as exc:
                print(f"Jira child story search error for {epic_key}: {exc}")

        return []


def fuzzy_match_score(s1: str, s2: str) -> float:
    s1_clean = re.sub(r"[^a-zA-Z0-9]", "", s1.lower())
    s2_clean = re.sub(r"[^a-zA-Z0-9]", "", s2.lower())
    return SequenceMatcher(None, s1_clean, s2_clean).ratio()


def find_best_match(rmi_name: str, jira_items: List[Dict], threshold: float = 0.6) -> Optional[Dict]:
    best_match = None
    best_score = threshold

    for item in jira_items:
        score = fuzzy_match_score(rmi_name, item["summary"])
        if score > best_score:
            best_match = dict(item)
            best_match["match_score"] = score
            best_score = score

    return best_match


def get_rmi_sheet_names(workbook) -> List[str]:
    return [sheet_name for sheet_name in workbook.sheetnames if "RMI" in sheet_name]


def extract_eligible_rmis() -> Dict[str, List[Dict]]:
    wb = load_workbook(EPIC_PLAN_FILE, data_only=True, read_only=True)
    eligible_rmis: Dict[str, List[Dict]] = {}

    for sheet_name in get_rmi_sheet_names(wb):
        ws = wb[sheet_name]
        sheet_rmis: List[Dict] = []

        for row_idx in range(3, ws.max_row + 1):
            dates = {
                date_name: ws[f"{col_letter}{row_idx}"].value
                for col_letter, date_name in DATE_COLS.items()
            }
            if not any(dates.values()):
                continue

            rmi_name = ws[f"{RMI_NAME_COL}{row_idx}"].value
            if not isinstance(rmi_name, str) or not rmi_name.strip():
                continue

            estimates = {
                estimate_name: ws[f"{col_letter}{row_idx}"].value
                for col_letter, estimate_name in ESTIMATE_COLS.items()
            }
            sheet_rmis.append(
                {
                    "sheet": sheet_name,
                    "row": row_idx,
                    "name": rmi_name.strip(),
                    "dates": dates,
                    "estimates": estimates,
                }
            )

        if sheet_rmis:
            eligible_rmis[sheet_name] = sheet_rmis

    wb.close()
    return eligible_rmis


def match_rmis_to_jira(
    eligible_rmis: Dict[str, List[Dict]], jira_client: JiraAPIClient
) -> Tuple[List[Dict], List[Dict]]:
    matched_rmis: List[Dict] = []
    unmatched_rmis: List[Dict] = []

    total = sum(len(items) for items in eligible_rmis.values())
    seen = 0

    for sheet_name, rmis in eligible_rmis.items():
        print(f"\nProcessing {sheet_name}...")
        for rmi in rmis:
            seen += 1
            print(f"  [{seen}/{total}] Searching Jira for: {rmi['name']}")
            jira_items = jira_client.search_by_text(rmi["name"], JIRA_PROJECT_KEYS, max_results=10)
            best_match = find_best_match(rmi["name"], jira_items, threshold=0.5)
            if best_match:
                rmi["jira_match"] = best_match
                matched_rmis.append(rmi)
                print(f"    Matched {best_match['key']} ({best_match['summary']})")
            else:
                unmatched_rmis.append(rmi)
                print("    No Jira match found")

    return matched_rmis, unmatched_rmis


def write_headers(ws, headers: List[str], fill_color: str) -> None:
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")


def mark_missing_required_fields(ws, row_idx: int, column_indexes: List[int]) -> None:
    for col_idx in column_indexes:
        cell = ws.cell(row=row_idx, column=col_idx)
        if cell.value in (None, ""):
            cell.fill = MISSING_FIELD_FILL


def populate_ipp_file(matched_rmis: List[Dict]) -> None:
    wb = load_workbook(IPP_OUTPUT_FILE) if IPP_OUTPUT_FILE.exists() else Workbook()

    if "Matched RMIs" in wb.sheetnames:
        ws = wb["Matched RMIs"]
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)
    else:
        ws = wb.create_sheet("Matched RMIs", 0)

    headers = [
        "Product",
        "RMI Name",
        "Row Source",
        "Jira Key",
        "Jira Summary",
        "Issue Type",
        "Status",
        "Priority",
        "Story Points",
        "Match Score",
        "Jira Start Date",
        "Jira Due Date",
        "Jira Original Estimate",
        "Jira Original Estimate Seconds",
        "Jira Aggregate Original Estimate Seconds",
        "Start Date",
        "Dev End",
        "SQA HO",
        "Prod Date",
        "Man Days",
        "Optimistic (50%)",
        "Pessimistic (10%)",
        "Est Formula",
        "TK Target",
    ]
    write_headers(ws, headers, "4472C4")

    row_idx = 2
    for rmi in matched_rmis:
        jira_match = rmi["jira_match"]
        row_data = [
            rmi["sheet"],
            rmi["name"],
            rmi["row"],
            jira_match.get("key", ""),
            jira_match.get("summary", ""),
            jira_match.get("issue_type", ""),
            jira_match.get("status", ""),
            jira_match.get("priority", ""),
            jira_match.get("story_points", ""),
            round(jira_match.get("match_score", 0), 3),
            jira_match.get("jira_start_date", ""),
            jira_match.get("jira_due_date", ""),
            jira_match.get("jira_original_estimate", ""),
            jira_match.get("jira_original_estimate_seconds", ""),
            jira_match.get("jira_aggregate_original_estimate_seconds", ""),
            rmi["dates"].get("Start Date", ""),
            rmi["dates"].get("Dev End", ""),
            rmi["dates"].get("SQA HO", ""),
            rmi["dates"].get("Prod Date", ""),
            rmi["estimates"].get("Man Days", ""),
            rmi["estimates"].get("Optimistic (50%)", ""),
            rmi["estimates"].get("Pessimistic (10%)", ""),
            rmi["estimates"].get("Est Formula", ""),
            rmi["estimates"].get("TK's TARGET", ""),
        ]
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
        mark_missing_required_fields(ws, row_idx, [11, 12, 13, 14])
        row_idx += 1

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 45
    ws.column_dimensions["J"].width = 12

    wb.save(IPP_OUTPUT_FILE)


def populate_story_sheet(matched_rmis: List[Dict], jira_client: JiraAPIClient) -> None:
    wb = load_workbook(IPP_OUTPUT_FILE) if IPP_OUTPUT_FILE.exists() else Workbook()

    if STORY_SHEET_NAME in wb.sheetnames:
        ws = wb[STORY_SHEET_NAME]
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)
    else:
        ws = wb.create_sheet(STORY_SHEET_NAME)

    headers = [
        "Product",
        "RMI Name",
        "Epic Key",
        "Epic Summary",
        "Story Key",
        "Story Summary",
        "Story Type",
        "Status",
        "Priority",
        "Jira Start Date",
        "Jira Due Date",
        "Jira Original Estimate",
        "Jira Original Estimate Seconds",
        "Jira Aggregate Original Estimate Seconds",
    ]
    write_headers(ws, headers, "70AD47")

    row_idx = 2
    for rmi in matched_rmis:
        epic = rmi["jira_match"]
        stories = jira_client.get_child_stories(epic["key"])
        for story in stories:
            row_data = [
                rmi["sheet"],
                rmi["name"],
                epic.get("key", ""),
                epic.get("summary", ""),
                story.get("key", ""),
                story.get("summary", ""),
                story.get("issue_type", ""),
                story.get("status", ""),
                story.get("priority", ""),
                story.get("jira_start_date", ""),
                story.get("jira_due_date", ""),
                story.get("jira_original_estimate", ""),
                story.get("jira_original_estimate_seconds", ""),
                story.get("jira_aggregate_original_estimate_seconds", ""),
            ]
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
            mark_missing_required_fields(ws, row_idx, [10, 11, 12, 13])
            row_idx += 1

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 35
    ws.column_dimensions["J"].width = 16
    ws.column_dimensions["K"].width = 16
    ws.column_dimensions["L"].width = 22

    wb.save(IPP_OUTPUT_FILE)


def create_unmatched_report(unmatched_rmis: List[Dict]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Unmatched RMIs"
    headers = [
        "Product",
        "RMI Name",
        "Row Source",
        "Dates Defined",
        "Start Date",
        "Dev End",
        "SQA HO",
        "Prod Date",
        "Man Days",
        "TK Target",
    ]
    write_headers(ws, headers, "C55A11")

    row_idx = 2
    for rmi in unmatched_rmis:
        dates_defined = ", ".join(key for key, value in rmi["dates"].items() if value)
        row_data = [
            rmi["sheet"],
            rmi["name"],
            rmi["row"],
            dates_defined,
            rmi["dates"].get("Start Date", ""),
            rmi["dates"].get("Dev End", ""),
            rmi["dates"].get("SQA HO", ""),
            rmi["dates"].get("Prod Date", ""),
            rmi["estimates"].get("Man Days", ""),
            rmi["estimates"].get("TK's TARGET", ""),
        ]
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
        row_idx += 1

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["D"].width = 25

    wb.save(UNMATCHED_REPORT_FILE)


def main() -> None:
    print(f"Using Jira config from: {ENV_CONFIG['ENV_PATH']}")
    eligible_rmis = extract_eligible_rmis()
    print(f"Eligible RMI count: {sum(len(items) for items in eligible_rmis.values())}")

    jira_client = JiraAPIClient(JIRA_SITE, JIRA_EMAIL, JIRA_API_TOKEN)
    if not jira_client.test_connection():
        raise SystemExit(1)

    matched_rmis, unmatched_rmis = match_rmis_to_jira(eligible_rmis, jira_client)
    populate_ipp_file(matched_rmis)
    populate_story_sheet(matched_rmis, jira_client)
    create_unmatched_report(unmatched_rmis)

    print("\nCompleted.")
    print(f"Matched RMIs: {len(matched_rmis)}")
    print(f"Unmatched RMIs: {len(unmatched_rmis)}")
    print(f"Output workbook: {IPP_OUTPUT_FILE}")
    print(f"Unmatched report: {UNMATCHED_REPORT_FILE}")


if __name__ == "__main__":
    main()
