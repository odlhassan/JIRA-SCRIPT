import sqlite3
import unittest
from pathlib import Path
from tempfile import TemporaryDirectory
from unittest.mock import patch

from openpyxl import Workbook

import generate_rmi_gantt_html as report


class GenerateRmiGanttHtmlTests(unittest.TestCase):
    def create_db(self, path: Path) -> None:
        conn = sqlite3.connect(path)
        conn.executescript(
            """
            CREATE TABLE source_rmi_rows (
                sheet_name TEXT NOT NULL,
                row_number INTEGER NOT NULL,
                roadmap_item TEXT NOT NULL,
                jira_id TEXT NOT NULL,
                man_days TEXT,
                man_days_value REAL,
                optimistic_50 TEXT,
                optimistic_50_value REAL,
                pessimistic_10 TEXT,
                pessimistic_10_value REAL,
                est_formula TEXT,
                est_formula_value REAL,
                tk_target TEXT,
                tk_target_value REAL,
                PRIMARY KEY (sheet_name, row_number)
            );
            CREATE TABLE epics (
                epic_key TEXT PRIMARY KEY,
                summary TEXT,
                issue_type TEXT,
                status TEXT,
                priority TEXT,
                jira_start_date TEXT,
                jira_due_date TEXT,
                jira_original_estimate TEXT,
                jira_original_estimate_seconds INTEGER,
                jira_aggregate_original_estimate_seconds INTEGER
            );
            CREATE TABLE stories (
                story_key TEXT PRIMARY KEY,
                epic_key TEXT NOT NULL,
                summary TEXT,
                issue_type TEXT,
                status TEXT,
                priority TEXT,
                jira_start_date TEXT,
                jira_due_date TEXT,
                jira_original_estimate TEXT,
                jira_original_estimate_seconds INTEGER,
                jira_aggregate_original_estimate_seconds INTEGER
            );
            CREATE TABLE story_descendants (
                issue_key TEXT PRIMARY KEY,
                parent_story_key TEXT NOT NULL,
                summary TEXT,
                issue_type TEXT,
                is_subtask INTEGER NOT NULL DEFAULT 0,
                status TEXT,
                priority TEXT,
                jira_start_date TEXT,
                jira_due_date TEXT,
                jira_original_estimate TEXT,
                jira_original_estimate_seconds INTEGER,
                jira_aggregate_original_estimate_seconds INTEGER
            );
            CREATE TABLE worklogs (
                worklog_id TEXT PRIMARY KEY,
                issue_key TEXT NOT NULL,
                author_display_name TEXT,
                started TEXT,
                time_spent TEXT,
                time_spent_seconds INTEGER
            );
            CREATE TABLE run_errors (
                error_scope TEXT NOT NULL,
                issue_key TEXT,
                sheet_name TEXT,
                row_number INTEGER,
                message TEXT NOT NULL
            );
            """
        )
        conn.execute(
            """
            INSERT INTO source_rmi_rows
            (sheet_name, row_number, roadmap_item, jira_id, man_days, man_days_value, optimistic_50, optimistic_50_value, pessimistic_10, pessimistic_10_value, est_formula, est_formula_value, tk_target, tk_target_value)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                "OmniConnect RMI",
                3,
                "Streaming",
                "O2-793",
                "=SUM(M3:U3)",
                84,
                "=V3 - (V3 * 0.5)",
                42,
                "=V3 + (W3 * 1.5)",
                105,
                "=(W3+(4*V3)+X3)/6",
                80.5,
                "=Y3/2",
                40.25,
            ),
        )
        conn.execute(
            """
            INSERT INTO epics
            (epic_key, summary, issue_type, status, priority, jira_start_date, jira_due_date, jira_original_estimate, jira_original_estimate_seconds, jira_aggregate_original_estimate_seconds)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            ("O2-793", "Real-time Tag Streaming", "Epic", "Resolved!", "Medium", "2025-12-23", "2026-02-06", "10w 2d", 1497600, 3009600),
        )
        conn.execute(
            """
            INSERT INTO stories
            (story_key, epic_key, summary, issue_type, status, priority, jira_start_date, jira_due_date, jira_original_estimate, jira_original_estimate_seconds, jira_aggregate_original_estimate_seconds)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            ("O2-788", "O2-793", "Research/URS", "Story", "Resolved!", "Medium", "2026-01-01", "2026-01-16", "3w 1d", 460800, 806400),
        )
        conn.execute(
            """
            INSERT INTO story_descendants
            (issue_key, parent_story_key, summary, issue_type, is_subtask, status, priority, jira_start_date, jira_due_date, jira_original_estimate, jira_original_estimate_seconds, jira_aggregate_original_estimate_seconds)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            ("O2-1314", "O2-788", "streaming RnD", "Sub-task", 1, "Resolved!", "Medium", "2026-01-06", "2026-01-07", "2d", 57600, 57600),
        )
        conn.execute(
            """
            INSERT INTO story_descendants
            (issue_key, parent_story_key, summary, issue_type, is_subtask, status, priority, jira_start_date, jira_due_date, jira_original_estimate, jira_original_estimate_seconds, jira_aggregate_original_estimate_seconds)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            ("O2-1315", "O2-788", "fix streaming defect", "Bug Subtask", 1, "Resolved!", "Medium", "2026-01-08", "2026-01-09", "1d", 28800, 28800),
        )
        conn.executemany(
            """
            INSERT INTO worklogs
            (worklog_id, issue_key, author_display_name, started, time_spent, time_spent_seconds)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            [
                ("44528", "O2-1314", "Mohsin Junaid", "2026-01-06T07:45:00.000+0500", "1d", 28800),
                ("44529", "O2-1314", "Mohsin Junaid", "2026-01-07T07:46:00.000+0500", "1d", 28800),
            ],
        )
        conn.execute(
            "INSERT INTO run_errors (error_scope, issue_key, sheet_name, row_number, message) VALUES (?, ?, ?, ?, ?)",
            ("worklog_lookup", "O2-2000", "OmniConnect RMI", 9, "rate limit"),
        )
        conn.commit()
        conn.close()

    def test_generate_html_report_writes_gantt_and_drilldown(self):
        with TemporaryDirectory() as temp_dir:
            db_path = Path(temp_dir) / "rmi.db"
            html_path = Path(temp_dir) / "rmi.html"
            self.create_db(db_path)
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "OmniConnect RMI"
            worksheet["D3"] = "Streaming"
            worksheet["E3"] = "O2-793"
            worksheet["E3"].hyperlink = "https://octopusdtlsupport.atlassian.net/browse/O2-793"
            workbook.save(Path(temp_dir) / report.DEFAULT_WORKBOOK_NAME)

            output = report.generate_html_report(db_path, html_path)

            self.assertEqual(output, html_path)
            html = html_path.read_text(encoding="utf-8")
            self.assertIn("RMI Jira Gantt Dashboard", html)
            self.assertIn("Total # of RMI Epics", html)
            self.assertIn(">Most Likely<", html)
            self.assertIn(">Optimistic<", html)
            self.assertIn(">Pessimistic<", html)
            self.assertIn(">Calculated<", html)
            self.assertIn(">TK Approved<", html)
            self.assertIn(">Idle Hours/Days<", html)
            self.assertIn(">Jira Original Estimate<", html)
            self.assertIn(">Story Estimates<", html)
            self.assertIn("All Products", html)
            self.assertIn("Digital Log", html)
            self.assertIn("Fintech Fuel", html)
            self.assertIn("OmniChat", html)
            self.assertIn("OmniConnect", html)
            self.assertIn('data-product-filter-card="true"', html)
            self.assertIn('data-product-summary="all"', html)
            self.assertIn('role="button" tabindex="0" aria-pressed="true" aria-label="Filter page by All Products"', html)
            self.assertIn("Total TK Approved", html)
            self.assertIn('class="product-summary-duration duration-value"', html)
            self.assertIn('class="metric-value duration-value" data-seconds="2419200.0" data-hours="672 h" data-days="84 d">672 h</span>', html)
            self.assertIn('class="product-summary-duration duration-value" data-seconds="1159200.0" data-hours="322 h" data-days="40 d">322 h</span>', html)
            self.assertNotIn('<div class="metric-label">Stories</div>', html)
            self.assertNotIn('<div class="metric-label">Descendants</div>', html)
            self.assertNotIn('<div class="metric-label">Worklogs</div>', html)
            self.assertNotIn('<div class="metric-label">Errors</div>', html)
            self.assertIn("Real-time Tag Streaming", html)
            self.assertIn("Research/URS", html)
            self.assertIn("streaming RnD", html)
            self.assertIn("Mohsin Junaid", html)
            self.assertIn("Run Errors", html)
            self.assertIn("<svg", html)
            self.assertIn("Duration unit selector", html)
            self.assertIn("Total Leaves", html)
            self.assertIn('id="capacity-leaves-label"', html)
            self.assertIn('data-label-hours="Total Leaves (Hours)"', html)
            self.assertIn('data-label-days="Total Leaves (Days)"', html)
            self.assertIn("Total Availability", html)
            self.assertIn('id="capacity-leaves"', html)
            self.assertIn('id="availability-value"', html)
            self.assertIn("Capacity minus leaves", html)
            self.assertIn("Remaining availability after TK Approved", html)
            self.assertLess(html.index("Capacity Calculator"), html.index("Total # of RMI Epics"))
            self.assertIn("Month Story Estimate Analysis", html)
            self.assertIn("Month scope estimate bar chart", html)
            self.assertIn("Excluded Epics", html)
            self.assertIn('id="tk-jira-only-enabled"', html)
            self.assertIn("Only Jira Populated Epics", html)
            self.assertEqual(html.count('class="tk-month-exclusion-col"'), 3)
            self.assertIn(".tk-month-exclusion-table-frame {", html)
            self.assertIn("height: 320px;", html)
            self.assertIn("overflow-y: auto;", html)
            self.assertIn(".tk-month-exclusion-table .tk-month-exclusion-col {", html)
            self.assertIn("width: 33.333%;", html)
            self.assertIn(".tk-month-exclusion-reason {", html)
            self.assertIn("Excluded Epic Details", html)
            self.assertIn("For epics started in", html)
            self.assertIn("For epics delivered in", html)
            self.assertIn("data-open-excluded-epic", html)
            self.assertIn('<option value="2026-02" selected>Feb 2026</option>', html)
            self.assertIn("const startKey = parseMonthKey(epic.start_date);", html)
            self.assertIn("const list = Array.isArray(config.epics) ? config.epics : [];", html)
            self.assertIn("function syncTkMonthUi()", html)
            self.assertIn("function currentTkApprovedSeconds()", html)
            self.assertIn("function updateIdleCapacityCard()", html)
            self.assertIn("const tkJiraOnlyToggleInput = document.getElementById('tk-jira-only-enabled');", html)
            self.assertIn("let tkJiraOnly = true;", html)
            self.assertIn("function currentExecutiveScopeLabel()", html)
            self.assertIn("function updateProductSummaryCards()", html)
            self.assertIn("return Boolean(epic && epic.jira_populated);", html)
            self.assertIn("const productSummaryCards = Array.from(document.querySelectorAll('.product-summary-card'));", html)
            self.assertEqual(html.count("const rowToggleButtons = Array.from(document.querySelectorAll('.row-toggle'));"), 1)
            self.assertIn("let activeProducts = new Set(['all']);", html)
            self.assertIn("function selectedProductKeys()", html)
            self.assertIn("function productInScope(product)", html)
            self.assertIn("function currentProductLabel()", html)
            self.assertIn("function epicMatchesMonthScope(epic, monthKey)", html)
            self.assertIn("function epicsInMonthScope(monthKey)", html)
            self.assertIn("openTkInScopeDrawer = function () {", html)
            self.assertIn("title: 'Epics contributing to TK Approved'", html)
            self.assertIn("const idleCapacityCard = document.querySelector('.metric-card[data-metric-key=\"idle_capacity_seconds\"]');", html)
            self.assertIn("const idleSeconds = availabilitySeconds - tkSeconds;", html)
            self.assertIn("tkMonthToggleInput.addEventListener(eventName, syncTkMonthUi);", html)
            self.assertIn("tkStartMonthToggleInput.addEventListener(eventName, syncTkMonthUi);", html)
            self.assertIn("tkJiraOnlyToggleInput.addEventListener(eventName, syncTkMonthUi);", html)
            self.assertIn("syncTkMonthUi();", html)
            self.assertIn("epic.reasons.join('\\n')", html)
            self.assertIn('data-month-analysis-slot="selected"', html)
            self.assertIn("productSummaryCards.forEach((card) => {", html)
            self.assertIn("const handleSelect = () => setProduct(card.dataset.productSummary || 'all');", html)
            self.assertIn("countByProduct", html)
            self.assertIn("RMIs/Epics • Total TK Approved", html)
            self.assertIn("records = records.filter((e) => e.jira_populated);", html)
            self.assertNotIn("TK Approved (in ", html)
            self.assertNotIn("TK Approved (outside", html)
            self.assertNotIn("tk-outside-month", html)
            self.assertIn("function openTkOutsideDrawer()", html)
            self.assertIn("Epic Search", html)
            self.assertIn("Filter by epic key, title, product, status, or priority", html)
            self.assertIn('data-search="O2-793 Real-time Tag Streaming Streaming OmniConnect Resolved! Medium"', html)
            self.assertIn('class="gantt-epic-row"', html)
            self.assertIn("data-unit=\"hours\"", html)
            self.assertIn("data-unit=\"days\"", html)
            self.assertIn("Gantt View", html)
            self.assertIn("Table View", html)
            self.assertIn("Hierarchical Table", html)
            self.assertIn("Table row color legend", html)
            self.assertIn(">Epics</span>", html)
            self.assertIn(">Stories</span>", html)
            self.assertIn(">Subtasks</span>", html)
            self.assertIn(">Bug Subtasks</span>", html)
            self.assertIn("Toggle", html)
            self.assertIn(">Jira</th>", html)
            self.assertNotIn("Epic Key", html)
            self.assertIn("class=\"hierarchy-table epic-table\"", html)
            self.assertIn("class=\"hierarchy-table epic-table story-table\"", html)
            self.assertIn("epic-summary-row epic-summary-row-epic", html)
            self.assertIn("story-summary-row story-summary-row-story", html)
            self.assertIn("descendant-summary-row descendant-summary-row-subtask", html)
            self.assertIn("descendant-summary-row descendant-summary-row-bug-subtask", html)
            self.assertIn("id=\"gantt-view\" hidden", html)
            self.assertIn("id=\"table-view\"", html)
            self.assertIn("jira-link-button", html)
            self.assertIn("https://octopusdtlsupport.atlassian.net/browse/O2-793", html)
            self.assertIn("sticky-title", html)
            self.assertIn("sticky-epic", html)
            self.assertIn("summary-condition", html)
            self.assertIn("TK Approved", html)
            self.assertIn("Jira Original Estimate", html)
            self.assertNotIn("<strong>Original Estimate</strong>", html)
            self.assertIn("Story Estimates", html)
            self.assertIn("Mismatch:", html)
            self.assertIn("below Jira original estimate", html)
            self.assertIn("summary-condition under", html)
            self.assertIn('data-hours="672.00 h"', html)
            self.assertIn('data-days="84.00 d"', html)
            self.assertIn('data-hours="322.00 h"', html)
            self.assertIn('"OmniConnect": {"epic_count": 1.0', html)
            self.assertIn('"most_likely_seconds": 2419200.0', html)
            self.assertIn('"story_estimate_seconds": 460800.0', html)
            self.assertIn('"stories": [{"story_key": "O2-788"', html)
            self.assertIn('"subtasks": [{"issue_key": "O2-1314"', html)
            self.assertNotIn("=SUM(M3:U3)", html)
            self.assertNotIn("=Y3/2", html)
            self.assertIn("Bars are drawn only for epics where both Jira start date and Jira due date are available", html)

    def test_epic_search_text_includes_key_title_and_status_fields(self):
        epic = {
            "jira_id": "O2-793",
            "epic_summary": "Real-time Tag Streaming",
            "roadmap_item": "Streaming",
            "product": "OmniConnect",
            "epic_status": "Resolved!",
            "epic_priority": "Medium",
        }

        self.assertEqual(
            report.epic_search_text(epic),
            "O2-793 Real-time Tag Streaming Streaming OmniConnect Resolved! Medium",
        )

    def test_build_summary_counts_nested_records(self):
        source_rows = [
            {
                "stories": [
                    {"descendants": [{"worklogs": [{"time_spent_seconds": 7200}, {"time_spent_seconds": 3600}]}]},
                    {"descendants": []},
                ]
            }
        ]
        summary = report.build_summary(source_rows, [{"message": "x"}])

        self.assertEqual(summary["epic_count"], 1)
        self.assertEqual(summary["story_count"], 2)
        self.assertEqual(summary["descendant_count"], 1)
        self.assertEqual(summary["worklog_count"], 2)
        self.assertEqual(summary["total_worklog_seconds"], 10800)
        self.assertEqual(summary["error_count"], 1)

    def test_build_epic_metric_summary_rolls_up_by_product(self):
        source_rows = [
            {
                "product": "OmniConnect",
                "man_days_value": 10,
                "optimistic_50_value": 8,
                "pessimistic_10_value": 12,
                "est_formula_value": 9,
                "tk_target_value": 7,
                "epic_original_estimate_seconds": 36000,
                "stories": [{"jira_original_estimate_seconds": 7200, "jira_aggregate_original_estimate_seconds": 7200}],
            },
            {
                "product": "Fintech Fuel",
                "man_days_value": 5,
                "optimistic_50_value": 4,
                "pessimistic_10_value": 6,
                "est_formula_value": 4.5,
                "tk_target_value": 3.5,
                "epic_original_estimate_seconds": 18000,
                "stories": [{"jira_original_estimate_seconds": 3600, "jira_aggregate_original_estimate_seconds": 3600}],
            },
        ]

        summary = report.build_epic_metric_summary(source_rows)

        self.assertEqual(summary["all"]["epic_count"], 2)
        self.assertEqual(summary["OmniConnect"]["epic_count"], 1)
        self.assertEqual(summary["Fintech Fuel"]["epic_count"], 1)
        self.assertEqual(summary["all"]["most_likely_seconds"], 15 * 28800)
        self.assertEqual(summary["all"]["optimistic_seconds"], 12 * 28800)
        self.assertEqual(summary["all"]["pessimistic_seconds"], 18 * 28800)
        self.assertEqual(summary["all"]["calculated_seconds"], 13.5 * 28800)
        self.assertEqual(summary["all"]["tk_approved_seconds"], 10.5 * 28800)
        self.assertEqual(summary["all"]["jira_original_estimate_seconds"], 54000)
        self.assertEqual(summary["all"]["story_estimate_seconds"], 10800)
        self.assertEqual(summary["all"]["tk_approved_seconds"], 10.5 * 28800)

    def test_build_epic_detail_records_include_nested_story_and_subtask_details(self):
        source_rows = [
            {
                "jira_id": "O2-793",
                "epic_summary": "Real-time Tag Streaming",
                "roadmap_item": "Streaming",
                "product": "OmniConnect",
                "man_days_value": 84,
                "optimistic_50_value": 42,
                "pessimistic_10_value": 105,
                "est_formula_value": 80.5,
                "epic_status": "Resolved!",
                "epic_priority": "Medium",
                "epic_start_date": "2025-12-23",
                "epic_due_date": "2026-02-06",
                "tk_target_value": 40.25,
                "epic_original_estimate_seconds": 1497600,
                "stories": [
                    {
                        "story_key": "O2-788",
                        "summary": "Research/URS",
                        "status": "Resolved!",
                        "priority": "Medium",
                        "jira_start_date": "2026-01-01",
                        "jira_due_date": "2026-01-16",
                        "jira_original_estimate_seconds": 460800,
                        "descendants": [
                            {
                                "issue_key": "O2-1314",
                                "summary": "streaming RnD",
                                "is_subtask": 1,
                                "status": "Resolved!",
                                "priority": "Medium",
                                "jira_start_date": "2026-01-06",
                                "jira_due_date": "2026-01-07",
                                "jira_original_estimate_seconds": 57600,
                            },
                            {
                                "issue_key": "O2-1400",
                                "summary": "Bug",
                                "is_subtask": 0,
                                "jira_start_date": "2026-01-08",
                                "jira_due_date": "2026-01-09",
                                "jira_original_estimate_seconds": 28800,
                            },
                        ],
                    }
                ],
            }
        ]

        details = report.build_epic_detail_records(source_rows)

        self.assertEqual(len(details), 1)
        self.assertEqual(details[0]["jira_id"], "O2-793")
        self.assertTrue(details[0]["jira_populated"])
        self.assertEqual(details[0]["most_likely_seconds"], 84 * 28800)
        self.assertEqual(details[0]["optimistic_seconds"], 42 * 28800)
        self.assertEqual(details[0]["pessimistic_seconds"], 105 * 28800)
        self.assertEqual(details[0]["calculated_seconds"], 80.5 * 28800)
        self.assertEqual(len(details[0]["stories"]), 1)
        self.assertEqual(details[0]["stories"][0]["story_key"], "O2-788")
        self.assertEqual(details[0]["stories"][0]["estimate_seconds"], 460800)
        self.assertEqual(details[0]["stories"][0]["status"], "Resolved!")
        self.assertEqual(details[0]["stories"][0]["priority"], "Medium")
        self.assertEqual(len(details[0]["stories"][0]["subtasks"]), 1)
        self.assertEqual(details[0]["stories"][0]["subtasks"][0]["issue_key"], "O2-1314")
        self.assertEqual(details[0]["stories"][0]["subtasks"][0]["estimate_seconds"], 57600)
        self.assertEqual(details[0]["stories"][0]["subtasks"][0]["status"], "Resolved!")
        self.assertEqual(details[0]["stories"][0]["subtasks"][0]["priority"], "Medium")

    def test_build_rmi_schedule_records_returns_sorted_records_with_stories(self):
        """build_rmi_schedule_records returns one record per epic, sorted by product then start_date."""
        source_rows = [
            {
                "roadmap_item": "Streaming",
                "jira_id": "O2-793",
                "product": "OmniConnect",
                "epic_summary": "Real-time Tag Streaming",
                "epic_status": "Resolved!",
                "epic_start_date": "2025-12-23",
                "epic_due_date": "2026-02-06",
                "man_days_value": 84,
                "tk_target_value": 40.25,
                "stories": [
                    {
                        "jira_start_date": "2026-01-01",
                        "jira_due_date": "2026-01-16",
                        "jira_original_estimate_seconds": 460800,
                        "descendants": [
                            {
                                "is_subtask": 1,
                                "jira_start_date": "2026-01-06",
                                "jira_due_date": "2026-01-07",
                                "jira_original_estimate_seconds": 57600,
                            },
                        ],
                    }
                ],
            },
            {
                "roadmap_item": "Payments",
                "jira_id": "FF-100",
                "product": "Fintech Fuel",
                "epic_summary": "Payments Module",
                "epic_status": "In Progress",
                "epic_start_date": "2026-03-01",
                "epic_due_date": "2026-05-15",
                "man_days_value": 50,
                "tk_target_value": 30,
                "stories": [],
            },
        ]

        records = report.build_rmi_schedule_records(source_rows)

        self.assertEqual(len(records), 2)
        # Fintech Fuel sorts before OmniConnect
        self.assertEqual(records[0]["product"], "Fintech Fuel")
        self.assertEqual(records[0]["roadmap_item"], "Payments")
        self.assertTrue(records[0]["jira_populated"])
        self.assertEqual(records[0]["most_likely_days"], 50)
        self.assertEqual(records[0]["tk_approved_days"], 30)
        self.assertEqual(records[0]["stories"], [])
        # OmniConnect second
        self.assertEqual(records[1]["product"], "OmniConnect")
        self.assertEqual(records[1]["jira_id"], "O2-793")
        self.assertTrue(records[1]["jira_populated"])
        self.assertEqual(records[1]["most_likely_days"], 84)
        self.assertEqual(records[1]["tk_approved_days"], 40.25)
        self.assertEqual(len(records[1]["stories"]), 1)
        self.assertEqual(records[1]["stories"][0]["estimate_seconds"], 460800)
        self.assertEqual(len(records[1]["stories"][0]["subtasks"]), 1)
        self.assertEqual(records[1]["stories"][0]["subtasks"][0]["estimate_seconds"], 57600)

    def test_epic_has_jira_population_requires_joined_jira_data(self):
        self.assertTrue(
            report.epic_has_jira_population(
                {
                    "jira_id": "O2-793",
                    "epic_summary": "Real-time Tag Streaming",
                    "stories": [],
                }
            )
        )
        self.assertTrue(
            report.epic_has_jira_population(
                {
                    "jira_id": "O2-794",
                    "epic_original_estimate_seconds": 3600,
                    "stories": [],
                }
            )
        )
        self.assertFalse(
            report.epic_has_jira_population(
                {
                    "jira_id": "O2-999",
                    "epic_summary": "",
                    "epic_status": "",
                    "epic_start_date": "",
                    "epic_due_date": "",
                    "epic_original_estimate_seconds": 0,
                    "stories": [],
                }
            )
        )

    def test_available_years_from_schedule_records_includes_current_year(self):
        """available_years_from_schedule_records always includes the current year."""
        records = [
            {
                "start_date": "2025-06-01",
                "due_date": "2025-12-31",
                "stories": [
                    {"start_date": "2025-07-01", "due_date": "2025-08-01", "subtasks": []},
                ],
            }
        ]
        years = report.available_years_from_schedule_records(records)
        from datetime import date as _date

        self.assertIn(_date.today().year, years)
        self.assertIn(2025, years)
        self.assertEqual(years, sorted(years))

    def test_initial_schedule_year_prefers_latest_observed_year_when_current_missing(self):
        """initial_schedule_year avoids empty first paint when the current year has no schedule data."""
        current_year = report.date.today().year
        prior_year = current_year - 1
        records = [
            {
                "start_date": f"{prior_year}-01-10",
                "due_date": f"{prior_year}-03-20",
                "stories": [
                    {"start_date": f"{prior_year}-02-01", "due_date": f"{prior_year}-02-10", "subtasks": []},
                ],
            }
        ]

        selected_year = report.initial_schedule_year(records)

        self.assertEqual(selected_year, prior_year)

    def test_render_rmi_schedule_table_rows_prepopulates_body_and_footer(self):
        """render_rmi_schedule_table_rows returns visible initial tbody and tfoot HTML."""
        records = [
            {
                "roadmap_item": "Streaming",
                "jira_id": "O2-793",
                "jira_url": "/browse/O2-793",
                "product": "OmniConnect",
                "status": "Resolved!",
                "most_likely_days": 84,
                "tk_approved_days": 40.25,
                "stories": [
                    {
                        "start_date": "2026-01-01",
                        "due_date": "2026-01-16",
                        "estimate_seconds": 460800,
                        "subtasks": [
                            {
                                "start_date": "2026-01-06",
                                "due_date": "2026-01-07",
                                "estimate_seconds": 57600,
                            }
                        ],
                    }
                ],
            }
        ]

        body_html, foot_html = report.render_rmi_schedule_table_rows(records, 2026, unit="hours")

        self.assertIn('class="rmi-sched-product-group"', body_html)
        self.assertIn('class="rmi-sched-epic-row"', body_html)
        self.assertIn('Streaming', body_html)
        self.assertIn('/browse/O2-793', body_html)
        self.assertIn('>672<', body_html)
        self.assertIn('>322<', body_html)
        self.assertIn('>128<', foot_html)
        self.assertIn('Grand Total', foot_html)

    def test_render_rmi_schedule_table_rows_totals_match_visible_row_rounding(self):
        """Subtotal and grand total cells should add the displayed epic values, not raw decimals."""
        records = [
            {
                "roadmap_item": "Epic A",
                "jira_id": "O2-901",
                "jira_url": "",
                "product": "OmniConnect",
                "status": "Open",
                "most_likely_days": 10.3125,
                "tk_approved_days": 5.3125,
                "stories": [],
            },
            {
                "roadmap_item": "Epic B",
                "jira_id": "O2-902",
                "jira_url": "",
                "product": "OmniConnect",
                "status": "Open",
                "most_likely_days": 10.3125,
                "tk_approved_days": 5.3125,
                "stories": [],
            },
        ]

        body_html, foot_html = report.render_rmi_schedule_table_rows(records, 2026, unit="hours")

        self.assertIn(">83<", body_html)
        self.assertIn(">43<", body_html)
        self.assertIn(">166<", body_html)
        self.assertIn(">86<", body_html)
        self.assertIn(">166<", foot_html)
        self.assertIn(">86<", foot_html)

    def test_load_report_data_reads_epic_jira_url_from_neighbor_workbook(self):
        """Epic links come from the workbook Jira ID hyperlink, not just the Jira key."""
        with TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            db_path = temp_path / "rmi.db"
            self.create_db(db_path)

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "OmniConnect RMI"
            worksheet["D3"] = "Streaming"
            worksheet["E3"] = "O2-793"
            worksheet["E3"].hyperlink = "https://octopusdtlsupport.atlassian.net/browse/O2-793"
            workbook.save(temp_path / report.DEFAULT_WORKBOOK_NAME)

            report_data = report.load_report_data(db_path)

            self.assertEqual(
                report_data["source_rows"][0]["jira_url"],
                "https://octopusdtlsupport.atlassian.net/browse/O2-793",
            )

    def test_load_report_data_ignores_locked_workbook(self):
        """A locked workbook should not crash HTML generation."""
        with TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            db_path = temp_path / "rmi.db"
            self.create_db(db_path)
            workbook_path = temp_path / report.DEFAULT_WORKBOOK_NAME
            workbook_path.write_bytes(b"placeholder")

            with patch.object(report, "load_workbook", side_effect=PermissionError("locked")):
                report_data = report.load_report_data(db_path)

            self.assertEqual(report_data["source_rows"][0]["jira_url"], "")

    def test_render_epic_table_view_hides_epic_jira_link_without_workbook_url(self):
        """Epic-level Jira buttons stay disabled when no workbook URL is available."""
        html = report.render_epic_table_view(
            [
                {
                    "product": "OmniConnect",
                    "jira_id": "O2-793",
                    "jira_url": "",
                    "roadmap_item": "Streaming",
                    "epic_summary": "Real-time Tag Streaming",
                    "epic_status": "Resolved!",
                    "epic_priority": "Medium",
                    "epic_start_date": "2025-12-23",
                    "epic_due_date": "2026-02-06",
                    "man_days_value": 84,
                    "optimistic_50_value": 42,
                    "pessimistic_10_value": 105,
                    "est_formula_value": 80.5,
                    "tk_target_value": 40.25,
                    "epic_original_estimate_seconds": 1497600,
                    "stories": [],
                }
            ]
        )

        self.assertIn('title="Jira link unavailable"', html)
        self.assertNotIn('href="/browse/O2-793"', html)


if __name__ == "__main__":
    unittest.main()
