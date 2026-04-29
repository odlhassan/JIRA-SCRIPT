import sqlite3
import unittest
from pathlib import Path
from tempfile import TemporaryDirectory
from unittest.mock import Mock, patch

from openpyxl import Workbook

import extract_rmi_jira_to_sqlite as extract


class ExtractRmiJiraToSqliteTests(unittest.TestCase):
    def create_workbook(self, path: Path) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Omni RMI"
        ws["D3"] = "Roadmap Item A"
        ws["E3"] = "EPIC-1"
        ws["V3"] = 12
        ws["W3"] = 8
        ws["X3"] = 16
        ws["Y3"] = "=W3+X3"
        ws["Z3"] = 10
        ws["D4"] = "Roadmap Item B"
        ws["V4"] = 3
        ws["E4"] = None

        other = wb.create_sheet("Backlog")
        other["D3"] = "Should Be Ignored"
        other["E3"] = "EPIC-IGNORE"
        wb.save(path)

    def test_extract_source_rmi_rows_keeps_workbook_rows_without_jira_ids(self):
        with TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "source.xlsx"
            self.create_workbook(workbook_path)

            rows = extract.extract_source_rmi_rows(workbook_path)

            self.assertEqual(len(rows), 2)
            row = rows[0]
            self.assertEqual(row.sheet_name, "Omni RMI")
            self.assertEqual(row.row_number, 3)
            self.assertEqual(row.roadmap_item, "Roadmap Item A")
            self.assertEqual(row.jira_id, "EPIC-1")
            self.assertEqual(row.man_days, 12)
            self.assertEqual(row.man_days_value, 12)
            self.assertEqual(row.optimistic_50, 8)
            self.assertEqual(row.optimistic_50_value, 8)
            self.assertEqual(row.pessimistic_10, 16)
            self.assertEqual(row.pessimistic_10_value, 16)
            self.assertEqual(row.est_formula, "=W3+X3")
            self.assertEqual(row.est_formula_value, None)
            self.assertEqual(row.tk_target, 10)
            self.assertEqual(row.tk_target_value, 10)
            blank_jira_row = rows[1]
            self.assertEqual(blank_jira_row.row_number, 4)
            self.assertEqual(blank_jira_row.roadmap_item, "Roadmap Item B")
            self.assertEqual(blank_jira_row.jira_id, "")
            self.assertEqual(blank_jira_row.man_days, 3)
            self.assertEqual(blank_jira_row.man_days_value, 3)

    def test_normalize_jira_id_extracts_key_from_browse_url(self):
        normalized = extract.normalize_jira_id(
            "https://octopusdtlsupport.atlassian.net/browse/O2-793"
        )

        self.assertEqual(normalized, "O2-793")

    def test_map_issue_maps_estimates_dates_and_parent(self):
        issue = {
            "key": "EPIC-1",
            "fields": {
                "summary": "Epic summary",
                "issuetype": {"name": "Epic", "subtask": False},
                "status": {"name": "To Do"},
                "priority": {"name": "High"},
                "customfield_10015": "2026-04-20",
                "duedate": "2026-04-30",
                "timetracking": {"originalEstimate": "5d", "originalEstimateSeconds": 144000},
                "timeoriginalestimate": 144000,
                "aggregatetimeoriginalestimate": 288000,
                "parent": {"key": "PARENT-1", "fields": {"summary": "Parent summary"}},
            },
        }

        mapped = extract.JiraAPIClient.map_issue(issue)

        self.assertEqual(mapped["key"], "EPIC-1")
        self.assertEqual(mapped["jira_start_date"], "2026-04-20")
        self.assertEqual(mapped["jira_due_date"], "2026-04-30")
        self.assertEqual(mapped["jira_original_estimate"], "5d")
        self.assertEqual(mapped["jira_original_estimate_seconds"], 144000)
        self.assertEqual(mapped["jira_aggregate_original_estimate_seconds"], 288000)
        self.assertEqual(mapped["parent_key"], "PARENT-1")

    def test_sqlite_writer_reset_is_idempotent(self):
        with TemporaryDirectory() as temp_dir:
            db_path = Path(temp_dir) / "out.db"
            writer = extract.SQLiteWriter(db_path)
            writer.setup_schema()
            writer.reset_tables()
            writer.insert_source_row(
                extract.SourceRmiRow(
                    "Omni RMI",
                    3,
                    "Roadmap Item",
                    "EPIC-1",
                    1,
                    1,
                    2,
                    2,
                    3,
                    3,
                    "f",
                    4,
                    4,
                    5,
                )
            )
            writer.insert_error("epic_lookup", "Missing", issue_key="EPIC-404")
            writer.commit()
            self.assertEqual(writer.count_rows("source_rmi_rows"), 1)
            self.assertEqual(writer.count_rows("run_errors"), 1)

            writer.reset_tables()
            writer.commit()
            self.assertEqual(writer.count_rows("source_rmi_rows"), 0)
            self.assertEqual(writer.count_rows("run_errors"), 0)
            writer.close()

    def test_run_extraction_persists_epics_stories_descendants_and_worklogs(self):
        with TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "source.xlsx"
            db_path = Path(temp_dir) / "out.db"
            self.create_workbook(workbook_path)
            progress_messages: list[str] = []

            client = Mock()
            client.test_connection.return_value = True
            client.get_issue.side_effect = [
                {
                    "key": "EPIC-1",
                    "summary": "Epic summary",
                    "issue_type": "Epic",
                    "status": "To Do",
                    "priority": "High",
                    "jira_start_date": "2026-04-20",
                    "jira_due_date": "2026-04-30",
                    "jira_original_estimate": "5d",
                    "jira_original_estimate_seconds": 144000,
                    "jira_aggregate_original_estimate_seconds": 288000,
                }
            ]
            client.get_child_stories.return_value = [
                {
                    "key": "STORY-1",
                    "summary": "Story summary",
                    "issue_type": "Story",
                    "status": "In Progress",
                    "priority": "Medium",
                    "jira_start_date": "2026-04-21",
                    "jira_due_date": "2026-04-25",
                    "jira_original_estimate": "2d",
                    "jira_original_estimate_seconds": 57600,
                    "jira_aggregate_original_estimate_seconds": 86400,
                }
            ]
            client.get_story_children.return_value = [
                {
                    "key": "SUB-1",
                    "summary": "Subtask summary",
                    "issue_type": "Sub-task",
                    "is_subtask": True,
                    "status": "Done",
                    "priority": "Low",
                    "jira_start_date": "2026-04-22",
                    "jira_due_date": "2026-04-23",
                    "jira_original_estimate": "4h",
                    "jira_original_estimate_seconds": 14400,
                    "jira_aggregate_original_estimate_seconds": 14400,
                }
            ]
            client.get_issue_worklogs.return_value = [
                {
                    "id": "10001",
                    "author": {"displayName": "Engineer 1"},
                    "started": "2026-04-22T09:00:00.000+0500",
                    "timeSpent": "2h",
                    "timeSpentSeconds": 7200,
                },
                {
                    "id": "10002",
                    "author": {"displayName": "Engineer 2"},
                    "started": "2026-04-22T12:00:00.000+0500",
                    "timeSpent": "1h",
                    "timeSpentSeconds": 3600,
                },
            ]

            with patch.object(extract, "JiraAPIClient", return_value=client):
                summary = extract.run_extraction(
                    workbook_path=workbook_path,
                    db_path=db_path,
                    env_config={
                        "JIRA_SITE": "example",
                        "JIRA_EMAIL": "user@example.com",
                        "JIRA_API_TOKEN": "token",
                    },
                    progress_callback=progress_messages.append,
                )

            self.assertEqual(summary["eligible_rows"], 2)
            self.assertEqual(summary["epics_fetched"], 1)
            self.assertEqual(summary["stories_fetched"], 1)
            self.assertEqual(summary["descendants_fetched"], 1)
            self.assertEqual(summary["worklogs_fetched"], 2)
            self.assertEqual(summary["errors"], 0)
            self.assertTrue(any("Scanning workbook rows" in message for message in progress_messages))
            self.assertTrue(any("Workbook scan complete: 2 eligible rows found" in message for message in progress_messages))
            self.assertTrue(any("Processing [1/2] Omni RMI row 3: EPIC-1 - Roadmap Item A" in message for message in progress_messages))
            self.assertTrue(any("Processing [2/2] Omni RMI row 4: No Jira ID - Roadmap Item B" in message for message in progress_messages))
            self.assertTrue(any("No Jira ID in workbook; stored source row only for Roadmap Item B" in message for message in progress_messages))
            self.assertTrue(any("Found 1 child stories for EPIC-1" in message for message in progress_messages))
            self.assertTrue(any("Stored 2 worklogs for SUB-1" in message for message in progress_messages))
            self.assertTrue(any("Extraction complete:" in message for message in progress_messages))
            client.get_issue.assert_called_once_with("EPIC-1")

            connection = sqlite3.connect(db_path)
            self.assertEqual(connection.execute("SELECT COUNT(*) FROM epics").fetchone()[0], 1)
            self.assertEqual(connection.execute("SELECT COUNT(*) FROM source_rmi_rows").fetchone()[0], 2)
            self.assertEqual(connection.execute("SELECT COUNT(*) FROM stories").fetchone()[0], 1)
            self.assertEqual(connection.execute("SELECT COUNT(*) FROM story_descendants").fetchone()[0], 1)
            self.assertEqual(connection.execute("SELECT COUNT(*) FROM worklogs").fetchone()[0], 2)
            self.assertEqual(
                connection.execute("SELECT jira_id FROM source_rmi_rows WHERE row_number = 3").fetchone()[0],
                "EPIC-1",
            )
            self.assertEqual(
                connection.execute("SELECT jira_id FROM source_rmi_rows WHERE row_number = 4").fetchone()[0],
                "",
            )
            connection.close()

    def test_run_extraction_records_invalid_jira_lookup_in_run_errors(self):
        with TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "source.xlsx"
            db_path = Path(temp_dir) / "out.db"
            self.create_workbook(workbook_path)

            client = Mock()
            client.test_connection.return_value = True
            response = Mock()
            response.status_code = 404
            error = extract.requests.exceptions.HTTPError("404 Client Error")
            error.response = response
            client.get_issue.side_effect = error

            with patch.object(extract, "JiraAPIClient", return_value=client):
                summary = extract.run_extraction(
                    workbook_path=workbook_path,
                    db_path=db_path,
                    env_config={
                        "JIRA_SITE": "example",
                        "JIRA_EMAIL": "user@example.com",
                        "JIRA_API_TOKEN": "token",
                    },
                )

            self.assertEqual(summary["epics_fetched"], 0)
            self.assertEqual(summary["errors"], 1)
            connection = sqlite3.connect(db_path)
            row = connection.execute(
                "SELECT error_scope, issue_key, sheet_name, row_number, message FROM run_errors"
            ).fetchone()
            self.assertEqual(row[0], "epic_lookup")
            self.assertEqual(row[1], "EPIC-1")
            self.assertEqual(row[2], "Omni RMI")
            self.assertEqual(row[3], 3)
            self.assertIn("404", row[4])
            connection.close()


if __name__ == "__main__":
    unittest.main()
