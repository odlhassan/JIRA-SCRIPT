from __future__ import annotations

import argparse
import base64
import re
import sqlite3
from dataclasses import dataclass
from pathlib import Path
from typing import Callable, Dict, Iterable, List, Optional

import requests
from openpyxl import load_workbook


BASE_DIR = Path(__file__).parent
REPORTS_DIR = BASE_DIR / "IPP Meeting Reports"
DEFAULT_WORKBOOK_PATH = REPORTS_DIR / "Epic Estimates Approved Plan.xlsx"
DEFAULT_DB_PATH = REPORTS_DIR / "rmi_jira_extract.db"
ENV_CANDIDATES = [
    BASE_DIR / ".env",
    BASE_DIR / "Documentation" / ".env",
]

SOURCE_COLUMNS = {
    "roadmap_item": "D",
    "jira_id": "E",
    "man_days": "V",
    "optimistic_50": "W",
    "pessimistic_10": "X",
    "est_formula": "Y",
    "tk_target": "Z",
}
JIRA_START_DATE_FIELD = "customfield_10015"
DEFAULT_SHEET_FILTER = "RMI"
SCAN_START_ROW = 3


def load_env_config(required_keys: Optional[Iterable[str]] = None) -> Dict[str, str]:
    env_path = next((path for path in ENV_CANDIDATES if path.exists()), None)
    if env_path is None:
        raise FileNotFoundError("Could not find a .env file with Jira credentials.")

    config: Dict[str, str] = {}
    for raw_line in env_path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        config[key.strip()] = value.strip()

    required = list(required_keys or ["JIRA_SITE", "JIRA_EMAIL", "JIRA_API_TOKEN"])
    missing = [key for key in required if not config.get(key)]
    if missing:
        raise ValueError(f"Missing Jira settings in {env_path}: {', '.join(missing)}")

    config["ENV_PATH"] = str(env_path)
    return config


@dataclass
class SourceRmiRow:
    sheet_name: str
    row_number: int
    roadmap_item: str
    jira_id: str
    man_days: object
    man_days_value: object
    optimistic_50: object
    optimistic_50_value: object
    pessimistic_10: object
    pessimistic_10_value: object
    est_formula: object
    est_formula_value: object
    tk_target: object
    tk_target_value: object


class JiraAPIClient:
    issue_fields = [
        "summary",
        "issuetype",
        "status",
        "priority",
        JIRA_START_DATE_FIELD,
        "duedate",
        "timetracking",
        "timeoriginalestimate",
        "aggregatetimeoriginalestimate",
        "parent",
        "subtasks",
    ]

    def __init__(self, site: str, email: str, token: str):
        self.site = site
        self.email = email
        self.token = token
        self.base_url_v3 = f"https://{site}.atlassian.net/rest/api/3"
        self.search_url = f"{self.base_url_v3}/search/jql"
        self.session = requests.Session()

        auth_str = f"{email}:{token}"
        auth_b64 = base64.b64encode(auth_str.encode("utf-8")).decode("utf-8")
        self.session.headers.update(
            {
                "Authorization": f"Basic {auth_b64}",
                "Accept": "application/json",
                "Content-Type": "application/json",
            }
        )

    @staticmethod
    def map_issue(issue: Dict) -> Dict:
        fields = issue.get("fields") or {}
        timetracking = fields.get("timetracking") or {}
        parent = fields.get("parent") or {}
        parent_fields = parent.get("fields") or {}
        return {
            "key": issue.get("key"),
            "summary": fields.get("summary"),
            "issue_type": ((fields.get("issuetype") or {}).get("name")),
            "is_subtask": ((fields.get("issuetype") or {}).get("subtask")) or False,
            "status": ((fields.get("status") or {}).get("name")),
            "priority": ((fields.get("priority") or {}).get("name")),
            "jira_start_date": fields.get(JIRA_START_DATE_FIELD),
            "jira_due_date": fields.get("duedate"),
            "jira_original_estimate": timetracking.get("originalEstimate"),
            "jira_original_estimate_seconds": timetracking.get("originalEstimateSeconds")
            or fields.get("timeoriginalestimate"),
            "jira_aggregate_original_estimate_seconds": fields.get("aggregatetimeoriginalestimate"),
            "parent_key": parent.get("key"),
            "parent_summary": parent_fields.get("summary"),
        }

    def test_connection(self) -> bool:
        try:
            response = self.session.get(f"{self.base_url_v3}/myself", timeout=15)
            response.raise_for_status()
            return True
        except requests.exceptions.RequestException:
            return False

    def get_issue(self, issue_key: str) -> Dict:
        response = self.session.get(
            f"{self.base_url_v3}/issue/{issue_key}",
            params={"fields": ",".join(self.issue_fields)},
            timeout=20,
        )
        response.raise_for_status()
        return self.map_issue(response.json())

    def _search_issues(self, jql: str, fields: Optional[List[str]] = None, max_results: int = 100) -> List[Dict]:
        response = self.session.post(
            self.search_url,
            json={
                "jql": jql,
                "maxResults": max_results,
                "fields": fields or self.issue_fields,
            },
            timeout=30,
        )
        response.raise_for_status()
        return [self.map_issue(issue) for issue in response.json().get("issues", [])]

    def get_child_stories(self, epic_key: str, max_results: int = 100) -> List[Dict]:
        queries = [
            f'"Epic Link" = {epic_key} ORDER BY Rank ASC',
            f'parent = {epic_key} ORDER BY Rank ASC',
            f'parentEpic = {epic_key} AND issuekey != {epic_key} ORDER BY Rank ASC',
        ]
        for jql in queries:
            stories = self._search_issues(jql, max_results=max_results)
            if stories:
                return stories
        return []

    def get_story_children(self, story_key: str, max_results: int = 100) -> List[Dict]:
        return self._search_issues(f"parent = {story_key} ORDER BY Rank ASC", max_results=max_results)

    def get_issue_worklogs(self, issue_key: str) -> List[Dict]:
        start_at = 0
        max_results = 100
        worklogs: List[Dict] = []

        while True:
            response = self.session.get(
                f"{self.base_url_v3}/issue/{issue_key}/worklog",
                params={"startAt": start_at, "maxResults": max_results},
                timeout=30,
            )
            response.raise_for_status()
            payload = response.json()
            chunk = payload.get("worklogs", [])
            worklogs.extend(chunk)
            start_at += len(chunk)
            total = payload.get("total", len(worklogs))
            if start_at >= total or not chunk:
                return worklogs


def get_rmi_sheet_names(workbook, sheet_contains: str) -> List[str]:
    return [sheet_name for sheet_name in workbook.sheetnames if sheet_contains in sheet_name]


def normalize_jira_id(value: object) -> str:
    text = str(value or "").strip()
    if not text:
        return ""

    browse_match = re.search(r"/browse/([A-Z][A-Z0-9_]*-\d+)", text, flags=re.IGNORECASE)
    if browse_match:
        return browse_match.group(1).upper()

    key_match = re.search(r"\b([A-Z][A-Z0-9_]*-\d+)\b", text, flags=re.IGNORECASE)
    if key_match:
        return key_match.group(1).upper()

    return text


def extract_source_rmi_rows(
    workbook_path: Path,
    sheet_contains: str = DEFAULT_SHEET_FILTER,
    limit: Optional[int] = None,
) -> List[SourceRmiRow]:
    workbook = load_workbook(workbook_path, data_only=False, read_only=True)
    workbook_values = load_workbook(workbook_path, data_only=True, read_only=True)
    extracted_rows: List[SourceRmiRow] = []

    try:
        for sheet_name in get_rmi_sheet_names(workbook, sheet_contains):
            worksheet = workbook[sheet_name]
            worksheet_values = workbook_values[sheet_name]
            for row_number in range(SCAN_START_ROW, worksheet.max_row + 1):
                roadmap_item = worksheet[f"{SOURCE_COLUMNS['roadmap_item']}{row_number}"].value
                if not isinstance(roadmap_item, str) or not roadmap_item.strip():
                    continue
                jira_id_value = normalize_jira_id(worksheet[f"{SOURCE_COLUMNS['jira_id']}{row_number}"].value)

                extracted_rows.append(
                    SourceRmiRow(
                        sheet_name=sheet_name,
                        row_number=row_number,
                        roadmap_item=roadmap_item.strip(),
                        jira_id=jira_id_value,
                        man_days=worksheet[f"{SOURCE_COLUMNS['man_days']}{row_number}"].value,
                        man_days_value=worksheet_values[f"{SOURCE_COLUMNS['man_days']}{row_number}"].value,
                        optimistic_50=worksheet[f"{SOURCE_COLUMNS['optimistic_50']}{row_number}"].value,
                        optimistic_50_value=worksheet_values[f"{SOURCE_COLUMNS['optimistic_50']}{row_number}"].value,
                        pessimistic_10=worksheet[f"{SOURCE_COLUMNS['pessimistic_10']}{row_number}"].value,
                        pessimistic_10_value=worksheet_values[f"{SOURCE_COLUMNS['pessimistic_10']}{row_number}"].value,
                        est_formula=worksheet[f"{SOURCE_COLUMNS['est_formula']}{row_number}"].value,
                        est_formula_value=worksheet_values[f"{SOURCE_COLUMNS['est_formula']}{row_number}"].value,
                        tk_target=worksheet[f"{SOURCE_COLUMNS['tk_target']}{row_number}"].value,
                        tk_target_value=worksheet_values[f"{SOURCE_COLUMNS['tk_target']}{row_number}"].value,
                    )
                )
                if limit is not None and len(extracted_rows) >= limit:
                    return extracted_rows
    finally:
        workbook.close()
        workbook_values.close()

    return extracted_rows


class SQLiteWriter:
    def __init__(self, db_path: Path):
        self.db_path = db_path
        self.connection = sqlite3.connect(db_path)
        self.connection.row_factory = sqlite3.Row

    def setup_schema(self) -> None:
        self.connection.executescript(
            """
            CREATE TABLE IF NOT EXISTS source_rmi_rows (
                sheet_name TEXT NOT NULL,
                row_number INTEGER NOT NULL,
                roadmap_item TEXT NOT NULL,
                jira_id TEXT NOT NULL,
                man_days TEXT,
                man_days_value REAL,
                optimistic_50 TEXT,
                optimistic_50_value REAL,
                pessimistic_10 TEXT,
                pessimistic_10_value REAL,
                est_formula TEXT,
                est_formula_value REAL,
                tk_target TEXT,
                tk_target_value REAL,
                PRIMARY KEY (sheet_name, row_number)
            );

            CREATE TABLE IF NOT EXISTS epics (
                epic_key TEXT PRIMARY KEY,
                summary TEXT,
                issue_type TEXT,
                status TEXT,
                priority TEXT,
                jira_start_date TEXT,
                jira_due_date TEXT,
                jira_original_estimate TEXT,
                jira_original_estimate_seconds INTEGER,
                jira_aggregate_original_estimate_seconds INTEGER
            );

            CREATE TABLE IF NOT EXISTS stories (
                story_key TEXT PRIMARY KEY,
                epic_key TEXT NOT NULL,
                summary TEXT,
                issue_type TEXT,
                status TEXT,
                priority TEXT,
                jira_start_date TEXT,
                jira_due_date TEXT,
                jira_original_estimate TEXT,
                jira_original_estimate_seconds INTEGER,
                jira_aggregate_original_estimate_seconds INTEGER
            );

            CREATE TABLE IF NOT EXISTS story_descendants (
                issue_key TEXT PRIMARY KEY,
                parent_story_key TEXT NOT NULL,
                summary TEXT,
                issue_type TEXT,
                is_subtask INTEGER NOT NULL DEFAULT 0,
                status TEXT,
                priority TEXT,
                jira_start_date TEXT,
                jira_due_date TEXT,
                jira_original_estimate TEXT,
                jira_original_estimate_seconds INTEGER,
                jira_aggregate_original_estimate_seconds INTEGER
            );

            CREATE TABLE IF NOT EXISTS worklogs (
                worklog_id TEXT PRIMARY KEY,
                issue_key TEXT NOT NULL,
                author_display_name TEXT,
                started TEXT,
                time_spent TEXT,
                time_spent_seconds INTEGER
            );

            CREATE TABLE IF NOT EXISTS run_errors (
                error_scope TEXT NOT NULL,
                issue_key TEXT,
                sheet_name TEXT,
                row_number INTEGER,
                message TEXT NOT NULL
            );
            """
        )
        existing_columns = {
            row[1] for row in self.connection.execute("PRAGMA table_info(source_rmi_rows)").fetchall()
        }
        for column_name in [
            "man_days_value",
            "optimistic_50_value",
            "pessimistic_10_value",
            "est_formula_value",
            "tk_target_value",
        ]:
            if column_name not in existing_columns:
                self.connection.execute(f"ALTER TABLE source_rmi_rows ADD COLUMN {column_name} REAL")
        self.connection.commit()

    def reset_tables(self) -> None:
        for table_name in [
            "source_rmi_rows",
            "epics",
            "stories",
            "story_descendants",
            "worklogs",
            "run_errors",
        ]:
            self.connection.execute(f"DELETE FROM {table_name}")
        self.connection.commit()

    def insert_source_row(self, row: SourceRmiRow) -> None:
        self.connection.execute(
            """
            INSERT INTO source_rmi_rows (
                sheet_name, row_number, roadmap_item, jira_id, man_days, man_days_value,
                optimistic_50, optimistic_50_value, pessimistic_10, pessimistic_10_value,
                est_formula, est_formula_value, tk_target, tk_target_value
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                row.sheet_name,
                row.row_number,
                row.roadmap_item,
                row.jira_id,
                None if row.man_days is None else str(row.man_days),
                row.man_days_value,
                None if row.optimistic_50 is None else str(row.optimistic_50),
                row.optimistic_50_value,
                None if row.pessimistic_10 is None else str(row.pessimistic_10),
                row.pessimistic_10_value,
                None if row.est_formula is None else str(row.est_formula),
                row.est_formula_value,
                None if row.tk_target is None else str(row.tk_target),
                row.tk_target_value,
            ),
        )

    def upsert_epic(self, epic: Dict) -> None:
        self.connection.execute(
            """
            INSERT INTO epics (
                epic_key, summary, issue_type, status, priority, jira_start_date,
                jira_due_date, jira_original_estimate, jira_original_estimate_seconds,
                jira_aggregate_original_estimate_seconds
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(epic_key) DO UPDATE SET
                summary=excluded.summary,
                issue_type=excluded.issue_type,
                status=excluded.status,
                priority=excluded.priority,
                jira_start_date=excluded.jira_start_date,
                jira_due_date=excluded.jira_due_date,
                jira_original_estimate=excluded.jira_original_estimate,
                jira_original_estimate_seconds=excluded.jira_original_estimate_seconds,
                jira_aggregate_original_estimate_seconds=excluded.jira_aggregate_original_estimate_seconds
            """,
            (
                epic.get("key"),
                epic.get("summary"),
                epic.get("issue_type"),
                epic.get("status"),
                epic.get("priority"),
                epic.get("jira_start_date"),
                epic.get("jira_due_date"),
                epic.get("jira_original_estimate"),
                epic.get("jira_original_estimate_seconds"),
                epic.get("jira_aggregate_original_estimate_seconds"),
            ),
        )

    def upsert_story(self, epic_key: str, story: Dict) -> None:
        self.connection.execute(
            """
            INSERT INTO stories (
                story_key, epic_key, summary, issue_type, status, priority,
                jira_start_date, jira_due_date, jira_original_estimate,
                jira_original_estimate_seconds, jira_aggregate_original_estimate_seconds
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(story_key) DO UPDATE SET
                epic_key=excluded.epic_key,
                summary=excluded.summary,
                issue_type=excluded.issue_type,
                status=excluded.status,
                priority=excluded.priority,
                jira_start_date=excluded.jira_start_date,
                jira_due_date=excluded.jira_due_date,
                jira_original_estimate=excluded.jira_original_estimate,
                jira_original_estimate_seconds=excluded.jira_original_estimate_seconds,
                jira_aggregate_original_estimate_seconds=excluded.jira_aggregate_original_estimate_seconds
            """,
            (
                story.get("key"),
                epic_key,
                story.get("summary"),
                story.get("issue_type"),
                story.get("status"),
                story.get("priority"),
                story.get("jira_start_date"),
                story.get("jira_due_date"),
                story.get("jira_original_estimate"),
                story.get("jira_original_estimate_seconds"),
                story.get("jira_aggregate_original_estimate_seconds"),
            ),
        )

    def upsert_story_descendant(self, parent_story_key: str, issue: Dict) -> None:
        self.connection.execute(
            """
            INSERT INTO story_descendants (
                issue_key, parent_story_key, summary, issue_type, is_subtask,
                status, priority, jira_start_date, jira_due_date,
                jira_original_estimate, jira_original_estimate_seconds,
                jira_aggregate_original_estimate_seconds
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(issue_key) DO UPDATE SET
                parent_story_key=excluded.parent_story_key,
                summary=excluded.summary,
                issue_type=excluded.issue_type,
                is_subtask=excluded.is_subtask,
                status=excluded.status,
                priority=excluded.priority,
                jira_start_date=excluded.jira_start_date,
                jira_due_date=excluded.jira_due_date,
                jira_original_estimate=excluded.jira_original_estimate,
                jira_original_estimate_seconds=excluded.jira_original_estimate_seconds,
                jira_aggregate_original_estimate_seconds=excluded.jira_aggregate_original_estimate_seconds
            """,
            (
                issue.get("key"),
                parent_story_key,
                issue.get("summary"),
                issue.get("issue_type"),
                1 if issue.get("is_subtask") else 0,
                issue.get("status"),
                issue.get("priority"),
                issue.get("jira_start_date"),
                issue.get("jira_due_date"),
                issue.get("jira_original_estimate"),
                issue.get("jira_original_estimate_seconds"),
                issue.get("jira_aggregate_original_estimate_seconds"),
            ),
        )

    def replace_issue_worklogs(self, issue_key: str, worklogs: List[Dict]) -> None:
        self.connection.execute("DELETE FROM worklogs WHERE issue_key = ?", (issue_key,))
        for worklog in worklogs:
            author = worklog.get("author") or {}
            self.connection.execute(
                """
                INSERT INTO worklogs (
                    worklog_id, issue_key, author_display_name, started, time_spent, time_spent_seconds
                ) VALUES (?, ?, ?, ?, ?, ?)
                """,
                (
                    str(worklog.get("id")),
                    issue_key,
                    author.get("displayName"),
                    worklog.get("started"),
                    worklog.get("timeSpent"),
                    worklog.get("timeSpentSeconds"),
                ),
            )

    def insert_error(
        self,
        error_scope: str,
        message: str,
        issue_key: Optional[str] = None,
        sheet_name: Optional[str] = None,
        row_number: Optional[int] = None,
    ) -> None:
        self.connection.execute(
            """
            INSERT INTO run_errors (error_scope, issue_key, sheet_name, row_number, message)
            VALUES (?, ?, ?, ?, ?)
            """,
            (error_scope, issue_key, sheet_name, row_number, message),
        )

    def count_rows(self, table_name: str) -> int:
        cursor = self.connection.execute(f"SELECT COUNT(*) FROM {table_name}")
        return int(cursor.fetchone()[0])

    def commit(self) -> None:
        self.connection.commit()

    def close(self) -> None:
        self.connection.close()


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Extract Jira data for RMI workbook rows into SQLite.")
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK_PATH)
    parser.add_argument("--db", type=Path, default=DEFAULT_DB_PATH)
    parser.add_argument("--sheet-contains", default=DEFAULT_SHEET_FILTER)
    parser.add_argument("--limit", type=int, default=None)
    return parser


def run_extraction(
    workbook_path: Path,
    db_path: Path,
    env_config: Dict[str, str],
    sheet_contains: str = DEFAULT_SHEET_FILTER,
    limit: Optional[int] = None,
    progress_callback: Optional[Callable[[str], None]] = None,
) -> Dict[str, int]:
    def report(message: str) -> None:
        if progress_callback is not None:
            progress_callback(message)

    report(f"Scanning workbook rows from {workbook_path} (sheet filter: {sheet_contains!r})")
    source_rows = extract_source_rmi_rows(workbook_path, sheet_contains=sheet_contains, limit=limit)
    report(f"Workbook scan complete: {len(source_rows)} eligible rows found")
    jira_scoped_rows = [row for row in source_rows if row.jira_id]
    jira_client: Optional[JiraAPIClient] = None
    if jira_scoped_rows:
        report("Creating Jira client")
        jira_client = JiraAPIClient(
            env_config["JIRA_SITE"],
            env_config["JIRA_EMAIL"],
            env_config["JIRA_API_TOKEN"],
        )
        report(f"Testing Jira connection for site {env_config['JIRA_SITE']}")
        if not jira_client.test_connection():
            raise RuntimeError("Failed to connect to Jira.")
        report("Jira connection OK")
    else:
        report("No Jira-populated workbook rows found; skipping Jira client setup")

    db_path.parent.mkdir(parents=True, exist_ok=True)
    report(f"Preparing SQLite database at {db_path}")
    writer = SQLiteWriter(db_path)
    writer.setup_schema()
    writer.reset_tables()
    report("SQLite schema ready and target tables reset")

    try:
        for index, source_row in enumerate(source_rows, start=1):
            report(
                "Processing "
                f"[{index}/{len(source_rows)}] {source_row.sheet_name} row {source_row.row_number}: "
                f"{source_row.jira_id or 'No Jira ID'} - {source_row.roadmap_item}"
            )
            writer.insert_source_row(source_row)
            if not source_row.jira_id:
                report(f"  No Jira ID in workbook; stored source row only for {source_row.roadmap_item}")
                continue
            try:
                assert jira_client is not None
                epic = jira_client.get_issue(source_row.jira_id)
            except requests.exceptions.RequestException as exc:
                report(f"  Epic lookup failed for {source_row.jira_id}: {exc}")
                writer.insert_error(
                    "epic_lookup",
                    str(exc),
                    issue_key=source_row.jira_id,
                    sheet_name=source_row.sheet_name,
                    row_number=source_row.row_number,
                )
                continue

            if not epic.get("key"):
                report(f"  Epic lookup returned no Jira key for {source_row.jira_id}")
                writer.insert_error(
                    "epic_lookup",
                    "Jira issue lookup returned no issue key.",
                    issue_key=source_row.jira_id,
                    sheet_name=source_row.sheet_name,
                    row_number=source_row.row_number,
                )
                continue

            writer.upsert_epic(epic)
            report(
                f"  Epic fetched: {epic['key']} | status={epic.get('status') or 'Unknown'} | "
                f"type={epic.get('issue_type') or 'Unknown'}"
            )

            try:
                stories = jira_client.get_child_stories(epic["key"])
            except requests.exceptions.RequestException as exc:
                report(f"  Story lookup failed for {epic['key']}: {exc}")
                writer.insert_error("story_lookup", str(exc), issue_key=epic["key"])
                continue

            report(f"  Found {len(stories)} child stories for {epic['key']}")
            for story in stories:
                writer.upsert_story(epic["key"], story)
                report(
                    f"    Story {story['key']}: {story.get('summary') or 'No summary'} "
                    f"| status={story.get('status') or 'Unknown'}"
                )
                try:
                    descendants = jira_client.get_story_children(story["key"])
                except requests.exceptions.RequestException as exc:
                    report(f"    Descendant lookup failed for {story['key']}: {exc}")
                    writer.insert_error("descendant_lookup", str(exc), issue_key=story["key"])
                    continue

                report(f"    Found {len(descendants)} child issues under {story['key']}")
                for descendant in descendants:
                    writer.upsert_story_descendant(story["key"], descendant)
                    report(
                        f"      Fetching worklogs for {descendant['key']} "
                        f"({descendant.get('issue_type') or 'Child'})"
                    )
                    try:
                        worklogs = jira_client.get_issue_worklogs(descendant["key"])
                    except requests.exceptions.RequestException as exc:
                        report(f"      Worklog lookup failed for {descendant['key']}: {exc}")
                        writer.insert_error("worklog_lookup", str(exc), issue_key=descendant["key"])
                        continue
                    writer.replace_issue_worklogs(descendant["key"], worklogs)
                    report(f"      Stored {len(worklogs)} worklogs for {descendant['key']}")

        writer.commit()
        summary = {
            "eligible_rows": len(source_rows),
            "epics_fetched": writer.count_rows("epics"),
            "stories_fetched": writer.count_rows("stories"),
            "descendants_fetched": writer.count_rows("story_descendants"),
            "worklogs_fetched": writer.count_rows("worklogs"),
            "errors": writer.count_rows("run_errors"),
        }
        report(
            "Extraction complete: "
            f"eligible_rows={summary['eligible_rows']}, "
            f"epics_fetched={summary['epics_fetched']}, "
            f"stories_fetched={summary['stories_fetched']}, "
            f"descendants_fetched={summary['descendants_fetched']}, "
            f"worklogs_fetched={summary['worklogs_fetched']}, "
            f"errors={summary['errors']}"
        )
        return summary
    finally:
        writer.close()


def main() -> None:
    parser = build_arg_parser()
    args = parser.parse_args()
    env_config = load_env_config()
    summary = run_extraction(
        workbook_path=args.workbook,
        db_path=args.db,
        env_config=env_config,
        sheet_contains=args.sheet_contains,
        limit=args.limit,
    )

    print(f"Using Jira config from: {env_config['ENV_PATH']}")
    print(f"Workbook: {args.workbook}")
    print(f"SQLite DB: {args.db}")
    print(
        "Summary: "
        f"eligible_rows={summary['eligible_rows']}, "
        f"epics_fetched={summary['epics_fetched']}, "
        f"stories_fetched={summary['stories_fetched']}, "
        f"descendants_fetched={summary['descendants_fetched']}, "
        f"worklogs_fetched={summary['worklogs_fetched']}, "
        f"errors={summary['errors']}"
    )


if __name__ == "__main__":
    main()
