from __future__ import annotations

import argparse
from datetime import date, datetime
from difflib import SequenceMatcher
from html import escape
from pathlib import Path
import shutil
import tempfile
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font


STORY_SHEET = "April Estimates"
EPIC_SHEET = "April Deadlines"

APRIL_START = date(2026, 4, 1)
APRIL_END = date(2026, 4, 30)

BUCKETS = ["Pre-April", "April", "After April"]
UNDATED_BUCKET = "Undated"

PROJECT_LABELS = {
    "OmniConnect-2025": "OmniConnect",
    "Omni Chat": "OmniChat",
    "Digital Log.": "Digital Log",
    "Fintech Fuel": "Fintech Fuel",
    "Subscription Management": "Subscription Management",
}

PROJECT_ORDER = [
    "OmniConnect",
    "Fintech Fuel",
    "OmniChat",
    "Digital Log",
    "Subscription Management",
]

PLAN_SHEETS = [
    "OmniConnect RMI",
    "Fintech Fuel RMI",
    "OmniChat RMI",
    "Digital Log RMI",
]

PLAN_SHEET_LABELS = {
    "OmniConnect RMI": "OmniConnect",
    "Fintech Fuel RMI": "Fintech Fuel",
    "OmniChat RMI": "OmniChat",
    "Digital Log RMI": "Digital Log",
}

DEFAULT_SOURCE = Path("IPP Meeting Reports") / "April Estimates and Deadlines (Jira export).xlsx"
DEFAULT_OUTPUT_DIR = Path("IPP Meeting Reports") / "TK Approved for IPP"
DEFAULT_PRODUCT_BREAKDOWN_SOURCE = Path("IPP Meeting Reports") / "Epic Estimates Approved Plan.xlsx"
DEFAULT_IPP_SOURCE = Path("IPP Meeting Reports") / "IPP Meeting Work Items Estimates.xlsx"

RAW_PROJECT_LABELS = {
    "OmniConnect": "OmniConnect-2025",
    "OmniChat": "Omni Chat",
    "Digital Log": "Digital Log.",
    "Fintech Fuel": "Fintech Fuel",
    "Subscription Management": "Subscription Management",
}

STAGE_COLUMN_MAP = {
    "AD": "Process Design",
    "AE": "Research/URS",
    "AF": "Research/DDS",
    "AG": "Development",
    "AH": "SQA",
    "AI": "Process Testing",
    "AJ": "User Manual",
    "AK": "Regression SQA",
    "AL": "Production Release",
}

STAGE_DUE_DATE_COLUMNS = {
    "Process Design": "Z",
    "Research/URS": "AA",
    "Research/DDS": "AA",
    "Development": "AA",
    "SQA": "AB",
    "Process Testing": "AB",
    "User Manual": "AB",
    "Regression SQA": "AB",
    "Production Release": "AC",
}

STAGE_ALIASES = {
    "Process Design": ["process design", "prc design"],
    "Research/URS": ["research/urs", "research urs", "r/urs", "urs"],
    "Research/DDS": ["research/dds", "research dds", "r/dds", "dds"],
    "Development": ["development", "dev"],
    "SQA": ["sqa"],
    "Process Testing": ["process testing", "prc test", "testing"],
    "User Manual": ["user manual", "documentation", "doc"],
    "Regression SQA": ["regression sqa", "reg sqa"],
    "Production Release": ["production release", "release"],
}

EXTRA_STORY_ALIASES = {
    "QA Handover": ["qa handover"],
    "Bug Fixing": ["bug fixing", "bugfixing", "bug fixes", "bug fix"],
}


def safe_float(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    if isinstance(value, str):
        text = value.strip().replace(",", "")
        if not text:
            return 0.0
        try:
            return float(text)
        except ValueError:
            return 0.0
    return 0.0


def normalize_excel_date(value: Any) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        try:
            return datetime.fromisoformat(text).date()
        except ValueError:
            return None
    return None


def normalize_project_name(value: Any) -> str:
    project = str(value or "").strip()
    return PROJECT_LABELS.get(project, project)


def denormalize_project_name(value: Any) -> str:
    project = str(value or "").strip()
    return RAW_PROJECT_LABELS.get(project, project)


def classify_due_bucket(value: date | datetime | None) -> str:
    if value is None:
        return UNDATED_BUCKET
    due_date = value.date() if isinstance(value, datetime) else value
    if due_date < APRIL_START:
        return "Pre-April"
    if due_date <= APRIL_END:
        return "April"
    return "After April"


def format_hours(value: float) -> str:
    return f"{value:,.2f}"


def format_days(value: float) -> str:
    return f"{value:,.3f}"


def format_date(value: date | None) -> str:
    if value is None:
        return ""
    return value.strftime("%d-%b-%Y")


def format_snapshot_number(value: float) -> str:
    return f"{value:,.0f}"


def autosize_worksheet(worksheet) -> None:
    for column_cells in worksheet.columns:
        max_length = max((len(str(cell.value)) for cell in column_cells if cell.value is not None), default=0)
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 42)


def iter_sheet_records(worksheet) -> Iterable[dict[str, Any]]:
    header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [str(value).strip() if value is not None else "" for value in header_row]

    for row_index, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(value not in (None, "") for value in row):
            continue
        record = {headers[index]: row[index] for index in range(len(headers)) if headers[index]}
        record["_row_number"] = row_index
        yield record


def load_ipp_sheet_records(ipp_source_path: Path, sheet_name: str) -> list[dict[str, Any]]:
    workbook = load_excel_workbook(ipp_source_path, data_only=True)
    if sheet_name not in workbook.sheetnames:
        manual_candidate = ipp_source_path.with_name(f"{ipp_source_path.stem}_manually_filled{ipp_source_path.suffix}")
        if manual_candidate.exists():
            manual_workbook = load_excel_workbook(manual_candidate, data_only=True)
            if sheet_name in manual_workbook.sheetnames:
                workbook = manual_workbook
    worksheet = workbook[sheet_name]
    return list(iter_sheet_records(worksheet))


def seconds_to_hours(seconds: Any) -> float:
    return safe_float(seconds) / 3600


def seconds_to_days(seconds: Any) -> float:
    return safe_float(seconds) / 28800


def aggregate_optional_numeric(records: Iterable[dict[str, Any]], field_name: str) -> float | None:
    values = [safe_float(record.get(field_name)) for record in records if record.get(field_name) not in (None, "")]
    if not values:
        return None
    return sum(values)


def load_excel_workbook(path: Path, data_only: bool = True):
    try:
        return load_workbook(path, data_only=data_only)
    except PermissionError:
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as temp_file:
            temp_path = Path(temp_file.name)
        try:
            shutil.copy2(path, temp_path)
            return load_workbook(temp_path, data_only=data_only)
        finally:
            temp_path.unlink(missing_ok=True)


def clean_match_text(value: Any) -> str:
    return "".join(character.lower() for character in str(value or "") if character.isalnum())


def similarity_score(left: Any, right: Any) -> float:
    return SequenceMatcher(None, clean_match_text(left), clean_match_text(right)).ratio()


def parse_row_source(row_source: Any) -> tuple[str, int] | None:
    text = str(row_source or "").strip()
    if "!" not in text:
        return None
    sheet_name, cell_ref = text.split("!", 1)
    row_text = "".join(character for character in cell_ref if character.isdigit())
    if not row_text:
        return None
    return sheet_name, int(row_text)


def evaluate_plan_number(worksheet, cell_reference: str) -> float:
    value = worksheet[cell_reference].value
    if value in (None, ""):
        return 0.0
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    text = str(value).strip()
    if not text.startswith("="):
        return safe_float(text)

    row_number = worksheet[cell_reference].row
    column = worksheet[cell_reference].column_letter
    if column == "U":
        return sum(evaluate_plan_number(worksheet, f"{stage_col}{row_number}") for stage_col in ["L", "M", "N", "O", "P", "Q", "R", "S", "T"])
    if column == "V":
        base = evaluate_plan_number(worksheet, f"U{row_number}")
        return base * 0.5
    if column == "W":
        base = evaluate_plan_number(worksheet, f"U{row_number}")
        if "V" in text and "SUBSTITUTE(SUBSTITUTE($V$2" in text:
            return base + evaluate_plan_number(worksheet, f"V{row_number}") * 0.5
        return base * 1.1
    if column == "X":
        most_likely = evaluate_plan_number(worksheet, f"U{row_number}")
        optimistic = evaluate_plan_number(worksheet, f"V{row_number}")
        pessimistic = evaluate_plan_number(worksheet, f"W{row_number}")
        return (optimistic + (4 * most_likely) + pessimistic) / 6
    if column == "Y":
        return evaluate_plan_number(worksheet, f"X{row_number}") / 2
    if column == "AD":
        return evaluate_plan_number(worksheet, f"L{row_number}") if worksheet[f"Y{row_number}"].value not in (None, "") and worksheet[f"L{row_number}"].value not in (None, "") else 0.0
    if column == "AE":
        return evaluate_plan_number(worksheet, f"Y{row_number}") * 0.05 if worksheet[f"Y{row_number}"].value not in (None, "") and worksheet[f"M{row_number}"].value not in (None, "") else 0.0
    if column == "AF":
        return evaluate_plan_number(worksheet, f"Y{row_number}") * 0.10 if worksheet[f"Y{row_number}"].value not in (None, "") and worksheet[f"N{row_number}"].value not in (None, "") else 0.0
    if column == "AI":
        return evaluate_plan_number(worksheet, f"Q{row_number}") if worksheet[f"Y{row_number}"].value not in (None, "") and worksheet[f"Q{row_number}"].value not in (None, "") else 0.0
    if column == "AJ":
        return evaluate_plan_number(worksheet, f"Y{row_number}") * 0.10 if worksheet[f"Y{row_number}"].value not in (None, "") and worksheet[f"R{row_number}"].value not in (None, "") else 0.0
    if column == "AK":
        return evaluate_plan_number(worksheet, f"Y{row_number}") * 0.10 if worksheet[f"Y{row_number}"].value not in (None, "") and worksheet[f"S{row_number}"].value not in (None, "") else 0.0
    if column == "AL":
        return evaluate_plan_number(worksheet, f"T{row_number}") if worksheet[f"Y{row_number}"].value not in (None, "") and worksheet[f"T{row_number}"].value not in (None, "") else 0.0
    if column == "AH":
        if worksheet[f"Y{row_number}"].value in (None, ""):
            return 0.0
        dev_value = worksheet[f"O{row_number}"].value
        sqa_value = worksheet[f"P{row_number}"].value
        if dev_value in (None, "") and sqa_value in (None, ""):
            return 0.0
        if sqa_value in (None, ""):
            return 0.0
        base_remainder = evaluate_plan_number(worksheet, f"Y{row_number}") - sum(
            evaluate_plan_number(worksheet, f"{stage_col}{row_number}") for stage_col in ["AD", "AE", "AF", "AI", "AJ", "AK", "AL"]
        )
        if dev_value in (None, ""):
            return base_remainder
        return base_remainder * 15 / 65
    if column == "AG":
        if worksheet[f"Y{row_number}"].value in (None, ""):
            return 0.0
        dev_value = worksheet[f"O{row_number}"].value
        sqa_value = worksheet[f"P{row_number}"].value
        if dev_value in (None, ""):
            return 0.0
        base_remainder = evaluate_plan_number(worksheet, f"Y{row_number}") - sum(
            evaluate_plan_number(worksheet, f"{stage_col}{row_number}") for stage_col in ["AD", "AE", "AF", "AI", "AJ", "AK", "AL"]
        )
        if sqa_value in (None, ""):
            return base_remainder
        return base_remainder - evaluate_plan_number(worksheet, f"AH{row_number}")
    return safe_float(text.lstrip("="))


def evaluate_plan_date(worksheet, cell_reference: str) -> date | None:
    value = worksheet[cell_reference].value
    parsed_value = normalize_excel_date(value)
    if parsed_value is not None:
        return parsed_value
    text = str(value or "").strip()
    if text.startswith('=DATEVALUE("') and text.endswith('")'):
        return normalize_excel_date(text.split('"')[1])
    if text.startswith("=") and "+" in text:
        left, right = text[1:].split("+", 1)
        base_date = evaluate_plan_date(worksheet, left.strip())
        if base_date is None:
            return None
        return base_date.fromordinal(base_date.toordinal() + int(safe_float(right)))
    return None


def plan_row_has_dates_or_effort(worksheet, row_number: int) -> bool:
    return bool(
        str(worksheet[f"D{row_number}"].value or "").strip()
        and (
            evaluate_plan_number(worksheet, f"Y{row_number}") > 0
            or evaluate_plan_date(worksheet, f"Z{row_number}") is not None
            or evaluate_plan_date(worksheet, f"AC{row_number}") is not None
        )
    )


def find_story_stage(story_summary: str) -> str | None:
    normalized = clean_match_text(story_summary)
    for stage_name, aliases in STAGE_ALIASES.items():
        if any(clean_match_text(alias) in normalized or normalized in clean_match_text(alias) for alias in aliases):
            return stage_name
    return None


def find_matching_story_for_stage(stage_name: str, story_records: list[dict[str, Any]], used_keys: set[str]) -> dict[str, Any] | None:
    best_match = None
    best_score = 0.0
    for record in story_records:
        story_key = str(record.get("Story Key") or "").strip()
        if story_key in used_keys:
            continue
        summary = str(record.get("Story Summary") or "").strip()
        score = max(similarity_score(summary, alias) for alias in STAGE_ALIASES[stage_name])
        if score > best_score:
            best_match = record
            best_score = score
    if best_score >= 0.45:
        return best_match
    return None


def build_plan_epic_index(ipp_source_path: Path) -> dict[tuple[str, int], dict[str, Any]]:
    index: dict[tuple[str, int], dict[str, Any]] = {}
    for record in load_ipp_sheet_records(ipp_source_path, "Matched RMIs"):
        parsed = parse_row_source(record.get("Row Source"))
        if parsed is not None:
            index[parsed] = record
    return index


def build_plan_story_index(ipp_source_path: Path) -> dict[str, list[dict[str, Any]]]:
    stories_by_epic: dict[str, list[dict[str, Any]]] = {}
    for record in load_ipp_sheet_records(ipp_source_path, "Matched Epic Stories"):
        epic_key = str(record.get("Epic Key") or "").strip()
        stories_by_epic.setdefault(epic_key, []).append(record)
    return stories_by_epic


def write_story_source_sheet(worksheet, story_rows: list[dict[str, Any]]) -> None:
    worksheet.append(
        [
            "Work Item Type",
            "Work Item Title",
            "Work Item Key",
            "Parent key",
            "Parent summary",
            "Project name",
            "Status",
            "Original Estimates (Hours)",
            "Original Estimates (Days)",
            "Due date",
        ]
    )
    for row in story_rows:
        worksheet.append(
            [
                "Story",
                row["work_item_title"],
                row["work_item_key"],
                row["parent_key"],
                row["parent_summary"],
                denormalize_project_name(row["product"]),
                row["status"],
                row["estimate_hours"],
                row["estimate_days"],
                row["due_date"],
            ]
        )
    autosize_worksheet(worksheet)


def write_epic_source_sheet(worksheet, epic_rows: list[dict[str, Any]]) -> None:
    worksheet.append(["Work Item Type", "Work Item Title", "Work Item Key", "Project name", "Status", "Due date"])
    for row in epic_rows:
        worksheet.append(
            [
                "Epic",
                row["work_item_title"],
                row["work_item_key"],
                denormalize_project_name(row["product"]),
                row["status"],
                row["due_date"],
            ]
        )
    autosize_worksheet(worksheet)


def build_story_rows_from_ipp(ipp_source_path: Path) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for record in load_ipp_sheet_records(ipp_source_path, "Matched Epic Stories"):
        due_date = normalize_excel_date(record.get("Jira Due Date"))
        estimate_seconds = record.get("Jira Original Estimate Seconds")
        row = {
            "product": normalize_project_name(record.get("Product")),
            "project_name_raw": denormalize_project_name(normalize_project_name(record.get("Product"))),
            "work_item_key": str(record.get("Story Key") or "").strip(),
            "work_item_title": str(record.get("Story Summary") or "").strip(),
            "parent_key": str(record.get("Epic Key") or "").strip(),
            "parent_summary": str(record.get("Epic Summary") or record.get("RMI Name") or "").strip(),
            "status": str(record.get("Status") or "").strip(),
            "estimate_hours": seconds_to_hours(estimate_seconds),
            "estimate_days": seconds_to_days(estimate_seconds),
            "due_date": due_date,
            "bucket": classify_due_bucket(due_date),
            "source_sheet": "Matched Epic Stories",
            "source_row": record.get("_row_number"),
        }
        if row["work_item_key"] or row["work_item_title"] or row["estimate_hours"] or row["estimate_days"] or row["due_date"]:
            rows.append(row)
    return sorted(rows, key=lambda row: (row["due_date"] or date.max, row["product"], row["work_item_key"], row["work_item_title"]))


def build_epic_rows_from_ipp(ipp_source_path: Path) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for record in load_ipp_sheet_records(ipp_source_path, "Matched RMIs"):
        due_date = normalize_excel_date(record.get("Jira Due Date"))
        row = {
            "product": normalize_project_name(record.get("Product")),
            "project_name_raw": denormalize_project_name(normalize_project_name(record.get("Product"))),
            "work_item_key": str(record.get("Jira Key") or "").strip(),
            "work_item_title": str(record.get("Jira Summary") or record.get("RMI Name") or "").strip(),
            "status": str(record.get("Status") or "").strip(),
            "due_date": due_date,
            "bucket": classify_due_bucket(due_date),
            "source_sheet": "Matched RMIs",
            "source_row": record.get("_row_number"),
        }
        if row["work_item_key"] or row["work_item_title"] or row["due_date"]:
            rows.append(row)
    return sorted(rows, key=lambda row: (row["due_date"] or date.max, row["product"], row["work_item_key"]))


def sync_source_workbook_from_ipp(ipp_source_path: Path, source_path: Path) -> None:
    if not ipp_source_path.exists():
        return
    epic_rows = build_epic_rows_from_ipp(ipp_source_path)
    story_rows = build_story_rows_from_ipp(ipp_source_path)

    workbook = Workbook()
    epic_sheet = workbook.active
    epic_sheet.title = EPIC_SHEET
    write_epic_source_sheet(epic_sheet, epic_rows)

    story_sheet = workbook.create_sheet(STORY_SHEET)
    write_story_source_sheet(story_sheet, story_rows)

    source_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(source_path)


def build_story_rows_from_plan(plan_source_path: Path, ipp_source_path: Path) -> list[dict[str, Any]]:
    workbook = load_excel_workbook(plan_source_path, data_only=False)
    epic_index = build_plan_epic_index(ipp_source_path)
    stories_by_epic = build_plan_story_index(ipp_source_path)
    rows: list[dict[str, Any]] = []

    for sheet_name in PLAN_SHEETS:
        worksheet = workbook[sheet_name]
        product = PLAN_SHEET_LABELS[sheet_name]
        for row_number in range(3, worksheet.max_row + 1):
            if not plan_row_has_dates_or_effort(worksheet, row_number):
                continue
            epic_record = epic_index.get((sheet_name, row_number))
            if epic_record is None:
                continue

            epic_key = str(epic_record.get("Jira Key") or "").strip()
            epic_summary = str(epic_record.get("Jira Summary") or epic_record.get("RMI Name") or worksheet[f"D{row_number}"].value or "").strip()
            jira_story_records = stories_by_epic.get(epic_key, [])
            used_story_keys: set[str] = set()

            for stage_column, stage_name in STAGE_COLUMN_MAP.items():
                estimate_days = evaluate_plan_number(worksheet, f"{stage_column}{row_number}")
                if estimate_days <= 0:
                    continue
                matched_story = find_matching_story_for_stage(stage_name, jira_story_records, used_story_keys)
                story_key = str(matched_story.get("Story Key") or "").strip() if matched_story else ""
                if story_key:
                    used_story_keys.add(story_key)
                due_date_column = STAGE_DUE_DATE_COLUMNS[stage_name]
                due_date = normalize_excel_date(matched_story.get("Jira Due Date")) if matched_story else None
                if due_date is None:
                    due_date = evaluate_plan_date(worksheet, f"{due_date_column}{row_number}") or evaluate_plan_date(worksheet, f"AC{row_number}")
                rows.append(
                    {
                        "product": product,
                        "project_name_raw": denormalize_project_name(product),
                        "work_item_key": story_key,
                        "work_item_title": str(matched_story.get("Story Summary") or stage_name).strip() if matched_story else stage_name,
                        "parent_key": epic_key,
                        "parent_summary": epic_summary,
                        "status": str(matched_story.get("Status") or "Planned").strip() if matched_story else "Planned",
                        "estimate_hours": estimate_days * 8,
                        "estimate_days": estimate_days,
                        "due_date": due_date,
                        "bucket": classify_due_bucket(due_date),
                        "source_sheet": sheet_name,
                        "source_row": row_number,
                    }
                )

            for story_record in jira_story_records:
                story_key = str(story_record.get("Story Key") or "").strip()
                if story_key in used_story_keys:
                    continue
                story_summary = str(story_record.get("Story Summary") or "").strip()
                if find_story_stage(story_summary) is not None or any(
                    clean_match_text(alias) in clean_match_text(story_summary)
                    for aliases in EXTRA_STORY_ALIASES.values()
                    for alias in aliases
                ):
                    due_date = normalize_excel_date(story_record.get("Jira Due Date")) or evaluate_plan_date(worksheet, f"AC{row_number}")
                    rows.append(
                        {
                            "product": product,
                            "project_name_raw": denormalize_project_name(product),
                            "work_item_key": story_key,
                            "work_item_title": story_summary,
                            "parent_key": epic_key,
                            "parent_summary": epic_summary,
                            "status": str(story_record.get("Status") or "").strip(),
                            "estimate_hours": 0.0,
                            "estimate_days": 0.0,
                            "due_date": due_date,
                            "bucket": classify_due_bucket(due_date),
                            "source_sheet": sheet_name,
                            "source_row": row_number,
                        }
                    )

    return sorted(rows, key=lambda row: (row["due_date"] or date.max, row["product"], row["parent_key"], row["work_item_title"]))


def build_epic_rows_from_plan(plan_source_path: Path, ipp_source_path: Path) -> list[dict[str, Any]]:
    workbook = load_excel_workbook(plan_source_path, data_only=False)
    epic_index = build_plan_epic_index(ipp_source_path)
    rows: list[dict[str, Any]] = []

    for sheet_name in PLAN_SHEETS:
        worksheet = workbook[sheet_name]
        product = PLAN_SHEET_LABELS[sheet_name]
        for row_number in range(3, worksheet.max_row + 1):
            if not plan_row_has_dates_or_effort(worksheet, row_number):
                continue
            epic_record = epic_index.get((sheet_name, row_number))
            if epic_record is None:
                continue
            due_date = evaluate_plan_date(worksheet, f"AC{row_number}") or normalize_excel_date(epic_record.get("Jira Due Date"))
            rows.append(
                {
                    "product": product,
                    "project_name_raw": denormalize_project_name(product),
                    "work_item_key": str(epic_record.get("Jira Key") or "").strip(),
                    "work_item_title": str(epic_record.get("Jira Summary") or epic_record.get("RMI Name") or worksheet[f"D{row_number}"].value or "").strip(),
                    "status": str(epic_record.get("Status") or "").strip(),
                    "due_date": due_date,
                    "bucket": classify_due_bucket(due_date),
                    "source_sheet": sheet_name,
                    "source_row": row_number,
                }
            )

    return sorted(rows, key=lambda row: (row["due_date"] or date.max, row["product"], row["work_item_key"]))


def sync_source_workbook_from_plan(
    plan_source_path: Path,
    ipp_source_path: Path,
    source_path: Path,
) -> None:
    if not ipp_source_path.exists():
        return
    epic_rows = build_epic_rows_from_plan(plan_source_path, ipp_source_path)
    story_rows = build_story_rows_from_plan(plan_source_path, ipp_source_path)

    workbook = Workbook()
    epic_sheet = workbook.active
    epic_sheet.title = EPIC_SHEET
    write_epic_source_sheet(epic_sheet, epic_rows)

    story_sheet = workbook.create_sheet(STORY_SHEET)
    write_story_source_sheet(story_sheet, story_rows)

    source_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(source_path)


def extract_story_rows(source_path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(source_path, data_only=True)
    worksheet = workbook[STORY_SHEET]
    rows: list[dict[str, Any]] = []

    for record in iter_sheet_records(worksheet):
        if str(record.get("Work Item Type", "")).strip().lower() != "story":
            continue

        due_date = normalize_excel_date(record.get("Due date"))
        row = {
            "product": normalize_project_name(record.get("Project name")),
            "project_name_raw": str(record.get("Project name") or "").strip(),
            "work_item_key": str(record.get("Work Item Key") or "").strip(),
            "work_item_title": str(record.get("Work Item Title") or "").strip(),
            "parent_key": str(record.get("Parent key") or "").strip(),
            "parent_summary": str(record.get("Parent summary") or "").strip(),
            "status": str(record.get("Status") or "").strip(),
            "estimate_hours": safe_float(record.get("Original Estimates (Hours)")),
            "estimate_days": safe_float(record.get("Original Estimates (Days)")),
            "due_date": due_date,
            "bucket": classify_due_bucket(due_date),
            "source_sheet": STORY_SHEET,
            "source_row": record["_row_number"],
        }
        if row["work_item_key"] or row["work_item_title"] or row["estimate_hours"] or row["estimate_days"] or row["due_date"]:
            rows.append(row)

    return rows


def extract_epic_rows(source_path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(source_path, data_only=True)
    worksheet = workbook[EPIC_SHEET]
    rows: list[dict[str, Any]] = []

    for record in iter_sheet_records(worksheet):
        if str(record.get("Work Item Type", "")).strip().lower() != "epic":
            continue

        due_date = normalize_excel_date(record.get("Due date"))
        row = {
            "product": normalize_project_name(record.get("Project name")),
            "project_name_raw": str(record.get("Project name") or "").strip(),
            "work_item_key": str(record.get("Work Item Key") or "").strip(),
            "work_item_title": str(record.get("Work Item Title") or "").strip(),
            "status": str(record.get("Status") or "").strip(),
            "due_date": due_date,
            "bucket": classify_due_bucket(due_date),
            "source_sheet": EPIC_SHEET,
            "source_row": record["_row_number"],
        }
        if row["work_item_key"] or row["work_item_title"] or row["due_date"]:
            rows.append(row)

    return rows


def sort_products(products: Iterable[str]) -> list[str]:
    ordered: list[str] = []
    seen: set[str] = set()
    for product in PROJECT_ORDER:
        if product in products:
            ordered.append(product)
            seen.add(product)
    for product in sorted(products):
        if product not in seen:
            ordered.append(product)
    return ordered


def build_bucket_totals(story_rows: list[dict[str, Any]], epic_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    totals = {
        bucket: {
            "bucket": bucket,
            "story_count": 0,
            "parent_epic_keys": set(),
            "story_hours": 0.0,
            "story_days": 0.0,
            "epic_count": 0,
        }
        for bucket in BUCKETS
    }

    for row in story_rows:
        bucket = row["bucket"]
        if bucket in totals:
            totals[bucket]["story_count"] += 1
            if row["parent_key"]:
                totals[bucket]["parent_epic_keys"].add(row["parent_key"])
            totals[bucket]["story_hours"] += row["estimate_hours"]
            totals[bucket]["story_days"] += row["estimate_days"]

    for row in epic_rows:
        bucket = row["bucket"]
        if bucket in totals:
            totals[bucket]["epic_count"] += 1

    ordered_totals = []
    for bucket in BUCKETS:
        totals[bucket]["parent_epic_count"] = len(totals[bucket]["parent_epic_keys"])
        del totals[bucket]["parent_epic_keys"]
        ordered_totals.append(totals[bucket])
    return ordered_totals


def build_product_breakdown(story_rows: list[dict[str, Any]], epic_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    products = sort_products({row["product"] for row in story_rows} | {row["product"] for row in epic_rows})
    summary_rows: list[dict[str, Any]] = []

    for product in products:
        summary = {"product": product}
        for bucket in BUCKETS:
            summary[f"{bucket}_story_count"] = 0
            summary[f"{bucket}_story_hours"] = 0.0
            summary[f"{bucket}_story_days"] = 0.0
            summary[f"{bucket}_epic_count"] = 0

        for row in story_rows:
            if row["product"] != product or row["bucket"] not in BUCKETS:
                continue
            bucket = row["bucket"]
            summary[f"{bucket}_story_count"] += 1
            summary[f"{bucket}_story_hours"] += row["estimate_hours"]
            summary[f"{bucket}_story_days"] += row["estimate_days"]

        for row in epic_rows:
            if row["product"] != product or row["bucket"] not in BUCKETS:
                continue
            bucket = row["bucket"]
            summary[f"{bucket}_epic_count"] += 1

        summary_rows.append(summary)

    return summary_rows


def build_product_breakdown_totals(product_rows: list[dict[str, Any]]) -> dict[str, Any]:
    total = {"product": "Total"}
    for bucket in BUCKETS:
        total[f"{bucket}_story_count"] = sum(row[f"{bucket}_story_count"] for row in product_rows)
        total[f"{bucket}_story_hours"] = sum(row[f"{bucket}_story_hours"] for row in product_rows)
        total[f"{bucket}_story_days"] = sum(row[f"{bucket}_story_days"] for row in product_rows)
        total[f"{bucket}_epic_count"] = sum(row[f"{bucket}_epic_count"] for row in product_rows)
    return total


def build_product_snapshot_rows(plan_source_path: Path, ipp_source_path: Path | None = None) -> list[dict[str, Any]]:
    workbook = load_excel_workbook(plan_source_path, data_only=False)
    snapshot_rows: list[dict[str, Any]] = []

    for sheet_name in PLAN_SHEETS:
        worksheet = workbook[sheet_name]
        product = PLAN_SHEET_LABELS[sheet_name]
        product_snapshot = {
            "product": product,
            "epic_count": 0,
            "most_likely": 0.0,
            "optimistic": 0.0,
            "pessimistic": 0.0,
            "calculated_man_days": 0.0,
            "tk_approved": 0.0,
        }

        for row_number in range(3, worksheet.max_row + 1):
            if not plan_row_has_dates_or_effort(worksheet, row_number):
                continue
            product_snapshot["epic_count"] += 1
            product_snapshot["most_likely"] += evaluate_plan_number(worksheet, f"U{row_number}")
            product_snapshot["optimistic"] += evaluate_plan_number(worksheet, f"V{row_number}")
            product_snapshot["pessimistic"] += evaluate_plan_number(worksheet, f"W{row_number}")
            product_snapshot["calculated_man_days"] += evaluate_plan_number(worksheet, f"X{row_number}")
            product_snapshot["tk_approved"] += evaluate_plan_number(worksheet, f"Y{row_number}")

        snapshot_rows.append(product_snapshot)

    return snapshot_rows


def build_product_snapshot_totals(snapshot_rows: list[dict[str, Any]]) -> dict[str, Any]:
    return {
        "product": "Total",
        "epic_count": sum(row["epic_count"] for row in snapshot_rows),
        "most_likely": sum(row["most_likely"] for row in snapshot_rows),
        "optimistic": sum(row["optimistic"] for row in snapshot_rows),
        "pessimistic": sum(row["pessimistic"] for row in snapshot_rows),
        "calculated_man_days": sum(row["calculated_man_days"] for row in snapshot_rows),
        "tk_approved": sum(row["tk_approved"] for row in snapshot_rows),
    }


def build_high_level_stats(bucket_totals: list[dict[str, Any]], snapshot_totals: dict[str, Any], source_path: Path, product_breakdown_source: Path) -> dict[str, dict[str, Any]]:
    total_planned_hours = sum(row["story_hours"] for row in bucket_totals)
    tk_approved_days = snapshot_totals.get("tk_approved", 0.0)
    tk_approved_hours = tk_approved_days * 8
    story_hours_source = f"{source_path.name} -> Original Estimates (Hours)"
    snapshot_hours_source = f"{product_breakdown_source.name} -> Column Y (TK Approved man-days)"
    return {
        "Total Planned Hours": {
            "hours": total_planned_hours,
            "days": total_planned_hours / 8 if total_planned_hours else 0.0,
            "source": [story_hours_source, f"Derived from {story_hours_source} / 8"],
        },
        "TK Approved Hours": {
            "hours": tk_approved_hours,
            "days": tk_approved_days,
            "source": [snapshot_hours_source, "Converted to hours using 8 hours per man-day"],
        },
        "Total Variance": {
            "hours": total_planned_hours - tk_approved_hours,
            "days": total_planned_hours / 8 - tk_approved_days if total_planned_hours or tk_approved_days else 0.0,
            "source": [story_hours_source, snapshot_hours_source, "Derived diff after converting TK Approved man-days to hours"],
        },
        "April Planned Hours": {
            "hours": next((row["story_hours"] for row in bucket_totals if row["bucket"] == "April"), 0.0),
            "days": next((row["story_hours"] for row in bucket_totals if row["bucket"] == "April"), 0.0) / 8,
            "source": [story_hours_source, f"Derived from {story_hours_source} / 8"],
        },
        "Future Planned Hours": {
            "hours": next((row["story_hours"] for row in bucket_totals if row["bucket"] == "After April"), 0.0),
            "days": next((row["story_hours"] for row in bucket_totals if row["bucket"] == "After April"), 0.0) / 8,
            "source": [story_hours_source, f"Derived from {story_hours_source} / 8"],
        },
    }


def render_high_level_stats_cards(stats: dict[str, dict[str, Any]]) -> str:
    return "".join(
        f"""
        <div class=\"metric-card\">
          <div class=\"metric-label\">{escape(label)}</div>
          <div class=\"metric-value\">{format_hours(values['hours'])} hrs</div>
          <div class=\"metric-meta\">{format_days(values['days'])} man-days</div>
          <div class=\"metric-meta\">{'<br>'.join(escape(source) for source in values['source'])}</div>
        </div>
        """
        for label, values in stats.items()
    )


def ordered_story_rows(rows: list[dict[str, Any]], bucket: str) -> list[dict[str, Any]]:
    selected = [row for row in rows if row["bucket"] == bucket]
    return sorted(selected, key=lambda row: (row["due_date"] or date.max, row["product"], row["work_item_key"]))


def ordered_epic_rows(rows: list[dict[str, Any]], bucket: str) -> list[dict[str, Any]]:
    selected = [row for row in rows if row["bucket"] == bucket]
    return sorted(selected, key=lambda row: (row["due_date"] or date.max, row["product"], row["work_item_key"]))


def write_story_rows_sheet(worksheet, rows: list[dict[str, Any]]) -> None:
    worksheet.append(
        [
            "Product",
            "Project Name Raw",
            "Story Key",
            "Story Title",
            "Parent Key",
            "Parent Summary",
            "Status",
            "Estimate Hours",
            "Estimate Days",
            "Due Date",
            "Bucket",
            "Source Sheet",
            "Source Row",
        ]
    )
    for cell in worksheet[1]:
        cell.font = Font(bold=True)

    for row in rows:
        worksheet.append(
            [
                row["product"],
                row["project_name_raw"],
                row["work_item_key"],
                row["work_item_title"],
                row["parent_key"],
                row["parent_summary"],
                row["status"],
                row["estimate_hours"],
                row["estimate_days"],
                format_date(row["due_date"]),
                row["bucket"],
                row["source_sheet"],
                row["source_row"],
            ]
        )

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    autosize_worksheet(worksheet)


def write_epic_rows_sheet(worksheet, rows: list[dict[str, Any]]) -> None:
    worksheet.append(
        [
            "Product",
            "Project Name Raw",
            "Epic Key",
            "Epic Title",
            "Status",
            "Due Date",
            "Bucket",
            "Source Sheet",
            "Source Row",
        ]
    )
    for cell in worksheet[1]:
        cell.font = Font(bold=True)

    for row in rows:
        worksheet.append(
            [
                row["product"],
                row["project_name_raw"],
                row["work_item_key"],
                row["work_item_title"],
                row["status"],
                format_date(row["due_date"]),
                row["bucket"],
                row["source_sheet"],
                row["source_row"],
            ]
        )

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    autosize_worksheet(worksheet)


def write_underlying_workbook(output_path: Path, story_rows: list[dict[str, Any]], epic_rows: list[dict[str, Any]]) -> None:
    workbook = Workbook()
    story_sheet = workbook.active
    story_sheet.title = "Story Effort"
    write_story_rows_sheet(story_sheet, story_rows)

    epic_sheet = workbook.create_sheet("Epic Deliveries")
    write_epic_rows_sheet(epic_sheet, epic_rows)
    workbook.save(output_path)


def write_summary_workbook(
    output_path: Path,
    bucket_totals: list[dict[str, Any]],
    product_rows: list[dict[str, Any]],
    total_row: dict[str, Any],
    high_level_stats: dict[str, dict[str, Any]],
) -> None:
    workbook = Workbook()

    bucket_sheet = workbook.active
    bucket_sheet.title = "Bucket Summary"
    bucket_sheet.append(["Bucket", "Story Count", "Story Hours", "Story Days", "Epic Deliveries"])
    for cell in bucket_sheet[1]:
        cell.font = Font(bold=True)
    for row in bucket_totals:
        bucket_sheet.append([row["bucket"], row["story_count"], row["story_hours"], row["story_days"], row["epic_count"]])
    bucket_sheet.freeze_panes = "A2"
    bucket_sheet.auto_filter.ref = bucket_sheet.dimensions
    autosize_worksheet(bucket_sheet)

    stats_sheet = workbook.create_sheet("High Level Stats")
    stats_sheet.append(["Metric", "Hours", "Man Days", "Source"])
    for cell in stats_sheet[1]:
        cell.font = Font(bold=True)
    for label, values in high_level_stats.items():
        stats_sheet.append([label, values["hours"], values["days"], "\n".join(values["source"])])
    stats_sheet.freeze_panes = "A2"
    stats_sheet.auto_filter.ref = stats_sheet.dimensions
    autosize_worksheet(stats_sheet)

    product_sheet = workbook.create_sheet("Product Breakdown")
    headers = ["Product"]
    for bucket in BUCKETS:
        headers.extend([f"{bucket} Story Count", f"{bucket} Hours", f"{bucket} Days", f"{bucket} Deliveries"])
    product_sheet.append(headers)
    for cell in product_sheet[1]:
        cell.font = Font(bold=True)
    for row in product_rows + [total_row]:
        values = [row["product"]]
        for bucket in BUCKETS:
            values.extend(
                [
                    row[f"{bucket}_story_count"],
                    row[f"{bucket}_story_hours"],
                    row[f"{bucket}_story_days"],
                    row[f"{bucket}_epic_count"],
                ]
            )
        product_sheet.append(values)
    for cell in product_sheet[product_sheet.max_row]:
        cell.font = Font(bold=True)
    product_sheet.freeze_panes = "A2"
    product_sheet.auto_filter.ref = product_sheet.dimensions
    autosize_worksheet(product_sheet)

    workbook.save(output_path)


def write_delivery_workbook(output_path: Path, epic_rows: list[dict[str, Any]]) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)

    for bucket in BUCKETS:
        worksheet = workbook.create_sheet(bucket[:31])
        write_epic_rows_sheet(worksheet, ordered_epic_rows(epic_rows, bucket))

    workbook.save(output_path)


def render_effort_cards(bucket_totals: list[dict[str, Any]]) -> str:
    colors = {"Pre-April": "#334155", "April": "#b45309", "After April": "#0f766e"}
    return "".join(
        f"""
        <div class="metric-card">
          <div class="metric-label">{escape(row['bucket'])} Story Effort</div>
          <div class="metric-value" style="color:{colors[row['bucket']]}">{format_hours(row['story_hours'])} hrs</div>
          <div class="metric-meta">{row['parent_epic_count']} epics</div>
          <div class="metric-meta">{format_days(row['story_days'])} man-days</div>
        </div>
        """
        for row in bucket_totals
    )


def render_delivery_cards(bucket_totals: list[dict[str, Any]]) -> str:
    colors = {"Pre-April": "#334155", "April": "#b45309", "After April": "#0f766e"}
    return "".join(
        f"""
        <div class="metric-card">
          <div class="metric-label">{escape(row['bucket'])} Epic Deliveries</div>
          <div class="metric-value" style="color:{colors[row['bucket']]}">{row['epic_count']}</div>
          <div class="metric-meta">Epics due in this bucket</div>
        </div>
        """
        for row in bucket_totals
    )


def render_product_breakdown_table(product_rows: list[dict[str, Any]], total_row: dict[str, Any]) -> str:
    html = []
    for row in product_rows + [total_row]:
        row_class = ' class="total-row"' if row["product"] == "Total" else ""
        cells = [f"<td>{escape(row['product'])}</td>"]
        for bucket in BUCKETS:
            cells.extend(
                [
                    f"<td>{row[f'{bucket}_story_count']}</td>",
                    f"<td>{format_hours(row[f'{bucket}_story_hours'])}</td>",
                    f"<td>{format_days(row[f'{bucket}_story_days'])}</td>",
                    f"<td>{row[f'{bucket}_epic_count']}</td>",
                ]
            )
        html.append(f"<tr{row_class}>{''.join(cells)}</tr>")
    return "".join(html)


def render_product_snapshot_table(snapshot_rows: list[dict[str, Any]], totals: dict[str, Any]) -> str:
    html = []
    for row in snapshot_rows + [totals]:
        row_class = ' class="snapshot-total-row"' if row["product"] == "Total" else ""
        product_cell = "<td class=\"snapshot-product total-label\">Total</td>" if row["product"] == "Total" else f"<td class=\"snapshot-product\">{escape(row['product'])}</td>"
        html.append(
            f"""
            <tr{row_class}>
              <td class="snapshot-epics">{row['epic_count']}</td>
              {product_cell}
              <td class="snapshot-most">{format_snapshot_number(row['most_likely'])}</td>
              <td class="snapshot-opt">{format_snapshot_number(row['optimistic'])}</td>
              <td class="snapshot-pess">{format_snapshot_number(row['pessimistic'])}</td>
              <td class="snapshot-days">{format_snapshot_number(row['calculated_man_days'])}</td>
              <td class="snapshot-approved">{format_snapshot_number(row['tk_approved'])}</td>
            </tr>
            """
        )
    return "".join(html)


def summarize_due_dates(due_dates: list[date | None]) -> str:
    valid_dates = sorted({due_date for due_date in due_dates if due_date is not None})
    if not valid_dates:
        return ""
    if len(valid_dates) == 1:
        return format_date(valid_dates[0])
    return f"{format_date(valid_dates[0])} to {format_date(valid_dates[-1])}"


def build_story_bucket_groups(rows: list[dict[str, Any]], bucket: str) -> list[dict[str, Any]]:
    grouped_by_product: dict[str, dict[tuple[str, str], dict[str, Any]]] = {}

    for row in ordered_story_rows(rows, bucket):
        product_groups = grouped_by_product.setdefault(row["product"], {})
        parent_key = row["parent_key"] or row["parent_summary"] or row["work_item_key"]
        parent_summary = row["parent_summary"] or row["parent_key"] or row["work_item_title"] or row["work_item_key"]
        group_key = (parent_key, parent_summary)
        if group_key not in product_groups:
            product_groups[group_key] = {
                "parent_key": parent_key,
                "parent_summary": parent_summary,
                "estimate_hours": 0.0,
                "estimate_days": 0.0,
                "due_dates": [],
            }
        product_groups[group_key]["estimate_hours"] += row["estimate_hours"]
        product_groups[group_key]["estimate_days"] += row["estimate_days"]
        product_groups[group_key]["due_dates"].append(row["due_date"])

    groups = []
    for product in sort_products(grouped_by_product.keys()):
        epic_rows = list(grouped_by_product[product].values())
        epic_rows.sort(key=lambda row: ((min([d for d in row["due_dates"] if d is not None], default=date.max)), row["parent_summary"]))
        groups.append(
            {
                "product": product,
                "epics": epic_rows,
                "epic_count": len(epic_rows),
                "estimate_hours": sum(row["estimate_hours"] for row in epic_rows),
                "estimate_days": sum(row["estimate_days"] for row in epic_rows),
            }
        )
    return groups


def render_story_bucket_accordion(rows: list[dict[str, Any]], bucket: str, empty_message: str) -> str:
    grouped_rows = build_story_bucket_groups(rows, bucket)
    if not grouped_rows:
        return f'<div class="empty-state">{escape(empty_message)}</div>'

    html = []
    for index, group in enumerate(grouped_rows):
        open_attr = " open" if index == 0 else ""
        html.append(
            f"""
            <details class="accordion"{open_attr}>
              <summary class="accordion-summary">
                <span class="accordion-title">{escape(group['product'])}</span>
                <span class="accordion-meta">{group['epic_count']} epics</span>
                <span class="accordion-meta">{format_hours(group['estimate_hours'])} hrs</span>
                <span class="accordion-meta">{format_days(group['estimate_days'])} days</span>
              </summary>
              <table>
                <thead><tr><th>Epic Name</th><th>Hours</th><th>Days</th><th>Due Date</th></tr></thead>
                <tbody>
                  {
                    "".join(
                        f"<tr><td>{escape(epic['parent_summary'])}</td><td>{format_hours(epic['estimate_hours'])}</td><td>{format_days(epic['estimate_days'])}</td><td>{escape(summarize_due_dates(epic['due_dates']))}</td></tr>"
                        for epic in group["epics"]
                    )
                  }
                </tbody>
              </table>
            </details>
            """
        )
    return "".join(html)


def render_epic_table(rows: list[dict[str, Any]], empty_message: str) -> str:
    if not rows:
        return f'<tr><td colspan="5" class="empty-cell">{escape(empty_message)}</td></tr>'
    return "".join(
        f"""
        <tr>
          <td>{escape(row['product'])}</td>
          <td>{escape(row['work_item_key'])}</td>
          <td>{escape(row['work_item_title'])}</td>
          <td>{escape(row['status'])}</td>
          <td>{format_date(row['due_date'])}</td>
        </tr>
        """
        for row in rows
    )


def render_delivery_gantt(rows: list[dict[str, Any]], title: str, empty_message: str) -> str:
    dated_rows = [row for row in rows if row["due_date"] is not None]
    if not dated_rows:
        return f'<div class="empty-state">{escape(empty_message)}</div>'

    timeline_start = min(row["due_date"] for row in dated_rows)
    timeline_end = max(row["due_date"] for row in dated_rows)
    total_days = max((timeline_end - timeline_start).days + 1, 1)

    width = 1380
    left_margin = 520
    right_margin = 36
    header_height = 62
    product_header_height = 28
    row_height = 34
    plot_width = width - left_margin - right_margin
    grouped_rows: list[tuple[str, list[dict[str, Any]]]] = []
    for product in sort_products({row["product"] for row in dated_rows}):
        product_rows = [row for row in dated_rows if row["product"] == product]
        if product_rows:
            grouped_rows.append((product, product_rows))
    total_content_height = sum(product_header_height + len(product_rows) * row_height for _, product_rows in grouped_rows)
    height = header_height + total_content_height + 24

    product_colors = {
        "OmniConnect": "#0f766e",
        "Fintech Fuel": "#b45309",
        "OmniChat": "#7c3aed",
        "Digital Log": "#2563eb",
        "Subscription Management": "#334155",
    }

    def day_x(value: date) -> float:
        offset = (value - timeline_start).days
        return left_margin + (offset / total_days) * plot_width

    ticks = []
    current = timeline_start
    while current <= timeline_end:
        x = day_x(current)
        ticks.append(
            f"""
            <line x1="{x:.2f}" y1="28" x2="{x:.2f}" y2="{height - 16}" class="gantt-grid"></line>
            <text x="{x + 4:.2f}" y="22" class="gantt-axis">{current.strftime('%d-%b')}</text>
            """
        )
        current = current.fromordinal(current.toordinal() + 7)

    rows_svg = []
    current_y = header_height
    for product, product_rows in grouped_rows:
        color = product_colors.get(product, "#475569")
        rows_svg.append(
            f"""
            <g>
              <text x="16" y="{current_y + 18:.2f}" class="gantt-group">{escape(product)}</text>
              <text x="{left_margin - 12:.2f}" y="{current_y + 18:.2f}" text-anchor="end" class="gantt-group-meta">{len(product_rows)} epics</text>
            </g>
            """
        )
        current_y += product_header_height
        for row in product_rows:
            y = current_y
            x = day_x(row["due_date"])
            rows_svg.append(
                f"""
                <g>
                  <title>{escape(row['product'])}: {escape(row['work_item_title'])} ({format_date(row['due_date'])})</title>
                  <text x="40" y="{y + 14:.2f}" class="gantt-item">{escape(row['work_item_title'])}</text>
                  <rect x="{left_margin}" y="{y + 2:.2f}" width="{plot_width}" height="16" rx="8" class="gantt-track"></rect>
                  <line x1="{x:.2f}" y1="{y:.2f}" x2="{x:.2f}" y2="{y + 20:.2f}" stroke="{color}" stroke-width="3"></line>
                  <circle cx="{x:.2f}" cy="{y + 10:.2f}" r="6" fill="{color}"></circle>
                </g>
                """
            )
            current_y += row_height

    return f"""
    <div class="footnote">Timeline markers indicate epic due dates within this bucket.</div>
    <svg viewBox="0 0 {width} {height}" role="img" aria-label="{escape(title)}">
      {''.join(ticks)}
      {''.join(rows_svg)}
    </svg>
    """


def render_html(
    output_path: Path,
    source_path: Path,
    product_breakdown_source: Path,
    bucket_totals: list[dict[str, Any]],
    snapshot_rows: list[dict[str, Any]],
    snapshot_totals: dict[str, Any],
    high_level_stats: dict[str, dict[str, Any]],
    story_rows: list[dict[str, Any]],
    epic_rows: list[dict[str, Any]],
) -> None:
    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>TK Approved for IPP</title>
  <style>
    :root {{
      --bg: #eef3f9;
      --panel: #ffffff;
      --text: #102033;
      --muted: #5c6f83;
      --line: #d7dde8;
      --shadow: 0 18px 36px rgba(16, 32, 51, 0.08);
    }}
    * {{ box-sizing: border-box; }}
    body {{ margin: 0; font-family: "Segoe UI", Tahoma, sans-serif; background: linear-gradient(180deg, #e7eef8 0%, var(--bg) 24%, var(--bg) 100%); color: var(--text); }}
    .page {{ width: min(1460px, calc(100% - 40px)); margin: 0 auto; padding: 32px 0 56px; }}
    header {{ display: grid; gap: 10px; margin-bottom: 24px; }}
    h1 {{ margin: 0; font-size: 2.1rem; letter-spacing: -0.02em; }}
    h2 {{ margin: 0 0 10px; font-size: 1.15rem; }}
    .subtext {{ color: var(--muted); line-height: 1.5; max-width: 1080px; }}
    .metric-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(260px, 1fr)); gap: 14px; margin: 18px 0 0; }}
    .metric-card, .panel {{ background: var(--panel); border: 1px solid rgba(215, 221, 232, 0.85); border-radius: 18px; box-shadow: var(--shadow); }}
    .metric-card {{ padding: 18px 20px; }}
    .metric-label {{ color: var(--muted); font-size: 0.92rem; margin-bottom: 8px; }}
    .metric-value {{ font-size: 1.8rem; font-weight: 700; letter-spacing: -0.02em; }}
    .metric-meta {{ color: var(--muted); margin-top: 6px; font-size: 0.92rem; }}
    .panel {{ padding: 18px 20px 22px; margin-bottom: 18px; }}
    .section-heading {{ margin: 28px 0 12px; font-size: 1.45rem; letter-spacing: -0.02em; }}
    .section-intro {{ color: var(--muted); margin: -4px 0 16px; line-height: 1.5; }}
    table {{ width: 100%; border-collapse: collapse; margin-top: 10px; font-size: 0.94rem; }}
    th, td {{ padding: 10px 12px; border-bottom: 1px solid var(--line); text-align: left; vertical-align: top; }}
    th {{ background: #f8fbff; font-size: 0.84rem; letter-spacing: 0.02em; text-transform: uppercase; color: #4f6278; }}
    tbody tr:hover {{ background: #f9fbff; }}
    .empty-cell {{ color: var(--muted); text-align: center; background: #fbfdff; }}
    .empty-state {{ padding: 16px; border: 1px dashed #cbd5e1; border-radius: 12px; color: var(--muted); background: #fbfdff; }}
    .total-row td {{ font-weight: 700; border-top: 3px solid #111827; }}
    .file-list {{ margin: 6px 0 0; padding-left: 18px; color: var(--muted); }}
    .bucket-grid {{ display: grid; grid-template-columns: 1fr; gap: 18px; }}
    .footnote {{ color: var(--muted); font-size: 0.9rem; line-height: 1.5; }}
    .accordion {{ border: 1px solid var(--line); border-radius: 14px; background: #fbfdff; overflow: hidden; margin-top: 12px; }}
    .accordion-summary {{ list-style: none; display: flex; gap: 16px; align-items: center; justify-content: space-between; padding: 14px 16px; cursor: pointer; font-weight: 600; background: #f8fbff; }}
    .accordion-summary::-webkit-details-marker {{ display: none; }}
    .accordion-title {{ flex: 1; }}
    .accordion-meta {{ color: var(--muted); font-weight: 500; white-space: nowrap; }}
    .snapshot-table {{ margin-top: 0; }}
    .snapshot-table th {{ font-size: 0.92rem; text-transform: none; color: #000; padding: 14px 16px; }}
    .snapshot-table td {{ padding: 14px 16px; font-size: 1rem; }}
    .snapshot-epics {{ color: #6b7280; font-style: italic; text-align: center; }}
    .snapshot-product {{ font-weight: 700; font-size: 1.05rem; }}
    .snapshot-most-header {{ background: #dddddd; }}
    .snapshot-opt-header {{ background: #dceccf; }}
    .snapshot-pess-header {{ background: #f8dfcf; }}
    .snapshot-days-header {{ background: #d7e8f8; }}
    .snapshot-approved-header {{ color: #ff0000; }}
    .snapshot-opt {{ color: #6b8e23; text-align: center; }}
    .snapshot-pess {{ color: #d95f02; text-align: center; }}
    .snapshot-days {{ color: #4a74c9; text-align: center; }}
    .snapshot-approved {{ color: #ff0000; text-align: center; }}
    .snapshot-most {{ text-align: center; }}
    .snapshot-total-row td {{ font-weight: 700; font-style: italic; border-top: 3px solid #111827; }}
    .total-label {{ text-align: center; }}
    .top-metric-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(220px, 1fr)); gap: 14px; margin-top: 18px; }}
    .metric-card {{ padding: 18px 20px; }}
    svg {{ width: 100%; height: auto; display: block; margin-top: 12px; }}
    .gantt-grid {{ stroke: #d7dde8; stroke-dasharray: 3 5; }}
    .gantt-axis {{ font-size: 11px; fill: #627487; }}
    .gantt-product {{ font-size: 13px; font-weight: 700; fill: #102033; }}
    .gantt-item {{ font-size: 12px; fill: #314154; }}
    .gantt-group {{ font-size: 14px; font-weight: 700; fill: #102033; }}
    .gantt-group-meta {{ font-size: 11px; fill: #627487; }}
    .gantt-track {{ fill: #eef3f9; }}
    @media (max-width: 900px) {{
      .page {{ width: min(100% - 20px, 1460px); }}
      table {{ display: block; overflow-x: auto; }}
      .accordion-summary {{ flex-wrap: wrap; }}
    }}
  </style>
</head>
<body>
  <div class="page">
    <header>
      <h1>TK Approved for IPP</h1>
      <div class="subtext">
        Source workbook: <strong>{escape(str(source_path))}</strong>. Story effort is read from <strong>{escape(STORY_SHEET)}</strong>
        using <strong>Original Estimates (Hours)</strong>, <strong>Original Estimates (Days)</strong>, and story <strong>Due date</strong>.
        Delivery counts are read from <strong>{escape(EPIC_SHEET)}</strong> using epic <strong>Due date</strong>.
      </div>
      <div class="subtext">
        Leadership buckets in this report:
        <strong>Pre-April</strong> for due dates before 01-Apr-2026,
        <strong>April</strong> for due dates from 01-Apr-2026 through 30-Apr-2026,
        and <strong>After April</strong> for due dates after 30-Apr-2026.
      </div>
      <div class="subtext">
        Output files generated alongside this dashboard:
        <ul class="file-list">
          <li>TK Approved for IPP - Underlying Data.xlsx</li>
          <li>TK Approved for IPP - Summary Data.xlsx</li>
          <li>TK Approved for IPP - Delivery Buckets.xlsx</li>
        </ul>
      </div>
    </header>

    <h2 class="section-heading">Estimates</h2>
    <div class="section-intro">Story-derived effort figures grouped for leadership into pre-April, April, and post-April views.</div>

    <section class="panel">
      <h2>Summary Metrics</h2>
      <div class="footnote">High-level planned and TK-approved hour metrics for the current report. Man-days are calculated from hours using 8h/day, and the source workbook/column names are shown on each metric card.</div>
      <section class="top-metric-grid">{render_high_level_stats_cards(high_level_stats)}</section>
    </section>

    <section class="panel">
      <h2>Product Breakdown</h2>
      <div class="footnote">
        Snapshot values in this section are sourced from <strong>{escape(str(product_breakdown_source))}</strong>.
        Column mapping: U to Most-likely, V to Optimistic (50%), W to Pessimistic (10%), X to Calculated Man-Days, and Y to TK Approved.
      </div>
      <table class="snapshot-table">
        <thead>
          <tr>
            <th># of Epics</th>
            <th>Product</th>
            <th class="snapshot-most-header">Most-likely</th>
            <th class="snapshot-opt-header">Optimistic (50%)</th>
            <th class="snapshot-pess-header">Pessimistic (10%)</th>
            <th class="snapshot-days-header">Calculated Man-Days</th>
            <th class="snapshot-approved-header">TK Approved</th>
          </tr>
        </thead>
        <tbody>{render_product_snapshot_table(snapshot_rows, snapshot_totals)}</tbody>
      </table>
    </section>

    <section class="bucket-grid">
      <section class="panel">
        <h2>Pre-April Story Effort</h2>
        {render_story_bucket_accordion(story_rows, "Pre-April", "No pre-April stories found.")}
      </section>
      <section class="panel">
        <h2>April Story Effort</h2>
        {render_story_bucket_accordion(story_rows, "April", "No April stories found.")}
      </section>
      <section class="panel">
        <h2>After April Story Effort</h2>
        {render_story_bucket_accordion(story_rows, "After April", "No post-April stories found.")}
      </section>
    </section>

    <h2 class="section-heading">Deliveries</h2>
    <div class="section-intro">Epic due dates grouped into the same three leadership buckets, with details listed below.</div>

    <section class="panel">
      <h2>Delivery Summary</h2>
      <div class="footnote">Use these figures for deliveries planned. Only epic rows contribute to delivery counts.</div>
      <section class="metric-grid">{render_delivery_cards(bucket_totals)}</section>
    </section>

    <section class="bucket-grid">
      <section class="panel">
        <h2>Pre-April Deliveries</h2>
        {render_delivery_gantt(ordered_epic_rows(epic_rows, "Pre-April"), "Pre-April Deliveries Timeline", "No pre-April epic deliveries found.")}
      </section>
      <section class="panel">
        <h2>April Deliveries</h2>
        {render_delivery_gantt(ordered_epic_rows(epic_rows, "April"), "April Deliveries Timeline", "No April epic deliveries found.")}
      </section>
      <section class="panel">
        <h2>After April Deliveries</h2>
        {render_delivery_gantt(ordered_epic_rows(epic_rows, "After April"), "After April Deliveries Timeline", "No post-April epic deliveries found.")}
      </section>
    </section>
  </div>
</body>
</html>
"""
    output_path.write_text(html, encoding="utf-8")


def generate_dashboard(
    source_path: Path,
    output_dir: Path,
    product_breakdown_source: Path = DEFAULT_PRODUCT_BREAKDOWN_SOURCE,
    ipp_source_path: Path | None = None,
) -> dict[str, Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    if ipp_source_path is not None:
        sync_source_workbook_from_ipp(ipp_source_path, source_path)

    story_rows = extract_story_rows(source_path)
    epic_rows = extract_epic_rows(source_path)
    bucket_totals = build_bucket_totals(story_rows, epic_rows)
    snapshot_rows = build_product_snapshot_rows(product_breakdown_source)
    snapshot_totals = build_product_snapshot_totals(snapshot_rows)
    product_rows = build_product_breakdown(story_rows, epic_rows)
    total_row = build_product_breakdown_totals(product_rows)

    html_path = output_dir / "TK Approved for IPP.html"
    underlying_path = output_dir / "TK Approved for IPP - Underlying Data.xlsx"
    summary_path = output_dir / "TK Approved for IPP - Summary Data.xlsx"
    delivery_path = output_dir / "TK Approved for IPP - Delivery Buckets.xlsx"

    high_level_stats = build_high_level_stats(bucket_totals, snapshot_totals, source_path, product_breakdown_source)
    render_html(html_path, source_path, product_breakdown_source, bucket_totals, snapshot_rows, snapshot_totals, high_level_stats, story_rows, epic_rows)
    write_underlying_workbook(underlying_path, story_rows, epic_rows)
    write_summary_workbook(summary_path, bucket_totals, product_rows, total_row, high_level_stats)
    write_delivery_workbook(delivery_path, epic_rows)

    return {
        "html": html_path,
        "underlying": underlying_path,
        "summary": summary_path,
        "deliveries": delivery_path,
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate the TK Approved for IPP dashboard from the Jira export workbook.")
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE, help="Source Excel workbook path.")
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR, help="Directory for generated outputs.")
    parser.add_argument("--product-breakdown-source", type=Path, default=DEFAULT_PRODUCT_BREAKDOWN_SOURCE, help="Product breakdown workbook path.")
    parser.add_argument("--ipp-source", type=Path, default=DEFAULT_IPP_SOURCE, help="IPP workbook path used to refresh the report source files.")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    outputs = generate_dashboard(
        args.source.resolve(),
        args.output_dir.resolve(),
        args.product_breakdown_source.resolve(),
        args.ipp_source.resolve(),
    )
    for name, path in outputs.items():
        print(f"{name}: {path}")


if __name__ == "__main__":
    main()
