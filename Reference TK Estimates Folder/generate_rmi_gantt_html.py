from __future__ import annotations

import argparse
import json
import sqlite3
import sys
from datetime import date, datetime, timedelta
from functools import lru_cache
from html import escape
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import load_workbook


BASE_DIR = Path(__file__).parent
REPORTS_DIR = BASE_DIR / "IPP Meeting Reports"
DEFAULT_DB_PATH = REPORTS_DIR / "rmi_jira_extract.db"
DEFAULT_HTML_PATH = REPORTS_DIR / "rmi_jira_gantt.html"
DEFAULT_WORKBOOK_NAME = "Epic Estimates Approved Plan.xlsx"
SOURCE_JIRA_ID_COLUMN = "E"

PRODUCT_COLORS = {
    "OmniConnect": "#0f766e",
    "Fintech Fuel": "#b45309",
    "OmniChat": "#2563eb",
    "Digital Log": "#7c3aed",
}


@lru_cache(maxsize=None)
def jira_issue_url(issue_key: Any) -> str:
    issue_key_text = str(issue_key or "").strip()
    return f"/browse/{issue_key_text}" if issue_key_text else "#"


def workbook_url_from_cell(cell: Any) -> str:
    """Return a hyperlink target or URL-like text from a workbook cell."""
    hyperlink = getattr(cell, "hyperlink", None)
    if hyperlink:
        target = str(getattr(hyperlink, "target", "") or getattr(hyperlink, "location", "") or "").strip()
        if target:
            return target
    value = str(getattr(cell, "value", "") or "").strip()
    if value.startswith(("http://", "https://", "/browse/")):
        return value
    return ""


@lru_cache(maxsize=4)
def load_workbook_jira_url_map(workbook_path_text: str) -> dict[tuple[str, int], str]:
    """Map workbook sheet/row pairs to Jira URLs from the Jira ID column."""
    workbook_path = Path(workbook_path_text)
    if not workbook_path.exists():
        return {}
    try:
        workbook = load_workbook(workbook_path, data_only=False, read_only=False)
    except (PermissionError, OSError) as exc:
        print(
            f"warning: unable to read workbook hyperlinks from {workbook_path}: {exc}",
            file=sys.stderr,
        )
        return {}
    try:
        url_map: dict[tuple[str, int], str] = {}
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            for row_number in range(3, worksheet.max_row + 1):
                url = workbook_url_from_cell(worksheet[f"{SOURCE_JIRA_ID_COLUMN}{row_number}"])
                if url:
                    url_map[(sheet_name, row_number)] = url
        return url_map
    finally:
        workbook.close()


def parse_iso_date(value: Any) -> date | None:
    text = str(value or "").strip()
    if not text:
        return None
    candidate = text[:10]
    try:
        return date.fromisoformat(candidate)
    except ValueError:
        return None


def parse_iso_datetime(value: Any) -> datetime | None:
    text = str(value or "").strip()
    if not text:
        return None
    normalized = text
    if len(text) >= 5 and text[-5] in "+-" and text[-3] != ":":
        normalized = f"{text[:-2]}:{text[-2:]}"
    try:
        return datetime.fromisoformat(normalized)
    except ValueError:
        try:
            return datetime.fromisoformat(text[:19])
        except ValueError:
            return None


def format_date(value: date | None) -> str:
    """Format a date as DD-Mon-YYYY or empty string."""
    if value is None:
        return ""
    return value.strftime("%d-%b-%Y")


def format_short_date(value: date | None) -> str:
    """Format a date as DD Mon or empty string."""
    if value is None:
        return ""
    return value.strftime("%d %b")


def format_month_key_label(month_key: str) -> str:
    """Convert a YYYY-MM key to a human-readable 'Mon YYYY' label."""
    text = str(month_key or "").strip()
    if len(text) != 7 or text[4] != "-":
        return text
    year_text, month_text = text.split("-", 1)
    try:
        year = int(year_text)
        month = int(month_text)
    except ValueError:
        return text
    if month < 1 or month > 12:
        return text
    return date(year, month, 1).strftime("%b %Y")


def month_key_range(start_key: str, end_key: str) -> List[str]:
    """Return every YYYY-MM key from *start_key* through *end_key* inclusive."""
    if not start_key and not end_key:
        return []
    if not start_key:
        return [end_key] if end_key else []
    if not end_key:
        return [start_key]
    try:
        start = date(int(start_key[:4]), int(start_key[5:7]), 1)
        end = date(int(end_key[:4]), int(end_key[5:7]), 1)
    except ValueError:
        return []
    if start > end:
        start, end = end, start
    keys: List[str] = []
    cursor = start
    while cursor <= end:
        keys.append(cursor.strftime("%Y-%m"))
        if cursor.month == 12:
            cursor = date(cursor.year + 1, 1, 1)
        else:
            cursor = date(cursor.year, cursor.month + 1, 1)
    return keys


def available_month_keys_from_epic_details(epic_detail_records: List[dict[str, Any]]) -> List[str]:
    """Collect every month key spanned by any epic's start-through-due range."""
    keys = set()
    for epic in epic_detail_records:
        start_date = str(epic.get("start_date") or "").strip()
        due_date = str(epic.get("due_date") or "").strip()
        start_key = start_date[:7] if len(start_date) >= 7 and start_date[4] == "-" else ""
        due_key = due_date[:7] if len(due_date) >= 7 and due_date[4] == "-" else ""
        keys.update(month_key_range(start_key, due_key))
    return sorted(keys)


def format_datetime(value: datetime | None) -> str:
    """Format a datetime as DD-Mon-YYYY HH:MM or empty string."""
    if value is None:
        return ""
    return value.strftime("%d-%b-%Y %H:%M")


def format_hours(seconds: Any) -> str:
    """Format seconds as hours with two decimal places."""
    if seconds in (None, ""):
        return "0.00 h"
    return f"{float(seconds) / 3600:,.2f} h"


def format_days(seconds: Any) -> str:
    """Format seconds as man-days (8 h) with two decimal places."""
    if seconds in (None, ""):
        return "0.00 d"
    return f"{float(seconds) / 28800:,.2f} d"


def format_hours_compact(seconds: Any) -> str:
    """Format seconds as compact integer hours."""
    if seconds in (None, ""):
        return "0 h"
    return f"{round(float(seconds) / 3600):,.0f} h"


def format_days_compact(seconds: Any) -> str:
    """Format seconds as compact integer man-days."""
    if seconds in (None, ""):
        return "0 d"
    return f"{round(float(seconds) / 28800):,.0f} d"


def duration_span(seconds: Any, css_class: str = "duration-value") -> str:
    """Return an HTML span with data-seconds, data-hours, data-days attributes."""
    if seconds in (None, ""):
        seconds = 0
    return (
        f'<span class="{css_class}" data-seconds="{float(seconds)}" '
        f'data-hours="{format_hours(seconds)}" data-days="{format_days(seconds)}">{escape(format_hours(seconds))}</span>'
    )


def duration_span_compact(seconds: Any, css_class: str = "duration-value") -> str:
    """Return a compact HTML duration span."""
    if seconds in (None, ""):
        seconds = 0
    return (
        f'<span class="{css_class}" data-seconds="{float(seconds)}" '
        f'data-hours="{format_hours_compact(seconds)}" data-days="{format_days_compact(seconds)}">'
        f'{escape(format_hours_compact(seconds))}</span>'
    )


def duration_span_logged(seconds: Any) -> str:
    """Return a duration span followed by 'logged'."""
    return f"{duration_span(seconds)} logged"


def format_source_value(value: Any) -> str:
    """Format a source spreadsheet value for display."""
    numeric = parse_numeric(value)
    if numeric is None:
        return escape(str(value or "").strip())
    if numeric.is_integer():
        return escape(str(int(numeric)))
    return escape(f"{numeric:.2f}".rstrip("0").rstrip("."))


def man_day_span(value: Any) -> str:
    """Convert a man-days value to a duration span in seconds."""
    numeric = parse_numeric(value)
    if numeric is None:
        return ""
    return duration_span(numeric * 28800)


def unit_caption(hours_text: str, days_text: str) -> str:
    """Return a unit caption span that switches between hours/days labels."""
    return (
        f'<span class="unit-caption" data-label-hours="{escape(hours_text)}" '
        f'data-label-days="{escape(days_text)}">{escape(hours_text)}</span>'
    )


def parse_numeric(value: Any) -> float | None:
    """Parse a numeric value from various types."""
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    text = str(value).strip().replace(",", "")
    try:
        return float(text)
    except ValueError:
        return None


def render_comparison_note(note_text: str, state: str) -> str:
    """Render a small info badge comparing TK target to Jira estimate."""
    escaped_note = escape(note_text)
    return (
        f'<span class="comparison-note {state}" data-note="{escaped_note}" '
        f'title="{escaped_note}" aria-label="{escaped_note}">i</span>'
    )


def render_tk_target_comparison(epic: dict[str, Any]) -> str:
    """Render a comparison note for a single epic's TK target vs Jira estimate."""
    tk_target_value = parse_numeric(epic.get("tk_target_value"))
    epic_original_seconds = parse_numeric(epic.get("epic_original_estimate_seconds"))
    if tk_target_value is None or epic_original_seconds is None:
        return render_comparison_note("No numeric TK target available", "neutral")
    epic_original_hours = epic_original_seconds / 3600
    delta_hours = tk_target_value - epic_original_hours
    if abs(delta_hours) < 0.01:
        return render_comparison_note("Matches Jira original estimate", "match")
    if delta_hours > 0:
        return render_comparison_note(f"+{delta_hours:.2f} h above Jira original estimate", "over")
    return render_comparison_note(f"{abs(delta_hours):.2f} h below Jira original estimate", "under")


def tk_target_state(epic: dict[str, Any]) -> str:
    """Return the TK target comparison state: neutral, match, over, or under."""
    tk_target_value = parse_numeric(epic.get("tk_target_value"))
    epic_original_seconds = parse_numeric(epic.get("epic_original_estimate_seconds"))
    if tk_target_value is None or epic_original_seconds is None:
        return "neutral"
    delta_hours = tk_target_value - (epic_original_seconds / 3600)
    if abs(delta_hours) < 0.01:
        return "match"
    return "over" if delta_hours > 0 else "under"


def product_name_from_sheet(sheet_name: str) -> str:
    """Extract the product name from the sheet name by removing ' RMI' suffix."""
    return sheet_name.replace(" RMI", "").strip()


def issue_type_key(value: Any) -> str:
    """Normalize an issue type value for comparison."""
    return str(value or "").strip().lower().replace("-", " ").replace("_", " ")


def descendant_row_variant(descendant: dict[str, Any]) -> str:
    """Classify a descendant row as a bug-subtask or subtask."""
    issue_type = issue_type_key(descendant.get("issue_type"))
    return "bug-subtask" if "bug" in issue_type else "subtask"


def epic_search_text(epic: dict[str, Any]) -> str:
    """Build a combined search-text string for an epic."""
    parts = [
        str(epic.get("jira_id") or "").strip(),
        str(epic.get("epic_summary") or "").strip(),
        str(epic.get("roadmap_item") or "").strip(),
        str(epic.get("product") or "").strip(),
        str(epic.get("epic_status") or "").strip(),
        str(epic.get("epic_priority") or "").strip(),
    ]
    return " ".join(part for part in parts if part)


def epic_has_jira_population(epic: dict[str, Any]) -> bool:
    """Return True when a source epic row has Jira-populated epic/story data."""
    populated_fields = (
        "epic_summary",
        "epic_issue_type",
        "epic_status",
        "epic_priority",
        "epic_start_date",
        "epic_due_date",
        "epic_original_estimate",
    )
    if any(str(epic.get(field) or "").strip() for field in populated_fields):
        return True
    if (parse_numeric(epic.get("epic_original_estimate_seconds")) or 0.0) > 0:
        return True
    return bool(epic.get("stories"))


def wrap_svg_text(text: str, max_chars: int = 44) -> List[str]:
    """Word-wrap text for SVG tspan rendering."""
    words = str(text or "").split()
    if not words:
        return [""]
    lines: List[str] = []
    current = words[0]
    for word in words[1:]:
        candidate = f"{current} {word}"
        if len(candidate) <= max_chars:
            current = candidate
        else:
            lines.append(current)
            current = word
    lines.append(current)
    return lines


def load_report_data(db_path: Path) -> dict[str, Any]:
    """Load all report data from the SQLite database."""
    connection = sqlite3.connect(db_path)
    connection.row_factory = sqlite3.Row
    try:
        source_rows = [
            dict(row)
            for row in connection.execute(
                """
                SELECT
                    srr.sheet_name,
                    srr.row_number,
                    srr.roadmap_item,
                    srr.jira_id,
                    srr.man_days,
                    srr.man_days_value,
                    srr.optimistic_50,
                    srr.optimistic_50_value,
                    srr.pessimistic_10,
                    srr.pessimistic_10_value,
                    srr.est_formula,
                    srr.est_formula_value,
                    srr.tk_target,
                    srr.tk_target_value,
                    e.summary AS epic_summary,
                    e.issue_type AS epic_issue_type,
                    e.status AS epic_status,
                    e.priority AS epic_priority,
                    e.jira_start_date AS epic_start_date,
                    e.jira_due_date AS epic_due_date,
                    e.jira_original_estimate AS epic_original_estimate,
                    e.jira_original_estimate_seconds AS epic_original_estimate_seconds,
                    e.jira_aggregate_original_estimate_seconds AS epic_aggregate_original_estimate_seconds
                FROM source_rmi_rows srr
                LEFT JOIN epics e ON e.epic_key = srr.jira_id
                ORDER BY srr.sheet_name, srr.row_number
                """
            )
        ]
        stories_by_epic: Dict[str, List[dict[str, Any]]] = {}
        for row in connection.execute(
            """
            SELECT *
            FROM stories
            ORDER BY epic_key, jira_start_date, jira_due_date, story_key
            """
        ):
            item = dict(row)
            item["descendants"] = []
            stories_by_epic.setdefault(item["epic_key"], []).append(item)

        descendants_by_story: Dict[str, List[dict[str, Any]]] = {}
        for row in connection.execute(
            """
            SELECT
                sd.*,
                COALESCE(SUM(w.time_spent_seconds), 0) AS total_logged_seconds,
                COUNT(w.worklog_id) AS worklog_count
            FROM story_descendants sd
            LEFT JOIN worklogs w ON w.issue_key = sd.issue_key
            GROUP BY
                sd.issue_key,
                sd.parent_story_key,
                sd.summary,
                sd.issue_type,
                sd.is_subtask,
                sd.status,
                sd.priority,
                sd.jira_start_date,
                sd.jira_due_date,
                sd.jira_original_estimate,
                sd.jira_original_estimate_seconds,
                sd.jira_aggregate_original_estimate_seconds
            ORDER BY sd.parent_story_key, sd.jira_start_date, sd.jira_due_date, sd.issue_key
            """
        ):
            item = dict(row)
            descendants_by_story.setdefault(item["parent_story_key"], []).append(item)

        worklogs_by_issue: Dict[str, List[dict[str, Any]]] = {}
        for row in connection.execute(
            """
            SELECT *
            FROM worklogs
            ORDER BY issue_key, started, worklog_id
            """
        ):
            item = dict(row)
            worklogs_by_issue.setdefault(item["issue_key"], []).append(item)

        errors = [dict(row) for row in connection.execute("SELECT * FROM run_errors ORDER BY sheet_name, row_number, issue_key")]
    finally:
        connection.close()

    for stories in stories_by_epic.values():
        for story in stories:
            story["descendants"] = descendants_by_story.get(story["story_key"], [])
            for descendant in story["descendants"]:
                descendant["worklogs"] = worklogs_by_issue.get(descendant["issue_key"], [])

    for source_row in source_rows:
        source_row["product"] = product_name_from_sheet(source_row["sheet_name"])
        source_row["stories"] = stories_by_epic.get(source_row["jira_id"], [])
    workbook_url_map = load_workbook_jira_url_map(str(db_path.with_name(DEFAULT_WORKBOOK_NAME)))
    for source_row in source_rows:
        source_row["jira_url"] = workbook_url_map.get((source_row["sheet_name"], source_row["row_number"]), "")

    return {
        "source_rows": source_rows,
        "errors": errors,
    }


def build_summary(source_rows: List[dict[str, Any]], errors: List[dict[str, Any]]) -> dict[str, Any]:
    story_count = 0
    descendant_count = 0
    worklog_count = 0
    total_worklog_seconds = 0
    for epic in source_rows:
        story_count += len(epic["stories"])
        for story in epic["stories"]:
            descendant_count += len(story["descendants"])
            for descendant in story["descendants"]:
                worklog_count += len(descendant["worklogs"])
                total_worklog_seconds += sum(item.get("time_spent_seconds") or 0 for item in descendant["worklogs"])
    return {
        "epic_count": len(source_rows),
        "story_count": story_count,
        "descendant_count": descendant_count,
        "worklog_count": worklog_count,
        "total_worklog_seconds": total_worklog_seconds,
        "error_count": len(errors),
    }


def build_epic_metric_summary(source_rows: List[dict[str, Any]]) -> dict[str, dict[str, float]]:
    metric_keys = (
        "epic_count",
        "most_likely_seconds",
        "optimistic_seconds",
        "pessimistic_seconds",
        "calculated_seconds",
        "tk_approved_seconds",
      "idle_capacity_seconds",
        "jira_original_estimate_seconds",
        "story_estimate_seconds",
        "subtask_estimate_seconds",
    )

    def empty_totals() -> dict[str, float]:
        return {key: 0.0 for key in metric_keys}

    summary: dict[str, dict[str, float]] = {"all": empty_totals()}

    for epic in source_rows:
        product = str(epic.get("product") or "").strip() or "Unassigned"
        product_totals = summary.setdefault(product, empty_totals())
        scoped_totals = (summary["all"], product_totals)
        for totals in scoped_totals:
            totals["epic_count"] += 1
            totals["most_likely_seconds"] += (parse_numeric(epic.get("man_days_value")) or 0.0) * 28800
            totals["optimistic_seconds"] += (parse_numeric(epic.get("optimistic_50_value")) or 0.0) * 28800
            totals["pessimistic_seconds"] += (parse_numeric(epic.get("pessimistic_10_value")) or 0.0) * 28800
            totals["calculated_seconds"] += (parse_numeric(epic.get("est_formula_value")) or 0.0) * 28800
            totals["tk_approved_seconds"] += (parse_numeric(epic.get("tk_target_value")) or 0.0) * 28800
            totals["jira_original_estimate_seconds"] += parse_numeric(epic.get("epic_original_estimate_seconds")) or 0.0
            totals["story_estimate_seconds"] += epic_story_rollup_seconds(epic)
            totals["subtask_estimate_seconds"] += epic_subtask_rollup_seconds(epic)

    return summary


def build_epic_detail_records(source_rows: List[dict[str, Any]]) -> List[dict[str, Any]]:
    """Return a lightweight list of epic records for client-side month filtering and drawer details."""
    records: List[dict[str, Any]] = []
    for epic in source_rows:
        tk_days = parse_numeric(epic.get("tk_target_value")) or 0.0
        most_likely_days = parse_numeric(epic.get("man_days_value")) or 0.0
        optimistic_days = parse_numeric(epic.get("optimistic_50_value")) or 0.0
        pessimistic_days = parse_numeric(epic.get("pessimistic_10_value")) or 0.0
        calculated_days = parse_numeric(epic.get("est_formula_value")) or 0.0
        jira_original_seconds = parse_numeric(epic.get("epic_original_estimate_seconds")) or 0.0
        jira_populated = epic_has_jira_population(epic)
        story_records: List[dict[str, Any]] = []
        for story in epic.get("stories") or []:
            subtask_records: List[dict[str, Any]] = []
            for descendant in story.get("descendants") or []:
                if not descendant.get("is_subtask"):
                    continue
                subtask_records.append(
                    {
                        "issue_key": str(descendant.get("issue_key") or ""),
                        "title": str(descendant.get("summary") or ""),
                    "status": str(descendant.get("status") or ""),
                    "priority": str(descendant.get("priority") or ""),
                        "start_date": str(descendant.get("jira_start_date") or ""),
                        "due_date": str(descendant.get("jira_due_date") or ""),
                        "estimate_seconds": parse_numeric(descendant.get("jira_original_estimate_seconds")) or 0.0,
                        "jira_url": jira_issue_url(descendant.get("issue_key")),
                    }
                )
            story_records.append(
                {
                    "story_key": str(story.get("story_key") or ""),
                    "title": str(story.get("summary") or ""),
                  "status": str(story.get("status") or ""),
                  "priority": str(story.get("priority") or ""),
                    "start_date": str(story.get("jira_start_date") or ""),
                    "due_date": str(story.get("jira_due_date") or ""),
                    "estimate_seconds": parse_numeric(story.get("jira_original_estimate_seconds")) or 0.0,
                    "jira_url": jira_issue_url(story.get("story_key")),
                    "subtasks": subtask_records,
                }
            )
        records.append(
            {
                "jira_id": str(epic.get("jira_id") or ""),
                "title": str(epic.get("epic_summary") or epic.get("roadmap_item") or epic.get("jira_id") or ""),
                "product": (str(epic.get("product") or "").strip() or "Unassigned"),
                "status": str(epic.get("epic_status") or ""),
                "priority": str(epic.get("epic_priority") or ""),
                "start_date": str(epic.get("epic_start_date") or ""),
                "due_date": str(epic.get("epic_due_date") or ""),
                "jira_populated": jira_populated,
                "most_likely_seconds": most_likely_days * 28800,
                "optimistic_seconds": optimistic_days * 28800,
                "pessimistic_seconds": pessimistic_days * 28800,
                "calculated_seconds": calculated_days * 28800,
                "tk_approved_seconds": tk_days * 28800,
                "jira_original_estimate_seconds": jira_original_seconds,
                "story_estimate_seconds": epic_story_rollup_seconds(epic),
                "subtask_estimate_seconds": epic_subtask_rollup_seconds(epic),
                "story_count": len(epic.get("stories") or []),
                "jira_url": str(epic.get("jira_url") or ""),
                "stories": story_records,
            }
        )
    return records


def build_story_detail_records(source_rows: List[dict[str, Any]]) -> List[dict[str, Any]]:
    """Return a flat list of story records across all epics for the estimate drawer."""
    records: List[dict[str, Any]] = []
    for epic in source_rows:
        product = str(epic.get("product") or "").strip() or "Unassigned"
        for story in epic.get("stories", []):
            estimate = parse_numeric(story.get("jira_original_estimate_seconds")) or 0.0
            if estimate <= 0:
                continue
            records.append(
                {
                    "story_key": str(story.get("story_key") or ""),
                    "title": str(story.get("summary") or ""),
                    "product": product,
                    "epic_key": str(story.get("epic_key") or ""),
                    "status": str(story.get("status") or ""),
                    "priority": str(story.get("priority") or ""),
                    "start_date": str(story.get("jira_start_date") or ""),
                    "due_date": str(story.get("jira_due_date") or ""),
                    "estimate_seconds": estimate,
                    "jira_url": jira_issue_url(story.get("story_key")),
                }
            )
    return records


def build_subtask_detail_records(source_rows: List[dict[str, Any]]) -> List[dict[str, Any]]:
    """Return a flat list of subtask records across all epics/stories for the estimate drawer."""
    records: List[dict[str, Any]] = []
    for epic in source_rows:
        product = str(epic.get("product") or "").strip() or "Unassigned"
        for story in epic.get("stories", []):
            for descendant in story.get("descendants", []):
                if not descendant.get("is_subtask"):
                    continue
                estimate = parse_numeric(descendant.get("jira_original_estimate_seconds")) or 0.0
                if estimate <= 0:
                    continue
                records.append(
                    {
                        "issue_key": str(descendant.get("issue_key") or ""),
                        "title": str(descendant.get("summary") or ""),
                        "product": product,
                        "parent_story_key": str(descendant.get("parent_story_key") or ""),
                        "epic_key": str(epic.get("jira_id") or ""),
                        "status": str(descendant.get("status") or ""),
                        "priority": str(descendant.get("priority") or ""),
                        "start_date": str(descendant.get("jira_start_date") or ""),
                        "due_date": str(descendant.get("jira_due_date") or ""),
                        "estimate_seconds": estimate,
                        "jira_url": jira_issue_url(descendant.get("issue_key")),
                    }
                )
    return records


def build_rmi_schedule_records(source_rows: List[dict[str, Any]]) -> List[dict[str, Any]]:
    """Build per-epic records for the RMI Estimation & Scheduling table.

    Each record carries workbook estimation values (most_likely_days, tk_approved_days)
    and the full stories/subtasks tree so the client-side JS can bucket estimates into
    monthly columns using the same logic as the Month Story Estimate Analysis panel.
    """
    records: List[dict[str, Any]] = []
    for epic in source_rows:
        product = str(epic.get("product") or "").strip() or "Unassigned"
        most_likely_days = parse_numeric(epic.get("man_days_value")) or 0.0
        tk_approved_days = parse_numeric(epic.get("tk_target_value")) or 0.0
        story_records: List[dict[str, Any]] = []
        for story in epic.get("stories", []):
            subtask_records: List[dict[str, Any]] = []
            for descendant in story.get("descendants", []):
                if not descendant.get("is_subtask"):
                    continue
                subtask_records.append({
                    "start_date": str(descendant.get("jira_start_date") or ""),
                    "due_date": str(descendant.get("jira_due_date") or ""),
                    "estimate_seconds": parse_numeric(descendant.get("jira_original_estimate_seconds")) or 0.0,
                })
            story_records.append({
                "start_date": str(story.get("jira_start_date") or ""),
                "due_date": str(story.get("jira_due_date") or ""),
                "estimate_seconds": parse_numeric(story.get("jira_original_estimate_seconds")) or 0.0,
                "subtasks": subtask_records,
            })
        records.append({
            "roadmap_item": str(epic.get("roadmap_item") or epic.get("epic_summary") or ""),
            "jira_id": str(epic.get("jira_id") or ""),
            "jira_url": str(epic.get("jira_url") or ""),
            "jira_populated": epic_has_jira_population(epic),
            "product": product,
            "status": str(epic.get("epic_status") or ""),
            "start_date": str(epic.get("epic_start_date") or ""),
            "due_date": str(epic.get("epic_due_date") or ""),
            "most_likely_days": most_likely_days,
            "tk_approved_days": tk_approved_days,
            "stories": story_records,
        })
    records.sort(key=lambda r: (r["product"], r.get("start_date") or "9999"))
    return records


def available_years_from_schedule_records(records: List[dict[str, Any]]) -> List[int]:
    """Collect every distinct year from epic/story/subtask dates plus the current year."""
    years: set[int] = {date.today().year}
    for epic in records:
        for date_field in ("start_date", "due_date"):
            text = str(epic.get(date_field) or "")[:4]
            if text.isdigit():
                years.add(int(text))
        for story in epic.get("stories", []):
            for date_field in ("start_date", "due_date"):
                text = str(story.get(date_field) or "")[:4]
                if text.isdigit():
                    years.add(int(text))
            for subtask in story.get("subtasks", []):
                for date_field in ("start_date", "due_date"):
                    text = str(subtask.get(date_field) or "")[:4]
                    if text.isdigit():
                        years.add(int(text))
    return sorted(years)


def observed_schedule_years(records: List[dict[str, Any]]) -> List[int]:
    """Collect years that are actually present in epic/story/subtask dates."""
    years: set[int] = set()
    for epic in records:
        for date_field in ("start_date", "due_date"):
            text = str(epic.get(date_field) or "")[:4]
            if text.isdigit():
                years.add(int(text))
        for story in epic.get("stories", []):
            for date_field in ("start_date", "due_date"):
                text = str(story.get(date_field) or "")[:4]
                if text.isdigit():
                    years.add(int(text))
            for subtask in story.get("subtasks", []):
                for date_field in ("start_date", "due_date"):
                    text = str(subtask.get(date_field) or "")[:4]
                    if text.isdigit():
                        years.add(int(text))
    return sorted(years)


def initial_schedule_year(records: List[dict[str, Any]]) -> int:
    """Pick the first visible schedule year, preferring the current year when it exists."""
    current_year = date.today().year
    years = observed_schedule_years(records)
    if current_year in years:
        selected_year = current_year
    elif years:
        selected_year = max(years)
    else:
        selected_year = current_year
    return selected_year


def schedule_month_key(value: Any) -> str:
    """Convert a date-like value to a YYYY-MM key or empty string."""
    text = str(value or "").strip()
    month_key = ""
    if len(text) >= 7 and text[4] == "-":
        month_key = text[:7]
    return month_key


def is_cross_month_schedule_range(item: dict[str, Any]) -> bool:
    """Return True when a record spans two distinct non-empty month keys."""
    start_key = schedule_month_key(item.get("start_date"))
    due_key = schedule_month_key(item.get("due_date"))
    crosses_month = bool(start_key and due_key and start_key != due_key)
    return crosses_month


def round_half_up(value: float) -> int:
    """Round positive numbers the same way JS Math.round does for this report."""
    rounded = int(value + 0.5)
    return rounded


def schedule_display_number(value: float, unit: str, source: str) -> int:
    """Return the visible integer shown in a schedule table cell."""
    numeric = float(value or 0)
    if numeric <= 0:
        return 0
    if source == "days":
        display_value = numeric if unit == "days" else numeric * 8
        return round_half_up(display_value)
    divisor = 28800 if unit == "days" else 3600
    return round_half_up(numeric / divisor)


def format_schedule_number(value: float, unit: str, source: str) -> str:
    """Format schedule values without unit suffixes for table cells."""
    visible_value = schedule_display_number(value, unit, source)
    return f"{visible_value:,}" if visible_value > 0 else ""


def bucket_rmi_schedule_months(epic: dict[str, Any]) -> dict[str, float]:
    """Bucket story and subtask estimates into YYYY-MM totals for the schedule table."""
    totals: dict[str, float] = {}
    for story in epic.get("stories", []):
        story_estimate = parse_numeric(story.get("estimate_seconds")) or 0.0
        if is_cross_month_schedule_range(story):
            for subtask in story.get("subtasks", []):
                estimate = parse_numeric(subtask.get("estimate_seconds")) or 0.0
                if estimate <= 0:
                    continue
                start_key = schedule_month_key(subtask.get("start_date"))
                due_key = schedule_month_key(subtask.get("due_date"))
                month_key = ""
                if start_key and due_key and start_key == due_key:
                    month_key = start_key
                elif due_key:
                    month_key = due_key
                else:
                    month_key = start_key
                if not month_key:
                    continue
                totals[month_key] = totals.get(month_key, 0.0) + estimate
            continue
        if story_estimate <= 0:
            continue
        start_key = schedule_month_key(story.get("start_date"))
        due_key = schedule_month_key(story.get("due_date"))
        month_key = ""
        if start_key and due_key and start_key == due_key:
            month_key = start_key
        elif due_key:
            month_key = due_key
        else:
            month_key = start_key
        if not month_key:
            continue
        totals[month_key] = totals.get(month_key, 0.0) + story_estimate
    return totals


def render_rmi_schedule_table_rows(records: List[dict[str, Any]], year: int, unit: str = "hours") -> tuple[str, str]:
    """Render initial tbody and tfoot HTML for the schedule table."""
    month_keys = [f"{year}-{month:02d}" for month in range(1, 13)]
    by_product: Dict[str, List[dict[str, Any]]] = {}
    for epic in records:
        by_product.setdefault(str(epic.get("product") or "Unassigned"), []).append(epic)
    product_order = sorted(by_product)
    body_rows: List[str] = []
    grand_total_months = [0.0] * 12
    grand_most_likely = 0
    grand_tk_approved = 0
    for product in product_order:
        color = PRODUCT_COLORS.get(product, "#475569")
        body_rows.append(
            f'<tr class="rmi-sched-product-group"><td></td><td class="rmi-sched-group-label" style="border-left-color:{escape(color)}">{escape(product)}</td>{"<td></td>" * 16}</tr>'
        )
        subtotal_months = [0.0] * 12
        subtotal_most_likely = 0
        subtotal_tk_approved = 0
        for row_number, epic in enumerate(by_product[product], start=1):
            buckets = bucket_rmi_schedule_months(epic)
            status = str(epic.get("status") or "")
            status_lower = status.lower()
            jira_url = str(epic.get("jira_url") or "")
            jira_link = ""
            if jira_url and jira_url != "#":
                jira_link = (
                    f' <a class="rmi-sched-jira-link" href="{escape(jira_url)}" target="_blank" '
                    f'rel="noopener" title="Open in Jira">J</a>'
                )
            row_cells = [
                f"<td>{row_number}</td>",
                f'<td class="rmi-sched-cell-rmi">{escape(str(epic.get("roadmap_item") or ""))}{jira_link}</td>',
                f"<td>{escape(product)}</td>",
                f'<td><span class="rmi-sched-status-pill" data-status-lower="{escape(status_lower)}">{escape(status or "—")}</span></td>',
                f'<td>{format_schedule_number(parse_numeric(epic.get("most_likely_days")) or 0.0, unit, "days")}</td>',
                f'<td>{format_schedule_number(parse_numeric(epic.get("tk_approved_days")) or 0.0, unit, "days")}</td>',
            ]
            subtotal_most_likely += schedule_display_number(
                parse_numeric(epic.get("most_likely_days")) or 0.0,
                unit,
                "days",
            )
            subtotal_tk_approved += schedule_display_number(
                parse_numeric(epic.get("tk_approved_days")) or 0.0,
                unit,
                "days",
            )
            for month_index, month_key in enumerate(month_keys):
                value = buckets.get(month_key, 0.0)
                subtotal_months[month_index] += value
                row_cells.append(f'<td>{format_schedule_number(value, unit, "seconds")}</td>')
            body_rows.append(f'<tr class="rmi-sched-epic-row" data-product="{escape(product)}">{"".join(row_cells)}</tr>')
        subtotal_cells = [
            "<td></td>",
            f'<td style="border-left-color:{escape(color)}">{escape(product)} Subtotal</td>',
            "<td></td>",
            "<td></td>",
            f'<td>{subtotal_most_likely:,}</td>' if subtotal_most_likely > 0 else "<td></td>",
            f'<td>{subtotal_tk_approved:,}</td>' if subtotal_tk_approved > 0 else "<td></td>",
        ]
        for month_index, value in enumerate(subtotal_months):
            grand_total_months[month_index] += value
            subtotal_cells.append(f'<td>{format_schedule_number(value, unit, "seconds")}</td>')
        grand_most_likely += subtotal_most_likely
        grand_tk_approved += subtotal_tk_approved
        body_rows.append(f'<tr class="rmi-sched-product-subtotal">{"".join(subtotal_cells)}</tr>')
    foot_cells = [
        f"<td>{len(records)}</td>",
        "<td>Grand Total</td>",
        "<td></td>",
        "<td></td>",
        f'<td>{grand_most_likely:,}</td>' if grand_most_likely > 0 else "<td></td>",
        f'<td>{grand_tk_approved:,}</td>' if grand_tk_approved > 0 else "<td></td>",
    ]
    for value in grand_total_months:
        foot_cells.append(f'<td>{format_schedule_number(value, unit, "seconds")}</td>')
    body_html = "".join(body_rows)
    foot_html = f'<tr class="rmi-sched-grand-total">{"".join(foot_cells)}</tr>'
    return body_html, foot_html


def render_rmi_schedule_table(initial_body_html: str = "", initial_foot_html: str = "") -> str:
    """Render the static HTML shell for the RMI Estimation & Scheduling table.

    Row data, year filtering, and monthly bucketing are handled entirely client-side
    via the embedded rmiScheduleRecords JSON and JS.
    """
    month_headers = "".join(
        f'<th class="rmi-sched-month" data-month-index="{i + 1}">{m}</th>'
        for i, m in enumerate(["Jan", "Feb", "Mar", "Apr", "May", "Jun",
                                "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"])
    )
    return f"""
    <section class="panel rmi-schedule-panel" id="rmi-schedule-section">
      <div class="rmi-schedule-header-bar">
        <h2>RMI Estimation &amp; Scheduling</h2>
        <div class="rmi-schedule-controls">
          <label class="tk-month-toggle" for="rmi-jira-only-toggle">
            <input id="rmi-jira-only-toggle" class="tk-month-toggle-input" type="checkbox" checked>
            <span class="tk-month-toggle-track" aria-hidden="true"><span class="tk-month-toggle-thumb"></span></span>
            <span class="tk-month-toggle-text">Only Jira Populated Epics</span>
          </label>
          <label for="rmi-schedule-year" class="rmi-schedule-year-label">Year</label>
          <select id="rmi-schedule-year" class="rmi-schedule-year-select"></select>
        </div>
      </div>
      <div class="rmi-sched-product-cards" id="rmi-sched-product-cards"></div>
      <div class="table-frame">
        <table class="rmi-schedule-table" id="rmi-schedule-table">
          <thead>
            <tr class="rmi-sched-header-groups">
              <th></th>
              <th></th>
              <th></th>
              <th></th>
              <th colspan="2" class="rmi-sched-group-estimation">Estimation</th>
              <th colspan="12" class="rmi-sched-group-scheduling">Scheduling</th>
            </tr>
            <tr class="rmi-sched-header-cols">
              <th class="rmi-sched-col-num">#</th>
              <th class="rmi-sched-col-rmi">RMI</th>
              <th class="rmi-sched-col-product">Product</th>
              <th class="rmi-sched-col-status">Status</th>
              <th class="rmi-sched-col-ml">Most&nbsp;likely</th>
              <th class="rmi-sched-col-tk">TK&nbsp;Approved</th>
              {month_headers}
            </tr>
          </thead>
          <tbody id="rmi-schedule-body">{initial_body_html}</tbody>
          <tfoot id="rmi-schedule-foot">{initial_foot_html}</tfoot>
        </table>
      </div>
    </section>
    """


def build_gantt_rows(source_rows: List[dict[str, Any]]) -> List[dict[str, Any]]:
    rows: List[dict[str, Any]] = []
    for epic in source_rows:
        epic_start = parse_iso_date(epic.get("epic_start_date"))
        epic_due = parse_iso_date(epic.get("epic_due_date"))
        if epic_start and epic_due:
            tk_target_value = parse_numeric(epic.get("tk_target_value"))
            tk_target_hours = format_hours((tk_target_value or 0) * 28800) if tk_target_value is not None else "N/A"
            rows.append(
                {
                    "product": epic["product"],
                    "jira_id": epic["jira_id"],
                    "label": epic.get("epic_summary") or epic["roadmap_item"] or epic["jira_id"],
                    "search_text": epic_search_text(epic),
                    "start": epic_start,
                    "end": epic_due,
                    "meta": (
                        f"{epic.get('epic_status') or 'Unknown'} | "
                        f"TK Approved {tk_target_hours} | "
                        f"{format_date(epic_start)} to {format_date(epic_due)}"
                    ),
                }
            )
    rows.sort(key=lambda item: (item["product"], item["start"], item["end"], item["label"]))
    return rows


def render_gantt(rows: List[dict[str, Any]]) -> str:
    if not rows:
        return '<div class="empty-state">No epics have both Jira start and due dates, so no gantt bars could be drawn.</div>'

    timeline_start = min(row["start"] for row in rows)
    timeline_end = max(row["end"] for row in rows)
    total_days = max((timeline_end - timeline_start).days + 1, 1)
    width = 1680
    left_margin = 650
    right_margin = 40
    top_margin = 92
    row_height = 62
    section_bottom_padding = 24
    plot_width = width - left_margin - right_margin
    title_x = 18
    meta_x = 18
    title_max_chars = 44
    axis_week_y = 46
    axis_month_y = 22
    grid_start_y = 54

    def x_for_day(value: date) -> float:
        offset = (value - timeline_start).days
        return left_margin + (offset / total_days) * plot_width

    product_sections: List[str] = []
    products = sorted({row["product"] for row in rows})
    for product in products:
        product_rows = [row for row in rows if row["product"] == product]
        current_y = top_margin
        elements: List[str] = []
        month_elements: List[str] = []
        week_elements: List[str] = []
        month_cursor = date(timeline_start.year, timeline_start.month, 1)
        while month_cursor <= timeline_end:
            next_month = date(month_cursor.year + (1 if month_cursor.month == 12 else 0), 1 if month_cursor.month == 12 else month_cursor.month + 1, 1)
            visible_start = max(month_cursor, timeline_start)
            visible_end = min(next_month, timeline_end + timedelta(days=1))
            start_x = x_for_day(visible_start)
            end_x = x_for_day(visible_end)
            width_x = max(end_x - start_x, 0)
            month_class = "gantt-month-band alt" if month_cursor.month % 2 == 0 else "gantt-month-band"
            month_elements.append(
                f"""
                <rect x="{start_x:.2f}" y="{grid_start_y:.2f}" width="{width_x:.2f}" height="100%" class="{month_class}"></rect>
                <line x1="{start_x:.2f}" y1="{grid_start_y - 14:.2f}" x2="{start_x:.2f}" y2="100%" class="gantt-month-line"></line>
                <text x="{start_x + 6:.2f}" y="{axis_month_y:.2f}" class="gantt-month-axis">{month_cursor.strftime('%b %Y')}</text>
                """
            )
            month_cursor = next_month

        week_cursor = timeline_start - timedelta(days=timeline_start.weekday())
        while week_cursor <= timeline_end:
            if week_cursor >= timeline_start:
                x = x_for_day(week_cursor)
                week_elements.append(
                    f"""
                    <line x1="{x:.2f}" y1="{grid_start_y:.2f}" x2="{x:.2f}" y2="100%" class="gantt-grid"></line>
                    <text x="{x + 4:.2f}" y="{axis_week_y:.2f}" class="gantt-axis">{escape(format_short_date(week_cursor))}</text>
                    """
                )
            week_cursor += timedelta(days=7)

        for row in product_rows:
            start_x = x_for_day(row["start"])
            end_x = x_for_day(row["end"] + timedelta(days=1))
            bar_width = max(end_x - start_x, 6)
            end_label_x = max(start_x + 34, end_x - 4)
            color = PRODUCT_COLORS.get(row["product"], "#475569")
            title_lines = wrap_svg_text(row["label"], max_chars=title_max_chars)
            title_tspans = "".join(
                f'<tspan x="{title_x}" dy="{0 if index == 0 else 15}">{escape(line)}</tspan>'
                for index, line in enumerate(title_lines[:3])
            )
            bar_y = current_y + 20
            bar_label_y = current_y + 47
            elements.append(
                f"""
                <g class="gantt-epic-row" data-product="{escape(row['product'])}" data-search="{escape(row['search_text'])}">
                  <title>{escape(row['jira_id'])} | {escape(row['label'])} | {escape(row['meta'])}</title>
                  <text x="{title_x}" y="{current_y + 8:.2f}" class="gantt-label">{title_tspans}</text>
                  <text x="{meta_x}" y="{current_y + 41:.2f}" class="gantt-meta-left">{escape(row['meta'])}</text>
                  <rect x="{left_margin}" y="{bar_y - 10:.2f}" width="{plot_width}" height="18" rx="9" class="gantt-track"></rect>
                  <rect x="{start_x:.2f}" y="{bar_y - 10:.2f}" width="{bar_width:.2f}" height="18" rx="9" class="gantt-bar epic-bar" fill="{color}"></rect>
                  <text x="{start_x:.2f}" y="{bar_label_y:.2f}" class="gantt-date-label" text-anchor="start">{escape(format_short_date(row['start']))}</text>
                  <text x="{end_label_x:.2f}" y="{bar_label_y:.2f}" class="gantt-date-label" text-anchor="end">{escape(format_short_date(row['end']))}</text>
                </g>
                """
            )
            current_y += row_height

        height = current_y + section_bottom_padding
        product_sections.append(
            f"""
            <section class="gantt-product-section" data-product="{escape(product)}">
              <h3>{escape(product)}</h3>
              <svg viewBox="0 0 {width} {height}" role="img" aria-label="{escape(product)} epic gantt chart">
                {''.join(month_elements)}
                {''.join(week_elements)}
                {''.join(elements)}
              </svg>
            </section>
            """
        )

    return f"""
    <div class="footnote">Bars are drawn only for epics where both Jira start date and Jira due date are available. The chart is grouped by product and each epic row shows status, TK Approved hours, and the start-to-due schedule.</div>
    <div class="gantt-product-grid">
      {''.join(product_sections)}
    </div>
    """


def render_metric_cards(epic_metric_summary: dict[str, dict[str, float]]) -> str:
    current = epic_metric_summary["all"]

    def render_card(key: str, label: str, meta: str, metric_type: str, accent: str, extra: str = "") -> str:
        value_html = (
            f'<div class="metric-value">{int(current[key]):,}</div>'
            if metric_type == "count"
            else duration_span_compact(current[key], css_class="metric-value duration-value")
        )
        return f"""
        <section class="metric-card {accent}" data-metric-key="{key}" data-metric-type="{metric_type}"{extra}>
          <div class="metric-label">{escape(label)}</div>
          <div class="metric-value-wrap">
            {value_html}
          </div>
          <div class="metric-meta">{escape(meta)}</div>
        </section>
        """

    epic_count_card = render_card(
        "epic_count",
        "Total # of RMI Epics",
        "Epic parents in the selected product scope",
        "count",
        "metric-card-teal",
    )

    # Joined 4-card estimate group (Optimistic -> Most Likely -> Pessimistic -> Calculated).
    # Colors step from light to dark in a single blue hue so the cards read as one visual unit.
    estimate_cards = [
        ("optimistic_seconds", "Optimistic", "Workbook optimistic total for the selected epic set", "metric-estimate-step-1"),
        ("most_likely_seconds", "Most Likely", "Workbook most likely total for the selected epic set", "metric-estimate-step-2"),
        ("pessimistic_seconds", "Pessimistic", "Workbook pessimistic total for the selected epic set", "metric-estimate-step-3"),
        ("calculated_seconds", "Calculated", "Workbook calculated estimate total for the selected epic set", "metric-estimate-step-4"),
    ]
    estimate_group = (
        '<div class="estimate-cards-group" role="group" aria-label="Estimation range">'
        + "".join(render_card(key, label, meta, "duration", f"metric-estimate-card {step}") for key, label, meta, step in estimate_cards)
        + "</div>"
    )

    tk_approved_card = render_card(
        "tk_approved_seconds",
        "TK Approved",
        "TK approved total for the selected epic set. Click to view contributing epics.",
        "duration",
        "metric-card-emerald metric-card-hero metric-card-clickable",
        ' data-metric-context="tk-in-scope" role="button" tabindex="0" aria-label="Show TK Approved epics"',
    )

    idle_capacity_card = render_card(
      "idle_capacity_seconds",
      "Idle Hours/Days",
      "Total Availability minus TK Approved for the current scope.",
      "duration",
      "metric-card-cyan",
    )

    jira_original_card = render_card(
        "jira_original_estimate_seconds",
        "Epic Estimates",
        "Epic-level Jira original estimate total. Click to view contributing epics.",
        "duration",
        "metric-card-indigo metric-card-clickable",
        ' data-metric-context="epic-estimates" role="button" tabindex="0" aria-label="Show Epic Estimates breakdown"',
    )

    aggregate_card = render_card(
        "story_estimate_seconds",
        "Story Estimates",
        "Story-level original estimate total (excludes subtasks). Click to view contributing stories.",
        "duration",
        "metric-card-slate metric-card-clickable",
        ' data-metric-context="story-estimates" role="button" tabindex="0" aria-label="Show Story Estimates breakdown"',
    )

    subtask_estimate_card = render_card(
        "subtask_estimate_seconds",
        "Subtask Estimates",
        "Subtask-level original estimate total. Click to view contributing subtasks.",
        "duration",
        "metric-card-violet metric-card-clickable",
        ' data-metric-context="subtask-estimates" role="button" tabindex="0" aria-label="Show Subtask Estimates breakdown"',
    )

    return (
        epic_count_card
        + estimate_group
        + tk_approved_card
      + idle_capacity_card
        + jira_original_card
        + aggregate_card
        + subtask_estimate_card
    )


def render_product_tk_cards(epic_metric_summary: dict[str, dict[str, float]]) -> str:
    ordered_products = ["all", "Digital Log", "Fintech Fuel", "OmniChat", "OmniConnect"]
    labels = {"all": "All Products"}
    cards: List[str] = []
    for product in ordered_products:
        totals = epic_metric_summary.get(product, {})
        label = labels.get(product, product)
        color = PRODUCT_COLORS.get(product, "#102033") if product != "all" else "#102033"
        cards.append(
            f"""
            <section class="product-summary-card{' active' if product == 'all' else ''}" data-product-summary="{escape(product)}" data-product-filter-card="true" style="--product-accent: {color};" role="button" tabindex="0" aria-pressed="{'true' if product == 'all' else 'false'}" aria-label="Filter page by {escape(label)}">
              <div class="product-summary-label">{escape(label)}</div>
              <div class="product-summary-value">
                {duration_span_compact(totals.get("tk_approved_seconds", 0), css_class="product-summary-duration duration-value")}
              </div>
              <div class="product-summary-meta">Total TK Approved</div>
            </section>
            """
        )
    return f'<section class="product-summary-grid">{"".join(cards)}</section>'


def render_label_value_table(
    pairs: List[tuple[str, str]],
    *,
    columns: int = 3,
    table_class: str = "pivot-detail-table",
) -> str:
    if not pairs:
        return '<div class="empty-state tight">No fields available.</div>'

    rows: List[str] = []
    for start in range(0, len(pairs), columns):
        chunk = pairs[start : start + columns]
        headers = "".join(f"<th>{label}</th>" for label, _ in chunk)
        values = "".join(f"<td>{value}</td>" for _, value in chunk)
        filler_count = columns - len(chunk)
        if filler_count > 0:
            headers += "<th></th>" * filler_count
            values += "<td></td>" * filler_count
        rows.append(f"<tr>{headers}</tr><tr>{values}</tr>")
    return f'<table class="{table_class}"><tbody>{"".join(rows)}</tbody></table>'


def render_product_filter_toolbar(source_rows: List[dict[str, Any]]) -> str:
    products = sorted({str(row.get("product") or "").strip() for row in source_rows if str(row.get("product") or "").strip()})
    if not products:
        return ""

    buttons = [
        '<button class="product-toggle active" type="button" data-product="all" aria-pressed="true">All Products</button>'
    ]
    buttons.extend(
        f'<button class="product-toggle" type="button" data-product="{escape(product)}" aria-pressed="false">{escape(product)}</button>'
        for product in products
    )
    return f"""
    <div class="product-toolbar" role="tablist" aria-label="Product filter">
      <span class="product-toolbar-label">Products</span>
      {''.join(buttons)}
    </div>
    """


def render_tk_month_toolbar(month_keys: List[str]) -> str:
    selected_key = ""
    if month_keys:
        current_key = date.today().strftime("%Y-%m")
    selected_key = current_key if current_key in month_keys else month_keys[-1]
    options_html = (
        "".join(
            f'<option value="{escape(key)}"{" selected" if key == selected_key else ""}>{escape(format_month_key_label(key))}</option>'
            for key in month_keys
        )
        if month_keys
        else '<option value="">No months available</option>'
    )
    return f"""
    <div class="tk-month-toolbar" role="group" aria-label="TK Approved month filter">
      <label class="tk-month-toggle" for="tk-start-month-enabled">
        <input id="tk-start-month-enabled" class="tk-month-toggle-input" type="checkbox" aria-describedby="tk-start-month-toggle-desc">
        <span class="tk-month-toggle-track" aria-hidden="true"><span class="tk-month-toggle-thumb"></span></span>
        <span class="tk-month-toggle-text" id="tk-start-month-toggle-desc">For epics started in</span>
      </label>
      <label class="tk-month-toggle" for="tk-month-enabled">
        <input id="tk-month-enabled" class="tk-month-toggle-input" type="checkbox" aria-describedby="tk-month-toggle-desc">
        <span class="tk-month-toggle-track" aria-hidden="true"><span class="tk-month-toggle-thumb"></span></span>
        <span class="tk-month-toggle-text" id="tk-month-toggle-desc">For epics delivered in</span>
      </label>
      <label class="tk-month-toggle" for="tk-through-month-enabled">
        <input id="tk-through-month-enabled" class="tk-month-toggle-input" type="checkbox" aria-describedby="tk-through-month-toggle-desc">
        <span class="tk-month-toggle-track" aria-hidden="true"><span class="tk-month-toggle-thumb"></span></span>
        <span class="tk-month-toggle-text" id="tk-through-month-toggle-desc">Any Work Done Through</span>
      </label>
      <select id="tk-month-select" class="tk-month-select" aria-label="Target analysis month" disabled>{options_html}</select>
      <label class="tk-month-toggle" for="tk-jira-only-enabled">
        <input id="tk-jira-only-enabled" class="tk-month-toggle-input" type="checkbox" checked aria-describedby="tk-jira-only-toggle-desc">
        <span class="tk-month-toggle-track" aria-hidden="true"><span class="tk-month-toggle-thumb"></span></span>
        <span class="tk-month-toggle-text" id="tk-jira-only-toggle-desc">Only Jira Populated Epics</span>
      </label>
      <span class="tk-month-status" data-tk-month-status aria-live="polite"></span>
    </div>
    """


def render_tk_month_story_analysis_panel() -> str:
    """Render the month-scope estimate panel that compares adjacent months."""
    return """
    <section class="panel tk-month-analysis-panel" id="tk-month-analysis" hidden aria-live="polite">
      <div class="tk-month-analysis-header">
        <div>
          <h2>Month Story Estimate Analysis</h2>
          <div class="footnote">For epics in the active analysis scope, single-month stories contribute directly. Stories spanning multiple months fall back to subtask estimates; if no usable subtasks exist, the epic is excluded and listed below. With all toggles off, the chart shows all available months. Through mode uses inclusive epic month overlap.</div>
        </div>
        <div class="tk-month-analysis-header-side">
          <div class="tk-month-analysis-status" data-tk-month-analysis-status></div>
          <button class="tk-month-analysis-action" type="button" data-open-tk-analysis-drawer>See Epics</button>
        </div>
      </div>
      <div class="tk-month-analysis-cards">
      </div>
      <div class="tk-month-chart" role="img" aria-label="Month scope estimate bar chart">
        <div class="tk-month-chart-bars">
          <section class="tk-month-chart-bar-card" data-month-analysis-slot="previous">
            <div class="tk-month-chart-value duration-value" data-month-analysis-chart-value data-seconds="0" data-hours="0 h" data-days="0 d">0 h</div>
            <div class="tk-month-chart-bar-track">
              <div class="tk-month-chart-bar-fill" data-month-analysis-bar style="height: 0%"></div>
            </div>
            <div class="tk-month-chart-label" data-month-analysis-label>Previous Month</div>
          </section>
          <section class="tk-month-chart-bar-card featured" data-month-analysis-slot="selected">
            <div class="tk-month-chart-value duration-value" data-month-analysis-chart-value data-seconds="0" data-hours="0 h" data-days="0 d">0 h</div>
            <div class="tk-month-chart-bar-track">
              <div class="tk-month-chart-bar-fill" data-month-analysis-bar style="height: 0%"></div>
            </div>
            <div class="tk-month-chart-label" data-month-analysis-label>Selected Month</div>
          </section>
          <section class="tk-month-chart-bar-card" data-month-analysis-slot="next">
            <div class="tk-month-chart-value duration-value" data-month-analysis-chart-value data-seconds="0" data-hours="0 h" data-days="0 d">0 h</div>
            <div class="tk-month-chart-bar-track">
              <div class="tk-month-chart-bar-fill" data-month-analysis-bar style="height: 0%"></div>
            </div>
            <div class="tk-month-chart-label" data-month-analysis-label>Next Month</div>
          </section>
        </div>
      </div>
      <div class="tk-month-analysis-summary">
        <div class="tk-month-analysis-pill">
          <span>Included epics</span>
          <strong data-month-analysis-included>0</strong>
        </div>
        <div class="tk-month-analysis-pill excluded">
          <span>Excluded epics</span>
          <strong data-month-analysis-excluded>0</strong>
        </div>
      </div>
      <div class="tk-month-exclusion-block">
        <h3>Excluded Epics</h3>
        <div class="footnote">The table lists epics omitted from the chart because their cross-month story ranges could not be resolved into usable subtask estimates, or because no usable story/subtask estimate fell into the adjacent month window.</div>
        <div class="table-frame tk-month-exclusion-table-frame">
          <table class="tk-month-exclusion-table">
            <colgroup>
              <col class="tk-month-exclusion-col">
              <col class="tk-month-exclusion-col">
              <col class="tk-month-exclusion-col">
            </colgroup>
            <thead>
              <tr><th>Epic</th><th>Product</th><th>Reason</th></tr>
            </thead>
            <tbody data-month-analysis-exclusion-body>
              <tr><td colspan="3" class="empty-state tight">No excluded epics for the current month.</td></tr>
            </tbody>
          </table>
        </div>
      </div>
    </section>
    """


def working_days_2026() -> List[dict[str, Any]]:
    """Return a list of 2026 months with the count of Mon-Fri working days in each.

    Holidays are intentionally not subtracted here; the capacity calculator surfaces
    raw business-day availability so the consumer can decide how to discount it.
    """
    months: List[dict[str, Any]] = []
    for month_index in range(1, 13):
        start = date(2026, month_index, 1)
        end = date(2027, 1, 1) if month_index == 12 else date(2026, month_index + 1, 1)
        total = 0
        cursor = start
        one_day = timedelta(days=1)
        while cursor < end:
            if cursor.weekday() < 5:
                total += 1
            cursor += one_day
        months.append(
            {
                "key": f"2026-{month_index:02d}",
                "label": start.strftime("%B %Y"),
                "working_days": total,
            }
        )
    return months


def render_capacity_calculator() -> str:
    return """
    <div class="capacity-calculator" role="group" aria-label="Capacity calculator">
      <div class="capacity-calc-title">Capacity Calculator</div>
      <div class="capacity-calc-body">
        <label class="capacity-field">
          <span class="capacity-field-label">No. of Employees</span>
          <input id="capacity-employees" class="capacity-input" type="number" inputmode="numeric" min="0" step="1" value="1" aria-label="Number of employees">
        </label>
        <label class="capacity-field">
          <span class="capacity-field-label">Month of 2026</span>
          <select id="capacity-month" class="capacity-select" aria-label="Month of 2026"></select>
        </label>
        <label class="capacity-field">
          <span class="capacity-field-label" id="capacity-leaves-label" data-label-hours="Total Leaves (Hours)" data-label-days="Total Leaves (Days)">Total Leaves (Hours)</span>
          <input id="capacity-leaves" class="capacity-input" type="number" inputmode="decimal" min="0" step="1" value="0" aria-label="Total leaves in hours" data-seconds="0">
        </label>
        <section class="metric-card metric-card-capacity capacity-result-card" aria-live="polite">
          <div class="metric-label">Total Capacity</div>
          <div class="metric-value-wrap">
            <span class="metric-value duration-value" id="capacity-value" data-seconds="0" data-hours="0.00 h" data-days="0.00 d">0.00 h</span>
          </div>
          <div class="metric-meta" data-capacity-meta>Man-hours/days available</div>
        </section>
        <section class="metric-card metric-card-emerald capacity-result-card" aria-live="polite">
          <div class="metric-label">Total Availability</div>
          <div class="metric-value-wrap">
            <span class="metric-value duration-value" id="availability-value" data-seconds="0" data-hours="0.00 h" data-days="0.00 d">0.00 h</span>
          </div>
          <div class="metric-meta" data-availability-meta>Capacity minus leaves</div>
        </section>
      </div>
    </div>
    """


def render_drawer_modal() -> str:
    return """
    <aside class="drawer-modal" id="epic-drawer" hidden aria-hidden="true" role="dialog" aria-labelledby="epic-drawer-title" aria-modal="true">
      <div class="drawer-backdrop" data-drawer-close></div>
      <div class="drawer-panel" role="document">
        <header class="drawer-header">
          <div class="drawer-heading">
            <div class="drawer-eyebrow" data-drawer-eyebrow></div>
            <h2 id="epic-drawer-title">Selected Epics</h2>
            <div class="drawer-subtitle" data-drawer-subtitle></div>
          </div>
          <button class="drawer-close" type="button" data-drawer-close aria-label="Close drawer">&times;</button>
        </header>
        <div class="drawer-summary" data-drawer-summary></div>
        <div class="drawer-body" data-drawer-body></div>
      </div>
    </aside>
    """


def render_search_toolbar() -> str:
    return """
    <div class="search-toolbar" role="search">
      <label class="search-toolbar-label" for="epic-search">Epic Search</label>
      <input
        id="epic-search"
        class="search-input"
        type="search"
        placeholder="Filter by epic key, title, product, status, or priority"
        aria-label="Filter epics in gantt and table views"
        autocomplete="off"
      >
      <button class="search-clear" type="button" data-clear-search aria-label="Clear epic search">Clear</button>
    </div>
    <div class="search-status" data-search-status aria-live="polite"></div>
    """


def render_table_legend() -> str:
    items = [
        ("legend-epic", "Epics"),
        ("legend-story", "Stories"),
        ("legend-subtask", "Subtasks"),
        ("legend-bug-subtask", "Bug Subtasks"),
    ]
    chips = "".join(
        f'<span class="table-legend-item"><span class="table-legend-swatch {css_class}" aria-hidden="true"></span>{escape(label)}</span>'
        for css_class, label in items
    )
    return f'<div class="table-legend" aria-label="Table row color legend">{chips}</div>'


def render_worklog_table(worklogs: List[dict[str, Any]]) -> str:
    if not worklogs:
        return '<div class="empty-state tight">No worklogs found for this issue.</div>'
    rows = []
    for item in worklogs:
        rows.append(
            f"""
            <tr>
              <td>{escape(str(item.get('worklog_id') or ''))}</td>
              <td>{escape(str(item.get('author_display_name') or ''))}</td>
              <td>{escape(format_datetime(parse_iso_datetime(item.get('started'))))}</td>
              <td>{escape(str(item.get('time_spent') or ''))}</td>
              <td>{escape(format_hours(item.get('time_spent_seconds')))}</td>
            </tr>
            """
        )
    return f"""
    <table>
      <thead>
        <tr><th>Worklog ID</th><th>Author</th><th>Started</th><th>Time Spent</th><th>Hours</th></tr>
      </thead>
      <tbody>{''.join(rows)}</tbody>
    </table>
    """


def story_total_logged_seconds(story: dict[str, Any]) -> float:
    return float(sum(descendant.get("total_logged_seconds") or 0 for descendant in story["descendants"]))


def epic_total_logged_seconds(epic: dict[str, Any]) -> float:
    return float(sum(story_total_logged_seconds(story) for story in epic["stories"]))


def epic_story_rollup_seconds(epic: dict[str, Any]) -> float:
    """Return sum of story-level original estimates only, excluding subtask estimates."""
    return float(sum(parse_numeric(story.get("jira_original_estimate_seconds")) or 0 for story in epic["stories"]))


def epic_subtask_rollup_seconds(epic: dict[str, Any]) -> float:
    """Return sum of subtask original estimates across all stories in the epic."""
    total = 0.0
    for story in epic["stories"]:
        for descendant in story.get("descendants", []):
            if descendant.get("is_subtask"):
                total += parse_numeric(descendant.get("jira_original_estimate_seconds")) or 0.0
    return total


def render_epic_rollup_mismatch(epic: dict[str, Any]) -> str:
    rollup_seconds = epic_story_rollup_seconds(epic)
    jira_original_seconds = parse_numeric(epic.get("epic_original_estimate_seconds"))
    if jira_original_seconds is None:
        return render_comparison_note("No Jira original estimate available", "neutral")

    delta_seconds = rollup_seconds - jira_original_seconds
    if abs(delta_seconds) < 1:
        return render_comparison_note("Matches Jira original estimate", "match")

    delta_hours = abs(delta_seconds) / 3600
    if delta_seconds > 0:
        return render_comparison_note(f"Mismatch: {delta_hours:.2f} h above Jira original estimate", "over")
    return render_comparison_note(f"Mismatch: {delta_hours:.2f} h below Jira original estimate", "under")


def epic_rollup_state(epic: dict[str, Any]) -> str:
    rollup_seconds = epic_story_rollup_seconds(epic)
    jira_original_seconds = parse_numeric(epic.get("epic_original_estimate_seconds"))
    if jira_original_seconds is None:
        return "neutral"

    delta_seconds = rollup_seconds - jira_original_seconds
    if abs(delta_seconds) < 1:
        return "match"
    return "over" if delta_seconds > 0 else "under"


def render_story_detail_table(story: dict[str, Any]) -> str:
    descendants = story["descendants"]
    if not descendants:
        return '<div class="empty-state tight">No descendant issues were fetched for this story.</div>'

    rows: List[str] = [
        """
        <table class="hierarchy-table epic-table descendant-table">
          <colgroup>
            <col style="width: 42px">
            <col style="width: 54px">
            <col style="width: 280px">
            <col style="width: 110px">
            <col style="width: 110px">
            <col style="width: 120px">
            <col style="width: 120px">
            <col style="width: 120px">
            <col style="width: 120px">
            <col style="width: 140px">
            <col style="width: 150px">
            <col style="width: 170px">
          </colgroup>
          <tbody>
        """
    ]

    for index, descendant in enumerate(descendants):
        detail_id = f"descendant-detail-{escape(str(descendant.get('issue_key') or index))}"
        has_worklogs = bool(descendant["worklogs"])
        title = escape(str(descendant.get("summary") or ""))
        row_variant = descendant_row_variant(descendant)
        rows.append(
            f"""
            <tr class="descendant-summary-row descendant-summary-row-{row_variant}">
              <td class="sticky-col sticky-toggle">
                <button class="row-toggle descendant-toggle" type="button" data-target="{detail_id}" aria-controls="{detail_id}" aria-expanded="false"{" disabled" if not has_worklogs else ""}>{'+' if has_worklogs else ''}</button>
              </td>
              <td class="sticky-col sticky-link">{render_issue_link(descendant.get('issue_key'), f'Open {descendant.get("issue_key") or ""} in Jira')}</td>
              <td class="sticky-col sticky-title">
                <div class="epic-title-block descendant-title-block">
                  <div class="epic-title-text">{title}</div>
                  <div class="epic-metadata">
                    <span><span class="label">Type</span>{escape(str(descendant.get('issue_type') or ''))}</span>
                    <span><span class="label">Logged</span>{duration_span_logged(descendant.get('total_logged_seconds'))}</span>
                    <span><span class="label">Status</span><span class="epic-status">{escape(str(descendant.get('status') or ''))}</span></span>
                    <span><span class="label">Priority</span><span class="epic-priority">{escape(str(descendant.get('priority') or ''))}</span></span>
                  </div>
                </div>
              </td>
              <td>{escape(format_date(parse_iso_date(descendant.get('jira_start_date'))))}</td>
              <td>{escape(format_date(parse_iso_date(descendant.get('jira_due_date'))))}</td>
              <td></td>
              <td></td>
              <td></td>
              <td></td>
              <td></td>
              <td>{duration_span(descendant.get('jira_original_estimate_seconds'))}</td>
              <td>{escape(str(descendant.get('worklog_count') or 0))} worklogs</td>
            </tr>
            <tr class="descendant-detail-row" id="{detail_id}" hidden>
              <td colspan="12">
                <div class="descendant-detail-body">
                  {render_worklog_table(descendant["worklogs"])}
                </div>
              </td>
            </tr>
            """
        )

    rows.append(
        """
          </tbody>
        </table>
        """
    )
    return "".join(rows)


def render_story_table(stories: List[dict[str, Any]]) -> str:
    if not stories:
        return '<div class="empty-state">No child stories were fetched for this epic.</div>'

    rows = [
        """
        <table class="hierarchy-table epic-table story-table">
          <colgroup>
            <col style="width: 42px">
            <col style="width: 54px">
            <col style="width: 280px">
            <col style="width: 110px">
            <col style="width: 110px">
            <col style="width: 120px">
            <col style="width: 120px">
            <col style="width: 120px">
            <col style="width: 120px">
            <col style="width: 140px">
            <col style="width: 150px">
            <col style="width: 170px">
          </colgroup>
          <tbody>
        """
    ]

    for index, story in enumerate(stories):
        detail_id = f"story-detail-{escape(str(story.get('story_key') or index))}"
        has_descendants = bool(story["descendants"])
        title = escape(str(story.get("summary") or ""))
        rows.append(
            f"""
            <tr class="story-summary-row story-summary-row-story">
              <td class="sticky-col sticky-toggle">
                <button class="row-toggle story-toggle" type="button" data-target="{detail_id}" aria-controls="{detail_id}" aria-expanded="false"{" disabled" if not has_descendants else ""}>{'+' if has_descendants else ''}</button>
              </td>
              <td class="sticky-col sticky-link">{render_issue_link(story.get('story_key'), f'Open {story.get("story_key") or ""} in Jira')}</td>
              <td class="sticky-col sticky-title">
                <div class="epic-title-block story-title-block">
                  <div class="epic-title-text">{title}</div>
                  <div class="epic-metadata">
                    <span><span class="label">Type</span>{escape(str(story.get('issue_type') or ''))}</span>
                    <span><span class="label">Logged</span>{duration_span_logged(story_total_logged_seconds(story))}</span>
                    <span><span class="label">Status</span><span class="epic-status">{escape(str(story.get('status') or ''))}</span></span>
                    <span><span class="label">Priority</span><span class="epic-priority">{escape(str(story.get('priority') or ''))}</span></span>
                  </div>
                </div>
              </td>
              <td>{escape(format_date(parse_iso_date(story.get('jira_start_date'))))}</td>
              <td>{escape(format_date(parse_iso_date(story.get('jira_due_date'))))}</td>
              <td></td>
              <td></td>
              <td></td>
              <td></td>
              <td></td>
              <td>{duration_span(story.get('jira_original_estimate_seconds'))}</td>
              <td>{duration_span(story.get('jira_aggregate_original_estimate_seconds'))}</td>
            </tr>
            <tr class="story-detail-row" id="{detail_id}" hidden>
              <td colspan="12">
                <div class="story-detail-body">
                  {render_story_detail_table(story)}
                </div>
              </td>
            </tr>
            """
        )

    rows.append(
        """
          </tbody>
        </table>
        """
    )
    return "".join(rows)


def render_issue_link(issue_key: Any, label: str, issue_url: Any = None) -> str:
    issue_key_text = str(issue_key or "").strip()
    resolved_issue_url = jira_issue_url(issue_key_text) if issue_url is None else str(issue_url).strip()
    if not resolved_issue_url or resolved_issue_url == "#":
        return '<span class="jira-link-button disabled" aria-hidden="true" title="Jira link unavailable">Go</span>'
    return (
        f'<a class="jira-link-button" href="{escape(resolved_issue_url)}" target="_blank" rel="noreferrer" '
        f'aria-label="{escape(label)}" title="{escape(label)}">'
        '<svg viewBox="0 0 24 24" aria-hidden="true" focusable="false">'
        '<path d="M14 5h5v5" />'
        '<path d="M10 14L19 5" />'
        '<path d="M19 14v4a1 1 0 0 1-1 1h-12a1 1 0 0 1-1-1v-12a1 1 0 0 1 1-1h4" />'
        "</svg>"
        "</a>"
    )


def render_epic_table_view(source_rows: List[dict[str, Any]]) -> str:
    if not source_rows:
        return '<div class="empty-state">No epic rows exist in the database.</div>'

    rows: List[str] = [
        """
        <table class="hierarchy-table epic-table">
          <colgroup>
            <col style="width: 42px">
            <col style="width: 54px">
            <col style="width: 280px">
            <col style="width: 110px">
            <col style="width: 110px">
            <col style="width: 120px">
            <col style="width: 120px">
            <col style="width: 120px">
            <col style="width: 120px">
            <col style="width: 140px">
            <col style="width: 150px">
            <col style="width: 170px">
          </colgroup>
          <thead>
            <tr class="hierarchy-header" role="row">
              <th class="sticky-col sticky-toggle">Toggle</th>
              <th class="sticky-col sticky-link">Jira</th>
              <th class="sticky-col sticky-title">Epic Title</th>
              <th>Start</th>
              <th>Due</th>
              <th>Most Likely</th>
              <th>Optimistic</th>
              <th>Pessimistic</th>
              <th>Calculated</th>
              <th>TK Approved</th>
              <th>Jira Original Estimate</th>
              <th>Story Estimates</th>
            </tr>
          </thead>
          <tbody>
        """
    ]

    for index, epic in enumerate(source_rows):
        is_open = index == 0
        detail_id = f"epic-detail-{index}"
        title = escape(str(epic.get('epic_summary') or epic['roadmap_item']))
        search_text = escape(epic_search_text(epic))
        rows.append(
            f"""
            <tr class="epic-summary-row epic-summary-row-epic" data-product="{escape(epic['product'])}" data-search="{search_text}">
              <td class="sticky-col sticky-toggle">
                <button class="row-toggle" type="button" data-target="{detail_id}" aria-controls="{detail_id}" aria-expanded="{str(is_open).lower()}">{'-' if is_open else '+'}</button>
              </td>
              <td class="sticky-col sticky-link">{render_issue_link(epic['jira_id'], f'Open {epic["jira_id"]} in Jira', epic.get('jira_url'))}</td>
              <td class="sticky-col sticky-title">
                <div class="epic-title-block">
                  <div class="epic-title-text">{title}</div>
                  <div class="epic-metadata">
                    <span><span class="label">Logged</span>{duration_span_logged(epic_total_logged_seconds(epic))}</span>
                    <span><span class="label">Status</span><span class="epic-status">{escape(str(epic.get('epic_status') or ''))}</span></span>
                    <span><span class="label">Priority</span><span class="epic-priority">{escape(str(epic.get('epic_priority') or ''))}</span></span>
                  </div>
                </div>
              </td>
              <td>{escape(format_date(parse_iso_date(epic.get('epic_start_date'))))}</td>
              <td>{escape(format_date(parse_iso_date(epic.get('epic_due_date'))))}</td>
              <td>{man_day_span(epic.get('man_days_value'))}</td>
              <td>{man_day_span(epic.get('optimistic_50_value'))}</td>
              <td>{man_day_span(epic.get('pessimistic_10_value'))}</td>
              <td>{man_day_span(epic.get('est_formula_value'))}</td>
              <td class="summary-condition {tk_target_state(epic)}">
                {man_day_span(epic.get('tk_target_value'))}
                {render_tk_target_comparison(epic)}
              </td>
              <td>{duration_span(epic.get('epic_original_estimate_seconds'))}</td>
              <td class="summary-condition {epic_rollup_state(epic)}">
                {duration_span(epic_story_rollup_seconds(epic))}
                {render_epic_rollup_mismatch(epic)}
              </td>
            </tr>
            <tr class="epic-detail-row" id="{detail_id}" data-product="{escape(epic['product'])}" data-search="{search_text}"{'' if is_open else ' hidden'}>
              <td colspan="12">
                <div class="epic-detail-body">
                  <section class="pivot-section">
                    <h3>Stories</h3>
                    {render_story_table(epic['stories'])}
                  </section>
                </div>
              </td>
            </tr>
            """
        )

    rows.append("""
          </tbody>
        </table>
    """)
    return "".join(rows)


def render_error_table(errors: List[dict[str, Any]]) -> str:
    if not errors:
        return '<div class="empty-state">No extraction errors are present in the database.</div>'
    rows = []
    for error in errors:
        rows.append(
            f"""
            <tr>
              <td>{escape(str(error.get('error_scope') or ''))}</td>
              <td>{escape(str(error.get('issue_key') or ''))}</td>
              <td>{escape(str(error.get('sheet_name') or ''))}</td>
              <td>{escape(str(error.get('row_number') or ''))}</td>
              <td>{escape(str(error.get('message') or ''))}</td>
            </tr>
            """
        )
    return f"""
    <table>
      <thead><tr><th>Scope</th><th>Issue Key</th><th>Sheet</th><th>Row</th><th>Message</th></tr></thead>
      <tbody>{''.join(rows)}</tbody>
    </table>
    """


def render_html(report_data: dict[str, Any], db_path: Path) -> str:
    source_rows = report_data["source_rows"]
    errors = report_data["errors"]
    epic_metric_summary = build_epic_metric_summary(source_rows)
    epic_metric_summary_json = json.dumps(epic_metric_summary)
    epic_detail_records = build_epic_detail_records(source_rows)
    month_keys = available_month_keys_from_epic_details(epic_detail_records)
    epic_detail_records_json = json.dumps(epic_detail_records)
    story_detail_records_json = json.dumps(build_story_detail_records(source_rows))
    subtask_detail_records_json = json.dumps(build_subtask_detail_records(source_rows))
    capacity_months_json = json.dumps(working_days_2026())
    rmi_schedule_records = build_rmi_schedule_records(source_rows)
    rmi_schedule_initial_year = initial_schedule_year(rmi_schedule_records)
    rmi_schedule_initial_body_html, rmi_schedule_initial_foot_html = render_rmi_schedule_table_rows(
      rmi_schedule_records,
      rmi_schedule_initial_year,
    )
    rmi_schedule_records_json = json.dumps(rmi_schedule_records)
    rmi_schedule_years_json = json.dumps(available_years_from_schedule_records(rmi_schedule_records))
    gantt_rows = build_gantt_rows(source_rows)
    return f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>RMI Jira Gantt</title>
  <style>
    :root {{
      --bg: #e8eef5;
      --panel: #ffffff;
      --panel-soft: #f7fafd;
      --text: #0d1b2a;
      --muted: #516174;
      --line: #d0dbe6;
      --shadow: 0 1px 2px rgba(16,32,51,0.04), 0 4px 12px rgba(16,32,51,0.06), 0 16px 40px rgba(16,32,51,0.07);
      --shadow-sm: 0 1px 2px rgba(16,32,51,0.04), 0 4px 10px rgba(16,32,51,0.05);
      --shadow-lg: 0 2px 4px rgba(16,32,51,0.04), 0 8px 24px rgba(16,32,51,0.08), 0 28px 56px rgba(16,32,51,0.10);
      --radius-sm: 10px;
      --radius-md: 16px;
      --radius-lg: 22px;
      --ring: rgba(37, 99, 235, 0.26);
      --gutter: clamp(16px, 2.5vw, 40px);
      --row-epic: #f0e8ff;
      --row-epic-hover: #e6d8ff;
      --row-story: #ddf0fd;
      --row-story-hover: #c9e7fb;
      --row-subtask: #d8fae8;
      --row-subtask-hover: #c2f5d6;
      --row-bug-subtask: #fde8e8;
      --row-bug-subtask-hover: #fcd0d0;
    }}
    * {{ box-sizing: border-box; }}
    body {{ margin: 0; font-family: "Inter", "Segoe UI", system-ui, -apple-system, sans-serif;
      overflow-x: clip;
      -webkit-font-smoothing: antialiased;
      text-rendering: optimizeLegibility;
      background:
        radial-gradient(ellipse 90% 50% at top left, rgba(15,118,110,0.09), transparent 42%),
        radial-gradient(ellipse 70% 40% at top right, rgba(190,24,93,0.08), transparent 40%),
        radial-gradient(ellipse 100% 80% at center, rgba(37,99,235,0.03), transparent 70%),
        linear-gradient(180deg, #f2f7fc 0%, var(--bg) 100%);
      color: var(--text);
    }}
    .page {{ width: 100%; max-width: none; margin: 0; padding: 28px var(--gutter) 44px; }}
    header {{ margin-bottom: 18px; max-width: 1760px; }}
    h1 {{ margin: 0 0 8px; font-size: 2.2rem; letter-spacing: -0.04em; font-weight: 800; }}
    h2 {{ margin: 0 0 10px; font-size: 1.2rem; font-weight: 700; letter-spacing: -0.01em; }}
    h3 {{ margin: 0 0 10px; font-size: 1rem; font-weight: 700; }}
    .subtext {{ color: var(--muted); line-height: 1.55; max-width: 1160px; }}
    .metric-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(220px, 1fr)); gap: 14px; margin: 18px 0 22px; }}
    .product-summary-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(220px, 1fr)); gap: 14px; margin: -4px 0 22px; }}
    .metric-card, .panel {{ background: var(--panel); border: 1px solid rgba(200, 214, 228, 0.80); border-radius: var(--radius-lg); box-shadow: var(--shadow); }}
    .metric-card {{
      padding: 18px 20px;
      position: relative;
      overflow: hidden;
      border-color: rgba(255, 255, 255, 0.16);
      color: #fff;
      box-shadow: var(--shadow), inset 0 1px 0 rgba(255, 255, 255, 0.22);
    }}
    .metric-card::before {{
      content: "";
      position: absolute;
      inset: auto -18px -38px auto;
      width: 112px;
      height: 112px;
      border-radius: 999px;
      background: rgba(255, 255, 255, 0.14);
      pointer-events: none;
    }}
    .metric-card-teal {{ background: linear-gradient(135deg, #0f766e 0%, #14b8a6 100%); }}
    .metric-card-blue {{ background: linear-gradient(135deg, #1d4ed8 0%, #38bdf8 100%); }}
    .metric-card-amber {{ background: linear-gradient(135deg, #b45309 0%, #f59e0b 100%); }}
    .metric-card-rose {{ background: linear-gradient(135deg, #be123c 0%, #fb7185 100%); }}
    .metric-card-violet {{ background: linear-gradient(135deg, #6d28d9 0%, #a78bfa 100%); }}
    .metric-card-cyan {{ background: linear-gradient(135deg, #0f766e 0%, #67e8f9 100%); color: #062a30; }}
    .metric-card-emerald {{ background: linear-gradient(135deg, #047857 0%, #34d399 100%); }}
    .metric-card-indigo {{ background: linear-gradient(135deg, #4338ca 0%, #818cf8 100%); }}
    .metric-card-slate {{ background: linear-gradient(135deg, #334155 0%, #64748b 100%); }}
    .metric-label {{ color: rgba(255, 255, 255, 0.82); font-size: 0.92rem; margin-bottom: 8px; }}
    .metric-value-wrap {{ min-height: 42px; display: flex; align-items: center; }}
    .metric-value {{ font-size: 1.9rem; font-weight: 700; letter-spacing: -0.03em; color: #fff; }}
    .metric-meta {{ color: rgba(255, 255, 255, 0.82); margin-top: 6px; font-size: 0.92rem; max-width: 28ch; }}

    /* === Joined estimate cards (Optimistic -> Most Likely -> Pessimistic -> Calculated) === */
    /* Rendered as a single grouped row that spans the full metric-grid width.
       Individual cards share a container so borders/shadows form a single visual unit,
       with the blue hue stepping light -> dark to reinforce the sequence. */
    .estimate-cards-group {{
      grid-column: 1 / -1;
      display: grid;
      grid-template-columns: repeat(4, minmax(0, 1fr));
      gap: 0;
      border-radius: 18px;
      overflow: hidden;
      box-shadow: var(--shadow);
      border: 1px solid rgba(29, 78, 216, 0.18);
      position: relative;
    }}
    .estimate-cards-group::before {{
      content: "Estimation Range";
      position: absolute;
      top: 10px;
      left: 14px;
      font-size: 0.64rem;
      font-weight: 800;
      letter-spacing: 0.14em;
      text-transform: uppercase;
      color: rgba(255, 255, 255, 0.68);
      z-index: 2;
      pointer-events: none;
    }}
    .estimate-cards-group .metric-estimate-card {{
      border-radius: 0;
      box-shadow: none;
      border: 0;
      padding: 28px 20px 20px;
      position: relative;
    }}
    .estimate-cards-group .metric-estimate-card::before {{
      display: none; /* hide the decorative circle from base .metric-card */
    }}
    .estimate-cards-group .metric-estimate-card + .metric-estimate-card {{
      border-left: 1px solid rgba(255, 255, 255, 0.18);
    }}
    .estimate-cards-group .metric-estimate-card::after {{
      content: "";
      position: absolute;
      top: 44px;
      right: -9px;
      width: 18px;
      height: 18px;
      border-top: 2px solid rgba(255, 255, 255, 0.45);
      border-right: 2px solid rgba(255, 255, 255, 0.45);
      transform: rotate(45deg);
      pointer-events: none;
      z-index: 3;
    }}
    .estimate-cards-group .metric-estimate-card:last-child::after {{
      display: none;
    }}
    .estimate-cards-group .metric-label {{ color: rgba(255, 255, 255, 0.92); font-weight: 700; }}
    .estimate-cards-group .metric-meta {{ color: rgba(255, 255, 255, 0.78); }}
    /* Light -> dark blue progression */
    .metric-estimate-step-1 {{ background: linear-gradient(180deg, #7cb2fb 0%, #60a5fa 100%); }}
    .metric-estimate-step-2 {{ background: linear-gradient(180deg, #4f93f7 0%, #3b82f6 100%); }}
    .metric-estimate-step-3 {{ background: linear-gradient(180deg, #2f6fe6 0%, #1d4ed8 100%); }}
    .metric-estimate-step-4 {{ background: linear-gradient(180deg, #1d3fae 0%, #1e3a8a 100%); }}

    /* === TK Approved hero card (largest) === */
    .metric-card-hero {{
      grid-column: span 2;
      padding: 22px 26px 24px;
      min-height: 156px;
    }}
    .metric-card-hero .metric-label {{
      font-size: 1rem;
      font-weight: 800;
      letter-spacing: 0.02em;
      margin-bottom: 10px;
    }}
    .metric-card-hero .metric-value-wrap {{ min-height: 60px; }}
    .metric-card-hero .metric-value {{
      font-size: 2.6rem;
      font-weight: 800;
    }}
    .metric-card-hero .metric-meta {{
      font-size: 0.95rem;
      max-width: 44ch;
    }}
    .metric-card-hero::before {{
      width: 150px;
      height: 150px;
      inset: auto -24px -50px auto;
    }}
    .product-summary-card {{
      background: linear-gradient(180deg, #ffffff 0%, #f7fafd 100%);
      border: 1px solid rgba(200, 214, 228, 0.80);
      border-left: 5px solid var(--product-accent);
      border-radius: var(--radius-lg);
      box-shadow: var(--shadow), inset 0 1px 0 rgba(255, 255, 255, 0.90);
      padding: 16px 18px;
      cursor: pointer;
      transition: transform 0.15s ease, box-shadow 0.15s ease, border-color 0.15s ease, background 0.15s ease;
    }}
    .product-summary-card:hover {{
      transform: translateY(-2px);
      box-shadow: var(--shadow-lg);
    }}
    .product-summary-card:focus-visible {{
      outline: 3px solid rgba(37, 99, 235, 0.24);
      outline-offset: 2px;
    }}
    .product-summary-card.active {{
      border-color: color-mix(in srgb, var(--product-accent) 45%, white);
      background: linear-gradient(180deg, color-mix(in srgb, var(--product-accent) 10%, white) 0%, color-mix(in srgb, var(--product-accent) 4%, #f8fbff) 100%);
      box-shadow: 0 0 0 2px color-mix(in srgb, var(--product-accent) 18%, white), 0 20px 36px rgba(16, 32, 51, 0.12);
    }}
    .product-summary-label {{
      color: var(--text);
      font-size: 0.92rem;
      font-weight: 800;
      margin-bottom: 8px;
    }}
    .product-summary-value {{
      min-height: 38px;
      display: flex;
      align-items: center;
    }}
    .product-summary-duration {{
      font-size: 1.55rem;
      font-weight: 800;
      color: var(--product-accent);
      letter-spacing: -0.03em;
    }}
    .product-summary-meta {{
      color: var(--muted);
      font-size: 0.88rem;
      margin-top: 4px;
    }}
    .panel {{ padding: 20px 22px 24px; margin-bottom: 20px; box-shadow: var(--shadow), inset 0 1px 0 rgba(255, 255, 255, 0.88); }}
    .panel.inset {{ box-shadow: none; background: var(--panel-soft); margin-bottom: 14px; }}
    .tk-approved-panel {{
      border-color: rgba(59, 130, 246, 0.22);
      background:
        radial-gradient(circle at top right, rgba(59, 130, 246, 0.10), transparent 28%),
        linear-gradient(180deg, #fdfefe 0%, #f6f9ff 100%);
      box-shadow: 0 14px 30px rgba(37, 99, 235, 0.08);
    }}
    .tk-approved-panel h3 {{
      font-size: 1.08rem;
      letter-spacing: 0.01em;
      margin-bottom: 14px;
    }}
    .tk-approved-grid {{
      grid-template-columns: repeat(auto-fit, minmax(240px, 1fr));
      gap: 14px;
    }}
    .tk-approved-card {{
      min-height: 92px;
      padding: 16px 16px;
      font-size: 1rem;
    }}
    .tk-approved-card strong {{
      font-size: 0.78rem;
    }}
    .tk-approved-card.primary {{
      grid-column: span 2;
      min-height: 120px;
      border-color: rgba(245, 158, 11, 0.55);
      background: linear-gradient(180deg, #fffdf5 0%, #fff7e8 100%);
    }}
    .tk-caption-legend {{
      margin: -2px 0 14px;
      padding: 10px 14px;
      border: 1px solid rgba(59, 130, 246, 0.18);
      border-radius: 12px;
      background: rgba(255, 255, 255, 0.72);
      color: #274060;
      font-weight: 700;
      line-height: 1.6;
    }}
    .tk-caption {{
      display: block;
      margin-top: 6px;
      color: #1e3a8a;
      font-size: 0.98rem;
      font-weight: 800;
    }}
    .unit-caption {{
      display: block;
    }}
    .tk-approved-card.primary .tk-caption {{
      color: #9a3412;
      font-size: 1.02rem;
    }}
    .view-toolbar {{ display: flex; gap: 10px; flex-wrap: wrap; margin: 10px 0 18px; }}
    .search-toolbar {{
      display: flex;
      gap: 12px;
      flex-wrap: wrap;
      align-items: center;
      margin: 18px 0 12px;
      padding: 14px 18px;
      border: 1px solid rgba(200, 214, 228, 0.75);
      border-radius: var(--radius-md);
      background: rgba(255, 255, 255, 0.90);
      box-shadow: var(--shadow-sm), inset 0 1px 0 rgba(255, 255, 255, 0.98);
    }}
    .search-toolbar-label {{
      font-size: 0.82rem;
      font-weight: 800;
      text-transform: uppercase;
      letter-spacing: 0.04em;
      color: #52657a;
    }}
    .search-input {{
      flex: 1 1 360px;
      min-width: 240px;
      padding: 11px 14px;
      border: 1px solid rgba(192, 206, 218, 0.90);
      border-radius: var(--radius-sm);
      background: #fff;
      color: var(--text);
      font: inherit;
      box-shadow: 0 1px 3px rgba(16, 32, 51, 0.06);
      transition: border-color 0.15s ease, box-shadow 0.15s ease;
    }}
    .search-input:focus {{
      outline: none;
      border-color: #2563eb;
      box-shadow: 0 0 0 3px var(--ring), 0 1px 3px rgba(16, 32, 51, 0.06);
    }}
    .search-clear {{
      appearance: none;
      border: 1px solid rgba(192, 206, 218, 0.90);
      background: #fff;
      color: #203141;
      border-radius: 999px;
      padding: 9px 14px;
      font-size: 0.84rem;
      font-weight: 700;
      cursor: pointer;
      box-shadow: 0 1px 3px rgba(16, 32, 51, 0.07), inset 0 1px 0 rgba(255, 255, 255, 0.95);
      transition: transform 0.12s ease, box-shadow 0.12s ease;
    }}
    .search-clear:hover {{ transform: translateY(-1px); box-shadow: 0 4px 10px rgba(16, 32, 51, 0.10); }}
    .search-status {{
      min-height: 1.25rem;
      margin: 0 0 12px;
      color: #5c6f83;
      font-size: 0.9rem;
      font-weight: 700;
    }}
    .view-toggle {{
      appearance: none; border: 1px solid rgba(200, 214, 228, 0.90); background: #fff; color: var(--text);
      border-radius: 999px; padding: 10px 16px; font-weight: 700; cursor: pointer;
      box-shadow: 0 1px 3px rgba(16, 32, 51, 0.07), inset 0 1px 0 rgba(255, 255, 255, 0.95);
      transition: transform 0.15s ease, background 0.15s ease, border-color 0.15s ease, box-shadow 0.15s ease;
    }}
    .view-toggle:hover {{ transform: translateY(-1px); box-shadow: 0 4px 12px rgba(16, 32, 51, 0.10), inset 0 1px 0 rgba(255, 255, 255, 0.95); }}
    .view-toggle.active {{ background: #0d1b2a; color: #fff; border-color: #0d1b2a; box-shadow: 0 2px 8px rgba(13, 27, 42, 0.30), inset 0 1px 0 rgba(255, 255, 255, 0.10); }}
    .view-section[hidden] {{ display: none; }}
    #table-view.panel {{
      padding: 0;
      border: 0;
      border-radius: 0;
      box-shadow: none;
      background: transparent;
      margin-left: calc(-1 * var(--gutter));
      margin-right: calc(-1 * var(--gutter));
      width: calc(100% + (2 * var(--gutter)));
    }}
    #table-view h2,
    #table-view .footnote,
    #table-view .product-toolbar,
    #table-view .table-legend {{
      margin-left: var(--gutter);
      margin-right: var(--gutter);
    }}
    #gantt-view.panel {{
      margin-left: calc(-1 * var(--gutter));
      margin-right: calc(-1 * var(--gutter));
      width: calc(100% + (2 * var(--gutter)));
      border-radius: 0;
      border-left: 0;
      border-right: 0;
    }}
    #gantt-view h2,
    #gantt-view .product-toolbar {{
      margin-left: var(--gutter);
      margin-right: var(--gutter);
    }}
    .unit-toolbar {{ display: flex; gap: 10px; flex-wrap: wrap; margin: 0 0 12px; }}
    .unit-toggle {{
      appearance: none; border: 1px solid rgba(200, 214, 228, 0.90); background: #fff; color: var(--text);
      border-radius: 999px; padding: 8px 14px; font-weight: 700; cursor: pointer;
      box-shadow: 0 1px 3px rgba(16, 32, 51, 0.07), inset 0 1px 0 rgba(255, 255, 255, 0.95);
      transition: transform 0.15s ease, background 0.15s ease, border-color 0.15s ease, box-shadow 0.15s ease;
    }}
    .unit-toggle:hover {{ transform: translateY(-1px); box-shadow: 0 4px 10px rgba(16, 32, 51, 0.10); }}
    .unit-toggle.active {{ background: #1d4ed8; color: #fff; border-color: #1d4ed8; box-shadow: 0 2px 8px rgba(29, 78, 216, 0.32), inset 0 1px 0 rgba(255, 255, 255, 0.15); }}
    .footnote {{ color: var(--muted); line-height: 1.5; font-size: 0.92rem; }}
    .chip-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(210px, 1fr)); gap: 12px; }}
    .info-chip {{ background: #fff; border: 1px solid rgba(200, 214, 228, 0.80); border-radius: var(--radius-sm); padding: 14px 16px; display: grid; gap: 4px; min-height: 68px; box-shadow: var(--shadow-sm), inset 0 1px 0 rgba(255, 255, 255, 0.90); }}
    .info-chip strong {{ color: var(--muted); font-size: 0.82rem; text-transform: uppercase; letter-spacing: 0.03em; }}
    .tk-target-chip.match {{ border-color: #22c55e; background: #f0fdf4; }}
    .tk-target-chip.over {{ border-color: #ef4444; background: #fef2f2; }}
    .tk-target-chip.under {{ border-color: #f59e0b; background: #fffbeb; }}
    .tk-target-chip.neutral {{ border-color: #cbd5e1; background: #fbfdff; }}
    .comparison-note {{
      position: relative;
      display: inline-flex;
      align-items: center;
      justify-content: center;
      width: 16px;
      height: 16px;
      margin-left: 6px;
      border: 1px solid currentColor;
      border-radius: 999px;
      font-size: 0.68rem;
      font-weight: 800;
      line-height: 1;
      cursor: help;
      vertical-align: middle;
      background: #fff;
    }}
    .comparison-note::after {{
      content: attr(data-note);
      position: absolute;
      left: 50%;
      bottom: calc(100% + 8px);
      transform: translateX(-50%);
      min-width: 220px;
      max-width: 320px;
      padding: 8px 10px;
      border: 1px solid #cbd5e1;
      border-radius: 8px;
      background: #fff;
      box-shadow: 0 10px 24px rgba(16, 32, 51, 0.14);
      color: #102033;
      font-size: 0.74rem;
      font-weight: 700;
      line-height: 1.35;
      white-space: normal;
      text-align: left;
      opacity: 0;
      visibility: hidden;
      pointer-events: none;
      z-index: 20;
    }}
    .comparison-note:hover::after,
    .comparison-note:focus-visible::after {{
      opacity: 1;
      visibility: visible;
    }}
    .comparison-note.match {{ color: #15803d; }}
    .comparison-note.over {{ color: #b91c1c; }}
    .comparison-note.under {{ color: #b45309; }}
    .comparison-note.neutral {{ color: #64748b; }}
    .epic-accordion, .story-accordion, .nested-accordion {{ border: 1px solid var(--line); border-radius: 16px; background: var(--panel); overflow: hidden; margin-top: 14px; }}
    .epic-summary, .story-summary, .nested-accordion summary {{
      list-style: none; display: grid; gap: 12px; align-items: center; cursor: pointer;
    }}
    .hierarchy-table {{
      --epic-cols: 42px 54px minmax(340px, 2.4fr) 120px 120px 130px 140px 140px 140px 140px 170px 170px;
      --epic-frozen-width: 436px;
      width: max-content;
      min-width: 100%;
    }}
    .epic-summary {{
      grid-template-columns: 42px minmax(var(--epic-frozen-width), 2.8fr) 120px 120px 130px 140px 140px 140px 140px 170px 170px;
      width: max-content;
      min-width: 100%;
      padding: 16px 18px; background: #f8fbff;
    }}
    .story-summary {{
      grid-template-columns: 110px minmax(240px, 2fr) 110px 100px 100px 120px 110px 110px 110px;
      padding: 12px 14px; background: #fff;
    }}
    .nested-accordion summary {{
      grid-template-columns: minmax(260px, 2fr) 120px 120px;
      padding: 12px 14px; background: #fff;
    }}
    .epic-summary::-webkit-details-marker, .story-summary::-webkit-details-marker, .nested-accordion summary::-webkit-details-marker {{ display: none; }}
    .epic-toggle {{ font-size: 0; color: var(--muted); }}
    .epic-toggle::before {{ content: "\\25BE"; font-size: 1.15rem; }}
    .story-key {{ font-weight: 800; letter-spacing: 0.01em; }}
    .epic-title, .story-title {{ font-weight: 700; }}
    .epic-meta, .story-meta, .summary-meta {{ color: var(--muted); white-space: nowrap; }}
    .epic-meta-stack {{
      display: grid;
      gap: 4px;
      align-items: start;
      white-space: normal;
      min-width: 0;
    }}
    .epic-meta-stack .duration-value {{
      color: var(--text);
      font-weight: 700;
    }}
    .summary-condition {{
      padding: 10px 12px;
      border: 1px solid transparent;
      border-radius: 12px;
      background: rgba(255, 255, 255, 0.78);
    }}
    .summary-condition.match {{ border-color: rgba(34, 197, 94, 0.45); background: #f0fdf4; }}
    .summary-condition.over {{ border-color: rgba(239, 68, 68, 0.45); background: #fef2f2; }}
    .summary-condition.under {{ border-color: rgba(245, 158, 11, 0.5); background: #fffbeb; }}
    .summary-condition.neutral {{ border-color: rgba(203, 213, 225, 0.65); background: #fbfdff; }}
    .summary-condition {{
      white-space: normal;
      min-width: 0;
    }}
    .epic-body, .story-body {{ padding: 0 18px 18px; }}
    .hierarchy-header {{
      display: grid;
      grid-template-columns: var(--epic-cols);
      width: max-content;
      min-width: 100%;
      gap: 12px;
      padding: 13px 18px;
      position: sticky;
      top: 0;
      z-index: 6;
      background: linear-gradient(180deg, #edf3f9 0%, #e3ecf5 100%);
      border-bottom: 1px solid rgba(193, 208, 222, 0.90);
      font-size: 0.82rem;
      font-weight: 800;
      text-transform: uppercase;
      letter-spacing: 0.04em;
      color: #3d5673;
    }}
    .sticky-col {{
      position: sticky;
      z-index: 3;
      background: inherit;
    }}
    .sticky-toggle {{ left: 18px; }}
    .sticky-link {{ left: 70px; }}
    .sticky-title {{
      left: 134px;
      box-shadow: 18px 0 20px -18px rgba(16, 32, 51, 0.18);
    }}
    .epic-summary > .epic-toggle {{ display: none; }}
    .epic-primary {{
      display: contents;
    }}
    .sticky-epic {{
      position: sticky;
      left: 0;
      z-index: 4;
      background: inherit;
      box-shadow: 14px 0 16px -14px rgba(16, 32, 51, 0.18);
    }}
    .epic-link-cell {{ display: flex; align-items: center; padding: 0 10px; }}
    .jira-link-button {{
      width: 34px;
      height: 34px;
      display: inline-flex;
      align-items: center;
      justify-content: center;
      border: 1px solid rgba(29, 78, 216, 0.22);
      border-radius: 9px;
      background: #fff;
      color: #1d4ed8;
      text-decoration: none;
      box-shadow: 0 1px 3px rgba(16, 32, 51, 0.08), inset 0 1px 0 rgba(255, 255, 255, 0.90);
      transition: transform 0.15s ease, border-color 0.15s ease, background 0.15s ease, box-shadow 0.15s ease;
    }}
    .jira-link-button:hover {{
      transform: translateY(-1px);
      border-color: rgba(29, 78, 216, 0.55);
      background: #eff6ff;
      box-shadow: 0 4px 10px rgba(29, 78, 216, 0.14), inset 0 1px 0 rgba(255, 255, 255, 0.90);
    }}
    .jira-link-button.disabled {{
      color: #94a3b8;
      border-color: rgba(148, 163, 184, 0.35);
      background: #f8fafc;
    }}
    .jira-link-button svg {{
      width: 17px;
      height: 17px;
      stroke: currentColor;
      stroke-width: 2;
      stroke-linecap: round;
      stroke-linejoin: round;
      fill: none;
      margin: 0;
    }}
    .nested-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(180px, 1fr)); gap: 10px; margin-bottom: 12px; }}
    table {{ width: 100%; border-collapse: collapse; margin-top: 10px; font-size: 0.93rem; }}
    th, td {{ padding: 10px 12px; border-bottom: 1px solid var(--line); text-align: left; vertical-align: top; }}
    th {{ background: #f8fbff; text-transform: uppercase; font-size: 0.81rem; letter-spacing: 0.03em; color: #4f6278; }}
    tbody tr:hover {{ background: #f9fbff; }}
    .empty-state {{ padding: 16px; border: 1px dashed rgba(192, 210, 226, 0.80); border-radius: var(--radius-sm); color: var(--muted); background: rgba(251, 253, 255, 0.90); }}
    .empty-state.tight {{ margin-top: 10px; padding: 12px; }}
    svg {{ width: 100%; height: auto; display: block; margin-top: 12px; }}
    .gantt-grid {{ stroke: #d8e1eb; stroke-dasharray: 3 5; }}
    .gantt-axis {{ font-size: 10px; fill: #6b7d90; font-weight: 700; }}
    .gantt-month-axis {{ font-size: 11px; fill: #102033; font-weight: 800; }}
    .gantt-month-line {{ stroke: #b9c9da; stroke-width: 1.2; }}
    .gantt-month-band {{ fill: rgba(148, 163, 184, 0.04); }}
    .gantt-month-band.alt {{ fill: rgba(59, 130, 246, 0.05); }}
    .gantt-track {{ fill: #edf2f7; }}
    .gantt-label {{ font-size: 12px; font-weight: 800; fill: #203141; }}
    .gantt-meta {{ font-size: 11px; fill: #627487; }}
    .gantt-meta-left {{ font-size: 11px; fill: #5f7387; }}
    .gantt-date-label {{ font-size: 10px; fill: #51667a; font-weight: 700; }}
    .epic-bar {{ fill-opacity: 0.95; }}
    .story-bar {{ fill-opacity: 0.70; }}
    .child-bar {{ fill-opacity: 0.45; }}
    .gantt-product-grid {{
      display: grid;
      gap: 16px;
    }}
    .gantt-product-section {{
      border: 1px solid rgba(200, 214, 228, 0.75);
      border-radius: var(--radius-md);
      padding: 16px 16px 10px;
      background: rgba(252, 253, 255, 0.96);
      box-shadow: var(--shadow-sm), inset 0 1px 0 rgba(255, 255, 255, 0.90);
    }}
    .gantt-product-section h3 {{
      margin-bottom: 12px;
    }}
    .gantt-product-section[data-hidden="true"] {{
      display: none;
    }}
    .table-frame {{
      width: 100%;
      overflow: auto;
      border: 0;
      border-radius: 0;
      background: transparent;
      box-shadow: none;
      padding: 0 20px 0;
    }}
    .epic-table {{
      width: max(100%, 1466px);
      border-collapse: collapse;
      table-layout: fixed;
      background: #fff;
    }}
    .epic-table .hierarchy-header {{
      display: table-row;
      width: auto;
      min-width: 0;
      position: static;
      background: transparent;
      border: 0;
    }}
    .epic-table th,
    .epic-table td {{
      padding: 10px 12px;
      border: 1px solid #d7e0ea;
      text-align: left;
      vertical-align: top;
    }}
    .epic-table thead th {{
      position: sticky;
      z-index: 18;
      background: linear-gradient(180deg, #d4e2f0 0%, #c0d2e6 100%);
      border-bottom-color: #a6bcd0;
      box-shadow: inset 0 -1px 0 #a6bcd0, inset 0 1px 0 rgba(255, 255, 255, 0.28);
      text-transform: uppercase;
      font-size: 0.78rem;
      letter-spacing: 0.05em;
      color: #3d5673;
    }}
    .epic-table .sticky-col {{
      position: sticky;
      z-index: 3;
      background: #fff;
    }}
    .epic-table thead .sticky-col {{
      background: linear-gradient(180deg, #dbe5f1 0%, #c7d4e4 100%);
      z-index: 22;
    }}
    .epic-table .sticky-toggle {{ left: 0; width: 42px; }}
    .epic-table .sticky-link {{ left: 42px; width: 54px; }}
    .epic-table .sticky-title {{
      left: 96px;
      width: 280px;
      min-width: 280px;
      max-width: 280px;
      box-shadow: 14px 0 16px -14px rgba(16, 32, 51, 0.18);
    }}
    .epic-summary-row {{
      background: #f8fbff;
    }}
    .epic-summary-row:hover {{
      background: #f3f8ff;
    }}
    .epic-detail-row td {{
      padding: 0;
      background: #fff;
    }}
    .epic-detail-body {{
      padding: 0;
      background: #fff;
    }}
    .story-table {{
      margin-top: 0;
      border-top: 1px solid #d7e0ea;
    }}
    .story-summary-row {{
      background: #fcfdff;
    }}
    .story-summary-row:hover {{
      background: #f5f8fd;
    }}
    .story-summary-row .sticky-col {{
      background: #fcfdff;
    }}
    .story-summary-row:hover .sticky-col {{
      background: #f5f8fd;
    }}
    .story-summary-row .sticky-title {{
      left: 96px;
    }}
    .story-summary-row td {{
      font-size: 0.9rem;
    }}
    .story-detail-row td {{
      padding: 0;
      background: #fbfcfe;
    }}
    .story-detail-body {{
      padding: 0;
      border-top: 1px solid #e2e8f0;
      background: #fbfcfe;
    }}
    .story-toggle[disabled] {{
      opacity: 0.35;
      cursor: default;
    }}
    .descendant-table {{
      width: 100%;
      margin-top: 0;
      border-top: 0;
    }}
    .descendant-summary-row {{
      background: #ffffff;
    }}
    .descendant-summary-row:hover {{
      background: #f7faff;
    }}
    .descendant-summary-row .sticky-col {{
      background: #ffffff;
    }}
    .descendant-summary-row:hover .sticky-col {{
      background: #f7faff;
    }}
    .descendant-summary-row .sticky-title {{
      left: 96px;
    }}
    .descendant-summary-row td {{
      font-size: 0.88rem;
    }}
    .descendant-detail-row td {{
      padding: 0;
      background: #fcfdff;
    }}
    .descendant-detail-body {{
      padding: 10px 12px 12px 96px;
      border-top: 1px solid #e2e8f0;
      background: #fcfdff;
    }}
    .descendant-toggle[disabled] {{
      opacity: 0.35;
      cursor: default;
    }}
    .row-toggle {{
      appearance: none;
      width: 28px;
      height: 28px;
      border: 1px solid rgba(192, 206, 218, 0.90);
      border-radius: 7px;
      background: #fff;
      color: #315b8a;
      font-size: 1rem;
      font-weight: 800;
      line-height: 1;
      cursor: pointer;
      box-shadow: 0 1px 3px rgba(16, 32, 51, 0.08), inset 0 1px 0 rgba(255, 255, 255, 0.90);
      transition: background 0.12s ease, box-shadow 0.12s ease;
    }}
    .row-toggle:hover {{ background: #f0f5fb; }}
    .epic-title-block {{
      display: grid;
      gap: 6px;
      min-width: 0;
    }}
    .epic-title-text {{
      font-size: 0.98rem;
      font-weight: 800;
      line-height: 1.3;
      color: #102033;
      white-space: normal;
      overflow-wrap: anywhere;
      word-break: break-word;
      hyphens: auto;
    }}
    .product-toolbar {{
      display: flex;
      gap: 8px;
      flex-wrap: wrap;
      margin: 0 0 12px;
      padding: 10px 14px;
      border: 1px solid rgba(200, 214, 228, 0.80);
      border-radius: var(--radius-sm);
      background: rgba(248, 250, 255, 0.92);
      box-shadow: var(--shadow-sm), inset 0 1px 0 rgba(255, 255, 255, 0.90);
    }}
    .product-toolbar-label {{
      align-self: center;
      margin-right: 6px;
      font-size: 0.82rem;
      font-weight: 800;
      text-transform: uppercase;
      letter-spacing: 0.04em;
      color: #52657a;
    }}
    .product-toggle {{
      appearance: none;
      border: 1px solid rgba(192, 206, 218, 0.90);
      background: #fff;
      color: #203141;
      border-radius: 999px;
      padding: 7px 12px;
      font-size: 0.84rem;
      font-weight: 700;
      cursor: pointer;
      box-shadow: 0 1px 2px rgba(16, 32, 51, 0.06), inset 0 1px 0 rgba(255, 255, 255, 0.95);
      transition: transform 0.12s ease, background 0.12s ease, border-color 0.12s ease, box-shadow 0.12s ease;
    }}
    .product-toggle:hover {{ transform: translateY(-1px); box-shadow: 0 3px 8px rgba(16, 32, 51, 0.10); }}
    .product-toggle.active {{
      background: #0d1b2a;
      color: #fff;
      border-color: #0d1b2a;
      box-shadow: 0 2px 6px rgba(13, 27, 42, 0.30), inset 0 1px 0 rgba(255, 255, 255, 0.10);
    }}
    .epic-metadata {{
      display: flex;
      flex-wrap: wrap;
      gap: 8px;
      align-items: flex-start;
      font-size: 0.78rem;
      font-weight: 700;
      color: #5b7086;
      white-space: normal;
    }}
    .epic-metadata .label {{
      text-transform: uppercase;
      letter-spacing: 0.04em;
      font-size: 0.72rem;
      color: #7a8d9d;
    }}
    .epic-metadata > span {{
      display: grid;
      gap: 2px;
      padding-right: 8px;
      margin-right: 4px;
      border-right: 1px solid #dbe3ec;
    }}
    .epic-metadata > span:last-child {{
      border-right: 0;
      padding-right: 0;
      margin-right: 0;
    }}
    .table-legend {{
      display: flex;
      gap: 10px;
      flex-wrap: wrap;
      margin: 12px 20px 12px;
      padding: 12px 14px;
      border: 1px solid rgba(200, 214, 228, 0.75);
      border-radius: var(--radius-sm);
      background: rgba(248, 250, 255, 0.90);
      box-shadow: var(--shadow-sm);
    }}
    .table-legend-item {{
      display: inline-flex;
      align-items: center;
      gap: 8px;
      padding: 6px 10px;
      border-radius: 999px;
      background: #fff;
      border: 1px solid rgba(210, 222, 232, 0.90);
      font-size: 0.84rem;
      font-weight: 700;
      color: #203141;
      box-shadow: 0 1px 2px rgba(16, 32, 51, 0.05), inset 0 1px 0 rgba(255, 255, 255, 0.90);
    }}
    .table-legend-swatch {{
      width: 12px;
      height: 12px;
      border-radius: 999px;
      border: 1px solid rgba(16, 32, 51, 0.12);
      flex: 0 0 auto;
    }}
    .legend-epic {{ background: var(--row-epic); }}
    .legend-story {{ background: var(--row-story); }}
    .legend-subtask {{ background: var(--row-subtask); }}
    .legend-bug-subtask {{ background: var(--row-bug-subtask); }}
    .epic-status,
    .epic-priority {{
      font-size: 0.82rem;
      font-weight: 700;
      color: #5b7086;
    }}
    .epic-summary-row-epic,
    .epic-summary-row-epic .sticky-col {{
      background: var(--row-epic);
    }}
    .epic-summary-row-epic:hover,
    .epic-summary-row-epic:hover .sticky-col {{
      background: var(--row-epic-hover);
    }}
    .story-summary-row-story,
    .story-summary-row-story .sticky-col {{
      background: var(--row-story);
    }}
    .story-summary-row-story:hover,
    .story-summary-row-story:hover .sticky-col {{
      background: var(--row-story-hover);
    }}
    .descendant-summary-row-subtask,
    .descendant-summary-row-subtask .sticky-col {{
      background: var(--row-subtask);
    }}
    .descendant-summary-row-subtask:hover,
    .descendant-summary-row-subtask:hover .sticky-col {{
      background: var(--row-subtask-hover);
    }}
    .descendant-summary-row-bug-subtask,
    .descendant-summary-row-bug-subtask .sticky-col {{
      background: var(--row-bug-subtask);
    }}
    .descendant-summary-row-bug-subtask:hover,
    .descendant-summary-row-bug-subtask:hover .sticky-col {{
      background: var(--row-bug-subtask-hover);
    }}
    .epic-summary-row[data-hidden="true"],
    .epic-detail-row[data-hidden="true"] {{
      display: none;
    }}
    @media (max-width: 900px) {{
      .page {{ width: 100%; max-width: none; padding: 24px 12px 40px; }}
      .estimate-cards-group {{ grid-template-columns: repeat(2, minmax(0, 1fr)); }}
      .estimate-cards-group .metric-estimate-card::after {{ display: none; }}
      .estimate-cards-group .metric-estimate-card:nth-child(even) {{ border-left: 1px solid rgba(255, 255, 255, 0.18); }}
      .estimate-cards-group .metric-estimate-card:nth-child(n+3) {{ border-top: 1px solid rgba(255, 255, 255, 0.18); }}
      .metric-card-hero {{ grid-column: auto; padding: 18px 20px; min-height: 0; }}
      .metric-card-hero .metric-value {{ font-size: 2rem; }}
      .metric-card-hero .metric-value-wrap {{ min-height: 42px; }}
      table {{ display: block; overflow-x: auto; }}
      .epic-summary, .story-summary, .nested-accordion summary, .hierarchy-header {{ grid-template-columns: 1fr; width: 100%; min-width: 0; }}
      .hierarchy-table {{ width: 100%; min-width: 0; }}
      .sticky-col {{ position: static; box-shadow: none; }}
      .epic-table .hierarchy-header {{ display: table-row; }}
      .epic-table .sticky-col {{ position: static; box-shadow: none; }}
      .epic-table thead th {{ position: static; }}
      .epic-metadata {{ display: grid; }}
      .story-detail-body {{ padding-left: 0; }}
      .descendant-detail-body {{ padding-left: 12px; }}
      #table-view.panel,
      #gantt-view.panel {{
        margin-left: -12px;
        margin-right: -12px;
        width: calc(100% + 24px);
      }}
      #table-view h2,
      #table-view .footnote,
      #table-view .product-toolbar,
      #table-view .table-legend,
      #gantt-view h2,
      #gantt-view .product-toolbar {{
        margin-left: 12px;
        margin-right: 12px;
      }}
      .search-toolbar,
      .search-status {{
        margin-left: 0;
        margin-right: 0;
      }}
      .table-frame {{
        padding: 0;
      }}
    }}

    /* === TK Approved month filter toolbar === */
    .tk-month-toolbar {{
      display: flex;
      align-items: center;
      gap: 14px;
      flex-wrap: wrap;
      margin: -4px 0 20px;
      padding: 14px 18px;
      border: 1px solid rgba(200, 214, 228, 0.75);
      border-radius: var(--radius-md);
      background: rgba(255, 255, 255, 0.90);
      box-shadow: var(--shadow-sm), inset 0 1px 0 rgba(255, 255, 255, 0.98);
    }}
    .tk-month-toggle {{
      display: inline-flex;
      align-items: center;
      gap: 10px;
      cursor: pointer;
      user-select: none;
      font-weight: 700;
      color: #2a3a4d;
      font-size: 0.92rem;
    }}
    .tk-month-toggle-input {{
      position: absolute;
      opacity: 0;
      width: 0;
      height: 0;
      pointer-events: none;
    }}
    .tk-month-toggle-track {{
      position: relative;
      display: inline-block;
      width: 42px;
      height: 24px;
      background: #c2ceda;
      border-radius: 999px;
      transition: background 0.18s ease;
      flex-shrink: 0;
      box-shadow: inset 0 1px 3px rgba(16, 32, 51, 0.16);
    }}
    .tk-month-toggle-thumb {{
      position: absolute;
      top: 3px;
      left: 3px;
      width: 18px;
      height: 18px;
      border-radius: 50%;
      background: #ffffff;
      box-shadow: 0 1px 3px rgba(15, 23, 42, 0.20), 0 2px 8px rgba(15, 23, 42, 0.14);
      transition: transform 0.22s cubic-bezier(0.34, 1.56, 0.64, 1);
    }}
    .tk-month-toggle-input:checked + .tk-month-toggle-track {{
      background: #2563eb;
    }}
    .tk-month-toggle-input:checked + .tk-month-toggle-track .tk-month-toggle-thumb {{
      transform: translateX(18px);
      background: linear-gradient(135deg, #fff 0%, #eef5ff 100%);
    }}
    .tk-month-toggle-input:focus-visible + .tk-month-toggle-track {{
      outline: 3px solid rgba(37, 99, 235, 0.30);
      outline-offset: 2px;
    }}
    .tk-month-select {{
      appearance: none;
      -webkit-appearance: none;
      padding: 9px 32px 9px 14px;
      border: 1px solid rgba(192, 206, 218, 0.90);
      border-radius: var(--radius-sm);
      background: #fff url("data:image/svg+xml;charset=UTF-8,%3Csvg viewBox='0 0 20 20' xmlns='http://www.w3.org/2000/svg'%3E%3Cpath d='M5.5 8l4.5 4.5L14.5 8' stroke='%234a5f78' stroke-width='1.6' fill='none' stroke-linecap='round' stroke-linejoin='round'/%3E%3C/svg%3E") no-repeat right 8px center;
      background-size: 16px 16px;
      color: var(--text);
      font: inherit;
      font-weight: 600;
      min-width: 140px;
      cursor: pointer;
      box-shadow: 0 1px 3px rgba(16, 32, 51, 0.07), inset 0 1px 0 rgba(255, 255, 255, 0.90);
      transition: border-color 0.15s ease, box-shadow 0.15s ease, opacity 0.15s ease;
    }}
    .tk-month-select:focus {{
      outline: none;
      border-color: #2563eb;
      box-shadow: 0 0 0 3px var(--ring), 0 1px 3px rgba(16, 32, 51, 0.07);
    }}
    .tk-month-select:disabled {{
      opacity: 0.55;
      cursor: not-allowed;
    }}
    .tk-month-status {{
      color: #52657a;
      font-size: 0.88rem;
      font-weight: 600;
    }}
    .tk-month-analysis-panel {{
      margin-top: -6px;
      margin-bottom: 18px;
      border-color: rgba(29, 78, 216, 0.18);
      background:
        radial-gradient(circle at top left, rgba(14, 165, 233, 0.10), transparent 24%),
        linear-gradient(180deg, #ffffff 0%, #f8fbff 100%);
    }}
    .tk-month-analysis-header {{
      display: flex;
      justify-content: space-between;
      align-items: flex-start;
      gap: 16px;
      margin-bottom: 16px;
    }}
    .tk-month-analysis-header-side {{
      min-width: 220px;
      display: grid;
      gap: 10px;
      justify-items: stretch;
    }}
    .tk-month-analysis-status {{
      padding: 10px 12px;
      border-radius: 12px;
      border: 1px solid rgba(191, 219, 254, 0.95);
      background: rgba(255, 255, 255, 0.82);
      color: #1e3a8a;
      font-size: 0.9rem;
      font-weight: 700;
      line-height: 1.45;
    }}
    .tk-month-analysis-action {{
      appearance: none;
      border: 1px solid rgba(37, 99, 235, 0.24);
      border-radius: 12px;
      background: linear-gradient(180deg, rgba(239, 246, 255, 0.94) 0%, rgba(219, 234, 254, 0.96) 100%);
      color: #1d4ed8;
      font: inherit;
      font-weight: 800;
      padding: 10px 14px;
      cursor: pointer;
      transition: transform 0.18s ease, box-shadow 0.18s ease, border-color 0.18s ease;
    }}
    .tk-month-analysis-action:hover {{
      transform: translateY(-1px);
      box-shadow: 0 10px 20px rgba(37, 99, 235, 0.12);
      border-color: rgba(37, 99, 235, 0.34);
    }}
    .tk-month-analysis-action:focus-visible {{
      outline: 2px solid rgba(37, 99, 235, 0.28);
      outline-offset: 2px;
    }}
    .tk-month-analysis-cards {{
      display: grid;
      grid-template-columns: repeat(auto-fit, minmax(180px, 1fr));
      gap: 14px;
      margin-bottom: 18px;
    }}
    .tk-month-analysis-card {{
      padding: 18px 20px;
      border: 1px solid rgba(200, 214, 228, 0.80);
      border-radius: var(--radius-lg);
      background: linear-gradient(180deg, #ffffff 0%, #f6f9fd 100%);
      box-shadow: var(--shadow), inset 0 1px 0 rgba(255, 255, 255, 0.90);
    }}
    .tk-month-analysis-card.featured {{
      border-color: rgba(37, 99, 235, 0.30);
      background: linear-gradient(180deg, #eff6ff 0%, #dbeafe 100%);
    }}
    .tk-month-analysis-card-label {{
      color: #334155;
      font-size: 0.88rem;
      font-weight: 800;
      margin-bottom: 8px;
    }}
    .tk-month-analysis-card-value {{
      color: #0f172a;
      font-size: 2rem;
      font-weight: 800;
      letter-spacing: -0.04em;
    }}
    .tk-month-analysis-card-meta {{
      margin-top: 8px;
      color: #64748b;
      font-size: 0.88rem;
      line-height: 1.45;
    }}
    .tk-month-chart {{
      padding: 22px 20px 18px;
      border: 1px solid rgba(200, 214, 228, 0.75);
      border-radius: var(--radius-lg);
      background: linear-gradient(180deg, rgba(246, 249, 252, 0.97) 0%, rgba(255, 255, 255, 0.99) 100%);
      box-shadow: var(--shadow-sm);
    }}
    .tk-month-chart-bars {{
      display: grid;
      grid-template-columns: repeat(auto-fit, minmax(120px, 1fr));
      gap: 16px;
      align-items: end;
      min-height: 280px;
    }}
    .tk-month-chart-bar-card {{
      min-height: 240px;
      display: flex;
      flex-direction: column;
      align-items: center;
      gap: 10px;
      padding-top: 10px;
    }}
    .tk-month-chart-bar-card.featured .tk-month-chart-bar-track {{
      border-color: rgba(59, 130, 246, 0.40);
      background: linear-gradient(180deg, rgba(219, 234, 254, 0.85) 0%, rgba(239, 246, 255, 0.70) 100%);
    }}
    .tk-month-chart-value {{
      min-height: 1.4rem;
      color: #0f172a;
      font-size: 1rem;
      font-weight: 800;
    }}
    .tk-month-chart-bar-track {{
      width: min(100%, 148px);
      height: 180px;
      padding: 10px;
      display: flex;
      align-items: flex-end;
      border: 1px solid rgba(203, 213, 225, 0.95);
      border-radius: 18px 18px 10px 10px;
      background: linear-gradient(180deg, rgba(241, 245, 249, 0.85) 0%, rgba(255, 255, 255, 0.98) 100%);
    }}
    .tk-month-chart-bar-fill {{
      width: 100%;
      min-height: 6px;
      border-radius: 12px 12px 8px 8px;
      background: linear-gradient(180deg, #60a5fa 0%, #2563eb 100%);
      box-shadow: 0 12px 22px rgba(37, 99, 235, 0.22);
      transition: height 0.25s ease;
    }}
    .tk-month-chart-bar-card.featured .tk-month-chart-bar-fill {{
      background: linear-gradient(180deg, #34d399 0%, #059669 100%);
      box-shadow: 0 12px 22px rgba(5, 150, 105, 0.24);
    }}
    .tk-month-chart-label {{
      color: #334155;
      font-size: 0.9rem;
      font-weight: 800;
      text-align: center;
    }}
    .tk-month-analysis-summary {{
      display: flex;
      flex-wrap: wrap;
      gap: 12px;
      margin: 16px 0 18px;
    }}
    .tk-month-analysis-pill {{
      min-width: 180px;
      padding: 12px 14px;
      border-radius: 14px;
      border: 1px solid rgba(191, 219, 254, 0.95);
      background: rgba(239, 246, 255, 0.92);
      display: grid;
      gap: 4px;
    }}
    .tk-month-analysis-pill.excluded {{
      border-color: rgba(251, 191, 36, 0.95);
      background: rgba(255, 251, 235, 0.95);
    }}
    .tk-month-analysis-pill span {{
      color: #64748b;
      font-size: 0.82rem;
      font-weight: 800;
      text-transform: uppercase;
      letter-spacing: 0.04em;
    }}
    .tk-month-analysis-pill strong {{
      color: #0f172a;
      font-size: 1.35rem;
      letter-spacing: -0.03em;
    }}
    .tk-month-exclusion-table-frame {{
      max-width: 100%;
      height: 320px;
      overflow-x: hidden;
      overflow-y: auto;
      scrollbar-gutter: stable;
    }}
    .tk-month-exclusion-table {{
      width: 100%;
      border-collapse: collapse;
      table-layout: fixed;
    }}
    .tk-month-exclusion-table .tk-month-exclusion-col {{
      width: 33.333%;
    }}
    .tk-month-exclusion-table th,
    .tk-month-exclusion-table td {{
      padding: 12px 14px;
      border-bottom: 1px solid #e2e8f0;
      text-align: left;
      vertical-align: top;
      white-space: normal;
      overflow-wrap: anywhere;
      word-break: break-word;
    }}
    .tk-month-exclusion-table th {{
      color: #475569;
      font-size: 0.8rem;
      font-weight: 800;
      text-transform: uppercase;
      letter-spacing: 0.05em;
      background: rgba(248, 250, 252, 0.96);
      position: sticky;
      top: 0;
      z-index: 1;
    }}
    .tk-month-exclusion-table td {{
      color: #1e293b;
      font-size: 0.92rem;
      line-height: 1.55;
    }}
    .tk-month-exclusion-link {{
      appearance: none;
      border: 0;
      padding: 0;
      background: transparent;
      color: #1d4ed8;
      font: inherit;
      font-weight: 800;
      text-align: left;
      cursor: pointer;
    }}
    .tk-month-exclusion-link:hover {{
      text-decoration: underline;
    }}
    .tk-month-exclusion-link:focus-visible {{
      outline: 2px solid rgba(37, 99, 235, 0.28);
      outline-offset: 3px;
      border-radius: 4px;
    }}
    .tk-month-exclusion-reason {{
      margin: 0;
      white-space: pre-line;
      overflow-wrap: anywhere;
      word-break: break-word;
    }}

    /* === Capacity calculator === */
    .capacity-calculator {{
      margin: 4px 0 22px;
      padding: 18px 20px;
      border: 1px solid rgba(200, 214, 228, 0.75);
      border-radius: var(--radius-md);
      background: rgba(255, 255, 255, 0.90);
      box-shadow: var(--shadow-sm), inset 0 1px 0 rgba(255, 255, 255, 0.98);
    }}
    .capacity-calc-title {{
      font-size: 0.78rem;
      font-weight: 800;
      letter-spacing: 0.1em;
      text-transform: uppercase;
      color: #52657a;
      margin-bottom: 12px;
    }}
    .capacity-calc-body {{
      display: grid;
      grid-template-columns: minmax(140px, 1fr) minmax(160px, 1fr) minmax(140px, 1fr) minmax(220px, 1.2fr) minmax(220px, 1.2fr);
      gap: 14px;
      align-items: stretch;
    }}
    .capacity-field {{
      display: flex;
      flex-direction: column;
      gap: 6px;
      min-width: 0;
    }}
    .capacity-field-label {{
      font-size: 0.76rem;
      font-weight: 800;
      letter-spacing: 0.05em;
      text-transform: uppercase;
      color: #52657a;
    }}
    .capacity-input,
    .capacity-select {{
      appearance: none;
      -webkit-appearance: none;
      padding: 10px 12px;
      border: 1px solid rgba(192, 206, 218, 0.90);
      border-radius: var(--radius-sm);
      background: #fff;
      color: var(--text);
      font: inherit;
      font-weight: 600;
      box-shadow: 0 1px 3px rgba(16, 32, 51, 0.06), inset 0 1px 0 rgba(255, 255, 255, 0.80);
      transition: border-color 0.15s ease, box-shadow 0.15s ease;
      min-width: 0;
      width: 100%;
    }}
    .capacity-select {{
      padding-right: 30px;
      background: #fff url("data:image/svg+xml;charset=UTF-8,%3Csvg viewBox='0 0 20 20' xmlns='http://www.w3.org/2000/svg'%3E%3Cpath d='M5.5 8l4.5 4.5L14.5 8' stroke='%234a5f78' stroke-width='1.6' fill='none' stroke-linecap='round' stroke-linejoin='round'/%3E%3C/svg%3E") no-repeat right 8px center;
      background-size: 16px 16px;
      cursor: pointer;
    }}
    .capacity-input:focus,
    .capacity-select:focus {{
      outline: none;
      border-color: #2563eb;
      box-shadow: 0 0 0 3px var(--ring), 0 1px 3px rgba(16, 32, 51, 0.06);
    }}
    .capacity-result-card {{
      padding: 14px 18px 16px;
      min-height: 0;
    }}
    .metric-card-capacity {{
      background: linear-gradient(135deg, #047857 0%, #0f766e 60%, #115e59 100%);
    }}
    .capacity-result-card .metric-label {{ font-weight: 800; }}
    .capacity-result-card .metric-value {{ font-size: 1.7rem; }}
    .capacity-result-card .metric-value-wrap {{ min-height: 36px; }}
    @media (max-width: 900px) {{
      .capacity-calc-body {{ grid-template-columns: 1fr; }}
    }}

    /* === Clickable metric card state === */
    .metric-card-clickable {{
      cursor: pointer;
      transition: transform 0.18s ease, box-shadow 0.18s ease;
      position: relative;
    }}
    .metric-card-clickable:hover {{
      transform: translateY(-2px);
      box-shadow: 0 22px 42px rgba(16, 32, 51, 0.14), 0 6px 12px rgba(16, 32, 51, 0.08);
    }}
    .metric-card-clickable:focus-visible {{
      outline: 3px solid #ffffff;
      outline-offset: 2px;
    }}
    .metric-card-clickable::after {{
      content: "\u2192";
      position: absolute;
      top: 14px;
      right: 16px;
      width: 26px;
      height: 26px;
      display: inline-flex;
      align-items: center;
      justify-content: center;
      border-radius: 999px;
      background: rgba(255, 255, 255, 0.22);
      color: rgba(255, 255, 255, 0.95);
      font-size: 1rem;
      font-weight: 700;
      line-height: 1;
      transition: transform 0.18s ease, background 0.18s ease;
      pointer-events: none;
    }}
    .metric-card-clickable:hover::after {{
      transform: translateX(3px);
      background: rgba(255, 255, 255, 0.32);
    }}
    .metric-card.tk-month-active {{
      box-shadow: 0 0 0 2px rgba(255, 255, 255, 0.32), 0 18px 36px rgba(16, 32, 51, 0.16);
    }}

    /* === Drawer modal === */
    .drawer-modal[hidden] {{ display: none; }}
    .drawer-modal {{
      position: fixed;
      inset: 0;
      z-index: 1000;
    }}
    .drawer-backdrop {{
      position: absolute;
      inset: 0;
      background: rgba(15, 23, 42, 0.52);
      backdrop-filter: blur(2px);
      animation: drawerFadeIn 0.18s ease-out;
    }}
    .drawer-panel {{
      position: absolute;
      top: 0;
      right: 0;
      bottom: 0;
      width: min(620px, 100%);
      background: #ffffff;
      box-shadow: -24px 0 60px rgba(15, 23, 42, 0.22);
      padding: 22px 26px 28px;
      overflow-y: auto;
      display: flex;
      flex-direction: column;
      gap: 16px;
      animation: drawerSlideIn 0.26s cubic-bezier(0.2, 0.7, 0.2, 1);
    }}
    @keyframes drawerSlideIn {{
      from {{ transform: translateX(100%); }}
      to {{ transform: translateX(0); }}
    }}
    @keyframes drawerFadeIn {{
      from {{ opacity: 0; }}
      to {{ opacity: 1; }}
    }}
    .drawer-header {{
      display: flex;
      justify-content: space-between;
      align-items: flex-start;
      gap: 14px;
      padding-bottom: 14px;
      border-bottom: 1px solid #e2e8ef;
    }}
    .drawer-heading {{ min-width: 0; flex: 1; }}
    .drawer-eyebrow {{
      font-size: 0.72rem;
      font-weight: 800;
      letter-spacing: 0.1em;
      text-transform: uppercase;
      color: #6b7a8c;
      margin-bottom: 4px;
    }}
    .drawer-header h2 {{
      margin: 0 0 6px;
      font-size: 1.32rem;
      color: #102033;
      letter-spacing: -0.01em;
    }}
    .drawer-subtitle {{
      color: #5c6f83;
      font-size: 0.92rem;
    }}
    .drawer-close {{
      appearance: none;
      border: 1px solid #d7e0ea;
      background: #fff;
      color: #102033;
      width: 34px;
      height: 34px;
      border-radius: 10px;
      font-size: 1.4rem;
      line-height: 1;
      cursor: pointer;
      transition: background 0.15s ease, color 0.15s ease, border-color 0.15s ease;
      flex-shrink: 0;
    }}
    .drawer-close:hover {{ background: #102033; color: #fff; border-color: #102033; }}
    .drawer-summary {{
      display: grid;
      grid-template-columns: repeat(auto-fit, minmax(160px, 1fr));
      gap: 10px;
      padding: 12px 14px;
      border: 1px solid #e2e8ef;
      border-radius: 12px;
      background: linear-gradient(180deg, #f7faff 0%, #eef4fb 100%);
    }}
    .drawer-summary-cell {{
      display: flex;
      flex-direction: column;
      gap: 4px;
    }}
    .drawer-summary-cell span {{
      font-size: 0.72rem;
      font-weight: 800;
      letter-spacing: 0.08em;
      text-transform: uppercase;
      color: #5c6f83;
    }}
    .drawer-summary-cell strong {{
      font-size: 1.12rem;
      color: #102033;
    }}
    .drawer-body {{
      display: flex;
      flex-direction: column;
      gap: 10px;
    }}
    .drawer-empty {{
      padding: 28px 16px;
      text-align: center;
      color: #5c6f83;
      border: 1px dashed #d7e0ea;
      border-radius: 12px;
      background: #fafcff;
    }}
    .drawer-empty.tight {{
      padding: 14px 12px;
      font-size: 0.92rem;
    }}
    .drawer-note-card {{
      border: 1px solid #fcd34d;
      border-left: 4px solid #d97706;
      border-radius: 12px;
      padding: 14px 16px;
      background: linear-gradient(180deg, #fffdf4 0%, #fff7df 100%);
      display: grid;
      gap: 8px;
    }}
    .drawer-note-title {{
      color: #92400e;
      font-size: 0.82rem;
      font-weight: 800;
      letter-spacing: 0.08em;
      text-transform: uppercase;
    }}
    .drawer-note-list {{
      margin: 0;
      padding-left: 18px;
      color: #78350f;
      line-height: 1.55;
    }}
    .drawer-group {{
      display: grid;
      gap: 10px;
    }}
    .drawer-group-title {{
      color: #334155;
      font-size: 0.84rem;
      font-weight: 800;
      letter-spacing: 0.08em;
      text-transform: uppercase;
    }}
    .drawer-nested-list {{
      display: grid;
      gap: 10px;
      padding-left: 18px;
      border-left: 2px solid #dbeafe;
      margin-left: 12px;
    }}
    .drawer-epic-card {{
      border: 1px solid #e2e8ef;
      border-left: 4px solid #2563eb;
      border-radius: 12px;
      padding: 12px 14px;
      background: #ffffff;
      display: flex;
      flex-direction: column;
      gap: 6px;
      transition: border-color 0.15s ease, box-shadow 0.15s ease, transform 0.15s ease;
    }}
    .drawer-epic-card:hover {{
      box-shadow: 0 10px 22px rgba(15, 23, 42, 0.08);
      transform: translateY(-1px);
    }}
    .drawer-epic-head {{
      display: flex;
      justify-content: space-between;
      align-items: center;
      gap: 10px;
      flex-wrap: wrap;
    }}
    .drawer-epic-key {{
      font-family: "JetBrains Mono", ui-monospace, "SFMono-Regular", Menlo, Consolas, monospace;
      font-size: 0.82rem;
      font-weight: 700;
      color: #1d4ed8;
      text-decoration: none;
      padding: 2px 8px;
      background: rgba(37, 99, 235, 0.08);
      border-radius: 6px;
    }}
    .drawer-epic-key:hover {{
      background: rgba(37, 99, 235, 0.16);
      text-decoration: underline;
    }}
    .drawer-epic-product {{
      font-size: 0.72rem;
      font-weight: 800;
      letter-spacing: 0.06em;
      text-transform: uppercase;
      color: #5c6f83;
    }}
    .drawer-epic-title {{
      font-weight: 700;
      color: #102033;
      line-height: 1.4;
    }}
    .drawer-epic-meta {{
      display: flex;
      flex-wrap: wrap;
      gap: 10px 16px;
      font-size: 0.86rem;
      color: #4a5f78;
    }}
    .drawer-epic-meta strong {{
      display: block;
      font-size: 0.68rem;
      font-weight: 800;
      letter-spacing: 0.08em;
      text-transform: uppercase;
      color: #8a98aa;
      margin-bottom: 2px;
    }}
    .drawer-epic-tk {{
      display: flex;
      justify-content: space-between;
      align-items: center;
      padding-top: 8px;
      border-top: 1px dashed #e2e8ef;
      margin-top: 4px;
    }}
    .drawer-epic-tk-label {{
      font-size: 0.72rem;
      font-weight: 800;
      letter-spacing: 0.08em;
      text-transform: uppercase;
      color: #5c6f83;
    }}
    .drawer-epic-tk-value {{
      font-weight: 800;
      color: #047857;
      font-size: 1rem;
    }}
    body.drawer-open {{ overflow: hidden; }}

    @media (max-width: 900px) {{
      .drawer-panel {{ padding: 18px 18px 22px; }}
      .tk-month-toolbar {{ padding: 10px 12px; }}
      .tk-month-analysis-header {{ flex-direction: column; }}
      .tk-month-analysis-status {{ width: 100%; min-width: 0; }}
      .tk-month-analysis-cards,
      .tk-month-chart-bars {{ grid-template-columns: 1fr; }}
      .tk-month-chart-bar-track {{ width: 100%; max-width: none; height: 160px; }}
    }}

    /* --- RMI Estimation & Scheduling Table --- */
    .rmi-schedule-panel {{
      margin-bottom: var(--gutter);
    }}
    .rmi-schedule-header-bar {{
      display: flex;
      align-items: center;
      justify-content: space-between;
      flex-wrap: wrap;
      gap: 12px;
      margin-bottom: 12px;
    }}
    .rmi-schedule-header-bar h2 {{ margin: 0; }}
    .rmi-schedule-controls {{
      display: flex;
      align-items: center;
      gap: 8px;
    }}
    .rmi-schedule-year-label {{
      font-size: 0.82rem;
      font-weight: 600;
      color: var(--muted);
    }}
    .rmi-schedule-year-select {{
      font-size: 0.82rem;
      padding: 4px 10px;
      border: 1px solid var(--line);
      border-radius: 6px;
      background: var(--panel);
      cursor: pointer;
    }}
    .rmi-schedule-panel .table-frame {{
      overflow-x: auto;
      overflow-y: auto;
      max-width: 100%;
      max-height: 80vh;
    }}
    .rmi-schedule-table {{
      border-collapse: separate;
      border-spacing: 0;
      font-size: 0.78rem;
      min-width: 1200px;
    }}
    .rmi-schedule-table th,
    .rmi-schedule-table td {{
      padding: 6px 8px;
      border: 1px solid var(--line);
      text-align: center;
      white-space: nowrap;
      overflow: hidden;
      text-overflow: ellipsis;
    }}
    .rmi-schedule-table th {{ font-weight: 700; }}
    /* --- Sticky header rows (vertical scroll) --- */
    .rmi-schedule-table thead th {{
      position: sticky;
      z-index: 2;
      background: #f1f5f9;
    }}
    .rmi-sched-header-groups th {{
      top: 0;
      background: var(--panel-soft) !important;
    }}
    .rmi-sched-header-cols th {{
      /* Must match actual rendered height of the first header row */
      top: var(--rmi-sched-row1-h, 31px);
    }}
    /* --- Sticky first 6 columns (horizontal scroll) --- */
    .rmi-schedule-table th:nth-child(-n+6),
    .rmi-schedule-table td:nth-child(-n+6) {{
      position: sticky;
      z-index: 2;
    }}
    /* Corner cells: sticky both directions need highest z-index */
    .rmi-schedule-table thead th:nth-child(-n+6) {{
      z-index: 4;
    }}
    .rmi-schedule-table th:nth-child(1),
    .rmi-schedule-table td:nth-child(1) {{ left: 0; min-width: 36px; width: 36px; background: var(--panel); text-align: center; color: var(--muted); font-size: 0.7rem; }}
    .rmi-schedule-table th:nth-child(2),
    .rmi-schedule-table td:nth-child(2) {{ left: 36px; min-width: 360px; width: 360px; white-space: normal; word-break: break-word; background: var(--panel); }}
    .rmi-schedule-table th:nth-child(3),
    .rmi-schedule-table td:nth-child(3) {{ left: 396px; min-width: 90px; width: 90px; background: var(--panel); }}
    .rmi-schedule-table th:nth-child(4),
    .rmi-schedule-table td:nth-child(4) {{ left: 486px; min-width: 80px; width: 80px; background: var(--panel); }}
    .rmi-schedule-table th:nth-child(5),
    .rmi-schedule-table td:nth-child(5) {{ left: 566px; min-width: 70px; width: 70px; background: var(--panel); }}
    .rmi-schedule-table th:nth-child(6),
    .rmi-schedule-table td:nth-child(6) {{ left: 636px; min-width: 70px; width: 70px; background: var(--panel); border-right: 2px solid #94a3b8; }}
    .rmi-sched-header-cols th {{ background: #f1f5f9 !important; }}
    .rmi-sched-header-groups th {{
      background: var(--panel-soft) !important;
      border-bottom: none;
    }}
    .rmi-sched-header-groups th.rmi-sched-group-estimation {{
      background: #1e3a5f !important;
      color: #fff !important;
    }}
    .rmi-sched-header-groups th.rmi-sched-group-scheduling {{
      background: #d97706 !important;
      color: #fff !important;
    }}
    .rmi-sched-header-cols th {{
      background: #f1f5f9 !important;
      font-size: 0.74rem;
      text-transform: uppercase;
      letter-spacing: 0.04em;
    }}
    .rmi-sched-col-rmi {{
      text-align: left !important;
    }}
    .rmi-sched-month {{ min-width: 62px; width: 62px; }}
    .rmi-schedule-table td.rmi-sched-cell-rmi {{
      text-align: left;
      font-weight: 600;
    }}
    .rmi-sched-jira-link {{
      display: inline-flex;
      align-items: center;
      justify-content: center;
      width: 18px;
      height: 18px;
      margin-left: 4px;
      vertical-align: middle;
      border-radius: 3px;
      background: #0052cc;
      color: #fff;
      font-size: 10px;
      font-weight: 800;
      text-decoration: none;
      line-height: 1;
    }}
    .rmi-sched-jira-link:hover {{ background: #0747a6; }}
    .rmi-sched-product-group td {{
      background: #f8fafc !important;
      font-weight: 800;
      font-size: 0.76rem;
      text-transform: uppercase;
      letter-spacing: 0.06em;
      border-top: 2px solid var(--line);
    }}
    .rmi-sched-product-group td:first-child {{
      border-left: 4px solid var(--muted);
    }}
    .rmi-sched-product-group td.rmi-sched-group-label {{
      text-align: left;
    }}
    .rmi-sched-product-subtotal td {{
      background: #f1f5f9 !important;
      font-weight: 700;
      font-size: 0.76rem;
      border-top: 1px solid var(--line);
    }}
    .rmi-sched-product-subtotal td:first-child {{
      border-left: 4px solid var(--muted);
      text-align: right;
      padding-right: 12px;
    }}
    .rmi-sched-grand-total td {{
      background: #e2e8f0 !important;
      font-weight: 800;
      font-size: 0.78rem;
      border-top: 2px solid #94a3b8;
    }}
    .rmi-sched-grand-total td:first-child {{
      text-align: right;
      padding-right: 12px;
    }}
    .rmi-sched-epic-row:nth-child(even) td {{ background: #fafbfd !important; }}
    .rmi-sched-epic-row:hover td {{ background: #eef2ff !important; }}
    .rmi-sched-status-pill {{
      display: inline-block;
      padding: 2px 8px;
      border-radius: 10px;
      font-size: 0.68rem;
      font-weight: 700;
      text-transform: capitalize;
      background: #e2e8f0;
      color: #334155;
    }}
    .rmi-sched-status-pill[data-status-lower="done"],
    .rmi-sched-status-pill[data-status-lower="delivered"] {{
      background: #d1fae5;
      color: #065f46;
    }}
    .rmi-sched-status-pill[data-status-lower="in progress"] {{
      background: #dbeafe;
      color: #1e40af;
    }}
    .rmi-sched-status-pill[data-status-lower="to do"] {{
      background: #fef3c7;
      color: #92400e;
    }}
    /* --- RMI Schedule Product Filter Cards --- */
    .rmi-sched-product-cards {{
      display: grid;
      grid-template-columns: repeat(auto-fit, minmax(160px, 1fr));
      gap: 10px;
      margin-bottom: 14px;
    }}
    .rmi-sched-pcard {{
      display: flex;
      flex-direction: column;
      gap: 2px;
      padding: 12px 16px;
      border: 2px solid var(--line);
      border-radius: 12px;
      background: var(--panel);
      cursor: pointer;
      transition: border-color 0.15s, box-shadow 0.15s;
      user-select: none;
      border-left: 4px solid var(--product-accent, var(--line));
    }}
    .rmi-sched-pcard:hover {{
      box-shadow: 0 2px 8px rgba(0,0,0,0.06);
    }}
    .rmi-sched-pcard:focus-visible {{
      outline: 2px solid var(--product-accent, #2563eb);
      outline-offset: 2px;
    }}
    .rmi-sched-pcard.active {{
      border-color: var(--product-accent, #102033);
      box-shadow: 0 2px 12px rgba(0,0,0,0.10);
    }}
    .rmi-sched-pcard-label {{
      font-size: 0.78rem;
      font-weight: 700;
      color: var(--text);
    }}
    .rmi-sched-pcard-value {{
      font-size: 1.3rem;
      font-weight: 800;
      color: var(--product-accent, var(--text));
      font-style: italic;
    }}
    .rmi-sched-pcard-meta {{
      font-size: 0.68rem;
      color: var(--muted);
    }}
  </style>
</head>
<body>
  <div class="page">
    <header>
      <h1>RMI Jira Gantt Dashboard</h1>
      <div class="subtext">
        This page is generated directly from <strong>{escape(str(db_path))}</strong>. It combines workbook source values from the `source_rmi_rows`
        table with fetched Jira epic, story, descendant, and worklog data so the timeline and drill-down sections stay aligned with the latest extraction run.
      </div>
    </header>

    {render_capacity_calculator()}
    <section class="metric-grid">{render_metric_cards(epic_metric_summary)}</section>
    {render_product_tk_cards(epic_metric_summary)}
    {render_tk_month_toolbar(month_keys)}
    {render_tk_month_story_analysis_panel()}

    {render_rmi_schedule_table(rmi_schedule_initial_body_html, rmi_schedule_initial_foot_html)}

    {render_search_toolbar()}

    <div class="unit-toolbar" role="tablist" aria-label="Duration unit selector">
      <button class="unit-toggle active" type="button" data-unit="hours" aria-pressed="true">Hours</button>
      <button class="unit-toggle" type="button" data-unit="days" aria-pressed="false">Days</button>
    </div>

    <div class="view-toolbar" role="tablist" aria-label="View selector">
      <button class="view-toggle" type="button" data-view="gantt-view" aria-pressed="false">Gantt View</button>
      <button class="view-toggle active" type="button" data-view="table-view" aria-pressed="true">Table View</button>
    </div>

    <section class="panel view-section" id="gantt-view" hidden>
      <h2>Timeline</h2>
      {render_product_filter_toolbar(source_rows)}
      {render_gantt(gantt_rows)}
    </section>

    <section class="panel view-section" id="table-view">
      <h2>Hierarchical Table</h2>
      <div class="footnote">Epic parents expand to reveal stories, and each story expands again to show descendant issues and worklogs.</div>
      {render_table_legend()}
      {render_product_filter_toolbar(source_rows)}
      <div class="table-frame">
        {render_epic_table_view(source_rows)}
      </div>
    </section>

    <section class="panel">
      <h2>Run Errors</h2>
      {render_error_table(errors)}
    </section>
  </div>
  {render_drawer_modal()}
  <script>
    (() => {{
      const viewButtons = Array.from(document.querySelectorAll('.view-toggle'));
      const unitButtons = Array.from(document.querySelectorAll('.unit-toggle'));
      const productButtons = Array.from(document.querySelectorAll('.product-toggle'));
      const productSummaryCards = Array.from(document.querySelectorAll('.product-summary-card'));
      const rowToggleButtons = Array.from(document.querySelectorAll('.row-toggle'));
      const views = Array.from(document.querySelectorAll('.view-section'));
      const durationNodes = Array.from(document.querySelectorAll('.duration-value'));
      const unitCaptions = Array.from(document.querySelectorAll('.unit-caption'));
      const metricCards = Array.from(document.querySelectorAll('.metric-card[data-metric-key]'));
      const epicSummaryRows = Array.from(document.querySelectorAll('.epic-summary-row'));
      const ganttProductSections = Array.from(document.querySelectorAll('.gantt-product-section'));
      const ganttEpicRows = Array.from(document.querySelectorAll('.gantt-epic-row'));
      const searchInput = document.getElementById('epic-search');
      const searchClearButton = document.querySelector('[data-clear-search]');
      const searchStatus = document.querySelector('[data-search-status]');
      const epicMetrics = {epic_metric_summary_json};
      const epicDetails = {epic_detail_records_json};
      const storyDetails = {story_detail_records_json};
      const subtaskDetails = {subtask_detail_records_json};
      const capacityMonths = {capacity_months_json};
      const rmiScheduleRecords = {rmi_schedule_records_json};
      const rmiScheduleYears = {rmi_schedule_years_json};
      const capacityEmployeesInput = document.getElementById('capacity-employees');
      const capacityLeavesInput = document.getElementById('capacity-leaves');
      const capacityLeavesLabel = document.getElementById('capacity-leaves-label');
      const capacityMonthSelect = document.getElementById('capacity-month');
      const capacityValueNode = document.getElementById('capacity-value');
      const availabilityValueNode = document.getElementById('availability-value');
      const idleCapacityCard = document.querySelector('.metric-card[data-metric-key="idle_capacity_seconds"]');
      const capacityMetaNode = document.querySelector('[data-capacity-meta]');
      const availabilityMetaNode = document.querySelector('[data-availability-meta]');
      const tkStartMonthToggleInput = document.getElementById('tk-start-month-enabled');
      const tkMonthToggleInput = document.getElementById('tk-month-enabled');
      const tkThroughMonthToggleInput = document.getElementById('tk-through-month-enabled');
      const tkJiraOnlyToggleInput = document.getElementById('tk-jira-only-enabled');
      const tkMonthSelect = document.getElementById('tk-month-select');
      const tkMonthStatusNode = document.querySelector('[data-tk-month-status]');
      const tkMonthAnalysisPanel = document.getElementById('tk-month-analysis');
      const tkMonthAnalysisStatusNode = document.querySelector('[data-tk-month-analysis-status]');
      const tkMonthAnalysisIncludedNode = document.querySelector('[data-month-analysis-included]');
      const tkMonthAnalysisExcludedNode = document.querySelector('[data-month-analysis-excluded]');
      const tkMonthAnalysisExclusionBody = document.querySelector('[data-month-analysis-exclusion-body]');
      const tkMonthAnalysisCardsContainer = document.querySelector('.tk-month-analysis-cards');
      const tkMonthChartBarsContainer = document.querySelector('.tk-month-chart-bars');
      const tkMonthAnalysisDrawerButton = document.querySelector('[data-open-tk-analysis-drawer]');
      const tkApprovedCard = document.querySelector('.metric-card[data-metric-key="tk_approved_seconds"]');
      const drawerEl = document.getElementById('epic-drawer');
      const drawerEyebrowNode = drawerEl ? drawerEl.querySelector('[data-drawer-eyebrow]') : null;
      const drawerTitleNode = drawerEl ? drawerEl.querySelector('#epic-drawer-title') : null;
      const drawerSubtitleNode = drawerEl ? drawerEl.querySelector('[data-drawer-subtitle]') : null;
      const drawerSummaryNode = drawerEl ? drawerEl.querySelector('[data-drawer-summary]') : null;
      const drawerBodyNode = drawerEl ? drawerEl.querySelector('[data-drawer-body]') : null;
      let activeUnit = 'hours';
      let activeProducts = new Set(['all']);
      let activeQuery = '';
      let tkStartMonthEnabled = false;
      let tkMonthEnabled = false;
      let tkThroughMonthEnabled = false;
      let tkJiraOnly = true;
      let tkMonthKey = '';

      function setEpicExpansion(button, expanded) {{
        const detailRow = document.getElementById(button.dataset.target);
        button.setAttribute('aria-expanded', String(expanded));
        button.textContent = expanded ? '-' : '+';
        if (detailRow) {{
          detailRow.hidden = !expanded;
        }}
      }}

      function formatLeaveInputValue(seconds, unit) {{
        const value = unit === 'days' ? seconds / 28800 : seconds / 3600;
        if (Math.abs(value - Math.round(value)) < 0.001) {{
          return String(Math.round(value));
        }}
        return value.toFixed(2).replace(/\\.00$/, '').replace(/(\\.\\d*[1-9])0$/, '$1');
      }}

      function syncLeavesSecondsFromInput() {{
        if (!capacityLeavesInput) return 0;
        const raw = Math.max(0, Number(capacityLeavesInput.value || 0));
        const seconds = activeUnit === 'days' ? raw * 28800 : raw * 3600;
        capacityLeavesInput.dataset.seconds = String(seconds);
        return seconds;
      }}

      function updateLeavesFieldDisplay(unit) {{
        if (!capacityLeavesInput) return;
        const seconds = Math.max(0, Number(capacityLeavesInput.dataset.seconds || 0));
        capacityLeavesInput.value = formatLeaveInputValue(seconds, unit);
        capacityLeavesInput.step = unit === 'days' ? '0.5' : '1';
        capacityLeavesInput.setAttribute('aria-label', unit === 'days' ? 'Total leaves in days' : 'Total leaves in hours');
        if (capacityLeavesLabel) {{
          capacityLeavesLabel.textContent = unit === 'days'
            ? (capacityLeavesLabel.dataset.labelDays || 'Total Leaves (Days)')
            : (capacityLeavesLabel.dataset.labelHours || 'Total Leaves (Hours)');
        }}
      }}

      function setUnits(unit) {{
        syncLeavesSecondsFromInput();
        activeUnit = unit;
        unitButtons.forEach((button) => {{
          const active = button.dataset.unit === unit;
          button.classList.toggle('active', active);
          button.setAttribute('aria-pressed', String(active));
        }});
        durationNodes.forEach((node) => {{
          const seconds = Number(node.dataset.seconds || '0');
          const value = unit === 'days' ? node.dataset.days : node.dataset.hours;
          node.textContent = value || (unit === 'days' ? `${{(seconds / 28800).toFixed(2)}} d` : `${{(seconds / 3600).toFixed(2)}} h`);
        }});
        document.querySelectorAll('.duration-value:not([data-unit-bound])').forEach((node) => {{
          node.setAttribute('data-unit-bound', '1');
          durationNodes.push(node);
          const seconds = Number(node.dataset.seconds || '0');
          const value = unit === 'days' ? node.dataset.days : node.dataset.hours;
          node.textContent = value || (unit === 'days' ? `${{(seconds / 28800).toFixed(2)}} d` : `${{(seconds / 3600).toFixed(2)}} h`);
        }});
        unitCaptions.forEach((node) => {{
          const label = unit === 'days' ? node.dataset.labelDays : node.dataset.labelHours;
          if (label) {{
            node.textContent = label;
          }}
        }});
        updateLeavesFieldDisplay(unit);
        if (typeof renderRmiScheduleTable === 'function') {{
          try {{ renderRmiScheduleTable(); }} catch (e) {{ /* schedule consts not yet initialized */ }}
        }}
      }}

      function formatMetricDuration(seconds, unit, compact = false) {{
        const value = unit === 'days' ? seconds / 28800 : seconds / 3600;
        const suffix = unit === 'days' ? ' d' : ' h';
        return compact
          ? `${{Math.round(value).toLocaleString()}}${{suffix}}`
          : `${{value.toLocaleString(undefined, {{ minimumFractionDigits: 2, maximumFractionDigits: 2 }})}}${{suffix}}`;
      }}

      function selectedProductKeys() {{
        return activeProducts.has('all') ? ['all'] : Array.from(activeProducts);
      }}

      function productInScope(product) {{
        return activeProducts.has('all') || activeProducts.has(product);
      }}

      function currentProductLabel() {{
        const selected = selectedProductKeys();
        if (selected.length === 1 && selected[0] === 'all') return 'All Products';
        if (selected.length === 1) return selected[0];
        return `${{selected.length}} products selected`;
      }}

      function isJiraPopulatedEpic(epic) {{
        return Boolean(epic && epic.jira_populated);
      }}

      function currentExecutiveScopeLabel() {{
        const productLabel = currentProductLabel();
        return tkJiraOnly ? `${{productLabel}} · Jira populated only` : productLabel;
      }}

      function executiveMetricKeys() {{
        return [
          'epic_count',
          'most_likely_seconds',
          'optimistic_seconds',
          'pessimistic_seconds',
          'calculated_seconds',
          'tk_approved_seconds',
          'jira_original_estimate_seconds',
          'story_estimate_seconds',
          'subtask_estimate_seconds',
        ];
      }}

      function emptyScopedMetrics() {{
        return Object.fromEntries(executiveMetricKeys().map((key) => [key, 0]));
      }}

      function scopedEpicDetails() {{
        return epicDetails.filter((epic) => productInScope(epic.product) && (!tkJiraOnly || isJiraPopulatedEpic(epic)));
      }}

      function scopedEpicKeySet() {{
        return new Set(scopedEpicDetails().map((epic) => epic.jira_id).filter(Boolean));
      }}

      function scopedMetrics() {{
        const totals = emptyScopedMetrics();
        scopedEpicDetails().forEach((epic) => {{
          totals.epic_count += 1;
          totals.most_likely_seconds += Number(epic.most_likely_seconds || 0);
          totals.optimistic_seconds += Number(epic.optimistic_seconds || 0);
          totals.pessimistic_seconds += Number(epic.pessimistic_seconds || 0);
          totals.calculated_seconds += Number(epic.calculated_seconds || 0);
          totals.tk_approved_seconds += Number(epic.tk_approved_seconds || 0);
          totals.jira_original_estimate_seconds += Number(epic.jira_original_estimate_seconds || 0);
          totals.story_estimate_seconds += Number(epic.story_estimate_seconds || 0);
          totals.subtask_estimate_seconds += Number(epic.subtask_estimate_seconds || 0);
        }});
        return totals;
      }}

      function updateProductSummaryCards() {{
        const scopedEpics = scopedEpicDetails();
        productSummaryCards.forEach((card) => {{
          const product = card.dataset.productSummary || 'all';
          const relevantEpics = product === 'all'
            ? scopedEpics
            : scopedEpics.filter((epic) => epic.product === product);
          const totalTk = relevantEpics.reduce((acc, epic) => acc + (Number(epic.tk_approved_seconds) || 0), 0);
          const valueNode = card.querySelector('.product-summary-duration');
          const metaNode = card.querySelector('.product-summary-meta');
          if (valueNode) {{
            valueNode.dataset.seconds = String(totalTk);
            valueNode.dataset.hours = formatMetricDuration(totalTk, 'hours', true);
            valueNode.dataset.days = formatMetricDuration(totalTk, 'days', true);
          }}
          if (metaNode) {{
            metaNode.textContent = `${{relevantEpics.length.toLocaleString()}} RMIs/Epics • Total TK Approved`;
          }}
          const active = product === 'all' ? activeProducts.has('all') : activeProducts.has(product);
          card.classList.toggle('active', active);
          card.setAttribute('aria-pressed', String(active));
        }});
      }}

      function currentTkApprovedSeconds() {{
        if (!tkApprovedCard) return 0;
        const valueNode = tkApprovedCard.querySelector('.metric-value');
        if (!valueNode) return 0;
        return Number(valueNode.dataset.seconds || 0);
      }}

      function updateIdleCapacityCard() {{
        if (!idleCapacityCard || !availabilityValueNode) return;
        const valueNode = idleCapacityCard.querySelector('.metric-value');
        const metaNode = idleCapacityCard.querySelector('.metric-meta');
        if (!valueNode) return;

        const availabilitySeconds = Number(availabilityValueNode.dataset.seconds || 0);
        const tkSeconds = currentTkApprovedSeconds();
        const idleSeconds = availabilitySeconds - tkSeconds;

        valueNode.dataset.seconds = String(idleSeconds);
        valueNode.dataset.hours = formatMetricDuration(idleSeconds, 'hours', true);
        valueNode.dataset.days = formatMetricDuration(idleSeconds, 'days', true);

        if (metaNode) {{
          metaNode.textContent = idleSeconds >= 0
            ? 'Remaining availability after TK Approved'
            : 'TK Approved exceeds total availability';
        }}
      }}

      function updateMetricCards() {{
        const metrics = scopedMetrics();
        metricCards.forEach((card) => {{
          const key = card.dataset.metricKey;
          const metricType = card.dataset.metricType;
          const valueNode = card.querySelector('.metric-value');
          const rawValue = Number(metrics[key] || 0);
          if (!valueNode) {{
            return;
          }}
          if (metricType === 'count') {{
            valueNode.textContent = Math.round(rawValue).toLocaleString();
            return;
          }}
          if (anyTkMonthFilterActive() && key === 'tk_approved_seconds') {{
            return;
          }}
          valueNode.dataset.seconds = String(rawValue);
          valueNode.dataset.hours = formatMetricDuration(rawValue, 'hours', true);
          valueNode.dataset.days = formatMetricDuration(rawValue, 'days', true);
        }});
        updateProductSummaryCards();
        updateIdleCapacityCard();
        setUnits(activeUnit);
      }}

      function parseMonthKey(iso) {{
        if (!iso) return '';
        const match = /^(\\d{{4}})-(\\d{{2}})/.exec(String(iso));
        return match ? `${{match[1]}}-${{match[2]}}` : '';
      }}

      const MONTH_NAMES = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec'];

      function monthLabel(key) {{
        if (!key) return '';
        const [year, month] = key.split('-');
        const idx = Number(month) - 1;
        if (Number.isNaN(idx) || idx < 0 || idx > 11) return key;
        return `${{MONTH_NAMES[idx]}} ${{year}}`;
      }}

      function shiftMonthKey(key, offset) {{
        if (!key) return '';
        const [yearText, monthText] = key.split('-');
        const year = Number(yearText);
        const monthIndex = Number(monthText) - 1;
        if (Number.isNaN(year) || Number.isNaN(monthIndex)) return '';
        const shifted = new Date(year, monthIndex + offset, 1);
        return `${{shifted.getFullYear()}}-${{String(shifted.getMonth() + 1).padStart(2, '0')}}`;
      }}

      function isCrossMonthRange(item) {{
        const startKey = parseMonthKey(item.start_date);
        const dueKey = parseMonthKey(item.due_date);
        return Boolean(startKey && dueKey && startKey !== dueKey);
      }}

      function bucketMonthKey(item, allowedKeys) {{
        const startKey = parseMonthKey(item.start_date);
        const dueKey = parseMonthKey(item.due_date);
        let resolvedKey = '';
        if (startKey && dueKey && startKey === dueKey) {{
          resolvedKey = startKey;
        }} else if (dueKey) {{
          resolvedKey = dueKey;
        }} else {{
          resolvedKey = startKey;
        }}
        return allowedKeys.includes(resolvedKey) ? resolvedKey : '';
      }}

      function formatIsoDayLabel(iso) {{
        if (!iso) return '--';
        const match = /^(\\d{{4}})-(\\d{{2}})-(\\d{{2}})/.exec(String(iso));
        if (!match) return iso;
        const idx = Number(match[2]) - 1;
        if (Number.isNaN(idx) || idx < 0 || idx > 11) return iso;
        return `${{match[3]}}-${{MONTH_NAMES[idx]}}-${{match[1]}}`;
      }}

      function sumTkSeconds(list) {{
        return list.reduce((acc, epic) => acc + (Number(epic.tk_approved_seconds) || 0), 0);
      }}

      function currentMonthKey() {{
        const today = new Date();
        return `${{today.getFullYear()}}-${{String(today.getMonth() + 1).padStart(2, '0')}}`;
      }}

      function availableMonthKeys() {{
        const set = new Set();
        scopedEpicDetails().forEach((epic) => {{
          const dueKey = parseMonthKey(epic.due_date);
          const startKey = parseMonthKey(epic.start_date);
          if (tkThroughMonthEnabled) {{
            monthKeyRange(startKey, dueKey).forEach((k) => set.add(k));
          }} else {{
            if (dueKey) set.add(dueKey);
            if (startKey) set.add(startKey);
          }}
        }});
        return Array.from(set).sort();
      }}

      function populateMonthOptions() {{
        if (!tkMonthSelect) return;
        const keys = availableMonthKeys();
        tkMonthSelect.innerHTML = keys.length
          ? keys.map((key) => `<option value="${{key}}">${{monthLabel(key)}}</option>`).join('')
          : '<option value="">No months available</option>';
        if (keys.length) {{
          if (!tkMonthKey || !keys.includes(tkMonthKey)) {{
            const currentKey = currentMonthKey();
            tkMonthKey = keys.includes(currentKey) ? currentKey : keys[0];
          }}
          tkMonthSelect.value = tkMonthKey;
        }} else {{
          tkMonthKey = '';
        }}
      }}

      function effectiveTkMonthKey() {{
        const keys = availableMonthKeys();
        if (!keys.length) return '';
        if (anyTkMonthFilterActive()) {{
          if (!tkMonthKey || !keys.includes(tkMonthKey)) {{
            const currentKey = currentMonthKey();
            return keys.includes(currentKey) ? currentKey : keys[0];
          }}
          return tkMonthKey;
        }}
        const currentKey = currentMonthKey();
        return keys.includes(currentKey) ? currentKey : keys[0];
      }}

      function monthScopeDescription(monthKey) {{
        const label = monthLabel(monthKey);
        if (tkStartMonthEnabled && tkMonthEnabled) return `started and delivered in ${{label}}`;
        if (tkStartMonthEnabled) return `started in ${{label}}`;
        if (tkThroughMonthEnabled) return `through ${{label}}`;
        if (tkMonthEnabled) return `delivered in ${{label}}`;
        return `started or delivered in ${{label}}`;
      }}

      function anyTkMonthFilterActive() {{
        return tkStartMonthEnabled || tkMonthEnabled || tkThroughMonthEnabled;
      }}

      function monthKeyRange(startKey, endKey) {{
        if (!startKey && !endKey) return [];
        if (!startKey) return endKey ? [endKey] : [];
        if (!endKey) return [startKey];
        const [sy, sm] = startKey.split('-').map(Number);
        const [ey, em] = endKey.split('-').map(Number);
        if ([sy, sm, ey, em].some((v) => Number.isNaN(v))) return [];
        let sDate = new Date(sy, sm - 1, 1);
        let eDate = new Date(ey, em - 1, 1);
        if (sDate > eDate) {{ const tmp = sDate; sDate = eDate; eDate = tmp; }}
        const keys = [];
        const cursor = new Date(sDate.getFullYear(), sDate.getMonth(), 1);
        while (cursor <= eDate) {{
          keys.push(`${{cursor.getFullYear()}}-${{String(cursor.getMonth() + 1).padStart(2, '0')}}`);
          cursor.setMonth(cursor.getMonth() + 1);
        }}
        return keys;
      }}

      function epicMatchesMonthScope(epic, monthKey) {{
        if (!monthKey) return false;
        const startKey = parseMonthKey(epic.start_date);
        const dueKey = parseMonthKey(epic.due_date);
        if (tkThroughMonthEnabled) {{
          const sk = startKey || dueKey || '';
          const dk = dueKey || startKey || '';
          return sk <= monthKey && monthKey <= dk;
        }}
        if (tkStartMonthEnabled && tkMonthEnabled) {{
          return startKey === monthKey && dueKey === monthKey;
        }}
        if (tkStartMonthEnabled) {{
          return startKey === monthKey;
        }}
        if (tkMonthEnabled) {{
          return dueKey === monthKey;
        }}
        return startKey === monthKey || dueKey === monthKey;
      }}

      function epicsInMonthScope(monthKey) {{
        return scopedEpicDetails().filter((epic) => epicMatchesMonthScope(epic, monthKey));
      }}

      function epicsOutsideMonthScope(monthKey) {{
        return scopedEpicDetails().filter((epic) => !epicMatchesMonthScope(epic, monthKey));
      }}

      function monthWindowEntries(centerKey) {{
        return [
          {{ slot: 'previous', key: shiftMonthKey(centerKey, -1), label: monthLabel(shiftMonthKey(centerKey, -1)) }},
          {{ slot: 'selected', key: centerKey, label: monthLabel(centerKey) }},
          {{ slot: 'next', key: shiftMonthKey(centerKey, 1), label: monthLabel(shiftMonthKey(centerKey, 1)) }},
        ];
      }}

      function buildTkMonthEstimateAnalysis() {{
        const scopeMonthKey = effectiveTkMonthKey();
        const allMonthsMode = !anyTkMonthFilterActive();
        const entries = allMonthsMode
          ? availableMonthKeys().map((key) => ({{ slot: key, key, label: monthLabel(key) }}))
          : monthWindowEntries(scopeMonthKey);
        const allowedKeys = entries.map((entry) => entry.key).filter(Boolean);
        const scopeEpics = allMonthsMode ? scopedEpicDetails() : epicsInMonthScope(scopeMonthKey);
        const totalsByKey = Object.fromEntries(allowedKeys.map((key) => [key, 0]));
        const includedEpics = [];
        const excludedEpics = [];

        scopeEpics.forEach((epic) => {{
          const epicTotals = Object.fromEntries(allowedKeys.map((key) => [key, 0]));
          const reasons = [];
          let contributed = false;
          const stories = Array.isArray(epic.stories) ? epic.stories : [];

          if (!stories.length) {{
            reasons.push('No stories available for month estimate analysis.');
          }}

          stories.forEach((story) => {{
            const storyKey = story.story_key || story.title || 'Unknown story';
            const storyEstimate = Number(story.estimate_seconds) || 0;
            if (isCrossMonthRange(story)) {{
              const subtasks = Array.isArray(story.subtasks) ? story.subtasks : [];
              const usableSubtasks = subtasks.filter((subtask) => {{
                const subtaskEstimate = Number(subtask.estimate_seconds) || 0;
                const bucketKey = bucketMonthKey(subtask, allowedKeys);
                return subtaskEstimate > 0 && Boolean(bucketKey);
              }});
              if (!usableSubtasks.length) {{
                const startLabel = monthLabel(parseMonthKey(story.start_date));
                const dueLabel = monthLabel(parseMonthKey(story.due_date));
                reasons.push(`Story ${{storyKey}} spans ${{startLabel || '--'}} to ${{dueLabel || '--'}} but has no usable subtask estimates.`);
                return;
              }}
              usableSubtasks.forEach((subtask) => {{
                const bucketKey = bucketMonthKey(subtask, allowedKeys);
                const estimateSeconds = Number(subtask.estimate_seconds) || 0;
                if (!bucketKey || estimateSeconds <= 0) return;
                epicTotals[bucketKey] += estimateSeconds;
                contributed = true;
              }});
              return;
            }}

            const bucketKey = bucketMonthKey(story, allowedKeys);
            if (!bucketKey || storyEstimate <= 0) {{
              return;
            }}
            epicTotals[bucketKey] += storyEstimate;
            contributed = true;
          }});

          if (reasons.length || !contributed) {{
            const epicReasons = reasons.length
              ? reasons
              : ['No story or subtask estimates fall within the previous, selected, or next month window.'];
            excludedEpics.push({{
              jira_id: epic.jira_id,
              title: epic.title,
              product: epic.product,
              jira_url: epic.jira_url,
              reasons: epicReasons,
            }});
            return;
          }}

          Object.keys(epicTotals).forEach((key) => {{
            totalsByKey[key] += epicTotals[key];
          }});
          includedEpics.push(epic);
        }});

        return {{
          mode: allMonthsMode ? 'all' : 'scoped',
          entries: entries.map((entry) => ({{
            ...entry,
            seconds: Number(totalsByKey[entry.key] || 0),
          }})),
          scopeMonthKey,
          scopeEpics,
          includedEpics,
          excludedEpics,
        }};
      }}

      function analysisCardHtml(card) {{
        const featuredClass = card.featured ? ' featured' : '';
        const seconds = Number(card.seconds || 0);
        const hoursText = formatMetricDuration(seconds, 'hours', true);
        const daysText = formatMetricDuration(seconds, 'days', true);
        if (card.type === 'count') {{
          return `
            <section class="tk-month-analysis-card${{featuredClass}}">
              <div class="tk-month-analysis-card-label">${{escapeHtml(card.label)}}</div>
              <div class="tk-month-analysis-card-value">${{Number(card.value || 0)}}</div>
              <div class="tk-month-analysis-card-meta">${{escapeHtml(card.meta || '')}}</div>
            </section>
          `;
        }}
        return `
          <section class="tk-month-analysis-card${{featuredClass}}">
            <div class="tk-month-analysis-card-label">${{escapeHtml(card.label)}}</div>
            <div class="tk-month-analysis-card-value duration-value" data-seconds="${{seconds}}" data-hours="${{hoursText}}" data-days="${{daysText}}">${{hoursText}}</div>
            <div class="tk-month-analysis-card-meta">${{escapeHtml(card.meta || '')}}</div>
          </section>
        `;
      }}

      function analysisBarCardHtml(entry, highlight) {{
        const seconds = Number(entry.seconds || 0);
        const hoursText = formatMetricDuration(seconds, 'hours', true);
        const daysText = formatMetricDuration(seconds, 'days', true);
        return `
          <div class="tk-month-chart-bar${{highlight ? ' selected' : ''}}">
            <div class="tk-month-chart-bar-track"><div class="tk-month-chart-bar-fill" data-month-analysis-bar style="height:0%"></div></div>
            <div class="tk-month-chart-bar-label" data-month-analysis-label>${{escapeHtml(entry.label)}}</div>
            <div class="tk-month-chart-bar-value duration-value" data-month-analysis-chart-value data-seconds="${{seconds}}" data-hours="${{hoursText}}" data-days="${{daysText}}">${{hoursText}}</div>
          </div>
        `;
      }}

      function renderMonthAnalysisCards(analysis) {{
        if (!tkMonthAnalysisCardsContainer) return;
        const peakEntry = analysis.entries.reduce((best, entry) => (
          (Number(entry.seconds) || 0) > (Number(best.seconds) || 0) ? entry : best
        ), analysis.entries[0] || {{ key: '', seconds: 0 }});
        const totalSeconds = analysis.entries.reduce((sum, entry) => sum + (Number(entry.seconds) || 0), 0);
        let cards;
        if (analysis.mode === 'all') {{
          cards = [
            {{ label: 'All Available Months', seconds: totalSeconds, meta: 'Estimate allocated across every available month in the scoped epic ranges.', featured: true }},
            {{ label: 'Months In Chart', value: analysis.entries.length, meta: 'Inclusive month buckets generated from epic date ranges.', type: 'count' }},
            {{ label: 'Peak Month', seconds: Number(peakEntry.seconds || 0), meta: peakEntry.label ? `Highest allocated month: ${{peakEntry.label}}` : 'No estimates allocated yet.' }},
          ];
        }} else {{
          cards = analysis.entries.map((entry) => ({{
            label: `${{entry.label}} estimates`,
            seconds: Number(entry.seconds || 0),
            meta: 'Estimate allocated from story/subtask analysis',
            featured: entry.slot === 'selected',
            slot: entry.slot,
          }}));
        }}
        tkMonthAnalysisCardsContainer.innerHTML = cards.map((card) => analysisCardHtml(card)).join('');
      }}

      function renderMonthAnalysisBars(analysis) {{
        if (!tkMonthChartBarsContainer) return;
        const peakEntry = analysis.entries.reduce((best, entry) => (
          (Number(entry.seconds) || 0) > (Number(best.seconds) || 0) ? entry : best
        ), analysis.entries[0] || {{ key: '', seconds: 0 }});
        tkMonthChartBarsContainer.innerHTML = analysis.entries
          .map((entry) => analysisBarCardHtml(entry, analysis.mode === 'scoped' ? entry.slot === 'selected' : entry.key === peakEntry.key))
          .join('');
        const maxSeconds = Math.max(...analysis.entries.map((entry) => Number(entry.seconds) || 0), 0);
        Array.from(tkMonthChartBarsContainer.querySelectorAll('[data-month-analysis-bar]')).forEach((barNode, index) => {{
          const entry = analysis.entries[index];
          if (!entry) return;
          const height = maxSeconds > 0 ? Math.max(6, ((Number(entry.seconds) || 0) / maxSeconds) * 100) : 0;
          barNode.style.height = `${{height}}%`;
        }});
      }}

      function renderTkMonthEstimateAnalysis() {{
        if (!tkMonthAnalysisPanel) return;
        const analysis = buildTkMonthEstimateAnalysis();
        if (!analysis.entries.length) {{
          tkMonthAnalysisPanel.hidden = true;
          if (tkMonthAnalysisStatusNode) tkMonthAnalysisStatusNode.textContent = '';
          if (tkMonthAnalysisDrawerButton) tkMonthAnalysisDrawerButton.disabled = true;
          return;
        }}
        const productLabel = currentExecutiveScopeLabel();
        const scopeDescription = analysis.mode === 'all' ? 'across all available months' : monthScopeDescription(analysis.scopeMonthKey);
        renderMonthAnalysisCards(analysis);
        renderMonthAnalysisBars(analysis);
        if (tkMonthAnalysisIncludedNode) tkMonthAnalysisIncludedNode.textContent = String(analysis.includedEpics.length);
        if (tkMonthAnalysisExcludedNode) tkMonthAnalysisExcludedNode.textContent = String(analysis.excludedEpics.length);
        if (tkMonthAnalysisStatusNode) {{
          tkMonthAnalysisStatusNode.textContent = analysis.mode === 'all'
            ? `${{productLabel}} · ${{analysis.scopeEpics.length}} epic${{analysis.scopeEpics.length === 1 ? '' : 's'}} across all available months · ${{analysis.includedEpics.length}} included · ${{analysis.excludedEpics.length}} excluded`
            : `${{productLabel}} · ${{analysis.scopeEpics.length}} epic${{analysis.scopeEpics.length === 1 ? '' : 's'}} ${{scopeDescription}} · ${{analysis.includedEpics.length}} included · ${{analysis.excludedEpics.length}} excluded`;
        }}
        if (tkMonthAnalysisExclusionBody) {{
          tkMonthAnalysisExclusionBody.innerHTML = analysis.excludedEpics.length
            ? analysis.excludedEpics.map((epic) => {{
                const epicLabel = `<button type="button" class="tk-month-exclusion-link" data-open-excluded-epic="${{escapeHtml(epic.jira_id)}}" aria-label="Open exclusion details for ${{escapeHtml(epic.jira_id)}}">${{escapeHtml(epic.jira_id)}}</button> <span>${{escapeHtml(epic.title || '')}}</span>`;
                return `
                  <tr>
                    <td>${{epicLabel}}</td>
                    <td>${{escapeHtml(epic.product || '--')}}</td>
                    <td><div class="tk-month-exclusion-reason">${{escapeHtml(epic.reasons.join('\\n'))}}</div></td>
                  </tr>
                `;
              }}).join('')
            : '<tr><td colspan="3" class="empty-state tight">No excluded epics for the current analysis.</td></tr>';
        }}
        if (tkMonthAnalysisDrawerButton) tkMonthAnalysisDrawerButton.disabled = false;
        tkMonthAnalysisPanel.hidden = false;
        setUnits(activeUnit);
      }}

      function updateTkMonthCards() {{
        if (!tkApprovedCard) return;
        const labelNode = tkApprovedCard.querySelector('.metric-label');
        const metaNode = tkApprovedCard.querySelector('.metric-meta');
        const valueNode = tkApprovedCard.querySelector('.metric-value');
        const scopeMonthKey = effectiveTkMonthKey();
        const productLabel = currentExecutiveScopeLabel();
        if (!anyTkMonthFilterActive() || !scopeMonthKey) {{
          if (labelNode) labelNode.textContent = 'TK Approved';
          if (metaNode) metaNode.textContent = 'TK approved total for the selected epic set';
          tkApprovedCard.classList.remove('tk-month-active');
          if (tkMonthStatusNode) {{
            tkMonthStatusNode.textContent = anyTkMonthFilterActive() ? `${{productLabel}} · no months available` : `${{productLabel}} · all available months`;
          }}
          renderTkMonthEstimateAnalysis();
          updateIdleCapacityCard();
          return;
        }}
        const inMonth = epicsInMonthScope(scopeMonthKey);
        const inSeconds = sumTkSeconds(inMonth);
        const scopeDescription = monthScopeDescription(scopeMonthKey);
        if (labelNode) labelNode.textContent = `TK Approved (${{scopeDescription}})`;
        if (metaNode) metaNode.textContent = `${{inMonth.length}} epic${{inMonth.length === 1 ? '' : 's'}} ${{scopeDescription}}`;
        if (valueNode) {{
          valueNode.dataset.seconds = String(inSeconds);
          valueNode.dataset.hours = formatMetricDuration(inSeconds, 'hours', true);
          valueNode.dataset.days = formatMetricDuration(inSeconds, 'days', true);
        }}
        tkApprovedCard.classList.add('tk-month-active');
        if (tkMonthStatusNode) {{
          tkMonthStatusNode.textContent = `${{productLabel}} · ${{inMonth.length}} epic${{inMonth.length === 1 ? '' : 's'}} ${{scopeDescription}}`;
        }}
        renderTkMonthEstimateAnalysis();
        updateIdleCapacityCard();
        setUnits(activeUnit);
      }}

      function syncTkMonthUi() {{
        tkThroughMonthEnabled = Boolean(tkThroughMonthToggleInput && tkThroughMonthToggleInput.checked);
        if (tkThroughMonthEnabled) {{
          if (tkStartMonthToggleInput) {{
            tkStartMonthToggleInput.checked = false;
            tkStartMonthToggleInput.disabled = true;
          }}
          if (tkMonthToggleInput) {{
            tkMonthToggleInput.checked = false;
            tkMonthToggleInput.disabled = true;
          }}
        }} else {{
          if (tkStartMonthToggleInput) tkStartMonthToggleInput.disabled = false;
          if (tkMonthToggleInput) tkMonthToggleInput.disabled = false;
        }}
        tkStartMonthEnabled = Boolean(tkStartMonthToggleInput && tkStartMonthToggleInput.checked);
        tkMonthEnabled = Boolean(tkMonthToggleInput && tkMonthToggleInput.checked);
        tkJiraOnly = Boolean(tkJiraOnlyToggleInput && tkJiraOnlyToggleInput.checked);
        populateMonthOptions();
        if (tkMonthSelect) {{
          tkMonthSelect.disabled = !anyTkMonthFilterActive();
          if (tkMonthKey && !tkMonthSelect.disabled) tkMonthSelect.value = tkMonthKey;
        }}
        updateMetricCards();
        updateTkMonthCards();
      }}

      function escapeHtml(text) {{
        return String(text == null ? '' : text)
          .replace(/&/g, '&amp;')
          .replace(/</g, '&lt;')
          .replace(/>/g, '&gt;')
          .replace(/"/g, '&quot;')
          .replace(/'/g, '&#39;');
      }}

      function drawerEpicCardHtml(epic) {{
        const tkSeconds = Number(epic.tk_approved_seconds) || 0;
        const tkValue = activeUnit === 'days'
          ? `${{(tkSeconds / 28800).toLocaleString(undefined, {{ minimumFractionDigits: 2, maximumFractionDigits: 2 }})}} d`
          : `${{(tkSeconds / 3600).toLocaleString(undefined, {{ minimumFractionDigits: 2, maximumFractionDigits: 2 }})}} h`;
        const hasUrl = epic.jira_url && epic.jira_url !== '#';
        const jiraKey = hasUrl
          ? `<a class="drawer-epic-key" href="${{escapeHtml(epic.jira_url)}}" target="_blank" rel="noreferrer">${{escapeHtml(epic.jira_id)}}</a>`
          : `<span class="drawer-epic-key">${{escapeHtml(epic.jira_id)}}</span>`;
        return `
          <article class="drawer-epic-card" data-epic-key="${{escapeHtml(epic.jira_id)}}">
            <div class="drawer-epic-head">
              ${{jiraKey}}
              <span class="drawer-epic-product">${{escapeHtml(epic.product)}}</span>
            </div>
            <div class="drawer-epic-title">${{escapeHtml(epic.title)}}</div>
            <div class="drawer-epic-meta">
              <span><strong>Start</strong>${{escapeHtml(formatIsoDayLabel(epic.start_date))}}</span>
              <span><strong>Due</strong>${{escapeHtml(formatIsoDayLabel(epic.due_date))}}</span>
              <span><strong>Status</strong>${{escapeHtml(epic.status || '--')}}</span>
              <span><strong>Priority</strong>${{escapeHtml(epic.priority || '--')}}</span>
              <span><strong>Stories</strong>${{Number(epic.story_count || 0)}}</span>
            </div>
            <div class="drawer-epic-tk">
              <span class="drawer-epic-tk-label">TK Approved</span>
              <span class="drawer-epic-tk-value">${{escapeHtml(tkValue)}}</span>
            </div>
          </article>
        `;
      }}

      function drawerEstimateValue(seconds) {{
        return activeUnit === 'days'
          ? `${{(seconds / 28800).toLocaleString(undefined, {{ minimumFractionDigits: 2, maximumFractionDigits: 2 }})}} d`
          : `${{(seconds / 3600).toLocaleString(undefined, {{ minimumFractionDigits: 2, maximumFractionDigits: 2 }})}} h`;
      }}

      function drawerEpicEstimateCardHtml(epic) {{
        const seconds = Number(epic.jira_original_estimate_seconds) || 0;
        const hasUrl = epic.jira_url && epic.jira_url !== '#';
        const jiraKey = hasUrl
          ? `<a class="drawer-epic-key" href="${{escapeHtml(epic.jira_url)}}" target="_blank" rel="noreferrer">${{escapeHtml(epic.jira_id)}}</a>`
          : `<span class="drawer-epic-key">${{escapeHtml(epic.jira_id)}}</span>`;
        return `
          <article class="drawer-epic-card">
            <div class="drawer-epic-head">
              ${{jiraKey}}
              <span class="drawer-epic-product">${{escapeHtml(epic.product)}}</span>
            </div>
            <div class="drawer-epic-title">${{escapeHtml(epic.title)}}</div>
            <div class="drawer-epic-meta">
              <span><strong>Start</strong>${{escapeHtml(formatIsoDayLabel(epic.start_date))}}</span>
              <span><strong>Due</strong>${{escapeHtml(formatIsoDayLabel(epic.due_date))}}</span>
              <span><strong>Status</strong>${{escapeHtml(epic.status || '--')}}</span>
              <span><strong>Priority</strong>${{escapeHtml(epic.priority || '--')}}</span>
            </div>
            <div class="drawer-epic-tk">
              <span class="drawer-epic-tk-label">Epic Estimate</span>
              <span class="drawer-epic-tk-value">${{escapeHtml(drawerEstimateValue(seconds))}}</span>
            </div>
          </article>
        `;
      }}

      function drawerStoryCardHtml(story) {{
        const seconds = Number(story.estimate_seconds) || 0;
        const hasUrl = story.jira_url && story.jira_url !== '#';
        const jiraKey = hasUrl
          ? `<a class="drawer-epic-key" href="${{escapeHtml(story.jira_url)}}" target="_blank" rel="noreferrer">${{escapeHtml(story.story_key)}}</a>`
          : `<span class="drawer-epic-key">${{escapeHtml(story.story_key)}}</span>`;
        return `
          <article class="drawer-epic-card">
            <div class="drawer-epic-head">
              ${{jiraKey}}
              <span class="drawer-epic-product">${{escapeHtml(story.product)}}</span>
            </div>
            <div class="drawer-epic-title">${{escapeHtml(story.title)}}</div>
            <div class="drawer-epic-meta">
              <span><strong>Epic</strong>${{escapeHtml(story.epic_key || '--')}}</span>
              <span><strong>Start</strong>${{escapeHtml(formatIsoDayLabel(story.start_date))}}</span>
              <span><strong>Due</strong>${{escapeHtml(formatIsoDayLabel(story.due_date))}}</span>
              <span><strong>Status</strong>${{escapeHtml(story.status || '--')}}</span>
              <span><strong>Priority</strong>${{escapeHtml(story.priority || '--')}}</span>
            </div>
            <div class="drawer-epic-tk">
              <span class="drawer-epic-tk-label">Story Estimate</span>
              <span class="drawer-epic-tk-value">${{escapeHtml(drawerEstimateValue(seconds))}}</span>
            </div>
          </article>
        `;
      }}

      function drawerSubtaskCardHtml(subtask) {{
        const seconds = Number(subtask.estimate_seconds) || 0;
        const hasUrl = subtask.jira_url && subtask.jira_url !== '#';
        const jiraKey = hasUrl
          ? `<a class="drawer-epic-key" href="${{escapeHtml(subtask.jira_url)}}" target="_blank" rel="noreferrer">${{escapeHtml(subtask.issue_key)}}</a>`
          : `<span class="drawer-epic-key">${{escapeHtml(subtask.issue_key)}}</span>`;
        return `
          <article class="drawer-epic-card">
            <div class="drawer-epic-head">
              ${{jiraKey}}
              <span class="drawer-epic-product">${{escapeHtml(subtask.product)}}</span>
            </div>
            <div class="drawer-epic-title">${{escapeHtml(subtask.title)}}</div>
            <div class="drawer-epic-meta">
              <span><strong>Story</strong>${{escapeHtml(subtask.parent_story_key || '--')}}</span>
              <span><strong>Epic</strong>${{escapeHtml(subtask.epic_key || '--')}}</span>
              <span><strong>Start</strong>${{escapeHtml(formatIsoDayLabel(subtask.start_date))}}</span>
              <span><strong>Due</strong>${{escapeHtml(formatIsoDayLabel(subtask.due_date))}}</span>
              <span><strong>Status</strong>${{escapeHtml(subtask.status || '--')}}</span>
              <span><strong>Priority</strong>${{escapeHtml(subtask.priority || '--')}}</span>
            </div>
            <div class="drawer-epic-tk">
              <span class="drawer-epic-tk-label">Subtask Estimate</span>
              <span class="drawer-epic-tk-value">${{escapeHtml(drawerEstimateValue(seconds))}}</span>
            </div>
          </article>
        `;
      }}

      function drawerExcludedStorySectionHtml(epic, story) {{
        const storyCard = drawerStoryCardHtml({{
          ...story,
          product: epic.product,
          epic_key: epic.jira_id,
        }});
        const subtasks = Array.isArray(story.subtasks) ? story.subtasks : [];
        const subtaskCards = subtasks.length
          ? subtasks.map((subtask) => drawerSubtaskCardHtml({{
              ...subtask,
              product: epic.product,
              epic_key: epic.jira_id,
              parent_story_key: story.story_key,
            }})).join('')
          : '<div class="drawer-empty tight">No subtasks found for this story.</div>';
        return `
          <section class="drawer-group">
            <div class="drawer-group-title">Story</div>
            ${{storyCard}}
            <div class="drawer-group-title">Subtasks</div>
            <div class="drawer-nested-list">${{subtaskCards}}</div>
          </section>
        `;
      }}

      function openExcludedEpicDrawer(epicKey) {{
        if (!drawerEl || !epicKey) return;
        const analysis = buildTkMonthEstimateAnalysis();
        const excluded = analysis.excludedEpics.find((item) => item.jira_id === epicKey);
        const epic = scopedEpicDetails().find((item) => item.jira_id === epicKey);
        if (!excluded || !epic) return;

        const productLabel = currentExecutiveScopeLabel();
        const stories = Array.isArray(epic.stories) ? epic.stories : [];
        const subtaskCount = stories.reduce((total, story) => total + ((Array.isArray(story.subtasks) ? story.subtasks.length : 0)), 0);

        if (drawerEyebrowNode) drawerEyebrowNode.textContent = 'Excluded Epic Details';
        if (drawerTitleNode) drawerTitleNode.textContent = `${{epic.jira_id}} excluded from ${{monthLabel(tkMonthKey)}} analysis`;
        if (drawerSubtitleNode) drawerSubtitleNode.textContent = `Scope: ${{productLabel}}`;
        if (drawerSummaryNode) {{
          drawerSummaryNode.innerHTML = `
            <div class="drawer-summary-cell"><span>Reasons</span><strong>${{excluded.reasons.length}}</strong></div>
            <div class="drawer-summary-cell"><span>Stories</span><strong>${{stories.length}}</strong></div>
            <div class="drawer-summary-cell"><span>Subtasks</span><strong>${{subtaskCount}}</strong></div>
          `;
        }}
        if (drawerBodyNode) {{
          const epicCard = drawerEpicCardHtml(epic);
          const reasonsBlock = `
            <section class="drawer-note-card">
              <div class="drawer-note-title">Exclusion Reasons</div>
              <ul class="drawer-note-list">${{excluded.reasons.map((reason) => `<li>${{escapeHtml(reason)}}</li>`).join('')}}</ul>
            </section>
          `;
          const storySections = stories.length
            ? stories.map((story) => drawerExcludedStorySectionHtml(epic, story)).join('')
            : '<div class="drawer-empty">No stories found for this epic.</div>';
          drawerBodyNode.innerHTML = reasonsBlock + epicCard + storySections;
        }}

        drawerEl.hidden = false;
        drawerEl.setAttribute('aria-hidden', 'false');
        document.body.classList.add('drawer-open');
      }}

      function openEstimateDrawer(type) {{
        if (!drawerEl) return;
        const productLabel = currentExecutiveScopeLabel();
        const allowedEpicKeys = scopedEpicKeySet();

        if (type === 'epic-estimates') {{
          const scope = scopedEpicDetails();
          const withEstimate = scope.filter((e) => (Number(e.jira_original_estimate_seconds) || 0) > 0);
          withEstimate.sort((a, b) => (Number(b.jira_original_estimate_seconds) || 0) - (Number(a.jira_original_estimate_seconds) || 0));
          const totalSeconds = withEstimate.reduce((acc, e) => acc + (Number(e.jira_original_estimate_seconds) || 0), 0);
          if (drawerEyebrowNode) drawerEyebrowNode.textContent = 'Epic Estimates';
          if (drawerTitleNode) drawerTitleNode.textContent = 'Epics with a Jira original estimate';
          if (drawerSubtitleNode) drawerSubtitleNode.textContent = `Scope: ${{productLabel}}`;
          if (drawerSummaryNode) {{
            drawerSummaryNode.innerHTML = `
              <div class="drawer-summary-cell"><span>Epics</span><strong>${{withEstimate.length}}</strong></div>
              <div class="drawer-summary-cell"><span>Total Estimate</span><strong>${{escapeHtml(formatMetricDuration(totalSeconds, activeUnit))}}</strong></div>
            `;
          }}
          if (drawerBodyNode) {{
            drawerBodyNode.innerHTML = withEstimate.length
              ? withEstimate.map(drawerEpicEstimateCardHtml).join('')
              : '<div class="drawer-empty">No epics with estimates found.</div>';
          }}
        }} else if (type === 'story-estimates') {{
          const scope = storyDetails.filter((s) => productInScope(s.product) && (!tkJiraOnly || allowedEpicKeys.has(s.epic_key)));
          scope.sort((a, b) => (Number(b.estimate_seconds) || 0) - (Number(a.estimate_seconds) || 0));
          const totalSeconds = scope.reduce((acc, s) => acc + (Number(s.estimate_seconds) || 0), 0);
          if (drawerEyebrowNode) drawerEyebrowNode.textContent = 'Story Estimates';
          if (drawerTitleNode) drawerTitleNode.textContent = 'Stories with a Jira original estimate';
          if (drawerSubtitleNode) drawerSubtitleNode.textContent = `Scope: ${{productLabel}}`;
          if (drawerSummaryNode) {{
            drawerSummaryNode.innerHTML = `
              <div class="drawer-summary-cell"><span>Stories</span><strong>${{scope.length}}</strong></div>
              <div class="drawer-summary-cell"><span>Total Estimate</span><strong>${{escapeHtml(formatMetricDuration(totalSeconds, activeUnit))}}</strong></div>
            `;
          }}
          if (drawerBodyNode) {{
            drawerBodyNode.innerHTML = scope.length
              ? scope.map(drawerStoryCardHtml).join('')
              : '<div class="drawer-empty">No stories with estimates found.</div>';
          }}
        }} else if (type === 'subtask-estimates') {{
          const scope = subtaskDetails.filter((s) => productInScope(s.product) && (!tkJiraOnly || allowedEpicKeys.has(s.epic_key)));
          scope.sort((a, b) => (Number(b.estimate_seconds) || 0) - (Number(a.estimate_seconds) || 0));
          const totalSeconds = scope.reduce((acc, s) => acc + (Number(s.estimate_seconds) || 0), 0);
          if (drawerEyebrowNode) drawerEyebrowNode.textContent = 'Subtask Estimates';
          if (drawerTitleNode) drawerTitleNode.textContent = 'Subtasks with a Jira original estimate';
          if (drawerSubtitleNode) drawerSubtitleNode.textContent = `Scope: ${{productLabel}}`;
          if (drawerSummaryNode) {{
            drawerSummaryNode.innerHTML = `
              <div class="drawer-summary-cell"><span>Subtasks</span><strong>${{scope.length}}</strong></div>
              <div class="drawer-summary-cell"><span>Total Estimate</span><strong>${{escapeHtml(formatMetricDuration(totalSeconds, activeUnit))}}</strong></div>
            `;
          }}
          if (drawerBodyNode) {{
            drawerBodyNode.innerHTML = scope.length
              ? scope.map(drawerSubtaskCardHtml).join('')
              : '<div class="drawer-empty">No subtasks with estimates found.</div>';
          }}
        }}

        drawerEl.hidden = false;
        drawerEl.setAttribute('aria-hidden', 'false');
        document.body.classList.add('drawer-open');
      }}

      function openDrawer(config) {{
        const list = Array.isArray(config.epics) ? config.epics : [];
        const totalSeconds = sumTkSeconds(list);
        if (drawerEyebrowNode) drawerEyebrowNode.textContent = config.eyebrow || '';
        if (drawerTitleNode) drawerTitleNode.textContent = config.title || 'Selected Epics';
        if (drawerSubtitleNode) drawerSubtitleNode.textContent = config.subtitle || '';
        if (drawerSummaryNode) {{
          drawerSummaryNode.innerHTML = `
            <div class="drawer-summary-cell"><span>Epics</span><strong>${{list.length}}</strong></div>
            <div class="drawer-summary-cell"><span>TK Approved Total</span><strong>${{escapeHtml(formatMetricDuration(totalSeconds, activeUnit))}}</strong></div>
          `;
        }}
        if (drawerBodyNode) {{
          drawerBodyNode.innerHTML = list.length
            ? list.map(drawerEpicCardHtml).join('')
            : '<div class="drawer-empty">No epics match this selection.</div>';
        }}
        drawerEl.hidden = false;
        drawerEl.setAttribute('aria-hidden', 'false');
        document.body.classList.add('drawer-open');
      }}

      function closeDrawer() {{
        if (!drawerEl) return;
        drawerEl.hidden = true;
        drawerEl.setAttribute('aria-hidden', 'true');
        document.body.classList.remove('drawer-open');
      }}

      function openTkInScopeDrawer() {{
        const scopeMonthKey = effectiveTkMonthKey();
        const productLabel = currentExecutiveScopeLabel();
        if (scopeMonthKey) {{
          const scopeDescription = monthScopeDescription(scopeMonthKey);
          const inMonth = epicsInMonthScope(scopeMonthKey);
          openDrawer({{
            eyebrow: 'TK Approved · month scope',
            title: `Epics ${{scopeDescription}}`,
            subtitle: `Scope: ${{productLabel}}`,
            epics: inMonth,
          }});
        }} else {{
          const scope = scopedEpicDetails();
          openDrawer({{
            eyebrow: 'TK Approved Total',
            title: 'Epics contributing to TK Approved',
            subtitle: `Scope: ${{productLabel}}`,
            epics: scope,
          }});
        }}
      }}

      function openTkOutsideDrawer() {{
        const scopeMonthKey = effectiveTkMonthKey();
        if (!scopeMonthKey || !(tkStartMonthEnabled || tkMonthEnabled)) return;
        const productLabel = currentExecutiveScopeLabel();
        const scopeDescription = monthScopeDescription(scopeMonthKey);
        const outside = epicsOutsideMonthScope(scopeMonthKey);
        openDrawer({{
          eyebrow: 'TK Approved · outside month scope',
          title: `Epics outside ${{scopeDescription}}`,
          subtitle: `Scope: ${{productLabel}}`,
          epics: outside,
        }});
      }}

      function populateCapacityMonths() {{
        if (!capacityMonthSelect || !Array.isArray(capacityMonths) || !capacityMonths.length) return;
        capacityMonthSelect.innerHTML = capacityMonths
          .map((entry) => `<option value="${{entry.key}}" data-working-days="${{entry.working_days}}">${{entry.label}} (${{entry.working_days}} working days)</option>`)
          .join('');
        const today = new Date();
        const currentKey = today.getFullYear() === 2026
          ? `2026-${{String(today.getMonth() + 1).padStart(2, '0')}}`
          : capacityMonths[0].key;
        const match = capacityMonths.find((entry) => entry.key === currentKey);
        capacityMonthSelect.value = match ? currentKey : capacityMonths[0].key;
      }}

      function updateCapacityResult() {{
        if (!capacityValueNode || !availabilityValueNode || !capacityMonthSelect || !capacityEmployeesInput || !capacityLeavesInput) return;
        const employees = Math.max(0, Math.floor(Number(capacityEmployeesInput.value || 0)));
        const leaveSeconds = syncLeavesSecondsFromInput();
        const entry = capacityMonths.find((m) => m.key === capacityMonthSelect.value) || capacityMonths[0] || {{ working_days: 0, label: '' }};
        const workingDays = Number(entry.working_days) || 0;
        const totalSeconds = employees * workingDays * 28800; // 8 working hours per day
        const availabilitySeconds = Math.max(0, totalSeconds - leaveSeconds);
        capacityValueNode.dataset.seconds = String(totalSeconds);
        capacityValueNode.dataset.hours = formatMetricDuration(totalSeconds, 'hours');
        capacityValueNode.dataset.days = formatMetricDuration(totalSeconds, 'days');
        availabilityValueNode.dataset.seconds = String(availabilitySeconds);
        availabilityValueNode.dataset.hours = formatMetricDuration(availabilitySeconds, 'hours');
        availabilityValueNode.dataset.days = formatMetricDuration(availabilitySeconds, 'days');
        if (capacityMetaNode) {{
          capacityMetaNode.textContent = employees === 0
            ? `${{workingDays}} working days in ${{entry.label}}`
            : `${{employees}} employee${{employees === 1 ? '' : 's'}} × ${{workingDays}} working days × 8 h`;
        }}
        if (availabilityMetaNode) {{
          availabilityMetaNode.textContent = 'Capacity minus leaves';
        }}
        updateIdleCapacityCard();
        setUnits(activeUnit);
      }}

      function setView(viewId) {{
        viewButtons.forEach((button) => {{
          const active = button.dataset.view === viewId;
          button.classList.toggle('active', active);
          button.setAttribute('aria-pressed', String(active));
        }});
        views.forEach((view) => {{
          view.hidden = view.id !== viewId;
        }});
      }}

      function updateFilters() {{
        let visibleEpicCount = 0;

        productButtons.forEach((button) => {{
          const product = button.dataset.product || 'all';
          const active = product === 'all' ? activeProducts.has('all') : activeProducts.has(product);
          button.classList.toggle('active', active);
          button.setAttribute('aria-pressed', String(active));
        }});

        epicSummaryRows.forEach((row) => {{
          const text = (row.dataset.search || '').toLowerCase();
          const matchesProduct = productInScope(row.dataset.product);
          const matchesQuery = !activeQuery || text.includes(activeQuery);
          const visible = matchesProduct && matchesQuery;
          row.hidden = !visible;
          row.dataset.hidden = visible ? 'false' : 'true';
          const detailRow = row.nextElementSibling;
          if (detailRow && detailRow.classList.contains('epic-detail-row')) {{
            if (!visible) {{
              detailRow.hidden = true;
            }} else {{
              const button = row.querySelector('.row-toggle');
              detailRow.hidden = button ? button.getAttribute('aria-expanded') !== 'true' : true;
            }}
            detailRow.dataset.hidden = visible ? 'false' : 'true';
          }}
          if (visible) {{
            visibleEpicCount += 1;
          }}
        }});

        const visibleProducts = new Set();
        ganttEpicRows.forEach((row) => {{
          const text = (row.dataset.search || '').toLowerCase();
          const matchesProduct = productInScope(row.dataset.product);
          const matchesQuery = !activeQuery || text.includes(activeQuery);
          const visible = matchesProduct && matchesQuery;
          row.style.display = visible ? '' : 'none';
          if (visible) {{
            visibleProducts.add(row.dataset.product);
          }}
        }});

        ganttProductSections.forEach((section) => {{
          const visible = visibleProducts.has(section.dataset.product);
          section.dataset.hidden = visible ? 'false' : 'true';
        }});

        updateMetricCards();
        updateTkMonthCards();
        if (searchStatus) {{
          searchStatus.textContent = activeQuery
            ? `${{visibleEpicCount}} epic${{visibleEpicCount === 1 ? '' : 's'}} match "${{searchInput ? searchInput.value.trim() : ''}}".`
            : '';
        }}
      }}

      function setProduct(product) {{
        if (product === 'all') {{
          activeProducts = new Set(['all']);
          updateFilters();
          return;
        }}

        if (activeProducts.has('all')) {{
          activeProducts = new Set();
        }}

        if (activeProducts.has(product)) {{
          activeProducts.delete(product);
        }} else {{
          activeProducts.add(product);
        }}

        if (!activeProducts.size) {{
          activeProducts = new Set(['all']);
        }}

        updateFilters();
      }}

      function setSearch(query) {{
        activeQuery = query.trim().toLowerCase();
        updateFilters();
      }}

      viewButtons.forEach((button) => {{
        button.addEventListener('click', () => setView(button.dataset.view));
      }});
      unitButtons.forEach((button) => {{
        button.addEventListener('click', () => setUnits(button.dataset.unit));
      }});
      rowToggleButtons.forEach((button) => {{
        button.textContent = button.getAttribute('aria-expanded') === 'true' ? '-' : '+';
        button.addEventListener('click', () => {{
          const expanded = button.getAttribute('aria-expanded') === 'true';
          setEpicExpansion(button, !expanded);
        }});
      }});
      productButtons.forEach((button) => {{
        button.addEventListener('click', () => setProduct(button.dataset.product));
      }});
      productSummaryCards.forEach((card) => {{
        const handleSelect = () => setProduct(card.dataset.productSummary || 'all');
        card.addEventListener('click', handleSelect);
        card.addEventListener('keydown', (event) => {{
          if (event.key === 'Enter' || event.key === ' ') {{
            event.preventDefault();
            handleSelect();
          }}
        }});
      }});
      if (searchInput) {{
        searchInput.addEventListener('input', () => setSearch(searchInput.value));
      }}
      if (searchClearButton && searchInput) {{
        searchClearButton.addEventListener('click', () => {{
          searchInput.value = '';
          setSearch('');
          searchInput.focus();
        }});
      }}

      populateMonthOptions();

      if (tkStartMonthToggleInput) {{
        ['change', 'input', 'click'].forEach((eventName) => {{
          tkStartMonthToggleInput.addEventListener(eventName, syncTkMonthUi);
        }});
      }}

      if (tkMonthToggleInput) {{
        ['change', 'input', 'click'].forEach((eventName) => {{
          tkMonthToggleInput.addEventListener(eventName, syncTkMonthUi);
        }});
      }}

      if (tkThroughMonthToggleInput) {{
        ['change', 'input', 'click'].forEach((eventName) => {{
          tkThroughMonthToggleInput.addEventListener(eventName, syncTkMonthUi);
        }});
      }}

      if (tkJiraOnlyToggleInput) {{
        ['change', 'input', 'click'].forEach((eventName) => {{
          tkJiraOnlyToggleInput.addEventListener(eventName, syncTkMonthUi);
        }});
      }}

      if (tkMonthAnalysisDrawerButton) {{
        tkMonthAnalysisDrawerButton.addEventListener('click', () => {{
          const analysis = buildTkMonthEstimateAnalysis();
          const productLabel = currentExecutiveScopeLabel();
          const scopeDescription = analysis.mode === 'all' ? 'across all available months' : monthScopeDescription(analysis.scopeMonthKey);
          if (drawerEyebrowNode) drawerEyebrowNode.textContent = `${{productLabel}} · Month Estimate Analysis`;
          if (drawerTitleNode) drawerTitleNode.textContent = `Epics ${{scopeDescription}}`;
          if (drawerSubtitleNode) drawerSubtitleNode.textContent = `${{analysis.includedEpics.length}} included · ${{analysis.excludedEpics.length}} excluded`;
          if (drawerSummaryNode) drawerSummaryNode.innerHTML = '';
          if (drawerBodyNode) {{
            const includedHtml = analysis.includedEpics.length
              ? '<h3 style="margin:0 0 .5rem">Included Epics</h3>' + analysis.includedEpics.map((epic) => drawerEpicCardHtml(epic)).join('')
              : '';
            const excludedHtml = analysis.excludedEpics.length
              ? '<h3 style="margin:1rem 0 .5rem">Excluded Epics</h3>' + analysis.excludedEpics.map((epic) => drawerEpicCardHtml(epic)).join('')
              : '';
            drawerBodyNode.innerHTML = includedHtml + excludedHtml;
          }}
          if (drawerEl) {{ drawerEl.hidden = false; }}
          setUnits(activeUnit);
        }});
      }}

      if (tkMonthSelect) {{
        tkMonthSelect.addEventListener('change', () => {{
          tkMonthKey = tkMonthSelect.value;
          updateTkMonthCards();
        }});
      }}

      syncTkMonthUi();
      populateCapacityMonths();
      updateCapacityResult();

      if (capacityEmployeesInput) {{
        capacityEmployeesInput.addEventListener('input', updateCapacityResult);
        capacityEmployeesInput.addEventListener('change', updateCapacityResult);
      }}
      if (capacityLeavesInput) {{
        capacityLeavesInput.addEventListener('input', updateCapacityResult);
        capacityLeavesInput.addEventListener('change', updateCapacityResult);
      }}
      if (capacityMonthSelect) {{
        capacityMonthSelect.addEventListener('change', updateCapacityResult);
      }}

      if (tkMonthAnalysisExclusionBody) {{
        tkMonthAnalysisExclusionBody.addEventListener('click', (event) => {{
          const target = event.target instanceof Element ? event.target.closest('[data-open-excluded-epic]') : null;
          if (!target) return;
          openExcludedEpicDrawer(target.getAttribute('data-open-excluded-epic') || '');
        }});
      }}

      openTkInScopeDrawer = function () {{
        const productLabel = currentExecutiveScopeLabel();
        const scope = scopedEpicDetails();
        openDrawer({{
          eyebrow: 'TK Approved Total',
          title: 'Epics contributing to TK Approved',
          subtitle: `Scope: ${{productLabel}}`,
          epics: scope,
        }});
      }};

      if (tkApprovedCard) {{
        tkApprovedCard.addEventListener('click', openTkInScopeDrawer);
        tkApprovedCard.addEventListener('keydown', (event) => {{
          if (event.key === 'Enter' || event.key === ' ') {{
            event.preventDefault();
            openTkInScopeDrawer();
          }}
        }});
      }}

      ['epic-estimates', 'story-estimates', 'subtask-estimates'].forEach((context) => {{
        const card = document.querySelector(`.metric-card[data-metric-context="${{context}}"]`);
        if (card) {{
          card.addEventListener('click', () => openEstimateDrawer(context));
          card.addEventListener('keydown', (event) => {{
            if (event.key === 'Enter' || event.key === ' ') {{
              event.preventDefault();
              openEstimateDrawer(context);
            }}
          }});
        }}
      }});

      if (drawerEl) {{
        drawerEl.querySelectorAll('[data-drawer-close]').forEach((el) => {{
          el.addEventListener('click', closeDrawer);
        }});
      }}

      document.addEventListener('keydown', (event) => {{
        if (event.key === 'Escape' && drawerEl && !drawerEl.hidden) {{
          closeDrawer();
        }}
      }});

      /* ---- RMI Estimation & Scheduling Table ---- */
      const PRODUCT_COLORS_JS = {json.dumps(PRODUCT_COLORS)};
      const rmiScheduleBody = document.getElementById('rmi-schedule-body');
      const rmiScheduleFoot = document.getElementById('rmi-schedule-foot');
      const rmiScheduleYearSelect = document.getElementById('rmi-schedule-year');
      const rmiProductCardsContainer = document.getElementById('rmi-sched-product-cards');
      const rmiJiraOnlyToggle = document.getElementById('rmi-jira-only-toggle');
      let rmiSelectedYear = {rmi_schedule_initial_year};
      let rmiActiveProduct = 'all';
      let rmiJiraOnly = true;

      function rmiFilteredRecords() {{
        let records = rmiScheduleRecords;
        if (rmiJiraOnly) {{
          records = records.filter((e) => e.jira_populated);
        }}
        if (rmiActiveProduct !== 'all') {{
          records = records.filter((e) => e.product === rmiActiveProduct);
        }}
        return records;
      }}

      function rmiScheduleInit() {{
        if (!rmiScheduleYearSelect || !rmiScheduleYears.length) return;
        rmiScheduleYears.forEach((year) => {{
          const opt = document.createElement('option');
          opt.value = String(year);
          opt.textContent = String(year);
          if (year === rmiSelectedYear) opt.selected = true;
          rmiScheduleYearSelect.appendChild(opt);
        }});
        rmiScheduleYearSelect.addEventListener('change', () => {{
          rmiSelectedYear = Number(rmiScheduleYearSelect.value);
          renderRmiScheduleTable();
        }});
        if (rmiJiraOnlyToggle) {{
          rmiJiraOnlyToggle.addEventListener('change', () => {{
            rmiJiraOnly = rmiJiraOnlyToggle.checked;
            renderRmiProductCards();
            renderRmiScheduleTable();
          }});
        }}
        renderRmiProductCards();
        renderRmiScheduleTable();
        // Measure first header row height for sticky offset of second row
        const headerGroupRow = document.querySelector('.rmi-sched-header-groups');
        if (headerGroupRow) {{
          const h = headerGroupRow.getBoundingClientRect().height;
          document.querySelector('.rmi-schedule-table').style.setProperty('--rmi-sched-row1-h', h + 'px');
        }}
      }}

      function renderRmiProductCards() {{
        if (!rmiProductCardsContainer) return;
        const base = rmiJiraOnly
          ? rmiScheduleRecords.filter((e) => e.jira_populated)
          : rmiScheduleRecords;
        const tkByProduct = {{}};
        const countByProduct = {{}};
        let totalTk = 0;
        base.forEach((epic) => {{
          const tk = (epic.tk_approved_days || 0) * 28800;
          tkByProduct[epic.product] = (tkByProduct[epic.product] || 0) + tk;
          countByProduct[epic.product] = (countByProduct[epic.product] || 0) + 1;
          totalTk += tk;
        }});
        const products = Object.keys(tkByProduct).sort();
        const totalCount = base.length;
        const allColor = '#102033';
        const allHours = formatMetricDuration(totalTk, 'hours', true);
        const allDays = formatMetricDuration(totalTk, 'days', true);
        let html = `<section class="rmi-sched-pcard active" data-rmi-pcard="all" style="--product-accent:${{allColor}}" role="button" tabindex="0" aria-pressed="true">`;
        html += `<div class="rmi-sched-pcard-label">All Products</div>`;
        html += `<div class="rmi-sched-pcard-value duration-value" data-seconds="${{totalTk}}" data-hours="${{allHours}}" data-days="${{allDays}}">${{allHours}}</div>`;
        html += `<div class="rmi-sched-pcard-meta">${{totalCount.toLocaleString()}} RMIs/Epics • Total TK Approved</div></section>`;
        products.forEach((product) => {{
          const color = PRODUCT_COLORS_JS[product] || '#475569';
          const seconds = tkByProduct[product] || 0;
          const count = countByProduct[product] || 0;
          const hText = formatMetricDuration(seconds, 'hours', true);
          const dText = formatMetricDuration(seconds, 'days', true);
          html += `<section class="rmi-sched-pcard" data-rmi-pcard="${{escapeHtml(product)}}" style="--product-accent:${{color}}" role="button" tabindex="0" aria-pressed="false">`;
          html += `<div class="rmi-sched-pcard-label">${{escapeHtml(product)}}</div>`;
          html += `<div class="rmi-sched-pcard-value duration-value" data-seconds="${{seconds}}" data-hours="${{hText}}" data-days="${{dText}}">${{hText}}</div>`;
          html += `<div class="rmi-sched-pcard-meta">${{count.toLocaleString()}} RMIs/Epics • Total TK Approved</div></section>`;
        }});
        rmiProductCardsContainer.innerHTML = html;
        rmiProductCardsContainer.querySelectorAll('.rmi-sched-pcard').forEach((card) => {{
          const handleClick = () => {{
            rmiActiveProduct = card.dataset.rmiPcard || 'all';
            rmiProductCardsContainer.querySelectorAll('.rmi-sched-pcard').forEach((c) => {{
              const isActive = c.dataset.rmiPcard === rmiActiveProduct;
              c.classList.toggle('active', isActive);
              c.setAttribute('aria-pressed', String(isActive));
            }});
            renderRmiScheduleTable();
          }};
          card.addEventListener('click', handleClick);
          card.addEventListener('keydown', (e) => {{ if (e.key === 'Enter' || e.key === ' ') {{ e.preventDefault(); handleClick(); }} }});
        }});
        setUnits(activeUnit);
      }}

      function rmiBucketEpicMonths(epic) {{
        /** Bucket an epic's stories/subtasks into YYYY-MM keys and return a totals map (seconds). */
        const totals = {{}};
        const stories = Array.isArray(epic.stories) ? epic.stories : [];
        stories.forEach((story) => {{
          const storyEstimate = Number(story.estimate_seconds) || 0;
          if (isCrossMonthRange(story)) {{
            const subtasks = Array.isArray(story.subtasks) ? story.subtasks : [];
            subtasks.forEach((subtask) => {{
              const est = Number(subtask.estimate_seconds) || 0;
              if (est <= 0) return;
              const sk = parseMonthKey(subtask.start_date);
              const dk = parseMonthKey(subtask.due_date);
              let mk = '';
              if (sk && dk && sk === dk) mk = sk;
              else if (dk) mk = dk;
              else mk = sk;
              if (!mk) return;
              totals[mk] = (totals[mk] || 0) + est;
            }});
            return;
          }}
          if (storyEstimate <= 0) return;
          const sk = parseMonthKey(story.start_date);
          const dk = parseMonthKey(story.due_date);
          let mk = '';
          if (sk && dk && sk === dk) mk = sk;
          else if (dk) mk = dk;
          else mk = sk;
          if (!mk) return;
          totals[mk] = (totals[mk] || 0) + storyEstimate;
        }});
        return totals;
      }}

      function rmiFormatValue(seconds) {{
        /** Format seconds as days (integer) or hours depending on active unit. */
        if (!seconds) return '';
        if (activeUnit === 'days') return Math.round(seconds / 28800).toLocaleString();
        return Math.round(seconds / 3600).toLocaleString();
      }}

      function rmiFormatDays(days) {{
        /** Format a workbook man-days value for display using the active unit. */
        const visible = rmiDisplayDays(days);
        return visible ? visible.toLocaleString() : '';
      }}

      function rmiDisplayDays(days) {{
        /** Return the visible integer shown for a workbook man-days value. */
        if (!days) return 0;
        if (activeUnit === 'days') return Math.round(days);
        return Math.round(days * 8);
      }}

      function renderRmiScheduleTable() {{
        if (!rmiScheduleBody || !rmiScheduleFoot) return;
        const yearStr = String(rmiSelectedYear);
        const monthKeys = Array.from({{length: 12}}, (_, i) => `${{yearStr}}-${{String(i + 1).padStart(2, '0')}}`);
        const filtered = rmiFilteredRecords();
        const productOrder = [];
        const byProduct = {{}};
        filtered.forEach((epic) => {{
          if (!byProduct[epic.product]) {{
            byProduct[epic.product] = [];
            productOrder.push(epic.product);
          }}
          byProduct[epic.product].push(epic);
        }});
        productOrder.sort();
        let bodyHtml = '';
        const grandTotalMonths = new Array(12).fill(0);
        let grandMl = 0;
        let grandTk = 0;
        productOrder.forEach((product) => {{
          const color = PRODUCT_COLORS_JS[product] || '#475569';
          bodyHtml += `<tr class="rmi-sched-product-group"><td></td><td class="rmi-sched-group-label" style="border-left-color:${{color}}">${{escapeHtml(product)}}</td>${{Array(16).fill('<td></td>').join('')}}</tr>`;
          const epics = byProduct[product];
          const subtotalMonths = new Array(12).fill(0);
          let subtotalMl = 0;
          let subtotalTk = 0;
          let rowNum = 0;
          epics.forEach((epic) => {{
            rowNum++;
            const buckets = rmiBucketEpicMonths(epic);
            const statusLower = (epic.status || '').toLowerCase();
            const jiraLink = epic.jira_url && epic.jira_url !== '#'
              ? ` <a class="rmi-sched-jira-link" href="${{epic.jira_url}}" target="_blank" rel="noopener" title="Open in Jira">J</a>`
              : '';
            let cells = `<td>${{rowNum}}</td>`;
            cells += `<td class="rmi-sched-cell-rmi">${{escapeHtml(epic.roadmap_item)}}${{jiraLink}}</td>`;
            cells += `<td>${{escapeHtml(epic.product)}}</td>`;
            cells += `<td><span class="rmi-sched-status-pill" data-status-lower="${{escapeHtml(statusLower)}}">${{escapeHtml(epic.status || '—')}}</span></td>`;
            cells += `<td>${{rmiFormatDays(epic.most_likely_days)}}</td>`;
            cells += `<td>${{rmiFormatDays(epic.tk_approved_days)}}</td>`;
            subtotalMl += rmiDisplayDays(epic.most_likely_days);
            subtotalTk += rmiDisplayDays(epic.tk_approved_days);
            monthKeys.forEach((mk, idx) => {{
              const val = buckets[mk] || 0;
              subtotalMonths[idx] += val;
              cells += `<td>${{rmiFormatValue(val)}}</td>`;
            }});
            bodyHtml += `<tr class="rmi-sched-epic-row" data-product="${{escapeHtml(epic.product)}}">${{cells}}</tr>`;
          }});
          // Subtotal row
          let subtotalCells = `<td></td><td style="border-left-color:${{color}}">${{escapeHtml(product)}} Subtotal</td><td></td><td></td>`;
          subtotalCells += `<td>${{subtotalMl ? subtotalMl.toLocaleString() : ''}}</td>`;
          subtotalCells += `<td>${{subtotalTk ? subtotalTk.toLocaleString() : ''}}</td>`;
          subtotalMonths.forEach((val, idx) => {{
            grandTotalMonths[idx] += val;
            subtotalCells += `<td>${{rmiFormatValue(val)}}</td>`;
          }});
          grandMl += subtotalMl;
          grandTk += subtotalTk;
          bodyHtml += `<tr class="rmi-sched-product-subtotal">${{subtotalCells}}</tr>`;
        }});
        rmiScheduleBody.innerHTML = bodyHtml;
        // Grand total
        const totalEpics = filtered.length;
        let footCells = `<td>${{totalEpics}}</td><td>Grand Total</td><td></td><td></td>`;
        footCells += `<td>${{grandMl ? grandMl.toLocaleString() : ''}}</td>`;
        footCells += `<td>${{grandTk ? grandTk.toLocaleString() : ''}}</td>`;
        grandTotalMonths.forEach((val) => {{
          footCells += `<td>${{rmiFormatValue(val)}}</td>`;
        }});
        rmiScheduleFoot.innerHTML = `<tr class="rmi-sched-grand-total">${{footCells}}</tr>`;
      }}

      setUnits('hours');
      setProduct('all');
      setView('table-view');
      syncTkMonthUi();
      rmiScheduleInit();
    }})();
  </script>
</body>
</html>
"""


def generate_html_report(db_path: Path, output_path: Path) -> Path:
    report_data = load_report_data(db_path)
    html = render_html(report_data, db_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(html, encoding="utf-8")
    return output_path


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate an HTML gantt dashboard from the RMI Jira SQLite database.")
    parser.add_argument("--db", type=Path, default=DEFAULT_DB_PATH, help="SQLite database produced by extract_rmi_jira_to_sqlite.py")
    parser.add_argument("--output", type=Path, default=DEFAULT_HTML_PATH, help="HTML output path")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    output_path = generate_html_report(args.db.resolve(), args.output.resolve())
    print(f"html: {output_path}")


if __name__ == "__main__":
    main()
