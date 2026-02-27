"""
Generate IPP Meeting dashboard HTML from Epics Planner DB + Jira exports.
"""
from __future__ import annotations

import argparse
import json
import os
import sqlite3
from datetime import date, datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

from export_ipp_phase_breakdown import (
    PHASE_COLUMNS,
    SMALL_MIN_WIDTH_PCT,
    _clamp_percent,
    _compute_phase_geometry_for_record,
    _compute_roadmap_axis,
    _normalize_jira_link,
    _parse_iso_date,
    _to_number,
)

DEFAULT_HTML_OUTPUT = "ipp_meeting_dashboard.html"
DEFAULT_TEMPLATE = "ipp_meeting_dashboard_template.html"
DEFAULT_SETTINGS_DB = "assignee_hours_capacity.db"
DEFAULT_WORK_ITEMS_XLSX = "1_jira_work_items_export.xlsx"
PHASE_NAMES = ["Research/URS", "DDS", "Development", "SQA", "User Manual", "Production"]

PLAN_KEY_BY_PHASE = {
    "Research/URS": "research_urs_plan",
    "DDS": "dds_plan",
    "Development": "development_plan",
    "SQA": "sqa_plan",
    "User Manual": "",
    "Production": "production_plan",
}


def _as_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _resolve_path(path_value: str, base_dir: Path) -> Path:
    path = Path(path_value)
    if path.is_absolute():
        return path
    return base_dir / path


def _to_float(value: object) -> float | None:
    text = _as_text(value)
    if not text:
        return None
    try:
        out = float(text)
    except ValueError:
        return None
    return out if out == out else None


def _normalize_yes_no(value: object) -> str:
    text = _as_text(value).lower()
    if text in {"yes", "y", "true", "1"}:
        return "Yes"
    return "No"


def _safe_json_dict(value: object) -> dict[str, object]:
    text = _as_text(value)
    if not text:
        return {}
    try:
        parsed = json.loads(text)
    except Exception:
        return {}
    return parsed if isinstance(parsed, dict) else {}


def _load_epics_from_db(db_path: Path) -> tuple[list[dict[str, object]], list[dict[str, object]]]:
    if not db_path.exists():
        return [], []
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    try:
        table_exists = conn.execute(
            "SELECT 1 FROM sqlite_master WHERE type='table' AND name='epics_management'"
        ).fetchone()
        if not table_exists:
            return [], []
        schema = conn.execute("PRAGMA table_info(epics_management)").fetchall()
        column_names = {str(item[1]) for item in schema}
        if "ipp_meeting_planned" not in column_names:
            return [], []
        remarks_select = "remarks" if "remarks" in column_names else "'' AS remarks"
        rows = conn.execute(
            f"""
            SELECT
                epic_key, project_key, project_name, product_category, epic_name,
                description, {remarks_select}, originator, priority, plan_status, ipp_meeting_planned, jira_url,
                epic_plan_json, research_urs_plan_json, dds_plan_json,
                development_plan_json, sqa_plan_json, production_plan_json
            FROM epics_management
            ORDER BY lower(project_name) ASC, lower(product_category) ASC, lower(epic_name) ASC, epic_key ASC
            """
        ).fetchall()
    finally:
        conn.close()

    all_rows: list[dict[str, object]] = []
    selected_rows: list[dict[str, object]] = []
    for row in rows:
        item = {
            "epic_key": _as_text(row["epic_key"]).upper(),
            "project_key": _as_text(row["project_key"]).upper(),
            "project_name": _as_text(row["project_name"]),
            "product_category": _as_text(row["product_category"]),
            "epic_name": _as_text(row["epic_name"]),
            "description": _as_text(row["description"]),
            "remarks": _as_text(row["remarks"]),
            "originator": _as_text(row["originator"]),
            "priority": _as_text(row["priority"]),
            "plan_status": _as_text(row["plan_status"]),
            "jira_url": _as_text(row["jira_url"]),
            "ipp_meeting_planned": _normalize_yes_no(row["ipp_meeting_planned"]),
            "plans": {
                "epic_plan": _safe_json_dict(row["epic_plan_json"]),
                "research_urs_plan": _safe_json_dict(row["research_urs_plan_json"]),
                "dds_plan": _safe_json_dict(row["dds_plan_json"]),
                "development_plan": _safe_json_dict(row["development_plan_json"]),
                "sqa_plan": _safe_json_dict(row["sqa_plan_json"]),
                "production_plan": _safe_json_dict(row["production_plan_json"]),
            },
        }
        all_rows.append(item)
        if item["ipp_meeting_planned"] == "Yes":
            selected_rows.append(item)
    return selected_rows, all_rows


def _is_story_issue_type(issue_type: str) -> bool:
    normalized = _as_text(issue_type).lower()
    return "story" in normalized


def _load_jira_rows_by_epic(
    work_items_path: Path,
) -> tuple[dict[str, dict[str, object]], dict[str, list[dict[str, object]]]]:
    epic_rows: dict[str, dict[str, object]] = {}
    stories_by_epic: dict[str, list[dict[str, object]]] = {}
    if not work_items_path.exists():
        return epic_rows, stories_by_epic
    wb = load_workbook(work_items_path, read_only=True, data_only=True)
    try:
        ws = wb.active
        header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header:
            return epic_rows, stories_by_epic
        headers = [_as_text(item) for item in header]
        idx = {name: pos for pos, name in enumerate(headers)}
        required = ["issue_key"]
        if any(name not in idx for name in required):
            return epic_rows, stories_by_epic

        for row in ws.iter_rows(min_row=2, values_only=True):
            issue_key = _as_text(row[idx["issue_key"]]).upper()
            if not issue_key:
                continue
            issue_type = _as_text(row[idx.get("jira_issue_type", -1)] if "jira_issue_type" in idx else "")
            if not issue_type:
                issue_type = _as_text(row[idx.get("work_item_type", -1)] if "work_item_type" in idx else "")

            epic_key = _as_text(row[idx.get("epic_key", -1)] if "epic_key" in idx else "").upper()
            parent_issue_key = _as_text(row[idx.get("parent_issue_key", -1)] if "parent_issue_key" in idx else "").upper()
            linked_epic_key = epic_key or parent_issue_key
            estimate = _to_float(row[idx.get("original_estimate_hours", -1)] if "original_estimate_hours" in idx else "")
            logged = _to_float(row[idx.get("total_hours_logged", -1)] if "total_hours_logged" in idx else "")
            progress_pct = None
            if estimate and estimate > 0 and logged is not None:
                progress_pct = round(min(100.0, (logged / estimate) * 100.0), 2)

            if "epic" in issue_type.lower():
                epic_rows[issue_key] = {
                    "issue_key": issue_key,
                    "project_key": _as_text(row[idx.get("project_key", -1)] if "project_key" in idx else "").upper(),
                    "summary": _as_text(row[idx.get("summary", -1)] if "summary" in idx else ""),
                    "status": _as_text(row[idx.get("status", -1)] if "status" in idx else ""),
                    "assignee": _as_text(row[idx.get("assignee", -1)] if "assignee" in idx else ""),
                    "jira_url": _as_text(row[idx.get("jira_url", -1)] if "jira_url" in idx else ""),
                    "start_date": _as_text(row[idx.get("start_date", -1)] if "start_date" in idx else ""),
                    "end_date": _as_text(row[idx.get("end_date", -1)] if "end_date" in idx else ""),
                    "actual_end_date": _as_text(row[idx.get("actual_end_date", -1)] if "actual_end_date" in idx else ""),
                    "ipp_actual_date": _as_text(
                        row[idx.get("IPP Actual Date (Production Date)", -1)] if "IPP Actual Date (Production Date)" in idx else ""
                    ),
                    "ipp_remarks": _as_text(row[idx.get("IPP Remarks", -1)] if "IPP Remarks" in idx else ""),
                    "original_estimate_hours": estimate,
                    "total_hours_logged": logged,
                    "progress_pct": progress_pct,
                }
                continue

            if not _is_story_issue_type(issue_type):
                continue
            if not linked_epic_key:
                continue
            stories_by_epic.setdefault(linked_epic_key, []).append(
                {
                    "story_key": issue_key,
                    "story_type": issue_type,
                    "story_name": _as_text(row[idx.get("summary", -1)] if "summary" in idx else ""),
                    "story_status": _as_text(row[idx.get("status", -1)] if "status" in idx else ""),
                    "story_assignee": _as_text(row[idx.get("assignee", -1)] if "assignee" in idx else ""),
                    "story_jira_url": _as_text(row[idx.get("jira_url", -1)] if "jira_url" in idx else ""),
                    "story_start_date": _as_text(row[idx.get("start_date", -1)] if "start_date" in idx else ""),
                    "story_end_date": _as_text(row[idx.get("end_date", -1)] if "end_date" in idx else ""),
                    "story_actual_end_date": _as_text(row[idx.get("actual_end_date", -1)] if "actual_end_date" in idx else ""),
                    "story_planned_hours": estimate,
                    "story_logged_hours": logged,
                    "story_progress_pct": progress_pct,
                }
            )
    finally:
        wb.close()
    for epic_key, story_rows in stories_by_epic.items():
        stories_by_epic[epic_key] = sorted(
            story_rows,
            key=lambda item: (
                _as_text(item.get("story_start_date")) or "9999-12-31",
                _as_text(item.get("story_end_date")) or "9999-12-31",
                _as_text(item.get("story_key")),
            ),
        )
    return epic_rows, stories_by_epic


def _phase_record_from_plan(phase_name: str, plan: dict[str, object]) -> dict[str, object]:
    start_iso = _as_text(plan.get("start_date"))
    end_iso = _as_text(plan.get("due_date"))
    start_date = _parse_iso_date(start_iso)
    end_date = _parse_iso_date(end_iso)
    mandays_num = _to_number(plan.get("man_days"))
    mandays_text = "" if mandays_num is None else str(mandays_num).rstrip("0").rstrip(".")

    warning = ""
    state = "no_entry"
    if start_iso or end_iso or mandays_text:
        if start_date and end_date and start_date <= end_date:
            state = "planned"
        else:
            state = "invalid"
            warning = "missing_or_invalid_date_range"
            if start_date and end_date and start_date > end_date:
                warning = "start_after_end"

    raw = ""
    if start_iso or end_iso or mandays_text:
        raw = f"{start_iso or '-'} to {end_iso or '-'} ({mandays_text or '-'} md)"

    return {
        "name": phase_name,
        "state": state,
        "state_label": state.replace("_", " "),
        "warning": warning,
        "start_iso": start_iso if start_date else "",
        "end_iso": end_iso if end_date else "",
        "start_date": start_date,
        "end_date": end_date,
        "mandays_text": mandays_text,
        "mandays_num": mandays_num,
        "raw": raw,
    }


def _build_records(
    selected_epics: list[dict[str, object]],
    jira_rows_by_epic: dict[str, dict[str, object]],
    jira_stories_by_epic: dict[str, list[dict[str, object]]],
) -> list[dict[str, object]]:
    records: list[dict[str, object]] = []
    for i, epic in enumerate(selected_epics, start=1):
        epic_key = _as_text(epic.get("epic_key")).upper()
        jira = jira_rows_by_epic.get(epic_key, {})
        plans = epic.get("plans") if isinstance(epic.get("plans"), dict) else {}
        epic_plan = plans.get("epic_plan") if isinstance(plans.get("epic_plan"), dict) else {}

        db_epic_start_iso = _as_text(epic_plan.get("start_date"))
        db_epic_end_iso = _as_text(epic_plan.get("due_date"))
        jira_epic_start_iso = _as_text(jira.get("start_date"))
        jira_epic_end_iso = _as_text(jira.get("end_date"))
        epic_start_iso = db_epic_start_iso or jira_epic_start_iso
        epic_end_iso = db_epic_end_iso or jira_epic_end_iso
        epic_actual_iso = _as_text(jira.get("ipp_actual_date")) or _as_text(jira.get("actual_end_date"))
        db_epic_mandays = _to_number(epic_plan.get("man_days"))
        db_epic_planned_hours = round(db_epic_mandays * 8.0, 4) if db_epic_mandays is not None else None

        phases = []
        for phase_name in PHASE_NAMES:
            plan_key = PLAN_KEY_BY_PHASE.get(phase_name, "")
            plan = plans.get(plan_key) if plan_key and isinstance(plans.get(plan_key), dict) else {}
            phases.append(_phase_record_from_plan(phase_name, plan))

        total_mandays = sum((p.get("mandays_num") or 0.0) for p in phases)
        epic_start_date = _parse_iso_date(epic_start_iso)
        epic_end_date = _parse_iso_date(epic_end_iso)
        epic_actual_date = _parse_iso_date(epic_actual_iso)
        has_valid_epic_plan = bool(epic_start_date and epic_end_date and epic_start_date <= epic_end_date)

        product = _as_text(epic.get("product_category")) or _as_text(epic.get("project_name")) or "Unmapped"
        remarks = _as_text(epic.get("remarks"))
        jira_link = _as_text(epic.get("jira_url")) or _as_text(jira.get("jira_url")) or _normalize_jira_link(epic_key)

        source_sheet = _as_text(epic.get("_record_source")) or "Epics Planner DB"
        story_rows = jira_stories_by_epic.get(epic_key, [])

        records.append(
            {
                "source_sheet": source_sheet,
                "row_number": i,
                "base": {
                    "Product": product,
                    "Epic/RMI": epic_key,
                    "Epic/RMI Jira Link": jira_link,
                    "Epic Planned Start Date": epic_start_iso,
                    "Epic Planned End Date": epic_end_iso,
                    "Epic Planned Start Date (DB)": db_epic_start_iso,
                    "Epic Planned End Date (DB)": db_epic_end_iso,
                    "Epic Planned Start Date (Jira Excel)": jira_epic_start_iso,
                    "Epic Planned End Date (Jira Excel)": jira_epic_end_iso,
                    "Epic Actual Date (Production Date)": epic_actual_iso,
                    "Remarks": remarks,
                },
                "phases": phases,
                "total_mandays": total_mandays,
                "computed_has_valid_epic_plan": "Yes" if has_valid_epic_plan else "No",
                "epic_start_date": epic_start_date,
                "epic_end_date": epic_end_date,
                "epic_actual_date": epic_actual_date,
                "jira_status": _as_text(jira.get("status")),
                "jira_assignee": _as_text(jira.get("assignee")),
                "db_epic_planned_mandays": db_epic_mandays,
                "db_epic_planned_hours": db_epic_planned_hours,
                "jira_original_estimate_hours": jira.get("original_estimate_hours"),
                "jira_total_hours_logged": jira.get("total_hours_logged"),
                "jira_progress_pct": jira.get("progress_pct"),
                "epic_name": _as_text(epic.get("epic_name")) or _as_text(jira.get("summary")),
                "project_name": _as_text(epic.get("project_name")),
                "plan_status": _as_text(epic.get("plan_status")),
                "priority": _as_text(epic.get("priority")),
                "stories": story_rows,
            }
        )
    return records


def _rows_for_payload(records: list[dict[str, object]]) -> list[dict[str, object]]:
    if not records:
        return []
    global_max_mandays = max((p["mandays_num"] or 0.0 for r in records for p in r["phases"]), default=0.0)
    roadmap_axis = _compute_roadmap_axis(records)

    out_rows: list[dict[str, object]] = []
    for r in records:
        if roadmap_axis["has_axis"] and r["computed_has_valid_epic_plan"] == "Yes":
            axis_start = roadmap_axis["axis_start"]
            axis_span_days = roadmap_axis["axis_span_days"]
            left = _clamp_percent(((r["epic_start_date"] - axis_start).days / max(1, axis_span_days - 1)) * 100.0)
            right = _clamp_percent(((r["epic_end_date"] - axis_start).days / max(1, axis_span_days - 1)) * 100.0)
            bar_left = round(left, 4)
            bar_width = round(max(SMALL_MIN_WIDTH_PCT, right - left), 4)
        else:
            bar_left = ""
            bar_width = ""

        actual_left = ""
        if roadmap_axis["has_axis"] and isinstance(r.get("epic_actual_date"), date):
            axis_start = roadmap_axis["axis_start"]
            axis_span_days = roadmap_axis["axis_span_days"]
            actual_pct = _clamp_percent(((r["epic_actual_date"] - axis_start).days / max(1, axis_span_days - 1)) * 100.0)
            actual_left = round(actual_pct, 4)

        mini = _compute_phase_geometry_for_record(r, global_max_mandays)
        out_rows.append(
            {
                "source_sheet": r["source_sheet"],
                "row_number": r["row_number"],
                "product": r["base"]["Product"],
                "epic_rmi": r["base"]["Epic/RMI"],
                "epic_name": r.get("epic_name", ""),
                "project_name": r.get("project_name", ""),
                "plan_status": r.get("plan_status", ""),
                "priority": r.get("priority", ""),
                "jira_link": r["base"]["Epic/RMI Jira Link"],
                "epic_planned_start_date": r["base"]["Epic Planned Start Date"],
                "epic_planned_end_date": r["base"]["Epic Planned End Date"],
                "epic_planned_start_date_db": r["base"]["Epic Planned Start Date (DB)"],
                "epic_planned_end_date_db": r["base"]["Epic Planned End Date (DB)"],
                "epic_planned_start_date_jira": r["base"]["Epic Planned Start Date (Jira Excel)"],
                "epic_planned_end_date_jira": r["base"]["Epic Planned End Date (Jira Excel)"],
                "epic_actual_date": r["base"]["Epic Actual Date (Production Date)"],
                "remarks": r["base"]["Remarks"],
                "computed_total_mandays": round(r["total_mandays"], 4),
                "jira_status": r.get("jira_status", ""),
                "jira_assignee": r.get("jira_assignee", ""),
                "db_epic_planned_mandays": r.get("db_epic_planned_mandays"),
                "epic_planned_hours_db": r.get("db_epic_planned_hours"),
                "epic_planned_hours_jira": r.get("jira_original_estimate_hours"),
                "jira_original_estimate_hours": r.get("jira_original_estimate_hours"),
                "jira_total_hours_logged": r.get("jira_total_hours_logged"),
                "jira_progress_pct": r.get("jira_progress_pct"),
                "phase_data": {
                    p["name"]: {
                        "start": p["start_iso"],
                        "end": p["end_iso"],
                        "mandays": p["mandays_text"],
                        "raw": p["raw"],
                        "state": p["state"],
                        "warning": p["warning"],
                    }
                    for p in r["phases"]
                },
                "stories": r.get("stories", []),
                "roadmap": {
                    "valid": r["computed_has_valid_epic_plan"] == "Yes",
                    "axis_start_iso": roadmap_axis["axis_start"].isoformat() if roadmap_axis["has_axis"] else "",
                    "axis_end_iso": roadmap_axis["axis_end"].isoformat() if roadmap_axis["has_axis"] else "",
                    "axis_span_days": roadmap_axis["axis_span_days"] if roadmap_axis["has_axis"] else 0,
                    "today_in_range": bool(roadmap_axis.get("today_in_range")),
                    "today_left_pct": roadmap_axis.get("today_left_pct", ""),
                    "bar_left_pct": bar_left,
                    "bar_width_pct": bar_width,
                    "actual_left_pct": actual_left,
                    "week_ticks": roadmap_axis.get("week_ticks", []),
                },
                "mini_gantt": {
                    "has_dated_phases": bool(mini["has_dated_phases"]),
                    "axis_start_iso": mini["axis_start"].isoformat() if isinstance(mini["axis_start"], date) else "",
                    "axis_end_iso": mini["axis_end"].isoformat() if isinstance(mini["axis_end"], date) else "",
                    "axis_span_days": mini["axis_span_days"],
                    "timeline_width_px": mini["timeline_width_px"],
                    "scroll_enabled": bool(mini["scroll_enabled"]),
                    "week_ticks": mini["week_ticks"],
                    "today_in_range": bool(mini["today_in_range"]),
                    "today_left_pct": mini["today_left_pct"],
                    "phases": mini["phases"],
                },
            }
        )
    return out_rows


def _build_payload(rows: list[dict[str, object]], db_path: Path, work_items_path: Path) -> dict[str, object]:
    return {
        "generated_at": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC"),
        "source_workbook": str(work_items_path),
        "source_sheet": "Epics Planner DB + Jira WorkItems Export",
        "settings_db_path": str(db_path),
        "phase_names": [name for name, _ in PHASE_COLUMNS],
        "rows": rows,
    }


def build_payload_from_sources(base_dir: Path) -> dict[str, object]:
    settings_db_value = os.getenv("JIRA_ASSIGNEE_HOURS_CAPACITY_DB_PATH", DEFAULT_SETTINGS_DB).strip() or DEFAULT_SETTINGS_DB
    settings_db_path = _resolve_path(settings_db_value, base_dir)
    work_items_value = os.getenv("JIRA_EXPORT_XLSX_PATH", DEFAULT_WORK_ITEMS_XLSX).strip() or DEFAULT_WORK_ITEMS_XLSX
    work_items_path = _resolve_path(work_items_value, base_dir)

    selected_epics, all_epics = _load_epics_from_db(settings_db_path)
    jira_rows, jira_stories = _load_jira_rows_by_epic(work_items_path)
    epics_to_render = selected_epics

    records = _build_records(epics_to_render, jira_rows, jira_stories)
    rows = _rows_for_payload(records)
    payload = _build_payload(rows, settings_db_path, work_items_path)
    payload["selection_mode"] = "selected_only"
    payload["selected_epics_count"] = len(selected_epics)
    payload["total_epics_count"] = len(all_epics)
    return payload


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Generate IPP Meeting dashboard HTML from Epics Planner DB + Jira exports."
    )
    parser.add_argument("--input-xlsx", default="", help="Deprecated (kept for compatibility).")
    parser.add_argument("--output-html", default="", help="Dashboard output HTML path override.")
    parser.add_argument("--output-dir", default="", help="Deprecated (kept for compatibility).")
    return parser.parse_args()


def main() -> None:
    args = _parse_args()
    base_dir = Path(__file__).resolve().parent

    output_html_value = args.output_html.strip() or os.getenv("IPP_PHASE_DASHBOARD_HTML_PATH", "").strip() or DEFAULT_HTML_OUTPUT
    output_html_path = _resolve_path(output_html_value, base_dir)

    template_path = base_dir / DEFAULT_TEMPLATE
    if not template_path.exists():
        raise FileNotFoundError(f"Dashboard template not found: {template_path}")

    payload = build_payload_from_sources(base_dir)

    json_blob = json.dumps(payload, default=str)
    json_blob = json_blob.replace("</script", "<\\/script").replace("</SCRIPT", "<\\/SCRIPT")

    template = template_path.read_text(encoding="utf-8")
    token = "__IPP_PHASE_DATA__"
    if token not in template:
        raise ValueError(f"Template missing data placeholder token: {token}")
    html = template.replace(token, json_blob)
    output_html_path.write_text(html, encoding="utf-8")

    print(f"Settings DB: {payload.get('settings_db_path', '')}")
    print(f"Jira work-items export: {payload.get('source_workbook', '')}")
    print(
        "Selected epics: "
        f"{payload.get('selected_epics_count', 0)} / total epics in DB: {payload.get('total_epics_count', 0)} "
        f"(mode={payload.get('selection_mode', '')})"
    )
    print(f"Output HTML: {output_html_path}")


if __name__ == "__main__":
    main()
