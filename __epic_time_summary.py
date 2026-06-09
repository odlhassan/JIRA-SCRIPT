"""
Reads the Jira export CSV, looks up the epic for every parent story via the
Jira REST API, then writes an Excel workbook showing total time spent (hours
and days) per epic.

Usage:
    python __epic_time_summary.py
"""

import csv
import os
import requests
import base64
from collections import defaultdict
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Configuration ─────────────────────────────────────────────────────────────
CSV_PATH = r"c:\Users\hmalik\Downloads\Jira (9).csv"
OUTPUT_PATH = r"c:\Users\hmalik\Downloads\epic_time_summary_9.xlsx"
WORKING_HOURS_PER_DAY = 8   # 1 day = 8 h for the "days" column

# Jira credentials (read from .env next to this script)
_ENV = {}
_env_file = Path(__file__).parent / ".env"
if _env_file.exists():
    for line in _env_file.read_text().splitlines():
        line = line.strip()
        if line and not line.startswith("#") and "=" in line:
            k, _, v = line.partition("=")
            _ENV[k.strip()] = v.strip()

JIRA_SITE  = _ENV.get("JIRA_SITE", "")
JIRA_EMAIL = _ENV.get("JIRA_EMAIL", "")
JIRA_TOKEN = _ENV.get("JIRA_API_TOKEN", "")
JIRA_BASE  = f"https://{JIRA_SITE}.atlassian.net"

AUTH = base64.b64encode(f"{JIRA_EMAIL}:{JIRA_TOKEN}".encode()).decode()
HEADERS = {
    "Authorization": f"Basic {AUTH}",
    "Accept": "application/json",
}

# ── Step 1: read CSV ──────────────────────────────────────────────────────────
print(f"Reading {CSV_PATH} …")
rows = []
with open(CSV_PATH, newline="", encoding="utf-8-sig") as f:
    reader = csv.DictReader(f)
    for row in reader:
        rows.append(row)

print(f"  {len(rows)} subtask rows loaded.")

# Collect per-parent aggregates  {parent_key -> {seconds, project_key}}
parent_seconds: dict[str, int] = defaultdict(int)
parent_project: dict[str, str] = {}

for row in rows:
    parent_key = (row.get("Parent key") or "").strip()
    project_key = (row.get("Project key") or "").strip()
    raw_spent = (row.get("Σ Time Spent") or "0").strip()
    try:
        seconds = int(raw_spent)
    except ValueError:
        seconds = 0

    if parent_key:
        parent_seconds[parent_key] += seconds
        parent_project[parent_key] = project_key

print(f"  {len(parent_seconds)} unique parent (story) keys.")

# ── Step 2: look up epics for each story via Jira API ────────────────────────
# We request the parent field (next-gen) and customfield_10014 (classic epic link).
# For classic projects the epic is stored in customfield_10014 (Epic Link key).
# For next-gen / team-managed the story's parent IS the epic.

FIELDS = "summary,parent,customfield_10014,issuetype"

story_to_epic_key: dict[str, str] = {}
story_to_epic_name: dict[str, str] = {}

_epic_name_cache: dict[str, str] = {}

def fetch_epic_name(epic_key: str) -> str:
    if epic_key in _epic_name_cache:
        return _epic_name_cache[epic_key]
    url = f"{JIRA_BASE}/rest/api/3/issue/{epic_key}?fields=summary"
    try:
        r = requests.get(url, headers=HEADERS, timeout=15)
        if r.ok:
            name = r.json().get("fields", {}).get("summary", epic_key)
        else:
            name = epic_key
    except Exception:
        name = epic_key
    _epic_name_cache[epic_key] = name
    return name

total = len(parent_seconds)
print(f"\nFetching epic info for {total} stories from Jira …")

for idx, story_key in enumerate(sorted(parent_seconds.keys()), 1):
    print(f"  [{idx}/{total}] {story_key}", end=" … ", flush=True)
    url = f"{JIRA_BASE}/rest/api/3/issue/{story_key}?fields={FIELDS}"
    try:
        r = requests.get(url, headers=HEADERS, timeout=15)
    except Exception as exc:
        print(f"ERROR ({exc})")
        story_to_epic_key[story_key] = "Unknown"
        story_to_epic_name[story_key] = "Unknown"
        continue

    if not r.ok:
        print(f"HTTP {r.status_code}")
        story_to_epic_key[story_key] = "Unknown"
        story_to_epic_name[story_key] = "Unknown"
        continue

    data = r.json()
    fields = data.get("fields", {})
    issue_type = (fields.get("issuetype") or {}).get("name", "")

    # If this issue itself IS an epic, it is its own epic
    if issue_type.lower() == "epic":
        epic_key  = story_key
        epic_name = fields.get("summary", story_key)
        print(f"(is Epic) {epic_name}")
    # Classic: Epic Link field
    elif fields.get("customfield_10014"):
        epic_key  = fields["customfield_10014"]
        epic_name = fetch_epic_name(epic_key)
        print(f"classic→ {epic_key}: {epic_name}")
    # Next-gen / team-managed: parent is the epic
    elif fields.get("parent"):
        parent_info = fields["parent"]
        parent_type = (parent_info.get("fields") or {}).get("issuetype", {}).get("name", "")
        if parent_type.lower() == "epic":
            epic_key  = parent_info["key"]
            epic_name = (parent_info.get("fields") or {}).get("summary", epic_key)
            print(f"parent→ {epic_key}: {epic_name}")
        else:
            # Parent is a story/task — fetch it to find its epic
            grandparent_key = parent_info["key"]
            gp_url = f"{JIRA_BASE}/rest/api/3/issue/{grandparent_key}?fields=summary,parent,customfield_10014,issuetype"
            try:
                gpr = requests.get(gp_url, headers=HEADERS, timeout=15)
                if gpr.ok:
                    gpf = gpr.json().get("fields", {})
                    if gpf.get("customfield_10014"):
                        epic_key  = gpf["customfield_10014"]
                        epic_name = fetch_epic_name(epic_key)
                    elif gpf.get("parent"):
                        epic_key  = gpf["parent"]["key"]
                        epic_name = (gpf["parent"].get("fields") or {}).get("summary", epic_key)
                    else:
                        epic_key  = grandparent_key
                        epic_name = gpf.get("summary", grandparent_key)
                else:
                    epic_key  = grandparent_key
                    epic_name = grandparent_key
            except Exception:
                epic_key  = grandparent_key
                epic_name = grandparent_key
            print(f"grandparent→ {epic_key}: {epic_name}")
    else:
        epic_key  = "No Epic"
        epic_name = "No Epic"
        print("no epic found")

    story_to_epic_key[story_key]  = epic_key
    story_to_epic_name[story_key] = epic_name

# ── Step 3: aggregate by epic ─────────────────────────────────────────────────
epic_seconds: dict[str, int]   = defaultdict(int)
epic_names:   dict[str, str]   = {}
epic_project: dict[str, str]   = {}

for story_key, seconds in parent_seconds.items():
    ekey  = story_to_epic_key.get(story_key, "Unknown")
    ename = story_to_epic_name.get(story_key, "Unknown")
    proj  = parent_project.get(story_key, "")
    epic_seconds[ekey] += seconds
    epic_names[ekey]    = ename
    epic_project[ekey]  = proj

# Also build a detailed sheet: one row per story
detail_rows = []
for story_key in sorted(parent_seconds.keys()):
    secs  = parent_seconds[story_key]
    ekey  = story_to_epic_key.get(story_key, "Unknown")
    ename = story_to_epic_name.get(story_key, "Unknown")
    proj  = parent_project.get(story_key, "")
    detail_rows.append({
        "Project":        proj,
        "Epic Key":       ekey,
        "Epic Name":      ename,
        "Story Key":      story_key,
        "Time Spent (s)": secs,
        "Hours":          round(secs / 3600, 2),
        "Days":           round(secs / 3600 / WORKING_HOURS_PER_DAY, 2),
    })

# ── Step 4: write Excel ───────────────────────────────────────────────────────
print(f"\nWriting Excel to {OUTPUT_PATH} …")
wb = openpyxl.Workbook()

# ── Sheet 1: Summary by Epic ──────────────────────────────────────────────────
ws1 = wb.active
ws1.title = "Summary by Epic"

HDR_FILL  = PatternFill("solid", fgColor="1F497D")
HDR_FONT  = Font(color="FFFFFF", bold=True, size=11)
ALT_FILL  = PatternFill("solid", fgColor="DCE6F1")
BORDER    = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)
CENTER    = Alignment(horizontal="center", vertical="center")
LEFT      = Alignment(horizontal="left",   vertical="center", wrap_text=True)

headers1 = ["#", "Project", "Epic Key", "Epic Name", "Total Time (s)", "Total Hours", "Total Days"]
ws1.append(headers1)
for col, h in enumerate(headers1, 1):
    cell = ws1.cell(row=1, column=col)
    cell.fill   = HDR_FILL
    cell.font   = HDR_FONT
    cell.border = BORDER
    cell.alignment = CENTER

# Sort by project then epic key
summary_rows = sorted(epic_seconds.items(), key=lambda x: (epic_project.get(x[0],""), x[0]))
for i, (ekey, secs) in enumerate(summary_rows, 1):
    row = [
        i,
        epic_project.get(ekey, ""),
        ekey,
        epic_names.get(ekey, ""),
        secs,
        round(secs / 3600, 2),
        round(secs / 3600 / WORKING_HOURS_PER_DAY, 2),
    ]
    ws1.append(row)
    fill = ALT_FILL if i % 2 == 0 else PatternFill()
    for col in range(1, len(headers1) + 1):
        cell = ws1.cell(row=i + 1, column=col)
        cell.border    = BORDER
        cell.alignment = CENTER if col != 4 else LEFT
        if i % 2 == 0:
            cell.fill = ALT_FILL

# Column widths
col_widths1 = [5, 12, 18, 55, 18, 14, 12]
for col, w in enumerate(col_widths1, 1):
    ws1.column_dimensions[get_column_letter(col)].width = w
ws1.row_dimensions[1].height = 20
ws1.freeze_panes = "A2"

# ── Sheet 2: Story-level Detail ───────────────────────────────────────────────
ws2 = wb.create_sheet("Story Detail")
headers2 = ["#", "Project", "Epic Key", "Epic Name", "Story Key", "Time Spent (s)", "Hours", "Days"]
ws2.append(headers2)
for col, h in enumerate(headers2, 1):
    cell = ws2.cell(row=1, column=col)
    cell.fill   = HDR_FILL
    cell.font   = HDR_FONT
    cell.border = BORDER
    cell.alignment = CENTER

detail_rows_sorted = sorted(detail_rows, key=lambda r: (r["Project"], r["Epic Key"], r["Story Key"]))
for i, dr in enumerate(detail_rows_sorted, 1):
    row = [i, dr["Project"], dr["Epic Key"], dr["Epic Name"], dr["Story Key"],
           dr["Time Spent (s)"], dr["Hours"], dr["Days"]]
    ws2.append(row)
    for col in range(1, len(headers2) + 1):
        cell = ws2.cell(row=i + 1, column=col)
        cell.border    = BORDER
        cell.alignment = CENTER if col != 4 else LEFT
        if i % 2 == 0:
            cell.fill = ALT_FILL

col_widths2 = [5, 12, 18, 55, 18, 18, 10, 10]
for col, w in enumerate(col_widths2, 1):
    ws2.column_dimensions[get_column_letter(col)].width = w
ws2.row_dimensions[1].height = 20
ws2.freeze_panes = "A2"

wb.save(OUTPUT_PATH)
print(f"\nDone!  Excel saved to:\n  {OUTPUT_PATH}")
print(f"\nSummary:")
print(f"  Epics found:       {len(epic_seconds)}")
print(f"  Stories processed: {len(parent_seconds)}")
total_secs = sum(epic_seconds.values())
print(f"  Grand total time:  {total_secs/3600:.1f} hours  /  {total_secs/3600/WORKING_HOURS_PER_DAY:.1f} days")
